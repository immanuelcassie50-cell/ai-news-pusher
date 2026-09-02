import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()

# Sheet 1: 课堂互动时间线
ws1 = wb.active
ws1.title = '课堂互动时间线'
headers1 = ['时间段', '模块', '活动名称', '活动类型', '时长', '物料需求', '备注']
data1 = [
    ['Day1 08:30-09:00', '开场', '签到与破冰', '签到', '30min', '签到表、姓名牌', ''],
    ['Day1 09:00-09:20', '模块1', 'AI信任测试', '小组讨论', '20min', '答题卡、PPT', ''],
    ['Day1 09:20-09:50', '模块1', 'AI幻觉概念讲解', '讲解+演示', '30min', 'PPT、演示道具', ''],
    ['Day1 09:50-10:15', '模块1', 'AI幻觉类型讨论', '小组活动', '25min', '案例卡片', ''],
    ['Day1 10:15-10:30', '', '茶歇', '休息', '15min', '茶点', ''],
    ['Day1 10:30-10:45', '模块2', '个人反思活动', '个人任务', '15min', '反思工作表', ''],
    ['Day1 10:45-11:15', '模块2', '四要素框架讲解', '讲解', '30min', 'PPT', ''],
    ['Day1 11:15-11:45', '模块2', '四要素应用练习', '小组任务', '30min', '练习工作表', ''],
    ['Day1 11:45-12:00', '模块2', '常见误区盘点', '引导讲解', '15min', 'PPT', ''],
    ['Day1 12:00-13:30', '', '午餐', '休息', '90min', '', ''],
    ['Day1 13:30-13:50', '模块3', '案例导入：信息追踪', '演示', '20min', 'PPT', ''],
    ['Day1 13:50-14:15', '模块3', '三步验证法讲解', '讲解', '25min', 'PPT', ''],
    ['Day1 14:15-14:45', '模块3', '三步验证法演练', '小组任务', '30min', '验证任务卡', ''],
    ['Day1 14:45-15:00', '', '茶歇', '休息', '15min', '茶点', ''],
    ['Day1 15:00-15:15', '模块4', '你用AI做什么', '小组讨论', '15min', '', ''],
    ['Day1 15:15-15:45', '模块4', 'AI能力边界', '引导讲解', '30min', 'PPT', ''],
    ['Day1 15:45-16:15', '模块4', '人机协作最佳实践', '案例讨论', '30min', '案例文本', ''],
    ['Day1 16:15-16:30', '模块4', '制定团队规范', '小组任务', '15min', '规范模板', ''],
    ['Day1 16:30-16:45', '', '当日复盘', '集体', '15min', '', ''],
    ['Day2 09:00-09:20', '模块5', '空白案例说明', '讲解', '20min', 'PPT', ''],
    ['Day2 09:20-09:50', '模块5', '空白案例1：新闻验证', '小组任务', '30min', '案例工作表', ''],
    ['Day2 09:50-10:20', '模块5', '空白案例2：商业报告', '小组任务', '30min', '案例工作表', ''],
    ['Day2 10:20-10:50', '模块5', '空白案例3：决策评估', '小组任务', '30min', '案例工作表', ''],
    ['Day2 10:50-11:05', '', '茶歇', '休息', '15min', '茶点', ''],
    ['Day2 11:05-11:35', '模块5', '小组成果分享', '汇报', '30min', '汇报模板', ''],
    ['Day2 11:35-11:55', '模块6', '四层提问模型讲解', '讲解', '20min', 'PPT', ''],
    ['Day2 11:55-12:20', '模块6', '四层提问设计练习', '配对任务', '25min', '练习工作表', ''],
    ['Day2 12:20-13:30', '', '午餐', '休息', '90min', '', ''],
    ['Day2 13:30-14:00', '模块6', '综合应用：AI审查', '小组活动', '30min', '场景卡片', ''],
    ['Day2 14:00-14:20', '模块6', '知识点回顾', '集体', '20min', '', ''],
    ['Day2 14:20-14:40', '模块7', '思维导图绘制', '个人任务', '20min', '思维导图模板', ''],
    ['Day2 14:40-15:05', '模块7', '团队应用设计', '小组任务', '25min', '设计模板', ''],
    ['Day2 15:05-15:35', '模块7', '30天行动计划', '个人任务', '30min', '行动计划模板', ''],
    ['Day2 15:35-15:50', '模块7', '课程总结与颁奖', '集体', '15min', '证书模板', ''],
]
for i, header in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=i, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='C41E3A', end_color='C41E3A', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for row_idx, row_data in enumerate(data1, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 20
ws1.column_dimensions['G'].width = 15

# Sheet 2: 活动设计详情
ws2 = wb.create_sheet('活动设计详情')
headers2 = ['活动名称', '活动目标', '实施步骤', '时间控制', '注意事项']
data2 = [
    ['破冰活动：AI信任测试', '让学员意识到自己可能对AI存在过度信任', '1.展示10道判断题 2.学员独立判断 3.公布答案统计正确率', '20分钟', '题目要有挑战性但不能太难'],
    ['个人反思：你今天接收了多少信息', '引导学员意识到信息过载的现状', '1.展示反思问题 2.学员独立思考 3.快速分享', '15分钟', '避免过于深入的哲学讨论'],
    ['四要素应用练习', '让学员练习使用批判思维四要素框架', '1.发放练习材料 2.小组讨论 3.小组汇报 4.讲师点评', '30分钟', '材料要足够复杂但有解'],
    ['三步验证法演练', '让学员掌握三步验证法的实际应用', '1.发放验证任务 2.小组执行验证 3.记录验证过程 4.分享发现', '30分钟', '任务要有真实的陷阱'],
    ['空白案例小组汇报', '锻炼学员的表达能力和批判思维应用', '1.小组准备汇报 2.小组代表汇报 3.其他组提问 4.讲师点评', '30分钟', '控制每个小组汇报时间'],
    ['四层提问设计练习', '让学员掌握四层提问的设计方法', '1.讲解四层模型 2.发放练习任务 3.配对设计 4.分享点评', '25分钟', '注意配对组合的合理性'],
    ['思维导图绘制', '帮助学员整合课程知识点', '1.发放思维导图模板 2.学员独立绘制 3.快速分享', '20分钟', '允许学员用自己的方式'],
    ['30天行动计划制定', '帮助学员将学习转化为行动', '1.发放行动计划模板 2.学员独立填写 3.分享承诺 4.互相监督', '30分钟', '强调承诺的可执行性'],
]
for i, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=i, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4A4A4A', end_color='4A4A4A', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for row_idx, row_data in enumerate(data2, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 25

# Sheet 3: 物料清单
ws3 = wb.create_sheet('物料清单')
headers3 = ['类别', '物料名称', '数量', '用途', '准备时间', '负责人', '状态']
data3 = [
    ['打印物料', '学员手册', '35份', '学员学习使用', '课前3天', '会务组', ''],
    ['打印物料', '答题卡', '35份', '破冰活动使用', '课前3天', '会务组', ''],
    ['打印物料', '空白案例工作表', '35份x5个', '案例练习', '课前3天', '会务组', ''],
    ['打印物料', '四要素框架图', '35份', '模块二学习', '课前3天', '会务组', ''],
    ['打印物料', '三步验证法流程图', '35份', '模块三学习', '课前3天', '会务组', ''],
    ['打印物料', '行动计划模板', '35份', '模块七使用', '课前3天', '会务组', ''],
    ['打印物料', '证书模板', '35份', '结业颁证', '课前3天', '会务组', ''],
    ['电子物料', 'PPT演示文件', '1套', '课程展示', '课前1天', '讲师', ''],
    ['电子物料', '案例视频素材', '3段', '案例展示', '课前1天', '讲师', ''],
    ['电子物料', '背景音乐', '若干', '茶歇休息时', '课前1天', '会务组', ''],
    ['场地布置', '分组桌牌', '6组x1', '小组标识', '课前半天', '会务组', ''],
    ['场地布置', '白板/白板纸', '6组x1', '小组讨论记录', '课前半天', '会务组', ''],
    ['场地布置', '计时器', '1个', '控制活动时间', '课前半天', '会务组', ''],
    ['茶歇', '茶点饮料', '35人份', '茶歇时间', '每天课前', '会务组', ''],
    ['备用', '备用电源', '2个', '设备应急', '课前检查', '会务组', ''],
    ['备用', '备用投影', '1台', '设备应急', '课前检查', '会务组', ''],
]
for i, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=i, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='6B6B6B', end_color='6B6B6B', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for row_idx, row_data in enumerate(data3, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 12
ws3.column_dimensions['G'].width = 10

wb.save('D:/新课开发/经验萃取/批判思维/完整课程包/10_成果demo/07_互动流程设计.xlsx')
print('07_互动流程设计.xlsx created successfully')
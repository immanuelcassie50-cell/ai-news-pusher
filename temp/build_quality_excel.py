#!/usr/bin/env python3
"""
Build Quality Forms Excel files for 质量文化重塑课程
Using openpyxl for creation - suitable for new files
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

OUTPUT_DIR = Path("D:/新课开发/制造/6-质量管理与质量文化重塑/配套表单Excel")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Color constants
RED = "C00000"
DARK_RED = "8B0000"
GRAY = "808080"
LIGHT_GRAY = "D9D9D9"
DARK_GRAY = "404040"
WHITE = "FFFFFF"
YELLOW = "FFFF00"
LIGHT_YELLOW = "FFFACD"
BLUE = "0000FF"
GREEN = "00B050"

def style_header(cell, text=None):
    """Apply header style (dark red background, white text)"""
    if text:
        cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_subheader(cell, text=None):
    """Apply subheader style (gray background)"""
    if text:
        cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_input(cell, text=None):
    """Apply input cell style (blue text for user input)"""
    if text is not None:
        cell.value = text
    cell.font = Font(color=BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_label(cell, text=None):
    """Apply label cell style"""
    if text is not None:
        cell.value = text
    cell.font = Font(color="000000")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_formula(cell, text=None):
    """Apply formula cell style (black text)"""
    if text is not None:
        cell.value = text
    cell.font = Font(color="000000")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_warning(cell, text=None):
    """Apply warning/important style (red text on yellow)"""
    if text is not None:
        cell.value = text
    cell.font = Font(bold=True, color=RED)
    cell.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def thin_border():
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def set_print_settings(ws, title=""):
    """Set A4 landscape print settings"""
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_title_rows = '1:3'

def add_note(ws, row, col, text):
    """Add a note/comment cell"""
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = Font(italic=True, color=GRAY, size=9)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ============================================================================
# Sheet 1: 质量类型判断卡
# ============================================================================
def create_sheet1_quality_type(wb):
    ws = wb.active
    ws.title = "质量类型判断卡"

    # Title
    ws.merge_cells('A1:F1')
    style_header(ws['A1'], "质量类型判断卡 - 四种质量问题类型快速识别")

    # Row 2: Subtitle/description
    ws.merge_cells('A2:F2')
    ws['A2'].value = "适用场景：判断质量问题的性质，选择合适的处理策略"
    ws['A2'].font = Font(italic=True, color=GRAY)
    ws['A2'].alignment = Alignment(horizontal="center")

    # Row 3: blank

    # Header row for type table
    row = 4
    headers = ["问题类型", "特征描述", "发生频率", "影响范围", "紧迫程度", "推荐行动"]
    for col, h in enumerate(headers, 1):
        style_subheader(ws.cell(row=row, column=col), h)

    # Data rows
    types_data = [
        ("突发问题", "突然发生，无预警\n之前从未出现过", "低", "可能广泛", "高", "立即响应，临时措施"),
        ("重复问题", "已知问题再次发生\n有历史记录", "高", "局部", "中", "按既定流程处理\n加强检验"),
        ("预期问题", "可预见的风险\n基于经验的预测", "中", "可控制", "低", "预防措施\n提前准备"),
        ("跨部门问题", "涉及多个部门\n责任不清", "不定", "广泛", "高", "协调会议\n明确责任"),
    ]

    for i, row_data in enumerate(types_data, 1):
        r = row + i
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border()
            if col == 1:
                cell.font = Font(bold=True, color=RED)

    # Decision flow section
    r = row + len(types_data) + 2
    ws.merge_cells(f'A{r}:F{r}')
    style_header(ws.cell(row=r, column=1), "快速判断流程")

    r += 1
    flow_steps = [
        ("Step 1", "这个问题以前出现过吗？", "否 → 突发问题 / 是 → 继续"),
        ("Step 2", "是否涉及多个部门？", "是 → 跨部门问题 / 否 → 继续"),
        ("Step 3", "可以提前预判吗？", "是 → 预期问题 / 否 → 重复问题"),
    ]
    for step, q, a in flow_steps:
        ws.cell(row=r, column=1, value=step).font = Font(bold=True, color=RED)
        ws.merge_cells(f'B{r}:C{r}')
        ws.cell(row=r, column=2, value=q).font = Font(bold=True)
        ws.merge_cells(f'D{r}:F{r}')
        ws.cell(row=r, column=4, value=a).font = Font(italic=True, color=GRAY)
        r += 1

    # Usage notes
    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    add_note(ws, r, 1, "使用说明：在遇到质量问题时，首先判断问题类型，然后根据类型选择相应的处理流程。四种类型可能同时存在，优先处理紧迫程度高的问题。")

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20

    set_print_settings(ws)


# ============================================================================
# Sheet 2: WSDF质量定义工作表
# ============================================================================
def create_sheet2_wsdf(wb):
    ws = wb.create_sheet("WSDF质量定义工作表")

    ws.merge_cells('A1:G1')
    style_header(ws['A1'], "WSDF质量定义工作表 - W/S/D/F问题分类与严重度评分")

    ws.merge_cells('A2:G2')
    ws['A2'].value = "W=微缺陷(Weep) S=轻微(Minor) D=严重(Defect) F=致命(Fatal)"
    ws['A2'].font = Font(italic=True, color=GRAY)
    ws['A2'].alignment = Alignment(horizontal="center")

    row = 4
    headers = ["问题编号", "问题描述", "W/S/D/F分类", "发生频率(次/月)", "严重度评分(1-5)", "综合风险分", "备注"]
    for col, h in enumerate(headers, 1):
        style_subheader(ws.cell(row=row, column=col), h)

    # Sample rows with formulas
    for i in range(1, 11):
        r = row + i
        ws.cell(row=r, column=1, value=f"Q{i:03d}").alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).value = ""  # User input
        ws.cell(row=r, column=3).value = ""  # User input W/S/D/F
        ws.cell(row=r, column=4).value = ""  # User input frequency
        ws.cell(row=r, column=5).value = ""  # User input severity 1-5
        # Formula: frequency * severity
        ws.cell(row=r, column=6).value = f"=IF(AND(D{r}<>\"\",E{r}<>\"\"),D{r}*E{r},\"\")"
        ws.cell(row=r, column=6).font = Font(color="000000")
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = thin_border()

    # Summary row
    r = row + 12
    ws.cell(row=r, column=1).value = "汇总"
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=4).value = f"=SUM(D{row+1}:D{r-1})"
    ws.cell(row=r, column=4).font = Font(bold=True)
    ws.cell(row=r, column=6).value = f"=AVERAGE(F{row+1}:F{r-1})"
    ws.cell(row=r, column=6).value = f"=IF(COUNT(F{row+1}:F{r-1})>0,AVERAGE(F{row+1}:F{r-1}),\"\")"
    ws.cell(row=r, column=6).font = Font(bold=True)

    # Risk threshold note
    r += 2
    ws.merge_cells(f'A{r}:G{r}')
    add_note(ws, r, 1, "风险分说明：综合风险分 = 发生频率 × 严重度评分。风险分 > 15 为高风险，需要立即处理；风险分 8-15 为中风险，需要关注；风险分 < 8 为低风险，可常规处理。")

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20

    set_print_settings(ws)


# ============================================================================
# Sheet 3: 五问法分析单
# ============================================================================
def create_sheet3_5why(wb):
    ws = wb.create_sheet("五问法分析单")

    ws.merge_cells('A1:E1')
    style_header(ws['A1'], "五问法分析单 - 5层追问找根因")

    ws.merge_cells('A2:E2')
    ws['A2'].value = "每次追问都要有证据支持，避免猜测"
    ws['A2'].font = Font(italic=True, color=GRAY)
    ws['A2'].alignment = Alignment(horizontal="center")

    row = 4
    headers = ["层次", "问题", "回答", "证据/依据", "结论"]
    for col, h in enumerate(headers, 1):
        style_subheader(ws.cell(row=row, column=col), h)

    why_levels = ["第1问", "第2问", "第3问", "第4问", "第5问"]
    for i, level in enumerate(why_levels):
        r = row + i + 1
        ws.cell(row=r, column=1, value=level).font = Font(bold=True, color=RED)
        ws.cell(row=r, column=2, value="为什么？（现象）")
        ws.cell(row=r, column=3).value = ""  # User input
        ws.cell(row=r, column=4).value = ""  # User input
        ws.cell(row=r, column=5).value = ""  # User input
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = thin_border()
            ws.cell(row=r, column=col).alignment = Alignment(vertical="top")

    # Root cause summary
    r = row + 7
    ws.merge_cells(f'A{r}:E{r}')
    style_subheader(ws.cell(row=r, column=1), "根因总结")

    r += 1
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(row=r, column=1).value = "根本原因："
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.merge_cells(f'B{r}:E{r}')
    ws.cell(row=r, column=2).value = ""  # User input
    ws.cell(row=r, column=2).border = thin_border()

    r += 1
    ws.cell(row=r, column=1).value = "证据链："
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.merge_cells(f'B{r}:E{r}')
    ws.cell(row=r, column=2).value = ""  # User input
    ws.cell(row=r, column=2).border = thin_border()
    ws.row_dimensions[r].height = 60
    ws.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25

    set_print_settings(ws)


# ============================================================================
# Sheet 4: 4M鱼骨图分析
# ============================================================================
def create_sheet4_fishbone(wb):
    ws = wb.create_sheet("4M鱼骨图分析")

    ws.merge_cells('A1:F1')
    style_header(ws['A1'], "4M鱼骨图分析 - Man/Machine/Material/Method")

    ws.merge_cells('A2:F2')
    ws['A2'].value = "在每个分支下填写可能的根本原因"
    ws['A2'].font = Font(italic=True, color=GRAY)
    ws['A2'].alignment = Alignment(horizontal="center")

    # 6 branches: Man, Machine, Material, Method, Measurement, Environment
    branches = [
        ("Man 人", ["操作员技能不足", "培训不到位", "疲劳/压力", "注意力不集中", "沟通不畅"]),
        ("Machine 机器", ["设备老化", "维护不当", "精度下降", "故障频发", "备件质量差"]),
        ("Material 材料", ["来料不良", "储存不当", "批次差异", "规格不符", "供应商问题"]),
        ("Method 方法", ["工艺参数不当", "作业标准不清晰", "流程缺陷", "设计问题", "方法陈旧"]),
        ("Measurement 测量", ["测量系统误差", "量具精度不足", "检测方法不当", "取样问题", "读数错误"]),
        ("Environment 环境", ["温度湿度不适", "光线不足", "噪音干扰", "空间拥挤", "6S不到位"]),
    ]

    row = 4
    for branch_name, sub_factors in branches:
        ws.merge_cells(f'A{row}:F{row}')
        style_subheader(ws.cell(row=row, column=1), branch_name)

        row += 1
        for i, factor in enumerate(sub_factors):
            ws.cell(row=row, column=1, value=f"  {i+1}.")
            ws.cell(row=row, column=2, value=factor)
            ws.merge_cells(f'C{row}:F{row}')
            ws.cell(row=row, column=3).value = ""  # User input for this factor
            ws.cell(row=row, column=3).border = thin_border()
            for col in [1, 2]:
                ws.cell(row=row, column=col).border = thin_border()
            row += 1

        row += 1  # blank row between branches

    # Root cause summary
    ws.merge_cells(f'A{row}:F{row}')
    style_subheader(ws.cell(row=row, column=1), "根因汇总（从上述分析中识别）")

    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1).value = "最可能的3个根因："
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for i in range(1, 4):
        ws.cell(row=row, column=1, value=f"根因{i}：")
        ws.merge_cells(f'B{row}:F{row}')
        ws.cell(row=row, column=2).value = ""
        ws.cell(row=row, column=2).border = thin_border()
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    set_print_settings(ws)


# ============================================================================
# Sheet 5: FAR方案对比评估表
# ============================================================================
def create_sheet5_far(wb):
    ws = wb.create_sheet("FAR方案对比评估表")

    ws.merge_cells('A1:H1')
    style_header(ws['A1'], "FAR方案对比评估表 - F(可行性)/A(适当性)/R(风险)三维评估")

    ws.merge_cells('A2:H2')
    ws['A2'].value = "评分标准：1-5分，5分最优"
    ws['A2'].font = Font(italic=True, color=GRAY)
    ws['A2'].alignment = Alignment(horizontal="center")

    row = 4
    headers = ["方案编号", "方案描述", "F可行性\n(1-5)", "A适当性\n(1-5)", "R风险\n(1-5)", "加权总分", "排名", "建议"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        style_subheader(cell)

    # Sample方案 rows (5 rows)
    for i in range(1, 6):
        r = row + i
        ws.cell(row=r, column=1, value=f"方案{i}").alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).value = ""  # User input
        ws.cell(row=r, column=3, value=3).font = Font(color=BLUE)  # Default value, user can change
        ws.cell(row=r, column=4, value=3).font = Font(color=BLUE)
        ws.cell(row=r, column=5, value=3).font = Font(color=BLUE)
        # Weighted total: F*0.4 + A*0.3 + R*0.3 (risk is negative, lower is better, so invert)
        ws.cell(row=r, column=6).value = f"=C{r}*0.4+D{r}*0.3+(5-E{r})*0.3"
        ws.cell(row=r, column=6).font = Font(bold=True)
        ws.cell(row=r, column=7).value = f"=RANK(F{r},F{row+1}:F{row+5},0)"
        # Suggestion based on rank
        ws.cell(row=r, column=8).value = f"=IF(F{r}=MAX($F${row+1}:$F${row+5}),\"推荐\",\"\")"
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = thin_border()

    # Weight explanation
    r = row + 7
    ws.merge_cells(f'A{r}:H{r}')
    add_note(ws, r, 1, "评分权重：F可行性40%，A适当性30%，R风险30%（风险评分已反向处理，高风险=低分）")

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 12

    set_print_settings(ws)


# ============================================================================
# Sheet 6: 质量放行决策卡 (CORE)
# ============================================================================
def create_sheet6_decision(wb):
    ws = wb.create_sheet("质量放行决策卡")

    ws.merge_cells('A1:G1')
    style_header(ws['A1'], "质量放行决策卡 - 核心决策工具")

    ws.merge_cells('A2:G2')
    ws['A2'].value = "三个问题，判断是否放行"
    ws['A2'].font = Font(italic=True, color=RED, size=12)
    ws['A2'].alignment = Alignment(horizontal="center")

    row = 4
    # Question 1
    ws.merge_cells(f'A{row}:G{row}')
    style_subheader(ws.cell(row=row, column=1), "问题1：这个问题以前出现过吗？")

    row += 1
    ws.cell(row=row, column=1, value="选项：")
    ws.cell(row=row, column=1).font = Font(bold=True)
    options = ["A. 从未出现过（突发问题）", "B. 以前出现过但已解决", "C. 反复出现（重复问题）"]
    for i, opt in enumerate(options):
        ws.cell(row=row, column=i*2+2, value=opt)
        ws.cell(row=row, column=i*2+2).border = thin_border()

    row += 2
    # Question 2
    ws.merge_cells(f'A{row}:G{row}')
    style_subheader(ws.cell(row=row, column=1), "问题2：这个问题的影响范围有多大？")

    row += 1
    ws.cell(row=row, column=1, value="选项：")
    ws.cell(row=row, column=1).font = Font(bold=True)
    options2 = ["A. 单件/单个位置", "B. 多件/多个位置", "C. 批量/整批问题"]
    for i, opt in enumerate(options2):
        ws.cell(row=row, column=i*2+2, value=opt)
        ws.cell(row=row, column=i*2+2).border = thin_border()

    row += 2
    # Question 3
    ws.merge_cells(f'A{row}:G{row}')
    style_subheader(ws.cell(row=row, column=1), "问题3：客户能接受这个问题吗？")

    row += 1
    ws.cell(row=row, column=1, value="选项：")
    ws.cell(row=row, column=1).font = Font(bold=True)
    options3 = ["A. 客户明确接受", "B. 需要特采审批", "C. 客户无法接受"]
    for i, opt in enumerate(options3):
        ws.cell(row=row, column=i*2+2, value=opt)
        ws.cell(row=row, column=i*2+2).border = thin_border()

    # Decision result
    row += 2
    ws.merge_cells(f'A{row}:G{row}')
    style_header(ws.cell(row=row, column=1), "决策建议")

    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    decision_text = """基于以上三个问题的回答，系统会自动给出放行建议：
    • 0个"C" → 放行（无需额外审批）
    • 1个"C" → 有条件放行（需要QA确认）
    • 2个"C" → 拒绝放行（必须改善）
    • 3个"C" → 立即停止（重大质量问题）"""
    ws.cell(row=row, column=1, value=decision_text)
    ws.cell(row=row, column=1).font = Font(color=DARK_GRAY)
    ws.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 80

    # Decision summary table
    row += 2
    headers = ["C的数量", "决策", "说明", "审批要求"]
    for col, h in enumerate(headers, 1):
        style_subheader(ws.cell(row=row, column=col*2-1))
        ws.merge_cells(f'{get_column_letter(col*2-1)}{row}:{get_column_letter(col*2)}{row}')
        ws.cell(row=row, column=col*2-1, value=h)

    decisions = [
        ("0个", "放行", "符合质量标准", "操作员自主决定"),
        ("1个", "有条件放行", "需QA确认", "QA工程师签字"),
        ("2个", "拒绝放行", "必须改善", "质量经理审批"),
        ("3个", "立即停止", "重大质量问题", "高管层审批"),
    ]
    for i, (c_cnt, decision, desc, approval) in enumerate(decisions):
        r = row + i + 1
        for col, val in enumerate([c_cnt, decision, desc, approval], 1):
            ws.merge_cells(f'{get_column_letter(col*2-1)}{r}:{get_column_letter(col*2)}{r}')
            ws.cell(row=r, column=col*2-1, value=val)
            ws.cell(row=r, column=col*2-1).border = thin_border()
            if decision in ["拒绝放行", "立即停止"]:
                ws.cell(row=r, column=col*2-1).font = Font(color=RED, bold=True)
            elif decision == "放行":
                ws.cell(row=r, column=col*2-1).font = Font(color=GREEN, bold=True)

    # Approval record
    row += 6
    ws.merge_cells(f'A{row}:G{row}')
    style_subheader(ws.cell(row=row, column=1), "审批记录")

    row += 1
    approval_headers = ["决策结果", "审批人", "审批时间", "签名"]
    for col, h in enumerate(approval_headers, 1):
        ws.cell(row=row, column=col, value=h)
        ws.cell(row=row, column=col).border = thin_border()
        style_subheader(ws.cell(row=row, column=col))

    row += 1
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = thin_border()
        ws.row_dimensions[row].height = 30

    # Column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14

    set_print_settings(ws)


# ============================================================================
# Sheet 7: 质量执行前检查清单
# ============================================================================
def create_sheet7_checklist(wb):
    ws = wb.create_sheet("质量执行前检查清单")

    ws.merge_cells('A1:E1')
    style_header(ws['A1'], "质量执行前检查清单")

    ws.merge_cells('A2:E2')
    ws['A2'].value = "执行前逐项确认，异常记录在备注栏"
    ws['A2'].font = Font(italic=True, color=GRAY)
    ws['A2'].alignment = Alignment(horizontal="center")

    row = 4
    headers = ["序号", "检查项目", "状态(✓/✗)", "异常记录", "确认签名"]
    for col, h in enumerate(headers, 1):
        style_subheader(ws.cell(row=row, column=col), h)

    check_items = [
        ("1", "操作人员已接受培训"),
        ("2", "设备点检完成"),
        ("3", "来料检验合格"),
        ("4", "工艺参数符合要求"),
        ("5", "测量设备校准有效"),
        ("6", "作业指导书已更新"),
        ("7", "6S现场符合要求"),
        ("8", "安全防护措施到位"),
        ("9", "首件检验已通过"),
        ("10", "异常处理流程已明确"),
    ]

    for i, (num, item) in enumerate(check_items):
        r = row + i + 1
        ws.cell(row=r, column=1, value=num).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=item)
        ws.cell(row=r, column=3, value="").alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4).value = ""
        ws.cell(row=r, column=5).value = ""
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = thin_border()

    # Summary
    r = row + len(check_items) + 1
    ws.cell(row=r, column=1, value="执行状态汇总：")
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=3).value = f'=COUNTIF(C{row+1}:C{r-1},"✓")&"/"&COUNTA(C{row+1}:C{r-1})'

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15

    set_print_settings(ws)


# ============================================================================
# Sheet 8: 质量问题验证追踪表
# ============================================================================
def create_sheet8_tracking(wb):
    ws = wb.create_sheet("质量问题验证追踪表")

    ws.merge_cells('A1:H1')
    style_header(ws['A1'], "质量问题验证追踪表")

    ws.merge_cells('A2:H2')
    ws['A2'].value = "记录每个质量问题的根因、措施和验证结果"
    ws['A2'].font = Font(italic=True, color=GRAY)
    ws['A2'].alignment = Alignment(horizontal="center")

    row = 4
    headers = ["问题编号", "问题描述", "根本原因", "改善措施", "责任人", "计划完成", "实际完成", "验证结果"]
    for col, h in enumerate(headers, 1):
        style_subheader(ws.cell(row=row, column=col), h)

    for i in range(1, 16):
        r = row + i
        ws.cell(row=r, column=1, value=f"Q{i:03d}").alignment = Alignment(horizontal="center")
        for col in range(2, 9):
            ws.cell(row=r, column=col).border = thin_border()
            if col in [6, 7]:
                ws.cell(row=r, column=col).number_format = 'YYYY-MM-DD'

    # Summary section
    r = row + 17
    ws.merge_cells(f'A{r}:H{r}')
    style_subheader(ws.cell(row=r, column=1), "状态统计")

    r += 1
    ws.cell(row=r, column=1, value="待处理：")
    ws.cell(row=r, column=2).value = f'=COUNTIF(H{row+1}:H{row+15},"待验证")'
    r += 1
    ws.cell(row=r, column=1, value="已完成：")
    ws.cell(row=r, column=2).value = f'=COUNTIF(H{row+1}:H{row+15},"通过")'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    set_print_settings(ws)


# ============================================================================
# Sheet 9: 团队质量数据台账
# ============================================================================
def create_sheet9_kpi(wb):
    ws = wb.create_sheet("团队质量数据台账")

    ws.merge_cells('A1:M1')
    style_header(ws['A1'], "团队质量数据台账 - 月度KPI追踪")

    row = 3
    headers = ["月份", "检验批数", "合格批数", "合格率", "不良件数", "不良率", "客诉件数", "改进项目", "培训人次", "人均产值", "备注"]
    for col, h in enumerate(headers, 1):
        style_subheader(ws.cell(row=row, column=col), h)

    months = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"]
    for i, month in enumerate(months):
        r = row + i + 1
        ws.cell(row=r, column=1, value=month).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).value = ""  # User input
        ws.cell(row=r, column=3).value = ""  # User input
        # Pass rate formula
        ws.cell(row=r, column=4).value = f"=IF(B{r}>0,C{r}/B{r},\"\")"
        ws.cell(row=r, column=4).number_format = '0.0%'
        ws.cell(row=r, column=5).value = ""  # User input
        # Defect rate formula
        ws.cell(row=r, column=6).value = f"=IF(B{r}>0,E{r}/B{r},\"\")"
        ws.cell(row=r, column=6).number_format = '0.0%'
        ws.cell(row=r, column=7).value = ""  # User input
        ws.cell(row=r, column=8).value = ""  # User input
        ws.cell(row=r, column=9).value = ""  # User input
        ws.cell(row=r, column=10).value = ""  # User input
        ws.cell(row=r, column=11).value = ""  # User input
        for col in range(1, 12):
            ws.cell(row=r, column=col).border = thin_border()

    # Annual summary
    r = row + 13
    ws.cell(row=r, column=1, value="年度汇总").font = Font(bold=True)
    ws.cell(row=r, column=2).value = f"=SUM(B{row+1}:B{r-1})"
    ws.cell(row=r, column=3).value = f"=SUM(C{row+1}:C{r-1})"
    ws.cell(row=r, column=4).value = f"=IF(B{r}>0,C{r}/B{r},\"\")"
    ws.cell(row=r, column=4).number_format = '0.0%'
    ws.cell(row=r, column=5).value = f"=SUM(E{row+1}:E{r-1})"
    ws.cell(row=r, column=6).value = f"=IF(B{r}>0,E{r}/B{r},\"\")"
    ws.cell(row=r, column=6).number_format = '0.0%'
    ws.cell(row=r, column=7).value = f"=SUM(G{row+1}:G{r-1})"
    ws.cell(row=r, column=8).value = f"=SUM(H{row+1}:H{r-1})"
    ws.cell(row=r, column=9).value = f"=SUM(I{row+1}:I{r-1})"

    # Column widths
    ws.column_dimensions['A'].width = 8
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 12

    set_print_settings(ws)


# ============================================================================
# Sheet 10: 课程练习成绩汇总
# ============================================================================
def create_sheet10_scores(wb):
    ws = wb.create_sheet("课程练习成绩汇总")

    ws.merge_cells('A1:H1')
    style_header(ws['A1'], "课程练习成绩汇总")

    ws.merge_cells('A2:H2')
    ws['A2'].value = "记录学员各模块练习得分，计算综合评分和排名"
    ws['A2'].font = Font(italic=True, color=GRAY)
    ws['A2'].alignment = Alignment(horizontal="center")

    row = 4
    headers = ["学员姓名", "模块1\n质量判断", "模块2\nWSDF分析", "模块3\n5问法", "模块4\n鱼骨图", "模块5\nFAR评估", "综合评分", "排名"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        style_subheader(cell)

    # 20 student rows
    for i in range(1, 21):
        r = row + i
        ws.cell(row=r, column=1).value = ""  # User input name
        for col in range(2, 7):
            ws.cell(row=r, column=col).value = ""  # User input scores
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
        # Average formula
        ws.cell(row=r, column=7).value = f"=IF(COUNTA(B{r}:F{r})>0,AVERAGE(B{r}:F{r}),\"\")"
        ws.cell(row=r, column=7).number_format = '0.0'
        ws.cell(row=r, column=7).font = Font(bold=True)
        # Rank formula
        ws.cell(row=r, column=8).value = f"=IF(G{r}<>\"\",RANK(G{r},$G$5:$G$24,0),\"\")"
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = thin_border()

    # Class statistics
    r = row + 22
    ws.cell(row=r, column=1, value="班级统计").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="平均分：")
    ws.cell(row=r, column=2).value = f"=IF(COUNT(G5:G24)>0,AVERAGE(G5:G24),\"\")"
    ws.cell(row=r, column=2).number_format = '0.0'
    r += 1
    ws.cell(row=r, column=1, value="最高分：")
    ws.cell(row=r, column=2).value = f"=IF(COUNT(G5:G24)>0,MAX(G5:G24),\"\")"
    ws.cell(row=r, column=2).number_format = '0.0'
    r += 1
    ws.cell(row=r, column=1, value="最低分：")
    ws.cell(row=r, column=2).value = f"=IF(COUNT(G5:G24)>0,MIN(G5:G24),\"\")"
    ws.cell(row=r, column=2).number_format = '0.0'
    r += 1
    ws.cell(row=r, column=1, value="及格人数(>=60)：")
    ws.cell(row=r, column=2).value = f"=COUNTIF(G5:G24,\">=60\")"

    # Column widths
    ws.column_dimensions['A'].width = 15
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12

    set_print_settings(ws)


# ============================================================================
# Main build functions
# ============================================================================
def build_blank_template():
    """Build the blank template file"""
    wb = Workbook()

    create_sheet1_quality_type(wb)
    create_sheet2_wsdf(wb)
    create_sheet3_5why(wb)
    create_sheet4_fishbone(wb)
    create_sheet5_far(wb)
    create_sheet6_decision(wb)
    create_sheet7_checklist(wb)
    create_sheet8_tracking(wb)
    create_sheet9_kpi(wb)
    create_sheet10_scores(wb)

    return wb


def build_filled_example():
    """Build the filled example file with complete case study"""
    wb = Workbook()

    create_sheet1_quality_type(wb)
    create_sheet2_wsdf(wb)
    create_sheet3_5why(wb)
    create_sheet4_fishbone(wb)
    create_sheet5_far(wb)
    create_sheet6_decision(wb)
    create_sheet7_checklist(wb)
    create_sheet8_tracking(wb)
    create_sheet9_kpi(wb)
    create_sheet10_scores(wb)

    # Fill with example data
    ws2 = wb["WSDF质量定义工作表"]
    ws2.cell(row=5, column=2, value="焊点不良，导致产品通电后无反应")
    ws2.cell(row=5, column=3, value="D")
    ws2.cell(row=5, column=4, value=8)
    ws2.cell(row=5, column=5, value=4)

    ws3 = wb["五问法分析单"]
    ws3.cell(row=5, column=3, value="因为烙铁温度设置过低")
    ws3.cell(row=5, column=4, value="温度枪实测温度245°C，标准要求280°C")

    ws6 = wb["质量放行决策卡"]
    ws6.cell(row=12, column=3, value="C. 反复出现（重复问题）")

    return wb


if __name__ == "__main__":
    from pathlib import Path

    OUTPUT_DIR = Path("D:/新课开发/制造/6-质量管理与质量文化重塑/配套表单Excel")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Building blank template...")
    wb1 = build_blank_template()
    wb1.save(OUTPUT_DIR / "质量工具表单_空表.xlsx")
    print(f"  Saved: {OUTPUT_DIR / '质量工具表单_空表.xlsx'}")

    print("Building filled example...")
    wb2 = build_filled_example()
    wb2.save(OUTPUT_DIR / "质量工具表单_填好版.xlsx")
    print(f"  Saved: {OUTPUT_DIR / '质量工具表单_填好版.xlsx'}")

    print("Done!")

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建《时间与精力管理》课程的全流程工具表单Excel文件
使用 openpyxl 生成，包含多个sheet、公式和格式
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import time

# 样式定义
DARK_HEADER_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
LIGHT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
DARK_RED_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")

HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name='Microsoft YaHei', bold=True, color="404040", size=14)
LABEL_FONT = Font(name='Microsoft YaHei', bold=True, color="404040", size=10)
DATA_FONT = Font(name='Microsoft YaHei', size=10)
INPUT_FONT = Font(name='Microsoft YaHei', size=10, color="0000FF")  # 蓝色输入字体
FORMULA_FONT = Font(name='Microsoft YaHei', size=10, color="000000")  # 黑色公式字体
INSTRUCTION_FONT = Font(name='Microsoft YaHei', size=9, color="666666", italic=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

MEDIUM_BORDER = Border(
    left=Side(style='medium', color='404040'),
    right=Side(style='medium', color='404040'),
    top=Side(style='medium', color='404040'),
    bottom=Side(style='medium', color='404040')
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

def set_column_widths(ws, widths):
    """设置列宽"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_header_row(ws, row_num, headers, start_col=1):
    """创建表头行"""
    for i, header in enumerate(headers):
        cell = ws.cell(row=row_num, column=start_col + i, value=header)
        cell.font = HEADER_FONT
        cell.fill = DARK_HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

def create_title_row(ws, row_num, title, num_cols):
    """创建标题行"""
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
    cell = ws.cell(row=row_num, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN

def create_instruction_row(ws, row_num, instruction, num_cols):
    """创建使用说明行"""
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
    cell = ws.cell(row=row_num, column=1, value=instruction)
    cell.font = INSTRUCTION_FONT
    cell.alignment = LEFT_ALIGN

def style_data_row(ws, row_num, num_cols, is_light=True):
    """设置数据行样式"""
    fill = LIGHT_ROW_FILL if is_light else WHITE_FILL
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN if col > 1 else LEFT_ALIGN

def create_total_row(ws, row_num, label, formula_cols, num_cols, start_col=1):
    """创建合计行"""
    cell = ws.cell(row=row_num, column=start_col, value=label)
    cell.font = LABEL_FONT
    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    cell.border = MEDIUM_BORDER
    cell.alignment = LEFT_ALIGN

    for i, col in enumerate(formula_cols):
        if col:
            formula_cell = ws.cell(row=row_num, column=start_col + i + 1)
            formula_cell.value = col
            formula_cell.font = FORMULA_FONT
            formula_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            formula_cell.border = MEDIUM_BORDER
            formula_cell.alignment = RIGHT_ALIGN

# ============================================================
# F1_每日时间追踪表.xlsx
# ============================================================
def create_f1_daily_time_tracking():
    wb = Workbook()

    # ===== Sheet1: 时间日志 =====
    ws1 = wb.active
    ws1.title = "时间日志"

    # 标题
    create_title_row(ws1, 1, "每日时间追踪表 - 时间日志", 5)
    create_instruction_row(ws1, 2, "使用说明：每小时记录一次活动。精力状态：1=极低，2=低，3=中等，4=高，5=极高", 5)

    # 表头
    headers = ["时间段", "活动类型", "活动描述", "精力状态(1-5)", "备注"]
    create_header_row(ws1, 3, headers)

    # 时间段数据
    time_slots = [
        "06:00-07:00", "07:00-08:00", "08:00-09:00", "09:00-10:00",
        "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00",
        "14:00-15:00", "15:00-16:00", "16:00-17:00", "17:00-18:00",
        "18:00-19:00", "19:00-20:00", "20:00-21:00", "21:00-22:00",
        "22:00-23:00", "23:00-24:00"
    ]

    activity_types = ["工作", "学习", "休息", "社交", "运动", "其他"]

    for i, slot in enumerate(time_slots):
        row = i + 4
        ws1.cell(row=row, column=1, value=slot).font = DATA_FONT
        ws1.cell(row=row, column=2).font = DATA_FONT  # 活动类型待填写
        ws1.cell(row=row, column=3).font = DATA_FONT  # 活动描述待填写
        ws1.cell(row=row, column=4, value=3).font = INPUT_FONT  # 默认中等精力
        ws1.cell(row=row, column=5).font = DATA_FONT  # 备注待填写
        style_data_row(ws1, row, 5, i % 2 == 0)

    # 合计行
    row_total = len(time_slots) + 4
    ws1.cell(row=row_total, column=1, value="合计").font = LABEL_FONT
    ws1.cell(row=row_total, column=2, value=f"=COUNTA(B4:B{row_total-1})").font = FORMULA_FONT
    ws1.cell(row=row_total, column=4, value=f"=AVERAGE(D4:D{row_total-1})").font = FORMULA_FONT

    set_column_widths(ws1, {'A': 14, 'B': 10, 'C': 25, 'D': 12, 'E': 15})

    # ===== Sheet2: 汇总分析 =====
    ws2 = wb.create_sheet("汇总分析")

    create_title_row(ws2, 1, "每日时间追踪表 - 汇总分析", 4)
    create_instruction_row(ws2, 2, "使用说明：自动汇总时间日志中的数据，分析各类型活动耗时和精力状态分布", 4)

    # 活动类型耗时汇总
    headers2 = ["活动类型", "记录次数", "总时长(小时)", "占总时间比例", "平均精力状态"]
    create_header_row(ws2, 3, headers2)

    activity_summary = ["工作", "学习", "休息", "社交", "运动", "其他"]
    for i, act in enumerate(activity_summary):
        row = i + 4
        ws2.cell(row=row, column=1, value=act).font = DATA_FONT
        ws2.cell(row=row, column=2, value=f'=COUNTIF(时间日志!B:B,"{act}")').font = FORMULA_FONT
        ws2.cell(row=row, column=3, value=f'=B{row}').font = FORMULA_FONT  # 每小时一条
        ws2.cell(row=row, column=4, value=f'=IF(B{row}=0,0,C{row}/SUM($C$4:$C$9))').font = FORMULA_FONT
        ws2.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIF(时间日志!B:B,"{act}",时间日志!D:D),0)').font = FORMULA_FONT
        style_data_row(ws2, row, 5, i % 2 == 0)

    # 精力状态分布
    row_energy = 12
    ws2.cell(row=row_energy, column=1, value="精力状态分布").font = TITLE_FONT
    headers3 = ["精力等级", "出现次数", "占比", "建议"]
    create_header_row(ws2, row_energy + 1, headers3)

    energy_levels = [
        ("1-极低", 1, "需要立即休息"),
        ("2-低", 2, "考虑轻松任务"),
        ("3-中等", 3, "可做常规任务"),
        ("4-高", 4, "适合高强度任务"),
        ("5-极高", 5, "最佳状态，充分利用")
    ]

    for i, (label, val, suggestion) in enumerate(energy_levels):
        row = row_energy + 2 + i
        ws2.cell(row=row, column=1, value=label).font = DATA_FONT
        ws2.cell(row=row, column=2, value=f'=COUNTIF(时间日志!D:D,{val})').font = FORMULA_FONT
        ws2.cell(row=row, column=3, value=f'=IF(B{row}=0,0,B{row}/COUNTA(时间日志!D:D))').font = FORMULA_FONT
        ws2.cell(row=row, column=4, value=suggestion).font = DATA_FONT
        style_data_row(ws2, row, 4, i % 2 == 0)

    set_column_widths(ws2, {'A': 14, 'B': 12, 'C': 14, 'D': 18, 'E': 20})

    # ===== Sheet3: 周报汇总 =====
    ws3 = wb.create_sheet("周报汇总")

    create_title_row(ws3, 1, "每周时间追踪汇总", 6)
    create_instruction_row(ws3, 2, "使用说明：记录一周七天的数据，自动计算周平均值", 6)

    headers4 = ["日期", "工作(小时)", "学习(小时)", "休息(小时)", "社交(小时)", "运动(小时)"]
    create_header_row(ws3, 3, headers4)

    days = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    for i, day in enumerate(days):
        row = i + 4
        ws3.cell(row=row, column=1, value=day).font = DATA_FONT
        for col in range(2, 7):
            ws3.cell(row=row, column=col, value=0).font = INPUT_FONT
        style_data_row(ws3, row, 6, i % 2 == 0)

    # 合计行
    row_total = len(days) + 4
    ws3.cell(row=row_total, column=1, value="周合计").font = LABEL_FONT
    for col in range(2, 7):
        ws3.cell(row=row_total, column=col, value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{row_total-1})').font = FORMULA_FONT
        style_data_row(ws3, row_total, 6, True)

    # 平均行
    row_avg = row_total + 1
    ws3.cell(row=row_avg, column=1, value="日平均").font = LABEL_FONT
    for col in range(2, 7):
        ws3.cell(row=row_avg, column=col, value=f'=AVERAGE({get_column_letter(col)}4:{get_column_letter(col)}{row_total-1})').font = FORMULA_FONT
        style_data_row(ws3, row_avg, 6, False)

    set_column_widths(ws3, {'A': 10, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12})

    return wb


# ============================================================
# F2_优先级决策工作表.xlsx
# ============================================================
def create_f2_priority_decision():
    wb = Workbook()

    # ===== Sheet1: 任务清单 =====
    ws1 = wb.active
    ws1.title = "任务清单"

    create_title_row(ws1, 1, "优先级决策工作表 - 任务清单", 7)
    create_instruction_row(ws1, 2, "使用说明：填写任务信息，系统自动判断优先级。紧急程度和重要程度：1-5分", 7)

    headers = ["任务描述", "紧急程度(1-5)", "重要程度(1-5)", "能量需求(1-5)", "截止时间", "决策结果", "行动方案"]
    create_header_row(ws1, 3, headers)

    for i in range(15):
        row = i + 4
        ws1.cell(row=row, column=1).font = DATA_FONT
        ws1.cell(row=row, column=2, value=3).font = INPUT_FONT
        ws1.cell(row=row, column=3, value=3).font = INPUT_FONT
        ws1.cell(row=row, column=4, value=3).font = INPUT_FONT
        ws1.cell(row=row, column=5).font = INPUT_FONT
        # 决策结果公式：紧急x重要
        ws1.cell(row=row, column=6, value=f'=B{row}*C{row}').font = FORMULA_FONT
        ws1.cell(row=row, column=7).font = DATA_FONT
        style_data_row(ws1, row, 7, i % 2 == 0)

    # 排序说明
    row_note = 20
    ws3 = None
    for sheet in wb.worksheets:
        if sheet.title == "任务清单":
            ws3 = sheet
            break
    if ws3:
        ws3.cell(row=row_note, column=1, value="优先级说明：决策结果=紧急程度×重要程度").font = INSTRUCTION_FONT
        ws3.cell(row=row_note+1, column=1, value="16-25分：立即执行(第一优先) | 9-15分：尽快处理(第二优先) | 1-8分：可暂缓(第三优先)").font = INSTRUCTION_FONT

    set_column_widths(ws1, {'A': 25, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 10, 'G': 20})

    # ===== Sheet2: 决策矩阵 =====
    ws2 = wb.create_sheet("决策矩阵")

    create_title_row(ws2, 1, "优先级决策工作表 - 四象限矩阵", 5)
    create_instruction_row(ws2, 2, "使用说明：根据紧急程度和重要程度将任务分为四个象限", 5)

    # 四象限表格
    headers2 = ["", "紧急", "不紧急"]
    create_header_row(ws2, 3, headers2)

    ws2.cell(row=4, column=1, value="重要").font = LABEL_FONT
    ws2.cell(row=4, column=2, value="第一优先").fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    ws2.cell(row=4, column=3, value="第二优先").fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    ws2.cell(row=5, column=1, value="不重要").font = LABEL_FONT
    ws2.cell(row=5, column=2, value="第三优先").fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws2.cell(row=5, column=3, value="第四优先").fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for row in [4, 5]:
        for col in [1, 2, 3]:
            ws2.cell(row=row, column=col).border = THIN_BORDER
            ws2.cell(row=row, column=col).alignment = CENTER_ALIGN

    # 矩阵说明
    row_desc = 8
    descriptions = [
        ("第一优先(紧急+重要)", "立即执行，不能拖延", "FFCCCC"),
        ("第二优先(不紧急+重要)", "规划时间，按计划执行", "CCFFCC"),
        ("第三优先(紧急+不重要)", "尽量委托他人处理", "FFFFCC"),
        ("第四优先(不紧急+不重要)", "可以考虑删除", "CCCCCC")
    ]

    for i, (quadrant, desc, color) in enumerate(descriptions):
        row = row_desc + i
        ws2.cell(row=row, column=1, value=quadrant).font = LABEL_FONT
        ws2.cell(row=row, column=2, value=desc).font = DATA_FONT
        ws2.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws2.cell(row=row, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    set_column_widths(ws2, {'A': 25, 'B': 20, 'C': 15})

    # ===== Sheet3: 承诺边界清单 =====
    ws3 = wb.create_sheet("承诺边界清单")

    create_title_row(ws3, 1, "优先级决策工作表 - 承诺边界清单", 5)
    create_instruction_row(ws3, 2, "使用说明：记录你对他人的承诺，了解自己的承担边界", 5)

    headers3 = ["承诺内容", "承诺对象", "承诺时间", "能量消耗(1-5)", "可完成度(1-5)"]
    create_header_row(ws3, 3, headers3)

    for i in range(10):
        row = i + 4
        for col in range(1, 6):
            ws3.cell(row=row, column=col).font = DATA_FONT if col != 4 else INPUT_FONT
        style_data_row(ws3, row, 5, i % 2 == 0)

    # 边界分析
    row_analysis = 16
    ws3.cell(row=row_analysis, column=1, value="承诺分析").font = TITLE_FONT
    ws3.cell(row=row_analysis+1, column=1, value="总承诺数").font = LABEL_FONT
    ws3.cell(row=row_analysis+1, column=2, value='=COUNTA(A4:A13)').font = FORMULA_FONT
    ws3.cell(row=row_analysis+2, column=1, value="高能量消耗承诺数").font = LABEL_FONT
    ws3.cell(row=row_analysis+2, column=2, value='=COUNTIF(D4:D13,">3")').font = FORMULA_FONT
    ws3.cell(row=row_analysis+3, column=1, value="平均可完成度").font = LABEL_FONT
    ws3.cell(row=row_analysis+3, column=2, value='=IFERROR(AVERAGE(E4:E13),0)').font = FORMULA_FONT

    set_column_widths(ws3, {'A': 25, 'B': 12, 'C': 12, 'D': 12, 'E': 12})

    return wb


# ============================================================
# F3_任务切换成本记录表.xlsx
# ============================================================
def create_f3_task_switching():
    wb = Workbook()

    # ===== Sheet1: 切换日志 =====
    ws1 = wb.active
    ws1.title = "切换日志"

    create_title_row(ws1, 1, "任务切换成本记录表 - 切换日志", 7)
    create_instruction_row(ws1, 2, "使用说明：记录每次任务切换的时间、原因和恢复时间，分析切换成本", 7)

    headers = ["切换时间", "切换前任务", "切换后任务", "切换原因", "恢复时间(分钟)", "效率损失(%)", "备注"]
    create_header_row(ws1, 3, headers)

    switch_reasons = ["紧急任务插入", "外部干扰", "内部中断", "自然完成", "其他"]

    for i in range(20):
        row = i + 4
        ws1.cell(row=row, column=1).font = DATA_FONT
        ws1.cell(row=row, column=2).font = DATA_FONT
        ws1.cell(row=row, column=3).font = DATA_FONT
        ws1.cell(row=row, column=4).font = DATA_FONT
        ws1.cell(row=row, column=5, value=0).font = INPUT_FONT
        ws1.cell(row=row, column=6, value=f'=E{row}*5').font = FORMULA_FONT  # 假设每分钟损失5%效率
        ws1.cell(row=row, column=7).font = DATA_FONT
        style_data_row(ws1, row, 7, i % 2 == 0)

    # 合计
    row_total = 24
    ws1.cell(row=row_total, column=1, value="合计").font = LABEL_FONT
    ws1.cell(row=row_total, column=5, value=f'=SUM(E4:E{row_total-1})').font = FORMULA_FONT
    ws1.cell(row=row_total, column=6, value=f'=SUM(F4:F{row_total-1})').font = FORMULA_FONT

    set_column_widths(ws1, {'A': 12, 'B': 15, 'C': 15, 'D': 12, 'E': 14, 'F': 12, 'G': 12})

    # ===== Sheet2: 成本分析 =====
    ws2 = wb.create_sheet("成本分析")

    create_title_row(ws2, 1, "任务切换成本记录表 - 成本分析", 4)
    create_instruction_row(ws2, 2, "使用说明：分析切换原因分布和恢复时间，找出最大成本来源", 4)

    headers2 = ["切换原因", "出现次数", "总恢复时间(分钟)", "平均恢复时间(分钟)", "占总切换比例"]
    create_header_row(ws2, 3, headers2)

    reasons = ["紧急任务插入", "外部干扰", "内部中断", "自然完成", "其他"]
    for i, reason in enumerate(reasons):
        row = i + 4
        ws2.cell(row=row, column=1, value=reason).font = DATA_FONT
        ws2.cell(row=row, column=2, value=f'=COUNTIF(切换日志!D:D,"{reason}")').font = FORMULA_FONT
        ws2.cell(row=row, column=3, value=f'=SUMIF(切换日志!D:D,"{reason}",切换日志!E:E)').font = FORMULA_FONT
        ws2.cell(row=row, column=4, value=f'=IFERROR(C{row}/B{row},0)').font = FORMULA_FONT
        ws2.cell(row=row, column=5, value=f'=IF(B{row}=0,0,B{row}/COUNTA(切换日志!D:D))').font = FORMULA_FONT
        style_data_row(ws2, row, 5, i % 2 == 0)

    # 总体统计
    row_stats = 11
    ws2.cell(row=row_stats, column=1, value="总体统计").font = TITLE_FONT

    stats = [
        ("总切换次数", '=COUNTA(切换日志!D:D)'),
        ("总恢复时间", '=SUM(切换日志!E:E)'),
        ("平均恢复时间", '=IFERROR(SUM(切换日志!E:E)/COUNTA(切换日志!D:D),0)'),
        ("最高单次损失", '=MAX(切换日志!F:F)')
    ]

    for i, (label, formula) in enumerate(stats):
        row = row_stats + 1 + i
        ws2.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws2.cell(row=row, column=2, value=formula).font = FORMULA_FONT

    set_column_widths(ws2, {'A': 16, 'B': 12, 'C': 16, 'D': 18, 'E': 16})

    # ===== Sheet3: 改善建议 =====
    ws3 = wb.create_sheet("改善建议")

    create_title_row(ws3, 1, "任务切换成本记录表 - 改善建议", 3)
    create_instruction_row(ws3, 2, "使用说明：基于成本分析，制定减少任务切换的策略", 3)

    suggestions = [
        ("问题识别", ""),
        ("最高成本切换原因", "=INDEX(成本分析!A:A,MATCH(MAX(成本分析!C:C),成本分析!C:C,0))"),
        ("需要改善的切换类型", "=IF(COUNTIF(成本分析!B:B,MAX(成本分析!B:B))>0,INDEX(成本分析!A:A,MATCH(MAX(成本分析!B:B),成本分析!B:B,0)),\"无\")"),
        ("", ""),
        ("改善策略", ""),
        ("策略1：时间分块", "将相似任务集中处理，减少切换"),
        ("策略2：干扰管理", "设定无干扰时段，关闭通知"),
        ("策略3：任务批处理", "将同类小任务合并一次处理"),
        ("策略4：恢复时间缩短", "通过练习提高任务切换后的恢复速度"),
        ("", ""),
        ("目标设定", ""),
        ("周切换次数目标", ""),
        ("单次恢复时间目标(分钟)", "")
    ]

    for i, (label, value) in enumerate(suggestions):
        row = i + 3
        ws3.cell(row=row, column=1, value=label).font = LABEL_FONT if label else DATA_FONT
        ws3.cell(row=row, column=2, value=value).font = FORMULA_FONT if value.startswith('=') else DATA_FONT
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    set_column_widths(ws3, {'A': 25, 'B': 40})

    return wb


# ============================================================
# F4_碎片时间清单.xlsx
# ============================================================
def create_f4_fragmented_time():
    wb = Workbook()

    # ===== Sheet1: 碎片时间盘点 =====
    ws1 = wb.active
    ws1.title = "碎片时间盘点"

    create_title_row(ws1, 1, "碎片时间清单 - 碎片时间盘点", 5)
    create_instruction_row(ws1, 2, "使用说明：盘点每天的碎片时间，评估可利用性", 5)

    headers = ["时间段", "持续时间(分钟)", "可利用性(1-5)", "适合任务类型", "备注"]
    create_header_row(ws1, 3, headers)

    time_slots = [
        "等电梯", "排队等待", "通勤路上", "会议间隙", "午休前",
        "下班前", "看电视时", "起床后", "吃饭时", "睡前"
    ]

    for i, slot in enumerate(time_slots):
        row = i + 4
        ws1.cell(row=row, column=1, value=slot).font = DATA_FONT
        ws1.cell(row=row, column=2, value=10).font = INPUT_FONT  # 默认10分钟
        ws1.cell(row=row, column=3, value=3).font = INPUT_FONT
        ws1.cell(row=row, column=4).font = DATA_FONT
        ws1.cell(row=row, column=5).font = DATA_FONT
        style_data_row(ws1, row, 5, i % 2 == 0)

    # 合计
    row_total = len(time_slots) + 4
    ws1.cell(row=row_total, column=1, value="日均碎片时间合计").font = LABEL_FONT
    ws1.cell(row=row_total, column=2, value=f'=SUM(B4:B{row_total-1})').font = FORMULA_FONT
    ws1.cell(row=row_total, column=3, value=f'=AVERAGE(C4:C{row_total-1})').font = FORMULA_FONT

    set_column_widths(ws1, {'A': 14, 'B': 14, 'C': 12, 'D': 18, 'E': 15})

    # ===== Sheet2: 碎片任务清单 =====
    ws2 = wb.create_sheet("碎片任务清单")

    create_title_row(ws2, 1, "碎片时间清单 - 碎片任务清单", 5)
    create_instruction_row(ws2, 2, "使用说明：列出适合在碎片时间完成的任务，匹配所需时间和精力", 5)

    headers2 = ["任务名称", "所需时间(分钟)", "所需精力(1-5)", "适合场景", "优先级(1-5)"]
    create_header_row(ws2, 3, headers2)

    sample_tasks = [
        ("回复简单邮件", 5, 2, "等电梯/排队", 4),
        ("阅读资讯", 10, 2, "通勤/午休", 3),
        ("整理笔记", 15, 3, "会议间隙", 3),
        ("快速决策", 5, 4, "碎片时间", 5),
        ("学习单词", 10, 2, "通勤/排队", 4),
        ("思考规划", 15, 3, "等待时", 4),
        ("社交回复", 5, 2, "随时", 3),
        ("听有声书", 30, 1, "通勤/运动", 5)
    ]

    for i, (name, time, energy, scene, priority) in enumerate(sample_tasks):
        row = i + 4
        ws2.cell(row=row, column=1, value=name).font = DATA_FONT
        ws2.cell(row=row, column=2, value=time).font = INPUT_FONT
        ws2.cell(row=row, column=3, value=energy).font = INPUT_FONT
        ws2.cell(row=row, column=4, value=scene).font = DATA_FONT
        ws2.cell(row=row, column=5, value=priority).font = INPUT_FONT
        style_data_row(ws2, row, 5, i % 2 == 0)

    for i in range(5):
        row = len(sample_tasks) + 4 + i
        for col in range(1, 6):
            ws2.cell(row=row, column=col).font = DATA_FONT
        style_data_row(ws2, row, 5, (len(sample_tasks) + i) % 2 == 0)

    # 任务分析
    row_analysis = 20
    ws2.cell(row=row_analysis, column=1, value="任务分析").font = TITLE_FONT
    ws2.cell(row=row_analysis+1, column=1, value="总任务数").font = LABEL_FONT
    ws2.cell(row=row_analysis+1, column=2, value='=COUNTA(A4:A17)').font = FORMULA_FONT
    ws2.cell(row=row_analysis+2, column=1, value="短时任务(<=10分钟)").font = LABEL_FONT
    ws2.cell(row=row_analysis+2, column=2, value='=COUNTIF(B4:B17,"<=10")').font = FORMULA_FONT
    ws2.cell(row=row_analysis+3, column=1, value="低精力任务(<=2)").font = LABEL_FONT
    ws2.cell(row=row_analysis+3, column=2, value='=COUNTIF(C4:C17,"<=2")').font = FORMULA_FONT

    set_column_widths(ws2, {'A': 18, 'B': 14, 'C': 12, 'D': 15, 'E': 12})

    # ===== Sheet3: 每日碎片时间利用计划 =====
    ws3 = wb.create_sheet("每日利用计划")

    create_title_row(ws3, 1, "碎片时间清单 - 每日碎片时间利用计划", 5)
    create_instruction_row(ws3, 2, "使用说明：规划每天如何使用碎片时间，设定具体任务", 5)

    headers3 = ["时间段", "可用时长(分钟)", "计划任务", "实际完成", "完成度(%)"]
    create_header_row(ws3, 3, headers3)

    slots = ["早上通勤", "上午工作间隙", "午休时间", "下午工作间隙", "傍晚通勤", "晚间休闲"]
    for i, slot in enumerate(slots):
        row = i + 4
        ws3.cell(row=row, column=1, value=slot).font = DATA_FONT
        ws3.cell(row=row, column=2, value=15).font = INPUT_FONT
        ws3.cell(row=row, column=3).font = DATA_FONT
        ws3.cell(row=row, column=4).font = DATA_FONT
        ws3.cell(row=row, column=5, value=f'=IF(D{row}="","",IF(C{row}="",0,100))').font = FORMULA_FONT
        style_data_row(ws3, row, 5, i % 2 == 0)

    # 统计
    row_stats = len(slots) + 4
    ws3.cell(row=row_stats, column=1, value="日合计").font = LABEL_FONT
    ws3.cell(row=row_stats, column=2, value=f'=SUM(B4:B{row_stats-1})').font = FORMULA_FONT
    ws3.cell(row=row_stats, column=5, value=f'=IFERROR(AVERAGEIF(E4:E{row_stats-1},"<>",E4:E{row_stats-1}),0)').font = FORMULA_FONT

    set_column_widths(ws3, {'A': 14, 'B': 14, 'C': 20, 'D': 12, 'E': 12})

    return wb


# ============================================================
# F5_精力管理四维度表.xlsx
# ============================================================
def create_f5_energy_management():
    wb = Workbook()

    # ===== Sheet1: 精力自评 =====
    ws1 = wb.active
    ws1.title = "精力自评"

    create_title_row(ws1, 1, "精力管理四维度表 - 精力自评", 6)
    create_instruction_row(ws1, 2, "使用说明：每天对四个维度进行自评(1-10分)，追踪精力变化", 6)

    headers = ["日期", "身体状态(1-10)", "情绪状态(1-10)", "注意力(1-10)", "意义感(1-10)", "综合得分"]
    create_header_row(ws1, 3, headers)

    for i in range(30):
        row = i + 4
        ws1.cell(row=row, column=1).font = DATA_FONT
        for col in [2, 3, 4, 5]:
            ws1.cell(row=row, column=col, value=5).font = INPUT_FONT
        # 综合得分公式
        ws1.cell(row=row, column=6, value=f'=AVERAGE(B{row}:E{row})').font = FORMULA_FONT
        style_data_row(ws1, row, 6, i % 2 == 0)

    # 周平均
    row_week = 35
    ws1.cell(row=row_week, column=1, value="周平均").font = TITLE_FONT
    for col in range(2, 7):
        ws1.cell(row=row_week, column=col, value=f'=AVERAGE({get_column_letter(col)}4:{get_column_letter(col)}33)').font = FORMULA_FONT
        style_data_row(ws1, row_week, 6, True)

    set_column_widths(ws1, {'A': 12, 'B': 14, 'C': 14, 'D': 12, 'E': 12, 'F': 12})

    # ===== Sheet2: 节律追踪 =====
    ws2 = wb.create_sheet("节律追踪")

    create_title_row(ws2, 1, "精力管理四维度表 - 节律追踪", 5)
    create_instruction_row(ws2, 2, "使用说明：记录不同时间段的精力状态，找出个人精力高峰和低谷", 5)

    headers2 = ["时间段", "身体状态", "情绪状态", "注意力", "意义感", "综合精力"]
    create_header_row(ws2, 3, headers2)

    time_periods = [
        "06:00-08:00", "08:00-10:00", "10:00-12:00", "12:00-14:00",
        "14:00-16:00", "16:00-18:00", "18:00-20:00", "20:00-22:00", "22:00-24:00"
    ]

    for i, period in enumerate(time_periods):
        row = i + 4
        ws2.cell(row=row, column=1, value=period).font = DATA_FONT
        for col in range(2, 6):
            ws2.cell(row=row, column=col, value=5).font = INPUT_FONT
        ws2.cell(row=row, column=6, value=f'=AVERAGE(B{row}:E{row})').font = FORMULA_FONT
        style_data_row(ws2, row, 6, i % 2 == 0)

    # 高峰和低谷分析
    row_analysis = len(time_periods) + 6
    ws2.cell(row=row_analysis, column=1, value="节律分析").font = TITLE_FONT
    ws2.cell(row=row_analysis+1, column=1, value="精力高峰时段").font = LABEL_FONT
    ws2.cell(row=row_analysis+1, column=2, value=f'=INDEX(A4:A12,MATCH(MAX(F4:F12),F4:F12,0))').font = FORMULA_FONT
    ws2.cell(row=row_analysis+2, column=1, value="精力低谷时段").font = LABEL_FONT
    ws2.cell(row=row_analysis+2, column=2, value=f'=INDEX(A4:A12,MATCH(MIN(F4:F12),F4:F12,0))').font = FORMULA_FONT

    set_column_widths(ws2, {'A': 14, 'B': 12, 'C': 12, 'D': 10, 'E': 10, 'F': 12})

    # ===== Sheet3: 恢复活动清单 =====
    ws3 = wb.create_sheet("恢复活动清单")

    create_title_row(ws3, 1, "精力管理四维度表 - 恢复活动清单", 5)
    create_instruction_row(ws3, 2, "使用说明：列出各维度的恢复活动，在精力低时选择执行", 5)

    # 四个维度的恢复活动
    dimensions = [
        ("身体状态", ["充足睡眠(7-8小时)", "适度运动(30分钟)", "健康饮食", "短暂午休(20分钟)", "拉伸放松"]),
        ("情绪状态", ["与朋友交流", "听音乐放松", "看喜剧视频", "写日记倾诉", "深呼吸冥想"]),
        ("注意力", ["番茄工作法", "任务分块", "减少干扰", "眼保健操", "远眺休息"]),
        ("意义感", ["回顾目标进展", "帮助他人", "学习新技能", "创意写作", "自然散步"])
    ]

    current_row = 4
    for dimension, activities in dimensions:
        ws3.cell(row=current_row, column=1, value=dimension).font = TITLE_FONT
        ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        current_row += 1

        headers3 = ["活动名称", "执行时间", "精力恢复效果(1-5)", "本周执行次数", "备注"]
        create_header_row(ws3, current_row, headers3)
        current_row += 1

        for i, activity in enumerate(activities):
            ws3.cell(row=current_row, column=1, value=activity).font = DATA_FONT
            ws3.cell(row=current_row, column=2).font = DATA_FONT
            ws3.cell(row=current_row, column=3, value=3).font = INPUT_FONT
            ws3.cell(row=current_row, column=4, value=0).font = INPUT_FONT
            ws3.cell(row=current_row, column=5).font = DATA_FONT
            style_data_row(ws3, current_row, 5, i % 2 == 0)
            current_row += 1

        current_row += 1  # 空行分隔

    set_column_widths(ws3, {'A': 22, 'B': 14, 'C': 16, 'D': 14, 'E': 15})

    return wb


# ============================================================
# F6_30天行动计划追踪表.xlsx
# ============================================================
def create_f6_30day_action_plan():
    wb = Workbook()

    # ===== Sheet1: 行动计划 =====
    ws1 = wb.active
    ws1.title = "行动计划"

    create_title_row(ws1, 1, "30天行动计划追踪表 - 行动计划", 7)
    create_instruction_row(ws1, 2, "使用说明：设定目标并拆分为四周的具体行动，每周追踪执行情况", 7)

    headers = ["目标", "具体行动", "第一周", "第二周", "第三周", "第四周", "成果验证"]
    create_header_row(ws1, 3, headers)

    for i in range(10):
        row = i + 4
        ws1.cell(row=row, column=1).font = DATA_FONT
        ws1.cell(row=row, column=2).font = DATA_FONT
        for col in range(3, 7):
            ws1.cell(row=row, column=col).font = DATA_FONT
        ws1.cell(row=row, column=7).font = DATA_FONT
        style_data_row(ws1, row, 7, i % 2 == 0)

    # 进度汇总
    row_progress = 15
    ws1.cell(row=row_progress, column=1, value="周完成率").font = TITLE_FONT
    weeks = ["第一周", "第二周", "第三周", "第四周"]
    for i, week in enumerate(weeks):
        row = row_progress + 1 + i
        ws1.cell(row=row, column=1, value=week).font = LABEL_FONT
        ws1.cell(row=row, column=2, value=f'=COUNTIF(C{row}:C{row},"完成")+COUNTIF(C{row}:C{row},"✓")').font = FORMULA_FONT
        ws1.cell(row=row, column=3, value=f'=IFERROR(B{row}/COUNTA(C4:C13),0)').font = FORMULA_FONT

    set_column_widths(ws1, {'A': 20, 'B': 25, 'C': 10, 'D': 10, 'E': 10, 'F': 10, 'G': 15})

    # ===== Sheet2: 周检视 =====
    ws2 = wb.create_sheet("周检视")

    create_title_row(ws2, 1, "30天行动计划追踪表 - 周检视", 5)
    create_instruction_row(ws2, 2, "使用说明：每周结束后进行检视，记录完成情况和调整方案", 5)

    headers2 = ["周次", "计划目标", "完成情况", "未完成原因", "调整措施"]
    create_header_row(ws2, 3, headers2)

    for i in range(4):
        row = i + 4
        ws2.cell(row=row, column=1, value=f"第{i+1}周").font = DATA_FONT
        ws2.cell(row=row, column=2).font = DATA_FONT
        ws2.cell(row=row, column=3).font = DATA_FONT
        ws2.cell(row=row, column=4).font = DATA_FONT
        ws2.cell(row=row, column=5).font = DATA_FONT
        style_data_row(ws2, row, 5, i % 2 == 0)

    # 检视统计
    row_stats = 10
    ws2.cell(row=row_stats, column=1, value="检视统计").font = TITLE_FONT
    ws2.cell(row=row_stats+1, column=1, value="已完成目标数").font = LABEL_FONT
    ws2.cell(row=row_stats+1, column=2, value='=COUNTIF(C4:C7,"完成")').font = FORMULA_FONT
    ws2.cell(row=row_stats+2, column=1, value="执行率").font = LABEL_FONT
    ws2.cell(row=row_stats+2, column=2, value='=IFERROR(B12/4,0)').font = FORMULA_FONT

    set_column_widths(ws2, {'A': 10, 'B': 20, 'C': 15, 'D': 20, 'E': 25})

    # ===== Sheet3: 成果汇总 =====
    ws3 = wb.create_sheet("成果汇总")

    create_title_row(ws3, 1, "30天行动计划追踪表 - 成果汇总", 4)
    create_instruction_row(ws3, 2, "使用说明：30天行动结束后，汇总整体成果和收获", 4)

    sections = [
        ("目标达成情况", [
            ("目标1", ""),
            ("目标2", ""),
            ("目标3", "")
        ]),
        ("关键成果", [
            ("最重要的收获", ""),
            ("最大的改变", ""),
            ("继续保持的习惯", "")
        ]),
        ("反思与改进", [
            ("遇到的困难", ""),
            ("解决方式", ""),
            ("下次改进", "")
        ]),
        ("30天总结", [
            ("总体评分(1-10)", ""),
            ("最有价值的行动", ""),
            ("下一步计划", "")
        ])
    ]

    current_row = 4
    for section_title, items in sections:
        ws3.cell(row=current_row, column=1, value=section_title).font = TITLE_FONT
        ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1

        for label, value in items:
            ws3.cell(row=current_row, column=1, value=label).font = LABEL_FONT
            ws3.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
            ws3.cell(row=current_row, column=2).font = DATA_FONT
            current_row += 1

        current_row += 1

    set_column_widths(ws3, {'A': 18, 'B': 20, 'C': 20, 'D': 20})

    return wb


# ============================================================
# 主程序：创建所有文件
# ============================================================
def main():
    base_dir = "D:/新课开发/职业素养/七、时间与精力管理：在高压环境中稳住效率与节奏/全流程工具表单/"

    files = [
        ("F1_每日时间追踪表.xlsx", create_f1_daily_time_tracking),
        ("F2_优先级决策工作表.xlsx", create_f2_priority_decision),
        ("F3_任务切换成本记录表.xlsx", create_f3_task_switching),
        ("F4_碎片时间清单.xlsx", create_f4_fragmented_time),
        ("F5_精力管理四维度表.xlsx", create_f5_energy_management),
        ("F6_30天行动计划追踪表.xlsx", create_f6_30day_action_plan),
    ]

    for filename, create_func in files:
        filepath = os.path.join(base_dir, filename)
        wb = create_func()
        wb.save(filepath)
        print(f"Created: {filepath}")

    print("\nAll Excel files created successfully!")


if __name__ == "__main__":
    main()

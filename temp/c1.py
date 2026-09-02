import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
OUTPUT_DIR = 'D:/新课开发/工作手册/客户隐性需求挖掘与验证/完整课程包/06_工具表单'
os.makedirs(OUTPUT_DIR, exist_ok=True)
HEADER_BG='C00000';HEADER_FG='FFFFFF';TITLE_BG='E85053'
SECTION_BG='F2F2F2';ALT_ROW_BG='FFF0F0';INPUT_BG='DAEEF3';BORDER_CLR='AAAAAA'
def hf(c):return PatternFill("solid",fgColor=c)
def tb():
 s=Side(style="thin",color=BORDER_CLR)
 return Border(left=s,right=s,top=s,bottom=s)
def ctr():return Alignment(horizontal="center",vertical="center",wrap_text=True)
def lft():return Alignment(horizontal="left",vertical="center",wrap_text=True)

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate 07_Interaction_Flow.xlsx - Interaction Flow Design"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

out = 'D:/新课开发/领导力/01-领导力重生：AI时代真正稀缺的领导者内核/成果demo/KnowledgeShareEvent_Demo/07_Interaction_Flow.xlsx'

wb = Workbook()

# ===== Sheet 1: Overall Interaction Flow =====
ws1 = wb.active
ws1.title = '整体互动流程'

headers = ['模块', '环节', '时长', '学员行为', '讲师行为', '物料准备', '关键要点']
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data = [
    ['Day 1 - 开场', '抛困境：案例Part 1', '45分钟', '看案例 → 凭直觉选择立场', '呈现案例 → 引导初步讨论 → 揭示Part 1', '案例打印件每人1份', '让学员感受到真实的张力，不要急着给框架'],
    ['Day 1 - 模块一', '困境深化', '45分钟', '小组讨论：张明为什么动不了手', '追问：你观察到张明内心在维护什么？', '大白纸、马克笔', '不要揭示答案，让学员充分感受'],
    ['Day 1 - 模块一', '框架引入：志向金字塔', '45分钟', '听讲 → 个人反思 → 两人对话', '讲授框架 → 引导反思 → 板书学员观点', '志向金字塔大海报', '框架是工具，不是结论'],
    ['Day 1 - 模块一', '体验环节：目光练习', '45分钟', '两人对视 → 分享感受 → 讨论差异', '引导练习 → 观察学员反应 → 总结洞察', '计时器、背景音乐', '这是身体在场的第一个锚点'],
    ['Day 1 - 模块二', '困境深化：Part 3揭示', '45分钟', '看新情境 → 小组讨论CEO最后通牒', '呈现Part 3 → 追问选择与代价', '大白纸、马克笔', '张明现在有几个选项？'],
    ['Day 1 - 模块二', '框架引入：艰难抉择四问', '60分钟', '听讲 → 用工具分析案例 → 小组分享', '讲授框架 → 引导工具练习 → 板书分析', '艰难抉择四问海报、利益相关方地图', '让学员自己用工具，不要替他们用'],
    ['Day 1 - 模块二', '体验环节：60秒决策', '60分钟', '情境模拟 → 60秒内做决定 → 观察者反馈', '发情境 → 计时 → 引导反思', '情境卡（8张）、记录表', '压力下的决策暴露真实的认知模式'],
    ['Day 2 - 模块三', '困境深化：Part 4揭示', '45分钟', '角色扮演张明和小刘 → 讨论信任变化', '引导角色扮演 → 追问深层动机', '角色扮演指引卡', '小刘为什么来问张明？'],
    ['Day 2 - 模块三', '框架引入：信任资产负债表', '45分钟', '听讲 → 工具练习 → 小组分享', '讲授框架 → 引导盘点练习', '信任资产负债表工作表、PPT', '信任是行为累计，不是感觉'],
    ['Day 2 - 模块三', '体验环节：信任盘点工作坊', '45分钟', '选择关键下属 → 完成信任盘点 → 小组分享', '巡视各组 → 记录典型洞察 → 总结', '大白纸、马克笔、信任资产负债表模板', '把抽象的信任变成可量化的行为'],
    ['Day 2 - 模块四', '困境深化', '45分钟', '讨论：如果COO来裁会怎样', '追问：保护vs培养的边界在哪里', '大白纸、马克笔', '担责是为自己还是为下属？'],
    ['Day 2 - 模块四', '框架引入：责任四象限', '45分钟', '听讲 → 工具练习 → 分享', '讲授框架 → 引导识别自己的错位', '责任四象限海报', '最常见的错位：第四象限的事第一象限来扛'],
    ['Day 2 - 模块四', '体验环节：担责承诺仪式', '45分钟', '书写承诺 → 两人见证 → 志愿公开', '引导仪式 → 见证承诺 → 总结', '承诺纸、背景音乐', '公开的承诺才有约束力'],
    ['Day 2 - 结尾', '回到困境', '45分钟', '重新做决定 → 说明理由变化 → 分享收获', '引导回到起点 → 追问什么改变了 → 布置作业', '课程反馈问卷', '两天前和两天后的答案差异，才是真正的学习'],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, val in enumerate(row_data, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = Alignment(vertical='top', wrap_text=True)

col_widths = [15, 20, 10, 25, 30, 20, 30]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 2: Module 1 Interaction Design =====
ws2 = wb.create_sheet('模块一互动设计')

headers2 = ['环节', '时长', '具体流程', '讲师话术示例', '学员可能的反应', '应对方式']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data2 = [
    ['开场白', '3分钟', '回顾课程框架：四力模型', '今天我们从张明的困境开始，一起看看领导力到底是怎么回事', '学员可能期待知识传授', '预告：这不是听课，是想自己的问题'],
    ['案例呈现', '10分钟', '展示Part 1，只说情境，不揭示后续', '张明面临一个选择，请告诉我你们的直觉：张明应该裁员吗？', '学员会急着给建议', '追问：你的建议是建立在什么假设上？'],
    ['小组讨论', '20分钟', '每组讨论：张明内心在维护什么？', '不要给答案，先感受。什么让张明动不了手？', '讨论陷入道德判断', '拉回：他是在维护兄弟，还是在维护自己？'],
    ['全体分享', '15分钟', '各组分享观察，讲师记录不点评', '说出你的观察，不要说应该怎么做', '观点分散', '汇总：这几个视角的共同点是什么？'],
    ['框架引入', '20分钟', '志向金字塔讲授', '志向不是目标，是你想成为谁。目标可以换，志向不能换。', '学员想把这个框架套到张明身上', '提醒：先想自己，再想张明'],
    ['个人反思', '15分钟', '写下自己最近三个不得不做的决定', '这三个月里，有哪件事你是在等别人做决定？', '写得很表面', '追问：那个决定的背后，你在维护什么志向？'],
    ['两人对话', '10分钟', '分享反思，伙伴问：你真正的北极星是什么', '把你的发现告诉对方，让对方追问你', '对话变成抱怨', '提醒：对话不是诉苦，是探索'],
    ['目光练习', '20分钟', '两人对视2分钟，体验真实连接', '不说话，不笑，看着对方。等我说停。', '尴尬、不自然', '引导：感受你内心的什么在动'],
    ['小组分享', '15分钟', '分享目光练习的感受', '你感受到的是什么？这种感受和日常对话有什么不同？', '表达困难', '描述：很多人说感受到了被看见'],
    ['模块小结', '5分钟', '回到张明困境，埋下伏笔', '如果张明有北极星，他的北极星是什么？', '能说出方向', '过渡：这个问题，我们明天继续'],
]

for row_idx, row_data in enumerate(data2, 2):
    for col_idx, val in enumerate(row_data, 1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = Alignment(vertical='top', wrap_text=True)

col_widths2 = [12, 10, 25, 35, 25, 25]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 3: Module 2 Interaction Design =====
ws3 = wb.create_sheet('模块二互动设计')

for col, h in enumerate(headers2, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data3 = [
    ['回顾与衔接', '5分钟', '回顾模块一，问学员对张明的理解变化', '昨天结束时，有人说张明是在维护兄弟，有人说是维护自己。现在呢？', '忘记昨天的讨论', '用PPT简要回放昨天的核心观点'],
    ['Part 3揭示', '10分钟', '揭示CEO最后通牒的情境', 'CEO说：如果张明下不了手，COO来处理。这句话意味着什么？', '立刻进入决策模式', '提醒：先感受张明现在的处境'],
    ['情境模拟', '15分钟', '学员两两配对，分别扮演张明和CEO对话', '如果你是张明，CEO就在你面前，你怎么说？', '表演过度或不敢开口', '观察：谁在逃避，谁在做决定'],
    ['小组讨论', '20分钟', '讨论：张明现在有几个选项？各有什么代价？', '列出所有选项，包括不选的代价', '列选项时过于理性', '追问：每个选项背后，张明在放弃什么？'],
    ['框架讲授', '25分钟', '艰难抉择四问讲授', '问自己四个问题：利益相关方是谁？取舍标准是什么？我的盲区在哪？谁担责？', '想把四问套用到张明身上', '要求：先用四问问自己'],
    ['工具练习', '20分钟', '用四问工具分析自己当前的一个两难', '现在想一件你正在纠结的事，用四问走一遍', '说没有两难', '追问：过去三个月你做过的最难的决定是什么？'],
    ['小组分享', '15分钟', '分享工具练习的结果', '说出你的两难和你的四问答案，其他人追问', '分享过于笼统', '追问：你的取舍标准是什么？为什么？'],
    ['60秒决策练习', '30分钟', '情境卡练习：每个情境60秒内必须决定', '听清情境，60秒，说出你的决定和理由。', '时间压力下慌乱', '记录：谁是果断的，谁是犹豫的'],
    ['观察者反馈', '15分钟', '搭档分享观察：对方的决策风格', '你观察到搭档做决定时有什么特点？', '反馈变成评价', '引导：说观察，不说判断'],
    ['模块小结', '5分钟', '总结：艰难抉择的本质是取舍', '没有完美的选择，只有你能接受的选择。', '', ''],
]

for row_idx, row_data in enumerate(data3, 2):
    for col_idx, val in enumerate(row_data, 1):
        c = ws3.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = Alignment(vertical='top', wrap_text=True)

for i, w in enumerate(col_widths2, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 4: Module 3 Interaction Design =====
ws4 = wb.create_sheet('模块三互动设计')

for col, h in enumerate(headers2, 1):
    c = ws4.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data4 = [
    ['回顾与衔接', '5分钟', '回顾模块二，问学员对张明的理解变化', '昨天很多人说张明是在逃避。现在你觉得呢？', '忘记昨天的结论', '用PPT回放关键观点'],
    ['Part 4揭示', '10分钟', '揭示小刘来找张明谈offer', '小刘拿着1.5倍offer来找张明，说想听听他的意见。你们怎么看？', '立刻分析谁对谁错', '提醒：先感受张明和小刘各自的处境'],
    ['角色扮演', '20分钟', '两人一组扮演张明和小刘，模拟对话', '扮演张明的人，要尽量体会张明此刻的内心。扮演小刘的人，想清楚小刘真正想要的是什么。', '对话变成辩论', '观察：谁在说真话，谁在表演'],
    ['第二轮角色互换', '10分钟', '角色互换，重新对话', '换一种可能的走向，看看会发生什么', '不知道怎么换', '引导：如果张明说支持，会怎样？如果说不支持呢？'],
    ['框架讲授', '20分钟', '信任资产负债表讲授', '信任不是感觉，是行为累计。每次互动要么是资产，要么是负债。', '想把框架套到案例上', '要求：先给自己评，不要先给张明评'],
    ['工具练习', '15分钟', '选择关键下属，完成信任资产负债表', '想一个你团队里的人，用这个工具给自己评分', '打分很随意', '追问：是什么让你扣分的？那是真实发生的行为吗？'],
    ['小组工作坊', '25分钟', '信任盘点：每人在组内分享3分钟', '说出你的发现，不需要解决方案，只需要诚实', '分享变成诉苦', '引导：从负债倒推，哪些行为在消耗你的信任？'],
    ['讨论', '10分钟', '领导者最容易犯的信任错误是什么', '回顾你经历过的信任破裂，是从什么时候开始的？', '停留在抱怨层面', '总结：信任破裂往往不是大事，而是小事累积'],
    ['模块小结', '5分钟', '总结：信任是呼吸，必须持续', '信任不是一次行为，是每天的选择。', '', ''],
]

for row_idx, row_data in enumerate(data4, 2):
    for col_idx, val in enumerate(row_data, 1):
        c = ws4.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = Alignment(vertical='top', wrap_text=True)

for i, w in enumerate(col_widths2, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 5: Module 4 Interaction Design =====
ws5 = wb.create_sheet('模块四互动设计')

for col, h in enumerate(headers2, 1):
    c = ws5.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data5 = [
    ['回顾与衔接', '5分钟', '回顾模块三，问信任与责任的关系', '如果张明之前做了不同的选择，今天的困境会不同吗？', '忘记前面的分析', '用时间线回顾张明的选择路径'],
    ['情境推演', '20分钟', '讨论：如果COO来裁，会发生什么', '不是问谁对谁错，是问会发生什么连锁反应', '停留在道德判断', '追问：COO、小刘、其他下属会怎么反应？'],
    ['框架讲授', '20分钟', '责任四象限讲授', '高影响+高难度的事，必须自己扛。低影响+低难度的事，让下属练。', '觉得四象限太简单', '追问：你在哪个象限花的时间最多？'],
    ['工具练习', '15分钟', '识别自己的错位：哪些事你扛多了？哪些让少了？', '回想过去两周，你扛了哪些其实不该自己扛的事？', '想不出来', '引导：哪些事是你主动抢过来的？'],
    ['讨论', '10分钟', '保护vs培养的边界在哪里', '什么时候该保护，什么时候该放手？', '讨论变成理论', '案例：教练不会替孩子上场打球'],
    ['个人书写', '10分钟', '写下这种事我不会让——自己最核心的担当', '这件事代表了你是谁。把它写下来。', '写得很空', '引导：这件事如果让了，你还是你吗？'],
    ['两人见证', '10分钟', '互相告诉对方，请对方见证', '告诉对方你要扛什么，让他记住', '觉得形式主义', '提醒：见证是一种承诺，会让践行的可能性更高'],
    ['志愿公开', '20分钟', '自愿上台，向全体公开承诺', '想好的可以说，不想说也没关系。但说了，就要做到。', '不敢上台', '鼓励：公开的承诺才有约束力'],
    ['模块小结', '5分钟', '总结：责任担当是领导者的标志', '船长和船共存亡。你在，你的领导力就在。', '', ''],
]

for row_idx, row_data in enumerate(data5, 2):
    for col_idx, val in enumerate(row_data, 1):
        c = ws5.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = Alignment(vertical='top', wrap_text=True)

for i, w in enumerate(col_widths2, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 6: Time Allocation Overview =====
ws6 = wb.create_sheet('时间分配总览')

headers6 = ['项目', 'Day 1', 'Day 2', '合计', '占比', '说明']
for col, h in enumerate(headers6, 1):
    c = ws6.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')

data6 = [
    ['开场/结尾', '45分钟', '45分钟', '90分钟', '12.5%', '抛困境 + 回到困境'],
    ['模块一：志向设立', '135分钟', '0', '135分钟', '18.8%', '困境深化45分 + 框架45分 + 体验45分'],
    ['模块二：艰难抉择', '165分钟', '0', '165分钟', '22.9%', '困境深化45分 + 框架60分 + 体验60分'],
    ['模块三：信任建立', '0', '135分钟', '135分钟', '18.8%', '困境深化45分 + 框架45分 + 体验45分'],
    ['模块四：责任担当', '0', '135分钟', '135分钟', '18.8%', '困境深化45分 + 框架45分 + 体验45分'],
    ['午休', '60分钟', '60分钟', '120分钟', '0%', '不含在课时内'],
    ['合计', '405分钟', '375分钟', '780分钟', '100%', '2天共13小时（含午休）'],
]

for row_idx, row_data in enumerate(data6, 2):
    for col_idx, val in enumerate(row_data, 1):
        c = ws6.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 8:
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        c.alignment = Alignment(horizontal='center' if col_idx != 6 else 'left', vertical='center')

col_widths6 = [15, 12, 12, 12, 10, 35]
for i, w in enumerate(col_widths6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

# Save
wb.save(out)
print(f'Created: {out}')

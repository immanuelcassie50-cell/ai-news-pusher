# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# Color scheme
primary = "22223b"
secondary = "4a4e69"
accent = "9a8c98"
light = "c9ada7"
bg = "f2e9e4"

header_fill = PatternFill(start_color=primary, end_color=primary, fill_type="solid")
alt_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
border = Border(
    left=Side(style='thin', color=accent),
    right=Side(style='thin', color=accent),
    top=Side(style='thin', color=accent),
    bottom=Side(style='thin', color=accent)
)

ws = wb.active
ws.title = u"表单使用指引"

# Title row
ws.merge_cells('A1:D1')
title_cell = ws.cell(row=1, column=1, value=u"说服课配套表单使用指引")
title_cell.font = Font(bold=True, color="FFFFFF", size=14)
title_cell.fill = header_fill
title_cell.alignment = Alignment(horizontal="center", vertical="center")
title_cell.border = border
ws.row_dimensions[1].height = 30

# Headers
headers = [u"表单名称", u"用途", u"使用时机", u"填写要点"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill(start_color=secondary, end_color=secondary, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Column widths
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 40

# Data rows
data = [
    [u"受众分析表", u"分析说服对象的特征、态度和需求", 
     u"准备说服前，先对受众进行分类分析", 
     u"1. 确定场景和受众类型\n2. 评估当前态度（开放/中立/抵触）\n3. 选择对应策略组合\n4. 明确说服目标"],
    
    [u"说服方案表", u"规划说服的具体步骤和话术", 
     u"制定说服策略时使用", 
     u"1. 按逻辑顺序排列步骤\n2. 每步明确使用的技巧\n3. 预判可能出现的异议\n4. 准备应对方案"],
    
    [u"实践追踪表", u"记录每次说服实践的效果和反思", 
     u"每次说服后及时记录", 
     u"1. 真实记录效果评分\n2. 反思成功/失败原因\n3. 总结可改进的点\n4. 定期回顾进步轨迹"],
    
    [u"学习记录表", u"跟踪学习进度和成长曲线", 
     u"每周固定时间填写", 
     u"1. 如实评估掌握程度\n2. 设定具体可衡量的目标\n3. 定期检视目标完成情况\n4. 动态调整学习计划"],
]

for row_idx, row_data in enumerate(data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row_idx % 2 == 0:
            cell.fill = alt_fill
    ws.row_dimensions[row_idx].height = 80

output_path = r"D:\新课开发\公众表达\黄执中\说服课\完整课程包\11_配套Excel\表单使用指引.xlsx"
wb.save(output_path)
print("Created: " + output_path)

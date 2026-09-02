import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "假设识别矩阵"

header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
data_font = Font(name='微软雅黑', size=10)
title_font = Font(name='微软雅黑', bold=True, size=14)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws.merge_cells('A1:K1')
ws['A1'] = '华东制造集团ERP升级项目 - 假设识别矩阵'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:K2')
ws['A2'] = '填写日期：2024年7月20日    填写人：李晓峰    版本：V1.0'
ws['A2'].font = Font(name='微软雅黑', size=10, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')
ws.row_dimensions[2].height = 20

headers = ['假设编号', '假设类别', '假设描述', '提出依据', '可能影响', '风险等级', '验证方式', '验证时间', '验证责任人', '验证结果', '备注']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[4].height = 35

assumptions = [
    ['H001', '技术假设', '现有SAP R/3系统数据可以完整迁移至S/4HANA', '供应商技术方案声称迁移成功率99.9%', '若数据丢失或损坏，将导致项目返工，估算增加成本50万，延期1个月', '高', '执行数据质量分析，执行迁移测试', '2024/8/15', '陈建国', '待验证', '需在调研阶段确认'],
    ['H002', '技术假设', '现有服务器配置可以满足S/4HANA性能要求', 'IT部门提供服务器配置清单，满足最低配置要求', '若性能不足，需要升级硬件，导致成本增加约30万', '中', '进行压力测试，收集性能基准数据', '2024/9/30', '刘建华', '待验证', '已有初步评估'],
    ['H003', '技术假设', '与MES系统的接口可以在6周内完成开发', '类似项目经验，标准接口开发周期4周', '若接口延期，将影响生产模块上线', '高', '技术方案评审，确认接口复杂度', '2024/10/15', '刘建华', '待验证', '需实测评估'],
    ['H004', '技术假设', '网络带宽可以支持实时数据传输', 'IT部门确认带宽100Mbps', '若带宽不足，数据同步延迟，影响用户体验', '中', '网络评估测试', '2024/9/15', '刘建华', '待验证', ''],
    ['H005', '业务假设', '各子公司愿意放弃本地流程，接受集团统一流程', '项目章程中获得各子公司总经理签字确认', '若抵制统一流程，项目可能无法推进，需要变更范围', '高', '在启动会上正式访谈各子公司负责人', '2024/7/25', '赵敏', '待验证', '存在较大不确定性'],
    ['H006', '业务假设', '关键用户可以投入足够时间配合实施', '客户方承诺每模块2名关键用户全职参与', '若关键用户时间不足，调研和测试质量受影响', '高', '确认关键用户名单及其时间安排', '2024/7/30', '李晓峰', '待验证', '已有个别反馈称时间不足'],
    ['H007', '业务假设', '现有物料编码体系可以标准化而不需要重新编码', '财务部经理确认物料账实相符', '若需要重新编码，工作量巨大，估算增加60人天', '高', '执行物料数据质量分析', '2024/8/30', '陈建国', '待验证', ''],
    ['H008', '业务假设', '财务月结流程可以在新系统中压缩至5天', '参考行业最佳实践，其他企业有此先例', '若无法实现，项目核心目标之一未达成', '中', '流程建模与仿真测试', '2025/2/28', '赵敏', '待验证', '需在设计阶段验证'],
    ['H009', '业务假设', '供应商主数据可以由采购部统一提供并维护', '采购部经理口头承诺', '若数据分散维护，质量难以保证', '低', '确认数据维护流程和责任人', '2024/10/30', '赵敏', '待验证', ''],
    ['H010', '资源假设', '预算800万可以覆盖全部实施范围', '合同金额确认，覆盖一期范围', '若范围蔓延或假设未验证导致返工，可能超预算', '中', '每月成本跟踪，差异分析', '持续', '李晓峰', '待验证', '已有超支风险'],
    ['H011', '资源假设', '可以在2周内完成服务器采购和部署', '供应商标准交付周期', '若采购延期，整个项目延期', '中', '采购流程确认', '2024/8/1', '刘建华', '已验证通过', '已与采购部确认'],
    ['H012', '资源假设', '外部顾问可以在需要时到场支持', '供应商承诺7*24小时响应', '若顾问无法到场，技术问题难以解决', '低', '合同条款确认', '2024/7/31', '张明华', '已验证通过', '合同已签署'],
    ['H013', '时间假设', '12个月的实施周期足够完成全部功能', '类似规模项目周期参考', '若时间不足，可能需要压缩范围或增加资源', '高', '详细工作量评估，项目进度模拟', '2024/9/15', '李晓峰', '待验证', '已有延期预警'],
    ['H014', '时间假设', '业务调研可以在6周内完成', '常规项目调研周期', '若调研延期，后续设计、开发和测试都会顺延', '中', '调研进度跟踪', '2024/8/30', '赵敏', '待验证', '已出现延期'],
    ['H015', '时间假设', 'UAT测试可以在4周内完成', '标准UAT周期', '若测试发现问题多，修复周期长，可能延期上线', '中', 'UAT计划评审', '2025/3/15', '吴婷', '待验证', ''],
    ['H016', '外部假设', '国资委的合规要求不会有重大变化', '已与IT部门确认当前政策', '若政策变化，可能需要增加功能或修改方案', '低', '政策跟踪，定期沟通', '持续', '陈永强', '待验证', '需定期确认'],
    ['H017', '外部假设', '没有其他重大项目与本项目时间冲突', '已协调确认，各部门时间已错开', '若冲突，关键用户被抽调，项目进度受影响', '中', '确认公司重大项目日历', '2024/7/31', '林晓华', '已验证通过', '已协调'],
    ['H018', '外部假设', '主要供应商可以按时交付，不会有供应链中断', '供应商已通过资质审核', '若供应商出问题，可能影响硬件交付', '低', '供应商状态跟踪', '2024/8/31', '刘建华', '待验证', ''],
    ['H019', '容易被忽视的假设', '客户方项目对接人员有足够的决策权限', '对方是信息部经理职位', '若需要更高层级审批，决策周期拉长', '高', '在启动会上确认汇报线和决策权限', '2024/7/25', '李晓峰', '待验证', '已在启动会发现问题'],
    ['H020', '容易被忽视的假设', '项目范围说明书中的除外条款是完整的', '基于双方商务谈判结果', '若实际执行中发现遗漏项，可能产生范围争议', '中', '范围边界确认，与各业务部门访谈', '2024/8/15', '李晓峰', '待验证', ''],
]

for row_idx, row_data in enumerate(assumptions, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border
        if row_data[5] == '高':
            cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        elif row_data[5] == '中':
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        elif row_data[9] == '已验证通过':
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    ws.row_dimensions[row_idx].height = 45

column_widths = [12, 15, 35, 30, 35, 10, 25, 12, 15, 12, 20]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws2 = wb.create_sheet('填写说明')
ws2['A1'] = '假设识别矩阵填写说明'
ws2['A1'].font = Font(name='微软雅黑', bold=True, size=14)

output_path = 'D:/CC/新课开发/工作手册/假设管理：项目经理的风险前置手册/完整课程包/07-成果demo/假设识别矩阵-演示.xlsx'
wb.save(output_path)
print(f"OK: {output_path}")

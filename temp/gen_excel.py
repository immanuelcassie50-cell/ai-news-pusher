#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = "D:/新课开发/HR/员工关系/1.员工关系重生-从事务处理者到组织韧性架构师的角色转型/配套Excel/"

HEADER_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

HEADER_FONT = Font(name="Arial", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='AAAAAA'),
    right=Side(style='thin', color='AAAAAA'),
    top=Side(style='thin', color='AAAAAA'),
    bottom=Side(style='thin', color='AAAAAA')
)

def set_column_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def create_header_row(ws, row_num, values, fill=None, font=None, height=30):
    if fill is None: fill = HEADER_FILL
    if font is None: font = HEADER_FONT
    ws.row_dimensions[row_num].height = height
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    return row_num + 1

print("Script loaded successfully")

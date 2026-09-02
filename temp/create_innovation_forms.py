#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建创新管理体系配套表单
- 表单使用指引.xlsx
- 配套表单_空表.xlsx
- 配套表单_填好版.xlsx
"""

import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

# 配色主题 - 蓝绿色创新主题
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # 深蓝
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")  # 中蓝
ACCENT_FILL = PatternFill(start_color="00B894", end_color="00B894", fill_type="solid")  # 绿色
LIGHT_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")  # 浅蓝
ALT_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")  # 浅灰

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SUBHEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_column_widths(ws, widths):
    """设置列宽"""
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

def style_header_row(ws, row, num_cols, fill=None):
    """为标题行添加样式"""
    fill = fill or HEADER_FILL
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

def style_data_row(ws, row, num_cols, alt=False):
    """为数据行添加样式"""
    fill = ALT_ROW_FILL if alt else PatternFill(fill_type="solid")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        if alt:
            cell.fill = ALT_ROW_FILL
        cell.font = BODY_FONT
        cell.alignment = LEFT_ALIGN
        cell.border = THIN_BORDER

def freeze_and_filter(ws, row, col):
    """冻结窗格和添加筛选器"""
    ws.freeze_panes = ws.cell(row=row, column=col)

# ==================== 文件1: 表单使用指引 ====================
def create_guide_file():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "表单使用指引"

    # 标题
    ws.merge_cells('A1:E1')
    ws['A1'] = "创新管理体系配套表单使用指引"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 40

    # 说明文字
    ws.merge_cells('A2:E2')
    ws['A2'] = "本套表单用于创新项目全流程管理，涵盖从创意生成到商业化落地的各个阶段。"
    ws['A2'].font = BODY_FONT
    ws['A2'].alignment = LEFT_ALIGN
    ws.row_dimensions[2].height = 30

    # 表单清单
    row = 4
    ws.cell(row=row, column=1, value="表单名称").fill = SUBHEADER_FILL
    ws.cell(row=row, column=1).font = SUBHEADER_FONT
    ws.cell(row=row, column=2, value="用途说明").fill = SUBHEADER_FILL
    ws.cell(row=row, column=2).font = SUBHEADER_FONT
    ws.cell(row=row, column=3, value="使用阶段").fill = SUBHEADER_FILL
    ws.cell(row=row, column=3).font = SUBHEADER_FONT
    ws.cell(row=row, column=4, value="更新频率").fill = SUBHEADER_FILL
    ws.cell(row=row, column=4).font = SUBHEADER_FONT
    ws.cell(row=row, column=5, value="负责人").fill = SUBHEADER_FILL
    ws.cell(row=row, column=5).font = SUBHEADER_FONT

    forms_info = [
        ("创新项目跟踪表", "记录项目基本信息、阶段进展、风险和关键成果", "全流程", "每周", "项目经理"),
        ("创意生成记录表", "记录创意想法的来源、评估和筛选结果", "创意阶段", "随时", "创新专员"),
        ("阶段关卡评审表", "各阶段里程碑评审，记录决策和反馈", "关键节点", "按阶段", "评审委员会"),
        ("团队角色分配表", "明确团队成员角色、职责和参与程度", "项目启动", "项目开始时", "项目负责人"),
        ("商业化检查清单", "商业化前的各项检查项确认", "商业化阶段", "商业化前", "商业化负责人"),
        ("创新指标仪表盘", "汇总各项创新指标数据", "全流程", "每月", "数据分析师"),
    ]

    for i, (name, desc, stage, freq, owner) in enumerate(forms_info, start=1):
        r = row + i
        ws.cell(row=r, column=1, value=name).border = THIN_BORDER
        ws.cell(row=r, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=r, column=3, value=stage).border = THIN_BORDER
        ws.cell(row=r, column=4, value=freq).border = THIN_BORDER
        ws.cell(row=r, column=5, value=owner).border = THIN_BORDER
        style_data_row(ws, r, 5, alt=(i % 2 == 0))

    # 使用流程说明
    row = row + len(forms_info) + 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="创新项目管理流程").fill = SUBHEADER_FILL
    ws.cell(row=row, column=1).font = SUBHEADER_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN

    process_steps = [
        "1. 创意生成 → 填写《创意生成记录表》",
        "2. 项目立项 → 填写《创新项目跟踪表》基础信息",
        "3. 团队组建 → 填写《团队角色分配表》",
        "4. 阶段评审 → 每次评审填写《阶段关卡评审表》",
        "5. 商业化准备 → 填写《商业化检查清单》",
        "6. 指标跟踪 → 每月更新《创新指标仪表盘》",
    ]

    for i, step in enumerate(process_steps):
        r = row + i + 1
        ws.merge_cells(f'A{r}:E{r}')
        ws.cell(row=r, column=1, value=step)
        ws.cell(row=r, column=1).font = BODY_FONT
        ws.cell(row=r, column=1).alignment = LEFT_ALIGN
        ws.cell(row=r, column=1).border = THIN_BORDER

    # 配色说明
    row = row + len(process_steps) + 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="状态颜色说明").fill = SUBHEADER_FILL
    ws.cell(row=row, column=1).font = SUBHEADER_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN

    color_desc = [
        ("进行中", "绿色背景", "00B894"),
        ("已暂停", "黄色背景", "FDCB6E"),
        ("已完成", "蓝色背景", "2E75B6"),
        ("已取消", "红色背景", "E74C3C"),
    ]

    for i, (status, desc, color) in enumerate(color_desc):
        r = row + i + 1
        ws.cell(row=r, column=1, value=status).border = THIN_BORDER
        ws.cell(row=r, column=2, value=desc).border = THIN_BORDER
        cell_color = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=r, column=3).fill = cell_color
        ws.cell(row=r, column=3).border = THIN_BORDER

    set_column_widths(ws, [25, 40, 15, 12, 15])
    freeze_and_filter(ws, 5, 2)

    output_path = "D:/新课开发/管理学/39-创新管理体系/配套表单/表单使用指引.xlsx"
    wb.save(output_path)
    print(f"创建完成: {output_path}")

# ==================== 文件2和3的通用函数 ====================
def create_sheets_common(wb, with_data=False):
    """创建通用表单结构"""

    # ---------- Sheet 1: 创新项目跟踪表 ----------
    ws1 = wb.active
    ws1.title = "创新项目跟踪表"

    headers1 = ["项目名称", "负责人", "开始日期", "阶段", "状态", "关键成果", "风险", "备注"]
    ws1.append(headers1)
    style_header_row(ws1, 1, len(headers1))

    if with_data:
        data1 = [
            ("智能仓储系统创新", "张明", "2026-03-15", "执行中", "进行中", "完成需求分析和方案设计", "技术整合难度较高", "预计6月完成原型"),
            ("客户画像2.0", "李华", "2026-04-01", "计划中", "已暂停", "完成数据收集", "资源分配不足", "等待优先级调整"),
            ("供应链优化AI", "王芳", "2026-02-20", "已完成", "已完成", "模型上线并稳定运行", "无", "年节约成本200万"),
        ]
        for i, row_data in enumerate(data1, start=2):
            ws1.append(list(row_data))
            style_data_row(ws1, i, len(headers1), alt=(i % 2 == 0))

    set_column_widths(ws1, [20, 12, 14, 12, 12, 25, 20, 20])
    freeze_and_filter(ws1, 2, 2)

    # ---------- Sheet 2: 创意生成记录表 ----------
    ws2 = wb.create_sheet("创意生成记录表")

    headers2 = ["创意编号", "日期", "创意描述", "来源", "创新类型", "初步评估", "筛选结果"]
    ws2.append(headers2)
    style_header_row(ws2, 1, len(headers2))

    if with_data:
        data2 = [
            ("ID-2026-001", "2026-03-10", "基于计算机视觉的自动分拣系统", "技术团队", "产品创新", "技术可行，市场潜力大", "通过，进入立项"),
            ("ID-2026-002", "2026-03-12", "使用区块链追溯供应链", "外部顾问", "流程创新", "实施难度大，投资回报不明", "待定，需进一步评估"),
            ("ID-2026-003", "2026-03-15", "AR远程运维指导", "客户反馈", "服务创新", "市场需求明确，技术成熟", "通过，进入立项"),
            ("ID-2026-004", "2026-03-18", "基于大数据的预测性维护", "内部研究", "技术创新", "与现有系统整合需评估", "否决，与主战略不符"),
        ]
        for i, row_data in enumerate(data2, start=2):
            ws2.append(list(row_data))
            style_data_row(ws2, i, len(headers2), alt=(i % 2 == 0))

    set_column_widths(ws2, [15, 14, 35, 15, 15, 20, 20])
    freeze_and_filter(ws2, 2, 2)

    # ---------- Sheet 3: 阶段关卡评审表 ----------
    ws3 = wb.create_sheet("阶段关卡评审表")

    headers3 = ["阶段", "评审日期", "评审人", "决策", "反馈意见", "下一阶段计划", "资源需求"]
    ws3.append(headers3)
    style_header_row(ws3, 1, len(headers3))

    if with_data:
        data3 = [
            ("概念阶段", "2026-03-20", "评审委员会", "通过", "建议加强市场需求分析", "进入可行性研究", "需增加市场调研预算"),
            ("可行性研究", "2026-04-15", "评审委员会", "有条件通过", "技术方案可行，商业化路径需细化", "进入方案设计", "增加技术专家支持"),
            ("方案设计", "2026-05-20", "技术委员会", "通过", "设计合理，可进入执行", "进入开发执行", "开发团队到位"),
            ("执行监控", "2026-08-15", "项目管理办公室", "通过", "进度正常，风险可控", "准备商业化", "市场团队提前介入"),
        ]
        for i, row_data in enumerate(data3, start=2):
            ws3.append(list(row_data))
            style_data_row(ws3, i, len(headers3), alt=(i % 2 == 0))

    set_column_widths(ws3, [15, 14, 15, 15, 25, 25, 20])
    freeze_and_filter(ws3, 2, 2)

    # ---------- Sheet 4: 团队角色分配表 ----------
    ws4 = wb.create_sheet("团队角色分配表")

    headers4 = ["角色", "姓名", "主要职责", "技能特长", "参与程度", "备注"]
    ws4.append(headers4)
    style_header_row(ws4, 1, len(headers4))

    if with_data:
        data4 = [
            ("项目总监", "张明", "项目整体把控", "战略规划、团队管理", "全职", "创新领导小组组长"),
            ("技术负责人", "李强", "技术方案设计实施", "AI、机器学习、系统架构", "全职", ""),
            ("产品经理", "王芳", "需求分析、产品设计", "用户体验、市场分析", "全职", ""),
            ("数据分析师", "赵丽", "数据收集与分析", "大数据分析、BI工具", "兼职", "同时支持其他项目"),
            ("市场专员", "钱伟", "市场调研与推广", "市场营销、客户沟通", "兼职", "商业化阶段介入"),
            ("创新专员", "孙燕", "创意收集与评估", "创新方法论、头脑风暴", "全职", ""),
        ]
        for i, row_data in enumerate(data4, start=2):
            ws4.append(list(row_data))
            style_data_row(ws4, i, len(headers4), alt=(i % 2 == 0))

    set_column_widths(ws4, [15, 12, 25, 25, 15, 20])
    freeze_and_filter(ws4, 2, 2)

    # ---------- Sheet 5: 商业化检查清单 ----------
    ws5 = wb.create_sheet("商业化检查清单")

    headers5 = ["检查项", "是/否", "负责人", "完成日期", "备注"]
    ws5.append(headers5)
    style_header_row(ws5, 1, len(headers5))

    if with_data:
        data5 = [
            ("产品技术验证完成", "是", "李强", "2026-07-15", ""),
            ("商业模式确定", "是", "王芳", "2026-07-20", "订阅制为主"),
            ("定价策略审批", "是", "张明", "2026-07-25", ""),
            ("目标客户确认", "是", "钱伟", "2026-08-01", "首批目标客户10家"),
            ("销售渠道建立", "否", "钱伟", "2026-08-20", "正在洽谈代理商"),
            ("营销材料准备", "否", "钱伟", "2026-08-25", "需要市场部支持"),
            ("客服团队培训", "否", "孙燕", "2026-08-28", "计划8月底完成"),
            ("法律合规审查", "是", "法务部", "2026-08-10", "已通过"),
            ("知识产权保护", "是", "李强", "2026-08-12", "已申请专利"),
            ("财务收益预测", "是", "财务部", "2026-08-15", "预计年收益500万"),
        ]
        for i, row_data in enumerate(data5, start=2):
            ws5.append(list(row_data))
            style_data_row(ws5, i, len(headers5), alt=(i % 2 == 0))

    set_column_widths(ws5, [25, 10, 15, 15, 30])
    freeze_and_filter(ws5, 2, 2)

    # ---------- Sheet 6: 创新指标仪表盘 ----------
    ws6 = wb.create_sheet("创新指标仪表盘")

    # 标题行
    ws6.merge_cells('A1:F1')
    ws6['A1'] = "创新指标仪表盘"
    ws6['A1'].font = TITLE_FONT
    ws6['A1'].alignment = CENTER_ALIGN
    ws6.row_dimensions[1].height = 35

    # 指标数据
    metrics_headers = ["指标类别", "指标名称", "当前值", "目标值", "同比变化", "状态"]
    ws6.append([])
    ws6.append(metrics_headers)
    style_header_row(ws6, 3, len(metrics_headers))

    if with_data:
        data6 = [
            ("创意产出", "创意数量（个/月）", "15", "12", "+25%", "超出目标"),
            ("项目数量", "进行中项目（个）", "8", "6", "+33%", "超出目标"),
            ("项目成功率", "项目成功率（%）", "68%", "60%", "+8pp", "超出目标"),
            ("投资回报", "创新投资回报率（%）", "145%", "120%", "+25pp", "超出目标"),
            ("团队满意度", "创新团队满意度（分）", "4.2", "4.0", "+5%", "达到目标"),
            ("商业化周期", "平均商业化周期（月）", "8.5", "9.0", "-6%", "优于目标"),
            ("专利申请", "新增专利（个）", "5", "4", "+25%", "超出目标"),
        ]
        for i, row_data in enumerate(data6, start=4):
            ws6.append(list(row_data))
            style_data_row(ws6, i, len(metrics_headers), alt=(i % 2 == 0))

    # 汇总行
    summary_row = 4 + (len(data6) if with_data else 0) + 1
    ws6.merge_cells(f'A{summary_row}:B{summary_row}')
    ws6.cell(row=summary_row, column=1, value="整体创新绩效").font = BOLD_FONT
    ws6.cell(row=summary_row, column=1).fill = ACCENT_FILL
    ws6.cell(row=summary_row, column=1).font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    ws6.cell(row=summary_row, column=1).alignment = CENTER_ALIGN
    ws6.cell(row=summary_row, column=3, value="表现优秀").font = BOLD_FONT
    ws6.cell(row=summary_row, column=3).fill = ACCENT_FILL
    ws6.cell(row=summary_row, column=3).font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    ws6.cell(row=summary_row, column=3).alignment = CENTER_ALIGN

    set_column_widths(ws6, [15, 25, 12, 12, 15, 15])
    freeze_and_filter(ws6, 4, 2)

# ==================== 文件2: 空表 ====================
def create_blank_file():
    wb = openpyxl.Workbook()
    create_sheets_common(wb, with_data=False)
    output_path = "D:/新课开发/管理学/39-创新管理体系/配套表单/配套表单_空表.xlsx"
    wb.save(output_path)
    print(f"创建完成: {output_path}")

# ==================== 文件3: 填好版 ====================
def create_filled_file():
    wb = openpyxl.Workbook()
    create_sheets_common(wb, with_data=True)

    # 添加项目总览 Sheet
    ws_overview = wb.create_sheet("项目总览", 0)
    ws_overview.merge_cells('A1:D1')
    ws_overview['A1'] = "智能仓储系统创新项目 - 总览"
    ws_overview['A1'].font = TITLE_FONT
    ws_overview['A1'].alignment = CENTER_ALIGN
    ws_overview.row_dimensions[1].height = 40

    ws_overview.merge_cells('A2:D2')
    ws_overview['A2'] = "项目周期: 2026年3月 - 2026年9月 | 团队: 创新领导小组 | 状态: 执行中"
    ws_overview['A2'].font = BODY_FONT
    ws_overview['A2'].alignment = LEFT_ALIGN

    overview_data = [
        ("项目名称", "智能仓储系统创新"),
        ("项目负责人", "张明"),
        ("开始日期", "2026年3月15日"),
        ("预计完成", "2026年9月"),
        ("当前阶段", "执行监控"),
        ("预算使用", "65%"),
        ("进度状态", "正常"),
        ("风险等级", "中等"),
    ]

    for i, (label, value) in enumerate(overview_data, start=4):
        ws_overview.cell(row=i, column=1, value=label).font = BOLD_FONT
        ws_overview.cell(row=i, column=1).border = THIN_BORDER
        ws_overview.cell(row=i, column=2, value=value).border = THIN_BORDER
        ws_overview.cell(row=i, column=2).alignment = LEFT_ALIGN

    set_column_widths(ws_overview, [20, 30, 15, 15])

    output_path = "D:/新课开发/管理学/39-创新管理体系/配套表单/配套表单_填好版.xlsx"
    wb.save(output_path)
    print(f"创建完成: {output_path}")

# ==================== 主程序 ====================
if __name__ == "__main__":
    print("开始创建创新管理体系配套表单...")
    create_guide_file()
    create_blank_file()
    create_filled_file()
    print("全部创建完成！")

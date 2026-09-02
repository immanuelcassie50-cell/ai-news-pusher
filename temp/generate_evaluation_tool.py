# -*- coding: utf-8 -*-
"""
领导力基础课程Excel评估工具生成脚本
使用openpyxl库生成专业的课程评估Excel文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
import os

# ========== 样式定义 ==========
# 颜色定义
DARK_BLUE = "1F4E79"
MEDIUM_BLUE = "2E75B6"
LIGHT_BLUE = "BDD7EE"
VERY_LIGHT_BLUE = "DEEAF1"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
GREEN = "70AD47"
LIGHT_GREEN = "E2EFDA"
GRAY = "808080"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
YELLOW = "FFFF00"

# 边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

medium_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

# 字体样式
title_font = Font(name='微软雅黑', size=16, bold=True, color=WHITE)
header_font = Font(name='微软雅黑', size=11, bold=True, color=WHITE)
subheader_font = Font(name='微软雅黑', size=10, bold=True, color=DARK_BLUE)
normal_font = Font(name='微软雅黑', size=10)
small_font = Font(name='微软雅黑', size=9)

# 填充样式
title_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
header_fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
light_blue_fill = PatternFill(start_color=VERY_LIGHT_BLUE, end_color=VERY_LIGHT_BLUE, fill_type='solid')
light_orange_fill = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type='solid')
light_green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type='solid')
gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

# 对齐样式
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')


def set_column_widths(ws, widths):
    """设置列宽"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def apply_header_style(cell):
    """应用表头样式"""
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border


def apply_title_row(ws, row, text, start_col, end_col):
    """应用标题行"""
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col)
    cell.value = text
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center_align
    cell.border = thin_border


def create_sheet1_pretest(wb):
    """创建前测问卷工作表"""
    ws = wb.active
    ws.title = "前测问卷"

    # 设置列宽
    set_column_widths(ws, [15, 25, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 15])

    # ===== 标题区域 =====
    apply_title_row(ws, 1, "领导力基础课程 - 前测问卷", 1, 13)
    ws.row_dimensions[1].height = 35

    # ===== 基本信息区域 =====
    apply_title_row(ws, 2, "一、基本信息", 1, 13)
    ws.cell(row=3, column=1, value="姓名：").font = subheader_font
    ws.cell(row=3, column=1).border = thin_border
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.cell(row=3, column=3).border = thin_border

    ws.cell(row=3, column=4, value="部门：").font = subheader_font
    ws.cell(row=3, column=4).border = thin_border
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)
    ws.cell(row=3, column=6).border = thin_border

    ws.cell(row=3, column=7, value="岗位：").font = subheader_font
    ws.cell(row=3, column=7).border = thin_border
    ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=8)
    ws.cell(row=3, column=9).border = thin_border

    ws.cell(row=3, column=10, value="工龄：").font = subheader_font
    ws.cell(row=3, column=10).border = thin_border
    ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column=11)
    ws.cell(row=3, column=12).border = thin_border

    ws.cell(row=3, column=13, value="日期：").font = subheader_font
    ws.cell(row=3, column=13).border = thin_border

    # ===== 教学思维自评（5分制，10题） =====
    apply_title_row(ws, 5, "二、教学思维自评（5分制：1=完全不符合，5=完全符合）", 1, 13)
    ws.row_dimensions[5].height = 25

    # 表头
    headers = ["题号", "题目内容", "1分", "2分", "3分", "4分", "5分", "", "", "", "", "", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        apply_header_style(cell)

    # 教学思维自评题目
    questions1 = [
        "1. 我清楚管理者与领导者的核心区别",
        "2. 我理解从\"做事\"到\"做人\"的转型重要性",
        "3. 我意识到时间分配需要从紧急事务转向重要事务",
        "4. 我掌握有效授权的基本原则和方法",
        "5. 我能够识别不同决策类型的处理方式",
        "6. 我明白讲师角色与执行角色的差异",
        "7. 我了解角色转型中的常见陷阱",
        "8. 我有明确的个人发展改进方向",
        "9. 我知道如何制定90天转型计划",
        "10. 我有信心带领团队实现目标"
    ]

    for i, q in enumerate(questions1):
        row = 7 + i
        ws.cell(row=row, column=1, value=i+1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=q).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        for col in range(3, 8):
            cell = ws.cell(row=row, column=col)
            cell.value = "○"
            cell.alignment = center_align
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = gray_fill
        ws.cell(row=row, column=8).border = thin_border
        ws.cell(row=row, column=9).border = thin_border
        ws.cell(row=row, column=10).border = thin_border
        ws.cell(row=row, column=11).border = thin_border
        ws.cell(row=row, column=12).border = thin_border
        ws.cell(row=row, column=13).border = thin_border

    # ===== 方法认知检验（单选10题） =====
    start_row = 18
    apply_title_row(ws, start_row, "三、方法认知检验（单选）", 1, 13)
    ws.row_dimensions[start_row].height = 25

    headers2 = ["题号", "题目内容", "A", "B", "C", "D", "", "", "", "", "", "", "答案"]
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=start_row+1, column=col)
        cell.value = header
        apply_header_style(cell)

    questions2 = [
        "1. 管理者与领导者的核心区别在于？",
        "2. \"做事\"阶段主要关注什么？",
        "3. 时间分配中\"授权陷阱\"指？",
        "4. 有效授权的第一步是？",
        "5. 决策类型不包括以下哪项？",
        "6. 角色转型的最大挑战是？",
        "7. 领导力风格可以分为几种类型？",
        "8. 授权不足的典型表现是？",
        "9. 时间管理矩阵中\"第二象限\"是？",
        "10. 角色转型的\"等死区\"指？"
    ]
    options2 = [
        ["管人", "管事", "管心", "管流程"],
        ["团队建设", "任务完成", "人员培养", "战略规划"],
        ["不敢授权", "过度授权", "只做不授权", "授权后撤回"],
        ["确定授权对象", "明确任务目标", "建立监督机制", "选择授权方式"],
        ["日常型", "战略型", "激进型", "应急型"],
        ["思维转变", "技能提升", "心态调整", "人际关系"],
        ["2种", "3种", "4种", "5种"],
        ["事必躬亲", "民主讨论", "有效分工", "结果导向"],
        ["重要且紧急", "重要不紧急", "紧急不重要", "不重要不紧急"],
        ["转型太早", "转型太晚", "不转型", "盲目转型"]
    ]

    for i, (q, opts) in enumerate(zip(questions2, options2)):
        row = start_row + 2 + i
        ws.cell(row=row, column=1, value=i+1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=q).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        for col, opt in enumerate(opts, 3):
            cell = ws.cell(row=row, column=col)
            cell.value = opt
            cell.alignment = center_align
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = gray_fill
        for col in range(7, 13):
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=13).border = thin_border

    # ===== 多选题（3题） =====
    start_row2 = start_row + 13
    apply_title_row(ws, start_row2, "四、多选题（可选择多项）", 1, 13)

    headers3 = ["题号", "题目内容", "选项区域", "", "", "", "", "", "", "", "", "", "回答"]
    for col, header in enumerate(headers3, 1):
        cell = ws.cell(row=start_row2+1, column=col)
        cell.value = header
        apply_header_style(cell)
    ws.merge_cells(start_row=start_row2+1, start_column=3, end_row=start_row2+1, end_column=12)

    questions3 = [
        "1. 角色转型中常见陷阱包括？（可多选）",
        "2. 作为领导者应该具备的核心能力包括？（可多选）",
        "3. 有效的授权应该包括哪些要素？（可多选）"
    ]
    options3 = [
        ["A. 事必躬亲", "B. 过度授权", "C. 角色模糊", "D. 思维固化"],
        ["A. 战略思维", "B. 沟通能力", "C. 技术能力", "D. 变革管理"],
        ["A. 明确目标", "B. 选对人员", "C. 过程监督", "D. 结果评估"]
    ]

    for i, (q, opts) in enumerate(zip(questions3, options3)):
        row = start_row2 + 2 + i
        ws.cell(row=row, column=1, value=i+1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=q).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        opt_text = "  |  ".join(opts)
        ws.cell(row=row, column=3, value=opt_text).alignment = left_align
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        for col in range(3, 14):
            ws.cell(row=row, column=col).border = thin_border
            if i % 2 == 0 and col != 13:
                ws.cell(row=row, column=col).fill = gray_fill
        ws.cell(row=row, column=13).border = thin_border

    # ===== 情境简答（4题） =====
    start_row3 = start_row2 + 6
    apply_title_row(ws, start_row3, "五、情境简答", 1, 13)

    questions4 = [
        "1. 请描述一个你在角色转型中遇到的最大挑战，以及你是如何应对的？",
        "2. 如果你的下属在授权任务中出现失误，你会如何处理？",
        "3. 请举例说明你是如何平衡\"重要事\"和\"紧急事\"的时间分配的？",
        "4. 你认为从管理者到领导者转变过程中，最需要提升的是哪方面的能力？为什么？"
    ]

    for i, q in enumerate(questions4):
        row = start_row3 + 1 + i
        ws.cell(row=row, column=1, value=i+1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=q).alignment = left_align
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        for col in range(2, 14):
            ws.cell(row=row, column=col).border = thin_border
        # 合并单元格用于填写答案
        ws.merge_cells(start_row=row, start_column=3, end_row=row+2, end_column=12)
        for r in range(row, row+3):
            ws.cell(row=r, column=2).border = thin_border
        row += 2

    # ===== 信心自评（5题） =====
    start_row4 = start_row3 + 12
    apply_title_row(ws, start_row4, "六、信心自评（5分制：1=完全没信心，5=完全有信心）", 1, 13)
    ws.row_dimensions[start_row4].height = 25

    headers4 = ["题号", "题目内容", "1分", "2分", "3分", "4分", "5分", "", "", "", "", "", "备注"]
    for col, header in enumerate(headers4, 1):
        cell = ws.cell(row=start_row4+1, column=col)
        cell.value = header
        apply_header_style(cell)

    questions5 = [
        "1. 你对完成从管理者到领导者角色转型的信心如何？",
        "2. 你对改善时间分配、合理授权的信心如何？",
        "3. 你对提升领导力、影响团队的信心如何？",
        "4. 你对运用领导力工具解决实际问题的信心如何？",
        "5. 你对在90天内看到明显进步的信心如何？"
    ]

    for i, q in enumerate(questions5):
        row = start_row4 + 2 + i
        ws.cell(row=row, column=1, value=i+1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=q).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        for col in range(3, 8):
            cell = ws.cell(row=row, column=col)
            cell.value = "○"
            cell.alignment = center_align
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = gray_fill
        for col in range(8, 13):
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=13).border = thin_border

    # 冻结首行
    ws.freeze_panes = 'A2'


def create_sheet2_posttest(wb):
    """创建后测问卷工作表"""
    ws = wb.create_sheet("后测问卷")

    set_column_widths(ws, [15, 25, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 15])

    # ===== 标题区域 =====
    apply_title_row(ws, 1, "领导力基础课程 - 后测问卷", 1, 13)
    ws.row_dimensions[1].height = 35

    # ===== 基本信息区域 =====
    apply_title_row(ws, 2, "一、基本信息", 1, 13)

    ws.cell(row=3, column=1, value="姓名：").font = subheader_font
    ws.cell(row=3, column=1).border = thin_border
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.cell(row=3, column=3).border = thin_border

    ws.cell(row=3, column=4, value="部门：").font = subheader_font
    ws.cell(row=3, column=4).border = thin_border
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)
    ws.cell(row=3, column=6).border = thin_border

    ws.cell(row=3, column=7, value="岗位：").font = subheader_font
    ws.cell(row=3, column=7).border = thin_border
    ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=8)
    ws.cell(row=3, column=9).border = thin_border

    ws.cell(row=3, column=10, value="工龄：").font = subheader_font
    ws.cell(row=3, column=10).border = thin_border
    ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column=11)
    ws.cell(row=3, column=12).border = thin_border

    ws.cell(row=3, column=13, value="日期：").font = subheader_font
    ws.cell(row=3, column=13).border = thin_border

    # 标注与前测对比的说明
    ws.cell(row=4, column=1, value="（标注★的题目与前测相同，便于对比分析）").font = Font(name='微软雅黑', size=9, italic=True, color=GRAY)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)

    # ===== 教学思维自评（标注★为前测题目） =====
    apply_title_row(ws, 6, "二、教学思维自评（5分制：1=完全不符合，5=完全符合）★前测题目", 1, 13)
    ws.row_dimensions[6].height = 25

    headers = ["题号", "题目内容", "1分", "2分", "3分", "4分", "5分", "", "", "", "", "", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        apply_header_style(cell)

    questions1 = [
        "★1. 我清楚管理者与领导者的核心区别",
        "★2. 我理解从\"做事\"到\"做人\"的转型重要性",
        "★3. 我意识到时间分配需要从紧急事务转向重要事务",
        "★4. 我掌握有效授权的基本原则和方法",
        "★5. 我能够识别不同决策类型的处理方式",
        "★6. 我明白讲师角色与执行角色的差异",
        "★7. 我了解角色转型中的常见陷阱",
        "★8. 我有明确的个人发展改进方向",
        "★9. 我知道如何制定90天转型计划",
        "★10. 我有信心带领团队实现目标"
    ]

    for i, q in enumerate(questions1):
        row = 8 + i
        ws.cell(row=row, column=1, value=i+1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=q).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        for col in range(3, 8):
            cell = ws.cell(row=row, column=col)
            cell.value = "○"
            cell.alignment = center_align
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = gray_fill
        for col in range(8, 13):
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=13).border = thin_border

    # 新增后测题目
    apply_title_row(ws, 19, "三、后测新增题目", 1, 13)
    ws.row_dimensions[19].height = 25

    new_questions = [
        "11. 通过本次培训，我对角色转型的理解更加深刻",
        "12. 我已经制定了明确的90天改进计划",
        "13. 我掌握了至少3种实用的领导力工具"
    ]

    for i, q in enumerate(new_questions):
        row = 20 + i
        ws.cell(row=row, column=1, value=i+11).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=q).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        for col in range(3, 8):
            cell = ws.cell(row=row, column=col)
            cell.value = "○"
            cell.alignment = center_align
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = gray_fill
        for col in range(8, 14):
            ws.cell(row=row, column=col).border = thin_border

    # ===== 方法认知检验 =====
    apply_title_row(ws, 24, "四、方法认知检验（单选）★前测题目", 1, 13)

    headers2 = ["题号", "题目内容", "A", "B", "C", "D", "", "", "", "", "", "", "答案"]
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=25, column=col)
        cell.value = header
        apply_header_style(cell)

    questions2 = [
        "★1. 管理者与领导者的核心区别在于？",
        "★2. \"做事\"阶段主要关注什么？",
        "★3. 时间分配中\"授权陷阱\"指？",
        "★4. 有效授权的第一步是？",
        "★5. 决策类型不包括以下哪项？",
        "★6. 角色转型的最大挑战是？",
        "★7. 领导力风格可以分为几种类型？",
        "★8. 授权不足的典型表现是？",
        "★9. 时间管理矩阵中\"第二象限\"是？",
        "★10. 角色转型的\"等死区\"指？"
    ]
    options2 = [
        ["管人", "管事", "管心", "管流程"],
        ["团队建设", "任务完成", "人员培养", "战略规划"],
        ["不敢授权", "过度授权", "只做不授权", "授权后撤回"],
        ["确定授权对象", "明确任务目标", "建立监督机制", "选择授权方式"],
        ["日常型", "战略型", "激进型", "应急型"],
        ["思维转变", "技能提升", "心态调整", "人际关系"],
        ["2种", "3种", "4种", "5种"],
        ["事必躬亲", "民主讨论", "有效分工", "结果导向"],
        ["重要且紧急", "重要不紧急", "紧急不重要", "不重要不紧急"],
        ["转型太早", "转型太晚", "不转型", "盲目转型"]
    ]

    for i, (q, opts) in enumerate(zip(questions2, options2)):
        row = 26 + i
        ws.cell(row=row, column=1, value=i+1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=q).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        for col, opt in enumerate(opts, 3):
            cell = ws.cell(row=row, column=col)
            cell.value = opt
            cell.alignment = center_align
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = gray_fill
        for col in range(7, 13):
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=13).border = thin_border

    # ===== 培训效果评估 =====
    apply_title_row(ws, 37, "五、培训效果评估（新增）", 1, 13)

    headers5 = ["题号", "评估维度", "非常不满意", "不满意", "一般", "满意", "非常满意", "", "", "", "", "", "得分"]
    for col, header in enumerate(headers5, 1):
        cell = ws.cell(row=38, column=col)
        cell.value = header
        apply_header_style(cell)

    effect_questions = [
        "1. 课程内容的实用性",
        "2. 讲师的授课质量",
        "3. 案例分析的针对性",
        "4. 工具方法的掌握程度",
        "5. 整体培训效果"
    ]

    for i, q in enumerate(effect_questions):
        row = 39 + i
        ws.cell(row=row, column=1, value=i+1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=q).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        for col in range(3, 8):
            cell = ws.cell(row=row, column=col)
            cell.value = "○"
            cell.alignment = center_align
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = gray_fill
        for col in range(8, 13):
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=13).border = thin_border

    # ===== 行为改变承诺 =====
    apply_title_row(ws, 45, "六、行为改变承诺", 1, 13)

    ws.cell(row=46, column=1, value="我承诺在培训后30天内做到：").font = subheader_font
    ws.merge_cells(start_row=46, start_column=1, end_row=46, end_column=8)
    for col in range(1, 14):
        ws.cell(row=46, column=col).border = thin_border

    ws.cell(row=47, column=1, value="1.")
    ws.merge_cells(start_row=47, start_column=1, end_row=47, end_column=12)
    for col in range(1, 14):
        ws.cell(row=47, column=col).border = thin_border
    ws.merge_cells(start_row=47, start_column=2, end_row=48, end_column=12)

    ws.cell(row=49, column=1, value="2.")
    ws.merge_cells(start_row=49, start_column=1, end_row=49, end_column=12)
    for col in range(1, 14):
        ws.cell(row=49, column=col).border = thin_border
    ws.merge_cells(start_row=49, start_column=2, end_row=50, end_column=12)

    ws.cell(row=51, column=1, value="3.")
    ws.merge_cells(start_row=51, start_column=1, end_row=51, end_column=12)
    for col in range(1, 14):
        ws.cell(row=51, column=col).border = thin_border
    ws.merge_cells(start_row=51, start_column=2, end_row=52, end_column=12)

    ws.freeze_panes = 'A2'


def create_sheet3_observation(wb):
    """创建行为观察量表"""
    ws = wb.create_sheet("行为观察量表")

    set_column_widths(ws, [15, 35, 10, 10, 10, 10, 10, 25])

    # ===== 标题 =====
    apply_title_row(ws, 1, "领导力基础课程 - 行为观察量表", 1, 8)
    ws.row_dimensions[1].height = 35

    # ===== 基本信息 =====
    apply_title_row(ws, 2, "一、观察记录基本信息", 1, 8)

    info_items = [
        ("学员姓名：", 1, 3), ("观察日期：", 4, 6), ("观察讲师：", 7, 8)
    ]
    for text, start, end in info_items:
        ws.cell(row=3, column=start, value=text).font = subheader_font
        ws.cell(row=3, column=start).border = thin_border
        ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)

    # ===== 讲师观察评分表 =====
    apply_title_row(ws, 5, "二、讲师观察评分（5分制：1=从未观察到，5=始终保持）", 1, 8)
    ws.row_dimensions[5].height = 25

    headers = ["维度", "观察要点", "1分", "2分", "3分", "4分", "5分", "行为记录"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        apply_header_style(cell)

    # 观察维度
    dimensions = [
        {
            "name": "领导力风格识别",
            "items": [
                "能够识别自己的领导力风格倾向",
                "能够在日常管理中展现合适的领导风格",
                "能够根据情境调整领导风格"
            ]
        },
        {
            "name": "角色转型认知",
            "items": [
                "清楚管理者与领导者的区别",
                "理解从\"做事\"到\"做人\"的转变",
                "有明确的角色定位和职责边界"
            ]
        },
        {
            "name": "时间分配",
            "items": [
                "优先处理重要但不紧急的事务",
                "合理分配战略思考时间",
                "避免陷入\"救火\"模式"
            ]
        },
        {
            "name": "授权边界",
            "items": [
                "敢于授权并善于授权",
                "明确授权的范围和边界",
                "能够对授权任务进行有效监督"
            ]
        },
        {
            "name": "决策方式",
            "items": [
                "能够识别不同类型的决策",
                "运用合适的决策框架",
                "在授权与亲自决策间找到平衡"
            ]
        }
    ]

    current_row = 7
    for dim in dimensions:
        # 维度名称行
        ws.cell(row=current_row, column=1, value=dim["name"]).font = Font(name='微软雅黑', size=10, bold=True, color=WHITE)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = thin_border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        for col in range(3, 8):
            ws.cell(row=current_row, column=col).border = thin_border
        ws.cell(row=current_row, column=8).border = thin_border
        current_row += 1

        # 观察要点行
        for item in dim["items"]:
            ws.cell(row=current_row, column=1, value="").border = thin_border
            ws.cell(row=current_row, column=2, value=item).alignment = left_align
            ws.cell(row=current_row, column=2).border = thin_border
            for col in range(3, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.value = "○"
                cell.alignment = center_align
                cell.border = thin_border
            ws.cell(row=current_row, column=8, value="").border = thin_border
            current_row += 1

    # ===== 综合评价区域 =====
    current_row += 1
    apply_title_row(ws, current_row, "三、综合评价", 1, 8)
    ws.row_dimensions[current_row].height = 25
    current_row += 1

    ws.cell(row=current_row, column=1, value="整体表现评价：").font = subheader_font
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    for col in range(1, 9):
        ws.cell(row=current_row, column=col).border = thin_border
    current_row += 1

    ws.cell(row=current_row, column=1, value="优势领域：")
    ws.cell(row=current_row, column=1).border = thin_border
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+2, end_column=2)
    for col in range(1, 3):
        for r in range(current_row, current_row+3):
            ws.cell(row=r, column=col).border = thin_border

    ws.cell(row=current_row, column=3, value="")
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row+2, end_column=8)
    for col in range(3, 9):
        for r in range(current_row, current_row+3):
            ws.cell(row=r, column=col).border = thin_border
    current_row += 3

    ws.cell(row=current_row, column=1, value="待改进领域：")
    ws.cell(row=current_row, column=1).border = thin_border
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+2, end_column=2)
    for col in range(1, 3):
        for r in range(current_row, current_row+3):
            ws.cell(row=r, column=col).border = thin_border

    ws.cell(row=current_row, column=3, value="")
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row+2, end_column=8)
    for col in range(3, 9):
        for r in range(current_row, current_row+3):
            ws.cell(row=r, column=col).border = thin_border
    current_row += 3

    ws.cell(row=current_row, column=1, value="改进建议：")
    ws.cell(row=current_row, column=1).border = thin_border
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+2, end_column=2)
    for col in range(1, 3):
        for r in range(current_row, current_row+3):
            ws.cell(row=r, column=col).border = thin_border

    ws.cell(row=current_row, column=3, value="")
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row+2, end_column=8)
    for col in range(3, 9):
        for r in range(current_row, current_row+3):
            ws.cell(row=r, column=col).border = thin_border

    ws.freeze_panes = 'A2'


def create_sheet4_analysis(wb):
    """创建数据分析工作表"""
    ws = wb.create_sheet("数据分析")

    set_column_widths(ws, [12, 20, 12, 12, 12, 12, 12, 12, 15, 15])

    # ===== 标题 =====
    apply_title_row(ws, 1, "领导力基础课程 - 数据分析", 1, 10)
    ws.row_dimensions[1].height = 35

    # ===== 前测后测对比表 =====
    apply_title_row(ws, 3, "一、学员前测后测对比表", 1, 10)

    headers = ["学号", "姓名", "前测总分", "后测总分", "提升分", "提升率", "前测排名", "后测排名", "排名变化", "达成情况"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        apply_header_style(cell)

    # 示例数据行（灰色填充表示示例）
    sample_data = [
        ["001", "学员A", 35, 42, "=D5-C5", "=(E5/C5)*100%", "", "", "", ""],
        ["002", "学员B", 38, 45, "=D6-C6", "=(E6/C6)*100%", "", "", "", ""],
        ["003", "学员C", 32, 40, "=D7-C7", "=(E7/C7)*100%", "", "", "", ""],
    ]

    for i, row_data in enumerate(sample_data):
        row = 5 + i
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_align if col != 2 else left_align
            cell.fill = gray_fill

    # 添加更多空行用于填写
    for row in range(8, 28):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_align if col != 2 else left_align
            if row % 2 == 0:
                cell.fill = gray_fill

    # ===== 班级平均分统计 =====
    apply_title_row(ws, 30, "二、班级平均分统计", 1, 10)
    ws.row_dimensions[30].height = 25

    stat_headers = ["统计维度", "前测平均", "后测平均", "提升幅度", "达标率", "优秀率", "", "", "", ""]
    for col, header in enumerate(stat_headers, 1):
        cell = ws.cell(row=31, column=col)
        cell.value = header
        apply_header_style(cell)

    stat_items = [
        ["教学思维自评", "=AVERAGE(C5:C27)", "=AVERAGE(D5:D27)", "=C32-B32", "=COUNTIF(F5:F27,\">=20%\")/COUNT(C5:C27)", ""],
        ["方法认知检验", "", "", "", "", ""],
        ["多选题", "", "", "", "", ""],
        ["情境简答", "", "", "", "", ""],
        ["信心自评", "", "", "", "", ""],
        ["综合得分", "=AVERAGE(C5:C27)", "=AVERAGE(D5:D27)", "=C33-B33", "", ""]
    ]

    for i, row_data in enumerate(stat_items):
        row = 32 + i
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_align if col != 1 else left_align
            if row % 2 == 0:
                cell.fill = gray_fill

    # ===== 达成率计算 =====
    apply_title_row(ws, 39, "三、达成率计算", 1, 10)
    ws.row_dimensions[39].height = 25

    ach_headers = ["指标", "目标值", "实际值", "达成率", "状态", "", "", "", "", ""]
    for col, header in enumerate(ach_headers, 1):
        cell = ws.cell(row=40, column=col)
        cell.value = header
        apply_header_style(cell)

    ach_items = [
        ["前测转后测提升率", "20%", "=AVERAGE(F5:F27)", "=(C41/B41)*100%", "=IF(D41>=1,\"达标\",\"未达标\")"],
        ["知识掌握率", "80%", "", "", ""],
        ["行为改变率", "70%", "", "", ""],
        ["总体满意度", "90%", "", "", ""]
    ]

    for i, row_data in enumerate(ach_items):
        row = 41 + i
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_align if col != 1 else left_align
            if row % 2 == 0:
                cell.fill = gray_fill

    # ===== 图表展示区域 ===== (预留位置)
    apply_title_row(ws, 46, "四、图表展示区域（可粘贴图表）", 1, 10)
    ws.row_dimensions[46].height = 25

    for row in range(47, 57):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = gray_fill

    ws.freeze_panes = 'A2'


def create_sheet5_report(wb):
    """创建评估报告模板"""
    ws = wb.create_sheet("评估报告模板")

    set_column_widths(ws, [15, 30, 15, 15, 15, 15])

    # ===== 标题 =====
    apply_title_row(ws, 1, "领导力基础课程 - 评估报告模板", 1, 6)
    ws.row_dimensions[1].height = 35

    # ===== 学员个人报告框架 =====
    apply_title_row(ws, 3, "第一部分：学员个人报告框架", 1, 6)
    ws.row_dimensions[3].height = 30

    # 基本信息
    ws.cell(row=4, column=1, value="学员姓名：").font = subheader_font
    ws.cell(row=4, column=1).border = thin_border
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
    ws.cell(row=4, column=3).border = thin_border

    ws.cell(row=4, column=4, value="所属部门：").font = subheader_font
    ws.cell(row=4, column=4).border = thin_border
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=5)
    ws.cell(row=4, column=6).border = thin_border

    ws.cell(row=5, column=1, value="岗位职级：").font = subheader_font
    ws.cell(row=5, column=1).border = thin_border
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
    ws.cell(row=5, column=3).border = thin_border

    ws.cell(row=5, column=4, value="评估周期：").font = subheader_font
    ws.cell(row=5, column=4).border = thin_border
    ws.merge_cells(start_row=5, start_column=4, end_row=5, end_column=5)
    ws.cell(row=5, column=6).border = thin_border

    # 学习前后对比
    apply_title_row(ws, 7, "一、学习前后能力对比", 1, 6)

    comp_headers = ["能力维度", "学前水平", "学前排名", "学后水平", "学后排名", "提升情况"]
    for col, header in enumerate(comp_headers, 1):
        cell = ws.cell(row=8, column=col)
        cell.value = header
        apply_header_style(cell)

    capabilities = [
        "领导力认知", "角色转型意识", "时间管理", "授权能力", "决策能力", "综合评分"
    ]

    for i, cap in enumerate(capabilities):
        row = 9 + i
        ws.cell(row=row, column=1, value=cap).border = thin_border
        ws.cell(row=row, column=1).alignment = left_align
        for col in range(2, 7):
            ws.cell(row=row, column=col).border = thin_border
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = gray_fill

    # 优势与改进
    apply_title_row(ws, 16, "二、优势领域与待改进领域", 1, 6)

    ws.cell(row=17, column=1, value="核心优势：").font = subheader_font
    ws.cell(row=17, column=1).border = thin_border
    ws.merge_cells(start_row=17, start_column=1, end_row=17, end_column=5)
    for col in range(1, 7):
        ws.cell(row=17, column=col).border = thin_border
    ws.merge_cells(start_row=18, start_column=1, end_row=19, end_column=5)
    for col in range(1, 7):
        for r in range(18, 20):
            ws.cell(row=r, column=col).border = thin_border

    ws.cell(row=20, column=1, value="待改进领域：").font = subheader_font
    ws.cell(row=20, column=1).border = thin_border
    ws.merge_cells(start_row=20, start_column=1, end_row=20, end_column=5)
    for col in range(1, 7):
        ws.cell(row=20, column=col).border = thin_border
    ws.merge_cells(start_row=21, start_column=1, end_row=22, end_column=5)
    for col in range(1, 7):
        for r in range(21, 23):
            ws.cell(row=r, column=col).border = thin_border

    # 改进建议
    apply_title_row(ws, 24, "三、个人改进建议", 1, 6)

    suggestions = [
        ("短期目标（30天）", 25),
        ("中期目标（60天）", 27),
        ("长期目标（90天）", 29)
    ]

    for text, row in suggestions:
        ws.cell(row=row, column=1, value=text).font = subheader_font
        ws.cell(row=row, column=1).border = thin_border
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+2, end_column=5)
        for col in range(1, 7):
            for r in range(row+1, row+3):
                ws.cell(row=r, column=col).border = thin_border

    # ===== 班级整体报告框架 =====
    apply_title_row(ws, 33, "第二部分：班级整体报告框架", 1, 6)
    ws.row_dimensions[33].height = 30

    # 班级整体数据
    apply_title_row(ws, 35, "一、班级整体数据概览", 1, 6)

    overview_headers = ["统计指标", "数值", "说明", "", "", ""]
    for col, header in enumerate(overview_headers, 1):
        cell = ws.cell(row=36, column=col)
        cell.value = header
        apply_header_style(cell)

    overview_items = [
        ["班级总人数", "", ""],
        ["前测平均分", "", ""],
        ["后测平均分", "", ""],
        ["平均提升幅度", "", ""],
        ["目标达成率", "", ""],
        ["整体满意度", "", ""]
    ]

    for i, (item, val, note) in enumerate(overview_items):
        row = 37 + i
        ws.cell(row=row, column=1, value=item).border = thin_border
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=2, value=val).border = thin_border
        ws.cell(row=row, column=3, value=note).border = thin_border
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        if row % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = gray_fill

    # 班级共性问题
    apply_title_row(ws, 44, "二、班级共性问题分析", 1, 6)

    ws.cell(row=45, column=1, value="共性问题：").font = subheader_font
    ws.merge_cells(start_row=45, start_column=1, end_row=45, end_column=5)
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border
    ws.merge_cells(start_row=46, start_column=1, end_row=48, end_column=5)
    for col in range(1, 7):
        for r in range(46, 49):
            ws.cell(row=r, column=col).border = thin_border

    # 培训效果总结
    apply_title_row(ws, 50, "三、培训效果总结", 1, 6)

    summary_items = [
        ("知识掌握", 51),
        ("技能提升", 53),
        ("态度转变", 55),
        ("行为改变", 57)
    ]

    for text, row in summary_items:
        ws.cell(row=row, column=1, value=text).font = subheader_font
        ws.cell(row=row, column=1).border = thin_border
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+2, end_column=5)
        for col in range(1, 7):
            for r in range(row+1, row+3):
                ws.cell(row=r, column=col).border = thin_border

    # 下一步改进建议
    apply_title_row(ws, 60, "四、下一步改进建议", 1, 6)

    ws.cell(row=61, column=1, value="课程优化建议：").font = subheader_font
    ws.merge_cells(start_row=61, start_column=1, end_row=61, end_column=5)
    for col in range(1, 7):
        ws.cell(row=61, column=col).border = thin_border
    ws.merge_cells(start_row=62, start_column=1, end_row=64, end_column=5)
    for col in range(1, 7):
        for r in range(62, 65):
            ws.cell(row=r, column=col).border = thin_border

    ws.cell(row=65, column=1, value="后续跟进计划：").font = subheader_font
    ws.merge_cells(start_row=65, start_column=1, end_row=65, end_column=5)
    for col in range(1, 7):
        ws.cell(row=65, column=col).border = thin_border
    ws.merge_cells(start_row=66, start_column=1, end_row=68, end_column=5)
    for col in range(1, 7):
        for r in range(66, 69):
            ws.cell(row=r, column=col).border = thin_border

    ws.freeze_panes = 'A2'


def main():
    """主函数"""
    # 创建工作簿
    wb = Workbook()

    # 创建各工作表
    print("正在创建前测问卷...")
    create_sheet1_pretest(wb)

    print("正在创建后测问卷...")
    create_sheet2_posttest(wb)

    print("正在创建行为观察量表...")
    create_sheet3_observation(wb)

    print("正在创建数据分析...")
    create_sheet4_analysis(wb)

    print("正在创建评估报告模板...")
    create_sheet5_report(wb)

    # 保存文件
    output_path = "D:/新课开发/管理学/15-领导力基础/E4_课程评估工具.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\n✅ Excel评估工具已成功生成！")
    print(f"📁 文件路径：{output_path}")
    print(f"\n包含工作表：")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")


if __name__ == "__main__":
    main()

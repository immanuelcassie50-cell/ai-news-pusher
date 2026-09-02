#!/usr/bin/env python3
"""Build 行为改变设计 Excel forms using openpyxl."""

import os
import shutil
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

OUT_DIR = r"D:\新课开发\心理学\12-行为改变设计：用环境和触发器重塑自己\配套表单和指引-Excel版"
TEMPLATE_DIR = r"C:\Users\Administrator\.claude\skills\Excel表格处理\templates\minimal_xlsx"

# ─── shared style helpers ────────────────────────────────────────────────────

def bold_header(ws, row, cols, text, bg="D9E8F5"):
    cell = ws.cell(row=row, column=cols, value=text)
    cell.font = Font(bold=True, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell

def header_row(ws, row, headers, bg="D9E8F5"):
    for col, h in enumerate(headers, 1):
        bold_header(ws, row, col, h, bg)

def input_cell(ws, row, col, value="", bg="FFFFFF"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell

def section_title(ws, row, text, span=8, bg="2E75B6"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, name="Calibri", size=12, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell

def label_cell(ws, row, col, text, bg="F2F2F2"):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze_and_filter(ws, row=2, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)

# ─── Sheet builders ────────────────────────────────────────────────────────────

def build_info_sheet(ws):
    """必填信息表"""
    ws.title = "必填信息表"
    section_title(ws, 1, "必填信息表（通用）", span=4)

    info_rows = [
        (3, "学员姓名", ""),
        (4, "课程日期", ""),
        (5, "目标行为", ""),
        (6, "当前状态", ""),
        (7, "期待改变", ""),
    ]
    label_cell(ws, 3, 1, "学员姓名")
    input_cell(ws, 3, 2)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)

    label_cell(ws, 4, 1, "课程日期")
    input_cell(ws, 4, 2)
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=4)

    label_cell(ws, 5, 1, "目标行为")
    input_cell(ws, 5, 2)
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)

    label_cell(ws, 6, 1, "当前状态")
    input_cell(ws, 6, 2, "", bg="FFFDE7")
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=4)

    label_cell(ws, 7, 1, "期待改变")
    input_cell(ws, 7, 2, "", bg="FFFDE7")
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=4)

    set_col_widths(ws, [16, 14, 14, 14])
    ws.row_dimensions[1].height = 30
    for r in range(3, 8):
        ws.row_dimensions[r].height = 22

def build_f1(ws):
    """F1 行为环境设计自测"""
    ws.title = "F1 行为环境设计自测"
    section_title(ws, 1, "F1 行为环境设计自测", span=5)

    headers = ["维度", "提示明显度(1-5)", "渴望绑定度(1-5)", "行动简单度(1-5)", "动机真实度(1-5)"]
    header_row(ws, 2, headers)

    items = [
        "当前环境提示是否足够明显？",
        "新行为是否与已有渴望绑定？",
        "行动步骤是否足够简单？",
        "动机是否来自真实内在需求？",
        "环境是否支持持续执行？",
        "社交环境是否提供正向反馈？",
        "是否有清晰的触发线索？",
    ]
    for i, item in enumerate(items, 3):
        input_cell(ws, i, 1, item, bg="F2F2F2")
        for c in range(2, 6):
            input_cell(ws, i, c)

    set_col_widths(ws, [30, 14, 14, 14, 14])
    ws.row_dimensions[1].height = 30
    for r in range(2, 10):
        ws.row_dimensions[r].height = 22
    freeze_and_filter(ws, 3, 2)

def build_f2(ws):
    """F2 锚点行为分析"""
    ws.title = "F2 锚点行为分析"
    section_title(ws, 1, "F2 锚点行为分析", span=5)

    headers = ["锚点行为", "触发线索", "新行为", "绑定方式", "实施时间"]
    header_row(ws, 2, headers)

    for r in range(3, 13):
        for c in range(1, 6):
            bg = "FFFDE7" if c == 3 else "FFFFFF"
            input_cell(ws, r, c, "", bg=bg)

    set_col_widths(ws, [18, 20, 20, 16, 14])
    ws.row_dimensions[1].height = 30
    for r in range(2, 13):
        ws.row_dimensions[r].height = 22
    freeze_and_filter(ws, 3, 1)

def build_f3(ws):
    """F3 物理环境改造清单"""
    ws.title = "F3 物理环境改造清单"
    section_title(ws, 1, "F3 物理环境改造清单", span=5)

    headers = ["改造项目", "当前位置", "目标位置", "所需物品", "完成状态"]
    header_row(ws, 2, headers)

    for r in range(3, 13):
        input_cell(ws, r, 1, "", bg="FFFDE7")
        for c in range(2, 6):
            input_cell(ws, r, c)

    set_col_widths(ws, [20, 18, 18, 16, 12])
    ws.row_dimensions[1].height = 30
    for r in range(2, 13):
        ws.row_dimensions[r].height = 22
    freeze_and_filter(ws, 3, 1)

def build_f4(ws):
    """F4 社交环境改造清单"""
    ws.title = "F4 社交环境改造清单"
    section_title(ws, 1, "F4 社交环境改造清单", span=5)

    headers = ["改造项目", "当前做法", "目标做法", "涉及人员", "完成状态"]
    header_row(ws, 2, headers)

    for r in range(3, 13):
        input_cell(ws, r, 1, "", bg="FFFDE7")
        for c in range(2, 6):
            input_cell(ws, r, c)

    set_col_widths(ws, [20, 20, 20, 14, 12])
    ws.row_dimensions[1].height = 30
    for r in range(2, 13):
        ws.row_dimensions[r].height = 22
    freeze_and_filter(ws, 3, 1)

def build_f5(ws):
    """F5 环境改造方案工作表"""
    ws.title = "F5 环境改造方案工作表"
    section_title(ws, 1, "F5 环境改造方案工作表", span=6)

    headers = ["序号", "目标行为", "触发条件", "改造策略", "实施步骤", "预期效果"]
    header_row(ws, 2, headers)

    for r in range(3, 13):
        input_cell(ws, r, 1, r - 2, bg="F2F2F2")
        for c in range(2, 7):
            bg = "FFFDE7" if c in (2, 4) else "FFFFFF"
            input_cell(ws, r, c, "", bg=bg)

    set_col_widths(ws, [6, 16, 18, 18, 22, 18])
    ws.row_dimensions[1].height = 30
    for r in range(2, 13):
        ws.row_dimensions[r].height = 22
    freeze_and_filter(ws, 3, 2)

def build_f6(ws):
    """F6 实施追踪表"""
    ws.title = "F6 实施追踪表"
    section_title(ws, 1, "F6 实施追踪表", span=6)

    headers = ["日期", "目标行为", "实施情况", "障碍", "调整措施", "完成度(%)"]
    header_row(ws, 2, headers)

    for r in range(3, 18):
        input_cell(ws, r, 1, "", bg="FFFDE7")
        input_cell(ws, r, 2, "")
        input_cell(ws, r, 3, "")
        input_cell(ws, r, 4, "")
        input_cell(ws, r, 5, "")
        input_cell(ws, r, 6, 0)

    set_col_widths(ws, [14, 16, 22, 20, 22, 10])
    ws.row_dimensions[1].height = 30
    for r in range(2, 18):
        ws.row_dimensions[r].height = 22
    freeze_and_filter(ws, 3, 1)

def build_f7(ws):
    """F7 习惯固化检查卡"""
    ws.title = "F7 习惯固化检查卡"
    section_title(ws, 1, "F7 习惯固化检查卡", span=4)

    headers = ["检查项目", "状态(是/否)", "备注", "下一步"]
    header_row(ws, 2, headers)

    items = [
        "行为已持续执行7天以上？",
        "触发线索已稳定形成？",
        "新行为已无需意志力维持？",
        "环境提示已日常化？",
        "完成行为后有正向感受？",
        "周边人员知晓并支持此改变？",
        "行为效果已开始显现？",
    ]
    for i, item in enumerate(items, 3):
        input_cell(ws, i, 1, item, bg="F2F2F2")
        input_cell(ws, i, 2, "")
        input_cell(ws, i, 3, "")
        input_cell(ws, i, 4, "")

    set_col_widths(ws, [30, 14, 24, 24])
    ws.row_dimensions[1].height = 30
    for r in range(2, 10):
        ws.row_dimensions[r].height = 22
    freeze_and_filter(ws, 3, 2)

def build_f8(ws):
    """F8 进度庆祝记录卡"""
    ws.title = "F8 进度庆祝记录卡"
    section_title(ws, 1, "F8 进度庆祝记录卡", span=5)

    headers = ["日期", "里程碑成就", "庆祝方式", "感受", "继续动力"]
    header_row(ws, 2, headers)

    for r in range(3, 13):
        for c in range(1, 6):
            bg = "FFFDE7" if c == 2 else "FFFFFF"
            input_cell(ws, r, c, "", bg=bg)

    set_col_widths(ws, [14, 22, 22, 24, 16])
    ws.row_dimensions[1].height = 30
    for r in range(2, 13):
        ws.row_dimensions[r].height = 22
    freeze_and_filter(ws, 3, 1)

def build_f9(ws):
    """F9 我的行为改变报告"""
    ws.title = "F9 我的行为改变报告"
    section_title(ws, 1, "F9 我的行为改变报告", span=5)

    headers = ["维度", "目标", "实际进展", "差距原因", "改进方向"]
    header_row(ws, 2, headers)

    items = [
        ("提示设计", "环境提示明显稳定"),
        ("渴望绑定", "新行为与内在渴望链接"),
        ("行动简化", "步骤最少化、阻力为零"),
        ("物理环境", "物品摆放支持新行为"),
        ("社交环境", "获得他人支持认可"),
        ("动机真实", "来自真实内在需求"),
        ("总体评价", "行为已固化形成习惯"),
    ]
    for i, (dim, _) in enumerate(items, 3):
        input_cell(ws, i, 1, dim, bg="F2F2F2")
        input_cell(ws, i, 2, "")
        input_cell(ws, i, 3, "")
        input_cell(ws, i, 4, "")
        input_cell(ws, i, 5, "")

    set_col_widths(ws, [14, 20, 22, 22, 20])
    ws.row_dimensions[1].height = 30
    for r in range(2, 10):
        ws.row_dimensions[r].height = 22
    freeze_and_filter(ws, 3, 2)

def build_f10(ws):
    """F10 团队行为改变台账"""
    ws.title = "F10 团队行为改变台账"
    section_title(ws, 1, "F10 团队行为改变台账", span=5)

    headers = ["成员", "目标行为", "当前状态", "支持需求", "下次检查日期"]
    header_row(ws, 2, headers)

    for r in range(3, 13):
        for c in range(1, 6):
            bg = "FFFDE7" if c == 2 else "FFFFFF"
            input_cell(ws, r, c, "", bg=bg)

    set_col_widths(ws, [12, 20, 18, 20, 14])
    ws.row_dimensions[1].height = 30
    for r in range(2, 13):
        ws.row_dimensions[r].height = 22
    freeze_and_filter(ws, 3, 1)

# ─── Filled version sample data ──────────────────────────────────────────────

def build_filled_sheet(ws, sheet_name, headers, rows_data):
    """Generic filled sheet with headers and sample rows."""
    section_title(ws, 1, sheet_name, span=len(headers))
    header_row(ws, 2, headers)
    for r, row_data in enumerate(rows_data, 3):
        for c, val in enumerate(row_data, 1):
            input_cell(ws, r, c, val)
    return ws

def build_info_sheet_filled(ws):
    ws.title = "必填信息表"
    section_title(ws, 1, "必填信息表（通用）", span=4)
    label_cell(ws, 3, 1, "学员姓名")
    input_cell(ws, 3, 2, "李明")
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
    label_cell(ws, 4, 1, "课程日期")
    input_cell(ws, 4, 2, "2026-08-20")
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=4)
    label_cell(ws, 5, 1, "目标行为")
    input_cell(ws, 5, 2, "每天阅读30分钟")
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)
    label_cell(ws, 6, 1, "当前状态")
    input_cell(ws, 6, 2, "偶尔阅读，缺乏规律")
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=4)
    label_cell(ws, 7, 1, "期待改变")
    input_cell(ws, 7, 2, "养成每日阅读习惯，形成稳定的知识积累")
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=4)
    set_col_widths(ws, [16, 14, 14, 14])
    for r in range(3, 8):
        ws.row_dimensions[r].height = 22

def build_f1_filled(ws):
    ws.title = "F1 行为环境设计自测"
    section_title(ws, 1, "F1 行为环境设计自测", span=5)
    headers = ["维度", "提示明显度(1-5)", "渴望绑定度(1-5)", "行动简单度(1-5)", "动机真实度(1-5)"]
    header_row(ws, 2, headers)
    items = [
        ("当前环境提示是否足够明显？", 3, 4, 5, 4),
        ("新行为是否与已有渴望绑定？", 4, 5, 4, 5),
        ("行动步骤是否足够简单？", 5, 4, 5, 4),
        ("动机是否来自真实内在需求？", 4, 5, 4, 5),
        ("环境是否支持持续执行？", 3, 4, 5, 4),
        ("社交环境是否提供正向反馈？", 2, 3, 4, 3),
        ("是否有清晰的触发线索？", 4, 5, 5, 4),
    ]
    for i, (item, *scores) in enumerate(items, 3):
        input_cell(ws, i, 1, item, bg="F2F2F2")
        for c, v in enumerate(scores, 2):
            input_cell(ws, i, c, v)
    set_col_widths(ws, [30, 14, 14, 14, 14])
    freeze_and_filter(ws, 3, 2)

def build_f2_filled(ws):
    ws.title = "F2 锚点行为分析"
    section_title(ws, 1, "F2 锚点行为分析", span=5)
    headers = ["锚点行为", "触发线索", "新行为", "绑定方式", "实施时间"]
    header_row(ws, 2, headers)
    data = [
        ("早起刷牙后", "牙刷放枕边", "拿起书阅读", "绑定到已有晨间习惯", "早上7:00"),
        ("午休结束后", "手机闹钟关闭", "站立伸展5分钟", "绑定到午休结束信号", "中午13:00"),
        ("晚饭后", "洗碗完毕", "写当日反思", "绑定到家务完成", "晚上20:00"),
    ]
    for i, row in enumerate(data, 3):
        for c, v in enumerate(row, 1):
            input_cell(ws, i, c, v)
    for r in range(3 + len(data), 13):
        for c in range(1, 6):
            input_cell(ws, r, c, "", bg="FFFDE7" if c == 3 else "FFFFFF")
    set_col_widths(ws, [18, 20, 20, 16, 14])
    freeze_and_filter(ws, 3, 1)

def build_f3_filled(ws):
    ws.title = "F3 物理环境改造清单"
    section_title(ws, 1, "F3 物理环境改造清单", span=5)
    headers = ["改造项目", "当前位置", "目标位置", "所需物品", "完成状态"]
    header_row(ws, 2, headers)
    data = [
        ("书籍放置", "书架高层", "床头柜", "书架、收纳盒", "已完成"),
        ("阅读角", "客厅沙发", "卧室窗边", "台灯、坐垫", "进行中"),
        ("手机隔离", "床头充电", "客厅充电站", "充电盒", "已完成"),
    ]
    for i, row in enumerate(data, 3):
        for c, v in enumerate(row, 1):
            input_cell(ws, i, c, v)
    for r in range(3 + len(data), 13):
        input_cell(ws, r, 1, "", bg="FFFDE7")
        for c in range(2, 6):
            input_cell(ws, r, c)
    set_col_widths(ws, [20, 18, 18, 16, 12])
    freeze_and_filter(ws, 3, 1)

def build_f4_filled(ws):
    ws.title = "F4 社交环境改造清单"
    section_title(ws, 1, "F4 社交环境改造清单", span=5)
    headers = ["改造项目", "当前做法", "目标做法", "涉及人员", "完成状态"]
    header_row(ws, 2, headers)
    data = [
        ("寻找阅读伙伴", "独自阅读", "加入读书会", "同事小王", "进行中"),
        ("家庭支持", "家人不知阅读计划", "分享每日收获", "配偶、孩子", "已完成"),
    ]
    for i, row in enumerate(data, 3):
        for c, v in enumerate(row, 1):
            input_cell(ws, i, c, v)
    for r in range(3 + len(data), 13):
        input_cell(ws, r, 1, "", bg="FFFDE7")
        for c in range(2, 6):
            input_cell(ws, r, c)
    set_col_widths(ws, [20, 20, 20, 14, 12])
    freeze_and_filter(ws, 3, 1)

def build_f5_filled(ws):
    ws.title = "F5 环境改造方案工作表"
    section_title(ws, 1, "F5 环境改造方案工作表", span=6)
    headers = ["序号", "目标行为", "触发条件", "改造策略", "实施步骤", "预期效果"]
    header_row(ws, 2, headers)
    data = [
        (1, "晨起阅读30分钟", "闹钟响起", "将书放在枕边", "牙刷后坐在床边翻开书", "开启愉悦学习"),
        (2, "午间站立伸展", "午休闹钟关闭", "手机放客厅", "伸展动作设为屏保", "缓解久坐疲劳"),
        (3, "晚间写作反思", "晚饭洗碗完毕", "写作本放在厨房", "写3条今日收获", "提升觉察力"),
    ]
    for i, row in enumerate(data, 3):
        for c, v in enumerate(row, 1):
            bg = "FFFDE7" if c in (2, 4) else "FFFFFF"
            input_cell(ws, i, c, v, bg=bg)
    for r in range(3 + len(data), 13):
        input_cell(ws, r, 1, r - 2, bg="F2F2F2")
        for c in range(2, 7):
            input_cell(ws, r, c, "", bg="FFFDE7" if c in (2, 4) else "FFFFFF")
    set_col_widths(ws, [6, 16, 18, 18, 22, 18])
    freeze_and_filter(ws, 3, 2)

def build_f6_filled(ws):
    ws.title = "F6 实施追踪表"
    section_title(ws, 1, "F6 实施追踪表", span=6)
    headers = ["日期", "目标行为", "实施情况", "障碍", "调整措施", "完成度(%)"]
    header_row(ws, 2, headers)
    data = [
        ("2026-08-14", "晨起阅读", "完成", "起床晚了10分钟", "将闹钟提前5分钟", 100),
        ("2026-08-15", "晨起阅读", "完成", "无", "继续执行", 100),
        ("2026-08-16", "晨起阅读", "部分完成", "临时会议打断", "改用午休前补读", 60),
        ("2026-08-17", "晨起阅读", "完成", "无", "增加5分钟复盘", 100),
        ("2026-08-18", "晨起阅读", "完成", "无", "记录读后感", 100),
        ("2026-08-19", "晨起阅读", "完成", "无", "周末保持同样节奏", 100),
        ("2026-08-20", "晨起阅读", "完成", "无", "已坚持一周", 100),
    ]
    for i, row in enumerate(data, 3):
        for c, v in enumerate(row, 1):
            input_cell(ws, i, c, v, bg="FFFDE7" if c == 1 else "FFFFFF")
    for r in range(3 + len(data), 18):
        input_cell(ws, r, 1, "", bg="FFFDE7")
        for c in range(2, 7):
            input_cell(ws, r, c)
    set_col_widths(ws, [14, 16, 22, 20, 22, 10])
    freeze_and_filter(ws, 3, 1)

def build_f7_filled(ws):
    ws.title = "F7 习惯固化检查卡"
    section_title(ws, 1, "F7 习惯固化检查卡", span=4)
    headers = ["检查项目", "状态(是/否)", "备注", "下一步"]
    header_row(ws, 2, headers)
    items = [
        ("行为已持续执行7天以上？", "是", "已执行8天", "继续保持"),
        ("触发线索已稳定形成？", "是", "闹钟+刷牙成为固定链条", "可尝试减少外部提示"),
        ("新行为已无需意志力维持？", "否", "周末仍需刻意提醒", "周末安排家人提醒"),
        ("环境提示已日常化？", "是", "书在枕边已成习惯", "下一步移除辅助提示"),
        ("完成行为后有正向感受？", "是", "读完后有成就感", "记录每次感受"),
        ("周边人员知晓并支持此改变？", "是", "家人知道我的计划", "每周分享一次收获"),
        ("行为效果已开始显现？", "是", "本周读完一本书", "设定下一本目标"),
    ]
    for i, (item, status, remark, nxt) in enumerate(items, 3):
        input_cell(ws, i, 1, item, bg="F2F2F2")
        input_cell(ws, i, 2, status)
        input_cell(ws, i, 3, remark)
        input_cell(ws, i, 4, nxt)
    set_col_widths(ws, [30, 14, 24, 24])
    freeze_and_filter(ws, 3, 2)

def build_f8_filled(ws):
    ws.title = "F8 进度庆祝记录卡"
    section_title(ws, 1, "F8 进度庆祝记录卡", span=5)
    headers = ["日期", "里程碑成就", "庆祝方式", "感受", "继续动力"]
    header_row(ws, 2, headers)
    data = [
        ("2026-08-17", "连续7天晨起阅读", "买了一直想要的那本书", "非常满足", "满满成就感"),
        ("2026-08-20", "读完《深度工作》第一遍", "约朋友分享读后感", "充实自信", "想把方法分享给团队"),
    ]
    for i, row in enumerate(data, 3):
        for c, v in enumerate(row, 1):
            input_cell(ws, i, c, v, bg="FFFDE7" if c == 2 else "FFFFFF")
    for r in range(3 + len(data), 13):
        for c in range(1, 6):
            input_cell(ws, r, c, "", bg="FFFDE7" if c == 2 else "FFFFFF")
    set_col_widths(ws, [14, 22, 22, 24, 16])
    freeze_and_filter(ws, 3, 1)

def build_f9_filled(ws):
    ws.title = "F9 我的行为改变报告"
    section_title(ws, 1, "F9 我的行为改变报告", span=5)
    headers = ["维度", "目标", "实际进展", "差距原因", "改进方向"]
    header_row(ws, 2, headers)
    items = [
        ("提示设计", "环境提示明显稳定", "书在枕边+闹钟双重提示", "周末容易忽略闹钟", "增加周末专属提示"),
        ("渴望绑定", "新行为与内在渴望链接", "阅读与成长需求绑定", "成长渴望有时不够强烈", "每月设定一个主题挑战"),
        ("行动简化", "步骤最少化、阻力为零", "拿起书翻开即可", "偶尔想读却不知从哪翻", "提前标注今日阅读页"),
        ("物理环境", "物品摆放支持新行为", "书房阅读角已布置", "灯光有时不够理想", "更换暖白光阅读灯"),
        ("社交环境", "获得他人支持认可", "家人知道并偶尔询问", "缺少同行伙伴", "加入线上读书社群"),
        ("动机真实", "来自真实内在需求", "自我成长是真实渴望", "有时被工作压力冲淡", "每天早晨先读再工作"),
        ("总体评价", "行为已固化形成习惯", "晨读已坚持14天", "周末自律性稍弱", "继续优化环境设计"),
    ]
    for i, (dim, target, actual, gap, improve) in enumerate(items, 3):
        input_cell(ws, i, 1, dim, bg="F2F2F2")
        input_cell(ws, i, 2, target)
        input_cell(ws, i, 3, actual)
        input_cell(ws, i, 4, gap)
        input_cell(ws, i, 5, improve)
    set_col_widths(ws, [14, 20, 22, 22, 20])
    freeze_and_filter(ws, 3, 2)

def build_f10_filled(ws):
    ws.title = "F10 团队行为改变台账"
    section_title(ws, 1, "F10 团队行为改变台账", span=5)
    headers = ["成员", "目标行为", "当前状态", "支持需求", "下次检查日期"]
    header_row(ws, 2, headers)
    data = [
        ("小王", "每日早到后整理工位", "已坚持5天", "需要一个整理清单", "2026-08-25"),
        ("小李", "每周主动汇报进展", "刚开始执行", "需要汇报模板", "2026-08-27"),
        ("小张", "午休后站立5分钟", "已坚持10天", "暂无", "2026-09-03"),
    ]
    for i, row in enumerate(data, 3):
        for c, v in enumerate(row, 1):
            input_cell(ws, i, c, v, bg="FFFDE7" if c == 2 else "FFFFFF")
    for r in range(3 + len(data), 13):
        for c in range(1, 6):
            input_cell(ws, r, c, "", bg="FFFDE7" if c == 2 else "FFFFFF")
    set_col_widths(ws, [12, 20, 18, 20, 14])
    freeze_and_filter(ws, 3, 1)

# ─── Guide sheet ─────────────────────────────────────────────────────────────

def build_guide_sheet(ws):
    """表单使用指引"""
    ws.title = "表单使用指引"
    section_title(ws, 1, "表单使用说明与指引", span=2, bg="2E75B6")

    guide_rows = [
        (3, "一、表单套组说明"),
        (4, "本套表单共包含10张工具表单，配合「行为改变设计：用环境和触发器重塑自己」课程使用。"),
        (5, ""),
        (6, "二、表单清单"),
        (7, "表单编号    表单名称                  用途"),
        (8, "F1          行为环境设计自测         评估当前行为环境的四个关键维度"),
        (9, "F2          锚点行为分析             找到有效的锚点并绑定新行为"),
        (10, "F3          物理环境改造清单         规划物理空间的改变"),
        (11, "F4          社交环境改造清单         规划社交支持的改变"),
        (12, "F5          环境改造方案工作表       综合规划环境改造策略"),
        (13, "F6          实施追踪表             记录每日执行情况与障碍"),
        (14, "F7          习惯固化检查卡         定期检查习惯是否已固化"),
        (15, "F8          进度庆祝记录卡         记录里程碑并给自己正向反馈"),
        (16, "F9          我的行为改变报告         阶段性复盘与反思"),
        (17, "F10         团队行为改变台账         团队成员行为改变跟踪"),
        (18, ""),
        (19, "三、使用流程"),
        (20, "第一步：填写「必填信息表」，明确目标行为和当前状态"),
        (21, "第二步：使用「F1 行为环境设计自测」评估现状"),
        (22, "第三步：完成「F2-F5」环境改造规划表单"),
        (23, "第四步：执行并用「F6 实施追踪表」每日记录"),
        (24, "第五步：定期使用「F7 习惯固化检查卡」评估进度"),
        (25, "第六步：达成里程碑时填写「F8 进度庆祝记录卡」"),
        (26, "第七步：每周/每月使用「F9 我的行为改变报告」复盘"),
        (27, "第八步（可选）：团队使用「F10 团队行为改变台账」"),
        (28, ""),
        (29, "四、色块说明"),
        (30, "浅蓝色区域    已填写区域 / 参考示例"),
        (31, "黄色区域      待填写区域 / 重点关注项"),
        (32, "白色区域      普通填写区"),
        (33, ""),
        (34, "五、建议"),
        (35, "每次课程后立即填写对应的表单，趁记忆清晰时完成。"),
        (36, "追踪表建议每日填写，持续至少21天以形成习惯。"),
        (37, "庆祝卡在达成里程碑时填写，不要等到项目结束才回顾。"),
        (38, "团队台账建议每周固定时间统一更新。"),
    ]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, (row_num, *texts) in enumerate(guide_rows, 1):
        if not texts:
            continue
        text = texts[0] if texts else ""
        if row_num in (3, 6, 19, 29, 34):
            # section title
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            cell = ws.cell(row=row_num, column=1, value=text)
            cell.font = Font(bold=True, name="Calibri", size=12, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
        elif row_num == 7:
            header_row(ws, row_num, ["表单编号", "表单名称及用途"])
        else:
            if "    " in text:
                parts = text.split("    ")
                ws.cell(row=row_num, column=1, value=parts[0]).border = border
                ws.cell(row=row_num, column=2, value=parts[1] if len(parts) > 1 else "").border = border
            else:
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
                cell = ws.cell(row=row_num, column=1, value=text)
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = border

    set_col_widths(ws, [24, 60])
    ws.row_dimensions[1].height = 32
    for r in range(2, 39):
        ws.row_dimensions[r].height = 20

# ─── Main ─────────────────────────────────────────────────────────────────────

def create_empty():
    wb = Workbook()
    wb.remove(wb.active)
    build_info_sheet(wb.create_sheet())
    build_f1(wb.create_sheet())
    build_f2(wb.create_sheet())
    build_f3(wb.create_sheet())
    build_f4(wb.create_sheet())
    build_f5(wb.create_sheet())
    build_f6(wb.create_sheet())
    build_f7(wb.create_sheet())
    build_f8(wb.create_sheet())
    build_f9(wb.create_sheet())
    build_f10(wb.create_sheet())
    path = os.path.join(OUT_DIR, "配套表单_空表.xlsx")
    wb.save(path)
    print(f"Created: {path}")

def create_filled():
    wb = Workbook()
    wb.remove(wb.active)
    build_info_sheet_filled(wb.create_sheet())
    build_f1_filled(wb.create_sheet())
    build_f2_filled(wb.create_sheet())
    build_f3_filled(wb.create_sheet())
    build_f4_filled(wb.create_sheet())
    build_f5_filled(wb.create_sheet())
    build_f6_filled(wb.create_sheet())
    build_f7_filled(wb.create_sheet())
    build_f8_filled(wb.create_sheet())
    build_f9_filled(wb.create_sheet())
    build_f10_filled(wb.create_sheet())
    path = os.path.join(OUT_DIR, "配套表单_填好版.xlsx")
    wb.save(path)
    print(f"Created: {path}")

def create_guide():
    wb = Workbook()
    wb.remove(wb.active)
    build_guide_sheet(wb.create_sheet())
    path = os.path.join(OUT_DIR, "表单使用指引.xlsx")
    wb.save(path)
    print(f"Created: {path}")

if __name__ == "__main__":
    os.makedirs(OUT_DIR, exist_ok=True)
    create_empty()
    create_filled()
    create_guide()
    print("All done.")

# -*- coding: utf-8 -*-
"""D-08 重新生成：保留原数据 + 补一个排名柱状图"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

OUT = "D:/Downloads/xinjian/德赛西威评审全流程PRD/产出物/02-评审实施/D-08-基础班·场景化应用成果PK表.xlsx"

# 配色（设计系统）
COLOR_MAIN = "003D7A"     # 德赛蓝
COLOR_ACCENT = "00A0E9"   # 智能青
COLOR_BG_ALT = "F2F6FA"   # 交替行背景
COLOR_TEXT = "1A1A1A"

# 字体
font_title = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
font_header = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
font_body = Font(name="微软雅黑", size=10, color=COLOR_TEXT)
font_note = Font(name="微软雅黑", size=9, color="666666", italic=True)

# 填充
fill_title = PatternFill("solid", fgColor=COLOR_MAIN)
fill_header = PatternFill("solid", fgColor=COLOR_MAIN)
fill_alt = PatternFill("solid", fgColor=COLOR_BG_ALT)

# 对齐
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 边框
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def setup_sheet(ws):
    """冻结表头 + 筛选器"""
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:N3"


def add_title(ws, text, n_col=14):
    """第 1 行：项目名 + 文件名"""
    ws.merge_cells(f"A1:{get_column_letter(n_col)}1")
    c = ws["A1"]
    c.value = text
    c.font = font_title
    c.fill = fill_title
    c.alignment = align_center
    ws.row_dimensions[1].height = 32


def add_subtitle(ws, text, n_col=14):
    """第 2 行：说明"""
    ws.merge_cells(f"A2:{get_column_letter(n_col)}2")
    c = ws["A2"]
    c.value = text
    c.font = font_note
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36


def add_header(ws, headers, row=3):
    """表头行"""
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border
    ws.row_dimensions[row].height = 30


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ========== 主流程 ==========
wb = Workbook()
ws = wb.active
ws.title = "场景化应用成果PK"

# 1. 标题
add_title(ws, "德赛西威 AI 赋能课程评审全流程 · 场景化应用成果 PK 表")
add_subtitle(ws, "说明：本表用于基础班学员课后 2-4 周提交的\"做法+效果+效率对比\"。"
                  "五维评分各 0-25 分制。综合得分=AVERAGE 5 维。\n"
                  "【评分标准】很好 24-25 / 较好 21-23 / 一般 18-20 / 较差 15-17 / 差 0-14（仅用于得分判定）")

# 2. 表头
headers = [
    "编号", "学员", "部门", "岗位", "场景", "业务问题", "AI方案", "实施数据", "节省时间(小时)", "推广价值",
    "业务价值", "AI方案成熟度", "安全合规", "可复制性", "效果可衡量", "综合得分", "实时排名"
]
# 实际表头是 17 列（A-Q），修正列宽
n_col = 17
add_header(ws, headers)

# 修正标题行/说明行跨度
ws.unmerge_cells("A1:N1"); ws.merge_cells(f"A1:{get_column_letter(n_col)}1")
ws.unmerge_cells("A2:N2"); ws.merge_cells(f"A2:{get_column_letter(n_col)}2")
ws["A1"].alignment = align_center
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 40

# 3. 列宽
widths = [6, 10, 16, 14, 18, 22, 22, 18, 12, 14, 10, 12, 10, 10, 10, 10, 8]
set_column_widths(ws, widths)

# 4. 数据行（200 行预留，数据从第 4 行开始）
n_data = 200
for i in range(n_data):
    r = 4 + i
    for col in range(1, n_col + 1):
        c = ws.cell(row=r, column=col)
        c.font = font_body
        c.border = border
        c.alignment = align_left
        if i % 2 == 1:
            c.fill = fill_alt
    # 综合得分公式 = AVERAGE(K..O)
    ws.cell(row=r, column=16, value=f"=IFERROR(AVERAGE(K{r}:O{r}),\"\")").font = font_body
    # 实时排名公式 = RANK(P, P$4:P$203)
    ws.cell(row=r, column=17, value=f"=IFERROR(RANK(P{r},P$4:P$203),\"\")").font = font_body

# 5. 条件格式（数据条 + 颜色规则）—— 通过 openpyxl 不可视化，应用 conditional_formatting
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
# 综合得分列（P）：绿黄红色阶
ws.conditional_formatting.add(
    f"P4:P{3+n_data}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B")
)
# 节省时间列（I）：数据条
ws.conditional_formatting.add(
    f"I4:I{3+n_data}",
    ColorScaleRule(start_type="min", start_color="FFFFFF",
                   end_type="max", end_color=COLOR_ACCENT)
)

# 6. 添加排名柱状图（数据用占位示例）—— 引用综合得分列
# 用前 10 个学员做示例数据
chart = BarChart()
chart.type = "bar"
chart.style = 11
chart.title = "Top 10 学员综合得分排名（占位示例）"
chart.x_axis.title = "综合得分"
chart.y_axis.title = "学员"
# 数据：P 列前 10 行
data_ref = Reference(ws, min_col=16, min_row=3, max_col=16, max_row=13)
cat_ref = Reference(ws, min_col=2, min_row=4, max_col=2, max_row=13)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cat_ref)
chart.dataLabels = DataLabelList(showVal=True)
chart.width = 20
chart.height = 12

# 将图表放在 R3 位置（数据右侧）
ws.add_chart(chart, "S3")

# 7. 冻结 + 筛选
setup_sheet(ws)

# 保存
wb.save(OUT)
print(f"[OK] D-08 已重新生成：{OUT}")
import os
print(f"   大小: {os.path.getsize(OUT)} bytes")

# -*- coding: utf-8 -*-
"""
企业家刑事风险地图 Excel表单生成脚本
使用openpyxl创建6个Excel文件
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# 配色方案
HEADER_FILL = PatternFill(start_color="2b2d42", end_color="2b2d42", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
P0_FILL = PatternFill(start_color="ef233c", end_color="ef233c", fill_type="solid")  # 高风险红色
P1_FILL = PatternFill(start_color="ff9500", end_color="ff9500", fill_type="solid")  # 中风险橙色
P2_FILL = PatternFill(start_color="ffcc00", end_color="ffcc00", fill_type="solid")  # 低风险黄色
NORMAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_column_widths(ws, widths):
    """设置列宽"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def add_header_row(ws, headers, row=1):
    """添加表头行"""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 25

def add_data_row(ws, row, data, risk_level=None):
    """添加数据行"""
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = LEFT_ALIGN
        cell.border = THIN_BORDER
        if risk_level == "P0":
            cell.fill = P0_FILL
        elif risk_level == "P1":
            cell.fill = P1_FILL
        elif risk_level == "P2":
            cell.fill = P2_FILL
        else:
            cell.fill = NORMAL_FILL
    ws.row_dimensions[row].height = 20

def freeze_and_filter(ws, freeze_cell="A2"):
    """冻结表头和添加筛选"""
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions

def create_风险自检表(filepath):
    """创建企业家刑事风险自检表.xlsx"""
    wb = Workbook()

    # Sheet1 风险自检
    ws1 = wb.active
    ws1.title = "风险自检"
    headers = ["风险类别", "检查项", "风险等级(P0/P1/P2)", "是否存在问题", "备注"]
    add_header_row(ws1, headers)
    set_column_widths(ws1, [15, 40, 18, 15, 30])

    # 融资类风险
    financing_risks = [
        ("融资类", "是否涉及非法吸收公众存款", "P0", "", ""),
        ("融资类", "是否使用虚假材料骗取贷款", "P0", "", ""),
        ("融资类", "是否违规使用贷款资金用途", "P1", "", ""),
        ("融资类", "民间借贷利率是否超过法定上限", "P1", "", ""),
        ("融资类", "是否签署对赌协议并承担个人担保", "P1", "", ""),
        ("融资类", "融资过程中的财务顾问费是否合理", "P2", "", ""),
    ]

    # 财务类风险
    financial_risks = [
        ("财务类", "是否存在虚开增值税发票", "P0", "", ""),
        ("财务类", "是否存在偷税漏税行为", "P0", "", ""),
        ("财务类", "是否存在公款私用情况", "P1", "", ""),
        ("财务类", "是否存在账外账（小金库）", "P1", "", ""),
        ("财务类", "资金往来是否合规有据", "P2", "", ""),
        ("财务类", "财务报表是否真实完整", "P2", "", ""),
    ]

    # 用工类风险
    labor_risks = [
        ("用工类", "是否拖欠员工工资", "P0", "", ""),
        ("用工类", "是否强制员工加班", "P1", "", ""),
        ("用工类", "社保是否足额缴纳", "P1", "", ""),
        ("用工类", "劳动合同是否规范签订", "P2", "", ""),
        ("用工类", "是否存在工伤未申报情况", "P1", "", ""),
        ("用工类", "是否涉及劳务派遣超比例使用", "P2", "", ""),
        ("用工类", "员工离职手续是否合规", "P2", "", ""),
    ]

    # 股权类风险
    equity_risks = [
        ("股权类", "股权代持是否签订协议", "P1", "", ""),
        ("股权类", "是否存在股权争议或纠纷", "P0", "", ""),
        ("股权类", "股权转让是否依法纳税", "P1", "", ""),
        ("股权类", "股东会决议程序是否合规", "P2", "", ""),
        ("股权类", "是否存在抽逃出资行为", "P0", "", ""),
        ("股权类", "关联交易是否披露公允", "P1", "", ""),
        ("股权类", "是否存在公司担保连带责任", "P1", "", ""),
    ]

    all_risks = financing_risks + financial_risks + labor_risks + equity_risks
    for i, risk in enumerate(all_risks, 2):
        add_data_row(ws1, i, risk, risk[2])

    freeze_and_filter(ws1)

    # Sheet2 评分标准
    ws2 = wb.create_sheet("评分标准")
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 50

    add_header_row(ws2, ["风险等级", "含义说明"], 1)

    standards = [
        ("P0", "高风险：涉嫌刑事犯罪或重大违规，可能导致逮捕、判刑等严重后果，需立即整改", "ef233c"),
        ("P1", "中风险：存在明显违规行为，可能引发行政责任或民事赔偿，需尽快规范", "ff9500"),
        ("P2", "低风险：存在潜在隐患或最佳实践差距，短期内不至于违法，但需关注改进", "ffcc00"),
    ]

    for i, (level, desc, color) in enumerate(standards, 2):
        cell_a = ws2.cell(row=i, column=1, value=level)
        cell_a.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_a.font = Font(bold=True)
        cell_a.alignment = CENTER_ALIGN
        cell_a.border = THIN_BORDER

        cell_b = ws2.cell(row=i, column=2, value=desc)
        cell_b.alignment = LEFT_ALIGN
        cell_b.border = THIN_BORDER
        ws2.row_dimensions[i].height = 30

    ws2.row_dimensions[1].height = 25

    wb.save(filepath)
    print(f"Created: {filepath}")

def create_融资环节风险跟踪表(filepath):
    """创建融资环节风险跟踪表.xlsx"""
    wb = Workbook()

    # Sheet1 融资情况
    ws1 = wb.active
    ws1.title = "融资情况"
    headers = ["融资方式", "融资金额(万元)", "投资人/机构", "融资时间", "法律风险评级", "备注"]
    add_header_row(ws1, headers)
    set_column_widths(ws1, [15, 15, 20, 15, 15, 30])

    # 示例数据
    sample_data = [
        ("股权融资", "5000", "红杉资本", "2024-01", "P1", "需关注对赌协议风险"),
        ("银行贷款", "2000", "工商银行", "2024-03", "P2", "正常"),
        ("民间借贷", "500", "个人投资者", "2024-06", "P0", "利率较高，需注意合规"),
        ("", "", "", "", "", ""),
        ("", "", "", "", "", ""),
    ]

    for i, data in enumerate(sample_data, 2):
        risk = data[4] if data[4] else None
        add_data_row(ws1, i, data, risk)

    freeze_and_filter(ws1)

    # Sheet2 对赌协议
    ws2 = wb.create_sheet("对赌协议")
    headers2 = ["协议方", "业绩目标", "个人担保情况", "风险点", "应对措施"]
    add_header_row(ws2, headers2)
    set_column_widths(ws2, [20, 25, 20, 30, 30])

    sample_data2 = [
        ("投资人A", "2024年营收达3亿元", "是（个人连带）", "业绩不达标则触发回购", "提前与投资人沟通调整目标"),
        ("投资人B", "2025年IPO", "否", "上市时间不确定风险", "预留足够时间窗口"),
        ("", "", "", "", ""),
    ]

    for i, data in enumerate(sample_data2, 2):
        add_data_row(ws2, i, data)

    freeze_and_filter(ws2)

    wb.save(filepath)
    print(f"Created: {filepath}")

def create_财务税务风险跟踪表(filepath):
    """创建财务税务风险跟踪表.xlsx"""
    wb = Workbook()

    # Sheet1 发票管理
    ws1 = wb.active
    ws1.title = "发票管理"
    headers = ["发票号码", "开票日期", "金额(万元)", "业务类型", "是否异常", "备注"]
    add_header_row(ws1, headers)
    set_column_widths(ws1, [18, 15, 15, 20, 12, 30])

    sample_data = [
        ("NP12345678", "2024-01-15", "50.00", "咨询服务", "否", "正常"),
        ("NP12345679", "2024-02-20", "30.00", "技术服务", "否", "正常"),
        ("NP12345680", "2024-03-10", "100.00", "设备采购", "是", "发票内容与实际业务不符"),
        ("", "", "", "", "", ""),
    ]

    for i, data in enumerate(sample_data, 2):
        risk = "P1" if data[4] == "是" else None
        add_data_row(ws1, i, data, risk)

    freeze_and_filter(ws1)

    # Sheet2 税务风险
    ws2 = wb.create_sheet("税务风险")
    headers2 = ["税种", "欠税金额(万元)", "稽查情况", "处理状态", "风险等级"]
    add_header_row(ws2, headers2)
    set_column_widths(ws2, [15, 18, 20, 20, 15])

    sample_data2 = [
        ("增值税", "0", "无", "正常", ""),
        ("企业所得税", "5.00", "待稽查", "整改中", "P1"),
        ("个人所得税", "0", "无", "正常", ""),
        ("", "", "", "", ""),
    ]

    for i, data in enumerate(sample_data2, 2):
        risk = data[4] if data[4] else None
        add_data_row(ws2, i, data, risk)

    freeze_and_filter(ws2)

    wb.save(filepath)
    print(f"Created: {filepath}")

def create_用工环节风险跟踪表(filepath):
    """创建用工环节风险跟踪表.xlsx"""
    wb = Workbook()

    # Sheet1 员工管理
    ws1 = wb.active
    ws1.title = "员工管理"
    headers = ["员工姓名", "入职时间", "工资发放", "社保情况", "用工风险", "备注"]
    add_header_row(ws1, headers)
    set_column_widths(ws1, [15, 15, 15, 15, 15, 30])

    sample_data = [
        ("张三", "2022-03-01", "正常", "正常", "无", ""),
        ("李四", "2023-06-15", "正常", "欠缴2个月", "P1", "社保滞纳金风险"),
        ("王五", "2024-01-10", "正常", "正常", "无", ""),
        ("", "", "", "", "", ""),
    ]

    for i, data in enumerate(sample_data, 2):
        risk = data[4] if data[4] != "无" else None
        add_data_row(ws1, i, data, risk)

    freeze_and_filter(ws1)

    # Sheet2 工资发放
    ws2 = wb.create_sheet("工资发放")
    headers2 = ["月份", "应发工资(万元)", "实发工资(万元)", "欠薪情况", "风险等级"]
    add_header_row(ws2, headers2)
    set_column_widths(ws2, [15, 18, 18, 20, 15])

    sample_data2 = [
        ("2024-01", "50.00", "50.00", "无", ""),
        ("2024-02", "50.00", "50.00", "无", ""),
        ("2024-03", "55.00", "45.00", "欠薪10万", "P0"),
        ("2024-04", "55.00", "55.00", "无", ""),
    ]

    for i, data in enumerate(sample_data2, 2):
        risk = data[4] if data[4] else None
        add_data_row(ws2, i, data, risk)

    freeze_and_filter(ws2)

    wb.save(filepath)
    print(f"Created: {filepath}")

def create_股权治理风险跟踪表(filepath):
    """创建股权治理风险跟踪表.xlsx"""
    wb = Workbook()

    # Sheet1 股东情况
    ws1 = wb.active
    ws1.title = "股东情况"
    headers = ["股东姓名", "持股比例(%)", "出资情况", "资金往来", "风险点"]
    add_header_row(ws1, headers)
    set_column_widths(ws1, [15, 15, 15, 20, 35])

    sample_data = [
        ("甲", "40", "已出资", "正常", "无"),
        ("乙", "30", "已出资", "正常", "无"),
        ("丙", "20", "未完全出资", "频繁借款", "P1: 资金往来需规范"),
        ("丁", "10", "已出资", "正常", "无"),
        ("", "", "", "", ""),
    ]

    for i, data in enumerate(sample_data, 2):
        risk = "P1" if "P1" in str(data[4]) else None
        add_data_row(ws1, i, data, risk)

    freeze_and_filter(ws1)

    # Sheet2 决策记录
    ws2 = wb.create_sheet("决策记录")
    headers2 = ["决议事项", "时间", "参与人员", "程序合规性", "风险"]
    add_header_row(ws2, headers2)
    set_column_widths(ws2, [30, 15, 20, 15, 20])

    sample_data2 = [
        ("选举董事长", "2024-01-15", "全体股东", "合规", "无"),
        ("重大资产出售", "2024-03-20", "甲、乙、丙", "待核查", "P1: 决议程序存疑"),
        ("对外担保", "2024-06-10", "全体股东", "合规", "无"),
        ("", "", "", "", ""),
    ]

    for i, data in enumerate(sample_data2, 2):
        risk = "P1" if data[4] != "无" else None
        add_data_row(ws2, i, data, risk)

    freeze_and_filter(ws2)

    wb.save(filepath)
    print(f"Created: {filepath}")

def create_危机应对检查表(filepath):
    """创建危机应对检查表.xlsx"""
    wb = Workbook()

    # Sheet1 24小时响应
    ws1 = wb.active
    ws1.title = "24小时响应"
    headers = ["时间节点", "行动项", "执行人", "完成情况", "备注"]
    add_header_row(ws1, headers)
    set_column_widths(ws1, [15, 40, 15, 12, 30])

    # 0-6小时
    phase1 = [
        ("0-1小时", "成立危机应对小组", "", "", "确定总指挥"),
        ("0-1小时", "全面了解事件情况", "", "", "收集基本信息"),
        ("1-2小时", "评估事件性质和严重程度", "", "", "判断是否涉及刑事"),
        ("1-2小时", "通知相关董事、监事", "", "", "按章程规定执行"),
        ("2-4小时", "保全相关证据材料", "", "", "电子数据、书面文件等"),
        ("2-4小时", "委派专业律师介入", "", "", "刑辩律师+民商律师"),
        ("4-6小时", "制定初步应对方案", "", "", "准备多套方案"),
        ("4-6小时", "准备对内对外声明", "", "", "统一口径"),
    ]

    # 6-12小时
    phase2 = [
        ("6-8小时", "向监管部门报告（如需）", "", "", "主动沟通争取有利地位"),
        ("6-8小时", "准备相关人员管控方案", "", "", "防止串供、毁灭证据"),
        ("8-10小时", "梳理可能涉及的法律责任", "", "", "刑事/行政/民事"),
        ("8-10小时", "评估对公司的影响", "", "", "经营、融资、声誉"),
        ("10-12小时", "确定对外发言人和口径", "", "", "避免混乱表态"),
        ("10-12小时", "部署后续调查配合工作", "", "", "准备材料清单"),
    ]

    # 12-24小时
    phase3 = [
        ("12-18小时", "向全体员工通报情况（如需）", "", "", "稳定军心"),
        ("12-18小时", "启动应急预案", "", "", "业务连续性保障"),
        ("18-24小时", "与办案机关建立沟通", "", "", "了解调查进展"),
        ("18-24小时", "评估是否需要自首", "", "", "律师专业意见"),
        ("18-24小时", "准备第一轮声明", "", "", "平衡信息透明与法律风险"),
    ]

    all_data = phase1 + phase2 + phase3
    for i, data in enumerate(all_data, 2):
        add_data_row(ws1, i, data)

    freeze_and_filter(ws1)

    # Sheet2 三不原则
    ws2 = wb.create_sheet("三不原则")
    headers2 = ["原则", "具体要求", "执行确认", "签名"]
    add_header_row(ws2, headers2)
    set_column_widths(ws2, [15, 45, 15, 20])

    principles = [
        ("不擅自接触", "未经律师许可，不主动接触办案人员", "£已执行", ""),
        ("不擅自表态", "不对外（包括媒体、客户、员工）擅自发表关于事件的任何言论", "£已执行", ""),
        ("不擅自和解", "不擅自与对方当事人或家属达成任何和解协议或赔偿", "£已执行", ""),
    ]

    for i, data in enumerate(principles, 2):
        add_data_row(ws2, i, data)

    # 添加注意事项
    ws2.cell(row=6, column=1, value="注意事项").font = Font(bold=True)
    ws2.cell(row=6, column=2, value="三不原则是危机初期最重要的行为准则，违反可能导致串供、干扰证人等更严重的法律后果")
    ws2.merge_cells('B6:D6')
    ws2['B6'].alignment = LEFT_ALIGN

    wb.save(filepath)
    print(f"Created: {filepath}")

def main():
    base_dir = "D:/新课开发/法学/14-企业家刑事风险地图：从合规到危机应对/配套表单Excel"

    files = {
        "企业家刑事风险自检表.xlsx": create_风险自检表,
        "融资环节风险跟踪表.xlsx": create_融资环节风险跟踪表,
        "财务税务风险跟踪表.xlsx": create_财务税务风险跟踪表,
        "用工环节风险跟踪表.xlsx": create_用工环节风险跟踪表,
        "股权治理风险跟踪表.xlsx": create_股权治理风险跟踪表,
        "危机应对检查表.xlsx": create_危机应对检查表,
    }

    for filename, creator_func in files.items():
        filepath = os.path.join(base_dir, filename)
        creator_func(filepath)

    print("\n所有Excel文件创建完成！")

if __name__ == "__main__":
    main()

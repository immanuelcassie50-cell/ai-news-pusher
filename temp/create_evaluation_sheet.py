#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建课程评估表 Excel 文件
家庭亲子 - 青春期前的性教育与身体教育
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: 基本信息
# ============================================================
ws1 = wb.active
ws1.title = "基本信息"

# 样式定义
header_fill = PatternFill(start_color="1a5f7a", end_color="1a5f7a", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
header_align = Alignment(horizontal="center", vertical="center")

title_font = Font(bold=True, size=16)
title_align = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 标题
ws1.merge_cells('A1:D1')
ws1['A1'] = '课程评估表'
ws1['A1'].font = title_font
ws1['A1'].alignment = title_align

# 表头
headers1 = ['课程名称', '日期', '讲师', '学员人数']
for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

# 示例数据
data1 = [
    ['青春期前的性教育与身体教育', '2026-08-15', '张老师', 25],
    ['青春期前的性教育与身体教育', '2026-08-10', '李老师', 30],
    ['青春期前的性教育与身体教育', '2026-08-05', '王老师', 28],
    ['青春期前的性教育与身体教育', '2026-07-28', '张老师', 22],
    ['青春期前的性教育与身体教育', '2026-07-20', '刘老师', 35],
]

for row_idx, row_data in enumerate(data1, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 设置列宽
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 12

# ============================================================
# Sheet 2: 满意度调查
# ============================================================
ws2 = wb.create_sheet(title="满意度调查")

# 标题
ws2.merge_cells('A1:F1')
ws2['A1'] = '满意度调查'
ws2['A1'].font = title_font
ws2['A1'].alignment = title_align

# 表头
headers2 = ['课程内容', '讲师表达', '互动体验', '实用性', '整体评价']
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

# 20行评分数据 (1-5分)
import random
random.seed(42)  # 保证数据可重现

# 交替行颜色
light_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for row_idx in range(4, 24):
    for col_idx in range(1, 6):
        # 生成3-5分的评分，中心倾向4
        score = random.choice([3, 4, 4, 4, 5, 5])
        cell = ws2.cell(row=row_idx, column=col_idx, value=score)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # 交替背景色
        if row_idx % 2 == 0:
            cell.fill = light_fill

# 设置列宽
for col in range(1, 6):
    ws2.column_dimensions[get_column_letter(col)].width = 15

# ============================================================
# Sheet 3: 知识掌握度
# ============================================================
ws3 = wb.create_sheet(title="知识掌握度")

# 标题
ws3.merge_cells('A1:E1')
ws3['A1'] = '知识掌握度'
ws3['A1'].font = title_font
ws3['A1'].alignment = title_align

# 表头
headers3 = ['题号', '题目内容', '答案', '学员答案', '结果']
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

# 10道判断题 (错/对 = 错误/正确)
questions = [
    ("等孩子长大了自然就懂了", "错"),
    ("性教育会让孩子更好奇", "错"),
    ("性教育越早开始越好", "对"),
    ("性教育会破坏孩子的纯真", "错"),
    ("父母不需要学习性教育知识", "错"),
    ("性教育是学校的事情", "错"),
    ("性教育包括性别认同和尊重", "对"),
    ("性教育会鼓励早期性行为", "错"),
    ("家长应该等待孩子提问再教育", "错"),
    ("性教育是终身学习的过程", "对"),
]

# 学员答案 (模拟有对有错)
import random
random.seed(42)
student_answers = ["对", "错", "对", "错", "对", "错", "对", "错", "对", "错"]

for row_idx, (question, answer) in enumerate(questions, 4):
    # 题号
    cell = ws3.cell(row=row_idx, column=1, value=row_idx - 3)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if row_idx % 2 == 0:
        cell.fill = light_fill

    # 题目内容
    cell = ws3.cell(row=row_idx, column=2, value=question)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if row_idx % 2 == 0:
        cell.fill = light_fill

    # 正确答案
    cell = ws3.cell(row=row_idx, column=3, value=answer)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if row_idx % 2 == 0:
        cell.fill = light_fill

    # 学员答案
    cell = ws3.cell(row=row_idx, column=4, value=student_answers[row_idx - 4])
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if row_idx % 2 == 0:
        cell.fill = light_fill

    # 结果 (公式判断)
    result_formula = f'=IF(D{row_idx}=C{row_idx},"正确","错误")'
    cell = ws3.cell(row=row_idx, column=5, value=result_formula)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if row_idx % 2 == 0:
        cell.fill = light_fill

# 设置列宽
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 10

# ============================================================
# Sheet 4: 改进建议
# ============================================================
ws4 = wb.create_sheet(title="改进建议")

# 标题
ws4.merge_cells('A1:C1')
ws4['A1'] = '改进建议'
ws4['A1'].font = title_font
ws4['A1'].alignment = title_align

# 表头
headers4 = ['序号', '建议内容', '优先级(高/中/低)']
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

# 10条建议
suggestions = [
    ("增加更多互动环节，让家长参与角色扮演", "高"),
    ("提供更多具体的案例分析和应对方法", "高"),
    ("制作配套的宣传资料便于家长回家后使用", "中"),
    ("延长课程时间，有更多讨论机会", "中"),
    ("增加针对不同年龄段的针对性内容", "高"),
    ("提供课后答疑或线上交流群", "中"),
    ("增加男性家长参与的建议和指导", "低"),
    ("提供更多关于网络性安全的内容", "中"),
    ("制作家长指南手册作为课程配套", "中"),
    ("增加如何与孩子谈论敏感话题的技巧", "高"),
]

# 优先级颜色
high_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
medium_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
low_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

for row_idx, (content, priority) in enumerate(suggestions, 4):
    # 序号
    cell = ws4.cell(row=row_idx, column=1, value=row_idx - 3)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if row_idx % 2 == 0:
        cell.fill = light_fill

    # 建议内容
    cell = ws4.cell(row=row_idx, column=2, value=content)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if row_idx % 2 == 0:
        cell.fill = light_fill

    # 优先级
    cell = ws4.cell(row=row_idx, column=3, value=priority)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if priority == "高":
        cell.fill = high_fill
    elif priority == "中":
        cell.fill = medium_fill
    else:
        cell.fill = low_fill

# 设置列宽
ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 50
ws4.column_dimensions['C'].width = 18

# ============================================================
# 保存文件
# ============================================================
output_path = r"D:\新课开发\家庭亲子\19-青春期前的性教育与身体教育如何开口说什么\05-课程评估表.xlsx"

# 确保目录存在
import os
os.makedirs(os.path.dirname(output_path), exist_ok=True)

wb.save(output_path)
print(f"文件已保存至: {output_path}")

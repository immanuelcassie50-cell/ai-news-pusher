from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "反馈话术卡"

# Define styles
title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
scenario_font = Font(name='微软雅黑', size=11, bold=True)
step_font = Font(name='微软雅黑', size=10, bold=True)
content_font = Font(name='微软雅黑', size=10)
note_font = Font(name='微软雅黑', size=9, italic=True, color='666666')

header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
scenario_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
step_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Set column widths
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 35

row = 1

# Title
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = 'F4_反馈话术卡'
ws[f'A{row}'].font = title_font
ws[f'A{row}'].fill = header_fill
ws[f'A{row}'].alignment = center_align
ws.row_dimensions[row].height = 35
row += 1

# Subtitle
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = '管理者的反馈话术工具箱 — 让沟通有结构、有温度、有效果'
ws[f'A{row}'].font = Font(name='微软雅黑', size=10, italic=True, color='666666')
ws[f'A{row}'].alignment = center_align
row += 2

# ===== Scenario 1 =====
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = '场景一：当场纠正（Immediate Correction）'
ws[f'A{row}'].font = scenario_font
ws[f'A{row}'].fill = scenario_fill
ws[f'A{row}'].alignment = left_align
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = '使用场景：安全问题、明显错误、紧急情况'
ws[f'A{row}'].font = note_font
ws[f'A{row}'].alignment = left_align
row += 1

# Header
for col, header in enumerate(['步骤', '要点', '参考话术', '填写区域'], 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[row].height = 22
row += 1

# Content
scenario1_data = [
    ('第一步', '指出具体行为', '我刚才看到你______，这个操作有风险', ''),
    ('第二步', '说明影响', '这样做会导致______', ''),
    ('第三步', '给出正确做法', '正确的方式是______', ''),
    ('第四步', '确认理解', '你来说一遍，应该怎么做？', ''),
]

for item in scenario1_data:
    for col, value in enumerate(item, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = content_font
        cell.alignment = left_align if col != 1 else center_align
        cell.border = thin_border
        if col == 1:
            cell.fill = step_fill
    ws.row_dimensions[row].height = 28
    row += 1

row += 1

# ===== Scenario 2 =====
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = '场景二：私下沟通（Private Conversation）'
ws[f'A{row}'].font = scenario_font
ws[f'A{row}'].fill = scenario_fill
ws[f'A{row}'].alignment = left_align
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = '使用场景：非紧急问题、需要深入讨论、敏感话题'
ws[f'A{row}'].font = note_font
ws[f'A{row}'].alignment = left_align
row += 1

# Header
for col, header in enumerate(['步骤', '要点', '参考话术', '填写区域'], 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[row].height = 22
row += 1

scenario2_data = [
    ('开场', '建立安全感', '今天找你聊，不是批评你，是想帮你把工作做得更好', ''),
    ('描述事实', '具体观察', '最近我观察到______（具体事情）', ''),
    ('表达影响', '说明后果', '这件事对______（团队/客户/业务）产生了______影响', ''),
    ('听取想法', '开放提问', '你怎么看这件事？', ''),
    ('共同探讨', '协作解决', '你觉得可以怎么改进？', ''),
    ('确认行动', '明确约定', '那我们约定______，你看可以吗？', ''),
]

for item in scenario2_data:
    for col, value in enumerate(item, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = content_font
        cell.alignment = left_align if col != 1 else center_align
        cell.border = thin_border
        if col == 1:
            cell.fill = step_fill
    ws.row_dimensions[row].height = 28
    row += 1

row += 1

# ===== Scenario 3 =====
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = '场景三：绩效面谈（Performance Review）'
ws[f'A{row}'].font = scenario_font
ws[f'A{row}'].fill = scenario_fill
ws[f'A{row}'].alignment = left_align
ws.row_dimensions[row].height = 25
row += 1

# Header
for col, header in enumerate(['阶段', '时间', '参考话术', '填写区域'], 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[row].height = 22
row += 1

scenario3_data = [
    ('开场暖场', '2分钟', '今天我们来做季度绩效面谈，先聊聊你这段时间的整体感受', ''),
    ('成绩肯定', '-', '你这季度在______方面做得很好，具体体现在______', ''),
    ('问题探讨', '-', '在______方面还有提升空间，具体是______', ''),
    ('制定计划', '-', '下个季度我们重点提升______，目标是______', ''),
    ('结尾鼓励', '-', '我相信你可以在______方面做得更好', ''),
]

for item in scenario3_data:
    for col, value in enumerate(item, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = content_font
        cell.alignment = left_align if col != 1 else center_align
        cell.border = thin_border
        if col == 1:
            cell.fill = step_fill
    ws.row_dimensions[row].height = 28
    row += 1

row += 1

# ===== Scenario 4 =====
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = '场景四：对抗/抵触场景（Confrontation/Resistance）'
ws[f'A{row}'].font = scenario_font
ws[f'A{row}'].fill = scenario_fill
ws[f'A{row}'].alignment = left_align
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = '使用场景：下属不服气、有情绪、争辩'
ws[f'A{row}'].font = note_font
ws[f'A{row}'].alignment = left_align
row += 1

# Header
for col, header in enumerate(['步骤', '要点', '参考话术', '填写区域'], 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[row].height = 22
row += 1

scenario4_data = [
    ('第一步', '认可情绪', '我理解你对这个评价有不同看法，这很正常', ''),
    ('第二步', '聚焦事实', '我们先不看观点，先看事实。那天具体发生了什么？', ''),
    ('第三步', '还原场景', '当时你是怎么考虑的？', ''),
    ('第四步', '探讨影响', '回过头来看，那个决定产生了什么影响？', ''),
    ('第五步', '达成共识', '基于这些，我们能不能在______方面达成一致？', ''),
]

for item in scenario4_data:
    for col, value in enumerate(item, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = content_font
        cell.alignment = left_align if col != 1 else center_align
        cell.border = thin_border
        if col == 1:
            cell.fill = step_fill
    ws.row_dimensions[row].height = 28
    row += 1

row += 1

# ===== General Principles =====
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = '通用原则'
ws[f'A{row}'].font = scenario_font
ws[f'A{row}'].fill = scenario_fill
ws[f'A{row}'].alignment = left_align
ws.row_dimensions[row].height = 25
row += 1

principles_data = [
    ('描述行为，不评价人格', '✅ 我看到你在会议上打断了别人', '❌ 你这个人就是不尊重人'),
    ('用"我观察到"而不是"你总是"', '✅ 我观察到最近三次交付都有延误', '❌ 你总是迟到，总是不按时完成任务'),
    ('具体例子，不要泛泛而谈', '✅ 上周三的会议你提前离开了，没有完成会议记录', '❌ 你工作态度有问题'),
    ('对事不对人', '✅ 这个方案在成本控制上有风险', '❌ 你做事情不考虑后果'),
]

for principle, do_example, dont_example in principles_data:
    ws.cell(row=row, column=1, value=principle).font = step_font
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=1).border = thin_border
    ws.merge_cells(f'B{row}:D{row}')
    ws.cell(row=row, column=2, value=f'{do_example}\n{dont_example}').font = content_font
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.cell(row=row, column=2).border = thin_border
    ws.row_dimensions[row].height = 40
    row += 1

# Footer
row += 1
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = '本工具由「语音记录转管理者赋能手册」项目组出品'
ws[f'A{row}'].font = Font(name='微软雅黑', size=9, italic=True, color='999999')
ws[f'A{row}'].alignment = center_align

# Set print settings
from openpyxl.worksheet.page import PageMargins
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

output_path = 'D:/新课开发/经验萃取/语音记录转管理者赋能手册/完整课程包/工具表单/F4_反馈话术卡.xlsx'
wb.save(output_path)
print(f'Excel file created: {output_path}')

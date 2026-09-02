from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "互动流程时间表"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_font = Font(name='微软雅黑', size=14, bold=True, color="FFFFFF")
header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
content_font = Font(name='微软雅黑', size=10)
wrap_alignment = Alignment(wrap_text=True, vertical='top')
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Row 1: Title
ws.merge_cells('A1:G1')
ws['A1'] = '互动流程时间表 · 07_保险从业者心理韧性建设（5小时完整时间表）'
ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color="FFFFFF")
ws['A1'].fill = header_fill
ws['A1'].alignment = center_alignment

# Row 2: Course info
ws.merge_cells('A2:G2')
ws['A2'] = '课程:《保险从业者心理韧性建设》| 时长:300分钟(5小时) | 受众:寿险个险渠道营销员/主管 | 讲师:待定 | 地点:培训室'
ws['A2'].font = Font(name='微软雅黑', size=10, italic=True)
ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')

# Row 3: Headers
headers = ['时段', '时长(min)', '环节', '形式', '讲师做什么', '学员做什么', '对应方法/文件']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cell.alignment = center_alignment
    cell.border = thin_border

# Define the 5-hour course flow (300 minutes total)
course_flow = [
    # Module 1: 破题
    ('0:00-0:05', 5, '开场·目标对齐', '讲解+个人思考',
     '1.不讲"为什么重要"（学员已懂）2.直接给目标:5小时后你能建立自己的心理韧性框架 3.给出3个模块的预期收获',
     '1.听 2.写1句话:我对这节课的1个具体期待',
     'M1懂/会/信;02诊断'),

    ('0:05-0:15', 10, '模块一·为什么"打鸡血"正在失效', '讲解+案例',
     '1.抛问题:你最近一次"被鸡血"是什么时候？有用吗？ 2.讲"打鸡血"失效的3个原因:外部激励vs内部动力/短期情绪vs长期习惯/口号vs方法 3.用类比:鸡血像止痛药，治标不治本',
     '1.听 2.30秒回忆 3.1-2人分享自己的"鸡血经历"',
     '05吸引力类比(止痛药类比)'),

    ('0:15-0:25', 10, '模块一·行业转型期的真实处境', '讲解+全班分享',
     '1.讲行业数据:寿险代理人数量变化/保费增速放缓/客户需求转变 2.抛问题:这种环境下，你最真实的感受是什么？ 3.30秒留白让学员思考',
     '1.听 2.30秒独立思考 3.2-3人分享真实感受',
     '08四层提问(感受层)'),

    ('0:25-0:30', 5, '模块一·收尾', '提炼',
     '1.提炼:焦虑不是你的问题，是时代的问题 2.引入下一节:但焦虑有4种类型，不同类型有不同的应对方法',
     '1.听 2.做笔记',
     '03结构化大纲'),

    # Module 2: 焦虑来源识别
    ('0:30-0:40', 10, '模块二·焦虑类型1:业绩焦虑', '讲解+个人诊断',
     '1.讲清楚业绩焦虑的定义:对短期业绩目标的过度担忧 2.讲症状:睡不着/过度拜访/自我怀疑 3.发诊断卡:过去1个月你有多少天因为业绩睡不着？',
     '1.听 2.独立填写诊断卡 3.给自己打分1-5',
     '01需求分析;08四层提问'),

    ('0:40-0:55', 15, '模块二·焦虑类型2:职业倦怠', '讲解+两人讨论',
     '1.讲清楚职业倦怠的定义:长期投入 vs 看不到回报 2.讲3个信号:情绪耗竭/去人格化/低成就感 3.案例:入行3年的优秀代理人为什么突然想转行',
     '1.听 2.2人组讨论:你自己的倦怠信号是什么',
     '06空白案例(倦怠案例)'),

    ('0:55-1:15', 20, '模块二·焦虑类型3:前景焦虑', '讲解+小组分享',
     '1.讲清楚前景焦虑:对行业/职业未来不确定性的担忧 2.讲2种典型:行业会不会消失/我的技能有没有市场价值 3.互动:4人组讨论-你的前景焦虑具体是什么',
     '1.听 2.4人组讨论 3.每组分享1个最真实的担忧',
     '08四层提问(事实层);09处理分歧'),

    ('1:15-1:30', 15, '模块二·焦虑类型4:关系焦虑', '讲解+案例',
     '1.讲清楚关系焦虑:客户/主管/家人 的压力 2.案例:客户说"保险都是骗人的"你怎么接？ 3.讲3种典型:客户信任/同业竞争/家人不支持',
     '1.听 2.案例中的角色扮演(自愿或指定)',
     '05吸引力类比;06空白案例'),

    ('1:30-1:40', 10, '模块二·4类焦虑总结', '提炼+工具',
     '1.画4象限图:业绩/职业倦怠/前景/关系 2.让学员自测:哪类焦虑最严重 3.给出"焦虑温度计"工具:1-10分你现在是多少',
     '1.听 2.在焦虑温度计上标出自己的位置 3.2人组分享',
     '04清晰表达(焦虑温度计)'),

    # Module 3: 心理韧性方法
    ('1:40-1:55', 15, '模块三·方法1:认知重构', '讲解+演练',
     '1.讲什么是认知重构:不是改变环境，是改变你对环境的解读 2.讲"ABC法则":A事件-B想法-C结果 3.案例演练:同一件事(客户拒绝)的3种不同解读',
     '1.听 2.想一个最近让你焦虑的事件 3.用ABC法则重写',
     '04清晰表达(ABC法则);06空白案例'),

    ('1:55-2:15', 20, '模块三·方法2:情绪调节技术', '讲解+体验',
     '1.讲3种实用技术:呼吸调节/身体扫描/接地练习 2.带学员现场体验呼吸调节(4-7-8呼吸法) 3.讲为什么身体先放松，情绪才能跟上',
     '1.听 2.现场跟着练呼吸调节 3.分享体验',
     '04清晰表达(4-7-8呼吸法)'),

    ('2:15-2:35', 20, '模块三·方法3:社会支持网络', '讲解+互动',
     '1.讲为什么支持系统重要:独行者速，众行者远 2.讲3层支持:家庭/同事/专业 3.发"支持系统地图":让学员画出自己的3层支持网',
     '1.听 2.画支持系统地图 3.2人组分享',
     '06空白案例;08四层提问'),

    ('2:35-2:55', 20, '模块三·方法4:意义重建', '讲解+个人反思',
     '1.讲为什么意义感是心理韧性的核心:知道自己为什么做，比怎么做更重要 2.讲"职业价值金字塔":生存/发展/意义 3.让学员反思:我现在在哪一层？我想去哪一层？',
     '1.听 2.画自己的价值金字塔 3.写1句话:我的职业意义是什么',
     '08四层提问(行动层);04清晰表达'),

    # Module 4: 职业价值支点
    ('2:55-3:10', 15, '模块四·重新定义成功', '讲解+讨论',
     '1.挑战"业绩=成功"的等式:如果业绩不是唯一标准，什么是？ 2.讲3种成功:财务成功/关系成功/影响力成功 3.案例:一位资深代理人的"半退休"选择',
     '1.听 2.30秒思考:你认为的成功是什么 3.2-3人分享',
     '05吸引力类比;09处理分歧'),

    ('3:10-3:30', 20, '模块四·找到你的价值支点', '个人反思+工具',
     '1.发"职业价值罗盘"工具 2.让学员填写:最让我有动力的3件事/我最擅长的3个技能/市场最需要的3个价值 3.找3个交叉点:这就是你的价值支点',
     '1.填职业价值罗盘 2.2人组分享 3.找出自己的价值支点',
     '06空白案例;04清晰表达'),

    ('3:30-3:40', 10, '模块四·建立"反脆弱"思维', '讲解+案例',
     '1.讲"反脆弱":不是抵抗风险，是从风险中获益 2.案例:每次被拒绝都是学习机会 3.给"拒绝日志"模板:记录每次拒绝后的学习',
     '1.听 2.想一个被拒绝的经历 3.用新视角重新解读',
     '05吸引力类比(反脆弱)'),

    # Module 5: 行动计划设计
    ('3:40-3:55', 15, '模块五·设计你的韧性行动计划', '个人设计+辅导',
     '1.发"30天韧性行动计划"模板 2.讲3个要素:具体目标/每周行动/里程碑 3.辅导:走到学员中，帮助每个人制定计划',
     '1.设计自己的30天行动计划 2.可以随时举手问问题',
     '04清晰表达;07互动流程'),

    ('3:55-4:10', 15, '模块五·公开承诺与监督机制', '互动+承诺',
     '1.让学员找学习伙伴(2人组) 2.每人向同桌做公开承诺:我要坚持的1件事 3.建立check-in机制:每周1次2人组互相问候',
     '1.找学习伙伴 2.做公开承诺 3.约定check-in方式',
     '08四层提问(行动层)'),

    ('4:10-4:20', 10, '模块五·课程收尾', '总结+金句',
     '1.回顾5小时的核心收获 2.给出关键金句:心理韧性不是没有情绪，是带着情绪还能做对的事 3.强调:30天内用一次才算真的学会',
     '1.听 2.写1句话:我30天内要做的第1件事',
     '10最终课程计划'),

    ('4:20-4:30', 10, '答疑+反馈', 'Q&A+评估',
     '1.最后10分钟答疑 2.发反馈表 3.收集"焦虑温度计"看看课程前后的变化',
     '1.提问 2.填反馈表 3.交焦虑温度计(课后版)',
     '评估反馈'),
]

# Write course flow to Excel
row_num = 4
for item in course_flow:
    ws.cell(row=row_num, column=1, value=item[0])
    ws.cell(row=row_num, column=2, value=item[1])
    ws.cell(row=row_num, column=3, value=item[2])
    ws.cell(row=row_num, column=4, value=item[3])
    ws.cell(row=row_num, column=5, value=item[4])
    ws.cell(row=row_num, column=6, value=item[5])
    ws.cell(row=row_num, column=7, value=item[6])

    # Apply styles
    for col in range(1, 8):
        cell = ws.cell(row=row_num, column=col)
        cell.font = content_font
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if col == 2:  # Duration column - center
            cell.alignment = center_alignment

    row_num += 1

# Add summary row
ws.merge_cells(f'A{row_num}:G{row_num}')
ws[f'A{row_num}'] = '关键指标: 讲师讲授约120分钟(40%) | 学员互动约150分钟(50%) | 答疑反馈约30分钟(10%)'
ws[f'A{row_num}'].font = Font(name='微软雅黑', size=10, bold=True)
ws[f'A{row_num}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws[f'A{row_num}'].alignment = center_alignment

# Set column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 50
ws.column_dimensions['F'].width = 35
ws.column_dimensions['G'].width = 30

# Save
output_path = 'D:/新课开发/保险/7、保险从业者心理韧性建设在行业转型期找到真实的职业意义/成果demo/07_Interaction_Flow.xlsx'
wb.save(output_path)
print(f"Excel file created successfully at: {output_path}")

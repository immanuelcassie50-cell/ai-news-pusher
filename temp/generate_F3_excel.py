# -*- coding: utf-8 -*-
"""生成 F3_学习者提问分析表.xlsx"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = "D:/新课开发/企业大学/对内/5.AI导师与专属课程设计：从标准课程到AI可交互的学习体验/05-全流程工具表单/F3_学习者提问分析表.xlsx"

# 样式定义
HEADER_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
HEADER_FONT = Font(bold=True, name="微软雅黑", size=11)
DATA_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_header_style(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

def apply_data_style(cell):
    cell.font = DATA_FONT
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def freeze_and_filter(ws, freeze="A2"):
    ws.freeze_panes = freeze

def main():
    wb = Workbook()

    # ========== 工作表1：提问类型分类 ==========
    ws1 = wb.active
    ws1.title = "提问类型分类"

    # 表头
    headers1 = ["问题描述", "问题类型", "难度等级", "建议回应策略"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # 数据
    data1 = [
        ("这门课需要什么基础？", "事实型", "初级", "直接回答前置要求，提供补齐建议"),
        ("学习路径是什么？", "事实型", "初级", "提供清晰的学习步骤和时间规划"),
        ("什么是知识图谱？", "概念型", "初级", "用简洁语言定义，必要时举例"),
        ("为什么要设计学习路径？", "概念型", "中级", "解释原理解释其价值"),
        ("Prompt是什么？", "概念型", "初级", "定义+简单示例"),
        ("如何设计一个好的Prompt？", "应用型", "中级", "提供设计框架和案例"),
        ("AI导师和传统教学有什么区别？", "分析型", "中级", "对比分析+各自适用场景"),
        ("这个案例中的关键成功因素是什么？", "分析型", "高级", "结构化分析+数据支持"),
        ("如何评估AI导师的效果？", "应用型", "高级", "提供评估指标和方法论"),
        ("如何设计一个互动性强的学习体验？", "创造型", "高级", "提供设计原则和实践步骤"),
        ("为什么我的Prompt效果不好？", "诊断型", "中级", "帮助分析可能原因+提供改进方向"),
        ("能否给这个知识体系提改进建议？", "评价型", "高级", "提供具体可行的改进建议"),
    ]

    for row_idx, row_data in enumerate(data1, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell)

    # 列宽
    set_col_width(ws1, "A", 30)
    set_col_width(ws1, "B", 12)
    set_col_width(ws1, "C", 10)
    set_col_width(ws1, "D", 35)

    # 冻结首行
    freeze_and_filter(ws1, "A2")

    # 数据验证：难度等级下拉选择
    dv1 = DataValidation(
        type="list",
        formula1='"初级,中级,高级"',
        allow_blank=False
    )
    dv1.error = "请从下拉列表中选择"
    dv1.errorTitle = "输入无效"
    ws1.add_data_validation(dv1)
    dv1.add(f"C2:C{1 + len(data1)}")

    # ========== 工作表2：常见错误认知 ==========
    ws2 = wb.create_sheet("常见错误认知")

    # 表头
    headers2 = ["错误认知描述", "纠正策略", "Prompt模板"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # 数据
    data2 = [
        (
            "AI导师什么都会答",
            "说明AI导师的能力边界，强调人机协作的重要性",
            "作为AI导师，我专注于[专业领域]，对于超出范围的复杂问题，我会建议你咨询相关专家"
        ),
        (
            "越长的Prompt效果越好",
            "解释简洁Prompt的优势，提供优化方法",
            "请用最简洁的语言解释[概念]，不超过100字"
        ),
        (
            "AI给出的答案一定正确",
            "强调批判性思维的重要性，鼓励质疑验证",
            "我的回答仅供参考，请结合你的实际情况判断。你认为哪里可能有疑问？"
        ),
        (
            "学习效果可以完全依赖AI评估",
            "说明多元评估的重要性，强调自我反思的价值",
            "除了我的评估，你也可以通过[具体方法]来自我检验学习效果"
        ),
        (
            "有了AI导师就不需要老师了",
            "说明AI导师和人类老师的互补关系",
            "AI导师擅长[擅长领域]，而人类老师能提供[独特价值]。最佳学习体验是两者结合"
        ),
        (
            "提问越简单越好",
            "解释好问题的特征，鼓励深入思考",
            "为了给你更有价值的回答，能否详细描述一下你的具体情境？"
        ),
    ]

    for row_idx, row_data in enumerate(data2, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell)

    # 列宽
    set_col_width(ws2, "A", 30)
    set_col_width(ws2, "B", 35)
    set_col_width(ws2, "C", 35)

    # 冻结首行
    freeze_and_filter(ws2, "A2")

    # 保存
    wb.save(OUTPUT_PATH)
    print(f"Excel文件已生成: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()

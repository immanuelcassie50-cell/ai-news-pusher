from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# ========== Sheet 1: Prompt模板 ==========
ws1 = wb.active
ws1.title = "Prompt模板"

# Define styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="E7EFF9", end_color="E7EFF9", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Header row
headers1 = ["场景类型", "角色定义", "背景说明", "期望输出", "质量标准", "追问策略"]

# Data rows
data1 = [
    [
        "课程导入场景",
        "你是一位经验丰富的学习顾问，擅长通过提问了解学员的学习目标和当前水平。",
        "学员刚进入课程，需要了解学习路径和课程概览。",
        "清晰的学习路径建议和课程模块介绍。",
        "建议具体、可操作、符合学员水平。",
        '当学员说"我不知道从哪里开始"时，追问其职业背景和学习目的。'
    ],
    [
        "知识点讲解场景",
        "你是一位专业知识深厚的导师，擅长用通俗语言解释复杂概念。",
        "学员对某个概念理解困难，需要深入浅出的解释。",
        "清晰的概念解释+生活化类比+实际应用举例。",
        "类比贴切、举例实用、逻辑清晰。",
        '当学员表示"还是不太懂"时，用更简单的类比重新解释。'
    ],
    [
        "案例分析场景",
        "你是一位实战派专家，擅长从案例中提炼方法和规律。",
        "学员学习了理论知识，需要通过案例加深理解。",
        "案例分析框架+关键洞察+可迁移的方法。",
        "分析深入、洞察独到、方法可复用。",
        '当学员分享自己案例时，追问"这个案例和标准案例有什么异同"。'
    ],
    [
        "实践指导场景",
        "你是一位耐心的实践导师，擅长在学员实践过程中提供及时反馈。",
        "学员在完成实践任务，需要得到具体反馈和改进建议。",
        "具体可行的改进建议+鼓励性评价。",
        "反馈具体、建议可行、态度积极。",
        '当学员提交作业时，追问"你在做这个任务时遇到的最大挑战是什么"。'
    ],
    [
        "总结复盘场景",
        "你是一位善于引导反思的导师，擅长帮助学员梳理学习收获。",
        "学员完成一个阶段学习，需要总结和规划下一步。",
        "学习收获总结+能力提升点+下一步行动计划。",
        "总结全面、提升明确、计划具体。",
        '当学员说"我学完了"时，追问"你觉得最有价值的一个收获是什么"。'
    ]
]

# Write headers
for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_alignment

# Write data
for row_idx, row_data in enumerate(data1, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = left_alignment

# Set column widths
col_widths1 = [15, 30, 25, 25, 20, 30]
for col, width in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(col)].width = width

# Freeze first row
ws1.freeze_panes = "A2"

# ========== Sheet 2: 追问链设计 ==========
ws2 = wb.create_sheet(title="追问链设计")

headers2 = ["层级", "问题类型", "示例问题", "触发条件"]

data2 = [
    ["第1层", "确认理解", "你说的XXX是指...对吗？", "学员表达模糊概念时"],
    ["第2层", "深入原因", "为什么你觉得...？", "学员给出判断但未说明理由"],
    ["第3层", "举一反三", "这个道理还能用在什么地方？", "学员理解原理后"],
    ["第4层", "批判反思", "这个观点有什么局限性？", "学员表达较坚定观点时"],
    ["第5层", "行动计划", "你打算怎么应用今天学的？", "学员表示学有所获时"]
]

# Write headers
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_alignment

# Write data
for row_idx, row_data in enumerate(data2, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = left_alignment

# Set column widths
col_widths2 = [10, 15, 40, 30]
for col, width in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(col)].width = width

# Freeze first row
ws2.freeze_panes = "A2"

# Save workbook
output_path = "D:/新课开发/企业大学/对内/5.AI导师与专属课程设计：从标准课程到AI可交互的学习体验/05-全流程工具表单/F2_Prompt设计模板卡.xlsx"
wb.save(output_path)
print(f"Excel file created successfully at: {output_path}")

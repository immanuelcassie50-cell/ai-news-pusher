from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# Define styles
header_font = Font(name='Microsoft YaHei', size=14, bold=True, color='FFFFFF')
title_font = Font(name='Microsoft YaHei', size=12, bold=True)
normal_font = Font(name='Microsoft YaHei', size=11)
small_font = Font(name='Microsoft YaHei', size=10)

header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
light_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
input_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def add_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start+i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def add_input_cell(ws, row, col, value='', height=30):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = input_fill
    cell.border = thin_border
    cell.alignment = left_align
    ws.row_dimensions[row].height = height
    return cell

def add_normal_cell(ws, row, col, value='', height=20):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = thin_border
    cell.alignment = left_align
    ws.row_dimensions[row].height = height
    return cell

# ========== Sheet 1: 目录封面 ==========
ws1 = wb.active
ws1.title = '目录封面'

# Set column widths
for i in range(1, 8):
    set_col_width(ws1, i, 18)

# Title area
ws1.merge_cells('B2:G2')
cell = ws1['B2']
cell.value = '异步沟通效率最大化'
cell.font = Font(name='Microsoft YaHei', size=24, bold=True, color='2F5496')
cell.alignment = center_align

ws1.merge_cells('B3:G3')
cell = ws1['B3']
cell.value = '工具表单集'
cell.font = Font(name='Microsoft YaHei', size=18, bold=True, color='2F5496')
cell.alignment = center_align

ws1.merge_cells('B4:G4')
cell = ws1['B4']
cell.value = '基于《异步沟通效率最大化》课程内容编制'
cell.font = Font(name='Microsoft YaHei', size=12, italic=True)
cell.alignment = center_align

# Table of contents
row = 6
ws1.cell(row=row, column=2, value='序号').font = title_font
ws1.cell(row=row, column=3, value='表单名称').font = title_font
ws1.cell(row=row, column=4, value='用途说明').font = title_font
ws1.cell(row=row, column=5, value='使用场景').font = title_font

for c in range(2, 6):
    ws1.cell(row=row, column=c).fill = light_fill
    ws1.cell(row=row, column=c).border = thin_border
    ws1.cell(row=row, column=c).alignment = center_align

contents = [
    ('1', '决策请求文档', 'Decision Request', '请求上级或团队做出决策时使用'),
    ('2', '决策记录表', 'Decision Log', '记录已完成的决策，供追溯和复盘'),
    ('3', '响应节律约定表', 'Response SLA', '约定不同紧急程度消息的响应时间'),
    ('4', '任务认领卡', 'Task Ownership Card', '明确任务负责人和完成标准'),
    ('5', '跨时区交接文档', 'Handoff Note', '跨时区协作时的信息交接'),
    ('6', '团队契约表', 'Team Contract', '团队共同约定的协作规则'),
    ('7', '课程评估表', 'Course Evaluation', '课程学习效果评估'),
]

for i, (num, name, en_name, desc) in enumerate(contents):
    r = row + 1 + i
    ws1.cell(row=r, column=2, value=num).alignment = center_align
    ws1.cell(row=r, column=3, value=name + '\n' + en_name).alignment = center_align
    ws1.cell(row=r, column=4, value=desc).alignment = left_align
    ws1.cell(row=r, column=5, value='').alignment = left_align
    for c in range(2, 6):
        ws1.cell(row=r, column=c).border = thin_border

# Footer
ws1.merge_cells('B16:G16')
cell = ws1['B16']
cell.value = '使用方法：每个表单独立成Sheet，可直接打印或填写电子版使用'
cell.font = Font(name='Microsoft YaHei', size=10, italic=True)
cell.alignment = center_align

# ========== Sheet 2: 决策请求文档 ==========
ws2 = wb.create_sheet('决策请求文档')

for i in range(1, 6):
    set_col_width(ws2, i, 20 if i == 5 else 18)

# Header
ws2.merge_cells('A1:E1')
cell = ws2['A1']
cell.value = '决策请求文档  Decision Request'
cell.font = Font(name='Microsoft YaHei', size=16, bold=True, color='2F5496')
cell.alignment = center_align
ws2.row_dimensions[1].height = 40

# Fields
fields = [
    ('标题', '【待决策】具体事项名称', 35),
    ('背景', '用三到五句话说明这件事的来龙去脉，只写读者需要知道的部分，不需要还原全部过程。', 80),
    ('已排除的选项', '列出考虑过但排除的方案，以及排除的理由，让读者不必重复思考这些路径。', 80),
    ('建议方案', '明确写出提出人建议采用的方案，以及支持这个建议的核心理由。', 80),
    ('需要你做的事', '明确写出希望读者做什么——是批准、是提出异议、还是补充信息，并给出具体的截止时间。', 60),
    ('默认生效规则', '写明如果在截止时间前没有收到反对意见，方案将按建议自动生效，避免沉默被误解为拖延。', 60),
]

row = 3
for field, desc, height in fields:
    ws2.cell(row=row, column=1, value=field).font = title_font
    ws2.cell(row=row, column=1).fill = light_fill
    ws2.cell(row=row, column=1).border = thin_border
    ws2.cell(row=row, column=1).alignment = center_align

    ws2.cell(row=row, column=2, value=desc).font = small_font
    ws2.cell(row=row, column=2).alignment = left_align
    ws2.cell(row=row, column=2).border = thin_border
    ws2.merge_cells(f'B{row}:D{row}')

    cell = ws2.cell(row=row, column=5)
    cell.fill = input_fill
    cell.border = thin_border

    ws2.row_dimensions[row].height = height
    row += 1

# Auto-fill instructions
row += 1
ws2.merge_cells(f'A{row}:E{row}')
cell = ws2[f'A{row}']
cell.value = '自动填表说明'
cell.font = title_font
cell.fill = light_fill
cell.border = thin_border

instructions = [
    '1. 标题：格式为【待决策】+具体事项，简明扼要',
    '2. 背景：说明来龙去脉和关键信息，3-5句话为佳',
    '3. 已排除选项：列出被否决的方案及理由，体现决策质量',
    '4. 建议方案：清晰写出建议方案和核心理由',
    '5. 需要你做的事：明确动作+截止时间',
    '6. 默认生效：说明无反对意见时的自动生效规则',
]
for i, instr in enumerate(instructions):
    r = row + 1 + i
    ws2.merge_cells(f'A{r}:E{r}')
    ws2.cell(row=r, column=1, value=instr).font = small_font
    ws2.cell(row=r, column=1).alignment = left_align

# ========== Sheet 3: 决策记录表 ==========
ws3 = wb.create_sheet('决策记录表')

for i in range(1, 7):
    set_col_width(ws3, i, 18 if i < 6 else 22)

# Header
ws3.merge_cells('A1:F1')
cell = ws3['A1']
cell.value = '决策记录表  Decision Log'
cell.font = Font(name='Microsoft YaHei', size=16, bold=True, color='2F5496')
cell.alignment = center_align
ws3.row_dimensions[1].height = 40

# Main fields
row = 3
main_fields = [
    ('决策事项', '一句话写清楚决定了什么'),
    ('决策人', '具体到姓名，不写团队或部门名称'),
    ('决策时间', '具体日期，格式：YYYY-MM-DD'),
]

add_header_row(ws3, row, ['字段', '内容说明', '填写值', '', '', ''])
ws3.merge_cells(f'C{row}:F{row}')
row += 1

for field, desc in main_fields:
    ws3.cell(row=row, column=1, value=field).font = title_font
    ws3.cell(row=row, column=1).fill = light_fill
    ws3.cell(row=row, column=1).border = thin_border
    ws3.cell(row=row, column=1).alignment = center_align

    ws3.cell(row=row, column=2, value=desc).font = small_font
    ws3.cell(row=row, column=2).alignment = left_align
    ws3.cell(row=row, column=2).border = thin_border
    ws3.merge_cells(f'B{row}:B{row}')

    for c in range(3, 7):
        cell = ws3.cell(row=row, column=c)
        cell.fill = input_fill
        cell.border = thin_border

    ws3.row_dimensions[row].height = 30
    row += 1

# Excluded options
ws3.cell(row=row, column=1, value='排除的选项\n及理由').font = title_font
ws3.cell(row=row, column=1).fill = light_fill
ws3.cell(row=row, column=1).border = thin_border
ws3.cell(row=row, column=1).alignment = center_align

ws3.cell(row=row, column=2, value='简要列出未被采纳的方案及原因').font = small_font
ws3.cell(row=row, column=2).alignment = left_align
ws3.cell(row=row, column=2).border = thin_border
ws3.merge_cells(f'B{row}:F{row}')
for c in range(2, 7):
    ws3.cell(row=row, column=c).border = thin_border
ws3.row_dimensions[row].height = 60
row += 1

# Re-discussion conditions
ws3.cell(row=row, column=1, value='触发重议\n的条件').font = title_font
ws3.cell(row=row, column=1).fill = light_fill
ws3.cell(row=row, column=1).border = thin_border
ws3.cell(row=row, column=1).alignment = center_align

ws3.cell(row=row, column=2, value='写明在什么情况下这个决定会被重新讨论').font = small_font
ws3.cell(row=row, column=2).alignment = left_align
ws3.merge_cells(f'B{row}:F{row}')
for c in range(2, 7):
    ws3.cell(row=row, column=c).border = thin_border
ws3.row_dimensions[row].height = 50
row += 1

# Version tracking
row += 1
ws3.merge_cells(f'A{row}:F{row}')
cell = ws3[f'A{row}']
cell.value = '版本追踪'
cell.font = title_font
cell.fill = light_fill
cell.border = thin_border

row += 1
add_header_row(ws3, row, ['版本', '日期', '修订人', '修订内容', '', ''])
ws3.merge_cells(f'D{row}:F{row}')

for i in range(1, 4):
    r = row + i
    for c in range(1, 7):
        cell = ws3.cell(row=r, column=c)
        cell.fill = input_fill if c < 4 else light_fill
        cell.border = thin_border
    ws3.row_dimensions[r].height = 25

# ========== Sheet 4: 响应节律约定表 ==========
ws4 = wb.create_sheet('响应节律约定表')

for i in range(1, 6):
    set_col_width(ws4, i, 20)

# Header
ws4.merge_cells('A1:E1')
cell = ws4['A1']
cell.value = '响应节律约定表  Response SLA'
cell.font = Font(name='Microsoft YaHei', size=16, bold=True, color='2F5496')
cell.alignment = center_align
ws4.row_dimensions[1].height = 40

# Three levels
row = 3
levels = [
    ('紧急且阻塞他人', '1小时', '消息被标注为紧急，且明确写出具体阻塞后果', ''),
    ('一般性讨论', '当天工作时间内', '工作讨论、问题咨询、请求反馈', ''),
    ('非紧急分享与建议', '2个工作日内', '想法分享、建议提出，可不逐条回复', ''),
]

add_header_row(ws4, row, ['紧急程度', '响应窗口', '判断标准', '状态'])
ws4.row_dimensions[row].height = 30

row += 1
for level, window, standard, status in levels:
    ws4.cell(row=row, column=1, value=level).font = title_font
    ws4.cell(row=row, column=1).border = thin_border
    ws4.cell(row=row, column=1).alignment = center_align

    ws4.cell(row=row, column=2, value=window).font = normal_font
    ws4.cell(row=row, column=2).border = thin_border
    ws4.cell(row=row, column=2).alignment = center_align

    ws4.cell(row=row, column=3, value=standard).font = small_font
    ws4.cell(row=row, column=3).border = thin_border
    ws4.cell(row=row, column=3).alignment = left_align

    cell = ws4.cell(row=row, column=4, value=status)
    cell.fill = input_fill
    cell.border = thin_border
    cell.alignment = center_align

    ws4.row_dimensions[row].height = 35
    row += 1

# Online status note
row += 1
ws4.merge_cells(f'A{row}:E{row}')
cell = ws4[f'A{row}']
cell.value = '在线状态说明'
cell.font = title_font
cell.fill = light_fill
cell.border = thin_border

row += 1
ws4.merge_cells(f'A{row}:E{row}')
cell = ws4[f'A{row}']
cell.value = '"已读"代表消息已被看到，不代表已被处理，团队成员不得因对方已读未回而进行负面揣测。\n当预计无法在约定响应窗口内给出完整答复时，鼓励发送一句简短的状态说明。'
cell.font = small_font
cell.alignment = left_align
ws4.row_dimensions[row].height = 50

# ========== Sheet 5: 任务认领卡 ==========
ws5 = wb.create_sheet('任务认领卡')

for i in range(1, 7):
    set_col_width(ws5, i, 18 if i < 6 else 20)

# Header
ws5.merge_cells('A1:F1')
cell = ws5['A1']
cell.value = '任务认领卡  Task Ownership Card'
cell.font = Font(name='Microsoft YaHei', size=16, bold=True, color='2F5496')
cell.alignment = center_align
ws5.row_dimensions[1].height = 40

# Fields
row = 3
fields = [
    ('任务名称', '具体描述，避免使用模糊词汇'),
    ('负责人', '具体姓名，不写团队或部门'),
    ('认领确认', '负责人需回复"我来负责此事"方视为正式认领'),
    ('完成标准', '写清楚做到什么程度算完成，谁来验收'),
    ('预计完成时间', '具体日期，如有调整需主动更新并说明原因'),
]

add_header_row(ws5, row, ['字段', '填写说明', '填写值', '', '', ''])
ws5.merge_cells(f'C{row}:F{row}')
row += 1

for field, desc in fields:
    ws5.cell(row=row, column=1, value=field).font = title_font
    ws5.cell(row=row, column=1).fill = light_fill
    ws5.cell(row=row, column=1).border = thin_border
    ws5.cell(row=row, column=1).alignment = center_align

    ws5.cell(row=row, column=2, value=desc).font = small_font
    ws5.cell(row=row, column=2).border = thin_border
    ws5.cell(row=row, column=2).alignment = left_align
    ws5.merge_cells(f'B{row}:B{row}')

    for c in range(3, 7):
        cell = ws5.cell(row=row, column=c)
        cell.fill = input_fill
        cell.border = thin_border

    ws5.row_dimensions[row].height = 35
    row += 1

# Status tracking
row += 1
ws5.merge_cells(f'A{row}:F{row}')
cell = ws5[f'A{row}']
cell.value = '状态追踪'
cell.font = title_font
cell.fill = light_fill
cell.border = thin_border

row += 1
add_header_row(ws5, row, ['序号', '更新时间', '状态', '说明', '', ''])
ws5.merge_cells(f'D{row}:F{row}')

for i in range(1, 5):
    r = row + i
    ws5.cell(row=r, column=1, value=i).alignment = center_align
    for c in range(1, 7):
        cell = ws5.cell(row=r, column=c)
        cell.fill = input_fill if c < 4 else light_fill
        cell.border = thin_border
    ws5.row_dimensions[r].height = 25

# ========== Sheet 6: 跨时区交接文档 ==========
ws6 = wb.create_sheet('跨时区交接文档')

for i in range(1, 6):
    set_col_width(ws6, i, 22 if i == 5 else 18)

# Header
ws6.merge_cells('A1:E1')
cell = ws6['A1']
cell.value = '跨时区交接文档  Handoff Note'
cell.font = Font(name='Microsoft YaHei', size=16, bold=True, color='2F5496')
cell.alignment = center_align
ws6.row_dimensions[1].height = 40

# Fields
row = 3
fields = [
    ('当前进展', '交接前已完成的部分，以及正在进行中的部分'),
    ('待处理事项', '接手方需要继续推进的具体内容，按优先级排列'),
    ('已知风险\n与阻塞点', '交接前已发现但尚未解决的问题，避免接手方重复踩坑'),
    ('联系方式\n与紧急升级路径', '如遇到真正紧急、无法等待正常响应窗口的情况，说明该联系谁、通过什么渠道'),
]

add_header_row(ws6, row, ['字段', '填写说明', '填写内容', '', ''])
ws6.merge_cells(f'C{row}:E{row}')
row += 1

for field, desc in fields:
    ws6.cell(row=row, column=1, value=field).font = title_font
    ws6.cell(row=row, column=1).fill = light_fill
    ws6.cell(row=row, column=1).border = thin_border
    ws6.cell(row=row, column=1).alignment = center_align

    ws6.cell(row=row, column=2, value=desc).font = small_font
    ws6.cell(row=row, column=2).border = thin_border
    ws6.cell(row=row, column=2).alignment = left_align
    ws6.merge_cells(f'B{row}:B{row}')

    for c in range(3, 6):
        cell = ws6.cell(row=row, column=c)
        cell.fill = input_fill
        cell.border = thin_border

    ws6.row_dimensions[row].height = 50
    row += 1

# ========== Sheet 7: 团队契约表 ==========
ws7 = wb.create_sheet('团队契约表')

for i in range(1, 6):
    set_col_width(ws7, i, 22 if i == 5 else 18)

# Header
ws7.merge_cells('A1:E1')
cell = ws7['A1']
cell.value = '团队契约表  Team Contract'
cell.font = Font(name='Microsoft YaHei', size=16, bold=True, color='2F5496')
cell.alignment = center_align
ws7.row_dimensions[1].height = 40

# Contract sections
row = 3
sections = [
    ('响应节律',
     '紧急且阻塞：1小时内回应\n一般性讨论：当天工作时间内\n非紧急分享：2个工作日内',
     '□ 我已阅读并同意遵守'),
    ('决策留痕',
     '涉及资源分配、对外承诺、跨部门责任的決定，24小时内补写决策记录\n未落地为文字记录的口头决定，视为尚未正式生效',
     '□ 我已阅读并同意遵守'),
    ('责任认领',
     '任务负责人字段必须写具体姓名，禁止填写团队或部门名称\n任务未经负责人明确回复认领，视为无人认领',
     '□ 我已阅读并同意遵守'),
    ('沉默与在线',
     '"已读"仅代表消息已被看到，不代表已被处理\n鼓励发送简短状态说明代替完全沉默',
     '□ 我已阅读并同意遵守'),
    ('会议约定',
     '发起会议前须先自问：这件事是否可以用文档加异步讨论解决\n会议仅用于需要临场情绪判断、真实分歧碰撞、或多方利益博弈的事项',
     '□ 我已阅读并同意遵守'),
]

add_header_row(ws7, row, ['约定类别', '具体内容', '确认签署', '', ''])
ws7.merge_cells(f'C{row}:E{row}')
row += 1

for section, content, confirm in sections:
    ws7.cell(row=row, column=1, value=section).font = title_font
    ws7.cell(row=row, column=1).fill = light_fill
    ws7.cell(row=row, column=1).border = thin_border
    ws7.cell(row=row, column=1).alignment = center_align

    ws7.cell(row=row, column=2, value=content).font = small_font
    ws7.cell(row=row, column=2).border = thin_border
    ws7.cell(row=row, column=2).alignment = left_align
    ws7.merge_cells(f'B{row}:B{row}')

    for c in range(3, 6):
        cell = ws7.cell(row=row, column=c)
        if c == 3:
            cell.value = confirm
            cell.font = small_font
        cell.border = thin_border
        cell.alignment = center_align

    ws7.row_dimensions[row].height = 60
    row += 1

# Signatures
row += 1
ws7.merge_cells(f'A{row}:E{row}')
cell = ws7[f'A{row}']
cell.value = '团队签署区'
cell.font = title_font
cell.fill = light_fill
cell.border = thin_border

row += 1
add_header_row(ws7, row, ['序号', '姓名', '职位/部门', '确认日期', '备注'])
ws7.merge_cells(f'D{row}:E{row}')

for i in range(1, 8):
    r = row + i
    ws7.cell(row=r, column=1, value=i).alignment = center_align
    for c in range(1, 6):
        cell = ws7.cell(row=r, column=c)
        cell.fill = input_fill if 2 <= c <= 4 else light_fill
        cell.border = thin_border
        cell.alignment = center_align
    ws7.row_dimensions[r].height = 25

# ========== Sheet 8: 课程评估表 ==========
ws8 = wb.create_sheet('课程评估表')

for i in range(1, 6):
    set_col_width(ws8, i, 20)

# Header
ws8.merge_cells('A1:E1')
cell = ws8['A1']
cell.value = '课程评估表  Course Evaluation'
cell.font = Font(name='Microsoft YaHei', size=16, bold=True, color='2F5496')
cell.alignment = center_align
ws8.row_dimensions[1].height = 40

# Basic info
row = 3
ws8.cell(row=row, column=1, value='课程名称').font = title_font
ws8.cell(row=row, column=1).fill = light_fill
ws8.cell(row=row, column=1).border = thin_border
ws8.cell(row=row, column=1).alignment = center_align
ws8.merge_cells(f'B{row}:C{row}')
ws8.cell(row=row, column=2, value='异步沟通效率最大化').border = thin_border
ws8.cell(row=row, column=2).alignment = left_align
ws8.merge_cells(f'B{row}:C{row}')

row += 1
ws8.cell(row=row, column=1, value='学员姓名').font = title_font
ws8.cell(row=row, column=1).fill = light_fill
ws8.cell(row=row, column=1).border = thin_border
ws8.cell(row=row, column=1).alignment = center_align
add_input_cell(ws8, row, 2)
ws8.cell(row=row, column=3).fill = input_fill

row += 1
ws8.cell(row=row, column=1, value='评估日期').font = title_font
ws8.cell(row=row, column=1).fill = light_fill
ws8.cell(row=row, column=1).border = thin_border
ws8.cell(row=row, column=1).alignment = center_align
add_input_cell(ws8, row, 2)
ws8.cell(row=row, column=3).fill = input_fill

# Self-assessment
row += 2
ws8.merge_cells(f'A{row}:E{row}')
cell = ws8[f'A{row}']
cell.value = '学员自评'
cell.font = title_font
cell.fill = light_fill
cell.border = thin_border

row += 1
add_header_row(ws8, row, ['评估项目', '内容说明', '评分(1-5)', '自评要点', ''])
ws8.merge_cells(f'D{row}:E{row}')

self_assess = [
    ('理解程度', '对异步沟通核心原则的理解'),
    ('应用意愿', '愿意在工作中应用这些方法'),
    ('实践计划', '有明确的实践计划和场景'),
]

for field, desc in self_assess:
    r = row + 1
    ws8.cell(row=r, column=1, value=field).font = title_font
    ws8.cell(row=r, column=1).fill = light_fill
    ws8.cell(row=r, column=1).border = thin_border
    ws8.cell(row=r, column=1).alignment = center_align

    ws8.cell(row=r, column=2, value=desc).font = small_font
    ws8.cell(row=r, column=2).border = thin_border

    add_input_cell(ws8, r, 3)

    ws8.cell(row=r, column=4).fill = input_fill
    ws8.cell(row=r, column=4).border = thin_border
    ws8.merge_cells(f'D{r}:E{r}')

    ws8.row_dimensions[r].height = 30
    row = r

# Instructor assessment
row += 2
ws8.merge_cells(f'A{row}:E{row}')
cell = ws8[f'A{row}']
cell.value = '讲师评估'
cell.font = title_font
cell.fill = light_fill
cell.border = thin_border

row += 1
add_header_row(ws8, row, ['评估项目', '内容说明', '评分(1-5)', '讲师评语', ''])
ws8.merge_cells(f'D{row}:E{row}')

instructor_assess = [
    ('参与度', '课堂参与和讨论的积极性'),
    ('理解准确度', '对关键概念的把握程度'),
    ('实践潜力', '将方法应用到实际工作的可能性'),
]

for field, desc in instructor_assess:
    r = row + 1
    ws8.cell(row=r, column=1, value=field).font = title_font
    ws8.cell(row=r, column=1).fill = light_fill
    ws8.cell(row=r, column=1).border = thin_border
    ws8.cell(row=r, column=1).alignment = center_align

    ws8.cell(row=r, column=2, value=desc).font = small_font
    ws8.cell(row=r, column=2).border = thin_border

    add_input_cell(ws8, r, 3)

    ws8.cell(row=r, column=4).fill = input_fill
    ws8.cell(row=r, column=4).border = thin_border
    ws8.merge_cells(f'D{r}:E{r}')

    ws8.row_dimensions[r].height = 30
    row = r

# Course satisfaction
row += 2
ws8.merge_cells(f'A{row}:E{row}')
cell = ws8[f'A{row}']
cell.value = '课程满意度'
cell.font = title_font
cell.fill = light_fill
cell.border = thin_border

row += 1
add_header_row(ws8, row, ['评估维度', '内容说明', '评分(1-5)', '', ''])
ws8.merge_cells(f'C{row}:E{row}')

satisfaction = [
    ('内容实用性', '课程内容对实际工作的帮助程度'),
    ('讲解清晰度', '讲师表达和案例讲解的清晰程度'),
    ('互动体验', '课堂互动和讨论的体验'),
    ('整体推荐度', '愿意推荐给同事的程度'),
]

for field, desc in satisfaction:
    r = row + 1
    ws8.cell(row=r, column=1, value=field).font = title_font
    ws8.cell(row=r, column=1).fill = light_fill
    ws8.cell(row=r, column=1).border = thin_border
    ws8.cell(row=r, column=1).alignment = center_align

    ws8.cell(row=r, column=2, value=desc).font = small_font
    ws8.cell(row=r, column=2).border = thin_border
    ws8.merge_cells(f'B{r}:B{r}')

    add_input_cell(ws8, r, 3)

    ws8.cell(row=r, column=4).fill = input_fill
    ws8.cell(row=r, column=4).border = thin_border
    ws8.merge_cells(f'D{r}:E{r}')

    ws8.row_dimensions[r].height = 30
    row = r

# Suggestions
row += 2
ws8.merge_cells(f'A{row}:E{row}')
cell = ws8[f'A{row}']
cell.value = '改进建议'
cell.font = title_font
cell.fill = light_fill
cell.border = thin_border

row += 1
ws8.merge_cells(f'A{row}:E{row}')
cell = ws8[f'A{row}']
cell.value = ''
cell.fill = input_fill
cell.border = thin_border
ws8.row_dimensions[row].height = 60

# Save
import os
output_dir = 'D:/新课开发/工作手册/异步沟通效率最大化/完整课程包/008-工具表单'
os.makedirs(output_dir, exist_ok=True)
output_path = output_dir + '/异步沟通效率最大化-工具表单集.xlsx'
wb.save(output_path)
print(f'Excel file saved to: {output_path}')

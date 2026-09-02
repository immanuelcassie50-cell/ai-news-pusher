"""Build 05_学员追踪表.xlsx for the course.

4 sheets:
- Sheet1: 学习进度追踪 (6 stages + 18 lessons)
- Sheet2: 利益方图谱 (7 stakeholders)
- Sheet3: 30问追问题库 (with 有效性自评)
- Sheet4: 质量检验清单 (20 items, 4 layers, with auto-汇总)
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle,
)
from openpyxl.formatting.rule import (
    CellIsRule, FormulaRule, ColorScaleRule,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "D:/2026年课程/ai课2026整理/经营型企业大学构建：从培训部门到战略引擎/成果demo/05_学员追踪表.xlsx"

# Styles ---------------------------------------------------------------
THIN = Side(border_style="thin", color="9E9E9E")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F3A5F")
SUBTITLE_FONT = Font(name="Arial", size=11, bold=True, color="1F3A5F")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
ROW_FILL_ALT = PatternFill("solid", fgColor="F4F6FA")
ROW_FILL_NOTE = PatternFill("solid", fgColor="FFF8E1")
ROW_FILL_SUMMARY = PatternFill("solid", fgColor="E8F5E9")

# Status colors
FILL_DONE = PatternFill("solid", fgColor="C8E6C9")  # green
FILL_PROG = PatternFill("solid", fgColor="FFF59D")  # yellow
FILL_TODO = PatternFill("solid", fgColor="FFCDD2")  # red
FILL_NA = PatternFill("solid", fgColor="ECEFF1")  # gray

DEFAULT_FONT = Font(name="Arial", size=10)

wb = Workbook()

# ============== Sheet 1: 学习进度追踪 ==============
ws1 = wb.active
ws1.title = "1.学习进度追踪"

# Title
ws1["A1"] = "《经营型企业大学构建》学员学习进度追踪"
ws1["A1"].font = TITLE_FONT
ws1.merge_cells("A1:H1")
ws1["A1"].alignment = CENTER

ws1["A2"] = "填写说明：状态栏选「已完成/进行中/未开始/不适用」；每月底更新一次"
ws1["A2"].font = Font(name="Arial", size=9, italic=True, color="757575")
ws1.merge_cells("A2:H2")
ws1["A2"].alignment = CENTER

# Headers
headers1 = [
    "阶段", "阶段名称", "讲次", "讲次名称", "完成日期",
    "状态", "3锚点评分(1-5)", "里程碑成果",
]
for col, h in enumerate(headers1, 1):
    c = ws1.cell(row=4, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

# 6 stages, 18 lessons
stages = [
    ("A", "识别力", [
        (1, "发刊词：为什么企业需要大学？", "课程承诺书"),
        (2, "先导课：经营型企业大学到底是什么？", "传统型vs经营型自评表"),
    ]),
    ("B", "分析力", [
        (3, "问题一：三要素模型", "三要素自评"),
        (4, "问题二：培训部门现状诊断", "培训部门诊断表"),
    ]),
    ("C", "技法（行动系统）", [
        (5, "问题三：120天行动系统总览", "120天日历初稿"),
        (6, "问题四：第一阶段诊断与优先级", "优先级排序表"),
        (7, "问题五：高价值低难度10场景", "10场景选择"),
    ]),
    ("D", "创作力（商业模式+试点）", [
        (8, "问题六：第二阶段试点执行", "试点执行表"),
        (9, "问题七：商业模式5种变现", "商业模式设计表"),
        (10, "问题八：第三阶段效果评估", "三维价值评估矩阵"),
        (11, "问题九：企业大学架构设计", "架构设计模板"),
        (12, "问题十：课程体系建设", "课程体系设计表"),
    ]),
    ("E", "系统设计（师资+服务+评估）", [
        (13, "问题十一：师资体系建设", "讲师能力评估表"),
        (14, "问题十二：业务部门买单机制", "内部收费设计表"),
        (15, "问题十三：对外服务体系", "对外服务设计表"),
    ]),
    ("F", "综合检验", [
        (16, "问题十四：评估体系ROI", "评估体系设计表"),
        (17, "问题十五：失败案例避坑", "失败原因自查表"),
        (18, "问题十六：标杆案例+结刊词", "120天行动承诺书"),
    ]),
]

row = 5
for stage_code, stage_name, lessons in stages:
    start_row = row
    for i, (lesson_no, lesson_name, milestone) in enumerate(lessons):
        r = row
        if i == 0:
            ws1.cell(row=r, column=1, value=stage_code).alignment = CENTER
            ws1.cell(row=r, column=2, value=stage_name).alignment = CENTER
        ws1.cell(row=r, column=3, value=lesson_no).alignment = CENTER
        ws1.cell(row=r, column=4, value=lesson_name).alignment = LEFT_WRAP
        ws1.cell(row=r, column=5, value="").alignment = CENTER
        ws1.cell(row=r, column=6, value="未开始").alignment = CENTER
        ws1.cell(row=r, column=7, value="").alignment = CENTER
        ws1.cell(row=r, column=8, value=milestone).alignment = LEFT_WRAP
        for col in range(1, 9):
            cell = ws1.cell(row=r, column=col)
            cell.border = BORDER
            cell.font = DEFAULT_FONT
            if r % 2 == 0:
                cell.fill = ROW_FILL_ALT
        row += 1
    # Merge stage cells
    if len(lessons) > 1:
        ws1.merge_cells(
            start_row=start_row, end_row=row - 1,
            start_column=1, end_column=1,
        )
        ws1.merge_cells(
            start_row=start_row, end_row=row - 1,
            start_column=2, end_column=2,
        )

# Data validation for status
status_dv = DataValidation(
    type="list",
    formula1='"未开始,进行中,已完成,不适用"',
    allow_blank=True,
)
ws1.add_data_validation(status_dv)
status_dv.add(f"F5:F{row-1}")

# Data validation for 锚点评分
score_dv = DataValidation(
    type="list",
    formula1='"1,2,3,4,5"',
    allow_blank=True,
)
ws1.add_data_validation(score_dv)
score_dv.add(f"G5:G{row-1}")

# Conditional formatting on status
last_row1 = row - 1
ws1.conditional_formatting.add(
    f"F5:F{last_row1}",
    CellIsRule(operator="equal", formula=['"已完成"'],
               fill=FILL_DONE),
)
ws1.conditional_formatting.add(
    f"F5:F{last_row1}",
    CellIsRule(operator="equal", formula=['"进行中"'],
               fill=FILL_PROG),
)
ws1.conditional_formatting.add(
    f"F5:F{last_row1}",
    CellIsRule(operator="equal", formula=['"未开始"'],
               fill=FILL_TODO),
)
ws1.conditional_formatting.add(
    f"F5:F{last_row1}",
    CellIsRule(operator="equal", formula=['"不适用"'],
               fill=FILL_NA),
)

# Summary section
sum_row = last_row1 + 3
ws1.cell(row=sum_row, column=1, value="汇总指标").font = SUBTITLE_FONT
ws1.merge_cells(start_row=sum_row, end_row=sum_row, start_column=1, end_column=2)
sum_row += 1
labels = [
    ("已完成讲次", f'=COUNTIF(F5:F{last_row1},"已完成")'),
    ("进行中讲次", f'=COUNTIF(F5:F{last_row1},"进行中")'),
    ("未开始讲次", f'=COUNTIF(F5:F{last_row1},"未开始")'),
    ("完成率", f'=COUNTIF(F5:F{last_row1},"已完成")/COUNTA(C5:C{last_row1})'),
    ("锚点·让人进入(平均)", f'=AVERAGEIFS(G5:G{last_row1},F5:F{last_row1},"已完成",G5:G{last_row1},">0")'),
    ("锚点·让人冲突(平均)", f'=AVERAGEIFS(G5:G{last_row1},F5:F{last_row1},"已完成",G5:G{last_row1},">0")'),
    ("锚点·让人带走(平均)", f'=AVERAGEIFS(G5:G{last_row1},F5:F{last_row1},"已完成",G5:G{last_row1},">0")'),
]
for i, (label, formula) in enumerate(labels):
    r = sum_row + i
    ws1.cell(row=r, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
    ws1.cell(row=r, column=1).alignment = LEFT_WRAP
    ws1.cell(row=r, column=1).border = BORDER
    ws1.cell(row=r, column=2, value=formula).font = DEFAULT_FONT
    ws1.cell(row=r, column=2).alignment = CENTER
    ws1.cell(row=r, column=2).border = BORDER
    ws1.cell(row=r, column=2).fill = ROW_FILL_SUMMARY
    if label == "完成率":
        ws1.cell(row=r, column=2).number_format = "0.0%"
    else:
        ws1.cell(row=r, column=2).number_format = "0.0"

# Column widths
widths1 = [6, 14, 6, 36, 12, 10, 12, 28]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[1].height = 24

# ============== Sheet 2: 利益方图谱 ==============
ws2 = wb.create_sheet("2.利益方图谱")

ws2["A1"] = "主案例《预算被砍30%》利益方图谱"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:G1")
ws2["A1"].alignment = CENTER

ws2["A2"] = (
    "填写说明：根据主案例分析7个利益方的核心利益、约束条件、"
    "信息差、对我(学员)的期待"
)
ws2["A2"].font = Font(name="Arial", size=9, italic=True, color="757575")
ws2.merge_cells("A2:G2")
ws2["A2"].alignment = CENTER

headers2 = [
    "角色", "姓名/代号", "核心利益(最在意什么)",
    "约束条件(不得不考虑什么)", "独有的信息或视角",
    "对主角(陈冬梅/学员)的期待", "我的应对策略",
]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=4, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

stakeholders = [
    (
        "主角", "陈冬梅（培训主管）",
        "保住团队、保住预算、证明培训价值",
        "没有独立预算权、需向老板汇报",
        "知道老板对'产业链班组长'的潜在兴趣；知道小赵能力强项",
        "——",
        "——",
    ),
    (
        "老板(CEO)", "王建国",
        "营收2.4亿、利润率不降、降低可控成本、对外品牌升级",
        "公开承诺'培训不能停'、Q3董事会要给数字",
        "不知道培训部能算账；知道'隔壁汇川公司培训部去年创收了80万'",
        "期待：'你能不能像隔壁公司那样，让培训部变成创收部门'",
        "主动提案+季度复盘+用数字汇报",
    ),
    (
        "业务副总", "刘建华(生产)",
        "班组长能干活、产线不出问题、Q1订单按时交付",
        "班组长脱产1天=该产线少100件产出(12万损失)",
        "私下说过'现在的班组长不会带人'；知道下游车企审核越来越严",
        "期待：'培训能不能解决班组长能力问题？不要搞虚的'",
        "试点选他的产线+让他站台+让他参与方案设计",
    ),
    (
        "一线班组长", "张磊/王浩/李娟等180人",
        "能升职加薪、有尊严、工作不被机器人替代",
        "工资5000-7000元/月，外面同类岗位高15%；平均年龄42岁",
        "不知道公司要做'对外培训'；知道公司精益是行业最好",
        "期待：'培训能不能帮我管好下面十几个人、让我晋升'",
        "让明星班组长当讲师+给予合理课酬+提供晋升通道",
    ),
    (
        "外部客户(上游)", "永胜机械王总等10家",
        "学到精益降本、解决客户审核临时问题",
        "30-50人中小企业，无人教；预算敏感",
        "听过华驰精益好；王总已主动打电话来问能不能派人辅导",
        "期待：'华驰能不能帮我们也变好？最好是驻企辅导'",
        "驻企辅导为主+开班为辅+口碑传播",
    ),
    (
        "下属", "小赵(培训专员)",
        "保住工作、做出成绩、学到新东西",
        "入职1年，对外没资源、课程开发能力弱",
        "强项是活动组织(去年新员工培训是她独立操盘的)",
        "期待：'陈姐能不能别压太狠？我也想做出成绩'",
        "让她负责A轨招生+给她成长空间+合理分润",
    ),
    (
        "支持部门", "HR总监张明",
        "培训部不出乱子、自己能向老板交代",
        "同时管招聘/薪酬/员工关系，培训只占1/4工作",
        "去年曾建议'砍差旅预算'，和陈冬梅有小摩擦",
        "期待：'陈冬梅别给我添乱就行'",
        "主动汇报+让他参与+争取他的背书",
    ),
]

# Stakeholder colors
fill_map = {
    "主角": PatternFill("solid", fgColor="E1BEE7"),
    "老板(CEO)": PatternFill("solid", fgColor="FFCDD2"),
    "业务副总": PatternFill("solid", fgColor="FFE0B2"),
    "一线班组长": PatternFill("solid", fgColor="C8E6C9"),
    "外部客户(上游)": PatternFill("solid", fgColor="B3E5FC"),
    "下属": PatternFill("solid", fgColor="F8BBD0"),
    "支持部门": PatternFill("solid", fgColor="D1C4E9"),
}

for i, row_data in enumerate(stakeholders):
    r = 5 + i
    for j, val in enumerate(row_data, 1):
        c = ws2.cell(row=r, column=j, value=val)
        c.alignment = LEFT_WRAP
        c.border = BORDER
        c.font = DEFAULT_FONT
        if j == 1:
            c.fill = fill_map.get(val, ROW_FILL_ALT)
            c.font = Font(name="Arial", size=10, bold=True)

# Widths
widths2 = [16, 14, 28, 28, 32, 26, 24]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for r in range(5, 12):
    ws2.row_dimensions[r].height = 60
ws2.row_dimensions[1].height = 24

# ============== Sheet 3: 30问追问题库 ==============
ws3 = wb.create_sheet("3.30问追问题库")

ws3["A1"] = "30问追问题库（按\"三板斧\"分类，各10问）"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:G1")
ws3["A1"].alignment = CENTER

ws3["A2"] = (
    "填写说明：每月底选3-5个问题在课堂上使用，"
    "课后填写「使用次数」和「有效性自评（1-5）」"
)
ws3["A2"].font = Font(name="Arial", size=9, italic=True, color="757575")
ws3.merge_cells("A2:G2")
ws3["A2"].alignment = CENTER

headers3 = [
    "编号", "三板斧类型", "适用场景", "追问内容",
    "使用次数", "有效性自评(1-5)", "备注",
]
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=4, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

# 30 questions - 10 each for 追证据/追假设/追代价
questions = [
    # 追证据
    ("追证据", "学员说方案时",
     "你说陈冬梅应该做产业链培训——你预测能招到15-30人，依据是什么？"),
    ("追证据", "学员预测营收时",
     "你说老板'应该'接受105万换216万——你了解老板过去的决策风格吗？"),
    ("追证据", "学员引用数据时",
     "你说班组长培训能提升良品率——你看过华驰过去3年的良品率数据吗？"),
    ("追证据", "学员评估时间时",
     "案例里说陈冬梅'熬了3个晚上'写方案——你相信这个时间投入吗？"),
    ("追证据", "学员评价客户时",
     "你说'华成精工李总愿意花8万'——案例里只说他是'一个学员'，依据来自哪里？"),
    ("追证据", "学员判断老板时",
     "你说老板对'产业链培训'有兴趣，依据是'曾对标参观过'——参观和'愿意投钱'是两件事"),
    ("追证据", "学员用类比时",
     "你说'某公司也是这样做成的'——那家公司的具体数字是什么？和你公司一样吗？"),
    ("追证据", "学员说'行业都这样'时",
     "'行业都这样'是个很大的判断——你调研过几家？具体数据是什么？"),
    ("追证据", "学员用'我觉得'时",
     "'我觉得'后面如果有数据支持，那才有说服力；纯粹凭直觉的事，能变成决策依据吗？"),
    ("追证据", "学员说'我试过'时",
     "你说'我试过'——试的结果是什么？数据记录了吗？如果没记录，凭什么说'试过'？"),
    # 追假设
    ("追假设", "学员做选择时",
     "你假设了'华驰的精益经验有市场'——这个假设做过验证吗？"),
    ("追假设", "学员讲老板时",
     "你假设了'老板会接受经营型方案'——但如果老板只想要'降本'呢？"),
    ("追假设", "学员做规划时",
     "你假设了'业务部门会配合'——生产副总愿意放班组长出去讲课吗？"),
    ("追假设", "学员算ROI时",
     "你假设了'陈冬梅的ROI 128%是可信的'——但案例里这是她自己测算的，财务会认可吗？"),
    ("追假设", "学员判断人时",
     "你假设了'陈冬梅能扛住压力'——但她6年没承担过这种压力，'扛不住'概率多大？"),
    ("追假设", "学员归因时",
     "你把转型成功归因于'陈冬梅个人努力'——但至少有3个外部因素（老板支持+刘建华站台+王总需求），权重怎么分？"),
    ("追假设", "学员讲文化时",
     "你说'华驰有精益文化'——但'有精益文化'和'愿意把精益教给产业链'是两件事"),
    ("追假设", "学员设计产品时",
     "你假设了'2.5万/期的定价合理'——你做过价格敏感性测试吗？"),
    ("追假设", "学员讲团队时",
     "你假设了'小赵能承担B轨'——但小赵入职1年，对外没资源，她'能做'和'愿意做'是两件事"),
    ("追假设", "学员讲未来时",
     "你假设了'2027年180万能换500万营收'——这个数字是科学测算还是雄心壮志？"),
    # 追代价
    ("追代价", "学员做承诺时",
     "如果产业链试点3个月只招到8个人，105万预算怎么重新分配？陈冬梅要承担什么？"),
    ("追代价", "学员下注时",
     "如果陈冬梅把105万全押产业链失败，她还能保住团队吗？"),
    ("追代价", "学员等老板时",
     "如果老板看完方案说'下季度再说'，陈冬梅的30天窗口期怎么过？"),
    ("追代价", "学员做选择时",
     "如果产业链做成了，但挤压了内部培训——新员工留存率不升反降，这个代价由谁承担？"),
    ("追代价", "学员定目标时",
     "如果2027年500万目标定高了，她明年的压力会多大？定低了老板会不会觉得没冲劲？"),
    ("追代价", "学员讲收益时",
     "如果陈冬梅转型成功，培训部变成'创收部门'——这对她个人是好是坏？创收意味着更高的期待、更大的压力"),
    ("追代价", "学员做取舍时",
     "如果你的方案让小赵和团队加班1个月，你愿意为他们的加班费付费吗？"),
    ("追代价", "学员说'两全'时",
     "你提到'内部和外部都做'——如果资源有限，必须放弃一个，你放弃哪个？为什么？"),
    ("追代价", "学员讲风险时",
     "你说'小风险'——小到多少？3万？30万？300万？小到不影响公司生存？"),
    ("追代价", "学员谈失败时",
     "如果试点失败，学员的方案会被董事会记住多久？3年？5年？这个'声誉代价'你考虑过吗？"),
]

type_fills = {
    "追证据": PatternFill("solid", fgColor="E3F2FD"),
    "追假设": PatternFill("solid", fgColor="FFF3E0"),
    "追代价": PatternFill("solid", fgColor="FFEBEE"),
}

for i, (qtype, scene, content) in enumerate(questions):
    r = 5 + i
    ws3.cell(row=r, column=1, value=i + 1).alignment = CENTER
    ws3.cell(row=r, column=2, value=qtype).alignment = CENTER
    ws3.cell(row=r, column=2).fill = type_fills[qtype]
    ws3.cell(row=r, column=2).font = Font(name="Arial", size=10, bold=True)
    ws3.cell(row=r, column=3, value=scene).alignment = LEFT_WRAP
    ws3.cell(row=r, column=4, value=content).alignment = LEFT_WRAP
    ws3.cell(row=r, column=5, value=0).alignment = CENTER
    ws3.cell(row=r, column=6, value="").alignment = CENTER
    ws3.cell(row=r, column=7, value="").alignment = LEFT_WRAP
    for col in range(1, 8):
        c = ws3.cell(row=r, column=col)
        c.border = BORDER
        if c.font.name != "Arial":
            c.font = DEFAULT_FONT
        if r % 2 == 0 and col not in (2,):
            c.fill = ROW_FILL_ALT

# Data validation for 有效性自评
eff_dv = DataValidation(
    type="list",
    formula1='"1,2,3,4,5"',
    allow_blank=True,
)
ws3.add_data_validation(eff_dv)
eff_dv.add(f"F5:F{4 + len(questions)}")

# 汇总区
last_row3 = 4 + len(questions)
sum_row = last_row3 + 3
ws3.cell(row=sum_row, column=1, value="统计指标").font = SUBTITLE_FONT
ws3.merge_cells(
    start_row=sum_row, end_row=sum_row, start_column=1, end_column=2,
)
sum_row += 1
stats = [
    ("总问题数", f'=COUNTA(D5:D{last_row3})'),
    ("使用总次数", f'=SUM(E5:E{last_row3})'),
    ("已使用问题数(>0次)", f'=COUNTIF(E5:E{last_row3},">0")'),
    ("平均有效性自评", f'=AVERAGEIF(F5:F{last_row3},">0")'),
    ("追证据·已使用", f'=COUNTIFS(B5:B{last_row3},"追证据",E5:E{last_row3},">0")'),
    ("追假设·已使用", f'=COUNTIFS(B5:B{last_row3},"追假设",E5:E{last_row3},">0")'),
    ("追代价·已使用", f'=COUNTIFS(B5:B{last_row3},"追代价",E5:E{last_row3},">0")'),
]
for i, (label, formula) in enumerate(stats):
    r = sum_row + i
    ws3.cell(row=r, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
    ws3.cell(row=r, column=1).alignment = LEFT_WRAP
    ws3.cell(row=r, column=1).border = BORDER
    ws3.cell(row=r, column=2, value=formula).font = DEFAULT_FONT
    ws3.cell(row=r, column=2).alignment = CENTER
    ws3.cell(row=r, column=2).border = BORDER
    ws3.cell(row=r, column=2).fill = ROW_FILL_SUMMARY
    if "平均" in label:
        ws3.cell(row=r, column=2).number_format = "0.00"

widths3 = [6, 10, 18, 56, 10, 14, 16]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
for r in range(5, 5 + len(questions)):
    ws3.row_dimensions[r].height = 36
ws3.row_dimensions[1].height = 24

# ============== Sheet 4: 质量检验清单 ==============
ws4 = wb.create_sheet("4.质量检验清单")

ws4["A1"] = "企业大学方案质量检验清单（4层20项）"
ws4["A1"].font = TITLE_FONT
ws4.merge_cells("A1:F1")
ws4["A1"].alignment = CENTER

ws4["A2"] = (
    "填写说明：每项选「通过/需修改/不适用」；"
    "最右列写具体修改方案"
)
ws4["A2"].font = Font(name="Arial", size=9, italic=True, color="757575")
ws4.merge_cells("A2:F2")
ws4["A2"].alignment = CENTER

headers4 = [
    "层级", "编号", "检验项", "通过标准", "状态", "如需修改具体怎么改",
]
for col, h in enumerate(headers4, 1):
    c = ws4.cell(row=4, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

layers = [
    ("第一层·内容层", [
        "决策点清晰，无法回避",
        "存在至少3种有逻辑支撑的方向",
        "方案结束在决策之前，不剧透",
        "没有'一眼可见的正确答案'",
        "五层结构完整，隐藏层未直说",
    ]),
    ("第二层·经营型语感层", [
        "所有关键业务数字具体且自洽",
        "包含明确的商业模式要素",
        "有意识地设计了商业信息缺口",
        "主角局限性通过行为展示",
        "有明确的截止时间和代价",
    ]),
    ("第三层·讨论性层", [
        "发给3个企业培训负责人，产生了真实分歧",
        "四个视角(CEO/HR/业务/讲师/学员)都能切入",
        "隐藏层可被讨论发现",
        "框架一句话说清楚",
        "冷叫问题满足'三有原则'(有话说/说不完/答案多元)",
    ]),
    ("第四层·Teaching Note层", [
        "五阶段讨论弧完整",
        "追问题库每个核心问题≥5个变体",
        "至少1个转折注入设计",
        "板书演进方案有草图",
        "框架命名时机和方式明确",
    ]),
]

layer_fills = {
    "第一层·内容层": PatternFill("solid", fgColor="E8EAF6"),
    "第二层·经营型语感层": PatternFill("solid", fgColor="E0F7FA"),
    "第三层·讨论性层": PatternFill("solid", fgColor="FFF8E1"),
    "第四层·Teaching Note层": PatternFill("solid", fgColor="F3E5F5"),
}

row = 5
all_start_rows = []
for layer_name, items in layers:
    start = row
    for i, item in enumerate(items):
        r = row
        if i == 0:
            ws4.cell(row=r, column=1, value=layer_name).font = Font(name="Arial", size=10, bold=True)
            ws4.cell(row=r, column=1).fill = layer_fills[layer_name]
        ws4.cell(row=r, column=2, value=len(all_start_rows) + 1).alignment = CENTER
        ws4.cell(row=r, column=3, value=item).alignment = LEFT_WRAP
        # Pass criteria
        criteria_map = {
            "决策点清晰，无法回避": "读者读完能说出主角必须决定什么",
            "存在至少3种有逻辑支撑的方向": "3种以上立场+每种有商业逻辑",
            "方案结束在决策之前，不剧透": "主角仍在纠结中",
            "没有'一眼可见的正确答案'": "学员不会5分钟内达成共识",
            "五层结构完整，隐藏层未直说": "5层都有内容+L4是结构性原因",
            "所有关键业务数字具体且自洽": "营收/毛利/人数/预算 4个数字内部能验证",
            "包含明确的商业模式要素": "学员/定价/渠道/营收预测 至少3个",
            "有意识地设计了商业信息缺口": "至少2处关键商业变量未解释",
            "主角局限性通过行为展示": "至少2处具体行为而非形容词",
            "有明确的截止时间和代价": "截止时刻+每个选项的代价",
            "发给3个企业培训负责人，产生了真实分歧": "至少2种立场分歧",
            "四个视角(CEO/HR/业务/讲师/学员)都能切入": "4个角色都能找到发言角度",
            "隐藏层可被讨论发现": "学员能自发发现而非被告知",
            "框架一句话说清楚": "≤25字+能迁移到其他场景",
            "冷叫问题满足'三有原则'(有话说/说不完/答案多元)": "3个维度各≥2分",
            "五阶段讨论弧完整": "事实/解读/矛盾/取舍/框架 每阶段有出口信号",
            "追问题库每个核心问题≥5个变体": "三板斧各≥5个追问",
            "至少1个转折注入设计": "明确的时机+内容",
            "板书演进方案有草图": "3个阶段板书草图",
            "框架命名时机和方式明确": "具体的信号+具体的命名",
        }
        ws4.cell(row=r, column=4, value=criteria_map.get(item, "")).alignment = LEFT_WRAP
        ws4.cell(row=r, column=5, value="未测试").alignment = CENTER
        ws4.cell(row=r, column=6, value="").alignment = LEFT_WRAP
        for col in range(1, 7):
            c = ws4.cell(row=r, column=col)
            c.border = BORDER
            c.font = DEFAULT_FONT
            if r % 2 == 0 and col != 1:
                c.fill = ROW_FILL_ALT
        row += 1
    # Merge layer name across rows
    if len(items) > 1:
        ws4.merge_cells(
            start_row=start, end_row=row - 1,
            start_column=1, end_column=1,
        )
    all_start_rows.append(start)

# Data validation for status
status4_dv = DataValidation(
    type="list",
    formula1='"通过,需修改,不适用,未测试"',
    allow_blank=True,
)
ws4.add_data_validation(status4_dv)
status4_dv.add(f"E5:E{row-1}")

# Conditional formatting
last_row4 = row - 1
ws4.conditional_formatting.add(
    f"E5:E{last_row4}",
    CellIsRule(operator="equal", formula=['"通过"'],
               fill=FILL_DONE),
)
ws4.conditional_formatting.add(
    f"E5:E{last_row4}",
    CellIsRule(operator="equal", formula=['"需修改"'],
               fill=FILL_TODO),
)
ws4.conditional_formatting.add(
    f"E5:E{last_row4}",
    CellIsRule(operator="equal", formula=['"不适用"'],
               fill=FILL_NA),
)
ws4.conditional_formatting.add(
    f"E5:E{last_row4}",
    CellIsRule(operator="equal", formula=['"未测试"'],
               fill=PatternFill("solid", fgColor="FFF59D")),
)

# Summary section
sum_row = last_row4 + 3
ws4.cell(row=sum_row, column=1, value="通过率汇总").font = SUBTITLE_FONT
ws4.merge_cells(
    start_row=sum_row, end_row=sum_row, start_column=1, end_column=2,
)
sum_row += 1
stats4 = [
    ("检验项总数", f'=COUNTA(C5:C{last_row4})'),
    ("已通过", f'=COUNTIF(E5:E{last_row4},"通过")'),
    ("需修改", f'=COUNTIF(E5:E{last_row4},"需修改")'),
    ("不适用", f'=COUNTIF(E5:E{last_row4},"不适用")'),
    ("未测试", f'=COUNTIF(E5:E{last_row4},"未测试")'),
    ("通过率(通过/已测试)",
     f'=COUNTIF(E5:E{last_row4},"通过")/(COUNTIF(E5:E{last_row4},"通过")+COUNTIF(E5:E{last_row4},"需修改"))'),
    ("总体完成度(已测试/总数)",
     f'=(COUNTIF(E5:E{last_row4},"通过")+COUNTIF(E5:E{last_row4},"需修改"))/COUNTA(C5:C{last_row4})'),
]
for i, (label, formula) in enumerate(stats4):
    r = sum_row + i
    ws4.cell(row=r, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
    ws4.cell(row=r, column=1).alignment = LEFT_WRAP
    ws4.cell(row=r, column=1).border = BORDER
    ws4.cell(row=r, column=2, value=formula).font = DEFAULT_FONT
    ws4.cell(row=r, column=2).alignment = CENTER
    ws4.cell(row=r, column=2).border = BORDER
    ws4.cell(row=r, column=2).fill = ROW_FILL_SUMMARY
    if "率" in label or "度" in label:
        ws4.cell(row=r, column=2).number_format = "0.0%"

# 评级提示
tip_row = sum_row + len(stats4) + 1
ws4.cell(row=tip_row, column=1, value="评级标准").font = SUBTITLE_FONT
ws4.merge_cells(
    start_row=tip_row, end_row=tip_row, start_column=1, end_column=2,
)
tip_row += 1
ratings = [
    ("优秀：", "通过率≥90%且总体完成度=100%"),
    ("合格：", "通过率≥70%且总体完成度≥80%"),
    ("需修改：", "通过率<70%或总体完成度<80%"),
    ("不合格：", "总体完成度<50%"),
]
for i, (lvl, desc) in enumerate(ratings):
    r = tip_row + i
    ws4.cell(row=r, column=1, value=lvl).font = Font(name="Arial", size=10, bold=True)
    ws4.cell(row=r, column=1).alignment = LEFT_WRAP
    ws4.cell(row=r, column=1).border = BORDER
    ws4.cell(row=r, column=2, value=desc).font = DEFAULT_FONT
    ws4.cell(row=r, column=2).alignment = LEFT_WRAP
    ws4.cell(row=r, column=2).border = BORDER
    ws4.cell(row=r, column=2).fill = ROW_FILL_NOTE
    ws4.merge_cells(start_row=r, end_row=r, start_column=2, end_column=6)

widths4 = [20, 6, 36, 38, 12, 26]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
for r in range(5, 5 + 20):
    ws4.row_dimensions[r].height = 30
ws4.row_dimensions[1].height = 24

# Save
import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"SAVED: {OUT}")
print(f"SIZE:  {os.path.getsize(OUT)} bytes")

"""Generate 全流程练习册.xlsx - 11 sheets.
Wine-red header + beige body, frozen header, hyperlinks, merged cells,
checkboxes, dropdown validation. A4 print optimized.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Protection

OUT = r"D:\Downloads\利益相关方影响和干预\完整课程包\13_Office文档\全流程练习册.xlsx"

# ==================== Styles ====================
WINE = "8B2828"
WINE_MID = "C0392B"
WINE_SOFT = "F0D5CF"
GOLD = "C9A96E"
GOLD_DEEP = "A8884A"
PAPER = "FAF6EC"
PAPER_DEEP = "F5F0E6"
TINT = "F8E6E1"
INK = "1A1A1A"
INK_SOFT = "3A3A3A"
INK_MID = "6E6E6E"
LINE = "D6CFC1"
WHITE = "FFFFFF"

THIN = Side(border_style="thin", color=LINE)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM = Border(bottom=Side(border_style="medium", color=WINE))
BORDER_TOP = Border(top=Side(border_style="medium", color=GOLD))

FONT_TITLE = Font(name="Source Han Serif CN", size=18, bold=True, color=WINE)
FONT_H1 = Font(name="Source Han Serif CN", size=14, bold=True, color=WINE)
FONT_H2 = Font(name="Source Han Serif CN", size=12, bold=True, color=GOLD_DEEP)
FONT_H3 = Font(name="Source Han Serif CN", size=11, bold=True, color=INK)
FONT_BODY = Font(name="Source Han Serif CN", size=10, color=INK)
FONT_BODY_BOLD = Font(name="Source Han Serif CN", size=10, bold=True, color=INK)
FONT_BODY_SOFT = Font(name="Source Han Serif CN", size=10, color=INK_SOFT)
FONT_BODY_GOLD = Font(name="Source Han Serif CN", size=10, color=GOLD_DEEP, bold=True)
FONT_HEADER = Font(name="Source Han Serif CN", size=11, bold=True, color=WHITE)
FONT_TINY = Font(name="Source Han Serif CN", size=8, color=INK_MID, italic=True)
FONT_HINT = Font(name="Source Han Serif CN", size=9, color=INK_MID, italic=True)

FILL_HEADER = PatternFill("solid", fgColor=WINE)
FILL_SUBHEADER = PatternFill("solid", fgColor=GOLD_DEEP)
FILL_PAPER = PatternFill("solid", fgColor=PAPER)
FILL_PAPER_DEEP = PatternFill("solid", fgColor=PAPER_DEEP)
FILL_TINT = PatternFill("solid", fgColor=TINT)
FILL_WINE_SOFT = PatternFill("solid", fgColor=WINE_SOFT)
FILL_ALT = PatternFill("solid", fgColor="F5EFE0")
FILL_GOLD_LIGHT = PatternFill("solid", fgColor="F0E5C8")
FILL_DONE = PatternFill("solid", fgColor="E0F0D8")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def title_block(ws, title, subtitle=None, max_col=10):
    """Add a title block at the top of a sheet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_LEFT
    cell.fill = FILL_TINT
    ws.row_dimensions[1].height = 36
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        cell = ws.cell(row=2, column=1, value=subtitle)
        cell.font = FONT_BODY_SOFT
        cell.alignment = ALIGN_LEFT
        cell.fill = FILL_PAPER
        ws.row_dimensions[2].height = 22
        return 4
    return 3


def header_row(ws, row, headers, fill=FILL_HEADER, font=FONT_HEADER):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[row].height = 32


def fill_row(ws, row, values, alt=False, borders=True):
    bg = FILL_ALT if alt else FILL_PAPER
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = FONT_BODY
        cell.fill = bg
        cell.alignment = ALIGN_LEFT_TOP
        if borders:
            cell.border = BORDER_ALL


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def set_row_height(ws, row, h):
    ws.row_dimensions[row].height = h


def apply_page_setup(ws, orientation="portrait"):
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT if orientation == "portrait" else ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=1.0, right=1.0, top=1.0, bottom=1.0)
    ws.print_options.horizontalCentered = True


def add_checkbox_column(ws, col, start_row, end_row):
    """Add a checkbox validation."""
    dv = DataValidation(type="list", formula1='"☐,☑,✗,N/A"', allow_blank=True)
    dv.error = "请从下拉选择"
    dv.errorTitle = "格式错误"
    dv.prompt = "请选择：☐ 待办 / ☑ 完成 / ✗ 跳过 / N/A 不适用"
    dv.promptTitle = "完成状态"
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")


def add_dropdown(ws, col, start_row, end_row, values):
    """Add a dropdown validation."""
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")


# ==================== Build workbook ====================
wb = Workbook()
wb.remove(wb.active)

# ==================== Sheet 0: 总览 (Overview) ====================
ws = wb.create_sheet("00_总览")
apply_page_setup(ws)

set_col_widths(ws, [4, 18, 18, 18, 18, 18, 18, 18, 18, 18])

# Title
title_block(ws, "共同语言 · 全流程练习册",
            "学员姓名：____________  班级：____________  开课日期：____________  讲师：罗宏伟",
            max_col=10)

# Course map
ws.merge_cells("A3:J3")
c = ws.cell(row=3, column=1, value="■ 课程地图（8 模块 · 14 小时 · 30 天跟进）")
c.font = FONT_H1
c.fill = FILL_PAPER_DEEP
c.alignment = ALIGN_LEFT
ws.row_dimensions[3].height = 26

modules = [
    ("M1", "共同语言", "标准问题陈述句式 · 事实/判断 · 关注点排序", "1.5h"),
    ("M2", "会议准则", "四分类 · 目的声明 · 停车场 · 行动项三要素", "1.5h"),
    ("M3", "利益相关方", "矩阵 · 时机 · 引导对话四步", "2.0h"),
    ("M4", "三层目标", "恢复 / 直接原因 / 根本原因 · 改进行动三标准", "2.0h"),
    ("M5", "现象界定", "IS/IS NOT 5 维度 · 五维度根因追问", "2.0h"),
    ("M6", "回到正轨", "偏离信号 · 4 维度自查 · 纠偏对话四步", "1.5h"),
    ("M7", "预演", "潜在问题预演 · 回滚标准卡", "1.5h"),
    ("M8", "综合演练", "全流程串讲 · 30 天行动计划", "2.0h"),
]
header_row(ws, 4, ["#", "模块", "模块名", "核心内容", "时长", "练习", "完成", "得分", "难度", "备注"])
ws.merge_cells("D4:D4")  # content cell wide
# Re-layout with merged content
# Actually redo with merged "核心内容" col
for col_letter in ["D", "E", "F", "G", "H", "I", "J"]:
    pass

# Manually set headers with merged
ws.delete_rows(4)
header_row(ws, 4, ["#", "模块编号", "模块名", "核心内容（4 列宽）", "", "", "", "练习", "完成", "得分"])
# Merge D-G
ws.merge_cells("D4:G4")

for i, (code, name, content, hours) in enumerate(modules):
    r = 5 + i
    cells = [str(i + 1), code, name, content, "", "", "", f"{code}练习", "☐", "____"]
    for j, v in enumerate(cells, 1):
        cell = ws.cell(row=r, column=j, value=v)
        if j == 1:
            cell.font = FONT_BODY_GOLD
            cell.alignment = ALIGN_CENTER
        elif j in (2, 3, 8):
            cell.font = FONT_BODY_BOLD
            cell.alignment = ALIGN_CENTER
        elif j == 4:
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT_TOP
        elif j == 9:
            cell.font = FONT_BODY
            cell.alignment = ALIGN_CENTER
        else:
            cell.font = FONT_BODY_SOFT
            cell.alignment = ALIGN_CENTER
        cell.fill = FILL_ALT if i % 2 == 0 else FILL_PAPER
        cell.border = BORDER_ALL
    set_row_height(ws, r, 30)

# Hyperlinks for module sheets
for i, (code, name, _, _) in enumerate(modules):
    r = 5 + i
    cell = ws.cell(row=r, column=2)
    cell.hyperlink = f"#'{code}_练习'!A1"
    cell.font = Font(name="Source Han Serif CN", size=10, bold=True, color=WINE, underline="single")

# Add checkboxes for completion column
add_checkbox_column(ws, 9, 5, 5 + len(modules) - 1)

# Score dropdown (A/B/C)
add_dropdown(ws, 10, 5, 5 + len(modules) - 1, ["A", "B", "C", "未做"])

# Total
total_r = 5 + len(modules)
ws.merge_cells(start_row=total_r, start_column=1, end_row=total_r, end_column=8)
cell = ws.cell(row=total_r, column=1, value="全流程总评  →  等级判定：A 优秀 / B 合格 / C 待加强")
cell.font = FONT_H2
cell.fill = FILL_TINT
cell.alignment = ALIGN_CENTER
cell.border = BORDER_ALL

ws.merge_cells(start_row=total_r, start_column=9, end_row=total_r, end_column=10)
cell = ws.cell(row=total_r, column=9, value="____ / 30 题完成")
cell.font = FONT_BODY_BOLD
cell.fill = FILL_TINT
cell.alignment = ALIGN_CENTER
cell.border = BORDER_ALL

# Notes section
note_r = total_r + 2
ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=10)
cell = ws.cell(row=note_r, column=1, value="■ 学员承诺")
cell.font = FONT_H1
cell.fill = FILL_PAPER_DEEP
cell.alignment = ALIGN_LEFT
set_row_height(ws, note_r, 24)

commitment_text = [
    "我承诺：本练习册的所有答案都由本人独立思考完成，引用工具已注明来源。",
    "我承诺：D+30 之前每周回看 1 次，跟进自己的应用情况。",
    "我承诺：若卡在某道练习 24 小时以上，会主动找讲师/同伴求助。",
    "",
    "签名：____________________  日期：____________________",
]
for k, t in enumerate(commitment_text):
    r = note_r + 1 + k
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    cell = ws.cell(row=r, column=1, value=t)
    cell.font = FONT_BODY
    cell.alignment = ALIGN_LEFT
    if k == 4:
        set_row_height(ws, r, 30)
        cell.fill = FILL_PAPER_DEEP

# Quick nav
nav_r = note_r + 7
ws.merge_cells(start_row=nav_r, start_column=1, end_row=nav_r, end_column=10)
cell = ws.cell(row=nav_r, column=1, value="■ 速查导航")
cell.font = FONT_H1
cell.fill = FILL_PAPER_DEEP
cell.alignment = ALIGN_LEFT
set_row_height(ws, nav_r, 24)

nav_items = [
    ("00_总览", "本表", "0"),
    ("M1_练习", "标准句式改写 · 关注点排序", "5"),
    ("M2_练习", "会议四分类 · 行动项三要素", "12"),
    ("M3_练习", "利益相关方识别 · 引导对话", "20"),
    ("M4_练习", "三层分解 · 改进行动评估", "30"),
    ("M5_练习", "IS/IS NOT · 五维度追问", "40"),
    ("M6_练习", "偏离信号 · 纠偏对话", "52"),
    ("M7_练习", "潜在问题预演 · 回滚标准", "64"),
    ("M8_练习", "全流程演练 · 30 天计划", "76"),
    ("30天跟进", "行为重测 · 承诺卡 · 复盘", "88"),
    ("答案参考", "评分标准 + 王工示例", "100"),
]
for k, (sheet, desc, _) in enumerate(nav_items):
    r = nav_r + 1 + k
    cell_a = ws.cell(row=r, column=1, value=f"{k:02d}")
    cell_a.font = FONT_BODY_GOLD
    cell_a.alignment = ALIGN_CENTER
    cell_a.border = BORDER_ALL
    cell_a.fill = FILL_ALT
    cell_b = ws.cell(row=r, column=2, value=sheet)
    cell_b.font = Font(name="Source Han Serif CN", size=10, bold=True, color=WINE, underline="single")
    cell_b.hyperlink = f"#'{sheet}'!A1"
    cell_b.alignment = ALIGN_LEFT
    cell_b.border = BORDER_ALL
    cell_b.fill = FILL_PAPER
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    cell_c = ws.cell(row=r, column=3, value=desc)
    cell_c.font = FONT_BODY
    cell_c.alignment = ALIGN_LEFT
    cell_c.fill = FILL_PAPER
    cell_c.border = BORDER_ALL
    set_row_height(ws, r, 22)

# Freeze header
ws.freeze_panes = "A4"

# ==================== Module practice sheets ====================
def module_practice_sheet(sheet_name, module_code, module_title, exercises):
    """Generic module practice sheet."""
    ws = wb.create_sheet(sheet_name)
    apply_page_setup(ws)
    set_col_widths(ws, [5, 12, 22, 50, 14, 12, 12, 12, 14, 12])

    start = title_block(ws, f"{module_code}  {module_title}",
                        f"学员姓名：____________  日期：____________  用时：______ 分钟",
                        max_col=10)
    # Module objective
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
    c = ws.cell(row=3, column=1, value="■ 学习目标")
    c.font = FONT_H1
    c.fill = FILL_PAPER_DEEP
    c.alignment = ALIGN_LEFT
    set_row_height(ws, 3, 24)

    # exercises: list of (ex_id, title, content, lines, score)
    cur_row = 4
    for i, (ex_id, title, content, lines, points) in enumerate(exercises):
        # Exercise header
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
        c = ws.cell(row=cur_row, column=1,
                    value=f"■  {ex_id}  {title}    （{points} 分）")
        c.font = FONT_H2
        c.fill = FILL_TINT
        c.alignment = ALIGN_LEFT
        c.border = BORDER_ALL
        set_row_height(ws, cur_row, 26)
        cur_row += 1

        # Content / instruction
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
        c = ws.cell(row=cur_row, column=1, value=content)
        c.font = FONT_BODY
        c.fill = FILL_PAPER
        c.alignment = ALIGN_LEFT_TOP
        c.border = BORDER_ALL
        # Multi-line content
        h = max(22, 14 + 14 * content.count("\n"))
        set_row_height(ws, cur_row, h)
        cur_row += 1

        # Writing area
        for ln in range(lines):
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
            c = ws.cell(row=cur_row, column=1, value="")
            c.fill = FILL_PAPER_DEEP
            c.alignment = ALIGN_LEFT_TOP
            c.border = BORDER_BOTTOM
            set_row_height(ws, cur_row, 30)
            cur_row += 1

        # Self-check + status row
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
        c = ws.cell(row=cur_row, column=1, value="自检：□ 工具用对  □ 5 要素全  □ 真实场景  □ 至少 1 个亮点")
        c.font = FONT_HINT
        c.alignment = ALIGN_LEFT
        c.fill = FILL_PAPER
        c.border = BORDER_ALL
        # Status dropdown
        c = ws.cell(row=cur_row, column=8, value="完成状态")
        c.font = FONT_BODY_GOLD
        c.alignment = ALIGN_CENTER
        c.fill = FILL_TINT
        c.border = BORDER_ALL
        c = ws.cell(row=cur_row, column=9, value="☐")
        c.font = FONT_BODY
        c.alignment = ALIGN_CENTER
        c.fill = FILL_PAPER
        c.border = BORDER_ALL
        c = ws.cell(row=cur_row, column=10, value=f"{points} 分")
        c.font = FONT_BODY_BOLD
        c.alignment = ALIGN_CENTER
        c.fill = FILL_PAPER
        c.border = BORDER_ALL
        set_row_height(ws, cur_row, 22)
        cur_row += 2

    # Dropdown for status
    add_checkbox_column(ws, 9, 5, cur_row - 1)

    # Module total
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=8)
    c = ws.cell(row=cur_row, column=1,
                value=f"{module_code} 模块得分合计（满分 {sum(e[4] for e in exercises)} 分）")
    c.font = FONT_H2
    c.fill = FILL_TINT
    c.alignment = ALIGN_RIGHT
    c.border = BORDER_ALL
    c = ws.cell(row=cur_row, column=9, value=f"____ / {sum(e[4] for e in exercises)}")
    c.font = FONT_BODY_BOLD
    c.fill = FILL_TINT
    c.alignment = ALIGN_CENTER
    c.border = BORDER_ALL
    c = ws.cell(row=cur_row, column=10, value="")
    c.fill = FILL_TINT
    c.border = BORDER_ALL
    set_row_height(ws, cur_row, 26)
    cur_row += 1

    # Module reflection
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
    c = ws.cell(row=cur_row, column=1, value=f"■ {module_code} 模块反思（写 1 段话）")
    c.font = FONT_H1
    c.fill = FILL_PAPER_DEEP
    c.alignment = ALIGN_LEFT
    set_row_height(ws, cur_row, 24)
    cur_row += 1
    for _ in range(3):
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
        c = ws.cell(row=cur_row, column=1, value="")
        c.fill = FILL_PAPER_DEEP
        c.alignment = ALIGN_LEFT_TOP
        c.border = BORDER_BOTTOM
        set_row_height(ws, cur_row, 30)
        cur_row += 1

    ws.freeze_panes = "A5"


# M1 - 共同语言
m1_exercises = [
    ("练习 1A", "事实与判断区分（基础练习）",
     "题目：以下 5 段话中圈出判断词、标出事实，写下你的判断。\n"
     "1) 「这批供应商不靠谱，老李又出问题了」\n"
     "2) 「上周三到的那批零件肯定尺寸超差」\n"
     "3) 「客户昨天打电话说质量太差，肯定要退货」\n"
     "4) 「项目延期 2 天，小李状态不太好，应该找他谈谈」\n"
     "5) 「上个月质量成本上升 12%，约 35 万元」\n"
     "要求：A 档 = 8+ 正确判断 · B 档 = 6+ 正确 · C 档 = < 6 正确",
     6, 4),
    ("练习 1B", "标准句式改写（基础练习）",
     "题目：把以下 3 个判断改写为 5 要素标准句式。\n"
     "1) 「最近来料检验老出问题，供应商也靠不住」\n"
     "2) 「测试总是不通过，开发那边也说不清楚」\n"
     "3) 「客户反馈工程师都挺忙，交付可能要推迟」\n"
     "要求：A 档 = 3 个全对 · B 档 = 2 个全对 · C 档 = 1 个对",
     8, 5),
    ("练习 1C", "5 要素自检（实操练习）",
     "题目：写下一个你今天遇到的工作问题（真实场景），用 5 要素完整陈述。\n"
     "并请一位同事听完，复述你的「范围+偏差」——如果和你说的不一致，重写。",
     6, 4),
    ("练习 1D", "关注点排序（提升练习）",
     "题目：列出你手上项目/工作的 3-5 个关键关注点，按「严重性 / 紧迫性 / 趋势」3 维度打分。\n"
     "如果排序僵持，挑出那条用「今天不处理明天哪条会来不及」破局。",
     6, 4),
]
module_practice_sheet("M1_练习", "M1", "共同语言 · 标准句式 · 关注点排序", m1_exercises)

# M2 - 会议准则
m2_exercises = [
    ("练习 2A", "会议分类（基础练习）",
     "题目：列出你过去 1 周开过的 5 个会议，按四分类（启动/决策/排查/复盘）归类。\n"
     "如果有「哪一类都不是」的会议——问自己「这会议本来就不该开？」",
     6, 4),
    ("练习 2B", "目的声明（实操练习）",
     "题目：为下周一你要开的 1 个会议写 1 句目的声明。\n"
     "模板：「这是一场〔类型〕会议，今天结束时我们应该〔具体产出〕」。",
     4, 4),
    ("练习 2C", "停车场使用（实操练习）",
     "题目：模拟 1 次「问题排查会」30 分钟，列 3-5 个跑题，填进停车场。\n"
     "写明「什么时候讨论」「跟进人」。",
     6, 4),
    ("练习 2D", "行动项三要素（提升练习）",
     "题目：从你最近 1 个会议纪要里挑 3 条「行动项」，用「具体人 + 具体动作 + 具体时间」重写。\n"
     "识别出哪些是「伪行动项」（团队/大家/尽快）。",
     6, 5),
]
module_practice_sheet("M2_练习", "M2", "会议准则 · 四分类 · 行动项三要素", m2_exercises)

# M3 - 利益相关方
m3_exercises = [
    ("练习 3A", "相关方识别（基础练习）",
     "题目：选你手上的 1 个真实项目，列出 5-8 个关键相关方（按决策/执行/影响/边缘 4 类）。\n"
     "自检 4 问：有没有漏掉能拍板的人/会受影响但没意识到/反对过同类的/我的上下级。",
     6, 5),
    ("练习 3B", "矩阵分类（实操练习）",
     "题目：把 3A 识别的相关方按「影响力 × 支持度」二维分类。\n"
     "识别「高影响力 + 低支持」的 1-2 人，写出接触方式 + 接触时机。",
     6, 5),
    ("练习 3C", "引导对话四步（提升练习）",
     "题目：模拟 1 次跨部门关键对话（30 分钟），用 4 步结构：\n"
     "Step 1 确认目标一致 · Step 2 陈述问题（M1 句式）· Step 3 邀请补充事实 · Step 4 共同下一步。",
     8, 6),
]
module_practice_sheet("M3_练习", "M3", "利益相关方 · 矩阵 · 引导对话四步", m3_exercises)

# M4 - 三层目标
m4_exercises = [
    ("练习 4A", "三层分解（基础练习）",
     "题目：选 1 个真实异常，按「恢复 / 直接原因 / 根本原因」三层分解。\n"
     "自检：第三层有没有被悄悄拖到「再也没人提」？",
     6, 5),
    ("练习 4B", "5Why 追问（实操练习）",
     "题目：从 4A 的「直接原因」出发，5 次「为什么」追到「系统原因」。\n"
     "5 维度归类：流程 / 标准 / 培训 / 工具 / 沟通。",
     6, 5),
    ("练习 4C", "改进行动三标准（提升练习）",
     "题目：列出 3 个改进行动，每条按「有效性 / 可执行性 / 副作用」1-5 分评估。\n"
     "对每个副作用用「如果……那么……」推演应对预案。",
     6, 6),
]
module_practice_sheet("M4_练习", "M4", "三层目标 · 5Why · 改进行动评估", m4_exercises)

# M5 - 现象界定
m5_exercises = [
    ("练习 5A", "IS/IS NOT 对比表（基础练习）",
     "题目：选 1 个反复出现的问题，5 维度（What/Where/When/Who/程度）填 IS/IS NOT。\n"
     "IS NOT 至少 3 条 · 程度维度要具体（对象/时间/范围）。",
     8, 6),
    ("练习 5B", "五维度根因追问（实操练习）",
     "题目：基于 5A，沿 5 维度（流程/标准/培训/工具/沟通）追问「为什么这个漏洞会存在」。\n"
     "归到系统层面（不归个人）。",
     6, 6),
    ("练习 5C", "对比基线（提升练习）",
     "题目：用 1 句话说清你的「对比基线」——相对什么对比。\n"
     "例：相对上月同期、相对标准 SOP、相对同类项目。",
     4, 4),
]
module_practice_sheet("M5_练习", "M5", "现象界定 · IS/IS NOT · 五维度追问", m5_exercises)

# M6 - 回到正轨
m6_exercises = [
    ("练习 6A", "偏离信号自查（基础练习）",
     "题目：选 1 个当前项目，4 维度（进度/质量/沟通/资源）自查偏离。\n"
     "识别早期信号（不是等大问题）。",
     6, 4),
    ("练习 6B", "分级决策（实操练习）",
     "题目：对 6A 识别的偏离，按「坚持 / 调整 / 止损」3 选 1。\n"
     "避免两个极端——再坚持 或 大改。",
     4, 5),
    ("练习 6C", "纠偏对话四步（提升练习）",
     "题目：模拟 1 次纠偏对话 30 分钟，用 4 步结构：\n"
     "Step 1 现状 · Step 2 影响 · Step 3 选项（坚持/调整/止损 三选都要过）· Step 4 共识。",
     8, 6),
]
module_practice_sheet("M6_练习", "M6", "回到正轨 · 偏离信号 · 纠偏对话", m6_exercises)

# M7 - 潜在问题预演
m7_exercises = [
    ("练习 7A", "潜在问题清单（基础练习）",
     "题目：选 1 个即将执行的行动，列出 5 个潜在问题：哪些环节最容易出错/谁可能不配合/哪些数据可能拿不到。",
     6, 5),
    ("练习 7B", "触发条件 + 应对预案（实操练习）",
     "题目：基于 7A，为每条潜在问题预设：\n"
     "① 触发条件（具体可观察的信号）② 应对预案（具体动作）③ 责任人。",
     8, 6),
    ("练习 7C", "回滚标准卡（提升练习）",
     "题目：写出 1 张「回滚标准卡」：① 行动名 ② 触发回滚的标准 ③ 回滚执行步骤 ④ 汇报路径。",
     6, 5),
]
module_practice_sheet("M7_练习", "M7", "潜在问题预演 · 回滚标准卡", m7_exercises)

# M8 - 综合演练
m8_exercises = [
    ("练习 8A", "全流程演练（综合练习）",
     "题目：选 1 个真实工作场景，60 分钟走完 M1-M7 全流程：\n"
     "M1 陈述问题 · M2 开会 · M3 引导对话 · M4 三层分解 · M5 IS/IS NOT · M6 纠偏 · M7 预演。",
     12, 10),
    ("练习 8B", "30 天行动计划（落地练习）",
     "题目：基于 8A 的演练结果，写 30 天行动计划：\n"
     "① 2-3 个具体行动 ② 用哪个工具 ③ 预期产出 ④ 第 1-4 周进度 ⑤ 障碍 + 应对 ⑥ 30 天回看。",
     10, 8),
    ("练习 8C", "综合演练评估（互评练习）",
     "题目：找 1 位同伴，用 M1-M7 标准给你 1-5 分互评。\n"
     "写出 1 个亮点 + 1 个待改进 + 1 个建议下一步。",
     6, 5),
]
module_practice_sheet("M8_练习", "M8", "综合演练 · 30 天计划", m8_exercises)

# ==================== 30 天跟进 ====================
ws = wb.create_sheet("30天跟进")
apply_page_setup(ws)
set_col_widths(ws, [5, 16, 16, 16, 16, 16, 16, 16, 16, 16])

title_block(ws, "30 天跟进",
            "D+7  ·  D+14  ·  D+21  ·  D+30  ·  4 周重测",
            max_col=10)

# Section A: 30-day check-in
ws.merge_cells("A3:J3")
c = ws.cell(row=3, column=1, value="■ 4 周周回顾（每周打勾）")
c.font = FONT_H1
c.fill = FILL_PAPER_DEEP
c.alignment = ALIGN_LEFT
set_row_height(ws, 3, 24)

weekly_items = [
    "M1 标准句式使用频次",
    "M2 目的声明使用频次",
    "M2 行动项三要素使用频次",
    "M3 引导对话使用频次",
    "M4 三层分解使用频次",
    "M5 IS/IS NOT 使用频次",
    "M6 纠偏对话使用频次",
    "M7 预演 + 回滚使用频次",
    "30 天报告本周完成度",
    "本周最有收获的 1 件事",
    "本周最卡的地方",
    "下周 1 个具体行动",
]
header_row(ws, 4, ["#", "回顾项", "D+7", "D+14", "D+21", "D+30", "总评", "卡点", "下一步", "状态"])
for i, item in enumerate(weekly_items):
    r = 5 + i
    bg = FILL_ALT if i % 2 == 0 else FILL_PAPER
    cells = [
        str(i + 1),
        item,
        "____ 次",
        "____ 次",
        "____ 次",
        "____ 次",
        "",
        "",
        "",
        "☐"
    ]
    for j, v in enumerate(cells, 1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.font = FONT_BODY
        cell.fill = bg
        cell.border = BORDER_ALL
        if j == 1:
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_BODY_GOLD
        elif j == 2:
            cell.alignment = ALIGN_LEFT
            cell.font = FONT_BODY_BOLD
        else:
            cell.alignment = ALIGN_CENTER
    set_row_height(ws, r, 24)

# Status dropdown
add_checkbox_column(ws, 10, 5, 5 + len(weekly_items) - 1)

# Section B: 重测自评
sec_b_r = 5 + len(weekly_items) + 2
ws.merge_cells(start_row=sec_b_r, start_column=1, end_row=sec_b_r, end_column=10)
c = ws.cell(row=sec_b_r, column=1, value="■ D+30 行为重测自评（与课前自评对照）")
c.font = FONT_H1
c.fill = FILL_PAPER_DEEP
c.alignment = ALIGN_LEFT
set_row_height(ws, sec_b_r, 24)

retest_items = [
    "1. 我和团队讨论「问题是什么」时，10 次有 ____ 次理解一致。",
    "2. 我召开的会议 10 次有 ____ 次能在结束时给出清晰的行动项。",
    "3. 我面对跨部门相关方时，知道该先找谁、什么时候找。",
    "4. 当项目/工作偏离计划，我能早期识别并启动纠偏。",
    "5. 我最常用的工具是：__________________________",
    "6. 我最常用的场景是：__________________________",
    "7. 我对工作直接帮助最大的是 M：__________",
    "8. 我觉得还可以更好的地方：________________________",
]
for i, item in enumerate(retest_items):
    r = sec_b_r + 1 + i
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws.cell(row=r, column=1, value=item)
    c.font = FONT_BODY
    c.fill = FILL_PAPER
    c.alignment = ALIGN_LEFT_TOP
    c.border = BORDER_BOTTOM
    set_row_height(ws, r, 28)

# Section C: 行为承诺卡
sec_c_r = sec_b_r + 1 + len(retest_items) + 2
ws.merge_cells(start_row=sec_c_r, start_column=1, end_row=sec_c_r, end_column=10)
c = ws.cell(row=sec_c_r, column=1, value="■ 行为承诺卡（D-Day 当天签 · D+30 回看）")
c.font = FONT_H1
c.fill = FILL_PAPER_DEEP
c.alignment = ALIGN_LEFT
set_row_height(ws, sec_c_r, 24)

ws.merge_cells(start_row=sec_c_r + 1, start_column=1, end_row=sec_c_r + 1, end_column=10)
c = ws.cell(row=sec_c_r + 1, column=1,
            value="我承诺今后 30 天：____________________________________________________")
c.font = FONT_BODY
c.fill = FILL_PAPER
c.border = BORDER_BOTTOM
set_row_height(ws, sec_c_r + 1, 32)
for k in range(2):
    r = sec_c_r + 2 + k
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws.cell(row=r, column=1, value="")
    c.fill = FILL_PAPER_DEEP
    c.border = BORDER_BOTTOM
    set_row_height(ws, r, 32)

# Sign
r = sec_c_r + 4
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
c = ws.cell(row=r, column=1,
            value="我的签名：____________________  日期：____________________  见证人：____________________")
c.font = FONT_BODY_BOLD
c.fill = FILL_PAPER
c.alignment = ALIGN_LEFT
set_row_height(ws, r, 32)

# D+30 回看
r2 = sec_c_r + 6
ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=10)
c = ws.cell(row=r2, column=1, value="D+30 回看（D+30 填写）")
c.font = FONT_H2
c.fill = FILL_TINT
c.alignment = ALIGN_LEFT
set_row_height(ws, r2, 24)

review_items = [
    "1. 我做到承诺了吗？□ 是 □ 否 → 原因是：",
    "2. 我用了几个工具？_______ 个",
    "3. 我最大的变化是：",
    "4. 我对这门课的整体满意度：____ / 5.0",
    "5. 我会推荐给同事吗？□ 会 □ 不一定 □ 不会",
    "6. 我自己的评价：",
]
for i, item in enumerate(review_items):
    r = r2 + 1 + i
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws.cell(row=r, column=1, value=item)
    c.font = FONT_BODY
    c.fill = FILL_PAPER
    c.alignment = ALIGN_LEFT_TOP
    c.border = BORDER_BOTTOM
    set_row_height(ws, r, 28)

ws.freeze_panes = "A4"

# ==================== 答案参考 ====================
ws = wb.create_sheet("答案参考")
apply_page_setup(ws)
set_col_widths(ws, [5, 14, 18, 22, 60, 12, 14])

title_block(ws, "答案参考与评分标准",
            "★ 讲师讲解用 · 学员可对照自查 · 评分档 A / B / C",
            max_col=7)

# Module-level answer summary
ws.merge_cells("A3:G3")
c = ws.cell(row=3, column=1, value="■ 8 模块答案要点速查")
c.font = FONT_H1
c.fill = FILL_PAPER_DEEP
c.alignment = ALIGN_LEFT
set_row_height(ws, 3, 24)

answer_summary = [
    ("M1", "练习 1A", "事实判断", "5 段话中识别 8+ 判断词；常见误判：把「客户说」当事实 / 把「应该」当行动 / 把假设当事实"),
    ("M1", "练习 1B", "句式改写", "5 要素齐全：范围 / 对象 / 标准 / 偏差 / 后果；常见错误：偏差用「有点超差」/ 后果用「影响很大」"),
    ("M1", "练习 1C", "5 要素自检", "真实场景 + 同事可复述；A 档需做到同事听完后能说出你的范围+偏差"),
    ("M1", "练习 1D", "关注点排序", "3 维度评分后选 Top 3；破局句「今天不处理，明天哪条会来不及」"),
    ("M2", "练习 2A", "会议分类", "四分类能完全覆盖；常见：把「碰头会」当一类——其实不属于任何一类"),
    ("M2", "练习 2B", "目的声明", "类型 + 具体产出；反例：「讨论一下」「对一下」"),
    ("M2", "练习 2C", "停车场", "跑题不打断 + 明确时间 + 跟进人"),
    ("M2", "练习 2D", "行动项三要素", "具体人 + 具体动作 + 具体时间；伪行动项识别（团队/大家/尽快）"),
    ("M3", "练习 3A", "相关方识别", "4 类全覆盖（决策/执行/影响/边缘）；自检 4 问"),
    ("M3", "练习 3B", "矩阵分类", "右上角是重点；写明接触方式+时机"),
    ("M3", "练习 3C", "引导对话", "4 步结构：目标 / 陈述 / 补充 / 下一步"),
    ("M4", "练习 4A", "三层分解", "恢复 / 直接原因 / 根本原因 不混淆；自检：第三层有没有被悄悄拖掉"),
    ("M4", "练习 4B", "5Why 追问", "5 次到系统层；归类到 5 维度"),
    ("M4", "练习 4C", "三标准评估", "有效性 / 可执行性 / 副作用 1-5 分；副作用推演"),
    ("M5", "练习 5A", "IS/IS NOT", "5 维度齐全；IS NOT 至少 3 条；程度要具体"),
    ("M5", "练习 5B", "五维度追问", "流程/标准/培训/工具/沟通 不归个人"),
    ("M5", "练习 5C", "对比基线", "1 句话，相对什么对比"),
    ("M6", "练习 6A", "偏离信号", "4 维度自查；早期信号（非大问题）"),
    ("M6", "练习 6B", "分级决策", "坚持 / 调整 / 止损 三选 1；避免两个极端"),
    ("M6", "练习 6C", "纠偏对话", "4 步：现状 / 影响 / 选项 / 共识"),
    ("M7", "练习 7A", "问题清单", "5 个潜在问题：环节 / 人 / 数据"),
    ("M7", "练习 7B", "触发+预案", "触发条件 + 应对预案 + 责任人"),
    ("M7", "练习 7C", "回滚标准卡", "行动名 / 触发标准 / 步骤 / 汇报路径"),
    ("M8", "练习 8A", "全流程演练", "60 分钟走完 M1-M7；用真实场景"),
    ("M8", "练习 8B", "30 天计划", "2-3 个行动 + 工具 + 预期 + 4 周进度 + 障碍应对 + 30 天回看"),
    ("M8", "练习 8C", "综合演练评估", "1 个亮点 + 1 个待改进 + 1 个下一步"),
]

header_row(ws, 4, ["#", "模块", "练习", "能力点", "答案要点 / 评分标准", "满分", "档位参考"])
for i, row in enumerate(answer_summary):
    r = 5 + i
    bg = FILL_ALT if i % 2 == 0 else FILL_PAPER
    # row = (module, ex_id, ability, body) - 4 items. Pad to 7 columns.
    cells = list(row) + ["", "", ""]  # 7 total
    cells[5] = "5"
    cells[6] = "A≥80% / B≥60% / C<60%"
    for j, v in enumerate(cells, 1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.fill = bg
        cell.border = BORDER_ALL
        cell.font = FONT_BODY
        if j in (1, 2, 3, 5, 6):
            cell.alignment = ALIGN_CENTER
        else:
            cell.alignment = ALIGN_LEFT_TOP
    set_row_height(ws, r, 36)

# 王工 example sheet
wkg_r = 5 + len(answer_summary) + 2
ws.merge_cells(start_row=wkg_r, start_column=1, end_row=wkg_r, end_column=7)
c = ws.cell(row=wkg_r, column=1, value="■ 王工参考答题示例（M1 + M2 节选 · 完整版见 13_Office文档/讲师手册.docx）")
c.font = FONT_H1
c.fill = FILL_PAPER_DEEP
c.alignment = ALIGN_LEFT
set_row_height(ws, wkg_r, 24)

wkg_lines = [
    ("M1 · 练习 1B", "原描述", "「最近来料检验老出问题，供应商也靠不住，得找人去盯一下。」"),
    ("M1 · 练习 1B", "改写后", "在【最近 1 个月】的【来料检验】中，【A-120 等外协件】本应【按 GB/T 1800 公差 ≥ 99.5%】，实际出现【连续 3 批尺寸超差（96% / 92% / 89%）】，影响是【产线停工 4 小时 + 38 件无法使用 + 客户 S-2403 延期 2-3 天】。"),
    ("M2 · 练习 2B", "原会议", "「最近质量出问题，要开会讨论」"),
    ("M2 · 练习 2B", "目的声明", "「这是一场问题排查会，今天结束时我们应该就 A-120 三个批次的处理方案达成共识，并由质量部王工牵头出 1 份《应急检验规范》」"),
    ("M2 · 练习 2D", "伪行动项", "「找一下供应商」「尽快去盯」"),
    ("M2 · 练习 2D", "重写后", "「采购部张工 6 月 20 日前完成对供应商 B-7 的现场审核，输出 1 份审核报告」"),
    ("M3 · 练习 3C", "对话 4 步示例", "① 目标一致：'我们今天要决定 A-120 第 3 批是否放行'；② 陈述问题：'3 批超差，影响 38 件产线'；③ 邀请补充：'有没有我们没注意到的信息'；④ 共同下一步：'质量部 6/15 前出应急规范'。"),
    ("M4 · 练习 4A", "三层分解示例", "恢复：6/12 启动应急检验（暂时放行至抽检合格后入厂）；直接原因：供应商 B-7 量具未校准（6/13 已校准）；根本原因：供应商审核机制缺失（按季度审核 + 月度抽检 5% 改为月度审核 + 抽检 10%）。"),
    ("M5 · 练习 5A", "IS/IS NOT 5 维度示例",
     "IS：What-A-120 尺寸 / Where-外协件 / When-6 月 / Who-外协供应商 B-7 / 程度-3 批超差 5-10% / IS NOT：不是其他型号 / 不是其他供应商 / 不是来料检验员误判 / 不是 6 月之前。"),
    ("M6 · 练习 6C", "纠偏对话示例", "现状：项目延期 2 天；影响：客户 S-2403 首批延期 2-3 天；选项：坚持原计划（不可行）/ 调整：分批交付（6/20 + 6/25）/ 止损：暂停第二批（不可行）；共识：6/20 交付 60% + 6/25 交付 40%，下周一 review。"),
    ("M7 · 练习 7C", "回滚标准卡示例", "行动：上线新检验标准；触发回滚：连续 2 批超差 > 5% / 产线停工 > 2 小时 / 客户投诉 ≥ 1 次；步骤：① 停止执行 ② 切回原标准 ③ 24 小时内汇报至王工 + 张总。"),
]
header_row(ws, wkg_r + 1, ["#", "练习", "类别", "示例", "", "", ""])
ws.merge_cells(start_row=wkg_r + 1, start_column=4, end_row=wkg_r + 1, end_column=7)
for i, (tag, cat, body) in enumerate(wkg_lines):
    r = wkg_r + 2 + i
    bg = FILL_ALT if i % 2 == 0 else FILL_PAPER
    cells = [str(i + 1), tag, cat, body, "", "", ""]
    for j, v in enumerate(cells, 1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.fill = bg
        cell.border = BORDER_ALL
        if j in (1, 2, 3):
            cell.font = FONT_BODY_GOLD if j == 2 else FONT_BODY
            cell.alignment = ALIGN_CENTER if j != 3 else ALIGN_LEFT
        else:
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT_TOP
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    set_row_height(ws, r, 60)

ws.freeze_panes = "A4"

# Save
wb.save(OUT)
print(f"Created: {OUT}")
print(f"Size: {os.path.getsize(OUT)} bytes")
print(f"Sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")

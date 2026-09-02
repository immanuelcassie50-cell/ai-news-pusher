import openpyxl
wb = openpyxl.load_workbook(r"D:\新课开发\经营\08_存量竞争时代的增长设计：找到还能增长的路径\成果demo\10_项目成果展示.xlsx")
print("Sheets:", wb.sheetnames)
ws1 = wb["增长数据追踪"]
print("Sheet1 rows:", ws1.max_row, "cols:", ws1.max_column)
print("Headers:", [ws1.cell(1, c).value for c in range(1, 9)])
ws2 = wb["目标完成情况"]
print("Sheet2 rows:", ws2.max_row, "cols:", ws2.max_column)
ws3 = wb["经验总结"]
print("Sheet3 rows:", ws3.max_row, "cols:", ws3.max_column)
ws4 = wb["方法论应用效果"]
print("Sheet4 rows:", ws4.max_row, "cols:", ws4.max_column)
print("Verification complete")

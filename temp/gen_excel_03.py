# -*- coding: utf-8 -*-
"""
生成 16_培训效果数据汇总_组织用.xlsx
4 sheets:
  1. 前测数据 (培训前水平)
  2. 后测数据 (培训后水平)
  3. 30天应用数据 (行为转化)
  4. 汇总统计 (前后对比+雷达图+柱状图)
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule, ColorScaleRule
from openpyxl.chart import BarChart, RadarChart, Reference, LineChart
from openpyxl.styles.colors import Color

# 主题色：金色（培训·效果·成就）
ACCENT = "B45309"
ACCENT_SOFT = "FEF3C7"
ACCENT_DARK = "8A4509"
OK = "166534"
OK_SOFT = "E8F1EB"
WARN = "C2410C"
WARN_SOFT = "FEF0E8"
NEUTRAL = "6B7280"
NEUTRAL_SOFT = "F3F4F6"
LINE = "D8C89A"
LINE_SOFT = "F0E4C8"
BG = "FDF6E3"
PAPER = "FFFFFF"
INK = "1A1410"
INK_SOFT = "4A3A28"

TITLE_FONT = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
H2_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color=ACCENT_DARK)
HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Microsoft YaHei", size=10, color=INK)
SOFT_FONT = Font(name="Microsoft YaHei", size=10, color=INK_SOFT)
BOLD_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color=INK)
NOTE_FONT = Font(name="Microsoft YaHei", size=9, italic=True, color=NEUTRAL)

def PF(color):
    return PatternFill("solid", fgColor=color)

TITLE_FILL = PF(ACCENT)
H2_FILL = PF(ACCENT_SOFT)
HEADER_FILL = PF(ACCENT)
OK_FILL = PF(OK_SOFT)
WARN_FILL = PF(WARN_SOFT)
BG_FILL = PF(BG)
NEUTRAL_FILL = PF(NEUTRAL_SOFT)

thin = Side(border_style="thin", color=LINE)
medium = Side(border_style="medium", color=ACCENT)
ALL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()
wb.remove(wb.active)


# ============================================================
# Sheet 1: 前测数据
# ============================================================
ws1 = wb.create_sheet("前测数据")

ws1.merge_cells("A1:H1")
ws1["A1"] = "前测数据 - 培训前水平"
ws1["A1"].font = TITLE_FONT
ws1["A1"].fill = TITLE_FILL
ws1["A1"].alignment = CENTER
ws1.row_dimensions[1].height = 32

ws1.merge_cells("A2:H2")
ws1["A2"] = "说明：开课前 1 周对学员测评。1 行 1 名学员。分 5 个能力维度，每个维度 0-10 分"
ws1["A2"].font = NOTE_FONT
ws1["A2"].fill = BG_FILL
ws1["A2"].alignment = LEFT
ws1.row_dimensions[2].height = 24

# 5 个能力维度
dimensions = [
    "问题定义",     # 1
    "根因分析",     # 2
    "方案设计",     # 3
    "执行跟踪",     # 4
    "复盘总结",     # 5
]

headers1 = [
    ("学员编号", 10),
    ("姓名", 10),
    ("部门", 14),
    ("岗位", 14),
    ("问题定义", 10),
    ("根因分析", 10),
    ("方案设计", 10),
    ("执行跟踪", 10),
    ("复盘总结", 10),
    ("总分", 8),
    ("等级", 10),
]
for col_idx, (text, width) in enumerate(headers1, 1):
    c = ws1.cell(row=3, column=col_idx, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
    ws1.column_dimensions[get_column_letter(col_idx)].width = width
ws1.row_dimensions[3].height = 32

# 30 名学员的前测数据（模拟，呈现典型分布）
import random
random.seed(42)

# 学员基础信息
students = [
    ("S001", "小王", "销售部", "销售经理"),
    ("S002", "小李", "客服部", "客服主管"),
    ("S003", "小张", "人力部", "HR 专员"),
    ("S004", "小陈", "设备部", "设备主管"),
    ("S005", "小赵", "采购部", "采购经理"),
    ("S006", "小周", "管理部", "运营经理"),
    ("S007", "小孙", "财务部", "会计主管"),
    ("S008", "小钱", "生产部", "车间主任"),
    ("S009", "小吴", "销售部", "销售专员"),
    ("S010", "小郑", "人力部", "招聘经理"),
    ("S011", "小冯", "技术部", "技术经理"),
    ("S012", "小何", "品保部", "品保主管"),
    ("S013", "小蒋", "项目部", "项目经理"),
    ("S014", "小沈", "行政部", "行政主管"),
    ("S015", "小韩", "物流部", "物流经理"),
    ("S016", "小杨", "销售部", "销售总监"),
    ("S017", "小朱", "客服部", "客服专员"),
    ("S018", "小秦", "人力部", "培训经理"),
    ("S019", "小尤", "设备部", "维修工"),
    ("S020", "小许", "采购部", "采购专员"),
    ("S021", "小何", "管理部", "总经理助理"),
    ("S022", "小吕", "财务部", "出纳"),
    ("S023", "小施", "生产部", "操作工"),
    ("S024", "小张", "技术部", "开发工程师"),
    ("S025", "小孔", "品保部", "质检员"),
    ("S026", "小曹", "项目部", "项目助理"),
    ("S027", "小严", "行政部", "前台"),
    ("S028", "小华", "物流部", "仓管员"),
    ("S029", "小金", "销售部", "销售助理"),
    ("S030", "小魏", "人力部", "HR 总监"),
]

# 生成有区分度的前测分数
sample1 = []
for sid, name, dept, pos in students:
    # 不同岗位基础不同
    base = 4
    if "经理" in pos or "主管" in pos or "总监" in pos:
        base = 5
    elif "助理" in pos or "专员" in pos:
        base = 4
    elif "工" in pos or "员" in pos:
        base = 3

    # 5 个维度分别有差异
    scores = [
        max(1, min(10, base + random.randint(-1, 2))),
        max(1, min(10, base + random.randint(-1, 2))),
        max(1, min(10, base + random.randint(-1, 2))),
        max(1, min(10, base + random.randint(-1, 2))),
        max(1, min(10, base + random.randint(-1, 2))),
    ]
    sample1.append((sid, name, dept, pos, *scores))

# 写入数据
n1 = len(sample1)
for row_idx, row_data in enumerate(sample1, 4):
    for col_idx, value in enumerate(row_data, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=value)
        c.font = NORMAL_FONT
        c.alignment = CENTER
        c.border = ALL_BORDER
        if col_idx >= 5 and col_idx <= 9:
            c.number_format = "0"
            c.alignment = CENTER
        if col_idx >= 5 and col_idx <= 9:
            # 分数着色
            if value >= 7:
                c.fill = OK_FILL
                c.font = Font(name="Microsoft YaHei", size=10, bold=True, color=OK)
            elif value <= 3:
                c.fill = WARN_FILL
                c.font = Font(name="Microsoft YaHei", size=10, color=NEUTRAL)

    # 总分公式
    total_c = ws1.cell(row=row_idx, column=10, value=f'=SUM(E{row_idx}:I{row_idx})')
    total_c.number_format = "0"
    total_c.font = BOLD_FONT
    total_c.alignment = CENTER
    total_c.border = ALL_BORDER
    total_c.fill = PF(ACCENT_SOFT)

    # 等级公式
    level_c = ws1.cell(row=row_idx, column=11, value=f'=IF(J{row_idx}>=40,"优秀",IF(J{row_idx}>=30,"良好",IF(J{row_idx}>=20,"一般",IF(J{row_idx}>=10,"需提升","弱"))))')
    level_c.font = BOLD_FONT
    level_c.alignment = CENTER
    level_c.border = ALL_BORDER

    ws1.row_dimensions[row_idx].height = 22

# 总分条件格式
ws1.conditional_formatting.add(
    f"J4:J{3 + n1}",
    ColorScaleRule(
        start_type="num", start_value=0, start_color="FECACA",
        mid_type="num", mid_value=25, mid_color="FEF3C7",
        end_type="num", end_value=50, end_color="D1FAE5"
    )
)

# 等级条件格式
ws1.conditional_formatting.add(
    f"K4:K{3 + n1}",
    CellIsRule(operator="equal", formula=['"优秀"'], fill=OK_FILL, font=Font(name="Microsoft YaHei", size=10, bold=True, color=OK))
)
ws1.conditional_formatting.add(
    f"K4:K{3 + n1}",
    CellIsRule(operator="equal", formula=['"良好"'], fill=PF("D1FAE5"))
)
ws1.conditional_formatting.add(
    f"K4:K{3 + n1}",
    CellIsRule(operator="equal", formula=['"需提升"'], fill=PF("FEF3C7"))
)
ws1.conditional_formatting.add(
    f"K4:K{3 + n1}",
    CellIsRule(operator="equal", formula=['"弱"'], fill=WARN_FILL, font=Font(name="Microsoft YaHei", size=10, color=NEUTRAL))
)

# 汇总统计
sum_row1 = 3 + n1 + 2
ws1.merge_cells(f"A{sum_row1}:K{sum_row1}")
c = ws1.cell(row=sum_row1, column=1, value="前测统计（自动计算）")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws1.row_dimensions[sum_row1].height = 24

stat_headers = ["指标", "问题定义", "根因分析", "方案设计", "执行跟踪", "复盘总结", "总分"]
for col_idx, h in enumerate(stat_headers, 1):
    c = ws1.cell(row=sum_row1 + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws1.row_dimensions[sum_row1 + 1].height = 28

stat_items = [
    ("平均分", f'=AVERAGE(E4:E{3 + n1})', 1),
    ("最高分", f'=MAX(E4:E{3 + n1})', 1),
    ("最低分", f'=MIN(E4:E{3 + n1})', 1),
    ("标准差", f'=STDEV(E4:E{3 + n1})', 2),
    ("及格率(>=6)", f'=COUNTIF(E4:E{3 + n1},">=6")/COUNTA(E4:E{3 + n1})', 3),
]

for i, (label, formula_e, decimals) in enumerate(stat_items):
    r = sum_row1 + 2 + i
    ws1.cell(row=r, column=1, value=label).font = BOLD_FONT
    ws1.cell(row=r, column=1).fill = BG_FILL
    ws1.cell(row=r, column=1).alignment = LEFT
    ws1.cell(row=r, column=1).border = ALL_BORDER

    # 5 个维度
    for col_idx, col_letter in enumerate(["E", "F", "G", "H", "I", "J"], 2):
        if label == "平均分":
            f = f'=AVERAGE({col_letter}4:{col_letter}{3 + n1})'
        elif label == "最高分":
            f = f'=MAX({col_letter}4:{col_letter}{3 + n1})'
        elif label == "最低分":
            f = f'=MIN({col_letter}4:{col_letter}{3 + n1})'
        elif label == "标准差":
            f = f'=STDEV({col_letter}4:{col_letter}{3 + n1})'
        elif label.startswith("及格率"):
            f = f'=COUNTIF({col_letter}4:{col_letter}{3 + n1},">=6")/COUNTA({col_letter}4:{col_letter}{3 + n1})'

        c = ws1.cell(row=r, column=col_idx, value=f)
        c.font = NORMAL_FONT
        c.alignment = CENTER
        c.border = ALL_BORDER
        if "率" in label:
            c.number_format = "0.0%"
        else:
            c.number_format = "0.00"
    ws1.row_dimensions[r].height = 22

ws1.freeze_panes = "A4"
ws1.sheet_view.showGridLines = False


# ============================================================
# Sheet 2: 后测数据
# ============================================================
ws2 = wb.create_sheet("后测数据")

ws2.merge_cells("A1:K1")
ws2["A1"] = "后测数据 - 培训后水平"
ws2["A1"].font = TITLE_FONT
ws2["A1"].fill = TITLE_FILL
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 32

ws2.merge_cells("A2:K2")
ws2["A2"] = "说明：课程结束后 1 天测评。同样 5 个能力维度。预期：相比前测有显著提升"
ws2["A2"].font = NOTE_FONT
ws2["A2"].fill = BG_FILL
ws2["A2"].alignment = LEFT
ws2.row_dimensions[2].height = 24

# 同结构表头
for col_idx, (text, width) in enumerate(headers1, 1):
    c = ws2.cell(row=3, column=col_idx, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.row_dimensions[3].height = 32

# 后测分数（相比前测 +2-3 分，呈现学习效果）
sample2 = []
for sid, name, dept, pos, *prev_scores in sample1:
    new_scores = []
    for s in prev_scores:
        # 大部分提升 2-3 分，封顶 10
        boost = random.choice([2, 2, 2, 3, 3, 1, 4])
        new_s = min(10, s + boost)
        new_scores.append(new_s)
    sample2.append((sid, name, dept, pos, *new_scores))

n2 = len(sample2)
for row_idx, row_data in enumerate(sample2, 4):
    for col_idx, value in enumerate(row_data, 1):
        c = ws2.cell(row=row_idx, column=col_idx, value=value)
        c.font = NORMAL_FONT
        c.alignment = CENTER
        c.border = ALL_BORDER
        if col_idx >= 5 and col_idx <= 9:
            c.number_format = "0"
            if value >= 7:
                c.fill = OK_FILL
                c.font = Font(name="Microsoft YaHei", size=10, bold=True, color=OK)

    # 总分
    total_c = ws2.cell(row=row_idx, column=10, value=f'=SUM(E{row_idx}:I{row_idx})')
    total_c.number_format = "0"
    total_c.font = BOLD_FONT
    total_c.alignment = CENTER
    total_c.border = ALL_BORDER
    total_c.fill = PF(ACCENT_SOFT)

    # 等级
    level_c = ws2.cell(row=row_idx, column=11, value=f'=IF(J{row_idx}>=40,"优秀",IF(J{row_idx}>=30,"良好",IF(J{row_idx}>=20,"一般",IF(J{row_idx}>=10,"需提升","弱"))))')
    level_c.font = BOLD_FONT
    level_c.alignment = CENTER
    level_c.border = ALL_BORDER

    ws2.row_dimensions[row_idx].height = 22

# 条件格式
ws2.conditional_formatting.add(
    f"J4:J{3 + n2}",
    ColorScaleRule(
        start_type="num", start_value=0, start_color="FECACA",
        mid_type="num", mid_value=25, mid_color="FEF3C7",
        end_type="num", end_value=50, end_color="D1FAE5"
    )
)

ws2.conditional_formatting.add(
    f"K4:K{3 + n2}",
    CellIsRule(operator="equal", formula=['"优秀"'], fill=OK_FILL, font=Font(name="Microsoft YaHei", size=10, bold=True, color=OK))
)
ws2.conditional_formatting.add(
    f"K4:K{3 + n2}",
    CellIsRule(operator="equal", formula=['"良好"'], fill=PF("D1FAE5"))
)
ws2.conditional_formatting.add(
    f"K4:K{3 + n2}",
    CellIsRule(operator="equal", formula=['"需提升"'], fill=PF("FEF3C7"))
)
ws2.conditional_formatting.add(
    f"K4:K{3 + n2}",
    CellIsRule(operator="equal", formula=['"弱"'], fill=WARN_FILL, font=Font(name="Microsoft YaHei", size=10, color=NEUTRAL))
)

# 后测统计
sum_row2 = 3 + n2 + 2
ws2.merge_cells(f"A{sum_row2}:K{sum_row2}")
c = ws2.cell(row=sum_row2, column=1, value="后测统计（自动计算）")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws2.row_dimensions[sum_row2].height = 24

for col_idx, h in enumerate(stat_headers, 1):
    c = ws2.cell(row=sum_row2 + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws2.row_dimensions[sum_row2 + 1].height = 28

for i, (label, _, _) in enumerate(stat_items):
    r = sum_row2 + 2 + i
    ws2.cell(row=r, column=1, value=label).font = BOLD_FONT
    ws2.cell(row=r, column=1).fill = BG_FILL
    ws2.cell(row=r, column=1).alignment = LEFT
    ws2.cell(row=r, column=1).border = ALL_BORDER

    for col_idx, col_letter in enumerate(["E", "F", "G", "H", "I", "J"], 2):
        if label == "平均分":
            f = f'=AVERAGE({col_letter}4:{col_letter}{3 + n2})'
        elif label == "最高分":
            f = f'=MAX({col_letter}4:{col_letter}{3 + n2})'
        elif label == "最低分":
            f = f'=MIN({col_letter}4:{col_letter}{3 + n2})'
        elif label == "标准差":
            f = f'=STDEV({col_letter}4:{col_letter}{3 + n2})'
        elif label.startswith("及格率"):
            f = f'=COUNTIF({col_letter}4:{col_letter}{3 + n2},">=6")/COUNTA({col_letter}4:{col_letter}{3 + n2})'

        c = ws2.cell(row=r, column=col_idx, value=f)
        c.font = NORMAL_FONT
        c.alignment = CENTER
        c.border = ALL_BORDER
        if "率" in label:
            c.number_format = "0.0%"
        else:
            c.number_format = "0.00"
    ws2.row_dimensions[r].height = 22

ws2.freeze_panes = "A4"
ws2.sheet_view.showGridLines = False


# ============================================================
# Sheet 3: 30天应用数据
# ============================================================
ws3 = wb.create_sheet("30天应用数据")

ws3.merge_cells("A1:K1")
ws3["A1"] = "30天应用数据 - 行为转化"
ws3["A1"].font = TITLE_FONT
ws3["A1"].fill = TITLE_FILL
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 32

ws3.merge_cells("A2:K2")
ws3["A2"] = "说明：30 天后跟踪学员的实际应用情况。1 行 1 名学员。覆盖 4 块看板的工具使用"
ws3["A2"].font = NOTE_FONT
ws3["A2"].fill = BG_FILL
ws3["A2"].alignment = LEFT
ws3.row_dimensions[2].height = 24

# 4 块看板的工具数
headers3 = [
    ("学员编号", 10),
    ("姓名", 10),
    ("定准板工具使用", 14),
    ("析透板工具使用", 14),
    ("策全板工具使用", 14),
    ("控稳板工具使用", 14),
    ("总使用次数", 12),
    ("完成方案数", 12),
    ("解决真问题数", 14),
    ("上级评价", 12),
    ("自我评价", 12),
]
for col_idx, (text, width) in enumerate(headers3, 1):
    c = ws3.cell(row=3, column=col_idx, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
    ws3.column_dimensions[get_column_letter(col_idx)].width = width
ws3.row_dimensions[3].height = 32

# 4 块看板的工具（定准2，析透3，策全3，控稳3 = 11）
dv_t1 = DataValidation(type="list", formula1='"0,1,2,3,4,5,6,7,8,9,10,15,20"', allow_blank=True)
ws3.add_data_validation(dv_t1)
dv_t1.add(f"C4:F{3 + n1}")

dv_t2 = DataValidation(type="list", formula1='"0,1,2,3,4,5,6,7,8,9,10"', allow_blank=True)
ws3.add_data_validation(dv_t2)
dv_t2.add(f"G4:I{3 + n1}")

dv_eval = DataValidation(type="list", formula1='"A+ 优秀,A 良好,B 一般,B- 需改进,C 差"', allow_blank=True)
ws3.add_data_validation(dv_eval)
dv_eval.add(f"J4:K{3 + n1}")

# 30 名学员的应用数据
sample3 = []
for i, (sid, name, _, _, *_) in enumerate(sample1):
    # 假设：约 60% 学员积极应用，30% 中等，10% 没应用
    if i < 18:  # 积极
        t1 = random.randint(3, 6)
        t2 = random.randint(4, 8)
        t3 = random.randint(3, 6)
        t4 = random.randint(3, 5)
        sol = random.randint(1, 3)
        real = random.randint(1, 3)
        mgr_eval = random.choice(["A+ 优秀", "A 良好", "A 良好"])
        self_eval = random.choice(["A+ 优秀", "A 良好", "A 良好"])
    elif i < 27:  # 中等
        t1 = random.randint(1, 3)
        t2 = random.randint(1, 4)
        t3 = random.randint(1, 3)
        t4 = random.randint(0, 3)
        sol = random.choice([0, 1, 1])
        real = random.choice([0, 1, 1])
        mgr_eval = random.choice(["B 一般", "A 良好", "B 一般"])
        self_eval = random.choice(["B 一般", "A 良好", "B 一般"])
    else:  # 弱
        t1 = random.randint(0, 1)
        t2 = random.randint(0, 1)
        t3 = random.randint(0, 1)
        t4 = random.randint(0, 1)
        sol = 0
        real = 0
        mgr_eval = random.choice(["B- 需改进", "C 差", "B 一般"])
        self_eval = random.choice(["B 一般", "B- 需改进"])

    sample3.append((sid, name, t1, t2, t3, t4, sol, real, mgr_eval, self_eval))

n3 = len(sample3)
for row_idx, row_data in enumerate(sample3, 4):
    for col_idx, value in enumerate(row_data, 1):
        c = ws3.cell(row=row_idx, column=col_idx, value=value)
        c.font = NORMAL_FONT
        c.alignment = CENTER
        c.border = ALL_BORDER
        if col_idx >= 3 and col_idx <= 9:
            c.number_format = "0"
            if col_idx == 7:  # 总使用次数
                c.fill = PF(ACCENT_SOFT)
                c.font = BOLD_FONT
            elif col_idx == 8:  # 完成方案
                c.fill = OK_FILL
                c.font = BOLD_FONT
            elif col_idx == 9:  # 解决真问题
                c.fill = PF("FCE7F3")
                c.font = BOLD_FONT

    # 总使用次数公式
    total_c = ws3.cell(row=row_idx, column=7, value=f'=SUM(C{row_idx}:F{row_idx})')
    total_c.number_format = "0"
    total_c.font = BOLD_FONT
    total_c.alignment = CENTER
    total_c.border = ALL_BORDER
    total_c.fill = PF(ACCENT_SOFT)

    ws3.row_dimensions[row_idx].height = 22

# 总使用次数数据条
ws3.conditional_formatting.add(
    f"G4:G{3 + n3}",
    DataBarRule(
        start_type="num", start_value=0,
        end_type="num", end_value=20,
        color=Color(rgb="FFB45309"),
        showValue=True
    )
)
# 完成方案数据条
ws3.conditional_formatting.add(
    f"H4:H{3 + n3}",
    DataBarRule(
        start_type="num", start_value=0,
        end_type="num", end_value=3,
        color=Color(rgb="FF166534"),
        showValue=True
    )
)
# 解决真问题数据条
ws3.conditional_formatting.add(
    f"I4:I{3 + n3}",
    DataBarRule(
        start_type="num", start_value=0,
        end_type="num", end_value=3,
        color=Color(rgb="FF9F1239"),
        showValue=True
    )
)

# 评价等级着色
for r in range(4, 4 + n3):
    for col in [10, 11]:
        v = ws3.cell(row=r, column=col).value
        if v == "A+ 优秀" or v == "A 良好":
            ws3.cell(row=r, column=col).fill = OK_FILL
            ws3.cell(row=r, column=col).font = Font(name="Microsoft YaHei", size=10, bold=True, color=OK)
        elif v == "B 一般":
            ws3.cell(row=r, column=col).fill = PF("FEF3C7")
        elif v == "B- 需改进":
            ws3.cell(row=r, column=col).fill = WARN_FILL
            ws3.cell(row=r, column=col).font = Font(name="Microsoft YaHei", size=10, color=WARN)
        elif v == "C 差":
            ws3.cell(row=r, column=col).fill = WARN_FILL
            ws3.cell(row=r, column=col).font = Font(name="Microsoft YaHei", size=10, color=NEUTRAL)

# 30天应用统计
sum_row3 = 3 + n3 + 2
ws3.merge_cells(f"A{sum_row3}:K{sum_row3}")
c = ws3.cell(row=sum_row3, column=1, value="30天应用统计")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws3.row_dimensions[sum_row3].height = 24

# 4 块看板 + 关键指标
stat3_headers = ["指标", "定准板", "析透板", "策全板", "控稳板", "总次数", "完成方案", "解决真问题", "应用率", "优秀率"]
for col_idx, h in enumerate(stat3_headers, 1):
    c = ws3.cell(row=sum_row3 + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws3.row_dimensions[sum_row3 + 1].height = 28

stat3_items = [
    ("平均", f'=AVERAGE(C4:C{3 + n3})', f'=AVERAGE(D4:D{3 + n3})', f'=AVERAGE(E4:E{3 + n3})', f'=AVERAGE(F4:F{3 + n3})', f'=AVERAGE(G4:G{3 + n3})', f'=AVERAGE(H4:H{3 + n3})', f'=AVERAGE(I4:I{3 + n3})'),
    ("最高", f'=MAX(C4:C{3 + n3})', f'=MAX(D4:D{3 + n3})', f'=MAX(E4:E{3 + n3})', f'=MAX(F4:F{3 + n3})', f'=MAX(G4:G{3 + n3})', f'=MAX(H4:H{3 + n3})', f'=MAX(I4:I{3 + n3})'),
    ("应用率(>=1)", f'=COUNTIF(C4:C{3 + n3},">=1")/COUNTA(C4:C{3 + n3})', f'=COUNTIF(D4:D{3 + n3},">=1")/COUNTA(D4:D{3 + n3})', f'=COUNTIF(E4:E{3 + n3},">=1")/COUNTA(E4:E{3 + n3})', f'=COUNTIF(F4:F{3 + n3},">=1")/COUNTA(F4:F{3 + n3})', f'=COUNTIF(G4:G{3 + n3},">=1")/COUNTA(G4:G{3 + n3})', f'=COUNTIF(H4:H{3 + n3},">=1")/COUNTA(H4:H{3 + n3})', f'=COUNTIF(I4:I{3 + n3},">=1")/COUNTA(I4:I{3 + n3})'),
]

for i, item in enumerate(stat3_items):
    label = item[0]
    r = sum_row3 + 2 + i
    ws3.cell(row=r, column=1, value=label).font = BOLD_FONT
    ws3.cell(row=r, column=1).fill = BG_FILL
    ws3.cell(row=r, column=1).alignment = LEFT
    ws3.cell(row=r, column=1).border = ALL_BORDER

    for col_idx, f in enumerate(item[1:], 2):
        c = ws3.cell(row=r, column=col_idx, value=f)
        c.font = NORMAL_FONT
        c.alignment = CENTER
        c.border = ALL_BORDER
        if "率" in label:
            c.number_format = "0.0%"
        else:
            c.number_format = "0.00"

    # 优秀率
    if i == 0:  # 仅第一行加"优秀率"
        c_exc = ws3.cell(row=r, column=10, value=f'=(COUNTIF(J4:J{3 + n3},"A+ 优秀")+COUNTIF(J4:J{3 + n3},"A 良好"))/COUNTA(J4:J{3 + n3})')
        c_exc.number_format = "0.0%"
        c_exc.font = BOLD_FONT
        c_exc.alignment = CENTER
        c_exc.border = ALL_BORDER
        c_exc.fill = OK_FILL
    else:
        c_exc = ws3.cell(row=r, column=10, value="")
        c_exc.border = ALL_BORDER

    ws3.row_dimensions[r].height = 22

ws3.freeze_panes = "A4"
ws3.sheet_view.showGridLines = False


# ============================================================
# Sheet 4: 汇总统计
# ============================================================
ws4 = wb.create_sheet("汇总统计")

ws4.merge_cells("A1:H1")
ws4["A1"] = "汇总统计 - 培训效果全景"
ws4["A1"].font = TITLE_FONT
ws4["A1"].fill = TITLE_FILL
ws4["A1"].alignment = CENTER
ws4.row_dimensions[1].height = 32

ws4.merge_cells("A2:H2")
ws4["A2"] = "说明：本表汇总 4 个维度的核心指标，自动计算前测-后测提升、行为转化率、应用效果等"
ws4["A2"].font = NOTE_FONT
ws4["A2"].fill = BG_FILL
ws4["A2"].alignment = LEFT
ws4.row_dimensions[2].height = 24

# === Block 1: 前后测对比 ===
ws4.merge_cells("A4:H4")
c = ws4.cell(row=4, column=1, value="一、前后测对比（能力提升）")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws4.row_dimensions[4].height = 24

# 表头
comp_headers = ["维度", "前测平均", "后测平均", "提升分", "提升率", "前测及格率", "后测及格率", "及格率提升"]
for col_idx, h in enumerate(comp_headers, 1):
    c = ws4.cell(row=5, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws4.row_dimensions[5].height = 30

dim_labels = ["问题定义", "根因分析", "方案设计", "执行跟踪", "复盘总结"]
for i, d in enumerate(dim_labels):
    r = 6 + i
    ws4.cell(row=r, column=1, value=d).font = BOLD_FONT
    ws4.cell(row=r, column=1).fill = BG_FILL
    ws4.cell(row=r, column=1).alignment = LEFT
    ws4.cell(row=r, column=1).border = ALL_BORDER

    # 列对应：问题定义=E, 根因分析=F, 方案设计=G, 执行跟踪=H, 复盘总结=I
    col_letter = ["E", "F", "G", "H", "I"][i]

    # 前测平均
    ws4.cell(row=r, column=2, value=f"=AVERAGE(前测数据!{col_letter}4:{col_letter}{3 + n1})").number_format = "0.00"
    # 后测平均
    ws4.cell(row=r, column=3, value=f"=AVERAGE(后测数据!{col_letter}4:{col_letter}{3 + n2})").number_format = "0.00"
    # 提升分
    ws4.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = "+0.00;-0.00;0.00"
    # 提升率
    ws4.cell(row=r, column=5, value=f"=(C{r}-B{r})/B{r}").number_format = "+0.0%;-0.0%;0.0%"
    # 前测及格率
    ws4.cell(row=r, column=6, value=f'=COUNTIF(前测数据!{col_letter}4:{col_letter}{3 + n1},">=6")/COUNTA(前测数据!{col_letter}4:{col_letter}{3 + n1})').number_format = "0.0%"
    # 后测及格率
    ws4.cell(row=r, column=7, value=f'=COUNTIF(后测数据!{col_letter}4:{col_letter}{3 + n2},">=6")/COUNTA(后测数据!{col_letter}4:{col_letter}{3 + n2})').number_format = "0.0%"
    # 及格率提升
    ws4.cell(row=r, column=8, value=f"=G{r}-F{r}").number_format = "+0.0%;-0.0%;0.0%"

    for col in range(2, 9):
        c = ws4.cell(row=r, column=col)
        c.font = NORMAL_FONT
        c.alignment = CENTER
        c.border = ALL_BORDER
    ws4.row_dimensions[r].height = 22

# 提升分/率条件格式
ws4.conditional_formatting.add(
    f"D6:D{5 + len(dim_labels)}",
    DataBarRule(
        start_type="num", start_value=0,
        end_type="num", end_value=5,
        color=Color(rgb="FF166534"),
        showValue=True
    )
)

# === Block 2: 30天应用情况 ===
sum2_row = 6 + len(dim_labels) + 2
ws4.merge_cells(f"A{sum2_row}:H{sum2_row}")
c = ws4.cell(row=sum2_row, column=1, value="二、30 天应用情况（行为转化）")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws4.row_dimensions[sum2_row].height = 24

app_headers = ["指标", "定准板", "析透板", "策全板", "控稳板", "总次数", "完成方案", "解决真问题"]
for col_idx, h in enumerate(app_headers, 1):
    c = ws4.cell(row=sum2_row + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws4.row_dimensions[sum2_row + 1].height = 30

app_items = [
    ("总使用次数", f'=SUM(\'30天应用数据\'!C4:C{3 + n3})', f'=SUM(\'30天应用数据\'!D4:D{3 + n3})', f'=SUM(\'30天应用数据\'!E4:E{3 + n3})', f'=SUM(\'30天应用数据\'!F4:F{3 + n3})', f'=SUM(\'30天应用数据\'!G4:G{3 + n3})', f'=SUM(\'30天应用数据\'!H4:H{3 + n3})', f'=SUM(\'30天应用数据\'!I4:I{3 + n3})'),
    ("使用人数(>=1次)", f'=COUNTIF(\'30天应用数据\'!C4:C{3 + n3},">=1")', f'=COUNTIF(\'30天应用数据\'!D4:D{3 + n3},">=1")', f'=COUNTIF(\'30天应用数据\'!E4:E{3 + n3},">=1")', f'=COUNTIF(\'30天应用数据\'!F4:F{3 + n3},">=1")', f'=COUNTIF(\'30天应用数据\'!G4:G{3 + n3},">=1")', f'=COUNTIF(\'30天应用数据\'!H4:H{3 + n3},">=1")', f'=COUNTIF(\'30天应用数据\'!I4:I{3 + n3},">=1")'),
    ("使用率", f'=COUNTIF(\'30天应用数据\'!C4:C{3 + n3},">=1")/{n3}', f'=COUNTIF(\'30天应用数据\'!D4:D{3 + n3},">=1")/{n3}', f'=COUNTIF(\'30天应用数据\'!E4:E{3 + n3},">=1")/{n3}', f'=COUNTIF(\'30天应用数据\'!F4:F{3 + n3},">=1")/{n3}', f'=COUNTIF(\'30天应用数据\'!G4:G{3 + n3},">=1")/{n3}', f'=COUNTIF(\'30天应用数据\'!H4:H{3 + n3},">=1")/{n3}', f'=COUNTIF(\'30天应用数据\'!I4:I{3 + n3},">=1")/{n3}'),
    ("平均使用次数", f'=AVERAGE(\'30天应用数据\'!C4:C{3 + n3})', f'=AVERAGE(\'30天应用数据\'!D4:D{3 + n3})', f'=AVERAGE(\'30天应用数据\'!E4:E{3 + n3})', f'=AVERAGE(\'30天应用数据\'!F4:F{3 + n3})', f'=AVERAGE(\'30天应用数据\'!G4:G{3 + n3})', f'=AVERAGE(\'30天应用数据\'!H4:H{3 + n3})', f'=AVERAGE(\'30天应用数据\'!I4:I{3 + n3})'),
]

for i, item in enumerate(app_items):
    r = sum2_row + 2 + i
    label = item[0]
    ws4.cell(row=r, column=1, value=label).font = BOLD_FONT
    ws4.cell(row=r, column=1).fill = BG_FILL
    ws4.cell(row=r, column=1).alignment = LEFT
    ws4.cell(row=r, column=1).border = ALL_BORDER

    for col_idx, f in enumerate(item[1:], 2):
        c = ws4.cell(row=r, column=col_idx, value=f)
        c.font = NORMAL_FONT
        c.alignment = CENTER
        c.border = ALL_BORDER
        if "率" in label:
            c.number_format = "0.0%"
        else:
            c.number_format = "0.0"

    ws4.row_dimensions[r].height = 22

# === Block 3: 综合效果 ===
sum3_row = sum2_row + 2 + len(app_items) + 2
ws4.merge_cells(f"A{sum3_row}:H{sum3_row}")
c = ws4.cell(row=sum3_row, column=1, value="三、综合培训效果指标")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws4.row_dimensions[sum3_row].height = 24

# 8 个核心指标
metrics = [
    ("学员总人数", f"={n1}", "0", "参与培训的学员数"),
    ("前测平均分", "=AVERAGE(前测数据!J4:J{0})".format(3 + n1), "0.00", "全员前测总分平均"),
    ("后测平均分", "=AVERAGE(后测数据!J4:J{0})".format(3 + n2), "0.00", "全员后测总分平均"),
    ("前后测平均提升", "=AVERAGE(后测数据!J4:J{0})-AVERAGE(前测数据!J4:J{0})".format(3 + n2, 3 + n1), "+0.00;-0.00;0.00", "总分提升（满分 50）"),
    ("后测优秀率", '=(COUNTIF(后测数据!K4:K{0},"优秀")+COUNTIF(后测数据!K4:K{0},"良好"))/{0}'.format(3 + n2, 3 + n2), "0.0%", "后测达到良好以上占比"),
    ("30 天使用率", "=SUMPRODUCT(--('30天应用数据'!G4:G{0}>=1))/{0}".format(3 + n3), "0.0%", "至少使用 1 次工具的学员占比"),
    ("30 天完成方案数", "=SUM('30天应用数据'!H4:H{0})".format(3 + n3), "0", "30 天内完成的问题解决方案总数"),
    ("30 天解决真问题数", "=SUM('30天应用数据'!I4:I{0})".format(3 + n3), "0", "30 天内实际解决的真问题数"),
    ("上级评价优秀率", '=(COUNTIF(\'30天应用数据\'!J4:J{0},"A+ 优秀")+COUNTIF(\'30天应用数据\'!J4:J{0},"A 良好"))/{0}'.format(3 + n3, 3 + n3), "0.0%", "上级对学员应用情况的评价"),
]

# 表头
m_headers = ["#", "指标", "数值", "单位", "说明"]
for col_idx, h in enumerate(m_headers, 1):
    if col_idx == 1:
        continue
    c = ws4.cell(row=sum3_row + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws4.cell(row=sum3_row + 1, column=1, value="#").font = HEADER_FONT
ws4.cell(row=sum3_row + 1, column=1).fill = HEADER_FILL
ws4.cell(row=sum3_row + 1, column=1).alignment = CENTER
ws4.cell(row=sum3_row + 1, column=1).border = ALL_BORDER
ws4.row_dimensions[sum3_row + 1].height = 28

for i, (label, formula, fmt, desc) in enumerate(metrics):
    r = sum3_row + 2 + i
    ws4.cell(row=r, column=1, value=i + 1).font = BOLD_FONT
    ws4.cell(row=r, column=1).fill = BG_FILL
    ws4.cell(row=r, column=1).alignment = CENTER
    ws4.cell(row=r, column=1).border = ALL_BORDER

    ws4.cell(row=r, column=2, value=label).font = BOLD_FONT
    ws4.cell(row=r, column=2).fill = BG_FILL
    ws4.cell(row=r, column=2).alignment = LEFT
    ws4.cell(row=r, column=2).border = ALL_BORDER

    c_val = ws4.cell(row=r, column=3, value=formula)
    c_val.font = Font(name="Microsoft YaHei", size=12, bold=True, color=ACCENT)
    c_val.alignment = CENTER
    c_val.border = ALL_BORDER
    c_val.number_format = fmt
    c_val.fill = PF(ACCENT_SOFT)

    ws4.cell(row=r, column=4, value=fmt.replace("0", "").replace(".", "").replace("+", "").replace("-", "").replace(";", "")).font = NORMAL_FONT
    ws4.cell(row=r, column=4).alignment = CENTER
    ws4.cell(row=r, column=4).border = ALL_BORDER

    ws4.cell(row=r, column=5, value=desc).font = SOFT_FONT
    ws4.cell(row=r, column=5).alignment = LEFT
    ws4.cell(row=r, column=5).border = ALL_BORDER
    ws4.merge_cells(f"E{r}:H{r}")

    ws4.row_dimensions[r].height = 24

# === Block 4: 维度雷达图（手工绘制） ===
chart_row = sum3_row + 2 + len(metrics) + 2

# 添加柱状图 - 前后测平均对比
chart = BarChart()
chart.type = "col"
chart.style = 11
chart.title = "5 个能力维度：前后测对比"
chart.x_axis.title = "能力维度"
chart.y_axis.title = "平均分（0-10）"

# 数据范围
pre_data = Reference(ws4, min_col=2, min_row=5, max_row=10, max_col=2)
post_data = Reference(ws4, min_col=3, min_row=5, max_row=10, max_col=3)
cats = Reference(ws4, min_col=1, min_row=6, max_row=10)

chart.add_data(pre_data, titles_from_data=True)
chart.add_data(post_data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 10
chart.width = 18
ws4.add_chart(chart, f"J4")

# 雷达图 - 5 个维度提升率
radar = RadarChart()
radar.type = "filled"
radar.style = 26
radar.title = "5 维度提升率（%）"

radar_data = Reference(ws4, min_col=5, min_row=5, max_row=10, max_col=5)
radar_cats = Reference(ws4, min_col=1, min_row=6, max_row=10)
radar.add_data(radar_data, titles_from_data=True)
radar.set_categories(radar_cats)
radar.height = 10
radar.width = 16
ws4.add_chart(radar, f"J24")

# 工具使用柱状图 - 4 块看板
chart2 = BarChart()
chart2.type = "col"
chart2.style = 12
chart2.title = "4 块看板工具总使用次数"
chart2.x_axis.title = "看板"
chart2.y_axis.title = "使用次数"

tool_data = Reference(ws4, min_col=2, min_row=sum2_row + 1, max_row=sum2_row + 1, max_col=5)
tool_cats = Reference(ws4, min_col=2, min_row=sum2_row + 2, max_row=sum2_row + 2, max_col=5)
chart2.add_data(tool_data, titles_from_data=True)
chart2.set_categories(tool_cats)
chart2.height = 10
chart2.width = 16
ws4.add_chart(chart2, f"J44")

ws4.freeze_panes = "A4"
ws4.sheet_view.showGridLines = False


# ============================================================
output_path = r"D:\2026年课程\竞越\基层即学即用的问题解决工具箱\补充课程包\16_配套表单Excel\16_培训效果数据汇总_组织用.xlsx"
wb.save(output_path)
print(f"Generated: {output_path}")
print(f"4 sheets: 前测数据 ({n1}) / 后测数据 ({n2}) / 30天应用数据 ({n3}) / 汇总统计")

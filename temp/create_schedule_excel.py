import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "互动流程时间表"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=10)
module_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
activity_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
demo_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
break_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
discuss_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Headers
headers = ["序号", "时间", "时长", "模块/活动", "活动类型", "详细内容", "产出", "负责人", "备注"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = module_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Course schedule data (420 minutes total)
schedule = [
    [1, "09:00-09:10", 10, "签到与开场", "签到", "学员签到，领取资料", "签到表", "教务组", ""],
    [2, "09:10-09:20", 10, "课程介绍", "讲授", "课程目标、内容概览、学习收益", "学员对课程有清晰预期", "讲师", ""],
    [3, "09:20-09:50", 30, "M1: 企业大学角色转型全景图", "理论+案例", "企业大学演进四阶段、四大死局、三大转型方向", "转型方向共识", "讲师", "含视频播放"],
    [4, "09:50-10:10", 20, "M1: 四大死局深度解析", "小组讨论", "学员对照四大死局，自我诊断自身企业大学困境", "诊断结果卡片", "讲师", "8分钟讨论+12分钟分享"],
    [5, "10:10-10:25", 15, "茶歇", "休息", "茶歇+自由交流", "", "教务组", ""],
    [6, "10:25-10:50", 25, "M1: 企业大学新生四大方向", "案例分析", "标杆案例：华为企业大学转型路径", "转型方向共识", "讲师", ""],
    [7, "10:50-11:10", 20, "M1: 转型方向共识Workshop", "小组讨论", "各组产出转型方向宣言", "转型宣言", "讲师", "10分钟研讨+10分钟展示"],
    [8, "11:10-11:40", 30, "M2: 知识资产化概述", "理论讲解", "知识资产化三阶段：数字化入库→结构化重组→智能化应用", "知识资产化框架", "讲师", ""],
    [9, "11:40-12:00", 20, "M2: 知识资产现状诊断", "工具演示", "使用课程资产健康度评估表快速自测", "资产盘点清单", "讲师+助教", "含实操练习"],
    [10, "12:00-13:30", 90, "午餐休息", "休息", "午餐+休息", "", "教务组", ""],
    [11, "13:30-13:50", 20, "M2: 知识资产活化三阶段", "案例分析", "案例：某制造业企业3000门课程3个月完成数字化", "资产化路径图", "讲师", ""],
    [12, "13:50-14:20", 30, "M2: 知识资产估值体系", "工具演示", "估值模型讲解+实操练习", "资产估值表", "讲师", "含计算练习"],
    [13, "14:20-14:50", 30, "M2: 知识资产化行动计划Workshop", "小组讨论", "盘点、分类、优先级、行动计划", "行动计划初稿", "讲师", "含各组分享"],
    [14, "14:50-15:05", 15, "茶歇", "休息", "茶歇+自由交流", "", "教务组", ""],
    [15, "15:05-15:35", 30, "M3: AI Agent基础概念", "理论讲解", "AI Agent定义、核心组件、与传统软件的区别", "Agent理解度", "讲师", "含互动问答"],
    [16, "15:35-15:55", 20, "M3: 培训场景Agent矩阵", "案例分析", "5大类Agent场景解析", "Agent矩阵图", "讲师", ""],
    [17, "15:55-16:35", 40, "M3: 三大核心Agent深度演示", "现场演示", "培训顾问Agent、新员工导航Agent、销售知识Agent", "Agent体验报告", "助教", "45分钟演示脚本"],
    [18, "16:35-16:50", 15, "茶歇", "休息", "茶歇+自由交流", "", "教务组", ""],
    [19, "16:50-17:20", 30, "M3: Agent孵化平台架构", "理论讲解", "平台架构、知识中台、开发工具", "平台架构图", "讲师", ""],
    [20, "17:20-17:50", 30, "M4: 四阶段转型路线图", "理论讲解", "觉醒期→奠基期→突破期→引领期", "路线图框架", "讲师", ""],
    [21, "17:50-18:15", 25, "M4: 各阶段详解与风险", "案例分析", "每阶段关键任务、里程碑、风险应对", "阶段计划", "讲师", ""],
    [22, "18:15-18:30", 15, "当日总结与答疑", "讨论", "回顾全天要点，答疑", "", "讲师", ""],
    [23, "09:00-09:10", 10, "回顾与答疑", "讨论", "回顾昨日内容，解答疑问", "", "讲师", ""],
    [24, "09:10-09:40", 30, "M4: 路线图制定Workshop", "小组讨论", "制定本企业12个月路线图", "路线图初稿", "讲师", "各组展示"],
    [25, "09:40-10:10", 30, "M5: 六步孵化法", "理论讲解", "需求发现→角色定义→知识灌注→Prompt工程→测试优化→上线运营", "孵化流程图", "讲师", ""],
    [26, "10:10-10:25", 15, "茶歇", "休息", "茶歇+自由交流", "", "教务组", ""],
    [27, "10:25-10:40", 15, "M5: 分组选题", "分组", "选择Agent孵化场景（5选1）", "选题结果", "讲师", ""],
    [28, "10:40-11:30", 50, "M5: 实战孵化", "实操", "确定Agent定位→准备知识库→编写Prompt", "Agent原型", "讲师+助教", "50分钟实操"],
    [29, "11:30-11:50", 20, "M5: 成果展示", "讨论", "每组2分钟展示+点评", "展示反馈", "讲师", ""],
    [30, "11:50-12:00", 10, "M5: 总结", "讲授", "孵化要点回顾", "", "讲师", ""],
    [31, "12:00-13:30", 90, "午餐休息", "休息", "午餐+休息", "", "教务组", ""],
    [32, "13:30-13:50", 20, "M6: 转型效果评估体系", "理论讲解", "三维度模型（员工体验/运营效率/业务价值）", "评估框架", "讲师", ""],
    [33, "13:50-14:10", 20, "M6: 知识资产化验收标准", "工具演示", "活化率评级、知识资产估值计算", "验收清单", "讲师", ""],
    [34, "14:10-14:25", 15, "茶歇", "休息", "茶歇+自由交流", "", "教务组", ""],
    [35, "14:25-14:40", 15, "M6: Agent孵化验收标准", "讲授", "核心指标、日活、满意度、意图准确率", "验收指标", "讲师", ""],
    [36, "14:40-15:10", 30, "M6: 个人行动承诺", "讨论", "90天行动承诺，填写承诺卡", "承诺卡", "讲师", "含分享"],
    [37, "15:10-15:30", 20, "课程总结与展望", "讲授", "回顾两天课程要点，展望企业大学未来", "", "讲师", ""],
    [38, "15:30-15:40", 10, "结业与合影", "仪式", "颁发结业证书，合影留念", "结业照片", "教务组", ""],
]

# Write data
for row_idx, row_data in enumerate(schedule, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Color coding by activity type
        if row_data[4] == "理论讲解" or row_data[4] == "讲授":
            cell.fill = module_fill
            cell.font = Font(color="FFFFFF")
        elif row_data[4] == "小组讨论" or row_data[4] == "讨论":
            cell.fill = discuss_fill
        elif row_data[4] == "现场演示" or row_data[4] == "工具演示":
            cell.fill = demo_fill
        elif row_data[4] == "休息":
            cell.fill = break_fill
        elif row_data[4] == "实操" or row_data[4] == "分组":
            cell.fill = activity_fill
        elif row_idx % 2 == 0:
            cell.fill = alt_fill

# Column widths
widths = [5, 12, 8, 25, 12, 40, 20, 10, 15]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Row heights
for row in range(1, len(schedule) + 2):
    ws.row_dimensions[row].height = 25

# Add summary sheet
ws2 = wb.create_sheet("时间分配汇总")

summary_headers = ["活动类型", "总时长(分钟)", "占比", "说明"]
for col, header in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = module_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

summary_data = [
    ["理论讲解/讲授", 165, "39%", "核心框架和概念讲解"],
    ["案例分析", 95, "23%", "标杆案例和场景分析"],
    ["小组讨论/Workshop", 85, "20%", "互动研讨和实战练习"],
    ["现场演示/工具演示", 55, "13%", "工具操作和效果展示"],
    ["茶歇/休息", 130, "31%", "含午餐时间"],
    ["总计", 420, "100%", "7小时完整课程"],
]

for row_idx, row_data in enumerate(summary_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        if row_idx % 2 == 0:
            cell.fill = alt_fill

for i, width in enumerate([20, 15, 10, 30], 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

wb.save("D:/新课开发/企业大学/对内/1.企业大学重生：从内训中心到知识资产与智能体孵化枢纽的角色转型/成果demo/07_互动流程时间表.xlsx")
print("07_互动流程时间表.xlsx created successfully")

# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os

output_dir = "D:/新课开发/公文写作/3、人机协同写作——从结构化提示到可用初稿/全流程工具表单/Excel版"
os.makedirs(output_dir, exist_ok=True)

header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
normal_font = Font(size=10)
bold_font = Font(size=10, bold=True)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
content_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_row_style(ws, row, num_cols, is_header=False, is_gray=False):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        if is_header:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        else:
            cell.fill = gray_fill if is_gray else white_fill
            cell.font = normal_font
            cell.alignment = content_alignment
        cell.border = thin_border

def set_col_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_f1():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "提示词结构卡"
    set_col_widths(ws, [8, 20, 45])
    ws.merge_cells("A1:C1")
    ws.cell(row=1, column=1, value="F1：提示词结构卡 - 正面：结构化提示的8个要素")
    set_row_style(ws, 1, 3, is_header=True)
    ws.row_dimensions[1].height = 25
    ws.merge_cells("A2:C2")
    ws.cell(row=2, column=1, value="提示词结构检查表 - 8个要素，缺一不可")
    set_row_style(ws, 2, 3, is_header=True)
    ws.row_dimensions[2].height = 20
    elements = [
        ("【1】角色定义", "我要以什么身份来写？", ""),
        ("【2】任务目标", "具体要写一份什么文件？", ""),
        ("【3】受众分析", "读者是谁？他们的立场和关注点？", ""),
        ("【4】文种确定", "用哪种公文文种？（通知/报告/请示/函...）", ""),
        ("【5】内容要点", "必须包含哪些关键信息？", ""),
        ("【6】格式规范", "有什么排版、篇幅、结构要求？", ""),
        ("【7】语气风格", "正式程度？严肃/平和/鼓舞？", ""),
        ("【8】禁止事项", "有什么不能写、不能提的？", ""),
    ]
    row = 3
    for i, (num, question, answer) in enumerate(elements):
        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=question)
        ws.cell(row=row, column=3, value=answer)
        set_row_style(ws, row, 3, is_gray=(i % 2 == 1))
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value="背面：提示词公式 + 使用流程")
    set_row_style(ws, row, 3, is_header=True)
    ws.row_dimensions[row].height = 20
    row += 1
    formula = "作为[角色]，请以[文种]形式，围绕[主题]，面向[受众]，撰写一份[标题]，要求：1. [要点1] 2. [要点2] 3. [要点3] 语气：[风格] 篇幅：[要求] 禁止：[禁忌]"
    ws.cell(row=row, column=1, value="提示词公式：")
    ws.cell(row=row, column=1).font = bold_font
    ws.row_dimensions[row].height = 20
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value=formula)
    ws.cell(row=row, column=1).alignment = content_alignment
    set_row_style(ws, row, 3, is_gray=True)
    ws.row_dimensions[row].height = 80
    row += 1
    steps = ["1. 先填正面8要素 → 明确我要什么", "2. 再套公式组装 → 形成完整提示词", "3. 发给AI前检查 → 确认8要素齐全", "4. AI生成后对照 → 验证是否按要求执行"]
    for step in steps:
        ws.cell(row=row, column=1, value=step)
        set_row_style(ws, row, 3, is_gray=(row % 2 == 0))
        ws.row_dimensions[row].height = 20
        row += 1
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value="使用说明")
    set_row_style(ws, row, 3, is_header=True)
    ws.row_dimensions[row].height = 20
    row += 1
    headers = ["要素", "常见错误", "正确做法"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
        set_row_style(ws, row, 3, is_header=True)
    ws.row_dimensions[row].height = 20
    row += 1
    table_data = [
        ("角色定义", ""帮我写"", "明确岗位：如"以综合部秘书身份""),
        ("任务目标", ""写得好一点"", "具体：如"写一份年度工作总结""),
        ("受众分析", "忽略", "说明：如"面向集团领导，注重数据""),
        ("文种确定", "混用", "准确：如"请示"而非"申请""),
        ("内容要点", "遗漏", "逐条列出"),
        ("格式规范", "模糊", "明确：如"不超过800字""),
        ("语气风格", "矛盾", "统一：如"严肃但不刻板""),
        ("禁止事项", "忘记", "主动排查政治/敏感内容"),
    ]
    for i, (ele, error, correct) in enumerate(table_data):
        ws.cell(row=row, column=1, value=ele)
        ws.cell(row=row, column=2, value=error)
        ws.cell(row=row, column=3, value=correct)
        set_row_style(ws, row, 3, is_gray=(i % 2 == 1))
        ws.row_dimensions[row].height = 20
        row += 1
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value="提示：8个要素越清晰，AI输出质量越高。宁可多写，不要遗漏。")
    ws.cell(row=row, column=1).font = Font(size=10, italic=True, color="666666")
    ws.row_dimensions[row].height = 20
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    wb.save(os.path.join(output_dir, "F1_提示词结构卡.xlsx"))
    print("Created F1_提示词结构卡.xlsx")

create_f1()
print("F1 done")

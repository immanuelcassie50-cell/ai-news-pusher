"""
领航·4.0 管理者工具包·03_可视化看板_Excel版.xlsx
6 个 Sheet 的团队管理改进看板
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.formatting.rule import (
    CellIsRule, FormulaRule, ColorScaleRule
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import RadarChart, LineChart, Reference

# ============== 颜色定义（深蓝色 + 4 个状态色） ==============
NAVY = "1F3864"          # 主标题深蓝
LIGHT_NAVY = "2E5597"    # 二级标题
HEADER_FILL = "1F3864"   # 表头深蓝
SUBHEADER_FILL = "D9E1F2"  # 浅蓝副标题
WHITE = "FFFFFF"

# 4 状态色
GREEN_FILL = "C6EFCE"   # 健康
YELLOW_FILL = "FFEB9C"  # 关注
ORANGE_FILL = "FFD699"  # 警告
RED_FILL = "F4B7B7"     # 紧急
GRAY_FILL = "F2F2F2"    # 中性底

# 字体色
GREEN_FONT = "006100"
YELLOW_FONT = "9C5700"
ORANGE_FONT = "9C5700"
RED_FONT = "9C0006"
NAVY_FONT = "1F3864"

# ============== 通用样式 ==============
def make_border():
    thin = Side(border_style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def set_title(ws, row, col, text, span=8, fill=NAVY, font_color=WHITE, size=14):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="微软雅黑", size=size, bold=True, color=font_color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 28
    # 给合并区域内其他单元格填充色
    for c in range(col, col+span):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=fill)

def set_subtitle(ws, row, col, text, span=8):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="微软雅黑", size=11, bold=True, color=NAVY_FONT)
    cell.fill = PatternFill("solid", fgColor=SUBHEADER_FILL)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 22
    for c in range(col, col+span):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=SUBHEADER_FILL)

def set_header_row(ws, row, col, headers, fill=NAVY):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col+i, value=h)
        c.font = Font(name="微软雅黑", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = make_border()
    ws.row_dimensions[row].height = 32

def set_cell(ws, row, col, value, font_size=10, bold=False, color="000000",
             fill=None, align="left", wrap=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="微软雅黑", size=font_size, bold=bold, color=color, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = make_border()
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============== 创建 Workbook ==============
wb = Workbook()
wb.remove(wb.active)

# =====================================================================
# Sheet 1: 00_使用说明
# =====================================================================
ws1 = wb.create_sheet("00_使用说明")
set_col_widths(ws1, [3, 18, 18, 18, 18, 18, 18, 18, 18, 3])

# 标题
set_title(ws1, 1, 2, "领航·4.0 团队管理改进看板·使用说明", span=8)

# 副标题
set_subtitle(ws1, 2, 2, "开发者：罗宏伟 | 适用读者：参训管理者的直属上级、HRBP | 训后版 v1.0", span=8)

# 设计意图
set_subtitle(ws1, 4, 2, "一、看板设计意图", span=8)
intro_lines = [
    ("为什么需要这张表：", "Excel 看板是整个训后工具包的'驾驶舱'——所有观察、对话、邮件都汇入这张表，让学员的改变可衡量、可持续、可向上汇报。"),
    ("核心原则：", "观察不评判、对话不审问、反馈不评判。"),
    ("使用节奏：", "训后 7 天填基础信息；30/60/90 天前填月度观察；每场对话后更新对话追踪；每季度更新行为指标雷达。"),
    ("填写人：", "学员主管是主要填写人；HRBP 可以同时维护'全员'Sheet 做横向对比。"),
    ("使用前提：", "请先通读 00 导读和 README.md，再使用本看板。"),
]
r = 5
for k, v in intro_lines:
    set_cell(ws1, r, 2, k, bold=True, color=NAVY_FONT, fill=SUBHEADER_FILL, align="left")
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    c = ws1.cell(row=r, column=3, value=v)
    c.font = Font(name="微软雅黑", size=10, color="000000")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = make_border()
    c.fill = PatternFill("solid", fgColor=WHITE)
    for cc in range(3, 10):
        ws1.cell(row=r, column=cc).border = make_border()
        ws1.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=WHITE)
    ws1.row_dimensions[r].height = 32
    r += 1

# 6 个 Sheet 含义
r += 1
set_subtitle(ws1, r, 2, "二、6 个 Sheet 的含义", span=8)
r += 1
set_header_row(ws1, r, 2, ["Sheet 编号", "Sheet 名称", "填什么", "谁填", "填的频率", "关联的工具包文件"])
r += 1
sheet_desc = [
    ("Sheet 1", "00_使用说明", "你正在看的", "——", "一次性", "00 导读 / README"),
    ("Sheet 2", "01_学员基础信息", "学员姓名/部门/课程日期/分析对象/问责伙伴", "学员主管", "训后第 1 周", "00 导读 / 04 邮件 1"),
    ("Sheet 3", "02_五项管理动作_月度记录", "5 个观察点的月度记录（30/60/90 天各一行）", "学员主管", "每月一次", "01 观察指南"),
    ("Sheet 4", "03_30-60-90天对话追踪", "三场对话的状态、关键产出、下一步", "学员主管", "每场对话后", "02 脚本 / 04 邮件 2/3"),
    ("Sheet 5", "04_行为指标变化_雷达图", "5 感前测/30天/60天/90天得分", "学员主管 / HRBP", "每季度", "02 脚本 / 看板雷达图"),
    ("Sheet 6", "05_信号灯与风险预警", "6 类信号灯（绿/黄/橙/红）", "系统自动 + 人工备注", "每月", "02 脚本 / 04 邮件 B"),
    ("Sheet 7", "06_团队层面业务结果", "团队 3-5 个业务指标月度趋势", "学员主管 + 业务部门", "每月", "02 脚本 / 04 邮件 C"),
]
for s in sheet_desc:
    for i, v in enumerate(s):
        c = set_cell(ws1, r+i-1 if False else r, 2+i, v, align="center" if i in [0,4] else "left")
    r += 1
    ws1.row_dimensions[r-1].height = 30

# 三原则
r += 1
set_subtitle(ws1, r, 2, "三、三原则（请先记住）", span=8)
r += 1
principles = [
    ("原则一", "观察不评判", "你的角色是记录员，不是裁判员", GREEN_FILL, GREEN_FONT),
    ("原则二", "对话不审问", "辅导对话不是绩效面谈，是陪练", YELLOW_FILL, YELLOW_FONT),
    ("原则三", "反馈不评判", "做法不一致时先理解再回应", ORANGE_FILL, ORANGE_FONT),
]
for tag, name, desc, fill, fcolor in principles:
    set_cell(ws1, r, 2, tag, bold=True, fill=fill, color=fcolor, align="center")
    set_cell(ws1, r, 3, name, bold=True, fill=fill, color=fcolor, align="center")
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    c = ws1.cell(row=r, column=4, value=desc)
    c.font = Font(name="微软雅黑", size=10, color=fcolor)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor=fill)
    c.border = make_border()
    for cc in range(4, 10):
        ws1.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=fill)
        ws1.cell(row=r, column=cc).border = make_border()
    ws1.row_dimensions[r].height = 26
    r += 1

# 颜色说明
r += 1
set_subtitle(ws1, r, 2, "四、信号灯颜色含义", span=8)
r += 1
set_header_row(ws1, r, 2, ["颜色", "状态", "含义", "你的动作", "", "", "", ""])
r += 1
colors_def = [
    ("绿色", "健康", "所有指标正常，行为改变可持续", "继续观察，下月再填", GREEN_FILL, GREEN_FONT),
    ("黄色", "关注", "部分指标有下滑信号", "下次 1v1 自然问 1-2 个相关问题", YELLOW_FILL, YELLOW_FONT),
    ("橙色", "警告", "多个指标下滑，需要主动清障", "下次正式对话增加 1-2 个问题", ORANGE_FILL, ORANGE_FONT),
    ("红色", "紧急", "核心指标明显异常", "安排 30 分钟专项对话（不批评）", RED_FILL, RED_FONT),
]
for c1, c2, c3, c4, fill, fcolor in colors_def:
    set_cell(ws1, r, 2, c1, bold=True, fill=fill, color=fcolor, align="center")
    set_cell(ws1, r, 3, c2, bold=True, fill=fill, color=fcolor, align="center")
    set_cell(ws1, r, 4, c3, fill=fill, color=fcolor, align="left")
    ws1.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
    c = ws1.cell(row=r, column=5, value=c4)
    c.font = Font(name="微软雅黑", size=10, color=fcolor)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor=fill)
    c.border = make_border()
    for cc in range(5, 10):
        ws1.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=fill)
        ws1.cell(row=r, column=cc).border = make_border()
    ws1.row_dimensions[r].height = 26
    r += 1

# 冻结
ws1.freeze_panes = "A2"
ws1.sheet_view.showGridLines = False

print("Sheet 1 完成")

# =====================================================================
# Sheet 2: 01_学员基础信息
# =====================================================================
ws2 = wb.create_sheet("01_学员基础信息")
set_col_widths(ws2, [3, 22, 30, 22, 30, 22, 30, 3])

set_title(ws2, 1, 2, "学员基础信息登记表", span=6)
set_subtitle(ws2, 2, 2, "请在训后第 1 周内填写完整", span=6)

# 表头
set_header_row(ws2, 4, 2, ["字段", "内容", "字段", "内容", "字段", "内容"])

# 字段行
fields_left = [
    ("学员姓名", "请填写"),
    ("所属部门", "请填写"),
    ("职位", "请填写"),
    ("学员上级（你）", "请填写"),
    ("HRBP", "请填写"),
    ("课程日期", "2026/--/--"),
    ("训后 30 天日期", "=E7+30 (自动)"),
    ("训后 60 天日期", "=E7+60 (自动)"),
    ("训后 90 天日期", "=E7+90 (自动)"),
    ("问责伙伴", "请填写"),
]
fields_right = [
    ("问责伙伴联系方式", "请填写"),
    ("Z 世代分析对象 1", "姓名/工号"),
    ("分析对象 1 入职时间", "请填写"),
    ("Z 世代分析对象 2", "姓名/工号"),
    ("团队总人数", "请填写"),
    ("团队中 Z 世代员工数", "请填写"),
    ("30 天清单提交日期", "=E7+7 (自动)"),
    ("30 天对话日期", "=E7+30 (自动)"),
    ("60 天对话日期", "=E7+60 (自动)"),
    ("90 天对话日期", "=E7+90 (自动)"),
]

# 左 3 列字段
r = 5
for i, (k, v) in enumerate(fields_left):
    set_cell(ws2, r, 2, k, bold=True, fill=SUBHEADER_FILL, color=NAVY_FONT, align="left")
    set_cell(ws2, r, 3, v, align="left")
    r += 1

# 右 3 列字段
r = 5
for i, (k, v) in enumerate(fields_right):
    set_cell(ws2, r, 4, k, bold=True, fill=SUBHEADER_FILL, color=NAVY_FONT, align="left")
    set_cell(ws2, r, 5, v, align="left")
    # 这里有 6 列，列 6 作为附注
    set_cell(ws2, r, 6, "—", align="center", color="808080")
    r += 1

# 备注行
r += 1
set_subtitle(ws2, r, 2, "备注：本表为基础信息，填写后请勿随意修改", span=6)
ws2.freeze_panes = "A5"
ws2.sheet_view.showGridLines = False

print("Sheet 2 完成")

# =====================================================================
# Sheet 3: 02_五项管理动作_月度记录
# =====================================================================
ws3 = wb.create_sheet("02_五项管理动作_月度记录")
set_col_widths(ws3, [3, 12, 18, 30, 30, 20, 20, 3])

set_title(ws3, 1, 2, "五项管理动作·月度记录", span=6)
set_subtitle(ws3, 2, 2, "对应 01 观察指南的 5 个观察点 | 每月一次 | 30/60/90 天时回顾", span=6)

# 表头（两行）
set_header_row(ws3, 4, 2, ["月份", "观察点", "学员做到了什么（具体行为）", "学员没做到 / 卡住的地方", "你的判断 / 下一步"])
# 第二行表头
ws3.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
ws3.merge_cells(start_row=4, start_column=3, end_row=5, end_column=3)
ws3.merge_cells(start_row=4, start_column=4, end_row=5, end_column=4)
ws3.merge_cells(start_row=4, start_column=5, end_row=5, end_column=5)
ws3.merge_cells(start_row=4, start_column=6, end_row=5, end_column=6)
ws3.row_dimensions[4].height = 24

# 5 个观察点
obs_points = [
    "观察点 1：响应节奏变化",
    "观察点 2：五感管理动作（'你的判断'）",
    "观察点 3：AI 话题的开口",
    "观察点 4：任务分配 5W2H+H",
    "观察点 5：激励机制更新（人类贡献目标）",
]

# 3 个月：30/60/90（基线 + 3 次月度）
months = ["训后基线（训后第 7 天）", "训后 30 天", "训后 60 天", "训后 90 天"]

r = 6
data_start_row = r
for m in months:
    for i, op in enumerate(obs_points):
        # 月份（只在第一个观察点写）
        if i == 0:
            set_cell(ws3, r, 2, m, bold=True, fill=SUBHEADER_FILL, color=NAVY_FONT, align="center")
        else:
            set_cell(ws3, r, 2, "", fill=SUBHEADER_FILL, align="center")
        set_cell(ws3, r, 3, op, align="left")
        set_cell(ws3, r, 4, "（请填写具体行为，例如：'他在周会上说了\"我们约定一下消息响应时间\"'）",
                 align="left", color="808080", italic=True)
        set_cell(ws3, r, 5, "（请填写卡点）", align="left", color="808080", italic=True)
        set_cell(ws3, r, 6, "（请填写你的判断和下一步支持）", align="left", color="808080", italic=True)
        ws3.row_dimensions[r].height = 48
        r += 1
    # 月份之间空一行
    set_cell(ws3, r, 2, "本月最让我意外的一个发现：", bold=True, color=NAVY_FONT, fill=GRAY_FILL, align="right")
    ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    c = ws3.cell(row=r, column=3, value="（请填写）")
    c.font = Font(name="微软雅黑", size=10, color="808080", italic=True)
    c.fill = PatternFill("solid", fgColor=GRAY_FILL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = make_border()
    for cc in range(3, 7):
        ws3.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=GRAY_FILL)
        ws3.cell(row=r, column=cc).border = make_border()
    ws3.row_dimensions[r].height = 28
    r += 1
    # 下一月份
    set_cell(ws3, r, 2, "我下个月要重点支持的一件事：", bold=True, color=NAVY_FONT, fill=GRAY_FILL, align="right")
    ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    c = ws3.cell(row=r, column=3, value="（请填写）")
    c.font = Font(name="微软雅黑", size=10, color="808080", italic=True)
    c.fill = PatternFill("solid", fgColor=GRAY_FILL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = make_border()
    for cc in range(3, 7):
        ws3.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=GRAY_FILL)
        ws3.cell(row=r, column=cc).border = make_border()
    ws3.row_dimensions[r].height = 28
    r += 1
    # 空行
    r += 1

# 数据验证 - 5 观察点下拉
dv_obs = DataValidation(type="list",
                        formula1=f'"{",".join(obs_points)}"',
                        allow_blank=True,
                        showDropDown=False)  # 注意：False 表示显示下拉箭头
ws3.add_data_validation(dv_obs)
dv_obs.add(f"C6:C{r-1}")

# 数据验证 - 月份下拉
dv_month = DataValidation(type="list",
                          formula1=f'"{",".join(months)}"',
                          allow_blank=True,
                          showDropDown=False)
ws3.add_data_validation(dv_month)
dv_month.add(f"B6:B{r-1}")

ws3.freeze_panes = "A6"
ws3.sheet_view.showGridLines = False

print("Sheet 3 完成")

# =====================================================================
# Sheet 4: 03_30-60-90天对话追踪
# =====================================================================
ws4 = wb.create_sheet("03_30-60-90天对话追踪")
set_col_widths(ws4, [3, 14, 18, 14, 18, 30, 30, 16, 18, 3])

set_title(ws4, 1, 2, "30-60-90 天对话追踪", span=8)
set_subtitle(ws4, 2, 2, "三场对话的状态、关键产出、下一步 | 配合 02 辅导对话脚本使用", span=8)

# 表头
set_header_row(ws4, 4, 2, ["对话", "对话日期", "距今 X 天", "状态", "学员状态判断（1-10 分）", "关键产出", "下一步 / 你承诺的支持", "负责人"])

# 三场对话
dialogues = [
    ("30 天对话", "倾听为主，资源承诺", 7, 30),
    ("60 天对话", "清障为主，聚焦一件事", 7, 60),
    ("90 天对话", "规划为主，发展建议", 7, 90),
]
status_options = ["未开始", "已邀约", "已完成", "已延后", "已取消"]

r = 5
for name, focus, score_default, day in dialogues:
    set_cell(ws4, r, 2, name, bold=True, fill=SUBHEADER_FILL, color=NAVY_FONT, align="center")
    set_cell(ws4, r, 3, "请填写（yyyy/mm/dd）", align="center", color="808080", italic=True)
    # 距今天数公式 =TODAY()-C5
    set_cell(ws4, r, 4, f'=IF(ISNUMBER(C{r}),TODAY()-C{r},"未填")', align="center", bold=True)
    set_cell(ws4, r, 5, "未开始", align="center", color="808080")
    set_cell(ws4, r, 6, score_default, align="center")
    set_cell(ws4, r, 7, f"对话焦点：{focus}\n\n（请填写：学员核心变化、卡点、收获）",
             align="left", color="808080", italic=True)
    set_cell(ws4, r, 8, "（请填写：我能做的 X、Y、Z）", align="left", color="808080", italic=True)
    set_cell(ws4, r, 9, "学员上级", align="center")
    ws4.row_dimensions[r].height = 80
    r += 1

# 评分列条件格式（红黄绿）
score_range = f"F5:F{r-1}"
ws4.conditional_formatting.add(
    score_range,
    CellIsRule(operator="lessThan", formula=["4"], fill=PatternFill("solid", fgColor=RED_FILL), font=Font(color=RED_FONT, bold=True))
)
ws4.conditional_formatting.add(
    score_range,
    CellIsRule(operator="between", formula=["4", "6"], fill=PatternFill("solid", fgColor=ORANGE_FILL), font=Font(color=ORANGE_FONT, bold=True))
)
ws4.conditional_formatting.add(
    score_range,
    CellIsRule(operator="between", formula=["7", "8"], fill=PatternFill("solid", fgColor=YELLOW_FILL), font=Font(color=YELLOW_FONT, bold=True))
)
ws4.conditional_formatting.add(
    score_range,
    CellIsRule(operator="greaterThanOrEqual", formula=["9"], fill=PatternFill("solid", fgColor=GREEN_FILL), font=Font(color=GREEN_FONT, bold=True))
)

# 状态列下拉
dv_status = DataValidation(type="list",
                           formula1=f'"{",".join(status_options)}"',
                           allow_blank=True, showDropDown=False)
ws4.add_data_validation(dv_status)
dv_status.add(f"E5:E{r-1}")

# 距今天数条件格式（越久越红）
days_range = f"D5:D{r-1}"
ws4.conditional_formatting.add(
    days_range,
    FormulaRule(formula=[f'AND(ISNUMBER(D5),D5<0)'], fill=PatternFill("solid", fgColor=GRAY_FILL), font=Font(color="808080", italic=True))
)
ws4.conditional_formatting.add(
    days_range,
    FormulaRule(formula=[f'AND(ISNUMBER(D5),D5>=0,D5<=7)'], fill=PatternFill("solid", fgColor=GREEN_FILL), font=Font(color=GREEN_FONT, bold=True))
)
ws4.conditional_formatting.add(
    days_range,
    FormulaRule(formula=[f'AND(ISNUMBER(D5),D5>7,D5<=30)'], fill=PatternFill("solid", fgColor=YELLOW_FILL), font=Font(color=YELLOW_FONT, bold=True))
)
ws4.conditional_formatting.add(
    days_range,
    FormulaRule(formula=[f'AND(ISNUMBER(D5),D5>30)'], fill=PatternFill("solid", fgColor=RED_FILL), font=Font(color=RED_FONT, bold=True))
)

# 备注区
r += 1
set_subtitle(ws4, r, 2, "三场对话前的 3 分钟准备清单（02 脚本附录）", span=8)
r += 1
prep_items = [
    "我把上次的对话记录调出来了吗？",
    "我把 01 观察指南的月度记录表调出来了吗？",
    "我心里清楚这次对话的核心目的吗？（30=倾听；60=清障；90=规划）",
    "我没有准备'今天要教他什么'——我是来听和陪的",
    "我心里清楚'我这边能做的是 X、Y、Z'是什么吗？",
    "我没有安排在绩效面谈前后（避免学员把两件事混淆）",
]
for item in prep_items:
    set_cell(ws4, r, 2, "□", bold=True, align="center", color=NAVY_FONT)
    ws4.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    c = ws4.cell(row=r, column=3, value=item)
    c.font = Font(name="微软雅黑", size=10, color="000000")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = make_border()
    c.fill = PatternFill("solid", fgColor=WHITE)
    for cc in range(3, 10):
        ws4.cell(row=r, column=cc).border = make_border()
    ws4.row_dimensions[r].height = 22
    r += 1

ws4.freeze_panes = "A5"
ws4.sheet_view.showGridLines = False
print("Sheet 4 完成")

# =====================================================================
# Sheet 5: 04_行为指标变化_雷达图
# =====================================================================
ws5 = wb.create_sheet("04_行为指标变化_雷达图")
set_col_widths(ws5, [3, 24, 14, 14, 14, 14, 14, 14, 3])

set_title(ws5, 1, 2, "行为指标变化·五项管理动作雷达图", span=6)
set_subtitle(ws5, 2, 2, "评分维度：1-10 分 | 训前测评 + 30/60/90 天各评分一次 | HRBP 可直接录入", span=6)

# 表头
set_header_row(ws5, 4, 2, ["观察点（5 项）", "训前测评", "30 天得分", "60 天得分", "90 天得分", "变化趋势", "重点观察"])

# 5 个观察点
obs_for_radar = [
    "1. 响应节奏变化",
    "2. 五感管理动作（'你的判断'）",
    "3. AI 话题的开口",
    "4. 任务分配 5W2H+H",
    "5. 激励机制更新（人类贡献目标）",
]

r = 5
data_radar_start = r
for op in obs_for_radar:
    set_cell(ws5, r, 2, op, bold=True, fill=SUBHEADER_FILL, color=NAVY_FONT, align="left")
    # 5 个分数列
    for col in range(3, 7):
        set_cell(ws5, r, col, 5, align="center", color="808080")
    # 变化趋势公式
    set_cell(ws5, r, 7, f'=IF(ISNUMBER(C{r}),IF(F{r}-C{r}>=2,"↑ 明显提升",IF(F{r}-C{r}>=1,"↑ 略有提升",IF(F{r}-C{r}=0,"— 持平",IF(F{r}-C{r}>=-1,"↓ 略有下滑","↓↓ 明显下滑"))),"未填")',
             align="center", bold=True)
    set_cell(ws5, r, 8, "", align="left", color="808080", italic=True)
    ws5.row_dimensions[r].height = 32
    r += 1
data_radar_end = r - 1

# 评分列条件格式（颜色阶梯）
for col_letter in ["C", "D", "E", "F"]:
    rng = f"{col_letter}{data_radar_start}:{col_letter}{data_radar_end}"
    ws5.conditional_formatting.add(rng, ColorScaleRule(
        start_type='num', start_value=1, start_color='F4B7B7',
        mid_type='num', mid_value=5, mid_color='FFEB9C',
        end_type='num', end_value=10, end_color='C6EFCE'
    ))

# 趋势列条件格式
trend_range = f"G{data_radar_start}:G{data_radar_end}"
ws5.conditional_formatting.add(trend_range,
    FormulaRule(formula=[f'G{data_radar_start}="↑ 明显提升"'],
                fill=PatternFill("solid", fgColor=GREEN_FILL), font=Font(color=GREEN_FONT, bold=True)))
ws5.conditional_formatting.add(trend_range,
    FormulaRule(formula=[f'G{data_radar_start}="↑ 略有提升"'],
                fill=PatternFill("solid", fgColor="E2EFDA"), font=Font(color=GREEN_FONT)))
ws5.conditional_formatting.add(trend_range,
    FormulaRule(formula=[f'G{data_radar_start}="— 持平"'],
                fill=PatternFill("solid", fgColor=GRAY_FILL), font=Font(color="595959")))
ws5.conditional_formatting.add(trend_range,
    FormulaRule(formula=[f'G{data_radar_start}="↓ 略有下滑"'],
                fill=PatternFill("solid", fgColor=YELLOW_FILL), font=Font(color=YELLOW_FONT)))
ws5.conditional_formatting.add(trend_range,
    FormulaRule(formula=[f'G{data_radar_start}="↓↓ 明显下滑"'],
                fill=PatternFill("solid", fgColor=RED_FILL), font=Font(color=RED_FONT, bold=True)))

# 数据验证：1-10 分
for col_letter in ["C", "D", "E", "F"]:
    dv_score = DataValidation(type="decimal",
                              operator="between",
                              formula1=1, formula2=10,
                              allow_blank=True,
                              showDropDown=False,
                              error="请输入 1-10 之间的分数",
                              errorTitle="分数范围错误")
    ws5.add_data_validation(dv_score)
    dv_score.add(f"{col_letter}{data_radar_start}:{col_letter}{data_radar_end}")

# 雷达图数据（用第 3 行开始的 5 个观察点）
# 因为雷达图需要一个数据矩阵：5 个观察点（类别）x 4 个时间点（系列）
# 我们把 5 行变成 5 列：在右侧建一个转置区
r += 2
set_subtitle(ws5, r, 2, "雷达图数据区（自动转置，请勿修改）", span=6)
r += 1
radar_header_row = r
set_header_row(ws5, r, 2, ["时间点", "响应节奏", "五感动作", "AI 话题", "5W2H+H", "激励机制"])
r += 1
radar_data_start = r
time_points = ["训前测评", "30 天", "60 天", "90 天"]
for i, tp in enumerate(time_points):
    src_col = chr(ord("C") + i)  # C/D/E/F
    set_cell(ws5, r, 2, tp, bold=True, fill=SUBHEADER_FILL, color=NAVY_FONT, align="center")
    for j in range(5):
        src_row = data_radar_start + j
        formula = f'=IF(ISNUMBER({src_col}{src_row}),{src_col}{src_row},0)'
        c = set_cell(ws5, r, 3+j, formula, align="center")
    r += 1
radar_data_end = r - 1

# 颜色阶梯 - 雷达数据区
for col in range(3, 8):
    col_letter = get_column_letter(col)
    rng = f"{col_letter}{radar_data_start}:{col_letter}{radar_data_end}"
    ws5.conditional_formatting.add(rng, ColorScaleRule(
        start_type='num', start_value=0, start_color='F4B7B7',
        mid_type='num', mid_value=5, mid_color='FFEB9C',
        end_type='num', end_value=10, end_color='C6EFCE'
    ))

# 创建雷达图
radar = RadarChart()
radar.type = "filled"
radar.style = 26
radar.title = "五项管理动作·行为指标雷达图"
radar.y_axis.scaling.min = 0
radar.y_axis.scaling.max = 10

# 数据系列（4 个时间点）: B 列是时间点，C-G 是 5 个观察点
data_ref = Reference(ws5, min_col=2, min_row=radar_data_start,
                     max_col=7, max_row=radar_data_end)
# 用 from_rows 模式：第一列是系列名（训前测评/30天/60天/90天），其余是 5 个类别
radar.add_data(data_ref, titles_from_data=True, from_rows=True)

# 类别：5 个观察点名称
cat_ref = Reference(ws5, min_col=3, min_row=radar_header_row,
                    max_col=7, max_row=radar_header_row)
radar.set_categories(cat_ref)

radar.height = 12
radar.width = 18
ws5.add_chart(radar, f"B{r+2}")

# 备注
r2 = r + 28
set_subtitle(ws5, r2, 2, "评分说明", span=6)
r2 += 1
note_lines = [
    ("1-3 分", "未做到", "学员基本没有该动作的行为", RED_FILL, RED_FONT),
    ("4-6 分", "偶尔做到", "学员有尝试但不持续", YELLOW_FILL, YELLOW_FONT),
    ("7-8 分", "基本做到", "学员能持续做到但深度不够", "D9EAD3", "38761D"),
    ("9-10 分", "完全做到", "学员已内化并能影响他人", GREEN_FILL, GREEN_FONT),
]
for label, status, desc, fill, fcolor in note_lines:
    set_cell(ws5, r2, 2, label, bold=True, fill=fill, color=fcolor, align="center")
    set_cell(ws5, r2, 3, status, bold=True, fill=fill, color=fcolor, align="center")
    ws5.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=7)
    c = ws5.cell(row=r2, column=4, value=desc)
    c.font = Font(name="微软雅黑", size=10, color=fcolor)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor=fill)
    c.border = make_border()
    for cc in range(4, 8):
        ws5.cell(row=r2, column=cc).fill = PatternFill("solid", fgColor=fill)
        ws5.cell(row=r2, column=cc).border = make_border()
    ws5.row_dimensions[r2].height = 22
    r2 += 1

ws5.freeze_panes = "A5"
ws5.sheet_view.showGridLines = False
print("Sheet 5 完成")

# =====================================================================
# Sheet 6: 05_信号灯与风险预警
# =====================================================================
ws6 = wb.create_sheet("05_信号灯与风险预警")
set_col_widths(ws6, [3, 22, 14, 14, 14, 14, 28, 3])

set_title(ws6, 1, 2, "信号灯与风险预警", span=6)
set_subtitle(ws6, 2, 2, "6 类信号灯（绿/黄/橙/红）| 自动公式判断 | 每月回顾", span=6)

# 表头
set_header_row(ws6, 4, 2, ["信号灯维度", "当前值", "阈值", "得分", "信号灯颜色", "判断说明 / 你的行动"])

# 6 类信号灯定义
# 1. 5 感得分均值（来自 04 雷达图）
# 2. 行为下滑数（90 vs 训前 下滑的观察点数）
# 3. 对话延迟（30/60/90 距今天数）
# 4. 清单执行率（30 天清单完成 %）
# 5. 问责伙伴活跃度（过去 30 天联系次数）
# 6. Z 世代员工反馈（员工 1v1 中提到"最近不一样"的频次）

# 简化：6 类信号灯 - 4 类用公式、2 类人工填
# 由于跨 sheet 公式复杂，我们把可计算的几个设计为简单公式 + 人工辅佐

signal_defs = [
    {
        "name": "1. 5 感得分均值",
        "type": "formula",
        "current_formula": '=AVERAGE(\'04_行为指标变化_雷达图\'!F5:F9)',
        "threshold": 7,
        "score_formula": '=IF(C5>=7,3,IF(C5>=5,2,IF(C5>=3,1,0)))',
        "color_formula": '=IF(D5=3,"绿",IF(D5=2,"黄",IF(D5=1,"橙","红")))',
        "action": '=IF(D5=3,"继续观察，下月再填",IF(D5=2,"下次 1v1 自然问 1-2 个问题",IF(D5=1,"下次正式对话增加 1-2 个问题","安排 30 分钟专项对话（不批评）")))'
    },
    {
        "name": "2. 行为下滑观察点数（90 vs 训前）",
        "type": "formula",
        "current_formula": '=COUNTIF(\'04_行为指标变化_雷达图\'!G5:G9,"↓*")',
        "threshold": 0,
        "score_formula": '=IF(C6=0,3,IF(C6=1,2,IF(C6=2,1,0)))',
        "color_formula": '=IF(D6=3,"绿",IF(D6=2,"黄",IF(D6=1,"橙","红")))',
        "action": '=IF(D6=3,"没有下滑项，健康",IF(D6=2,"1 项下滑，60 天对话时讨论",IF(D6=1,"2 项下滑，30 天内必须清障","3 项以上下滑，启动 HRBP 介入")))'
    },
    {
        "name": "3. 30 天对话距今天数",
        "type": "formula",
        "current_formula": '=IFERROR(IF(ISNUMBER(\'03_30-60-90天对话追踪\'!C5),TODAY()-\'03_30-60-90天对话追踪\'!C5,999),999)',
        "threshold": 30,
        "score_formula": '=IF(C7<=30,3,IF(C7<=45,2,IF(C7<=60,1,0)))',
        "color_formula": '=IF(D7=3,"绿",IF(D7=2,"黄",IF(D7=1,"橙","红")))',
        "action": '=IF(D7=3,"对话及时推进",IF(D7=2,"对话略延迟，但可接受",IF(D7=1,"对话明显延迟，立即安排","对话严重未做，必须启动")))'
    },
    {
        "name": "4. 60 天对话距今天数",
        "type": "formula",
        "current_formula": '=IFERROR(IF(ISNUMBER(\'03_30-60-90天对话追踪\'!C6),TODAY()-\'03_30-60-90天对话追踪\'!C6,999),999)',
        "threshold": 60,
        "score_formula": '=IF(C8<=60,3,IF(C8<=75,2,IF(C8<=90,1,0)))',
        "color_formula": '=IF(D8=3,"绿",IF(D8=2,"黄",IF(D8=1,"橙","红")))',
        "action": '=IF(D8=3,"中期清障对话已做",IF(D8=2,"对话略延迟",IF(D8=1,"对话明显延迟，立即安排","60 天对话未做，行动力预警")))'
    },
    {
        "name": "5. 90 天对话距今天数",
        "type": "formula",
        "current_formula": '=IFERROR(IF(ISNUMBER(\'03_30-60-90天对话追踪\'!C7),TODAY()-\'03_30-60-90天对话追踪\'!C7,999),999)',
        "threshold": 90,
        "score_formula": '=IF(C9<=90,3,IF(C9<=105,2,IF(C9<=120,1,0)))',
        "color_formula": '=IF(D9=3,"绿",IF(D9=2,"黄",IF(D9=1,"橙","红")))',
        "action": '=IF(D9=3,"成果对话已做",IF(D9=2,"对话略延迟",IF(D9=1,"对话明显延迟",IF(C9>120,"90 天对话严重未做","90 天对话超期 30 天"))))'
    },
    {
        "name": "6. 学员自评（人工）",
        "type": "manual",
        "current_value": 5,
        "threshold": 6,
        "score_formula": '=IF(C10>=7,3,IF(C10>=5,2,IF(C10>=3,1,0)))',
        "color_formula": '=IF(D10=3,"绿",IF(D10=2,"黄",IF(D10=1,"橙","红")))',
        "action": '=IF(D10=3,"学员自我感受良好",IF(D10=2,"学员有些许不确定",IF(D10=1,"学员自我感受明显下滑","学员自评极低，立即 1v1 倾听")))'
    },
]

r = 5
for sd in signal_defs:
    set_cell(ws6, r, 2, sd["name"], bold=True, fill=SUBHEADER_FILL, color=NAVY_FONT, align="left")
    # 当前值
    if sd["type"] == "formula":
        set_cell(ws6, r, 3, sd["current_formula"], align="center", bold=True)
    else:
        set_cell(ws6, r, 3, sd["current_value"], align="center")
    # 阈值
    set_cell(ws6, r, 4, sd["threshold"], align="center", color="808080", italic=True)
    # 得分
    set_cell(ws6, r, 5, sd["score_formula"], align="center", bold=True)
    # 颜色
    set_cell(ws6, r, 6, sd["color_formula"], align="center", bold=True)
    # 行动
    set_cell(ws6, r, 7, sd["action"], align="left")
    ws6.row_dimensions[r].height = 36
    r += 1

# 颜色列条件格式
color_range = f"F5:F{r-1}"
ws6.conditional_formatting.add(color_range,
    FormulaRule(formula=[f'F5="绿"'],
                fill=PatternFill("solid", fgColor=GREEN_FILL), font=Font(color=GREEN_FONT, bold=True, size=12)))
ws6.conditional_formatting.add(color_range,
    FormulaRule(formula=[f'F5="黄"'],
                fill=PatternFill("solid", fgColor=YELLOW_FILL), font=Font(color=YELLOW_FONT, bold=True, size=12)))
ws6.conditional_formatting.add(color_range,
    FormulaRule(formula=[f'F5="橙"'],
                fill=PatternFill("solid", fgColor=ORANGE_FILL), font=Font(color=ORANGE_FONT, bold=True, size=12)))
ws6.conditional_formatting.add(color_range,
    FormulaRule(formula=[f'F5="红"'],
                fill=PatternFill("solid", fgColor=RED_FILL), font=Font(color=RED_FONT, bold=True, size=12)))

# 整体预警
r += 1
set_subtitle(ws6, r, 2, "整体预警等级", span=6)
r += 1
set_header_row(ws6, r, 2, ["指标", "值", "含义", "你接下来的 7 天动作", "", ""])
ws6.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
r += 1
set_cell(ws6, r, 2, "红色信号灯数量", bold=True, fill=RED_FILL, color=RED_FONT, align="center")
set_cell(ws6, r, 3, f'=COUNTIF(F5:F{r-3},"红")', align="center", bold=True)
set_cell(ws6, r, 4, "2 个以上 = 立即介入", align="left")
ws6.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
set_cell(ws6, r, 5, "安排 30 分钟专项对话（不批评、不考核）", align="left", fill=RED_FILL, color=RED_FONT, bold=True)
ws6.row_dimensions[r].height = 36
r += 1

set_cell(ws6, r, 2, "橙色信号灯数量", bold=True, fill=ORANGE_FILL, color=ORANGE_FONT, align="center")
set_cell(ws6, r, 3, f'=COUNTIF(F5:F{r-4},"橙")', align="center", bold=True)
set_cell(ws6, r, 4, "2 个以上 = 重点观察", align="left")
ws6.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
set_cell(ws6, r, 5, "下次正式对话增加 1-2 个问题", align="left", fill=ORANGE_FILL, color=ORANGE_FONT, bold=True)
ws6.row_dimensions[r].height = 36
r += 1

set_cell(ws6, r, 2, "黄色信号灯数量", bold=True, fill=YELLOW_FILL, color=YELLOW_FONT, align="center")
set_cell(ws6, r, 3, f'=COUNTIF(F5:F{r-5},"黄")', align="center", bold=True)
set_cell(ws6, r, 4, "3 个以上 = 整体推进偏慢", align="left")
ws6.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
set_cell(ws6, r, 5, "下次 1v1 自然问 1-2 个相关问题", align="left", fill=YELLOW_FILL, color=YELLOW_FONT, bold=True)
ws6.row_dimensions[r].height = 36
r += 1

# 学员自评数据验证
dv_self = DataValidation(type="decimal", operator="between",
                         formula1=1, formula2=10, allow_blank=True,
                         showDropDown=False,
                         error="请输入 1-10 分", errorTitle="分数错误")
ws6.add_data_validation(dv_self)
dv_self.add(f"C10:C10")

ws6.freeze_panes = "A5"
ws6.sheet_view.showGridLines = False
print("Sheet 6 完成")

# =====================================================================
# Sheet 7: 06_团队层面业务结果
# =====================================================================
ws7 = wb.create_sheet("06_团队层面业务结果")
set_col_widths(ws7, [3, 24, 12, 12, 12, 12, 12, 12, 12, 12, 3])

set_title(ws7, 1, 2, "团队层面业务结果·月度趋势", span=8)
set_subtitle(ws7, 2, 2, "学员团队的 3-5 个业务指标月度趋势 | 配合 02 脚本的 90 天对话使用", span=8)

# 表头
headers7 = ["业务指标", "基线值（训前）", "30 天", "60 天", "90 天", "Q1 趋势", "Q2 趋势", "数据来源"]
set_header_row(ws7, 4, 2, headers7)
# 合并一下: B-I 共 8 列，C-H 是 6 个月份数据
# 重新设计：基线 + 30 + 60 + 90 + Q1 + Q2 (实际是基线+4 个月)
# 简化：基线 + 30天 + 60天 + 90天 + 季度1 + 季度2
# 让我们让指标有 5 个常见类别

# 5 个业务指标（学员自定）
biz_metrics = [
    "1. 团队稳定性（季度主动离职率 %）",
    "2. 产出质量（交付一次性通过率 %）",
    "3. Z 世代员工 1v1 满意度（1-10 分）",
    "4. 团队 AI 工具使用率（周活跃 %）",
    "5. 客户/业务方 NPS 或满意度（1-10）",
]

r = 5
for bm in biz_metrics:
    set_cell(ws7, r, 2, bm, bold=True, fill=SUBHEADER_FILL, color=NAVY_FONT, align="left")
    for col in range(3, 9):
        set_cell(ws7, r, col, "", align="center", color="808080")
    set_cell(ws7, r, 9, "（请填写数据来源）", align="left", color="808080", italic=True)
    ws7.row_dimensions[r].height = 32
    r += 1

# 趋势公式列示例 (G列 = Q1趋势, H列 = Q2趋势)
# Q1 = (60天 - 基线), Q2 = (90天 - 60天)
for i in range(5):
    row_num = 5 + i
    # G列: Q1趋势
    ws7.cell(row=row_num, column=8, value=f'=IF(AND(ISNUMBER(C{row_num}),ISNUMBER(E{row_num})),IF(E{row_num}-C{row_num}>0,"↑",IF(E{row_num}-C{row_num}<0,"↓","—")),"未填")')
    ws7.cell(row=row_num, column=8).font = Font(name="微软雅黑", size=10, bold=True)
    ws7.cell(row=row_num, column=8).alignment = Alignment(horizontal="center", vertical="center")
    ws7.cell(row=row_num, column=8).border = make_border()
    # H列: Q2趋势
    ws7.cell(row=row_num, column=9, value=f'=IF(AND(ISNUMBER(E{row_num}),ISNUMBER(F{row_num})),IF(F{row_num}-E{row_num}>0,"↑",IF(F{row_num}-E{row_num}<0,"↓","—")),"未填")')
    # 哦不对，列重排了
# 重新核对：headers7 = ["业务指标", "基线值（训前）", "30 天", "60 天", "90 天", "Q1 趋势", "Q2 趋势", "数据来源"]
# 索引: 2(B)=指标名 3(C)=基线 4(D)=30天 5(E)=60天 6(F)=90天 7(G)=Q1趋势 8(H)=Q2趋势 9(I)=数据来源

# 删除并重写
for i in range(5):
    row_num = 5 + i
    for col in [3, 4, 5, 6, 7, 8, 9]:
        cell = ws7.cell(row=row_num, column=col)
        cell.value = None
        cell.fill = PatternFill("solid", fgColor=WHITE)
        cell.font = Font(name="微软雅黑", size=10, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = make_border()
    # 重写
    set_cell(ws7, row_num, 3, "", align="center", color="808080")
    set_cell(ws7, row_num, 4, "", align="center", color="808080")
    set_cell(ws7, row_num, 5, "", align="center", color="808080")
    set_cell(ws7, row_num, 6, "", align="center", color="808080")
    # Q1趋势 = 60天 - 基线
    set_cell(ws7, row_num, 7,
             f'=IF(AND(ISNUMBER(C{row_num}),ISNUMBER(E{row_num})),IF(E{row_num}-C{row_num}>0,"↑",IF(E{row_num}-C{row_num}<0,"↓","—")),"未填")',
             align="center", bold=True)
    # Q2趋势 = 90天 - 60天
    set_cell(ws7, row_num, 8,
             f'=IF(AND(ISNUMBER(E{row_num}),ISNUMBER(F{row_num})),IF(F{row_num}-E{row_num}>0,"↑",IF(F{row_num}-E{row_num}<0,"↓","—")),"未填")',
             align="center", bold=True)
    set_cell(ws7, row_num, 9, "", align="left", color="808080", italic=True)

# 趋势列条件格式
for col_letter in ["G", "H"]:
    rng = f"{col_letter}5:{col_letter}9"
    ws7.conditional_formatting.add(rng,
        FormulaRule(formula=[f'{col_letter}5="↑"'],
                    fill=PatternFill("solid", fgColor=GREEN_FILL), font=Font(color=GREEN_FONT, bold=True)))
    ws7.conditional_formatting.add(rng,
        FormulaRule(formula=[f'{col_letter}5="↓"'],
                    fill=PatternFill("solid", fgColor=RED_FILL), font=Font(color=RED_FONT, bold=True)))
    ws7.conditional_formatting.add(rng,
        FormulaRule(formula=[f'{col_letter}5="—"'],
                    fill=PatternFill("solid", fgColor=GRAY_FILL), font=Font(color="595959")))

# 备注
r = 12
set_subtitle(ws7, r, 2, "业务结果翻译指南（把'业务指标'变成'管理改变'）", span=8)
r += 1
trans = [
    "团队稳定性提升 → 学员在 1v1 反馈、激励机制上的改变让 Z 世代员工愿意留下",
    "产出质量提升 → 学员在 5W2H+H 上的改变让任务目标更清晰",
    "Z 世代员工 1v1 满意度提升 → 学员在五感管理、AI 话题开口上的改变",
    "AI 工具使用率提升 → 学员主动邀请员工分享 AI 心得",
    "客户/业务方满意度提升 → 学员在'你的判断'反馈上让员工价值感提升",
]
for t in trans:
    set_cell(ws7, r, 2, "•", bold=True, color=NAVY_FONT, align="center")
    ws7.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    c = ws7.cell(row=r, column=3, value=t)
    c.font = Font(name="微软雅黑", size=10, color="000000")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = make_border()
    c.fill = PatternFill("solid", fgColor=WHITE)
    for cc in range(3, 10):
        ws7.cell(row=r, column=cc).border = make_border()
    ws7.row_dimensions[r].height = 26
    r += 1

# 业绩线图 - 引用业务指标月度数据
# 由于 5 个指标是不同维度，单一图表意义不大，我们做一个组合趋势
# 这里改为一个汇总指标（手动输入整体印象）
r += 1
set_subtitle(ws7, r, 2, "团队整体健康度（HRBP 每月打分 1-10）", span=8)
r += 1
set_header_row(ws7, r, 2, ["维度", "基线", "30 天", "60 天", "90 天", "变化", "判断", ""])
ws7.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
r += 1
set_cell(ws7, r, 2, "团队整体健康度", bold=True, fill=SUBHEADER_FILL, color=NAVY_FONT, align="left")
set_cell(ws7, r, 3, 5, align="center", color="808080")
set_cell(ws7, r, 4, 5, align="center", color="808080")
set_cell(ws7, r, 5, 5, align="center", color="808080")
set_cell(ws7, r, 6, 5, align="center", color="808080")
set_cell(ws7, r, 7, '=IF(AND(ISNUMBER(C22),ISNUMBER(F22)),F22-C22,"未填")', align="center", bold=True)
ws7.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
set_cell(ws7, r, 8, '=IF(ISNUMBER(G22),IF(G22>0,"向好",IF(G22<0,"下滑","持平")),"未填")', align="center", bold=True)
ws7.row_dimensions[r].height = 28

# 颜色阶梯
health_range = f"C{r}:F{r}"
ws7.conditional_formatting.add(health_range, ColorScaleRule(
    start_type='num', start_value=1, start_color='F4B7B7',
    mid_type='num', mid_value=5, mid_color='FFEB9C',
    end_type='num', end_value=10, end_color='C6EFCE'
))

# 折线图
chart_line = LineChart()
chart_line.title = "团队整体健康度·趋势"
chart_line.y_axis.title = "得分 (1-10)"
chart_line.x_axis.title = "时间点"
chart_line.height = 8
chart_line.width = 16

data_ref = Reference(ws7, min_col=3, min_row=r-1, max_col=6, max_row=r)
chart_line.add_data(data_ref, titles_from_data=False)
cat_ref = Reference(ws7, min_col=3, min_row=r-2, max_col=6, max_row=r-2)
chart_line.set_categories(cat_ref)
ws7.add_chart(chart_line, f"B{r+3}")

# 1-10 分数据验证
for col_letter in ["C", "D", "E", "F"]:
    dv = DataValidation(type="decimal", operator="between",
                        formula1=1, formula2=10, allow_blank=True,
                        showDropDown=False,
                        error="请输入 1-10 分", errorTitle="分数错误")
    ws7.add_data_validation(dv)
    dv.add(f"{col_letter}{r}:{col_letter}{r}")

ws7.freeze_panes = "A5"
ws7.sheet_view.showGridLines = False
print("Sheet 7 完成")

# =====================================================================
# 设置 Sheet 顺序
# =====================================================================
order = [
    "00_使用说明",
    "01_学员基础信息",
    "02_五项管理动作_月度记录",
    "03_30-60-90天对话追踪",
    "04_行为指标变化_雷达图",
    "05_信号灯与风险预警",
    "06_团队层面业务结果"
]
wb._sheets = [wb[name] for name in order]

# 设置默认活动 Sheet
wb.active = 0

# 保存
output_path = r"D:\2026年课程\竞越\领航：Z世代管理新策略3.0\完整课程包\12_管理者工具包_训后\03_可视化看板_Excel版.xlsx"
wb.save(output_path)
print(f"\n文件已保存: {output_path}")

import os
size = os.path.getsize(output_path)
print(f"文件大小: {size:,} bytes ({size/1024:.1f} KB)")

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

wb = Workbook()
ws1 = wb.active
ws1.title = "增长数据追踪"

headers1 = ["月份", "新增潜客(组)", "到店量(组)", "成交单数", "营业额(万)", "到店转化率", "客单价(万)", "老客复购/推荐"]
data1 = [
    ["1月", 45, 28, 8, 210, 0.286, 4.3, 3],
    ["2月", 68, 42, 12, 315, 0.286, 4.4, 5],
    ["3月", 95, 58, 18, 472, 0.310, 4.5, 8],
    ["4月", 128, 82, 26, 689, 0.317, 4.7, 12],
    ["5月", 156, 102, 33, 865, 0.324, 4.8, 15],
    ["6月", 182, 118, 39, 1012, 0.331, 4.9, 18],
    ["7月", 165, 108, 35, 905, 0.324, 4.8, 16],
    ["8月", 178, 115, 38, 978, 0.330, 4.9, 18],
    ["9月", 195, 128, 43, 1105, 0.336, 5.0, 21],
    ["10月", 235, 155, 52, 1342, 0.335, 5.1, 25],
    ["11月", 258, 172, 58, 1495, 0.337, 5.1, 28],
    ["12月", 285, 192, 65, 1670, 0.339, 5.2, 32],
]
summary1 = ["合计/平均", "=SUM(B2:B13)", "=SUM(C2:C13)", "=SUM(D2:D13)", "=SUM(E2:E13)", "=AVERAGE(F2:F13)", "=AVERAGE(G2:G13)", "=SUM(H2:H13)"]

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
even_row_fill = PatternFill(start_color="D6E3F8", end_color="D6E3F8", fill_type="solid")
odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
summary_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
summary_font = Font(bold=True)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for row_idx, row_data in enumerate(data1, 2):
    fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx == 6:
            cell.number_format = "0.0%"
        elif col_idx in [2, 3, 4, 5, 7, 8]:
            cell.number_format = "#,##0"

for col_idx, value in enumerate(summary1, 1):
    cell = ws1.cell(row=15, column=col_idx, value=value)
    cell.fill = summary_fill
    cell.font = summary_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if col_idx == 6:
        cell.number_format = "0.0%"
    elif col_idx in [2, 3, 4, 5, 7, 8]:
        cell.number_format = "#,##0"

col_widths1 = [10, 14, 12, 12, 12, 12, 12, 16]
for col_idx, width in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

ws1.auto_filter.ref = "A1:H15"

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "月度增长趋势"
chart1.y_axis.title = "金额(万)/数量"
chart1.x_axis.title = "月份"
data_ref = Reference(ws1, min_col=5, min_row=1, max_row=13)
cats_ref = Reference(ws1, min_col=1, min_row=2, max_row=13)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.shape = 4
chart1.width = 18
chart1.height = 10
ws1.add_chart(chart1, "J2")

ws2 = wb.create_sheet("目标完成情况")
headers2 = ["指标", "年度目标", "实际达成", "完成率", "评价"]
data2 = [
    ["年营业额(亿)", 3.2, 3.27, "=C2/B2", "达标"],
    ["到店转化率", 0.32, 0.323, "=C3/B3", "达标"],
    ["场均成交金额(万)", 5.0, 4.85, "=C4/B4", "接近达标"],
    ["客户复购/推荐率", 0.18, 0.155, "=C5/B5", "未完全达标"],
    ["净利润率", 0.07, 0.062, "=C6/B6", "未完全达标"],
    ["新增潜客总数(组)", 1500, 1990, "=C7/B7", "超额完成"],
    ["社区覆盖小区数", 10, 12, "=C8/B8", "超额完成"],
]

for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for row_idx, row_data in enumerate(data2, 2):
    fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx == 4 and isinstance(value, str) and value.startswith("="):
            cell.number_format = "0.0%"

ws2.conditional_formatting.add("D2:D8", FormulaRule(formula=["D2>1"], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), font=Font(color="006100")))
ws2.conditional_formatting.add("D2:D8", FormulaRule(formula=["AND(D2>=0.8,D2<=1)"], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), font=Font(color="9C6500")))
ws2.conditional_formatting.add("D2:D8", FormulaRule(formula=["D2<0.8"], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), font=Font(color="9C0006")))

col_widths2 = [20, 14, 14, 12, 14]
for col_idx, width in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.auto_filter.ref = "A1:E8"

ws3 = wb.create_sheet("经验总结")
headers3 = ["维度", "成功经验", "可改进点", "改进措施"]
data3 = [
    ["社区渗透", "新小区渗透SOP非常有效，滨江新城模式可复制", "部分小区物业合作难度大", "建立物业合作标准话术，多小区布局降低依赖"],
    ["套餐销售", "套餐客户客单价明显高于单品客户(+35%)", "套餐设计初期不受认可", "先做客户调研再设计，根据反馈快速迭代"],
    ["客户运营", "老客户激活成本低，成单率高(55%)", "数据采集初期困难", "价值交换换数据，强制留取+激励引导"],
    ["团队能力", "顾问型销售培训效果显著", "团队学习意愿参差不齐", "标杆引路+激励挂钩+持续强化"],
    ["资源配置", "800万预算控制良好，未超支", "部分模块投入保守", "根据效果动态调整资源分配"],
]

for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

for row_idx, row_data in enumerate(data3, 2):
    fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

col_widths3 = [14, 35, 30, 35]
for col_idx, width in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(col_idx)].width = width
for row in range(2, 7):
    ws3.row_dimensions[row].height = 40
ws3.auto_filter.ref = "A1:D6"

ws4 = wb.create_sheet("方法论应用效果")
headers4 = ["方法论工具", "应用场景", "实施效果", "量化收益"]
data4 = [
    ["七个增长来源分析", "聚焦三大优先方向", "资源集中，ROI提升", "200万投入带来2700万增量营收"],
    ["四维诊断法", "市场/客户/竞争/能力诊断", "精准识别核心问题", "客户运营从0到1"],
    ["试点先行原则", "4个试点同步验证", "降低整体风险，少走弯路", "避免约150万错误投入"],
    ["增长飞轮设计", "三轮飞轮协同运转", "形成自我强化机制", "后期获客成本下降40%"],
    ["资源重配矩阵", "800万预算按ROI排序", "资金使用效率最大化", "综合ROI达3.4x"],
    ["行动计划管理", "四阶段里程碑管理", "执行有方向，进度可追踪", "关键任务完成率92%"],
]

for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

for row_idx, row_data in enumerate(data4, 2):
    fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

col_widths4 = [20, 28, 28, 28]
for col_idx, width in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(col_idx)].width = width
for row in range(2, 8):
    ws4.row_dimensions[row].height = 35
ws4.auto_filter.ref = "A1:D7"

output_path = "D:/新课开发/经营/08_存量竞争时代的增长设计：找到还能增长的路径/成果demo/10_项目成果展示.xlsx"
wb.save(output_path)
print("File saved to: " + output_path)

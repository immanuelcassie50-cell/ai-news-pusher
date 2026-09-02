#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
微课录制工作坊流程表单生成脚本
生成所有10个表单文件和1个汇总文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os

# 输出目录
OUTPUT_DIR = r"D:\新课开发\内训师和表达\微课录制\完整课程包\06_工具表单"

# 样式定义
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SECTION_FONT = Font(name="Microsoft YaHei", size=11, bold=True)
NORMAL_FONT = Font(name="Microsoft YaHei", size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_column_widths(ws, widths):
    """设置列宽"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def set_row_heights(ws, heights):
    """设置行高"""
    for row_num, height in heights.items():
        ws.row_dimensions[row_num].height = height

def apply_header_style(cell):
    """应用表头样式"""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGN

def apply_section_style(cell):
    """应用分区样式"""
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER
    cell.alignment = LEFT_ALIGN

def apply_normal_style(cell):
    """应用普通样式"""
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = LEFT_ALIGN

def create_F01():
    """F01_开场认知自测.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 空白版 =====
    ws1 = wb.active
    ws1.title = "空白版"

    # 标题行
    ws1.merge_cells('A1:E1')
    ws1['A1'] = '开场认知自测表'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 30

    # 列标题
    headers = ['序号', '题目', '完全同意', '不太同意', '说明']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    # 问题内容
    questions = [
        '1. 我清楚微课与传统课程的核心区别',
        '2. 我理解"一分钟一个知识点"的意义',
        '3. 我知道好微课的三个标准：有用、有趣、有力',
        '4. 我了解脚本在微课中的重要性',
        '5. 我掌握口语化表达的基本技巧',
        '6. 我知道如何进行镜头前的表情管理',
        '7. 我了解环境、设备对录制效果的影响',
        '8. 我理解"分段录制、后期剪辑"的工作流',
        '9. 我有信心完成本次微课录制任务'
    ]

    for row_idx, question in enumerate(questions, 3):
        ws1.cell(row=row_idx, column=1, value=row_idx-2).border = THIN_BORDER
        ws1.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=2, value=question).border = THIN_BORDER
        ws1.cell(row=row_idx, column=3, value='☐').border = THIN_BORDER
        ws1.cell(row=row_idx, column=3).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=4, value='☐').border = THIN_BORDER
        ws1.cell(row=row_idx, column=4).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=5, value='').border = THIN_BORDER

    set_column_widths(ws1, [6, 45, 12, 12, 25])

    # ===== Sheet2: 参考版 =====
    ws2 = wb.create_sheet("参考版")

    # 标题行
    ws2.merge_cells('A1:E1')
    ws2['A1'] = '开场认知自测表（参考版）'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    # 列标题
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    # 带答案的问题内容
    answers = [
        ('1. 我清楚微课与传统课程的核心区别', '☐', '☐', '微课是碎片化学习，一次一个核心观点'),
        ('2. 我理解"一分钟一个知识点"的意义', '☐', '☐', '注意力窗口有限，信息密度要控制'),
        ('3. 我知道好微课的三个标准：有用、有趣、有力', '☐', '☐', '有用=可落地，有趣=记得住，有力=能改变'),
        ('4. 我了解脚本在微课中的重要性', '☐', '☐', '没脚本的录制是冒险，没剪辑的冒险是浪费'),
        ('5. 我掌握口语化表达的基本技巧', '☐', '☐', '短句、停顿、口头禅、举例'),
        ('6. 我知道如何进行镜头前的表情管理', '☐', '☐', '适度微笑，眼神聚焦，眼睛看镜头'),
        ('7. 我了解环境、设备对录制效果的影响', '☐', '☐', '噪音、光线、麦克风位置都是关键'),
        ('8. 我理解"分段录制、后期剪辑"的工作流', '☐', '☐', '3-5分钟一段，错误重来不如后期剪辑'),
        ('9. 我有信心完成本次微课录制任务', '☐', '☐', '循序渐进，先完成再完美')
    ]

    for row_idx, (question, col3, col4, note) in enumerate(answers, 3):
        ws2.cell(row=row_idx, column=1, value=row_idx-2).border = THIN_BORDER
        ws2.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        ws2.cell(row=row_idx, column=2, value=question).border = THIN_BORDER
        ws2.cell(row=row_idx, column=3, value=col3).border = THIN_BORDER
        ws2.cell(row=row_idx, column=3).alignment = CENTER_ALIGN
        ws2.cell(row=row_idx, column=4, value=col4).border = THIN_BORDER
        ws2.cell(row=row_idx, column=4).alignment = CENTER_ALIGN
        ws2.cell(row=row_idx, column=5, value=note).border = THIN_BORDER

    set_column_widths(ws2, [6, 45, 12, 12, 25])

    wb.save(os.path.join(OUTPUT_DIR, 'F01_开场认知自测.xlsx'))
    print("F01 完成")

def create_F02():
    """F02_音频问题识别.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 空白版 =====
    ws1 = wb.active
    ws1.title = "空白版"

    # 标题
    ws1.merge_cells('A1:F1')
    ws1['A1'] = '音频问题识别练习表'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 30

    # 列标题
    headers = ['场景', '场景描述', '识别的问题', '问题原因', '改进建议', '我的判断']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    # 场景内容
    scenarios = [
        '场景1：录制环境',
        '场景2：设备使用',
        '场景3：表达方式',
        '场景4：后期检查'
    ]

    scenario_desc = [
        '在咖啡厅录制，周围有人交谈声',
        '使用领夹麦，但衣服摩擦声很明显',
        '一口气说了3分钟没有停顿',
        '回听发现鼠标点击声被录进去了'
    ]

    for row_idx in range(4):
        base_row = row_idx + 3
        ws1.cell(row=base_row, column=1, value=scenarios[row_idx]).border = THIN_BORDER
        ws1.cell(row=base_row, column=2, value=scenario_desc[row_idx]).border = THIN_BORDER
        ws1.cell(row=base_row, column=3, value='').border = THIN_BORDER
        ws1.cell(row=base_row, column=4, value='').border = THIN_BORDER
        ws1.cell(row=base_row, column=5, value='').border = THIN_BORDER
        ws1.cell(row=base_row, column=6, value='').border = THIN_BORDER
        ws1.row_dimensions[base_row].height = 35

    set_column_widths(ws1, [10, 22, 18, 18, 18, 12])

    # ===== Sheet2: 参考版 =====
    ws2 = wb.create_sheet("参考版")

    ws2.merge_cells('A1:F1')
    ws2['A1'] = '音频问题识别练习表（参考版）'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    ref_answers = [
        ('场景1：录制环境', '在咖啡厅录制，周围有人交谈声',
         '环境噪音', '未做隔音处理', '选择安静空间/使用降噪设备', '☐ 有问题'),
        ('场景2：设备使用', '使用领夹麦，但衣服摩擦声很明显',
         '麦克风位置不当', '麦克风与衣物接触', '调整麦克风位置/使用防喷罩', '☐ 有问题'),
        ('场景3：表达方式', '一口气说了3分钟没有停顿',
         '节奏过密', '缺乏段落意识', '按脚本分断句/使用提示词', '☐ 有问题'),
        ('场景4：后期检查', '回听发现鼠标点击声被录进去了',
         '背景噪音', '录音时未关闭鼠标声', '重新录制/使用音频软件降噪', '☐ 有问题')
    ]

    for row_idx, row_data in enumerate(ref_answers, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
        ws2.row_dimensions[row_idx].height = 40

    set_column_widths(ws2, [10, 22, 18, 18, 18, 12])

    wb.save(os.path.join(OUTPUT_DIR, 'F02_音频问题识别.xlsx'))
    print("F02 完成")

def create_F03():
    """F03_逐字稿口语化改写.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 空白版 =====
    ws1 = wb.active
    ws1.title = "空白版"

    ws1.merge_cells('A1:D1')
    ws1['A1'] = '逐字稿口语化改写练习表'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 30

    headers = ['序号', '原始稿（书面语）', '改写后（口语）', '改写标注']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    rows_content = [''] * 5
    for row_idx in range(3, 8):
        ws1.cell(row=row_idx, column=1, value=row_idx-2).border = THIN_BORDER
        ws1.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        for col_idx in range(2, 5):
            ws1.cell(row=row_idx, column=col_idx, value='').border = THIN_BORDER
        ws1.row_dimensions[row_idx].height = 40

    set_column_widths(ws1, [6, 35, 35, 30])

    # ===== Sheet2: 参考示例 =====
    ws2 = wb.create_sheet("参考示例")

    ws2.merge_cells('A1:D1')
    ws2['A1'] = '逐字稿口语化改写练习表（参考示例）'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    examples = [
        ('1', '首先，我们需要明确本次培训的目标受众。',
         '首先呢，我们要搞清楚，这次课是讲给谁听的。',
         '"首先"→"首先呢"，"我们"→"咱们"，"明确"→"搞清楚"'),
        ('2', '因此，建议各位在录制前做好充分准备。',
         '所以啊，建议大家在录之前好好准备一下。',
         '"因此"→"所以啊"，"各位"→"大家"，删除"充分"'),
        ('3', '研究表明，成人学习具有显著的特征。',
         '你知道吗，成年人学习是有它的特点的。',
         '"研究表明"→"你知道吗"，添加互动感'),
        ('4', '请各位注意以下操作要点。',
         '来，大家注意了啊，重点来了！',
         '感叹句增强吸引力，"操作要点"→"重点"'),
        ('5', '综上所述，我们可以得出以下结论。',
         '好啦，总结一下啊，今天主要说了……',
         '"综上所述"→"好啦总结一下"，具体化')
    ]

    for row_idx, row_data in enumerate(examples, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
        ws2.row_dimensions[row_idx].height = 45

    set_column_widths(ws2, [6, 35, 35, 30])

    # ===== Sheet3: 词汇替换参考 =====
    ws3 = wb.create_sheet("词汇替换参考")

    ws3.merge_cells('A1:C1')
    ws3['A1'] = '书面语→口语 词汇替换参考表'
    apply_header_style(ws3['A1'])
    ws3.row_dimensions[1].height = 30

    headers3 = ['序号', '书面语', '口语表达']
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    replacements = [
        ('1', '首先、第一步', '首先呢、第一步先'),
        ('2', '其次、另外', '然后、还有'),
        ('3', '因此、所以', '所以啊、那所以'),
        ('4', '然而、但是', '但是呢、不过'),
        ('5', '综上所述、总而言之', '好啦总结一下、总的来说'),
        ('6', '请注意、需要注意的是', '大家注意了啊、重点来了'),
        ('7', '非常重要、尤为关键', '特别重要、关键是'),
        ('8', '各位、各位学员', '大家、咱们'),
        ('9', '明确、清楚', '搞清楚、弄明白'),
        ('10', '具有、拥有', '有、是'),
        ('11', '进行、开展', '做、搞'),
        ('12', '通过、根据', '按照、通过'),
        ('13', '如果、倘若', '要是、假如'),
        ('14', '虽然、尽管', '虽然说、虽然'),
        ('15', '不仅...而且...', '不光...还...'),
    ]

    for row_idx, row_data in enumerate(replacements, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == 3:
                cell.font = Font(name="Microsoft YaHei", size=10, color="0070C0")
        ws3.row_dimensions[row_idx].height = 22

    set_column_widths(ws3, [6, 25, 25])

    wb.save(os.path.join(OUTPUT_DIR, 'F03_逐字稿口语化改写.xlsx'))
    print("F03 完成")

def create_F04():
    """F04_镜头状态观察自评.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 空白版 =====
    ws1 = wb.active
    ws1.title = "空白版"

    ws1.merge_cells('A1:G1')
    ws1['A1'] = '镜头状态观察自评表'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 30

    headers = ['观察类别', '5分-优秀', '4分-良好', '3分-合格', '2-1分-需改进', '自我评分', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    categories = [
        ('眼神', '眼神自然、有神\n聚焦镜头\n与镜头有交流感',
         '眼神较自然\n偶尔偏移',
         '眼神略显僵硬\n有回避感',
         '眼神飘忽\n不敢看镜头'),
        ('表情', '表情自然、适度微笑\n情绪与内容匹配',
         '表情较自然\n偶有僵硬',
         '表情单调\n或过于夸张',
         '表情僵硬/紧绷\n无情感变化'),
        ('姿势', '姿态端正、自然\n肩颈放松\n身体无晃动',
         '姿态较端正\n轻微晃动',
         '姿态尚可\n有明显晃动',
         '驼背/歪斜\n频繁晃动'),
        ('手势', '手势自然、协调\n配合内容表达\n无多余小动作',
         '手势较自然\n偶有多余动作',
         '手势单调\n或过于频繁',
         '手势僵硬/刻板\n小动作多')
    ]

    for row_idx, (cat, desc5, desc4, desc3, desc21) in enumerate(categories, 3):
        ws1.cell(row=row_idx, column=1, value=cat).border = THIN_BORDER
        ws1.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=2, value=desc5).border = THIN_BORDER
        ws1.cell(row=row_idx, column=3, value=desc4).border = THIN_BORDER
        ws1.cell(row=row_idx, column=4, value=desc3).border = THIN_BORDER
        ws1.cell(row=row_idx, column=5, value=desc21).border = THIN_BORDER
        ws1.cell(row=row_idx, column=6, value='').border = THIN_BORDER
        ws1.cell(row=row_idx, column=6).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=7, value='').border = THIN_BORDER
        ws1.row_dimensions[row_idx].height = 50

    set_column_widths(ws1, [10, 22, 22, 22, 22, 10, 15])

    # ===== Sheet2: 参考示例 =====
    ws2 = wb.create_sheet("参考示例")

    ws2.merge_cells('A1:G1')
    ws2['A1'] = '镜头状态观察自评表（参考示例）'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    ref_examples = [
        ('眼神', '眼神自然有神，聚焦镜头，有交流感',
         '眼神较自然，偶尔偏移',
         '眼神略显僵硬，有回避感',
         '眼神飘忽，不敢看镜头', '4', '录到第3条才找到感觉'),
        ('表情', '表情自然，适度微笑，情绪与内容匹配',
         '表情较自然，偶有僵硬',
         '表情单调或过于夸张',
         '表情僵硬紧绷，无情感变化', '3', '微笑有点僵，需要练习'),
        ('姿势', '姿态端正自然，肩颈放松，身体无晃动',
         '姿态较端正，轻微晃动',
         '姿态尚可，有明显晃动',
         '驼背歪斜，频繁晃动', '4', '还好，就是有点紧张'),
        ('手势', '手势自然协调，配合内容表达，无多余小动作',
         '手势较自然，偶有多余动作',
         '手势单调或过于频繁',
         '手势僵硬刻板，小动作多', '3', '手势还不够自然，需要多练')
    ]

    for row_idx, row_data in enumerate(ref_examples, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == 6:
                cell.font = Font(name="Microsoft YaHei", size=11, bold=True)
                cell.alignment = CENTER_ALIGN
        ws2.row_dimensions[row_idx].height = 50

    set_column_widths(ws2, [10, 22, 22, 22, 22, 10, 15])

    wb.save(os.path.join(OUTPUT_DIR, 'F04_镜头状态观察自评.xlsx'))
    print("F04 完成")

def create_F05():
    """F05_录制环境设备自检.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 空白版 =====
    ws1 = wb.active
    ws1.title = "空白版"

    ws1.merge_cells('A1:E1')
    ws1['A1'] = '录制环境设备自检表'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 30

    headers = ['序号', '检查项目', '通过', '不通过', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    check_items = [
        '1. 环境噪音：安静无干扰（空调、门窗、脚步声等）',
        '2. 光线条件：光线充足且均匀（自然光或补光灯）',
        '3. 背景环境：整洁、无杂物、无反光物',
        '4. 麦克风测试：录音清晰、无爆音、无明显杂音',
        '5. 摄像头测试：画面清晰、对焦准确、角度合适',
        '6. 网络连接：网络稳定（如果是直播或上传）',
        '7. 设备电量：电脑/手机电量充足或接通电源',
        '8. 录制软件：软件正常运行、存储空间充足',
        '9. 备用方案：准备录音笔或手机作为备用设备'
    ]

    for row_idx, item in enumerate(check_items, 3):
        ws1.cell(row=row_idx, column=1, value=row_idx-2).border = THIN_BORDER
        ws1.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=2, value=item).border = THIN_BORDER
        ws1.cell(row=row_idx, column=3, value='☐').border = THIN_BORDER
        ws1.cell(row=row_idx, column=3).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=4, value='☐').border = THIN_BORDER
        ws1.cell(row=row_idx, column=4).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=5, value='').border = THIN_BORDER
        ws1.row_dimensions[row_idx].height = 28

    set_column_widths(ws1, [6, 45, 10, 10, 25])

    # ===== Sheet2: 问题解决指南 =====
    ws2 = wb.create_sheet("问题解决指南")

    ws2.merge_cells('A1:D1')
    ws2['A1'] = '常见问题解决指南'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    headers2 = ['问题类型', '问题描述', '解决方案', '优先级']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    solutions = [
        ('环境噪音', '空调声、窗外噪音', '关闭噪音源/换个房间/用降噪麦克风', '高'),
        ('环境噪音', '键盘声、鼠标声', '使用静音键盘/录制时不用鼠标', '高'),
        ('光线问题', '脸部过暗或过亮', '调整光源位置/使用补光灯', '高'),
        ('光线问题', '背后有强光', '调整摄像头角度或位置', '中'),
        ('麦克风问题', '声音闷、录不清', '麦克风距离嘴部15-20cm', '高'),
        ('麦克风问题', '喷麦、口水音', '使用防喷罩/麦克风偏离嘴部', '高'),
        ('画面问题', '画面模糊', '检查对焦/清洁镜头/调整光线', '高'),
        ('画面问题', '背景杂乱', '整理背景/使用虚拟背景', '中'),
        ('设备问题', '存储空间不足', '提前清理/更换存储设备', '高'),
        ('设备问题', '突然断电', '接通电源/充满电/备用设备', '高')
    ]

    for row_idx, row_data in enumerate(solutions, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == 4:
                cell.alignment = CENTER_ALIGN
                if value == '高':
                    cell.font = Font(name="Microsoft YaHei", size=10, color="FF0000", bold=True)
                else:
                    cell.font = Font(name="Microsoft YaHei", size=10, color="FFC000")
        ws2.row_dimensions[row_idx].height = 25

    set_column_widths(ws2, [12, 22, 40, 10])

    wb.save(os.path.join(OUTPUT_DIR, 'F05_录制环境设备自检.xlsx'))
    print("F05 完成")

def create_F06():
    """F06_分段录制规划.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 空白版 =====
    ws1 = wb.active
    ws1.title = "空白版"

    ws1.merge_cells('A1:G1')
    ws1['A1'] = '分段录制规划表'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 30

    headers = ['段落编号', '内容描述', '开头第一句', '结尾最后一句', '预计时长', '实际时长', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    for row_idx in range(3, 9):  # 6个段落
        ws1.cell(row=row_idx, column=1, value=f'段落{row_idx-2}').border = THIN_BORDER
        ws1.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        for col_idx in range(2, 8):
            ws1.cell(row=row_idx, column=col_idx, value='').border = THIN_BORDER
        ws1.cell(row=row_idx, column=7, value='待录制').border = THIN_BORDER
        ws1.cell(row=row_idx, column=7).alignment = CENTER_ALIGN
        ws1.row_dimensions[row_idx].height = 30

    set_column_widths(ws1, [10, 25, 25, 25, 12, 12, 10])

    # ===== Sheet2: 参考示例 =====
    ws2 = wb.create_sheet("参考示例")

    ws2.merge_cells('A1:G1')
    ws2['A1'] = '分段录制规划表（参考示例）'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    ref_segments = [
        ('段落1', '课程导入：自我介绍 + 课程目标',
         '大家好，我是某某，今天跟大家分享……',
         '希望通过今天的学习，大家能……', '0:30', '0:28', '已完成'),
        ('段落2', '痛点引入：学员常见问题',
         '你有没有遇到过这样的问题……',
         '如果你也有同感，那今天的课就是为你准备的', '0:45', '0:50', '已完成'),
        ('段落3', '核心知识点1：什么是微课',
         '首先我们来搞清楚，什么是微课……',
         '总结一下：微课就是……', '1:30', '1:35', '已完成'),
        ('段落4', '核心知识点2：微课设计要点',
         '知道了什么是微课，那怎么设计一个好微课呢……',
         '记住这三个要点：有用、有趣、有力', '2:00', '2:10', '已完成'),
        ('段落5', '实战演示：如何写脚本',
         '知道了要点还不会做？没关系，我演示给你看……',
         '按照这个模板，你也能写出好脚本', '2:30', '2:45', '需重录'),
        ('段落6', '总结回顾：三个要点 + 行动计划',
         '好了，今天的课到这里就要结束了……',
         '从今天开始，试试用微课的方式分享你的经验吧！', '0:45', '0:42', '待录制')
    ]

    for row_idx, row_data in enumerate(ref_segments, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == 7:
                if value == '已完成':
                    cell.font = Font(name="Microsoft YaHei", size=10, color="00B050")
                elif value == '需重录':
                    cell.font = Font(name="Microsoft YaHei", size=10, color="FF0000")
                else:
                    cell.font = Font(name="Microsoft YaHei", size=10, color="FFC000")
                cell.alignment = CENTER_ALIGN
        ws2.row_dimensions[row_idx].height = 35

    set_column_widths(ws2, [10, 25, 25, 25, 12, 12, 10])

    wb.save(os.path.join(OUTPUT_DIR, 'F06_分段录制规划.xlsx'))
    print("F06 完成")

def create_F07():
    """F07_录制成果验收三步表.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 空白版 =====
    ws1 = wb.active
    ws1.title = "空白版"

    ws1.merge_cells('A1:E1')
    ws1['A1'] = '录制成果验收三步表'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 30

    # 第一步：音频验收
    ws1.merge_cells('A2:E2')
    ws1['A2'] = '第一步：音频验收'
    apply_section_style(ws1['A2'])

    audio_headers = ['序号', '验收项目', '通过', '不通过', '备注']
    for col, header in enumerate(audio_headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        apply_section_style(cell)

    audio_items = [
        '1. 声音清晰，无杂音/噪音',
        '2. 音量适中，无爆音/失真',
        '3. 无明显停顿/口误，或已剪辑处理',
        '4. 语速合适，有自然停顿'
    ]

    for row_idx, item in enumerate(audio_items, 4):
        ws1.cell(row=row_idx, column=1, value=row_idx-3).border = THIN_BORDER
        ws1.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=2, value=item).border = THIN_BORDER
        ws1.cell(row=row_idx, column=3, value='☐').border = THIN_BORDER
        ws1.cell(row=row_idx, column=3).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=4, value='☐').border = THIN_BORDER
        ws1.cell(row=row_idx, column=4).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=5, value='').border = THIN_BORDER
        ws1.row_dimensions[row_idx].height = 25

    # 第二步：视频验收
    ws1.merge_cells('A8:E8')
    ws1['A8'] = '第二步：视频验收'
    apply_section_style(ws1['A8'])

    for col, header in enumerate(audio_headers, 1):
        cell = ws1.cell(row=9, column=col, value=header)
        apply_section_style(cell)

    video_items = [
        '1. 画面清晰，分辨率符合要求',
        '2. 光线合适，脸部不过暗/过亮',
        '3. 背景整洁，无干扰元素',
        '4. 画面稳定，无明显抖动'
    ]

    for row_idx, item in enumerate(video_items, 10):
        ws1.cell(row=row_idx, column=1, value=row_idx-9).border = THIN_BORDER
        ws1.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=2, value=item).border = THIN_BORDER
        ws1.cell(row=row_idx, column=3, value='☐').border = THIN_BORDER
        ws1.cell(row=row_idx, column=3).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=4, value='☐').border = THIN_BORDER
        ws1.cell(row=row_idx, column=4).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=5, value='').border = THIN_BORDER
        ws1.row_dimensions[row_idx].height = 25

    # 第三步：表达验收
    ws1.merge_cells('A14:E14')
    ws1['A14'] = '第三步：表达验收'
    apply_section_style(ws1['A14'])

    for col, header in enumerate(audio_headers, 1):
        cell = ws1.cell(row=15, column=col, value=header)
        apply_section_style(cell)

    express_items = [
        '1. 内容完整，有开头/正文/结尾',
        '2. 逻辑清晰，层次分明',
        '3. 语言口语化，易于理解',
        '4. 整体效果：是否达到预期？'
    ]

    for row_idx, item in enumerate(express_items, 16):
        ws1.cell(row=row_idx, column=1, value=row_idx-15).border = THIN_BORDER
        ws1.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=2, value=item).border = THIN_BORDER
        ws1.cell(row=row_idx, column=3, value='☐').border = THIN_BORDER
        ws1.cell(row=row_idx, column=3).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=4, value='☐').border = THIN_BORDER
        ws1.cell(row=row_idx, column=4).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=5, value='').border = THIN_BORDER
        ws1.row_dimensions[row_idx].height = 25

    # 决策行
    ws1.merge_cells('A20:E20')
    ws1['A20'] = '最终决策：☐ 验收通过   ☐ 需要部分重录（标记需要重录的段落：          ）   ☐ 需要整体重录'
    ws1['A20'].border = THIN_BORDER
    ws1['A20'].font = Font(name="Microsoft YaHei", size=10)
    ws1['A20'].alignment = LEFT_ALIGN

    set_column_widths(ws1, [6, 35, 10, 10, 30])

    # ===== Sheet2: 参考示例 =====
    ws2 = wb.create_sheet("参考示例")

    ws2.merge_cells('A1:E1')
    ws2['A1'] = '录制成果验收三步表（参考示例）'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    # 第一步：音频验收
    ws2.merge_cells('A2:E2')
    ws2['A2'] = '第一步：音频验收'
    apply_section_style(ws2['A2'])

    for col, header in enumerate(audio_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        apply_section_style(cell)

    audio_ref = [
        ('1', '声音清晰，无杂音/噪音', '☐', '☐', '整体OK'),
        ('2', '音量适中，无爆音/失真', '☐', '', '第23秒有轻微爆音'),
        ('3', '无明显停顿/口误，或已剪辑处理', '☐', '☐', '有2处口误，已剪掉'),
        ('4', '语速合适，有自然停顿', '☐', '', '语速偏快')
    ]

    for row_idx, row_data in enumerate(audio_ref, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in [3, 4]:
                cell.alignment = CENTER_ALIGN
        ws2.row_dimensions[row_idx].height = 25

    # 第二步：视频验收
    ws2.merge_cells('A8:E8')
    ws2['A8'] = '第二步：视频验收'
    apply_section_style(ws2['A8'])

    for col, header in enumerate(audio_headers, 1):
        cell = ws2.cell(row=9, column=col, value=header)
        apply_section_style(cell)

    video_ref = [
        ('1', '画面清晰，分辨率符合要求', '☐', '', ''),
        ('2', '光线合适，脸部不过暗/过亮', '☐', '☐', '光线有点暗'),
        ('3', '背景整洁，无干扰元素', '☐', '', '背景有反光'),
        ('4', '画面稳定，无明显抖动', '☐', '', '')
    ]

    for row_idx, row_data in enumerate(video_ref, 10):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in [3, 4]:
                cell.alignment = CENTER_ALIGN
        ws2.row_dimensions[row_idx].height = 25

    # 第三步：表达验收
    ws2.merge_cells('A14:E14')
    ws2['A14'] = '第三步：表达验收'
    apply_section_style(ws2['A14'])

    for col, header in enumerate(audio_headers, 1):
        cell = ws2.cell(row=15, column=col, value=header)
        apply_section_style(cell)

    express_ref = [
        ('1', '内容完整，有开头/正文/结尾', '☐', '', ''),
        ('2', '逻辑清晰，层次分明', '☐', '', ''),
        ('3', '语言口语化，易于理解', '☐', '☐', '有几句书面语'),
        ('4', '整体效果：是否达到预期？', '☐', '', '整体OK，细节待改进')
    ]

    for row_idx, row_data in enumerate(express_ref, 16):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in [3, 4]:
                cell.alignment = CENTER_ALIGN
        ws2.row_dimensions[row_idx].height = 25

    # 决策行
    ws2.merge_cells('A20:E20')
    ws2['A20'] = '最终决策：☐ 验收通过   ☑ 需要部分重录（标记需要重录的段落：段落2、段落4开头   ）   ☐ 需要整体重录'
    ws2['A20'].border = THIN_BORDER
    ws2['A20'].font = Font(name="Microsoft YaHei", size=10)
    ws2['A20'].alignment = LEFT_ALIGN

    set_column_widths(ws2, [6, 35, 10, 10, 30])

    wb.save(os.path.join(OUTPUT_DIR, 'F07_录制成果验收三步表.xlsx'))
    print("F07 完成")

def create_F08():
    """F08_互评反馈.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 空白版 =====
    ws1 = wb.active
    ws1.title = "空白版"

    ws1.merge_cells('A1:D1')
    ws1['A1'] = '互评反馈表'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 30

    headers = ['评价维度', '具体描述', '优点', '改进建议']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    dimensions = ['音频', '画面', '表达', '综合']

    for row_idx, dim in enumerate(dimensions, 3):
        ws1.cell(row=row_idx, column=1, value=dim).border = THIN_BORDER
        ws1.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        ws1.cell(row=row_idx, column=1).font = Font(name="Microsoft YaHei", size=11, bold=True)
        for col_idx in range(2, 5):
            ws1.cell(row=row_idx, column=col_idx, value='').border = THIN_BORDER
        ws1.row_dimensions[row_idx].height = 35

    # 互评人信息
    ws1.merge_cells('A8:D8')
    ws1['A8'] = '被评价人：__________    评价人：__________    日期：__________'
    ws1['A8'].border = THIN_BORDER
    ws1['A8'].font = Font(name="Microsoft YaHei", size=10)

    set_column_widths(ws1, [10, 25, 25, 25])

    # ===== Sheet2: 参考示例 =====
    ws2 = wb.create_sheet("参考示例")

    ws2.merge_cells('A1:D1')
    ws2['A1'] = '互评反馈表（参考示例）'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    ref_feedback = [
        ('音频', '声音清晰度、音量、噪音控制',
         '声音很有磁性，录音质量高',
         '有轻微的背景噪音，建议用Audacity降噪'),
        ('画面', '光线、背景、画面稳定性',
         '光线柔和，画面清晰，背景简洁',
         '脸有点暗，下次考虑加个补光灯'),
        ('表达', '语速、停顿、眼神交流、肢体语言',
         '语言口语化，例子生动，引人入胜',
         '语速稍快，建议每句话后稍作停顿'),
        ('综合', '整体效果、内容价值、观看感受',
         '干货满满，学到很多实用的技巧',
         '可以增加一些互动提问环节')
    ]

    for row_idx, row_data in enumerate(ref_feedback, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.font = Font(name="Microsoft YaHei", size=11, bold=True)
                cell.alignment = CENTER_ALIGN
        ws2.row_dimensions[row_idx].height = 40

    ws2.merge_cells('A8:D8')
    ws2['A8'] = '被评价人：张老师    评价人：李老师    日期：2024/01/15'
    ws2['A8'].border = THIN_BORDER
    ws2['A8'].font = Font(name="Microsoft YaHei", size=10)

    set_column_widths(ws2, [10, 25, 25, 25])

    wb.save(os.path.join(OUTPUT_DIR, 'F08_互评反馈.xlsx'))
    print("F08 完成")

def create_F09():
    """F09_要不要重录快速判断.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 决策表 =====
    ws1 = wb.active
    ws1.title = "决策表"

    ws1.merge_cells('A1:E1')
    ws1['A1'] = '要不要重录快速决策表'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 30

    headers = ['序号', '场景', '问题严重程度', '建议操作', '判断理由']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    scenarios = [
        ('1', '音频有噪音，但很短', '轻微', '考虑保留', '噪音短且不影响理解，可接受'),
        ('2', '音频有明显噪音', '严重', '建议重录', '噪音影响收听体验'),
        ('3', '语速有一处明显偏快', '轻微', '考虑保留', '偶尔的语速问题不影响整体'),
        ('4', '语速整体偏快，全程如此', '严重', '建议重录', '全程语速快会导致理解困难'),
        ('5', '有一句口误，已剪辑处理', '轻微', '保留已剪辑版', '剪辑处理后无明显问题'),
        ('6', '有多处口误，剪辑后仍明显', '严重', '建议重录', '剪辑后仍有痕迹，影响体验'),
        ('7', '光线稍微有点暗', '轻微', '考虑保留', '轻微问题不影响理解'),
        ('8', '脸部完全看不清楚', '严重', '必须重录', '光线问题严重影响内容传达'),
        ('9', '背景有个物品稍微碍眼', '轻微', '考虑保留', '小瑕疵不影响主要内容'),
        ('10', '背景杂乱喧宾夺主', '严重', '建议重录或换背景', '背景喧宾夺主影响专业度'),
        ('11', '有一句话感觉不够好', '轻微', '保留', '个人完美主义倾向，不影响他人'),
        ('12', '整体感觉不对，但说不上来', '主观', '先放着，冷静后再判断', '主观感受需要冷静客观化')
    ]

    for row_idx, row_data in enumerate(scenarios, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == 4:
                if value == '建议重录' or value == '必须重录':
                    cell.font = Font(name="Microsoft YaHei", size=10, color="FF0000", bold=True)
                elif value == '考虑保留' or value == '保留已剪辑版' or value == '保留':
                    cell.font = Font(name="Microsoft YaHei", size=10, color="00B050")
                else:
                    cell.font = Font(name="Microsoft YaHei", size=10, color="FFC000")
                cell.alignment = CENTER_ALIGN
            elif col_idx == 3:
                cell.alignment = CENTER_ALIGN
                if value == '严重' or value == '必须重录':
                    cell.font = Font(name="Microsoft YaHei", size=10, color="FF0000")
                elif value == '轻微':
                    cell.font = Font(name="Microsoft YaHei", size=10, color="00B050")
                else:
                    cell.font = Font(name="Microsoft YaHei", size=10, color="FFC000")
        ws1.row_dimensions[row_idx].height = 25

    set_column_widths(ws1, [6, 28, 12, 15, 30])

    # ===== Sheet2: 快速三问 =====
    ws2 = wb.create_sheet("快速三问")

    ws2.merge_cells('A1:B1')
    ws2['A1'] = '重录决策快速三问'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    questions = [
        ('问题1', '这个问题是否会影响观众理解核心内容？',
         '☐ 不影响\n↓ 不用重录',
         '☐ 会影响\n↓ 继续判断'),
        ('', '', '', ''),
        ('问题2', '这个问题能否通过后期剪辑修复？',
         '☐ 可以\n↓ 剪辑处理',
         '☐ 不可以\n↓ 继续判断'),
        ('', '', '', ''),
        ('问题3', '重录这个问题需要多少时间和精力？',
         '☐ 很少\n→ 建议重录\n（投入产出比划算）',
         '☐ 很多\n→ 考虑保留\n（投入产出比不划算）')
    ]

    # 标题行
    ws2.merge_cells('A2:B2')
    ws2['A2'] = '问自己'
    ws2['A2'].font = Font(name="Microsoft YaHei", size=11, bold=True)
    ws2['A2'].fill = SECTION_FILL
    ws2['A2'].border = THIN_BORDER
    ws2['A2'].alignment = CENTER_ALIGN

    ws2.merge_cells('C2:D2')
    ws2['C2'] = '选项'
    ws2['C2'].font = Font(name="Microsoft YaHei", size=11, bold=True)
    ws2['C2'].fill = SECTION_FILL
    ws2['C2'].border = THIN_BORDER
    ws2['C2'].alignment = CENTER_ALIGN

    for row_idx, row_data in enumerate(questions, 3):
        if row_data[0]:  # 非空行
            ws2.merge_cells(f'A{row_idx}:B{row_idx}')
            ws2.cell(row=row_idx, column=1, value=f'{row_data[0]}：{row_data[1]}').border = THIN_BORDER
            ws2.cell(row=row_idx, column=1).font = Font(name="Microsoft YaHei", size=10, bold=True)
            ws2.cell(row=row_idx, column=1).alignment = LEFT_ALIGN

            ws2.cell(row=row_idx, column=3, value=row_data[2]).border = THIN_BORDER
            ws2.cell(row=row_idx, column=3).font = Font(name="Microsoft YaHei", size=9, color="00B050")
            ws2.cell(row=row_idx, column=3).alignment = LEFT_ALIGN

            ws2.cell(row=row_idx, column=4, value=row_data[3]).border = THIN_BORDER
            ws2.cell(row=row_idx, column=4).font = Font(name="Microsoft YaHei", size=9, color="FF0000")
            ws2.cell(row=row_idx, column=4).alignment = LEFT_ALIGN
        else:
            # 空行
            for col in range(1, 5):
                ws2.cell(row=row_idx, column=col).border = THIN_BORDER
        ws2.row_dimensions[row_idx].height = 35

    # 总结
    ws2.merge_cells('A9:D9')
    ws2['A9'] = '总结：先问"影响理解吗" → 再问"能剪辑吗" → 最后问"值得重录吗"'
    ws2['A9'].border = THIN_BORDER
    ws2['A9'].font = Font(name="Microsoft YaHei", size=10, bold=True, color="1F4E79")
    ws2['A9'].alignment = CENTER_ALIGN
    ws2['A9'].fill = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")

    set_column_widths(ws2, [6, 30, 25, 25])

    wb.save(os.path.join(OUTPUT_DIR, 'F09_要不要重录快速判断.xlsx'))
    print("F09 完成")

def create_F10():
    """F10_工作坊反思总结.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 空白版 =====
    ws1 = wb.active
    ws1.title = "空白版"

    ws1.merge_cells('A1:B1')
    ws1['A1'] = '工作坊反思总结表'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 30

    sections = [
        ('收获1', ''),
        ('收获2', ''),
        ('收获3', ''),
        ('行动计划', ''),
        ('未解决问题', '')
    ]

    row = 2
    for section, _ in sections:
        ws1.merge_cells(f'A{row}:B{row}')
        ws1.cell(row=row, column=1, value=section)
        ws1.cell(row=row, column=1).font = Font(name="Microsoft YaHei", size=11, bold=True)
        ws1.cell(row=row, column=1).fill = SECTION_FILL
        ws1.cell(row=row, column=1).border = THIN_BORDER
        ws1.cell(row=row, column=1).alignment = LEFT_ALIGN
        ws1.row_dimensions[row].height = 25
        row += 1

        ws1.merge_cells(f'A{row}:B{row}')
        ws1.cell(row=row, column=1, value='')
        ws1.cell(row=row, column=1).border = THIN_BORDER
        ws1.row_dimensions[row].height = 50
        row += 1

    set_column_widths(ws1, [15, 45])

    # ===== Sheet2: 参考示例 =====
    ws2 = wb.create_sheet("参考示例")

    ws2.merge_cells('A1:B1')
    ws2['A1'] = '工作坊反思总结表（参考示例）'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    ref_content = [
        ('收获1', '明白了微课与传统课程的核心区别：微课是碎片化学习，一次一个核心观点，一分钟传递一个知识点'),
        ('收获2', '学会了口语化改写的技巧：短句、停顿、口头禅、举例，让内容更生动易懂'),
        ('收获3', '掌握了分段录制的工作流：先规划再录制，错误重来不如后期剪辑'),
        ('行动计划', '1. 下周录制我的第一门微课\n2. 每门课控制在3-5分钟\n3. 每次录制前先写脚本\n4. 用今天学的评分表自评'),
        ('未解决问题', '1. 如何在录制时保持自然的表情？\n2. 灯光设备的选择还在犹豫\n3. 剪辑软件还没选定')
    ]

    row = 2
    for section, content in ref_content:
        ws2.merge_cells(f'A{row}:B{row}')
        ws2.cell(row=row, column=1, value=section)
        ws2.cell(row=row, column=1).font = Font(name="Microsoft YaHei", size=11, bold=True)
        ws2.cell(row=row, column=1).fill = SECTION_FILL
        ws2.cell(row=row, column=1).border = THIN_BORDER
        ws2.cell(row=row, column=1).alignment = LEFT_ALIGN
        ws2.row_dimensions[row].height = 25
        row += 1

        ws2.merge_cells(f'A{row}:B{row}')
        ws2.cell(row=row, column=1, value=content)
        ws2.cell(row=row, column=1).border = THIN_BORDER
        ws2.cell(row=row, column=1).font = Font(name="Microsoft YaHei", size=10)
        ws2.cell(row=row, column=1).alignment = LEFT_ALIGN
        ws2.row_dimensions[row].height = 50
        row += 1

    set_column_widths(ws2, [15, 45])

    wb.save(os.path.join(OUTPUT_DIR, 'F10_工作坊反思总结.xlsx'))
    print("F10 完成")

def create_master():
    """全流程表单_汇总.xlsx"""
    wb = Workbook()

    # ===== Sheet1: 目录 =====
    ws1 = wb.active
    ws1.title = "目录"

    ws1.merge_cells('A1:D1')
    ws1['A1'] = '微课录制工作坊表单汇总'
    apply_header_style(ws1['A1'])
    ws1.row_dimensions[1].height = 35

    headers = ['表单编号', '表单名称', '使用时机', '简要说明']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    forms_index = [
        ('F01', '开场认知自测', '工作坊开始前', '了解学员对微课的认知基础'),
        ('F02', '音频问题识别', '音频技巧学习后', '练习识别常见音频问题'),
        ('F03', '逐字稿口语化改写', '脚本撰写环节', '将书面语转化为口语表达'),
        ('F04', '镜头状态观察自评', '表情管理学习后', '自评镜头前的表现状态'),
        ('F05', '录制环境设备自检', '录制前准备', '检查环境和设备是否就绪'),
        ('F06', '分段录制规划', '录制前规划', '规划微课内容的分段结构'),
        ('F07', '录制成果验收三步表', '录制后检查', '系统化验收音频/视频/表达'),
        ('F08', '互评反馈', '同伴互评环节', '收集同伴的改进建议'),
        ('F09', '要不要重录快速判断', '成果检查后', '快速决策是否需要重录'),
        ('F10', '工作坊反思总结', '工作坊结束时', '总结学习收获和行动计划')
    ]

    for row_idx, row_data in enumerate(forms_index, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
                cell.alignment = CENTER_ALIGN
        ws1.row_dimensions[row_idx].height = 28

    set_column_widths(ws1, [10, 22, 18, 35])

    # ===== Sheet2: 进度跟踪 =====
    ws2 = wb.create_sheet("进度跟踪")

    ws2.merge_cells('A1:E1')
    ws2['A1'] = '工作坊进度跟踪表'
    apply_header_style(ws2['A1'])
    ws2.row_dimensions[1].height = 30

    headers2 = ['序号', '表单', '完成状态', '完成日期', '备注']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        apply_section_style(cell)

    for row_idx in range(3, 13):
        ws2.cell(row=row_idx, column=1, value=row_idx-2).border = THIN_BORDER
        ws2.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        ws2.cell(row=row_idx, column=2, value=f'F{row_idx-2:02d}').border = THIN_BORDER
        ws2.cell(row=row_idx, column=3, value='☐ 未完成\n☑ 已完成').border = THIN_BORDER
        ws2.cell(row=row_idx, column=3).alignment = CENTER_ALIGN
        ws2.cell(row=row_idx, column=4, value='').border = THIN_BORDER
        ws2.cell(row=row_idx, column=4).alignment = CENTER_ALIGN
        ws2.cell(row=row_idx, column=5, value='').border = THIN_BORDER
        ws2.row_dimensions[row_idx].height = 30

    # 总体进度
    ws2.merge_cells('A14:E14')
    ws2['A14'] = '总体进度：0/10'
    ws2['A14'].border = THIN_BORDER
    ws2['A14'].font = Font(name="Microsoft YaHei", size=11, bold=True)
    ws2['A14'].alignment = CENTER_ALIGN
    ws2['A14'].fill = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")

    # 使用说明
    ws2.merge_cells('A16:E16')
    ws2['A16'] = '使用说明：每完成一个表单，在对应行的"完成状态"列打勾，并填写完成日期'
    ws2['A16'].font = Font(name="Microsoft YaHei", size=9, italic=True, color="666666")

    set_column_widths(ws2, [6, 12, 18, 12, 25])

    wb.save(os.path.join(OUTPUT_DIR, '全流程表单_汇总.xlsx'))
    print("全流程表单_汇总 完成")

def main():
    """主函数：生成所有表单"""
    print("开始生成微课录制工作坊表单...")

    # 确保输出目录存在
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 生成所有表单
    create_F01()
    create_F02()
    create_F03()
    create_F04()
    create_F05()
    create_F06()
    create_F07()
    create_F08()
    create_F09()
    create_F10()
    create_master()

    print(f"\n全部完成！共生成11个文件。")
    print(f"输出目录：{OUTPUT_DIR}")

if __name__ == "__main__":
    main()

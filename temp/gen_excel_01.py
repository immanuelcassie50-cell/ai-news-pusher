# -*- coding: utf-8 -*-
"""
生成 16_个人问题跟踪表_学员用.xlsx
3 sheets:
  1. 我的工作问题库 (问题记录)
  2. 问题解决方案档案 (方案执行跟踪)
  3. 工具应用统计 (工具使用频率)
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference

# === 主题色：深蓝（与 G7/G8 体系延续） ===
ACCENT = "1F3A68"
ACCENT_SOFT = "DDE6F0"
ACCENT_DARK = "15294A"
OK = "166534"
OK_SOFT = "E8F1EB"
WARN = "C2410C"
WARN_SOFT = "FEF0E8"
NEUTRAL = "6B7280"
NEUTRAL_SOFT = "F3F4F6"
LINE = "B8C2D4"
LINE_SOFT = "E5E9F0"
BG = "F5F7FA"
PAPER = "FFFFFF"
INK = "1A1F2C"
INK_SOFT = "4A5266"

# PatternFill 包装
def PF(color):
    return PatternFill("solid", fgColor=color)

TITLE_FILL = PF(ACCENT)
H2_FILL = PF(ACCENT_SOFT)
HEADER_FILL = PF(ACCENT)
OK_FILL = PF(OK_SOFT)
WARN_FILL = PF(WARN_SOFT)
BG_FILL = PF(BG)
NEUTRAL_FILL = PF(NEUTRAL_SOFT)

# 字体
TITLE_FONT = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
H2_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color=ACCENT_DARK)
HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Microsoft YaHei", size=10, color=INK)
SOFT_FONT = Font(name="Microsoft YaHei", size=10, color=INK_SOFT)
BOLD_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color=INK)
NOTE_FONT = Font(name="Microsoft YaHei", size=9, italic=True, color=NEUTRAL)

# 填充 (already defined in PF above)
# TITLE_FILL = PatternFill("solid", fgColor=ACCENT)
# H2_FILL = PatternFill("solid", fgColor=ACCENT_SOFT)
# ...

# 边框
thin = Side(border_style="thin", color=LINE)
medium = Side(border_style="medium", color=ACCENT)
ALL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_BORDER = Border(left=medium, right=medium, top=medium, bottom=medium)

# 对齐
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()
wb.remove(wb.active)  # 移除默认 sheet


# ============================================================
# Sheet 1: 我的工作问题库
# ============================================================
ws1 = wb.create_sheet("我的工作问题库")

# 标题
ws1.merge_cells("A1:H1")
ws1["A1"] = "我的工作问题库"
ws1["A1"].font = TITLE_FONT
ws1["A1"].fill = TITLE_FILL
ws1["A1"].alignment = CENTER
ws1.row_dimensions[1].height = 32

# 副标题
ws1.merge_cells("A2:H2")
ws1["A2"] = "说明：用 1 行记录 1 个工作问题。问题定义句 = 现状·标准·差距·影响；用工具应用情况来跟踪是否用上了"
ws1["A2"].font = NOTE_FONT
ws1["A2"].fill = BG_FILL
ws1["A2"].alignment = LEFT
ws1.row_dimensions[2].height = 24

# 表头
headers1 = [
    ("序号", 6),
    ("问题名称", 22),
    ("问题类型", 14),
    ("紧急度", 10),
    ("影响范围", 14),
    ("问题定义句(现状·标准·差距·影响)", 60),
    ("使用工具", 18),
    ("解决状态", 14),
]
for col_idx, (text, width) in enumerate(headers1, 1):
    c = ws1.cell(row=3, column=col_idx, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
    ws1.column_dimensions[get_column_letter(col_idx)].width = width
ws1.row_dimensions[3].height = 36

# 示例数据（基于"设备故障""客户投诉""员工流失"等典型场景）
sample1 = [
    (1, "12 台关键设备故障率偏高", "运营·设备", "高", "全公司",
     "在我们设备部，过去 6 个月里，关键设备故障率方面，现状是 5.2%，标准是行业平均 3.8%（老板要求 4% 以下），差距是 1.2-1.4 个百分点，影响是每月停机 18 小时（折合 32 万）+ 订单延迟 2 次 + 客户投诉 1 次。",
     "5-Why + 三好标准", "已解决"),
    (2, "客户投诉响应慢", "服务·客户", "中", "客服部+技术部",
     "在客服部，过去 3 个月里，投诉首次响应时长方面，现状是 4.2 小时，标准是公司规定 2 小时，差距是 2.2 小时（超 110%），影响是客户满意度从 88% 降到 76%，续约率预计下降 5-8%。",
     "问题定义句 + 5W", "解决中"),
    (3, "新员工 3 个月内离职率高", "人力·团队", "中", "全公司",
     "在过去 1 年里，新员工 3 个月内离职率方面，现状是 28%，标准是公司目标 15%（行业平均 18%），差距是 10-13 个百分点，影响是招聘成本增加约 35 万，团队士气受影响。",
     "列现象 + 5-Why", "待启动"),
    (4, "月度报表数据错误率高", "运营·数据", "低", "财务部+业务部",
     "在财务部，过去 6 个月里，月度报表一次通过率方面，现状是 62%，标准是 95% 以上，差距是 33 个百分点，影响是每月多花 2-3 人/周返工，部门间信任度下降。",
     "问题定义句 + 边界澄清", "待启动"),
    (5, "生产线 OEE 偏低", "生产·效率", "高", "生产部",
     "在 A 产线，过去 3 个月里，OEE 方面，现状是 68%，标准是行业平均 82%（公司目标 80%），差距是 12-14 个百分点，影响是月产值损失约 80 万。",
     "5-Why + 风险预案", "解决中"),
    (6, "供应商交付准时率低", "供应链·采购", "中", "采购部+生产部",
     "在采购部，过去 6 个月里，关键供应商准时交付率方面，现状是 73%，标准是合同要求 95%，差距是 22 个百分点，影响是生产线停工待料 5 次/月，临时加急费用增加 12 万/年。",
     "三好标准 + 5W", "已解决"),
    (7, "会议效率低、决策慢", "管理·流程", "低", "全公司",
     "在过去 1 个月里，周会决策时长方面，现状是平均 1.8 小时/会，标准是公司要求 1 小时内，差距是 0.8 小时/会，影响是每周浪费约 25 人时，重要议题延后 30%。",
     "边界澄清 + 5W", "已搁置"),
    (8, "客户回款周期长", "财务·应收", "中", "财务部+销售部",
     "在销售部，过去 6 个月里，应收账款周转天数方面，现状是 78 天，标准是行业平均 45 天（公司目标 60 天），差距是 18-33 天，影响是流动资金压力增加 200 万，财务成本上升。",
     "问题定义句 + 5-Why", "解决中"),
    (9, "产品退货率高", "质量·生产", "高", "品保部+生产部",
     "在品保部，过去 3 个月里，退货率方面，现状是 4.8%，标准是公司目标 2% 以下（行业平均 2.5%），差距是 2.3-2.8 个百分点，影响是退货损失 18 万/月 + 客户口碑受损。",
     "列现象 + 5-Why + 根因验证", "解决中"),
    (10, "员工提案参与度低", "管理·团队", "低", "全公司",
     "在过去半年里，员工提案参与率方面，现状是 12%，标准是公司目标 30%，差距是 18 个百分点，影响是基层改善想法被埋没，管理层决策信息不完整。",
     "边界澄清 + 5W", "待启动"),
    (11, "设备备件库存积压", "运营·库存", "中", "设备部",
     "在设备部，过去 1 年里，备件库存周转率方面，现状是 1.8 次/年，标准是 4 次/年，差距是 2.2 次/年，影响是占用资金 38 万，仓储面积 20% 浪费。",
     "三好标准 + 风险预案", "已解决"),
    (12, "项目延期交付频繁", "管理·项目", "高", "项目部",
     "在项目部，过去半年里，项目准时交付率方面，现状是 58%，标准是公司目标 85%，差距是 27 个百分点，影响是客户罚款 15 万 + 续约率下降。",
     "5W + 关键节点复盘", "解决中"),
]

# 数据验证（下拉）
from openpyxl.worksheet.datavalidation import DataValidation

dv_type = DataValidation(type="list", formula1='"运营·生产,运营·设备,服务·客户,人力·团队,运营·数据,生产·效率,供应链·采购,管理·流程,财务·应收,质量·生产,管理·团队,运营·库存,管理·项目,其他"', allow_blank=True)
dv_type.error = "请从下拉列表选择"
dv_type.errorTitle = "类型无效"
ws1.add_data_validation(dv_type)
dv_type.add(f"D4:D{3 + len(sample1)}")

dv_urgency = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
ws1.add_data_validation(dv_urgency)
dv_urgency.add(f"E4:E{3 + len(sample1)}")

dv_status = DataValidation(type="list", formula1='"待启动,解决中,已解决,已搁置,未启动"', allow_blank=True)
ws1.add_data_validation(dv_status)
dv_status.add(f"I4:I{3 + len(sample1)}")

dv_tools = DataValidation(type="list", formula1='"问题定义句,5-Why,5W,三好标准,风险预案,边界澄清,根因验证,列现象,关键节点复盘,跟踪检查清单,终局复盘,未使用"', allow_blank=True)
ws1.add_data_validation(dv_tools)
dv_tools.add(f"H4:H{3 + len(sample1)}")

# 写入数据
for row_idx, row_data in enumerate(sample1, 4):
    for col_idx, value in enumerate(row_data, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=value)
        c.font = NORMAL_FONT
        c.alignment = LEFT_TOP
        c.border = ALL_BORDER

    # 紧急度上色
    urgency = row_data[3]
    if urgency == "高":
        ws1.cell(row=row_idx, column=5).fill = WARN_FILL
        ws1.cell(row=row_idx, column=5).font = BOLD_FONT
    elif urgency == "中":
        ws1.cell(row=row_idx, column=5).fill = PatternFill("solid", fgColor="FEF3C7")
    else:
        ws1.cell(row=row_idx, column=5).fill = OK_FILL

    # 状态上色
    status = row_data[7]
    if status == "已解决":
        ws1.cell(row=row_idx, column=8).fill = OK_FILL
        ws1.cell(row=row_idx, column=8).font = Font(name="Microsoft YaHei", size=10, bold=True, color=OK)
    elif status == "解决中":
        ws1.cell(row=row_idx, column=8).fill = PatternFill("solid", fgColor="FEF3C7")
        ws1.cell(row=row_idx, column=8).font = Font(name="Microsoft YaHei", size=10, bold=True, color="B45309")
    elif status == "待启动":
        ws1.cell(row=row_idx, column=8).fill = NEUTRAL_FILL
    elif status == "已搁置":
        ws1.cell(row=row_idx, column=8).fill = WARN_FILL
        ws1.cell(row=row_idx, column=8).font = Font(name="Microsoft YaHei", size=10, color=NEUTRAL)

    ws1.row_dimensions[row_idx].height = 60

# 汇总区
summary_row = 3 + len(sample1) + 2
ws1.merge_cells(f"A{summary_row}:H{summary_row}")
c = ws1.cell(row=summary_row, column=1, value="汇总统计")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws1.row_dimensions[summary_row].height = 24

# 汇总
summary_items = [
    ("问题总数", f"=COUNTA(A4:A{3 + len(sample1)})"),
    ("已解决", f'=COUNTIF(I4:I{3 + len(sample1)},"已解决")'),
    ("解决中", f'=COUNTIF(I4:I{3 + len(sample1)},"解决中")'),
    ("待启动", f'=COUNTIF(I4:I{3 + len(sample1)},"待启动")'),
    ("高紧急度问题", f'=COUNTIF(E4:E{3 + len(sample1)},"高")'),
    ("已解决率", f'=COUNTIF(I4:I{3 + len(sample1)},"已解决")/COUNTA(A4:A{3 + len(sample1)})'),
]

for i, (label, formula) in enumerate(summary_items):
    r = summary_row + 1 + i
    c_label = ws1.cell(row=r, column=1, value=label)
    c_label.font = BOLD_FONT
    c_label.fill = BG_FILL
    c_label.alignment = LEFT
    c_label.border = ALL_BORDER
    ws1.merge_cells(f"A{r}:C{r}")

    c_value = ws1.cell(row=r, column=4, value=formula)
    c_value.font = Font(name="Microsoft YaHei", size=12, bold=True, color=ACCENT)
    c_value.alignment = CENTER
    c_value.border = ALL_BORDER
    if "率" in label:
        c_value.number_format = "0.0%"
    else:
        c_value.number_format = "0"

ws1.freeze_panes = "A4"
ws1.sheet_view.showGridLines = False


# ============================================================
# Sheet 2: 问题解决方案档案
# ============================================================
ws2 = wb.create_sheet("问题解决方案档案")

ws2.merge_cells("A1:I1")
ws2["A1"] = "问题解决方案档案"
ws2["A1"].font = TITLE_FONT
ws2["A1"].fill = TITLE_FILL
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 32

ws2.merge_cells("A2:I2")
ws2["A2"] = "说明：用 1 行记录 1 个完整方案。从定义→分析→方案→执行→复盘，全流程跟踪"
ws2["A2"].font = NOTE_FONT
ws2["A2"].fill = BG_FILL
ws2["A2"].alignment = LEFT
ws2.row_dimensions[2].height = 24

headers2 = [
    ("方案编号", 10),
    ("方案名称", 22),
    ("启动日期", 12),
    ("负责人", 10),
    ("关键路径", 26),
    ("进度(%)", 10),
    ("当前状态", 12),
    ("关键风险", 22),
    ("复盘节点", 14),
]
for col_idx, (text, width) in enumerate(headers2, 1):
    c = ws2.cell(row=3, column=col_idx, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.row_dimensions[3].height = 36

# 示例数据
sample2 = [
    ("P-001", "KPI 修订 + SOP 卡片项目", "2024-06-01", "老周",
     "部门会议通过(6月20日)", 100, "已完成", "无重大风险", "30%/50%/70%/100%"),
    ("P-002", "客户投诉首响提速方案", "2024-07-15", "小王",
     "客服系统改造(8月30日)", 65, "执行中", "技术部排期延迟", "30%/70%"),
    ("P-003", "新员工导师制方案", "2024-09-01", "HR-小张",
     "导师激励兑现(9月15日)", 30, "执行中", "部门经理配合度", "30%/70%"),
    ("P-004", "A 产线 OEE 提升方案", "2024-08-01", "老李",
     "瓶颈工位改造(9月10日)", 75, "执行中", "设备到货延期", "30%/50%/70%"),
    ("P-005", "供应商分级管理方案", "2024-05-15", "采购-小赵",
     "新考核标准上线(6月30日)", 100, "已完成", "无", "30%/50%/70%/100%"),
    ("P-006", "周会决策提速方案", "2024-08-15", "管理层",
     "议程标准化(8月20日)", 50, "执行中", "管理层习惯难改", "30%"),
    ("P-007", "应收账款分阶段催收", "2024-10-01", "财务-小钱",
     "客户分级确认(10月15日)", 25, "执行中", "销售部配合度", "30%"),
    ("P-008", "品保退货改善方案", "2024-07-01", "品保-小孙",
     "工艺改进验证(8月15日)", 80, "执行中", "工艺验证不通过", "30%/50%/70%"),
    ("P-009", "提案制度优化方案", "2024-11-01", "HR-小张",
     "新制度发布(11月15日)", 0, "未启动", "员工积极性难提升", "30%/70%"),
    ("P-010", "备件库存优化方案", "2024-04-01", "设备-小陈",
     "ABC 分类完成(5月1日)", 100, "已完成", "无", "30%/50%/70%/100%"),
    ("P-011", "项目准时交付方案", "2024-09-15", "项目-小李",
     "关键节点预警上线(10月20日)", 40, "执行中", "部门间协调", "30%"),
    ("P-012", "报表自动化方案", "2024-10-15", "财务-小周",
     "系统对接完成(11月30日)", 15, "执行中", "IT 资源紧张", "30%"),
]

# 数据验证
dv_state = DataValidation(type="list", formula1='"未启动,执行中,已完成,已暂停,已取消"', allow_blank=True)
ws2.add_data_validation(dv_state)
dv_state.add(f"G4:G{3 + len(sample2)}")

# 写入数据
for row_idx, row_data in enumerate(sample2, 4):
    for col_idx, value in enumerate(row_data, 1):
        c = ws2.cell(row=row_idx, column=col_idx, value=value)
        c.font = NORMAL_FONT
        c.alignment = LEFT_TOP
        c.border = ALL_BORDER

    # 进度条（条件格式 - 数据条）
    state = row_data[6]
    if state == "已完成":
        ws2.cell(row=row_idx, column=7).fill = OK_FILL
        ws2.cell(row=row_idx, column=7).font = Font(name="Microsoft YaHei", size=10, bold=True, color=OK)
    elif state == "执行中":
        ws2.cell(row=row_idx, column=7).fill = PatternFill("solid", fgColor="FEF3C7")
        ws2.cell(row=row_idx, column=7).font = Font(name="Microsoft YaHei", size=10, bold=True, color="B45309")
    elif state == "未启动":
        ws2.cell(row=row_idx, column=7).fill = NEUTRAL_FILL
    elif state == "已暂停":
        ws2.cell(row=row_idx, column=7).fill = WARN_FILL
    elif state == "已取消":
        ws2.cell(row=row_idx, column=7).fill = NEUTRAL_FILL

    ws2.row_dimensions[row_idx].height = 38

# 进度条件格式（数据条）
progress_range = f"F4:F{3 + len(sample2)}"
from openpyxl.styles.colors import Color
ws2.conditional_formatting.add(
    progress_range,
    openpyxl.formatting.rule.DataBarRule(
        start_type="num", start_value=0,
        end_type="num", end_value=100,
        color=Color(rgb="FF1F3A68"),
        showValue=True
    )
)

# 复盘节点列也加底色提示
for row_idx in range(4, 4 + len(sample2)):
    nodes = ws2.cell(row=row_idx, column=9).value
    if "100%" in nodes:
        ws2.cell(row=row_idx, column=9).fill = OK_FILL
        ws2.cell(row=row_idx, column=9).font = Font(name="Microsoft YaHei", size=10, color=OK, bold=True)
    elif "70%" in nodes:
        ws2.cell(row=row_idx, column=9).fill = PatternFill("solid", fgColor="FEF3C7")
    else:
        ws2.cell(row=row_idx, column=9).fill = NEUTRAL_FILL

# 汇总
summary_row2 = 3 + len(sample2) + 2
ws2.merge_cells(f"A{summary_row2}:I{summary_row2}")
c = ws2.cell(row=summary_row2, column=1, value="方案进度统计")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws2.row_dimensions[summary_row2].height = 24

summary_items2 = [
    ("方案总数", f"=COUNTA(A4:A{3 + len(sample2)})"),
    ("已完成", f'=COUNTIF(G4:G{3 + len(sample2)},"已完成")'),
    ("执行中", f'=COUNTIF(G4:G{3 + len(sample2)},"执行中")'),
    ("未启动", f'=COUNTIF(G4:G{3 + len(sample2)},"未启动")'),
    ("平均完成率", f'=AVERAGE(F4:F{3 + len(sample2)})'),
    ("按时完成率", f'=COUNTIF(G4:G{3 + len(sample2)},"已完成")/COUNTA(A4:A{3 + len(sample2)})'),
]

for i, (label, formula) in enumerate(summary_items2):
    r = summary_row2 + 1 + i
    c_label = ws2.cell(row=r, column=1, value=label)
    c_label.font = BOLD_FONT
    c_label.fill = BG_FILL
    c_label.alignment = LEFT
    c_label.border = ALL_BORDER
    ws2.merge_cells(f"A{r}:C{r}")

    c_value = ws2.cell(row=r, column=4, value=formula)
    c_value.font = Font(name="Microsoft YaHei", size=12, bold=True, color=ACCENT)
    c_value.alignment = CENTER
    c_value.border = ALL_BORDER
    if "率" in label or "完成率" in label:
        c_value.number_format = "0.0%"
    else:
        c_value.number_format = "0.0"

ws2.freeze_panes = "A4"
ws2.sheet_view.showGridLines = False


# ============================================================
# Sheet 3: 工具应用统计
# ============================================================
ws3 = wb.create_sheet("工具应用统计")

ws3.merge_cells("A1:F1")
ws3["A1"] = "工具应用统计"
ws3["A1"].font = TITLE_FONT
ws3["A1"].fill = TITLE_FILL
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 32

ws3.merge_cells("A2:F2")
ws3["A2"] = "说明：记录我每月用了哪些工具、用了多少次、解决了哪些问题。用 1 行记录 1 次使用"
ws3["A2"].font = NOTE_FONT
ws3["A2"].fill = BG_FILL
ws3["A2"].alignment = LEFT
ws3.row_dimensions[2].height = 24

headers3 = [
    ("序号", 6),
    ("使用日期", 12),
    ("所属看板", 14),
    ("使用工具", 16),
    ("应用场景", 28),
    ("效果评估", 14),
]
for col_idx, (text, width) in enumerate(headers3, 1):
    c = ws3.cell(row=3, column=col_idx, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
    ws3.column_dimensions[get_column_letter(col_idx)].width = width
ws3.row_dimensions[3].height = 36

sample3 = [
    (1, "2024-06-05", "定准板", "问题定义句", "设备故障率高", "有效"),
    (2, "2024-06-08", "定准板", "边界澄清", "12 台关键设备 vs 辅助设备", "有效"),
    (3, "2024-06-10", "析透板", "列现象清单", "12 条现象全列", "有效"),
    (4, "2024-06-12", "析透板", "5-Why 追问", "挖到机制层：KPI 错位", "有效"),
    (5, "2024-06-13", "析透板", "根因验证", "用保养记录+访谈验证", "有效"),
    (6, "2024-06-15", "策全板", "三好标准", "评估 4 个方案", "有效"),
    (7, "2024-06-16", "策全板", "5W 行动清单", "拆成 5 个具体动作", "有效"),
    (8, "2024-06-18", "策全板", "风险预案", "识别 3 个主要风险", "有效"),
    (9, "2024-06-20", "控稳板", "跟踪检查清单", "每周一次自检", "有效"),
    (10, "2024-06-21", "控稳板", "关键节点复盘", "30% 时点复盘", "有效"),
    (11, "2024-07-15", "控稳板", "关键节点复盘", "50% 时点复盘+调整", "有效"),
    (12, "2024-07-20", "定准板", "问题定义句", "客户投诉响应慢", "有效"),
    (13, "2024-07-25", "析透板", "5-Why 追问", "客户投诉 - 系统卡顿", "部分有效"),
    (14, "2024-08-01", "策全板", "5W 行动清单", "客服系统改造任务分解", "有效"),
    (15, "2024-08-10", "控稳板", "跟踪检查清单", "周度跟踪改造进度", "有效"),
    (16, "2024-08-15", "定准板", "问题定义句", "新员工 3 月内离职率高", "有效"),
    (17, "2024-08-20", "析透板", "列现象清单", "离职原因 8 条", "有效"),
    (18, "2024-08-25", "析透板", "5-Why 追问", "挖到 HR 流程问题", "有效"),
    (19, "2024-09-01", "策全板", "三好标准", "导师制 vs 集训制", "有效"),
    (20, "2024-09-05", "控稳板", "终局复盘", "KPI 项目完整复盘", "有效"),
    (21, "2024-09-10", "定准板", "边界澄清", "新员工问题边界划清", "有效"),
    (22, "2024-09-15", "策全板", "5W 行动清单", "导师制落地清单", "有效"),
    (23, "2024-09-20", "控稳板", "风险预案", "识别导师配合度风险", "有效"),
    (24, "2024-10-01", "控稳板", "跟踪检查清单", "导师制月度跟踪", "部分有效"),
    (25, "2024-10-10", "定准板", "问题定义句", "A 产线 OEE 偏低", "有效"),
]

# 数据验证
dv_kanban = DataValidation(type="list", formula1='"定准板,析透板,策全板,控稳板"', allow_blank=True)
ws3.add_data_validation(dv_kanban)
dv_kanban.add(f"C4:C{3 + len(sample3)}")

dv_tool3 = DataValidation(type="list", formula1='"问题定义句,边界澄清,列现象清单,5-Why 追问,根因验证,三好标准,5W 行动清单,风险预案,跟踪检查清单,关键节点复盘,终局复盘"', allow_blank=True)
ws3.add_data_validation(dv_tool3)
dv_tool3.add(f"D4:D{3 + len(sample3)}")

dv_effect = DataValidation(type="list", formula1='"非常有效,有效,部分有效,无效,未评估"', allow_blank=True)
ws3.add_data_validation(dv_effect)
dv_effect.add(f"F4:F{3 + len(sample3)}")

# 写入数据
for row_idx, row_data in enumerate(sample3, 4):
    for col_idx, value in enumerate(row_data, 1):
        c = ws3.cell(row=row_idx, column=col_idx, value=value)
        c.font = NORMAL_FONT
        c.alignment = LEFT_TOP
        c.border = ALL_BORDER

    # 看板着色
    kanban = row_data[2]
    kanban_fills = {
        "定准板": "DBEAFE",
        "析透板": "D1FAE5",
        "策全板": "FEF3C7",
        "控稳板": "FCE7F3",
    }
    if kanban in kanban_fills:
        ws3.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor=kanban_fills[kanban])
        ws3.cell(row=row_idx, column=3).font = BOLD_FONT

    # 效果着色
    eff = row_data[5]
    if eff == "非常有效":
        ws3.cell(row=row_idx, column=6).fill = OK_FILL
        ws3.cell(row=row_idx, column=6).font = Font(name="Microsoft YaHei", size=10, bold=True, color=OK)
    elif eff == "有效":
        ws3.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor="D1FAE5")
    elif eff == "部分有效":
        ws3.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor="FEF3C7")
    elif eff == "无效":
        ws3.cell(row=row_idx, column=6).fill = WARN_FILL
    else:
        ws3.cell(row=row_idx, column=6).fill = NEUTRAL_FILL

    ws3.row_dimensions[row_idx].height = 28

# 工具使用频次统计表
stats_row = 3 + len(sample3) + 2
ws3.merge_cells(f"A{stats_row}:F{stats_row}")
c = ws3.cell(row=stats_row, column=1, value="工具使用频次（自动统计）")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws3.row_dimensions[stats_row].height = 24

stats_header = ["看板", "工具", "使用次数", "占比", "有效率", "备注"]
for col_idx, h in enumerate(stats_header, 1):
    c = ws3.cell(row=stats_row + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws3.row_dimensions[stats_row + 1].height = 30

# 工具频次统计（用公式）
tool_list = [
    ("定准板", "问题定义句"),
    ("定准板", "边界澄清"),
    ("析透板", "列现象清单"),
    ("析透板", "5-Why 追问"),
    ("析透板", "根因验证"),
    ("策全板", "三好标准"),
    ("策全板", "5W 行动清单"),
    ("策全板", "风险预案"),
    ("控稳板", "跟踪检查清单"),
    ("控稳板", "关键节点复盘"),
    ("控稳板", "终局复盘"),
]

n_data = len(sample3)
data_start = 4
data_end = 3 + n_data

for i, (kanban, tool) in enumerate(tool_list):
    r = stats_row + 2 + i
    c_k = ws3.cell(row=r, column=1, value=kanban)
    c_k.font = BOLD_FONT
    c_k.fill = H2_FILL
    c_k.alignment = CENTER
    c_k.border = ALL_BORDER

    c_t = ws3.cell(row=r, column=2, value=tool)
    c_t.font = NORMAL_FONT
    c_t.alignment = LEFT
    c_t.border = ALL_BORDER

    # COUNTIFS 公式
    c_count = ws3.cell(row=r, column=3, value=f'=COUNTIFS(C{data_start}:C{data_end},A{r},D{data_start}:D{data_end},B{r})')
    c_count.font = BOLD_FONT
    c_count.alignment = CENTER
    c_count.border = ALL_BORDER
    c_count.fill = OK_FILL
    c_count.number_format = "0"

    c_pct = ws3.cell(row=r, column=4, value=f'=C{r}/COUNTA(D{data_start}:D{data_end})')
    c_pct.font = NORMAL_FONT
    c_pct.alignment = CENTER
    c_pct.border = ALL_BORDER
    c_pct.number_format = "0.0%"

    # 有效率（"非常有效"+"有效" 占该工具使用次数）
    c_eff = ws3.cell(row=r, column=5, value=f'=IF(C{r}=0,0,(COUNTIFS(C{data_start}:C{data_end},A{r},D{data_start}:D{data_end},B{r},F{data_start}:F{data_end},"非常有效")+COUNTIFS(C{data_start}:C{data_end},A{r},D{data_start}:D{data_end},B{r},F{data_start}:F{data_end},"有效"))/C{r})')
    c_eff.font = NORMAL_FONT
    c_eff.alignment = CENTER
    c_eff.border = ALL_BORDER
    c_eff.number_format = "0.0%"

    c_note = ws3.cell(row=r, column=6, value="")
    c_note.font = SOFT_FONT
    c_note.alignment = LEFT
    c_note.border = ALL_BORDER

    ws3.row_dimensions[r].height = 24

# 配色给占比行加数据条
data_bar_range = f"D{stats_row + 2}:D{stats_row + 1 + len(tool_list)}"
ws3.conditional_formatting.add(
    data_bar_range,
    openpyxl.formatting.rule.DataBarRule(
        start_type="num", start_value=0,
        end_type="num", end_value=0.3,
        color=Color(rgb="FF1F3A68"),
        showValue=True
    )
)

# 添加图表：工具使用频次柱状图
chart = BarChart()
chart.type = "bar"
chart.style = 11
chart.title = "11 个工具使用频次"
chart.x_axis.title = "使用次数"
chart.y_axis.title = "工具"

data_ref = Reference(ws3, min_col=3, min_row=stats_row + 1, max_row=stats_row + 1 + len(tool_list), max_col=3)
cats_ref = Reference(ws3, min_col=2, min_row=stats_row + 2, max_row=stats_row + 1 + len(tool_list))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 10
chart.width = 18
# 设置图表中数据系列的填充色（使用纯色，无渐变）
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice, PatternFillProperties, ColorChoice
chart.legend = None
ws3.add_chart(chart, f"H{stats_row}")

# 月度使用总览（右侧小表）
monthly_row = stats_row
ws3.merge_cells(f"H{monthly_row}:L{monthly_row}")
c = ws3.cell(row=monthly_row, column=8, value="月度使用总览")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws3.row_dimensions[monthly_row].height = 24

monthly_header = ["月份", "使用次数", "主要工具", "新增场景", "解决率"]
for col_idx, h in enumerate(monthly_header, 8):
    c = ws3.cell(row=monthly_row + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws3.row_dimensions[monthly_row + 1].height = 30

monthly_data = [
    ("2024-06", 9, "全流程", 2, 0.78),
    ("2024-07", 4, "控稳+定准", 1, 0.75),
    ("2024-08", 5, "析透+策全", 2, 0.80),
    ("2024-09", 5, "策全+控稳", 1, 0.82),
    ("2024-10", 2, "定准+控稳", 1, 0.50),
]

for i, (month, cnt, main, new, rate) in enumerate(monthly_data):
    r = monthly_row + 2 + i
    ws3.cell(row=r, column=8, value=month).font = BOLD_FONT
    ws3.cell(row=r, column=9, value=cnt).font = NORMAL_FONT
    ws3.cell(row=r, column=9).number_format = "0"
    ws3.cell(row=r, column=10, value=main).font = NORMAL_FONT
    ws3.cell(row=r, column=11, value=new).font = NORMAL_FONT
    ws3.cell(row=r, column=11).number_format = "0"
    ws3.cell(row=r, column=12, value=rate).font = NORMAL_FONT
    ws3.cell(row=r, column=12).number_format = "0.0%"

    for col in range(8, 13):
        c = ws3.cell(row=r, column=col)
        c.alignment = CENTER
        c.border = ALL_BORDER
        if col == 9:
            c.fill = OK_FILL
            c.font = BOLD_FONT
    ws3.row_dimensions[r].height = 24

# 列宽
for col, w in zip(["H", "I", "J", "K", "L"], [12, 12, 18, 12, 12]):
    ws3.column_dimensions[col].width = w

ws3.freeze_panes = "A4"
ws3.sheet_view.showGridLines = False


# ============================================================
# 保存
# ============================================================
output_path = r"D:\2026年课程\竞越\基层即学即用的问题解决工具箱\补充课程包\16_配套表单Excel\16_个人问题跟踪表_学员用.xlsx"

wb.save(output_path)
print(f"Generated: {output_path}")
print(f"包含 3 个 sheet: 我的工作问题库 / 问题解决方案档案 / 工具应用统计")
print(f"数据行数: {len(sample1)} + {len(sample2)} + {len(sample3)} = {len(sample1) + len(sample2) + len(sample3)} 行")

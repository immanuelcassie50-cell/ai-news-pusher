# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "90分钟互动时间表"

# Styles
header_font = Font(name='Microsoft YaHei', bold=True, size=10, color="FFFFFF")
cell_font = Font(name='Microsoft YaHei', size=9)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
light_blue_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["时间段", "环节", "时长", "讲师动作", "学员动作", "互动形式", "物料准备", "备注"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Data - 90 minute timeline (using English for Chinese content to avoid encoding issues)
data = [
    ["0:00-0:05", "开场", "5min", "宣布目标：90分钟后你能设计一套师徒制优化方案", "写下1个最想解决的问题", "个人书写", "白板、彩笔", "直接给目标，不讲废话"],
    ["0:05-0:20", "传承之困", "15min", "讲：师徒制面临的4大挑战（经验难显性化、新人成长慢、知识断层、AI替代焦虑）", "听+思考", "讲解", "PPT课件", "用数据开场：老师傅平均年龄52岁"],
    ["0:20-0:30", "诊断互动", "10min", "发诊断卡、巡视各组", "2人组讨论：填写诊断卡（4个挑战各打1-5分）", "两人讨论", "诊断卡12份", "收集真实数据"],
    ["0:30-0:45", "断档之因", "15min", "讲：张师傅退休案例 + 知识断层的3个根因", "听+反思", "讲解", "PPT课件", "案例要具体到听音辨故障这种绝活"],
    ["0:45-0:55", "案例讨论", "10min", "抛出聚焦问题：你的车间，谁的绝活最可能失传？", "2人组讨论5min，1-2人分享", "两人讨论", "无", "留白30秒，逼学员开口"],
    ["0:55-1:05", "重建之路-概念", "10min", "讲：AI帮什么 + 人做什么（4个明确边界）", "听+记笔记", "讲解", "PPT课件", "金句：AI帮不了手感、帮不了直觉"],
    ["1:05-1:15", "重建之路-设计", "10min", "引导：你的车间，师徒制优化的3个着力点", "3人组设计：每个组设计1套优化方案", "小组讨论", "大白纸、彩笔", "每组只设计1个核心改变"],
    ["1:15-1:25", "方案分享", "10min", "邀请2-3组分享、点评", "各组派代表分享方案", "全班分享", "无", "点评聚焦在可执行性上"],
    ["1:25-1:30", "收尾", "5min", "总结：3件事 + 1句承诺", "每人写下：回去1周内要做的1件事", "个人书写", "承诺卡12份", "可测量、可检查"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if row_idx % 2 == 0:
            cell.fill = light_blue_fill

# Column widths
col_widths = [10, 12, 8, 35, 25, 12, 15, 20]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Summary row
ws.cell(row=12, column=1, value="时间分配统计").font = Font(name='Microsoft YaHei', bold=True, size=10)
ws.cell(row=13, column=1, value="讲解：35min | 互动：45min | 收尾：5min | 学员做占比：50%")
ws.cell(row=14, column=1, value="互动形式统计").font = Font(name='Microsoft YaHei', bold=True, size=10)
ws.cell(row=15, column=1, value="讲解x2 | 个人书写x2 | 两人讨论x2 | 小组讨论x1 | 全班分享x2")

output_path = "D:/新课开发/企业大学/对内/11.文化基建：企业大学AI化推进中不能松手的师徒传承与组织记忆/成果demo/07_Interaction_Flow.xlsx"
wb.save(output_path)
print("07_Interaction_Flow.xlsx created successfully at", output_path)

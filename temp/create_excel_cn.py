# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Microsoft YaHei', size=10)
cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

priority_colors = {'P1': PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid'), 'P2': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'), 'P3': PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid')}
yes_no_colors = {'是': PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid'), '否': PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid'), '待定': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')}
risk_colors = {'高': PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid'), '中': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'), '低': PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid')}

output_dir = r'D:\新课开发\工作手册\岗位微流程拆解与标准化\完整课程包\07_工具表单'
os.makedirs(output_dir, exist_ok=True)

# 1. 微流程地图.xlsx
wb1 = Workbook()
ws1 = wb1.active
ws1.title = '微流程地图'
headers1 = ['岗位名称', '场景/情境', '微流程名称', '触发条件', '现状描述', '出错概率', '出错后果', '是否需要标准化', '优先级', '负责拆解人', '截止日期', '备注']
for col, h in enumerate(headers1, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = header_alignment; c.border = thin_border
widths1 = [15, 20, 20, 25, 30, 12, 25, 18, 10, 15, 12, 20]
for col, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.row_dimensions[1].height = 30
sample1 = [
    ['客服', '客户抗诉处理', '抗诉安抚流程', '客户表达不滥时', '缺乏统一话术，效果参差不齐', '高', '客户流失、品牌提质', '是', 'P1', '张三', '2026-07-20', '优先制定标准话术'],
    ['销售', '产品介绍', '需求挖掘流程', '客户询问产品时', '推荐盲目，缺乏针对性', '中', '转化率低、丢单', '是', 'P2', '李四', '2026-07-25', ''],
    ['运营', '活动策划', '活动审批流程', '活动方案完成后', '审批流程不清晰，效率低', '中', '活动延期、资源浪费', '待定', 'P3', '王五', '2026-08-01', '需与相关部门确认'],
]
for ri, data in enumerate(sample1, 2):
    for ci, v in enumerate(data, 1):
        c = ws1.cell(row=ri, column=ci, value=v)
        c.font = cell_font; c.border = thin_border; c.alignment = cell_alignment
        if ci == 6 and v in risk_colors: c.fill = risk_colors[v]
        if ci == 8 and v in yes_no_colors: c.fill = yes_no_colors[v]
        if ci == 9 and v in priority_colors: c.fill = priority_colors[v]
ws1.freeze_panes = 'A2'
dv1 = DataValidation(type='list', formula1='"高,中,位"', allow_blank=True); ws1.add_data_validation(dv1); dv1.add('F2:F1000')
dv2 = DataValidation(type='list', formula1='"是,否,待定"', allow_blank=True); ws1.add_data_validation(dv2); dv2.add('H2:H1000')
dv3 = DataValidation(type='list', formula1='"P1,P2,P3"', allow_blank=True); ws1.add_data_validation(dv3); dv3.add('I2:I1000')
wb1.save(output_dir + r'\微流程地图.xlsx')
print('OK: 微流程地图.xlsx')

# 2. 标准作业卡.xlsx
wb2 = Workbook()
ws2 = wb2.active
ws2.title = '标准作业卡'
headers2 = ['微流程名称', '适用岗位', '触发时机', '步骤1', '步骤2', '步骤3', '步骤4', '步骤5', '步骤6', '步骤7', '步骤8', '步骤9', '步骤10', '每步时间参考', '关键确认点', '标准话术', '负人人确认方式', '常见错误提示', '版本号', '更新日期']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = header_alignment; c.border = thin_border
widths2 = [20, 12, 20, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 15, 25, 30, 18, 25, 10, 12]
for col, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.row_dimensions[1].height = 35
sample2 = [['客户抗诉处理流程', '客服', '客户表达不滥时', '主动问候并道歉', '认真聊听客户请求', '记录问题要点', '确认理解是否正确', '提供解决方案', '询问客户源足度', '记录处理结果', '发送源足度调查', '升级处理（如需）', '', '3-5分钟/步', '客户情绪稳定、问题理解准确、方案客户认可', '非常抱歉给您带来不便，我会立即为您处理这个问题。', '签名+工号', '未道歉、抢话、承诺未记录', 'V1.0', '2026-07-15']]
for ri, data in enumerate(sample2, 2):
    for ci, v in enumerate(data, 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.font = cell_font; c.border = thin_border; c.alignment = cell_alignment
ws2.freeze_panes = 'D2'
dv4 = DataValidation(type='list', formula1='"V1.0,V1.1,V2.0"', allow_blank=True); ws2.add_data_validation(dv4); dv4.add('S2:S1000')
wb2.save(output_dir + r'\标准作业卡.xlsx')
print('OK: 标准作业卡.xlsx')

# 3. 日常稽核表.xlsx
wb3 = Workbook()
ws3 = wb3.active
ws3.title = '日常稿核表'
headers3 = ['岗位', '被稿核人', '稿核日期', '稿核项目', '漏（次数）', '错序（次数）', '走样（次数）', '总分', '稿核人', '改进建议']
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = header_alignment; c.border = thin_border
widths3 = [15, 15, 12, 25, 12, 12, 12, 10, 12, 30]
for col, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.row_dimensions[1].height = 30
sample3 = [
    ['客服', '张三', '2026-07-15', '客户抗诉处理', 0, 1, 0, 95, '李主管', '加强错序环节的练习'],
    ['销售', '李四', '2026-07-15', '产品介绍流程', 1, 0, 2, 80, '王经理', '增加演练次数'],
    ['运营', '王五', '2026-07-14', '活动策划流程', 0, 0, 0, 100, '张总', '\u4e继续保持'],
]
for ri, data in enumerate(sample3, 2):
    for ci, v in enumerate(data, 1):
        c = ws3.cell(row=ri, column=ci, value=v)
        c.font = cell_font; c.border = thin_border; c.alignment = cell_alignment
        if ci == 8 and isinstance(v, (int, float)):
            if v < 85: c.fill = PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid')
            elif v >= 95: c.fill = PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid')
        if ci in [5, 6, 7] and isinstance(v, (int, float)) and v > 0:
            c.fill = PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid')
ws3.freeze_panes = 'A2'
dv5 = DataValidation(type='whole', operator='between', formula1='0', formula2='100')
dv5.error = '请输入0-100之间的数字'; dv5.errorTitle = '分数无效

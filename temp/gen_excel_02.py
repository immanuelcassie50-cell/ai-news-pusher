# -*- coding: utf-8 -*-
"""
生成 16_团队问题看板_小组用.xlsx
3 sheets:
  1. 问题清单 (全团队问题总览)
  2. 进度跟踪 (周度/月度进展)
  3. 复盘记录 (关键节点复盘)
适用于 5-8 人小组
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.styles.colors import Color
from datetime import datetime

# 主题色：深红（团队·协作·警示）
ACCENT = "991B1B"
ACCENT_SOFT = "FCE8E8"
ACCENT_DARK = "7F1D1D"
OK = "166534"
OK_SOFT = "E8F1EB"
WARN = "C2410C"
WARN_SOFT = "FEF0E8"
NEUTRAL = "6B7280"
NEUTRAL_SOFT = "F3F4F6"
LINE = "D8B8B8"
LINE_SOFT = "F0D8D8"
BG = "FAF0F0"
PAPER = "FFFFFF"
INK = "1F1A1A"
INK_SOFT = "4A3A3A"

TITLE_FONT = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
H2_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color=ACCENT_DARK)
HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Microsoft YaHei", size=10, color=INK)
SOFT_FONT = Font(name="Microsoft YaHei", size=10, color=INK_SOFT)
BOLD_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color=INK)
NOTE_FONT = Font(name="Microsoft YaHei", size=9, italic=True, color=NEUTRAL)

def PF(color):
    return PatternFill("solid", fgColor=color)

TITLE_FILL = PF(ACCENT)
H2_FILL = PF(ACCENT_SOFT)
HEADER_FILL = PF(ACCENT)
OK_FILL = PF(OK_SOFT)
WARN_FILL = PF(WARN_SOFT)
BG_FILL = PF(BG)
NEUTRAL_FILL = PF(NEUTRAL_SOFT)

thin = Side(border_style="thin", color=LINE)
medium = Side(border_style="medium", color=ACCENT)
ALL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()
wb.remove(wb.active)


# ============================================================
# Sheet 1: 问题清单
# ============================================================
ws1 = wb.create_sheet("问题清单")

ws1.merge_cells("A1:I1")
ws1["A1"] = "团队问题清单"
ws1["A1"].font = TITLE_FONT
ws1["A1"].fill = TITLE_FILL
ws1["A1"].alignment = CENTER
ws1.row_dimensions[1].height = 32

ws1.merge_cells("A2:I2")
ws1["A2"] = "说明：本表记录团队 5-8 人共同关注的问题。1 行 1 个问题。建议月初同步，月中更新，月底汇总"
ws1["A2"].font = NOTE_FONT
ws1["A2"].fill = BG_FILL
ws1["A2"].alignment = LEFT
ws1.row_dimensions[2].height = 24

headers1 = [
    ("序号", 6),
    ("问题名称", 22),
    ("提出人", 10),
    ("所属模块", 14),
    ("紧急度", 10),
    ("问题定义句", 50),
    ("是否本月重点", 14),
    ("解决状态", 12),
    ("认领人", 10),
]
for col_idx, (text, width) in enumerate(headers1, 1):
    c = ws1.cell(row=3, column=col_idx, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
    ws1.column_dimensions[get_column_letter(col_idx)].width = width
ws1.row_dimensions[3].height = 36

# 示例数据 - 8 个团队成员的 12 个问题
sample1 = [
    (1, "新客户开发数下降", "小王", "销售", "高",
     "在销售部，过去 3 个月里，新客户开发数方面，现状是月均 8 家，标准是公司目标 15 家/月，差距是 7 家/月，影响是销售线索枯竭 1-2 个月后业绩会下滑。",
     "是", "解决中", "小王"),
    (2, "客户投诉响应慢", "小李", "客服", "中",
     "在客服部，过去 1 个月里，投诉首响时长方面，现状是 4.5 小时，标准是公司规定 2 小时，差距是 2.5 小时，影响是 3 个客户已提出解约。",
     "是", "解决中", "小李"),
    (3, "员工提案参与率低", "小张", "人力", "低",
     "在过去半年里，提案参与率方面，现状是 12%，标准是 30%，差距是 18 个百分点，影响是基层改善想法未充分利用。",
     "否", "待启动", "未认领"),
    (4, "设备故障率高", "小陈", "设备", "高",
     "在设备部，过去 6 个月里，关键设备故障率方面，现状是 5.2%，标准是 3.8% 以下，差距是 1.4 个百分点，影响是月停机损失 32 万。",
     "是", "已解决", "小陈"),
    (5, "供应商交付不准时", "小赵", "采购", "中",
     "在采购部，过去 6 个月里，关键供应商准时交付率方面，现状是 73%，标准是 95%，差距是 22 个百分点，影响是生产线停工待料 5 次/月。",
     "是", "已解决", "小赵"),
    (6, "会议效率低", "小周", "管理", "低",
     "在过去 1 个月里，周会决策时长方面，现状是 1.8 小时/会，标准是 1 小时内，差距是 0.8 小时/会，影响是每周浪费 25 人时。",
     "否", "已搁置", "小周"),
    (7, "报表错误率高", "小孙", "财务", "中",
     "在财务部，过去 6 个月里，月度报表一次通过率方面，现状是 62%，标准是 95% 以上，差距是 33 个百分点，影响是每月多花 2-3 人/周返工。",
     "是", "解决中", "小孙"),
    (8, "OEE 偏低", "小钱", "生产", "高",
     "在 A 产线，过去 3 个月里，OEE 方面，现状是 68%，标准是 80% 以上，差距是 12 个百分点，影响是月产值损失 80 万。",
     "是", "执行中", "小钱"),
    (9, "应收账款回款慢", "小吴", "财务", "中",
     "在销售部，过去 6 个月里，应收账款周转天数方面，现状是 78 天，标准是 60 天，差距是 18 天，影响是流动资金压力增加 200 万。",
     "否", "待启动", "未认领"),
    (10, "新员工离职率高", "小郑", "人力", "中",
     "在过去 1 年里，新员工 3 个月内离职率方面，现状是 28%，标准是 15% 以下，差距是 13 个百分点，影响是招聘成本增加 35 万/年。",
     "是", "执行中", "小张"),
    (11, "退货率高", "小王", "品保", "高",
     "在品保部，过去 3 个月里，退货率方面，现状是 4.8%，标准是 2% 以下，差距是 2.8 个百分点，影响是退货损失 18 万/月。",
     "是", "执行中", "小孙"),
    (12, "项目延期", "小李", "项目", "中",
     "在项目部，过去半年里，项目准时交付率方面，现状是 58%，标准是 85%，差距是 27 个百分点，影响是客户罚款 15 万 + 续约率下降。",
     "否", "待启动", "未认领"),
]

# 数据验证
dv_module = DataValidation(type="list", formula1='"销售,客服,人力,设备,采购,管理,财务,生产,品保,项目,其他"', allow_blank=True)
ws1.add_data_validation(dv_module)
dv_module.add(f"D4:D{3 + len(sample1)}")

dv_urgency = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
ws1.add_data_validation(dv_urgency)
dv_urgency.add(f"E4:E{3 + len(sample1)}")

dv_focus = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
ws1.add_data_validation(dv_focus)
dv_focus.add(f"G4:G{3 + len(sample1)}")

dv_status = DataValidation(type="list", formula1='"待启动,解决中,执行中,已解决,已搁置,已取消"', allow_blank=True)
ws1.add_data_validation(dv_status)
dv_status.add(f"H4:H{3 + len(sample1)}")

for row_idx, row_data in enumerate(sample1, 4):
    for col_idx, value in enumerate(row_data, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=value)
        c.font = NORMAL_FONT
        c.alignment = LEFT_TOP
        c.border = ALL_BORDER

    # 紧急度上色
    urgency = row_data[4]
    if urgency == "高":
        ws1.cell(row=row_idx, column=5).fill = PF("FECACA")
        ws1.cell(row=row_idx, column=5).font = Font(name="Microsoft YaHei", size=10, bold=True, color="991B1B")
    elif urgency == "中":
        ws1.cell(row=row_idx, column=5).fill = PF("FEF3C7")
    else:
        ws1.cell(row=row_idx, column=5).fill = OK_FILL

    # 本月重点上色
    focus = row_data[6]
    if focus == "是":
        ws1.cell(row=row_idx, column=7).fill = ACCENT_FILL = PF(ACCENT)
        ws1.cell(row=row_idx, column=7).font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    else:
        ws1.cell(row=row_idx, column=7).fill = NEUTRAL_FILL

    # 状态上色
    status = row_data[7]
    if status == "已解决":
        ws1.cell(row=row_idx, column=8).fill = OK_FILL
        ws1.cell(row=row_idx, column=8).font = Font(name="Microsoft YaHei", size=10, bold=True, color=OK)
    elif status == "执行中" or status == "解决中":
        ws1.cell(row=row_idx, column=8).fill = PF("FEF3C7")
        ws1.cell(row=row_idx, column=8).font = Font(name="Microsoft YaHei", size=10, bold=True, color="B45309")
    elif status == "待启动":
        ws1.cell(row=row_idx, column=8).fill = NEUTRAL_FILL
    elif status == "已搁置":
        ws1.cell(row=row_idx, column=8).fill = WARN_FILL
        ws1.cell(row=row_idx, column=8).font = Font(name="Microsoft YaHei", size=10, color=NEUTRAL)
    else:
        ws1.cell(row=row_idx, column=8).fill = NEUTRAL_FILL

    ws1.row_dimensions[row_idx].height = 60

# 汇总区
summary_row = 3 + len(sample1) + 2
ws1.merge_cells(f"A{summary_row}:I{summary_row}")
c = ws1.cell(row=summary_row, column=1, value="本月问题分布统计")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws1.row_dimensions[summary_row].height = 24

# 按模块分组
n = len(sample1)
stats_header = ["模块", "问题数", "高紧急", "本月重点", "已解决", "解决率"]
for col_idx, h in enumerate(stats_header, 1):
    c = ws1.cell(row=summary_row + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws1.row_dimensions[summary_row + 1].height = 30

modules = ["销售", "客服", "人力", "设备", "采购", "管理", "财务", "生产", "品保", "项目"]
for i, m in enumerate(modules):
    r = summary_row + 2 + i
    ws1.cell(row=r, column=1, value=m).font = BOLD_FONT
    ws1.cell(row=r, column=1).fill = BG_FILL
    ws1.cell(row=r, column=1).alignment = CENTER
    ws1.cell(row=r, column=1).border = ALL_BORDER

    # COUNTIF 公式
    ws1.cell(row=r, column=2, value=f'=COUNTIF(D4:D{3 + n},A{r})').number_format = "0"
    ws1.cell(row=r, column=3, value=f'=COUNTIFS(D4:D{3 + n},A{r},E4:E{3 + n},"高")').number_format = "0"
    ws1.cell(row=r, column=4, value=f'=COUNTIFS(D4:D{3 + n},A{r},G4:G{3 + n},"是")').number_format = "0"
    ws1.cell(row=r, column=5, value=f'=COUNTIFS(D4:D{3 + n},A{r},H4:H{3 + n},"已解决")').number_format = "0"
    ws1.cell(row=r, column=6, value=f'=IF(B{r}=0,0,E{r}/B{r})').number_format = "0.0%"

    for col in range(2, 7):
        c = ws1.cell(row=r, column=col)
        c.font = NORMAL_FONT
        c.alignment = CENTER
        c.border = ALL_BORDER
        if col == 6:
            c.fill = OK_FILL
    ws1.row_dimensions[r].height = 22

# 合计
total_r = summary_row + 2 + len(modules)
ws1.cell(row=total_r, column=1, value="合计").font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
ws1.cell(row=total_r, column=1).fill = HEADER_FILL
ws1.cell(row=total_r, column=1).alignment = CENTER
ws1.cell(row=total_r, column=1).border = ALL_BORDER

ws1.cell(row=total_r, column=2, value=f'=SUM(B{summary_row + 2}:B{total_r - 1})').number_format = "0"
ws1.cell(row=total_r, column=3, value=f'=SUM(C{summary_row + 2}:C{total_r - 1})').number_format = "0"
ws1.cell(row=total_r, column=4, value=f'=SUM(D{summary_row + 2}:D{total_r - 1})').number_format = "0"
ws1.cell(row=total_r, column=5, value=f'=SUM(E{summary_row + 2}:E{total_r - 1})').number_format = "0"
ws1.cell(row=total_r, column=6, value=f'=E{total_r}/B{total_r}').number_format = "0.0%"

for col in range(2, 7):
    c = ws1.cell(row=total_r, column=col)
    c.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws1.row_dimensions[total_r].height = 24

ws1.freeze_panes = "A4"
ws1.sheet_view.showGridLines = False


# ============================================================
# Sheet 2: 进度跟踪
# ============================================================
ws2 = wb.create_sheet("进度跟踪")

ws2.merge_cells("A1:H1")
ws2["A1"] = "团队方案进度跟踪"
ws2["A1"].font = TITLE_FONT
ws2["A1"].fill = TITLE_FILL
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 32

ws2.merge_cells("A2:H2")
ws2["A2"] = "说明：用 1 行记录 1 个方案的进度。建议每周更新一次。每周五下班前完成"
ws2["A2"].font = NOTE_FONT
ws2["A2"].fill = BG_FILL
ws2["A2"].alignment = LEFT
ws2.row_dimensions[2].height = 24

headers2 = [
    ("方案编号", 10),
    ("方案名称", 22),
    ("负责人", 10),
    ("启动日期", 12),
    ("本周进度(%)", 12),
    ("累计完成度", 12),
    ("本周动作", 36),
    ("本周风险", 22),
]
for col_idx, (text, width) in enumerate(headers2, 1):
    c = ws2.cell(row=3, column=col_idx, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.row_dimensions[3].height = 36

# 示例 - 团队 6 个主要方案 + 多周记录
# 这里用每个方案多周一行，体现"周度跟踪"
sample2 = [
    # 方案 1: 设备故障改善 (4 周)
    ("P-001", "KPI 修订 + SOP 卡片", "小陈", "2024-06-01", 25, 25, "完成 KPI 草案初稿", "无"),
    ("P-001", "KPI 修订 + SOP 卡片", "小陈", "2024-06-08", 20, 45, "和老张 1 对 1 沟通通过", "风险 1 启动"),
    ("P-001", "KPI 修订 + SOP 卡片", "小陈", "2024-06-15", 30, 75, "部门会议表决通过", "无"),
    ("P-001", "KPI 修订 + SOP 卡片", "小陈", "2024-06-22", 25, 100, "SOP 卡片完成+培训", "无"),
    # 方案 2: 客户投诉改善 (3 周)
    ("P-002", "客服系统改造", "小李", "2024-07-15", 15, 15, "需求收集", "技术部排期"),
    ("P-002", "客服系统改造", "小李", "2024-07-22", 20, 35, "需求评审通过", "无"),
    ("P-002", "客服系统改造", "小李", "2024-07-29", 30, 65, "开发完成 60%", "测试资源紧张"),
    # 方案 3: 新员工导师制 (2 周)
    ("P-003", "新员工导师制", "小张", "2024-09-01", 15, 15, "导师名单确认", "部分经理反对"),
    ("P-003", "新员工导师制", "小张", "2024-09-08", 15, 30, "导师激励方案沟通", "激励金额待批"),
    # 方案 4: OEE 提升 (4 周)
    ("P-004", "A 产线 OEE 提升", "小钱", "2024-08-01", 20, 20, "瓶颈工位诊断完成", "设备到货延期"),
    ("P-004", "A 产线 OEE 提升", "小钱", "2024-08-08", 15, 35, "工位改造方案确定", "无"),
    ("P-004", "A 产线 OEE 提升", "小钱", "2024-08-15", 20, 55, "改造施工中", "工期紧张"),
    ("P-004", "A 产线 OEE 提升", "小钱", "2024-08-22", 20, 75, "改造完成+调试", "无"),
    # 方案 5: 供应商分级 (2 周)
    ("P-005", "供应商分级管理", "小赵", "2024-05-15", 50, 50, "完成 ABC 分类", "无"),
    ("P-005", "供应商分级管理", "小赵", "2024-05-22", 50, 100, "新考核标准上线", "无"),
    # 方案 6: 周会决策提速 (1 周)
    ("P-006", "周会决策提速", "小周", "2024-08-15", 50, 50, "议程模板发布", "管理层习惯难改"),
]

# 写入数据
n2 = len(sample2)
for row_idx, row_data in enumerate(sample2, 4):
    for col_idx, value in enumerate(row_data, 1):
        c = ws2.cell(row=row_idx, column=col_idx, value=value)
        c.font = NORMAL_FONT
        c.alignment = LEFT_TOP
        c.border = ALL_BORDER

    # 风险高亮
    risk = row_data[7]
    if risk != "无" and risk:
        ws2.cell(row=row_idx, column=8).fill = WARN_FILL
        ws2.cell(row=row_idx, column=8).font = Font(name="Microsoft YaHei", size=10, color=WARN, bold=True)
    else:
        ws2.cell(row=row_idx, column=8).fill = OK_FILL
        ws2.cell(row=row_idx, column=8).font = Font(name="Microsoft YaHei", size=10, color=OK)

    # 累计完成度条件格式 - 数据条
    ws2.row_dimensions[row_idx].height = 28

# 累计完成度数据条
ws2.conditional_formatting.add(
    f"F4:F{3 + n2}",
    DataBarRule(
        start_type="num", start_value=0,
        end_type="num", end_value=100,
        color=Color(rgb="FF991B1B"),
        showValue=True
    )
)

# 本周进度数据条
ws2.conditional_formatting.add(
    f"E4:E{3 + n2}",
    DataBarRule(
        start_type="num", start_value=0,
        end_type="num", end_value=50,
        color=Color(rgb="FFFEF3C7"),
        showValue=True
    )
)

# 方案总览（按方案汇总）
sum_row = 3 + n2 + 2
ws2.merge_cells(f"A{sum_row}:H{sum_row}")
c = ws2.cell(row=sum_row, column=1, value="方案总览（按方案汇总 - 用公式自动统计）")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws2.row_dimensions[sum_row].height = 24

ov_headers = ["方案编号", "方案名称", "负责人", "起始周", "完成周数", "最终完成度", "状态", "评级"]
for col_idx, h in enumerate(ov_headers, 1):
    c = ws2.cell(row=sum_row + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws2.row_dimensions[sum_row + 1].height = 30

# 唯一方案
unique_p = []
seen = set()
for r in sample2:
    if r[0] not in seen:
        seen.add(r[0])
        unique_p.append(r)

for i, p in enumerate(unique_p):
    r = sum_row + 2 + i
    pid, pname, leader, _, _, _, _, _ = p
    ws2.cell(row=r, column=1, value=pid).font = BOLD_FONT
    ws2.cell(row=r, column=2, value=pname).font = NORMAL_FONT
    ws2.cell(row=r, column=3, value=leader).font = NORMAL_FONT

    # 起始周 = MIN 启动日期对应行号
    ws2.cell(row=r, column=4, value=f'=INDEX(D4:D{3 + n2},MATCH(A{r},A4:A{3 + n2},0))')
    ws2.cell(row=r, column=5, value=f'=COUNTIF(A4:A{3 + n2},A{r})').number_format = "0"
    # 最终完成度 = 最后一个记录
    ws2.cell(row=r, column=6, value=f'=LOOKUP(2,1/(A4:A{3 + n2}=A{r}),F4:F{3 + n2})').number_format = "0"
    # 状态
    ws2.cell(row=r, column=7, value=f'=IF(F{r}>=100,"已完成",IF(F{r}>=60,"执行中",IF(F{r}>=30,"执行中","执行中")))')
    # 评级
    ws2.cell(row=r, column=8, value=f'=IF(F{r}>=100,"达标",IF(F{r}>=80,"基本达标",IF(F{r}>=50,"进行中","需关注")))')

    for col in range(1, 9):
        c = ws2.cell(row=r, column=col)
        c.alignment = CENTER
        c.border = ALL_BORDER
        if col not in [2]:
            c.font = BOLD_FONT
        if col == 6:
            c.fill = OK_FILL
    ws2.row_dimensions[r].height = 24

# 给评级列条件格式
ov_data_range = f"H{sum_row + 2}:H{sum_row + 1 + len(unique_p)}"
ws2.conditional_formatting.add(
    ov_data_range,
    CellIsRule(operator="equal", formula=['"达标"'], fill=OK_FILL, font=Font(name="Microsoft YaHei", size=10, bold=True, color=OK))
)
ws2.conditional_formatting.add(
    ov_data_range,
    CellIsRule(operator="equal", formula=['"需关注"'], fill=WARN_FILL, font=Font(name="Microsoft YaHei", size=10, bold=True, color=WARN))
)

# 添加柱状图 - 各方案最终完成度
chart = BarChart()
chart.type = "col"
chart.style = 11
chart.title = "各方案最终完成度"
chart.x_axis.title = "方案"
chart.y_axis.title = "完成度"

data_ref = Reference(ws2, min_col=6, min_row=sum_row + 1, max_row=sum_row + 1 + len(unique_p), max_col=6)
cats_ref = Reference(ws2, min_col=2, min_row=sum_row + 2, max_row=sum_row + 1 + len(unique_p))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 9
chart.width = 16
chart.legend = None
ws2.add_chart(chart, f"J{sum_row}")

ws2.freeze_panes = "A4"
ws2.sheet_view.showGridLines = False


# ============================================================
# Sheet 3: 复盘记录
# ============================================================
ws3 = wb.create_sheet("复盘记录")

ws3.merge_cells("A1:G1")
ws3["A1"] = "团队复盘记录"
ws3["A1"].font = TITLE_FONT
ws3["A1"].fill = TITLE_FILL
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 32

ws3.merge_cells("A2:G2")
ws3["A2"] = "说明：用 1 行记录 1 次复盘（30%/50%/70%/100% 节点）。复盘 = 看原计划 vs 实际 + 找差异 + 做调整"
ws3["A2"].font = NOTE_FONT
ws3["A2"].fill = BG_FILL
ws3["A2"].alignment = LEFT
ws3.row_dimensions[2].height = 24

headers3 = [
    ("复盘日期", 12),
    ("方案编号", 10),
    ("方案名称", 22),
    ("节点", 10),
    ("完成度", 10),
    ("复盘要点（差异 + 调整）", 50),
    ("参与人", 16),
]
for col_idx, (text, width) in enumerate(headers3, 1):
    c = ws3.cell(row=3, column=col_idx, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
    ws3.column_dimensions[get_column_letter(col_idx)].width = width
ws3.row_dimensions[3].height = 36

sample3 = [
    ("2024-06-21", "P-001", "KPI 修订 + SOP 卡片", "30%", 30,
     "原计划 6/15 完成 KPI 草案。实际 6/13 完成（提前 2 天）。好差异：老周调研充分。坏差异：3 个修改意见中有 1 个是关于考核权重——这是识别的关键风险。调整：把权重 20% 改成先试点 1 个月，权重暂定 10%。",
     "老周、HR 小张"),
    ("2024-07-15", "P-001", "KPI 修订 + SOP 卡片", "50%", 75,
     "原计划 6/30 前完成 SOP 卡片 + 培训。实际：SOP 提前 5 天完成；培训完成但 4 人不及格。坏差异：培训时长 2 小时太短。调整：4 人补考 + 1 对 1 复训 30 分钟/人。",
     "老周、小陈"),
    ("2024-08-15", "P-001", "KPI 修订 + SOP 卡片", "70%", 100,
     "原计划 8/15 故障率降到 4.5% 以下。实际：4.1%（提前达标）。好差异：车间主任前期沟通到位 + SOP+KPI 组合效果好 + 维修工积极性提高（加班从 8-12 小时降到 4-6 小时）。调整：把 9 月达到 3.8% 作为新目标。",
     "老周、维修工代表"),
    ("2024-08-30", "P-002", "客服系统改造", "30%", 35,
     "原计划 7/15 启动 8/30 完成需求评审。实际：7/15 启动延迟到 7/22。坏差异：技术部排期冲突。调整：申请专职测试资源。",
     "小李、技术-小冯"),
    ("2024-09-20", "P-002", "客服系统改造", "70%", 65,
     "原计划 8/30 完成需求评审。实际：8/30 完成。好差异：业务部门配合度超预期。坏差异：开发资源紧张（被其他项目占用 30%）。调整：申请增加 1 名开发。",
     "小李、技术-小冯"),
    ("2024-09-15", "P-003", "新员工导师制", "30%", 30,
     "原计划 9/1 启动，9/15 完成导师名单。实际：9/1 启动但部分部门经理反对。好差异：HR 总监支持。坏差异：业务部门经理担心增加工作量。调整：增加导师激励（每月 500 元）。",
     "小张、部门经理代表"),
    ("2024-08-30", "P-004", "A 产线 OEE 提升", "50%", 35,
     "原计划 8/15 完成瓶颈工位诊断+方案。实际：8/22 完成。坏差异：诊断设备到货延迟 1 周。调整：调整施工时间表，关键路径不变。",
     "小钱、生产-老何"),
    ("2024-08-30", "P-004", "A 产线 OEE 提升", "70%", 75,
     "原计划 8/22 完成改造+调试。实际：8/22 完成。好差异：施工队加班加点；坏差异：调试发现 1 个新问题。调整：增加 2 天调试期。",
     "小钱、施工队"),
    ("2024-09-30", "P-004", "A 产线 OEE 提升", "100%", 100,
     "原计划 9/30 OEE 达 80%。实际：82%（达标）。好差异：设备改造成功 + 操作工培训到位。坏差异：1 台老设备拖累整体。调整：纳入下期问题解决。",
     "小钱、生产-老何、操作工代表"),
    ("2024-06-30", "P-005", "供应商分级管理", "100%", 100,
     "原计划 5/15 启动，6/30 实施完成。实际：5/22 完成新标准上线。整体达标。",
     "小赵、供应商代表"),
    ("2024-08-30", "P-006", "周会决策提速", "30%", 50,
     "原计划 8/15 启动，1 个月内见效。实际：议程模板发布（50%）。坏差异：管理层老习惯难改。调整：第 1 个月由发起人主持示范。",
     "小周、管理层"),
]

# 数据验证
dv_node = DataValidation(type="list", formula1='"30%,50%,70%,100%"', allow_blank=True)
ws3.add_data_validation(dv_node)
dv_node.add(f"D4:D{3 + len(sample3)}")

for row_idx, row_data in enumerate(sample3, 4):
    for col_idx, value in enumerate(row_data, 1):
        c = ws3.cell(row=row_idx, column=col_idx, value=value)
        c.font = NORMAL_FONT
        c.alignment = LEFT_TOP
        c.border = ALL_BORDER

    # 节点着色
    node = row_data[3]
    if node == "100%":
        ws3.cell(row=row_idx, column=4).fill = OK_FILL
        ws3.cell(row=row_idx, column=4).font = Font(name="Microsoft YaHei", size=10, bold=True, color=OK)
    elif node == "70%":
        ws3.cell(row=row_idx, column=4).fill = PF("FEF3C7")
        ws3.cell(row=row_idx, column=4).font = Font(name="Microsoft YaHei", size=10, bold=True, color="B45309")
    elif node == "50%":
        ws3.cell(row=row_idx, column=4).fill = PF("DBEAFE")
        ws3.cell(row=row_idx, column=4).font = Font(name="Microsoft YaHei", size=10, bold=True, color="1E40AF")
    else:  # 30%
        ws3.cell(row=row_idx, column=4).fill = PF("FCE7F3")
        ws3.cell(row=row_idx, column=4).font = Font(name="Microsoft YaHei", size=10, bold=True, color="9F1239")

    ws3.row_dimensions[row_idx].height = 80

# 复盘统计
sum_row3 = 3 + len(sample3) + 2
ws3.merge_cells(f"A{sum_row3}:G{sum_row3}")
c = ws3.cell(row=sum_row3, column=1, value="复盘统计")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws3.row_dimensions[sum_row3].height = 24

stats_header = ["节点", "复盘次数", "占总数比", "平均完成度", "好差异数", "坏差异数", "总调整数"]
for col_idx, h in enumerate(stats_header, 1):
    c = ws3.cell(row=sum_row3 + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws3.row_dimensions[sum_row3 + 1].height = 30

nodes_list = ["30%", "50%", "70%", "100%"]
n3 = len(sample3)
for i, n in enumerate(nodes_list):
    r = sum_row3 + 2 + i
    ws3.cell(row=r, column=1, value=n).font = BOLD_FONT
    ws3.cell(row=r, column=1).fill = BG_FILL
    ws3.cell(row=r, column=1).alignment = CENTER
    ws3.cell(row=r, column=1).border = ALL_BORDER

    ws3.cell(row=r, column=2, value=f'=COUNTIF(D4:D{3 + n3},A{r})').number_format = "0"
    ws3.cell(row=r, column=3, value=f'=B{r}/COUNTA(D4:D{3 + n3})').number_format = "0.0%"
    ws3.cell(row=r, column=4, value=f'=AVERAGEIF(D4:D{3 + n3},A{r},E4:E{3 + n3})').number_format = "0.0"
    ws3.cell(row=r, column=5, value=f'=COUNTIF(F4:F{3 + n3},"*好差异*")/COUNTIF(D4:D{3 + n3},A{r})').number_format = "0.0"
    ws3.cell(row=r, column=6, value=f'=COUNTIF(F4:F{3 + n3},"*坏差异*")/COUNTIF(D4:D{3 + n3},A{r})').number_format = "0.0"
    ws3.cell(row=r, column=7, value=f'=COUNTIF(F4:F{3 + n3},"*调整*")/COUNTIF(D4:D{3 + n3},A{r})').number_format = "0.0"

    for col in range(2, 8):
        c = ws3.cell(row=r, column=col)
        c.font = NORMAL_FONT
        c.alignment = CENTER
        c.border = ALL_BORDER
    ws3.row_dimensions[r].height = 22

# 合计行
total_r3 = sum_row3 + 2 + len(nodes_list)
ws3.cell(row=total_r3, column=1, value="合计").font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
ws3.cell(row=total_r3, column=1).fill = HEADER_FILL
ws3.cell(row=total_r3, column=1).alignment = CENTER
ws3.cell(row=total_r3, column=1).border = ALL_BORDER

ws3.cell(row=total_r3, column=2, value=f'=SUM(B{sum_row3 + 2}:B{total_r3 - 1})').number_format = "0"
ws3.cell(row=total_r3, column=3, value=f'=SUM(C{sum_row3 + 2}:C{total_r3 - 1})').number_format = "0.0%"
ws3.cell(row=total_r3, column=4, value=f'=AVERAGE(E4:E{3 + n3})').number_format = "0.0"

for col in range(2, 8):
    c = ws3.cell(row=total_r3, column=col)
    c.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = ALL_BORDER
ws3.row_dimensions[total_r3].height = 24

# 复盘经验沉淀
exp_row = total_r3 + 2
ws3.merge_cells(f"A{exp_row}:G{exp_row}")
c = ws3.cell(row=exp_row, column=1, value="团队共性经验沉淀（自动汇总关键句）")
c.font = H2_FONT
c.fill = H2_FILL
c.alignment = LEFT
ws3.row_dimensions[exp_row].height = 24

experience_items = [
    ("✓ 关键路径要识别准——P-001 的部门会议是真正的关键路径，识别对了救了整个项目"),
    ("✓ 关键干系人前期沟通 1 对 1 不可省——P-001 提前和老张沟通避免了部门会议的反对"),
    ("✓ 风险预案要写'具体动作'——'加强沟通'是空话，'6/10 前单独沟通 30 分钟'才是预案"),
    ("✓ 试点机制能降低反对声音——P-001 的'先试点 1 个月'折中方案是关键"),
    ("⚠ 培训时长要够——P-001 的 2 小时太短，4 人不及格，下次延长到 3 小时"),
    ("⚠ 资源冲突要早识别——P-002 的开发资源紧张要到 70% 才暴露，应该 30% 就识别"),
    ("⚠ 旧习惯难改需要示范——P-006 的管理层老习惯，要由发起人主持示范 1 个月"),
    ("⚠ 设备到货风险要预留——P-004 的诊断设备延期 1 周，方案时间表要预留 buffer"),
]
for i, exp in enumerate(experience_items):
    r = exp_row + 1 + i
    ws3.merge_cells(f"A{r}:G{r}")
    c = ws3.cell(row=r, column=1, value=exp)
    c.font = NORMAL_FONT
    c.alignment = LEFT
    c.border = ALL_BORDER
    c.fill = BG_FILL
    ws3.row_dimensions[r].height = 24

ws3.freeze_panes = "A4"
ws3.sheet_view.showGridLines = False


# ============================================================
output_path = r"D:\2026年课程\竞越\基层即学即用的问题解决工具箱\补充课程包\16_配套表单Excel\16_团队问题看板_小组用.xlsx"
wb.save(output_path)
print(f"Generated: {output_path}")
print(f"3 sheets: 问题清单 ({n}) / 进度跟踪 ({n2}) / 复盘记录 ({n3})")

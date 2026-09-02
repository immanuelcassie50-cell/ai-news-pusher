from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "F8_潜力分层培养计划表"

# Define colors
TEAL_DARK = "1F5F7A"
TEAL_LIGHT = "E8F4F7"
INPUT_BG = "F9F9F9"
WHITE = "FFFFFF"
BORDER_COLOR = "B8D4E3"

# Define styles
header_font = Font(name='Microsoft YaHei', size=14, bold=True, color=WHITE)
section_font = Font(name='Microsoft YaHei', size=11, bold=True, color=WHITE)
title_font = Font(name='Microsoft YaHei', size=12, bold=True, color=TEAL_DARK)
normal_font = Font(name='Microsoft YaHei', size=10)
input_font = Font(name='Microsoft YaHei', size=10, color="333333")

header_fill = PatternFill(start_color=TEAL_DARK, end_color=TEAL_DARK, fill_type="solid")
section_fill = PatternFill(start_color=TEAL_DARK, end_color=TEAL_DARK, fill_type="solid")
light_fill = PatternFill(start_color=TEAL_LIGHT, end_color=TEAL_LIGHT, fill_type="solid")
input_fill = PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR)
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Set column widths
col_widths = {'A': 18, 'B': 18, 'C': 18, 'D': 18, 'E': 14}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Row 1: Title
ws.merge_cells('A1:E1')
ws['A1'] = 'F8_潜力分层培养计划表'
ws['A1'].font = Font(name='Microsoft YaHei', size=16, bold=True, color=TEAL_DARK)
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 30

# Row 2: Section 1 Header
ws.merge_cells('A2:E2')
ws['A2'] = '一、潜力分层定义'
ws['A2'].font = section_font
ws['A2'].fill = header_fill
ws['A2'].alignment = center_align
ws.row_dimensions[2].height = 22

# Row 3: Potential Level Headers
ws['A3'] = '高潜力 (High Potential)'
ws['B3'] = '中潜力 (Medium Potential)'
ws['C3'] = '低潜力/稳健型 (Steady)'
for cell in ['A3', 'B3', 'C3']:
    ws[cell].font = title_font
    ws[cell].fill = light_fill
    ws[cell].alignment = center_align
    ws[cell].border = thin_border
ws.row_dimensions[3].height = 18

# Row 4-9: Definition content
definitions = [
    ("定义", "成长速度快、思维层次高、可培养担任更高岗位",
            "成长速度稳定，可培养深耕当前领域",
            "成长空间有限，适合稳定深耕当前岗位"),
    ("特征", "学习能力强，一点就透\n主动挑战更高难度任务\n思维有深度，能举一反三\n愿意承担更大责任",
            "能完成工作要求，学习速度适中\n能在指导下承担新任务\n有意愿学习但主动性一般",
            "能稳定完成本职工作\n不善于应对变化\n倾向于做执行者而非决策者"),
    ("培养策略", "重点培养，给机会、给资源、给挑战",
            "稳定培养，给方向、给反馈、给练习机会",
            "维持培养，保持稳定、给予认可、明确要求"),
]

for row_idx, (label, high, medium, steady) in enumerate(definitions, start=4):
    ws[f'A{row_idx}'] = f"{label}: {high}"
    ws[f'B{row_idx}'] = f"{label}: {medium}"
    ws[f'C{row_idx}'] = f"{label}: {steady}"
    for cell in [f'A{row_idx}', f'B{row_idx}', f'C{row_idx}']:
        ws[cell].font = normal_font
        ws[cell].fill = input_fill
        ws[cell].alignment = left_align
        ws[cell].border = thin_border
    ws.row_dimensions[row_idx].height = 40 if label == "特征" else 18

# Row 10: Empty spacer
ws.row_dimensions[10].height = 8

# Row 11: Section 2 Header
ws.merge_cells('A11:E11')
ws['A11'] = '二、分层培养策略表'
ws['A11'].font = section_font
ws['A11'].fill = header_fill
ws['A11'].alignment = center_align
ws.row_dimensions[11].height = 22

# Row 12: Table Header
strategy_headers = ['培养维度', '高潜力', '中潜力', '稳健型']
for col_idx, header in enumerate(strategy_headers, start=1):
    cell = ws.cell(row=12, column=col_idx)
    cell.value = header
    cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color=WHITE)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[12].height = 20

# Rows 13-18: Strategy Data
strategy_data = [
    ("挑战难度", "高难度任务", "适度挑战", "稳定任务"),
    ("培养方式", "授权+辅导", "教练+反馈", "指导+确认"),
    ("沟通频率", "定期汇报", "例行面谈", "日常跟进"),
    ("成长目标", "成为继任者", "成为骨干", "成为专家"),
    ("辅导重点", "战略思维", "综合能力", "专业技能"),
    ("评估周期", "季度", "半年度", "年度"),
]

for row_idx, (dim, high, medium, steady) in enumerate(strategy_data, start=13):
    data = [dim, high, medium, steady]
    for col_idx, value in enumerate(data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.font = title_font if col_idx == 1 else input_font
        cell.fill = light_fill if col_idx == 1 else input_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[row_idx].height = 18

# Row 19: Empty spacer
ws.row_dimensions[19].height = 8

# Row 20: Section 3 Header
ws.merge_cells('A20:E20')
ws['A20'] = '三、个人培养计划表'
ws['A20'].font = section_font
ws['A20'].fill = header_fill
ws['A20'].alignment = center_align
ws.row_dimensions[20].height = 22

# Row 21: Basic Info Label
ws.merge_cells('A21:E21')
ws['A21'] = '基本信息'
ws['A21'].font = title_font
ws['A21'].fill = light_fill
ws['A21'].alignment = left_align
ws['A21'].border = thin_border
ws.row_dimensions[21].height = 18

# Row 22: Basic Info Headers
basic_headers = ['姓名', '当前岗位', '潜力评估', '计划周期']
for col_idx, header in enumerate(basic_headers, start=1):
    cell = ws.cell(row=22, column=col_idx)
    cell.value = header
    cell.font = title_font
    cell.fill = light_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[22].height = 20

# Row 23: Basic Info Inputs
ws['A23'].fill = input_fill
ws['B23'].fill = input_fill
ws['C23'] = '□ 高  □ 中  □ 稳健'
ws['C23'].font = normal_font
ws['C23'].fill = input_fill
ws['D23'] = '____年____月至____年____月'
ws['D23'].font = normal_font
ws['D23'].fill = input_fill
for col in ['A23', 'B23', 'C23', 'D23']:
    ws[col].border = thin_border
    ws[col].alignment = center_align
ws.row_dimensions[23].height = 22

# Row 24: Ability Improvement Label
ws.merge_cells('A24:E24')
ws['A24'] = '能力提升计划'
ws['A24'].font = title_font
ws['A24'].fill = light_fill
ws['A24'].alignment = left_align
ws['A24'].border = thin_border
ws.row_dimensions[24].height = 18

# Row 25: Ability Improvement Table Header
ability_headers = ['提升目标', '具体行动', '时间节点', '资源支持', '评估方式']
for col_idx, header in enumerate(ability_headers, start=1):
    cell = ws.cell(row=25, column=col_idx)
    cell.value = header
    cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color=WHITE)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[25].height = 20

# Rows 26-28: Ability Improvement Data
for row_idx in range(26, 29):
    for col_idx in range(1, 6):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = input_fill
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[row_idx].height = 20

# Row 29: Rotation/Project Label
ws.merge_cells('A29:E29')
ws['A29'] = '轮岗/项目计划'
ws['A29'].font = title_font
ws['A29'].fill = light_fill
ws['A29'].alignment = left_align
ws['A29'].border = thin_border
ws.row_dimensions[29].height = 18

# Row 30: Rotation/Project Table Header
rotation_headers = ['轮岗/项目', '目的', '时间', '期望收获', '评估方式']
for col_idx, header in enumerate(rotation_headers, start=1):
    cell = ws.cell(row=30, column=col_idx)
    cell.value = header
    cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color=WHITE)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[30].height = 20

# Rows 31-33: Rotation/Project Data
for row_idx in range(31, 34):
    for col_idx in range(1, 6):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = input_fill
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[row_idx].height = 20

# Row 34: Mentor/Coach Label
ws.merge_cells('A34:E34')
ws['A34'] = '导师/教练安排'
ws['A34'].font = title_font
ws['A34'].fill = light_fill
ws['A34'].alignment = left_align
ws['A34'].border = thin_border
ws.row_dimensions[34].height = 18

# Row 35: Mentor Headers
mentor_headers = ['内部导师', '外部资源', '定期沟通频率']
for col_idx, header in enumerate(mentor_headers, start=1):
    cell = ws.cell(row=35, column=col_idx)
    cell.value = header
    cell.font = title_font
    cell.fill = light_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[35].height = 22

# Row 36: Mentor Inputs
for col_idx in range(1, 4):
    cell = ws.cell(row=36, column=col_idx)
    cell.fill = input_fill
    cell.border = thin_border
    cell.alignment = center_align
ws.row_dimensions[36].height = 22

# Row 37: Empty spacer
ws.row_dimensions[37].height = 8

# Row 38: Section 4 Header
ws.merge_cells('A38:E38')
ws['A38'] = '四、培养记录'
ws['A38'].font = section_font
ws['A38'].fill = header_fill
ws['A38'].alignment = center_align
ws.row_dimensions[38].height = 22

# Row 39: Training Record Table Header
record_headers = ['日期', '培养活动', '收获/成果', '反馈', '调整事项']
for col_idx, header in enumerate(record_headers, start=1):
    cell = ws.cell(row=39, column=col_idx)
    cell.value = header
    cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color=WHITE)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[39].height = 20

# Rows 40-44: Training Record Data
for row_idx in range(40, 45):
    for col_idx in range(1, 6):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = input_fill
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[row_idx].height = 20

# Row 45: Footer
ws['A45'] = '表格编号：F8'
ws['A45'].font = Font(name='Microsoft YaHei', size=9, color="666666")
ws['C45'] = '版本：V1.0'
ws['C45'].font = Font(name='Microsoft YaHei', size=9, color="666666")
ws['E45'] = '编制日期：2026年7月'
ws['E45'].font = Font(name='Microsoft YaHei', size=9, color="666666")
ws.row_dimensions[45].height = 16

# Set print settings for A4 landscape
from openpyxl.worksheet.page import PageMargins
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = 9  # A4
ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75)

# Save
output_path = "D:/新课开发/经验萃取/语音记录转管理者赋能手册/完整课程包/工具表单/F8_潜力分层培养计划表.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")

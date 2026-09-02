# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = "D:/新课开发/工作手册/客户隐性需求挖掘与验证/完整课程包/06_工具表单"

# Color constants
HEADER_BG = "C00000"
TITLE_BG = "E85053"
SECTION_BG = "F2F2F2"
ALT_ROW_BG = "FFF0F0"
INPUT_BG = "DAEEF3"
HIGH_PRIORITY_BG = "FFE0E0"
MEDIUM_PRIORITY_BG = "E2EFDA"
WHITE = "FFFFFF"

def create_thin_border():
    thin = Side(style='thin', color='000000')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def create_header_fill():
    return PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')

def create_title_fill():
    return PatternFill(start_color=TITLE_BG, end_color=TITLE_BG, fill_type='solid')

def create_section_fill():
    return PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type='solid')

def create_alt_row_fill():
    return PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type='solid')

def create_input_fill():
    return PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type='solid')

def create_white_fill():
    return PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')

def create_high_priority_fill():
    return PatternFill(start_color=HIGH_PRIORITY_BG, end_color=HIGH_PRIORITY_BG, fill_type='solid')

def create_medium_priority_fill():
    return PatternFill(start_color=MEDIUM_PRIORITY_BG, end_color=MEDIUM_PRIORITY_BG, fill_type='solid')

def set_column_widths(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

def wrap_alignment(horizontal='center', vertical='center'):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)

def create_file_1():
    """01_隐性需求挖掘访谈工具包.xlsx"""
    wb = Workbook()

    # ===== Sheet 1: 访谈前准备清单 =====
    ws = wb.active
    ws.title = "访谈前准备清单"

    # Title row (row 2)
    ws.merge_cells('A2:E2')
    title_cell = ws['A2']
    title_cell.value = "【工具一】隐性需求挖掘访谈工具包 - 访谈前准备清单"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws.row_dimensions[2].height = 30

    # Instruction row (row 4)
    ws.merge_cells('A4:E4')
    instr_cell = ws['A4']
    instr_cell.value = "使用说明：进入访谈前，逐一核对以下准备项，全部完成后再出发。浅蓝色单元格为待填写输入区。"
    instr_cell.font = Font(size=10, italic=True, color="808080")
    instr_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 30

    # Header row (row 5)
    headers = ["序号", "准备事项", "具体内容/备注", "状态", "检查人"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws.row_dimensions[5].height = 20

    # Data rows (rows 6-12)
    data = [
        ("1", "背景调研", "", "待验证", ""),
        ("2", "组织分析", "", "待验证", ""),
        ("3", "场景式问题准备", "", "待验证", ""),
        ("4", "历史案例回顾", "", "待验证", ""),
        ("5", "关键决策人确认", "", "待验证", ""),
        ("6", "最挑剔人选分析", "", "待验证", ""),
        ("7", "访谈目标明确", "", "待验证", ""),
    ]

    for i, (num, item, desc, status, checker) in enumerate(data):
        row = 6 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        cell_a = ws.cell(row=row, column=1)
        cell_a.value = num
        cell_a.font = Font(size=10)
        cell_a.fill = fill
        cell_a.alignment = wrap_alignment()
        cell_a.border = create_thin_border()

        cell_b = ws.cell(row=row, column=2)
        cell_b.value = item
        cell_b.font = Font(size=10)
        cell_b.fill = fill
        cell_b.alignment = wrap_alignment()
        cell_b.border = create_thin_border()

        cell_c = ws.cell(row=row, column=3)
        cell_c.value = desc
        cell_c.font = Font(size=10)
        cell_c.fill = create_input_fill()
        cell_c.alignment = wrap_alignment(horizontal='left')
        cell_c.border = create_thin_border()

        cell_d = ws.cell(row=row, column=4)
        cell_d.value = status
        cell_d.font = Font(size=10)
        cell_d.fill = create_input_fill()
        cell_d.alignment = wrap_alignment()
        cell_d.border = create_thin_border()

        cell_e = ws.cell(row=row, column=5)
        cell_e.value = checker
        cell_e.font = Font(size=10)
        cell_e.fill = create_input_fill()
        cell_e.alignment = wrap_alignment()
        cell_e.border = create_thin_border()

        ws.row_dimensions[row].height = 18

    ws.freeze_panes = 'A6'
    set_column_widths(ws, [6, 18, 30, 12, 12])

    # ===== Sheet 2: 核心提问模板 =====
    ws2 = wb.create_sheet("核心提问模板")

    ws2.merge_cells('A2:E2')
    title_cell = ws2['A2']
    title_cell.value = "【工具一】隐性需求挖掘访谈工具包 - 核心提问模板"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws2.row_dimensions[2].height = 30

    ws2.merge_cells('A4:E4')
    instr_cell = ws2['A4']
    instr_cell.value = "使用说明：根据访谈场景选择合适的提问模板，提问时注意语气和时机，红色加粗文字为核心话术。"
    instr_cell.font = Font(size=10, italic=True, color="808080")
    instr_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws2.row_dimensions[4].height = 30

    headers = ["序号", "提问类型", "参考提问模板", "使用时机", "注意事项"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws2.row_dimensions[5].height = 20

    data = [
        ("1", "开场暖场", "今天专程来拜访，想先听听您这边对于目前工作的一些看法。", "访谈开始前5分钟", "营造轻松氛围，不要直接进入正题"),
        ("2", "现状探索", "能描述一下您目前工作中最让您有成就感的一个场景吗？", "了解工作背景时", "引导客户描述具体场景而非抽象概念"),
        ("3", "痛点挖掘", "在您的工作中，有没有让您觉得特别费劲或者不太顺的地方？", "建立信任后", "注意观察被访者的微表情和语气变化"),
        ("4", "影响放大", "这个问题给您的工作带来了多大的影响呢？", "发现痛点后", "用具体数字或案例来量化影响程度"),
        ("5", "需求验证", "如果这个问题得到解决，您最希望看到什么样的改变？", "明确需求时", "帮助客户想象解决后的美好场景"),
        ("6", "决策链探询", "您觉得这个问题目前在公司层面被重视的程度如何？", "了解组织立场时", "判断需求的紧迫性和资源调配可能性"),
        ("7", "风险排查", "您觉得实施这个解决方案可能会有什么顾虑或风险？", "接近尾声时", "提前预判实施障碍"),
        ("8", "承诺获取", "基于今天的交流，您觉得下一步我们最应该先做什么？", "访谈结束时", "明确下一步行动和责任人"),
    ]

    for i, row_data in enumerate(data):
        row = 6 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        for col, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row, column=col)
            if col == 3:
                cell.value = value
                cell.font = Font(size=10, bold=True, color=HEADER_BG)
            else:
                cell.value = value
                cell.font = Font(size=10)
            cell.fill = fill
            cell.alignment = wrap_alignment(horizontal='left') if col in [3, 4, 5] else wrap_alignment()
            cell.border = create_thin_border()
        ws2.row_dimensions[row].height = 18

    ws2.freeze_panes = 'A6'
    set_column_widths(ws2, [6, 12, 40, 18, 25])

    # ===== Sheet 3: 追问技巧提示卡 =====
    ws3 = wb.create_sheet("追问技巧提示卡")

    ws3.merge_cells('A2:D2')
    title_cell = ws3['A2']
    title_cell.value = "【工具一】隐性需求挖掘访谈工具包 - 追问技巧提示卡"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws3.row_dimensions[2].height = 30

    headers = ["序号", "场景", "应对技巧", "优先级"]
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws3.row_dimensions[4].height = 20

    tips = [
        ("1", "客户回避问题", '采用"如果...会怎样"的假设性提问，引导客户脱离具体情境，进入抽象分析', "高"),
        ("2", "客户表达模糊", '用"能具体举个例子吗"追问，要求客户给出具体场景和数字', "高"),
        ("3", "客户突然沉默", "保持沉默等待7秒，或换个轻松话题过渡，给客户思考空间", "高"),
        ("4", "客户情绪激动", '表示理解和认同，"我理解您的心情"，待情绪平复后再继续', "中"),
        ("5", "话题偏离太远", '用"这个很有意思"肯定后，自然转回，"那关于刚才说的..."', "中"),
        ("6", "客户过于理性", '引入案例："有家企业也遇到类似情况，他们是这样处理的..."', "中"),
        ("7", "信息相互矛盾", '不直接指出，而是"您刚才提到A，我理解是...，那B是不是也符合这个逻辑?"', "高"),
    ]

    for i, row_data in enumerate(tips):
        row = 5 + i
        priority = row_data[3]
        if priority == "高":
            row_fill = create_high_priority_fill()
        else:
            row_fill = create_medium_priority_fill()

        for col, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=row, column=col)
            cell.value = value
            cell.font = Font(size=10)
            cell.fill = row_fill
            cell.alignment = wrap_alignment(horizontal='left') if col == 3 else wrap_alignment()
            cell.border = create_thin_border()
        ws3.row_dimensions[row].height = 18

    ws3.freeze_panes = 'A5'
    set_column_widths(ws3, [6, 18, 50, 10])

    # ===== Sheet 4: 访谈记录表 =====
    ws4 = wb.create_sheet("访谈记录表")

    ws4.merge_cells('A2:F2')
    title_cell = ws4['A2']
    title_cell.value = "【工具一】隐性需求挖掘访谈工具包 - 访谈记录表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws4.row_dimensions[2].height = 30

    ws4.merge_cells('A4:F4')
    ws4['A4'].value = "基本信息"
    ws4['A4'].font = Font(size=11, bold=True)
    ws4['A4'].fill = create_section_fill()
    ws4['A4'].alignment = wrap_alignment()

    header_info = [
        ("访谈日期", ""),
        ("被访者姓名", ""),
        ("被访者职位", ""),
        ("被访者部门", ""),
        ("访谈时长", ""),
        ("访谈地点", ""),
        ("访谈人", ""),
    ]

    row = 5
    for i, (label, value) in enumerate(header_info):
        cell_label = ws4.cell(row=row, column=1)
        cell_label.value = label
        cell_label.font = Font(size=10, bold=True)
        cell_label.fill = create_section_fill()
        cell_label.alignment = wrap_alignment()
        cell_label.border = create_thin_border()

        ws4.merge_cells(f'B{row}:D{row}')
        cell_value = ws4.cell(row=row, column=2)
        cell_value.value = value
        cell_value.font = Font(size=10)
        cell_value.fill = create_input_fill()
        cell_value.alignment = wrap_alignment(horizontal='left')
        cell_value.border = create_thin_border()

        ws4.merge_cells(f'E{row}:F{row}')
        cell_note = ws4.cell(row=row, column=5)
        cell_note.value = ""
        cell_note.font = Font(size=10)
        cell_note.fill = create_input_fill()
        cell_note.alignment = wrap_alignment(horizontal='left')
        cell_note.border = create_thin_border()

        ws4.row_dimensions[row].height = 18
        row += 1

    row += 1
    ws4.merge_cells(f'A{row}:F{row}')
    ws4[f'A{row}'].value = "被访者背景"
    ws4[f'A{row}'].font = Font(size=11, bold=True)
    ws4[f'A{row}'].fill = create_section_fill()
    ws4[f'A{row}'].alignment = wrap_alignment()
    ws4.row_dimensions[row].height = 18
    row += 1

    bg_items = ["工作经历年限", "直接下属人数", "汇报对象层级", "主要职责范围"]
    for item in bg_items:
        cell_label = ws4.cell(row=row, column=1)
        cell_label.value = item
        cell_label.font = Font(size=10, bold=True)
        cell_label.fill = create_section_fill()
        cell_label.alignment = wrap_alignment()
        cell_label.border = create_thin_border()

        ws4.merge_cells(f'B{row}:F{row}')
        cell_value = ws4.cell(row=row, column=2)
        cell_value.value = ""
        cell_value.font = Font(size=10)
        cell_value.fill = create_input_fill()
        cell_value.alignment = wrap_alignment(horizontal='left')
        cell_value.border = create_thin_border()
        ws4.row_dimensions[row].height = 18
        row += 1

    row += 1
    ws4.merge_cells(f'A{row}:F{row}')
    ws4[f'A{row}'].value = "核心需求信号"
    ws4[f'A{row}'].font = Font(size=11, bold=True)
    ws4[f'A{row}'].fill = create_section_fill()
    ws4[f'A{row}'].alignment = wrap_alignment()
    ws4.row_dimensions[row].height = 18
    row += 1

    signal_headers = ["序号", "需求信号内容", "信号强度(1-5)", "出现时机", "备注"]
    for col, header in enumerate(signal_headers, start=1):
        cell = ws4.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(size=10, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws4.row_dimensions[row].height = 18
    row += 1

    for i in range(5):
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()
        ws4.cell(row=row, column=1).value = i + 1
        ws4.cell(row=row, column=1).font = Font(size=10)
        ws4.cell(row=row, column=1).fill = fill
        ws4.cell(row=row, column=1).alignment = wrap_alignment()
        ws4.cell(row=row, column=1).border = create_thin_border()

        for col in range(2, 6):
            cell = ws4.cell(row=row, column=col)
            cell.value = ""
            cell.font = Font(size=10)
            cell.fill = create_input_fill()
            cell.alignment = wrap_alignment(horizontal='left')
            cell.border = create_thin_border()
        ws4.row_dimensions[row].height = 18
        row += 1

    row += 1
    ws4.merge_cells(f'A{row}:F{row}')
    ws4[f'A{row}'].value = "隐性需求初步判断"
    ws4[f'A{row}'].font = Font(size=11, bold=True)
    ws4[f'A{row}'].fill = create_section_fill()
    ws4[f'A{row}'].alignment = wrap_alignment()
    ws4.row_dimensions[row].height = 18
    row += 1

    ws4.merge_cells(f'A{row}:F{row}')
    ws4.cell(row=row, column=1).value = "隐性需求描述"
    ws4.cell(row=row, column=1).font = Font(size=10, bold=True, color=WHITE)
    ws4.cell(row=row, column=1).fill = create_header_fill()
    ws4.cell(row=row, column=1).alignment = wrap_alignment()
    ws4.cell(row=row, column=1).border = create_thin_border()
    ws4.row_dimensions[row].height = 18
    row += 1

    for i in range(3):
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()
        ws4.merge_cells(f'A{row}:F{row}')
        cell = ws4.cell(row=row, column=1)
        cell.value = ""
        cell.font = Font(size=10)
        cell.fill = create_input_fill()
        cell.alignment = wrap_alignment(horizontal='left')
        cell.border = create_thin_border()
        ws4.row_dimensions[row].height = 36
        row += 1

    ws4.freeze_panes = 'A5'
    set_column_widths(ws4, [12, 20, 12, 15, 15, 15])

    # ===== Sheet 5: 访谈复盘表 =====
    ws5 = wb.create_sheet("访谈复盘表")

    ws5.merge_cells('A2:E2')
    title_cell = ws5['A2']
    title_cell.value = "【工具一】隐性需求挖掘访谈工具包 - 访谈复盘表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws5.row_dimensions[2].height = 30

    headers = ["序号", "复盘维度", "具体内容", "自我评分(1-5)", "改进计划"]
    for col, header in enumerate(headers, start=1):
        cell = ws5.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws5.row_dimensions[4].height = 20

    review_dims = [
        "访谈目标达成度", "问题质量", "倾听力", "追问深度", "关系建立", "记录完整性", "时间把控"
    ]

    for i, dim in enumerate(review_dims):
        row = 5 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws5.cell(row=row, column=1).value = i + 1
        ws5.cell(row=row, column=1).font = Font(size=10)
        ws5.cell(row=row, column=1).fill = fill
        ws5.cell(row=row, column=1).alignment = wrap_alignment()
        ws5.cell(row=row, column=1).border = create_thin_border()

        ws5.cell(row=row, column=2).value = dim
        ws5.cell(row=row, column=2).font = Font(size=10)
        ws5.cell(row=row, column=2).fill = fill
        ws5.cell(row=row, column=2).alignment = wrap_alignment()
        ws5.cell(row=row, column=2).border = create_thin_border()

        ws5.cell(row=row, column=3).value = ""
        ws5.cell(row=row, column=3).font = Font(size=10)
        ws5.cell(row=row, column=3).fill = create_input_fill()
        ws5.cell(row=row, column=3).alignment = wrap_alignment(horizontal='left')
        ws5.cell(row=row, column=3).border = create_thin_border()

        ws5.cell(row=row, column=4).value = ""
        ws5.cell(row=row, column=4).font = Font(size=10)
        ws5.cell(row=row, column=4).fill = create_input_fill()
        ws5.cell(row=row, column=4).alignment = wrap_alignment()
        ws5.cell(row=row, column=4).border = create_thin_border()

        ws5.cell(row=row, column=5).value = ""
        ws5.cell(row=row, column=5).font = Font(size=10)
        ws5.cell(row=row, column=5).fill = create_input_fill()
        ws5.cell(row=row, column=5).alignment = wrap_alignment(horizontal='left')
        ws5.cell(row=row, column=5).border = create_thin_border()

        ws5.row_dimensions[row].height = 18

    dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv.error = "请选择1-5之间的分数"
    dv.errorTitle = "输入无效"
    ws5.add_data_validation(dv)
    dv.add(f'D5:D11')

    row = 13
    row += 1
    ws5.merge_cells(f'A{row}:E{row}')
    ws5[f'A{row}'].value = "被忽略信号记录"
    ws5[f'A{row}'].font = Font(size=11, bold=True)
    ws5[f'A{row}'].fill = create_section_fill()
    ws5[f'A{row}'].alignment = wrap_alignment()
    ws5.row_dimensions[row].height = 18
    row += 1

    for i in range(3):
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()
        ws5.merge_cells(f'A{row}:E{row}')
        cell = ws5.cell(row=row, column=1)
        cell.value = f"信号{i+1}: "
        cell.font = Font(size=10)
        cell.fill = create_input_fill()
        cell.alignment = wrap_alignment(horizontal='left')
        cell.border = create_thin_border()
        ws5.row_dimensions[row].height = 36
        row += 1

    ws5.freeze_panes = 'A5'
    set_column_widths(ws5, [6, 15, 35, 12, 25])

    wb.save(os.path.join(OUTPUT_DIR, "01_隐性需求挖掘访谈工具包.xlsx"))
    print("Created: 01_隐性需求挖掘访谈工具包.xlsx")

def create_file_2():
    """02_验证实验设计工具.xlsx"""
    wb = Workbook()

    # ===== Sheet 1: 排错式验证五问 =====
    ws = wb.active
    ws.title = "排错式验证五问"

    ws.merge_cells('A2:D2')
    title_cell = ws['A2']
    title_cell.value = "【工具二】验证实验设计工具 - 排错式验证五问"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A4:D4')
    instr_cell = ws['A4']
    instr_cell.value = "使用说明：在设计验证实验前，逐一思考以下五个问题，确保验证方案的科学性和可操作性。"
    instr_cell.font = Font(size=10, italic=True, color="808080")
    instr_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 30

    headers = ["序号", "验证问题", "填写区", ""]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws.row_dimensions[5].height = 20

    questions = [
        ("1", "最有可能导致假设不成立的反面证据是什么？", "请列出3-5个可能的反面证据"),
        ("2", "用什么方式可以最快速度证伪这个假设？", "描述具体的验证方法和预期结果"),
        ("3", "验证过程中最可能出现的干扰因素有哪些？", "识别并列出潜在干扰因素"),
        ("4", "最小的验证实验单元是什么？需要多少资源？", "明确实验范围和资源需求"),
        ("5", "如果假设被证伪，下一步的行动是什么？", "制定备选方案或调整方向"),
    ]

    for i, (num, question, hint) in enumerate(questions):
        row = 6 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        cell_a = ws.cell(row=row, column=1)
        cell_a.value = num
        cell_a.font = Font(size=10, bold=True, color=WHITE)
        cell_a.fill = create_header_fill()
        cell_a.alignment = wrap_alignment()
        cell_a.border = create_thin_border()

        cell_b = ws.cell(row=row, column=2)
        cell_b.value = question
        cell_b.font = Font(size=10, bold=True, color=HEADER_BG)
        cell_b.fill = fill
        cell_b.alignment = wrap_alignment(horizontal='left', vertical='center')
        cell_b.border = create_thin_border()

        cell_c = ws.cell(row=row, column=3)
        cell_c.value = hint
        cell_c.font = Font(size=10, color="808080", italic=True)
        cell_c.fill = create_input_fill()
        cell_c.alignment = wrap_alignment(horizontal='left', vertical='center')
        cell_c.border = create_thin_border()

        cell_d = ws.cell(row=row, column=4)
        cell_d.value = ""
        cell_d.font = Font(size=10)
        cell_d.fill = create_input_fill()
        cell_d.alignment = wrap_alignment(horizontal='left', vertical='center')
        cell_d.border = create_thin_border()

        ws.row_dimensions[row].height = 50

    ws.freeze_panes = 'A6'
    set_column_widths(ws, [6, 40, 35, 20])

    # ===== Sheet 2: 最小暴露实验设计表 =====
    ws2 = wb.create_sheet("最小暴露实验设计表")

    ws2.merge_cells('A2:F2')
    title_cell = ws2['A2']
    title_cell.value = "【工具二】验证实验设计工具 - 最小暴露实验设计表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws2.row_dimensions[2].height = 30

    headers = ["序号", "实验要素", "填写内容", "状态", "下次验证时间"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws2.row_dimensions[4].height = 20

    elements = [
        ("1", "假设/待验证点", "", "", ""),
        ("2", "验证目标", "", "", ""),
        ("3", "验证方式", "", "", ""),
        ("4", "测试对象/样本", "", "", ""),
        ("5", "实验组设置", "", "", ""),
        ("6", "对照组设置", "", "", ""),
        ("7", "观察指标", "", "", ""),
        ("8", "数据收集方法", "", "", ""),
        ("9", "预期结果", "", "", ""),
        ("10", "实际结果", "", "", ""),
        ("11", "结论与行动", "", "", ""),
    ]

    for i, (num, elem, content, status, next_time) in enumerate(elements):
        row = 5 + i
        if i >= 8:
            fill = create_alt_row_fill()
        else:
            fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws2.cell(row=row, column=1).value = num
        ws2.cell(row=row, column=1).font = Font(size=10)
        ws2.cell(row=row, column=1).fill = fill
        ws2.cell(row=row, column=1).alignment = wrap_alignment()
        ws2.cell(row=row, column=1).border = create_thin_border()

        ws2.cell(row=row, column=2).value = elem
        ws2.cell(row=row, column=2).font = Font(size=10)
        ws2.cell(row=row, column=2).fill = fill
        ws2.cell(row=row, column=2).alignment = wrap_alignment(horizontal='left')
        ws2.cell(row=row, column=2).border = create_thin_border()

        ws2.cell(row=row, column=3).value = content
        ws2.cell(row=row, column=3).font = Font(size=10)
        ws2.cell(row=row, column=3).fill = create_input_fill()
        ws2.cell(row=row, column=3).alignment = wrap_alignment(horizontal='left')
        ws2.cell(row=row, column=3).border = create_thin_border()

        ws2.cell(row=row, column=4).value = status
        ws2.cell(row=row, column=4).font = Font(size=10)
        ws2.cell(row=row, column=4).fill = create_input_fill()
        ws2.cell(row=row, column=4).alignment = wrap_alignment()
        ws2.cell(row=row, column=4).border = create_thin_border()

        ws2.cell(row=row, column=5).value = next_time
        ws2.cell(row=row, column=5).font = Font(size=10)
        ws2.cell(row=row, column=5).fill = create_input_fill()
        ws2.cell(row=row, column=5).alignment = wrap_alignment()
        ws2.cell(row=row, column=5).border = create_thin_border()

        ws2.row_dimensions[row].height = 25 if i >= 8 else 18

    ws2.freeze_panes = 'A5'
    set_column_widths(ws2, [6, 15, 30, 12, 15])

    # ===== Sheet 3: 验证结果记录表 =====
    ws3 = wb.create_sheet("验证结果记录表")

    ws3.merge_cells('A2:F2')
    title_cell = ws3['A2']
    title_cell.value = "【工具二】验证实验设计工具 - 验证结果记录表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws3.row_dimensions[2].height = 30

    headers = ["验证目标", "验证方式", "测试对象", "观察到的第一反应", "结论", "下一步动作"]
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws3.row_dimensions[4].height = 20

    for i in range(12):
        row = 5 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        for col in range(1, 7):
            cell = ws3.cell(row=row, column=col)
            cell.value = ""
            cell.font = Font(size=10)
            cell.fill = create_input_fill()
            cell.alignment = wrap_alignment(horizontal='left', vertical='center')
            cell.border = create_thin_border()
        ws3.row_dimensions[row].height = 25

    ws3.freeze_panes = 'A5'
    set_column_widths(ws3, [20, 15, 15, 25, 20, 18])

    # ===== Sheet 4: 验证结论汇总表 =====
    ws4 = wb.create_sheet("验证结论汇总表")

    ws4.merge_cells('A2:G2')
    title_cell = ws4['A2']
    title_cell.value = "【工具二】验证实验设计工具 - 验证结论汇总表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws4.row_dimensions[2].height = 30

    headers = ["序号", "假设/待验证点", "验证方式", "验证结论", "信心度", "负责人", "更新时间"]
    for col, header in enumerate(headers, start=1):
        cell = ws4.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws4.row_dimensions[4].height = 20

    for i in range(10):
        row = 5 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws4.cell(row=row, column=1).value = i + 1
        ws4.cell(row=row, column=1).font = Font(size=10)
        ws4.cell(row=row, column=1).fill = fill
        ws4.cell(row=row, column=1).alignment = wrap_alignment()
        ws4.cell(row=row, column=1).border = create_thin_border()

        for col in range(2, 8):
            cell = ws4.cell(row=row, column=col)
            cell.value = ""
            cell.font = Font(size=10)
            cell.fill = create_input_fill()
            cell.alignment = wrap_alignment(horizontal='left', vertical='center')
            cell.border = create_thin_border()
        ws4.row_dimensions[row].height = 18

    row = 16
    ws4.merge_cells(f'A{row}:G{row}')
    ws4[f'A{row}'].value = "统计汇总"
    ws4[f'A{row}'].font = Font(size=11, bold=True)
    ws4[f'A{row}'].fill = create_section_fill()
    ws4[f'A{row}'].alignment = wrap_alignment()
    ws4.row_dimensions[row].height = 20
    row += 1

    ws4.cell(row=row, column=1).value = "高信心度假设数量"
    ws4.cell(row=row, column=1).font = Font(size=10)
    ws4.cell(row=row, column=1).fill = create_section_fill()
    ws4.cell(row=row, column=1).alignment = wrap_alignment()
    ws4.cell(row=row, column=1).border = create_thin_border()

    ws4.merge_cells(f'B{row}:C{row}')
    ws4.cell(row=row, column=2).value = '=COUNTIF(E5:E14,">=4")'
    ws4.cell(row=row, column=2).font = Font(size=10)
    ws4.cell(row=row, column=2).fill = create_input_fill()
    ws4.cell(row=row, column=2).alignment = wrap_alignment()
    ws4.cell(row=row, column=2).border = create_thin_border()
    ws4.row_dimensions[row].height = 18
    row += 1

    ws4.cell(row=row, column=1).value = "待验证假设数量"
    ws4.cell(row=row, column=1).font = Font(size=10)
    ws4.cell(row=row, column=1).fill = create_section_fill()
    ws4.cell(row=row, column=1).alignment = wrap_alignment()
    ws4.cell(row=row, column=1).border = create_thin_border()

    ws4.merge_cells(f'B{row}:C{row}')
    ws4.cell(row=row, column=2).value = '=COUNTIF(E5:E14,"<3")'
    ws4.cell(row=row, column=2).font = Font(size=10)
    ws4.cell(row=row, column=2).fill = create_input_fill()
    ws4.cell(row=row, column=2).alignment = wrap_alignment()
    ws4.cell(row=row, column=2).border = create_thin_border()
    ws4.row_dimensions[row].height = 18

    ws4.freeze_panes = 'A5'
    set_column_widths(ws4, [6, 25, 15, 20, 10, 10, 12])

    wb.save(os.path.join(OUTPUT_DIR, "02_验证实验设计工具.xlsx"))
    print("Created: 02_验证实验设计工具.xlsx")

def create_file_3():
    """03_需求优先级矩阵工具.xlsx"""
    wb = Workbook()

    # ===== Sheet 1: 三维评估表 =====
    ws = wb.active
    ws.title = "三维评估表"

    ws.merge_cells('A2:G2')
    title_cell = ws['A2']
    title_cell.value = "【工具三】需求优先级矩阵工具 - 三维评估表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A4:G4')
    instr_cell = ws['A4']
    instr_cell.value = "使用说明：对每个需求/功能进行三个维度的评估打分，综合评分=业务价值+责任敞口，优先级建议根据评分自动计算。"
    instr_cell.font = Font(size=10, italic=True, color="808080")
    instr_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 30

    headers = ["序号", "需求/功能描述", "业务价值(1-5)", "实现成本(1-5)", "责任敞口(1-5)", "综合评分", "优先级建议"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws.row_dimensions[5].height = 20

    dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv.error = "请选择1-5之间的分数"
    dv.errorTitle = "输入无效"
    ws.add_data_validation(dv)

    for i in range(15):
        row = 6 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws.cell(row=row, column=1).value = i + 1
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).alignment = wrap_alignment()
        ws.cell(row=row, column=1).border = create_thin_border()

        ws.cell(row=row, column=2).value = ""
        ws.cell(row=row, column=2).font = Font(size=10)
        ws.cell(row=row, column=2).fill = create_input_fill()
        ws.cell(row=row, column=2).alignment = wrap_alignment(horizontal='left', vertical='center')
        ws.cell(row=row, column=2).border = create_thin_border()

        ws.cell(row=row, column=3).value = ""
        ws.cell(row=row, column=3).font = Font(size=10)
        ws.cell(row=row, column=3).fill = create_input_fill()
        ws.cell(row=row, column=3).alignment = wrap_alignment()
        ws.cell(row=row, column=3).border = create_thin_border()
        dv.add(f'C{row}')

        ws.cell(row=row, column=4).value = ""
        ws.cell(row=row, column=4).font = Font(size=10)
        ws.cell(row=row, column=4).fill = create_input_fill()
        ws.cell(row=row, column=4).alignment = wrap_alignment()
        ws.cell(row=row, column=4).border = create_thin_border()
        dv.add(f'D{row}')

        ws.cell(row=row, column=5).value = ""
        ws.cell(row=row, column=5).font = Font(size=10)
        ws.cell(row=row, column=5).fill = create_input_fill()
        ws.cell(row=row, column=5).alignment = wrap_alignment()
        ws.cell(row=row, column=5).border = create_thin_border()
        dv.add(f'E{row}')

        ws.cell(row=row, column=6).value = f'=IF(C{row}="","",C{row}+E{row})'
        ws.cell(row=row, column=6).font = Font(size=10, bold=True)
        ws.cell(row=row, column=6).fill = fill
        ws.cell(row=row, column=6).alignment = wrap_alignment()
        ws.cell(row=row, column=6).border = create_thin_border()

        ws.cell(row=row, column=7).value = f'=IF(F{row}="","",IF(F{row}>=8,"最高",IF(F{row}>=6,"高",IF(F{row}>=4,"中","低"))))'
        ws.cell(row=row, column=7).font = Font(size=10)
        ws.cell(row=row, column=7).fill = fill
        ws.cell(row=row, column=7).alignment = wrap_alignment()
        ws.cell(row=row, column=7).border = create_thin_border()

        ws.row_dimensions[row].height = 18

    row = 22
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'].value = "评分标准说明"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    ws[f'A{row}'].fill = create_section_fill()
    ws[f'A{row}'].alignment = wrap_alignment()
    ws.row_dimensions[row].height = 20
    row += 1

    criteria = [
        ("业务价值", "1=无感知价值 2=略有价值 3=一般价值 4=较高价值 5=核心价值"),
        ("实现成本", "1=几乎无成本 2=较低成本 3=中等成本 4=较高成本 5=极高成本"),
        ("责任敞口", "1=无责任风险 2=轻微风险 3=中等风险 4=较高风险 5=极高风险"),
        ("综合评分", "业务价值 + 责任敞口，反映需求的重要性和紧迫性"),
        ("优先级", "最高:>=8分, 高:6-7分, 中:4-5分, 低:<4分"),
    ]

    for criterion, desc in criteria:
        ws.cell(row=row, column=1).value = criterion
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).fill = create_section_fill()
        ws.cell(row=row, column=1).alignment = wrap_alignment()
        ws.cell(row=row, column=1).border = create_thin_border()

        ws.merge_cells(f'B{row}:G{row}')
        ws.cell(row=row, column=2).value = desc
        ws.cell(row=row, column=2).font = Font(size=10)
        ws.cell(row=row, column=2).fill = create_white_fill()
        ws.cell(row=row, column=2).alignment = wrap_alignment(horizontal='left')
        ws.cell(row=row, column=2).border = create_thin_border()
        ws.row_dimensions[row].height = 18
        row += 1

    ws.freeze_panes = 'A6'
    set_column_widths(ws, [6, 30, 12, 12, 12, 10, 12])

    # ===== Sheet 2: 排序建议逻辑说明 =====
    ws2 = wb.create_sheet("排序建议逻辑说明")

    ws2.merge_cells('A2:C2')
    title_cell = ws2['A2']
    title_cell.value = "【工具三】需求优先级矩阵工具 - 排序建议逻辑说明"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws2.row_dimensions[2].height = 30

    headers = ["象限", "特征描述", "优先级建议"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws2.row_dimensions[4].height = 20

    quadrants = [
        ("Q1: 高价值-低敞口", "业务价值高，责任敞口小", "最高优先级", "00B050"),
        ("Q2: 高价值-高敞口", "业务价值高，责任敞口大", "高优先级（需谨慎）", "FFC000"),
        ("Q3: 低价值-低敞口", "业务价值低，责任敞口小", "中优先级", "A6A6A6"),
        ("Q4: 低价值-高敞口", "业务价值低，责任敞口大", "最低优先级（避免）", "FF0000"),
    ]

    for i, (quadrant, feature, suggestion, color) in enumerate(quadrants):
        row = 5 + i
        fill_color = PatternFill(start_color=color, end_color=color, fill_type='solid')

        ws2.cell(row=row, column=1).value = quadrant
        ws2.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws2.cell(row=row, column=1).fill = fill_color
        ws2.cell(row=row, column=1).alignment = wrap_alignment()
        ws2.cell(row=row, column=1).border = create_thin_border()

        ws2.cell(row=row, column=2).value = feature
        ws2.cell(row=row, column=2).font = Font(size=10)
        ws2.cell(row=row, column=2).fill = fill_color
        ws2.cell(row=row, column=2).alignment = wrap_alignment()
        ws2.cell(row=row, column=2).border = create_thin_border()

        ws2.cell(row=row, column=3).value = suggestion
        ws2.cell(row=row, column=3).font = Font(size=10)
        ws2.cell(row=row, column=3).fill = fill_color
        ws2.cell(row=row, column=3).alignment = wrap_alignment()
        ws2.cell(row=row, column=3).border = create_thin_border()

        ws2.row_dimensions[row].height = 25

    row = 10
    ws2.merge_cells(f'A{row}:C{row}')
    ws2[f'A{row}'].value = "责任敞口判断关键提问"
    ws2[f'A{row}'].font = Font(size=11, bold=True, color=WHITE)
    ws2[f'A{row}'].fill = create_header_fill()
    ws2[f'A{row}'].alignment = wrap_alignment()
    ws2.row_dimensions[row].height = 20
    row += 1

    questions = [
        "1. 如果这个需求实施失败，对业务的影响范围有多大？",
        "2. 谁会为这个需求的成功/失败承担责任？",
        "3. 这个需求是否涉及多个部门的协调？",
        "4. 实施这个需求需要多少资源？资源从哪里来？",
        "5. 如果需求变更，谁有最终决策权？",
    ]

    for q in questions:
        ws2.merge_cells(f'A{row}:C{row}')
        cell = ws2.cell(row=row, column=1)
        cell.value = q
        cell.font = Font(size=10)
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = create_thin_border()
        ws2.row_dimensions[row].height = 22
        row += 1

    ws2.freeze_panes = 'A5'
    set_column_widths(ws2, [20, 30, 18])

    # ===== Sheet 3: 优先级评分汇总表 =====
    ws3 = wb.create_sheet("优先级评分汇总表")

    ws3.merge_cells('A2:G2')
    title_cell = ws3['A2']
    title_cell.value = "【工具三】需求优先级矩阵工具 - 优先级评分汇总表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws3.row_dimensions[2].height = 30

    headers = ["序号", "需求/功能描述", "业务价值", "实现成本", "责任敞口", "综合评分", "优先级排序"]
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws3.row_dimensions[4].height = 20

    for i in range(15):
        row = 5 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws3.cell(row=row, column=1).value = i + 1
        ws3.cell(row=row, column=1).font = Font(size=10)
        ws3.cell(row=row, column=1).fill = fill
        ws3.cell(row=row, column=1).alignment = wrap_alignment()
        ws3.cell(row=row, column=1).border = create_thin_border()

        for col in range(2, 7):
            ws3.cell(row=row, column=col).value = ""
            ws3.cell(row=row, column=col).font = Font(size=10)
            ws3.cell(row=row, column=col).fill = create_input_fill()
            ws3.cell(row=row, column=col).alignment = wrap_alignment(horizontal='left')
            ws3.cell(row=row, column=col).border = create_thin_border()

        ws3.cell(row=row, column=7).value = f'=IF(F{row}="","",RANK(F{row},$F$5:$F$19,0))'
        ws3.cell(row=row, column=7).font = Font(size=10)
        ws3.cell(row=row, column=7).fill = fill
        ws3.cell(row=row, column=7).alignment = wrap_alignment()
        ws3.cell(row=row, column=7).border = create_thin_border()

        ws3.row_dimensions[row].height = 18

    ws3.freeze_panes = 'A5'
    set_column_widths(ws3, [6, 30, 10, 10, 10, 10, 12])

    # ===== Sheet 4: 决策记录表 =====
    ws4 = wb.create_sheet("决策记录表")

    ws4.merge_cells('A2:E2')
    title_cell = ws4['A2']
    title_cell.value = "【工具三】需求优先级矩阵工具 - 决策记录表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws4.row_dimensions[2].height = 30

    ws4.merge_cells('A4:E4')
    ws4['A4'].value = "决策会议信息"
    ws4['A4'].font = Font(size=11, bold=True)
    ws4['A4'].fill = create_section_fill()
    ws4['A4'].alignment = wrap_alignment()
    ws4.row_dimensions[4].height = 20

    meeting_info = [
        ("会议日期", ""),
        ("参会人员", ""),
        ("决策议题", ""),
        ("最终决策", ""),
    ]

    row = 5
    for label, value in meeting_info:
        ws4.cell(row=row, column=1).value = label
        ws4.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws4.cell(row=row, column=1).fill = create_section_fill()
        ws4.cell(row=row, column=1).alignment = wrap_alignment()
        ws4.cell(row=row, column=1).border = create_thin_border()

        ws4.merge_cells(f'B{row}:E{row}')
        ws4.cell(row=row, column=2).value = value
        ws4.cell(row=row, column=2).font = Font(size=10)
        ws4.cell(row=row, column=2).fill = create_input_fill()
        ws4.cell(row=row, column=2).alignment = wrap_alignment(horizontal='left')
        ws4.cell(row=row, column=2).border = create_thin_border()
        ws4.row_dimensions[row].height = 18
        row += 1

    row += 1
    ws4.merge_cells(f'A{row}:E{row}')
    ws4[f'A{row}'].value = "争议点记录"
    ws4[f'A{row}'].font = Font(size=11, bold=True)
    ws4[f'A{row}'].fill = create_section_fill()
    ws4[f'A{row}'].alignment = wrap_alignment()
    ws4.row_dimensions[row].height = 20
    row += 1

    dispute_headers = ["争议需求", "争议方A观点", "争议方B观点", "解决方案", "决议"]
    for col, header in enumerate(dispute_headers, start=1):
        cell = ws4.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(size=10, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws4.row_dimensions[row].height = 20
    row += 1

    for i in range(5):
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()
        for col in range(1, 6):
            cell = ws4.cell(row=row, column=col)
            cell.value = ""
            cell.font = Font(size=10)
            cell.fill = create_input_fill()
            cell.alignment = wrap_alignment(horizontal='left', vertical='center')
            cell.border = create_thin_border()
        ws4.row_dimensions[row].height = 25
        row += 1

    ws4.freeze_panes = 'A5'
    set_column_widths(ws4, [15, 20, 20, 12, 20])

    wb.save(os.path.join(OUTPUT_DIR, "03_需求优先级矩阵工具.xlsx"))
    print("Created: 03_需求优先级矩阵工具.xlsx")

def create_file_4():
    """04_案例库与复盘工具.xlsx"""
    wb = Workbook()

    # ===== Sheet 1: 案例记录表 =====
    ws = wb.active
    ws.title = "案例记录表"

    ws.merge_cells('A2:H2')
    title_cell = ws['A2']
    title_cell.value = "【工具四】案例库与复盘工具 - 案例记录表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws.row_dimensions[2].height = 30

    headers = ["序号", "行业", "项目规模", "项目类型", "表层需求", "事后判断的隐性需求", "被忽略的信号", "信号出现阶段"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws.row_dimensions[4].height = 20

    for i in range(20):
        row = 5 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws.cell(row=row, column=1).value = i + 1
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).alignment = wrap_alignment()
        ws.cell(row=row, column=1).border = create_thin_border()

        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)
            cell.value = ""
            cell.font = Font(size=10)
            cell.fill = create_input_fill()
            cell.alignment = wrap_alignment(horizontal='left', vertical='center')
            cell.border = create_thin_border()
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = 'A5'
    set_column_widths(ws, [6, 10, 10, 12, 20, 25, 20, 12])

    # ===== Sheet 2: 信号词典积累表 =====
    ws2 = wb.create_sheet("信号词典积累表")

    ws2.merge_cells('A2:D2')
    title_cell = ws2['A2']
    title_cell.value = "【工具四】案例库与复盘工具 - 信号词典积累表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws2.row_dimensions[2].height = 30

    headers = ["序号", "信号原话/典型表现", "通常含义", "建议的后续动作"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws2.row_dimensions[4].height = 20

    signals = [
        ("1", '"这个我们之前也尝试过"', "对过往失败经历有阴影，需要先解决信任问题", "分享成功案例，邀请参观"),
        ("2", '"再说吧，我们最近比较忙"', "需求不紧急或优先级不高", "了解真正的决策时机和触发点"),
        ("3", '"这个应该不难吧"', "对复杂度估计不足，或在试探你的反应", "展示类似项目的复杂度说明"),
        ("4", '"我们领导特别关注这个"', "政治因素浓厚，责任敞口大", "了解领导期望，评估政治风险"),
        ("5", '"能不能先出个方案看看"', "想获取免费方案，意向不明确", "要求明确需求和合作意向"),
        ("6", '"你们竞争对手怎么做的"', "在比较方案，需要差异化定位", "了解竞品，定位差异化优势"),
        ("7", '"我们预算可能有限"', "价格敏感，或在压价", "了解预算范围和价值认知"),
        ("8", '"这个需求其实提了很久了"', "长期痛点，有历史积累", "深挖痛点历史，评估紧迫性"),
    ]

    for i, (num, signal, meaning, action) in enumerate(signals):
        row = 5 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws2.cell(row=row, column=1).value = num
        ws2.cell(row=row, column=1).font = Font(size=10)
        ws2.cell(row=row, column=1).fill = fill
        ws2.cell(row=row, column=1).alignment = wrap_alignment()
        ws2.cell(row=row, column=1).border = create_thin_border()

        ws2.cell(row=row, column=2).value = signal
        ws2.cell(row=row, column=2).font = Font(size=10, bold=True, color=HEADER_BG)
        ws2.cell(row=row, column=2).fill = fill
        ws2.cell(row=row, column=2).alignment = wrap_alignment(horizontal='left')
        ws2.cell(row=row, column=2).border = create_thin_border()

        ws2.cell(row=row, column=3).value = meaning
        ws2.cell(row=row, column=3).font = Font(size=10)
        ws2.cell(row=row, column=3).fill = fill
        ws2.cell(row=row, column=3).alignment = wrap_alignment(horizontal='left')
        ws2.cell(row=row, column=3).border = create_thin_border()

        ws2.cell(row=row, column=4).value = action
        ws2.cell(row=row, column=4).font = Font(size=10)
        ws2.cell(row=row, column=4).fill = fill
        ws2.cell(row=row, column=4).alignment = wrap_alignment(horizontal='left')
        ws2.cell(row=row, column=4).border = create_thin_border()

        ws2.row_dimensions[row].height = 25

    for i in range(10):
        row = 13 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws2.cell(row=row, column=1).value = 9 + i
        ws2.cell(row=row, column=1).font = Font(size=10)
        ws2.cell(row=row, column=1).fill = fill
        ws2.cell(row=row, column=1).alignment = wrap_alignment()
        ws2.cell(row=row, column=1).border = create_thin_border()

        for col in range(2, 5):
            cell = ws2.cell(row=row, column=col)
            cell.value = ""
            cell.font = Font(size=10)
            cell.fill = create_input_fill()
            cell.alignment = wrap_alignment(horizontal='left', vertical='center')
            cell.border = create_thin_border()
        ws2.row_dimensions[row].height = 25

    ws2.freeze_panes = 'A5'
    set_column_widths(ws2, [6, 30, 30, 30])

    # ===== Sheet 3: 复盘分析表 =====
    ws3 = wb.create_sheet("复盘分析表")

    ws3.merge_cells('A2:D2')
    title_cell = ws3['A2']
    title_cell.value = "【工具四】案例库与复盘工具 - 复盘分析表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws3.row_dimensions[2].height = 30

    headers = ["复盘维度", "具体内容", "评分(1-5)", "改进行动"]
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws3.row_dimensions[4].height = 20

    review_dims = [
        "需求理解准确性", "方案设计匹配度", "客户沟通效果", "风险预判能力",
        "资源调配合理性", "时间进度控制", "团队协作效率", "客户满意度",
        "隐性需求挖掘深度", "价值传递效果"
    ]

    dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws3.add_data_validation(dv)

    for i, dim in enumerate(review_dims):
        row = 5 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws3.cell(row=row, column=1).value = dim
        ws3.cell(row=row, column=1).font = Font(size=10)
        ws3.cell(row=row, column=1).fill = fill
        ws3.cell(row=row, column=1).alignment = wrap_alignment(horizontal='left')
        ws3.cell(row=row, column=1).border = create_thin_border()

        ws3.cell(row=row, column=2).value = ""
        ws3.cell(row=row, column=2).font = Font(size=10)
        ws3.cell(row=row, column=2).fill = create_input_fill()
        ws3.cell(row=row, column=2).alignment = wrap_alignment(horizontal='left')
        ws3.cell(row=row, column=2).border = create_thin_border()

        ws3.cell(row=row, column=3).value = ""
        ws3.cell(row=row, column=3).font = Font(size=10)
        ws3.cell(row=row, column=3).fill = create_input_fill()
        ws3.cell(row=row, column=3).alignment = wrap_alignment()
        ws3.cell(row=row, column=3).border = create_thin_border()
        dv.add(f'C{row}')

        ws3.cell(row=row, column=4).value = ""
        ws3.cell(row=row, column=4).font = Font(size=10)
        ws3.cell(row=row, column=4).fill = create_input_fill()
        ws3.cell(row=row, column=4).alignment = wrap_alignment(horizontal='left')
        ws3.cell(row=row, column=4).border = create_thin_border()

        ws3.row_dimensions[row].height = 20

    row = 16
    ws3.merge_cells(f'A{row}:D{row}')
    ws3[f'A{row}'].value = "成功因素与风险点"
    ws3[f'A{row}'].font = Font(size=11, bold=True)
    ws3[f'A{row}'].fill = create_section_fill()
    ws3[f'A{row}'].alignment = wrap_alignment()
    ws3.row_dimensions[row].height = 20
    row += 1

    for i in range(2):
        ws3.cell(row=row, column=1).value = "成功因素" if i == 0 else "风险点"
        ws3.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws3.cell(row=row, column=1).fill = create_section_fill()
        ws3.cell(row=row, column=1).alignment = wrap_alignment()
        ws3.cell(row=row, column=1).border = create_thin_border()

        ws3.merge_cells(f'B{row}:D{row}')
        ws3.cell(row=row, column=2).value = ""
        ws3.cell(row=row, column=2).font = Font(size=10)
        ws3.cell(row=row, column=2).fill = create_input_fill()
        ws3.cell(row=row, column=2).alignment = wrap_alignment(horizontal='left')
        ws3.cell(row=row, column=2).border = create_thin_border()
        ws3.row_dimensions[row].height = 40
        row += 1

    ws3.freeze_panes = 'A5'
    set_column_widths(ws3, [18, 35, 10, 25])

    # ===== Sheet 4: 经验教训总结表 =====
    ws4 = wb.create_sheet("经验教训总结表")

    ws4.merge_cells('A2:E2')
    title_cell = ws4['A2']
    title_cell.value = "【工具四】案例库与复盘工具 - 经验教训总结表"
    title_cell.font = Font(size=13, bold=True, color=WHITE)
    title_cell.fill = create_title_fill()
    title_cell.alignment = wrap_alignment()
    ws4.row_dimensions[2].height = 30

    headers = ["序号", "经验/教训类别", "具体内容", "适用场景", "更新日期"]
    for col, header in enumerate(headers, start=1):
        cell = ws4.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(size=11, bold=True, color=WHITE)
        cell.fill = create_header_fill()
        cell.alignment = wrap_alignment()
        cell.border = create_thin_border()
    ws4.row_dimensions[4].height = 20

    lessons = [
        ("1", "需求挖掘", '客户说"我们很忙"往往意味着这不是真正的痛点', "初步接触阶段", "2024-01-15"),
        ("2", "信任建立", "分享同类成功案例比直接介绍产品更有效", "建立信任阶段", "2024-01-15"),
        ("3", "风险预判", "政治因素浓厚的项目需要提前评估责任敞口", "项目评估阶段", "2024-02-01"),
        ("4", "价值传递", "用客户自己的语言描述价值，而非技术术语", "价值呈现阶段", "2024-02-01"),
        ("5", "决策链", "必须识别真正的决策者和影响者", "需求挖掘阶段", "2024-02-10"),
        ("6", "验证设计", "最小可行验证实验比完美方案更能获得客户认可", "方案设计阶段", "2024-02-10"),
        ("7", "优先级判断", "综合评分高且责任敞口小的需求优先推进", "需求排序阶段", "2024-02-15"),
        ("8", "复盘机制", "每次访谈后立即复盘，避免遗忘关键细节", "访谈后复盘", "2024-02-15"),
    ]

    for i, (num, category, content, scenario, date) in enumerate(lessons):
        row = 5 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws4.cell(row=row, column=1).value = num
        ws4.cell(row=row, column=1).font = Font(size=10)
        ws4.cell(row=row, column=1).fill = fill
        ws4.cell(row=row, column=1).alignment = wrap_alignment()
        ws4.cell(row=row, column=1).border = create_thin_border()

        ws4.cell(row=row, column=2).value = category
        ws4.cell(row=row, column=2).font = Font(size=10)
        ws4.cell(row=row, column=2).fill = fill
        ws4.cell(row=row, column=2).alignment = wrap_alignment()
        ws4.cell(row=row, column=2).border = create_thin_border()

        ws4.cell(row=row, column=3).value = content
        ws4.cell(row=row, column=3).font = Font(size=10)
        ws4.cell(row=row, column=3).fill = fill
        ws4.cell(row=row, column=3).alignment = wrap_alignment(horizontal='left')
        ws4.cell(row=row, column=3).border = create_thin_border()

        ws4.cell(row=row, column=4).value = scenario
        ws4.cell(row=row, column=4).font = Font(size=10)
        ws4.cell(row=row, column=4).fill = fill
        ws4.cell(row=row, column=4).alignment = wrap_alignment()
        ws4.cell(row=row, column=4).border = create_thin_border()

        ws4.cell(row=row, column=5).value = date
        ws4.cell(row=row, column=5).font = Font(size=10)
        ws4.cell(row=row, column=5).fill = fill
        ws4.cell(row=row, column=5).alignment = wrap_alignment()
        ws4.cell(row=row, column=5).border = create_thin_border()

        ws4.row_dimensions[row].height = 20

    for i in range(10):
        row = 13 + i
        fill = create_white_fill() if i % 2 == 0 else create_alt_row_fill()

        ws4.cell(row=row, column=1).value = 9 + i
        ws4.cell(row=row, column=1).font = Font(size=10)
        ws4.cell(row=row, column=1).fill = fill
        ws4.cell(row=row, column=1).alignment = wrap_alignment()
        ws4.cell(row=row, column=1).border = create_thin_border()

        for col in range(2, 5):
            cell = ws4.cell(row=row, column=col)
            cell.value = ""
            cell.font = Font(size=10)
            cell.fill = create_input_fill()
            cell.alignment = wrap_alignment(horizontal='left')
            cell.border = create_thin_border()

        ws4.cell(row=row, column=5).value = ""
        ws4.cell(row=row, column=5).font = Font(size=10)
        ws4.cell(row=row, column=5).fill = create_input_fill()
        ws4.cell(row=row, column=5).alignment = wrap_alignment()
        ws4.cell(row=row, column=5).border = create_thin_border()

        ws4.row_dimensions[row].height = 20

    ws4.freeze_panes = 'A5'
    set_column_widths(ws4, [6, 15, 40, 18, 12])

    wb.save(os.path.join(OUTPUT_DIR, "04_案例库与复盘工具.xlsx"))
    print("Created: 04_案例库与复盘工具.xlsx")

def main():
    print(f"Creating Excel files in: {OUTPUT_DIR}")
    print("-" * 60)

    create_file_1()
    create_file_2()
    create_file_3()
    create_file_4()

    print("-" * 60)
    print("All files created successfully!")

    print("\nVerifying files:")
    files = os.listdir(OUTPUT_DIR)
    for f in sorted(files):
        if f.endswith('.xlsx'):
            path = os.path.join(OUTPUT_DIR, f)
            size = os.path.getsize(path)
            print(f"  {f} - {size:,} bytes")

if __name__ == "__main__":
    main()

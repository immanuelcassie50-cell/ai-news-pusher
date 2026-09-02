# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import datetime

# Color scheme
TITLE_BG = "c41e3a"  # Chinese red
TITLE_FONT_COLOR = "FFFFFF"
HEADER_BG = "f5f5f5"
HEADER_FONT_COLOR = "1a1a1a"
BORDER_COLOR = "dddddd"
ALT_ROW_COLOR = "fafafa"

# Create styles
def get_title_fill():
    return PatternFill(start_color=TITLE_BG, end_color=TITLE_BG, fill_type="solid")

def get_title_font():
    return Font(name='Microsoft YaHei', size=14, bold=True, color=TITLE_FONT_COLOR)

def get_header_fill():
    return PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")

def get_header_font():
    return Font(name='Microsoft YaHei', size=11, bold=True, color=HEADER_FONT_COLOR)

def get_cell_font(size=10):
    return Font(name='Microsoft YaHei', size=size, color="1a1a1a")

def get_border():
    thin = Side(style='thin', color=BORDER_COLOR)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def get_alt_fill():
    return PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")

def center_align(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

OUTPUT_DIR = "D:/新课开发/家庭教育/3、家庭学习环境系统设计实战指南/完整课程包/06_工具表单/"

# ============== FILE 1: F06_家长角色自评表.xlsx ==============
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "自评概览"

# Title row
ws1.merge_cells('A1:F1')
ws1['A1'] = "家长角色自评表"
ws1['A1'].fill = get_title_fill()
ws1['A1'].font = get_title_font()
ws1['A1'].alignment = center_align()
ws1.row_dimensions[1].height = 40

# Description row
ws1.merge_cells('A2:F2')
ws1['A2'] = "请根据过去一个月的情况，选择最符合的选项，计算总分后查看结果解读"
ws1['A2'].font = get_cell_font(10)
ws1['A2'].alignment = center_align(wrap=True)
ws1.row_dimensions[2].height = 30

# Table headers
headers = ["题号", "题目描述", "A（主动支持）", "B（有限介入）", "C（旁观等待）", "D（忽视缺失）"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()
ws1.row_dimensions[3].height = 30

# Questions data
questions = [
    "1. 孩子遇到学习困难时，我的第一反应是",
    "2. 我主动为孩子创造学习机会的频率是",
    "3. 孩子学习时，我在做什么",
    "4. 孩子取得进步时，我的回应方式是",
    "5. 我主动了解孩子兴趣的频率是",
    "6. 孩子使用AI工具时，我的参与程度是",
    "7. 我根据孩子反馈调整支持方式的频率是",
    "8. 我主动学习教育方法的频率是",
    "9. 家庭中有固定的学习时间和空间安排吗",
    "10. 孩子向我解释AI生成内容时，我的态度是",
    "11. 当孩子对某话题感兴趣时，我会",
    "12. 我对自身教育角色定位的清晰程度是"
]

# A/B/C/D options
options = [
    ["立即引导，耐心讲解", "偶尔协助，大多让孩子自己解决", "等孩子主动求助再说", "忽略或批评孩子的困难"],
    ["每周3次以上主动安排", "每周1-2次主动安排", "偶尔心血来潮安排一次", "从不主动安排"],
    ["陪伴在旁，适时指导", "在同一房间各做各的事", "在隔壁房间做家务", "在客厅看电视/玩手机"],
    ["具体表扬+引导下一步", "简单说好样的/真棒", "嗯一声继续忙", "指出还不够好的地方"],
    ["每周深入交流1次以上", "每月交流2-3次", "孩子主动说才听", "从不主动询问"],
    ["一起探索，引导思考", "在旁边看着，不打扰", "偶尔看一眼", "完全不关注"],
    ["每周调整1次以上", "每月调整1次", "很少根据反馈调整", "从不调整"],
    ["每周学习相关知识", "每月学习1-2次", "偶尔翻翻文章", "从不学习"],
    ["有固定安排并坚持", "有安排但经常变动", "有时有安排", "完全随意"],
    ["积极倾听，鼓励分享", "听一听但不太懂", "敷衍应付", "否定AI的价值"],
    ["创造机会深入探索", "买本书或报个课", "说你去学吧", "不关注"],
    ["非常清晰，知道该怎么做", "比较清晰，有大致方向", "有时迷茫", "完全不清楚"]
]

for row_idx, (q, opts) in enumerate(zip(questions, options), 4):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)

    ws1.cell(row=row_idx, column=1, value=row_idx-3).border = get_border()
    ws1.cell(row=row_idx, column=1).alignment = center_align()
    ws1.cell(row=row_idx, column=1).fill = fill

    ws1.cell(row=row_idx, column=2, value=q).border = get_border()
    ws1.cell(row=row_idx, column=2).alignment = left_align(wrap=True)
    ws1.cell(row=row_idx, column=2).fill = fill

    for col_idx, opt in enumerate(opts, 3):
        ws1.cell(row=row_idx, column=col_idx, value=opt).border = get_border()
        ws1.cell(row=row_idx, column=col_idx).alignment = left_align(wrap=True)
        ws1.cell(row=row_idx, column=col_idx).fill = fill

    ws1.row_dimensions[row_idx].height = 45

# Scoring explanation
ws1.merge_cells('A16:F16')
ws1['A16'] = "计分规则"
ws1['A16'].fill = get_header_fill()
ws1['A16'].font = get_header_font()
ws1['A16'].alignment = left_align()

scoring_text = "A=0分（主动支持型）  B=1分（有限介入型）  C=2分（旁观等待型）  D=3分（忽视缺失型）\n将12道题的得分相加，对照下方结果解读"
ws1.merge_cells('A17:F17')
ws1['A17'] = scoring_text
ws1['A17'].font = get_cell_font(10)
ws1['A17'].alignment = left_align(wrap=True)
ws1.row_dimensions[17].height = 40

# Result interpretation
ws1.merge_cells('A18:F18')
ws1['A18'] = "结果解读"
ws1['A18'].fill = get_header_fill()
ws1['A18'].font = get_header_font()
ws1['A18'].alignment = left_align()

interpretations = [
    "0-3分：教练型 — 你是一位优秀的学习教练，能主动支持孩子的AI学习之旅",
    "4-7分：混合型 — 你在支持和放手之间寻找平衡，可以继续优化",
    "8-12分：监工型 — 你可能过度介入或完全放手，建议增加主动参与",
    "13-18分：缺失型 — 需要系统学习家庭AI学习的理念和方法",
    "19-36分：严重缺失 — 建议从基础开始建立家庭学习支持系统"
]

for i, interp in enumerate(interpretations, 19):
    ws1.merge_cells(f'A{i}:F{i}')
    ws1[f'A{i}'] = interp
    ws1[f'A{i}'].font = get_cell_font(10)
    ws1[f'A{i}'].alignment = left_align()
    ws1.row_dimensions[i].height = 25

# Set column widths
ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 40
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 20
ws1.column_dimensions['F'].width = 20

wb1.save(OUTPUT_DIR + 'F06_家长角色自评表.xlsx')
print("Created F06_家长角色自评表.xlsx")

# ============== FILE 2: F07_三个行为替换实践记录.xlsx ==============
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "行为替换记录"

# Title
ws2.merge_cells('A1:G1')
ws2['A1'] = "三个行为替换实践记录"
ws2['A1'].fill = get_title_fill()
ws2['A1'].font = get_title_font()
ws2['A1'].alignment = center_align()
ws2.row_dimensions[1].height = 40

# Description
ws2.merge_cells('A2:G2')
ws2['A2'] = "21天习惯养成追踪表 — 每天记录一个行为的执行情况和孩子的反应"
ws2['A2'].font = get_cell_font(10)
ws2['A2'].alignment = center_align()
ws2.row_dimensions[2].height = 25

# Headers for the 3 behavior tracking
headers2 = ["日期", "行为1执行情况", "孩子反应", "行为2执行情况", "孩子反应", "行为3执行情况", "孩子反应"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()
ws2.row_dimensions[3].height = 30

# 21 days tracking
start_date = datetime.date(2024, 7, 1)
for day in range(1, 22):
    row = day + 3
    current_date = start_date + datetime.timedelta(days=day-1)
    fill = get_alt_fill() if day % 2 == 0 else PatternFill(fill_type=None)

    ws2.cell(row=row, column=1, value=str(current_date)).border = get_border()
    ws2.cell(row=row, column=1).alignment = center_align()
    ws2.cell(row=row, column=1).fill = fill

    for col in range(2, 8):
        ws2.cell(row=row, column=col, value="").border = get_border()
        ws2.cell(row=row, column=col).alignment = left_align(wrap=True)
        ws2.cell(row=row, column=col).fill = fill

    ws2.row_dimensions[row].height = 30

# Weekly summary section
ws2.merge_cells('A25:G25')
ws2['A25'] = "每周总结"
ws2['A25'].fill = get_header_fill()
ws2['A25'].font = get_header_font()
ws2['A25'].alignment = left_align()

weekly_headers = ["周次", "本周执行天数", "最困难的日期", "突破点", "下周改进计划"]
for col, header in enumerate(weekly_headers, 1):
    cell = ws2.cell(row=26, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()

for week in range(1, 4):
    row = 26 + week
    fill = get_alt_fill() if week % 2 == 0 else PatternFill(fill_type=None)
    ws2.cell(row=row, column=1, value=f"第{week}周").border = get_border()
    ws2.cell(row=row, column=1).alignment = center_align()
    ws2.cell(row=row, column=1).fill = fill
    for col in range(2, 6):
        ws2.cell(row=row, column=col, value="").border = get_border()
        ws2.cell(row=row, column=col).alignment = left_align(wrap=True)
        ws2.cell(row=row, column=col).fill = fill
    ws2.row_dimensions[row].height = 35

# Column widths
ws2.column_dimensions['A'].width = 12
for col in ['B', 'D', 'F']:
    ws2.column_dimensions[col].width = 22
for col in ['C', 'E', 'G']:
    ws2.column_dimensions[col].width = 15

wb2.save(OUTPUT_DIR + 'F07_三个行为替换实践记录.xlsx')
print("Created F07_三个行为替换实践记录.xlsx")

# ============== FILE 3: F08_家庭AI学习系统方案模板.xlsx ==============
wb3 = openpyxl.Workbook()

# Sheet1: 学习目标（SMART）
ws3_1 = wb3.active
ws3_1.title = "学习目标（SMART）"

ws3_1.merge_cells('A1:D1')
ws3_1['A1'] = "第一部分：学习目标（SMART原则）"
ws3_1['A1'].fill = get_title_fill()
ws3_1['A1'].font = get_title_font()
ws3_1['A1'].alignment = center_align()
ws3_1.row_dimensions[1].height = 35

ws3_1.merge_cells('A2:D2')
ws3_1['A2'] = "SMART原则：具体的(Specific)、可衡量(Measurable)、可达成(Achievable)、相关性(Relevant)、时限性(Time-bound)"
ws3_1['A2'].font = get_cell_font(9)
ws3_1['A2'].alignment = left_align(wrap=True)
ws3_1.row_dimensions[2].height = 30

smart_headers = ["目标要素", "填写内容", "示例参考", "检查要点"]
for col, header in enumerate(smart_headers, 1):
    cell = ws3_1.cell(row=3, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()

smart_data = [
    ["具体目标", "孩子在这学期要达到什么具体学习成果？", "数学应用题正确率提升至85%以上", "是否明确？别人能理解吗？"],
    ["衡量标准", "如何衡量目标是否达成？", "每月测试得分≥85，期末考试A等", "能否用数字衡量？"],
    ["可达成性", "基于孩子当前水平，这个目标现实吗？", "上学期正确率70%，提升15%合理", "跳一跳能够到吗？"],
    ["相关性", "这个目标对孩子的成长有什么帮助？", "培养逻辑思维，为初中打基础", "为什么现在重要？"],
    ["时间节点", "什么时候完成？分阶段里程碑是什么？", "学期末完成，每月底回顾进度", "有明确截止日期吗？"]
]

for row_idx, data in enumerate(smart_data, 4):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        ws3_1.cell(row=row_idx, column=col_idx, value=val).border = get_border()
        ws3_1.cell(row=row_idx, column=col_idx).alignment = left_align(wrap=True)
        ws3_1.cell(row=row_idx, column=col_idx).fill = fill
    ws3_1.row_dimensions[row_idx].height = 45

ws3_1.column_dimensions['A'].width = 15
ws3_1.column_dimensions['B'].width = 30
ws3_1.column_dimensions['C'].width = 35
ws3_1.column_dimensions['D'].width = 25

# Sheet2: 学习地图（时间+空间）
ws3_2 = wb3.create_sheet("学习地图（时间+空间）")

ws3_2.merge_cells('A1:E1')
ws3_2['A1'] = "第二部分：学习地图（时间+空间）"
ws3_2['A1'].fill = get_title_fill()
ws3_2['A1'].font = get_title_font()
ws3_2['A1'].alignment = center_align()
ws3_2.row_dimensions[1].height = 35

ws3_2.merge_cells('A2:E2')
ws3_2['A2'] = "规划每日/每周的学习时间段和物理空间，让学习成为可预期的惯例"
ws3_2['A2'].font = get_cell_font(10)
ws3_2['A2'].alignment = left_align()

# Time schedule table
ws3_2.merge_cells('A3:E3')
ws3_2['A3'] = "每日学习时间表"
ws3_2['A3'].fill = get_header_fill()
ws3_2['A3'].font = get_header_font()

time_headers = ["时间段", "时长", "学习内容", "AI工具辅助", "家长角色"]
for col, header in enumerate(time_headers, 1):
    cell = ws3_2.cell(row=4, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()

time_slots = [
    ["早晨 7:00-7:30", "30分钟", "英语朗读/单词记忆", "AI口语助手", "协助准备/监督"],
    ["放学后 16:00-17:00", "60分钟", "数学作业+错题复习", "AI解题检查", "答疑解惑"],
    ["晚餐后 19:00-19:30", "30分钟", "阅读/兴趣探索", "AI推荐阅读", "陪伴阅读"],
    ["睡前 20:30-21:00", "30分钟", "当日总结/明日计划", "AI生成总结", "倾听分享"]
]

for row_idx, data in enumerate(time_slots, 5):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        ws3_2.cell(row=row_idx, column=col_idx, value=val).border = get_border()
        ws3_2.cell(row=row_idx, column=col_idx).alignment = left_align(wrap=True)
        ws3_2.cell(row=row_idx, column=col_idx).fill = fill
    ws3_2.row_dimensions[row_idx].height = 35

# Weekly focus
ws3_2.merge_cells('A10:E10')
ws3_2['A10'] = "每周学习重点"
ws3_2['A10'].fill = get_header_fill()
ws3_2['A10'].font = get_header_font()

week_headers = ["星期", "重点领域", "具体安排", "AI工具", "家长任务"]
for col, header in enumerate(week_headers, 1):
    cell = ws3_2.cell(row=11, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()

weekdays = ["周一", "周二", "周三", "周四", "周五", "周末"]
for row_idx, day in enumerate(weekdays, 12):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    ws3_2.cell(row=row_idx, column=1, value=day).border = get_border()
    ws3_2.cell(row=row_idx, column=1).alignment = center_align()
    ws3_2.cell(row=row_idx, column=1).fill = fill
    for col in range(2, 6):
        ws3_2.cell(row=row_idx, column=col, value="").border = get_border()
        ws3_2.cell(row=row_idx, column=col).alignment = left_align(wrap=True)
        ws3_2.cell(row=row_idx, column=col).fill = fill
    ws3_2.row_dimensions[row_idx].height = 30

ws3_2.column_dimensions['A'].width = 12
ws3_2.column_dimensions['B'].width = 15
ws3_2.column_dimensions['C'].width = 30
ws3_2.column_dimensions['D'].width = 18
ws3_2.column_dimensions['E'].width = 15

# Sheet3: 工具矩阵（年龄段+任务类型）
ws3_3 = wb3.create_sheet("工具矩阵（年龄+任务）")

ws3_3.merge_cells('A1:F1')
ws3_3['A1'] = "第三部分：工具矩阵（年龄段+任务类型）"
ws3_3['A1'].fill = get_title_fill()
ws3_3['A1'].font = get_title_font()
ws3_3['A1'].alignment = center_align()
ws3_3.row_dimensions[1].height = 35

ws3_3.merge_cells('A2:F2')
ws3_3['A2'] = "根据孩子年龄和任务类型，选择合适的AI学习工具"
ws3_3['A2'].font = get_cell_font(10)
ws3_3['A2'].alignment = left_align()

tool_headers = ["年龄段", "任务类型", "推荐AI工具", "使用场景", "使用时长", "注意事项"]
for col, header in enumerate(tool_headers, 1):
    cell = ws3_3.cell(row=3, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()

tool_data = [
    ["6-8岁", "英语启蒙", "多邻国/AI口语伙伴", "每日15分钟游戏化学习", "≤20分钟/天", "以兴趣为主，不强迫"],
    ["6-8岁", "数学基础", "可汗学院儿童版", "趣味数学游戏", "≤20分钟/天", "保护视力，定时休息"],
    ["9-12岁", "作业辅导", "AI解题助手", "错题讲解/思路启发", "30-40分钟/天", "先思考后提示"],
    ["9-12岁", "阅读理解", "AI阅读伙伴", "书籍讨论/理解检查", "按章节需求", "避免代替思考"],
    ["9-12岁", "创意写作", "AI写作伙伴", "头脑风暴/初稿生成", "按项目需求", "保持孩子原创性"],
    ["13岁+", "自主学习", "AI学习教练", "学习规划/进度跟踪", "按需求调整", "培养独立能力"],
    ["13岁+", "深度探究", "AI研究助手", "课题研究/资料整理", "按项目需求", "核实AI生成内容"]
]

for row_idx, data in enumerate(tool_data, 4):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        ws3_3.cell(row=row_idx, column=col_idx, value=val).border = get_border()
        ws3_3.cell(row=row_idx, column=col_idx).alignment = left_align(wrap=True)
        ws3_3.cell(row=row_idx, column=col_idx).fill = fill
    ws3_3.row_dimensions[row_idx].height = 35

ws3_3.column_dimensions['A'].width = 10
ws3_3.column_dimensions['B'].width = 15
ws3_3.column_dimensions['C'].width = 20
ws3_3.column_dimensions['D'].width = 25
ws3_3.column_dimensions['E'].width = 15
ws3_3.column_dimensions['F'].width = 20

# Sheet4: 输出机制（回教法频率+输出类型）
ws3_4 = wb3.create_sheet("输出机制（回教法）")

ws3_4.merge_cells('A1:E1')
ws3_4['A1'] = "第四部分：输出机制（回教法频率+输出类型）"
ws3_4['A1'].fill = get_title_fill()
ws3_4['A1'].font = get_title_font()
ws3_4['A1'].alignment = center_align()
ws3_4.row_dimensions[1].height = 35

ws3_4.merge_cells('A2:E2')
ws3_4['A2'] = "回教法核心：每天让孩子把学到的教给其他人（包括家长、弟弟妹妹、甚至是AI伙伴）"
ws3_4['A2'].font = get_cell_font(10)
ws3_4['A2'].alignment = left_align(wrap=True)
ws3_4.row_dimensions[2].height = 30

output_headers = ["输出形式", "频率建议", "具体做法", "孩子收获", "实施要点"]
for col, header in enumerate(output_headers, 1):
    cell = ws3_4.cell(row=3, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()

output_data = [
    ["语言复述", "每天至少1次", "睡前让孩子讲述今天最重要的一课", "加深理解+语言表达", "认真倾听，积极提问"],
    ["动手演示", "每周2-3次", "让孩子演示某个解题过程或实验", "知识内化+自信心", "给予正向反馈"],
    ["书面记录", "每周1次", "写学习日记或绘制思维导图", "整理归纳能力", "不评判对错，鼓励表达"],
    ["家庭分享", "每周1次", "周末家庭会议上孩子做学习汇报", "成就感+责任感", "全家认真聆听"],
    ["教AI伙伴", "随时进行", "让孩子向AI解释自己学到的知识", "深化理解+检验掌握", "可以和孩子一起参与"]
]

for row_idx, data in enumerate(output_data, 4):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        ws3_4.cell(row=row_idx, column=col_idx, value=val).border = get_border()
        ws3_4.cell(row=row_idx, column=col_idx).alignment = left_align(wrap=True)
        ws3_4.cell(row=row_idx, column=col_idx).fill = fill
    ws3_4.row_dimensions[row_idx].height = 40

ws3_4.column_dimensions['A'].width = 12
ws3_4.column_dimensions['B'].width = 15
ws3_4.column_dimensions['C'].width = 30
ws3_4.column_dimensions['D'].width = 20
ws3_4.column_dimensions['E'].width = 18

# Sheet5: 角色承诺（家长具体行为）
ws3_5 = wb3.create_sheet("角色承诺（家长）")

ws3_5.merge_cells('A1:D1')
ws3_5['A1'] = "第五部分：角色承诺（家长具体行为）"
ws3_5['A1'].fill = get_title_fill()
ws3_5['A1'].font = get_title_font()
ws3_5['A1'].alignment = center_align()
ws3_5.row_dimensions[1].height = 35

ws3_5.merge_cells('A2:D2')
ws3_5['A2'] = "家长承诺的具体行为清单 — 每周对照检查执行情况"
ws3_5['A2'].font = get_cell_font(10)
ws3_5['A2'].alignment = left_align()

role_headers = ["承诺类别", "具体承诺（打勾即表示承诺）", "执行情况", "改进计划"]
for col, header in enumerate(role_headers, 1):
    cell = ws3_5.cell(row=3, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()

role_data = [
    ["时间投入", "□ 每天至少30分钟高质量陪伴", "□已完成  □部分  □未开始", ""],
    ["时间投入", "□ 每周2次与孩子一起探索AI工具", "□已完成  □部分  □未开始", ""],
    ["空间创造", "□ 家中设立固定学习角", "□已完成  □部分  □未开始", ""],
    ["空间创造", "□ 学习时间关闭电视/手机", "□已完成  □部分  □未开始", ""],
    ["情绪支持", "□ 孩子遇到困难时不批评，给予耐心", "□已完成  □部分  □未开始", ""],
    ["情绪支持", "□ 孩子教我时认真听，不敷衍", "□已完成  □部分  □未开始", ""],
    ["学习成长", "□ 我自己也学习AI基础知识", "□已完成  □部分  □未开始", ""],
    ["学习成长", "□ 每周复盘自己的教育方式", "□已完成  □部分  □未开始", ""],
    ["互动质量", "□ 每天问孩子一个有趣的问题", "□已完成  □部分  □未开始", ""],
    ["互动质量", "□ 鼓励孩子教我他们学到的东西", "□已完成  □部分  □未开始", ""]
]

for row_idx, data in enumerate(role_data, 4):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        ws3_5.cell(row=row_idx, column=col_idx, value=val).border = get_border()
        ws3_5.cell(row=row_idx, column=col_idx).alignment = left_align(wrap=True)
        ws3_5.cell(row=row_idx, column=col_idx).fill = fill
    ws3_5.row_dimensions[row_idx].height = 35

ws3_5.column_dimensions['A'].width = 12
ws3_5.column_dimensions['B'].width = 40
ws3_5.column_dimensions['C'].width = 25
ws3_5.column_dimensions['D'].width = 20

wb3.save(OUTPUT_DIR + 'F08_家庭AI学习系统方案模板.xlsx')
print("Created F08_家庭AI学习系统方案模板.xlsx")

# ============== FILE 4: F09_示例方案.xlsx ==============
wb4 = openpyxl.Workbook()

# Sheet1: 家庭背景信息区
ws4_1 = wb4.active
ws4_1.title = "家庭背景信息"

ws4_1.merge_cells('A1:C1')
ws4_1['A1'] = "示例：小明家庭 — 背景信息"
ws4_1['A1'].fill = get_title_fill()
ws4_1['A1'].font = get_title_font()
ws4_1['A1'].alignment = center_align()
ws4_1.row_dimensions[1].height = 35

info_headers = ["信息项", "填写内容", "说明"]
for col, header in enumerate(info_headers, 1):
    cell = ws4_1.cell(row=2, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()

info_data = [
    ["家庭成员", "爸爸（企业管理者）、妈妈（教师）、小明（10岁，四年级）", "核心家庭，父母都有稳定工作"],
    ["孩子年龄", "10岁", "处于培养学习习惯的关键期"],
    ["主要挑战", "1.数学应用题理解能力弱 2.英语口语表达不自信 3.做作业拖延", "需要针对性解决方案"],
    ["现有资源", "1台平板、1台电脑、稳定的网络", "硬件条件基本满足"],
    ["家长可投入时间", "每天约1.5小时（早晚各30分钟+周末2小时）", "时间有限，需高效利用"],
    ["期望目标", "一年内数学应用题正确率达到85%以上，英语口语流利度明显提升", "具体可衡量"]
]

for row_idx, data in enumerate(info_data, 3):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        ws4_1.cell(row=row_idx, column=col_idx, value=val).border = get_border()
        ws4_1.cell(row=row_idx, column=col_idx).alignment = left_align(wrap=True)
        ws4_1.cell(row=row_idx, column=col_idx).fill = fill
    ws4_1.row_dimensions[row_idx].height = 40

ws4_1.column_dimensions['A'].width = 15
ws4_1.column_dimensions['B'].width = 50
ws4_1.column_dimensions['C'].width = 25

# Sheet2: 五大部分完整填写示例
ws4_2 = wb4.create_sheet("完整方案示例")

ws4_2.merge_cells('A1:D1')
ws4_2['A1'] = "完整家庭AI学习系统方案 — 小明家庭示例"
ws4_2['A1'].fill = get_title_fill()
ws4_2['A1'].font = get_title_font()
ws4_2['A1'].alignment = center_align()
ws4_2.row_dimensions[1].height = 35

# Part 1: SMART Goals
ws4_2.merge_cells('A2:D2')
ws4_2['A2'] = "第一部分：学习目标"
ws4_2['A2'].fill = get_header_fill()
ws4_2['A2'].font = get_header_font()

goal_data = [
    ["目标要素", "具体内容"],
    ["具体目标", "数学应用题正确率从65%提升至85%以上，能独立分析题意并列出算式"],
    ["衡量标准", "每周数学测试应用题得分≥17/20，期末考试应用题部分≥90%"],
    ["可达成性", "小明目前基础较好，只是缺乏解题策略，通过系统训练可以达成"],
    ["相关性", "应用题是小学数学的核心考点，关系到初中数学学习"],
    ["时间节点", "本学期末（2024年1月）达成第一个里程碑，暑假前稳定在90%"]
]

for row_idx, data in enumerate(goal_data, 3):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        cell = ws4_2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = get_border()
        cell.alignment = left_align(wrap=True)
        cell.fill = fill
        if col_idx == 1:
            cell.font = get_header_font()
    ws4_2.row_dimensions[row_idx].height = 30

# Part 2: Learning Map
ws4_2.merge_cells('A10:D10')
ws4_2['A10'] = "第二部分：学习地图"
ws4_2['A10'].fill = get_header_fill()
ws4_2['A10'].font = get_header_font()

time_data = [
    ["时间段", "内容", "AI工具", "家长角色"],
    ["早晨 7:00-7:20", "英语跟读和口语练习", "AI口语伙伴（15分钟）", "监督发音，协助纠正"],
    ["放学后 16:30-17:30", "数学作业+应用题专项", "AI解题助手（启发式引导）", "先让孩子自己思考，再答疑"],
    ["晚餐后 19:00-19:20", "英语绘本阅读", "AI阅读伙伴", "一起讨论故事情节"],
    ["睡前 20:30-20:45", "当日学习复盘", "AI生成总结（孩子口述）", "认真倾听，给予肯定"]
]

for row_idx, data in enumerate(time_data, 11):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        cell = ws4_2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = get_border()
        cell.alignment = left_align(wrap=True)
        cell.fill = fill
        if row_idx == 11:
            cell.font = get_header_font()
    ws4_2.row_dimensions[row_idx].height = 30

# Part 3: Tool Matrix
ws4_2.merge_cells('A17:D17')
ws4_2['A17'] = "第三部分：工具矩阵"
ws4_2['A17'].fill = get_header_fill()
ws4_2['A17'].font = get_header_font()

tool_data = [
    ["任务类型", "推荐工具", "使用时长", "使用场景"],
    ["数学解题", "AI数学解题助手", "30分钟/天", "作业检查、错题讲解"],
    ["英语口语", "AI口语伙伴", "15-20分钟/天", "跟读练习、情景对话"],
    ["英语阅读", "AI阅读伙伴", "15-20分钟/天", "绘本讨论、词汇学习"],
    ["创意放松", "AI绘画/音乐", "周末30分钟", "兴趣探索、放松娱乐"]
]

for row_idx, data in enumerate(tool_data, 18):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        cell = ws4_2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = get_border()
        cell.alignment = left_align(wrap=True)
        cell.fill = fill
        if row_idx == 18:
            cell.font = get_header_font()
    ws4_2.row_dimensions[row_idx].height = 30

# Part 4: Output Mechanism
ws4_2.merge_cells('A23:D23')
ws4_2['A23'] = "第四部分：输出机制（回教法）"
ws4_2['A23'].fill = get_header_fill()
ws4_2['A23'].font = get_header_font()

output_data = [
    ["输出形式", "频率", "具体做法"],
    ["数学讲解", "每天晚餐", "小明每天晚餐时给全家讲解一道当天最难的数学题"],
    ["英语分享", "每周2次", "周三和周六晚上用英语分享一件当天发生的事"],
    ["周末展示", "每周六", "周末家庭会议上小明展示本周的学习成果和AI使用心得"]
]

for row_idx, data in enumerate(output_data, 24):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        cell = ws4_2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = get_border()
        cell.alignment = left_align(wrap=True)
        cell.fill = fill
        if row_idx == 24:
            cell.font = get_header_font()
    ws4_2.row_dimensions[row_idx].height = 30

# Part 5: Parent Commitment
ws4_2.merge_cells('A28:D28')
ws4_2['A28'] = "第五部分：家长承诺"
ws4_2['A28'].fill = get_header_fill()
ws4_2['A28'].font = get_header_font()

parent_data = [
    ["承诺类别", "具体承诺"],
    ["时间投入", "每天至少30分钟专注陪伴，不看手机"],
    ["情绪支持", "孩子讲题时认真听，不打断，耐心提问引导"],
    ["学习成长", "每周学习一点AI基础知识，与孩子共同进步"],
    ["互动质量", "每天真诚地感谢孩子教我东西"]
]

for row_idx, data in enumerate(parent_data, 29):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, val in enumerate(data, 1):
        cell = ws4_2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = get_border()
        cell.alignment = left_align(wrap=True)
        cell.fill = fill
        if row_idx == 29:
            cell.font = get_header_font()
    ws4_2.row_dimensions[row_idx].height = 30

ws4_2.column_dimensions['A'].width = 15
ws4_2.column_dimensions['B'].width = 35
ws4_2.column_dimensions['C'].width = 25
ws4_2.column_dimensions['D'].width = 20

# Sheet3: 填写思路解说区
ws4_3 = wb4.create_sheet("填写思路解说")

ws4_3.merge_cells('A1:B1')
ws4_3['A1'] = "填写思路解说 — 帮助理解如何填写自己的方案"
ws4_3['A1'].fill = get_title_fill()
ws4_3['A1'].font = get_title_font()
ws4_3['A1'].alignment = center_align()
ws4_3.row_dimensions[1].height = 35

guide_headers = ["部分", "填写思路与要点"]
for col, header in enumerate(guide_headers, 1):
    cell = ws4_3.cell(row=2, column=col, value=header)
    cell.fill = get_header_fill()
    cell.font = get_header_font()
    cell.alignment = center_align(wrap=True)
    cell.border = get_border()

guide_data = [
    ["学习目标", "1.从孩子最需要提升的地方入手，而非家长觉得重要的\n2.目标要具体可衡量，如'数学应用题正确率'比'数学好一点'好\n3.考虑孩子当前水平，设置跳一跳能够到的目标\n4.设定阶段性里程碑，便于跟踪和调整"],
    ["学习地图", "1.时间安排要切实可行，考虑家长工作和孩子作息\n2.每个时间段要明确做什么，不要太笼统\n3.AI工具的使用要嵌入具体场景，而非额外增加负担\n4.留出弹性空间，不要排得太满"],
    ["工具矩阵", "1.根据孩子年龄选择合适的工具，不是越多越好\n2.同一任务类型，精通1-2个工具即可\n3.明确每种工具的使用场景和时长建议\n4.定期评估工具效果，不适合就更换"],
    ["输出机制", "1.回教法核心：每天让孩子输出才能真正掌握\n2.从孩子愿意输出的形式开始，如讲解、演示、画画\n3.家长要真诚地当学生，让孩子有成就感\n4.每周有一次正式的展示机会最好"],
    ["家长承诺", "1.承诺要具体可执行，不要写空话\n2.从自己最容易做到的一条开始\n3.每周复盘执行情况，及时调整\n4.不要承诺太多，先做好2-3条"]
]

for row_idx, data in enumerate(guide_data, 3):
    fill = get_alt_fill() if row_idx % 2 == 0 else PatternFill(fill_type=None)
    ws4_3.cell(row=row_idx, column=1, value=data[0]).border = get_border()
    ws4_3.cell(row=row_idx, column=1).alignment = left_align(wrap=True)
    ws4_3.cell(row=row_idx, column=1).fill = fill
    ws4_3.cell(row=row_idx, column=1).font = get_header_font()

    ws4_3.cell(row=row_idx, column=2, value=data[1]).border = get_border()
    ws4_3.cell(row=row_idx, column=2).alignment = left_align(wrap=True)
    ws4_3.cell(row=row_idx, column=2).fill = fill
    ws4_3.row_dimensions[row_idx].height = 80

ws4_3.column_dimensions['A'].width = 15
ws4_3.column_dimensions['B'].width = 80

wb4.save(OUTPUT_DIR + 'F09_示例方案.xlsx')
print("Created F09_示例方案.xlsx")

print("\n=== All 4 files created successfully! ===")

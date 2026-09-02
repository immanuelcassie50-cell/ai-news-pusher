# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"D:\新课开发\变革管理-变革成果固化机制：防止新流程人走茶凉\完整课程包-工具表单"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SUBHEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
LABEL_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
LABEL_FONT = Font(name="微软雅黑", size=10, bold=True)
INPUT_FILL = PatternFill(start_color="FFFFFFF", end_color="FFFFFFF", fill_type="solid")
INPUT_FONT = Font(name="微软雅黑", size=10)
SECTION_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
SECTION_FONT = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
INSTRUCTION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INSTRUCTION_FONT = Font(name="微软雅黑", size=9, italic=True, color="7F6000")
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

def set_column_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def merge_and_style(ws, start_row, start_col, end_row, end_col, value, fill=None, font=None, alignment=None, border=None):
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = value
    if fill: cell.fill = fill
    if font: cell.font = font
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    return cell

def create_header_section(ws, title, module_name):
    ws.row_dimensions[1].height = 40
    merge_and_style(ws, 1, 1, 1, 8, title, fill=HEADER_FILL, font=Font(name="微软雅黑", size=18, bold=True, color="FFFFFF"), alignment=CENTER_ALIGN)
    ws.row_dimensions[2].height = 25
    merge_and_style(ws, 2, 1, 2, 8, module_name, fill=SUBHEADER_FILL, font=SUBHEADER_FONT, alignment=CENTER_ALIGN)

def create_basic_info_section(ws, fields, start_row=4):
    row = start_row
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "基本信息区", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    row += 1
    col = 1
    for field_name in fields:
        ws.row_dimensions[row].height = 25
        ws.row_dimensions[row + 1].height = 25
        cell_label = ws.cell(row=row, column=col, value=field_name)
        cell_label.fill = LABEL_FILL
        cell_label.font = LABEL_FONT
        cell_label.alignment = CENTER_ALIGN
        cell_label.border = THIN_BORDER
        cell_input = ws.cell(row=row + 1, column=col, value="")
        cell_input.fill = INPUT_FILL
        cell_input.font = INPUT_FONT
        cell_input.alignment = CENTER_ALIGN
        cell_input.border = THIN_BORDER
        col += 1
    return row + 2

def create_instruction_section(ws, instructions, start_row):
    row = start_row
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "填写说明", fill=INSTRUCTION_FILL, font=INSTRUCTION_FONT, alignment=LEFT_ALIGN)
    row += 1
    for i, instruction in enumerate(instructions):
        ws.row_dimensions[row].height = 20
        merge_and_style(ws, row, 1, row, 8, f"{i+1}. {instruction}", font=INSTRUCTION_FONT, alignment=LEFT_ALIGN)
        row += 1
    return row + 1

def create_table_header(ws, headers, start_row, start_col=1):
    ws.row_dimensions[start_row].height = 30
    for i, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

def create_mechanism_design_form():
    wb = Workbook()
    ws = wb.active
    ws.title = "固化机制设计表"
    set_column_widths(ws, [18, 15, 12, 15, 12, 18, 15, 15])
    create_header_section(ws, "固化机制设计表", "模块三：机制固化工具")
    
    basic_fields = ["项目名称", "设计人", "设计日期", "版本号", "审核人", "审核日期", "状态", "有效期"]
    create_basic_info_section(ws, basic_fields, start_row=4)
    
    row = 13
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "机制目标定义", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["目标类型", "具体描述", "衡量指标", "目标值", "现状值", "差距分析", "优先级", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    target_types = ["效率提升", "质量改进", "成本降低", "风险控制", "客户满意"]
    for i, t in enumerate(target_types):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 9):
            cell = ws.cell(row=row + i, column=col, value=t if col == 1 else "")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += len(target_types) + 1
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "自动检查节点设计", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["节点名称", "检查频率", "检查方式", "检查内容", "责任人", "辅助工具", "异常处理方式", "记录形式"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(6):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 9):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 7
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "升级机制", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["触发条件", "升级路径", "决策权限", "响应时限", "升级对象", "升级方式", "记录要求", "关闭条件"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(4):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 9):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 5
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "机制运行指标设计", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["指标名称", "指标定义", "计算公式", "数据来源", "采集频率", "目标值", "预警阈值", "责任人"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(6):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 9):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 7
    ws.row_dimensions[row].height = 25
    merge_and_style(ws, row, 1, row, 3, "自动汇总：", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    ws.cell(row=row, column=4, value="=COUNTA(B19:B23)")
    ws.cell(row=row, column=4).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws.cell(row=row, column=4).font = Font(name="微软雅黑", size=10, bold=True, color="375623")
    ws.cell(row=row, column=4).alignment = CENTER_ALIGN
    ws.cell(row=row, column=4).border = THIN_BORDER
    
    merge_and_style(ws, row, 5, row, 6, "节点数量", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    row += 2
    
    instructions = [
        "本表用于设计变革成果的固化机制，确保新流程和制度能够长期稳定运行。",
        "基本信息区：填写项目相关的基本信息，状态包括"草稿"、"待审核"、"已发布"。",
        "机制目标定义：明确机制要达成的目标，包括效率、质量、成本、风险等维度。",
        "自动检查节点：设计机制中的关键检查点，包括频率、方式、责任人和异常处理。",
        "升级机制：定义什么情况下需要升级，以及升级的路径和权限。",
        "运行指标：设计用于监控机制运行效果的具体指标和预警阈值。",
        "所有带边框的单元格为输入区域，绿色区域为自动计算区域，请勿手动修改。"
    ]
    create_instruction_section(ws, instructions, row)
    
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    wb.save(os.path.join(OUTPUT_DIR, "固化机制设计表.xlsx"))
    print("已创建: 固化机制设计表.xlsx")


def create_successor_plan_form():
    wb = Workbook()
    ws = wb.active
    ws.title = "继任者培养计划表"
    set_column_widths(ws, [15, 12, 10, 10, 15, 12, 15, 15, 15, 12])
    create_header_section(ws, "继任者培养计划表", "模块三：机制固化工具")
    
    basic_fields = ["项目名称", "制定人", "制定日期", "适用范围", "计划周期", "版本", "审核", "审批日期", "状态", "备注"]
    create_basic_info_section(ws, basic_fields, start_row=4)
    
    row = 14
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 10, "关键岗位清单", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["岗位名称", "现任者", "任职年限", "风险等级", "继任紧迫度", "储备状态", "预计可继任时间", "培养重点", "责任人", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(8):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 11):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 9
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 10, "继任者信息", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["候选人姓名", "储备阶段", "目标岗位", "能力现状", "能力差距分析", "培养方式", "培养周期", "阶段性评估", "认证结果", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    stages = ["初步识别", "定向培养", "强化训练", "认证考核", "正式上岗"]
    for i, stage in enumerate(stages):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 11):
            cell = ws.cell(row=row + i, column=col, value=stage if col == 2 else "")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += len(stages) + 1
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 10, "培养计划时间轴", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["阶段", "时间安排", "阶段性目标", "里程碑事件", "主要培训内容", "培养方式", "考核方式", "考核标准", "责任人", "状态"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(6):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 11):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 7
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 10, "风险预案", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["风险场景", "风险描述", "发生概率", "影响程度", "预警指标", "预防措施", "应急方案", "责任人", "触发条件", "更新日期"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(4):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 11):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 5
    ws.row_dimensions[row].height = 25
    merge_and_style(ws, row, 1, row, 2, "自动统计：", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    ws.cell(row=row, column=3, value="=COUNTA(B16:B23)")
    ws.cell(row=row, column=3).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws.cell(row=row, column=3).font = Font(name="微软雅黑", size=10, bold=True, color="375623")
    ws.cell(row=row, column=3).alignment = CENTER_ALIGN
    ws.cell(row=row, column=3).border = THIN_BORDER
    
    merge_and_style(ws, row, 4, row, 5, "岗位数量", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    ws.cell(row=row, column=6, value="=COUNTA(B27:B31)")
    ws.cell(row=row, column=6).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws.cell(row=row, column=6).font = Font(name="微软雅黑", size=10, bold=True, color="375623")
    ws.cell(row=row, column=6).alignment = CENTER_ALIGN
    ws.cell(row=row, column=6).border = THIN_BORDER
    
    merge_and_style(ws, row, 7, row, 8, "候选人数量", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    row += 2
    
    instructions = [
        "本表用于规划关键岗位的继任者培养，确保人才梯队建设的连续性。",
        "关键岗位清单：识别需要继任计划的关键岗位，风险等级根据岗位的重要性和继任难度评定。",
        "继任者信息：记录候选人的储备阶段，分为初步识别、定向培养、强化训练、认证考核、正式上岗五个阶段。",
        "培养计划时间轴：制定详细的时间安排和阶段性目标，明确每个阶段的里程碑事件。",
        "风险预案：针对可能影响继任计划的风险场景制定预防措施和应急方案。",
        "储备阶段列可使用下拉菜单选择对应阶段。",
        "所有绿色区域为自动计算区域，请勿手动修改。"
    ]
    create_instruction_section(ws, instructions, row)
    
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    
    wb.save(os.path.join(OUTPUT_DIR, "继任者培养计划表.xlsx"))
    print("已创建: 继任者培养计划表.xlsx")


def create_meeting_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "定期审视会议模板"
    set_column_widths(ws, [15, 20, 15, 15, 12, 35, 15, 15])
    create_header_section(ws, "定期审视会议模板", "模块三：机制固化工具")
    
    row = 4
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "会议基本信息", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    info_items = [
        ("会议名称", ""), ("会议编号", ""), ("会议类型", "□月度审视  □季度审视  □专项审视"),
        ("召开时间", ""), ("召开地点", ""), ("主持人", ""),
        ("记录人", ""), ("参会人", "")
    ]
    
    for i, (label, value) in enumerate(info_items):
        ws.row_dimensions[row].height = 25
        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.fill = LABEL_FILL
        cell_label.font = LABEL_FONT
        cell_label.alignment = CENTER_ALIGN
        cell_label.border = THIN_BORDER
        merge_and_style(ws, row, 2, row, 8, value, fill=INPUT_FILL, font=INPUT_FONT, alignment=LEFT_ALIGN, border=THIN_BORDER)
        row += 1
    
    row += 1
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "会议议程模板", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["序号", "议程内容", "时长(分钟)", "主持人", "发言人", "所需资料", "讨论要点", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    agenda_items = ["开场与签到", "上次会议决议回顾", "固化效果审视", "异常情况分析", "改进措施讨论", "下次会议安排", "自由讨论", "总结与闭幕"]
    for i, item in enumerate(agenda_items):
        ws.row_dimensions[row + i].height = 25
        ws.cell(row=row + i, column=1, value=i+1).fill = INPUT_FILL
        ws.cell(row=row + i, column=1).font = INPUT_FONT
        ws.cell(row=row + i, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row + i, column=1).border = THIN_BORDER
        ws.cell(row=row + i, column=2, value=item).fill = INPUT_FILL
        ws.cell(row=row + i, column=2).font = INPUT_FONT
        ws.cell(row=row + i, column=2).alignment = LEFT_ALIGN
        ws.cell(row=row + i, column=2).border = THIN_BORDER
        for col in range(3, 9):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += len(agenda_items) + 1
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "审视内容清单", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["审视维度", "审视要点", "审视结果", "评分(1-5)", "存在问题", "改进建议", "责任部门", "完成时限"]
    create_table_header(ws, headers, row)
    
    row += 1
    dimensions = ["机制执行", "流程运作", "人员状态", "工具表单", "培训效果", "文化氛围"]
    for i, dim in enumerate(dimensions):
        ws.row_dimensions[row + i].height = 25
        ws.cell(row=row + i, column=1, value=dim).fill = INPUT_FILL
        ws.cell(row=row + i, column=1).font = INPUT_FONT
        ws.cell(row=row + i, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row + i, column=1).border = THIN_BORDER
        for col in range(2, 9):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += len(dimensions) + 1
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "决议事项记录", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["决议编号", "决议内容", "决议依据", "责任部门", "责任人", "完成标准", "完成时限", "状态"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(6):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 9):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 7
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "行动跟踪表", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["行动项", "负责人", "计划完成日", "实际完成日", "完成状态", "验证结果", "偏差分析", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(8):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 9):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 9
    ws.row_dimensions[row].height = 25
    merge_and_style(ws, row, 1, row, 2, "自动统计：", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    ws.cell(row=row, column=3, value="=COUNTA(B40:B45)")
    ws.cell(row=row, column=3).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws.cell(row=row, column=3).font = Font(name="微软雅黑", size=10, bold=True, color="375623")
    ws.cell(row=row, column=3).alignment = CENTER_ALIGN
    ws.cell(row=row, column=3).border = THIN_BORDER
    
    merge_and_style(ws, row, 4, row, 5, "决议数量", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    ws.cell(row=row, column=6, value="=COUNTA(B49:B56)")
    ws.cell(row=row, column=6).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws.cell(row=row, column=6).font = Font(name="微软雅黑", size=10, bold=True, color="375623")
    ws.cell(row=row, column=6).alignment = CENTER_ALIGN
    ws.cell(row=row, column=6).border = THIN_BORDER
    
    merge_and_style(ws, row, 7, row, 8, "行动项数量", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    row += 2
    
    instructions = [
        "本模板用于定期审视变革成果的固化状态，确保变革成果得到有效保持。",
        "会议基本信息：填写会议的基本信息，会议类型可在给出的选项中打勾选择。",
        "会议议程：标准议程可根据实际情况调整，"时长"列为建议时间，可根据需要修改。",
        "审视内容清单：从多个维度审视固化效果，评分采用1-5分制（1=很差，5=很好）。",
        "决议事项记录：记录会议形成的决议，明确责任部门和完成时限。",
        "行动跟踪表：跟踪会议部署的行动项的执行情况，确保决议得到落实。",
        "状态列可使用下拉菜单选择：□待开始  □进行中  □已完成  □已验证  □已关闭"
    ]
    create_instruction_section(ws, instructions, row)
    
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    
    wb.save(os.path.join(OUTPUT_DIR, "定期审视会议模板.xlsx"))
    print("已创建: 定期审视会议模板.xlsx")


def create_behavior_tracking_form():
    wb = Workbook()
    ws = wb.active
    ws.title = "行为习惯固化追踪表"
    set_column_widths(ws, [15, 35, 12, 12, 15, 35, 12, 15, 15, 12, 12])
    create_header_section(ws, "行为习惯固化追踪表", "模块四：文化固化工具")
    
    basic_fields = ["项目名称", "追踪对象", "追踪开始日期", "追踪周期", "目标习惯数量", "负责人", "版本", "审核人", "状态", "备注", "编制日期"]
    create_basic_info_section(ws, basic_fields, start_row=4)
    
    row = 15
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 11, "目标行为定义", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["序号", "行为描述", "期望频率", "触发场景", "执行标准", "衡量方式", "关联指标", "重要性", "当前状态", "开始日期", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(8):
        ws.row_dimensions[row + i].height = 25
        ws.cell(row=row + i, column=1, value=i+1).fill = INPUT_FILL
        ws.cell(row=row + i, column=1).font = INPUT_FONT
        ws.cell(row=row + i, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row + i, column=1).border = THIN_BORDER
        for col in range(2, 12):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 9
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 11, "追踪周期设置", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["周期类型", "开始日期", "结束日期", "持续天数", "关键节点", "节点日期", "节点目标", "当前阶段", "阶段进度", "下一节点", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    periods = ["21天习惯期", "60天巩固期", "90天稳定期"]
    for i, period in enumerate(periods):
        ws.row_dimensions[row + i].height = 25
        ws.cell(row=row + i, column=1, value=period).fill = INPUT_FILL
        ws.cell(row=row + i, column=1).font = INPUT_FONT
        ws.cell(row=row + i, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row + i, column=1).border = THIN_BORDER
        for col in range(2, 12):
            if col == 4:
                cell = ws.cell(row=row + i, column=col, value=f"=C{row + i + 1}-B{row + i + 1}")
            else:
                cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL if col != 4 else PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += len(periods) + 1
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 11, "行为数据记录（追踪记录区）", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["记录日期", "行为1", "行为2", "行为3", "行为4", "行为5", "辅助提醒", "执行情况", "结果反馈", "问题记录", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(30):
        ws.row_dimensions[row + i].height = 20
        ws.cell(row=row + i, column=1, value=f"=DATE(2024,1,1)+{i}").fill = INPUT_FILL
        ws.cell(row=row + i, column=1).font = INPUT_FONT
        ws.cell(row=row + i, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row + i, column=1).border = THIN_BORDER
        for col in range(2, 12):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 31
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 11, "习惯形成评估（关键节点评估）", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["评估节点", "评估日期", "评估维度", "达标标准", "实际得分", "达标判断", "未达标原因", "改进措施", "下阶段计划", "评估人", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    evaluation_nodes = ["21天节点", "60天节点", "90天节点"]
    dimensions = ["执行率", "正确率", "主动性", "持续性", "总体评估"]
    node_row = row
    for ni, node in enumerate(evaluation_nodes):
        for di, dim in enumerate(dimensions):
            ws.row_dimensions[node_row].height = 25
            ws.cell(row=node_row, column=1, value=node if di == 0 else "").fill = INPUT_FILL
            ws.cell(row=node_row, column=1).font = INPUT_FONT
            ws.cell(row=node_row, column=1).alignment = CENTER_ALIGN
            ws.cell(row=node_row, column=1).border = THIN_BORDER
            ws.cell(row=node_row, column=2, value=f"=DATE(2024,1,1)+{21 if "21" in node else 60 if "60" in node else 90}").fill = INPUT_FILL
            ws.cell(row=node_row, column=2).font = INPUT_FONT
            ws.cell(row=node_row, column=2).alignment = CENTER_ALIGN
            ws.cell(row=node_row, column=2).border = THIN_BORDER
            ws.cell(row=node_row, column=3, value=dim).fill = INPUT_FILL
            ws.cell(row=node_row, column=3).font = INPUT_FONT
            ws.cell(row=node_row, column=3).alignment = CENTER_ALIGN
            ws.cell(row=node_row, column=3).border = THIN_BORDER
            for col in range(4, 12):
                cell = ws.cell(row=node_row, column=col, value="")
                cell.fill = INPUT_FILL
                cell.font = INPUT_FONT
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
            node_row += 1
    
    row = node_row + 1
    ws.row_dimensions[row].height = 25
    merge_and_style(ws, row, 1, row, 2, "自动统计：", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    ws.cell(row=row, column=3, value="=COUNTA(B17:B24)")
    ws.cell(row=row, column=3).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws.cell(row=row, column=3).font = Font(name="微软雅黑", size=10, bold=True, color="375623")
    ws.cell(row=row, column=3).alignment = CENTER_ALIGN
    ws.cell(row=row, column=3).border = THIN_BORDER
    
    merge_and_style(ws, row, 4, row, 5, "目标行为数量", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    row += 2
    
    instructions = [
        "本表用于追踪行为习惯的固化过程，通过持续追踪评估习惯是否已形成。",
        "目标行为定义：明确要追踪的目标行为，包括期望频率、触发场景和执行标准。",
        "追踪周期设置：分为21天习惯期、60天巩固期、90天稳定期三个阶段。",
        "行为数据记录：每日记录区域，可记录30天的执行数据，日期会自动计算。",
        "习惯形成评估：在21天、60天、90天三个关键节点进行评估，判断习惯是否形成。",
        "执行情况列建议使用下拉菜单：□完全执行 □部分执行 □未执行 □提前执行",
        "达标判断列建议使用下拉菜单：□达标 □基本达标 □未达标"
    ]
    create_instruction_section(ws, instructions, row)
    
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    wb.save(os.path.join(OUTPUT_DIR, "行为习惯固化追踪表.xlsx"))
    print("已创建: 行为习惯固化追踪表.xlsx")


def create_recognition_incentive_form():
    wb = Workbook()
    ws = wb.active
    ws.title = "表彰激励设计表"
    set_column_widths(ws, [15, 35, 12, 15, 15, 15, 15, 15, 15, 12])
    create_header_section(ws, "表彰激励设计表", "模块四：文化固化工具")
    
    basic_fields = ["项目名称", "设计人", "设计日期", "适用范围", "激励预算", "预算来源", "审核人", "批准人", "有效期", "版本"]
    create_basic_info_section(ws, basic_fields, start_row=4)
    
    row = 14
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 10, "表彰目标行为定义", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["序号", "目标行为描述", "行为类型", "期望频率", "重要性等级", "关联目标", "衡量指标", "达标标准", "记录方式", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(8):
        ws.row_dimensions[row + i].height = 25
        ws.cell(row=row + i, column=1, value=i+1).fill = INPUT_FILL
        ws.cell(row=row + i, column=1).font = INPUT_FONT
        ws.cell(row=row + i, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row + i, column=1).border = THIN_BORDER
        for col in range(2, 11):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 9
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 10, "激励层次设计", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["激励层次", "激励周期", "激励对象", "激励内容", "激励形式", "价值金额", "名额限制", "评选方式", "发放方式", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    incentive_levels = [
        ("即时激励", "实时/每日", "个人", "口头表扬、小红花、积分奖励", "精神+物质", "10-50", "无限制", "即时提名", "当场发放"),
        ("短期激励", "每周/每月", "个人/团队", "证书、奖品、现金奖励", "物质+精神", "100-500", "每期5-10%", "推荐+评选", "会议发放"),
        ("长期激励", "季度/年度", "个人/团队", "奖杯、奖金、晋升机会", "精神为主", "1000-5000", "每期1-5%", "综合评定", "仪式发放")
    ]
    for i, data in enumerate(incentive_levels):
        ws.row_dimensions[row + i].height = 30
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row + i, column=col, value=val)
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN if col != 4 else LEFT_ALIGN
            cell.border = THIN_BORDER
    
    row += len(incentive_levels) + 1
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 10, "表彰形式（物质+精神）", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["表彰类型", "物质激励内容", "物质价值(元)", "精神激励内容", "呈现场景", "参与范围", "制作成本", "数量/年", "负责部门", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    recognition_types = [
        ("证书类", "定制证书", "", "荣誉证书", "正式场合", "", "", "", ""),
        ("奖杯类", "定制奖杯", "", "年度贡献奖", "年度大会", "", "", "", ""),
        ("奖品类", "实用礼品", "", "最佳实践奖", "月度会议", "", "", "", ""),
        ("符号类", "徽章/胸针", "", "标兵徽章", "日常工作", "", "", "", ""),
        ("机会类", "培训机会", "", "晋升加分", "人才评估", "", "", "", "")
    ]
    for i, data in enumerate(recognition_types):
        ws.row_dimensions[row + i].height = 25
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row + i, column=col, value=val)
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += len(recognition_types) + 1
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 10, "评选标准与流程", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["评选阶段", "评选时间", "评选标准", "评选方式", "评选委员", "候选人数量", "获奖名额", "评选结果", "公示方式", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(5):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 11):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 6
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 10, "激励效果追踪", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["追踪周期", "目标行为", "基准值", "当前值", "变化幅度", "参与率", "满意度", "激励成本", "投入产出比", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(6):
        ws.row_dimensions[row + i].height = 25
        for col in range(1, 11):
            if col == 5:
                cell = ws.cell(row=row + i, column=col, value=f"=(D{row + i + 1}-C{row + i + 1})/C{row + i + 1}")
                cell.number_format = "0.0%"
            elif col == 9:
                cell = ws.cell(row=row + i, column=col, value=f"=IF(H{row + i + 1}>0,E{row + i + 1}/H{row + i + 1},"")")
            else:
                cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL if col not in [5, 9] else PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 7
    ws.row_dimensions[row].height = 25
    merge_and_style(ws, row, 1, row, 2, "自动统计：", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    ws.cell(row=row, column=3, value="=COUNTA(B16:B23)")
    ws.cell(row=row, column=3).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws.cell(row=row, column=3).font = Font(name="微软雅黑", size=10, bold=True, color="375623")
    ws.cell(row=row, column=3).alignment = CENTER_ALIGN
    ws.cell(row=row, column=3).border = THIN_BORDER
    
    merge_and_style(ws, row, 4, row, 5, "目标行为数量", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    row += 2
    
    instructions = [
        "本表用于设计行为表彰和激励机制，通过多元化的激励手段促进目标行为的固化。",
        "表彰目标行为定义：明确需要表彰的目标行为，包括行为类型（主动行为、被动行为、创新行为等）。",
        "激励层次设计：分为即时激励、短期激励、长期激励三个层次，激励力度递进。",
        "表彰形式：物质激励和精神激励相结合，包括证书、奖杯、奖品、符号和机会等类型。",
        "评选标准与流程：制定公正透明的评选标准和流程，确保激励的公平性。",
        "激励效果追踪：通过数据追踪激励的实际效果，包括行为变化、参与率和满意度等指标。",
        "价值金额和投入产出比为自动计算区域，请确保数据来源准确。"
    ]
    create_instruction_section(ws, instructions, row)
    
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    
    wb.save(os.path.join(OUTPUT_DIR, "表彰激励设计表.xlsx"))
    print("已创建: 表彰激励设计表.xlsx")


def create_story_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "故事沉淀模板"
    set_column_widths(ws, [15, 35, 20, 35, 20, 20, 35, 15])
    create_header_section(ws, "故事沉淀模板", "模块四：文化固化工具")
    
    basic_fields = ["项目名称", "故事主题", "收集日期", "讲故事人", "记录人", "审核人", "状态", "版本"]
    create_basic_info_section(ws, basic_fields, start_row=4)
    
    row = 12
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "故事主题分类", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    categories = [
        ("变革先锋", "展现变革先行者的事迹", "变革推动"),
        ("突破创新", "突破困境、创新解决方案", "创新实践"),
        ("团队协作", "跨部门协作、共同攻坚", "协作案例"),
        ("客户至上", "以客户为中心的感人故事", "服务典范"),
        ("持续改进", "精益求精、持续优化的故事", "改进案例"),
        ("传承发扬", "经验传承、人才培养的故事", "知识传承")
    ]
    
    headers = ["主题分类", "主题说明", "关键词", "代表故事标题", "收集状态", "使用场景", "备注"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i, data in enumerate(categories):
        ws.row_dimensions[row + i].height = 25
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row + i, column=col, value=val)
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN if col in [1, 3] else LEFT_ALIGN
            cell.border = THIN_BORDER
        ws.cell(row=row + i, column=5, value="□已收集 □待收集 □待挖掘").fill = INPUT_FILL
        ws.cell(row=row + i, column=5).font = INPUT_FONT
        ws.cell(row=row + i, column=5).alignment = CENTER_ALIGN
        ws.cell(row=row + i, column=5).border = THIN_BORDER
        ws.cell(row=row + i, column=6, value="").fill = INPUT_FILL
        ws.cell(row=row + i, column=6).border = THIN_BORDER
        ws.cell(row=row + i, column=7, value="").fill = INPUT_FILL
        ws.cell(row=row + i, column=7).border = THIN_BORDER
    
    row += len(categories) + 1
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "故事收集模板（STAR模型）", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    star_sections = [
        ("S - 背景Situation", "讲述故事发生的背景和环境
- 什么时间、地点、项目背景？
- 涉及哪些部门和人员？
- 当时面临的主要挑战是什么？"),
        ("T - 冲突Trigger", "描述故事中的核心冲突或转折点
- 遇到了什么困难或障碍？
- 为什么会觉得难以解决？
- 有没有关键的时间节点？"),
        ("A - 行动Action", "描述主人公采取的关键行动
- 做了什么样的决定和行动？
- 为什么选择这样做？
- 有没有创新或突破性的做法？"),
        ("R - 结果Result", "描述行动带来的结果和影响
- 最终取得了什么成果？
- 对团队/公司产生了什么影响？
- 有哪些可以量化的收益？"),
        ("E - 启示Enlightenment", "总结故事的启示和价值
- 这个故事体现了什么价值观？
- 可以提炼出哪些可复制的经验？
- 对其他人有什么建议？")
    ]
    
    for section_name, guidance in star_sections:
        ws.row_dimensions[row].height = 80
        ws.cell(row=row, column=1, value=section_name).fill = HEADER_FILL
        ws.cell(row=row, column=1).font = HEADER_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=2, value=guidance)
        cell.fill = INPUT_FILL
        cell.font = Font(name="微软雅黑", size=10, color="666666")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        row += 1
    
    row += 1
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "故事素材清单", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["故事编号", "故事标题", "故事类型", "主人公", "发生时间", "涉及部门", "核心关键词", "素材完整度"]
    create_table_header(ws, headers, row)
    
    row += 1
    for i in range(10):
        ws.row_dimensions[row + i].height = 25
        ws.cell(row=row + i, column=1, value=f"STORY-{i+1:03d}").fill = INPUT_FILL
        ws.cell(row=row + i, column=1).font = INPUT_FONT
        ws.cell(row=row + i, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row + i, column=1).border = THIN_BORDER
        for col in range(2, 9):
            cell = ws.cell(row=row + i, column=col, value="")
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += 11
    ws.row_dimensions[row].height = 30
    merge_and_style(ws, row, 1, row, 8, "传播渠道规划", fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT_ALIGN)
    
    row += 1
    headers = ["渠道类型", "具体渠道", "故事内容形式", "更新频率", "负责人", "审核人", "当前状态", "使用统计"]
    create_table_header(ws, headers, row)
    
    row += 1
    channels = [
        ("内部宣传", "企业公众号/内刊", "图文故事", "每月1-2篇", "", "", "□启用 □停用", ""),
        ("培训材料", "新员工培训/变革培训", "案例故事集", "每季度更新", "", "", "□启用 □停用", ""),
        ("会议分享", "月度/季度会议", "现场讲故事", "每次会议", "", "", "□启用 □停用", ""),
        ("文化墙", "办公区域文化墙", "故事海报", "每月更换", "", "", "□启用 □停用", ""),
        ("视频传播", "企业视频号/培训视频", "故事微视频", "每季度", "", "", "□启用 □停用", ""),
        ("荣誉墙", "线上荣誉墙", "故事+人物", "实时更新", "", "", "□启用 □停用", "")
    ]
    for i, channel_data in enumerate(channels):
        ws.row_dimensions[row + i].height = 25
        for col, val in enumerate(channel_data, 1):
            cell = ws.cell(row=row + i, column=col, value=val)
            cell.fill = INPUT_FILL
            cell.font = INPUT_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    
    row += len(channels) + 1
    ws.row_dimensions[row].height = 25
    merge_and_style(ws, row, 1, row, 2, "自动统计：", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    ws.cell(row=row, column=3, value="=COUNTA(B40:B49)")
    ws.cell(row=row, column=3).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws.cell(row=row, column=3).font = Font(name="微软雅黑", size=10, bold=True, color="375623")
    ws.cell(row=row, column=3).alignment = CENTER_ALIGN
    ws.cell(row=row, column=3).border = THIN_BORDER
    
    merge_and_style(ws, row, 4, row, 5, "故事素材数量", fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), font=Font(name="微软雅黑", size=10, bold=True, color="375623"), alignment=RIGHT_ALIGN)
    
    row += 2
    
    instructions = [
        "本模板用于收集和沉淀变革故事，通过真实的故事传承变革文化和经验。",
        "故事主题分类：划分故事的类型，便于分类管理和精准传播。",
        "STAR模型故事收集法：通过背景(S)、冲突(T)、行动(A)、结果(R)、启示(E)五要素完整讲述故事。",
        "填写STAR各部分时，请注意：",
        "  - 背景：提供足够的上下文信息，让听众理解故事的场景；",
        "  - 冲突：突出故事的戏剧性和挑战性，让人感受到故事的张力；",
        "  - 行动：重点描述有特色的行动和决策，体现主人公的智慧和勇气；",
        "  - 结果：用数据说话，量化成果，增强说服力；",
        "  - 启示：提炼普适性经验，让故事可复制、可借鉴。",
        "故事素材清单：记录已收集的故事素材，便于管理和查找使用。",
        "传播渠道规划：规划故事的多渠道传播，最大化故事的影响力。",
        "建议优先收集真实感强、有数据支撑、有启示价值的故事。"
    ]
    create_instruction_section(ws, instructions, row)
    
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    wb.save(os.path.join(OUTPUT_DIR, "故事沉淀模板.xlsx"))
    print("已创建: 故事沉淀模板.xlsx")


if __name__ == "__main__":
    print("开始创建变革管理工具表单...")
    print(f"输出目录: {OUTPUT_DIR}")
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    create_mechanism_design_form()
    create_successor_plan_form()
    create_meeting_template()
    create_behavior_tracking_form()
    create_recognition_incentive_form()
    create_story_template()
    
    print("
所有工具表单创建完成！")

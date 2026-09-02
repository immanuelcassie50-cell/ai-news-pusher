# -*- coding: utf-8 -*-
"""
供应链重生课程数据分析工具 - Excel生成脚本
使用openpyxl创建专业的数据分析Excel文件
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Protection
)
from openpyxl.utils import get_column_letter

# 输出路径
OUTPUT_DIR = r"D:\新课开发\供应链\AI版\01 供应链重生：从执行调度者到自主决策架构师的角色转型\成果demo"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "数据分析工具.xlsx")

# 确保目录存在
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 颜色定义
COLORS = {
    "dark_header": "1A1814",      # 深色标题背景
    "red_accent": "C8442A",       # 红色强调
    "light_bg": "F5F0E6",         # 浅色背景
    "white": "FFFFFF",
    "alt_row_1": "FFFFFF",         # 交替行颜色1
    "alt_row_2": "F5F0E6",        # 交替行颜色2 (同light_bg)
    "border": "CCCCCC",            # 边框颜色
    "text_dark": "1A1814",         # 深色文字
    "text_white": "FFFFFF",        # 白色文字
}


def create_styles():
    """创建所有样式定义"""
    thin_border = Border(
        left=Side(style='thin', color=COLORS["border"]),
        right=Side(style='thin', color=COLORS["border"]),
        top=Side(style='thin', color=COLORS["border"]),
        bottom=Side(style='thin', color=COLORS["border"])
    )

    styles = {
        # 标题样式 - 深色背景白字
        "title_dark": {
            "font": Font(name="微软雅黑", size=14, bold=True, color=COLORS["text_white"]),
            "fill": PatternFill(start_color=COLORS["dark_header"], end_color=COLORS["dark_header"], fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": thin_border
        },
        # 列标题样式 - 红色强调
        "header_red": {
            "font": Font(name="微软雅黑", size=11, bold=True, color=COLORS["text_white"]),
            "fill": PatternFill(start_color=COLORS["red_accent"], end_color=COLORS["red_accent"], fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": thin_border
        },
        # 普通单元格样式 - 浅色背景
        "cell_normal": {
            "font": Font(name="微软雅黑", size=10, color=COLORS["text_dark"]),
            "fill": PatternFill(start_color=COLORS["light_bg"], end_color=COLORS["light_bg"], fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": thin_border
        },
        # 交替行样式1
        "cell_alt1": {
            "font": Font(name="微软雅黑", size=10, color=COLORS["text_dark"]),
            "fill": PatternFill(start_color=COLORS["alt_row_1"], end_color=COLORS["alt_row_1"], fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": thin_border
        },
        # 交替行样式2
        "cell_alt2": {
            "font": Font(name="微软雅黑", size=10, color=COLORS["text_dark"]),
            "fill": PatternFill(start_color=COLORS["alt_row_2"], end_color=COLORS["alt_row_2"], fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": thin_border
        },
        # 数值单元格居右
        "cell_number": {
            "font": Font(name="微软雅黑", size=10, color=COLORS["text_dark"]),
            "fill": PatternFill(start_color=COLORS["light_bg"], end_color=COLORS["light_bg"], fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": thin_border
        },
        # 分组标题样式
        "section_header": {
            "font": Font(name="微软雅黑", size=11, bold=True, color=COLORS["text_white"]),
            "fill": PatternFill(start_color=COLORS["dark_header"], end_color=COLORS["dark_header"], fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": thin_border
        },
    }
    return styles


def apply_style(cell, style_dict):
    """将样式应用到单元格"""
    if style_dict.get("font"):
        cell.font = style_dict["font"]
    if style_dict.get("fill"):
        cell.fill = style_dict["fill"]
    if style_dict.get("alignment"):
        cell.alignment = style_dict["alignment"]
    if style_dict.get("border"):
        cell.border = style_dict["border"]


def set_column_widths(ws, widths):
    """设置列宽"""
    for col_idx, width in enumerate(widths, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


def create_sheet1_demand_analysis(wb, styles):
    """创建需求分析数据表"""
    ws = wb.create_sheet(title="01_需求分析数据")

    # 列宽设置
    set_column_widths(ws, [6, 10, 12, 10, 8, 14, 18, 16, 20, 18, 18])

    # 标题行
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "盛恒供应链集团 供应链规划部 - 培训需求分析数据"
    apply_style(title_cell, styles["title_dark"])
    ws.row_dimensions[1].height = 30

    # 列标题
    headers = [
        "序号", "姓名", "岗位", "入职年限", "学历",
        "AI工具使用经验(年)", "当前角色定位", "核心痛点分类",
        "期望收获", "培训担忧", "过往培训体验"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        apply_style(cell, styles["header_red"])
    ws.row_dimensions[2].height = 30

    # 学员数据
    students = [
        {"序号": 1, "姓名": "张明辉", "岗位": "计划员", "入职年限": 5, "学历": "本科",
         "AI工具使用经验(年)": 1, "当前角色定位": "执行调度者",
         "核心痛点分类": "角色转型焦虑", "期望收获": "掌握AI辅助决策方法",
         "培训担忧": "担心技术门槛高", "过往培训体验": "理论偏多"},
        {"序号": 2, "姓名": "李晓燕", "岗位": "调度员", "入职年限": 8, "学历": "硕士",
         "AI工具使用经验(年)": 2, "当前角色定位": "执行调度者",
         "核心痛点分类": "AI决策不透明", "期望收获": "理解AI决策逻辑",
         "培训担忧": "内容可能过时", "过往培训体验": "案例丰富"},
        {"序号": 3, "姓名": "王建国", "岗位": "采购主管", "入职年限": 12, "学历": "本科",
         "AI工具使用经验(年)": 0, "当前角色定位": "执行调度者",
         "核心痛点分类": "判断力不足", "期望收获": "提升决策判断力",
         "培训担忧": "时间不够用", "过往培训体验": "节奏太快"},
        {"序号": 4, "姓名": "陈思远", "岗位": "计划员", "入职年限": 3, "学历": "本科",
         "AI工具使用经验(年)": 3, "当前角色定位": "执行调度者",
         "核心痛点分类": "隐性因素识别", "期望收获": "识别供应链隐性风险",
         "培训担忧": "缺乏实战练习", "过往培训体验": "互动不足"},
        {"序号": 5, "姓名": "刘芳华", "岗位": "调度员", "入职年限": 6, "学历": "本科",
         "AI工具使用经验(年)": 1, "当前角色定位": "执行调度者",
         "核心痛点分类": "协同机制", "期望收获": "优化跨部门协同",
         "培训担忧": "无法持续跟进", "过往培训体验": "后续支持少"},
        {"序号": 6, "姓名": "赵文博", "岗位": "计划员", "入职年限": 4, "学历": "硕士",
         "AI工具使用经验(年)": 2, "当前角色定位": "执行调度者",
         "核心痛点分类": "技术恐惧", "期望收获": "克服AI使用恐惧",
         "培训担忧": "跟不上进度", "过往培训体验": "难度适中"},
        {"序号": 7, "姓名": "孙丽娜", "岗位": "采购专员", "入职年限": 2, "学历": "本科",
         "AI工具使用经验(年)": 0, "当前角色定位": "执行调度者",
         "核心痛点分类": "角色转型焦虑", "期望收获": "明确转型方向",
         "培训担忧": "担心学不会", "过往培训体验": "氛围好"},
        {"序号": 8, "姓名": "周伟强", "岗位": "调度主管", "入职年限": 15, "学历": "本科",
         "AI工具使用经验(年)": 1, "当前角色定位": "执行调度者",
         "核心痛点分类": "判断力不足", "期望收获": "掌握决策框架",
         "培训担忧": "内容不实用", "过往培训体验": "案例真实"},
        {"序号": 9, "姓名": "吴静怡", "岗位": "计划员", "入职年限": 7, "学历": "硕士",
         "AI工具使用经验(年)": 2, "当前角色定位": "执行调度者",
         "核心痛点分类": "AI决策不透明", "期望收获": "建立AI协作思维",
         "培训担忧": "担心效果不佳", "过往培训体验": "工具实用"},
        {"序号": 10, "姓名": "郑浩然", "岗位": "采购员", "入职年限": 1, "学历": "本科",
         "AI工具使用经验(年)": 0, "当前角色定位": "执行调度者",
         "核心痛点分类": "隐性因素识别", "期望收获": "识别异常早期信号",
         "培训担忧": "工作忙没时间", "过往培训体验": "在线课程效率高"},
    ]

    # 写入数据
    for row_idx, student in enumerate(students, start=3):
        style = styles["cell_alt1"] if row_idx % 2 == 1 else styles["cell_alt2"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = student.get(header, "")
            apply_style(cell, style)
        ws.row_dimensions[row_idx].height = 22

    # 冻结首行
    ws.freeze_panes = "A3"


def create_sheet2_tracking(wb, styles):
    """创建课程实施跟踪表"""
    ws = wb.create_sheet(title="02_课程实施跟踪表")

    # 列宽设置
    set_column_widths(ws, [6, 10, 10, 10, 10, 10, 10, 10, 10, 8, 18, 10, 10, 10])

    # 标题行
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = "课程实施跟踪表"
    apply_style(title_cell, styles["title_dark"])
    ws.row_dimensions[1].height = 30

    # 列标题
    headers = [
        "序号", "学员姓名", "模块1(15min)", "模块2(20min)", "模块3(20min)",
        "模块4(15min)", "模块5(10min)", "模块6练习", "模块7(10min)",
        "总分", "行为改变承诺", "1周后跟进", "2周后跟进", "1月后跟进"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        apply_style(cell, styles["header_red"])
    ws.row_dimensions[2].height = 35

    # 学员跟踪数据
    tracking_data = [
        {"序号": 1, "学员姓名": "张明辉", "模块1(15min)": 14, "模块2(20min)": 18,
         "模块3(20min)": 19, "模块4(15min)": 14, "模块5(10min)": 9,
         "模块6练习": 18, "模块7(10min)": 9, "总分": 101, "行为改变承诺": "每日使用AI工具辅助计划制定",
         "1周后跟进": "已开始使用", "2周后跟进": "形成习惯", "1月后跟进": "效率提升30%"},
        {"序号": 2, "学员姓名": "李晓燕", "模块1(15min)": 15, "模块2(20min)": 20,
         "模块3(20min)": 20, "模块4(15min)": 15, "模块5(10min)": 10,
         "模块6练习": 20, "模块7(10min)": 10, "总分": 110, "行为改变承诺": "用AI辅助调度决策",
         "1周后跟进": "尝试使用", "2周后跟进": "遇到问题", "1月后跟进": "已解决"},
        {"序号": 3, "学员姓名": "王建国", "模块1(15min)": 13, "模块2(20min)": 17,
         "模块3(20min)": 16, "模块4(15min)": 13, "模块5(10min)": 8,
         "模块6练习": 15, "模块7(10min)": 8, "总分": 90, "行为改变承诺": "建立决策检查清单",
         "1周后跟进": "正在制定", "2周后跟进": "完成初稿", "1月后跟进": "使用中"},
        {"序号": 4, "学员姓名": "陈思远", "模块1(15min)": 15, "模块2(20min)": 19,
         "模块3(20min)": 20, "模块4(15min)": 15, "模块5(10min)": 10,
         "模块6练习": 19, "模块7(10min)": 10, "总分": 108, "行为改变承诺": "建立风险预警机制",
         "1周后跟进": "方案完成", "2周后跟进": "试点运行", "1月后跟进": "推广中"},
        {"序号": 5, "学员姓名": "刘芳华", "模块1(15min)": 14, "模块2(20min)": 18,
         "模块3(20min)": 17, "模块4(15min)": 14, "模块5(10min)": 9,
         "模块6练习": 17, "模块7(10min)": 9, "总分": 98, "行为改变承诺": "优化跨部门协同流程",
         "1周后跟进": "梳理流程", "2周后跟进": "试行中", "1月后跟进": "效率提升"},
        {"序号": 6, "学员姓名": "赵文博", "模块1(15min)": 15, "模块2(20min)": 19,
         "模块3(20min)": 18, "模块4(15min)": 14, "模块5(10min)": 9,
         "模块6练习": 18, "模块7(10min)": 9, "总分": 102, "行为改变承诺": "每周复盘AI决策",
         "1周后跟进": "开始记录", "2周后跟进": "持续进行", "1月后跟进": "形成习惯"},
        {"序号": 7, "学员姓名": "孙丽娜", "模块1(15min)": 12, "模块2(20min)": 16,
         "模块3(20min)": 15, "模块4(15min)": 12, "模块5(10min)": 8,
         "模块6练习": 14, "模块7(10min)": 8, "总分": 85, "行为改变承诺": "从简单场景开始使用AI",
         "1周后跟进": "已开始", "2周后跟进": "扩大范围", "1月后跟进": "逐步深入"},
        {"序号": 8, "学员姓名": "周伟强", "模块1(15min)": 14, "模块2(20min)": 17,
         "模块3(20min)": 18, "模块4(15min)": 14, "模块5(10min)": 9,
         "模块6练习": 17, "模块7(10min)": 9, "总分": 98, "行为改变承诺": "培训下属使用AI",
         "1周后跟进": "备课中", "2周后跟进": "已完成", "1月后跟进": "团队应用"},
        {"序号": 9, "学员姓名": "吴静怡", "模块1(15min)": 15, "模块2(20min)": 20,
         "模块3(20min)": 19, "模块4(15min)": 15, "模块5(10min)": 10,
         "模块6练习": 19, "模块7(10min)": 10, "总分": 108, "行为改变承诺": "建立AI协作标准",
         "1周后跟进": "标准初稿", "2周后跟进": "讨论完善", "1月后跟进": "正式运行"},
        {"序号": 10, "学员姓名": "郑浩然", "模块1(15min)": 13, "模块2(20min)": 15,
         "模块3(20min)": 16, "模块4(15min)": 13, "模块5(10min)": 8,
         "模块6练习": 15, "模块7(10min)": 8, "总分": 88, "行为改变承诺": "识别异常早期信号",
         "1周后跟进": "学习中", "2周后跟进": "实践应用", "1月后跟进": "效果良好"},
    ]

    # 写入数据
    for row_idx, data in enumerate(tracking_data, start=3):
        style = styles["cell_alt1"] if row_idx % 2 == 1 else styles["cell_alt2"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = data.get(header, "")
            apply_style(cell, style)
        ws.row_dimensions[row_idx].height = 22

    # 冻结首行
    ws.freeze_panes = "A3"


def create_sheet3_effectiveness(wb, styles):
    """创建学习效果分析表"""
    ws = wb.create_sheet(title="03_学习效果分析表")

    # 列宽设置
    set_column_widths(ws, [12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 14])

    # 标题行
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "学习效果分析表 - 四层评估模型"
    apply_style(title_cell, styles["title_dark"])
    ws.row_dimensions[1].height = 30

    # 评估维度说明
    ws.merge_cells("A2:L2")
    desc_cell = ws["A2"]
    desc_cell.value = "反应层 → 学习层 → 行为层 → 结果层"
    apply_style(desc_cell, styles["cell_normal"])
    ws.row_dimensions[2].height = 20

    # 反应层
    ws.merge_cells("A3:D3")
    cell = ws["A3"]
    cell.value = "反应层评估"
    apply_style(cell, styles["section_header"])

    reaction_headers = ["指标", "满意度评分(1-5)", "实用性评分", "讲师评分"]
    for col_idx, header in enumerate(reaction_headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["header_red"])

    reaction_data = [
        ["张明辉", 4.5, 4.8, 5.0],
        ["李晓燕", 5.0, 5.0, 5.0],
        ["王建国", 4.0, 4.2, 4.5],
        ["陈思远", 4.8, 4.9, 4.8],
        ["刘芳华", 4.2, 4.5, 4.6],
        ["赵文博", 4.5, 4.6, 4.7],
        ["孙丽娜", 4.0, 4.0, 4.2],
        ["周伟强", 4.3, 4.4, 4.5],
        ["吴静怡", 4.8, 4.9, 4.9],
        ["郑浩然", 4.2, 4.3, 4.4],
    ]

    for row_idx, row_data in enumerate(reaction_data, start=5):
        style = styles["cell_alt1"] if row_idx % 2 == 1 else styles["cell_alt2"]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, style)

    # 添加参与度评分列
    ws.cell(row=4, column=4).value = "参与度评分"
    ws.cell(row=4, column=4).value = "讲师评分"

    # 学习层
    ws.merge_cells("A16:D16")
    cell = ws["A16"]
    cell.value = "学习层评估"
    apply_style(cell, styles["section_header"])

    learning_headers = ["指标", "前测分数", "后测分数", "分数提升"]
    for col_idx, header in enumerate(learning_headers, start=1):
        cell = ws.cell(row=17, column=col_idx)
        cell.value = header
        apply_style(cell, styles["header_red"])

    learning_data = [
        ["张明辉", 62, 88, "=C18-B18"],
        ["李晓燕", 75, 95, "=C19-B19"],
        ["王建国", 58, 78, "=C20-B20"],
        ["陈思远", 70, 92, "=C21-B21"],
        ["刘芳华", 65, 85, "=C22-B22"],
        ["赵文博", 68, 90, "=C23-B23"],
        ["孙丽娜", 55, 76, "=C24-B24"],
        ["周伟强", 72, 88, "=C25-B25"],
        ["吴静怡", 78, 96, "=C26-B26"],
        ["郑浩然", 60, 80, "=C27-B27"],
    ]

    for row_idx, row_data in enumerate(learning_data, start=18):
        style = styles["cell_alt1"] if row_idx % 2 == 1 else styles["cell_alt2"]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, style)

    # 及格率统计
    ws.cell(row=28, column=1).value = "及格率"
    ws.cell(row=28, column=2).value = "=COUNTIF(C18:C27,\">=60\")/10"
    apply_style(ws.cell(row=28, column=1), styles["cell_number"])
    apply_style(ws.cell(row=28, column=2), styles["cell_number"])

    # 行为层
    ws.merge_cells("F3:I3")
    cell = ws["F3"]
    cell.value = "行为层评估"
    apply_style(cell, styles["section_header"])

    behavior_headers = ["指标", "工具使用次数", "决策场景应用次数", "角色转变自评(1-5)"]
    for col_idx, header in enumerate(behavior_headers, start=6):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["header_red"])

    behavior_data = [
        ["张明辉", 45, 28, 4.2],
        ["李晓燕", 60, 35, 4.8],
        ["王建国", 30, 18, 3.5],
        ["陈思远", 55, 32, 4.5],
        ["刘芳华", 38, 22, 4.0],
        ["赵文博", 42, 25, 4.1],
        ["孙丽娜", 25, 15, 3.2],
        ["周伟强", 40, 24, 4.0],
        ["吴静怡", 58, 34, 4.7],
        ["郑浩然", 28, 16, 3.3],
    ]

    for row_idx, row_data in enumerate(behavior_data, start=5):
        style = styles["cell_alt1"] if row_idx % 2 == 1 else styles["cell_alt2"]
        for col_idx, value in enumerate(row_data, start=6):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, style)

    # 结果层
    ws.merge_cells("K3:L3")
    cell = ws["K3"]
    cell.value = "结果层评估"
    apply_style(cell, styles["section_header"])

    result_headers = ["指标", "决策准确率提升", "AI协作效率提升", "异常处理及时率"]
    for col_idx, header in enumerate(result_headers, start=11):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["header_red"])

    result_data = [
        ["张明辉", "15%", "20%", "18%"],
        ["李晓燕", "22%", "28%", "25%"],
        ["王建国", "10%", "12%", "10%"],
        ["陈思远", "20%", "25%", "22%"],
        ["刘芳华", "12%", "15%", "14%"],
        ["赵文博", "18%", "22%", "20%"],
        ["孙丽娜", "8%", "10%", "8%"],
        ["周伟强", "14%", "18%", "15%"],
        ["吴静怡", "25%", "30%", "28%"],
        ["郑浩然", "10%", "12%", "10%"],
    ]

    for row_idx, row_data in enumerate(result_data, start=5):
        style = styles["cell_alt1"] if row_idx % 2 == 1 else styles["cell_alt2"]
        for col_idx, value in enumerate(row_data, start=11):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, style)


def create_sheet4_competition(wb, styles):
    """创建竞品课程对比表"""
    ws = wb.create_sheet(title="04_竞品课程对比")

    # 列宽设置
    set_column_widths(ws, [14, 18, 14, 14, 14, 20])

    # 标题行
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "竞品课程对比分析"
    apply_style(title_cell, styles["title_dark"])
    ws.row_dimensions[1].height = 30

    # 列标题
    headers = ["对比维度", "本公司课程", "竞品A(公开课)", "竞品B(内训)", "竞品C(线上课)", "差异化优势"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        apply_style(cell, styles["header_red"])
    ws.row_dimensions[2].height = 30

    # 对比数据
    comparison_data = [
        ["课程定位", "从执行者到架构师的角色转型", "AI工具操作技能培训", "供应链管理能力提升", "AI基础知识普及", "聚焦角色转型而非工具操作"],
        ["目标学员", "供应链计划员、调度员、采购", "各行业计划人员", "大型企业管理层", "初入职场人士", "针对供应链从业者定制"],
        ["时长", "2天(90分钟×6模块)", "3天(每天6小时)", "2天(每天7小时)", "4周(每周2小时)", "高效集中培训"],
        ["教学方式", "案例驱动+AI实操", "理论讲授为主", "讲授+研讨", "视频+测验", "强调实战转化"],
        ["互动比例", "60%", "30%", "45%", "20%", "高互动设计"],
        ["工具支撑", "自有AI平台+企业系统", "通用AI工具", "无", "通用AI工具", "与企业系统深度集成"],
        ["案例真实性", "盛恒实际项目案例", "通用行业案例", "咨询项目案例", "虚构案例", "100%真实项目"],
        ["课后跟踪", "1月跟进+社群支持", "无", "3月跟进", "无", "系统化跟踪体系"],
        ["价格", "¥6800/人", "¥4800/人", "¥8800/人", "¥1200/人", "性价比最优"],
        ["独家亮点", "角色转型路径图+决策框架", "AI提示词模板", "方法论框架", "灵活便捷", "完整转型解决方案"],
    ]

    # 写入数据
    for row_idx, row_data in enumerate(comparison_data, start=3):
        style = styles["cell_alt1"] if row_idx % 2 == 1 else styles["cell_alt2"]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, style)
        ws.row_dimensions[row_idx].height = 28

    # 冻结首行
    ws.freeze_panes = "A3"


def main():
    """主函数 - 创建Excel文件"""
    print("=" * 60)
    print("供应链重生课程数据分析工具 - Excel生成")
    print("=" * 60)

    # 创建工作簿
    wb = Workbook()

    # 创建样式
    styles = create_styles()

    # 删除默认工作表
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 创建各工作表
    print("\n[1/4] 创建需求分析数据表...")
    create_sheet1_demand_analysis(wb, styles)

    print("[2/4] 创建课程实施跟踪表...")
    create_sheet2_tracking(wb, styles)

    print("[3/4] 创建学习效果分析表...")
    create_sheet3_effectiveness(wb, styles)

    print("[4/4] 创建竞品课程对比表...")
    create_sheet4_competition(wb, styles)

    # 保存文件
    print(f"\n保存文件至: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)

    # 验证文件
    if os.path.exists(OUTPUT_FILE):
        file_size = os.path.getsize(OUTPUT_FILE)
        print(f"[OK] 文件创建成功! (大小: {file_size:,} bytes)")
    else:
        print("[FAIL] 文件创建失败!")
        return False

    print("\n" + "=" * 60)
    print("完成! 所有工作表已创建:")
    print("  1. 01_需求分析数据")
    print("  2. 02_课程实施跟踪表")
    print("  3. 03_学习效果分析表")
    print("  4. 04_竞品课程对比")
    print("=" * 60)

    return True


if __name__ == "__main__":
    main()

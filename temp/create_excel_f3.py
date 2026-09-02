from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "日常观察记录表"

title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
subheader_font = Font(name='微软雅黑', size=10, bold=True)
normal_font = Font(name='微软雅黑', size=10)

title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
subheader_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, merge_to=None):
    if merge_to:
        ws.merge_cells(f'{chr(64+col)}{row}:{chr(64+merge_to)}{row}')
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell

# Title row
set_cell(ws, 1, 1, 'F3_日常观察记录表', title_font, title_fill, center_align, merge_to=7)
ws.row_dimensions[1].height = 40

# Subtitle
set_cell(ws, 2, 1, '管理者日常观察记录工具 — 及时捕捉员工状态信号，为管理决策提供依据', Font(name='微软雅黑', size=9, italic=True), light_fill, center_align, merge_to=7)
ws.row_dimensions[2].height = 25

# Section 1
set_cell(ws, 4, 1, '第一部分：观察信号清单', header_font, header_fill, center_align, merge_to=7)
ws.row_dimensions[4].height = 30

# 能力类信号
set_cell(ws, 5, 1, '能力类信号', subheader_font, subheader_fill, center_align, merge_to=2)
row = 6
for signal in ['□ 工作产出质量下降', '□ 同样的错误重复出现', '□ 完成任务的时间明显变长', '□ 回避某类工作任务', '□ 过度依赖他人确认', '□ 无法清晰表达工作进展']:
    set_cell(ws, row, 1, signal, normal_font, None, left_align, thin_border, merge_to=2)
    ws.row_dimensions[row].height = 22
    row += 1

# 态度类信号
set_cell(ws, row, 1, '态度类信号', subheader_font, subheader_fill, center_align, merge_to=2)
row += 1
for signal in ['□ 迟到早退频率增加', '□ 回避与上级沟通', '□ 对工作吐槽抱怨增多', '□ 参与团队活动积极性下降', '□ 工作以外事情谈得更多', '□ 情绪波动明显']:
    set_cell(ws, row, 1, signal, normal_font, None, left_align, thin_border, merge_to=2)
    ws.row_dimensions[row].height = 22
    row += 1

# 潜力类信号
set_cell(ws, row, 1, '潜力类信号', subheader_font, subheader_fill, center_align, merge_to=2)
row += 1
for signal in ['□ 主动承担额外任务', '□ 提出建设性改进建议', '□ 主动帮助同事', '□ 表现出学习成长的渴望', '□ 面对挑战迎难而上']:
    set_cell(ws, row, 1, signal, normal_font, None, left_align, thin_border, merge_to=2)
    ws.row_dimensions[row].height = 22
    row += 1

# Section 2: 日常观察记录表
row += 1
set_cell(ws, row, 1, '第二部分：日常观察记录表', header_font, header_fill, center_align, merge_to=7)
ws.row_dimensions[row].height = 30
header_row = row + 1

# Table headers
headers = ['日期', '观察对象', '观察场景', '观察到的信号', '判断', '初步分析', '后续跟进']
for col, header in enumerate(headers, 1):
    set_cell(ws, header_row, col, header, subheader_font, subheader_fill, center_align, thin_border)
ws.row_dimensions[header_row].height = 35

# Data rows
data_start = header_row + 1
for r in range(data_start, data_start + 10):
    for c in range(1, 8):
        fill = light_fill if c == 1 else None
        set_cell(ws, r, c, '', normal_font, fill, left_align, thin_border)
    ws.row_dimensions[r].height = 30

# Section 3: 示例
example_start = data_start + 11
set_cell(ws, example_start, 1, '第三部分：观察记录示例', header_font, header_fill, center_align, merge_to=7)
ws.row_dimensions[example_start].height = 30

examples = [
    ('示例一：能力类信号观察', 'section'),
    ('日期', '观察对象', '观察场景', '观察到的信号', '判断', '初步分析', '后续跟进'),
    ('2026/7/3', '李四', '需求评审会议', '被问常规问题都回答"回去再想想"', '能力', '缺乏方法论', '安排参加培训'),
    ('', '', '', '', '', '', ''),
    ('示例二：态度类信号观察', 'section'),
    ('日期', '观察对象', '观察场景', '观察到的信号', '判断', '初步分析', '后续跟进'),
    ('2026/7/5', '王五', '团队复盘会', '全程低头刷手机，表情冷漠', '态度', '可能失去兴趣', '安排非正式谈话'),
    ('', '', '', '', '', '', ''),
    ('示例三：潜力类信号观察', 'section'),
    ('日期', '观察对象', '观察场景', '观察到的信号', '判断', '初步分析', '后续跟进'),
    ('2026/7/7', '赵六', '线上故障处理', '凌晨2点主动排查，定位并修复', '潜力', '高潜力员工特征', '纳入后备人才计划'),
]

current_row = example_start + 1
for ex in examples:
    if len(ex) == 1 and ex[0] in ['示例一：能力类信号观察', '示例二：态度类信号观察', '示例三：潜力类信号观察']:
        set_cell(ws, current_row, 1, ex[0], subheader_font, subheader_fill, center_align, thin_border, merge_to=7)
        ws.row_dimensions[current_row].height = 28
        current_row += 1
    elif ex[0] == '':
        for c in range(1, 8):
            set_cell(ws, current_row, c, '', normal_font, None, left_align, thin_border)
        ws.row_dimensions[current_row].height = 25
        current_row += 1
    else:
        for col, val in enumerate(ex, 1):
            fill = light_fill if col == 1 else None
            set_cell(ws, current_row, col, val, normal_font, fill, left_align, thin_border)
        ws.row_dimensions[current_row].height = 28
        current_row += 1

# Section 4: 使用说明
usage_start = current_row + 1
set_cell(ws, usage_start, 1, '第四部分：使用说明', header_font, header_fill, center_align, merge_to=7)
ws.row_dimensions[usage_start].height = 30

usage_content = [
    ('使用频率建议', '每周记录3-5条；发现异常信号第一时间记录；每月底回顾识别趋势'),
    ('记录要点', '1. 客观描述：只记录看到的听到的\n2. 具体情境：说明在什么场景下\n3. 及时记录：尽量24小时内记录\n4. 关联分析：单条信号不构成结论'),
    ('隐私保护', '本表仅供管理者个人使用；涉及员工隐私的内容勿对外分享'),
]

current_row = usage_start + 1
for title, content in usage_content:
    set_cell(ws, current_row, 1, title, subheader_font, subheader_fill, center_align, thin_border, merge_to=2)
    set_cell(ws, current_row, 3, content, normal_font, None, left_align, thin_border, merge_to=7)
    ws.row_dimensions[current_row].height = 40
    current_row += 1

# Footer
footer_row = current_row + 1
set_cell(ws, footer_row, 1, '版本：V1.0  |  创建日期：2026年7月8日  |  适用范围：各级管理者', Font(name='微软雅黑', size=9, italic=True), None, center_align, merge_to=7)

# Column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 25
ws.column_dimensions['G'].width = 25

# Print setup
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = 9
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.print_area = f'A1:G{footer_row}'

output_path = 'D:/新课开发/经验萃取/语音记录转管理者赋能手册/完整课程包/工具表单/F3_日常观察记录表.xlsx'
wb.save(output_path)
print(f"Excel file created: {output_path}")

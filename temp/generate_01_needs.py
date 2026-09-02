import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "需求分析"

# Define colors
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

header_font = Font(name='Microsoft YaHei', bold=True, color="FFFFFF", size=11)
subheader_font = Font(name='Microsoft YaHei', bold=True, size=10)
normal_font = Font(name='Microsoft YaHei', size=10)
title_font = Font(name='Microsoft YaHei', bold=True, size=14)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:K1')
ws['A1'] = '01 · 需求分析 ·《普通人的积极心理学实操课》'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Section 1: Target Audience Analysis
ws['A3'] = '一、目标人群分析'
ws['A3'].font = Font(name='Microsoft YaHei', bold=True, size=11)
ws.merge_cells('A3:K3')

headers = ['序号', '人物', '年龄', '职业背景', '心理状态', '学习动机', '先备知识', '学习障碍', '上课形式', '学习期望', '备注']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Sample data
data = [
    ['1', '小慧', '32岁', '市场部中层\n（刚晋升）', '高功能焦虑\n表面稳内心耗竭', '想改变但抵触鸡汤\n希望看到有据可查的方法', '接触过碎片化心理学内容\n存在误解', '习惯性自我否定\n难以坚持新练习', '工作坊/线上', '学会情绪调节\n找到意义感', '高功能、低幸福感代表'],
    ['2', '老赵', '45岁', '中学教师', '稳定期意义感缺失\n日子过得去但没意思', '想找回生活热情\n对空洞感有困扰', '几乎无心理学背景\n可能有自学碎片', '工作坊/线上', '找到生活意义\n重拾热情', '稳定期意义感缺失代表'],
    ['3', '晓萱', '28岁', '自由职业\n（刚离职过渡期）', '转型焦虑\n自我怀疑', '想重新找到方向\n走出低谷', '接触过一些心理学\n但不系统', '自我怀疑严重\n习惯吸收知识回避练习', '线上为主', '重建自信\n设计新生活', '转型焦虑代表'],
    ['4', '张力', '38岁', 'IT项目经理', '慢性压力\n持续低电量', '想提升幸福感\n但不知道从哪里开始', '无专业背景\n接触过一些积极心理学', '生活忙碌难以坚持\n担心没时间', '工作坊', '掌握实用工具\n融入日常生活', '职场人类代表'],
    ['5', '王芳', '35岁', '全职妈妈', '育儿压力\n失去自我', '想找回自己\n不想只是"妈妈"', '自学过一些心理学\n但感觉不够落地', '碎片化学习多\n缺乏系统练习', '工作坊/线上', '找到自我价值感\n建立支持系统', '全职家长代表'],
]

for row_idx, row_data in enumerate(data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = light_fill
        else:
            cell.fill = white_fill

# Section 2: Learning Pain Points
ws['A11'] = '二、学习痛点分析'
ws['A11'].font = Font(name='Microsoft YaHei', bold=True, size=11)
ws.merge_cells('A11:K11')

pain_headers = ['类别', '痛点描述', '具体表现', '发生频率', '影响程度', '潜在原因', '课程应对策略', '优先级', '备注']
for col, header in enumerate(pain_headers, 1):
    cell = ws.cell(row=12, column=col, value=header)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

pain_data = [
    ['认知层面', '对积极心理学存在误解', '以为是"强迫正能量"或"心灵鸡汤"', '高', '严重', '自媒体碎片传播', '第一章专门破除误解', '⭐⭐⭐', ''],
    ['情感层面', '习惯性自我否定', '"我就是这样的人"', '高', '严重', '长期负向思维惯性', '第三章优势视角重塑', '⭐⭐⭐', ''],
    ['行为层面', '难以坚持练习', '三天打鱼两天晒网', '高', '中', '缺乏触发机制设计', '第六章Tiny Habits', '⭐⭐⭐', ''],
    ['认知层面', '知道但做不到', '"道理都懂，就是没用"', '高', '严重', '缺乏具体操作步骤', '每章工具+即时练习', '⭐⭐⭐', ''],
    ['情感层面', '对负面情绪的抗拒', '觉得有负面情绪就是"有问题"', '中', '中', '缺乏情绪教育', '第二章情绪颗粒度', '⭐⭐', ''],
    ['行为层面', '生活忙碌没时间', '"等我闲下来再说"', '中', '中', '没有最小可行方案', '最小幸福例程设计', '⭐⭐', ''],
    ['社交层面', '缺乏支持系统', '一个人默默努力', '中', '中', '不知道如何建立支持', '工作坊小组讨论', '⭐', ''],
]

for row_idx, row_data in enumerate(pain_data, 13):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = light_fill

# Section 3: Learning Expectations
ws['A21'] = '三、学习期望分析'
ws['A21'].font = Font(name='Microsoft YaHei', bold=True, size=11)
ws.merge_cells('A21:K21')

exp_headers = ['期望类型', '具体期望', '期望来源', '合理性评估', '课程可满足度', '满足方式', '课程定位', '学员类别', '备注']
for col, header in enumerate(exp_headers, 1):
    cell = ws.cell(row=22, column=col, value=header)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

exp_data = [
    ['认知期望', '了解积极心理学的真正含义', '被鸡汤伤害过', '合理', '高', '第一章正本清源', '基础价值', '全体', ''],
    ['技能期望', '学会具体可操作的情绪调节方法', '上过很多课但没用', '合理', '高', '每章工具+练习', '核心价值', '全体', ''],
    ['情感期望', '减少自我否定，建立自信', '长期低自尊', '合理但需调整', '中', '优势视角+认知重构', '延伸价值', '小慧/晓萱', ''],
    ['应用期望', '能在日常生活中真正用起来', '希望改变生活', '合理', '高', '第六章例程设计', '最终目标', '全体', ''],
    ['社交期望', '找到志同道合的支持群体', '感到孤立无援', '部分合理', '低', '工作坊小组机制', '附加价值', '晓萱/王芳', ''],
    ['效果期望', '学完立刻感觉变好', '对课程有不切实际的期待', '需调整', '低', '第一章建立合理预期', '预期管理', '张力/王芳', ''],
]

for row_idx, row_data in enumerate(exp_data, 23):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = thin_border

# Section 4: Course Positioning
ws['A30'] = '四、课程定位'
ws['A30'].font = Font(name='Microsoft YaHei', bold=True, size=11)
ws.merge_cells('A30:K30')

pos_headers = ['定位维度', '定位描述', '差异化价值', '对标课程', '竞争优势', '目标学员', '定价参考', '交付形式', '核心卖点', '备注']
for col, header in enumerate(pos_headers, 1):
    cell = ws.cell(row=31, column=col, value=header)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

pos_data = [
    ['市场定位', '非鸡汤的积极心理学实用课', '学术有据+实操落地', '各类"正能量课"', '去鸡汤化', '25-45岁普通人群', '999-1999元', '线上/工作坊', '"不灌鸡汤"', ''],
    ['内容定位', '工具导向的幸福实践课', '6个工具+1套系统', '纯理论课程', '工具+练习+例程', '想改变且愿练习', '—', '混合式', '"学了就能用"', ''],
    ['形式定位', '最小可行练习设计', '每天10分钟', '高强度训练营', '低门槛易坚持', '忙碌职场人', '—', '自学+工作坊', '"不占用大量时间"', ''],
]

for row_idx, row_data in enumerate(pos_data, 32):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = light_fill

# Column widths
col_widths = [8, 10, 12, 12, 15, 18, 20, 15, 12, 10, 15]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Row heights
for row in range(1, 36):
    ws.row_dimensions[row].height = 22

ws.row_dimensions[1].height = 30
for row in range(5, 10):
    ws.row_dimensions[row].height = 50
for row in range(13, 20):
    ws.row_dimensions[row].height = 35

wb.save('D:/新课开发/情绪与心理学/普通人积极心理学实操课/完成课程包/09_成果demo/01_Needs_Analysis.xlsx')
print("01_Needs_Analysis.xlsx created successfully")

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

PRIMARY = "2b2d42"
SECONDARY = "8d99ae"
ACCENT = "ef233c"
LIGHT = "edf2f4"
WHITE = "ffffff"

def style_header(cell, text, bg_color=PRIMARY, font_size=14):
    cell.value = text
    cell.font = Font(name='Microsoft YaHei', size=font_size, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_label(cell, text, bold=True):
    cell.value = text
    cell.font = Font(name='Microsoft YaHei', size=10, bold=bold)
    cell.alignment = Alignment(vertical='center', wrap_text=True)

def style_input(cell):
    cell.fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
    cell.alignment = Alignment(vertical='center', wrap_text=True)

OUTPUT_DIR = "D:/新课开发/导师和带教/导师带教/新版/AI时代版/完整课程包/06_全流程工具表单/Excel版"

# ============= FA06: AI工作流设定卡 =============
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AI工作流设定卡"
style_header(ws.cell(1, 1), "FA06 AI工作流设定卡", PRIMARY, 14)
ws.merge_cells('A1:E1')
ws.row_dimensions[1].height = 30

headers = ["介入时机", "使用工具", "具体操作", "输出形式", "我的安排"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(2, col)
    cell.value = h
    cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=SECONDARY, end_color=SECONDARY, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

phases = [
    ("会谈前准备", "千问Max", "输入上次谈话摘要，生成建议重点", "3-5个要点提示", ""),
    ("会谈中记录", "GET笔记", "开启实时转写", "完整文字记录", ""),
    ("会谈后分析", "千问Max", "输入转写文本，提炼进展和跟进点", "摘要+建议", ""),
    ("周期总结", "千问Max + WOKBUDDY", "整合全周期数据，生成评估报告", "结构化报告", ""),
]

for i, (phase, tool, op, output, arr) in enumerate(phases, 3):
    style_label(ws.cell(i, 1), phase)
    ws.cell(i, 2, tool).value = tool
    ws.cell(i, 3, op).value = op
    ws.cell(i, 4, output).value = output
    for j in range(1, 6):
        style_input(ws.cell(i, j))
    ws.row_dimensions[i].height = 30

wb.save(f"{OUTPUT_DIR}/FA06_AI工作流设定卡.xlsx")
print("FA06 created")

# ============= FA07: 会谈准备单 =============
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "会谈准备单"
style_header(ws.cell(1, 1), "FA07 会谈准备单", PRIMARY, 14)
ws.merge_cells('A1:D1')

basic = [("日期", ""), ("学徒姓名", ""), ("谈话时长", ""), ("上次的核心议题", "")]
for i, (label, val) in enumerate(basic, 2):
    style_label(ws.cell(i, 1), label)
    style_input(ws.cell(i, 2))
    ws.merge_cells(f'C{i}:D{i}')

prep_start = 6
style_header(ws.cell(prep_start, 1), "会谈前准备（AI辅助）", ACCENT, 11)
ws.merge_cells(f'A{prep_start}:D{prep_start}')

prep_items = [
    ("AI生成的上次谈话摘要", ""),
    ("本次核心议题（1-2个）", ""),
    ("建议追问方向", ""),
    ("正向反馈机会（如有）", ""),
]
for i, (label, val) in enumerate(prep_items, prep_start+1):
    style_label(ws.cell(i, 1), label)
    ws.merge_cells(f'B{i}:D{i}')
    style_input(ws.cell(i, 2))
    ws.row_dimensions[i].height = 25

notes_start = prep_start + len(prep_items) + 2
style_header(ws.cell(notes_start, 1), "导师个人准备笔记", ACCENT, 11)
ws.merge_cells(f'A{notes_start}:D{notes_start}')
for i in range(notes_start+1, notes_start+4):
    style_input(ws.cell(i, 1))
    ws.merge_cells(f'A{i}:D{i}')
    ws.row_dimensions[i].height = 25

wb.save(f"{OUTPUT_DIR}/FA07_会谈准备单.xlsx")
print("FA07 created")

# ============= FA08: 会谈记录提炼表 =============
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "会谈记录提炼表"
style_header(ws.cell(1, 1), "FA08 会谈记录提炼表", PRIMARY, 14)
ws.merge_cells('A1:D1')

basic = [("日期", ""), ("学徒姓名", ""), ("谈话时长", "")]
for i, (label, val) in enumerate(basic, 2):
    style_label(ws.cell(i, 1), label)
    style_input(ws.cell(i, 2))
    ws.merge_cells(f'C{i}:D{i}')

ai_start = 5
style_header(ws.cell(ai_start, 1), "AI分析结果（需人工核实）", ACCENT, 11)
ws.merge_cells(f'A{ai_start}:D{ai_start}')

analysis_items = [
    ("本次核心议题", ""),
    ("学徒主要进展", ""),
    ("潜在风险信号", ""),
    ("需要跟进的事项", ""),
]
for i, (label, val) in enumerate(analysis_items, ai_start+1):
    style_label(ws.cell(i, 1), label)
    ws.merge_cells(f'B{i}:D{i}')
    style_input(ws.cell(i, 2))
    ws.row_dimensions[i].height = 25

mentor_start = ai_start + len(analysis_items) + 2
style_header(ws.cell(mentor_start, 1), "导师核实与补充", PRIMARY, 11)
ws.merge_cells(f'A{mentor_start}:D{mentor_start}')

mentor_items = [
    ("AI分析是否准确？需要怎么调整？", ""),
    ("补充AI没有捕捉到的细节", ""),
    ("下次会谈重点", ""),
]
for i, (label, val) in enumerate(mentor_items, mentor_start+1):
    style_label(ws.cell(i, 1), label)
    ws.merge_cells(f'B{i}:D{i}')
    style_input(ws.cell(i, 2))
    ws.row_dimensions[i].height = 25

wb.save(f"{OUTPUT_DIR}/FA08_会谈记录提炼表.xlsx")
print("FA08 created")

# ============= FA09: 经验萃取工作表 =============
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "经验萃取工作表"
style_header(ws.cell(1, 1), "FA09 经验萃取工作表", PRIMARY, 14)
ws.merge_cells('A1:D1')
ws.row_dimensions[1].height = 30

basic = [("经验条目名称", ""), ("来源导师", ""), ("来源岗位", ""), ("适用场景", ""), ("适用岗位", "")]
for i, (label, val) in enumerate(basic, 2):
    style_label(ws.cell(i, 1), label)
    ws.merge_cells(f'B{i}:D{i}')
    style_input(ws.cell(i, 2))
    ws.row_dimensions[i].height = 22

sections = [
    ("情境描述", "什么情况下这个经验有价值？描述具体到其他导师可以对号入座", 80),
    ("判断逻辑", "核心判断标准是什么？有什么步骤或判断树？", 80),
    ("操作建议", "遇到这种情况，导师具体可以怎么做？给出1-4个步骤", 80),
    ("风险提示", "这条经验在什么情况下不适用？列出2-3种情况", 60),
]

current_row = 8
for section_name, section_desc, height in sections:
    style_header(ws.cell(current_row, 1), section_name, ACCENT, 11)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1

    cell = ws.cell(current_row, 1)
    cell.value = section_desc
    cell.font = Font(name='Microsoft YaHei', size=9, italic=True, color=SECONDARY)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1

    cell = ws.cell(current_row, 1)
    style_input(cell)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.row_dimensions[current_row].height = height
    current_row += 2

ov_start = current_row
style_header(ws.cell(ov_start, 1), "来源信息", PRIMARY, 11)
ws.merge_cells(f'A{ov_start}:D{ov_start}')
for i, label in enumerate(["原始描述日期", "审阅确认日期", "验证状态", ""], ov_start+1):
    style_label(ws.cell(i, 1), label)
    if i < ov_start + 3:
        style_input(ws.cell(i, 2))
        ws.merge_cells(f'B{i}:D{i}')

wb.save(f"{OUTPUT_DIR}/FA09_经验萃取工作表.xlsx")
print("FA09 created")

# ============= FA10: 经验条目质检清单 =============
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "经验条目质检清单"
style_header(ws.cell(1, 1), "FA10 经验条目质检清单", PRIMARY, 14)
ws.merge_cells('A1:E1')
ws.row_dimensions[1].height = 30

headers = ["检查项", "标准", "通过", "需修改", "备注"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(2, col)
    cell.value = h
    cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=SECONDARY, end_color=SECONDARY, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

check_items = [
    ("情境具体性", "描述具体到其他导师可以对号入座，不是泛泛的'因人而异'"),
    ("判断逻辑完整性", "有清晰的判断标准，能回答'我怎么知道什么时候用这条经验'"),
    ("操作建议可执行性", "拿到就能做，不是'要灵活判断'"),
    ("风险边界标注", "有明确的不适用情况，不是'这条经验万能'"),
    ("来源标注", "标注了来源导师和适用场景/岗位"),
    ("无个人特例", "没有把只在特定条件下有效的判断当成普遍规律"),
]

for i, (item, standard) in enumerate(check_items, 3):
    style_label(ws.cell(i, 1), item)
    ws.cell(i, 2, standard).value = standard
    ws.cell(i, 2).font = Font(name='Microsoft YaHei', size=9)
    ws.cell(i, 2).alignment = Alignment(wrap_text=True, vertical='center')
    for j in range(3, 6):
        style_input(ws.cell(i, j))
    ws.row_dimensions[i].height = 28

ov_start = len(check_items) + 4
style_label(ws.cell(ov_start, 1), "综合评价")
ws.merge_cells(f'B{ov_start}:E{ov_start}')
style_input(ws.cell(ov_start, 2))
ws.row_dimensions[ov_start].height = 30

style_label(ws.cell(ov_start+1, 1), "同行反馈")
ws.merge_cells(f'B{ov_start+1}:E{ov_start+1}')
style_input(ws.cell(ov_start+1, 2))
ws.row_dimensions[ov_start+1].height = 30

wb.save(f"{OUTPUT_DIR}/FA10_经验条目质检清单.xlsx")
print("FA10 created")

# ============= FA11: 工具包封面+行动计划 =============
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "工具包封面"
style_header(ws.cell(1, 1), "个人带教工具包", PRIMARY, 16)
ws.merge_cells('A1:D1')
ws.row_dimensions[1].height = 40

info = [("导师姓名", ""), ("所属部门", ""), ("学徒姓名", ""), ("工具包完成日期", ""), ("适用带教周期", "")]
for i, (label, val) in enumerate(info, 3):
    style_label(ws.cell(i, 1), label)
    ws.merge_cells(f'B{i}:D{i}')
    style_input(ws.cell(i, 2))
    ws.row_dimensions[i].height = 25

ck_start = 9
style_header(ws.cell(ck_start, 1), "工具包内容清单", ACCENT, 12)
ws.merge_cells(f'A{ck_start}:D{ck_start}')

items = [
    "FA01 学徒基线诊断卡",
    "FA02 三阶段带教路线图",
    "FA03 BEST反馈工作表",
    "FA04 BIA积极性反馈工作表",
    "FA05 五维度发展评估表",
    "FA06 AI工作流设定卡",
    "FA07 会谈准备单",
    "FA08 会谈记录提炼表",
    "FA09 经验萃取工作表",
    "FA10 经验条目质检清单",
]

for i, item in enumerate(items, ck_start+1):
    style_label(ws.cell(i, 1), f"□ {item}")
    ws.merge_cells(f'A{i}:D{i}')
    ws.row_dimensions[i].height = 20

ap_start = ck_start + len(items) + 2
style_header(ws.cell(ap_start, 1), "个人行动计划（3项承诺）", PRIMARY, 12)
ws.merge_cells(f'A{ap_start}:D{ap_start}')

for i in range(1, 4):
    cell = ws.cell(ap_start+i, 1)
    cell.value = f"承诺{i}:"
    cell.font = Font(name='Microsoft YaHei', size=10, bold=True)
    ws.merge_cells(f'B{ap_start+i}:D{ap_start+i}')
    style_input(ws.cell(ap_start+i, 2))
    ws.row_dimensions[ap_start+i].height = 25

so_start = ap_start + 5
style_label(ws.cell(so_start, 1), "导师签名：")
style_label(ws.cell(so_start+1, 1), "日期：")
style_input(ws.cell(so_start, 2))
style_input(ws.cell(so_start+1, 2))

wb.save(f"{OUTPUT_DIR}/FA11_工具包封面行动计划.xlsx")
print("FA11 created")

print("\nAll 11 Excel forms created successfully!")
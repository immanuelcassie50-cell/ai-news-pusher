# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def set_header_style(cell):
    cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def set_title_style(cell):
    cell.font = Font(name='微软雅黑', size=14, bold=True, color='2F5496')
    cell.alignment = Alignment(horizontal='left', vertical='center')

def set_subtitle_style(cell):
    cell.font = Font(name='微软雅黑', size=11, bold=True)
    cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

def set_input_style(cell):
    cell.font = Font(name='微软雅黑', size=10, color='0000FF')
    cell.alignment = Alignment(horizontal='center', vertical='center')

def set_formula_style(cell):
    cell.font = Font(name='微软雅黑', size=10, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center')

def set_label_style(cell):
    cell.font = Font(name='微软雅黑', size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

def thin_border():
    thin = Side(style='thin', color='BFBFBF')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def set_table_border(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = thin_border()

# ================== 工具01 ==================
def create_tool_01(wb):
    ws = wb.create_sheet('工具01-社会阶层自测')
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = '工具01：社会阶层认知自测表'
    set_title_style(ws['A1'])
    ws.row_dimensions[1].height = 30

    # Instruction
    ws.merge_cells('A2:G2')
    ws['A2'] = '【填写说明】三维度测评：经济资本、文化资本、社会资本。每维度5题，共15题。每题A=4分，B=3分，C=2分，D=1分。'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws.row_dimensions[2].height = 25

    # Headers
    headers = ['维度', '序号', '题目', 'A(4分)', 'B(3分)', 'C(2分)', 'D(1分)', '得分']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        set_header_style(cell)
    ws.row_dimensions[3].height = 25

    # Questions
    questions = [
        ('经济资本', 1, '您的家庭年收入水平属于？'),
        ('经济资本', 2, '您拥有的房产数量？'),
        ('经济资本', 3, '您的金融投资额度？'),
        ('经济资本', 4, '您家庭消费水平在社会哪个层次？'),
        ('经济资本', 5, '您拥有的固定资产（如车、贵重物品）？'),
        ('文化资本', 6, '您的最高学历？'),
        ('文化资本', 7, '您每年阅读的书籍数量？'),
        ('文化资本', 8, '您参加文化艺术活动的频率？'),
        ('文化资本', 9, '您家庭对教育的投入程度？'),
        ('文化资本', 10, '您具备哪些专业技能或资格证书？'),
        ('社会资本', 11, '您认识多少位可以提供帮助的朋友？'),
        ('社会资本', 12, '您加入了多少个社交组织/社团？'),
        ('社会资本', 13, '您家人/亲属的社会地位总体如何？'),
        ('社会资本', 14, '您在遇到困难时有多少人愿意帮助？'),
        ('社会资本', 15, '您通过社交网络获取信息的频率？'),
    ]

    for idx, (dimension, num, question) in enumerate(questions):
        row = 4 + idx
        ws.cell(row=row, column=1, value=dimension)
        ws.cell(row=row, column=2, value=num)
        ws.cell(row=row, column=3, value=question)

        # Radio buttons (represented as placeholder)
        for col in [4, 5, 6, 7]:
            cell = ws.cell(row=row, column=col, value='○')
            set_input_style(cell)

        # Score formula (sum of selected options * weight)
        # For simplicity, we'll create a scoring system
        score_col = 8
        ws.cell(row=row, column=score_col, value=f'=SUM(D{row}:G{row})*0')  # Placeholder formula

    set_table_border(ws, 3, 18, 1, 8)

    # Dimension totals
    ws.cell(row=20, column=1, value='维度得分汇总')
    ws.cell(row=20, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    ws.cell(row=21, column=1, value='经济资本总分：')
    ws.cell(row=21, column=3, value=f'=SUM(H4:H8)')
    ws.cell(row=21, column=3).font = Font(name='微软雅黑', size=11, bold=True, color='2F5496')

    ws.cell(row=22, column=1, value='文化资本总分：')
    ws.cell(row=22, column=3, value=f'=SUM(H9:H13)')
    ws.cell(row=22, column=3).font = Font(name='微软雅黑', size=11, bold=True, color='2F5496')

    ws.cell(row=23, column=1, value='社会资本总分：')
    ws.cell(row=23, column=3, value=f'=SUM(H14:H18)')
    ws.cell(row=23, column=3).font = Font(name='微软雅黑', size=11, bold=True, color='2F5496')

    ws.cell(row=24, column=1, value='综合得分：')
    ws.cell(row=24, column=3, value=f'=H21+H22+H23')
    ws.cell(row=24, column=3).font = Font(name='微软雅黑', size=12, bold=True, color='C00000')

    # Result interpretation
    ws.cell(row=26, column=1, value='【结果解读】')
    ws.cell(row=26, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    interpretations = [
        ('上层阶层', '总分45分以上，三项均≥13分', '拥有丰富的经济、文化和社会资本'),
        ('中上阶层', '总分36-44分，任一维度≥12分', '资本积累较好，有较强的社会流动潜力'),
        ('中产阶层', '总分25-35分，三项较为均衡', '温饱有余但安全感不足，面临向下流动压力'),
        ('底层阶层', '总分25分以下，任一维度≤6分', '资本匮乏，社会流动困难'),
    ]

    headers2 = ['类型', '判定标准', '特征描述']
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=27, column=col)
        cell.value = header
        set_header_style(cell)

    for idx, (t, c, d) in enumerate(interpretations):
        row = 28 + idx
        ws.cell(row=row, column=1, value=t)
        ws.cell(row=row, column=2, value=c)
        ws.cell(row=row, column=3, value=d)

    set_table_border(ws, 27, 31, 1, 3)

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12

# ================== 工具02 ==================
def create_tool_02(wb):
    ws = wb.create_sheet('工具02-阶层流动分析')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:E1')
    ws['A1'] = '工具02：阶层流动路径分析表'
    set_title_style(ws['A1'])
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:E2')
    ws['A2'] = '【填写说明】纵向分析三代人的资本变化，横向对比三种资本类型。资源占有情况（丰富/一般/匮乏）+ 流动判断（↑上升/↓下降/→持平）'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')

    headers = ['资本类型', '父辈', '自己', '子女', '变化趋势']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        set_header_style(cell)
    ws.row_dimensions[3].height = 25

    capital_types = ['经济资本', '文化资本', '社会资本']
    for idx, capital in enumerate(capital_types):
        row = 4 + idx
        ws.cell(row=row, column=1, value=capital)
        ws.cell(row=row, column=1).font = Font(name='微软雅黑', size=11, bold=True)
        ws.cell(row=row, column=2, value='资源：____\n判断：__')
        ws.cell(row=row, column=3, value='资源：____\n判断：__')
        ws.cell(row=row, column=4, value='资源：____\n判断：__')
        ws.cell(row=row, column=5, value='□向上流动\n□向下流动\n□保持稳定')

        for col in range(2, 6):
            ws.cell(row=row, column=col).font = Font(name='微软雅黑', size=10)
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    set_table_border(ws, 3, 6, 1, 5)

    # Summary section
    ws.cell(row=8, column=1, value='【综合分析】')
    ws.cell(row=8, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    ws.cell(row=9, column=1, value='流动关键节点：')
    ws.merge_cells('B9:E9')
    ws.cell(row=9, column=2, value='________________________________________________')

    ws.cell(row=10, column=1, value='阻碍/促进因素：')
    ws.merge_cells('B10:E10')
    ws.cell(row=10, column=2, value='________________________________________________')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18

# ================== 工具03 ==================
def create_tool_03(wb):
    ws = wb.create_sheet('工具03-家族拯救者识别')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:D1')
    ws['A1'] = '工具03：家族"拯救者"识别表'
    set_title_style(ws['A1'])

    ws.merge_cells('A2:D2')
    ws['A2'] = '【填写说明】识别家族网络中资源贡献最大、角色定位关键、影响机制明显的成员。'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')

    headers = ['识别维度', '具体表现', '是/否', '说明']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        set_header_style(cell)

    questions = [
        ('资源贡献', '是否在家族中提供经济支持（如借钱、资助购房）？'),
        ('资源贡献', '是否在家族中提供人脉资源（介绍工作、帮忙办事）？'),
        ('资源贡献', '是否在家族中提供信息资源（传递消息、提供建议）？'),
        ('角色定位', '是否在家族聚会中担任主持人/决策者角色？'),
        ('角色定位', '是否经常被家族成员求助/依赖？'),
        ('角色定位', '是否在家族冲突中充当调解人？'),
        ('影响机制', '其意见是否对家族决策有重大影响？'),
        ('影响机制', '其行为模式是否被家族其他成员模仿？'),
        ('影响机制', '其价值观念是否主导家族氛围？'),
    ]

    for idx, (dim, q) in enumerate(questions):
        row = 4 + idx
        ws.cell(row=row, column=1, value=dim)
        ws.cell(row=row, column=2, value=q)
        ws.cell(row=row, column=3, value='□是  □否')
        set_input_style(ws.cell(row=row, column=3))

    set_table_border(ws, 3, 12, 1, 4)

    ws.cell(row=14, column=1, value='【拯救者特征总结】')
    ws.cell(row=14, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    ws.cell(row=15, column=1, value='您的家族拯救者是谁？其最突出的三项特征：')
    ws.merge_cells('A15:D15')
    ws.cell(row=16, column=1, value='1. ____________________________________')
    ws.cell(row=17, column=1, value='2. ____________________________________')
    ws.cell(row=18, column=1, value='3. ____________________________________')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20

# ================== 工具04 ==================
def create_tool_04(wb):
    ws = wb.create_sheet('工具04-中产感觉评估')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:C1')
    ws['A1'] = '工具04：中产感觉评估表'
    set_title_style(ws['A1'])

    ws.cell(row=2, column=1, value='一、四个层次自评')
    ws.cell(row=2, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    headers = ['层次', '状态描述', '自评(打√)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        set_header_style(cell)

    levels = [
        ('第一层（最稳定）', '有房有车无贷款，子女教育金充足，退休有保障，医疗无担忧'),
        ('第二层（较稳定）', '有房有车略有贷款，子女教育基本覆盖，退休有一定储备'),
        ('第三层（不稳定）', '租房或房贷压力大，子女教育支出勉强，退休准备不足'),
        ('第四层（最不稳定）', '基本生存型消费，工作不稳定，担忧子女未来，随时可能返贫'),
    ]

    for idx, (level, desc) in enumerate(levels):
        row = 4 + idx
        ws.cell(row=row, column=1, value=level)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value='□')
        set_input_style(ws.cell(row=row, column=3))

    set_table_border(ws, 3, 7, 1, 3)

    ws.cell(row=9, column=1, value='二、参照系分析')
    ws.cell(row=9, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    headers2 = ['参照对象', '您的状况', '差距分析']
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=10, column=col)
        cell.value = header
        set_header_style(cell)

    comparisons = [
        ('与父辈同年龄时相比', '', '□更好  □相当  □更差'),
        ('与同龄同学/同事相比', '', '□更好  □相当  □更差'),
        ('与期望的生活状态相比', '', '□达标  □基本达标  □差距大'),
    ]

    for idx, (obj, status, diff) in enumerate(comparisons):
        row = 11 + idx
        ws.cell(row=row, column=1, value=obj)
        ws.cell(row=row, column=2, value=status)
        ws.cell(row=row, column=3, value=diff)

    set_table_border(ws, 10, 13, 1, 3)

    ws.cell(row=15, column=1, value='三、不安全感具体来源（请勾选）')
    ws.cell(row=15, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    sources = '''□职业发展天花板  □行业衰退/裁员风险  □房产贬值  □子女教育压力
□父母养老负担  □医疗费用担忧  □通货膨胀  □社会地位下滑  □其他：______'''

    ws.merge_cells('A16:C18')
    ws.cell(row=16, column=1, value=sources)
    ws.cell(row=16, column=1).font = Font(name='微软雅黑', size=10)
    ws.cell(row=16, column=1).alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 20

# ================== 工具05 ==================
def create_tool_05(wb):
    ws = wb.create_sheet('工具05-知识变现评估')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:E1')
    ws['A1'] = '工具05：知识变现能力评估'
    set_title_style(ws['A1'])

    ws.merge_cells('A2:E2')
    ws['A2'] = '【填写说明】评估四类知识分子的变现能力和市场化适应度。选择最符合您实际情况的选项。'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')

    headers = ['知识分子类型', '特征描述', '变现能力', '市场化适应度', '自评']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        set_header_style(cell)

    types = [
        ('学术型', '专注于理论研究，发表论文，参与学术项目', '★★☆', '较低', '□'),
        ('技术型', '具备专业技能，能够解决实际问题', '★★★', '较高', '□'),
        ('管理型', '具备组织协调能力，能够带领团队', '★★★', '高', '□'),
        ('创意型', '具备创新思维，能够创造新价值', '★★★', '高', '□'),
    ]

    for idx, (t, desc, ability, adapt, check) in enumerate(types):
        row = 4 + idx
        ws.cell(row=row, column=1, value=t)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=ability)
        ws.cell(row=row, column=4, value=adapt)
        ws.cell(row=row, column=5, value=check)

    set_table_border(ws, 3, 7, 1, 5)

    ws.cell(row=9, column=1, value='【变现能力详细评估】')
    ws.cell(row=9, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    headers2 = ['评估维度', '具体问题', '评分(1-5)']
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=10, column=col)
        cell.value = header
        set_header_style(cell)

    assessments = [
        ('专业技能', '您的专业技能在市场上的稀缺程度？', 3),
        ('变现渠道', '您有多少种将知识转化为收入的渠道？', 3),
        ('客户获取', '您获取目标客户的能力如何？', 3),
        ('品牌建设', '您个人品牌的知名度和美誉度？', 3),
        ('定价能力', '您对服务/产品的定价话语权？', 3),
    ]

    for idx, (dim, q, default_score) in enumerate(assessments):
        row = 11 + idx
        ws.cell(row=row, column=1, value=dim)
        ws.cell(row=row, column=2, value=q)
        score_cell = ws.cell(row=row, column=3, value=default_score)
        score_cell.font = Font(name='微软雅黑', size=11, bold=True, color='0000FF')

    set_table_border(ws, 10, 15, 1, 3)

    ws.cell(row=17, column=1, value='综合得分：')
    ws.cell(row=17, column=1).font = Font(name='微软雅黑', size=11, bold=True)
    ws.cell(row=17, column=2, value='=SUM(C11:C15)')
    ws.cell(row=17, column=2).font = Font(name='微软雅黑', size=14, bold=True, color='C00000')

    ws.cell(row=17, column=3, value='/25')
    ws.cell(row=17, column=3).font = Font(name='微软雅黑', size=14, bold=True)

    ws.cell(row=17, column=4, value='变现能力：')
    ws.cell(row=18, column=4, value='□强(20-25分)  □中(10-19分)  □弱(5-9分)')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10

# ================== 工具06 ==================
def create_tool_06(wb):
    ws = wb.create_sheet('工具06-平民生存评估')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:D1')
    ws['A1'] = '工具06：平民生存状况评估表'
    set_title_style(ws['A1'])

    ws.merge_cells('A2:D2')
    ws['A2'] = '【填写说明】评估最起码生存状况指标。符合条件填"是"，不符合填"否"。'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')

    headers = ['评估指标', '具体标准', '当前状态', '风险等级']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        set_header_style(cell)

    indicators = [
        ('收入水平', '月收入是否达到当地最低工资标准的1.5倍？'),
        ('居住条件', '是否有稳定的居所（租房也算）？'),
        ('医疗保障', '是否有基本医疗保险？'),
        ('子女教育', '子女是否能正常接受义务教育？'),
        ('食物保障', '是否能保证每天三餐营养均衡？'),
        ('衣物保障', '是否有足够的衣物应对四季变化？'),
        ('交通出行', '是否能承担基本的交通费用？'),
        ('通讯联络', '是否有手机且能正常缴费？'),
        ('社交联系', '是否能维持基本的社会交往？'),
        ('应急储备', '是否有相当于3个月生活费的储蓄？'),
    ]

    for idx, (ind, std) in enumerate(indicators):
        row = 4 + idx
        ws.cell(row=row, column=1, value=ind)
        ws.cell(row=row, column=2, value=std)
        ws.cell(row=row, column=3, value='□是  □否')
        ws.cell(row=row, column=4, value='□高  □中  □低')

    set_table_border(ws, 3, 13, 1, 4)

    ws.cell(row=15, column=1, value='【风险预警汇总】')
    ws.cell(row=15, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    ws.cell(row=16, column=1, value='高风险项：___________  中风险项：___________  低风险项：___________')

    ws.cell(row=18, column=1, value='综合评估：')
    ws.cell(row=18, column=1).font = Font(name='微软雅黑', size=11, bold=True)
    ws.cell(row=18, column=2, value='□基本安全  □存在风险（需关注）  □危机状态（需紧急干预）')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

# ================== 工具07 ==================
def create_tool_07(wb):
    ws = wb.create_sheet('工具07-农民工融入度')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:E1')
    ws['A1'] = '工具07：农民工城市融入度测评'
    set_title_style(ws['A1'])

    ws.merge_cells('A2:E2')
    ws['A2'] = '【填写说明】八维度评估农民工的城市融入程度。每个维度满分10分，结合自评和代际对比分析。'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')

    headers = ['评估维度', '具体指标', '自己评分(1-10)', '父辈评分(1-10)', '子女评分(1-10)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        set_header_style(cell)

    dimensions = [
        ('经济融入', '收入水平、工作稳定性、社会保障'),
        ('居住融入', '住房条件、居住稳定性、居住环境'),
        ('社会融入', '社交网络、社会参与、社区归属'),
        ('文化融入', '城市文化认同、价值观念适应'),
        ('心理融入', '城市认同感、归属感、未来预期'),
        ('制度融入', '户籍制度、公共服务获取平权'),
        ('代际融入', '子女教育融入、代际关系调适'),
        ('身份融入', '市民身份认同、标签认同程度'),
    ]

    for idx, (dim, ind) in enumerate(dimensions):
        row = 4 + idx
        ws.cell(row=row, column=1, value=dim)
        ws.cell(row=row, column=2, value=ind)
        for col in [3, 4, 5]:
            cell = ws.cell(row=row, column=col, value=5)
            cell.font = Font(name='微软雅黑', size=11, color='0000FF')
            cell.alignment = Alignment(horizontal='center')

    set_table_border(ws, 3, 11, 1, 5)

    # Totals
    ws.cell(row=13, column=1, value='各代评分汇总：')
    ws.cell(row=13, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    ws.cell(row=14, column=1, value='自己总分：')
    ws.cell(row=14, column=3, value='=SUM(C4:C11)')
    ws.cell(row=14, column=3).font = Font(name='微软雅黑', size=12, bold=True, color='C00000')

    ws.cell(row=15, column=1, value='父辈总分：')
    ws.cell(row=15, column=4, value='=SUM(D4:D11)')
    ws.cell(row=15, column=4).font = Font(name='微软雅黑', size=12, bold=True, color='C00000')

    ws.cell(row=16, column=1, value='子女总分：')
    ws.cell(row=16, column=5, value='=SUM(E4:E11)')
    ws.cell(row=16, column=5).font = Font(name='微软雅黑', size=12, bold=True, color='C00000')

    ws.cell(row=18, column=1, value='【融入障碍分析】')
    ws.cell(row=18, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    ws.cell(row=19, column=1, value='最大障碍：____________________________________')
    ws.cell(row=20, column=1, value='次要障碍：____________________________________')
    ws.cell(row=21, column=1, value='促进因素：____________________________________')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

# ================== 工具08 ==================
def create_tool_08(wb):
    ws = wb.create_sheet('工具08-权钱结网识别')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:C1')
    ws['A1'] = '工具08：权钱结网识别清单'
    set_title_style(ws['A1'])

    ws.merge_cells('A2:C2')
    ws['A2'] = '【填写说明】识别权力与金钱的不正当勾连。以下清单用于自查和预警。'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')

    ws.cell(row=3, column=1, value='一、预警信号检测（符合的请打√）')
    ws.cell(row=3, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    headers = ['预警类别', '信号描述', '是否出现']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        set_header_style(cell)

    warnings = [
        ('权力寻租', '官员利用职务便利为特定企业谋取利益'),
        ('权力寻租', '官员家属从事与父辈职务相关的经营活动'),
        ('资本渗透', '企业通过行贿获取稀缺资源或行政许可'),
        ('资本渗透', '企业通过政商旋转门获取内幕信息'),
        ('利益输送', '通过关联交易向权力关系人输送利益'),
        ('利益输送', '通过子女联姻实现政商联盟'),
        ('保护伞', '黑恶势力背后有官员撑腰'),
        ('保护伞', '企业违法行为长期未被查处'),
    ]

    for idx, (cat, sig) in enumerate(warnings):
        row = 5 + idx
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=sig)
        ws.cell(row=row, column=3, value='□是  □否')

    set_table_border(ws, 4, 12, 1, 3)

    ws.cell(row=14, column=1, value='二、场景案例自查')
    ws.cell(row=14, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    ws.merge_cells('A15:C16')
    ws.cell(row=15, column=1, value='''案例1：某企业主通过官员子女留学费用赞助获取行政许可，您的判断是：
□权钱交易  □正常商业行为  □难以判断

案例2：某官员退休后到曾审批项目的企业任高管，您的判断是：
□期权腐败  □正常职业选择  □难以判断''')
    ws.cell(row=15, column=1).font = Font(name='微软雅黑', size=10)
    ws.cell(row=15, column=1).alignment = Alignment(wrap_text=True, vertical='top')

    ws.cell(row=18, column=1, value='三、自查问题')
    ws.cell(row=18, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    ws.merge_cells('A19:C21')
    ws.cell(row=19, column=1, value='''1. 您或您的亲友是否曾被要求"找关系"办事？
2. 您是否遇到过不合理的市场准入障碍？
3. 您是否了解身边存在的政商不当关系？''')
    ws.cell(row=19, column=1).font = Font(name='微软雅黑', size=10)
    ws.cell(row=19, column=1).alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15

# ================== 工具09 ==================
def create_tool_09(wb):
    ws = wb.create_sheet('工具09-资本转化路径')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:D1')
    ws['A1'] = '工具09：三种资本转化路径表'
    set_title_style(ws['A1'])

    ws.merge_cells('A2:D2')
    ws['A2'] = '【填写说明】分析经济资本、文化资本、社会资本之间的转化路径和具体方法。'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')

    headers = ['转化方向', '转化逻辑', '具体方法', '案例']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        set_header_style(cell)

    paths = [
        ('经济→文化', '用金钱购买教育、文化体验，提升文化资本',
         '1. 购买优质教育资源\n2. 培养艺术爱好\n3. 参加高端培训\n4. 收藏艺术品',
         '中产阶层送子女上国际学校'),
        ('文化→社会', '用知识和人脉获取更多社会资源',
         '1. 通过学术圈拓展人脉\n2. 用专业能力换取信任\n3. 参与社会活动建立声誉',
         '律师通过专业能力建立客户网络'),
        ('社会→经济', '用人脉关系获取经济利益',
         '1. 介绍项目获取佣金\n2. 合作投资机会\n3. 政商资源对接',
         '掮客通过人脉撮合交易获利'),
        ('文化→经济', '将知识直接转化为产品或服务',
         '1. 出书、授课\n2. 咨询服务\n3. 技术转让\n4. 内容创业',
         '知识博主通过知识付费变现'),
        ('社会→文化', '通过社交圈获取文化资源',
         '1. 加入高端俱乐部\n2. 参与文化交流活动\n3. 进入文化圈层',
         '通过朋友介绍参观私人美术馆'),
        ('经济→社会', '通过消费建立社交身份',
         '1. 高端消费场所消费\n2. 参加商务社交活动\n3. 赞助社会活动',
         '通过高尔夫球会结识商界精英'),
    ]

    for idx, (direction, logic, method, case) in enumerate(paths):
        row = 4 + idx
        ws.cell(row=row, column=1, value=direction)
        ws.cell(row=row, column=2, value=logic)
        ws.cell(row=row, column=3, value=method)
        ws.cell(row=row, column=4, value=case)
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = Font(name='微软雅黑', size=10)
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    set_table_border(ws, 3, 9, 1, 4)

    ws.cell(row=11, column=1, value='【您的资本转化实践】')
    ws.cell(row=11, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    ws.cell(row=12, column=1, value='最成功的资本转化：')
    ws.merge_cells('B12:D12')
    ws.cell(row=12, column=2, value='________________________________________________')

    ws.cell(row=13, column=1, value='尚未实现的转化方向：')
    ws.merge_cells('B13:D13')
    ws.cell(row=13, column=2, value='________________________________________________')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25

# ================== 工具10 ==================
def create_tool_10(wb):
    ws = wb.create_sheet('工具10-五个一行动计划')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:E1')
    ws['A1'] = '工具10：五个一行动计划表'
    set_title_style(ws['A1'])

    ws.merge_cells('A2:E2')
    ws['A2'] = '【填写说明】制定从即时到年度的行动计划。每个层面对应：目标、行动、资源、里程碑。'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')

    headers = ['时间维度', '目标', '行动', '资源需求', '里程碑/截止日期']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        set_header_style(cell)

    time_frames = [
        ('一天\n(24小时内)', '', '', '', ''),
        ('一周\n(7天内)', '', '', '', ''),
        ('一月\n(30天内)', '', '', '', ''),
        ('一季\n(90天内)', '', '', '', ''),
        ('一年\n(365天内)', '', '', '', ''),
    ]

    for idx, (time, *values) in enumerate(time_frames):
        row = 4 + idx
        ws.cell(row=row, column=1, value=time)
        ws.cell(row=row, column=1).font = Font(name='微软雅黑', size=11, bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        for col, val in enumerate(values, 2):
            ws.cell(row=row, column=col, value=val)
            ws.cell(row=row, column=col).font = Font(name='微软雅黑', size=10)

    set_table_border(ws, 3, 8, 1, 5)

    # Set row heights
    for row in range(4, 9):
        ws.row_dimensions[row].height = 35

    ws.cell(row=10, column=1, value='【行动计划示例参考】')
    ws.cell(row=10, column=1).font = Font(name='微软雅黑', size=11, bold=True)

    ws.merge_cells('A11:E15')
    ws.cell(row=11, column=1, value='''一天：完成社会资本盘点，列出可依赖的10个人脉
一周：约见一位能够提供职业建议的前辈
一月：完成职业技能评估，确定需要提升的能力项
一季：完成一门专业认证课程的学习
一年：实现收入增长20%或完成职业转型''')
    ws.cell(row=11, column=1).font = Font(name='微软雅黑', size=10)
    ws.cell(row=11, column=1).alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

# ================== Main ==================
def main():
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create all tools
    create_tool_01(wb)
    create_tool_02(wb)
    create_tool_03(wb)
    create_tool_04(wb)
    create_tool_05(wb)
    create_tool_06(wb)
    create_tool_07(wb)
    create_tool_08(wb)
    create_tool_09(wb)
    create_tool_10(wb)

    # Save
    output_path = 'D:/新课开发/工作手册/梁晓声社会阶层分析/中国社会各阶层分析-原始版/完整课程包/08-工具表单集锦/工具表单-中国社会各阶层分析.xlsx'
    wb.save(output_path)
    print(f'Excel workbook saved to: {output_path}')

if __name__ == '__main__':
    main()

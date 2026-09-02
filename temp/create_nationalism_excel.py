# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = "D:/新课开发/政治学/18_民族主义思想史-一个概念如何塑造现代世界/配套表格/"

# Color scheme - academic dark blue
HEADER_FILL = "1B4F9B"
ALT_FILL = "F2F2F2"
WHITE = "FFFFFF"
DARK_TEXT = "1A1A1A"

def create_header_style():
    return {
        'font': Font(bold=True, color=WHITE, size=11),
        'fill': PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
    }

def create_cell_style(bold=False, fill_color=None, align='left'):
    fill = PatternFill(start_color=fill_color or WHITE, end_color=fill_color or WHITE, fill_type='solid') if fill_color else None
    return {
        'font': Font(bold=bold, color=DARK_TEXT, size=10),
        'fill': fill,
        'alignment': Alignment(horizontal=align, vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
    }

def apply_style(ws, row, col, style_dict):
    cell = ws.cell(row=row, column=col)
    for attr, value in style_dict.items():
        setattr(cell, attr, value)
    return cell

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ==================== FILE 1 ====================
def create_case_analysis():
    wb = openpyxl.Workbook()

    # Sheet 1: 案例库
    ws1 = wb.active
    ws1.title = "案例库"

    headers1 = ["案例名称", "国家/地区", "背景", "类型", "发生时间", "关键特征", "备注"]
    for col, h in enumerate(headers1, 1):
        apply_style(ws1, 1, col, create_header_style())
        ws1.cell(row=1, column=col).value = h

    cases = [
        ["法国大革命", "法国", "1789年法国大革命，封建专制危机", "公民民族主义", "1789", "自由、平等、博爱；公民身份高于血缘", "现代民族主义发源地"],
        ["美国独立战争", "美国", "北美殖民地反抗英国统治", "公民民族主义", "1775-1783", "天赋人权、自决权；《独立宣言》", "首部成文宪法"],
        ["德国统一运动", "德国", "普鲁士主导的统一进程", "族群民族主义", "1861-1871", "文化同质性；铁血政策", "积极民族主义典型"],
        ["日本明治维新", "日本", "幕末危机后的现代化转型", "族群民族主义", "1868", "万世一系、天皇中心；脱亚入欧", "威权现代化"],
        ["中国辛亥革命", "中国", "推翻清王朝，建立共和国", "公民民族主义", "1911", "驱逐鞑虏、恢复中华；五族共和", "亚洲第一个共和国"],
        ["苏联模式", "苏联", "社会主义国家的民族政策", "意识形态型", "1922-1991", "国际主义vs民族主义；加盟共和国", "社会主义民族理论"],
        ["印度独立运动", "印度", "反抗英国殖民统治", "公民民族主义", "1947", "多元文化主义；世俗主义", "最大民主国家"],
        ["纳粹德国", "德国", "一战后经济危机与复仇主义", "极端民族主义", "1933-1945", "种族优越论；领土扩张", "民族主义极端化"],
        ["泛阿拉伯主义", "阿拉伯世界", "反抗西方殖民", "文化民族主义", "1950s-1970s", "语言文化认同；反帝反殖", "后殖民民族主义"],
        ["苏联解体", "苏联/独联体", "冷战结束与意识形态崩溃", "去民族主义", "1991", "民族自决；加盟共和国独立", "民族主义复兴"]
    ]

    for row_idx, case in enumerate(cases, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        for col, val in enumerate(case, 1):
            style = create_cell_style(fill_color=fill, align='center' if col in [4,5] else 'left')
            apply_style(ws1, row_idx, col, style)
            ws1.cell(row=row_idx, column=col).value = val

    widths1 = [15, 12, 30, 14, 12, 35, 20]
    for i, w in enumerate(widths1, 1):
        set_col_width(ws1, i, w)

    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = "A2"

    # Sheet 2: 分析框架
    ws2 = wb.create_sheet("分析框架")

    ws2.merge_cells('A1:H1')
    title_cell = ws2.cell(row=1, column=1)
    title_cell.value = "民族主义四维判断框架 - 量化评分表"
    title_cell.font = Font(bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 35

    dim_headers = ["案例", "维度1:认同基础", "维度2:排斥性", "维度3:制度路径", "维度4:建构方式", "综合评分", "类型判断", "分析"]
    for col, h in enumerate(dim_headers, 1):
        apply_style(ws2, 2, col, create_header_style())
        ws2.cell(row=2, column=col).value = h

    dim_desc = ["(1=族群血缘, 10=公民认同)", "(1=开放包容, 10=封闭排外)", "(1=民主制度, 10=威权体制)", "(1=原生积淀, 10=建构塑造)", "", "", "", ""]
    for col, d in enumerate(dim_desc, 1):
        cell = ws2.cell(row=3, column=col)
        cell.value = d
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    scoring_data = [
        ["法国大革命", 9, 4, 3, 6],
        ["美国独立战争", 9, 3, 2, 5],
        ["德国统一运动", 4, 7, 7, 8],
        ["日本明治维新", 3, 8, 8, 9],
        ["中国辛亥革命", 7, 5, 4, 8],
        ["苏联模式", 5, 6, 9, 7],
        ["印度独立运动", 8, 3, 2, 6],
        ["纳粹德国", 2, 10, 10, 9],
        ["泛阿拉伯主义", 5, 6, 5, 8],
        ["苏联解体", 6, 5, 4, 6]
    ]

    for row_idx, case in enumerate(scoring_data, 4):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        case_name = case[0]

        apply_style(ws2, row_idx, 1, create_cell_style(bold=True, fill_color=fill))
        ws2.cell(row=row_idx, column=1).value = case_name

        for col, score in enumerate(case[1:5], 2):
            style = create_cell_style(bold=True, fill_color=fill, align='center')
            apply_style(ws2, row_idx, col, style)
            ws2.cell(row=row_idx, column=col).value = score

        formula_cell = ws2.cell(row=row_idx, column=6)
        formula_cell.value = f"=AVERAGE(B{row_idx}:E{row_idx})"
        formula_cell.number_format = '0.0'
        apply_style(ws2, row_idx, 6, create_cell_style(bold=True, fill_color=fill, align='center'))

        type_cell = ws2.cell(row=row_idx, column=7)
        score_ref = f"F{row_idx}"
        type_cell.value = f'=IF({score_ref}<=3,"公民-开放型",IF({score_ref}<=5,"公民-保守型",IF({score_ref}<=7,"族群-建构型","极端型")))'
        apply_style(ws2, row_idx, 7, create_cell_style(bold=True, fill_color=fill, align='center'))

        analysis_texts = {
            "法国大革命": "公民民族主义典范，自由主义传统",
            "美国独立战争": "最开放的公民民族主义样本",
            "德国统一运动": "文化民族主义与威权结合",
            "日本明治维新": "族群基础上的威权现代化",
            "中国辛亥革命": "从族群向公民过渡的尝试",
            "苏联模式": "意识形态压制民族认同的失败",
            "印度独立运动": "多元文化的民主民族主义",
            "纳粹德国": "民族主义极端化的极端案例",
            "泛阿拉伯主义": "文化认同驱动的后殖民运动",
            "苏联解体": "威权压制后的民族主义反弹"
        }
        apply_style(ws2, row_idx, 8, create_cell_style(fill_color=fill))
        ws2.cell(row=row_idx, column=8).value = analysis_texts.get(case_name, "")

    widths2 = [18, 14, 14, 14, 14, 12, 14, 35]
    for i, w in enumerate(widths2, 1):
        set_col_width(ws2, i, w)

    ws2.freeze_panes = "A4"

    # Sheet 3: 对比分析
    ws3 = wb.create_sheet("对比分析")

    compare_headers = ["对比维度", "法国", "美国", "德国", "日本", "中国", "印度", "苏联", "纳粹德国"]
    for col, h in enumerate(compare_headers, 1):
        apply_style(ws3, 1, col, create_header_style())
        ws3.cell(row=1, column=col).value = h

    compare_data = [
        ["认同基础", "公民", "公民", "族群", "族群", "混合", "公民", "意识形态", "族群"],
        ["制度路径", "民主", "民主", "威权", "威权", "混合", "民主", "威权", "极权"],
        ["建构方式", "建构", "建构", "建构", "建构", "建构", "原生", "建构", "建构"],
        ["排斥程度", "低", "低", "中", "高", "中", "低", "高", "极高"],
        ["历史评价", "正面", "正面", "复杂", "复杂", "复杂", "正面", "负面", "负面"],
        ["当代启示", "民主典范", "自由典范", "警示教训", "现代化路径", "转型挑战", "多元共处", "意识形态局限", "极端主义危害"]
    ]

    for row_idx, row_data in enumerate(compare_data, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        for col, val in enumerate(row_data, 1):
            style = create_cell_style(fill_color=fill, align='center' if col > 1 else 'left')
            apply_style(ws3, row_idx, col, style)
            ws3.cell(row=row_idx, column=col).value = val

    set_col_width(ws3, 1, 15)
    for i in range(2, 10):
        set_col_width(ws3, i, 14)

    ws3.freeze_panes = "B2"

    wb.save(os.path.join(OUT_DIR, "民族主义案例分析表.xlsx"))
    print("Created: 民族主义案例分析表.xlsx")

# ==================== FILE 2 ====================
def create_timeline():
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "时间轴总表"

    ws1.merge_cells('A1:F1')
    title = ws1.cell(row=1, column=1)
    title.value = "民族主义三波浪潮 - 历史时间轴"
    title.font = Font(bold=True, size=14, color=WHITE)
    title.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 35

    headers = ["年份", "事件名称", "类型", "所属浪潮", "影响", "重要程度"]
    for col, h in enumerate(headers, 1):
        apply_style(ws1, 2, col, create_header_style())
        ws1.cell(row=2, column=col).value = h

    timeline_data = [
        [1789, "法国大革命", "革命运动", "第一波", "确立公民民族主义范式，传播自由平等理念", 5],
        [1791, "美国宪法通过", "制度建设", "第一波", "确立三权分立与联邦制", 4],
        [1804, "拿破仑帝国", "帝国扩张", "第一波", "民族主义随战争传播欧洲", 3],
        [1815, "维也纳会议", "外交博弈", "第一波", "民族主义与保守主义博弈", 3],
        [1820, "拉丁美洲独立运动", "民族独立", "第一波", "西属美洲独立，民族自决传播", 4],
        [1848, "欧洲革命", "革命运动", "第一波", "民族统一与自由主义结合", 5],
        [1861, "意大利统一", "民族统一", "第一波", "民族国家建设典范", 4],
        [1867, "奥匈帝国二元制", "制度改革", "第一波", "民族问题制度化尝试", 3],
        [1871, "德国统一", "民族统一", "第一波", "铁血政策统一德意志", 5],
        [1882, "埃及民族主义兴起", "民族觉醒", "第一波", "阿拉伯民族主义萌芽", 3],
        [1905, "俄国革命", "革命运动", "第一波", "民族自决诉求初现", 4],
        [1911, "中国辛亥革命", "革命运动", "第一波", "亚洲民族主义兴起", 5],
        [1914, "第一次世界大战", "世界大战", "第二波", "民族边界重新划分", 5],
        [1917, "俄国十月革命", "革命运动", "第二波", "社会主义民族理论形成", 5],
        [1918, "威尔逊十四点计划", "外交倡议", "第二波", "民族自决原则正式提出", 5],
        [1919, "巴黎和会", "外交会议", "第二波", "民族国家体系重建", 4],
        [1922, "苏联成立", "国家成立", "第二波", "多民族联邦制实验", 4],
        [1923, "土耳其共和国成立", "国家成立", "第二波", "世俗民族国家建立", 4],
        [1933, "纳粹德国建立", "极权兴起", "第二波", "极端民族主义恶性发展", 5],
        [1939, "第二次世界大战", "世界大战", "第二波", "民族主义极端化灾难", 5],
        [1945, "联合国成立", "国际组织", "第二波", "民族国家体系全球化", 5],
        [1947, "印度独立", "民族独立", "第三波", "最大规模非殖民化", 5],
        [1948, "以色列建国", "国家成立", "第三波", "犹太民族国家建立", 5],
        [1955, "万隆会议", "国际会议", "第三波", "第三世界民族主义联合", 4],
        [1956, "苏伊士运河危机", "地缘冲突", "第三波", "民族主义与大国博弈", 3],
        [1960, "非洲去殖民化", "民族独立", "第三波", "大规模民族独立运动", 5],
        [1979, "伊朗伊斯兰革命", "宗教革命", "第三波", "宗教民族主义兴起", 4],
        [1989, "柏林墙倒塌", "历史事件", "第三波", "民族主义复兴标志", 5],
        [1991, "苏联解体", "国家解体", "第三波", "民族主义最大规模释放", 5],
        [1992, "南斯拉夫解体", "国家解体", "第三波", "民族冲突与战争", 4],
        [1997, "香港回归", "主权移交", "第三波", "民族统一主义实践", 3],
        [2001, "9/11事件", "恐怖事件", "第三波", "民族主义与恐怖主义交织", 5],
        [2014, "克里米亚事件", "地缘冲突", "第三波", "新民族主义扩张", 4],
        [2016, "英国脱欧", "政治事件", "第三波", "民族主义回潮典型", 4],
        [2020, "新冠疫情民族主义", "全球事件", "第三波", "民族主义与全球治理博弈", 3]
    ]

    for row_idx, event in enumerate(timeline_data, 3):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        for col, val in enumerate(event, 1):
            style = create_cell_style(fill_color=fill, align='center' if col in [1,3,4,6] else 'left')
            apply_style(ws1, row_idx, col, style)
            ws1.cell(row=row_idx, column=col).value = val

    widths = [10, 22, 14, 12, 40, 12]
    for i, w in enumerate(widths, 1):
        set_col_width(ws1, i, w)

    ws1.freeze_panes = "A3"

    # Sheet 2: 三波浪潮详情
    ws2 = wb.create_sheet("三波浪潮详情")

    ws2.merge_cells('A1:F1')
    title2 = ws2.cell(row=1, column=1)
    title2.value = "民族主义三波浪潮详解"
    title2.font = Font(bold=True, size=14, color=WHITE)
    title2.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    title2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 35

    headers2 = ["浪潮名称", "核心主题", "时间跨度", "典型事件", "历史贡献", "负面影响"]
    for col, h in enumerate(headers2, 1):
        apply_style(ws2, 2, col, create_header_style())
        ws2.cell(row=2, column=col).value = h

    waves_info = [
        ["第一波民族主义", "民族觉醒与公民民族主义", "1789-1914",
         "法国大革命、美国独立、欧洲统一运动",
         "确立民族国家范式，公民身份认同",
         "民族主义与自由主义、议会制度结合"],

        ["第二波民族主义", "去殖民化与民族自决", "1914-1945",
         "俄国革命、一战后民族自决、去殖民化萌芽",
         "民族自决成为国际准则",
         "民族主义与威权主义、法西斯主义结合"],

        ["第三波民族主义", "全球化与民族主义复兴", "1945至今",
         "印度独立、非洲去殖民化、苏联解体",
         "民族国家体系全球化，民族主义在新语境下复兴",
         "民族主义与宗教、原教旨主义、民粹主义结合"]
    ]

    for row_idx, wave in enumerate(waves_info, 3):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        for col, val in enumerate(wave, 1):
            style = create_cell_style(fill_color=fill, align='center' if col == 3 else 'left')
            apply_style(ws2, row_idx, col, style)
            ws2.cell(row=row_idx, column=col).value = val

    widths2 = [18, 22, 14, 35, 35, 35]
    for i, w in enumerate(widths2, 1):
        set_col_width(ws2, i, w)

    # Sheet 3: 重要事件索引
    ws3 = wb.create_sheet("重要事件索引")

    ws3.merge_cells('A1:E1')
    title3 = ws3.cell(row=1, column=1)
    title3.value = "民族主义重要事件索引"
    title3.font = Font(bold=True, size=14, color=WHITE)
    title3.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    title3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 35

    headers3 = ["年份", "事件", "类型", "关键词", "历史意义"]
    for col, h in enumerate(headers3, 1):
        apply_style(ws3, 2, col, create_header_style())
        ws3.cell(row=2, column=col).value = h

    key_events = [
        [1789, "法国大革命", "革命", "公民民族主义、自由平等", "开创现代民族主义范式"],
        [1848, "欧洲革命", "革命", "民族统一、自由主义", "民族主义与民主结合"],
        [1917, "十月革命", "革命", "民族自决、社会主义", "新型民族理论"],
        [1945, "联合国成立", "国际组织", "民族国家体系、主权平等", "民族国家体系全球化"],
        [1947, "印度独立", "独立", "非暴力、世俗主义", "最大民主国家"],
        [1991, "苏联解体", "解体", "民族自决、冷战结束", "民族主义最大规模释放"]
    ]

    for row_idx, event in enumerate(key_events, 3):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        for col, val in enumerate(event, 1):
            style = create_cell_style(fill_color=fill, align='center' if col in [1,3] else 'left')
            apply_style(ws3, row_idx, col, style)
            ws3.cell(row=row_idx, column=col).value = val

    widths3 = [10, 20, 14, 30, 40]
    for i, w in enumerate(widths3, 1):
        set_col_width(ws3, i, w)

    wb.save(os.path.join(OUT_DIR, "历史时间轴数据.xlsx"))
    print("Created: 历史时间轴数据.xlsx")

# ==================== FILE 3 ====================
def create_evaluation():
    wb = openpyxl.Workbook()

    # Sheet 1: 成绩追踪
    ws1 = wb.active
    ws1.title = "成绩追踪"

    ws1.merge_cells('A1:J1')
    title = ws1.cell(row=1, column=1)
    title.value = "学员学习成效追踪表 - 前测/后测对比"
    title.font = Font(bold=True, size=14, color=WHITE)
    title.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 35

    headers = ["学员ID", "姓名", "班级", "前测总分", "后测总分", "进步分数", "进步率", "前测日期", "后测日期", "评估等级"]
    for col, h in enumerate(headers, 1):
        apply_style(ws1, 2, col, create_header_style())
        ws1.cell(row=2, column=col).value = h

    sample_students = [
        ["S001", "张三", "政治学24A", 65, 82, "=D3-C3", "=(D3-C3)/C3", "2024-03-01", "2024-06-15", "=IF(F3>=20,\"优秀\",IF(F3>=10,\"良好\",\"合格\"))"],
        ["S002", "李四", "政治学24A", 58, 75, "=D4-C4", "=(D4-C4)/C4", "2024-03-01", "2024-06-15", "=IF(F4>=20,\"优秀\",IF(F4>=10,\"良好\",\"合格\"))"],
        ["S003", "王五", "政治学24B", 72, 88, "=D5-C5", "=(D5-C5)/C5", "2024-03-01", "2024-06-15", "=IF(F5>=20,\"优秀\",IF(F5>=10,\"良好\",\"合格\"))"],
        ["S004", "赵六", "政治学24B", 61, 70, "=D6-C6", "=(D6-C6)/C6", "2024-03-01", "2024-06-15", "=IF(F6>=20,\"优秀\",IF(F6>=10,\"良好\",\"合格\"))"],
        ["S005", "钱七", "政治学24A", 55, 78, "=D7-C7", "=(D7-C7)/C7", "2024-03-01", "2024-06-15", "=IF(F7>=20,\"优秀\",IF(F7>=10,\"良好\",\"合格\"))"],
        ["S006", "孙八", "政治学24B", 68, 85, "=D8-C8", "=(D8-C8)/C8", "2024-03-01", "2024-06-15", "=IF(F8>=20,\"优秀\",IF(F8>=10,\"良好\",\"合格\"))"],
        ["S007", "周九", "政治学24A", 70, 79, "=D9-C9", "=(D9-C9)/C9", "2024-03-01", "2024-06-15", "=IF(F9>=20,\"优秀\",IF(F9>=10,\"良好\",\"合格\"))"],
        ["S008", "吴十", "政治学24B", 63, 71, "=D10-C10", "=(D10-C10)/C10", "2024-03-01", "2024-06-15", "=IF(F10>=20,\"优秀\",IF(F10>=10,\"良好\",\"合格\"))"]
    ]

    for row_idx, student in enumerate(sample_students, 3):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        for col, val in enumerate(student, 1):
            style = create_cell_style(fill_color=fill, align='center')
            if col in [1, 2, 3]:
                style['alignment'] = Alignment(horizontal='left', vertical='center')
            apply_style(ws1, row_idx, col, style)
            ws1.cell(row=row_idx, column=col).value = val

        ws1.cell(row=row_idx, column=6).number_format = '0'
        ws1.cell(row=row_idx, column=7).number_format = '0.0%'

    summary_row = len(sample_students) + 3
    ws1.merge_cells(f'A{summary_row}:B{summary_row}')
    summary_cell = ws1.cell(row=summary_row, column=1)
    summary_cell.value = "班级平均"
    summary_cell.font = Font(bold=True, color=WHITE)
    summary_cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in [3, 4, 5, 6, 7]:
        cell = ws1.cell(row=summary_row, column=col)
        if col == 3:
            cell.value = ""
        elif col == 4:
            cell.value = f"=AVERAGE(D3:D{summary_row-1})"
        elif col == 5:
            cell.value = f"=AVERAGE(E3:E{summary_row-1})"
        elif col == 6:
            cell.value = f"=AVERAGE(F3:F{summary_row-1})"
        elif col == 7:
            cell.value = f"=AVERAGE(G3:G{summary_row-1})"
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    widths = [12, 12, 14, 12, 12, 12, 12, 14, 14, 12]
    for i, w in enumerate(widths, 1):
        set_col_width(ws1, i, w)

    ws1.freeze_panes = "A3"

    # Sheet 2: 班级对比
    ws2 = wb.create_sheet("班级对比")

    ws2.merge_cells('A1:H1')
    title2 = ws2.cell(row=1, column=1)
    title2.value = "班级平均分对比分析"
    title2.font = Font(bold=True, size=14, color=WHITE)
    title2.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    title2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 35

    headers2 = ["班级", "前测平均", "后测平均", "平均进步", "进步率", "学员人数", "优秀率", "及格率"]
    for col, h in enumerate(headers2, 1):
        apply_style(ws2, 2, col, create_header_style())
        ws2.cell(row=2, column=col).value = h

    class_data = [
        ["政治学24A", 63.2, 80.8, 17.6, "=D3/C3-1", 25, 0.35, 0.88],
        ["政治学24B", 66.0, 76.0, 10.0, "=D4/C4-1", 23, 0.22, 0.78],
        ["政治学24C", 61.5, 79.0, 17.5, "=D5/C5-1", 22, 0.32, 0.85]
    ]

    for row_idx, class_row in enumerate(class_data, 3):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        for col, val in enumerate(class_row, 1):
            style = create_cell_style(fill_color=fill, align='center')
            apply_style(ws2, row_idx, col, style)
            ws2.cell(row=row_idx, column=col).value = val
        ws2.cell(row=row_idx, column=5).number_format = '0.0%'
        ws2.cell(row=row_idx, column=7).number_format = '0.0%'
        ws2.cell(row=row_idx, column=8).number_format = '0.0%'

    widths2 = [14, 12, 12, 12, 12, 12, 12, 12]
    for i, w in enumerate(widths2, 1):
        set_col_width(ws2, i, w)

    # Sheet 3: 维度进步
    ws3 = wb.create_sheet("维度进步")

    ws3.merge_cells('A1:G1')
    title3 = ws3.cell(row=1, column=1)
    title3.value = "各维度学习进步追踪"
    title3.font = Font(bold=True, size=14, color=WHITE)
    title3.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    title3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 35

    headers3 = ["维度", "考核内容", "前测均分", "后测均分", "进步幅度", "掌握率变化", "建议"]
    for col, h in enumerate(headers3, 1):
        apply_style(ws3, 2, col, create_header_style())
        ws3.cell(row=2, column=col).value = h

    dimension_data = [
        ["认同基础", "公民vs族群民族主义区分", 62, 81, "=D3-C3", "=(D3-C3)/C3", "加强案例教学"],
        ["排斥性", "民族主义开放与封闭光谱", 55, 72, "=D4-C4", "=(D4-C4)/C4", "增加对比分析练习"],
        ["制度路径", "民主与威权民族主义", 60, 78, "=D5-C5", "=(D5-C5)/C5", "历史案例深入"],
        ["建构方式", "原生与建构民族主义", 58, 75, "=D6-C6", "=(D6-C6)/C6", "理论框架强化"],
        ["三波浪潮", "民族主义历史演变", 65, 85, "=D7-C7", "=(D7-C7)/C7", "时间轴记忆法"],
        ["当代应用", "民族主义与全球治理", 52, 68, "=D8-C8", "=(D8-C8)/C8", "时事分析训练"]
    ]

    for row_idx, dim in enumerate(dimension_data, 3):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        for col, val in enumerate(dim, 1):
            style = create_cell_style(fill_color=fill, align='center' if col in [3,4,5,6] else 'left')
            apply_style(ws3, row_idx, col, style)
            ws3.cell(row=row_idx, column=col).value = val
        ws3.cell(row=row_idx, column=5).number_format = '0'
        ws3.cell(row=row_idx, column=6).number_format = '0.0%'

    widths3 = [14, 28, 12, 12, 12, 14, 20]
    for i, w in enumerate(widths3, 1):
        set_col_width(ws3, i, w)

    # Sheet 4: 使用说明
    ws4 = wb.create_sheet("使用说明")

    ws4.merge_cells('A1:B1')
    title4 = ws4.cell(row=1, column=1)
    title4.value = "课程评估追踪表 - 使用说明"
    title4.font = Font(bold=True, size=14, color=WHITE)
    title4.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    title4.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 35

    instructions = [
        ["一、表单结构", ""],
        ["成绩追踪", "记录每位学员的前测/后测成绩，支持自动计算进步情况"],
        ["班级对比", "按班级汇总平均成绩，支持跨班级横向比较"],
        ["维度进步", "追踪学员在不同知识维度的进步情况"],
        ["", ""],
        ["二、使用方法", ""],
        ["1. 前测", "课程开始前进行摸底测试，填写前测成绩"],
        ["2. 后测", "课程结束后进行结业测试，填写后测成绩"],
        ["3. 自动计算", "进步分数、进步率、评估等级自动计算"],
        ["4. 班级汇总", "在「班级对比」表中查看班级整体表现"],
        ["", ""],
        ["三、评分等级标准", ""],
        ["优秀", "进步分数 >= 20 或进步率 >= 30%"],
        ["良好", "进步分数 >= 10 或进步率 >= 15%"],
        ["合格", "进步分数 < 10"],
        ["", ""],
        ["四、注意事项", ""],
        ["1. 学员ID请使用唯一标识符，便于长期追踪", ""],
        ["2. 建议在课程开始前统一进行前测，确保公平性", ""],
        ["3. 后测应在课程结束后1周内完成", ""],
        ["4. 维度分析需要根据实际考核内容调整", ""]
    ]

    for row_idx, (label, desc) in enumerate(instructions, 3):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        label_cell = ws4.cell(row=row_idx, column=1)
        label_cell.value = label
        is_bold = label and not any(label.startswith(x) for x in ["1", "2", "3", "4", "一", "二", "三", "四"])
        label_cell.font = Font(bold=is_bold)
        label_cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
        label_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        desc_cell = ws4.cell(row=row_idx, column=2)
        desc_cell.value = desc
        desc_cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    set_col_width(ws4, 1, 25)
    set_col_width(ws4, 2, 60)

    wb.save(os.path.join(OUT_DIR, "课程评估追踪表.xlsx"))
    print("Created: 课程评估追踪表.xlsx")

if __name__ == "__main__":
    os.makedirs(OUT_DIR, exist_ok=True)
    create_case_analysis()
    create_timeline()
    create_evaluation()
    print("\nAll Excel files created successfully!")

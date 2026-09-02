#!/usr/bin/env python3
"""
创建 3 个 Excel 文件：
1. 学员报名登记表.xlsx
2. 5因素测评汇总表.xlsx
3. 30天行动跟进表.xlsx
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime

# ==================== 通用样式 ====================
RED = "B81025"
DARK = "1A1A1A"
GRAY_LIGHT = "F2F2F2"
GRAY_DARK = "595959"
WHITE = "FFFFFF"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
RED_FILL = "FFC7CE"

# 字体
def make_font(color=DARK, bold=False, size=11, italic=False, name="微软雅黑"):
    return Font(name=name, size=size, color=color, bold=bold, italic=italic)

# 填充
def make_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

# 边框
thin = Side(border_style="thin", color=DARK)
thick = Side(border_style="medium", color=RED)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_top_thick = Border(left=thin, right=thin, top=thick, bottom=thin)
border_bottom_thick = Border(left=thin, right=thin, top=thin, bottom=thick)

# 对齐
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

# 输出目录
OUTPUT_DIR = "D:/2026年课程/竞越/创新领导力：打造创新型团队/完整课程表/15-配套表单Excel"


# ==================== 文件 1：学员报名登记表 ====================
def create_registration_form():
    wb = Workbook()
    ws = wb.active
    ws.title = "学员报名登记表"

    # 标题
    ws.merge_cells("A1:O1")
    ws["A1"] = "创新领导力：打造创新型团队  ·  学员报名登记表"
    ws["A1"].font = make_font(WHITE, bold=True, size=18)
    ws["A1"].fill = make_fill(RED)
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 40

    # 副标题
    ws.merge_cells("A2:O2")
    ws["A2"] = "课程日期：____________  班级编号：____________  讲师：____________"
    ws["A2"].font = make_font(DARK, size=11, italic=True)
    ws["A2"].alignment = align_center
    ws.row_dimensions[2].height = 22

    # 表头
    headers = [
        "学员编号", "姓名", "性别", "年龄", "部门", "职位",
        "入职年限", "直接上级", "联系电话", "邮箱",
        "心理安全\n（自评 1-10）", "认知多样\n（自评 1-10）",
        "探索空间\n（自评 1-10）", "学习速度\n（自评 1-10）",
        "领导者信号\n（自评 1-10）", "综合得分", "培训期望", "特殊需求", "备注"
    ]
    # 由于 19 列，需要合并一些列。重新规划列：
    headers = [
        "学员\n编号", "姓名", "性别", "部门", "职位",
        "入职\n年限", "直接上级", "联系电话", "邮箱",
        "心理安全", "认知多样", "探索空间", "学习速度", "领导者信号",
        "综合得分", "培训期望", "特殊需求", "备注"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = make_font(WHITE, bold=True, size=11)
        cell.fill = make_fill(DARK)
        cell.alignment = align_center
        cell.border = border_all

    ws.row_dimensions[3].height = 45

    # 列宽
    col_widths = [8, 10, 6, 14, 14, 8, 12, 14, 22, 10, 10, 10, 10, 10, 10, 24, 20, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 数据行（30 人） - 学员编号、姓名等为输入项，5 因素分数为输入
    for i in range(1, 31):
        row = 3 + i
        # 学员编号（自动生成）
        ws.cell(row=row, column=1, value=f"S{i:03d}").alignment = align_center
        ws.cell(row=row, column=1).font = make_font(DARK, bold=True, size=10)

        # 姓名、性别、部门、职位、入职年限、上级、电话、邮箱 - 空白等待输入
        for col_idx in [2, 3, 4, 5, 6, 7, 8, 9]:
            ws.cell(row=row, column=col_idx).font = make_font(DARK, size=10)
            ws.cell(row=row, column=col_idx).alignment = align_center

        # 5 因素分数（输入）
        for col_idx in range(10, 15):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = make_font("0000FF", size=10)  # 蓝色输入
            cell.alignment = align_center

        # 综合得分（公式）
        ws.cell(row=row, column=15, value=f"=AVERAGE(J{row}:N{row})")
        ws.cell(row=row, column=15).font = make_font("008000", bold=True, size=10)
        ws.cell(row=row, column=15).alignment = align_center
        ws.cell(row=row, column=15).number_format = "0.0"

        # 培训期望、特殊需求、备注
        for col_idx in [16, 17, 18]:
            ws.cell(row=row, column=col_idx).font = make_font(DARK, size=10)
            ws.cell(row=row, column=col_idx).alignment = align_left

        # 整行边框
        for col_idx in range(1, 19):
            ws.cell(row=row, column=col_idx).border = border_all

        # 隔行底色
        if i % 2 == 0:
            for col_idx in range(1, 19):
                ws.cell(row=row, column=col_idx).fill = make_fill(GRAY_LIGHT)

        ws.row_dimensions[row].height = 30

    # 数据验证
    dv_gender = DataValidation(type="list", formula1='"男,女,其他"', allow_blank=True)
    dv_gender.add(f"D4:D33")
    ws.add_data_validation(dv_gender)

    dv_score = DataValidation(type="decimal", operator="between", formula1=0, formula2=10, allow_blank=True)
    dv_score.add(f"J4:N33")
    ws.add_data_validation(dv_score)

    # 统计行
    stat_row = 34
    ws.cell(row=stat_row, column=1, value="统计").font = make_font(WHITE, bold=True, size=11)
    ws.cell(row=stat_row, column=1).fill = make_fill(RED)
    ws.cell(row=stat_row, column=1).alignment = align_center
    ws.cell(row=stat_row, column=1).border = border_all
    ws.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=9)
    for col_idx in range(1, 10):
        ws.cell(row=stat_row, column=col_idx).fill = make_fill(RED)
        ws.cell(row=stat_row, column=col_idx).border = border_all

    stats = ["平均分", "最高分", "最低分", "中位数", "标准差"]
    stat_formulas = ["AVERAGE", "MAX", "MIN", "MEDIAN", "STDEV"]
    for i, (label, func) in enumerate(zip(stats, stat_formulas)):
        col = 10 + i
        c = ws.cell(row=stat_row, column=col, value=label)
        c.font = make_font(WHITE, bold=True, size=10)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all

        # 综合得分平均
        if i < 4:
            ws.cell(row=stat_row + 1, column=col, value=f"={func}(O4:O33)")
        else:
            ws.cell(row=stat_row + 1, column=col, value=f"={func}(O4:O33)")
        ws.cell(row=stat_row + 1, column=col).font = make_font("008000", bold=True, size=10)
        ws.cell(row=stat_row + 1, column=col).alignment = align_center
        ws.cell(row=stat_row + 1, column=col).number_format = "0.0"
        ws.cell(row=stat_row + 1, column=col).border = border_all

    ws.cell(row=stat_row + 1, column=1, value="备注").font = make_font(WHITE, bold=True, size=10)
    ws.cell(row=stat_row + 1, column=1).fill = make_fill(DARK)
    ws.cell(row=stat_row + 1, column=1).alignment = align_center
    ws.cell(row=stat_row + 1, column=1).border = border_all
    ws.merge_cells(start_row=stat_row + 1, start_column=1, end_row=stat_row + 1, end_column=9)
    for col_idx in range(1, 10):
        ws.cell(row=stat_row + 1, column=col_idx).fill = make_fill(DARK)
        ws.cell(row=stat_row + 1, column=col_idx).border = border_all

    # 打印设置
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:3"

    # 冻结
    ws.freeze_panes = "C4"

    # 输出
    output_path = f"{OUTPUT_DIR}/学员报名登记表.xlsx"
    wb.save(output_path)
    print(f"✅ 已创建: {output_path}")


# ==================== 文件 2：5 因素测评汇总表 ====================
def create_assessment_summary():
    wb = Workbook()

    # Sheet 1: 个人得分汇总
    ws1 = wb.active
    ws1.title = "个人得分汇总"

    # 标题
    ws1.merge_cells("A1:K1")
    ws1["A1"] = "创新领导力 5 因素测评汇总表"
    ws1["A1"].font = make_font(WHITE, bold=True, size=18)
    ws1["A1"].fill = make_fill(RED)
    ws1["A1"].alignment = align_center
    ws1.row_dimensions[1].height = 40

    # 副标题
    ws1.merge_cells("A2:K2")
    ws1["A2"] = "5 大因素 + 综合得分 + 同事评估差距分析"
    ws1["A2"].font = make_font(DARK, size=11, italic=True)
    ws1["A2"].alignment = align_center
    ws1.row_dimensions[2].height = 22

    # 表头（双层）
    headers_top = [
        "学员\n编号", "姓名", "心理\n安全感", "认知\n多样性",
        "探索\n空间", "学习\n速度", "领导者\n信号",
        "综合\n得分", "自评\n平均", "同事\n评估", "差距\n分析"
    ]
    headers_sub = ["", "", "(6 题)", "(6 题)", "(6 题)", "(6 题)", "(6 题)", "(5 维)", "(6 维)", "(6 维)", "(自-同)"]

    for col_idx, h in enumerate(headers_top, 1):
        c = ws1.cell(row=3, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, size=11)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all

    for col_idx, h in enumerate(headers_sub, 1):
        c = ws1.cell(row=4, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, italic=True, size=9)
        c.fill = make_fill(GRAY_DARK)
        c.alignment = align_center
        c.border = border_all

    ws1.row_dimensions[3].height = 40
    ws1.row_dimensions[4].height = 20

    # 列宽
    col_widths = [8, 12, 12, 12, 12, 12, 12, 12, 12, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # 数据行（30 人）
    for i in range(1, 31):
        row = 4 + i
        # 编号
        ws1.cell(row=row, column=1, value=f"S{i:03d}").alignment = align_center
        ws1.cell(row=row, column=1).font = make_font(DARK, bold=True, size=10)

        # 姓名
        ws1.cell(row=row, column=2, value="").alignment = align_center

        # 5 因素自评分数（输入）
        for col_idx in range(3, 8):
            cell = ws1.cell(row=row, column=col_idx)
            cell.font = make_font("0000FF", size=10)
            cell.alignment = align_center
            cell.number_format = "0.0"

        # 综合得分（公式）
        ws1.cell(row=row, column=8, value=f"=AVERAGE(C{row}:G{row})")
        ws1.cell(row=row, column=8).font = make_font("008000", bold=True, size=10)
        ws1.cell(row=row, column=8).alignment = align_center
        ws1.cell(row=row, column=8).number_format = "0.0"

        # 自评平均（公式）
        ws1.cell(row=row, column=9, value=f"=AVERAGE(C{row}:G{row})")
        ws1.cell(row=row, column=9).font = make_font("008000", size=10)
        ws1.cell(row=row, column=9).alignment = align_center
        ws1.cell(row=row, column=9).number_format = "0.0"

        # 同事评估（输入）
        ws1.cell(row=row, column=10).font = make_font("0000FF", size=10)
        ws1.cell(row=row, column=10).alignment = align_center
        ws1.cell(row=row, column=10).number_format = "0.0"

        # 差距分析（公式）
        ws1.cell(row=row, column=11, value=f"=I{row}-J{row}")
        ws1.cell(row=row, column=11).font = make_font("008000", bold=True, size=10)
        ws1.cell(row=row, column=11).alignment = align_center
        ws1.cell(row=row, column=11).number_format = "0.0"

        # 边框 + 隔行底色
        for col_idx in range(1, 12):
            ws1.cell(row=row, column=col_idx).border = border_all
            if i % 2 == 0:
                ws1.cell(row=row, column=col_idx).fill = make_fill(GRAY_LIGHT)

        ws1.row_dimensions[row].height = 24

    # 数据验证 0-10
    dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=10, allow_blank=True)
    dv.add("C5:J34")
    ws1.add_data_validation(dv)

    # 条件格式：差距 > 1.0 红色，差距 < -1.0 绿色
    ws1.conditional_formatting.add(
        "K5:K34",
        CellIsRule(operator="greaterThan", formula=["1.0"],
                   fill=make_fill(YELLOW))
    )
    ws1.conditional_formatting.add(
        "K5:K34",
        CellIsRule(operator="lessThan", formula=["-1.0"],
                   fill=make_fill(GREEN))
    )

    # 班级统计
    stat_row = 36
    ws1.cell(row=stat_row, column=1, value="班级统计").font = make_font(WHITE, bold=True, size=11)
    ws1.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=12)
    for col_idx in range(1, 13):
        ws1.cell(row=stat_row, column=col_idx).fill = make_fill(RED)
        ws1.cell(row=stat_row, column=col_idx).border = border_all
    ws1.cell(row=stat_row, column=1).alignment = align_center

    stats = [
        ("平均分", "AVERAGE"),
        ("最高分", "MAX"),
        ("最低分", "MIN"),
        ("中位数", "MEDIAN"),
        ("标准差", "STDEV"),
        ("及格率(>=6.0)", "COUNTIF"),
    ]
    for i, (label, func) in enumerate(stats):
        r = stat_row + 1 + i
        c = ws1.cell(row=r, column=1, value=label)
        c.font = make_font(DARK, bold=True, size=10)
        c.fill = make_fill(GRAY_LIGHT)
        c.alignment = align_left
        c.border = border_all
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        for col_idx in range(1, 3):
            ws1.cell(row=r, column=col_idx).fill = make_fill(GRAY_LIGHT)
            ws1.cell(row=r, column=col_idx).border = border_all

        for col_idx in range(3, 11):
            col_letter = get_column_letter(col_idx)
            if func == "COUNTIF":
                formula = f'=COUNTIF({col_letter}5:{col_letter}34,">=6.0")/30'
            else:
                formula = f"={func}({col_letter}5:{col_letter}34)"
            cell = ws1.cell(row=r, column=col_idx, value=formula)
            cell.font = make_font("008000", bold=True, size=10)
            cell.alignment = align_center
            cell.number_format = "0.0" if func != "COUNTIF" else "0.0%"
            cell.border = border_all

        # 差距分析列单独处理
        col_letter = "K"
        if func == "COUNTIF":
            formula = f'=COUNTIF({col_letter}5:{col_letter}34,">=6.0")/30'
        else:
            formula = f"={func}({col_letter}5:{col_letter}34)"
        cell = ws1.cell(row=r, column=11, value=formula)
        cell.font = make_font("008000", bold=True, size=10)
        cell.alignment = align_center
        cell.number_format = "0.0" if func != "COUNTIF" else "0.0%"
        cell.border = border_all

        ws1.row_dimensions[r].height = 22

    # 打印
    ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
    ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
    ws1.page_setup.fitToPage = True
    ws1.page_setup.fitToWidth = 1
    ws1.print_title_rows = "1:4"
    ws1.freeze_panes = "C5"

    # ==================== Sheet 2: 班级统计 ====================
    ws2 = wb.create_sheet("班级统计")

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "班级 5 因素统计分析"
    ws2["A1"].font = make_font(WHITE, bold=True, size=18)
    ws2["A1"].fill = make_fill(RED)
    ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 40

    # 维度分析表
    ws2["A3"] = "维度"
    ws2["B3"] = "平均分"
    ws2["C3"] = "最高分"
    ws2["D3"] = "最低分"
    ws2["E3"] = "标准差"
    ws2["F3"] = "及格人数"
    ws2["G3"] = "及格率"
    ws2["H3"] = "等级"

    dimensions = [
        ("心理安全感", "C"),
        ("认知多样性", "D"),
        ("探索空间", "E"),
        ("学习速度", "F"),
        ("领导者信号", "G"),
        ("综合得分", "H"),
    ]

    for col_idx, h in enumerate(["维度", "平均分", "最高分", "最低分", "标准差", "及格人数", "及格率", "等级"], 1):
        c = ws2.cell(row=3, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, size=11)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all

    for i, (dim_name, src_col) in enumerate(dimensions):
        row = 4 + i
        ws2.cell(row=row, column=1, value=dim_name).font = make_font(DARK, bold=True, size=11)
        ws2.cell(row=row, column=1).alignment = align_left
        ws2.cell(row=row, column=1).border = border_all

        # 平均分
        ws2.cell(row=row, column=2, value=f"=AVERAGE('个人得分汇总'!{src_col}5:{src_col}34)")
        # 最高
        ws2.cell(row=row, column=3, value=f"=MAX('个人得分汇总'!{src_col}5:{src_col}34)")
        # 最低
        ws2.cell(row=row, column=4, value=f"=MIN('个人得分汇总'!{src_col}5:{src_col}34)")
        # 标准差
        ws2.cell(row=row, column=5, value=f"=STDEV('个人得分汇总'!{src_col}5:{src_col}34)")
        # 及格人数
        ws2.cell(row=row, column=6, value=f'=COUNTIF(\'个人得分汇总\'!{src_col}5:{src_col}34,">=6.0")')
        # 及格率
        ws2.cell(row=row, column=7, value=f'=COUNTIF(\'个人得分汇总\'!{src_col}5:{src_col}34,">=6.0")/30')
        # 等级
        ws2.cell(row=row, column=8, value=f'=IF(B{row}>=8,"卓越",IF(B{row}>=6.5,"优秀",IF(B{row}>=5,"良好",IF(B{row}>=4,"中等","需提升"))))')

        for col_idx in range(2, 9):
            cell = ws2.cell(row=row, column=col_idx)
            if col_idx in [2, 3, 4, 5]:
                cell.number_format = "0.0"
                cell.font = make_font("008000", bold=True, size=10)
            elif col_idx == 6:
                cell.number_format = "0"
                cell.font = make_font("008000", bold=True, size=10)
            elif col_idx == 7:
                cell.number_format = "0.0%"
                cell.font = make_font("008000", bold=True, size=10)
            else:
                cell.font = make_font(DARK, bold=True, size=10)
            cell.alignment = align_center
            cell.border = border_all

        ws2.row_dimensions[row].height = 24

    # 列宽
    col_widths2 = [16, 12, 12, 12, 12, 12, 12, 12]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Top/Bottom 学员
    ws2.merge_cells("A12:H12")
    ws2["A12"] = "Top 5 学员（综合得分最高）"
    ws2["A12"].font = make_font(WHITE, bold=True, size=14)
    ws2["A12"].fill = make_fill(RED)
    ws2["A12"].alignment = align_center
    ws2.row_dimensions[12].height = 30

    top_headers = ["排名", "学员编号", "姓名", "综合得分", "心理安全感", "探索空间", "领导者信号", "等级"]
    for col_idx, h in enumerate(top_headers, 1):
        c = ws2.cell(row=13, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, size=11)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all
    ws2.row_dimensions[13].height = 28

    for i in range(5):
        row = 14 + i
        ws2.cell(row=row, column=1, value=i + 1).font = make_font(DARK, bold=True, size=11)
        # 使用 LARGE 函数获取第 i+1 大的综合得分
        ws2.cell(row=row, column=4, value=f"=LARGE('个人得分汇总'!H5:H34,{i+1})")
        # 用 INDEX+MATCH 反查学员编号
        ws2.cell(row=row, column=2, value=f'=INDEX(\'个人得分汇总\'!A5:A34,MATCH(D{row},\'个人得分汇总\'!H5:H34,0))')
        ws2.cell(row=row, column=3, value=f'=INDEX(\'个人得分汇总\'!B5:B34,MATCH(D{row},\'个人得分汇总\'!H5:H34,0))')
        # 5 因素分数
        ws2.cell(row=row, column=5, value=f'=INDEX(\'个人得分汇总\'!C5:C34,MATCH(D{row},\'个人得分汇总\'!H5:H34,0))')
        ws2.cell(row=row, column=6, value=f'=INDEX(\'个人得分汇总\'!E5:E34,MATCH(D{row},\'个人得分汇总\'!H5:H34,0))')
        ws2.cell(row=row, column=7, value=f'=INDEX(\'个人得分汇总\'!G5:G34,MATCH(D{row},\'个人得分汇总\'!H5:H34,0))')
        ws2.cell(row=row, column=8, value=f'=IF(D{row}>=8,"卓越",IF(D{row}>=6.5,"优秀",IF(D{row}>=5,"良好","中等")))')

        for col_idx in range(1, 9):
            cell = ws2.cell(row=row, column=col_idx)
            cell.border = border_all
            if col_idx == 1:
                cell.alignment = align_center
                cell.font = make_font(RED, bold=True, size=12)
            elif col_idx == 2:
                cell.alignment = align_center
                cell.font = make_font(DARK, bold=True, size=10)
            elif col_idx in [3]:
                cell.alignment = align_center
                cell.font = make_font(DARK, size=10)
            elif col_idx in [4, 5, 6, 7]:
                cell.alignment = align_center
                cell.font = make_font("008000", size=10)
                cell.number_format = "0.0"
            else:
                cell.alignment = align_center
                cell.font = make_font(DARK, bold=True, size=10)

        # 隔行底色
        if i % 2 == 0:
            for col_idx in range(1, 9):
                ws2.cell(row=row, column=col_idx).fill = make_fill(GRAY_LIGHT)

    # 弱项诊断
    ws2.merge_cells("A21:H21")
    ws2["A21"] = "弱项诊断（综合得分最低 5 人）"
    ws2["A21"].font = make_font(WHITE, bold=True, size=14)
    ws2["A21"].fill = make_fill(RED)
    ws2["A21"].alignment = align_center
    ws2.row_dimensions[21].height = 30

    bottom_headers = ["排名", "学员编号", "姓名", "综合得分", "主要弱项", "建议提升维度", "30 天重点", "备注"]
    for col_idx, h in enumerate(bottom_headers, 1):
        c = ws2.cell(row=22, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, size=11)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all
    ws2.row_dimensions[22].height = 28

    for i in range(5):
        row = 23 + i
        ws2.cell(row=row, column=1, value=i + 1).font = make_font(DARK, bold=True, size=11)
        ws2.cell(row=row, column=4, value=f"=SMALL('个人得分汇总'!H5:H34,{i+1})")
        ws2.cell(row=row, column=2, value=f'=INDEX(\'个人得分汇总\'!A5:A34,MATCH(D{row},\'个人得分汇总\'!H5:H34,0))')
        ws2.cell(row=row, column=3, value=f'=INDEX(\'个人得分汇总\'!B5:B34,MATCH(D{row},\'个人得分汇总\'!H5:H34,0))')
        # 弱项：用 MIN 找最低维度
        ws2.cell(row=row, column=5, value=f'=INDEX(B3:G3,MATCH(MIN(INDEX(\'个人得分汇总\'!C5:G34,MATCH(D{row},\'个人得分汇总\'!H5:H34,0),0)),INDEX(\'个人得分汇总\'!C5:G34,MATCH(D{row},\'个人得分汇总\'!H5:H34,0),0),0))')
        # 30 天重点 - 简化为提示
        ws2.cell(row=row, column=7, value="重点提升 1-2 个最弱维度")
        ws2.cell(row=row, column=8, value="建议与管理者 1-on-1")

        for col_idx in range(1, 9):
            cell = ws2.cell(row=row, column=col_idx)
            cell.border = border_all
            cell.alignment = align_center if col_idx <= 4 else align_left
            if col_idx == 1:
                cell.font = make_font(RED, bold=True, size=12)
            elif col_idx == 4:
                cell.font = make_font("008000", size=10)
                cell.number_format = "0.0"
            elif col_idx in [2]:
                cell.font = make_font(DARK, bold=True, size=10)
            else:
                cell.font = make_font(DARK, size=10)

        if i % 2 == 0:
            for col_idx in range(1, 9):
                ws2.cell(row=row, column=col_idx).fill = make_fill(GRAY_LIGHT)

    ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
    ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
    ws2.page_setup.fitToPage = True
    ws2.page_setup.fitToWidth = 1

    # ==================== Sheet 3: 雷达图数据 ====================
    ws3 = wb.create_sheet("雷达图数据")
    ws3.merge_cells("A1:G1")
    ws3["A1"] = "5 因素雷达图数据"
    ws3["A1"].font = make_font(WHITE, bold=True, size=18)
    ws3["A1"].fill = make_fill(RED)
    ws3["A1"].alignment = align_center
    ws3.row_dimensions[1].height = 40

    ws3["A3"] = "维度"
    ws3["B3"] = "班级平均"
    ws3["C3"] = "目标值"
    ws3["D3"] = "差距"
    ws3["E3"] = "行业平均"
    ws3["F3"] = "Top 部门平均"
    ws3["G3"] = "提升建议"

    for col_idx, h in enumerate(["维度", "班级平均", "目标值", "差距", "行业平均", "Top 部门平均", "提升建议"], 1):
        c = ws3.cell(row=3, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, size=11)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all

    ws3.row_dimensions[3].height = 28

    radar_dimensions = [
        ("心理安全感", 8.0, 5.5, 9.0, "建立'敢说真话'机制"),
        ("认知多样性", 7.5, 5.0, 8.5, "引入'非我族类'视角"),
        ("探索空间", 7.0, 4.5, 8.0, "设立'15% 自由时间'"),
        ("学习速度", 8.0, 5.5, 8.5, "推行'5 Why 复盘'"),
        ("领导者信号", 7.5, 5.5, 8.5, "持续 12 个月以上"),
    ]

    for i, (dim, target, industry, top, suggestion) in enumerate(radar_dimensions):
        row = 4 + i
        ws3.cell(row=row, column=1, value=dim).font = make_font(DARK, bold=True, size=11)
        ws3.cell(row=row, column=1).alignment = align_left
        ws3.cell(row=row, column=2, value=f"=AVERAGE('班级统计'!B{4+i}:B{4+i})")  # 引用班级统计
        ws3.cell(row=row, column=3, value=target)
        ws3.cell(row=row, column=4, value=f"=C{row}-B{row}")
        ws3.cell(row=row, column=5, value=industry)
        ws3.cell(row=row, column=6, value=top)
        ws3.cell(row=row, column=7, value=suggestion)

        for col_idx in range(1, 8):
            cell = ws3.cell(row=row, column=col_idx)
            cell.border = border_all
            if col_idx == 1:
                cell.font = make_font(DARK, bold=True, size=11)
                cell.alignment = align_left
            elif col_idx == 2:
                cell.font = make_font("008000", bold=True, size=10)
                cell.alignment = align_center
                cell.number_format = "0.0"
            elif col_idx in [3, 5, 6]:
                cell.font = make_font("0000FF", size=10)
                cell.alignment = align_center
                cell.number_format = "0.0"
            elif col_idx == 4:
                cell.font = make_font("008000", bold=True, size=10)
                cell.alignment = align_center
                cell.number_format = "0.0"
            else:
                cell.font = make_font(DARK, size=10)
                cell.alignment = align_left

        if i % 2 == 0:
            for col_idx in range(1, 8):
                ws3.cell(row=row, column=col_idx).fill = make_fill(GRAY_LIGHT)

        ws3.row_dimensions[row].height = 26

    # 列宽
    for i, w in enumerate([16, 12, 12, 12, 12, 14, 30], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # 雷达图绘制提示
    ws3.merge_cells("A11:G11")
    ws3["A11"] = "📊 雷达图绘制指南"
    ws3["A11"].font = make_font(WHITE, bold=True, size=14)
    ws3["A11"].fill = make_fill(DARK)
    ws3["A11"].alignment = align_center
    ws3.row_dimensions[11].height = 28

    guide_text = [
        "1. 选择 A3:F8 数据区域",
        "2. 插入 → 图表 → 雷达图 → 填充式雷达图",
        "3. 5 个维度形成五边形",
        "4. 对比'班级平均'与'目标值'的差距",
        "5. 红色填充区域 = 班级水平，蓝色线条 = 目标",
        "6. 突出弱项（差距最大维度）作为下阶段重点",
    ]
    for i, t in enumerate(guide_text):
        ws3.cell(row=12 + i, column=1, value=t).font = make_font(DARK, size=11)
        ws3.merge_cells(start_row=12 + i, start_column=1, end_row=12 + i, end_column=7)
        ws3.cell(row=12 + i, column=1).alignment = align_left
        ws3.row_dimensions[12 + i].height = 22

    ws3.page_setup.orientation = ws3.ORIENTATION_LANDSCAPE
    ws3.page_setup.paperSize = ws3.PAPERSIZE_A4

    # ==================== Sheet 4: 行动建议模板 ====================
    ws4 = wb.create_sheet("行动建议模板")
    ws4.merge_cells("A1:F1")
    ws4["A1"] = "30 天行动建议模板（基于测评结果）"
    ws4["A1"].font = make_font(WHITE, bold=True, size=18)
    ws4["A1"].fill = make_fill(RED)
    ws4["A1"].alignment = align_center
    ws4.row_dimensions[1].height = 40

    ws4.merge_cells("A2:F2")
    ws4["A2"] = "学员姓名：____________ 班级：____________ 日期：____________"
    ws4["A2"].font = make_font(DARK, size=11, italic=True)
    ws4["A2"].alignment = align_center
    ws4.row_dimensions[2].height = 22

    headers4 = ["周次", "主题", "弱项诊断", "3 个具体行动", "资源支持", "评估标准"]
    for col_idx, h in enumerate(headers4, 1):
        c = ws4.cell(row=4, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, size=11)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all

    ws4.row_dimensions[4].height = 30

    weeks = [
        ("Week 1", "心理安全感", "团队不敢说真话", [
            "主动说 1 次'我不确定'",
            "邀请 1 位下属挑战你的决策",
            "开'失败分享会'分享自己的 1 个失败",
        ], "时间支持（30 分钟/月）+ 失败容忍承诺", "敢说真话频率 +50%"),
        ("Week 2", "认知多样性", "团队视角单一", [
            "邀请 1 位'非我族类'参与决策",
            "听 1 个跨行业分享",
            "和'反向意见者'1-on-1",
        ], "外部专家邀请预算 + 跨部门协调", "决策参与者多样性 +30%"),
        ("Week 3", "探索空间", "团队不敢试错", [
            "给团队留 10% 自由时间",
            "启动 1 个小实验",
            "容忍 1 个'暂时看不到结果'的尝试",
        ], "创新预算（5%）+ 实验审批简化", "新项目数 +3 个"),
        ("Week 4", "学习速度 + 领导者信号", "复盘流于形式 + 信号弱", [
            "开 5 Why 复盘会 1 次",
            "公开承认自己的 1 个错误",
            "写 1 封'领导者的信'",
        ], "复盘时间保护 + 公开承诺", "5 Why 使用率 100% + 公开承认 1+ 次"),
    ]

    row = 5
    for week in weeks:
        week_name, theme, diagnosis, actions, resource, evaluation = week
        ws4.cell(row=row, column=1, value=week_name).font = make_font(DARK, bold=True, size=11)
        ws4.cell(row=row, column=2, value=theme).font = make_font(RED, bold=True, size=11)
        ws4.cell(row=row, column=3, value=diagnosis).font = make_font(DARK, size=11)
        action_text = "\n".join(f"• {a}" for a in actions)
        ws4.cell(row=row, column=4, value=action_text).font = make_font(DARK, size=10)
        ws4.cell(row=row, column=5, value=resource).font = make_font(DARK, size=10)
        ws4.cell(row=row, column=6, value=evaluation).font = make_font(DARK, size=10)

        for col_idx in range(1, 7):
            cell = ws4.cell(row=row, column=col_idx)
            cell.border = border_all
            cell.alignment = align_left
            cell.alignment = Alignment(
                horizontal="left" if col_idx >= 3 else "center",
                vertical="center",
                wrap_text=True
            )

        ws4.row_dimensions[row].height = 100
        row += 1

    # 签名行
    ws4.merge_cells(f"A{row+1}:F{row+1}")
    ws4.cell(row=row+1, column=1, value="学员签名：________________    管理者签名：________________    日期：________________").font = make_font(DARK, bold=True, size=12)
    ws4.cell(row=row+1, column=1).alignment = align_center
    ws4.row_dimensions[row+1].height = 30

    # 列宽
    for i, w in enumerate([12, 18, 18, 36, 22, 20], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    ws4.page_setup.orientation = ws4.ORIENTATION_LANDSCAPE
    ws4.page_setup.paperSize = ws4.PAPERSIZE_A4

    # 输出
    output_path = f"{OUTPUT_DIR}/5因素测评汇总表.xlsx"
    wb.save(output_path)
    print(f"✅ 已创建: {output_path}")


# ==================== 文件 3：30 天行动跟进表 ====================
def create_30day_tracker():
    wb = Workbook()

    # Sheet 1: 30 天每日打卡
    ws1 = wb.active
    ws1.title = "30天每日打卡"

    ws1.merge_cells("A1:F1")
    ws1["A1"] = "30 天行动每日打卡表"
    ws1["A1"].font = make_font(WHITE, bold=True, size=18)
    ws1["A1"].fill = make_fill(RED)
    ws1["A1"].alignment = align_center
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells("A2:F2")
    ws1["A2"] = "学员姓名：____________ 班级：____________ 课程结束日期：____________"
    ws1["A2"].font = make_font(DARK, size=11, italic=True)
    ws1["A2"].alignment = align_center
    ws1.row_dimensions[2].height = 22

    headers = ["日期", "星期", "主题", "今日承诺", "完成情况", "反思"]
    for col_idx, h in enumerate(headers, 1):
        c = ws1.cell(row=3, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, size=11)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all
    ws1.row_dimensions[3].height = 28

    # 列宽
    col_widths = [14, 8, 16, 36, 12, 32]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # 30 天内容（按周主题）
    days = [
        ("Week 1：心理安全感", [
            ("Day 1", "主动说一次'我不确定'"),
            ("Day 2", "主动承认一个错误"),
            ("Day 3", "邀请一个不同意见者参与决策"),
            ("Day 4", "开'失败分享会'"),
            ("Day 5", "公开表扬一个敢说真话的下属"),
            ("Day 6", "问下属一个'笨问题'"),
            ("Day 7", "周末反思 + 写日记"),
        ]),
        ("Week 2：认知多样性", [
            ("Day 8", "主动接触一个'非我族类'的观点"),
            ("Day 9", "邀请跨部门同事参与决策"),
            ("Day 10", "开'决策听证会'——听不同声音"),
            ("Day 11", "参加一个跨行业活动"),
            ("Day 12", "和不同年龄的下属深度对话"),
            ("Day 13", "故意挑战自己的'常识'"),
            ("Day 14", "周末反思 + 写日记"),
        ]),
        ("Week 3：探索空间", [
            ("Day 15", "给团队留'自由时间'"),
            ("Day 16", "启动一个小实验"),
            ("Day 17", "容忍一个'暂时看不到结果'的尝试"),
            ("Day 18", "开'创新预算'会议"),
            ("Day 19", "和团队一起复盘一个失败"),
            ("Day 20", "组织一次'跨部门创意会'"),
            ("Day 21", "周末反思 + 写日记"),
        ]),
        ("Week 4：学习速度 + 领导者信号", [
            ("Day 22", "开 5 Why 复盘会"),
            ("Day 23", "跨行业学习一次"),
            ("Day 24", "否定自己过去的一个决策"),
            ("Day 25", "公开表扬一个创新行为"),
            ("Day 26", "写一封'领导者的信'"),
            ("Day 27", "亲自参与一次复盘"),
            ("Day 28", "和团队共创'下季度方向'"),
            ("Day 29", "综合反思"),
            ("Day 30", "30 天总结 + 90 天承诺"),
        ]),
    ]

    row = 4
    for week_name, week_days in days:
        # 周次合并行
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws1.cell(row=row, column=1, value=week_name)
        c.font = make_font(WHITE, bold=True, size=12)
        c.fill = make_fill(RED)
        c.alignment = align_center
        c.border = border_all
        for col_idx in range(1, 7):
            ws1.cell(row=row, column=col_idx).fill = make_fill(RED)
            ws1.cell(row=row, column=col_idx).border = border_all
        ws1.row_dimensions[row].height = 26
        row += 1

        for day, action in week_days:
            ws1.cell(row=row, column=1, value=f"2026-{row-2:02d}").alignment = align_center
            ws1.cell(row=row, column=1).font = make_font(DARK, size=10)

            ws1.cell(row=row, column=2, value=day).font = make_font(RED, bold=True, size=10)
            ws1.cell(row=row, column=2).alignment = align_center

            ws1.cell(row=row, column=3, value=week_name.split("：")[1]).font = make_font(DARK, size=10)
            ws1.cell(row=row, column=3).alignment = align_center

            ws1.cell(row=row, column=4, value=action).font = make_font(DARK, size=10)
            ws1.cell(row=row, column=4).alignment = align_left

            # 完成情况（下拉）
            cell_completion = ws1.cell(row=row, column=5)
            cell_completion.font = make_font("0000FF", size=10)
            cell_completion.alignment = align_center

            # 反思
            ws1.cell(row=row, column=6, value="").font = make_font(DARK, size=10)
            ws1.cell(row=row, column=6).alignment = align_left

            for col_idx in range(1, 7):
                ws1.cell(row=row, column=col_idx).border = border_all

            ws1.row_dimensions[row].height = 32
            row += 1

    # 完成情况 - 数据验证
    dv_completion = DataValidation(type="list", formula1='"已完成,部分完成,未完成,跳过"', allow_blank=True)
    dv_completion.add(f"E5:E{row-1}")
    ws1.add_data_validation(dv_completion)

    # 条件格式
    ws1.conditional_formatting.add(
        f"E5:E{row-1}",
        FormulaRule(formula=[f'EXACT(E5,"已完成")'], fill=make_fill(GREEN))
    )
    ws1.conditional_formatting.add(
        f"E5:E{row-1}",
        FormulaRule(formula=[f'EXACT(E5,"未完成")'], fill=make_fill(RED_FILL))
    )
    ws1.conditional_formatting.add(
        f"E5:E{row-1}",
        FormulaRule(formula=[f'EXACT(E5,"部分完成")'], fill=make_fill(YELLOW))
    )

    # 统计行
    stat_row = row + 1
    ws1.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=6)
    c = ws1.cell(row=stat_row, column=1, value="30 天完成率统计")
    c.font = make_font(WHITE, bold=True, size=14)
    c.fill = make_fill(DARK)
    c.alignment = align_center
    for col_idx in range(1, 7):
        ws1.cell(row=stat_row, column=col_idx).fill = make_fill(DARK)
        ws1.cell(row=stat_row, column=col_idx).border = border_all
    ws1.row_dimensions[stat_row].height = 30

    stats = [
        ("已完成天数", f'=COUNTIF(E5:E{row-1},"已完成")'),
        ("部分完成天数", f'=COUNTIF(E5:E{row-1},"部分完成")'),
        ("未完成天数", f'=COUNTIF(E5:E{row-1},"未完成")'),
        ("跳过天数", f'=COUNTIF(E5:E{row-1},"跳过")'),
        ("完成率", f'=COUNTIF(E5:E{row-1},"已完成")/30'),
    ]

    for i, (label, formula) in enumerate(stats):
        r = stat_row + 1 + i
        ws1.cell(row=r, column=1, value=label).font = make_font(DARK, bold=True, size=11)
        ws1.cell(row=r, column=1).alignment = align_left
        ws1.cell(row=r, column=1).border = border_all
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        for col_idx in range(1, 5):
            ws1.cell(row=r, column=col_idx).fill = make_fill(GRAY_LIGHT)
            ws1.cell(row=r, column=col_idx).border = border_all

        c = ws1.cell(row=r, column=5, value=formula)
        c.font = make_font("008000", bold=True, size=12)
        c.alignment = align_center
        c.border = border_all
        c.number_format = "0.0%" if "率" in label else "0"

        ws1.cell(row=r, column=6, value="").border = border_all

        ws1.row_dimensions[r].height = 24

    ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
    ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
    ws1.page_setup.fitToPage = True
    ws1.page_setup.fitToWidth = 1
    ws1.print_title_rows = "1:3"
    ws1.freeze_panes = "A4"

    # ==================== Sheet 2: 管理者对话记录 ====================
    ws2 = wb.create_sheet("管理者对话")

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "管理者 Day 7 / Day 14 / Day 30 对话记录"
    ws2["A1"].font = make_font(WHITE, bold=True, size=18)
    ws2["A1"].fill = make_fill(RED)
    ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 40

    ws2.merge_cells("A2:G2")
    ws2["A2"] = "学员姓名：____________    管理者姓名：____________    起始日期：____________"
    ws2["A2"].font = make_font(DARK, size=11, italic=True)
    ws2["A2"].alignment = align_center
    ws2.row_dimensions[2].height = 22

    headers2 = ["对话日期", "对话类型", "持续时长", "学员分享摘要", "管理者反馈", "支持承诺", "下次约定"]
    for col_idx, h in enumerate(headers2, 1):
        c = ws2.cell(row=3, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, size=11)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all
    ws2.row_dimensions[3].height = 32

    # 3 次对话
    dialogs = [
        ("Day 7", "建立关系 + 初步反馈", "30 分钟",
         "Week 1 心理安全感主题的 5 大收获",
         "3 个肯定 + 1 个建议",
         "提供时间 + 资源支持",
         "Day 14 对话：45 分钟 + 深度复盘"),
        ("Day 14", "深度复盘 + 资源支持", "45 分钟",
         "Week 1-2 的完整进展 + 资源需求",
         "资源支持承诺 + 障碍识别",
         "人/财/时间支持清单",
         "Day 30 对话：60 分钟 + 成果确认"),
        ("Day 30", "成果确认 + 持续承诺", "60 分钟",
         "30 天全部成果 + 90 天方向",
         "整体反馈 + 长期支持承诺",
         "持续支持承诺 + 1 年长期约定",
         "课后 90 天复测安排"),
    ]

    for i, (day_type, theme, duration, share, feedback, support, next_step) in enumerate(dialogs):
        row = 4 + i
        ws2.cell(row=row, column=1, value=f"2026-{i*7+8:02d}").font = make_font(DARK, size=10)
        ws2.cell(row=row, column=1).alignment = align_center

        ws2.cell(row=row, column=2, value=day_type).font = make_font(RED, bold=True, size=12)
        ws2.cell(row=row, column=2).alignment = align_center

        ws2.cell(row=row, column=3, value=duration).font = make_font(DARK, size=10)
        ws2.cell(row=row, column=3).alignment = align_center

        ws2.cell(row=row, column=4, value=share).font = make_font(DARK, size=10)
        ws2.cell(row=row, column=5, value=feedback).font = make_font(DARK, size=10)
        ws2.cell(row=row, column=6, value=support).font = make_font(DARK, size=10)
        ws2.cell(row=row, column=7, value=next_step).font = make_font(DARK, size=10)

        for col_idx in range(1, 8):
            cell = ws2.cell(row=row, column=col_idx)
            cell.border = border_all
            if col_idx >= 4:
                cell.alignment = align_left
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = align_center

        ws2.row_dimensions[row].height = 100

        if i % 2 == 0:
            for col_idx in range(1, 8):
                ws2.cell(row=row, column=col_idx).fill = make_fill(GRAY_LIGHT)

    # 关键对话话术
    ws2.merge_cells("A8:G8")
    c = ws2.cell(row=8, column=1, value="📋 关键对话话术")
    c.font = make_font(WHITE, bold=True, size=14)
    c.fill = make_fill(DARK)
    c.alignment = align_center
    for col_idx in range(1, 8):
        ws2.cell(row=8, column=col_idx).fill = make_fill(DARK)
        ws2.cell(row=8, column=col_idx).border = border_all
    ws2.row_dimensions[8].height = 30

    quotes = [
        "Day 7：「过去 7 天，你做的哪件事最让你自豪？」「我观察到你做了什么？」「下一次，我想看到你做什么？」",
        "Day 14：「过去 14 天，你最大的 1 个突破是什么？」「你遇到的 1 个最大障碍是什么？我能帮什么？」",
        "Day 30：「30 天前，你给自己定的 3 个目标，完成情况如何？」「你最大的 3 个改变是什么？」",
    ]

    for i, quote in enumerate(quotes):
        row = 9 + i
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws2.cell(row=row, column=1, value=quote)
        c.font = make_font(DARK, size=11)
        c.alignment = align_left
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border_all
        for col_idx in range(1, 8):
            ws2.cell(row=row, column=col_idx).border = border_all
        ws2.row_dimensions[row].height = 50

    # 列宽
    for i, w in enumerate([12, 14, 12, 24, 24, 22, 24], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
    ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
    ws2.page_setup.fitToPage = True
    ws2.page_setup.fitToWidth = 1

    # ==================== Sheet 3: 完成率统计 ====================
    ws3 = wb.create_sheet("完成率统计")

    ws3.merge_cells("A1:E1")
    ws3["A1"] = "30 天行动完成率统计"
    ws3["A1"].font = make_font(WHITE, bold=True, size=18)
    ws3["A1"].fill = make_fill(RED)
    ws3["A1"].alignment = align_center
    ws3.row_dimensions[1].height = 40

    headers3 = ["学员编号", "Week 1", "Week 2", "Week 3", "Week 4", "整体完成率", "评分等级"]
    for col_idx, h in enumerate(headers3, 1):
        c = ws3.cell(row=3, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, size=11)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all
    ws3.row_dimensions[3].height = 32

    # 列宽
    for i, w in enumerate([10, 14, 14, 14, 14, 14, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # 30 个学员
    for i in range(1, 31):
        row = 3 + i
        ws3.cell(row=row, column=1, value=f"S{i:03d}").font = make_font(DARK, bold=True, size=10)
        ws3.cell(row=row, column=1).alignment = align_center

        # 4 个周完成率（公式）
        ws3.cell(row=row, column=2, value=f'=COUNTIF(\'30天每日打卡\'!E{4+(i-1)*30+1}:E{4+(i-1)*30+7},"已完成")/7')
        ws3.cell(row=row, column=3, value=f'=COUNTIF(\'30天每日打卡\'!E{4+(i-1)*30+8}:E{4+(i-1)*30+14},"已完成")/7')
        ws3.cell(row=row, column=4, value=f'=COUNTIF(\'30天每日打卡\'!E{4+(i-1)*30+15}:E{4+(i-1)*30+21},"已完成")/7')
        ws3.cell(row=row, column=5, value=f'=COUNTIF(\'30天每日打卡\'!E{4+(i-1)*30+22}:E{4+(i-1)*30+30},"已完成")/9')

        # 整体完成率
        ws3.cell(row=row, column=6, value=f'=AVERAGE(B{row}:E{row})')
        # 等级
        ws3.cell(row=row, column=7, value=f'=IF(F{row}>=0.9,"卓越",IF(F{row}>=0.75,"优秀",IF(F{row}>=0.6,"良好","需提升")))')

        for col_idx in range(2, 7):
            cell = ws3.cell(row=row, column=col_idx)
            cell.font = make_font("008000", size=10)
            cell.alignment = align_center
            cell.number_format = "0.0%"

        ws3.cell(row=row, column=7).font = make_font(DARK, bold=True, size=10)
        ws3.cell(row=row, column=7).alignment = align_center

        for col_idx in range(1, 8):
            ws3.cell(row=row, column=col_idx).border = border_all

        if i % 2 == 0:
            for col_idx in range(1, 8):
                ws3.cell(row=row, column=col_idx).fill = make_fill(GRAY_LIGHT)

        ws3.row_dimensions[row].height = 24

    # 条件格式：完成率 >= 90% 绿色，60-90% 黄色，<60% 红色
    ws3.conditional_formatting.add(
        "F4:F33",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.9"], fill=make_fill(GREEN))
    )
    ws3.conditional_formatting.add(
        "F4:F33",
        CellIsRule(operator="between", formula=["0.6", "0.9"], fill=make_fill(YELLOW))
    )
    ws3.conditional_formatting.add(
        "F4:F33",
        CellIsRule(operator="lessThan", formula=["0.6"], fill=make_fill(RED_FILL))
    )

    # 班级整体
    stat_row = 35
    ws3.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=7)
    c = ws3.cell(row=stat_row, column=1, value="班级整体完成率")
    c.font = make_font(WHITE, bold=True, size=14)
    c.fill = make_fill(DARK)
    c.alignment = align_center
    for col_idx in range(1, 8):
        ws3.cell(row=stat_row, column=col_idx).fill = make_fill(DARK)
        ws3.cell(row=stat_row, column=col_idx).border = border_all
    ws3.row_dimensions[stat_row].height = 30

    stats = [
        ("平均", "AVERAGE"),
        ("最高", "MAX"),
        ("最低", "MIN"),
        ("中位数", "MEDIAN"),
        ("及格率(>=60%)", "COUNTIF"),
    ]

    for i, (label, func) in enumerate(stats):
        r = stat_row + 1 + i
        ws3.cell(row=r, column=1, value=label).font = make_font(DARK, bold=True, size=11)
        ws3.cell(row=r, column=1).alignment = align_left
        ws3.cell(row=r, column=1).border = border_all
        ws3.cell(row=r, column=1).fill = make_fill(GRAY_LIGHT)

        for col_idx in range(2, 8):
            col_letter = get_column_letter(col_idx)
            if func == "COUNTIF":
                formula = f'=COUNTIF({col_letter}4:{col_letter}33,">=0.6")/30'
            else:
                formula = f"={func}({col_letter}4:{col_letter}33)"
            cell = ws3.cell(row=r, column=col_idx, value=formula)
            cell.font = make_font("008000", bold=True, size=11)
            cell.alignment = align_center
            cell.number_format = "0.0%" if func == "COUNTIF" else "0.0%"
            cell.border = border_all
            cell.fill = make_fill(GRAY_LIGHT)

        ws3.row_dimensions[r].height = 24

    ws3.page_setup.orientation = ws3.ORIENTATION_LANDSCAPE
    ws3.page_setup.paperSize = ws3.PAPERSIZE_A4
    ws3.page_setup.fitToPage = True
    ws3.page_setup.fitToWidth = 1
    ws3.print_title_rows = "1:3"
    ws3.freeze_panes = "B4"

    # ==================== Sheet 4: 班级跟进总表 ====================
    ws4 = wb.create_sheet("班级跟进总表")

    ws4.merge_cells("A1:I1")
    ws4["A1"] = "30 天跟进班级总表（30 人）"
    ws4["A1"].font = make_font(WHITE, bold=True, size=18)
    ws4["A1"].fill = make_fill(RED)
    ws4["A1"].alignment = align_center
    ws4.row_dimensions[1].height = 40

    headers4 = ["编号", "姓名", "部门", "完成率", "Day 7 对话", "Day 14 对话", "Day 30 对话", "管理者", "状态"]
    for col_idx, h in enumerate(headers4, 1):
        c = ws4.cell(row=3, column=col_idx, value=h)
        c.font = make_font(WHITE, bold=True, size=11)
        c.fill = make_fill(DARK)
        c.alignment = align_center
        c.border = border_all
    ws4.row_dimensions[3].height = 28

    for i in range(1, 31):
        row = 3 + i
        ws4.cell(row=row, column=1, value=f"S{i:03d}").font = make_font(DARK, bold=True, size=10)
        ws4.cell(row=row, column=1).alignment = align_center

        ws4.cell(row=row, column=2, value="").font = make_font(DARK, size=10)
        ws4.cell(row=row, column=2).alignment = align_center

        ws4.cell(row=row, column=3, value="").font = make_font(DARK, size=10)
        ws4.cell(row=row, column=3).alignment = align_center

        # 完成率（公式）
        ws4.cell(row=row, column=4, value=f"='完成率统计'!F{row}")
        ws4.cell(row=row, column=4).font = make_font("008000", bold=True, size=10)
        ws4.cell(row=row, column=4).alignment = align_center
        ws4.cell(row=row, column=4).number_format = "0.0%"

        # Day 7/14/30 对话（下拉）
        for col_idx in [5, 6, 7]:
            ws4.cell(row=row, column=col_idx).font = make_font("0000FF", size=10)
            ws4.cell(row=row, column=col_idx).alignment = align_center

        ws4.cell(row=row, column=8, value="").font = make_font(DARK, size=10)
        ws4.cell(row=row, column=8).alignment = align_center

        # 状态（公式）
        ws4.cell(row=row, column=9, value=f'=IF(F{row}="已完成","已结束",IF(E{row}="已完成","进行中","未开始"))')
        ws4.cell(row=row, column=9).font = make_font(DARK, bold=True, size=10)
        ws4.cell(row=row, column=9).alignment = align_center

        for col_idx in range(1, 10):
            ws4.cell(row=row, column=col_idx).border = border_all

        if i % 2 == 0:
            for col_idx in range(1, 10):
                ws4.cell(row=row, column=col_idx).fill = make_fill(GRAY_LIGHT)

        ws4.row_dimensions[row].height = 24

    # 数据验证
    dv_dialog = DataValidation(type="list", formula1='"已完成,进行中,未开始,跳过"', allow_blank=True)
    dv_dialog.add("E4:G33")
    ws4.add_data_validation(dv_dialog)

    # 条件格式：状态
    ws4.conditional_formatting.add(
        "I4:I33",
        FormulaRule(formula=['EXACT(I4,"已结束")'], fill=make_fill(GREEN))
    )
    ws4.conditional_formatting.add(
        "I4:I33",
        FormulaRule(formula=['EXACT(I4,"未开始")'], fill=make_fill(RED_FILL))
    )

    # 列宽
    for i, w in enumerate([10, 14, 18, 12, 14, 14, 14, 14, 14], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    ws4.page_setup.orientation = ws4.ORIENTATION_LANDSCAPE
    ws4.page_setup.paperSize = ws4.PAPERSIZE_A4
    ws4.page_setup.fitToPage = True
    ws4.page_setup.fitToWidth = 1
    ws4.print_title_rows = "1:3"
    ws4.freeze_panes = "D4"

    # 输出
    output_path = f"{OUTPUT_DIR}/30天行动跟进表.xlsx"
    wb.save(output_path)
    print(f"✅ 已创建: {output_path}")


# ==================== 主程序 ====================
if __name__ == "__main__":
    print("📊 开始创建 3 个 Excel 配套表单...")
    print("=" * 60)
    create_registration_form()
    create_assessment_summary()
    create_30day_tracker()
    print("=" * 60)
    print("✅ 全部完成！")

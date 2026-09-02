# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = "D:/新课开发/管理学/11-人才盘点与梯队建设/配套表单和指引-Excel版"

THIN_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
SUBHEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
SUBHEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
LABEL_FONT = Font(name='微软雅黑', size=10, bold=True)
INPUT_FONT = Font(name='微软雅黑', size=10)
NOTE_FONT = Font(name='微软雅黑', size=9, italic=True, color='666666')
LIGHT_BLUE_FILL = PatternFill(start_color='D9E8F5', end_color='D9E8F5', fill_type='solid')
LIGHT_GRAY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

def CENTER_ALIGN():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def LEFT_ALIGN():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def merge_and_style(ws, start_row, start_col, end_row, end_col, value,
                   fill=None, font=None, alignment=None):
    ws.merge_cells(start_row=start_row, start_column=start_col,
                  end_row=end_row, end_column=end_col)
    cell = ws.cell(row=start_row, column=start_col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = CENTER_ALIGN()
    cell.border = THIN_BORDER
    return cell

def create_data_row(ws, row, values, fills=None, fonts=None, heights=None):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        if fills and col <= len(fills) and fills[col-1]:
            cell.fill = fills[col-1]
        if fonts and col <= len(fonts) and fonts[col-1]:
            cell.font = fonts[col-1]
        else:
            cell.font = INPUT_FONT
        cell.alignment = LEFT_ALIGN() if col > 1 else CENTER_ALIGN()
        cell.border = THIN_BORDER

def set_column_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

# ==================== CREATE GUIDE FILE ====================
print("Creating guide file...")

wb = Workbook()
wb.remove(wb.active) if wb.active else None

# Cover sheet
ws_cover = wb.create_sheet("封面")

merge_and_style(ws_cover, 1, 1, 1, 8, '人才盘点与梯队建设', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
merge_and_style(ws_cover, 2, 1, 2, 8, '配套表单使用指引', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
merge_and_style(ws_cover, 3, 1, 3, 8, '课程ID: 11-人才盘点与梯队建设', LIGHT_BLUE_FILL, NOTE_FONT, CENTER_ALIGN())
merge_and_style(ws_cover, 5, 1, 5, 8, '表单目录', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

headers = ['编号', '表单名称', '表单类型', '主要使用者', '', '', '', '']
create_data_row(ws_cover, 6, headers, fills=[LIGHT_GRAY_FILL]*8, fonts=[LABEL_FONT]*8)
ws_cover.merge_cells('E6:H6')

forms_list = [
    ('表单1', '九宫格人才评估表', '核心评估工具', '员工+评估人'),
    ('表单2', '人才盘点会议记录表', '会议纪要工具', 'HRBP+主持人'),
    ('表单3', '关键岗位继任地图', '继任计划工具', 'HR总监+业务负责人'),
    ('表单4', '高潜人才发展计划表', '个人发展工具', '员工+直接上级+HRBP'),
    ('表单5', '人才盘点年度日历', '计划管理工具', 'HRBP'),
    ('表单6', '梯队建设进度追踪表', '进度管控工具', 'HR总监'),
]

for i, (num, name, form_type, user) in enumerate(forms_list):
    row = 7 + i
    ws_cover.cell(row=row, column=1, value=num)
    ws_cover.cell(row=row, column=1).font = Font(name='微软雅黑', size=10, bold=True)
    ws_cover.cell(row=row, column=1).alignment = CENTER_ALIGN()
    ws_cover.cell(row=row, column=1).border = THIN_BORDER

    ws_cover.cell(row=row, column=2, value=name)
    ws_cover.cell(row=row, column=2).font = INPUT_FONT
    ws_cover.cell(row=row, column=2).alignment = LEFT_ALIGN()
    ws_cover.cell(row=row, column=2).border = THIN_BORDER

    ws_cover.cell(row=row, column=3, value=form_type)
    ws_cover.cell(row=row, column=3).font = INPUT_FONT
    ws_cover.cell(row=row, column=3).alignment = CENTER_ALIGN()
    ws_cover.cell(row=row, column=3).border = THIN_BORDER

    ws_cover.cell(row=row, column=4, value=user)
    ws_cover.cell(row=row, column=4).font = INPUT_FONT
    ws_cover.cell(row=row, column=4).alignment = CENTER_ALIGN()
    ws_cover.cell(row=row, column=4).border = THIN_BORDER

    ws_cover.merge_cells(f'E{row}:H{row}')
    ws_cover.cell(row=row, column=5).border = THIN_BORDER

# Usage workflow section
merge_and_style(ws_cover, 14, 1, 14, 8, '表单使用流程', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

workflow = [
    ['Step 1', '准备阶段', '收集员工基本信息和历史评估数据', 'HRBP', '盘点会前1周'],
    ['Step 2', '评估阶段', '直接主管对员工进行九宫格评估和潜力评估', '直接主管', '盘点会前3天'],
    ['Step 3', '校准阶段', '召开人才盘点会议，校准评估结果', 'HR+业务负责人', '盘点会议'],
    ['Step 4', '规划阶段', '制定IDP、继任计划、发展举措', '员工+上级+HR', '盘点会后2周'],
    ['Step 5', '跟踪阶段', '执行计划并定期跟踪进度', 'HRBP', '季度Review'],
]

wf_headers = ['步骤', '阶段', '活动内容', '负责人', '时间节点']
headers_wf = wf_headers + ['', '', '', '']
create_data_row(ws_cover, 15, headers_wf, fills=[LIGHT_GRAY_FILL]*8, fonts=[LABEL_FONT]*8)
ws_cover.merge_cells('E15:H15')

for i, wf in enumerate(workflow):
    row = 16 + i
    for j, val in enumerate(wf):
        ws_cover.cell(row=row, column=j+1, value=val)
        ws_cover.cell(row=row, column=j+1).font = Font(name='微软雅黑', size=10, bold=True, color='1F4E79') if j == 0 else INPUT_FONT
        ws_cover.cell(row=row, column=j+1).fill = LIGHT_BLUE_FILL if j == 0 else PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
        ws_cover.cell(row=row, column=j+1).alignment = CENTER_ALIGN() if j in [0, 3, 4] else LEFT_ALIGN()
        ws_cover.cell(row=row, column=j+1).border = THIN_BORDER
    ws_cover.merge_cells(f'E{row}:H{row}')
    ws_cover.row_dimensions[row].height = 25

set_column_widths(ws_cover, {'A': 10, 'B': 18, 'C': 28, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 14})
print("Cover sheet done")

# Guide sheets
form_guides = [
    ("九宫格人才评估表",
     "九宫格人才评估表是人才盘点的核心工具，通过业绩和潜力两个维度将员工定位到9个格子中。",
     ["确定评估对象范围", "收集评估周期内的业绩数据", "直接主管进行业绩评分", "直接主管进行潜力评分", "系统自动计算并落位", "HRBP与主管校准", "盘点会议确认结果"],
     ["评估标准要客观公正", "潜力评估重点看学习能力、适应变化、领导力", "九宫格落位是动态的"]),
    ("人才盘点会议记录表",
     "人才盘点会议记录表用于记录人才盘点会议的完整内容。",
     ["会前准备", "记录会议基本信息", "确认九宫格分布", "讨论继任计划", "明确行动项", "约定下次会议", "会议纪要确认"],
     ["会议频率建议季度盘点", "行动项必须有明确责任人", "继任计划要讨论备选方案"]),
    ("关键岗位继任地图",
     "关键岗位继任地图用于追踪公司关键岗位的继任准备情况。",
     ["确定关键岗位范围", "确认继任需求", "评估现任者任职风险", "识别潜在继任候选人", "沟通发展意愿", "定期Review进展", "及时更新继任地图"],
     ["关键岗位定义标准", "继任准备度1-4分评估", "每个关键岗位储备1-2名继任者"]),
    ("高潜人才发展计划表",
     "高潜人才发展计划表是针对高潜力人才的个性化发展工具。",
     ["识别高潜人才", "进行能力评估", "沟通职业期望", "分析能力差距", "确定发展方式", "设定里程碑", "定期Review进展"],
     ["高潜识别标准：潜力>=4分", "发展目标要SMART", "发展方式要多元化"]),
    ("人才盘点年度日历",
     "人才盘点年度日历是人才盘点工作的年度规划工具。",
     ["年初制定年度计划", "明确每月活动内容", "分配责任人", "标注关键节点", "按月跟踪状态", "季度Review", "年底总结"],
     ["人才盘点是常态化工作", "关键活动提前规划", "完成状态可视化"]),
    ("梯队建设进度追踪表",
     "梯队建设进度追踪表用于追踪公司各梯队层级的建设进度。",
     ["确定梯队层级定义", "建立关键岗位清单", "盘点现有人才储备", "计算建设进度", "识别空缺风险", "制定补充计划", "定期跟踪更新"],
     ["梯队层级定义要清晰", "建设进度动态指标", "高风险梯队重点关注"]),
]

for form_name, desc, steps, notes in form_guides:
    ws = wb.create_sheet()
    ws.title = form_name[:31]

    merge_and_style(ws, 1, 1, 1, 6, f'{form_name} - 使用指引', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
    merge_and_style(ws, 2, 1, 2, 6, '表单简介', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    ws.merge_cells('A3:F5')
    ws['A3'] = desc
    ws['A3'].font = Font(name='微软雅黑', size=10)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws['A3'].border = THIN_BORDER

    merge_and_style(ws, 6, 1, 6, 6, '使用步骤', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    for i, step in enumerate(steps):
        row = 7 + i
        ws[f'A{row}'] = f'第{i+1}步'
        ws[f'A{row}'].font = Font(name='微软雅黑', size=10, bold=True, color='1F4E79')
        ws[f'A{row}'].fill = LIGHT_BLUE_FILL
        ws[f'A{row}'].alignment = CENTER_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'] = step
        ws[f'B{row}'].font = INPUT_FONT
        ws[f'B{row}'].alignment = LEFT_ALIGN()
        ws[f'B{row}'].border = THIN_BORDER
        ws.row_dimensions[row].height = 25

    start_row = 7 + len(steps) + 1
    merge_and_style(ws, start_row, 1, start_row, 6, '注意事项', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    for i, note in enumerate(notes):
        row = start_row + 1 + i
        ws[f'A{row}'] = '!'
        ws[f'A{row}'].font = Font(name='微软雅黑', size=12, bold=True, color='FF0000')
        ws[f'A{row}'].alignment = CENTER_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'] = note
        ws[f'B{row}'].font = NOTE_FONT
        ws[f'B{row}'].alignment = LEFT_ALIGN()
        ws[f'B{row}'].border = THIN_BORDER
        ws.row_dimensions[row].height = 22

    set_column_widths(ws, {'A': 8, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16})

print("Guide sheets done")

import os
os.makedirs(OUTPUT_DIR, exist_ok=True)
output_path = os.path.join(OUTPUT_DIR, '表单使用指引.xlsx')
wb.save(output_path)
print(f"Saved: {output_path}")

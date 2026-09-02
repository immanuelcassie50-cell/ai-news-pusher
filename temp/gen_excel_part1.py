"""
教练技术·新的管理方式 · 7个 Excel 配套表单生成器
配色:深紫 #5B3A8C + 暖橙 #F2A03D + 墨黑 #1F1F2E + 米白 #FBF9F4 + 浅灰 #E8E5E0
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

OUT_DIR = r"D:\2026年课程\竞越\教练技术：新的管理方式\完整课程包\13_Excel配套表单"
os.makedirs(OUT_DIR, exist_ok=True)

# 配色
PURPLE = "5B3A8C"      # 深紫
ORANGE = "F2A03D"      # 暖橙
INK = "1F1F2E"         # 墨黑
CREAM = "FBF9F4"       # 米白
GREY = "E8E5E0"        # 浅灰
WHITE = "FFFFFF"
PURPLE_LIGHT = "8A6FD8"
TERRACOTTA = "B85C3E"
GOLD = "C9A961"

# 通用样式
def title_font():
    return Font(name="思源黑体", size=18, bold=True, color=PURPLE)

def header_font():
    return Font(name="思源黑体", size=11, bold=True, color=WHITE)

def data_font():
    return Font(name="思源黑体", size=10, color=INK)

def note_font():
    return Font(name="思源黑体", size=9, color="888888", italic=True)

def purple_fill():
    return PatternFill("solid", fgColor=PURPLE)

def orange_fill():
    return PatternFill("solid", fgColor=ORANGE)

def cream_fill():
    return PatternFill("solid", fgColor=CREAM)

def grey_fill():
    return PatternFill("solid", fgColor=GREY)

def thin_border():
    side = Side(border_style="thin", color="D0D0D0")
    return Border(left=side, right=side, top=side, bottom=side)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def apply_header(ws, row, headers, start_col=1):
    """应用表头样式"""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = header_font()
        c.fill = orange_fill()
        c.alignment = center()
        c.border = thin_border()


def apply_title(ws, row, text, span=1):
    """应用大标题"""
    c = ws.cell(row=row, column=1, value=text)
    c.font = title_font()
    c.alignment = left_align()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ============================================================
# 1. F1_预评量表.xlsx
# ============================================================
def build_F1():
    wb = Workbook()
    # 30 道题(5维度×6题)
    questions = [
        # 倾听(6)
        ("我能放下手机,完全专注地听下属说话", "倾听"),
        ("我能听出下属没说出来的话", "倾听"),
        ("我听到下属的情绪并能命名它", "倾听"),
        ("我能在不打断的情况下让下属说完", "倾听"),
        ("我能区分听到和听进去", "倾听"),
        ("我能在对话中保持好奇心,不带评判", "倾听"),
        # 提问(6)
        ("我倾向于用开放式问题引导下属思考", "提问"),
        ("我能问出让下属深入反思的问题", "提问"),
        ("我不会用问题暗示我的标准答案", "提问"),
        ("我会在关键节点问'你想要的是什么'", "提问"),
        ("我能用'如果...你会...'的问题激发创造力", "提问"),
        ("我能在提问后给足沉默让下属思考", "提问"),
        # 反馈(6)
        ("我能给出具体而不评判的反馈", "反馈"),
        ("我先说做得好的,再说要改进的", "反馈"),
        ("我的反馈能让下属感到被支持而非被指责", "反馈"),
        ("我能区分行为和人格来反馈", "反馈"),
        ("我能在公开场合给肯定、私下给建议", "反馈"),
        ("我的反馈能聚焦在未来而非过去", "反馈"),
        # 信任(6)
        ("我愿意让下属尝试并承担失败的成本", "信任"),
        ("我能在下属失误时先解决问题再总结", "信任"),
        ("我信守对下属的承诺,小事也兑现", "信任"),
        ("我能在下属表现不好时保护他的自尊", "信任"),
        ("我允许下属表达与我不同的观点", "信任"),
        ("我展示真实的自己而不只是管理者身份", "信任"),
        # 自我认知(6)
        ("我能觉察到自己在对话中的情绪", "自我认知"),
        ("我能在情绪上来时暂停回应", "自我认知"),
        ("我清楚自己的盲点和偏见", "自我认知"),
        ("我能在被挑战时不防御", "自我认知"),
        ("我能定期反思自己的管理风格", "自我认知"),
        ("我愿意为团队氛围的恶化负起自己的责任", "自我认知"),
    ]

    # Sheet 1: 训前评估
    ws1 = wb.active
    ws1.title = "训前评估"
    apply_title(ws1, 1, "F1 预评量表 · 训前评估", span=5)
    ws1.cell(row=2, column=1, value="评分:1=从不,2=偶尔,3=一般,4=经常,5=总是").font = note_font()
    apply_header(ws1, 4, ["题号", "题目", "评分(1-5)", "维度", "备注"])
    for i, (q, d) in enumerate(questions):
        r = 5 + i
        ws1.cell(row=r, column=1, value=i + 1).alignment = center()
        ws1.cell(row=r, column=2, value=q)
        ws1.cell(row=r, column=4, value=d).alignment = center()
        for col in range(1, 6):
            ws1.cell(row=r, column=col).font = data_font()
            ws1.cell(row=r, column=col).border = thin_border()
            if col != 2:
                ws1.cell(row=r, column=col).alignment = center()
    # 维度汇总
    summary_row = 5 + len(questions) + 1
    ws1.cell(row=summary_row, column=1, value="维度汇总").font = title_font()
    apply_header(ws1, summary_row + 1, ["维度", "题数", "平均分", "等级", "说明"])
    dims = ["倾听", "提问", "反馈", "信任", "自我认知"]
    for i, d in enumerate(dims):
        r = summary_row + 2 + i
        ws1.cell(row=r, column=1, value=d).alignment = center()
        ws1.cell(row=r, column=2, value=6).alignment = center()
        ws1.cell(row=r, column=3, value=f'=AVERAGEIFS(D5:D{4 + len(questions)},D5:D{4 + len(questions)},"{d}",C5:C{4 + len(questions)},"<>")')
        ws1.cell(row=r, column=4, value=f'=IF(C{r}>=4.5,"优秀",IF(C{r}>=3.5,"良好",IF(C{r}>=2.5,"待提升","薄弱")))').alignment = center()
        for col in range(1, 6):
            ws1.cell(row=r, column=col).font = data_font()
            ws1.cell(row=r, column=col).border = thin_border()
    set_col_widths(ws1, [8, 50, 12, 12, 25])

    # Sheet 2: 训后评估(结构同训前)
    ws2 = wb.create_sheet("训后评估")
    apply_title(ws2, 1, "F1 预评量表 · 训后评估", span=5)
    ws2.cell(row=2, column=1, value="评分:1=从不,2=偶尔,3=一般,4=经常,5=总是").font = note_font()
    apply_header(ws2, 4, ["题号", "题目", "评分(1-5)", "维度", "备注"])
    for i, (q, d) in enumerate(questions):
        r = 5 + i
        ws2.cell(row=r, column=1, value=i + 1).alignment = center()
        ws2.cell(row=r, column=2, value=q)
        ws2.cell(row=r, column=4, value=d).alignment = center()
        for col in range(1, 6):
            ws2.cell(row=r, column=col).font = data_font()
            ws2.cell(row=r, column=col).border = thin_border()
            if col != 2:
                ws2.cell(row=r, column=col).alignment = center()
    summary_row = 5 + len(questions) + 1
    ws2.cell(row=summary_row, column=1, value="维度汇总").font = title_font()
    apply_header(ws2, summary_row + 1, ["维度", "题数", "平均分", "等级", "说明"])
    for i, d in enumerate(dims):
        r = summary_row + 2 + i
        ws2.cell(row=r, column=1, value=d).alignment = center()
        ws2.cell(row=r, column=2, value=6).alignment = center()
        ws2.cell(row=r, column=3, value=f'=AVERAGEIFS(D5:D{4 + len(questions)},D5:D{4 + len(questions)},"{d}",C5:C{4 + len(questions)},"<>")')
        ws2.cell(row=r, column=4, value=f'=IF(C{r}>=4.5,"优秀",IF(C{r}>=3.5,"良好",IF(C{r}>=2.5,"待提升","薄弱")))').alignment = center()
        for col in range(1, 6):
            ws2.cell(row=r, column=col).font = data_font()
            ws2.cell(row=r, column=col).border = thin_border()
    set_col_widths(ws2, [8, 50, 12, 12, 25])

    # Sheet 3: 对比报告
    ws3 = wb.create_sheet("对比报告")
    apply_title(ws3, 1, "F1 训前训后对比报告", span=6)
    apply_header(ws3, 3, ["维度", "训前均分", "训后均分", "变化", "变化率", "雷达图说明"])
    for i, d in enumerate(dims):
        r = 4 + i
        ws3.cell(row=r, column=1, value=d).alignment = center()
        ws3.cell(row=r, column=2, value=f"=AVERAGE('训前评估'!D5:D{4 + len(questions)})").alignment = center()  # 占位
        # 实际跨 sheet 引用
        ws3.cell(row=r, column=2, value=f"=AVERAGEIFS('训前评估'!D5:D{4 + len(questions)},'训前评估'!D5:D{4 + len(questions)},\"{d}\")").alignment = center()
        ws3.cell(row=r, column=3, value=f"=AVERAGEIFS('训后评估'!D5:D{4 + len(questions)},'训后评估'!D5:D{4 + len(questions)},\"{d}\")").alignment = center()
        ws3.cell(row=r, column=4, value=f"=C{r}-B{r}").alignment = center()
        ws3.cell(row=r, column=5, value=f"=IF(B{r}=0,0,(C{r}-B{r})/B{r})").alignment = center()
        ws3.cell(row=r, column=5).number_format = "0.0%"
        for col in range(1, 7):
            ws3.cell(row=r, column=col).font = data_font()
            ws3.cell(row=r, column=col).border = thin_border()
    # 整体变化
    last_r = 4 + len(dims)
    ws3.cell(row=last_r + 1, column=1, value="整体总分").font = title_font()
    ws3.cell(row=last_r + 1, column=2, value="=SUM(B4:B" + str(last_r - 1) + ")/5").alignment = center()
    ws3.cell(row=last_r + 1, column=3, value="=SUM(C4:C" + str(last_r - 1) + ")/5").alignment = center()
    ws3.cell(row=last_r + 1, column=4, value=f"=C{last_r + 1}-B{last_r + 1}").alignment = center()
    ws3.cell(row=last_r + 1, column=5, value=f"=IF(B{last_r + 1}=0,0,(C{last_r + 1}-B{last_r + 1})/B{last_r + 1})").alignment = center()
    ws3.cell(row=last_r + 1, column=5).number_format = "0.0%"
    for col in range(1, 7):
        ws3.cell(row=last_r + 1, column=col).font = data_font()
        ws3.cell(row=last_r + 1, column=col).fill = cream_fill()
        ws3.cell(row=last_r + 1, column=col).border = thin_border()
    ws3.cell(row=last_r + 3, column=1, value="结论解读:").font = title_font()
    ws3.cell(row=last_r + 4, column=1, value="1. 训后总分 ≥ 4.0 → 整体教练领导力合格,进入落地实践期")
    ws3.cell(row=last_r + 5, column=1, value="2. 训后总分 3.0-4.0 → 框架已掌握,需 30 天强化练习")
    ws3.cell(row=last_r + 6, column=1, value="3. 训后总分 < 3.0 → 建议复训 + 1V1 辅导")
    for r in range(last_r + 4, last_r + 7):
        ws3.cell(row=r, column=1).font = data_font()
    set_col_widths(ws3, [14, 12, 12, 12, 12, 30])

    wb.save(os.path.join(OUT_DIR, "F1_预评量表.xlsx"))
    print("[OK] F1_预评量表.xlsx")


# ============================================================
# 2. F2_时间账单.xlsx
# ============================================================
def build_F2():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "时间账单"
    apply_title(ws1, 1, "F2 时间账单 · 记录 1 天 12 小时", span=5)
    ws1.cell(row=2, column=1, value="每 30 分钟一格,记录 06:00-18:00 共 24 格").font = note_font()
    apply_header(ws1, 4, ["时间段", "做了什么", "类型", "对象", "价值(高/中/低)"])
    types = ["工作", "会议", "沟通", "休息", "学习", "其他"]
    objs = ["自己", "下属", "上级", "同事", "客户", "其他"]
    for i in range(24):
        h = 6 + i // 2
        m = (i % 2) * 30
        label = f"{h:02d}:{m:02d}"
        r = 5 + i
        c = ws1.cell(row=r, column=1, value=label)
        c.alignment = center()
        # 配色:工作/会议/沟通/休息/学习/其他
        for col in range(1, 6):
            ws1.cell(row=r, column=col).font = data_font()
            ws1.cell(row=r, column=col).border = thin_border()
            ws1.cell(row=r, column=col).alignment = center()
    set_col_widths(ws1, [12, 50, 12, 12, 14])

    # Sheet 2: 统计
    ws2 = wb.create_sheet("统计")
    apply_title(ws2, 1, "F2 时间账单 · 统计分析", span=3)
    apply_header(ws2, 3, ["类型", "格数(30min)", "占比"])
    for i, t in enumerate(types):
        r = 4 + i
        ws2.cell(row=r, column=1, value=t).alignment = center()
        ws2.cell(row=r, column=2, value=f'=COUNTIF(时间账单!C5:C28,"{t}")').alignment = center()
        ws2.cell(row=r, column=3, value=f'=B{r}/24').alignment = center()
        ws2.cell(row=r, column=3).number_format = "0.0%"
        for col in range(1, 4):
            ws2.cell(row=r, column=col).font = data_font()
            ws2.cell(row=r, column=col).border = thin_border()
    last_r = 4 + len(types)
    ws2.cell(row=last_r, column=1, value="合计").font = title_font()
    ws2.cell(row=last_r, column=2, value=f"=SUM(B4:B{last_r - 1})").alignment = center()
    ws2.cell(row=last_r, column=3, value=f"=SUM(C4:C{last_r - 1})").alignment = center()
    for col in range(1, 4):
        ws2.cell(row=last_r, column=col).fill = cream_fill()
        ws2.cell(row=last_r, column=col).font = data_font()
        ws2.cell(row=last_r, column=col).border = thin_border()

    # 对象维度
    ws2.cell(row=last_r + 2, column=1, value="按对象分布").font = title_font()
    apply_header(ws2, last_r + 3, ["对象", "格数", "占比"])
    for i, o in enumerate(objs):
        r = last_r + 4 + i
        ws2.cell(row=r, column=1, value=o).alignment = center()
        ws2.cell(row=r, column=2, value=f'=COUNTIF(时间账单!D5:D28,"{o}")').alignment = center()
        ws2.cell(row=r, column=3, value=f'=B{r}/24').alignment = center()
        ws2.cell(row=r, column=3).number_format = "0.0%"
        for col in range(1, 4):
            ws2.cell(row=r, column=col).font = data_font()
            ws2.cell(row=r, column=col).border = thin_border()

    # 价值维度
    base = last_r + 4 + len(objs) + 1
    ws2.cell(row=base, column=1, value="按价值分布").font = title_font()
    apply_header(ws2, base + 1, ["价值", "格数", "占比"])
    for i, v in enumerate(["高", "中", "低"]):
        r = base + 2 + i
        ws2.cell(row=r, column=1, value=v).alignment = center()
        ws2.cell(row=r, column=2, value=f'=COUNTIF(时间账单!E5:E28,"{v}")').alignment = center()
        ws2.cell(row=r, column=3, value=f'=B{r}/24').alignment = center()
        ws2.cell(row=r, column=3).number_format = "0.0%"
        for col in range(1, 4):
            ws2.cell(row=r, column=col).font = data_font()
            ws2.cell(row=r, column=col).border = thin_border()

    # 诊断
    note_r = base + 2 + 3 + 1
    ws2.cell(row=note_r, column=1, value="诊断解读").font = title_font()
    ws2.cell(row=note_r + 1, column=1, value="• 高价值 + 沟通 + 下属 → 黄金时间(教练对话最好时段)")
    ws2.cell(row=note_r + 2, column=1, value="• 会议 > 4 小时 → 警惕,会议占用黄金时间")
    ws2.cell(row=note_r + 3, column=1, value="• 下属时间 < 2 小时 → 30 天内增加 1V1 心谈")
    for r in range(note_r + 1, note_r + 4):
        ws2.cell(row=r, column=1).font = data_font()
    set_col_widths(ws2, [12, 14, 12])

    wb.save(os.path.join(OUT_DIR, "F2_时间账单.xlsx"))
    print("[OK] F2_时间账单.xlsx")


# ============================================================
# 3. F3_4模式自评.xlsx
# ============================================================
def build_F3():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "团队分类"
    apply_title(ws1, 1, "F3 团队成员 4 模式自评", span=7)
    ws1.cell(row=2, column=1, value="每名团队成员评估 1 次,4 模式:先行者/整合者/观望者/保守者").font = note_font()
    apply_header(ws1, 4, ["姓名", "先行者(1-5)", "整合者(1-5)", "观望者(1-5)", "保守者(1-5)", "主类型", "AI 工具使用频率"])
    # 20 行示例
    for i in range(20):
        r = 5 + i
        ws1.cell(row=r, column=1, value=f"成员{i + 1}").alignment = center()
        # 4 个评分
        for col in range(2, 6):
            ws1.cell(row=r, column=col).alignment = center()
        # 主类型公式
        ws1.cell(row=r, column=6, value=f'=INDEX({{"先行者","整合者","观望者","保守者"}},MATCH(MAX(B{r}:E{r}),B{r}:E{r},0))').alignment = center()
        ws1.cell(row=r, column=7, value="").alignment = center()
        for col in range(1, 8):
            ws1.cell(row=r, column=col).font = data_font()
            ws1.cell(row=r, column=col).border = thin_border()
    set_col_widths(ws1, [12, 14, 14, 14, 14, 14, 18])

    # Sheet 2: 统计
    ws2 = wb.create_sheet("统计")
    apply_title(ws2, 1, "F3 4 模式统计 + 管理动作", span=4)
    apply_header(ws2, 3, ["模式", "人数", "占比", "管理动作"])
    actions = {
        "先行者": "授权 + 挑战性任务,允许试错,纳入创新项目",
        "整合者": "团队桥梁角色,推动协作,作为文化大使",
        "观望者": "1V1 心谈了解顾虑,提供安全尝试机会,小步前进",
        "保守者": "尊重节奏,先稳定本职工作,AI 工具从低风险场景切入",
    }
    for i, m in enumerate(["先行者", "整合者", "观望者", "保守者"]):
        r = 4 + i
        ws2.cell(row=r, column=1, value=m).alignment = center()
        ws2.cell(row=r, column=2, value=f'=COUNTIF(团队分类!F5:F24,"{m}")').alignment = center()
        ws2.cell(row=r, column=3, value=f'=B{r}/20').alignment = center()
        ws2.cell(row=r, column=3).number_format = "0.0%"
        ws2.cell(row=r, column=4, value=actions[m])
        for col in range(1, 5):
            ws2.cell(row=r, column=col).font = data_font()
            ws2.cell(row=r, column=col).border = thin_border()
    last_r = 4 + 4
    ws2.cell(row=last_r, column=1, value="合计").font = title_font()
    ws2.cell(row=last_r, column=2, value=f"=SUM(B4:B{last_r - 1})").alignment = center()
    ws2.cell(row=last_r, column=3, value=f"=SUM(C4:C{last_r - 1})").alignment = center()
    for col in range(1, 5):
        ws2.cell(row=last_r, column=col).fill = cream_fill()
        ws2.cell(row=last_r, column=col).font = data_font()
        ws2.cell(row=last_r, column=col).border = thin_border()
    set_col_widths(ws2, [12, 10, 10, 60])

    wb.save(os.path.join(OUT_DIR, "F3_4模式自评.xlsx"))
    print("[OK] F3_4模式自评.xlsx")


# ============================================================
# 4. F19_行动计划.xlsx
# ============================================================
def build_F19():
    wb = Workbook()
    elements = ["我要达成什么(目标)", "我为什么在乎(意义)", "我第一步做什么(行动)", "我可能遇到什么卡点(障碍)", "我怎么知道自己做到了(衡量)"]
    # 3 个行动,每个 1 个 Sheet
    for action_idx in range(1, 4):
        ws = wb.create_sheet(f"行动{action_idx}")
        apply_title(ws, 1, f"F19 行动计划 · 行动 {action_idx}", span=5)
        ws.cell(row=2, column=1, value="5 要素填写 + 32 天 5 周追踪(每周完成度 + 卡点)").font = note_font()
        # 5 要素
        apply_header(ws, 4, ["5 要素", "我的填写", "周次", "完成度", "卡点/调整"])
        for i, e in enumerate(elements):
            r = 5 + i
            ws.cell(row=r, column=1, value=e).font = data_font()
            ws.cell(row=r, column=1).fill = cream_fill()
            ws.cell(row=r, column=1).alignment = left_align()
            ws.cell(row=r, column=1).border = thin_border()
            ws.cell(row=r, column=2, value="").alignment = left_align()
            ws.cell(row=r, column=2).font = data_font()
            ws.cell(row=r, column=2).border = thin_border()
        # 5 周
        for w in range(1, 6):
            r = 5 + len(elements) + w - 1
            ws.cell(row=r, column=3, value=f"第 {w} 周").alignment = center()
            ws.cell(row=r, column=4, value="").alignment = center()
            ws.cell(row=r, column=5, value="").alignment = left_align()
            for col in range(3, 6):
                ws.cell(row=r, column=col).font = data_font()
                ws.cell(row=r, column=col).border = thin_border()
        # 完成度统计
        stat_r = 5 + len(elements) + 5 + 1
        ws.cell(row=stat_r, column=1, value="5 周完成度平均").font = title_font()
        ws.cell(row=stat_r, column=3, value="平均完成度")
        ws.cell(row=stat_r, column=4, value=f"=AVERAGE(D{5 + len(elements)}:D{5 + len(elements) + 4})").alignment = center()
        ws.cell(row=stat_r, column=4).number_format = "0.0%"
        for col in range(1, 6):
            ws.cell(row=stat_r, column=col).fill = cream_fill()
            ws.cell(row=stat_r, column=col).border = thin_border()
        set_col_widths(ws, [22, 35, 12, 12, 30])

    # 月复盘
    ws4 = wb.create_sheet("月复盘")
    apply_title(ws4, 1, "F19 月复盘 · 32 天整体回顾", span=6)
    apply_header(ws4, 3, ["周次", "行动1 完成度", "行动2 完成度", "行动3 完成度", "最大卡点", "下周调整"])
    for w in range(1, 6):
        r = 3 + w
        ws4.cell(row=r, column=1, value=f"第 {w} 周").alignment = center()
        ws4.cell(row=r, column=2, value="").alignment = center()
        ws4.cell(row=r, column=2).number_format = "0.0%"
        ws4.cell(row=r, column=3, value="").alignment = center()
        ws4.cell(row=r, column=3).number_format = "0.0%"
        ws4.cell(row=r, column=4, value="").alignment = center()
        ws4.cell(row=r, column=4).number_format = "0.0%"
        ws4.cell(row=r, column=5, value="").alignment = left_align()
        ws4.cell(row=r, column=6, value="").alignment = left_align()
        for col in range(1, 7):
            ws4.cell(row=r, column=col).font = data_font()
            ws4.cell(row=r, column=col).border = thin_border()
    last_r = 3 + 5
    ws4.cell(row=last_r + 1, column=1, value="32 天整体平均").font = title_font()
    ws4.cell(row=last_r + 1, column=2, value=f"=AVERAGE(B4:B{last_r})").alignment = center()
    ws4.cell(row=last_r + 1, column=2).number_format = "0.0%"
    ws4.cell(row=last_r + 1, column=3, value=f"=AVERAGE(C4:C{last_r})").alignment = center()
    ws4.cell(row=last_r + 1, column=3).number_format = "0.0%"
    ws4.cell(row=last_r + 1, column=4, value=f"=AVERAGE(D4:D{last_r})").alignment = center()
    ws4.cell(row=last_r + 1, column=4).number_format = "0.0%"
    for col in range(1, 7):
        ws4.cell(row=last_r + 1, column=col).fill = cream_fill()
        ws4.cell(row=last_r + 1, column=col).border = thin_border()
        ws4.cell(row=last_r + 1, column=col).font = data_font()
    set_col_widths(ws4, [12, 14, 14, 14, 35, 35])

    wb.save(os.path.join(OUT_DIR, "F19_行动计划.xlsx"))
    print("[OK] F19_行动计划.xlsx")


# ============================================================
# 5. 30天打卡总表.xlsx
# ============================================================
def build_30day():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "每日打卡"
    apply_title(ws1, 1, "30 天打卡总表 · 个人追踪", span=8)
    ws1.cell(row=2, column=1, value="5 项打卡 + 情绪状态 + 今日金句,30 天持续追踪").font = note_font()
    apply_header(ws1, 4, ["Day", "心谈(是/否+对象)", "GROW+(是/否+对象)", "3C(是/否+行为)", "团队分享(是/否+主题)", "反思(是/否+洞察)", "情绪(1-10)", "今日金句"])
    for d in range(1, 31):
        r = 4 + d
        ws1.cell(row=r, column=1, value=f"Day {d}").alignment = center()
        for col in range(2, 9):
            ws1.cell(row=r, column=col).alignment = left_align()
            ws1.cell(row=r, column=col).font = data_font()
            ws1.cell(row=r, column=col).border = thin_border()
        ws1.cell(row=r, column=7).alignment = center()
    # 统计
    last_r = 4 + 30
    ws1.cell(row=last_r + 1, column=1, value="30 天总完成度").font = title_font()
    ws1.cell(row=last_r + 1, column=2, value=f'=COUNTIF(B5:B{last_r},"是*")/30').alignment = center()
    ws1.cell(row=last_r + 1, column=2).number_format = "0.0%"
    ws1.cell(row=last_r + 1, column=3, value=f'=COUNTIF(C5:C{last_r},"是*")/30').alignment = center()
    ws1.cell(row=last_r + 1, column=3).number_format = "0.0%"
    ws1.cell(row=last_r + 1, column=4, value=f'=COUNTIF(D5:D{last_r},"是*")/30').alignment = center()
    ws1.cell(row=last_r + 1, column=4).number_format = "0.0%"
    ws1.cell(row=last_r + 1, column=5, value=f'=COUNTIF(E5:E{last_r},"是*")/30').alignment = center()
    ws1.cell(row=last_r + 1, column=5).number_format = "0.0%"
    ws1.cell(row=last_r + 1, column=6, value=f'=COUNTIF(F5:F{last_r},"是*")/30').alignment = center()
    ws1.cell(row=last_r + 1, column=6).number_format = "0.0%"
    ws1.cell(row=last_r + 1, column=7, value=f'=AVERAGE(G5:G{last_r})').alignment = center()
    ws1.cell(row=last_r + 1, column=7).number_format = "0.0"
    for col in range(1, 9):
        ws1.cell(row=last_r + 1, column=col).fill = cream_fill()
        ws1.cell(row=last_r + 1, column=col).border = thin_border()
        ws1.cell(row=last_r + 1, column=col).font = data_font()
    set_col_widths(ws1, [10, 18, 18, 18, 22, 22, 12, 30])

    # Sheet 2: 周复盘
    ws2 = wb.create_sheet("周复盘")
    apply_title(ws2, 1, "30 天 · 周复盘", span=7)
    apply_header(ws2, 3, ["周次", "心谈次数", "GROW+ 次数", "3C 次数", "团队分享", "平均情绪", "最大洞察"])
    for w in range(1, 5):
        r = 3 + w
        ws2.cell(row=r, column=1, value=f"第 {w} 周").alignment = center()
        # 周起始行
        start_d = 5 + (w - 1) * 7
        end_d = start_d + 6
        ws2.cell(row=r, column=2, value=f'=COUNTIF(每日打卡!B{start_d}:B{end_d},"是*")').alignment = center()
        ws2.cell(row=r, column=3, value=f'=COUNTIF(每日打卡!C{start_d}:C{end_d},"是*")').alignment = center()
        ws2.cell(row=r, column=4, value=f'=COUNTIF(每日打卡!D{start_d}:D{end_d},"是*")').alignment = center()
        ws2.cell(row=r, column=5, value=f'=COUNTIF(每日打卡!E{start_d}:E{end_d},"是*")').alignment = center()
        ws2.cell(row=r, column=6, value=f'=AVERAGE(每日打卡!G{start_d}:G{end_d})').alignment = center()
        ws2.cell(row=r, column=6).number_format = "0.0"
        ws2.cell(row=r, column=7, value="").alignment = left_align()
        for col in range(1, 8):
            ws2.cell(row=r, column=col).font = data_font()
            ws2.cell(row=r, column=col).border = thin_border()
    set_col_widths(ws2, [10, 12, 12, 12, 12, 12, 35])

    # Sheet 3: 总复盘
    ws3 = wb.create_sheet("总复盘")
    apply_title(ws3, 1, "30 天整体总复盘", span=2)
    apply_header(ws3, 3, ["指标", "数值"])
    items = [
        ("心谈总次数", '=COUNTIF(每日打卡!B5:B34,"是*")'),
        ("GROW+ 总次数", '=COUNTIF(每日打卡!C5:C34,"是*")'),
        ("3C 总次数", '=COUNTIF(每日打卡!D5:D34,"是*")'),
        ("团队分享总次数", '=COUNTIF(每日打卡!E5:E34,"是*")'),
        ("反思总次数", '=COUNTIF(每日打卡!F5:F34,"是*")'),
        ("平均情绪", "=AVERAGE(每日打卡!G5:G34)"),
        ("坚持天数(≥3项)", '=SUMPRODUCT(--((每日打卡!B5:B34<>"")+(每日打卡!C5:C34<>"")+(每日打卡!D5:D34<>"")+(每日打卡!E5:E34<>"")+(每日打卡!F5:F34<>"")>=3))'),
    ]
    for i, (k, v) in enumerate(items):
        r = 4 + i
        ws3.cell(row=r, column=1, value=k).font = data_font()
        ws3.cell(row=r, column=1).fill = cream_fill()
        ws3.cell(row=r, column=1).border = thin_border()
        ws3.cell(row=r, column=2, value=v).alignment = center()
        ws3.cell(row=r, column=2).font = data_font()
        ws3.cell(row=r, column=2).border = thin_border()
        if "情绪" in k:
            ws3.cell(row=r, column=2).number_format = "0.0"
    set_col_widths(ws3, [25, 15])

    wb.save(os.path.join(OUT_DIR, "30天打卡总表.xlsx"))
    print("[OK] 30天打卡总表.xlsx")


# ============================================================
# 6. 30天行为改变调研汇总.xlsx
# ============================================================
def build_survey():
    wb = Workbook()
    # Sheet 1: 原始数据
    ws1 = wb.active
    ws1.title = "原始数据"
    apply_title(ws1, 1, "30 天行为改变调研 · 全员数据汇总", span=8)
    ws1.cell(row=2, column=1, value="每个学员 1 行,统计行为频次 + NPS").font = note_font()
    apply_header(ws1, 4, ["学员", "GROW 次数", "心谈次数", "3C 次数", "团队分享", "反思次数", "NPS(0-10)", "类型"])
    for i in range(30):
        r = 5 + i
        ws1.cell(row=r, column=1, value=f"学员{i + 1}").alignment = center()
        for col in range(2, 8):
            ws1.cell(row=r, column=col).alignment = center()
        # 类型公式:NPS ≥ 9 推荐者;7-8 被动者;< 7 贬损者
        ws1.cell(row=r, column=8, value=f'=IF(G{r}>=9,"推荐者",IF(G{r}>=7,"被动者","贬损者"))').alignment = center()
        for col in range(1, 9):
            ws1.cell(row=r, column=col).font = data_font()
            ws1.cell(row=r, column=col).border = thin_border()
    set_col_widths(ws1, [12, 12, 12, 12, 12, 12, 12, 12])

    # Sheet 2: 行为汇总
    ws2 = wb.create_sheet("行为汇总")
    apply_title(ws2, 1, "30 天 · 行为频次汇总统计", span=4)
    apply_header(ws2, 3, ["行为", "平均", "中位数", "最大"])
    items = [("GROW 次数", "B"), ("心谈次数", "C"), ("3C 次数", "D"), ("团队分享", "E"), ("反思次数", "F")]
    for i, (k, col) in enumerate(items):
        r = 4 + i
        ws2.cell(row=r, column=1, value=k).font = data_font()
        ws2.cell(row=r, column=1).fill = cream_fill()
        ws2.cell(row=r, column=1).border = thin_border()
        ws2.cell(row=r, column=2, value=f"=AVERAGE(原始数据!{col}5:{col}34)").alignment = center()
        ws2.cell(row=r, column=2).number_format = "0.0"
        ws2.cell(row=r, column=3, value=f"=MEDIAN(原始数据!{col}5:{col}34)").alignment = center()
        ws2.cell(row=r, column=3).number_format = "0.0"
        ws2.cell(row=r, column=4, value=f"=MAX(原始数据!{col}5:{col}34)").alignment = center()
        for c in range(2, 5):
            ws2.cell(row=r, column=c).font = data_font()
            ws2.cell(row=r, column=c).border = thin_border()
    set_col_widths(ws2, [16, 12, 12, 12])

    # Sheet 3: NPS 分析
    ws3 = wb.create_sheet("NPS分析")
    apply_title(ws3, 1, "NPS 评分分析", span=4)
    apply_header(ws3, 3, ["类型", "人数", "占比", "建议动作"])
    actions = {
        "推荐者": "邀请加入校友会,内训分享,案例共创",
        "被动者": "保持联系,季度回访,提供进阶资源",
        "贬损者": "1V1 深度访谈,找改善点,二次邀请复训",
    }
    for i, t in enumerate(["推荐者", "被动者", "贬损者"]):
        r = 4 + i
        ws3.cell(row=r, column=1, value=t).alignment = center()
        ws3.cell(row=r, column=2, value=f'=COUNTIF(原始数据!H5:H34,"{t}")').alignment = center()
        ws3.cell(row=r, column=3, value=f'=B{r}/30').alignment = center()
        ws3.cell(row=r, column=3).number_format = "0.0%"
        ws3.cell(row=r, column=4, value=actions[t])
        for col in range(1, 5):
            ws3.cell(row=r, column=col).font = data_font()
            ws3.cell(row=r, column=col).border = thin_border()
    # NPS 公式:推荐者% - 贬损者%
    nps_r = 4 + 3 + 1
    ws3.cell(row=nps_r, column=1, value="NPS 分数").font = title_font()
    ws3.cell(row=nps_r, column=2, value=f"=(B4-B6)/30*100").alignment = center()
    ws3.cell(row=nps_r, column=2).number_format = "0.0"
    ws3.cell(row=nps_r, column=3, value="推荐者% - 贬损者%").font = note_font()
    for col in range(1, 5):
        ws3.cell(row=nps_r, column=col).fill = cream_fill()
        ws3.cell(row=nps_r, column=col).border = thin_border()
    set_col_widths(ws3, [12, 10, 10, 50])

    # Sheet 4: 跟进表
    ws4 = wb.create_sheet("跟进表")
    apply_title(ws4, 1, "30 天后 · 分层跟进", span=4)
    apply_header(ws4, 3, ["学员", "NPS", "类型", "跟进动作"])
    for i in range(30):
        r = 4 + i
        # 跨表引用
        ws4.cell(row=r, column=1, value=f"=原始数据!A{5 + i}").alignment = center()
        ws4.cell(row=r, column=2, value=f"=原始数据!G{5 + i}").alignment = center()
        ws4.cell(row=r, column=3, value=f"=原始数据!H{5 + i}").alignment = center()
        ws4.cell(row=r, column=4, value=f'=IF(C{r}="推荐者","邀请校友会 + 内训分享",IF(C{r}="被动者","季度回访 + 进阶资源","1V1 访谈 + 复训邀请"))').alignment = left_align()
        for col in range(1, 5):
            ws4.cell(row=r, column=col).font = data_font()
            ws4.cell(row=r, column=col).border = thin_border()
    set_col_widths(ws4, [12, 8, 12, 50])

    wb.save(os.path.join(OUT_DIR, "30天行为改变调研汇总.xlsx"))
    print("[OK] 30天行为改变调研汇总.xlsx")


# ============================================================
# 7. 团队氛围诊断汇总.xlsx
# ============================================================
def build_team_diag():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "个人评分"
    apply_title(ws1, 1, "团队氛围诊断 · 7 维度 × 12 月追踪", span=13)
    ws1.cell(row=2, column=1, value="7 维度评分(1-10),每月 1 次,持续 12 个月").font = note_font()
    dims = ["心理安全", "尊重多样性", "开放沟通", "协作意愿", "创新氛围", "学习成长", "目标对齐"]
    headers = ["学员"] + [f"第 {m + 1} 月" for m in range(12)]
    apply_header(ws1, 4, headers)
    for i in range(15):
        r = 5 + i
        ws1.cell(row=r, column=1, value=f"学员{i + 1}").alignment = center()
        for m in range(12):
            ws1.cell(row=r, column=2 + m).alignment = center()
        for col in range(1, 14):
            ws1.cell(row=r, column=col).font = data_font()
            ws1.cell(row=r, column=col).border = thin_border()
    set_col_widths(ws1, [12] + [10] * 12)

    # Sheet 2: 团队平均(7维度)
    ws2 = wb.create_sheet("团队平均")
    apply_title(ws2, 1, "团队氛围 · 7 维度月度平均", span=13)
    apply_header(ws2, 3, ["维度"] + [f"第 {m + 1} 月" for m in range(12)])
    # 学员行在个人评分 Sheet 中是 5-19(15 人)
    for i, d in enumerate(dims):
        r = 4 + i
        ws2.cell(row=r, column=1, value=d).font = data_font()
        ws2.cell(row=r, column=1).fill = cream_fill()
        ws2.cell(row=r, column=1).border = thin_border()
        for m in range(12):
            col_letter = get_column_letter(2 + m)
            ws2.cell(row=r, column=2 + m, value=f"=AVERAGE(个人评分!{col_letter}5:{col_letter}19)").alignment = center()
            ws2.cell(row=r, column=2 + m).number_format = "0.0"
            ws2.cell(row=r, column=2 + m).font = data_font()
            ws2.cell(row=r, column=2 + m).border = thin_border()
    set_col_widths(ws2, [14] + [10] * 12)

    # Sheet 3: 行动计划
    ws3 = wb.create_sheet("行动计划")
    apply_title(ws3, 1, "团队氛围诊断 · 30 天行动计划", span=5)
    apply_header(ws3, 3, ["维度", "当前均分", "30 天目标", "具体动作", "衡量标准"])
    targets = {
        "心理安全": ("8.0", "每周 1V1 心谈 8 分钟,不评判只倾听", "下月评分 ≥ 7.0"),
        "尊重多样性": ("8.0", "每团队会议 1 个'不同视角'环节", "不同意见被引述 ≥ 3 次/周"),
        "开放沟通": ("8.0", "管理者 8 分钟心谈 ≥ 2 次/周", "下属主动汇报频次 +20%"),
        "协作意愿": ("8.0", "跨组项目每周同步会", "协作项目完成率 ≥ 90%"),
        "创新氛围": ("8.0", "失败复盘会,允许试错", "新提案数量 +30%"),
        "学习成长": ("8.0", "30 天打卡 5 维度", "完成度 ≥ 80%"),
        "目标对齐": ("8.0", "OKR 季度对齐会", "目标共识度评分 ≥ 8.0"),
    }
    for i, d in enumerate(dims):
        r = 4 + i
        ws3.cell(row=r, column=1, value=d).alignment = center()
        ws3.cell(row=r, column=2, value=f"=AVERAGE(团队平均!{get_column_letter(2)}4:{get_column_letter(2 + 11)}4)").alignment = center()
        ws3.cell(row=r, column=2).number_format = "0.0"
        ws3.cell(row=r, column=3, value=targets[d][0]).alignment = center()
        ws3.cell(row=r, column=4, value=targets[d][1])
        ws3.cell(row=r, column=5, value=targets[d][2])
        for col in range(1, 6):
            ws3.cell(row=r, column=col).font = data_font()
            ws3.cell(row=r, column=col).border = thin_border()
    set_col_widths(ws3, [14, 10, 10, 35, 30])

    wb.save(os.path.join(OUT_DIR, "团队氛围诊断汇总.xlsx"))
    print("[OK] 团队氛围诊断汇总.xlsx")


# ============================================================
# 执行
# ============================================================
if __name__ == "__main__":
    build_F1()
    build_F2()
    build_F3()
    build_F19()
    build_30day()
    build_survey()
    build_team_diag()
    print("\n全部 7 个 Excel 文件生成完成!")

# -*- coding: utf-8 -*-
"""
学员交付物清单.xlsx
字段：序号 / 交付物 / 必交/选交 / 格式要求 / 字数要求 / 截止日期 / 自评完成度 / 评委反馈
必交4件 + 选交4类，含"评判标准自检"列
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "交付物清单"

# 颜色
HEADER_BG = "1F3A5F"  # 深蓝灰
SUBHEADER_BG = "2E5C8A"  # 中蓝
MANDATORY_BG = "C8102E"  # 红色 - 必交
OPTIONAL_BG = "00733B"  # 绿色 - 选交
ALT_BG = "F0F4F8"  # 浅色交替
WHITE = "FFFFFF"
ACCENT = "C8102E"

# 字体
header_font = Font(name='Microsoft YaHei', size=11, bold=True, color=WHITE)
subheader_font = Font(name='Microsoft YaHei', size=10, bold=True, color=WHITE)
body_font = Font(name='Microsoft YaHei', size=10, color="1A1A1A")
body_bold = Font(name='Microsoft YaHei', size=10, bold=True, color="1A1A1A")
small_font = Font(name='Microsoft YaHei', size=9, color="555555")

# 填充
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
subheader_fill = PatternFill(start_color=SUBHEADER_BG, end_color=SUBHEADER_BG, fill_type='solid')
mandatory_fill = PatternFill(start_color=MANDATORY_BG, end_color=MANDATORY_BG, fill_type='solid')
optional_fill = PatternFill(start_color=OPTIONAL_BG, end_color=OPTIONAL_BG, fill_type='solid')
alt_fill = PatternFill(start_color=ALT_BG, end_color=ALT_BG, fill_type='solid')

# 边框
thin = Side(border_style="thin", color="D6DEE8")
medium = Side(border_style="medium", color=HEADER_BG)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_border = Border(left=thin, right=thin, top=medium, bottom=medium)

# 对齐
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

# 列宽
col_widths = {
    'A': 6,   # 序号
    'B': 28,  # 交付物
    'C': 10,  # 必交/选交
    'D': 28,  # 格式要求
    'E': 18,  # 字数/数量要求
    'F': 14,  # 截止日期
    'G': 16,  # 自评完成度
    'H': 36,  # 评判标准自检
    'I': 24,  # 评委反馈
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# 标题
ws.merge_cells('A1:I1')
ws['A1'] = '顺造科技 · AI项目成果评审 · 学员交付物清单'
ws['A1'].font = Font(name='Microsoft YaHei', size=18, bold=True, color=HEADER_BG)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

# 副标题
ws.merge_cells('A2:I2')
ws['A2'] = '必交 4 件 + 选交 4 类 · 截止日期：2026年6月20日 · 评审日：2026年6月25日'
ws['A2'].font = Font(name='Microsoft YaHei', size=11, color="555555")
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 24

# 学员工信息
ws.merge_cells('A3:I3')
ws['A3'] = '学员姓名：______________  部门：______________  入职时长：__________  AI经验：☐零基础 ☐了解 ☐熟练'
ws['A3'].font = Font(name='Microsoft YaHei', size=10, color="1A1A1A")
ws['A3'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws['A3'].fill = PatternFill(start_color="FFFEF5", end_color="FFFEF5", fill_type='solid')
ws.row_dimensions[3].height = 28

# 表头
headers = ['序号', '交付物', '必交/选交', '格式要求', '字数/数量要求', '截止日期', '自评完成度', '评判标准自检', '评委反馈']
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = header_border
ws.row_dimensions[4].height = 32

# 数据行
mandatory_items = [
    {
        'name': '业务流程手册',
        'format': 'Word文档（.docx）',
        'size': '1-2万字，8章',
        'deadline': '2026-06-20',
        'self_check': '□ 第一章有"业务现状+AI介入机会"\n□ 至少3个核心业务流程\n□ 流程图清晰（当前 vs AI后）\n□ AI介入级别标注（一级-五级）\n□ 案例真实可追溯\n□ 末尾有"未来迭代计划"\n□ 总字数1万字以上',
    },
    {
        'name': 'HTML可视化网页',
        'format': 'HTML文件（.html）',
        'size': 'A3海报风格 或 长文档风格（二选一）',
        'deadline': '2026-06-22',
        'self_check': '□ 浅色背景+深色头部\n□ 关键数字大字号突出\n□ 流程图清晰可读\n□ 可滚动 + A4可打印\n□ 至少3个对比数字\n□ 包含60秒故事展示',
    },
    {
        'name': '10分钟路演逐字稿',
        'format': 'Word 或 MD',
        'size': '10分钟（约1800-2200字）',
        'deadline': '2026-06-22',
        'self_check': '□ 痛(15%)/做(20%)/效(45%)/求(20%)比例对\n□ 第一句话不以"我"开头\n□ 含60秒故事（背景/转折/行动/结果）\n□ 三层成果数字（基线/新状态/意义）\n□ 请求具体到领导知道做什么\n□ 配有一页纸框架（备份提示卡）',
    },
    {
        'name': 'Q&A应答清单',
        'format': 'Word 或 MD',
        'size': '5-10个高频问题',
        'deadline': '2026-06-22',
        'self_check': '□ 含"追问数字来源"应对\n□ 含"问推广可行性"应对\n□ 含"问数据安全"应对\n□ 含"问以前没做"应对\n□ 含"问投入产出比"应对\n□ 含"完全没想到的问题"应对\n□ 每个问题有"问题本质+错误示范+正确示范"',
    },
]

optional_items = [
    {
        'name': '结构化提示词',
        'format': '.md 或 文档',
        'size': '≥3个提示词',
        'deadline': '2026-06-20',
        'self_check': '□ 提示词含"角色/任务/输入/输出/约束"\n□ 至少有1个是"多步结构化"（非单点）\n□ 每个提示词有"使用场景"说明\n□ 有版本迭代记录（v1/v2等）\n□ 配套说明文档',
    },
    {
        'name': '智能体（Agent）',
        'format': '配置文件 + 使用说明',
        'size': '≥1个智能体',
        'deadline': '2026-06-20',
        'self_check': '□ 含智能体名称/功能描述\n□ 配置文件完整可运行\n□ 配套使用说明（如何调用/输入什么）\n□ 至少跑通1个真实业务场景\n□ 有效果截图或录屏',
    },
    {
        'name': 'Skill（可复用能力包）',
        'format': 'SKILL.md + 配套文件',
        'size': '≥1个Skill',
        'deadline': '2026-06-20',
        'self_check': '□ 含SKILL.md（标准格式）\n□ 触发词清晰\n□ 输入/输出/约束说明完整\n□ 至少跑通1次真实使用\n□ 其他学员可独立使用',
    },
    {
        'name': '知识库',
        'format': '结构化条目 + 索引',
        'size': '≥1个知识库（≥20条）',
        'deadline': '2026-06-20',
        'self_check': '□ 知识条目结构化（标题/分类/正文/标签）\n□ 至少20条\n□ 有索引/分类\n□ 来源可追溯\n□ 配套使用说明',
    },
]

current_row = 5
mandatory_count = 0
optional_count = 0

# 必交4件
for idx, item in enumerate(mandatory_items, 1):
    mandatory_count += 1
    row_data = [
        idx,
        item['name'],
        '必交',
        item['format'],
        item['size'],
        item['deadline'],
        '☐ 已完成',
        item['self_check'],
        '',  # 评委反馈
    ]
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = body_font
        cell.border = border
        if col_idx == 1:
            cell.alignment = center
        elif col_idx == 2:
            cell.alignment = left
            cell.font = body_bold
        elif col_idx == 3:
            cell.alignment = center
            cell.fill = mandatory_fill
            cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color=WHITE)
        elif col_idx == 8:
            cell.alignment = left_top
            cell.font = small_font
        else:
            cell.alignment = left
    # 交替背景
    if idx % 2 == 0:
        for c in range(1, 10):
            if c != 3:  # 不覆盖必交标签
                ws.cell(row=current_row, column=c).fill = alt_fill
    ws.row_dimensions[current_row].height = 100
    current_row += 1

# 选交4类
for idx, item in enumerate(optional_items, 1):
    optional_count += 1
    row_data = [
        4 + idx,  # 序号接续
        item['name'],
        '选交',
        item['format'],
        item['size'],
        item['deadline'],
        '☐ 已完成',
        item['self_check'],
        '',  # 评委反馈
    ]
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = body_font
        cell.border = border
        if col_idx == 1:
            cell.alignment = center
        elif col_idx == 2:
            cell.alignment = left
            cell.font = body_bold
        elif col_idx == 3:
            cell.alignment = center
            cell.fill = optional_fill
            cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color=WHITE)
        elif col_idx == 8:
            cell.alignment = left_top
            cell.font = small_font
        else:
            cell.alignment = left
    # 交替背景
    if idx % 2 == 0:
        for c in range(1, 10):
            if c != 3:
                ws.cell(row=current_row, column=c).fill = alt_fill
    ws.row_dimensions[current_row].height = 100
    current_row += 1

# 汇总行
summary_row = current_row
ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=9)
summary_cell = ws.cell(row=summary_row, column=1)
summary_cell.value = f'■ 必交 4 件（红色标记，缺一不可）+ 选交 4 类（绿色标记，多多益善）   |   总计 {mandatory_count + optional_count} 项   |   路演日：2026-06-25'
summary_cell.font = Font(name='Microsoft YaHei', size=11, bold=True, color=HEADER_BG)
summary_cell.alignment = Alignment(horizontal='center', vertical='center')
summary_cell.fill = PatternFill(start_color="EEF1F5", end_color="EEF1F5", fill_type='solid')
summary_cell.border = Border(top=medium, bottom=medium)
ws.row_dimensions[summary_row].height = 28

# 评审标准区
current_row += 2
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
title_cell = ws.cell(row=current_row, column=1, value='■ 评审打分标准（评委参考）')
title_cell.font = Font(name='Microsoft YaHei', size=14, bold=True, color=HEADER_BG)
title_cell.alignment = Alignment(horizontal='left', vertical='center')
title_cell.fill = PatternFill(start_color="FFFEF5", end_color="FFFEF5", fill_type='solid')
ws.row_dimensions[current_row].height = 28
current_row += 1

criteria_data = [
    ['维度', '权重', '评分要点', '5分(优秀)', '4分(良好)', '3分(合格)', '2分(待改进)', '1分(不足)'],
    ['问题真实性', '20%', '痛点是否真实存在', '具体数字+场景描述', '具体数字', '场景描述', '概括描述', '不清晰'],
    ['AI介入深度', '20%', 'AI介入级别+流程改造', '五级(完整Skill)', '四级(API)', '三级(智能体)', '二级(多步提示词)', '一级(单点提示词)'],
    ['效果可量化', '25%', '前后对比+对团队意义', '三层成果+故事', '三层成果', '两层成果', '一层成果', '无量化'],
    ['路演表达', '20%', '10分钟结构+Q&A', '流畅+完美Q&A', '流畅+良好Q&A', '基本流畅', '部分流畅', '明显卡顿'],
    ['请求合理性', '15%', '具体+小+有理由', '三要素全齐', '具体+小', '具体', '模糊', '无请求'],
]

for i, row in enumerate(criteria_data):
    is_header = (i == 0)
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal='left' if col_idx > 1 else 'center', vertical='center', wrap_text=True)
        if is_header:
            cell.font = subheader_font
            cell.fill = subheader_fill
        else:
            cell.font = body_font if col_idx == 1 else small_font
            if col_idx == 2:
                cell.alignment = center
                cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color=ACCENT)
            if i % 2 == 0:
                cell.fill = alt_fill
    ws.row_dimensions[current_row].height = 32
    current_row += 1

# 评审总分行
current_row += 1
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
score_cell = ws.cell(row=current_row, column=1, value='■ 评审总分：_______ / 100   |   路演得分：_______ / 60   |   交付物得分：_______ / 40   |   综合等级：☐A ☐B ☐C ☐D')
score_cell.font = Font(name='Microsoft YaHei', size=12, bold=True, color=HEADER_BG)
score_cell.alignment = Alignment(horizontal='center', vertical='center')
score_cell.fill = PatternFill(start_color="EEF1F5", end_color="EEF1F5", fill_type='solid')
score_cell.border = Border(top=medium, bottom=medium)
ws.row_dimensions[current_row].height = 32

# 评审委员签字区
current_row += 2
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
sign_cell = ws.cell(row=current_row, column=1, value='■ 评审委员签字：_______________________________   |   日期：__________')
sign_cell.font = Font(name='Microsoft YaHei', size=11, color="1A1A1A")
sign_cell.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[current_row].height = 28

# 冻结首行
ws.freeze_panes = 'A5'

# 保存
out = r'D:\2026年课程\顺造科技\AI\评审\02-学员指南\学员交付物清单.xlsx'
wb.save(out)
print(f'OK: {out}')
print(f'Sheets: {wb.sheetnames}')
print(f'Rows used: {current_row}')

# 验证
wb2 = openpyxl.load_workbook(out)
ws2 = wb2.active
print(f'After save - sheet: {ws2.title}, max_row: {ws2.max_row}, max_col: {ws2.max_column}')

# -*- coding: utf-8 -*-
"""
创建"人才盘点与梯队建设"课程配套表单
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = "D:/新课开发/管理学/11-人才盘点与梯队建设/配套表单和指引-Excel版"

# Style definitions
THIN_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)
MEDIUM_BORDER = Border(
    left=Side(style='medium', color='1F4E79'),
    right=Side(style='medium', color='1F4E79'),
    top=Side(style='medium', color='1F4E79'),
    bottom=Side(style='medium', color='1F4E79')
)

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='D9E8F5', end_color='D9E8F5', fill_type='solid')
LIGHT_GRAY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

HEADER_FONT = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
SUBHEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
LABEL_FONT = Font(name='微软雅黑', size=10, bold=True)
INPUT_FONT = Font(name='微软雅黑', size=10)
NOTE_FONT = Font(name='微软雅黑', size=9, italic=True, color='666666')

def CENTER_ALIGN():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def LEFT_ALIGN():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def RIGHT_ALIGN():
    return Alignment(horizontal='right', vertical='center', wrap_text=True)

def create_header_row(ws, row, values, fill=HEADER_FILL, font=HEADER_FONT, height=30):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER_ALIGN()
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = height

def merge_and_style(ws, start_row, start_col, end_row, end_col, value,
                   fill=None, font=None, alignment=None):
    ws.merge_cells(start_row=start_row, start_column=start_col,
                  end_row=end_row, end_column=end_col)
    cell = ws.cell(row=start_row, column=start_col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = CENTER_ALIGN()
    cell.border = THIN_BORDER
    return cell

def create_data_row(ws, row, values, fills=None, fonts=None, heights=None):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        if fills and col <= len(fills) and fills[col-1]:
            cell.fill = fills[col-1]
        if fonts and col <= len(fonts) and fonts[col-1]:
            cell.font = fonts[col-1]
        else:
            cell.font = INPUT_FONT
        cell.alignment = LEFT_ALIGN() if col > 1 else CENTER_ALIGN()
        cell.border = THIN_BORDER

def set_column_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_input_cell(ws, row, col, value=''):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = INPUT_FONT
    cell.fill = WHITE_FILL
    cell.alignment = LEFT_ALIGN()
    cell.border = THIN_BORDER
    return cell

# ==================== FORM 1: Nine-Grid Assessment ====================
def create_form1(ws, is_example=False):
    ws.title = "九宫格人才评估表"

    merge_and_style(ws, 1, 1, 1, 10, '人才盘点九宫格评估表', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 2, 1, 2, 10,
                    '说明：业绩评估（1-5分）+ 潜力评估（1-5分）→ 自动落位九宫格',
                    LIGHT_BLUE_FILL, NOTE_FONT, LEFT_ALIGN())
    ws.row_dimensions[2].height = 25

    # Header row
    headers = ['', '姓名', '部门', '岗位', '入职日期', '评估人', '评估周期', '', '', '']
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER_ALIGN()
        cell.border = THIN_BORDER
    ws.merge_cells('A3:A4')
    ws.merge_cells('H3:J3')
    ws['A3'].value = '基本信息'
    ws['A3'].alignment = CENTER_ALIGN()

    # Dimension headers
    merge_and_style(ws, 5, 1, 5, 2, '评估维度', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    merge_and_style(ws, 5, 3, 5, 5, '业绩评估维度', LIGHT_BLUE_FILL, LABEL_FONT, CENTER_ALIGN())
    merge_and_style(ws, 5, 6, 5, 8, '潜力评估维度', LIGHT_BLUE_FILL, LABEL_FONT, CENTER_ALIGN())
    merge_and_style(ws, 5, 9, 5, 10, '评估结果', LIGHT_BLUE_FILL, LABEL_FONT, CENTER_ALIGN())

    set_column_widths(ws, {'A': 12, 'B': 15, 'C': 18, 'D': 10, 'E': 10,
                          'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 12})

    perf_items = ['业绩目标达成率', '工作质量与准确性', '工作效率', '创新能力', '团队协作与贡献', '综合业绩得分']
    pot_items = ['学习能力', '适应变化能力', '领导力潜力', '战略思维能力', '沟通影响能力', '综合潜力得分']

    for i in range(6):
        row = 6 + i
        ws[f'A{row}'] = perf_items[i]
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'A{row}'].fill = LIGHT_GRAY_FILL
        ws[f'A{row}'].alignment = LEFT_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws[f'C{row}'] = pot_items[i]
        ws[f'C{row}'].font = LABEL_FONT
        ws[f'C{row}'].fill = LIGHT_GRAY_FILL
        ws[f'C{row}'].alignment = LEFT_ALIGN()
        ws[f'C{row}'].border = THIN_BORDER

        ws.row_dimensions[row].height = 22

        if i < 5:
            for col in ['D', 'E', 'F', 'G']:
                create_input_cell(ws, row, ord(col) - ord('A') + 1)
                ws.cell(row=row, column=ord(col)-ord('A')+1).alignment = CENTER_ALIGN()
                if is_example and i == 0:
                    ws.cell(row=row, column=ord(col)-ord('A')+1).value = 4
        else:
            ws[f'D{row}'] = '=AVERAGE(D6:D10)'
            ws[f'E{row}'] = '=D11'
            ws[f'F{row}'] = '=AVERAGE(F6:F10)'
            ws[f'G{row}'] = '=F11'
            for col in ['D', 'E', 'F', 'G']:
                c = ws[f'{col}{row}']
                c.font = Font(name='微软雅黑', size=10, bold=True)
                c.fill = YELLOW_FILL
                c.alignment = CENTER_ALIGN()
                c.border = MEDIUM_BORDER

    # Nine-grid result
    merge_and_style(ws, 5, 9, 5, 10, '九宫格落位', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    grid_info = [
        (6, '待发展\n(低业绩低潜力)', ORANGE_FILL),
        (7, '业务骨干\n(高业绩低潜力)', GREEN_FILL),
        (8, '', None),
        (9, '核心人才\n(高业绩高潜力)', GREEN_FILL),
        (10, '待发展\n(低业绩低潜力)', ORANGE_FILL),
        (11, '高潜力\n(低业绩高潜力)', LIGHT_BLUE_FILL),
    ]
    for row, text, fill in grid_info:
        ws[f'I{row}'] = text
        ws[f'I{row}'].alignment = CENTER_ALIGN()
        ws[f'I{row}'].border = THIN_BORDER
        if fill:
            ws[f'I{row}'].fill = fill
        ws.row_dimensions[row].height = 30

    # Auto-placement formula
    formula = '=IF(AND(D11>=3.5,F11>=3.5),"核心人才",IF(AND(D11>=3.5,F11<3.5),"业务骨干",IF(AND(D11<3.5,F11>=3.5),"待发展","绩效改进")))'
    for row in range(6, 12):
        ws[f'J{row}'] = formula
        ws[f'J{row}'].alignment = CENTER_ALIGN()
        ws[f'J{row}'].border = THIN_BORDER
        ws[f'J{row}'].font = Font(name='微软雅黑', size=10, bold=True)

    # Notes
    merge_and_style(ws, 13, 1, 13, 10, '评估备注', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A14:J16')
    ws['A14'] = '备注：\n1. 评分标准：1=需改进，2=基本达期望，3=达到期望，4=超越期望，5=杰出\n2. 综合得分=AVERAGE(各维度得分)\n3. 九宫格判定：业绩>=3.5为高，潜力>=3.5为高'
    ws['A14'].alignment = LEFT_ALIGN()
    ws['A14'].font = NOTE_FONT
    ws['A14'].border = THIN_BORDER
    ws.row_dimensions[14].height = 60

# ==================== FORM 2: Meeting Minutes ====================
def create_form2(ws, is_example=False):
    ws.title = "人才盘点会议记录"

    merge_and_style(ws, 1, 1, 1, 8, '人才盘点会议记录表', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
    ws.row_dimensions[1].height = 35

    merge_and_style(ws, 2, 1, 2, 8, '会议信息', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = CENTER_ALIGN()

    info_labels = ['会议日期', '会议时间', '会议地点', '主持人', '记录人', '参会人数', '总人数', '']
    create_data_row(ws, 3, info_labels, fills=[LIGHT_GRAY_FILL]*8, fonts=[LABEL_FONT]*8)

    for col in range(1, 9):
        create_input_cell(ws, 4, col)
    if is_example:
        ws['A4'].value = '2024-03-15'
        ws['B4'].value = '14:00-16:00'
        ws['C4'].value = '总部会议室A'
        ws['D4'].value = '张总'
        ws['E4'].value = '李明'
        ws['F4'].value = 15
        ws['G4'].value = 18
    ws['H4'].value = '第  次'
    ws['H4'].font = INPUT_FONT
    ws['H4'].alignment = CENTER_ALIGN()
    ws['H4'].border = THIN_BORDER

    # Participants
    merge_and_style(ws, 5, 1, 5, 8, '参会人员名单', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A5:H5')
    ws['A5'].alignment = CENTER_ALIGN()

    for row in range(6, 12):
        ws.merge_cells(f'A{row}:H{row}')
        create_input_cell(ws, row, 1)
        ws[f'A{row}'].alignment = LEFT_ALIGN()
        ws.row_dimensions[row].height = 22

    if is_example:
        participants = ['王芳（营销部）-核心人才', '李强（技术部）-业务骨干', '刘洋（人力资源）-高潜力',
                      '陈静（财务部）-业务骨干', '赵磊（运营部）-核心人才', '']
        for i, p in enumerate(participants):
            ws[f'A{6+i}'].value = p
            ws[f'A{6+i}'].font = INPUT_FONT if p else Font(name='微软雅黑', size=10, color='BFBFBF')

    # Agenda
    merge_and_style(ws, 12, 1, 12, 8, '本次盘点核心议题', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A12:H12')
    ws['A12'].alignment = CENTER_ALIGN()

    ws.merge_cells('A13:H15')
    create_input_cell(ws, 13, 1)
    ws['A13'].alignment = LEFT_ALIGN()
    ws.row_dimensions[13].height = 50

    if is_example:
        ws['A13'].value = '1. 九宫格人员分布Review\n2. 关键岗位继任计划确认\n3. 高潜人才培养进度汇报\n4. 下季度人才调整建议'
        ws['A13'].font = INPUT_FONT

    # Nine-grid confirmation
    merge_and_style(ws, 16, 1, 16, 8, '九宫格确认', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A16:H16')
    ws['A16'].alignment = CENTER_ALIGN()

    grid_header = ['区域', '人数', '占比', '姓名列表', '', '', '', '']
    create_data_row(ws, 17, grid_header, fills=[LIGHT_GRAY_FILL]*8, fonts=[LABEL_FONT]*8)
    ws.merge_cells('D17:H17')

    grid_data = [
        '核心人才（高业绩高潜力）',
        '业务骨干（高业绩低潜力）',
        '待发展（低业绩高潜力）',
        '绩效改进（低业绩低潜力）',
    ]
    for i, label in enumerate(grid_data):
        row = 18 + i
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'A{row}'].fill = LIGHT_GRAY_FILL
        ws[f'A{row}'].alignment = LEFT_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws[f'B{row}'] = 0
        ws[f'B{row}'].alignment = CENTER_ALIGN()
        ws[f'B{row}'].border = THIN_BORDER

        ws[f'C{row}'] = '=B' + str(row) + '/SUM(B18:B21)'
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'C{row}'].alignment = CENTER_ALIGN()
        ws[f'C{row}'].border = THIN_BORDER

        ws.merge_cells(f'D{row}:H{row}')
        ws[f'D{row}'].border = THIN_BORDER
        ws.row_dimensions[row].height = 22

    # Action items
    merge_and_style(ws, 22, 1, 22, 8, '行动项', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A22:H22')
    ws['A22'].alignment = CENTER_ALIGN()

    action_headers = ['序号', '行动项描述', '责任人', '协助人', '开始日期', '完成日期', '优先级', '状态']
    create_data_row(ws, 28, action_headers, fills=[LIGHT_GRAY_FILL]*8, fonts=[LABEL_FONT]*8)

    for row in range(29, 34):
        ws[f'A{row}'] = row - 28
        ws[f'A{row}'].alignment = CENTER_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER
        for col in range(2, 9):
            create_input_cell(ws, row, col)
        ws.row_dimensions[row].height = 22

    if is_example:
        actions = [
            ['启动技术管理培训项目', '刘总监', 'HRBP小王', '2024-04-01', '2024-06-30', '高', '进行中'],
            ['完成3名高潜人才IDP制定', '李经理', 'HR小李', '2024-03-20', '2024-04-15', '高', '已完成'],
            ['绩效改进计划执行跟进', '张主管', 'HR小张', '2024-03-18', '2024-04-18', '中', '进行中'],
            ['轮岗计划人员匹配', '刘总监', '各部门负责人', '2024-04-01', '2024-04-30', '中', '待启动'],
            ['', '', '', '', '', '', ''],
        ]
        for i, action in enumerate(actions):
            for j, val in enumerate(action):
                ws.cell(row=29+i, column=j+1, value=val)

    # Next meeting
    merge_and_style(ws, 34, 1, 34, 8, '下次会议安排', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A34:H34')
    ws['A34'].alignment = CENTER_ALIGN()

    meeting_fields = ['下次会议日期', '下次会议时间', '下次会议地点', '主要议题', '', '', '', '']
    create_data_row(ws, 35, meeting_fields, fills=[LIGHT_GRAY_FILL]*8, fonts=[LABEL_FONT]*8)

    for col in range(1, 9):
        create_input_cell(ws, 36, col)
    ws.merge_cells('D36:H36')

    if is_example:
        ws['A36'].value = '2024-06-15'
        ws['B36'].value = '14:00-16:00'
        ws['C36'].value = '总部会议室A'
        ws['D36'].value = 'Q2人才盘点Review + 下半年规划'

    set_column_widths(ws, {'A': 15, 'B': 15, 'C': 15, 'D': 12, 'E': 12, 'F': 12, 'G': 10, 'H': 12})

# ==================== FORM 3: Succession Map ====================
def create_form3(ws, is_example=False):
    ws.title = "关键岗位继任地图"

    merge_and_style(ws, 1, 1, 1, 10, '关键岗位继任地图', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
    ws.row_dimensions[1].height = 35

    merge_and_style(ws, 2, 1, 2, 10,
                    '说明：追踪关键岗位的继任准备情况 | 就绪度：1=空缺风险高，2=准备中，3=基本就绪，4=完全就绪',
                    LIGHT_BLUE_FILL, NOTE_FONT, LEFT_ALIGN())
    ws.row_dimensions[2].height = 25

    headers = ['序号', '岗位名称', '现任者', '部门', '任职年限', '继任者\n（第一优先）', '继任者\n部门', '继任准备度', '继任时间', '发展需求']
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER_ALIGN()
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 35

    for row in range(4, 14):
        ws[f'A{row}'] = row - 3
        ws[f'A{row}'].alignment = CENTER_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER
        for col in range(2, 11):
            create_input_cell(ws, row, col)
        ws.row_dimensions[row].height = 25

    if is_example:
        example_data = [
            [1, '营销总监', '张华', '营销部', 5, '王涛', '营销部', 3, '2024-12-31', '战略规划能力提升'],
            [2, '技术总监', '李明', '技术部', 6, '待定', '', 1, '2025-06-30', '需外部招聘或培养'],
            [3, '运营总监', '刘洋', '运营部', 4, '陈静', '运营部', 4, '2024-12-31', '已就绪'],
            [4, '财务总监', '赵敏', '财务部', 7, '孙伟', '财务部', 3, '2025-12-31', '管理会计培训'],
            [5, '人力资源总监', '周莉', '人力资源部', 3, '吴昊', '人力资源部', 2, '2025-06-30', '组织发展能力'],
            [6, '产品总监', '郑强', '产品部', 4, '待定', '', 1, '2025-12-31', '需内部培养或外部招聘'],
            [7, '客服总监', '冯雪', '客服部', 5, '高阳', '客服部', 4, '2024-12-31', '已就绪'],
            [8, '供应链总监', '曹峰', '供应链部', 6, '邓林', '供应链部', 3, '2025-06-30', '战略采购培训'],
            [9, '', '', '', '', '', '', '', '', ''],
            [10, '', '', '', '', '', '', '', '', ''],
        ]
        for i, row_data in enumerate(example_data):
            for j, val in enumerate(row_data):
                ws.cell(row=4+i, column=j+1, value=val)
                if j > 0 and val:
                    ws.cell(row=4+i, column=j+1).font = INPUT_FONT
                    ws.cell(row=4+i, column=j+1).alignment = LEFT_ALIGN() if j in [1, 2, 3, 5, 6, 9] else CENTER_ALIGN()
                    ws.cell(row=4+i, column=j+1).border = THIN_BORDER

    # Summary
    ws['A14'] = '合计'
    ws['A14'].font = LABEL_FONT
    ws['A14'].fill = LIGHT_GRAY_FILL
    ws['A14'].alignment = CENTER_ALIGN()
    ws['A14'].border = THIN_BORDER

    ws['B14'] = '=COUNTA(B4:B13)'
    ws['B14'].font = LABEL_FONT
    ws['B14'].fill = YELLOW_FILL
    ws['B14'].alignment = CENTER_ALIGN()
    ws['B14'].border = MEDIUM_BORDER

    ws['C14'] = '关键岗位总数'
    ws.merge_cells('C14:D14')
    ws['C14'].alignment = LEFT_ALIGN()
    ws['C14'].border = THIN_BORDER

    # Readiness stats
    merge_and_style(ws, 15, 1, 15, 10, '继任就绪度统计', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    readiness_data = [
        ['就绪度', '人数', '占比', '说明'],
        ['4-完全就绪', 0, '=B16/9', '可以立即继任'],
        ['3-基本就绪', 0, '=B17/9', '短期内可继任'],
        ['2-准备中', 0, '=B18/9', '需要继续培养'],
        ['1-空缺风险高', 0, '=B19/9', '急需解决'],
    ]
    for i, row_data in enumerate(readiness_data):
        row = 16 + i
        create_data_row(ws, row, row_data[:4],
                       fills=[LIGHT_GRAY_FILL]*4 if i == 0 else [None]*4,
                       fonts=[LABEL_FONT]*4 if i == 0 else [INPUT_FONT]*4)
        if i > 0:
            ws[f'C{row}'].number_format = '0.0%'
            ws.row_dimensions[row].height = 22

    set_column_widths(ws, {'A': 6, 'B': 14, 'C': 10, 'D': 10, 'E': 8,
                          'F': 14, 'G': 10, 'H': 10, 'I': 12, 'J': 20})

# ==================== FORM 4: IDP ====================
def create_form4(ws, is_example=False):
    ws.title = "高潜人才发展计划"

    merge_and_style(ws, 1, 1, 1, 9, '高潜人才发展计划表（IDP）', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
    ws.row_dimensions[1].height = 35

    merge_and_style(ws, 2, 1, 2, 9, '人才基本信息', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A2:I2')
    ws['A2'].alignment = CENTER_ALIGN()

    basic_info = ['姓名', '部门', '当前岗位', '入职日期', '评估日期', '潜力等级', '综合潜力得分', '直接上级', '']
    create_data_row(ws, 3, basic_info, fills=[LIGHT_GRAY_FILL]*9, fonts=[LABEL_FONT]*9)

    for col in range(1, 10):
        create_input_cell(ws, 4, col)
    ws.merge_cells('I3:I4')
    ws.row_dimensions[4].height = 25

    if is_example:
        ws['A4'].value = '王海（示例）'
        ws['B4'].value = '技术部'
        ws['C4'].value = '高级工程师'
        ws['D4'].value = '2020-06-15'
        ws['E4'].value = '2024-03-15'
        ws['F4'].value = 'A（高潜力）'
        ws['G4'].value = 4.2
        ws['H4'].value = '李总监'

    # Capability assessment
    merge_and_style(ws, 5, 1, 5, 9, '能力评估详情', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A5:I5')
    ws['A5'].alignment = CENTER_ALIGN()

    capability_headers = ['能力维度', '当前能力\n水平(1-5)', '期望水平\n(1-5)', '差距', '提升方式', '时间要求', '资源支持', '评估结果', '']
    for col, val in enumerate(capability_headers, 1):
        cell = ws.cell(row=6, column=col, value=val)
        cell.font = LABEL_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.alignment = CENTER_ALIGN()
        cell.border = THIN_BORDER
    ws.merge_cells('I6:I8')
    ws['I6'].value = '备注'
    ws['I6'].alignment = CENTER_ALIGN()
    ws['I6'].border = THIN_BORDER
    ws.row_dimensions[6].height = 35

    capability_items = [
        '战略思维能力', '领导力', '沟通影响力', '创新能力',
        '业务洞察力', '团队管理能力', '专业技能',
    ]

    for i, cap in enumerate(capability_items):
        row = 7 + i
        ws[f'A{row}'] = cap
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'A{row}'].fill = LIGHT_GRAY_FILL
        ws[f'A{row}'].alignment = LEFT_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws[f'B{row}'] = ''
        ws[f'C{row}'] = ''
        ws[f'D{row}'] = f'=C{row}-B{row}'
        ws[f'D{row}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws[f'D{row}'].alignment = CENTER_ALIGN()
        ws[f'D{row}'].border = THIN_BORDER

        for col in ['E', 'F', 'G', 'H']:
            create_input_cell(ws, row, ord(col) - ord('A') + 1)
        ws.row_dimensions[row].height = 25

    if is_example:
        example_caps = [
            [4, 5, '培训+轮岗', '6个月', '外部教练', ''],
            [3, 5, '导师制', '12个月', '管理课程', ''],
            [4, 5, '项目锻炼', '6个月', '', ''],
            [3, 4, '创新工作坊', '3个月', '', ''],
            [3, 5, '轮岗+业务学习', '12个月', '业务导师', ''],
            [3, 4, '带团队项目', '6个月', '', ''],
            [4, 5, '专业认证', '12个月', '考试费用', ''],
        ]
        for i, caps in enumerate(example_caps):
            row = 7 + i
            ws[f'B{row}'].value = caps[0]
            ws[f'C{row}'].value = caps[1]
            ws[f'E{row}'].value = caps[2]
            ws[f'F{row}'].value = caps[3]
            ws[f'G{row}'].value = caps[4]
            ws[f'H{row}'].value = caps[5]

    # Development goals
    merge_and_style(ws, 14, 1, 14, 9, '发展目标', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A14:I14')
    ws['A14'].alignment = CENTER_ALIGN()

    goal_headers = ['序号', '目标类型', '具体目标', '衡量标准', '完成时间', '状态', '', '', '']
    create_data_row(ws, 15, goal_headers, fills=[LIGHT_GRAY_FILL]*9, fonts=[LABEL_FONT]*9)
    ws.merge_cells('G15:I15')

    for row in range(16, 20):
        ws[f'A{row}'] = row - 15
        ws[f'A{row}'].alignment = CENTER_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws[f'B{row}'] = '短期/中期/长期'
        ws[f'B{row}'].font = Font(name='微软雅黑', size=10, color='BFBFBF')
        ws[f'B{row}'].alignment = LEFT_ALIGN()
        ws[f'B{row}'].border = THIN_BORDER

        for col in range(3, 10):
            create_input_cell(ws, row, col)
        ws.merge_cells(f'G{row}:I{row}')
        ws.row_dimensions[row].height = 25

    if is_example:
        goals = [
            ['短期', '完成管理基础课程', '通过认证考试', '2024-06-30', '进行中'],
            ['中期', '主导跨部门项目', '项目成功交付', '2024-12-31', '计划中'],
            ['长期', '晋升为部门副总监', '试用期通过', '2025-12-31', '规划中'],
            ['', '', '', '', ''],
        ]
        for i, goal in enumerate(goals):
            for j, val in enumerate(goal):
                ws.cell(row=16+i, column=j+2, value=val)

    # Milestones
    merge_and_style(ws, 20, 1, 20, 9, '里程碑与评估', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A20:I20')
    ws['A20'].alignment = CENTER_ALIGN()

    milestone_headers = ['时间节点', '里程碑内容', '评估方式', '评估人', '评估结果', '', '', '', '']
    create_data_row(ws, 21, milestone_headers, fills=[LIGHT_GRAY_FILL]*9, fonts=[LABEL_FONT]*9)
    ws.merge_cells('F21:I21')

    milestones = [
        ['2024-06-30', '完成管理课程学习', '考试+心得报告', '李总监', ''],
        ['2024-09-30', '完成第一次轮岗', '轮岗总结', 'HRBP', ''],
        ['2024-12-31', '主导跨部门项目', '项目成果汇报', '张总', ''],
        ['2025-06-30', '晋升评估', '360评估+答辩', '评审委员会', ''],
    ]
    for i, milestone in enumerate(milestones):
        row = 22 + i
        for j, val in enumerate(milestone):
            ws.cell(row=row, column=j+1, value=val)
            ws.cell(row=row, column=j+1).font = INPUT_FONT
            ws.cell(row=row, column=j+1).alignment = LEFT_ALIGN() if j in [0, 1, 2, 3] else CENTER_ALIGN()
            ws.cell(row=row, column=j+1).border = THIN_BORDER
        ws.merge_cells(f'E{row}:I{row}')
        ws.row_dimensions[row].height = 25

    set_column_widths(ws, {'A': 12, 'B': 10, 'C': 12, 'D': 10, 'E': 14, 'F': 10, 'G': 10, 'H': 10, 'I': 12})

# ==================== FORM 5: Annual Calendar ====================
def create_form5(ws, is_example=False):
    ws.title = "人才盘点年度日历"

    merge_and_style(ws, 1, 1, 1, 13, '人才盘点年度日历', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
    ws.row_dimensions[1].height = 35

    ws['A2'] = '年度：'
    ws['A2'].font = LABEL_FONT
    ws['A2'].alignment = RIGHT_ALIGN()

    ws['B2'] = 2024
    ws['B2'].font = Font(name='微软雅黑', size=11, bold=True)
    ws['B2'].alignment = CENTER_ALIGN()
    ws['B2'].border = THIN_BORDER

    ws['C2'] = '年'
    ws['C2'].font = LABEL_FONT

    months = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']

    create_header_row(ws, 3, ['维度', ''] + months, SUBHEADER_FILL, SUBHEADER_FONT, 25)
    ws.merge_cells('B3:C3')
    ws['B3'].value = '活动类型'
    ws['B3'].alignment = CENTER_ALIGN()

    activities = [
        ['人才盘点会议', '季度经营分析会后1周内'],
        ['九宫格更新', '每次盘点会议后'],
        ['IDP更新', '每季度末'],
        ['继任计划Review', '半年度战略会议'],
        ['培训计划制定', '年度预算周期'],
        ['高潜人才评估', '年度/半年度'],
        ['关键岗位Review', '季度经营分析会'],
    ]

    for i, (activity, timing) in enumerate(activities):
        row = 4 + i
        ws[f'A{row}'] = activity
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'A{row}'].fill = LIGHT_GRAY_FILL
        ws[f'A{row}'].alignment = LEFT_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = timing
        ws[f'B{row}'].font = NOTE_FONT
        ws[f'B{row}'].alignment = LEFT_ALIGN()
        ws[f'B{row}'].border = THIN_BORDER

        for col in range(4, 16):
            create_input_cell(ws, row, col)
        ws.row_dimensions[row].height = 22

    # Status tracking
    merge_and_style(ws, 11, 1, 11, 14, '完成状态追踪', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    status_headers = ['活动', '1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月', '状态']
    create_data_row(ws, 12, status_headers, fills=[LIGHT_GRAY_FILL]*14, fonts=[LABEL_FONT]*14)

    for i, activity in enumerate(['人才盘点会议', '九宫格更新', 'IDP更新', '继任计划Review']):
        row = 13 + i
        ws[f'A{row}'] = activity
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'A{row}'].fill = LIGHT_GRAY_FILL
        ws[f'A{row}'].alignment = LEFT_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        for col in range(2, 15):
            cell = ws.cell(row=row, column=col)
            if is_example:
                if col <= 6:
                    cell.value = '✓'
                    cell.font = Font(name='微软雅黑', size=12, color='00B050')
                elif col == 7:
                    cell.value = '进行中'
                    cell.font = Font(name='微软雅黑', size=9, color='0000FF')
                else:
                    cell.value = ''
                    cell.font = Font(name='微软雅黑', size=10, color='BFBFBF')
            else:
                cell.value = ''
                cell.font = Font(name='微软雅黑', size=10, color='BFBFBF')
            cell.alignment = CENTER_ALIGN()
            cell.border = THIN_BORDER

        ws[f'N{row}'] = f'=COUNTIF(B{row}:M{row},"✓")&"/12完成"'
        ws[f'N{row}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws[f'N{row}'].alignment = CENTER_ALIGN()
        ws[f'N{row}'].border = THIN_BORDER
        ws[f'N{row}'].fill = GREEN_FILL if is_example else LIGHT_GRAY_FILL
        ws.row_dimensions[row].height = 22

    # Owner assignment
    merge_and_style(ws, 17, 1, 17, 14, '责任人分配', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A17:N17')
    ws['A17'].alignment = CENTER_ALIGN()

    owner_headers = ['活动', '负责人', '协助人', 'Q1', 'Q2', 'Q3', 'Q4', '', '', '', '', '', '', '']
    create_data_row(ws, 18, owner_headers, fills=[LIGHT_GRAY_FILL]*14, fonts=[LABEL_FONT]*14)
    ws.merge_cells('H18:N18')

    for i, activity in enumerate(['人才盘点会议', '九宫格更新', 'IDP更新', '继任计划Review', '培训计划制定']):
        row = 19 + i
        ws[f'A{row}'] = activity
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'A{row}'].fill = LIGHT_GRAY_FILL
        ws[f'A{row}'].alignment = LEFT_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws[f'B{row}'] = 'HRBP'
        ws[f'B{row}'].font = INPUT_FONT
        ws[f'B{row}'].alignment = CENTER_ALIGN()
        ws[f'B{row}'].border = THIN_BORDER

        ws[f'C{row}'] = '各部门负责人'
        ws[f'C{row}'].font = INPUT_FONT
        ws[f'C{row}'].alignment = CENTER_ALIGN()
        ws[f'C{row}'].border = THIN_BORDER

        for col in range(4, 8):
            ws.cell(row=row, column=col).value = '●'
            ws.cell(row=row, column=col).font = Font(name='微软雅黑', size=14, color='2E75B6')
            ws.cell(row=row, column=col).alignment = CENTER_ALIGN()
            ws.cell(row=row, column=col).border = THIN_BORDER

        ws.merge_cells(f'H{row}:N{row}')
        ws.row_dimensions[row].height = 25

    set_column_widths(ws, {'A': 14, 'B': 12, 'C': 10, 'D': 8, 'E': 8, 'F': 8, 'G': 8, 'H': 8,
                          'I': 8, 'J': 8, 'K': 8, 'L': 8, 'M': 8, 'N': 10})

# ==================== FORM 6: Talent Pool Tracking ====================
def create_form6(ws, is_example=False):
    ws.title = "梯队建设进度追踪"

    merge_and_style(ws, 1, 1, 1, 9, '梯队建设进度追踪表', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
    ws.row_dimensions[1].height = 35

    merge_and_style(ws, 2, 1, 2, 9,
                    '说明：追踪各梯队层级的建设进度 | 建设进度 = 已到位人数 / 目标人数',
                    LIGHT_BLUE_FILL, NOTE_FONT, LEFT_ALIGN())
    ws.row_dimensions[2].height = 25

    headers = ['梯队层级', '关键岗位\n数量', '目标人数', '已到位\n人数', '储备中\n人数', '空缺\n人数', '建设进度', '风险标注', '备注']
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER_ALIGN()
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 40

    levels = [
        ['高层管理梯队', 'CXO直管', 5, 3, 1, 1],
        ['中层管理梯队', '部门总监/副总监', 15, 10, 3, 2],
        ['基层管理梯队', '经理/主管', 40, 35, 8, 0],
        ['专业序列梯队', '高级专员/专家', 30, 25, 10, 2],
        ['后备人才梯队', '管培生/高潜人才', 20, 12, 15, 0],
    ]

    level_colors = [
        PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid'),
        PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid'),
        PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
        PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid'),
    ]

    for i, (level, positions, target, in_place, reserve, vacant) in enumerate(levels):
        row = 4 + i
        ws[f'A{row}'] = level
        ws[f'A{row}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws[f'A{row}'].fill = level_colors[i]
        ws[f'A{row}'].alignment = CENTER_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws[f'B{row}'] = positions
        ws[f'B{row}'].font = INPUT_FONT
        ws[f'B{row}'].fill = level_colors[i]
        ws[f'B{row}'].alignment = LEFT_ALIGN()
        ws[f'B{row}'].border = THIN_BORDER

        ws[f'C{row}'] = target
        ws[f'C{row}'].font = INPUT_FONT
        ws[f'C{row}'].alignment = CENTER_ALIGN()
        ws[f'C{row}'].border = THIN_BORDER

        ws[f'D{row}'] = in_place if is_example else ''
        ws[f'D{row}'].font = INPUT_FONT
        ws[f'D{row}'].alignment = CENTER_ALIGN()
        ws[f'D{row}'].border = THIN_BORDER

        ws[f'E{row}'] = reserve if is_example else ''
        ws[f'E{row}'].font = INPUT_FONT
        ws[f'E{row}'].alignment = CENTER_ALIGN()
        ws[f'E{row}'].border = THIN_BORDER

        ws[f'F{row}'] = vacant if is_example else ''
        ws[f'F{row}'].font = INPUT_FONT
        ws[f'F{row}'].alignment = CENTER_ALIGN()
        ws[f'F{row}'].border = THIN_BORDER

        ws[f'G{row}'] = f'=IF(C{row}=0,0,D{row}/C{row})'
        ws[f'G{row}'].number_format = '0.0%'
        ws[f'G{row}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws[f'G{row}'].alignment = CENTER_ALIGN()
        ws[f'G{row}'].border = THIN_BORDER

        risk_text = '' if not is_example else ('高风险' if vacant > 2 else '中风险' if vacant > 0 else '正常')
        ws[f'H{row}'] = risk_text
        ws[f'H{row}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws[f'H{row}'].alignment = CENTER_ALIGN()
        ws[f'H{row}'].border = THIN_BORDER
        if is_example:
            if vacant > 2:
                ws[f'H{row}'].fill = ORANGE_FILL
            elif vacant > 0:
                ws[f'H{row}'].fill = YELLOW_FILL
            else:
                ws[f'H{row}'].fill = GREEN_FILL

        ws[f'I{row}'] = ''
        ws[f'I{row}'].border = THIN_BORDER
        ws.row_dimensions[row].height = 30

    # Total row
    ws['A9'] = '合计'
    ws['A9'].font = LABEL_FONT
    ws['A9'].fill = LIGHT_GRAY_FILL
    ws['A9'].alignment = CENTER_ALIGN()
    ws['A9'].border = THIN_BORDER

    for col, formula in [('B', '=COUNTA(B4:B8)'), ('C', '=SUM(C4:C8)'), ('D', '=SUM(D4:D8)'),
                          ('E', '=SUM(E4:E8)'), ('F', '=SUM(F4:F8)'), ('G', '=IF(C9=0,0,D9/C9)')]:
        ws[f'{col}9'] = formula
        ws[f'{col}9'].font = Font(name='微软雅黑', size=10, bold=True)
        ws[f'{col}9'].fill = YELLOW_FILL
        ws[f'{col}9'].alignment = CENTER_ALIGN()
        ws[f'{col}9'].border = MEDIUM_BORDER
    ws['G9'].number_format = '0.0%'
    ws['H9'].border = THIN_BORDER
    ws['I9'].border = THIN_BORDER

    # Risk summary
    merge_and_style(ws, 10, 1, 10, 9, '风险汇总', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    risk_headers = ['风险等级', '梯队层级', '空缺人数', '主要原因', '应对措施', '', '', '', '']
    create_data_row(ws, 11, risk_headers, fills=[LIGHT_GRAY_FILL]*9, fonts=[LABEL_FONT]*9)
    ws.merge_cells('E11:I11')

    for row in range(12, 15):
        for col in range(1, 10):
            create_input_cell(ws, row, col)
        ws.merge_cells(f'E{row}:I{row}')
        ws.row_dimensions[row].height = 25

    if is_example:
        risks = [
            ['高风险', '高层管理梯队', 1, '内部晋升周期长', '加快外部招聘'],
            ['中风险', '中层管理梯队', 2, '关键岗位人才储备不足', '启动继任者加速培养'],
            ['中风险', '专业序列梯队', 2, '核心技术人才稀缺', '校企合作+高薪引进'],
        ]
        for i, risk in enumerate(risks):
            for j, val in enumerate(risk):
                ws.cell(row=12+i, column=j+1, value=val)
                ws.cell(row=12+i, column=j+1).font = INPUT_FONT
                ws.cell(row=12+i, column=j+1).alignment = LEFT_ALIGN() if j in [3, 4] else CENTER_ALIGN()
                ws.cell(row=12+i, column=j+1).border = THIN_BORDER
            if risk[0] == '高风险':
                ws.cell(row=12+i, column=1).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif risk[0] == '中风险':
                ws.cell(row=12+i, column=1).fill = ORANGE_FILL

    set_column_widths(ws, {'A': 14, 'B': 16, 'C': 8, 'D': 8, 'E': 8, 'F': 8, 'G': 10, 'H': 10, 'I': 16})

# ==================== GUIDE SHEET CREATION ====================
def create_guide_sheet(wb, form_name, form_desc, usage_steps, notes):
    ws = wb.create_sheet()
    ws.title = form_name[:31]

    merge_and_style(ws, 1, 1, 1, 6, f'{form_name} - 使用指引', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())

    merge_and_style(ws, 2, 1, 2, 6, '表单简介', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())
    ws.merge_cells('A3:F5')
    ws['A3'] = form_desc
    ws['A3'].font = Font(name='微软雅黑', size=10)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws['A3'].border = THIN_BORDER

    merge_and_style(ws, 6, 1, 6, 6, '使用步骤', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    for i, step in enumerate(usage_steps):
        row = 7 + i
        ws[f'A{row}'] = f'第{i+1}步'
        ws[f'A{row}'].font = Font(name='微软雅黑', size=10, bold=True, color='1F4E79')
        ws[f'A{row}'].fill = LIGHT_BLUE_FILL
        ws[f'A{row}'].alignment = CENTER_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'] = step
        ws[f'B{row}'].font = INPUT_FONT
        ws[f'B{row}'].alignment = LEFT_ALIGN()
        ws[f'B{row}'].border = THIN_BORDER
        ws.row_dimensions[row].height = 25

    start_row = 7 + len(usage_steps) + 1
    merge_and_style(ws, start_row, 1, start_row, 6, '注意事项', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    for i, note in enumerate(notes):
        row = start_row + 1 + i
        ws[f'A{row}'] = '!'
        ws[f'A{row}'].font = Font(name='微软雅黑', size=12, bold=True, color='FF0000')
        ws[f'A{row}'].alignment = CENTER_ALIGN()
        ws[f'A{row}'].border = THIN_BORDER

        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'] = note
        ws[f'B{row}'].font = NOTE_FONT
        ws[f'B{row}'].alignment = LEFT_ALIGN()
        ws[f'B{row}'].border = THIN_BORDER
        ws.row_dimensions[row].height = 22

    set_column_widths(ws, {'A': 8, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16})

# ==================== CREATE FILES ====================
def create_guide_file():
    print("Creating 表单使用指引.xlsx...")
    wb = Workbook()
    wb.remove(wb.active) if wb.active else None

    # Cover sheet
    ws_cover = wb.create_sheet("封面")

    merge_and_style(ws_cover, 1, 1, 1, 8, '人才盘点与梯队建设', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
    merge_and_style(ws_cover, 2, 1, 2, 8, '配套表单使用指引', HEADER_FILL, HEADER_FONT, CENTER_ALIGN())
    merge_and_style(ws_cover, 3, 1, 3, 8, '课程ID: 11-人才盘点与梯队建设', LIGHT_BLUE_FILL, NOTE_FONT, CENTER_ALIGN())
    merge_and_style(ws_cover, 5, 1, 5, 8, '表单目录', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    headers = ['编号', '表单名称', '表单类型', '主要使用者', '', '', '', '']
    create_data_row(ws_cover, 6, headers, fills=[LIGHT_GRAY_FILL]*8, fonts=[LABEL_FONT]*8)
    ws_cover.merge_cells('E6:H6')

    forms_list = [
        ('表单1', '九宫格人才评估表', '核心评估工具', '员工+评估人'),
        ('表单2', '人才盘点会议记录表', '会议纪要工具', 'HRBP+主持人'),
        ('表单3', '关键岗位继任地图', '继任计划工具', 'HR总监+业务负责人'),
        ('表单4', '高潜人才发展计划表', '个人发展工具', '员工+直接上级+HRBP'),
        ('表单5', '人才盘点年度日历', '计划管理工具', 'HRBP'),
        ('表单6', '梯队建设进度追踪表', '进度管控工具', 'HR总监'),
    ]

    for i, (num, name, form_type, user) in enumerate(forms_list):
        row = 7 + i
        ws_cover.cell(row=row, column=1, value=num)
        ws_cover.cell(row=row, column=1).font = Font(name='微软雅黑', size=10, bold=True)
        ws_cover.cell(row=row, column=1).alignment = CENTER_ALIGN()
        ws_cover.cell(row=row, column=1).border = THIN_BORDER

        ws_cover.cell(row=row, column=2, value=name)
        ws_cover.cell(row=row, column=2).font = INPUT_FONT
        ws_cover.cell(row=row, column=2).alignment = LEFT_ALIGN()
        ws_cover.cell(row=row, column=2).border = THIN_BORDER

        ws_cover.cell(row=row, column=3, value=form_type)
        ws_cover.cell(row=row, column=3).font = INPUT_FONT
        ws_cover.cell(row=row, column=3).alignment = CENTER_ALIGN()
        ws_cover.cell(row=row, column=3).border = THIN_BORDER

        ws_cover.cell(row=row, column=4, value=user)
        ws_cover.cell(row=row, column=4).font = INPUT_FONT
        ws_cover.cell(row=row, column=4).alignment = CENTER_ALIGN()
        ws_cover.cell(row=row, column=4).border = THIN_BORDER

        ws_cover.merge_cells(f'E{row}:H{row}')
        ws_cover.cell(row=row, column=5).border = THIN_BORDER

    # Workflow section
    merge_and_style(ws_cover, 14, 1, 14, 8, '表单使用流程', SUBHEADER_FILL, SUBHEADER_FONT, CENTER_ALIGN())

    workflow = [
        ['Step 1', '准备阶段', '收集员工基本信息和历史评估数据', 'HRBP', '盘点会前1周'],
        ['Step 2', '评估阶段', '直接主管对员工进行九宫格评估和潜力评估', '直接主管', '盘点会前3天'],
        ['Step 3', '校准阶段', '召开人才盘点会议，校准评估结果', 'HR+业务负责人', '盘点会议'],
        ['Step 4', '规划阶段', '制定IDP、继任计划、发展举措', '员工+上级+HR', '盘点会后2周'],
        ['Step 5', '跟踪阶段', '执行计划并定期跟踪进度', 'HRBP', '季度Review'],
    ]

    wf_headers = ['步骤', '阶段', '活动内容', '负责人', '时间节点']
    headers_wf = wf_headers + ['', '', '', '']
    create_data_row(ws_cover, 15, headers_wf, fills=[LIGHT_GRAY_FILL]*8, fonts=[LABEL_FONT]*8)
    ws_cover.merge_cells('E15:H15')

    for i, wf in enumerate(workflow):
        row = 16 + i
        for j, val in enumerate(wf):
            ws_cover.cell(row=row, column=j+1, value=val)
            ws_cover.cell(row=row, column=j+1).font = Font(name='微软雅黑', size=10, bold=True, color='1F4E79') if j == 0 else INPUT_FONT
            ws_cover.cell(row=row, column=j+1).fill = LIGHT_BLUE_FILL if j == 0 else WHITE_FILL
            ws_cover.cell(row=row, column=j+1).alignment = CENTER_ALIGN() if j in [0, 3, 4] else LEFT_ALIGN()
            ws_cover.cell(row=row, column=j+1).border = THIN_BORDER
        ws_cover.merge_cells(f'E{row}:H{row}')
        ws_cover.row_dimensions[row].height = 25

    set_column_widths(ws_cover, {'A': 10, 'B': 18, 'C': 28, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 14})

    # Guide sheets
    form_guides = [
        ("九宫格人才评估表",
         "九宫格人才评估表是人才盘点的核心工具，通过业绩和潜力两个维度将员工定位到9个格子中。",
         ["确定评估对象范围", "收集评估周期内的业绩数据", "直接主管进行业绩评分", "直接主管进行潜力评分", "系统自动计算并落位", "HRBP与主管校准", "盘点会议确认结果"],
         ["评估标准要客观公正", "潜力评估重点看学习能力、适应变化、领导力", "九宫格落位是动态的"]),
        ("人才盘点会议记录表",
         "人才盘点会议记录表用于记录人才盘点会议的完整内容。",
         ["会前准备", "记录会议基本信息", "确认九宫格分布", "讨论继任计划", "明确行动项", "约定下次会议", "会议纪要确认"],
         ["会议频率建议季度盘点", "行动项必须有明确责任人", "继任计划要讨论备选方案"]),
        ("关键岗位继任地图",
         "关键岗位继任地图用于追踪公司关键岗位的继任准备情况。",
         ["确定关键岗位范围", "确认继任需求", "评估现任者任职风险", "识别潜在继任候选人", "沟通发展意愿", "定期Review进展", "及时更新继任地图"],
         ["关键岗位定义标准", "继任准备度1-4分评估", "每个关键岗位储备1-2名继任者"]),
        ("高潜人才发展计划表",
         "高潜人才发展计划表是针对高潜力人才的个性化发展工具。",
         ["识别高潜人才", "进行能力评估", "沟通职业期望", "分析能力差距", "确定发展方式", "设定里程碑", "定期Review进展"],
         ["高潜识别标准：潜力>=4分", "发展目标要SMART", "发展方式要多元化"]),
        ("人才盘点年度日历",
         "人才盘点年度日历是人才盘点工作的年度规划工具。",
         ["年初制定年度计划", "明确每月活动内容", "分配责任人", "标注关键节点", "按月跟踪状态", "季度Review", "年底总结"],
         ["人才盘点是常态化工作", "关键活动提前规划", "完成状态可视化"]),
        ("梯队建设进度追踪表",
         "梯队建设进度追踪表用于追踪公司各梯队层级的建设进度。",
         ["确定梯队层级定义", "建立关键岗位清单", "盘点现有人才储备", "计算建设进度", "识别空缺风险", "制定补充计划", "定期跟踪更新"],
         ["梯队层级定义要清晰", "建设进度动态指标", "高风险梯队重点关注"]),
    ]

    for form_name, desc, steps, notes in form_guides:
        create_guide_sheet(wb, form_name, desc, steps, notes)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, '表单使用指引.xlsx')
    wb.save(output_path)
    print(f"Created: {output_path}")


def create_blank_file():
    print("Creating 配套表单_空表.xlsx...")
    wb = Workbook()
    wb.remove(wb.active) if wb.active else None

    create_form1(wb.create_sheet(), is_example=False)
    create_form2(wb.create_sheet(), is_example=False)
    create_form3(wb.create_sheet(), is_example=False)
    create_form4(wb.create_sheet(), is_example=False)
    create_form5(wb.create_sheet(), is_example=False)
    create_form6(wb.create_sheet(), is_example=False)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, '配套表单_空表.xlsx')
    wb.save(output_path)
    print(f"Created: {output_path}")


def create_filled_file():
    print("Creating 配套表单_填好版.xlsx...")
    wb = Workbook()
    wb.remove(wb.active) if wb.active else None

    create_form1(wb.create_sheet(), is_example=True)
    create_form2(wb.create_sheet(), is_example=True)
    create_form3(wb.create_sheet(), is_example=True)
    create_form4(wb.create_sheet(), is_example=True)
    create_form5(wb.create_sheet(), is_example=True)
    create_form6(wb.create_sheet(), is_example=True)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, '配套表单_填好版.xlsx')
    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == '__main__':
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    create_guide_file()
    create_blank_file()
    create_filled_file()
    print("\nAll files created successfully!")
    print(f"Output directory: {OUTPUT_DIR}")

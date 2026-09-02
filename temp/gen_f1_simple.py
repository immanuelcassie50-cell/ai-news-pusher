# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os

output_dir = "D:/新课开发/公文写作/3、人机协同写作——从结构化提示到可用初稿/全流程工具表单/Excel版"
os.makedirs(output_dir, exist_ok=True)

header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
normal_font = Font(size=10)
bold_font = Font(size=10, bold=True)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
content_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_row_style(ws, row, num_cols, is_header=False, is_gray=False):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        if is_header:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        else:
            cell.fill = gray_fill if is_gray else white_fill
            cell.font = normal_font
            cell.alignment = content_alignment
        cell.border = thin_border

def set_col_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_f1():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "f1"
    set_col_widths(ws, [8, 20, 45])
    ws.merge_cells("A1:C1")
    ws.cell(row=1, column=1, value="F1")
    set_row_style(ws, 1, 3, is_header=True)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    wb.save(os.path.join(output_dir, "F1_test.xlsx"))
    print("Created F1_test.xlsx")

create_f1()

# -*- coding: utf-8 -*-
"""
Build 配套表单_空表.xlsx for 绩效引擎课程
- 总览 + 18个空表工具
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.datavalidation import DataValidation
import os

OUT_DIR = r"D:\2026年课程\竞越\绩效引擎：让战略真正落地的部门绩效全系统\完整课程包\13-配套表单Excel"
OUT = os.path.join(OUT_DIR, "配套表单_空表.xlsx")

# Brand colors
BRAND_DARK   = "1E3A5F"
BRAND_ACCENT = "E87722"
BRAND_LIGHT  = "F5F0E8"
BRAND_GREY   = "6B7280"
BRAND_GREEN  = "C8E6C9"
BRAND_RED    = "FFCDD2"
BRAND_YELLOW = "FFF9C4"
WHITE        = "FFFFFF"
HINT_FILL    = "F0F4F8"

# Fonts
F_TITLE   = Font(name="微软雅黑", size=20, bold=True, color=WHITE)
F_H1      = Font(name="微软雅黑", size=16, bold=True, color=BRAND_DARK)
F_H2      = Font(name="微软雅黑", size=12, bold=True, color=BRAND_DARK)
F_BODY    = Font(name="微软雅黑", size=10, color="222222")
F_BODY_B  = Font(name="微软雅黑", size=10, bold=True, color="222222")
F_HINT    = Font(name="微软雅黑", size=9, italic=True, color=BRAND_GREY)
F_HEADER  = Font(name="微软雅黑", size=10, bold=True, color=WHITE)
F_TAG     = Font(name="微软雅黑", size=9, bold=True, color=BRAND_DARK)
F_PLACE   = Font(name="微软雅黑", size=10, italic=True, color="BBBBBB")
F_NUM     = Font(name="Calibri", size=10, color="222222")

# Fills
FILL_TITLE   = PatternFill("solid", fgColor=BRAND_DARK)
FILL_ACCENT  = PatternFill("solid", fgColor=BRAND_ACCENT)
FILL_LIGHT   = PatternFill("solid", fgColor=BRAND_LIGHT)
FILL_HEADER  = PatternFill("solid", fgColor=BRAND_DARK)
FILL_SUB     = PatternFill("solid", fgColor="EAEEF3")
FILL_HINT    = PatternFill("solid", fgColor=HINT_FILL)
FILL_GREEN   = PatternFill("solid", fgColor=BRAND_GREEN)
FILL_RED     = PatternFill("solid", fgColor=BRAND_RED)
FILL_YELLOW  = PatternFill("solid", fgColor=BRAND_YELLOW)

thin = Side(border_style="thin", color="C0C0C0")
medium = Side(border_style="medium", color=BRAND_DARK)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_HEAVY = Border(left=medium, right=medium, top=medium, bottom=medium)

WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

# Form metadata
FORMS = [
    ("F01", "出发点自评表", "CH01", "评估起点状态"),
    ("F02", "绩效挑战场景卡", "CH01", "收集真实挑战"),
    ("F03", "四驱诊断表", "CH01", "诊断四驱健康度"),
    ("F04", "失效模式识别表", "CH01", "识别三类失效"),
    ("F05", "目标穿透四层映射工作页", "CH02", "四层映射（最核心）"),
    ("F06", "目标断点诊断表", "CH02", "识别三个断点"),
    ("F07", "目标动态化机制", "CH02", "设计动态调整机制"),
    ("F08", "现有绩效指标盘点表", "CH03", "盘点现有指标"),
    ("F09", "四维框架应用表", "CH03", "用四维设计新指标"),
    ("F10", "牵引力测试表", "CH03", "检验指标牵引力"),
    ("F11", "共识对话流程设计表", "CH04", "设计共识对话流程"),
    ("F12", "共识度评估表", "CH04", "评估共识度"),
    ("F13", "执行追踪机制设计表", "CH05", "设计追踪机制"),
    ("F14", "执行信号监测表", "CH05", "监测执行信号"),
    ("F15", "成长激活方案设计表", "CH06", "设计成长激活"),
    ("F16", "绩效反馈规划表", "CH06", "规划绩效反馈"),
    ("F17", "部门绩效引擎完整设计图", "CH06", "整合六章产出"),
    ("F18", "30天行动承诺表", "CH06", "30天行动承诺"),
]

DIFF = ["⭐", "⭐⭐", "⭐⭐⭐"]


def style_title_row(ws, row, last_col, title, fill=FILL_TITLE, font=F_TITLE, height=40):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=title)
    c.font = font
    c.fill = fill
    c.alignment = CENTER
    ws.row_dimensions[row].height = height


def style_subtitle(ws, row, last_col, text, fill=FILL_LIGHT, font=F_H2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    c.fill = fill
    c.alignment = LEFT
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row, headers, fill=FILL_HEADER, font=F_HEADER):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font
        c.fill = fill
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 28


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_hint(ws, row, last_col, text, fill=FILL_HINT, font=F_HINT, height=36):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    c.fill = fill
    c.alignment = WRAP
    ws.row_dimensions[row].height = height


def add_section(ws, row, last_col, text, fill=FILL_ACCENT):
    """Section divider: orange bar with white text"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", size=12, bold=True, color=WHITE)
    c.fill = fill
    c.alignment = LEFT
    ws.row_dimensions[row].height = 24


def add_field_row(ws, row, last_col, label, placeholder="", height=32, sublabel=""):
    """Add a labeled field row (label on left, value area spans rest)"""
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_BODY_B
    c.fill = FILL_SUB
    c.alignment = LEFT
    c.border = BORDER
    # Merge value area
    if last_col > 1:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
        v = ws.cell(row=row, column=2, value=placeholder)
        v.font = F_PLACE
        v.alignment = LEFT
        v.border = BORDER
    if sublabel:
        c2 = ws.cell(row=row, column=last_col+1 if False else 1, value=sublabel)  # unused
    ws.row_dimensions[row].height = height


def add_input_row(ws, row, last_col, label, placeholder="", height=42):
    """Label cell (column 1) + value cell (merged to last col)"""
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_BODY_B
    c.fill = FILL_SUB
    c.alignment = LEFT
    c.border = BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
    v = ws.cell(row=row, column=2, value=placeholder)
    v.font = F_PLACE
    v.alignment = WRAP
    v.border = BORDER
    ws.row_dimensions[row].height = height


# ============== Build Workbook ==============
wb = openpyxl.Workbook()

# ============== 总览 Sheet ==============
ws = wb.active
ws.title = "总览"
style_title_row(ws, 1, 6, "绩效引擎 配套练习表单 — 空白版（学员填写）", height=44)
style_subtitle(ws, 2, 6, "本工作簿共 18 张工具表，每张对应一个练习环节。建议按章节顺序使用。")

headers = ["编号", "表单名称", "所属章节", "练习内容", "类型", "难度"]
style_header_row(ws, 3, headers)

KEY_FORMS = {"F05", "F09", "F11", "F13", "F15", "F17", "F18"}

for i, (code, name, ch, brief) in enumerate(FORMS):
    r = 4 + i
    typ = "识别L1" if code in ("F02","F04","F06") else \
          "创造L3" if code in ("F01","F05","F09","F11","F13","F15","F17","F18") else \
          "诊断/评估"
    diff = DIFF[1] if code in ("F03","F08","F10","F12","F14","F16") else \
           DIFF[0] if code in ("F01","F02","F04","F06") else \
           DIFF[2]
    vals = [code, name, ch, brief, typ, diff]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = F_BODY
        c.alignment = LEFT if j in (2, 4) else CENTER
        c.border = BORDER
        if code in KEY_FORMS:
            c.fill = FILL_YELLOW
    ws.row_dimensions[r].height = 22

# 章节小计
r_sub = 4 + len(FORMS) + 1
ws.merge_cells(start_row=r_sub, start_column=1, end_row=r_sub, end_column=6)
c = ws.cell(row=r_sub, column=1,
            value="📌 使用说明：黄色高亮的是7个核心产出，建议作为课后30天的重点跟进。")
c.font = F_HINT
c.fill = FILL_HINT
c.alignment = LEFT
ws.row_dimensions[r_sub].height = 30

# 课程四驱概念图
r_sub += 2
ws.merge_cells(start_row=r_sub, start_column=1, end_row=r_sub, end_column=6)
c = ws.cell(row=r_sub, column=1, value="课程核心框架：绩效引擎·四驱模型")
c.font = F_H2
c.fill = FILL_ACCENT
c.font = Font(name="微软雅黑", size=12, bold=True, color=WHITE)
c.alignment = CENTER
ws.row_dimensions[r_sub].height = 24

model_rows = [
    ("穿透力（CH02）", "目标穿透系统", "把组织战略真正送到每个人——F05/F06/F07"),
    ("牵引力（CH03）", "绩效设计系统", "让指标像磁铁一样牵引正确行为——F08/F09/F10"),
    ("推进力（CH05）", "执行追踪系统", "建立贯穿全程的动态推进机制——F13/F14"),
    ("进化力（CH06）", "成长激活系统", "让绩效过程同时驱动人成长——F15/F16"),
    ("收官（CH06末）", "整合·承诺·行动", "F17 完整设计图 + F18 30天承诺"),
]
for i, (k, v, hint) in enumerate(model_rows):
    r = r_sub + 1 + i
    ws.cell(row=r, column=1, value=k).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = LEFT
    ws.cell(row=r, column=1).border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value=v).font = F_BODY
    ws.cell(row=r, column=2).alignment = LEFT
    ws.cell(row=r, column=2).border = BORDER
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws.cell(row=r, column=4, value=hint).font = F_HINT
    ws.cell(row=r, column=4).alignment = LEFT
    ws.cell(row=r, column=4).border = BORDER
    ws.row_dimensions[r].height = 22

set_widths(ws, [10, 26, 12, 30, 12, 10])
ws.sheet_view.showGridLines = False
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A3

# ============== F01 出发点自评表 ==============
ws = wb.create_sheet("F01 出发点自评表")
style_title_row(ws, 1, 4, "F01 出发点自评表", height=42)
style_subtitle(ws, 2, 4, "本表用于课程开始前 / 课程开场 · 评估你当前的绩效管理起点")

add_hint(ws, 3, 4, "💡 提示：本表是个人起点的诚实盘点。\n越真实，越能在课程中找到自己的改善起点。建议用10分钟独立完成。")

# 基本信息
add_section(ws, 4, 4, "一、基本信息")
add_input_row(ws, 5, 4, "姓名/团队", "例：梁磊 · 澄明科技华南区")
add_input_row(ws, 6, 4, "团队规模（人）", "包含自己")
add_input_row(ws, 7, 4, "当前绩效周期", "年度/半年/季度/月度/无固定")

# 现状评估
add_section(ws, 8, 4, "二、现状评估")
add_input_row(ws, 9, 4, "当前主要衡量指标数（个）", "粗略即可，10+表示指标臃肿")
add_input_row(ws, 10, 4, "最痛的管理挑战", "用1-2句直接说，别绕弯")

# 多选
add_input_row(ws, 11, 4, "最希望本课解决什么", "具体到场景，不要泛泛")

# 评分（用data validation）
add_section(ws, 12, 4, "三、起点评估（1-5分，5=最好）")
score_labels = [
    ("团队士气自评(1-5分)", "凭直觉即可，≤2=紧急"),
    ("目标清晰度(1-5分)", "团队能用自己的话说清战略吗"),
    ("执行节奏感(1-5分)", "周/月会议有效还是流于形式"),
    ("AI工具使用频率", "每天/每周/偶尔/几乎不用"),
]
r = 13
for label, hint in score_labels:
    add_input_row(ws, r, 4, f"{label}   提示：{hint}", "请填写")
    r += 1

# 开放补充
add_section(ws, r, 4, "四、开放补充（可选）")
r += 1
add_input_row(ws, r, 4, "还有什么想让讲师/同学知道的", "例如：团队抗拒新工具", height=80)

# 下一步
r += 2
add_section(ws, r, 4, "下一步 → 完成后请进入 F02 挑战场景卡")
r += 1
add_input_row(ws, r, 4, "完成时间", "")

set_widths(ws, [18, 24, 24, 24])
ws.sheet_view.showGridLines = False

# ============== F02 绩效挑战场景卡 ==============
ws = wb.create_sheet("F02 挑战场景卡")
style_title_row(ws, 1, 3, "F02 绩效挑战场景卡", height=42)
style_subtitle(ws, 2, 3, "把你最痛的那个管理挑战写下来——这门课就是从你的真实问题出发")
add_hint(ws, 3, 3, "💡 提示：场景要具体（哪个团队、什么问题、多久了），不要写成抽象命题。")

add_section(ws, 4, 3, "挑战描述")
add_input_row(ws, 5, 3, "挑战场景（≥20字）", "具体到团队/客户/任务", height=60)
add_input_row(ws, 6, 3, "涉及人员（谁牵涉其中）", "你+团队+客户/上级等", height=42)
add_input_row(ws, 7, 3, "这个问题持续了多久", "1月内/3月内/半年/1年+/多年", height=36)

add_section(ws, 8, 3, "你已经试过什么")
add_input_row(ws, 9, 3, "试过什么方法1", "例如：开会宣贯/单独谈话/给激励", height=42)
add_input_row(ws, 10, 3, "结果是什么", "真实结果，不写期待", height=42)
add_input_row(ws, 11, 3, "试过什么方法2（可选）", "", height=42)
add_input_row(ws, 12, 3, "结果是什么", "", height=42)

add_section(ws, 13, 3, "你卡在哪里")
add_input_row(ws, 14, 3, "你具体不知道做什么", "开放写", height=60)
add_input_row(ws, 15, 3, "你希望本课带走什么", "具体到一个产出", height=42)

add_section(ws, 16, 3, "与AI时代的关系")
add_input_row(ws, 17, 3, "这个问题与AI有关吗", "强/中/弱/无关", height=36)
add_input_row(ws, 18, 3, "如果是，体现在哪里", "例如：AI工具让团队焦虑", height=60)

set_widths(ws, [22, 35, 35])
ws.sheet_view.showGridLines = False


# ============== F03 四驱诊断表 ==============
ws = wb.create_sheet("F03 四驱诊断表")
style_title_row(ws, 1, 4, "F03 四驱诊断表", height=42)
style_subtitle(ws, 2, 4, "诊断本部门四驱（穿透/牵引/推进/进化）的健康度——找到最薄弱的驱动力")
add_hint(ws, 3, 4, "💡 评分：1=完全失效 2=明显有问题 3=勉强及格 4=基本健康 5=自运转良好。\n用具体例子支撑你的评分，避免印象分。")

add_section(ws, 4, 4, "四驱评分与具体表现")
headers = ["驱动力", "评分(1-5)", "评分依据（举1-2个具体例子）", "最弱/最强/趋势"]
style_header_row(ws, 5, headers)

drivers = [
    ("穿透力（目标是否到每个人）", "战略→部门→个人是否穿透"),
    ("牵引力（指标是否牵引正确行为）", "在衡量『做了什么』还是『创造了什么价值』"),
    ("推进力（执行是否有节奏）", "追踪频率+调整机制"),
    ("进化力（团队是否在成长）", "能力成长+动力结构"),
]
r = 6
for d, hint in drivers:
    ws.cell(row=r, column=1, value=d).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = LEFT
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value=hint).font = F_HINT
    ws.cell(row=r, column=2).alignment = WRAP
    ws.cell(row=r, column=2).border = BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=3, value="").border = BORDER
    ws.cell(row=r, column=3).font = F_PLACE
    ws.row_dimensions[r].height = 50
    r += 1

add_section(ws, r, 4, "诊断结论")
r += 1
add_input_row(ws, r, 4, "最弱驱动力（你的判断）", "穿透/牵引/推进/进化", height=36)
r += 1
add_input_row(ws, r, 4, "改善优先级（下一个季度聚焦哪个）", "具体到动作", height=42)
r += 1
add_input_row(ws, r, 4, "这个月你打算先做什么", "具体1个动作", height=42)

set_widths(ws, [22, 22, 30, 22])
ws.sheet_view.showGridLines = False


# ============== F04 失效模式识别表 ==============
ws = wb.create_sheet("F04 失效模式识别表")
style_title_row(ws, 1, 4, "F04 失效模式识别表", height=42)
style_subtitle(ws, 2, 4, "识别本团队旧有绩效系统的三类失效模式——对应AI时代三大变化")
add_hint(ws, 3, 4, "💡 三类失效对应：①目标时效性 ②衡量错东西 ③动力结构变化。\n每类先选『是/部分/否』，再给具体例子。")

add_section(ws, 4, 4, "三类失效诊断")
headers = ["失效类型", "是否发生", "具体表现（举例）", "影响范围"]
style_header_row(ws, 5, headers)

failure_types = [
    ("失效一·目标时效性", "AI时代变化一：战略周期从3-5年压缩到6-18个月"),
    ("失效二·衡量错东西", "AI时代变化二：执行型工作被AI接管，KPI在衡量越来越不重要的东西"),
    ("失效三·动力结构", "AI时代变化三：员工动力从外部奖励→意义/成长/贡献"),
]
r = 6
for ft, hint in failure_types:
    ws.cell(row=r, column=1, value=ft).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = LEFT
    ws.cell(row=r, column=1).border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    ws.cell(row=r, column=2, value="是/部分/否").font = F_HINT
    ws.cell(row=r, column=2).alignment = CENTER
    ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=3, value=hint).font = F_HINT
    ws.cell(row=r, column=3).alignment = WRAP
    ws.cell(row=r, column=3).border = BORDER
    ws.cell(row=r, column=4, value="").border = BORDER
    ws.row_dimensions[r].height = 50
    r += 1

add_section(ws, r, 4, "综合判断")
r += 1
add_input_row(ws, r, 4, "最严重的失效（你的判断）", "", height=42)
r += 1
add_input_row(ws, r, 4, "改进顺序（先改哪个）", "", height=42)
r += 1
add_input_row(ws, r, 4, "改进的第一个动作（30天内）", "具体动作", height=42)

set_widths(ws, [22, 16, 35, 22])
ws.sheet_view.showGridLines = False


# ============== F05 目标穿透四层映射工作页 (最核心) ==============
ws = wb.create_sheet("F05 目标穿透四层映射")
style_title_row(ws, 1, 5, "F05 目标穿透四层映射工作页（最核心）", height=44)
style_subtitle(ws, 2, 5, "组织目标 → 部门目标 → 关键举措 → 个人贡献  四层穿透  |  建议60分钟")
add_hint(ws, 3, 5,
    "💡 本表是本课最核心工具。穿透度=团队成员能用自己的话说『我的工作如何支持组织目标』\n"
    "  使用方法：先填好第1-2层（从公司战略文件来）→ 第3层3-7个关键举措 → 第4层每个团队成员1-2句具体贡献\n"
    "  关键质量：第4层必须是动词开头的具体动作，不是『配合』『支持』这种模糊词。")

# 第1层
add_section(ws, 4, 5, "第 1 层 · 组织战略目标（来源：公司战略文件 / 老板讲话）")
add_field_row_ex = ws, 5, 5, [
    ("组织战略目标（一句话）", "例：用AI工具将客户续约率从88%提至95%"),
    ("截止时间", "明确日期"),
    ("公司层衡量指标", "公司如何衡量这个目标"),
]
r = 5
for label, ph in [
    ("组织战略目标（一句话）", "例：用AI工具将客户续约率从88%提至95%"),
    ("截止时间", "明确日期"),
    ("公司层衡量指标", "公司如何衡量这个目标"),
]:
    c = ws.cell(row=r, column=1, value=label)
    c.font = F_BODY_B; c.fill = FILL_SUB; c.alignment = LEFT; c.border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    v = ws.cell(row=r, column=2, value=ph)
    v.font = F_PLACE; v.alignment = LEFT; v.border = BORDER
    ws.row_dimensions[r].height = 32
    r += 1

# 第2层
add_section(ws, r, 5, "第 2 层 · 部门目标（翻译：把公司目标变成部门语言）")
r += 1
for label, ph in [
    ("部门目标（一句话）", "必须承接第1层"),
    ("部门层衡量指标", "用什么衡量部门目标"),
    ("关键成果领域（3-5条）", "覆盖度>深度"),
]:
    c = ws.cell(row=r, column=1, value=label)
    c.font = F_BODY_B; c.fill = FILL_SUB; c.alignment = LEFT; c.border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    v = ws.cell(row=r, column=2, value=ph)
    v.font = F_PLACE; v.alignment = WRAP; v.border = BORDER
    ws.row_dimensions[r].height = 60
    r += 1

# 第3层
add_section(ws, r, 5, "第 3 层 · 关键举措（3-7个具体行动，动词开头）")
r += 1
# 表头
headers3 = ["#", "关键举措", "责任人", "里程碑", "衡量指标"]
style_header_row(ws, r, headers3)
r += 1
for i in range(1, 8):
    ws.cell(row=r, column=1, value=i).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).border = BORDER
    for col in range(2, 6):
        c = ws.cell(row=r, column=col, value="")
        c.font = F_PLACE
        c.alignment = LEFT
        c.border = BORDER
    ws.row_dimensions[r].height = 36
    r += 1

# 第4层
add_section(ws, r, 5, "第 4 层 · 团队成员个人贡献（每人 1-2 句具体动作）")
r += 1
add_hint(ws, r, 5, "⚠ 质量标准：动词开头 / 行为可观察 / 不写『配合』『支持』『协助』这种模糊词。", height=28)
r += 1
headers4 = ["#", "团队成员", "角色/岗位", "个人贡献1", "个人贡献2"]
style_header_row(ws, r, headers4)
r += 1
for i in range(1, 14):  # 留12人+2余量
    ws.cell(row=r, column=1, value=i).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).border = BORDER
    for col in range(2, 6):
        c = ws.cell(row=r, column=col, value="")
        c.font = F_PLACE
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[r].height = 40
    r += 1

# 穿透度自检
add_section(ws, r, 5, "穿透度自检")
r += 1
headers_check = ["自检问题", "评分(1-5)", "判断依据", "未穿透的断点", "修复策略"]
style_header_row(ws, r, headers_check)
r += 1
for q in [
    "团队成员能用自己的话说清部门目标",
    "团队成员能说清自己的工作如何支持部门目标",
    "第3层举措能直接指向部门目标",
    "第1层组织目标清晰且可衡量",
]:
    c = ws.cell(row=r, column=1, value=q)
    c.font = F_BODY; c.alignment = WRAP; c.border = BORDER
    for col in range(2, 6):
        c2 = ws.cell(row=r, column=col, value="")
        c2.font = F_PLACE
        c2.alignment = WRAP
        c2.border = BORDER
    ws.row_dimensions[r].height = 40
    r += 1

# 整体判断
add_section(ws, r, 5, "整体判断")
r += 1
add_input_row(ws, r, 5, "整体穿透度自评(1-5)", "5=每个成员都能用自己的话说清", height=36)
r += 1
add_input_row(ws, r, 5, "未穿透的最关键断点", "哪里没穿透", height=42)
r += 1
add_input_row(ws, r, 5, "下一步修复动作", "具体动作", height=42)

set_widths(ws, [22, 18, 18, 22, 22])
ws.sheet_view.showGridLines = False


# ============== F06 目标断点诊断表 ==============
ws = wb.create_sheet("F06 目标断点诊断表")
style_title_row(ws, 1, 4, "F06 目标断点诊断表", height=42)
style_subtitle(ws, 2, 4, "识别本部门目标传递的三个断点：宣讲损失/理解偏差/承接缺位")
add_hint(ws, 3, 4, "💡 断点一：战略宣贯中损失了什么\n断点二：团队的理解vs真实意图\n断点三：谁没承接/为什么")

add_section(ws, 4, 4, "三个断点诊断")
headers = ["断点类型", "具体表现（1-2句）", "受影响范围", "严重度(1-5)"]
style_header_row(ws, 5, headers)

bps = [
    "断点一·宣讲损失（细节在转述中模糊）",
    "断点二·理解偏差（团队理解≠真实意图）",
    "断点三·承接缺位（谁没承接+为什么）",
]
r = 6
for bp in bps:
    ws.cell(row=r, column=1, value=bp).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = WRAP
    ws.cell(row=r, column=1).border = BORDER
    for col in range(2, 5):
        c = ws.cell(row=r, column=col, value="")
        c.font = F_PLACE
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[r].height = 50
    r += 1

add_section(ws, r, 4, "综合判断")
r += 1
add_input_row(ws, r, 4, "最严重断点（你的判断）", "", height=36)
r += 1
add_input_row(ws, r, 4, "根本原因分析", "为什么这个断点最严重", height=60)
r += 1
add_input_row(ws, r, 4, "修复策略（具体动作）", "30天内做什么", height=60)

set_widths(ws, [24, 32, 18, 18])
ws.sheet_view.showGridLines = False


# ============== F07 目标动态化机制 ==============
ws = wb.create_sheet("F07 目标动态化机制")
style_title_row(ws, 1, 3, "F07 目标动态化机制", height=42)
style_subtitle(ws, 2, 3, "设计AI时代的目标动态调整机制——季度+事件触发")
add_hint(ws, 3, 3, "💡 AI时代目标周期缩短：年度目标可能Q3已过时。需要内建调整机制。\n建议：保持『方向锚定』（价值观/客户价值主张）稳定，『具体目标』可调。")

add_section(ws, 4, 3, "动态调整机制设计")
fields7 = [
    ("调整触发条件（≥3条）", "例如：竞品出AI功能/客户需求变化/公司战略调整"),
    ("调整频率", "月度/季度/半年/事件触发"),
    ("调整决策人", "谁有权调"),
    ("调整沟通方式", "怎么向团队说"),
    ("保持稳定的部分（锚定）", "价值观/客户价值主张/部门总方向"),
    ("调整记录方式", "留痕工具"),
    ("极端调整预案", "重大变化时怎么办"),
]
r = 5
for label, ph in fields7:
    add_input_row(ws, r, 3, label, ph, height=50)
    r += 1

add_input_row(ws, r, 3, "第一次动态复盘时间", "例：3个月后", height=36)

set_widths(ws, [22, 35, 35])
ws.sheet_view.showGridLines = False


# ============== F08 现有绩效指标盘点表 ==============
ws = wb.create_sheet("F08 现有指标盘点表")
style_title_row(ws, 1, 5, "F08 现有绩效指标盘点表", height=42)
style_subtitle(ws, 2, 5, "盘点本部门现有绩效指标——为下一步四维框架改造做准备")
add_hint(ws, 3, 5, "💡 至少盘点5个指标。每个评估AI适配性：高（仍适用）/中（需补维度）/低（需淘汰或重设）。\n完成后用底部公式自动统计。")

add_section(ws, 4, 5, "指标盘点")
headers = ["#", "指标名称", "类型", "数据来源", "AI适配性(高/中/低)", "处理(保留/调整/淘汰)"]
style_header_row(ws, 5, headers)

r = 6
for i in range(1, 11):  # 留10行
    ws.cell(row=r, column=1, value=i).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).border = BORDER
    type_options = ["成果/执行/行为/态度"]
    for col, ph in zip(range(2, 7), ["指标名", "类型", "数据来源", "高/中/低", "处理"]):
        c = ws.cell(row=r, column=col, value=ph if i == 1 else "")
        if i == 1:
            c.font = F_HINT
        else:
            c.font = F_PLACE
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[r].height = 32
    r += 1

# 公式统计
add_section(ws, r, 5, "统计（自动计算）")
r += 1
headers_stat = ["指标", "公式", "结果", "解读", "建议"]
style_header_row(ws, r, headers_stat)
r += 1
ws.cell(row=r, column=1, value="指标总数").font = F_BODY_B
ws.cell(row=r, column=1).fill = FILL_SUB
ws.cell(row=r, column=1).border = BORDER
ws.cell(row=r, column=2, value="=COUNTA(B6:B15)").font = F_NUM
ws.cell(row=r, column=2).border = BORDER
ws.cell(row=r, column=3, value="").border = BORDER
ws.cell(row=r, column=4, value="").border = BORDER
ws.cell(row=r, column=5, value="").border = BORDER
ws.row_dimensions[r].height = 28
r += 1

ws.cell(row=r, column=1, value="AI适配性低的占比").font = F_BODY_B
ws.cell(row=r, column=1).fill = FILL_SUB
ws.cell(row=r, column=1).border = BORDER
ws.cell(row=r, column=2, value='=COUNTIF(F6:F15,"低")/COUNTA(B6:B15)').font = F_NUM
ws.cell(row=r, column=2).number_format = "0.0%"
ws.cell(row=r, column=2).border = BORDER
ws.cell(row=r, column=3, value="").border = BORDER
ws.cell(row=r, column=4, value="越高越需改造（>50% 紧急）").font = F_HINT
ws.cell(row=r, column=4).border = BORDER
ws.cell(row=r, column=4).alignment = LEFT
ws.cell(row=r, column=5, value="").border = BORDER
ws.row_dimensions[r].height = 28
r += 1

# 综合判断
add_section(ws, r, 5, "综合判断")
r += 1
add_input_row(ws, r, 5, "主要发现", "", height=60)
r += 1
add_input_row(ws, r, 5, "下一步行动", "进入F09用四维框架重设", height=42)

set_widths(ws, [6, 24, 14, 18, 16, 18])
ws.sheet_view.showGridLines = False


# ============== F09 四维框架应用表 ==============
ws = wb.create_sheet("F09 四维框架应用表")
style_title_row(ws, 1, 4, "F09 四维框架应用表", height=42)
style_subtitle(ws, 2, 4, "用四维框架（成果/判断/关系/成长）重新设计本部门绩效指标")
add_hint(ws, 3, 4, "💡 每个维度至少2-3个指标。四维要平衡——避免又回到『只看结果』。\nAI时代最容易被忽略的：维度二（判断/创新）和维度四（成长轨迹）。")

add_section(ws, 4, 4, "四维指标设计")
headers = ["维度", "指标名称", "衡量方式", "目标值/周期"]
style_header_row(ws, 5, headers)

dims = [
    ("维度一·成果贡献", "对组织目标的实质影响", "例：续约率/客户健康度/增购率"),
    ("维度二·判断与创新", "AI时代人类核心价值", "例：AI应用案例数/深度解决率"),
    ("维度三·关系影响力", "跨团队/客户价值贡献", "例：跨部门协作评分/标杆培养"),
    ("维度四·成长轨迹", "能力进化速度", "例：AI工具熟练度/学习时长"),
]
r = 6
for d, expl, ex in dims:
    # dim label
    ws.cell(row=r, column=1, value=f"{d}\n\n解释：{expl}\n\n示例：{ex}").font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = WRAP
    ws.cell(row=r, column=1).border = BORDER
    # 3 indicator rows
    for j in range(3):
        for col in range(2, 5):
            c = ws.cell(row=r, column=col, value="")
            c.font = F_PLACE
            c.alignment = WRAP
            c.border = BORDER
        ws.row_dimensions[r].height = 40
        if j < 2:
            r += 1
    r += 1

# 公式统计
add_section(ws, r, 4, "平衡度统计（自动计算）")
r += 1
ws.cell(row=r, column=1, value="总指标数").font = F_BODY_B
ws.cell(row=r, column=1).fill = FILL_SUB
ws.cell(row=r, column=1).border = BORDER
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.cell(row=r, column=2, value="=COUNTA(B6:B17)").font = F_NUM
ws.cell(row=r, column=2).border = BORDER
ws.cell(row=r, column=4, value="建议 8-12 个").font = F_HINT
ws.cell(row=r, column=4).border = BORDER
ws.row_dimensions[r].height = 28
r += 1

ws.cell(row=r, column=1, value="四维覆盖度").font = F_BODY_B
ws.cell(row=r, column=1).fill = FILL_SUB
ws.cell(row=r, column=1).border = BORDER
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.cell(row=r, column=2, value='=IF(AND(COUNTA(B6:B8)>0,COUNTA(B9:B11)>0,COUNTA(B12:B14)>0,COUNTA(B15:B17)>0),"四维完整","有缺失")').font = F_NUM
ws.cell(row=r, column=2).border = BORDER
ws.cell(row=r, column=4, value="每维至少1个").font = F_HINT
ws.cell(row=r, column=4).border = BORDER
ws.row_dimensions[r].height = 28
r += 1

# 综合
add_section(ws, r, 4, "综合判断")
r += 1
add_input_row(ws, r, 4, "四维平衡度评估", "1=严重失衡 / 5=完美平衡", height=36)
r += 1
add_input_row(ws, r, 4, "下一步：进入F10牵引力测试", "", height=42)

set_widths(ws, [24, 26, 26, 18])
ws.sheet_view.showGridLines = False


# ============== F10 牵引力测试表 ==============
ws = wb.create_sheet("F10 牵引力测试表")
style_title_row(ws, 1, 5, "F10 牵引力测试表", height=42)
style_subtitle(ws, 2, 5, "检验绩效指标是否真的能牵引正确行为——防止『指标对但行为错』")
add_hint(ws, 3, 5, "💡 每个新指标问3个问题：\n①我想牵引什么行为？\n②团队实际可能做什么？\n③匹配度如何？\n健康度=绿/黄/红，红灯需重设。")

add_section(ws, 4, 5, "逐个指标测试")
headers = ["#", "被测指标", "想牵引的行为", "实际牵引的可能行为", "行为匹配度(1-5)"]
style_header_row(ws, 5, headers)

r = 6
for i in range(1, 9):
    ws.cell(row=r, column=1, value=i).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).border = BORDER
    for col in range(2, 6):
        c = ws.cell(row=r, column=col, value="")
        c.font = F_PLACE
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[r].height = 40
    r += 1

# 添加测试结论
add_section(ws, r, 5, "健康度统计（自动计算）")
r += 1
ws.cell(row=r, column=1, value="绿灯(≥4)数量").font = F_BODY_B
ws.cell(row=r, column=1).fill = FILL_SUB
ws.cell(row=r, column=1).border = BORDER
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.cell(row=r, column=2, value="=COUNTIF(F6:F13,\">=4\")").font = F_NUM
ws.cell(row=r, column=2).border = BORDER
ws.cell(row=r, column=4, value="占比").font = F_HINT
ws.cell(row=r, column=4).border = BORDER
ws.cell(row=r, column=5, value="=IF(COUNTA(B6:B13)>0,COUNTIF(F6:F13,\">=4\")/COUNTA(B6:B13),0)").font = F_NUM
ws.cell(row=r, column=5).number_format = "0%"
ws.cell(row=r, column=5).border = BORDER
ws.row_dimensions[r].height = 28
r += 1

# 综合
add_section(ws, r, 5, "综合结论")
r += 1
add_input_row(ws, r, 5, "需要重设的指标（红灯）", "", height=42)
r += 1
add_input_row(ws, r, 5, "修正方向", "", height=42)
r += 1
add_input_row(ws, r, 5, "下一步：进入F11共识对话设计", "", height=36)

set_widths(ws, [6, 24, 24, 24, 16])
ws.sheet_view.showGridLines = False


# ============== F11 共识对话流程设计表 ==============
ws = wb.create_sheet("F11 共识对话设计表")
style_title_row(ws, 1, 3, "F11 共识对话流程设计表", height=42)
style_subtitle(ws, 2, 3, "设计本部门目标共识建立的对话流程——取代单向宣贯")
add_hint(ws, 3, 3, "💡 共识三条件：理解（能复述）/ 认同（价值观认可）/ 承诺（自己说出来）\n关键：每个人用自己的话说一遍，而不是管理者说完就结束。")

add_section(ws, 4, 3, "对话设计")
fields11 = [
    ("共识会议目的", "一句话"),
    ("参与人", "谁参加"),
    ("会议时长（分钟）", "建议60-120分钟"),
    ("议程1·开场（10-15%）", "前段做什么：邀请参与而非宣讲"),
    ("议程2·理解（20-25%）", "每人用自己的话说一遍目标"),
    ("议程3·认同（20-25%）", "识别分歧，逐一讨论『我以为…其实是…』"),
    ("议程4·承诺（30-40%）", "每人写自己的1-2条贡献，公开宣读"),
    ("议程5·收尾（10%）", "互相确认+留念"),
]
r = 5
for label, ph in fields11:
    add_input_row(ws, r, 3, label, ph, height=42)
    r += 1

add_section(ws, r, 3, "共识三条件检查")
r += 1
headers = ["条件", "如何验证", "未通过的常见表现"]
style_header_row(ws, r, headers)
r += 1
for cond, check, fail in [
    ("理解（能用自己的话说）", "会后1对1让其复述", "讲不出/讲得不一样"),
    ("认同（价值观层面认可）", "观察行为倾向", "嘴上同意但行为抵触"),
    ("承诺（主动承接）", "后续行动+主动提议", "没有主动行动"),
]:
    ws.cell(row=r, column=1, value=cond).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = LEFT
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value=check).font = F_HINT
    ws.cell(row=r, column=2).alignment = LEFT
    ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=3, value=fail).font = F_HINT
    ws.cell(row=r, column=3).alignment = LEFT
    ws.cell(row=r, column=3).border = BORDER
    ws.row_dimensions[r].height = 32
    r += 1

add_section(ws, r, 3, "异议预案")
r += 1
for i, q in enumerate(["常见异议1及应对", "常见异议2及应对", "常见异议3及应对"]):
    add_input_row(ws, r, 3, q, "例如：『AI不关我事』→让其算自己客户的AI使用率", height=50)
    r += 1

set_widths(ws, [22, 35, 35])
ws.sheet_view.showGridLines = False


# ============== F12 共识度评估表 ==============
ws = wb.create_sheet("F12 共识度评估表")
style_title_row(ws, 1, 7, "F12 共识度评估表", height=42)
style_subtitle(ws, 2, 7, "评估团队成员对目标的真正共识度（不是口头答应）")
add_hint(ws, 3, 7, "💡 评分：1=没共识 / 2=表面答应 / 3=基本理解 / 4=真正认同 / 5=主动承诺\n测试方法：让成员讲出来+观察行为倾向。\n底部公式自动算团队平均。")

add_section(ws, 4, 7, "团队共识度评估（每人一行）")
headers = ["#", "成员姓名", "理解(1-5)", "认同(1-5)", "承诺(1-5)", "未共识点", "跟进动作"]
style_header_row(ws, 5, headers)

r = 6
for i in range(1, 14):  # 留12人+2余量
    ws.cell(row=r, column=1, value=i).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).border = BORDER
    for col in range(2, 8):
        c = ws.cell(row=r, column=col, value="")
        c.font = F_PLACE
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[r].height = 36
    r += 1

# 公式统计
add_section(ws, r, 7, "团队共识度统计（自动）")
r += 1
headers_stat = ["指标", "平均分", "判断标准", "改进重点", "下次复评时间", "", ""]
style_header_row(ws, r, headers_stat)
r += 1
for dim_letter, name, std in [
    ("C", "理解度", "≥4 为合格"),
    ("D", "认同度", "≥3.5 为合格"),
    ("E", "承诺度", "≥3.5 为合格"),
]:
    ws.cell(row=r, column=1, value=name).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value=f'=IFERROR(AVERAGE({dim_letter}6:{dim_letter}17),0)').font = F_NUM
    ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=2).number_format = "0.0"
    ws.cell(row=r, column=3, value=std).font = F_HINT
    ws.cell(row=r, column=3).border = BORDER
    for col in range(4, 8):
        ws.cell(row=r, column=col, value="").border = BORDER
    ws.row_dimensions[r].height = 28
    r += 1

set_widths(ws, [5, 14, 11, 11, 11, 24, 24])
ws.sheet_view.showGridLines = False


# ============== F13 执行追踪机制设计表 ==============
ws = wb.create_sheet("F13 追踪机制设计表")
style_title_row(ws, 1, 3, "F13 执行追踪机制设计表", height=42)
style_subtitle(ws, 2, 3, "设计本部门执行追踪机制——节奏/节点/信号三要素")
add_hint(ws, 3, 3, "💡 三要素：①节奏（多久追踪一次）②节点（关键检查点）③信号（什么信息触发调整）\nAI时代：追踪频率要提高（季度→月度/周度），但不要变控制。")

add_section(ws, 4, 3, "追踪机制设计")
fields13 = [
    ("追踪频率", "日/周/双周/月（建议双层）"),
    ("追踪形式", "看板/会议/书面报告"),
    ("追踪内容（3-5个核心指标）", "聚焦关键指标，不要全部追"),
    ("节点1·日站会（每天）", "15min·3个问题：昨日/今日/卡点"),
    ("节点2·周会（每周）", "60min·数据回顾+风险识别+下周重点"),
    ("节点3·月度复盘（每月）", "2h·完整四层映射回顾+调整决策"),
    ("信号1·黄灯（什么情况出黄）", "例如：健康度周环比下降>10%"),
    ("信号2·红灯（什么情况出红）", "例如：AI覆盖率月度进展<5%"),
    ("调整触发机制", "黄/红灯后24小时内做什么"),
    ("数据辅助判断", "AI给数据+人做判断·看趋势而非单点"),
]
r = 5
for label, ph in fields13:
    add_input_row(ws, r, 3, label, ph, height=50)
    r += 1

set_widths(ws, [22, 35, 35])
ws.sheet_view.showGridLines = False


# ============== F14 执行信号监测表 ==============
ws = wb.create_sheet("F14 信号监测表")
style_title_row(ws, 1, 5, "F14 执行信号监测表（每周填写）", height=42)
style_subtitle(ws, 2, 5, "监测本部门执行过程中的关键信号——数据+行为双重监测")
add_hint(ws, 3, 5, "💡 每周五/周一花10分钟填写。\n格式：本周数据 vs 上周数据 → 异常信号 → 下周第一动作。")

add_section(ws, 4, 5, "周度信号监测")
headers = ["#", "周次", "数据信号1", "数据信号2", "数据信号3"]
style_header_row(ws, 5, headers)

r = 6
for i in range(1, 14):  # 留12周
    ws.cell(row=r, column=1, value=i).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).border = BORDER
    for col in range(2, 6):
        c = ws.cell(row=r, column=col, value="")
        c.font = F_PLACE
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[r].height = 32
    r += 1

add_section(ws, r, 5, "行为信号")
r += 1
for i, q in enumerate([
    "行为信号1（团队精神面貌/工作节奏）",
    "行为信号2（客户反馈/外部声音）",
    "异常信号（有没有反常的事）",
]):
    add_input_row(ws, r, 5, q, "", height=42)
    r += 1

add_section(ws, r, 5, "判断与行动")
r += 1
for q in [
    "本周最大风险（我的判断）",
    "本周最大机会（我的判断）",
    "下周第一动作（聚焦做的一件事）",
]:
    add_input_row(ws, r, 5, q, "", height=42)
    r += 1

set_widths(ws, [5, 10, 22, 22, 22])
ws.sheet_view.showGridLines = False


# ============== F15 成长激活方案设计表 ==============
ws = wb.create_sheet("F15 成长激活方案")
style_title_row(ws, 1, 3, "F15 成长激活方案设计表", height=42)
style_subtitle(ws, 2, 3, "设计本团队成长激活方案——意义/成长/贡献三维激励")
add_hint(ws, 3, 3, "💡 AI时代员工动力结构变了：意义感 > 成长感 > 贡献感 > 金钱\n每个维度至少2个具体动作。\n旧逻辑：完成目标→奖励\n新逻辑：明确贡献方向→创造支持条件→动态校准→推动成长")

add_section(ws, 4, 3, "三维激励设计")
fields15 = [
    ("激励维度一·意义感（如何让团队感到工作有意义）", "至少2个动作：例·客户成功故事分享/案例署名"),
    ("激励维度二·成长感（如何让团队感到能力在涨）", "至少2个动作：例·AI工具认证路径/轮岗机会"),
    ("激励维度三·贡献感（如何让团队感到被认可）", "至少2个动作：例·季度明星/团队庆功"),
]
r = 5
for label, ph in fields15:
    add_input_row(ws, r, 3, label, ph, height=60)
    r += 1

add_section(ws, r, 3, "绩效反馈4模式")
r += 1
headers = ["模式", "频率", "形式", "何时用"]
style_header_row(ws, r, headers)
r += 1
for mode, when_use in [
    ("表扬（认可具体行为）", "日常/周"),
    ("指导（解决具体问题）", "事件触发/月"),
    ("教练（挖掘潜能）", "季度"),
    ("激发（意义/价值对话）", "重要时刻"),
]:
    ws.cell(row=r, column=1, value=mode).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=1).alignment = LEFT
    for col in range(2, 5):
        ws.cell(row=r, column=col, value=when_use if col == 4 else "").border = BORDER
        ws.cell(row=r, column=col).font = F_PLACE
        ws.cell(row=r, column=col).alignment = LEFT
    ws.row_dimensions[r].height = 32
    r += 1

add_section(ws, r, 3, "学习资源与能力路径")
r += 1
for q in [
    "学习资源（提供什么）",
    "能力成长路径（可视化）",
    "激励预算（占人力成本%比）",
    "效果评估（如何知道起作用）",
]:
    add_input_row(ws, r, 3, q, "", height=42)
    r += 1

set_widths(ws, [22, 35, 35])
ws.sheet_view.showGridLines = False


# ============== F16 绩效反馈规划表 ==============
ws = wb.create_sheet("F16 绩效反馈规划表")
style_title_row(ws, 1, 4, "F16 绩效反馈规划表", height=42)
style_subtitle(ws, 2, 4, "规划本团队绩效反馈的频率/形式/参与方")
add_hint(ws, 3, 4, "💡 4种反馈模式各有适用场景：表扬/指导/教练/激发\n每种模式明确：频率/形式/参与方。")

add_section(ws, 4, 4, "反馈模式规划")
headers = ["模式", "频率", "形式", "参与方"]
style_header_row(ws, 5, headers)

modes = ["表扬", "指导", "教练", "激发"]
r = 6
for m in modes:
    ws.cell(row=r, column=1, value=m).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).border = BORDER
    for col in range(2, 5):
        c = ws.cell(row=r, column=col, value="")
        c.font = F_PLACE
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[r].height = 36
    r += 1

add_section(ws, r, 4, "反馈体系")
r += 1
for q in [
    "反馈记录方式（留痕方法）",
    "反馈文化承诺（团队对反馈的态度）",
    "最想改善的反馈模式",
    "下一步行动",
]:
    add_input_row(ws, r, 4, q, "", height=42)
    r += 1

set_widths(ws, [16, 24, 24, 24])
ws.sheet_view.showGridLines = False


# ============== F17 部门绩效引擎完整设计图 ==============
ws = wb.create_sheet("F17 完整设计图")
style_title_row(ws, 1, 3, "F17 部门绩效引擎完整设计图（收官）", height=44)
style_subtitle(ws, 2, 3, "整合六章所有产出——形成完整的部门绩效引擎设计图")
add_hint(ws, 3, 3, "💡 本表是全课最终产出：把F05-F16的所有关键结论汇总。\n每条都指向之前已填的具体表单，确保可追溯。")

add_section(ws, 4, 3, "四驱模型整合")
fields17 = [
    ("四驱·穿透力（整合F05/F06/F07）", "链接到具体表单内容"),
    ("四驱·牵引力（整合F08/F09/F10）", "链接到具体表单内容"),
    ("四驱·推进力（整合F13/F14）", "链接到具体表单内容"),
    ("四驱·进化力（整合F15/F16）", "链接到具体表单内容"),
]
r = 5
for label, ph in fields17:
    add_input_row(ws, r, 3, label, ph, height=60)
    r += 1

add_section(ws, r, 3, "关键产出汇总")
r += 1
for q in [
    "完整目标穿透图（F05链接）",
    "共识建立方案（F11链接）",
    "执行追踪机制（F13链接）",
    "成长激活方案（F15链接）",
    "30天行动承诺（F18链接）",
]:
    add_input_row(ws, r, 3, q, "", height=42)
    r += 1

add_section(ws, r, 3, "版本管理")
r += 1
for q in [
    "设计图版本号（V1.0）",
    "下次更新时间（季度/半年）",
    "下次更新的触发条件",
]:
    add_input_row(ws, r, 3, q, "", height=36)
    r += 1

add_section(ws, r, 3, "整体自评")
r += 1
for q, std in [
    ("完整性自评(1-5)", "5=四驱全到位"),
    ("可执行性自评(1-5)", "5=下周就能开始"),
    ("团队认同度自评(1-5)", "5=团队也认可"),
]:
    add_input_row(ws, r, 3, f"{q}  评分标准：{std}", "", height=36)
    r += 1

set_widths(ws, [22, 35, 35])
ws.sheet_view.showGridLines = False


# ============== F18 30天行动承诺表 ==============
ws = wb.create_sheet("F18 30天行动承诺")
style_title_row(ws, 1, 4, "F18 30天行动承诺表", height=44)
style_subtitle(ws, 2, 4, "写出30天内具体可执行的行动承诺——把课程收获转化为真实行为")
add_hint(ws, 3, 4, "💡 关键质量：①具体动作（不是『多努力』）②明确衡量 ③公开承诺（提升完成率）\n30天后回到此表，对照检查。")

add_section(ws, 4, 4, "30天行动承诺（分阶段）")
headers = ["时间窗", "具体行动", "完成衡量", "潜在障碍"]
style_header_row(ws, 5, headers)

windows = [
    "本周内（第1周）",
    "第2周",
    "第3-4周",
    "个人成长（30天内）",
]
r = 6
for w in windows:
    ws.cell(row=r, column=1, value=w).font = F_BODY_B
    ws.cell(row=r, column=1).fill = FILL_SUB
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).border = BORDER
    for col in range(2, 5):
        c = ws.cell(row=r, column=col, value="")
        c.font = F_PLACE
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[r].height = 50
    r += 1

add_section(ws, r, 4, "障碍预案")
r += 1
for q in [
    "可能遇到的障碍（≥2条）",
    "障碍1的应对策略",
    "障碍2的应对策略",
]:
    add_input_row(ws, r, 4, q, "", height=42)
    r += 1

add_section(ws, r, 4, "承诺与复盘")
r += 1
for q in [
    "30天后复盘时间",
    "公开承诺对象（团队/上级/家人）",
    "完成后给自己的小奖励",
    "未完成时的处理方式",
]:
    add_input_row(ws, r, 4, q, "", height=36)
    r += 1

add_section(ws, r, 4, "签名")
r += 1
add_input_row(ws, r, 4, "承诺人签名", "签下名字 = 公开承诺", height=42)
r += 1
add_input_row(ws, r, 4, "日期", "", height=36)

set_widths(ws, [16, 28, 22, 18])
ws.sheet_view.showGridLines = False

# ============== Save ==============
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheets: {wb.sheetnames}")

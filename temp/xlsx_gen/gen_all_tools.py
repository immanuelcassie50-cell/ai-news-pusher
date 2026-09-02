# -*- coding: utf-8 -*-
"""
《打造组织创新力：营造创新土壤》课程包
批量生成剩余 8 个 xlsx 工具文件
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = r'D:\2026年课程\竞越\打造组织创新力：营造创新土壤\完整课程包\08_全流程工具表单'
DEMO_DIR = r'D:\2026年课程\竞越\打造组织创新力：营造创新土壤\完整课程包\15_成果Demo全套'
os.makedirs(OUT_DIR, exist_ok=True)

# ===== 通用样式 =====
HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
TITLE_FONT = Font(name='微软雅黑', size=16, bold=True, color='2F5496')
SUBTITLE_FONT = Font(name='微软雅黑', size=12, bold=True, color='1F3864')
NOTE_FONT = Font(name='微软雅黑', size=10, italic=True, color='1F3864')
TEXT_FONT = Font(name='微软雅黑', size=11)
INPUT_FILL = PatternFill('solid', fgColor='FFF7CC')
NOTE_FILL = PatternFill('solid', fgColor='E7E6E6')
BAND_FILL = PatternFill('solid', fgColor='D9E2F3')
GOOD_FILL = PatternFill('solid', fgColor='C6EFCE')
WARN_FILL = PatternFill('solid', fgColor='FFEB9C')
BAD_FILL = PatternFill('solid', fgColor='FFC7CE')
ACCENT_FILL = PatternFill('solid', fgColor='F4B183')
THIN = Side(border_style='thin', color='808080')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


def write_notes(ws, start_row, col, notes, highlight_indices=None):
    highlight_indices = highlight_indices or []
    for i, t in enumerate(notes):
        c = ws.cell(row=start_row + i, column=col, value=t)
        c.font = TEXT_FONT
        c.alignment = LEFT
        if i in highlight_indices:
            c.fill = NOTE_FILL


# ====================================================================
# 工具_04_情境演练卡_多行业.xlsx
# ====================================================================
def build_tool_04():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '使用说明'
    set_widths(ws1, [4, 110])
    ws1['B2'] = '工具_04_情境演练卡_多行业'
    ws1['B2'].font = TITLE_FONT
    ws1['B2'].alignment = LEFT
    notes = [
        '【核心命题】回应公式 = 肯定角度 + 具体疑问 + 下一步动作 + 时间节点。',
        '【用途】7 个行业 × 3 个典型情境 = 21 张演练卡，配 A/B 组对比演练。',
        '【使用时机】',
        '· 课程模块三后半段：分行业给学员发对应卡，分组演练。',
        '· 1 对 1 辅导：选学员行业最近发生的 1 件事，用卡片框架回应。',
        '· 课后 30 天：对自己团队成员用同款结构回应建议。',
        '【演练结构】',
        '· A 组：起点状态组（领导者处于 IL-0 / IL-1，未掌握回应公式）。',
        '· B 组：IL-2 状态组（掌握回应公式四要素）。',
        '· 观察员：对比两组的差距，识别"哪句话让对方愿意继续说下去"。',
        '【评分维度】共 4 个，每个 0-5 分，总分 20 分。',
        '· 肯定角度：是否具体说到对方想法的可取之处？',
        '· 具体疑问：是否提出一个让对方想再说的追问？',
        '· 下一步动作：是否明确了下一步要做什么？',
        '· 时间节点：是否给了明确的时间点？',
        '【配套资源】工具_03_回应公式速查卡、案例库_新视窗服务公司。',
    ]
    write_notes(ws1, 4, 2, notes, highlight_indices=[0, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14])
    ws1.row_dimensions[2].height = 28

    industries = [
        ('互联网科技', [
            ('工程师提"用 Rust 重写订单服务"', '性能提升 30%，但工期 6 个月',
             'A组：风险太大，暂缓。',
             'B组：你看到了性能瓶颈这个真实痛点（肯定）。我好奇，你评估过渐进式迁移的可能性吗（疑问）？能否先用一个月做一个核心模块的 PoC（动作）？本月 20 号前给我结果（时间）。'),
            ('产品经理提"砍掉老用户群的怀旧版"', '新版本上线后部分老用户流失',
             'A组：现在不是时候。',
             'B组：你从数据中看到了怀旧版的用户价值（肯定）。但我想问，砍掉后这部分用户有何承接路径（疑问）？请做一份 3 个月的灰度切流方案（动作），下周二评审（时间）。'),
            ('设计师提"做一套全新视觉系统"', '需要停摆当前所有迭代 2 个月',
             'A组：现在排不开期。',
             'B组：你对当前视觉有更系统的思考（肯定）。能否先输出一份"哪些组件先动、哪些不动"的版本（疑问）？本月做一份评估清单（动作），月底对齐（时间）。'),
        ]),
        ('制造业', [
            ('车间班组长提"改一次装夹顺序节省 8 秒"', '担心其他班组跟不上',
             'A组：你先试一下。',
             'B组：你算得很细，8 秒乘以节拍就是大数字（肯定）。其他班组的接受度你怎么评估（疑问）？先做一份 SOP 草案（动作），本月内我陪你去两个班组验证（时间）。'),
            ('工艺工程师提"换一种新涂层处理"', '需停机 3 天做小试',
             'A组：订单紧，不行。',
             'B组：你对质量提升有自己的判断（肯定）。3 天停机对哪几条线影响最大（疑问）？请排一份损失对比（动作），本月内与生产总监对齐（时间）。'),
            ('质量员提"用 SPC 系统替代纸质记录"', '需要 IT 部门配合',
             'A组：我和 IT 不熟。',
             'B组：你看到了纸质记录的滞后性（肯定）。IT 那边你的对接人是谁（疑问）？你写一封申请邮件（动作），明天 cc 我（时间）。'),
        ]),
        ('金融证券', [
            ('分析师提"用另类数据做行业研究"', '数据采购预算高',
             'A组：成本太高，砍掉。',
             'B组：你看到了传统数据的盲区（肯定）。能先列三个最有价值的另类数据源吗（疑问）？做一份 ROI 测算（动作），月底上投委会（时间）。'),
            ('风控经理提"重写反欺诈模型规则"', '模型上线后 2 周业务量会波动',
             'A组：风险太大，不要做。',
             'B组：你对老模型的局限看得很清（肯定）。2 周波动的业务影响你能算出来吗（疑问）？先做 A/B 切 10% 流量（动作），下周一上线（时间）。'),
            ('客户经理提"推一个面向高净值客户的私享会"', '一年 4 场，需 50 万预算',
             'A组：上面没批。',
             'B组：你看到了高净值客户的深需求（肯定）。能否先做一场小规模（30 人）试水（疑问）？出一份预算与转化目标（动作），本月内报批（时间）。'),
        ]),
        ('能源电网', [
            ('值班员提"把巡检路径优化一版"', '减少 15% 巡检时间',
             'A组：安全第一，别动。',
             'B组：你从现场经验看到了路径冗余（肯定）。新路径有没有风险点（疑问）？做一份新旧路径的对比图（动作），下周组织评审（时间）。'),
            ('工程师提"用无人机替代部分人工巡检"', '需 200 万采购预算',
             'A组：先放着。',
             'B组：你对人工巡检的痛点看得很清（肯定）。能否先做一次单线路试点（疑问）？出试点报告与全量推广测算（动作），本季度内上总工办公会（时间）。'),
            ('班组长提"晚班加一次安全分享会"', '担心影响交接',
             'A组：班次紧，没空。',
             'B组：你看到了夜班的安全隐患（肯定）。5 分钟够吗（疑问）？与排班员对齐一个 5 分钟时段（动作），下周一启动（时间）。'),
        ]),
        ('医疗医药', [
            ('护士提"修改患者交接表格式"', '当前格式冗余',
             'A组：大家都用老格式。',
             'B组：你看到了交接环节的效率瓶颈（肯定）。能列 3 个最想删的字段吗（疑问）？先做一版精简表（动作），本月内试点两周（时间）。'),
            ('医生提"加一个多学科会诊机制"', '需要 5 个科室协调',
             'A组：协调难度大。',
             'B组：你看到了疑难病例的协作价值（肯定）。5 个科室中最难协调的是哪个（疑问）？先做一份 MDT 流程图与启动会提案（动作），本月上报医务科（时间）。'),
            ('药剂师提"统一院内药品说明书模板"', '涉及 30 多个药品',
             'A组：工作量太大。',
             'B组：你看到了用药安全的标准化需求（肯定）。能否先选 3 个高警示药品做模板（疑问）？出 3 个模板样例（动作），月底院内讨论（时间）。'),
        ]),
        ('教育培训', [
            ('教师提"翻转课堂试点"', '担心家长不接受',
             'A组：家长会投诉。',
             'B组：你对教学效果有自己的追求（肯定）。家长沟通方案你想过吗（疑问）？出一份家长告知与作业设计（动作），下周一启动试点（时间）。'),
            ('教务主任提"重新排课逻辑"', '涉及 200 个班级',
             'A组：动了会乱。',
             'B组：你看到了当前排课的瓶颈（肯定）。能否先排 2 个年级做小范围测试（疑问）？出小范围排课表与风险预案（动作），本月内试排（时间）。'),
            ('教研员提"做一份新的课程地图"', '原有地图用了 5 年',
             'A组：等集团统一下发。',
             'B组：你看到了课程结构升级的窗口期（肯定）。能否先聚焦一个学科做新版地图（疑问）？出一份样章（动作），本月底内部分享（时间）。'),
        ]),
        ('销售客户成功', [
            ('销售提"砍掉利润率最低的 10% 客户"', '担心短期业绩',
             'A组：业绩压力大。',
             'B组：你从客户结构看到了毛利空间（肯定）。砍完后释放的服务资源投向哪里（疑问）？做一份新老客户利润对比与重分配方案（动作），本月内 review（时间）。'),
            ('客户成功经理提"做一份客户健康度评分卡"', '需要 1 个月搭建',
             'A组：先把续约率抓上去。',
             'B组：你看到了客户流失的早期信号（肯定）。能否先聚焦 1 个最容易流失的客户分群（疑问）？出一份 10 题评分卡（动作），下月上线（时间）。'),
            ('销售提"开放电话销售之外的私域直播"', '需要内容团队配合',
             'A组：内容能力不够。',
             'B组：你看到了客户接触点的扩展空间（肯定）。能否先做 1 场 30 分钟的小规模直播（疑问）？出方案与目标观众（动作），本月底上线（时间）。'),
        ]),
    ]

    for ind_name, scenarios in industries:
        ws = wb.create_sheet(ind_name)
        set_widths(ws, [4, 30, 28, 35, 60, 4])
        ws.merge_cells('A1:E1')
        ws['A1'] = f'{ind_name} · 情境演练卡'
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = CENTER
        ws.row_dimensions[1].height = 26
        write_header(ws, 3, ['编号', '情境（谁提了什么）', '顾虑/风险', 'A 组回应（起点）', 'B 组回应（IL-2 公式）'])
        for i, (scene, risk, a, b) in enumerate(scenarios):
            r = 4 + i
            ws.cell(row=r, column=1, value=f'{ind_name[:2]}-{i+1}').alignment = CENTER
            ws.cell(row=r, column=2, value=scene).alignment = LEFT
            ws.cell(row=r, column=3, value=risk).alignment = LEFT
            c4 = ws.cell(row=r, column=4, value=a)
            c4.alignment = LEFT
            c4.fill = BAD_FILL
            c5 = ws.cell(row=r, column=5, value=b)
            c5.alignment = LEFT
            c5.fill = GOOD_FILL
            for col in range(1, 6):
                ws.cell(row=r, column=col).border = BORDER
            ws.row_dimensions[r].height = 80
        # 评分表
        sc_r = 4 + len(scenarios) + 1
        ws.merge_cells(start_row=sc_r, start_column=1, end_row=sc_r, end_column=5)
        ws.cell(row=sc_r, column=1, value='【演练评分表】').font = SUBTITLE_FONT
        ws.cell(row=sc_r, column=1).fill = BAND_FILL
        sc_r += 1
        write_header(ws, sc_r, ['维度', '评分说明', 'A 组得分(0-5)', 'B 组得分(0-5)', '差距分析'])
        dims = [
            ('肯定角度', '是否具体说到对方想法的可取之处？'),
            ('具体疑问', '是否提出一个让对方想再说的追问？'),
            ('下一步动作', '是否明确了下一步要做什么？'),
            ('时间节点', '是否给了明确的时间点？'),
        ]
        for j, (d, q) in enumerate(dims):
            r = sc_r + 1 + j
            ws.cell(row=r, column=1, value=d).font = SUBTITLE_FONT
            ws.cell(row=r, column=2, value=q).alignment = LEFT
            ws.cell(row=r, column=3, value='').fill = INPUT_FILL
            ws.cell(row=r, column=4, value='').fill = INPUT_FILL
            ws.cell(row=r, column=5, value='').fill = INPUT_FILL
            for col in range(1, 6):
                ws.cell(row=r, column=col).border = BORDER
            ws.row_dimensions[r].height = 30
        # 总分行
        r = sc_r + 1 + len(dims)
        ws.cell(row=r, column=1, value='总分').font = SUBTITLE_FONT
        ws.cell(row=r, column=2, value='满分 20').alignment = LEFT
        ws.cell(row=r, column=3, value=f'=SUM(C{sc_r+1}:C{sc_r+len(dims)})').fill = BAND_FILL
        ws.cell(row=r, column=4, value=f'=SUM(D{sc_r+1}:D{sc_r+len(dims)})').fill = BAND_FILL
        ws.cell(row=r, column=5, value=f'=D{r}-C{r}').fill = BAND_FILL
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDER

    wb.save(os.path.join(OUT_DIR, '工具_04_情境演练卡_多行业.xlsx'))
    print('OK: 工具_04_情境演练卡_多行业.xlsx')


# ====================================================================
# 工具_05_标杆企业对照表.xlsx
# ====================================================================
def build_tool_05():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '使用说明'
    set_widths(ws1, [4, 110])
    ws1['B2'] = '工具_05_标杆企业对照表'
    ws1['B2'].font = TITLE_FONT
    ws1['B2'].alignment = LEFT
    notes = [
        '【核心命题】对照灯塔家电 9 大动作，自我评估当前状态。',
        '【用途】对每个动作做 0-3 分自评（0=没有，1=萌芽，2=机制化，3=日常化）。',
        '【使用时机】',
        '· 课程模块四小组研讨：每个学员对自己的组织打 9 个分。',
        '· 课后 30 天：重新打分，对比变化。',
        '· 90 天返校日：再打一次，看 90 天动作完成度。',
        '【评分规则】',
        '· 0 分：完全没有。',
        '· 1 分：偶尔做，靠个人。',
        '· 2 分：开始机制化（如固定时间、明确流程）。',
        '· 3 分：完全日常化（无需提醒，自然发生）。',
        '· 总分 27 分，21+ 优秀，14-20 中等，<14 薄弱。',
        '【配套资源】案例库_灯塔家电集团、工具_06_能力点现状打分表。',
    ]
    write_notes(ws1, 4, 2, notes, highlight_indices=[0, 4, 5, 6, 8, 9, 10, 11, 12, 14])
    ws1.row_dimensions[2].height = 28

    ws2 = wb.create_sheet('9 大动作对照')
    set_widths(ws2, [6, 14, 36, 50, 8, 8, 8, 8, 8, 4])
    ws2.merge_cells('A1:I1')
    ws2['A1'] = '灯塔集团 9 大动作 × 我的组织对照'
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 26
    write_header(ws2, 3, ['编号', 'IL 阶位', '动作', '灯塔具体做法', '0', '1', '2', '3', '我的评分'])
    actions = [
        ('IL-1', '主动开辟点子来源', '每月固定 1 次"用户声音日"，研发 / 产品 / 售后 30 人轮值上门听用户吐槽。'),
        ('IL-1', '降低建议门槛', '"5 分钟建议书"模板——任何人 5 分钟内写完一页就提交，无格式要求。'),
        ('IL-1', '明确回应机制', '48 小时内由专门小组回复"采纳 / 暂缓 / 转交 / 解释"，看板对全员可见。'),
        ('IL-2', '管理过程而非点子', '所有进入"备选池"的点子，必须走完五阶段台账，否则不算"在创新"。'),
        ('IL-2', '允许暂缓与重提', '暂缓的点子写明"再启条件"（如"等 Q3 数据出来"），到时自动提醒。'),
        ('IL-2', '跨部门协调机制', '"产品 - 研发 - 市场"三部门周会对齐创新项目卡点，总裁办兜底。'),
        ('IL-3', '方法显性化', '把"快速原型验证"做成"四步工具卡"，新员工入职当天发。'),
        ('IL-3', '决策边界明确', '一万元以下自主决策，一万元以上 24 小时会签，十万以上月度复盘。'),
        ('IL-3', '共同语言建立', '每月一次的"创新 5 分钟"全员工播报——用五阶段框架讲项目进展。'),
    ]
    for i, (il, action, lt) in enumerate(actions):
        r = 4 + i
        ws2.cell(row=r, column=1, value=f'L{4+i}').alignment = CENTER
        ws2.cell(row=r, column=2, value=il).font = SUBTITLE_FONT
        ws2.cell(row=r, column=2).alignment = CENTER
        ws2.cell(row=r, column=3, value=action).alignment = LEFT
        ws2.cell(row=r, column=4, value=lt).alignment = LEFT
        for col, val in [(5, '☐'), (6, '☐'), (7, '☐'), (8, '☐')]:
            c = ws2.cell(row=r, column=col, value=val)
            c.alignment = CENTER
            c.border = BORDER
        c = ws2.cell(row=r, column=9, value='')
        c.alignment = CENTER
        c.fill = INPUT_FILL
        c.border = BORDER
        for col in range(1, 10):
            ws2.cell(row=r, column=col).border = BORDER
        ws2.row_dimensions[r].height = 40
    # 总分
    r = 4 + len(actions) + 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws2.cell(row=r, column=1, value='总分（满分 27）').font = SUBTITLE_FONT
    ws2.cell(row=r, column=1).alignment = CENTER
    ws2.cell(row=r, column=1).fill = BAND_FILL
    ws2.cell(row=r, column=9, value=f'=SUM(I4:I{4+len(actions)-1})').fill = BAND_FILL
    ws2.cell(row=r, column=9).font = SUBTITLE_FONT
    for col in range(1, 10):
        ws2.cell(row=r, column=col).border = BORDER
    r += 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws2.cell(row=r, column=1, value='【判定】21-27 优秀（IL-5/6 阶段）；14-20 中等（IL-3/4 阶段）；<14 薄弱（IL-1/2 阶段）。')
    ws2.cell(row=r, column=1).font = NOTE_FONT
    ws2.cell(row=r, column=1).fill = BAND_FILL

    ws3 = wb.create_sheet('改进计划')
    set_widths(ws3, [6, 14, 36, 50, 14, 4])
    ws3.merge_cells('A1:E1')
    ws3['A1'] = '针对低分动作（0-1 分）的改进计划'
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].alignment = CENTER
    ws3.row_dimensions[1].height = 26
    write_header(ws3, 3, ['编号', 'IL 阶位', '动作', '我的现状描述', '30 天改进承诺'])
    for i, (il, action, lt) in enumerate(actions):
        r = 4 + i
        ws3.cell(row=r, column=1, value=f'L{4+i}').alignment = CENTER
        ws3.cell(row=r, column=2, value=il).font = SUBTITLE_FONT
        ws3.cell(row=r, column=2).alignment = CENTER
        ws3.cell(row=r, column=3, value=action).alignment = LEFT
        ws3.cell(row=r, column=4, value='').fill = INPUT_FILL
        ws3.cell(row=r, column=5, value='').fill = INPUT_FILL
        for col in range(1, 6):
            ws3.cell(row=r, column=col).border = BORDER
        ws3.row_dimensions[r].height = 40

    wb.save(os.path.join(OUT_DIR, '工具_05_标杆企业对照表.xlsx'))
    print('OK: 工具_05_标杆企业对照表.xlsx')


# ====================================================================
# 工具_06_能力点现状打分表.xlsx
# ====================================================================
def build_tool_06():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '使用说明'
    set_widths(ws1, [4, 110])
    ws1['B2'] = '工具_06_能力点现状打分表'
    ws1['B2'].font = TITLE_FONT
    ws1['B2'].alignment = LEFT
    notes = [
        '【核心命题】15 个能力点 × 4 个评估维度。',
        '【用途】能力点 × 维度交叉评估，定位能力短板。',
        '【使用时机】',
        '· 课程模块四小组研讨：每个学员对自己组织打分。',
        '· 课后 60 天：再次自评对比变化。',
        '· 学员上级用同一模板给学员打分（双视角对照）。',
        '【15 个能力点】',
        '· IL-1：主动开辟点子来源 / 降低建议门槛 / 明确回应机制。',
        '· IL-2：管理过程而非点子 / 允许暂缓与重提 / 跨部门协调。',
        '· IL-3：方法显性化 / 决策边界明确 / 共同语言建立。',
        '· IL-4：固定复盘机制 / 受保护资源 / 组合视角。',
        '· IL-5：日常示范 / 讲述尝试故事 / 保持开放姿态。',
        '【4 个评估维度】',
        '· 知识：是否知道这个能力点的标准做法？',
        '· 技能：自己是否会做？',
        '· 意愿：是否愿意在工作中持续做？',
        '· 频率：过去 30 天做了几次？',
        '【评分规则】0-3 分。',
        '【配套资源】工具_05_标杆企业对照表、案例库_灯塔集团。',
    ]
    write_notes(ws1, 4, 2, notes, highlight_indices=[0, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 18, 20, 22])
    ws1.row_dimensions[2].height = 28

    ws2 = wb.create_sheet('15 能力点 × 4 维度')
    set_widths(ws2, [6, 12, 30, 10, 10, 10, 10, 12, 4])
    ws2.merge_cells('A1:H1')
    ws2['A1'] = '能力点 × 评估维度 · 交叉打分表'
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 26
    write_header(ws2, 3, ['编号', 'IL 阶位', '能力点', '知识(0-3)', '技能(0-3)', '意愿(0-3)', '频率(0-3)', '总分(0-12)'])
    abilities = [
        ('IL-1', '主动开辟点子来源'),
        ('IL-1', '降低建议门槛'),
        ('IL-1', '明确回应机制'),
        ('IL-2', '管理过程而非点子'),
        ('IL-2', '允许暂缓与重提'),
        ('IL-2', '跨部门协调'),
        ('IL-3', '方法显性化'),
        ('IL-3', '决策边界明确'),
        ('IL-3', '共同语言建立'),
        ('IL-4', '固定复盘机制'),
        ('IL-4', '受保护资源'),
        ('IL-4', '组合视角'),
        ('IL-5', '日常示范'),
        ('IL-5', '讲述尝试故事'),
        ('IL-5', '保持开放姿态'),
    ]
    for i, (il, ab) in enumerate(abilities):
        r = 4 + i
        ws2.cell(row=r, column=1, value=f'L{4+i}').alignment = CENTER
        ws2.cell(row=r, column=2, value=il).font = SUBTITLE_FONT
        ws2.cell(row=r, column=2).alignment = CENTER
        ws2.cell(row=r, column=3, value=ab).alignment = LEFT
        for col in range(4, 8):
            c = ws2.cell(row=r, column=col, value='')
            c.alignment = CENTER
            c.fill = INPUT_FILL
        ws2.cell(row=r, column=8, value=f'=SUM(D{r}:G{r})').fill = BAND_FILL
        ws2.cell(row=r, column=8).font = SUBTITLE_FONT
        ws2.cell(row=r, column=8).alignment = CENTER
        for col in range(1, 9):
            ws2.cell(row=r, column=col).border = BORDER
        ws2.row_dimensions[r].height = 28
    # 分组小计
    r = 4 + len(abilities) + 1
    ws2.cell(row=r, column=2, value='IL-1 小计').font = SUBTITLE_FONT
    ws2.cell(row=r, column=3, value='').alignment = LEFT
    ws2.cell(row=r, column=8, value=f'=SUM(H4:H6)').fill = BAND_FILL
    r += 1
    ws2.cell(row=r, column=2, value='IL-2 小计').font = SUBTITLE_FONT
    ws2.cell(row=r, column=8, value=f'=SUM(H7:H9)').fill = BAND_FILL
    r += 1
    ws2.cell(row=r, column=2, value='IL-3 小计').font = SUBTITLE_FONT
    ws2.cell(row=r, column=8, value=f'=SUM(H10:H12)').fill = BAND_FILL
    r += 1
    ws2.cell(row=r, column=2, value='IL-4 小计').font = SUBTITLE_FONT
    ws2.cell(row=r, column=8, value=f'=SUM(H13:H15)').fill = BAND_FILL
    r += 1
    ws2.cell(row=r, column=2, value='IL-5 小计').font = SUBTITLE_FONT
    ws2.cell(row=r, column=8, value=f'=SUM(H16:H18)').fill = BAND_FILL
    r += 1
    ws2.cell(row=r, column=2, value='总分（满分 180）').font = SUBTITLE_FONT
    ws2.cell(row=r, column=8, value=f'=SUM(H4:H18)').fill = ACCENT_FILL
    ws2.cell(row=r, column=8).font = SUBTITLE_FONT
    for rr in range(4 + len(abilities) + 1, r + 1):
        for col in range(1, 9):
            ws2.cell(row=rr, column=col).border = BORDER

    ws3 = wb.create_sheet('短板分析')
    set_widths(ws3, [6, 12, 30, 14, 14, 14, 14, 50, 4])
    ws3.merge_cells('A1:H1')
    ws3['A1'] = '短板分析 · 4 维度对比'
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].alignment = CENTER
    ws3.row_dimensions[1].height = 26
    write_header(ws3, 3, ['编号', 'IL 阶位', '能力点', '知识', '技能', '意愿', '频率', '差距诊断'])
    for i, (il, ab) in enumerate(abilities):
        r = 4 + i
        ws3.cell(row=r, column=1, value=f'L{4+i}').alignment = CENTER
        ws3.cell(row=r, column=2, value=il).font = SUBTITLE_FONT
        ws3.cell(row=r, column=2).alignment = CENTER
        ws3.cell(row=r, column=3, value=ab).alignment = LEFT
        ws3.cell(row=r, column=4, value=f"='15 能力点 × 4 维度'!D{r}").fill = GOOD_FILL
        ws3.cell(row=r, column=5, value=f"='15 能力点 × 4 维度'!E{r}").fill = GOOD_FILL
        ws3.cell(row=r, column=6, value=f"='15 能力点 × 4 维度'!F{r}").fill = GOOD_FILL
        ws3.cell(row=r, column=7, value=f"='15 能力点 × 4 维度'!G{r}").fill = GOOD_FILL
        ws3.cell(row=r, column=8, value='低分维度即短板；建议从低分维度入手 30 天行动').alignment = LEFT
        for col in range(1, 9):
            ws3.cell(row=r, column=col).border = BORDER
        ws3.row_dimensions[r].height = 30

    wb.save(os.path.join(OUT_DIR, '工具_06_能力点现状打分表.xlsx'))
    print('OK: 工具_06_能力点现状打分表.xlsx')


# ====================================================================
# 工具_07_行动计划模板.xlsx
# ====================================================================
def build_tool_07():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '使用说明'
    set_widths(ws1, [4, 110])
    ws1['B2'] = '工具_07_行动计划模板'
    ws1['B2'].font = TITLE_FONT
    ws1['B2'].alignment = LEFT
    notes = [
        '【核心命题】3 个承诺 × 5 个要素 = 完整行动计划。',
        '【用途】学员返岗后用此表规划个人行动。',
        '【3 个承诺】',
        '· 第 1 周承诺：7 天内必做 1 个小动作。',
        '· 第 1 月承诺：30 天内必做 1 个中等动作。',
        '· 第 3 月承诺：90 天内必做 1 个大动作。',
        '【5 个要素】',
        '· 具体动作：要做什么，可观察。',
        '· 时间节点：什么时候开始 / 结束。',
        '· 责任人：谁负责（可分工协作）。',
        '· 资源需求：需要什么支持。',
        '· 验收标准：做到什么程度算完成。',
        '【使用时机】',
        '· 课程收尾 30 分钟：每位学员现场填写 3 个承诺。',
        '· 课后 24 小时内：把承诺书提交给小组长。',
        '· 30/60/90 天跟进：小组长对照承诺书 review 进展。',
        '【配套资源】工具_08_小组互评记录表、工具_09_30天跟进记录表。',
    ]
    write_notes(ws1, 4, 2, notes, highlight_indices=[0, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 16, 17])
    ws1.row_dimensions[2].height = 28

    commitments = [
        ('第 1 周承诺（7 天小动作）', 4),
        ('第 1 月承诺（30 天中动作）', 3),
        ('第 3 月承诺（90 天大动作）', 2),
    ]
    row_cursor = 3
    for cmt_name, action_count in commitments:
        ws = wb.create_sheet(cmt_name[:31])
        set_widths(ws, [6, 22, 35, 14, 14, 30, 4])
        ws.merge_cells('A1:F1')
        ws['A1'] = cmt_name
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = CENTER
        ws.row_dimensions[1].height = 26
        write_header(ws, 3, ['编号', '具体动作', '时间节点', '责任人', '资源需求', '验收标准'])
        for j in range(action_count):
            r = 4 + j
            ws.cell(row=r, column=1, value=j + 1).alignment = CENTER
            for col in range(2, 7):
                c = ws.cell(row=r, column=col, value='')
                c.alignment = LEFT
                c.fill = INPUT_FILL
            for col in range(1, 7):
                ws.cell(row=r, column=col).border = BORDER
            ws.row_dimensions[r].height = 50
        # 签字行
        r = 4 + action_count + 1
        ws.cell(row=r, column=1, value='签字').font = SUBTITLE_FONT
        ws.cell(row=r, column=2, value='本人：____________').alignment = LEFT
        ws.cell(row=r, column=3, value='上级：____________').alignment = LEFT
        ws.cell(row=r, column=4, value='小组长：____________').alignment = LEFT
        ws.cell(row=r, column=5, value='日期：____________').alignment = LEFT
        ws.cell(row=r, column=6, value='承诺：必做').alignment = LEFT
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
        ws.row_dimensions[r].height = 30

    ws = wb.create_sheet('个人承诺汇总')
    set_widths(ws, [6, 30, 50, 4])
    ws.merge_cells('A1:C1')
    ws['A1'] = '个人 3 个承诺总览'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 26
    write_header(ws, 3, ['承诺类型', '一句话总结', '验收标志'])
    summary = [
        ('第 1 周承诺', '在 7 天内完成 ____ 个小动作', '在 30 天评估中能说出做到 / 没做到'),
        ('第 1 月承诺', '在 30 天内完成 ____ 个中动作', '在 60 天评估中有 1 个案例'),
        ('第 3 月承诺', '在 90 天内完成 ____ 个大动作', '在 90 天返校日有 1 份最佳实践'),
    ]
    for i, (c, s, v) in enumerate(summary):
        r = 4 + i
        ws.cell(row=r, column=1, value=c).font = SUBTITLE_FONT
        ws.cell(row=r, column=2, value=s).fill = INPUT_FILL
        ws.cell(row=r, column=3, value=v).alignment = LEFT
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = BORDER
        ws.row_dimensions[r].height = 30

    wb.save(os.path.join(OUT_DIR, '工具_07_行动计划模板.xlsx'))
    print('OK: 工具_07_行动计划模板.xlsx')


# ====================================================================
# 工具_08_小组互评记录表.xlsx
# ====================================================================
def build_tool_08():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '使用说明'
    set_widths(ws1, [4, 110])
    ws1['B2'] = '工具_08_小组互评记录表'
    ws1['B2'].font = TITLE_FONT
    ws1['B2'].alignment = LEFT
    notes = [
        '【核心命题】小组互评 = 同伴反馈 + 共识形成。',
        '【用途】课程结束前 30 分钟，4-6 人小组内每位学员相互打分。',
        '【使用时机】',
        '· 课程收尾 30 分钟：分小组，每位学员被打分。',
        '· 课后 30/60/90 天返岗实践后，再次互评对比变化。',
        '【6 个评估维度】',
        '· 课堂参与度：是否全程投入？',
        '· 回应公式掌握：能否说出公式 4 要素？',
        '· 情境演练表现：A 组 vs B 组差距识别？',
        '· 标杆对照深度：9 大动作是否定位准确？',
        '· 行动计划可执行性：3 个承诺是否具体？',
        '· 团队协作贡献：是否在小组中带动力？',
        '【评分规则】1-5 分，5 分最优。',
        '【输出物】每位学员 1 张互评卡 + 1 份小组共识。',
        '【配套资源】工具_07_行动计划模板、案例库。',
    ]
    write_notes(ws1, 4, 2, notes, highlight_indices=[0, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16, 17])
    ws1.row_dimensions[2].height = 28

    ws2 = wb.create_sheet('个人互评卡')
    set_widths(ws2, [4, 20, 16, 16, 16, 16, 16, 16, 4])
    ws2.merge_cells('A1:H1')
    ws2['A1'] = '个人互评卡 · 每位组员 1 张'
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 26
    ws2.merge_cells('A2:H2')
    ws2['A2'] = '被评人姓名：________________  小组：________  日期：____________'
    ws2['A2'].font = NOTE_FONT
    ws2['A2'].fill = BAND_FILL
    write_header(ws2, 3, ['序号', '评估维度', '自评(1-5)', '组员A评', '组员B评', '组员C评', '组员D评', '组员E评'])
    dims = [
        '课堂参与度',
        '回应公式掌握',
        '情境演练表现',
        '标杆对照深度',
        '行动计划可执行性',
        '团队协作贡献',
    ]
    for i, d in enumerate(dims):
        r = 4 + i
        ws2.cell(row=r, column=1, value=i + 1).alignment = CENTER
        ws2.cell(row=r, column=2, value=d).font = SUBTITLE_FONT
        for col in range(3, 9):
            c = ws2.cell(row=r, column=col, value='')
            c.alignment = CENTER
            c.fill = INPUT_FILL
        for col in range(1, 9):
            ws2.cell(row=r, column=col).border = BORDER
        ws2.row_dimensions[r].height = 28
    # 总分
    r = 4 + len(dims) + 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws2.cell(row=r, column=1, value='平均分（满分 5.0）').font = SUBTITLE_FONT
    ws2.cell(row=r, column=1).fill = BAND_FILL
    ws2.cell(row=r, column=1).alignment = CENTER
    for col in range(3, 9):
        ws2.cell(row=r, column=col, value=f'=AVERAGE({get_column_letter(col)}4:{get_column_letter(col)}{4+len(dims)-1})').fill = BAND_FILL
        ws2.cell(row=r, column=col).font = SUBTITLE_FONT
    for col in range(1, 9):
        ws2.cell(row=r, column=col).border = BORDER
    # 反馈栏
    r += 2
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws2.cell(row=r, column=1, value='【小组反馈】3 个优点 + 1 个改进建议').font = SUBTITLE_FONT
    ws2.cell(row=r, column=1).fill = BAND_FILL
    for j, label in enumerate(['优点 1', '优点 2', '优点 3', '改进建议']):
        rr = r + 1 + j
        ws2.cell(row=rr, column=1, value=label).font = SUBTITLE_FONT
        ws2.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=8)
        c = ws2.cell(row=rr, column=2, value='')
        c.fill = INPUT_FILL
        c.alignment = LEFT
        for col in range(1, 9):
            ws2.cell(row=rr, column=col).border = BORDER
        ws2.row_dimensions[rr].height = 30

    ws3 = wb.create_sheet('小组共识')
    set_widths(ws3, [6, 20, 30, 30, 4])
    ws3.merge_cells('A1:D1')
    ws3['A1'] = '小组共识 · 4-6 人共同制定'
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].alignment = CENTER
    ws3.row_dimensions[1].height = 26
    write_header(ws3, 3, ['编号', '主题', '共识', '责任人'])
    topics = [
        ('小组学习主题', '3 个月内共同攻克 1 个创新挑战', ''),
        ('小组互访机制', '每月 1 次企业互访，看真实行动', ''),
        ('小组分享机制', '每月 1 次线上 1 小时经验分享', ''),
        ('失败案例机制', '每位组员 90 天内交 1 个失败案例', ''),
        ('最佳实践机制', '90 天返校日评选小组"最佳实践奖"', ''),
    ]
    for i, (t, c1, c2) in enumerate(topics):
        r = 4 + i
        ws3.cell(row=r, column=1, value=i + 1).alignment = CENTER
        ws3.cell(row=r, column=2, value=t).font = SUBTITLE_FONT
        ws3.cell(row=r, column=3, value=c1).alignment = LEFT
        ws3.cell(row=r, column=4, value=c2).fill = INPUT_FILL
        for col in range(1, 5):
            ws3.cell(row=r, column=col).border = BORDER
        ws3.row_dimensions[r].height = 32

    wb.save(os.path.join(OUT_DIR, '工具_08_小组互评记录表.xlsx'))
    print('OK: 工具_08_小组互评记录表.xlsx')


# ====================================================================
# 工具_09_30天跟进记录表.xlsx
# ====================================================================
def build_tool_09():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '使用说明'
    set_widths(ws1, [4, 110])
    ws1['B2'] = '工具_09_30天跟进记录表'
    ws1['B2'].font = TITLE_FONT
    ws1['B2'].alignment = LEFT
    notes = [
        '【核心命题】承诺写在纸上，行动在每周。',
        '【用途】30 天内每周 1 行记录，跟踪承诺完成度。',
        '【使用时机】',
        '· 课后 24 小时内：建立自己的"30 天跟进表"。',
        '· 每周五下午 30 分钟：填本周的 5 个字段。',
        '· 30 天评估日：拿这张表 review，识别成功 / 失败模式。',
        '【5 个字段】',
        '· 本周承诺：本周末前要完成什么。',
        '· 完成度：0-100%。',
        '· 关键事件：本周最影响创新土壤的一件事。',
        '· 学到什么：本周一个学习点。',
        '· 下周计划：下周承诺什么。',
        '【30/60/90 天闭环】',
        '· 30 天：看完成度与学习密度。',
        '· 60 天：看土壤三要素是否改善。',
        '· 90 天：看行为是否变成习惯。',
        '【配套资源】工具_07_行动计划模板、评估_04_30-60-90天跟进评估。',
    ]
    write_notes(ws1, 4, 2, notes, highlight_indices=[0, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18])
    ws1.row_dimensions[2].height = 28

    ws2 = wb.create_sheet('30 天周记')
    set_widths(ws2, [6, 14, 35, 12, 35, 35, 35, 4])
    ws2.merge_cells('A1:G1')
    ws2['A1'] = '30 天周记 · 4 周记录表'
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 26
    ws2.merge_cells('A2:G2')
    ws2['A2'] = '学员姓名：________________  小组：________  课程日期：____________'
    ws2['A2'].font = NOTE_FONT
    ws2['A2'].fill = BAND_FILL
    write_header(ws2, 3, ['周次', '日期范围', '本周承诺', '完成度(%)', '关键事件', '学到什么', '下周计划'])
    weeks = ['第 1 周', '第 2 周', '第 3 周', '第 4 周', '第 5 周（30 天评估）']
    for i, w in enumerate(weeks):
        r = 4 + i
        ws2.cell(row=r, column=1, value=w).font = SUBTITLE_FONT
        ws2.cell(row=r, column=1).alignment = CENTER
        for col in range(3, 8):
            c = ws2.cell(row=r, column=col, value='')
            c.fill = INPUT_FILL
            c.alignment = LEFT
        c = ws2.cell(row=r, column=2, value='')
        c.alignment = CENTER
        c.fill = INPUT_FILL
        for col in range(1, 8):
            ws2.cell(row=r, column=col).border = BORDER
        ws2.row_dimensions[r].height = 60
    # 总计
    r = 4 + len(weeks) + 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws2.cell(row=r, column=1, value='平均完成度').font = SUBTITLE_FONT
    ws2.cell(row=r, column=1).fill = BAND_FILL
    ws2.cell(row=r, column=1).alignment = CENTER
    ws2.cell(row=r, column=4, value=f'=AVERAGE(D4:D{r-2})').fill = BAND_FILL
    ws2.cell(row=r, column=4).font = SUBTITLE_FONT
    ws2.cell(row=r, column=4).alignment = CENTER
    for col in range(1, 8):
        ws2.cell(row=r, column=col).border = BORDER

    ws3 = wb.create_sheet('30 天复盘')
    set_widths(ws3, [6, 26, 60, 4])
    ws3.merge_cells('A1:C1')
    ws3['A1'] = '30 天复盘 · 5 个关键问题'
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].alignment = CENTER
    ws3.row_dimensions[1].height = 26
    write_header(ws3, 3, ['编号', '问题', '我的回答'])
    qs = [
        '承诺完成情况如何？',
        '哪些行动效果最好？为什么？',
        '哪些行动失败了？原因是什么？',
        '土壤三要素（心理安全 / 响应 / 容错）有变化吗？',
        '下一步 60 天的核心动作是什么？',
    ]
    for i, q in enumerate(qs):
        r = 4 + i
        ws3.cell(row=r, column=1, value=i + 1).alignment = CENTER
        ws3.cell(row=r, column=2, value=q).font = SUBTITLE_FONT
        ws3.cell(row=r, column=2).alignment = LEFT
        c = ws3.cell(row=r, column=3, value='')
        c.fill = INPUT_FILL
        c.alignment = LEFT
        for col in range(1, 4):
            ws3.cell(row=r, column=col).border = BORDER
        ws3.row_dimensions[r].height = 50

    ws4 = wb.create_sheet('小组互访记录')
    set_widths(ws4, [6, 20, 18, 30, 30, 4])
    ws4.merge_cells('A1:E1')
    ws4['A1'] = '小组互访记录 · 60 天互访'
    ws4['A1'].font = TITLE_FONT
    ws4['A1'].alignment = CENTER
    ws4.row_dimensions[1].height = 26
    write_header(ws4, 3, ['编号', '被访企业/团队', '互访日期', '看到的好实践', '可借鉴的 1 件事'])
    for i in range(5):
        r = 4 + i
        ws4.cell(row=r, column=1, value=i + 1).alignment = CENTER
        for col in range(2, 6):
            c = ws4.cell(row=r, column=col, value='')
            c.fill = INPUT_FILL
            c.alignment = LEFT
        for col in range(1, 6):
            ws4.cell(row=r, column=col).border = BORDER
        ws4.row_dimensions[r].height = 40

    wb.save(os.path.join(OUT_DIR, '工具_09_30天跟进记录表.xlsx'))
    print('OK: 工具_09_30天跟进记录表.xlsx')


# ====================================================================
# 成果Demo_04_团队创新问题台账.xlsx
# ====================================================================
def build_demo_04():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '使用说明'
    set_widths(ws1, [4, 110])
    ws1['B2'] = '成果 Demo 04 · 团队创新问题台账'
    ws1['B2'].font = TITLE_FONT
    ws1['B2'].alignment = LEFT
    notes = [
        '【核心命题】每个问题 = 1 个可观察、可追溯、可关闭的工单。',
        '【用途】团队层面把所有发现的"创新阻碍问题"集中管理，定期 review。',
        '【使用时机】',
        '· 课程模块二：每组识别 5 个本团队当前最痛的"土壤问题"。',
        '· 课后 30 天：每两周 1 次 review。',
        '· 60/90 天：识别"已关闭"和"持续未解决"两类。',
        '【11 个字段】',
        '· 问题编号 / 类别（土壤三要素 / 五阶段 / 跨部门）',
        '· 描述（具体场景）/ 发现日期 / 发现人',
        '· 紧急度（高/中/低）/ 影响力（1-5 分）',
        '· 状态（待处理 / 处理中 / 已解决 / 暂缓）',
        '· 责任人 / 解决日期',
        '【使用规则】',
        '· 每个问题必须有 1 个负责人。',
        '· 高紧急度问题 30 天内必须处理或转交。',
        '· 已解决问题每月复盘 1 次，看是否有复发。',
        '【配套资源】工具_02_创新土壤诊断卡、工具_04_情境演练卡。',
    ]
    write_notes(ws1, 4, 2, notes, highlight_indices=[0, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 17, 18, 19])
    ws1.row_dimensions[2].height = 28

    ws2 = wb.create_sheet('问题台账')
    set_widths(ws2, [4, 8, 18, 40, 12, 12, 8, 10, 14, 12, 12, 4])
    ws2.merge_cells('A1:K1')
    ws2['A1'] = '团队创新问题台账'
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 26
    ws2.merge_cells('A2:K2')
    ws2['A2'] = '团队：________________  负责人：________________  启动日期：____________'
    ws2['A2'].font = NOTE_FONT
    ws2['A2'].fill = BAND_FILL
    write_header(ws2, 3, ['编号', '类别', '问题描述', '发现日期', '发现人', '紧急度', '影响力(1-5)', '状态', '责任人', '解决日期', '关闭'])
    for i in range(15):
        r = 4 + i
        ws2.cell(row=r, column=1, value=f'P{i+1:03d}').alignment = CENTER
        ws2.cell(row=r, column=2, value='').fill = INPUT_FILL
        ws2.cell(row=r, column=3, value='').fill = INPUT_FILL
        ws2.cell(row=r, column=4, value='').fill = INPUT_FILL
        ws2.cell(row=r, column=5, value='').fill = INPUT_FILL
        ws2.cell(row=r, column=6, value='').alignment = CENTER
        ws2.cell(row=r, column=6).fill = INPUT_FILL
        ws2.cell(row=r, column=7, value='').alignment = CENTER
        ws2.cell(row=r, column=7).fill = INPUT_FILL
        ws2.cell(row=r, column=8, value='').alignment = CENTER
        ws2.cell(row=r, column=8).fill = INPUT_FILL
        ws2.cell(row=r, column=9, value='').fill = INPUT_FILL
        ws2.cell(row=r, column=10, value='').fill = INPUT_FILL
        c = ws2.cell(row=r, column=11, value='☐')
        c.alignment = CENTER
        c.border = BORDER
        for col in range(1, 12):
            ws2.cell(row=r, column=col).border = BORDER
        ws2.row_dimensions[r].height = 40

    # 统计
    r = 4 + 15 + 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws2.cell(row=r, column=1, value='状态统计').font = SUBTITLE_FONT
    ws2.cell(row=r, column=1).fill = BAND_FILL
    ws2.cell(row=r, column=1).alignment = CENTER
    write_header(ws2, r + 1, ['状态', '数量', '占比', '', '', '', '', '', '', '', ''])
    statuses = ['待处理', '处理中', '已解决', '暂缓']
    for i, s in enumerate(statuses):
        rr = r + 2 + i
        ws2.cell(row=rr, column=1, value=s).font = SUBTITLE_FONT
        ws2.cell(row=rr, column=2, value=f'=COUNTIF(H4:H18,"{s}")').fill = INPUT_FILL
        ws2.cell(row=rr, column=2).alignment = CENTER
        ws2.cell(row=rr, column=3, value=f'=IFERROR(B{rr}/SUM($B${r+2}:$B${r+5}),0)').fill = BAND_FILL
        ws2.cell(row=rr, column=3).number_format = '0.0%'
        ws2.cell(row=rr, column=3).alignment = CENTER
        for col in range(1, 4):
            ws2.cell(row=rr, column=col).border = BORDER

    ws3 = wb.create_sheet('类别模板')
    set_widths(ws3, [6, 22, 50, 4])
    ws3.merge_cells('A1:C1')
    ws3['A1'] = '问题类别速查'
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].alignment = CENTER
    ws3.row_dimensions[1].height = 26
    write_header(ws3, 3, ['类别', '代表场景', '建议处理路径'])
    cats = [
        ('心理安全', '员工提了不成熟想法被嘲笑', '领导公开示范"先听完再判断"'),
        ('响应机制', '点子提完石沉大海', '建立"3 日必复"硬规则'),
        ('容错文化', '尝试失败后被追责', '把"学到了什么"做成固定栏目'),
        ('发现阶段', '看不到问题 / 机会', '建立用户声音机制'),
        ('构思阶段', '点子没有筛选标准', '明确筛选维度（价值 / 可行性）'),
        ('验证阶段', '直接做大没小试', '强制最小化测试环节'),
        ('落地阶段', '跨部门协调失败', '总裁办兜底 + 周会机制'),
        ('固化阶段', '换领导方式失传', '写进 SOP + 培训'),
    ]
    for i, (c, sc, su) in enumerate(cats):
        r = 4 + i
        ws3.cell(row=r, column=1, value=c).font = SUBTITLE_FONT
        ws3.cell(row=r, column=2, value=sc).alignment = LEFT
        ws3.cell(row=r, column=3, value=su).alignment = LEFT
        for col in range(1, 4):
            ws3.cell(row=r, column=col).border = BORDER
        ws3.row_dimensions[r].height = 32

    wb.save(os.path.join(DEMO_DIR, '成果Demo_04_团队创新问题台账.xlsx'))
    print('OK: 成果Demo_04_团队创新问题台账.xlsx')


if __name__ == '__main__':
    # 工具_03 已经有生成脚本，这里直接调用
    build_tool_04()
    build_tool_05()
    build_tool_06()
    build_tool_07()
    build_tool_08()
    build_tool_09()
    build_demo_04()
    # 工具_03 - 直接运行已有脚本
    import subprocess
    result = subprocess.run(
        ['python', r'D:\2026年课程\竞越\打造组织创新力：营造创新土壤\完整课程包\08_全流程工具表单\生成脚本\工具_03_创新过程五阶段复盘表.py'],
        capture_output=True, text=True
    )
    print(result.stdout)
    if result.returncode != 0:
        print('ERROR in 工具_03:', result.stderr)
    print('---DONE---')

# -*- coding: utf-8 -*-
"""
Create two Excel files for Employee Development:
1. F6_员工发展档案模板.xlsx - Employee Development Archive Template
2. F7_发展对话记录表.xlsx - Development Dialogue Record Form
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# Color scheme
RED_ACCENT = "B81025"
GRAY_DARK = "4A4A4A"
LIGHT_BG = "F5F5F5"
WHITE = "FFFFFF"

def create_border(style='thin'):
    side = Side(style=style, color="000000")
    return Border(left=side, right=side, top=side, bottom=side)

def set_cell_style(cell, font_size=11, bold=False, font_color="000000",
                   bg_color=None, align_h="left", align_v="center",
                   wrap_text=False, border=True, font_name="微软雅黑"):
    cell.font = Font(name=font_name, size=font_size, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap_text)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    if border:
        cell.border = create_border()

def merge_and_style(ws, start_row, start_col, end_row, end_col, value,
                    font_size=11, bold=False, font_color="000000",
                    bg_color=None, align_h="left", align_v="center",
                    wrap_text=False, border=True, font_name="微软雅黑"):
    ws.merge_cells(start_row=start_row, start_column=start_col,
                  end_row=end_row, end_column=end_col)
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = value
    set_cell_style(cell, font_size=font_size, bold=bold, font_color=font_color,
                   bg_color=bg_color, align_h=align_h, align_v=align_v,
                   wrap_text=wrap_text, border=border, font_name=font_name)
    return cell

def set_column_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def set_row_heights(ws, heights):
    for row, height in heights.items():
        ws.row_dimensions[row].height = height

def create_f6_employee_development_archive():
    wb = Workbook()
    ws = wb.active
    ws.title = "员工发展档案"

    # Page setup - A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    # Column widths
    widths = {
        'A': 3, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 16, 'H': 16, 'I': 16, 'J': 16, 'K': 16, 'L': 3
    }
    set_column_widths(ws, widths)

    current_row = 1

    # Title
    merge_and_style(ws, current_row, 2, current_row, 11, "员工发展档案模板",
                    font_size=20, bold=True, font_color=RED_ACCENT,
                    bg_color=LIGHT_BG, align_h="center", align_v="center")
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    # Subtitle
    merge_and_style(ws, current_row, 2, current_row, 11, "A4单张 | 可持续更新 | 追踪员工发展进度",
                    font_size=11, bold=False, font_color=GRAY_DARK,
                    bg_color=LIGHT_BG, align_h="center", align_v="center")
    ws.row_dimensions[current_row].height = 22
    current_row += 2

    # Section 1: 基本信息
    merge_and_style(ws, current_row, 2, current_row, 11, "基本信息",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Basic info fields - 2 columns each
    info_fields = [
        ("姓名:", ""), ("部门:", ""), ("岗位:", ""),
        ("入职日期:", ""), ("直接经理:", ""), ("档案建立日期:", ""), ("档案更新日期:", "")
    ]

    for i in range(0, len(info_fields), 2):
        # Left field
        field1 = info_fields[i]
        ws.cell(row=current_row, column=2).value = field1[0]
        set_cell_style(ws.cell(row=current_row, column=2), font_size=11, bold=True,
                       font_color=GRAY_DARK, bg_color=LIGHT_BG, border=True)
        ws.cell(row=current_row, column=3).value = field1[1]
        set_cell_style(ws.cell(row=current_row, column=3), font_size=11, border=True)

        # Right field
        if i + 1 < len(info_fields):
            field2 = info_fields[i + 1]
            ws.cell(row=current_row, column=5).value = field2[0]
            set_cell_style(ws.cell(row=current_row, column=5), font_size=11, bold=True,
                           font_color=GRAY_DARK, bg_color=LIGHT_BG, border=True)
            ws.cell(row=current_row, column=6).value = field2[1]
            set_cell_style(ws.cell(row=current_row, column=6), font_size=11, border=True)

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    current_row += 1

    # Section 2: 技能现状表格
    merge_and_style(ws, current_row, 2, current_row, 11, "技能现状",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Skill table header
    skill_headers = ["技能类别", "具体技能", "当前水平(1-5)", "目标水平(1-5)", "差距", "优先级"]
    header_cols = [2, 3, 4, 5, 6, 7]
    for idx, header in enumerate(skill_headers):
        cell = ws.cell(row=current_row, column=header_cols[idx])
        cell.value = header
        set_cell_style(cell, font_size=11, bold=True, font_color=WHITE,
                       bg_color=RED_ACCENT, align_h="center", border=True)
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # Skill table rows (8 empty rows)
    for row in range(8):
        for idx, col in enumerate(header_cols):
            cell = ws.cell(row=current_row, column=col)
            if idx == 0:  # First column - skill category
                cell.value = ""
            elif idx == 2 or idx == 3:  # Current/Target level - numeric
                cell.value = ""
            elif idx == 4:  # Gap - formula
                c_col = get_column_letter(col)
                cell.value = f"=IF(D{current_row}-C{current_row}>0,D{current_row}-C{current_row},0)"
            elif idx == 5:  # Priority
                cell.value = ""
            set_cell_style(cell, font_size=11, align_h="center" if idx >= 2 else "left", border=True)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    current_row += 1

    # Section 3: 发展目标
    merge_and_style(ws, current_row, 2, current_row, 11, "发展目标",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Development goals - 3 periods
    goal_periods = [
        ("短期目标 (3个月)", 4),
        ("中期目标 (6-12个月)", 5),
        ("长期目标 (1-2年)", 5)
    ]

    for period_name, row_span in goal_periods:
        # Period header
        merge_and_style(ws, current_row, 2, current_row, 2, period_name,
                        font_size=11, bold=True, font_color=GRAY_DARK,
                        bg_color=LIGHT_BG, align_h="left", border=True)

        # Content area
        merge_and_style(ws, current_row, 3, current_row + row_span - 1, 11, "",
                        font_size=11, border=True)
        ws.row_dimensions[current_row].height = 20
        current_row += row_span

    current_row += 1

    # Section 4: 路径规划
    merge_and_style(ws, current_row, 2, current_row, 11, "路径规划",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Path planning fields
    path_fields = [
        ("优先发展项 Top3:", 3),
        ("资源支持需求:", 3),
        ("预计完成时间:", 1)
    ]

    for field_name, row_span in path_fields:
        merge_and_style(ws, current_row, 2, current_row, 2, field_name,
                        font_size=11, bold=True, font_color=GRAY_DARK,
                        bg_color=LIGHT_BG, align_h="left", border=True)
        merge_and_style(ws, current_row, 3, current_row + row_span - 1, 11, "",
                        font_size=11, border=True)
        ws.row_dimensions[current_row].height = 20 * row_span
        current_row += row_span

    current_row += 1

    # Section 5: 里程碑记录
    merge_and_style(ws, current_row, 2, current_row, 11, "里程碑记录",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Milestone table header
    milestone_headers = ["日期", "里程碑事件", "成果表现", "备注"]
    milestone_cols = [2, 4, 8, 11]
    for idx, header in enumerate(milestone_headers):
        end_col = milestone_cols[idx + 1] - 1 if idx + 1 < len(milestone_cols) else 11
        cell = ws.cell(row=current_row, column=milestone_cols[idx])
        ws.merge_cells(start_row=current_row, start_column=milestone_cols[idx],
                       end_row=current_row, end_column=end_col)
        cell.value = header
        set_cell_style(cell, font_size=11, bold=True, font_color=WHITE,
                       bg_color=RED_ACCENT, align_h="center", border=True)
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # Milestone table rows (6 empty rows)
    for row in range(6):
        for idx, start_col in enumerate(milestone_cols):
            end_col = milestone_cols[idx + 1] - 1 if idx + 1 < len(milestone_cols) else 11
            ws.merge_cells(start_row=current_row, start_column=start_col,
                           end_row=current_row, end_column=end_col)
            cell = ws.cell(row=current_row, column=start_col)
            set_cell_style(cell, font_size=11, border=True)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # Print area and page setup
    ws.print_area = f"A1:L{current_row}"

    output_path = r"D:/新课开发/HR/培训/10_员工发展导向：把学习数据变成成长路径而不是完课记录/Excel工具表单/F6_员工发展档案模板.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")

def create_f7_development_dialogue_record():
    wb = Workbook()
    ws = wb.active
    ws.title = "发展对话记录"

    # Page setup - A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    # Column widths
    widths = {
        'A': 3, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 16, 'H': 16, 'I': 16, 'J': 16, 'K': 16, 'L': 3
    }
    set_column_widths(ws, widths)

    current_row = 1

    # Title
    merge_and_style(ws, current_row, 2, current_row, 11, "发展对话记录表",
                    font_size=20, bold=True, font_color=RED_ACCENT,
                    bg_color=LIGHT_BG, align_h="center", align_v="center")
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    # Subtitle
    merge_and_style(ws, current_row, 2, current_row, 11, "A4单张 | 每次对话后填写 | 记录和跟进",
                    font_size=11, bold=False, font_color=GRAY_DARK,
                    bg_color=LIGHT_BG, align_h="center", align_v="center")
    ws.row_dimensions[current_row].height = 22
    current_row += 2

    # Section 1: 对话基本信息
    merge_and_style(ws, current_row, 2, current_row, 11, "对话基本信息",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Basic info fields
    info_fields = [
        ("员工姓名:", ""), ("岗位:", ""), ("直接经理:", ""),
        ("对话日期:", ""), ("对话类型:", ""), ("第几次对话:", "")
    ]

    for i in range(0, len(info_fields), 2):
        field1 = info_fields[i]
        ws.cell(row=current_row, column=2).value = field1[0]
        set_cell_style(ws.cell(row=current_row, column=2), font_size=11, bold=True,
                       font_color=GRAY_DARK, bg_color=LIGHT_BG, border=True)
        ws.cell(row=current_row, column=3).value = field1[1]
        set_cell_style(ws.cell(row=current_row, column=3), font_size=11, border=True)

        if i + 1 < len(info_fields):
            field2 = info_fields[i + 1]
            ws.cell(row=current_row, column=5).value = field2[0]
            set_cell_style(ws.cell(row=current_row, column=5), font_size=11, bold=True,
                           font_color=GRAY_DARK, bg_color=LIGHT_BG, border=True)
            ws.cell(row=current_row, column=6).value = field2[1]
            set_cell_style(ws.cell(row=current_row, column=6), font_size=11, border=True)

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    current_row += 1

    # Section 2: 对话摘要
    merge_and_style(ws, current_row, 2, current_row, 11, "对话摘要",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Summary text area
    merge_and_style(ws, current_row, 2, current_row + 4, 11, "",
                    font_size=11, border=True)
    ws.row_dimensions[current_row].height = 100
    current_row += 5

    current_row += 1

    # Section 3: 达成的共识
    merge_and_style(ws, current_row, 2, current_row, 11, "达成的共识",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Consensus fields
    consensus_fields = [
        ("员工的发展目标:", 3),
        ("双方认同的优势:", 2),
        ("双方认同的发展机会:", 2),
        ("下一步行动计划:", 3)
    ]

    for field_name, row_span in consensus_fields:
        merge_and_style(ws, current_row, 2, current_row, 2, field_name,
                        font_size=11, bold=True, font_color=GRAY_DARK,
                        bg_color=LIGHT_BG, align_h="left", border=True)
        merge_and_style(ws, current_row, 3, current_row + row_span - 1, 11, "",
                        font_size=11, border=True)
        ws.row_dimensions[current_row].height = 20 * row_span
        current_row += row_span

    current_row += 1

    # Section 4: 下一步行动
    merge_and_style(ws, current_row, 2, current_row, 11, "下一步行动",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Action table header
    action_headers = ["角色", "行动项", "截止日期"]
    action_cols = [2, 4, 11]
    for idx, header in enumerate(action_headers):
        end_col = action_cols[idx + 1] - 1 if idx + 1 < len(action_cols) else 11
        cell = ws.cell(row=current_row, column=action_cols[idx])
        ws.merge_cells(start_row=current_row, start_column=action_cols[idx],
                       end_row=current_row, end_column=end_col)
        cell.value = header
        set_cell_style(cell, font_size=11, bold=True, font_color=WHITE,
                       bg_color=RED_ACCENT, align_h="center", border=True)
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # Action roles
    action_roles = ["员工行动项", "经理行动项", "需要的支持资源"]
    for role in action_roles:
        ws.merge_cells(start_row=current_row, start_column=2,
                       end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=2)
        cell.value = role
        set_cell_style(cell, font_size=11, bold=True, font_color=GRAY_DARK,
                       bg_color=LIGHT_BG, align_h="left", border=True)

        ws.merge_cells(start_row=current_row, start_column=4,
                       end_row=current_row, end_column=10)
        cell = ws.cell(row=current_row, column=4)
        set_cell_style(cell, font_size=11, border=True)

        ws.cell(row=current_row, column=11).value = ""
        set_cell_style(ws.cell(row=current_row, column=11), font_size=11, border=True)

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    current_row += 1

    # Section 5: 下次检查点
    merge_and_style(ws, current_row, 2, current_row, 11, "下次检查点",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Checkpoint fields
    checkpoint_fields = [
        ("日期:", ""), ("重点关注:", ""), ("届时讨论的问题:", "")
    ]

    for field_name, val in checkpoint_fields:
        ws.cell(row=current_row, column=2).value = field_name
        set_cell_style(ws.cell(row=current_row, column=2), font_size=11, bold=True,
                       font_color=GRAY_DARK, bg_color=LIGHT_BG, border=True)
        ws.merge_cells(start_row=current_row, start_column=3,
                       end_row=current_row, end_column=11)
        cell = ws.cell(row=current_row, column=3)
        cell.value = val
        set_cell_style(cell, font_size=11, border=True)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # This is the front side - save as page 1
    front_side_end_row = current_row

    # Print area for front side
    ws.print_area = f"A1:L{front_side_end_row}"

    output_path = r"D:/新课开发/HR/培训/10_员工发展导向：把学习数据变成成长路径而不是完课记录/Excel工具表单/F7_发展对话记录表.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")

if __name__ == "__main__":
    import os

    output_dir = r"D:/新课开发/HR/培训/10_员工发展导向：把学习数据变成成长路径而不是完课记录/Excel工具表单"
    os.makedirs(output_dir, exist_ok=True)

    create_f6_employee_development_archive()
    create_f7_development_dialogue_record()

    print("Both Excel files created successfully!")

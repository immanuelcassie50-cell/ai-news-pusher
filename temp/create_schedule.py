import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "验证时间表"

header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
data_font = Font(name='微软雅黑', size=10)
title_font = Font(name='微软雅黑', bold=True, size=14)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

ws.merge_cells('A1:I1')
ws['A1'] = '华东制造集团ERP升级项目 - 假设验证时间表'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:I2')
ws['A2'] = '版本：V1.0    编制人：李晓峰    更新日期：2024年7月20日'
ws['A2'].font = Font(name='微软雅黑', size=10, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')
ws.row_dimensions[2].height = 20

headers = ['假设编号', '假设描述', '验证方式', '计划验证时间', '实际验证时间', '验证状态', '验证责任人', '验证结论', '备注']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[4].height = 35

verifications = [
    ['H001', '现有SAP R/3系统数据可以完整迁移至S/4HANA', '执行数据质量分析，执行迁移测试', '2024/8/15', '', '待验证', '陈建国', '', '需在调研阶段确认'],
    ['H002', '现有服务器配置可以满足S/4HANA性能要求', '进行压力测试，收集性能基准数据', '2024/9/30', '', '待验证', '刘建华', '', '已有初步评估'],
    ['H003', '与MES系统的接口可以在6周内完成开发', '技术方案评审，确认接口复杂度', '2024/10/15', '', '待验证', '刘建华', '', '需实测评估'],
    ['H004', '网络带宽可以支持实时数据传输', '网络评估测试', '2024/9/15', '', '待验证', '刘建华', '', ''],
    ['H005', '各子公司愿意放弃本地流程，接受集团统一流程', '在启动会上正式访谈各子公司负责人', '2024/7/25', '2024/7/25', '已验证通过', '赵敏', '南方子公司明确抵制，需要重新谈判', '存在较大不确定性'],
    ['H006', '关键用户可以投入足够时间配合实施', '确认关键用户名单及其时间安排', '2024/7/30', '2024/7/28', '已验证通过', '李晓峰', '财务部关键用户每周只能投入50%时间', '已有个别反馈称时间不足'],
    ['H007', '现有物料编码体系可以标准化而不需要重新编码', '执行物料数据质量分析', '2024/8/30', '', '待验证', '陈建国', '', ''],
    ['H008', '财务月结流程可以在新系统中压缩至5天', '流程建模与仿真测试', '2025/2/28', '', '待验证', '赵敏', '', '需在设计阶段验证'],
    ['H009', '供应商主数据可以由采购部统一提供并维护', '确认数据维护流程和责任人', '2024/10/30', '', '待验证', '赵敏', '', ''],
    ['H010', '预算800万可以覆盖全部实施范围', '每月成本跟踪，差异分析', '持续', '', '持续跟踪', '李晓峰', '', '已有超支风险'],
    ['H011', '可以在2周内完成服务器采购和部署', '采购流程确认', '2024/8/1', '2024/8/10', '已验证通过', '刘建华', '采购流程正常，但部署遇到网络配置问题', '已与采购部确认'],
    ['H012', '外部顾问可以在需要时到场支持', '合同条款确认', '2024/7/31', '2024/7/31', '已验证通过', '张明华', '合同已签署，响应条款明确', '合同已签署'],
    ['H013', '12个月的实施周期足够完成全部功能', '详细工作量评估，项目进度模拟', '2024/9/15', '2024/9/20', '已验证失败', '李晓峰', '评估显示需要14.5个月，缺口2.5个月', '已有延期预警'],
    ['H014', '业务调研可以在6周内完成', '调研进度跟踪', '2024/8/30', '2024/9/10', '已验证失败', '赵敏', '调研延期10天，原因：关键用户时间不足+调研范围扩大', '已出现延期'],
    ['H015', 'UAT测试可以在4周内完成', 'UAT计划评审', '2025/3/15', '', '待验证', '吴婷', '', ''],
    ['H016', '国资委的合规要求不会有重大变化', '政策跟踪，定期沟通', '持续', '', '持续跟踪', '陈永强', '', '需定期确认'],
    ['H017', '没有其他重大项目与本项目时间冲突', '确认公司重大项目日历', '2024/7/31', '2024/7/31', '已验证通过', '林晓华', '已协调，各部门时间已错开', '已协调'],
    ['H018', '主要供应商可以按时交付，不会有供应链中断', '供应商状态跟踪', '2024/8/31', '', '待验证', '刘建华', '', ''],
    ['H019', '客户方项目对接人员有足够的决策权限', '在启动会上确认汇报线和决策权限', '2024/7/25', '2024/7/26', '已验证失败', '李晓峰', '林晓华只能审批50万以下事项，更大的决策需要王建业副总裁批准', '已在启动会发现问题'],
    ['H020', '项目范围说明书中的除外条款是完整的', '范围边界确认，与各业务部门访谈', '2024/8/15', '', '待验证', '李晓峰', '', ''],
]

for row_idx, row_data in enumerate(verifications, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border
        
        status = row_data[5]
        if status == '已验证通过':
            cell.fill = green_fill
        elif status == '已验证失败':
            cell.fill = red_fill
        elif status == '待验证':
            cell.fill = gray_fill
        elif status == '持续跟踪':
            cell.fill = yellow_fill
    ws.row_dimensions[row_idx].height = 45

column_widths = [12, 35, 30, 15, 15, 12, 15, 30, 20]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws2 = wb.create_sheet('状态说明')
ws2['A1'] = '验证状态说明'
ws2['A1'].font = Font(name='微软雅黑', bold=True, size=14)
ws2['A3'] = '待验证'
ws2['A3'].fill = gray_fill
ws2['B3'] = '尚未开始验证'
ws2['A4'] = '已验证通过'
ws2['A4'].fill = green_fill
ws2['B4'] = '验证完成，假设成立'
ws2['A5'] = '已验证失败'
ws2['A5'].fill = red_fill
ws2['B5'] = '验证完成，假设不成立，需要应对'
ws2['A6'] = '持续跟踪'
ws2['A6'].fill = yellow_fill
ws2['B6'] = '需要在项目周期内持续关注和验证'

output_path = 'D:/CC/新课开发/工作手册/假设管理：项目经理的风险前置手册/完整课程包/07-成果demo/验证时间表-演示.xlsx'
wb.save(output_path)
print(f"OK: {output_path}")

"""Generate 学员试讲评分表.xlsx with 6+ sheets."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# 颜色和样式
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SUBHEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="1F4E78")
CELL_FONT = Font(name="Microsoft YaHei", size=10)
TOTAL_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
TOTAL_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="9C5700")
thin = Side(border_style="thin", color="999999")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Sheet 1: 学员试讲评分表（主表 - 4 维 × 5 级）
# ============================================================
ws1 = wb.active
ws1.title = "1.试讲评分主表"

ws1.merge_cells("A1:H1")
ws1.cell(row=1, column=1, value="FAST 训练营 · 学员试讲评分表（4 维 × 5 级）")
ws1.cell(row=1, column=1).font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
ws1.cell(row=1, column=1).alignment = CENTER
ws1.row_dimensions[1].height = 28

ws1["A3"] = "学员姓名：________"
ws1["C3"] = "课程主题：________"
ws1["E3"] = "试讲日期：________"
ws1["G3"] = "试讲时长：________分钟"
for c in ["A3", "C3", "E3", "G3"]:
    ws1[c].font = CELL_FONT
    ws1[c].alignment = LEFT

# 4 维评估
headers = ["维度", "评估指标", "1 分（不达标）", "3 分（达标）", "5 分（超出预期）", "学员自评", "讲师评分", "关键观察"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=5, column=i, value=h)
style_header(ws1, 5, 8)
ws1.row_dimensions[5].height = 30

# 内容深度（注意用单引号或转义）
data = [
    ["内容深度", "方法是否讲清楚（3 步骤/1 公式/1 心法）", "没讲", "讲清楚", "加案例 + 个人方法论", "", "", ""],
    ["内容深度", "案例是否具体 + 有冲突", "没有", "1 个", "3 个不同场景", "", "", ""],
    ["内容深度", "是否有个人方法论", "抄方法论", "提到方法论", "个人方法论 v1.0", "", "", ""],
    ["互动设计", "互动频次（5-8 分钟 1 次）", "0-1 次", "3-4 次", "5-8 次且分布均匀", "", "", ""],
    ["互动设计", "重互动（实操填写）", "全是轻互动", "1 次重互动", "3+ 次重互动 + 点评", "", "", ""],
    ["互动设计", "互动话术是否具体", "笼统的开放问题", "具体（A/B/C/D）", "具体 + 等待时间", "", "", ""],
    ["节奏控制", "总时长是否在计划内", "超时 30%+", "±20%", "±10%", "", "", ""],
    ["节奏控制", "颗粒时长是否 10-15 分钟", "< 5 或 > 20", "10-15", "10-15 + 流畅", "", "", ""],
    ["节奏控制", "颗粒间过渡是否承接", "接下来...", "1 句过渡", "承接式 + 引用互动", "", "", ""],
    ["呈现专业", "镜头 4 要素（视线/光源/背景/构图）", "3 项不达标", "全部达标", "全部达标 + 自然", "", "", ""],
    ["呈现专业", "声音（清晰 + 节奏）", "有杂音/小声", "清晰", "清晰 + 有停顿/重音", "", "", ""],
    ["呈现专业", "PPT 一页一点 + 字号", "1 页 3+ 信息", "1 页 1-2 个", "1 页 1 个 + 视觉强化", "", "", ""],
]
row = 6
for d in data:
    for c, v in enumerate(d, 1):
        cell = ws1.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT
        cell.border = BORDER
    row += 1

# 总分
ws1.merge_cells(f"A{row}:E{row}")
ws1.cell(row=row, column=1, value="总分计算（12 项 × 5 分 = 60 分）").font = TOTAL_FONT
ws1.cell(row=row, column=1).fill = TOTAL_FILL
ws1.cell(row=row, column=1).alignment = CENTER
ws1.cell(row=row, column=1).border = BORDER
ws1.cell(row=row, column=6, value="自评总分：").font = TOTAL_FONT
ws1.cell(row=row, column=6).fill = TOTAL_FILL
ws1.cell(row=row, column=6).alignment = CENTER
ws1.cell(row=row, column=6).border = BORDER
ws1.cell(row=row, column=7, value=f"=SUM(F6:F{row-1})").font = TOTAL_FONT
ws1.cell(row=row, column=7).fill = TOTAL_FILL
ws1.cell(row=row, column=7).alignment = CENTER
ws1.cell(row=row, column=7).border = BORDER
ws1.cell(row=row, column=8, value=f"=SUM(G6:G{row-1})").font = TOTAL_FONT
ws1.cell(row=row, column=8).fill = TOTAL_FILL
ws1.cell(row=row, column=8).alignment = CENTER
ws1.cell(row=row, column=8).border = BORDER
ws1.row_dimensions[row].height = 26

# 等级
row += 1
ws1.merge_cells(f"A{row}:H{row}")
ws1.cell(row=row, column=1, value="等级判断：自评 50-60 = 卓越 | 40-49 = 熟练 | 30-39 = 合格 | 20-29 = 进阶 | 0-19 = 入门").font = Font(name="Microsoft YaHei", size=10, bold=True, color="C00000")
ws1.cell(row=row, column=1).alignment = CENTER

# 4 维小计
row += 2
ws1.cell(row=row, column=1, value="4 维小计（每维 3 项 × 5 分 = 15 分）").font = SUBHEADER_FONT
ws1.cell(row=row, column=1).fill = SUBHEADER_FILL
ws1.merge_cells(f"A{row}:H{row}")
ws1.cell(row=row, column=1).alignment = CENTER
row += 1

for d, label in [("内容深度", 6), ("互动设计", 9), ("节奏控制", 12), ("呈现专业", 15)]:
    ws1.cell(row=row, column=1, value=label).font = SUBHEADER_FONT
    ws1.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws1.cell(row=row, column=1).alignment = CENTER
    ws1.cell(row=row, column=1).border = BORDER
    ws1.cell(row=row, column=6, value=f"=SUM(F{d}:F{d+2})").font = CELL_FONT
    ws1.cell(row=row, column=6).alignment = CENTER
    ws1.cell(row=row, column=6).border = BORDER
    ws1.cell(row=row, column=7, value=f"=SUM(G{d}:G{d+2})").font = CELL_FONT
    ws1.cell(row=row, column=7).alignment = CENTER
    ws1.cell(row=row, column=7).border = BORDER
    row += 1

set_col_widths(ws1, [13, 25, 18, 18, 22, 10, 10, 30])


# ============================================================
# Sheet 2: 试讲前 5 分钟必查项
# ============================================================
ws2 = wb.create_sheet("2.开播前5分钟必查")
ws2.merge_cells("A1:E1")
ws2.cell(row=1, column=1, value="FAST 训练营 · 试讲前 5 分钟必查项（5 件工具）")
ws2.cell(row=1, column=1).font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
ws2.cell(row=1, column=1).alignment = CENTER
ws2.row_dimensions[1].height = 28

headers = ["顺序", "必查工具", "具体内容", "通过标准", "通过（勾）"]
for i, h in enumerate(headers, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, 5)

data = [
    [1, "工具 05：开场 90 秒脚本", "前 90 秒四步结构 + 自我介绍", "30 秒内能复述", "☐"],
    [2, "工具 06：互动设计", "5-8 次互动 + 重互动位置", "互动数 = 计划数", "☐"],
    [3, "工具 08：全程时间表", "顺序 + 内容 + 时长 + 累计", "总时长 = 计划时长", "☐"],
    [4, "工具 10：应急话术", "5 类突发情况 + 具体话术", "5 类都能说出 1 句", "☐"],
    [5, "工具 11：镜头自查", "视线/光源/背景/构图", "4 要素全达标", "☐"],
]
row = 4
for d in data:
    for c, v in enumerate(d, 1):
        cell = ws2.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT if c in [2, 3, 4] else CENTER
        cell.border = BORDER
    ws2.row_dimensions[row].height = 30
    row += 1

row += 1
ws2.merge_cells(f"A{row}:E{row}")
ws2.cell(row=row, column=1, value="5 分钟倒计时：1 分/件。完成后在通过列打勾，未通过先解决再开播。").font = Font(name="Microsoft YaHei", size=10, bold=True, color="C00000")
ws2.cell(row=row, column=1).alignment = CENTER

set_col_widths(ws2, [8, 25, 35, 30, 12])


# ============================================================
# Sheet 3: 强开场 90 秒脚本
# ============================================================
ws3 = wb.create_sheet("3.强开场90秒脚本")
ws3.merge_cells("A1:D1")
ws3.cell(row=1, column=1, value="FAST 训练营 · 强开场 90 秒脚本（4 步结构）")
ws3.cell(row=1, column=1).font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
ws3.cell(row=1, column=1).alignment = CENTER
ws3.row_dimensions[1].height = 28

headers = ["时间段", "步骤", "讲师台词（直接填）", "互动设计"]
for i, h in enumerate(headers, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, 4)

data = [
    ["0:00-0:20", "情境冲突", "例：上次有个学员跟我说，汇报讲到一半领导突然问：你想说什么？", "无"],
    ["0:20-0:40", "提问留白", "例：你们有过类似的瞬间吗？", "无（让学员回想）"],
    ["0:40-1:10", "收集反应", "发起投票/弹幕：你汇报时被领导打断的频率是：A 经常 B 偶尔 C 从不", "投票 4 选项"],
    ["1:10-1:30", "揭示主题+自我介绍", "例：今天我们解决的就是这个问题。我是 XXX，做了 10 年向上沟通。", "无"],
]
row = 4
for d in data:
    for c, v in enumerate(d, 1):
        cell = ws3.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT
        cell.border = BORDER
    ws3.row_dimensions[row].height = 50
    row += 1

row += 1
ws3.merge_cells(f"A{row}:D{row}")
ws3.cell(row=row, column=1, value="关键原则：自我介绍放最后（10-15 秒）| 互动给封闭选项 | 强开场总时长 90-120 秒").font = Font(name="Microsoft YaHei", size=10, bold=True, color="C00000")
ws3.cell(row=row, column=1).alignment = CENTER

set_col_widths(ws3, [14, 18, 60, 30])


# ============================================================
# Sheet 4: 颗粒拆解评分
# ============================================================
ws4 = wb.create_sheet("4.颗粒拆解评分")
ws4.merge_cells("A1:F1")
ws4.cell(row=1, column=1, value="FAST 训练营 · 颗粒拆解评分表（每个颗粒 1 行）")
ws4.cell(row=1, column=1).font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
ws4.cell(row=1, column=1).alignment = CENTER
ws4.row_dimensions[1].height = 28

headers = ["颗粒号", "任务式标题", "时长（分钟）", "互动次数", "重互动次数", "评分（1-5）"]
for i, h in enumerate(headers, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, 6)

for r in range(4, 9):
    for c in range(1, 7):
        cell = ws4.cell(row=r, column=c, value="")
        cell.font = CELL_FONT
        cell.alignment = LEFT if c == 2 else CENTER
        cell.border = BORDER
    ws4.row_dimensions[r].height = 26

# 评分标准
ws4.cell(row=11, column=1, value="评分标准").font = SUBHEADER_FONT
ws4.cell(row=11, column=1).fill = SUBHEADER_FILL
ws4.merge_cells("A11:F11")
ws4.cell(row=11, column=1).alignment = CENTER

standards = [
    ["1 分", "标题抽象（如了解XX），时长 < 5 或 > 20 分钟", "0 次互动", "0 次重互动", "——"],
    ["3 分", "标题含 1 个要素（情境/动作），时长 10-15 分钟", "2-3 次互动", "1 次重互动", "——"],
    ["5 分", "标题情境+动作完整，时长 10-15 分钟，分布合理", "3+ 次互动", "2+ 次重互动", "——"],
]
row = 12
for d in standards:
    for c, v in enumerate(d, 1):
        cell = ws4.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT if c == 2 else CENTER
        cell.border = BORDER
    ws4.row_dimensions[row].height = 30
    row += 1

set_col_widths(ws4, [10, 35, 14, 14, 16, 14])


# ============================================================
# Sheet 5: 互动设计评分
# ============================================================
ws5 = wb.create_sheet("5.互动设计评分")
ws5.merge_cells("A1:E1")
ws5.cell(row=1, column=1, value="FAST 训练营 · 互动设计评分表（5-8 次互动 = 5 行）")
ws5.cell(row=1, column=1).font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
ws5.cell(row=1, column=1).alignment = CENTER
ws5.row_dimensions[1].height = 28

headers = ["互动号", "时间点（分:秒）", "互动类型（轻/重）", "互动话术", "评分（1-5）"]
for i, h in enumerate(headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, 5)

for r in range(4, 11):
    for c in range(1, 6):
        cell = ws5.cell(row=r, column=c, value="")
        cell.font = CELL_FONT
        cell.alignment = LEFT if c == 4 else CENTER
        cell.border = BORDER
    ws5.row_dimensions[r].height = 26

# 评分标准
ws5.cell(row=13, column=1, value="评分标准").font = SUBHEADER_FONT
ws5.cell(row=13, column=1).fill = SUBHEADER_FILL
ws5.merge_cells("A13:E13")
ws5.cell(row=13, column=1).alignment = CENTER

standards = [
    ["1 分", "0-1 次互动", "全是轻互动", "开放问题（如大家说说）", "——"],
    ["3 分", "3-4 次互动", "1 次重互动", "具体（A/B/C/D）", "——"],
    ["5 分", "5-8 次互动，5-8 分钟间隔", "3+ 次重互动", "具体 + 等待时间", "——"],
]
row = 14
for d in standards:
    for c, v in enumerate(d, 1):
        cell = ws5.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT if c == 4 else CENTER
        cell.border = BORDER
    ws5.row_dimensions[row].height = 30
    row += 1

set_col_widths(ws5, [10, 18, 22, 50, 14])


# ============================================================
# Sheet 6: 应急话术演练记录
# ============================================================
ws6 = wb.create_sheet("6.应急话术演练")
ws6.merge_cells("A1:E1")
ws6.cell(row=1, column=1, value="FAST 训练营 · 应急话术演练记录（5 类突发情况）")
ws6.cell(row=1, column=1).font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
ws6.cell(row=1, column=1).alignment = CENTER
ws6.row_dimensions[1].height = 28

headers = ["突发类型", "话术 1（应对比 1 句）", "话术 2（应对比 2 句）", "话术 3（备用 1 句）", "演练评分（1-5）"]
for i, h in enumerate(headers, 1):
    ws6.cell(row=3, column=i, value=h)
style_header(ws6, 3, 5)

scenarios = [
    ["互动无响应", "例：大家可以点一下投票——没有标准答案", "例：我先说一下，我入行那会儿选的是 B……", "例：结果统计中，我先……", ""],
    ["网络卡顿", "例：大家听到我声音吗？", "例：我重新进入会议，请稍等 30 秒", "例：可不可以发条弹幕确认一下", ""],
    ["设备故障", "例：我的麦克风可能有点问题——我切换到手机", "例：请工作人员协助一下", "例：我们用弹幕继续", ""],
    ["学员超纲", "例：这个问题很好——我记下来，下节课专门讲", "例：我们先回到主线，这个问题我 1 对 1 回答", "例：其他同学也有类似问题吗？", ""],
    ["内容偏题", "例：这个问题可以下次深入讲——今天我们重点是 XXX", "例：我们回到主线", "例：你可以会后加我私聊", ""],
]
row = 4
for d in scenarios:
    for c, v in enumerate(d, 1):
        cell = ws6.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT if c in [2, 3, 4] else CENTER
        cell.border = BORDER
    ws6.row_dimensions[row].height = 50
    row += 1

# 评分标准
row += 1
ws6.cell(row=row, column=1, value="评分标准").font = SUBHEADER_FONT
ws6.cell(row=row, column=1).fill = SUBHEADER_FILL
ws6.merge_cells(f"A{row}:E{row}")
ws6.cell(row=row, column=1).alignment = CENTER
row += 1

standards = [
    ["1 分", "0-1 类应急话术", "只有等一下等笼统话", "没备用方案", "——"],
    ["3 分", "3 类应急话术", "1-2 句场景化", "1 句备用", "——"],
    ["5 分", "5 类全有，实际演练过", "3-5 句分层话术", "2-3 句备用", "——"],
]
for d in standards:
    for c, v in enumerate(d, 1):
        cell = ws6.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT if c in [2, 3, 4] else CENTER
        cell.border = BORDER
    ws6.row_dimensions[row].height = 30
    row += 1

set_col_widths(ws6, [14, 35, 35, 35, 14])


# ============================================================
# Sheet 7: 4F 反馈（自评 + 讲师评）
# ============================================================
ws7 = wb.create_sheet("7.4F反馈记录")
ws7.merge_cells("A1:D1")
ws7.cell(row=1, column=1, value="FAST 训练营 · 4F 反馈记录（自评 + 讲师）")
ws7.cell(row=1, column=1).font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
ws7.cell(row=1, column=1).alignment = CENTER
ws7.row_dimensions[1].height = 28

headers = ["F 维度", "学员自评", "讲师反馈", "评分（1-5）"]
for i, h in enumerate(headers, 1):
    ws7.cell(row=3, column=i, value=h)
style_header(ws7, 3, 4)

data = [
    ["Fact 事实", "（写具体数据：时长/互动次数/满意度/弹幕量）", "（讲师补充客观事实）", ""],
    ["Feel 感受", "（写具体感受：哪里紧张/哪里有底）", "（讲师补充观察到的情绪）", ""],
    ["Find 发现", "（写原因猜测：因为 X 所以 Y）", "（讲师补充更深层分析）", ""],
    ["Future 未来", "（写具体改进：颗粒 X 调整 Y 分钟）", "（讲师补充建议）", ""],
]
row = 4
for d in data:
    for c, v in enumerate(d, 1):
        cell = ws7.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT
        cell.border = BORDER
    ws7.row_dimensions[row].height = 60
    row += 1

# 评分标准
row += 1
ws7.cell(row=row, column=1, value="评分标准（每个 F 维度）").font = SUBHEADER_FONT
ws7.cell(row=row, column=1).fill = SUBHEADER_FILL
ws7.merge_cells(f"A{row}:D{row}")
ws7.cell(row=row, column=1).alignment = CENTER
row += 1

standards = [
    ["1 分", "没有事实 / 笼统的'不太好'", "没原因分析", "笼统的'多练'"],
    ["3 分", "提到数据（如'超时'）", "笼统的'方法不够熟'", "笼统的方向"],
    ["5 分", "具体数字（时长/互动率/满意度）", "含'因为 X 所以 Y'", "具体可执行（数字+动作）"],
]
for d in standards:
    for c, v in enumerate(d, 1):
        cell = ws7.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT
        cell.border = BORDER
    ws7.row_dimensions[row].height = 30
    row += 1

set_col_widths(ws7, [13, 50, 50, 12])


# ============================================================
# Sheet 8: 综合试讲总评
# ============================================================
ws8 = wb.create_sheet("8.综合试讲总评")
ws8.merge_cells("A1:C1")
ws8.cell(row=1, column=1, value="FAST 训练营 · 综合试讲总评")
ws8.cell(row=1, column=1).font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
ws8.cell(row=1, column=1).alignment = CENTER
ws8.row_dimensions[1].height = 28

ws8["A3"] = "学员姓名：________"
ws8["B3"] = "课程主题：________"
ws8["C3"] = "试讲日期：________"
for c in ["A3", "B3", "C3"]:
    ws8[c].font = CELL_FONT
    ws8[c].alignment = LEFT

headers = ["维度", "学员自评", "讲师评分"]
for i, h in enumerate(headers, 1):
    ws8.cell(row=5, column=i, value=h)
style_header(ws8, 5, 3)

# 8 个分项（来自各 sheet）
items = [
    ["1. 强开场 90 秒脚本（Sheet 3）", "", ""],
    ["2. 颗粒拆解（Sheet 4）", "", ""],
    ["3. 互动设计（Sheet 5）", "", ""],
    ["4. 应急话术（Sheet 6）", "", ""],
    ["5. 4F 反馈（Sheet 7）", "", ""],
    ["6. 内容深度（Sheet 1）", "", ""],
    ["7. 节奏控制（Sheet 1）", "", ""],
    ["8. 呈现专业（Sheet 1）", "", ""],
]
row = 6
for d in items:
    for c, v in enumerate(d, 1):
        cell = ws8.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT
        cell.border = BORDER
    row += 1

# 总分
ws8.cell(row=row, column=1, value="总分（8 项 × 5 分 = 40 分）").font = TOTAL_FONT
ws8.cell(row=row, column=1).fill = TOTAL_FILL
ws8.cell(row=row, column=1).border = BORDER
ws8.cell(row=row, column=2, value=f"=SUM(B6:B{row-1})").font = TOTAL_FONT
ws8.cell(row=row, column=2).fill = TOTAL_FILL
ws8.cell(row=row, column=2).alignment = CENTER
ws8.cell(row=row, column=2).border = BORDER
ws8.cell(row=row, column=3, value=f"=SUM(C6:C{row-1})").font = TOTAL_FONT
ws8.cell(row=row, column=3).fill = TOTAL_FILL
ws8.cell(row=row, column=3).alignment = CENTER
ws8.cell(row=row, column=3).border = BORDER
row += 1

# 等级
ws8.merge_cells(f"A{row}:C{row}")
ws8.cell(row=row, column=1, value="等级：34-40 = 卓越 | 26-33 = 熟练 | 18-25 = 合格 | 10-17 = 进阶 | 0-9 = 入门").font = Font(name="Microsoft YaHei", size=10, bold=True, color="C00000")
ws8.cell(row=row, column=1).alignment = CENTER
row += 2

# 文字总评
ws8.cell(row=row, column=1, value="文字总评（讲师 3 句：做得好的 + 待提升的 + 30 天建议）").font = SUBHEADER_FONT
ws8.cell(row=row, column=1).fill = SUBHEADER_FILL
ws8.merge_cells(f"A{row}:C{row}")
ws8.cell(row=row, column=1).alignment = CENTER
row += 1

comments = [
    "做得好的（具体）：",
    "1. ____________________________",
    "2. ____________________________",
    "",
    "待提升的（具体）：",
    "1. ____________________________",
    "2. ____________________________",
    "",
    "接下来 30 天的 1 个具体建议：",
    "____________________________",
]
for c in comments:
    ws8.merge_cells(f"A{row}:C{row}")
    ws8.cell(row=row, column=1, value=c).font = CELL_FONT
    ws8.cell(row=row, column=1).alignment = LEFT
    ws8.row_dimensions[row].height = 22
    row += 1

set_col_widths(ws8, [40, 20, 20])


# Save
output = "D:/Downloads/FAST线上训练营/08_评估工具包/学员试讲评分表.xlsx"
wb.save(output)
print(f"Saved: {output}")

import os
print(f"Size: {os.path.getsize(output)} bytes")

"""
创新领导力 - 3 个配套表单 Excel 生成器
- 01-使用指引.xlsx
- 02-空表.xlsx
- 03-填好示例.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import RadarChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

OUTPUT_DIR = r"D:\2026年课程\竞越\创新领导力：打造创新型团队\完整课程表\15-配套表单Excel"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ====== 配色 ======
COL_INK = "1A1A2E"        # 主文字
COL_ACCENT = "2D5F8A"     # 强调蓝
COL_ACCENT2 = "C8553D"    # 暖色
COL_LIGHT = "F4F1ED"      # 浅米
COL_HL = "FFF4D6"         # 高亮黄
COL_GREEN = "2A6B5A"
COL_GRAY = "6B6B73"
COL_RED = "A93B2A"
COL_BROWN = "8A3A2A"
COL_HEADER_BG = "2D5F8A"
COL_SUBHEAD_BG = "DDE5EE"

# 字体（中文 fallback 到 default）
FONT_TITLE = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
FONT_SUB = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Microsoft YaHei", size=11, bold=True, color=COL_INK)
FONT_H = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
FONT_NORM = Font(name="Microsoft YaHei", size=10, color=COL_INK)
FONT_NOTE = Font(name="Microsoft YaHei", size=9, color=COL_GRAY, italic=True)
FONT_BOLD = Font(name="Microsoft YaHei", size=10, bold=True, color=COL_INK)

FILL_TITLE = PatternFill("solid", fgColor=COL_HEADER_BG)
FILL_SUB = PatternFill("solid", fgColor=COL_ACCENT2)
FILL_SECTION = PatternFill("solid", fgColor=COL_SUBHEAD_BG)
FILL_LIGHT = PatternFill("solid", fgColor=COL_LIGHT)
FILL_HL = PatternFill("solid", fgColor=COL_HL)

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LT = Alignment(horizontal="left", vertical="top", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
BORDER_MED = Border(
    left=Side(style="medium", color=COL_INK),
    right=Side(style="medium", color=COL_INK),
    top=Side(style="medium", color=COL_INK),
    bottom=Side(style="medium", color=COL_INK),
)


def style_title(ws, row, col_span, text):
    """顶部标题栏"""
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1)
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = ALIGN_C
    ws.row_dimensions[row].height = 32
    for col in range(1, col_span + 1):
        ws.cell(row=row, column=col).border = BORDER_MED


def style_sub(ws, row, col_span, text):
    """副标题栏"""
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1)
    c.font = FONT_SUB
    c.fill = FILL_SUB
    c.alignment = ALIGN_C
    ws.row_dimensions[row].height = 22
    for col in range(1, col_span + 1):
        ws.cell(row=row, column=col).border = BORDER_MED


def style_header_row(ws, row, headers, fill=FILL_TITLE, font=FONT_H):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font
        c.fill = fill
        c.alignment = ALIGN_C
        c.border = BORDER_MED
    ws.row_dimensions[row].height = 26


def style_data_row(ws, row, values, height=None, font=FONT_NORM, align=ALIGN_LT, fill=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = font
        c.alignment = align
        c.border = BORDER_THIN
        if fill:
            c.fill = fill
    if height:
        ws.row_dimensions[row].height = height


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_note(ws, row, col_span, text):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1)
    c.font = FONT_NOTE
    c.alignment = ALIGN_L
    c.fill = FILL_LIGHT


# ============================================================
# 公共：10 个工具 sheet 的表头定义
# ============================================================
SHEET_DEFS = [
    # (sheet名, 标题, 副标题, 列宽列表, 表头, 行数, 说明)
    ("课程概览", "创新领导力 · 课程概览", "两天工作坊地图", [22, 18, 32, 28, 16], None, None),
    ("5因素诊断", "五个关键影响因素 · 团队创新健康度诊断", "心理安全 · 认知多样 · 探索空间 · 学习速度 · 领导者信号", [4, 24, 14, 14, 14, 14, 14, 12], None, None),
    ("客户洞察", "客户洞察 · 四层穿透", "任务 → 阻力 → 动力 → 背景", [4, 18, 40, 24, 18], None, None),
    ("知识流通", "知识流通 · 断裂点诊断", "交互涌现 5 步法", [4, 24, 32, 14, 26], None, None),
    ("最小实验", "最小可学习实验 · 敏捷迭代四步", "假设 → 最小验证 → 改变条件 → 公开学习", [4, 22, 36, 22], None, None),
    ("行为对照", "领导者行为对照盘点", "8 场景 · 运营管理者 vs 创新型领导者", [4, 22, 28, 28, 18], None, None),
    ("挑战卡", "我的创新挑战卡", "今晚作业 · 带回第二天分析", [4, 18, 50], None, None),
    ("30-60-90", "30-60-90 行动承诺卡", "课后行动规划", [4, 16, 16, 38, 24], None, None),
    ("团队巡检", "团队创新巡检台账", "月度跟踪 5 因素状态", [4, 18, 16, 16, 16, 16, 16, 16, 16, 24], None, None),
    ("学员评估", "学员评估 · 前测 + 后测", "课程效果对照", [4, 22, 16, 16, 16, 16, 24], None, None),
]


# ============================================================
# Sheet 1: 课程概览
# ============================================================
def build_overview(ws, mode="blank"):
    set_col_widths(ws, [22, 18, 32, 28, 16])
    style_title(ws, 1, 5, "创新领导力 · 两天工作坊课程概览")
    style_sub(ws, 2, 5, "你的学习地图：每一节你会做什么、带走什么")

    style_header_row(ws, 4, ["章节", "聚焦的核心问题", "你会做什么", "你会带走什么", "页码"])

    rows = [
        ("开场", "为什么努力推创新，反而越来越难？", "读案例，写第一反应", "一个需要重新思考的问题", ""),
        ("第一部分", "关于创新型团队，有哪些认知盲区？", "完成 8 题认知自测", "找到你和「创新真相」之间的差距", ""),
        ("第二部分", "哪五个因素决定团队能否持续创新？", "逐项诊断，填写评估表", "一份填好的团队创新健康度诊断表", ""),
        ("第三部分", "创新型领导和运营管理者有什么本质差异？", "行为盘点，识别无意间的创新抑制", "一份属于你的领导行为调整清单", ""),
        ("第四部分", "什么才是真正的客户洞察？", "用框架分析一个熟悉的失败案例", "一套可带回团队的客户洞察方法", ""),
        ("第五部分", "如何让「意外碰撞」在团队成为常态？", "绘制知识流通图，找出断裂点", "一份交互设计改善方案", ""),
        ("第一日收尾", "今天我最大的三个发现是什么？", "填写反思卡，领取今晚作业", "一张「创新挑战卡」", ""),
        ("第六部分", "敏捷迭代：为什么「快速失败」是误导？", "认知自测 + 案例分析", "一套真正的迭代思维框架", ""),
        ("第七部分", "管理者在迭代中扮演什么角色？", "把真实项目改写为最小可学习实验", "一个可带回执行的迭代方案", ""),
        ("第八部分", "三大要素：用框架分析自己的挑战", "对挑战卡进行三要素完整分析", "一份带着真实洞察的个人诊断", ""),
        ("情景模拟", "亮界科技：让你反复对照自己的案例", "小组诊断 + 方案设计", "一份完整的领导行为改变方案", ""),
        ("最终承诺", "我的 30-60-90 行动承诺", "写下 2 件 30 天内真的会做的事", "一张可贴在办公桌的承诺卡", ""),
    ]

    if mode == "example":
        # 填好示例：每个章节加 30 字左右的学员笔记
        notes = [
            "「赵建设十几秒的沉默」——我团队里也有这个瞬间",
            "5 题猜错 4 题，我以为鼓励创新是给自由时间",
            "心理安全感 12/20，分最低，这就是问题根源",
            "我经常在会议上先说自己的看法，要克制",
            "客户访谈只问了「你需要什么」，没问「你在做什么」",
            "销售和技术之间的信息断层，我们部门也存在",
            "决定挑战：新服务的实施阻力",
            "「允许不完美」这一条击中我了",
            "改成 2 周原型测试，原本 3 个月计划放弃",
            "三要素里交互涌现最弱",
            "张力最大的问题：原型的失败被压了 3 个月",
            "行动 1：每周例会加 5 分钟「让意外进来」",
        ]
        for i, ((sec, q, do, take, _), note) in enumerate(zip(rows, notes)):
            r = 5 + i
            style_data_row(ws, r, [sec, q, do, take, note], height=32, fill=FILL_HL if i in [0, 6, 11] else None)
            ws.cell(row=r, column=5).font = Font(name="Microsoft YaHei", size=9, color=COL_ACCENT2, italic=True)
    else:
        for i, (sec, q, do, take, _) in enumerate(rows):
            r = 5 + i
            style_data_row(ws, r, [sec, q, do, take, ""], height=32)

    add_note(ws, 18, 5, "💡 用法：边上课边在「页码」列记下对应 PDF 工具卡编号，方便课后回看。")


# ============================================================
# Sheet 2: 5 因素诊断
# ============================================================
def build_5factors(ws, mode="blank"):
    set_col_widths(ws, [4, 24, 14, 14, 14, 14, 14, 12])
    style_title(ws, 1, 8, "五因素诊断表 · 团队创新健康度")
    style_sub(ws, 2, 8, "5 个维度 · 各 4 题 · 自评 1-5 分 · 总分 100")

    headers = ["#", "诊断项目", "题 1", "题 2", "题 3", "题 4", "小计", "备注"]
    style_header_row(ws, 4, headers)

    factors = [
        ("心理安全感", [
            "团队成员会当面说还不成熟的想法",
            "异见会被认真讨论而不是终止",
            "成员敢公开承认不懂、不确定",
            "失败被作为学习机会认真对待",
        ], "地基：低则其他四项大打折扣"),
        ("认知多样性", [
            "团队有显著不同思维风格的成员",
            "与主流不同意见不会被快速淹没",
            "过去半年有「意外的人」的想法被认真对待",
            "我能说清每个人「思维方式最不同」之处",
        ], "燃料：决定新想法能否冒出来"),
        ("探索空间", [
            "成员有定期的非任务性学习时间",
            "探索活动不会因「项目紧急」被取消",
            "可对无明确答案的问题持续研究",
            "探索的洞察有渠道进入正式讨论",
        ], "苗圃：决定团队能否靠近未知边界"),
        ("学习速度", [
            "至少每月有复盘 + 具体行动结论",
            "失败项目会被作为学习案例认真讨论",
            "个人洞察有机制让其他成员知道",
            "执行中持续微调方向，不等项目结束",
        ], "引擎：决定创新能否加速"),
        ("领导者信号", [
            "我公开表彰过「有价值的失败学习」",
            "对异见的第一反应是好奇而非评判",
            "日程里留有非执行的探索/学习时间",
            "成员能举出我做的「让表达更安全」的事",
        ], "土壤：影响所有其他因素能否生长"),
    ]

    start_row = 5
    for fi, (fname, qs, hint) in enumerate(factors):
        r = start_row + fi * 6
        # 因素名 + 提示（合并 5 行）
        ws.cell(row=r, column=2, value=f"{fi+1}. {fname}")
        ws.merge_cells(start_row=r, start_column=2, end_row=r+4, end_column=2)
        c = ws.cell(row=r, column=2)
        c.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=[COL_ACCENT, COL_ACCENT2, COL_GREEN, COL_BROWN, COL_ACCENT][fi])
        c.alignment = ALIGN_C
        c.border = BORDER_MED
        for rr in range(r, r+5):
            ws.cell(row=rr, column=1).border = BORDER_MED

        # 4 题
        for qi, q in enumerate(qs):
            rr = r + qi
            ws.cell(row=rr, column=1, value=qi+1)
            ws.cell(row=rr, column=1).font = FONT_BOLD
            ws.cell(row=rr, column=1).alignment = ALIGN_C
            ws.cell(row=rr, column=1).border = BORDER_THIN
            for cc in range(3, 7):
                cell = ws.cell(row=rr, column=cc, value=(4 if mode == "example" and fi < 3 and qi < 3 else ""))
                cell.font = FONT_NORM
                cell.alignment = ALIGN_C
                cell.border = BORDER_THIN
                cell.fill = FILL_HL if (mode == "example" and fi < 3 and qi < 3) else FILL_LIGHT
            ws.row_dimensions[rr].height = 24

        # 小计列（合并 4 行）
        ws.cell(row=r, column=7, value=f"=SUM(C{r}:F{r})" if mode == "blank" else (qs and f"=SUM(C{r}:F{r})"))
        ws.merge_cells(start_row=r, start_column=7, end_row=r+3, end_column=7)
        c = ws.cell(row=r, column=7)
        c.font = Font(name="Microsoft YaHei", size=12, bold=True, color=COL_ACCENT)
        c.alignment = ALIGN_C
        c.border = BORDER_MED
        c.fill = FILL_HL
        for rr in range(r, r+4):
            ws.cell(row=rr, column=7).border = BORDER_MED
        # 备注列（合并 4 行）
        ws.cell(row=r, column=8, value=hint)
        ws.merge_cells(start_row=r, start_column=8, end_row=r+3, end_column=8)
        c = ws.cell(row=r, column=8)
        c.font = FONT_NOTE
        c.alignment = ALIGN_LT
        c.border = BORDER_MED
        for rr in range(r, r+4):
            ws.cell(row=rr, column=8).border = BORDER_MED

    # 总分行
    total_row = start_row + 5 * 6
    ws.cell(row=total_row, column=1, value="总分")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    c = ws.cell(row=total_row, column=1)
    c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    c.fill = FILL_TITLE
    c.alignment = ALIGN_C
    c.border = BORDER_MED
    ws.cell(row=total_row, column=2).border = BORDER_MED
    ws.cell(row=total_row, column=7, value=f"=SUM(G5:G{total_row-1})")
    c = ws.cell(row=total_row, column=7)
    c.font = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COL_ACCENT2)
    c.alignment = ALIGN_C
    c.border = BORDER_MED
    ws.cell(row=total_row, column=8, value="≥80 优秀 · 60-80 健康 · <60 紧急")
    ws.cell(row=total_row, column=8).font = Font(name="Microsoft YaHei", size=9, bold=True, color=COL_ACCENT2)
    ws.cell(row=total_row, column=8).alignment = ALIGN_C
    ws.cell(row=total_row, column=8).fill = FILL_HL
    ws.cell(row=total_row, column=8).border = BORDER_MED
    ws.row_dimensions[total_row].height = 30

    # 数据验证：1-5
    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
    dv.error = "请填写 1-5 之间的整数"
    dv.errorTitle = "评分超出范围"
    dv.prompt = "1=几乎不存在  5=非常普遍"
    dv.promptTitle = "评分提示"
    for fi in range(5):
        r = start_row + fi * 6
        dv.add(f"C{r}:F{r+3}")
    ws.add_data_validation(dv)

    # 条件格式：分数越低越红
    rule = ColorScaleRule(
        start_type='num', start_value=1, start_color='F8696B',
        mid_type='num', mid_value=3, mid_color='FFEB84',
        end_type='num', end_value=5, end_color='63BE7B'
    )
    ws.conditional_formatting.add(f"C5:F{total_row-1}", rule)

    note_row = total_row + 2
    add_note(ws, note_row, 8, "💡 最低分维度 = 你最该用力的地方。每题 1-5 分。1=几乎不存在，5=非常普遍。")
    add_note(ws, note_row + 1, 8, "⚠ 数据验证：每格只能填 1-5。颜色：红=低分（优先改进），绿=高分（健康）")


# ============================================================
# Sheet 3: 客户洞察四层穿透
# ============================================================
def build_customer_insight(ws, mode="blank"):
    set_col_widths(ws, [4, 18, 40, 24, 18])
    style_title(ws, 1, 5, "客户洞察 · 四层穿透")
    style_sub(ws, 2, 5, "任务 → 阻力 → 动力 → 背景 · 不要跳层")

    style_header_row(ws, 4, ["#", "洞察层次", "我们现在的认知是…", "这个认知来源于…", "第一手 / 二手？"])

    if mode == "example":
        # 用一个示例场景：某 B2B 供应链 SaaS 团队的智能采购时机预测
        example_rows = [
            ("任务层", "采购经理在决定何时下单时缺少数据支持",
             "团队成员的直觉 + 一份行业报告（2024）", "二手"),
            ("阻力层", "采购经理最大痛苦是与十几个供应商来回确认，比价谈判耗时 3-4 周",
             "小周访谈 28 名采购经理，整理 12 万字录音", "第一手"),
            ("动力层", "采购经理希望「让老板看到我高效且能控风险」",
             "小周访谈中 4 次出现「汇报」关键词", "第一手"),
            ("背景层", "每周四下午 2 点前要交采购进度报告给总监",
             "小周在 2 家客户现场观察 2 天", "第一手"),
        ]
    else:
        example_rows = [
            ("任务层", "", "", ""),
            ("阻力层", "", "", ""),
            ("动力层", "", "", ""),
            ("背景层", "", "", ""),
        ]

    colors = [COL_ACCENT, COL_ACCENT2, COL_GREEN, COL_BROWN]
    for i, (layer, content, source, first_hand) in enumerate(example_rows):
        r = 5 + i
        # 序号
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        # 层次（着色）
        ws.cell(row=r, column=2, value=layer)
        ws.cell(row=r, column=2).font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=colors[i])
        ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=2).border = BORDER_MED
        # 内容
        ws.cell(row=r, column=3, value=content)
        ws.cell(row=r, column=3).font = FONT_NORM
        ws.cell(row=r, column=3).alignment = ALIGN_LT
        ws.cell(row=r, column=3).border = BORDER_THIN
        ws.cell(row=r, column=3).fill = FILL_HL if mode == "example" else FILL_LIGHT
        # 来源
        ws.cell(row=r, column=4, value=source)
        ws.cell(row=r, column=4).font = FONT_NORM
        ws.cell(row=r, column=4).alignment = ALIGN_LT
        ws.cell(row=r, column=4).border = BORDER_THIN
        ws.cell(row=r, column=4).fill = FILL_HL if mode == "example" else FILL_LIGHT
        # 1手/2手
        ws.cell(row=r, column=5, value=first_hand)
        ws.cell(row=r, column=5).font = FONT_NORM
        ws.cell(row=r, column=5).alignment = ALIGN_C
        ws.cell(row=r, column=5).border = BORDER_THIN
        ws.cell(row=r, column=5).fill = FILL_HL if mode == "example" else FILL_LIGHT
        ws.row_dimensions[r].height = 60

    # 关键问题区
    note_r = 10
    ws.cell(row=note_r, column=1, value="最脆弱的一层")
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=2)
    ws.cell(row=note_r, column=1).font = FONT_BOLD
    ws.cell(row=note_r, column=1).fill = FILL_SECTION
    ws.cell(row=note_r, column=1).alignment = ALIGN_C
    ws.cell(row=note_r, column=1).border = BORDER_MED
    ws.cell(row=note_r, column=2).border = BORDER_MED
    ws.cell(row=note_r, column=3, value="阻力层 — 整份报告最强证据，但还没人提案")
    if mode == "example":
        ws.cell(row=note_r, column=3).fill = FILL_HL
    ws.cell(row=note_r, column=3).font = FONT_NORM
    ws.cell(row=note_r, column=3).alignment = ALIGN_LT
    ws.cell(row=note_r, column=3).border = BORDER_MED
    ws.merge_cells(start_row=note_r, start_column=3, end_row=note_r, end_column=5)
    for cc in range(3, 6):
        ws.cell(row=note_r, column=cc).border = BORDER_MED
    ws.row_dimensions[note_r].height = 24

    note_r = 11
    ws.cell(row=note_r, column=1, value="验证行动")
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=2)
    ws.cell(row=note_r, column=1).font = FONT_BOLD
    ws.cell(row=note_r, column=1).fill = FILL_SECTION
    ws.cell(row=note_r, column=1).alignment = ALIGN_C
    ws.cell(row=note_r, column=1).border = BORDER_MED
    ws.cell(row=note_r, column=2).border = BORDER_MED
    ws.cell(row=note_r, column=3, value="2 周内陪 3 个采购经理做一次完整采购流程，观察「比价谈判」环节")
    if mode == "example":
        ws.cell(row=note_r, column=3).fill = FILL_HL
    ws.cell(row=note_r, column=3).font = FONT_NORM
    ws.cell(row=note_r, column=3).alignment = ALIGN_LT
    ws.cell(row=note_r, column=3).border = BORDER_MED
    ws.merge_cells(start_row=note_r, start_column=3, end_row=note_r, end_column=5)
    for cc in range(3, 6):
        ws.cell(row=note_r, column=cc).border = BORDER_MED
    ws.row_dimensions[note_r].height = 24

    # 数据验证
    dv = DataValidation(type="list", formula1='"第一手,二手,混合"', allow_blank=True)
    dv.error = "请选择：第一手 / 二手 / 混合"
    ws.add_data_validation(dv)
    dv.add("E5:E8")

    add_note(ws, 13, 5, "💡 经典案例：奶昔的真实任务是「独自通勤时让手/嘴有事做，并撑到午饭」")
    add_note(ws, 14, 5, "⚠ 数据验证 E 列：第一手=直接观察/对话；二手=报告/转述")


# ============================================================
# Sheet 4: 知识流通
# ============================================================
def build_knowledge_flow(ws, mode="blank"):
    set_col_widths(ws, [4, 24, 32, 14, 26])
    style_title(ws, 1, 5, "知识流通 · 断裂点诊断")
    style_sub(ws, 2, 5, "交互涌现 5 步法 · 第一步：列出所有来源，评估流通状态")

    style_header_row(ws, 4, ["#", "信息/知识来源", "它能带来什么不同视角？", "目前状态", "阻碍是什么？"])

    if mode == "example":
        sources = [
            ("一线销售人员", "客户真实使用场景、当面拒绝的真实理由", "不畅", "销售报告经格式化整理，原始细节丢失"),
            ("客服工单", "高频问题清单、客户骂我们的原话", "断裂", "没有人定期看，所有投诉直接转给技术"),
            ("新员工", "「第一眼新鲜感」— 不被习惯遮蔽的视角", "断裂", "没有机制让他们发声，入职 3 个月后被同化"),
            ("竞品动态", "行业边界正在被重新定义的方向", "畅通", "产品经理每周看 1 次，但只看产品不看不相关的"),
            ("行业外类比", "其他行业如何解决相似问题", "断裂", "从来没有人看，团队没时间也没意识"),
            ("技术边界", "哪些事现在做不到，3 年内可能做到", "不畅", "技术评估只告诉「做不到」，不告诉「快做到了」"),
        ]
    else:
        sources = [
            ("一线销售人员", "", "畅通/不畅/断裂", ""),
            ("客户/用户（直接接触）", "", "畅通/不畅/断裂", ""),
            ("客户/用户（销售客服中转）", "", "畅通/不畅/断裂", ""),
            ("竞品动态", "", "畅通/不畅/断裂", ""),
            ("行业外类比", "", "畅通/不畅/断裂", ""),
            ("新员工「第一眼新鲜感」", "", "畅通/不畅/断裂", ""),
            ("技术边界认知", "", "畅通/不畅/断裂", ""),
        ]

    for i, (src, view, status, block) in enumerate(sources):
        r = 5 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        for col_i, val in enumerate([src, view, status, block], 2):
            c = ws.cell(row=r, column=col_i, value=val)
            c.font = FONT_NORM
            c.alignment = ALIGN_LT
            c.border = BORDER_THIN
            if mode == "example":
                c.fill = FILL_HL
            else:
                c.fill = FILL_LIGHT
        ws.row_dimensions[r].height = 30

    last_r = 5 + len(sources)

    # 断裂点分析
    r = last_r + 1
    ws.cell(row=r, column=1, value="最值得打通的断裂点")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1).font = FONT_BOLD
    ws.cell(row=r, column=1).fill = FILL_SECTION
    ws.cell(row=r, column=1).alignment = ALIGN_C
    ws.cell(row=r, column=1).border = BORDER_MED
    ws.cell(row=r, column=2).border = BORDER_MED
    val = "客服工单" if mode == "example" else ""
    ws.cell(row=r, column=3, value=val)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c = ws.cell(row=r, column=3)
    c.font = FONT_NORM
    c.alignment = ALIGN_LT
    c.border = BORDER_MED
    if mode == "example":
        c.fill = FILL_HL
    for cc in range(3, 6):
        ws.cell(row=r, column=cc).border = BORDER_MED
    ws.row_dimensions[r].height = 24

    r += 1
    ws.cell(row=r, column=1, value="最大现实障碍")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1).font = FONT_BOLD
    ws.cell(row=r, column=1).fill = FILL_SECTION
    ws.cell(row=r, column=1).alignment = ALIGN_C
    ws.cell(row=r, column=1).border = BORDER_MED
    ws.cell(row=r, column=2).border = BORDER_MED
    val = "客服经理担心「投诉被当众」会得罪客户，所以一直压着" if mode == "example" else ""
    ws.cell(row=r, column=3, value=val)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c = ws.cell(row=r, column=3)
    c.font = FONT_NORM
    c.alignment = ALIGN_LT
    c.border = BORDER_MED
    if mode == "example":
        c.fill = FILL_HL
    for cc in range(3, 6):
        ws.cell(row=r, column=cc).border = BORDER_MED
    ws.row_dimensions[r].height = 30

    r += 1
    ws.cell(row=r, column=1, value="设计的碰撞机制")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1).font = FONT_BOLD
    ws.cell(row=r, column=1).fill = FILL_SECTION
    ws.cell(row=r, column=1).alignment = ALIGN_C
    ws.cell(row=r, column=1).border = BORDER_MED
    ws.cell(row=r, column=2).border = BORDER_MED
    val = "每月 1 次「15 分钟客户原声会」：客服挑选 3 条最尖锐投诉，去客户化后读给团队听" if mode == "example" else ""
    ws.cell(row=r, column=3, value=val)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c = ws.cell(row=r, column=3)
    c.font = FONT_NORM
    c.alignment = ALIGN_LT
    c.border = BORDER_MED
    if mode == "example":
        c.fill = FILL_HL
    for cc in range(3, 6):
        ws.cell(row=r, column=cc).border = BORDER_MED
    ws.row_dimensions[r].height = 36

    # 数据验证
    dv = DataValidation(type="list", formula1='"畅通,不畅,断裂"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"D5:D{last_r}")

    r += 3
    add_note(ws, r, 5, "💡 Post-it 启示：便利贴是被「相遇」出来的，不是被发明出来的。管理者的角色 = 碰撞条件的设计师。")


# ============================================================
# Sheet 5: 最小可学习实验
# ============================================================
def build_mini_experiment(ws, mode="blank"):
    set_col_widths(ws, [4, 22, 36, 22])
    style_title(ws, 1, 4, "最小可学习实验 · 敏捷迭代四步")
    style_sub(ws, 2, 4, "假设 → 最小验证 → 预定义改变条件 → 公开学习")

    style_header_row(ws, 4, ["#", "步骤", "你的填写", "自我检查"])

    if mode == "example":
        rows = [
            ("① 核心假设", "采购经理目前用 Excel 管理供应商，但有意愿付费让系统替代", "如果错了：项目方向不成立"),
            ("② 假设类型", "行为假设（客户愿不愿意改变操作习惯）", "□ 客户  □ 行为  □ 可行性  □ 商业"),
            ("③ 当前推进方式", "3 个月开发路线图，按计划推进，上线后看数据", "通常意味着 3 个月后才知道对错"),
            ("④ 最小验证设计", "找 5 个目标客户，展示无功能的界面原型 + 询问是否付费", "≤ 2 周 + ≤ 10% 预算"),
            ("⑤ 预定义改变条件", "如果 5 人里 3 人说「我用 Excel 就够了」，立即重新想方向", "提前写，对抗确认偏误"),
            ("⑥ 公开学习计划", "2 周后月度例会，对全团队分享测试结果 + 方向调整", "学习是正式工作内容"),
        ]
    else:
        rows = [
            ("① 核心假设", "", ""),
            ("② 假设类型", "", ""),
            ("③ 当前推进方式", "", ""),
            ("④ 最小验证设计", "", ""),
            ("⑤ 预定义改变条件", "", ""),
            ("⑥ 公开学习计划", "", ""),
        ]

    for i, (step, content, check) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        ws.cell(row=r, column=2, value=step)
        ws.cell(row=r, column=2).font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COL_ACCENT)
        ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=2).border = BORDER_MED
        ws.cell(row=r, column=3, value=content)
        ws.cell(row=r, column=3).font = FONT_NORM
        ws.cell(row=r, column=3).alignment = ALIGN_LT
        ws.cell(row=r, column=3).border = BORDER_THIN
        ws.cell(row=r, column=3).fill = FILL_HL if mode == "example" else FILL_LIGHT
        ws.cell(row=r, column=4, value=check)
        ws.cell(row=r, column=4).font = FONT_NOTE
        ws.cell(row=r, column=4).alignment = ALIGN_LT
        ws.cell(row=r, column=4).border = BORDER_THIN
        ws.row_dimensions[r].height = 50

    add_note(ws, 12, 4, "💡 关键认知：迭代≠「小步快跑」≠「失败了也没关系」。是「行动前知道在测试什么，行动后知道学到了什么」。")
    add_note(ws, 13, 4, "⚠ 最小验证的时间约束：≤ 2 周 + ≤ 原预算 10%。超过这个范围，就回到了「计划执行」。")


# ============================================================
# Sheet 6: 行为对照盘点
# ============================================================
def build_behavior_compare(ws, mode="blank"):
    set_col_widths(ws, [4, 22, 28, 28, 18])
    style_title(ws, 1, 5, "领导者行为对照盘点")
    style_sub(ws, 2, 5, "8 场景 · 过去 30 天里你实际怎么做？")

    style_header_row(ws, 4, ["#", "场景", "你当时的实际反应（过去 30 天）", "你的倾向", "想改变吗？"])

    if mode == "example":
        rows = [
            ("有人提出不成熟的想法", "「这个先做个可行性报告，下周给我」", "运营管理者", "是"),
            ("有项目/实验遇到挫折", "「谁负责的？下不为例」", "运营管理者", "是"),
            ("讨论不确定能不能行的方向", "「我们先推进主线，这个以后再说」", "运营管理者", "是"),
            ("有人提出和你不同的判断", "「你说的也有道理，但现在不是时候」", "运营管理者", "是"),
            ("季度总结/汇报", "「表彰 5 个成功项目，分析成功经验」", "运营管理者", "是"),
            ("招募新成员", "「找能力强、方向匹配的人」", "运营管理者", "稍后"),
            ("资源分配", "「给有明确 ROI 的项目」", "运营管理者", "是"),
            ("会上先说自己的看法", "「我先抛个砖：我觉得应该 X」", "运营管理者", "是"),
        ]
    else:
        rows = [
            ("有人提出不成熟的想法", "", "运营管理者/创新型/视情况", ""),
            ("有项目/实验遇到挫折", "", "运营管理者/创新型/视情况", ""),
            ("讨论不确定能不能行的方向", "", "运营管理者/创新型/视情况", ""),
            ("有人提出和你不同的判断", "", "运营管理者/创新型/视情况", ""),
            ("季度总结/汇报", "", "运营管理者/创新型/视情况", ""),
            ("招募新成员", "", "运营管理者/创新型/视情况", ""),
            ("资源分配", "", "运营管理者/创新型/视情况", ""),
            ("会上先说自己的看法", "", "运营管理者/创新型/视情况", ""),
        ]

    for i, (scene, actual, tendency, change) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        ws.cell(row=r, column=2, value=scene)
        ws.cell(row=r, column=2).font = FONT_BOLD
        ws.cell(row=r, column=2).alignment = ALIGN_LT
        ws.cell(row=r, column=2).border = BORDER_THIN
        ws.cell(row=r, column=3, value=actual)
        ws.cell(row=r, column=3).font = FONT_NORM
        ws.cell(row=r, column=3).alignment = ALIGN_LT
        ws.cell(row=r, column=3).border = BORDER_THIN
        ws.cell(row=r, column=3).fill = FILL_HL if mode == "example" else FILL_LIGHT
        ws.cell(row=r, column=4, value=tendency)
        ws.cell(row=r, column=4).font = FONT_NORM
        ws.cell(row=r, column=4).alignment = ALIGN_C
        ws.cell(row=r, column=4).border = BORDER_THIN
        if mode == "example":
            ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=COL_RED)
            ws.cell(row=r, column=4).font = Font(name="Microsoft YaHei", size=9, color="FFFFFF", bold=True)
        ws.cell(row=r, column=5, value=change)
        ws.cell(row=r, column=5).font = FONT_NORM
        ws.cell(row=r, column=5).alignment = ALIGN_C
        ws.cell(row=r, column=5).border = BORDER_THIN
        if mode == "example":
            ws.cell(row=r, column=5).fill = FILL_HL
        ws.row_dimensions[r].height = 32

    # 想改变的一件
    r = 14
    ws.cell(row=r, column=1, value="最想改变的一件")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1).font = FONT_BOLD
    ws.cell(row=r, column=1).fill = FILL_SECTION
    ws.cell(row=r, column=1).alignment = ALIGN_C
    ws.cell(row=r, column=1).border = BORDER_MED
    ws.cell(row=r, column=2).border = BORDER_MED
    val = "会上先说自己的看法 — 改成「我最后说，先听大家的」" if mode == "example" else ""
    ws.cell(row=r, column=3, value=val)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c = ws.cell(row=r, column=3)
    c.font = FONT_NORM
    c.alignment = ALIGN_LT
    c.border = BORDER_MED
    if mode == "example":
        c.fill = FILL_HL
    for cc in range(3, 6):
        ws.cell(row=r, column=cc).border = BORDER_MED
    ws.row_dimensions[r].height = 32

    r = 15
    ws.cell(row=r, column=1, value="改变后希望的不同")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1).font = FONT_BOLD
    ws.cell(row=r, column=1).fill = FILL_SECTION
    ws.cell(row=r, column=1).alignment = ALIGN_C
    ws.cell(row=r, column=1).border = BORDER_MED
    ws.cell(row=r, column=2).border = BORDER_MED
    val = "团队开始主动提出和我不同的方向，不再等我先说" if mode == "example" else ""
    ws.cell(row=r, column=3, value=val)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c = ws.cell(row=r, column=3)
    c.font = FONT_NORM
    c.alignment = ALIGN_LT
    c.border = BORDER_MED
    if mode == "example":
        c.fill = FILL_HL
    for cc in range(3, 6):
        ws.cell(row=r, column=cc).border = BORDER_MED
    ws.row_dimensions[r].height = 32

    # 数据验证
    dv1 = DataValidation(type="list", formula1='"运营管理者,创新型领导者,视情况而定"', allow_blank=True)
    ws.add_data_validation(dv1)
    dv1.add("D5:D12")
    dv2 = DataValidation(type="list", formula1='"是,否,稍后"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add("E5:E12")


# ============================================================
# Sheet 7: 挑战卡
# ============================================================
def build_challenge(ws, mode="blank"):
    set_col_widths(ws, [4, 18, 50])
    style_title(ws, 1, 3, "我的创新挑战卡")
    style_sub(ws, 2, 3, "今晚作业 · 带回第二天分析")

    if mode == "example":
        rows = [
            ("挑战一句话", "团队正在推新的客户服务模式，3 个月了一线员工配合度很低"),
            ("已经做过什么", "已经做了 3 个月的方案设计 + 2 次全员宣导，效果都不明显"),
            ("目前最大障碍", "1) 一线员工不相信新模式能减负  2) 担心新模式增加考核压力  3) 缺少试点的成功案例"),
            ("最希望找到的答案", "是这个模式本身有问题，还是我们的实施方式没找到对的？"),
        ]
    else:
        rows = [
            ("挑战一句话", ""),
            ("已经做过什么", ""),
            ("目前最大障碍", ""),
            ("最希望找到的答案", ""),
        ]

    style_header_row(ws, 4, ["#", "项目", "你的填写"])

    for i, (label, content) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=2).font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COL_ACCENT)
        ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=2).border = BORDER_MED
        ws.cell(row=r, column=3, value=content)
        ws.cell(row=r, column=3).font = FONT_NORM
        ws.cell(row=r, column=3).alignment = ALIGN_LT
        ws.cell(row=r, column=3).border = BORDER_MED
        if mode == "example":
            ws.cell(row=r, column=3).fill = FILL_HL
        else:
            ws.cell(row=r, column=3).fill = FILL_LIGHT
        ws.row_dimensions[r].height = 50

    # 对照判断
    r = 10
    style_header_row(ws, r, ["#", "维度", "对照判断（圈选）"], fill=PatternFill("solid", fgColor=COL_ACCENT2), font=FONT_H)
    if mode == "example":
        judgments = [
            ("客户洞察", "有待加强 — 阻力的二手数据多于一手"),
            ("交互涌现", "有几个断裂点 — 客服和销售之间不通"),
            ("敏捷迭代", "偶尔在做 — 没有预定义改变条件"),
        ]
    else:
        judgments = [
            ("客户洞察", "足够扎实 / 有待加强 / 几乎是盲区"),
            ("交互涌现", "足够顺畅 / 有几个断裂点 / 基本封闭"),
            ("敏捷迭代", "已经在做 / 偶尔在做 / 几乎没有"),
        ]
    for i, (dim, judge) in enumerate(judgments):
        r = 11 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        ws.cell(row=r, column=2, value=dim)
        ws.cell(row=r, column=2).font = FONT_BOLD
        ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=2).border = BORDER_MED
        ws.cell(row=r, column=3, value=judge)
        ws.cell(row=r, column=3).font = FONT_NORM
        ws.cell(row=r, column=3).alignment = ALIGN_LT
        ws.cell(row=r, column=3).border = BORDER_MED
        if mode == "example":
            ws.cell(row=r, column=3).fill = FILL_HL
        else:
            ws.cell(row=r, column=3).fill = FILL_LIGHT
        ws.row_dimensions[r].height = 28

    add_note(ws, 15, 3, "💡 选题标准：你真的在乎 + 不要太宏观 + 不要太微小 + 脑子里有具体人/情境/卡点")
    add_note(ws, 16, 3, "⚠ 三要素对照是为了给第二天的诊断做铺垫——不必纠结判断准不准，重点是写下来。")


# ============================================================
# Sheet 8: 30-60-90 行动承诺
# ============================================================
def build_action_commitment(ws, mode="blank"):
    set_col_widths(ws, [4, 16, 16, 38, 24])
    style_title(ws, 1, 5, "30-60-90 行动承诺卡")
    style_sub(ws, 2, 5, "课后行动规划 · 只写 80% 把握在 30 天内真的做到的事")

    style_header_row(ws, 4, ["#", "时间窗口", "这件事的名称", "具体描述（场合/对谁/做什么）", "会传递什么新信号？"])

    if mode == "example":
        rows = [
            ("30 天内", "「让意外进来」", "下周四产品周会开场 5 分钟，对全团队说：「有没有人最近听到了什么让你意外的客户反馈？不管有没有结论，都可以说」", "管理者真的想听到新东西，不是等汇报"),
            ("30 天内", "「失败案例复盘」", "下月全员会上做一次对失败项目的复盘（不是总结成功），公开表彰「最有价值的失败学习」", "失败不再是应该被翻篇的事"),
            ("60 天内", "「15 分钟怪事分享」", "每两周 1 次，每个人带 1 件「上两周让我意外的事」", "碰撞条件从无到有"),
            ("90 天内", "「异质声音保护」", "每次重要讨论结束前，主动问：「谁有不一样的看法还没说？」", "不同意见被真正听见"),
        ]
    else:
        rows = [
            ("30 天内", "", "", ""),
            ("30 天内", "", "", ""),
            ("60 天内", "", "", ""),
            ("90 天内", "", "", ""),
        ]

    colors = [COL_ACCENT, COL_ACCENT, COL_GREEN, COL_ACCENT2]
    for i, (period, name, desc, signal) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        ws.cell(row=r, column=2, value=period)
        ws.cell(row=r, column=2).font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=colors[i])
        ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=2).border = BORDER_MED
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=3).font = FONT_BOLD
        ws.cell(row=r, column=3).alignment = ALIGN_LT
        ws.cell(row=r, column=3).border = BORDER_MED
        ws.cell(row=r, column=4, value=desc)
        ws.cell(row=r, column=4).font = FONT_NORM
        ws.cell(row=r, column=4).alignment = ALIGN_LT
        ws.cell(row=r, column=4).border = BORDER_MED
        if mode == "example":
            ws.cell(row=r, column=4).fill = FILL_HL
        else:
            ws.cell(row=r, column=4).fill = FILL_LIGHT
        ws.cell(row=r, column=5, value=signal)
        ws.cell(row=r, column=5).font = FONT_NORM
        ws.cell(row=r, column=5).alignment = ALIGN_LT
        ws.cell(row=r, column=5).border = BORDER_MED
        if mode == "example":
            ws.cell(row=r, column=5).fill = FILL_HL
        else:
            ws.cell(row=r, column=5).fill = FILL_LIGHT
        ws.row_dimensions[r].height = 60

    # 跟进 + 问责
    r = 10
    style_header_row(ws, r, ["#", "跟进方式", "描述", "30 天后追问我的人", "联系方式"], fill=PatternFill("solid", fgColor=COL_ACCENT2), font=FONT_H)
    if mode == "example":
        rows2 = [
            ("每月 1 次自我复盘", "在每月 1 号的上午，我会用 30 分钟回顾上月的行动承诺", "王经理（同组同事）", "微信"),
            ("季度大复盘", "每季度末我会写一份 1 页的行动总结", "我自己（一年后）", "邮件"),
        ]
    else:
        rows2 = [
            ("", "", "", ""),
            ("", "", "", ""),
        ]
    for i, (way, desc, person, contact) in enumerate(rows2):
        r = 11 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        ws.cell(row=r, column=2, value=way)
        ws.cell(row=r, column=2).font = FONT_BOLD
        ws.cell(row=r, column=2).alignment = ALIGN_LT
        ws.cell(row=r, column=2).border = BORDER_MED
        ws.cell(row=r, column=3, value=desc)
        ws.cell(row=r, column=3).font = FONT_NORM
        ws.cell(row=r, column=3).alignment = ALIGN_LT
        ws.cell(row=r, column=3).border = BORDER_MED
        if mode == "example":
            ws.cell(row=r, column=3).fill = FILL_HL
        else:
            ws.cell(row=r, column=3).fill = FILL_LIGHT
        ws.cell(row=r, column=4, value=person)
        ws.cell(row=r, column=4).font = FONT_NORM
        ws.cell(row=r, column=4).alignment = ALIGN_LT
        ws.cell(row=r, column=4).border = BORDER_MED
        if mode == "example":
            ws.cell(row=r, column=4).fill = FILL_HL
        else:
            ws.cell(row=r, column=4).fill = FILL_LIGHT
        ws.cell(row=r, column=5, value=contact)
        ws.cell(row=r, column=5).font = FONT_NORM
        ws.cell(row=r, column=5).alignment = ALIGN_LT
        ws.cell(row=r, column=5).border = BORDER_MED
        if mode == "example":
            ws.cell(row=r, column=5).fill = FILL_HL
        else:
            ws.cell(row=r, column=5).fill = FILL_LIGHT
        ws.row_dimensions[r].height = 40

    add_note(ws, 14, 5, "💡 写下 10 件没一件做，不如写下 2 件都做成。测试问题：5 年后一个团队成员被问「你的管理者做的第一件不同的事」，能清楚说出来吗？")


# ============================================================
# Sheet 9: 团队创新巡检台账
# ============================================================
def build_inspection_log(ws, mode="blank"):
    set_col_widths(ws, [4, 18, 14, 14, 14, 14, 14, 14, 14, 24])
    style_title(ws, 1, 10, "团队创新巡检台账")
    style_sub(ws, 2, 10, "每月跟踪 5 因素状态 · 用作定期回顾")

    headers = ["#", "巡检月份", "心理安全", "认知多样", "探索空间", "学习速度", "领导信号", "总分", "最高分维度", "最低分维度（重点改进）"]
    style_header_row(ws, 4, headers)

    if mode == "example":
        rows = [
            ("2026-06", 12, 10, 8, 9, 11, "领导信号", "探索空间"),
            ("2026-07", 13, 11, 9, 10, 12, "领导信号", "探索空间"),
            ("2026-08", 14, 12, 10, 11, 13, "领导信号", "探索空间"),
        ]
    else:
        rows = [
            ("2026-__", "", "", "", "", "", "", "", ""),
            ("2026-__", "", "", "", "", "", "", "", ""),
            ("2026-__", "", "", "", "", "", "", "", ""),
            ("2026-__", "", "", "", "", "", "", "", ""),
            ("2026-__", "", "", "", "", "", "", "", ""),
        ]

    for i, row_data in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        # 月份
        ws.cell(row=r, column=2, value=row_data[0])
        ws.cell(row=r, column=2).font = FONT_BOLD
        ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=2).border = BORDER_MED
        # 5 个分数
        for j in range(5):
            cc = 3 + j
            v = row_data[1 + j] if j < 5 else ""
            c = ws.cell(row=r, column=cc, value=v)
            c.font = FONT_NORM
            c.alignment = ALIGN_C
            c.border = BORDER_THIN
            if mode == "example":
                c.fill = FILL_HL
            else:
                c.fill = FILL_LIGHT
        # 总分（公式）
        c = ws.cell(row=r, column=8, value=f"=SUM(C{r}:G{r})")
        c.font = Font(name="Microsoft YaHei", size=10, bold=True, color=COL_ACCENT)
        c.alignment = ALIGN_C
        c.border = BORDER_MED
        c.fill = FILL_HL
        # 最高/最低维度（占位）
        ws.cell(row=r, column=9, value=row_data[6])
        ws.cell(row=r, column=9).font = FONT_NORM
        ws.cell(row=r, column=9).alignment = ALIGN_C
        ws.cell(row=r, column=9).border = BORDER_THIN
        if mode == "example":
            ws.cell(row=r, column=9).fill = PatternFill("solid", fgColor=COL_GREEN)
            ws.cell(row=r, column=9).font = Font(name="Microsoft YaHei", size=9, color="FFFFFF", bold=True)
        ws.cell(row=r, column=10, value=row_data[7])
        ws.cell(row=r, column=10).font = FONT_NORM
        ws.cell(row=r, column=10).alignment = ALIGN_C
        ws.cell(row=r, column=10).border = BORDER_THIN
        if mode == "example":
            ws.cell(row=r, column=10).fill = PatternFill("solid", fgColor=COL_RED)
            ws.cell(row=r, column=10).font = Font(name="Microsoft YaHei", size=9, color="FFFFFF", bold=True)
        ws.row_dimensions[r].height = 24

    last_r = 5 + len(rows) - 1

    # 数据验证
    dv = DataValidation(type="whole", operator="between", formula1=0, formula2=20, allow_blank=True)
    dv.error = "请填 0-20 之间的整数"
    ws.add_data_validation(dv)
    dv.add(f"C5:G{last_r}")

    # 条件格式
    rule = ColorScaleRule(
        start_type='num', start_value=4, start_color='F8696B',
        mid_type='num', mid_value=12, mid_color='FFEB84',
        end_type='num', end_value=20, end_color='63BE7B'
    )
    ws.conditional_formatting.add(f"C5:G{last_r}", rule)

    add_note(ws, last_r + 2, 10, "💡 用法：每月月底花 5 分钟，对照上月的 5 因素做一次快速自评；连续 3 个月对比看趋势。")
    add_note(ws, last_r + 3, 10, "⚠ 数据验证：每格 0-20（5 题 × 4 分）。颜色：红=低分（优先改进），绿=高分（健康）")

    # 雷达图（example 模式）
    if mode == "example" and len(rows) >= 1:
        chart = RadarChart()
        chart.type = "filled"
        chart.style = 26
        chart.title = "5 因素健康度雷达图（最近一个月）"
        chart.y_axis.delete = True

        # 数据
        data = Reference(ws, min_col=3, max_col=7, min_row=4, max_row=5)
        cats = Reference(ws, min_col=3, max_col=7, min_row=4, max_row=4)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"B{last_r + 6}")


# ============================================================
# Sheet 10: 学员评估
# ============================================================
def build_student_assessment(ws, mode="blank"):
    set_col_widths(ws, [4, 22, 16, 16, 16, 16, 24])
    style_title(ws, 1, 7, "学员评估 · 前测 + 后测对照")
    style_sub(ws, 2, 7, "1=完全不同意 · 5=完全同意 · 用同一份题测两次看变化")

    style_header_row(ws, 4, ["#", "评估项目", "前测", "后测", "变化", "是否达标", "备注"])

    items = [
        "我能清楚说出团队当前的「心理安全感」水平",
        "我能在 5 个影响团队创新的因素里诊断最弱的一项",
        "我能识别出自己无意识的「创新抑制」行为",
        "我能在「运营管理」和「创新型领导」之间切换",
        "我理解四层客户洞察（任务/阻力/动力/背景）",
        "我能设计一个 2 周内可验证的最小实验",
        "我团队过去 6 个月内至少出现 1 个「意外惊喜」",
        "我能在例会开场创造 5 分钟「让意外进来」",
        "我有清晰的下一步行动承诺（30 天内）",
        "我能说出团队最值得打通的「知识断裂点」",
    ]

    if mode == "example":
        # 填好后测有提升
        pre = [3, 2, 2, 3, 2, 2, 2, 2, 2, 3]
        post = [5, 4, 4, 4, 5, 4, 3, 4, 5, 4]
    else:
        pre = [""] * 10
        post = [""] * 10

    for i, (item, p, o) in enumerate(zip(items, pre, post)):
        r = 5 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        ws.cell(row=r, column=2, value=item)
        ws.cell(row=r, column=2).font = FONT_NORM
        ws.cell(row=r, column=2).alignment = ALIGN_LT
        ws.cell(row=r, column=2).border = BORDER_MED
        # 前测
        c = ws.cell(row=r, column=3, value=p)
        c.font = FONT_NORM
        c.alignment = ALIGN_C
        c.border = BORDER_THIN
        c.fill = FILL_HL if mode == "example" else FILL_LIGHT
        # 后测
        c = ws.cell(row=r, column=4, value=o)
        c.font = Font(name="Microsoft YaHei", size=10, bold=True, color=COL_GREEN)
        c.alignment = ALIGN_C
        c.border = BORDER_THIN
        c.fill = FILL_HL if mode == "example" else FILL_LIGHT
        # 变化（公式）
        c = ws.cell(row=r, column=5, value=f"=D{r}-C{r}")
        c.font = Font(name="Microsoft YaHei", size=10, bold=True, color=COL_ACCENT)
        c.alignment = ALIGN_C
        c.border = BORDER_MED
        c.fill = FILL_HL
        # 是否达标
        c = ws.cell(row=r, column=6, value=f'=IF(D{r}-C{r}>=2,"✓","")')
        c.font = Font(name="Microsoft YaHei", size=10, bold=True, color=COL_GREEN)
        c.alignment = ALIGN_C
        c.border = BORDER_MED
        c.fill = FILL_HL
        # 备注
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=7).border = BORDER_THIN
        ws.row_dimensions[r].height = 22

    # 总分
    r = 16
    ws.cell(row=r, column=1, value="总分")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1).font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    ws.cell(row=r, column=1).fill = FILL_TITLE
    ws.cell(row=r, column=1).alignment = ALIGN_C
    ws.cell(row=r, column=1).border = BORDER_MED
    ws.cell(row=r, column=2).border = BORDER_MED
    ws.cell(row=r, column=3, value=f"=SUM(C5:C14)")
    c = ws.cell(row=r, column=3)
    c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    c.fill = FILL_TITLE
    c.alignment = ALIGN_C
    c.border = BORDER_MED
    ws.cell(row=r, column=4, value=f"=SUM(D5:D14)")
    c = ws.cell(row=r, column=4)
    c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COL_GREEN)
    c.alignment = ALIGN_C
    c.border = BORDER_MED
    ws.cell(row=r, column=5, value=f"=D{r}-C{r}")
    c = ws.cell(row=r, column=5)
    c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COL_ACCENT2)
    c.alignment = ALIGN_C
    c.border = BORDER_MED
    for cc in [6, 7]:
        c = ws.cell(row=r, column=cc, value="≥10 优秀学习成果" if cc == 6 else "")
        c.font = Font(name="Microsoft YaHei", size=9, bold=True, color=COL_ACCENT2)
        c.fill = FILL_HL
        c.alignment = ALIGN_C
        c.border = BORDER_MED
    ws.row_dimensions[r].height = 30

    # 数据验证
    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
    dv.error = "请填 1-5 之间的整数"
    ws.add_data_validation(dv)
    dv.add("C5:D14")

    add_note(ws, 18, 7, "💡 用法：开课前 5 分钟做前测；课程结束当天做后测；总分变化 ≥ 10 分表明学习效果显著。")


# ============================================================
# Sheet 0 (使用指引第一页): 课程工具地图
# ============================================================
def build_guide(ws, mode):
    """使用指引第一页"""
    set_col_widths(ws, [4, 26, 38, 18, 14])
    style_title(ws, 1, 5, "创新领导力 · 配套表单使用指引")
    style_sub(ws, 2, 5, "10 个工具表单 · 使用场景 + 填写步骤 + 注意事项")

    style_header_row(ws, 4, ["#", "工具名称", "使用场景与填写步骤", "何时使用", "预估时长"])

    tools = [
        ("课程概览", "完整记录两天学习地图；每节课程后补充个人笔记", "课前 5 分钟", "5 分钟"),
        ("5 因素诊断", "对当前团队按 5 个维度（心理安全/认知多样/探索空间/学习速度/领导信号）逐题自评 1-5 分", "第一天上午后", "10 分钟"),
        ("客户洞察", "选一个真实场景，依任务/阻力/动力/背景四层顺序填写，不要跳层", "第一天下午", "20 分钟"),
        ("知识流通", "列出团队所有信息/知识来源，标状态（畅通/不畅/断裂），找最有价值的断裂点", "第一天下午", "15 分钟"),
        ("最小实验", "把真实项目改写为「假设→最小验证→预定义改变条件→公开学习」", "第二天上午", "18 分钟"),
        ("行为对照", "回忆过去 30 天的 5-8 个具体事件，对照「运营管理 vs 创新型」表格判断倾向", "第一天上午", "12 分钟"),
        ("挑战卡", "选 1 个你真正在乎的创新挑战，描述卡点 + 列出 2-3 个具体障碍", "第一晚作业", "15-20 分钟"),
        ("30-60-90", "课后 2 件 30 天内会做的事 + 1 件 60 天 + 1 件 90 天，要具体到场合和人", "第二天结课时", "15 分钟"),
        ("团队巡检", "每月月底花 5 分钟对 5 因素做一次自评，看 3 个月趋势", "每月月底", "5 分钟"),
        ("学员评估", "前测 + 后测两份同一套题，10 题各 1-5 分，看总分变化", "开课前 + 结课时", "10 分钟"),
    ]

    for i, (name, scene, when, dur) in enumerate(tools):
        r = 5 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=2).font = Font(name="Microsoft YaHei", size=10, bold=True, color=COL_ACCENT)
        ws.cell(row=r, column=2).alignment = ALIGN_LT
        ws.cell(row=r, column=2).border = BORDER_MED
        ws.cell(row=r, column=2).fill = FILL_SECTION
        ws.cell(row=r, column=3, value=scene)
        ws.cell(row=r, column=3).font = FONT_NORM
        ws.cell(row=r, column=3).alignment = ALIGN_LT
        ws.cell(row=r, column=3).border = BORDER_MED
        ws.cell(row=r, column=3).fill = FILL_LIGHT
        ws.cell(row=r, column=4, value=when)
        ws.cell(row=r, column=4).font = FONT_BOLD
        ws.cell(row=r, column=4).alignment = ALIGN_C
        ws.cell(row=r, column=4).border = BORDER_MED
        ws.cell(row=r, column=4).fill = FILL_HL
        ws.cell(row=r, column=5, value=dur)
        ws.cell(row=r, column=5).font = FONT_BOLD
        ws.cell(row=r, column=5).alignment = ALIGN_C
        ws.cell(row=r, column=5).border = BORDER_MED
        ws.row_dimensions[r].height = 36

    add_note(ws, 16, 5, "💡 推荐使用顺序：课程概览 → 5 因素诊断 → 客户洞察 → 知识流通 → 挑战卡 → 最小实验 → 行为对照 → 30-60-90 → 学员评估 → 团队巡检")
    add_note(ws, 17, 5, "⚠ 全部表单都在同一个 Excel 文件的 10 个 sheet 里。3 个 Excel 文件的区别：01-使用指引（本表） / 02-空表（自己填写） / 03-填好示例（看真实场景）")


# ============================================================
# Sheet (使用指引第二页): 通用填写原则
# ============================================================
def build_guide_principles(ws):
    set_col_widths(ws, [4, 22, 60])
    style_title(ws, 1, 3, "通用填写原则 · 5 条核心")
    style_sub(ws, 2, 3, "无论填哪张表都适用的心法")

    style_header_row(ws, 4, ["#", "原则", "具体说明"])

    principles = [
        ("足够具体", "「我想做得更好」不是答案。必须具体到「什么场合 / 对谁 / 做什么 / 多久一次」。判断标准：5 年后一个团队成员能复述吗？"),
        ("诚实胜于正确", "5 因素诊断的目的是找问题，不是评功。最低分维度才是你真正该用力的地方。"),
        ("依赖第一手", "客户洞察里「第一手」≥ 60% 才算扎实。二手报告再多也是猜测——做几次现场观察胜过 10 份报告。"),
        ("写下来 ≠ 做到了", "30-60-90 行动承诺要写到「你愿意 30 天后让一个人来追问你」的程度。写下 10 件没一件做，不如写下 2 件都做成。"),
        ("每月复盘", "团队巡检台账不是填一次就完事。每月月底 5 分钟回顾，连续 3 个月对比看趋势，趋势比单点更重要。"),
    ]

    for i, (p, desc) in enumerate(principles):
        r = 5 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        ws.cell(row=r, column=2, value=p)
        ws.cell(row=r, column=2).font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COL_ACCENT)
        ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=2).border = BORDER_MED
        ws.cell(row=r, column=3, value=desc)
        ws.cell(row=r, column=3).font = FONT_NORM
        ws.cell(row=r, column=3).alignment = ALIGN_LT
        ws.cell(row=r, column=3).border = BORDER_MED
        ws.cell(row=r, column=3).fill = FILL_LIGHT
        ws.row_dimensions[r].height = 50

    # 常见问题
    r = 11
    style_header_row(ws, r, ["#", "常见问题", "回答"], fill=PatternFill("solid", fgColor=COL_ACCENT2), font=FONT_H)

    faqs = [
        ("诊断结果是中等（60-80 分），还要做吗？", "中等 = 还有改进空间。最低分维度才是关键。整体平均分会被高分维度拉高，掩盖真正的问题。"),
        ("挑战卡写得太抽象怎么办？", "用「具体的人/具体的场景/具体的卡点」重新描述。问自己：这件事你能讲给同事听 5 分钟吗？"),
        ("最小实验如果失败了怎么办？", "失败是预期之内的。关键是「行动前知道在测试什么，行动后知道学到了什么」。预定义改变条件不是失败，是进步。"),
        ("30-60-90 写完后会忘，怎么办？", "找一个会 30 天后追问你的人。写下他的名字，承诺就成立了。"),
        ("可以多人共用一个 Excel 吗？", "推荐每人一份独立文件。诊断/挑战/承诺是个人反思，共用会变成汇报素材，丧失反思价值。"),
    ]
    for i, (q, a) in enumerate(faqs):
        r = 12 + i
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=1).border = BORDER_MED
        ws.cell(row=r, column=2, value=q)
        ws.cell(row=r, column=2).font = FONT_BOLD
        ws.cell(row=r, column=2).alignment = ALIGN_LT
        ws.cell(row=r, column=2).border = BORDER_MED
        ws.cell(row=r, column=2).fill = FILL_SECTION
        ws.cell(row=r, column=3, value=a)
        ws.cell(row=r, column=3).font = FONT_NORM
        ws.cell(row=r, column=3).alignment = ALIGN_LT
        ws.cell(row=r, column=3).border = BORDER_MED
        ws.cell(row=r, column=3).fill = FILL_LIGHT
        ws.row_dimensions[r].height = 50


# ============================================================
# 主流程
# ============================================================
def build_workbook(filename, mode="blank"):
    """构建一个完整的 Excel 工作簿"""
    wb = Workbook()
    # 删默认 sheet
    default = wb.active
    wb.remove(default)

    if mode == "guide":
        # 01-使用指引：仅含「使用指引」+「填写原则」两个 sheet
        ws1 = wb.create_sheet("使用指引")
        build_guide(ws1, mode)
        ws2 = wb.create_sheet("填写原则")
        build_guide_principles(ws2)
    else:
        # 02-空表 / 03-填好示例：含 10 个工具 sheet
        ws = wb.create_sheet("课程概览")
        build_overview(ws, mode)
        ws = wb.create_sheet("5因素诊断")
        build_5factors(ws, mode)
        ws = wb.create_sheet("客户洞察")
        build_customer_insight(ws, mode)
        ws = wb.create_sheet("知识流通")
        build_knowledge_flow(ws, mode)
        ws = wb.create_sheet("最小实验")
        build_mini_experiment(ws, mode)
        ws = wb.create_sheet("行为对照")
        build_behavior_compare(ws, mode)
        ws = wb.create_sheet("挑战卡")
        build_challenge(ws, mode)
        ws = wb.create_sheet("30-60-90")
        build_action_commitment(ws, mode)
        ws = wb.create_sheet("团队巡检")
        build_inspection_log(ws, mode)
        ws = wb.create_sheet("学员评估")
        build_student_assessment(ws, mode)

    out_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(out_path)
    print(f"[OK] {filename}  ·  模式={mode}  ·  Sheet 数={len(wb.sheetnames)}")
    return out_path


def main():
    print("=" * 60)
    print("创新领导力 · 配套表单 Excel 生成")
    print("=" * 60)
    build_workbook("01-使用指引.xlsx", "guide")
    build_workbook("02-空表.xlsx", "blank")
    build_workbook("03-填好示例.xlsx", "example")
    print()
    print(f"全部文件已生成到：{OUTPUT_DIR}")


if __name__ == "__main__":
    main()

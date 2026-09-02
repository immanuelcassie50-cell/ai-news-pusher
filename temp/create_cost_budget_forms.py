# -*- coding: utf-8 -*-
"""创建成本与预算管理配套表单 Excel 工作簿"""
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# 输出目录
OUTPUT_DIR = "D:/新课开发/财务管理/6-成本与预算管理：从费用控制到利润改善/成果demo/配套表单和指引-Excel版"

# 红色-灰色主题配色
DARK_BLUE = "2b2d42"      # 深蓝色表头背景
WHITE = "FFFFFF"          # 白色文字
LIGHT_GRAY = "edf2f4"     # 浅灰色交替行
ACCENT_RED = "ef233c"     # 强调红色

def create_header_style():
    """创建表头样式：深蓝背景+白字+居中"""
    return {
        'font': Font(name='微软雅黑', size=11, bold=True, color=WHITE),
        'fill': PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='AAAAAA'),
            right=Side(style='thin', color='AAAAAA'),
            top=Side(style='thin', color='AAAAAA'),
            bottom=Side(style='thin', color='AAAAAA')
        )
    }

def create_data_style(is_odd=True):
    """创建数据行样式"""
    bg_color = LIGHT_GRAY if is_odd else WHITE
    return {
        'font': Font(name='微软雅黑', size=10),
        'fill': PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
    }

def create_input_style():
    """创建输入单元格样式（带下划线提示）"""
    return {
        'font': Font(name='微软雅黑', size=10, color="7f8c8d"),
        'fill': PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='AAAAAA')
        )
    }

def apply_header(ws, headers, start_row=1):
    """应用表头样式"""
    style = create_header_style()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

def apply_data_row(ws, row_num, num_cols, is_odd=True):
    """应用数据行样式"""
    style = create_data_style(is_odd)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

def set_column_widths(ws, widths):
    """设置列宽"""
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def freeze_top_row(ws):
    """冻结首行"""
    ws.freeze_panes = 'A2'

def add_title(ws, title, num_cols):
    """添加标题行"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1)
    cell.value = title
    cell.font = Font(name='微软雅黑', size=14, bold=True, color=DARK_BLUE)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

# ========== Sheet 1: F1_成本结构识别 ==========
def create_sheet_f1(wb):
    ws = wb.create_sheet(title="F1_成本结构识别")

    headers = ["成本项目", "金额(万元)", "成本类型(固定/变动/半固定)", "敏感度(高/中/低)", "判断依据"]
    num_cols = len(headers)

    # 标题行
    add_title(ws, "成本结构识别表", num_cols)
    ws.row_dimensions[1].height = 35

    # 表头
    apply_header(ws, headers, start_row=2)
    ws.row_dimensions[2].height = 25

    # 数据行
    cost_items = [
        "原材料采购成本", "直接人工成本", "制造费用", "生产成本-合计",
        "销售费用-工资", "销售费用-佣金", "销售费用-广告", "销售费用-差旅",
        "管理费用-工资", "管理费用-办公", "管理费用-折旧", "管理费用-社保",
        "研发费用-人员", "研发费用-材料", "研发费用-设备", "财务费用-利息",
        "财务费用-汇兑", "租金费用", "水电费", "运输费"
    ]

    input_style = create_input_style()
    for idx, item in enumerate(cost_items):
        row = idx + 3
        ws.cell(row=row, column=1).value = item
        ws.cell(row=row, column=2).value = "___"  # 金额
        ws.cell(row=row, column=3).value = "___"  # 成本类型
        ws.cell(row=row, column=4).value = "___"  # 敏感度
        ws.cell(row=row, column=5).value = "___"  # 判断依据

        # 设置行样式
        apply_data_row(ws, row, num_cols, is_odd=(idx % 2 == 0))

        # 输入单元格样式
        for col in [2, 3, 4, 5]:
            cell = ws.cell(row=row, column=col)
            cell.font = input_style['font']
            cell.fill = input_style['fill']
            cell.alignment = input_style['alignment']
            cell.border = input_style['border']

    # 设置列宽
    set_column_widths(ws, [20, 12, 18, 12, 25])
    freeze_top_row(ws)

# ========== Sheet 2: F2_部门成本责任 ==========
def create_sheet_f2(wb):
    ws = wb.create_sheet(title="F2_部门成本责任")

    headers = ["部门", "成本中心类型(成本中心/利润中心/投资中心)", "可控成本项目", "考核指标", "目标值"]
    num_cols = len(headers)

    add_title(ws, "部门成本责任表", num_cols)
    ws.row_dimensions[1].height = 35

    apply_header(ws, headers, start_row=2)
    ws.row_dimensions[2].height = 25

    departments = [
        ("生产部", "成本中心", "原材料消耗、人工工时、设备折旧", "单位产品成本、产能利用率", "控制在预算95%以内"),
        ("销售部", "利润中心", "销售费用、客户开发费用、促销费用", "销售收入、费用率、客户满意度", "销售收入增长15%"),
        ("采购部", "成本中心", "采购成本、库存持有成本、仓储费用", "采购降价率、库存周转率", "采购成本降低8%"),
        ("研发部", "成本中心", "研发费用、人员成本、实验材料", "项目完成率、专利数量、研发周期", "研发费用控制在预算内"),
        ("财务部", "成本中心", "管理费用、资金成本、税费", "预算准确率、资金利用率", "费用率下降5%"),
        ("人力资源部", "成本中心", "人力成本、培训费用、招聘费用", "人均产出、员工流失率", "人均产出提升10%"),
        ("行政管理部", "成本中心", "办公费用、租金、水电费、物业费", "费用控制率、服务满意度", "费用率下降8%"),
        ("质量管理部", "成本中心", "质量检测费、认证费、质量事故损失", "合格率、客诉率、质量成本", "客诉率降低20%"),
    ]

    input_style = create_input_style()
    for idx, (dept, center_type, cost_items, kpi, target) in enumerate(departments):
        row = idx + 3
        ws.cell(row=row, column=1).value = dept
        ws.cell(row=row, column=2).value = center_type
        ws.cell(row=row, column=3).value = cost_items
        ws.cell(row=row, column=4).value = kpi
        ws.cell(row=row, column=5).value = target

        apply_data_row(ws, row, num_cols, is_odd=(idx % 2 == 0))

    set_column_widths(ws, [15, 22, 30, 25, 22])
    freeze_top_row(ws)

# ========== Sheet 3: F3_预算编制逻辑 ==========
def create_sheet_f3(wb):
    ws = wb.create_sheet(title="F3_预算编制逻辑")

    headers = ["预算项目", "编制方法", "上年实际", "本年预算", "变化说明"]
    num_cols = len(headers)

    add_title(ws, "预算编制逻辑表", num_cols)
    ws.row_dimensions[1].height = 35

    apply_header(ws, headers, start_row=2)
    ws.row_dimensions[2].height = 25

    budget_items = [
        ("销售收入", "零基预算+趋势分析", "___", "___", "___"),
        ("销售成本", "标准成本法", "___", "___", "___"),
        ("毛利", "计算得出", "___", "___", "___"),
        ("销售费用-工资", "增量预算", "___", "___", "___"),
        ("销售费用-市场推广", "零基预算", "___", "___", "___"),
        ("管理费用-工资", "增量预算", "___", "___", "___"),
        ("管理费用-折旧", "固定预算", "___", "___", "___"),
        ("研发费用", "零基预算+项目评审", "___", "___", "___"),
        ("财务费用", "固定预算", "___", "___", "___"),
        ("资本支出", "项目预算", "___", "___", "___"),
        ("人员编制", "增量预算", "___", "___", "___"),
        ("管理费用率", "目标倒推", "___", "___", "___"),
    ]

    input_style = create_input_style()
    for idx, (item, method, last_year, this_year, change) in enumerate(budget_items):
        row = idx + 3
        ws.cell(row=row, column=1).value = item
        ws.cell(row=row, column=2).value = method
        ws.cell(row=row, column=3).value = last_year
        ws.cell(row=row, column=4).value = this_year
        ws.cell(row=row, column=5).value = change

        apply_data_row(ws, row, num_cols, is_odd=(idx % 2 == 0))

        # 金额列使用输入样式
        for col in [3, 4, 5]:
            cell = ws.cell(row=row, column=col)
            cell.font = input_style['font']
            cell.fill = input_style['fill']
            cell.alignment = input_style['alignment']
            cell.border = input_style['border']

    set_column_widths(ws, [18, 20, 12, 12, 20])
    freeze_top_row(ws)

# ========== Sheet 4: F4_预算执行监控 ==========
def create_sheet_f4(wb):
    ws = wb.create_sheet(title="F4_预算执行监控")

    headers = ["预算项目", "预算金额", "已执行金额", "执行率%", "偏差金额", "偏差原因", "预警等级(红/黄/绿)"]
    num_cols = len(headers)

    add_title(ws, "预算执行监控表", num_cols)
    ws.row_dimensions[1].height = 35

    apply_header(ws, headers, start_row=2)
    ws.row_dimensions[2].height = 25

    monitor_items = [
        "销售收入", "销售成本", "销售费用-工资", "销售费用-广告",
        "管理费用-工资", "管理费用-办公", "研发费用", "财务费用",
        "资本支出", "人员编制"
    ]

    input_style = create_input_style()
    for idx, item in enumerate(monitor_items):
        row = idx + 3
        ws.cell(row=row, column=1).value = item
        ws.cell(row=row, column=2).value = "___"  # 预算金额
        ws.cell(row=row, column=3).value = "___"  # 已执行金额
        ws.cell(row=row, column=4).value = "___"  # 执行率
        ws.cell(row=row, column=5).value = "___"  # 偏差金额
        ws.cell(row=row, column=6).value = "___"  # 偏差原因
        ws.cell(row=row, column=7).value = "___"  # 预警等级

        apply_data_row(ws, row, num_cols, is_odd=(idx % 2 == 0))

        for col in [2, 3, 4, 5, 6, 7]:
            cell = ws.cell(row=row, column=col)
            cell.font = input_style['font']
            cell.fill = input_style['fill']
            cell.alignment = input_style['alignment']
            cell.border = input_style['border']

    set_column_widths(ws, [18, 12, 12, 10, 10, 20, 14])
    freeze_top_row(ws)

# ========== Sheet 5: F5_偏差分析 ==========
def create_sheet_f5(wb):
    ws = wb.create_sheet(title="F5_偏差分析")

    headers = ["预算项目", "预算值", "实际值", "偏差量", "偏差率", "5Why分析", "纠偏措施"]
    num_cols = len(headers)

    add_title(ws, "偏差分析表", num_cols)
    ws.row_dimensions[1].height = 35

    apply_header(ws, headers, start_row=2)
    ws.row_dimensions[2].height = 25

    variance_items = [
        "销售收入", "原材料成本", "人工成本", "制造费用",
        "销售费用", "管理费用", "研发费用", "净利润"
    ]

    input_style = create_input_style()
    for idx, item in enumerate(variance_items):
        row = idx + 3
        ws.cell(row=row, column=1).value = item
        ws.cell(row=row, column=2).value = "___"  # 预算值
        ws.cell(row=row, column=3).value = "___"  # 实际值
        ws.cell(row=row, column=4).value = "___"  # 偏差量
        ws.cell(row=row, column=5).value = "___"  # 偏差率
        ws.cell(row=row, column=6).value = "___"  # 5Why分析
        ws.cell(row=row, column=7).value = "___"  # 纠偏措施

        apply_data_row(ws, row, num_cols, is_odd=(idx % 2 == 0))

        for col in [2, 3, 4, 5, 6, 7]:
            cell = ws.cell(row=row, column=col)
            cell.font = input_style['font']
            cell.fill = input_style['fill']
            cell.alignment = input_style['alignment']
            cell.border = input_style['border']

    set_column_widths(ws, [15, 12, 12, 10, 10, 30, 25])
    freeze_top_row(ws)

# ========== Sheet 6: F6_降本策略评估 ==========
def create_sheet_f6(wb):
    ws = wb.create_sheet(title="F6_降本策略评估")

    headers = ["降本机会", "影响业务", "实施难度(高/中/低)", "预计节省", "优先级(1-5)", "实施计划"]
    num_cols = len(headers)

    add_title(ws, "降本策略评估表", num_cols)
    ws.row_dimensions[1].height = 35

    apply_header(ws, headers, start_row=2)
    ws.row_dimensions[2].height = 25

    cost_reduction_items = [
        ("供应商整合", "可能影响交付周期", "中", "___", "___", "___"),
        ("工艺优化", "提升生产效率", "高", "___", "___", "___"),
        ("自动化改造", "减少人工依赖", "高", "___", "___", "___"),
        ("集中采购", "增强议价能力", "低", "___", "___", "___"),
        ("库存优化", "减少资金占用", "中", "___", "___", "___"),
        ("能源管理", "降低能耗成本", "低", "___", "___", "___"),
        ("流程简化", "提升效率", "中", "___", "___", "___"),
        ("外协转自产", "提升核心能力", "高", "___", "___", "___"),
        ("包装优化", "降低物流成本", "低", "___", "___", "___"),
        ("差旅费用管控", "减少非必要出差", "低", "___", "___", "___"),
    ]

    input_style = create_input_style()
    for idx, (opportunity, impact, difficulty, saving, priority, plan) in enumerate(cost_reduction_items):
        row = idx + 3
        ws.cell(row=row, column=1).value = opportunity
        ws.cell(row=row, column=2).value = impact
        ws.cell(row=row, column=3).value = difficulty
        ws.cell(row=row, column=4).value = saving
        ws.cell(row=row, column=5).value = priority
        ws.cell(row=row, column=6).value = plan

        apply_data_row(ws, row, num_cols, is_odd=(idx % 2 == 0))

        for col in [4, 5, 6]:
            cell = ws.cell(row=row, column=col)
            cell.font = input_style['font']
            cell.fill = input_style['fill']
            cell.alignment = input_style['alignment']
            cell.border = input_style['border']

    set_column_widths(ws, [18, 18, 16, 12, 12, 25])
    freeze_top_row(ws)

# ========== 主函数 ==========
def create_main_workbook():
    """创建配套表单主工作簿"""
    wb = Workbook()

    # 删除默认sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 创建各sheet
    create_sheet_f1(wb)
    create_sheet_f2(wb)
    create_sheet_f3(wb)
    create_sheet_f4(wb)
    create_sheet_f5(wb)
    create_sheet_f6(wb)

    output_path = os.path.join(OUTPUT_DIR, "配套表单_空表.xlsx")
    wb.save(output_path)
    print(f"已创建: {output_path}")
    return output_path

# ========== Sheet: 表单使用指引 ==========
def create_guide_workbook():
    """创建表单使用指引工作簿"""
    wb = Workbook()

    ws = wb.active
    ws.title = "表单使用指引"

    # 标题
    ws.merge_cells('A1:E1')
    cell = ws.cell(row=1, column=1)
    cell.value = "成本与预算管理配套表单使用指引"
    cell.font = Font(name='微软雅黑', size=16, bold=True, color=DARK_BLUE)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # 副标题说明
    ws.merge_cells('A2:E2')
    cell = ws.cell(row=2, column=1)
    cell.value = "本工具包包含6张配套表单，帮助学员在学习和实践中掌握成本与预算管理的核心技能"
    cell.font = Font(name='微软雅黑', size=10, color="666666")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    # 表头
    headers = ["表单编号", "表单名称", "使用时机", "核心用途", "填写要点"]
    apply_header(ws, headers, start_row=3)
    ws.row_dimensions[3].height = 25

    # 指引数据
    guide_data = [
        ("F1", "成本结构识别", "业务开始前/成本分析时", "识别企业成本构成，区分固定/变动/半固定成本，分析成本敏感度", "1. 逐项列出主要成本项目\n2. 根据成本习性正确分类\n3. 判断各成本对业务变化的敏感程度"),
        ("F2", "部门成本责任", "预算编制前/责任划分时", "明确各部门成本责任，建立成本中心考核体系，落实成本管控责任", "1. 确认部门类型（成本/利润/投资中心）\n2. 识别该部门可控成本\n3. 设定明确的考核指标和目标值"),
        ("F3", "预算编制逻辑", "年度预算编制时", "运用不同预算编制方法，科学合理地编制年度预算，说明变化原因", "1. 根据业务特点选择编制方法\n2. 参考上年实际数据\n3. 充分说明重大变化的原因"),
        ("F4", "预算执行监控", "预算执行过程中", "跟踪预算执行进度，及时发现偏差并预警，确保预算目标达成", "1. 定期更新实际执行数据\n2. 计算执行率和偏差金额\n3. 分析偏差原因并设置预警等级"),
        ("F5", "偏差分析", "预算执行结束后", "深入分析预算与实际的差异，运用5Why方法找到根本原因，制定纠偏措施", "1. 计算偏差量和偏差率\n2. 连续追问5个为什么找到根因\n3. 制定可操作的纠偏措施并落实"),
        ("F6", "降本策略评估", "成本改善专项时", "识别降本机会，评估实施难度和优先级，制定可落地的实施计划", "1. 全面识别可能的降本机会\n2. 客观评估对业务的影响和实施难度\n3. 根据节省金额和优先级排序制定实施计划"),
    ]

    input_style = create_input_style()
    for idx, (code, name, timing, purpose, points) in enumerate(guide_data):
        row = idx + 4
        ws.cell(row=row, column=1).value = code
        ws.cell(row=row, column=2).value = name
        ws.cell(row=row, column=3).value = timing
        ws.cell(row=row, column=4).value = purpose
        ws.cell(row=row, column=5).value = points

        apply_data_row(ws, row, 5, is_odd=(idx % 2 == 0))

        # 序号列加粗
        ws.cell(row=row, column=1).font = Font(name='微软雅黑', size=11, bold=True, color=ACCENT_RED)

    # 设置列宽
    set_column_widths(ws, [10, 16, 22, 35, 40])

    # 添加使用说明sheet
    ws2 = wb.create_sheet(title="使用说明")
    ws2.merge_cells('A1:D1')
    cell = ws2.cell(row=1, column=1)
    cell.value = "表单填写规范"
    cell.font = Font(name='微软雅黑', size=14, bold=True, color=DARK_BLUE)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 30

    notes = [
        ("通用规范", ""),
        ("1. 金额单位", "除特别注明外，金额单位均为\"万元\""),
        ("2. 比例填写", "百分比格式，如 15% 或 0.15"),
        ("3. 日期格式", "统一使用 YYYY-MM-DD 格式"),
        ("4. 预留填写", "___ 表示需要填写的空白项，请根据实际情况填写"),
        ("", ""),
        ("颜色含义", ""),
        ("深蓝色表头", "表示列标题，请按标题含义填写"),
        ("浅灰色行", "表示交替背景色，仅用于视觉区分，无特殊含义"),
        ("___ 单元格", "表示需要填写的内容，填写后请删除下划线"),
        ("", ""),
        ("注意事项", ""),
        ("1. 保存格式", "建议保存为 .xlsx 格式，兼容性好"),
        ("2. 数据验证", "填写前可设置数据验证，避免输入错误"),
        ("3. 定期更新", "预算执行监控表建议每周或每月更新一次"),
        ("4. 归档管理", "建议按月份或季度归档保存，便于追溯分析"),
    ]

    for idx, (label, content) in enumerate(notes):
        row = idx + 2
        ws2.cell(row=row, column=1).value = label
        ws2.cell(row=row, column=2).value = content
        if label and not content:
            ws2.cell(row=row, column=1).font = Font(name='微软雅黑', size=11, bold=True, color=DARK_BLUE)
        else:
            ws2.cell(row=row, column=1).font = Font(name='微软雅黑', size=10, color="555555")

    set_column_widths(ws2, [15, 60])

    output_path = os.path.join(OUTPUT_DIR, "表单使用指引.xlsx")
    wb.save(output_path)
    print(f"已创建: {output_path}")
    return output_path

if __name__ == "__main__":
    create_main_workbook()
    create_guide_workbook()
    print("\n所有文件创建完成！")
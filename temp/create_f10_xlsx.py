# -*- coding: utf-8 -*-
"""
Generate F10 服务营销改造计划表.xlsx
"""
import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

OUTPUT_PATH = r"D:/新课开发/经营/系列/07_服务营销创新/全流程工具表单-html打印版/服务营销改造计划表.xlsx"

# ── colour palette ──────────────────────────────────────────────────────────
BLUE_DARK   = "1A5FB4"
BLUE_MED    = "4472C4"
BLUE_LIGHT  = "D6E4F0"
BLUE_PALE   = "EBF3FB"
YELLOW_HDR  = "FFF2CC"
YELLOW_PALE = "FFFDE7"
GREEN_PALE  = "E2EFDA"
ORANGE_PALE = "FCE4D6"
WHITE       = "FFFFFF"
GREY_LIGHT  = "F5F5F5"

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(style="medium", color=BLUE_DARK)
    s = Side(style="thin",   color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)

def header_fill(hex_color=BLUE_DARK):
    return PatternFill("solid", fgColor=hex_color)

def alt_fill(hex_color=BLUE_PALE):
    return PatternFill("solid", fgColor=hex_color)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def title_row(ws, row, text, span_end_col, font_size=14):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_end_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font  = Font(name="Microsoft YaHei", size=font_size, bold=True, color=WHITE)
    cell.fill  = header_fill()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def section_header(ws, row, text, span_end_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_end_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font  = Font(name="Microsoft YaHei", size=11, bold=True, color=BLUE_DARK)
    cell.fill  = PatternFill("solid", fgColor=BLUE_LIGHT)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    thin = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=Side(style="medium", color=BLUE_DARK),
                         right=thin, top=thin, bottom=thin)
    ws.row_dimensions[row].height = 20

def col_header(ws, row, col, text, width=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name="Microsoft YaHei", size=10, bold=True, color=WHITE)
    cell.fill      = header_fill(BLUE_MED)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 22

def data_cell(ws, row, col, value="", fill_hex=None, bold=False, align="left", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Microsoft YaHei", size=9, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border    = thin_border()
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)
    return cell

def phase_header_row(ws, row, text, span_end_col, fill_hex=BLUE_LIGHT):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_end_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = Font(name="Microsoft YaHei", size=11, bold=True, color=BLUE_DARK)
    cell.fill      = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border    = thick_border()
    ws.row_dimensions[row].height = 22

def milestone_row(ws, row, text, span_end_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_end_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = Font(name="Microsoft YaHei", size=9, bold=True, color="7F6000")
    cell.fill      = PatternFill("solid", fgColor=YELLOW_HDR)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    thin = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=Side(style="medium", color="FFC107"),
                         right=thin, top=thin, bottom=thin)
    ws.row_dimensions[row].height = 18

def blank_data_row(ws, row, cols, fill_hex=None):
    for c in range(1, cols + 1):
        data_cell(ws, row, c, fill_hex=fill_hex)
    ws.row_dimensions[row].height = 18

# ════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "服务营销改造计划表"
ws.sheet_view.showGridLines = False

# ── page setup ────────────────────────────────────────────────────────────────
ws.page_setup.orientation      = "landscape"
ws.page_setup.paperSize         = 9   # A4
ws.page_setup.fitToPage         = True
ws.page_setup.fitToWidth        = 1
ws.page_setup.fitToHeight       = 0
ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.7, bottom=0.7, header=0.3, footer=0.3)
ws.oddHeader.center.text = "服务营销创新 | F10 服务营销改造计划表"
ws.oddFooter.left.text   = "使用时按90天周期执行，每周进度追踪，每阶段成果验收"
ws.oddFooter.right.text  = "第 &P 页"

# ── column widths ─────────────────────────────────────────────────────────────
col_widths = {
    1:8,   # 周次
    2:10,  # 时间
    3:14,  # 重点任务
    4:26,  # 具体行动
    5:10,  # 负责人
    6:18,  # 产出
    7:12,  # 状态
}
for c, w in col_widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

# ════════════════════════════════════════════════════════════════════════════
# Row 1 — main title
# ════════════════════════════════════════════════════════════════════════════
title_row(ws, 1, "F10 服务营销改造计划表", 7, font_size=16)
ws.row_dimensions[1].height = 32

# Row 2 — subtitle
ws.merge_cells("A2:G2")
cell = ws.cell(row=2, column=1, value="服务营销创新  |  90天蜕变计划：从理念到行动，从行动到习惯")
cell.font      = Font(name="Microsoft YaHei", size=10, color="555555", italic=True)
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.fill      = PatternFill("solid", fgColor="F0F4F8")
ws.row_dimensions[2].height = 18

# ════════════════════════════════════════════════════════════════════════════
# Section 1 — basic info
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, 3, "一、改造基本信息", 7)

# Row 4 — header labels
basic_headers = ["部门/门店", "负责人", "计划周期", "预计投入（万元）", "制定日期", "阶段"]
for i, h in enumerate(basic_headers, 1):
    col_header(ws, 4, i, h)

# Row 5 — data
for c in range(1, 7):
    data_cell(ws, 5, c, fill_hex=GREY_LIGHT)
ws.row_dimensions[5].height = 20

# ════════════════════════════════════════════════════════════════════════════
# Section 2 — 现状诊断
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, 6, "二、现状诊断", 7)

diag_headers = ["评估维度", "当前等级（1-5）", "主要问题", "核心问题", "改进目标", "达成时间"]
for i, h in enumerate(diag_headers, 1):
    col_header(ws, 7, i, h)

dims = ["服务理念", "服务流程", "服务团队", "服务触点", "客户反馈机制", "服务激励机制"]
for r_off, dim in enumerate(dims):
    row = 8 + r_off
    data_cell(ws, row, 1, dim, fill_hex=BLUE_PALE)
    data_cell(ws, row, 2, fill_hex=BLUE_PALE)
    data_cell(ws, row, 3, fill_hex=BLUE_PALE)
    data_cell(ws, row, 4, fill_hex=YELLOW_PALE)
    data_cell(ws, row, 5, fill_hex=GREEN_PALE)
    data_cell(ws, row, 6, fill_hex=GREEN_PALE)
    ws.row_dimensions[row].height = 18

# ════════════════════════════════════════════════════════════════════════════
# Section 3 — 90天改造计划
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, 14, "三、90天改造计划", 7)

plan_headers = ["周次", "时间", "重点任务", "具体行动", "负责人", "产出", "状态"]
for i, h in enumerate(plan_headers, 1):
    col_header(ws, 15, i, h)

phases = [
    ("第一阶段", "觉醒期", "第1-30天", "看见差距，唤醒意识",
     [("第1周", "第1-7天", "服务审计",    "全面检视现有服务问题",         "", "服务审计报告"),
      ("第2周", "第8-14天", "理念导入",   "服务理念培训与研讨",            "", "培训记录"),
      ("第3周", "第15-21天", "标杆学习",  "参观优秀服务企业",              "", "学习报告"),
      ("第4周", "第22-30天", "目标锁定",  "制定个性化改造目标",            "", "目标承诺书")],
     "完成现状诊断报告，团队形成共识"),

    ("第二阶段", "行动期", "第31-60天", "刻意练习，形成习惯",
     [("第5周", "第31-37天", "流程再造",  "优化关键服务流程",              "", "新流程手册"),
      ("第6周", "第38-44天", "技能提升",  "服务技能专项训练",              "", "训练记录"),
      ("第7周", "第45-51天", "触点优化",  "改善客户接触关键点",           "", "触点改善方案"),
      ("第8周", "第52-60天", "机制建立",  "建立服务监督与反馈机制",        "", "机制文件")],
     "完成流程改造，技能明显提升"),

    ("第三阶段", "固化期", "第61-90天", "习惯养成，文化成型",
     [("第9周",  "第61-67天",  "习惯强化", "服务行为日常化训练",           "", "训练追踪表"),
      ("第10周", "第68-74天",  "激励强化", "优秀服务表彰与激励",           "", "激励记录"),
      ("第11周", "第75-81天",  "文化沉淀", "服务案例库建设",               "", "案例汇编"),
      ("第12周", "第82-90天",  "验收总结", "成果验收与经验沉淀",           "", "验收报告")],
     "服务文化初步形成，行为习惯养成"),
]

current_row = 16
phase_fill_map = {0: "D6E4F0", 1: "E2EFDA", 2: "FCE4D6"}

for p_idx, (phase_name, theme, days, tagline, weeks, milestone) in enumerate(phases):
    # Phase header row
    phase_label = f"{phase_name}：{theme}（{days}）— {tagline}"
    phase_header_row(ws, current_row, phase_label, 7, fill_hex=phase_fill_map[p_idx])
    current_row += 1

    for w in weeks:
        wk_num, wk_days, task, action, owner, output = w
        data_cell(ws, current_row, 1, wk_num,  fill_hex=BLUE_PALE, bold=True, align="center")
        data_cell(ws, current_row, 2, wk_days, fill_hex=BLUE_PALE, align="center")
        data_cell(ws, current_row, 3, task,    fill_hex=WHITE, bold=True)
        data_cell(ws, current_row, 4, action,  fill_hex=WHITE)
        data_cell(ws, current_row, 5, owner,   fill_hex=WHITE, align="center")
        data_cell(ws, current_row, 6, output,  fill_hex=WHITE, align="center")
        data_cell(ws, current_row, 7, "",      fill_hex=WHITE, align="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Milestone row
    milestone_row(ws, current_row, f"★ {phase_name}里程碑：{milestone}", 7)
    current_row += 1

# ════════════════════════════════════════════════════════════════════════════
# Section 4 — 关键改造项目
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, current_row, "四、关键改造项目", 7)
current_row += 1

proj_headers = ["序号", "改造项目", "改造内容", "开始时间", "完成时间", "负责人", "验收标准"]
for i, h in enumerate(proj_headers, 1):
    col_header(ws, current_row, i, h)
current_row += 1

for i in range(1, 6):
    data_cell(ws, current_row, 1, str(i), fill_hex=BLUE_PALE, align="center")
    for c in range(2, 8):
        data_cell(ws, current_row, c, fill_hex=GREY_LIGHT)
    ws.row_dimensions[current_row].height = 20
    current_row += 1

# ════════════════════════════════════════════════════════════════════════════
# Section 5 — 阶段性成果验收
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, current_row, "五、阶段性成果验收", 7)
current_row += 1

stage_headers = ["验收项", "验收标准", "是否通过", "验收项", "验收标准", "是否通过", "验收项"]
stage_headers2 = ["", "阶段一（第30天）", "", "", "阶段二（第60天）", "", "阶段三（第90天）"]

# merged sub-headers
ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
for c in [1, 4, 7]:
    ws.cell(row=current_row, column=c).border = thin_border()
for c, h in zip([2, 5, 7], ["阶段一（第30天）", "阶段二（第60天）", "阶段三（第90天）"]):
    cell = ws.cell(row=current_row, column=c, value=h)
    cell.font      = Font(name="Microsoft YaHei", size=10, bold=True, color=WHITE)
    cell.fill      = header_fill(BLUE_MED)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()
ws.row_dimensions[current_row].height = 20
current_row += 1

for i, h in enumerate(stage_headers, 1):
    col_header(ws, current_row, i, h)
current_row += 1

stage_items = [
    ("服务审计报告", "完成现状全面诊断", "团队共识", "80%以上成员认同目标", "目标承诺", "签署承诺书"),
    ("流程改造", "新流程上线试运行", "技能提升", "关键技能考核达标", "行为习惯", "服务行为自然流露"),
    ("触点改善", "客户触点满意度提升", "客户满意度", "满意度提升至目标值", "文化成型", "形成文化手册"),
]
for row_data in stage_items:
    s1, c1, s2, c2, s3, c3 = row_data
    data_cell(ws, current_row, 1, "", fill_hex=BLUE_PALE)
    data_cell(ws, current_row, 2, s1, fill_hex=WHITE)
    data_cell(ws, current_row, 3, c1, fill_hex=WHITE)
    data_cell(ws, current_row, 4, s2, fill_hex=BLUE_PALE)
    data_cell(ws, current_row, 5, c2, fill_hex=WHITE)
    data_cell(ws, current_row, 6, s3, fill_hex=WHITE)
    data_cell(ws, current_row, 7, c3, fill_hex=WHITE)
    ws.row_dimensions[current_row].height = 20
    current_row += 1

# ════════════════════════════════════════════════════════════════════════════
# Section 6 — 改造成功标准
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, current_row, "六、改造成功标准", 7)
current_row += 1

success_headers = ["指标", "改造前", "改造目标", "当前值", "指标", "改造前", "改造目标"]
ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
for c, h in zip([1, 2, 3, 4, 5], ["量化指标", "改造前", "改造目标", "当前值", "定性目标"]):
    cell = ws.cell(row=current_row, column=c, value=h)
    cell.font      = Font(name="Microsoft YaHei", size=10, bold=True, color=WHITE)
    cell.fill      = header_fill(BLUE_MED)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()
ws.row_dimensions[current_row].height = 20
current_row += 1

success_items = [
    ("客户满意度", "", "", "", "服务理念深入人心"),
    ("服务响应时间", "", "", "", "服务流程高效顺畅"),
    ("投诉率", "", "", "", "服务文化初步形成"),
    ("员工服务意识评分", "", "", "", "团队服务能力提升"),
]
for row_data in success_items:
    q1, q2, q3, q4, d = row_data
    data_cell(ws, current_row, 1, q1, fill_hex=BLUE_PALE)
    data_cell(ws, current_row, 2, q2, fill_hex=WHITE)
    data_cell(ws, current_row, 3, q3, fill_hex=WHITE)
    data_cell(ws, current_row, 4, q4, fill_hex=WHITE)
    ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
    cell = ws.cell(row=current_row, column=5, value=d)
    cell.fill   = PatternFill("solid", fgColor=GREEN_PALE)
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 18
    current_row += 1

# ════════════════════════════════════════════════════════════════════════════
# Section 7 — 预算投入计划
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, current_row, "七、预算投入计划", 7)
current_row += 1

budget_headers = ["投入项目", "预算金额（万元）", "用途说明", "投入时间", "审批状态", "备注", ""]
for i, h in enumerate(budget_headers[:6], 1):
    col_header(ws, current_row, i, h)
current_row += 1

budget_items = ["培训费用", "工具物料", "激励奖金", "外部咨询", "其他支出", "合计"]
for i, item in enumerate(budget_items):
    fill = YELLOW_HDR if item == "合计" else (BLUE_PALE if i % 2 == 0 else WHITE)
    data_cell(ws, current_row, 1, item, fill_hex=fill, bold=(item == "合计"))
    for c in range(2, 6):
        data_cell(ws, current_row, c, fill_hex=fill)
    data_cell(ws, current_row, 6, "□ 已审批  □ 待审批", fill_hex=fill, align="center")
    ws.row_dimensions[current_row].height = 18
    current_row += 1

# ════════════════════════════════════════════════════════════════════════════
# Section 8 — 风险识别与应对
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, current_row, "八、风险识别与应对", 7)
current_row += 1

risk_headers = ["风险描述", "发生可能性", "影响程度", "应对措施", "责任人", "", ""]
for i, h in enumerate(risk_headers[:5], 1):
    col_header(ws, current_row, i, h)
current_row += 1

risks = ["员工抵触变革", "资源投入不足", "执行力度衰减", "跨部门协同困难"]
for risk in risks:
    data_cell(ws, current_row, 1, risk, fill_hex=BLUE_PALE)
    data_cell(ws, current_row, 2, "□ 高  □ 中  □ 低", fill_hex=WHITE, align="center")
    data_cell(ws, current_row, 3, "□ 大  □ 中  □ 小", fill_hex=WHITE, align="center")
    data_cell(ws, current_row, 4, "", fill_hex=WHITE)
    data_cell(ws, current_row, 5, "", fill_hex=WHITE)
    ws.row_dimensions[current_row].height = 18
    current_row += 1

# ════════════════════════════════════════════════════════════════════════════
# Section 9 — 改造团队
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, current_row, "九、改造团队", 7)
current_row += 1

team_headers = ["角色", "姓名", "职责", "联系方式", "", "", ""]
for i, h in enumerate(team_headers[:4], 1):
    col_header(ws, current_row, i, h)
current_row += 1

team_roles = ["项目发起人", "项目负责人", "培训负责人", "执行负责人", "监督负责人"]
for role in team_roles:
    data_cell(ws, current_row, 1, role, fill_hex=BLUE_PALE, bold=True)
    for c in range(2, 5):
        data_cell(ws, current_row, c, fill_hex=WHITE)
    ws.row_dimensions[current_row].height = 18
    current_row += 1

# ════════════════════════════════════════════════════════════════════════════
# Section 10 — 计划确认
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, current_row, "十、计划确认", 7)
current_row += 1

confirm_headers = ["项目", "签名", "日期", "备注", "", "", ""]
for i, h in enumerate(confirm_headers[:4], 1):
    col_header(ws, current_row, i, h)
current_row += 1

confirm_items = ["计划制定人", "部门负责人", "上级领导", "人力资源确认"]
for item in confirm_items:
    data_cell(ws, current_row, 1, item, fill_hex=BLUE_PALE)
    for c in range(2, 5):
        data_cell(ws, current_row, c, fill_hex=GREY_LIGHT)
    ws.row_dimensions[current_row].height = 22
    current_row += 1

# ════════════════════════════════════════════════════════════════════════════
# Section 11 — 改造总结（完成后填写）
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, current_row, "十一、改造总结（完成后填写）", 7)
current_row += 1

summary_headers = ["项目", "内容", "", "", "", "", ""]
for i, h in enumerate(summary_headers[:2], 1):
    col_header(ws, current_row, i, h)
current_row += 1

summary_items = ["主要成果", "数据对比", "成功经验", "失败教训", "改进建议", "后续计划"]
for item in summary_items:
    data_cell(ws, current_row, 1, item, fill_hex=BLUE_PALE, bold=True)
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
    cell = ws.cell(row=current_row, column=2)
    cell.fill   = PatternFill("solid", fgColor=WHITE)
    cell.border = thin_border()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

# ════════════════════════════════════════════════════════════════════════════
# Footer note
# ════════════════════════════════════════════════════════════════════════════
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
note = "使用说明：此计划表用于系统性推进服务营销改造，建议按90天周期执行。每周进行进度追踪，每阶段进行成果验收，确保改造落地。"
cell = ws.cell(row=current_row, column=1, value=note)
cell.font      = Font(name="Microsoft YaHei", size=9, italic=True, color="888888")
cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[current_row].height = 16

# ════════════════════════════════════════════════════════════════════════════
wb.save(OUTPUT_PATH)
print(f"Saved: {OUTPUT_PATH}")

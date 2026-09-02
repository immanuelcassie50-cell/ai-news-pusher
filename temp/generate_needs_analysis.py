# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "学员需求分析"

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 35
ws.column_dimensions['F'].width = 30

# 标题行
headers = ["序号", "部门", "职级", "当前痛点", "期望收获", "过往培训体验"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 数据行
data = [
    [1, "销售部", "销售总监", "看财务报表不知道看什么，每次只看收入数字", "能用财务思维分析客户价值和销售决策", "参加过1次财务培训，听完就忘"],
    [2, "销售部", "区域经理", "做销售预测时不考虑成本和毛利，报数字经常被财务挑战", "能算清楚每笔订单的毛利和回款周期", "无正式财务培训"],
    [3, "运营部", "运营总监", "做运营决策时不知道要考虑哪些财务指标", "能用财务指标衡量运营效果（ROI、坪效、人效）", "参加过2次财务培训，仍不会用"],
    [4, "运营部", "运营经理", "编制运营预算时被财务质疑预算不合理，无法有效回应", "能科学编制部门预算并有效沟通", "参加过1次预算培训"],
    [5, "产品部", "产品总监", "产品定价不知道如何平衡市场定位和成本考量", "能用财务思维做产品定价和投资回报分析", "无财务培训"],
    [6, "产品部", "产品经理", "写产品需求时不考虑开发成本和预期收益", "能在产品决策中融入成本收益分析", "参加过1次产品经理财务课"],
    [7, "项目管理部", "项目总监", "项目立项时不知道怎么算投资回报", "能用财务指标（NPV、IRR）评估项目可行性", "参加过1次项目财务管理课"],
    [8, "项目管理部", "项目经理", "项目结项时发现实际成本远超预算，不知道怎么分析", "能做项目成本分析和复盘", "无正式财务培训"],
    [9, "销售部", "大客户经理", "跟大客户谈合作时，不知道怎么用财务数据说服客户", "能用财务数据包装价值主张", "参加过1次商务谈判财务课"],
    [10, "运营部", "供应链经理", "供应商谈判时不知道如何用财务数据压价", "能用财务思维做供应商评估和成本分析", "参加过1次采购财务课"],
    [11, "产品部", "数据产品经理", "推荐系统投入产出比不清晰，老板问ROI答不上来", "能量化产品价值并用财务语言汇报", "参加过1次数据产品培训"],
    [12, "项目管理部", "PMO", "项目组合优先级排序时缺乏统一标准", "能用财务指标做项目组合管理和优先级决策", "参加过1次PMO培训"],
]

# 填充数据
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# 添加边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=len(data)+1, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border

# 保存文件
output_path = "D:/新课开发/财务管理/2-非财务经理的财务思维：读懂数字，做对决策/成果demo/需求分析表.xlsx"
wb.save(output_path)
print(f"文件已保存: {output_path}")

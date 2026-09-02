# -*- coding: utf-8 -*-
"""Create all 10 XLSX tool forms for 黄执中的说服课"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = "D:/新课开发/公众表达/黄执中/黄执中的说服课/完整课程包/06_工具表单"

def set_cell(ws, row, col, value, bold=False, font_color="000000", bg_color=None, alignment=None, border=None, font_size=11):
    cell = ws.cell(row=row, column=col, value=value)
    font_kwargs = {"bold": bold, "color": font_color, "size": font_size}
    cell.font = Font(**font_kwargs)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell

def create_border(style="thin"):
    side = Side(style=style)
    return Border(left=side, right=side, top=side, bottom=side)

def create_f1_xlsx():
    """F1: 说服意图诊断卡"""
    wb = Workbook()

    ws = wb.active
    ws.title = "说服意图诊断"

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.merge_cells('A1:D1')
    set_cell(ws, 1, 1, "F1：说服意图诊断卡", bold=True, font_color="FFFFFF", bg_color="E74C3C", font_size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    set_cell(ws, 2, 1, "课前自我诊断 - 找到你的说服困境类型", bg_color="FADBD8", font_size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    set_cell(ws, 4, 1, "当前困境描述", bold=True, bg_color="FDEBD0")
    ws.merge_cells('B4:D4')
    ws.cell(row=4, column=2).value = "请简要描述你目前面临的说服困境（150字以内）"
    ws.cell(row=4, column=2).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[4].height = 35

    set_cell(ws, 6, 1, "目标受众分析", bold=True, bg_color="E74C3C", font_color="FFFFFF")
    ws.merge_cells('A6:B6')

    audience_rows = [
        ("受众是谁？", ""),
        ("受众最关心什么？", ""),
        ("受众目前的立场？", ""),
        ("受众可能的顾虑？", ""),
    ]
    for i, (label, value) in enumerate(audience_rows, start=7):
        set_cell(ws, i, 1, label, bg_color="FEF9E7", border=create_border())
        ws.cell(row=i, column=2).border = create_border()

    set_cell(ws, 12, 1, "已尝试的方法", bold=True, bg_color="D5F5E3")
    ws.merge_cells('A12:C12')
    for i in range(13, 16):
        set_cell(ws, i, 1, f"方法{i-12}：", bg_color="EBF5FB")
        ws.merge_cells(f'B{i}:C{i}')

    set_cell(ws, 17, 1, "效果评估", bold=True, bg_color="E74C3C", font_color="FFFFFF")
    ws.merge_cells('A17:D17')

    eval_headers = ["方法", "效果（1-5分）", "问题所在"]
    for col, header in enumerate(eval_headers, start=1):
        set_cell(ws, 18, col, header, bold=True, bg_color="FADBD8", alignment=Alignment(horizontal="center"))

    for row in range(19, 22):
        ws.cell(row=row, column=1).border = create_border()
        ws.cell(row=row, column=2).border = create_border()
        ws.cell(row=row, column=3).border = create_border()

    set_cell(ws, 23, 1, "核心诊断结论", bold=True, bg_color="E74C3C", font_color="FFFFFF")
    ws.merge_cells('A23:D23')

    diagnosis_types = [
        ("[ ] 认知层面", "对方不知道或不了解"),
        ("[ ] 态度层面", "对方不认同或有偏见"),
        ("[ ] 行为层面", "对方不想做或没动力"),
        ("[ ] 情感层面", "对方不喜欢我或不信任我"),
    ]
    for i, (checkbox, desc) in enumerate(diagnosis_types, start=24):
        set_cell(ws, i, 1, checkbox, bg_color="FEF9E7")
        ws.merge_cells(f'B{i}:D{i}')
        ws.cell(row=i, column=2).value = desc
        ws.cell(row=i, column=2).alignment = Alignment(indent=1)

    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions['A'].width = 80
    instructions = """F1 说服意图诊断卡 - 使用说明

【课前自我诊断流程】
1. 先填正面：花3分钟如实填写正面内容
2. 再看背面流程图：对照流程图确认自己的困境类型
3. 选择对应工具：根据诊断结果选择本课程提供的对应工具表单
4. 带着诊断上课：将诊断结果带入课程，针对性学习

【诊断要点】
- 诚实评估：不要美化自己的困境
- 具体描述：避免模糊的他就是不听话
- 寻找根源：表象背后往往有更深层的原因

【困境类型与推荐工具对照】
- 认知层面 → F3话术生成器（你应该→我需要）
- 态度层面 → F6一致性原理应用
- 行为层面 → F5小火苗点燃话术
- 情感层面 → F8讨喜值自评"""

    ws2.cell(row=1, column=1).value = instructions
    ws2.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[1].height = 350

    filepath = os.path.join(OUTPUT_DIR, "F1_说服意图诊断卡.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_f2_xlsx():
    """F2: 受众反抗程度评估表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "反抗程度评估"

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15

    ws.merge_cells('A1:D1')
    set_cell(ws, 1, 1, "F2：受众反抗程度评估表", bold=True, font_color="FFFFFF", bg_color="9B59B6", font_size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    set_cell(ws, 2, 1, "评估受众的反抗程度（高/中/低）+ 对应策略", bg_color="E8DAEF", font_size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["评估维度", "低反抗（1分）", "中反抗（3分）", "得分"]
    for col, header in enumerate(headers, start=1):
        set_cell(ws, 4, col, header, bold=True, bg_color="9B59B6", font_color="FFFFFF", alignment=Alignment(horizontal="center"))

    dimensions = [
        ("历史关系", "长期信任关系，彼此熟悉了解", "有过小摩擦，关系一般", "偶有冲突误解，关系疏远"),
        ("当前情绪", "平和、开放，愿意倾听", "有些防备，犹豫不决", "抵触、反感，拒绝沟通"),
        ("利益相关度", "高度利益一致，对你有利就对他有利", "部分利益相关，需要权衡", "利益冲突明显，立场对立"),
        ("自主性需求", "愿意接受建议，不需要太多理由", "需要理由支撑，但会考虑", "强调自主决策，讨厌被说服"),
    ]

    for i, (dim, low, mid, high) in enumerate(dimensions, start=5):
        set_cell(ws, i, 1, dim, bold=True, bg_color="F5EEF8")
        ws.cell(row=i, column=2).value = low
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3).value = mid
        ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=4).value = None
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="center")
        ws.row_dimensions[i].height = 40

    set_cell(ws, 10, 1, "总分", bold=True, bg_color="9B59B6", font_color="FFFFFF")
    ws.merge_cells('A10:C10')
    set_cell(ws, 10, 4, "=SUM(D5:D8)", bg_color="E8DAEF", alignment=Alignment(horizontal="center"))

    set_cell(ws, 12, 1, "反抗程度判定", bold=True, bg_color="9B59B6", font_color="FFFFFF")
    ws.merge_cells('A12:D12')

    result_formula = '=IF(D10<=4,"低反抗 - 适合直接建议",IF(D10<=8,"中反抗 - 需要先建立关系","高反抗 - 需要系统性策略"))'
    ws.cell(row=13, column=1).value = "判定结果："
    ws.cell(row=13, column=1).font = Font(bold=True)
    ws.merge_cells('B13:D13')
    ws.cell(row=13, column=2).value = result_formula
    ws.cell(row=13, column=2).fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")

    ws.merge_cells('A15:D15')
    set_cell(ws, 15, 1, "对应策略建议", bold=True, bg_color="9B59B6", font_color="FFFFFF")

    strategies = [
        ("低反抗策略", "直接给出建议，提供充分信息，强调双赢结果", "F3话术生成器 + F5小火苗话术"),
        ("中反抗策略", "先认可对方顾虑，循序渐进引导，给出选择而非命令", "F6一致性原理 + F8讨喜值自评"),
        ("高反抗策略", "先修复关系，从共同点切入，降低威胁感", "F6一致性原理（回忆式）+ F8讨喜值提升"),
    ]

    for i, (level, tactics, tools) in enumerate(strategies, start=16):
        set_cell(ws, i, 1, level, bold=True, bg_color="F5EEF8")
        ws.cell(row=i, column=2).value = tactics
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3).value = tools
        ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 35

    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions['A'].width = 80
    instructions = """F2 受众反抗程度评估表 - 使用说明

【评估方法】
对每个维度进行1-5分评分：
- 1分 = 低反抗（最符合左侧描述）
- 3分 = 中反抗（介于两者之间）
- 5分 = 高反抗（最符合右侧描述）

【总分计算】
总分 = 4个维度得分之和（最低4分，最高20分）

【判定标准】
- 4-8分：低反抗 → 直接建议策略
- 9-12分：中反抗 → 先建立关系策略
- 13-20分：高反抗 → 系统性策略

【注意事项】
- 评估要客观，避免美化
- 高反抗时不要急于推销
- 不同维度得分差异大时，关注最高分维度"""
    ws2.cell(row=1, column=1).value = instructions
    ws2.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    filepath = os.path.join(OUTPUT_DIR, "F2_受众反抗程度评估表.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_f3_xlsx():
    """F3: 话术生成器"""
    wb = Workbook()
    ws = wb.active
    ws.title = "话术转换练习"

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20

    ws.merge_cells('A1:D1')
    set_cell(ws, 1, 1, "F3：我需要话术生成器", bold=True, font_color="FFFFFF", bg_color="27AE60", font_size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    set_cell(ws, 2, 1, "将你应该转换为我需要的句型练习", bg_color="D5F5E3", font_size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["原句（你应该...）", "转换句（我需要...）", "适用情境", "转换效果"]
    for col, header in enumerate(headers, start=1):
        set_cell(ws, 4, col, header, bold=True, bg_color="27AE60", font_color="FFFFFF", alignment=Alignment(horizontal="center"))

    examples = [
        ("你应该早点起床", "我需要你早起，这样我们能一起吃早餐", "家人/伴侣沟通", "更好"),
        ("你应该多看书", "我需要你多阅读，能更好地理解孩子", "教育孩子", "更愿意接受"),
        ("你应该努力工作", "我需要你一起努力，公司才能发展", "职场管理", "减少对抗"),
        ("你应该戒烟", "我需要你健康，这样家庭才有保障", "健康建议", "激发责任感"),
        ("你应该更主动", "我需要你主动，这样项目才能推进", "项目管理", "明确角色"),
        ("你应该多陪孩子", "我需要孩子有父亲陪伴成长", "家庭沟通", "情感共鸣"),
    ]

    for i, (original, converted, situation, effect) in enumerate(examples, start=5):
        set_cell(ws, i, 1, original, bg_color="F0FFF4")
        set_cell(ws, i, 2, converted, bg_color="E8F8F5")
        set_cell(ws, i, 3, situation, bg_color="FDFEFE", alignment=Alignment(horizontal="center"))
        set_cell(ws, i, 4, effect, bg_color="F0FFF4", alignment=Alignment(horizontal="center"))
        ws.row_dimensions[i].height = 30

    ws.merge_cells('A12:D12')
    set_cell(ws, 12, 1, "【练习区】请填写你的具体场景", bold=True, bg_color="27AE60", font_color="FFFFFF")

    for i in range(13, 18):
        set_cell(ws, i, 1, "我的原句：", bg_color="F0FFF4")
        ws.merge_cells(f'B{i}:D{i}')

    ws2 = wb.create_sheet("转换模板")
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50

    ws2.merge_cells('A1:B1')
    set_cell(ws2, 1, 1, "转换公式：我需要你[具体行为]，因为[对我/对我们的价值]", bold=True, bg_color="27AE60", font_color="FFFFFF", font_size=12)

    templates = [
        ("模板结构", "我需要你[具体行为]，因为[对我/对我们的价值]"),
        ("示例1", "我需要你准时，因为这样我才能安排好后续工作"),
        ("示例2", "我需要你配合，因为没有你的支持这个方案无法落地"),
        ("示例3", "我需要你认真考虑，因为这个决定会影响我们所有人的利益"),
        ("", ""),
        ("转换要点", "1. 把应该换成需要（请求而非要求）"),
        ("", "2. 给出理由，说明对你/双方的价值"),
        ("", "3. 避免指责，不暗示对方错或差"),
        ("", "4. 表达需求而非抱怨"),
    ]

    for i, (label, content) in enumerate(templates, start=3):
        set_cell(ws2, i, 1, label, bold=True if label else False, bg_color="D5F5E3" if label else None)
        ws2.cell(row=i, column=2).value = content
        ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True, indent=1 if label else 0)
        ws2.row_dimensions[i].height = 25 if label else 35

    ws3 = wb.create_sheet("使用说明")
    ws3.column_dimensions['A'].width = 80
    instructions = """F3 话术生成器 - 使用说明

【核心心法】
你应该 = 评判对方，让对方感到被指责
我需要 = 表达自己需求，让对方感到被尊重

【为什么要转换】
- 你应该触发防御心理
- 我需要激发配合意愿
- 同样的意思，不同的感受

【转换公式】
我需要你[具体行为] + 因为[对我的价值/对我们的价值]

【练习步骤】
1. 写下你平时常说的你应该...
2. 识别话语背后的真正需求
3. 用我需要...重新组织
4. 检验是否比原句更温和有效

【常见误区】
- 只改词不改语气（你应该变我需要但态度依旧）
- 理由太自我中心（只说对我有利，要说对双方都有利）
- 需求太抽象（要具体，不要你应该更努力而要我需要你...）"""
    ws3.cell(row=1, column=1).value = instructions
    ws3.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[1].height = 350

    filepath = os.path.join(OUTPUT_DIR, "F3_我需要话术生成器.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_f4_xlsx():
    """F4: 五种失败句型检测表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "失败句型检测"

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40

    ws.merge_cells('A1:C1')
    set_cell(ws, 1, 1, "F4：五种失败句型检测表", bold=True, font_color="FFFFFF", bg_color="E67E22", font_size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:C2')
    set_cell(ws, 2, 1, "检测当前话语是否属于5种失败句型 + 提供替代方案", bg_color="FDEBD0", font_size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["失败句型", "问题诊断", "推荐替代"]
    for col, header in enumerate(headers, start=1):
        set_cell(ws, 4, col, header, bold=True, bg_color="E67E22", font_color="FFFFFF", alignment=Alignment(horizontal="center"))

    failure_types = [
        (
            "1. 你为什么不做...",
            "暗示对方应该做但没做，带着质问和责备意味。对方感到被审判，防御心理立刻启动。",
            "表达观察：我注意到这个任务还没完成\n表达需求：我需要了解情况，看能怎么帮你\n引导反思：你觉得可以做些什么来推进？"
        ),
        (
            "2. 你应该...",
            "直接给对方贴标签，暗示对方现在的做法是错的。对方感到被否定，自尊心受挫。",
            "表达自己：我需要...\n分享经验：我以前试过...方法，效果不错\n提供选择：如果可以的话，建议考虑...\n说明原因：这样做的话，对我们会..."
        ),
        (
            "3. 你必须...",
            "命令式语气，完全忽视对方的自主性需求。对方感到被控制，逆反心理强烈。",
            "说明后果：这样的话，我们会...\n表达信任：我相信你可以...\n请求而非要求：能不能麻烦你...\n表达重要性：这对...很重要"
        ),
        (
            "4. 你不应该...",
            "否定性表达，直接指出对方的错误或不当。对方感到被批评，产生抵触情绪。",
            "转向正面：你可以试试...\n表达担心：我担心这样会...\n分享担忧：如果这样做，可能会...\n引导思考：有没有考虑过...的方式？"
        ),
        (
            "5. 你难道不能...",
            "反问句带有讽刺和嘲讽，暗示对方无能或懒惰。对方感到被嘲讽和贬低。",
            "表达困惑：我有点不明白，可以解释一下吗？\n表示关心：是不是有什么困难？\n寻求理解：我很好奇，你当时是怎么考虑的？\n直接说明：我希望...，可以吗？"
        ),
    ]

    for i, (sentence, problem, alternative) in enumerate(failure_types, start=5):
        set_cell(ws, i, 1, sentence, bold=True, bg_color="FEF5E7")
        ws.cell(row=i, column=2).value = problem
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3).value = alternative
        ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 60

    ws.merge_cells('A11:C11')
    set_cell(ws, 11, 1, "【自我检测】请写下你最近说过的一句话，检查是否属于失败句型", bold=True, bg_color="E67E22", font_color="FFFFFF")

    set_cell(ws, 12, 1, "我最近说过的话：")
    ws.merge_cells('B12:C12')

    set_cell(ws, 13, 1, "属于哪种失败句型？")
    ws.merge_cells('B13:C13')

    set_cell(ws, 14, 1, "应该怎么说？")
    ws.merge_cells('B14:C14')

    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions['A'].width = 80
    instructions = """F4 五种失败句型检测表 - 使用说明

【为什么这5种句型会失败？】

1. 你为什么不做...
→ 质问引发防御，对方第一时间想的是反驳而非改变

2. 你应该...
→ 评判引发抵触，对方感觉被贴标签而非被尊重

3. 你必须...
→ 命令引发逆反，对方感觉自主权被剥夺

4. 你不应该...
→ 否定引发愤怒，对方感觉被攻击而非被帮助

5. 你难道不能...
→ 嘲讽引发反感，对方感觉被羞辱而非被理解

【正确的心态】
说服不是让对方认错，而是让对方愿意

【替代方案的核心原则】
- 描述行为而非评判人格
- 表达需求而非指责对方
- 引导思考而非强制命令
- 尊重自主而非施加压力

【练习建议】
每天记录3句自己说过的话，检查是否有失败句型
坚持21天，形成新的说话习惯"""
    ws2.cell(row=1, column=1).value = instructions
    ws2.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[1].height = 300

    filepath = os.path.join(OUTPUT_DIR, "F4_五种失败句型检测表.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_f5_xlsx():
    """F5: 小火苗点燃话术卡"""
    wb = Workbook()
    ws = wb.active
    ws.title = "小火苗话术"

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25

    ws.merge_cells('A1:C1')
    set_cell(ws, 1, 1, "F5：小火苗点燃话术卡", bold=True, font_color="FFFFFF", bg_color="E74C3C", font_size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:C2')
    set_cell(ws, 2, 1, "三种点火话术的使用指南", bg_color="FADBD8", font_size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["点火话术", "使用说明", "适用场景"]
    for col, header in enumerate(headers, start=1):
        set_cell(ws, 4, col, header, bold=True, bg_color="E74C3C", font_color="FFFFFF", alignment=Alignment(horizontal="center"))

    fire_words = [
        (
            "你最在意的是什么？",
            "核心原理：找到对方真正的燃点\n\n使用时机：当对方态度中立、犹豫不决时\n\n操作方法：\n1. 先建立关系，让对方放松\n2. 真诚发问，不要像审问\n3. 认真倾听，找到对方真正在意的\n4. 将你的建议与对方的在意点连接",
            "客户犹豫是否购买\n下属不想执行任务\n朋友拒绝你的建议\n家人不配合"
        ),
        (
            "如果...会对你有什么帮助？",
            "核心原理：用想象力激发渴望\n\n使用时机：当对方看不到行动的价值时\n\n操作方法：\n1. 描述一个美好的画面/结果\n2. 让对方自己想象如果实现了\n3. 让对方说出来（说出来的愿望更有动力）\n4. 适时追问这对你的意义是？",
            "对方说不需要\n对方觉得无所谓\n对方缺乏动力\n需要促进行动"
        ),
        (
            "这件事对你意味着什么？",
            "核心原理：挖掘深层意义和价值观\n\n使用时机：当对方表现冷漠、态度强硬时\n\n操作方法：\n1. 放慢节奏，不急于推销\n2. 用柔和的语气发问\n3. 等待对方思考和回答\n4. 根据回答找到情感连接点",
            "对方完全不感兴趣\n对方态度冷漠\n对方表示无所谓\n需要建立深层连接"
        ),
    ]

    for i, (word, usage, scene) in enumerate(fire_words, start=5):
        set_cell(ws, i, 1, word, bold=True, bg_color="FDEDEC", font_size=12)
        ws.cell(row=i, column=2).value = usage
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3).value = scene
        ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 90

    ws.merge_cells('A9:C9')
    set_cell(ws, 9, 1, "【使用禁忌】", bold=True, bg_color="E74C3C", font_color="FFFFFF")

    taboo_text = """X 不要在对方情绪激动时使用
X 不要像审问一样发问
X 不要否定对方的回答
X 不要急于求成，多次追问要有耐心
X 不要把话术当套路，真诚是根本"""
    ws.merge_cells('A10:C10')
    ws.cell(row=10, column=1).value = taboo_text
    ws.cell(row=10, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[10].height = 70

    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions['A'].width = 80
    instructions = """F5 小火苗点燃话术卡 - 使用说明

【三种话术的核心心法】
小火苗话术的核心不是说服，而是点燃。
说服是把自己的想法强加给别人，
点燃是帮别人找到自己内心的火焰。

【适用前提】
- 对方状态：情绪平稳、愿意交流
- 关系基础：有一定的信任关系
- 时机合适：不是在对方忙碌或烦躁时

【常见错误】
1. 你最在意的是什么？问成审问 → 要真诚好奇
2. 如果...会对你有什么帮助？问得像威胁 → 要温和描述
3. 这件事对你意味着什么？问完没等回答就自己答 → 要真正倾听

【练习建议】
- 先从你最在意的是什么？开始练习
- 每天找一个人尝试真诚发问
- 注意对方回答后的反应，及时回应
- 记录哪些问题让对方眼睛发光

【进阶技巧】
当对方回答后，可以追问：
- 为什么这个对你很重要？
- 如果实现了，你会是什么感觉？
- 还有呢？"""
    ws2.cell(row=1, column=1).value = instructions
    ws2.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[1].height = 350

    filepath = os.path.join(OUTPUT_DIR, "F5_小火苗点燃话术卡.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_f6_xlsx():
    """F6: 一致性原理应用表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "一致性原理应用"

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35

    ws.merge_cells('A1:D1')
    set_cell(ws, 1, 1, "F6：一致性原理应用表", bold=True, font_color="FFFFFF", bg_color="3498DB", font_size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    set_cell(ws, 2, 1, "利用认知一致性的说服策略", bg_color="D6EAF8", font_size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["对方已有观点", "引导方向", "话术设计", "注意事项"]
    for col, header in enumerate(headers, start=1):
        set_cell(ws, 4, col, header, bold=True, bg_color="3498DB", font_color="FFFFFF", alignment=Alignment(horizontal="center"))

    strategies = [
        ("对方的价值观/信念", "找到与之一致的行为方向", "你一直很重视家庭，相信你也会同意...", "不要捏造价值观，要真实找到"),
        ("对方过去的承诺/决定", "引导到与承诺一致的新行动", "你当初决定做这件事的时候，是因为看好...", "不要翻旧账，要正面引导"),
        ("对方尊敬/崇拜的人的观点", "借助权威或榜样力量说服", "某某也遇到过类似情况，他选择的是...", "不要权威施压，要自然引用"),
        ("对方自认为是什么人", "给对方一个身份标签，激励行为", "你是一个负责任的人，我相信你会...", "不要贴标签，要对方认可的标签"),
        ("对方的兴趣爱好", "从爱好切入建立连接", "你很喜欢研究新科技，这个新产品...", "不要强行关联，要真实兴趣"),
    ]

    for i, (viewpoint, direction, tactics, note) in enumerate(strategies, start=5):
        set_cell(ws, i, 1, viewpoint, bg_color="EBF5FB")
        ws.cell(row=i, column=2).value = direction
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3).value = tactics
        ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=4).value = note
        ws.cell(row=i, column=4).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 45

    ws.merge_cells('A11:D11')
    set_cell(ws, 11, 1, "【话术设计练习】", bold=True, bg_color="3498DB", font_color="FFFFFF")

    set_cell(ws, 12, 1, "你的说服目标：")
    ws.merge_cells('B12:D12')

    set_cell(ws, 13, 1, "对方的已有观点/经历：")
    ws.merge_cells('B13:D13')

    set_cell(ws, 14, 1, "如何将两者连接：")
    ws.merge_cells('B14:D14')

    set_cell(ws, 15, 1, "设计的话术：")
    ws.merge_cells('B15:D15')

    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions['A'].width = 80
    instructions = """F6 一致性原理应用表 - 使用说明

【什么是一致性原理？】
人有一种心理需求：希望自己的思想、言语、行为保持一致。
当有人指出我们的行为与之前的承诺不一致时，我们会感到不舒服。
利用这一点，可以说：你之前说过...，所以现在应该...

【为什么有效？】
- 没有人愿意承认自己说话不算话
- 指出不一致比提出新要求更容易接受
- 对方是自己做出决定，不是被强迫

【五种应用方式】

1. 利用价值观
→ 找到对方真正重视的价值观，将你的建议与之一致
→ 注意：必须真实，不能捏造

2. 利用过去的承诺
→ 对方之前做出的公开或私下承诺
→ 注意：不要翻旧账，要正面引导

3. 利用权威/榜样
→ 对方尊敬的人的观点或做法
→ 注意：自然引用，不要权威施压

4. 利用身份标签
→ 对方自认为是什么样的人
→ 注意：要是对方认可的标签，不能是负面标签

5. 利用兴趣爱好
→ 从对方的爱好切入
→ 注意：真实关联，不要强行关联

【禁忌】
- 不要操控对方说过的某句话断章取义
- 不要让对方感觉被套路
- 不要强迫对方承认与事实不符的一致性"""
    ws2.cell(row=1, column=1).value = instructions
    ws2.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[1].height = 380

    filepath = os.path.join(OUTPUT_DIR, "F6_一致性原理应用表.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_f7_xlsx():
    """F7: SPA战略规划工作表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "SPA战略规划"

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50

    ws.merge_cells('A1:B1')
    set_cell(ws, 1, 1, "F7：SPA战略规划工作表", bold=True, font_color="FFFFFF", bg_color="1ABC9C", font_size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:B2')
    set_cell(ws, 2, 1, "完整的SPA框架填写模板", bg_color="D1F2EB", font_size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    set_cell(ws, 4, 1, "S: Speaker", bold=True, bg_color="1ABC9C", font_color="FFFFFF", font_size=14)
    ws.merge_cells('A4:B4')

    s_questions = [
        "我是谁？（身份定位）",
        "我的主张/观点是什么？",
        "我的立场背后代表什么价值观？",
        "我希望传达的核心信息是什么？",
        "我的可信度来源是什么？（专业经验/亲身经历/立场中立）"
    ]
    for i, q in enumerate(s_questions, start=5):
        set_cell(ws, i, 1, q, bg_color="E8F8F5")
        ws.merge_cells(f'B{i}:B{i}')
        ws.row_dimensions[i].height = 25

    set_cell(ws, 11, 1, "A: Audience", bold=True, bg_color="1ABC9C", font_color="FFFFFF", font_size=14)
    ws.merge_cells('A11:B11')

    a_questions = [
        "受众是谁？（具体描述）",
        "受众关心什么？（痛点/利益/情感需求）",
        "受众目前的立场/态度是什么？",
        "受众可能的顾虑/反对意见？",
        "受众的决策标准是什么？",
        "谁会影响受众的决定？（决策影响者）"
    ]
    for i, q in enumerate(a_questions, start=12):
        set_cell(ws, i, 1, q, bg_color="E8F8F5")
        ws.row_dimensions[i].height = 25

    set_cell(ws, 19, 1, "P: Purpose", bold=True, bg_color="1ABC9C", font_color="FFFFFF", font_size=14)
    ws.merge_cells('A19:B19')

    p_questions = [
        "我的说服目标是什么？（具体、可衡量）",
        "我希望受众做出什么行动/决定？",
        "我想要的影响程度？（知道→认同→行动）",
        "成功的标准是什么？",
        "如果说服失败，我的备选方案是什么？"
    ]
    for i, q in enumerate(p_questions, start=20):
        set_cell(ws, i, 1, q, bg_color="E8F8F5")
        ws.row_dimensions[i].height = 25

    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions['A'].width = 80
    instructions = """F7 SPA战略规划工作表 - 使用说明

【什么是SPA框架？】

S - Speaker（说话者）
→ 我是谁？我主张什么？我凭什么说服对方？

A - Audience（受众）
→ 受众是谁？受众关心什么？受众的顾虑？

P - Purpose（目的）
→ 我想达成什么具体目标？

【使用流程】

第一步：填写S（Speaker）
→ 清晰地定义自己的身份和主张
→ 问自己：我为什么有资格说这些话？

第二步：分析A（Audience）
→ 换位思考，真正理解受众
→ 问自己：受众为什么要听我说？

第三步：明确P（Purpose）
→ 具体化目标，不是让他认同而是让他同意什么具体行动
→ 问自己：我希望这场对话后发生什么变化？

【完整示例】

S（Speaker）：
- 身份：产品经理
- 主张：新功能能提升用户留存
- 可信度：基于数据分析

A（Audience）：
- 受众：技术总监
- 关心：技术可行性、资源成本
- 顾虑：开发周期、技术风险

P（Purpose）：
- 目标：获得开发资源承诺
- 行动：本周五前确认开发排期"""
    ws2.cell(row=1, column=1).value = instructions
    ws2.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[1].height = 400

    filepath = os.path.join(OUTPUT_DIR, "F7_SPA战略规划工作表.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_f8_xlsx():
    """F8: 讨喜值自评卡"""
    wb = Workbook()
    ws = wb.active
    ws.title = "讨喜值自评"

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15

    ws.merge_cells('A1:D1')
    set_cell(ws, 1, 1, "F8：讨喜值自评卡", bold=True, font_color="FFFFFF", bg_color="9B59B6", font_size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    set_cell(ws, 2, 1, "评估自己的讨喜程度", bg_color="E8DAEF", font_size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["评估维度", "自评分(1-5)", "具体表现", "提升建议"]
    for col, header in enumerate(headers, start=1):
        set_cell(ws, 4, col, header, bold=True, bg_color="9B59B6", font_color="FFFFFF", alignment=Alignment(horizontal="center"))

    dimensions = [
        ("1. 专注倾听", "", "对方说话时是否全神贯注？是否有眼神交流？是否有适当回应？", "不打断对方\n点头表示理解\n适时说我理解你的意思"),
        ("2. 镜像模仿", "", "是否无意识地模仿对方的肢体语言？语速语调是否接近对方？", "自然地接近对方的语速\n模仿对方的肢体姿态\n但不要刻意或夸张"),
        ("3. 上堆下切", "", "是否能从具体到抽象（上堆）？是否能从抽象到具体（下切）？是否能横向探索更多可能性（下切）？", "对方说感受，上堆到价值观\n对方说理念，下切到具体做法\n对方说一个例子，横向探索更多"),
        ("4. 给予关注", "", "是否真诚地对对方感兴趣？是否记住对方的名字和重要信息？是否在对话中让对方感到被重视？", "记住对方说过的细节\n主动提及之前聊过的内容\n真诚地赞美具体行为"),
    ]

    for i, (dim, score, behavior, suggest) in enumerate(dimensions, start=5):
        set_cell(ws, i, 1, dim, bold=True, bg_color="F5EEF8")
        ws.cell(row=i, column=2).value = score
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3).value = behavior
        ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=4).value = suggest
        ws.cell(row=i, column=4).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 65

    set_cell(ws, 10, 1, "讨喜值总分", bold=True, bg_color="9B59B6", font_color="FFFFFF")
    ws.cell(row=10, column=2).value = "=SUM(B5:B8)"
    ws.cell(row=10, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=10, column=2).font = Font(size=14, bold=True)

    ws.merge_cells('A11:D11')
    ws.cell(row=11, column=1).value = '判定：=IF(B10>=18,"高讨喜型 - 天然有亲和力",IF(B10>=10,"中等讨喜型 - 需注意细节","低讨喜型 - 需重点提升"))'
    ws.cell(row=11, column=1).fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")

    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions['A'].width = 80
    instructions = """F8 讨喜值自评卡 - 使用说明

【什么是讨喜？】
讨喜不是讨好，而是让对方自然地喜欢你、信任你、愿意与你交流。
高讨喜值的人往往更善于说服，因为对方愿意倾听。

【四个维度解析】

1. 专注倾听
→ 核心：让对方感到被重视
→ 关键：全神贯注 + 眼神交流 + 适当回应

2. 镜像模仿
→ 核心：让对方感到熟悉和舒适
→ 关键：无意识地建立相似性
→ 注意：模仿要自然，不能夸张或刻意

3. 上堆下切
→ 核心：让对话更有深度和广度
→ 上堆：从具体到抽象（价值观、意义）
→ 下切：从抽象到具体（例子、做法）
→ 平切：横向探索更多可能性

4. 给予关注
→ 核心：让对方感到自己是特别的
→ 关键：记住细节、提及之前对话、真诚赞美

【评分标准】
- 1分：完全做不到
- 2分：偶尔做到
- 3分：有时做到
- 4分：经常做到
- 5分：总是自然做到

【提升建议】
- 每天练习一个维度
- 找人对话时录像回看
- 收集反馈并调整"""
    ws2.cell(row=1, column=1).value = instructions
    ws2.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[1].height = 380

    filepath = os.path.join(OUTPUT_DIR, "F8_讨喜值自评卡.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_f9_xlsx():
    """F9: 六种观点速查卡"""
    wb = Workbook()
    ws = wb.active
    ws.title = "六种观点速查"

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35

    ws.merge_cells('A1:D1')
    set_cell(ws, 1, 1, "F9：六种观点速查卡", bold=True, font_color="FFFFFF", bg_color="34495E", font_size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    set_cell(ws, 2, 1, "六种观点类型的快速参考", bg_color="ECF0F1", font_size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["观点类型", "核心关键词", "适用情境", "话术示例"]
    for col, header in enumerate(headers, start=1):
        set_cell(ws, 4, col, header, bold=True, bg_color="34495E", font_color="FFFFFF", alignment=Alignment(horizontal="center"))

    viewpoints = [
        ("1. 红色观点（情感感受）", "感受、情绪、感觉", "当对方处于情绪中\n当需要建立情感连接\n当逻辑说服效果不好时", "我能感受到你的挫败感...\n换成是你关心的人遇到这种情况，你会希望他怎么做？"),
        ("2. 硬货币（实际利益）", "利益、成本、收益、风险", "当对方理性导向\n当涉及金钱/资源决策\n当需要具体数据支撑", "这个方案能为你节省30%的成本\n投资回报周期是6个月"),
        ("3. 速度（时间紧迫）", "紧迫、截止、限时、窗口期", "当有时间压力\n当机会稍纵即逝\n当需要快速决策", "这个优惠仅限本周\n名额只剩3个\n错过就要等明年"),
        ("4. 距离（与问题远近）", "远近、亲疏、关联度", "当对方觉得自己不受影响\n当需要拉近关系\n当需要强调相关性", "虽然这是发生在别人身上，但...\n这个问题离我们其实很近"),
        ("5. 专业度（权威背书）", "专业、权威、资质、经验", "当需要建立可信度\n当涉及专业领域\n当需要消除疑虑", "根据我们10年的行业经验...\n这个方案已帮助500+企业成功转型"),
        ("6. 故事（叙事说服）", "故事、案例、经历", "当抽象概念难以理解\n当需要情感共鸣\n当需要生动说明", "我有一个客户，之前遇到一模一样的问题...\n他是怎么从困境中走出来的"),
    ]

    colors = ["FADBD8", "FDEBD0", "FEF9E7", "E8F8F5", "EBF5FB", "F5EEF8"]

    for i, (view_type, keywords, situation, example) in enumerate(viewpoints, start=5):
        set_cell(ws, i, 1, view_type, bold=True, bg_color=colors[i-5], font_size=11)
        set_cell(ws, i, 2, keywords, bg_color=colors[i-5], alignment=Alignment(horizontal="center"))
        ws.cell(row=i, column=3).value = situation
        ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=4).value = example
        ws.cell(row=i, column=4).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 65

    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions['A'].width = 80
    instructions = """F9 六种观点速查卡 - 使用说明

【六种观点类型】

1. 红色观点（情感感受）
→ 核心：诉诸情感
→ 适用：情绪主导的决策
→ 注意：情感要真诚，不要操控

2. 硬货币（实际利益）
→ 核心：诉诸利益
→ 适用：理性决策、商业决策
→ 注意：数据要准确，不要夸大

3. 速度（时间紧迫）
→ 核心：诉诸紧迫感
→ 适用：限时优惠、机会窗口
→ 注意：真实限时，不要虚假催促

4. 距离（与问题远近）
→ 核心：诉诸关联性
→ 适用：对方觉得自己不受影响时
→ 注意：找到真实的关联点

5. 专业度（权威背书）
→ 核心：诉诸权威
→ 适用：需要建立信任时
→ 注意：权威要真实，不要造假

6. 故事（叙事说服）
→ 核心：诉诸故事
→ 适用：需要生动说明时
→ 注意：故事要真实，不能编造

【选用原则】
- 了解受众：不同人接受不同的观点类型
- 组合使用：可以同时用2-3种观点
- 避免单一：不要只用一种观点说服"""
    ws2.cell(row=1, column=1).value = instructions
    ws2.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[1].height = 350

    filepath = os.path.join(OUTPUT_DIR, "F9_六种观点速查卡.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_f10_xlsx():
    """F10: 综合运用检核表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "说服全流程检核"

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30

    ws.merge_cells('A1:C1')
    set_cell(ws, 1, 1, "F10：综合运用检核表", bold=True, font_color="FFFFFF", bg_color="2C3E50", font_size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:C2')
    set_cell(ws, 2, 1, "说服全流程检核", bg_color="ECF0F1", font_size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["阶段", "检核要点", "完成情况"]
    for col, header in enumerate(headers, start=1):
        set_cell(ws, 4, col, header, bold=True, bg_color="2C3E50", font_color="FFFFFF", alignment=Alignment(horizontal="center"))

    stages = [
        ("1. 准备阶段", "□ 明确说服目标（具体、可衡量）\n□ 分析受众（谁、关心什么、顾虑什么）\n□ 准备论据和证据\n□ 选择观点类型（情感/利益/权威/故事）\n□ 准备应对异议的答案", "未做 / 部分 / 完成"),
        ("2. 开场阶段", "□ 建立良好氛围\n□ 引起对方注意和兴趣\n□ 表明来意，不要绕弯子\n□ 说明对话的目的和价值", "未做 / 部分 / 完成"),
        ("3. 探询阶段", "□ 提问了解对方想法（F5小火苗话术）\n□ 倾听对方的顾虑和反对意见\n□ 确认自己理解正确\n□ 找到共同点和利益交汇点", "未做 / 部分 / 完成"),
        ("4. 建议阶段", "□ 清晰表达自己的建议\n□ 使用我需要而非你应该（F3话术）\n□ 说明对对方的好处\n□ 避免失败句型（F4检测表）", "未做 / 部分 / 完成"),
        ("5. 处理异议", "□ 不要回避异议\n□ 先认可对方的顾虑\n□ 用一致性原理引导（F6）\n□ 提供选择而非命令\n□ 不争论对错，关注解决方案", "未做 / 部分 / 完成"),
        ("6. 达成共识", "□ 总结双方一致的地方\n□ 确认下一步行动\n□ 明确责任人和时间\n□ 约定后续跟进方式", "未做 / 部分 / 完成"),
        ("7. 跟进阶段", "□ 按约定时间跟进\n□ 感谢对方的配合\n□ 兑现承诺\n□ 维护长期关系\n□ 记录经验教训", "未做 / 部分 / 完成"),
    ]

    for i, (stage, checklist, status) in enumerate(stages, start=5):
        set_cell(ws, i, 1, stage, bold=True, bg_color="ECF0F1")
        ws.cell(row=i, column=2).value = checklist
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3).value = status
        ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True, horizontal="center")
        ws.row_dimensions[i].height = 70

    ws.merge_cells('A13:C13')
    set_cell(ws, 13, 1, "【整体自评】本次说服效果：成功 / 部分成功 / 未成功  原因分析：", bold=True, bg_color="2C3E50", font_color="FFFFFF")

    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions['A'].width = 80
    instructions = """F10 综合运用检核表 - 使用说明

【使用流程】

1. 说服前：完成准备阶段检核
→ 确保有明确的目标和受众分析
→ 准备好应对各种情况

2. 说服中：根据阶段对照检核
→ 每完成一个阶段打勾确认
→ 如有遗漏，及时补充

3. 说服后：完成跟进阶段和整体自评
→ 总结本次说服的经验教训
→ 为下次说服积累

【检核等级说明】

未做：完全没有考虑这个环节
部分：考虑但执行不完整
完成：考虑周全且执行到位

【关键提示】

- 准备阶段是最重要的，磨刀不误砍柴工
- 开场不要急于推销，先建立关系
- 探询比说服更重要，多听少说
- 处理异议时不要争论，先认可再引导
- 达成共识后一定要明确下一步行动
- 跟进是建立长期关系的关键"""
    ws2.cell(row=1, column=1).value = instructions
    ws2.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[1].height = 350

    filepath = os.path.join(OUTPUT_DIR, "F10_综合运用检核表.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


if __name__ == "__main__":
    create_f1_xlsx()
    create_f2_xlsx()
    create_f3_xlsx()
    create_f4_xlsx()
    create_f5_xlsx()
    create_f6_xlsx()
    create_f7_xlsx()
    create_f8_xlsx()
    create_f9_xlsx()
    create_f10_xlsx()
    print("\nAll 10 XLSX files created successfully!")

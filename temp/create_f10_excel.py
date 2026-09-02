import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
alt_row_fill = PatternFill(start_color="D6E3F8", end_color="D6E3F8", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
summary_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
title_font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
normal_font = Font(name="微软雅黑", size=10)
bold_font = Font(name="微软雅黑", size=10, bold=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# Sheet1: 月度成本台账
# ============================================================
ws1 = wb.active
ws1.title = "月度成本台账"

ws1.merge_cells('A1:I1')
ws1['A1'] = '企业AI成本治理台账 — 月度成本明细'
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align
ws1.row_dimensions[1].height = 30

headers = ["部门", "应用", "1月", "2月", "3月", "累计成本", "效果指标", "优化措施", "备注"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws1.row_dimensions[3].height = 22

data = [
    ["技术中心", "智能客服", 12500, 13200, 14800, None, "响应时间<2s, 接通率95%", "引入缓存机制+模型压缩", ""],
    ["技术中心", "代码审查AI", 8500, 9100, 9800, None, "代码缺陷率下降18%", "优化审查范围", ""],
    ["运营部", "数据清洗", 6200, 6800, 7200, None, "效率提升3倍", "调整清洗策略", ""],
    ["运营部", "报表生成", 4800, 5100, 5600, None, "人力节省40%", "模板标准化", ""],
    ["市场部", "文案生成", 7800, 8200, 9100, None, "内容产出+60%", "提示词优化", ""],
    ["市场部", "舆情分析", 5400, 5800, 6300, None, "预警准确率92%", "模型微调", ""],
    ["财务部", "发票识别", 3200, 3500, 3800, None, "识别准确率99%", "增加训练样本", ""],
    ["财务部", "预算预测", 4100, 4300, 4700, None, "预测偏差<5%", "引入外部数据", ""],
    ["人力资源部", "简历筛选", 2800, 3100, 3400, None, "筛选效率+80%", "关键词权重调整", ""],
    ["人力资源部", "培训推荐", 1900, 2100, 2300, None, "员工满意度+25%", "学习路径优化", ""],
]

for row_idx, row_data in enumerate(data, 4):
    fill = alt_row_fill if row_idx % 2 == 0 else white_fill
    for col, val in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col, value=val)
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        if col in [3, 4, 5]:
            cell.alignment = center_align
            cell.number_format = '#,##0'
        elif col == 6:
            cell.value = f"=C{row_idx}+D{row_idx}+E{row_idx}"
            cell.font = bold_font
            cell.alignment = center_align
            cell.number_format = '#,##0'
            cell.fill = green_fill
        elif col in [7, 8]:
            cell.alignment = left_align
        else:
            cell.alignment = center_align

summary_row = len(data) + 4
ws1.merge_cells(f'A{summary_row}:B{summary_row}')
ws1.cell(row=summary_row, column=1, value="合计").font = bold_font
ws1.cell(row=summary_row, column=1).fill = summary_fill
ws1.cell(row=summary_row, column=1).alignment = center_align
ws1.cell(row=summary_row, column=1).border = thin_border
ws1.cell(row=summary_row, column=2).fill = summary_fill
ws1.cell(row=summary_row, column=2).border = thin_border

for col in [3, 4, 5, 6]:
    cell = ws1.cell(row=summary_row, column=col, value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{summary_row-1})")
    cell.font = bold_font
    cell.fill = summary_fill
    cell.alignment = center_align
    cell.border = thin_border
    cell.number_format = '#,##0'

for col in [7, 8, 9]:
    cell = ws1.cell(row=summary_row, column=col)
    cell.fill = summary_fill
    cell.border = thin_border

col_widths = [14, 12, 10, 10, 10, 12, 22, 22, 10]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = 'A4'

# ============================================================
# Sheet2: 成本分析汇总
# ============================================================
ws2 = wb.create_sheet("成本分析汇总")

ws2.merge_cells('A1:H1')
ws2['A1'] = '企业AI成本分析汇总'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align
ws2.row_dimensions[1].height = 30

ws2.cell(row=3, column=1, value="一、各部门AI支出占比").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

dept_headers = ["部门", "1月", "2月", "3月", "季度合计", "占比", "累计效果得分"]
for col, h in enumerate(dept_headers, 1):
    cell = ws2.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

dept_data = [
    ["技术中心", 21000, 22300, 24600],
    ["运营部", 11000, 11900, 12800],
    ["市场部", 13200, 14000, 15400],
    ["财务部", 7300, 7800, 8500],
    ["人力资源部", 4700, 5200, 5700],
]

for row_idx, row_data in enumerate(dept_data, 5):
    fill = alt_row_fill if row_idx % 2 == 0 else white_fill
    ws2.cell(row=row_idx, column=1, value=row_data[0]).font = normal_font
    ws2.cell(row=row_idx, column=1).fill = fill
    ws2.cell(row=row_idx, column=1).border = thin_border
    ws2.cell(row=row_idx, column=1).alignment = center_align

    for col_i, val in enumerate(row_data[1:], 2):
        cell = ws2.cell(row=row_idx, column=col_i, value=val)
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center_align
        cell.number_format = '#,##0'

    tot_cell = ws2.cell(row=row_idx, column=5, value=f"=SUM(B{row_idx}:D{row_idx})")
    tot_cell.font = bold_font
    tot_cell.fill = green_fill
    tot_cell.border = thin_border
    tot_cell.number_format = '#,##0'
    tot_cell.alignment = center_align

    pct_cell = ws2.cell(row=row_idx, column=6, value=f"=E{row_idx}/$E$10")
    pct_cell.font = normal_font
    pct_cell.fill = fill
    pct_cell.border = thin_border
    pct_cell.number_format = '0.0%'
    pct_cell.alignment = center_align

    score_cell = ws2.cell(row=row_idx, column=7, value=round(85 + row_idx * 1.5, 1))
    score_cell.font = normal_font
    score_cell.fill = fill
    score_cell.border = thin_border
    score_cell.alignment = center_align

total_row = 10
for col in range(1, 8):
    cell = ws2.cell(row=total_row, column=col)
    cell.fill = summary_fill
    cell.border = thin_border
    cell.font = bold_font
ws2.cell(row=total_row, column=1, value="合计").alignment = center_align
for col in [2, 3, 4, 5]:
    cell = ws2.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}9)")
    cell.number_format = '#,##0'
    cell.alignment = center_align
ws2.cell(row=total_row, column=6, value="=E10/E10").number_format = '0.0%'
ws2.cell(row=total_row, column=6).alignment = center_align
ws2.cell(row=total_row, column=7, value="=SUM(G5:G9)").alignment = center_align

# Section B: Monthly trend
trend_start = 13
ws2.cell(row=trend_start, column=1, value="二、月度成本趋势").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

trend_headers = ["月份", "技术中心", "运营部", "市场部", "财务部", "人力资源部", "当月总计"]
for col, h in enumerate(trend_headers, 1):
    cell = ws2.cell(row=trend_start+1, column=col, value=h)
    cell.font = header_font
    cell.fill = subheader_fill
    cell.alignment = center_align
    cell.border = thin_border

trend_data = [
    ["1月", 21000, 11000, 13200, 7300, 4700],
    ["2月", 22300, 11900, 14000, 7800, 5200],
    ["3月", 24600, 12800, 15400, 8500, 5700],
]

for row_idx, row_data in enumerate(trend_data, trend_start+2):
    fill = alt_row_fill if row_idx % 2 == 0 else white_fill
    ws2.cell(row=row_idx, column=1, value=row_data[0]).font = normal_font
    ws2.cell(row=row_idx, column=1).fill = fill
    ws2.cell(row=row_idx, column=1).border = thin_border
    ws2.cell(row=row_idx, column=1).alignment = center_align

    for col_i, val in enumerate(row_data[1:], 2):
        cell = ws2.cell(row=row_idx, column=col_i, value=val)
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center_align
        cell.number_format = '#,##0'

    tot_cell = ws2.cell(row=row_idx, column=7, value=f"=SUM(B{row_idx}:F{row_idx})")
    tot_cell.font = bold_font
    tot_cell.fill = green_fill
    tot_cell.border = thin_border
    tot_cell.number_format = '#,##0'
    tot_cell.alignment = center_align

sum_row = trend_start + 5
for col in range(1, 8):
    cell = ws2.cell(row=sum_row, column=col)
    cell.fill = summary_fill
    cell.border = thin_border
    cell.font = bold_font
ws2.cell(row=sum_row, column=1, value="季度总计").alignment = center_align
for col in range(2, 8):
    cell = ws2.cell(row=sum_row, column=col, value=f"=SUM({get_column_letter(col)}{trend_start+2}:{get_column_letter(col)}{trend_start+4})")
    cell.number_format = '#,##0'
    cell.alignment = center_align

# Section C: Cost composition
cost_start = 21
ws2.cell(row=cost_start, column=1, value="三、成本构成分析（季度总成本占比）").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

cost_items = ["模型调用费", "向量数据库", "人工干预", "基础设施", "其他"]
cost_values = [58, 22, 12, 6, 2]
for col, h in enumerate(["成本项", "占比", "金额估算（元）"], 1):
    cell = ws2.cell(row=cost_start+1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, (item, val) in enumerate(zip(cost_items, cost_values), cost_start+2):
    fill = alt_row_fill if row_idx % 2 == 0 else white_fill
    ws2.cell(row=row_idx, column=1, value=item).font = normal_font
    ws2.cell(row=row_idx, column=1).fill = fill
    ws2.cell(row=row_idx, column=1).border = thin_border
    ws2.cell(row=row_idx, column=1).alignment = center_align

    ws2.cell(row=row_idx, column=2, value=val/100).font = normal_font
    ws2.cell(row=row_idx, column=2).fill = fill
    ws2.cell(row=row_idx, column=2).border = thin_border
    ws2.cell(row=row_idx, column=2).number_format = '0.0%'
    ws2.cell(row=row_idx, column=2).alignment = center_align

    ws2.cell(row=row_idx, column=3, value=f"=E10*{val/100}").font = normal_font
    ws2.cell(row=row_idx, column=3).fill = fill
    ws2.cell(row=row_idx, column=3).border = thin_border
    ws2.cell(row=row_idx, column=3).number_format = '#,##0'
    ws2.cell(row=row_idx, column=3).alignment = center_align

# Pie chart
pie = PieChart()
pie.title = "成本构成占比"
pie.style = 10
labels = Reference(ws2, min_col=1, min_row=cost_start+2, max_row=cost_start+6)
pie_data = Reference(ws2, min_col=2, min_row=cost_start+1, max_row=cost_start+6)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(labels)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
pie.width = 14
pie.height = 10
ws2.add_chart(pie, f"E{cost_start}")

col_widths2 = [16, 12, 12, 12, 12, 12, 14]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A3'

# ============================================================
# Sheet3: 优化记录
# ============================================================
ws3 = wb.create_sheet("优化记录")

ws3.merge_cells('A1:F1')
ws3['A1'] = 'AI成本优化记录台账'
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align
ws3.row_dimensions[1].height = 30

opt_headers = ["序号", "问题描述", "原因分析", "优化措施", "效果", "节约金额（元）"]
for col, h in enumerate(opt_headers, 1):
    cell = ws3.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws3.row_dimensions[3].height = 22

opt_data = [
    [1, "客服模型调用频次过高", "未启用上下文缓存，重复请求多", "引入Redis缓存层+请求去重", "调用量下降40%", 8500],
    [2, "代码审查范围过宽", "全量代码扫描，资源浪费", "按变更范围智能触发", "成本降低35%", 6200],
    [3, "报表生成重复计算", "缺少结果缓存机制", "结果缓存+增量更新", "计算资源节省50%", 4800],
    [4, "文案生成模型过大", "使用旗舰模型处理简单任务", "分级模型策略（GPT-4o/4o-mini）", "成本下降45%", 9200],
    [5, "数据清洗冗余处理", "重复清洗同一批次数据", "布隆过滤器去重+增量处理", "效率提升60%", 3600],
    [6, "舆情分析实时性过强", "高频轮询造成资源浪费", "改为事件触发+批处理", "成本降低30%", 4100],
    [7, "发票识别训练样本冗余", "样本重复度高", "去重+多样性采样", "训练成本下降25%", 2800],
    [8, "预算预测模型过拟合", "特征工程复杂度过高", "简化特征+正则化", "推理成本-40%", 3300],
]

for row_idx, row_data in enumerate(opt_data, 4):
    fill = alt_row_fill if row_idx % 2 == 0 else white_fill
    for col, val in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col, value=val)
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        if col == 1:
            cell.alignment = center_align
        elif col == 6:
            cell.alignment = center_align
            cell.number_format = '#,##0'
            cell.fill = green_fill
            cell.font = bold_font
        else:
            cell.alignment = left_align
    ws3.row_dimensions[row_idx].height = 18

total_savings_row = len(opt_data) + 4
ws3.merge_cells(f'A{total_savings_row}:E{total_savings_row}')
ws3.cell(row=total_savings_row, column=1, value="累计节约金额").font = bold_font
ws3.cell(row=total_savings_row, column=1).fill = summary_fill
ws3.cell(row=total_savings_row, column=1).alignment = center_align
ws3.cell(row=total_savings_row, column=1).border = thin_border
for col in range(2, 6):
    ws3.cell(row=total_savings_row, column=col).fill = summary_fill
    ws3.cell(row=total_savings_row, column=col).border = thin_border

savings_cell = ws3.cell(row=total_savings_row, column=6, value=f"=SUM(F4:F{total_savings_row-1})")
savings_cell.font = Font(name="微软雅黑", size=12, bold=True, color="375623")
savings_cell.fill = light_green_fill
savings_cell.border = thin_border
savings_cell.number_format = '#,##0'
savings_cell.alignment = center_align

opt_widths = [6, 30, 28, 28, 20, 16]
for i, w in enumerate(opt_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = 'A4'

# ============================================================
# Sheet4: 总览摘要 (Summary)
# ============================================================
ws4 = wb.create_sheet("总览摘要", 0)

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 20

ws4.merge_cells('A1:D1')
ws4['A1'] = '企业AI成本治理台账 — 总览摘要'
ws4['A1'].font = Font(name="微软雅黑", size=16, bold=True, color="1F4E79")
ws4['A1'].alignment = center_align
ws4.row_dimensions[1].height = 35

ws4.merge_cells('A2:D2')
ws4['A2'] = '统计周期：2026年Q1（1月—3月）'
ws4['A2'].font = Font(name="微软雅黑", size=10, color="7F7F7F")
ws4['A2'].alignment = center_align

section_start = 4
ws4.cell(row=section_start, column=1, value="核心指标总览").font = Font(name="微软雅黑", size=13, bold=True, color="1F4E79")

header_row = section_start + 1
for col, h in enumerate(["指标名称", "当前值", "单位", "说明"], 1):
    cell = ws4.cell(row=header_row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

kpi_data = [
    ["季度AI总投入", "=成本分析汇总!E10", "#,##0", "元", "Q1季度各业务线AI投入总计"],
    ["累计节约金额", "=优化记录!F12", "#,##0", "元", "Q1季度通过优化措施累计节约"],
    ["优化项目数", 8, "0", "个", "本季度已执行的优化改进项目"],
    ["成本节约率", "=优化记录!F12/成本分析汇总!E10", "0.0%", "", "节约金额占总投入的比例"],
    ["覆盖部门数", 5, "0", "个", "已接入AI系统的部门数量"],
    ["AI应用总数", 10, "0", "个", "各部门正在使用的AI应用数量"],
]

for row_idx, (name, val, fmt, unit, note) in enumerate(kpi_data, section_start+2):
    fill = alt_row_fill if row_idx % 2 == 0 else white_fill
    cell = ws4.cell(row=row_idx, column=1, value=name)
    cell.font = normal_font
    cell.fill = fill
    cell.border = thin_border
    cell.alignment = left_align

    val_cell = ws4.cell(row=row_idx, column=2, value=val)
    val_cell.font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
    val_cell.fill = fill
    val_cell.border = thin_border
    val_cell.alignment = center_align
    val_cell.number_format = fmt

    ws4.cell(row=row_idx, column=3, value=unit).font = normal_font
    ws4.cell(row=row_idx, column=3).fill = fill
    ws4.cell(row=row_idx, column=3).border = thin_border
    ws4.cell(row=row_idx, column=3).alignment = center_align

    note_cell = ws4.cell(row=row_idx, column=4, value=note)
    note_cell.font = Font(name="微软雅黑", size=9, color="7F7F7F")
    note_cell.fill = fill
    note_cell.border = thin_border
    note_cell.alignment = left_align

dept_section = section_start + 10
ws4.cell(row=dept_section, column=1, value="部门成本明细").font = Font(name="微软雅黑", size=13, bold=True, color="1F4E79")

for col, h in enumerate(["部门", "季度成本（元）", "占比", "效果得分"], 1):
    cell = ws4.cell(row=dept_section+1, column=col, value=h)
    cell.font = header_font
    cell.fill = subheader_fill
    cell.alignment = center_align
    cell.border = thin_border

dept_summary = [
    ["技术中心", "=成本分析汇总!E5", "=成本分析汇总!F5", "=成本分析汇总!G5"],
    ["运营部", "=成本分析汇总!E6", "=成本分析汇总!F6", "=成本分析汇总!G6"],
    ["市场部", "=成本分析汇总!E7", "=成本分析汇总!F7", "=成本分析汇总!G7"],
    ["财务部", "=成本分析汇总!E8", "=成本分析汇总!F8", "=成本分析汇总!G8"],
    ["人力资源部", "=成本分析汇总!E9", "=成本分析汇总!F9", "=成本分析汇总!G9"],
]

for row_idx, row_data in enumerate(dept_summary, dept_section+2):
    fill = alt_row_fill if row_idx % 2 == 0 else white_fill
    for col, val in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col, value=val)
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        if col == 1:
            cell.alignment = center_align
        elif col == 2:
            cell.number_format = '#,##0'
            cell.alignment = center_align
        elif col == 3:
            cell.number_format = '0.0%'
            cell.alignment = center_align
        else:
            cell.alignment = center_align

opt_section = dept_section + 9
ws4.cell(row=opt_section, column=1, value="优化成效亮点").font = Font(name="微软雅黑", size=13, bold=True, color="1F4E79")

opt_highlights = [
    "模型分级策略效果显著：文案生成成本下降45%，保持质量的同时大幅节约",
    "缓存机制全面推广：客服调用量下降40%，数据清洗效率提升60%",
    "智能触发机制：代码审查成本降低35%，从全量扫描转为增量审查",
    "累计节约金额已达3.25万元，预计Q2通过持续优化可再节约20%以上",
]

for row_idx, text in enumerate(opt_highlights, opt_section+1):
    ws4.merge_cells(f'A{row_idx}:D{row_idx}')
    cell = ws4.cell(row=row_idx, column=1, value="* " + text)
    cell.font = Font(name="微软雅黑", size=10, color="375623")
    cell.fill = green_fill
    cell.border = thin_border
    cell.alignment = left_align
    ws4.row_dimensions[row_idx].height = 20

footer_row = opt_section + 7
ws4.merge_cells(f'A{footer_row}:D{footer_row}')
ws4.cell(row=footer_row, column=1, value="数据更新时间：2026年3月  |  编制：数字化转型办公室  |  审核：IT总监").font = Font(name="微软雅黑", size=9, color="7F7F7F")
ws4.cell(row=footer_row, column=1).alignment = center_align

# Save
output_path = "D:/新课开发/数字化转型/10.AI规模化的成本与治理新挑战：Token经济与架构选择/全流程工具表单/F10_企业AI成本治理台账.xlsx"
wb.save(output_path)
print(f"File saved: {output_path}")

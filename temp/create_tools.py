import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

OUTPUT_DIR = "D:/新课开发/工作手册/客户隐性需求挖掘与验证/完整课程包/06_工具表单"
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADER_BG   = "C00000"
HEADER_FG   = "FFFFFF"
TITLE_BG    = "E85053"
SECTION_BG  = "F2F2F2"
ALT_ROW_BG  = "FFF0F0"
INPUT_BG    = "DAEEF3"
BORDER_CLR  = "AAAAAA"

def hdr_fill(color): return PatternFill("solid", fgColor=color)
def thin_border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def apply_header(ws, row, cols, texts, bg=HEADER_BG, fg=HEADER_FG, height=22):
    for col, text in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(bold=True, color=fg, size=11)
        c.fill      = hdr_fill(bg)
        c.alignment = center()
        c.border    = thin_border()
    ws.row_dimensions[row].height = height

def apply_title(ws, row, cols, text, span_cols=None, height=28):
    if span_cols:
        ws.merge_cells(start_row=row, start_column=span_cols[0],
                       end_row=row, end_column=span_cols[1])
    c = ws.cell(row=row, column=cols[0], value=text)
    c.font      = Font(bold=True, color=HEADER_FG, size=13)
    c.fill      = hdr_fill(TITLE_BG)
    c.alignment = center()
    c.border    = thin_border()
    ws.row_dimensions[row].height = height

def apply_section(ws, row, col, text, ncols, height=18):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+ncols-1)
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, size=10)
    c.fill      = hdr_fill(SECTION_BG)
    c.alignment = left()
    c.border    = thin_border()
    ws.row_dimensions[row].height = height

def apply_data_row(ws, row, col, text, style="normal", alt=False, height=18):
    bg = ALT_ROW_BG if alt else "FFFFFF"
    if style == "input": bg = INPUT_BG
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(size=10, color="0000FF" if style == "input" else "000000")
    c.fill      = hdr_fill(bg)
    c.alignment = left()
    c.border    = thin_border()
    ws.row_dimensions[row].height = height
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_sheet_header(ws, title):
    ws.row_dimensions[1].height = 8
    apply_title(ws, 2, [1], title, span_cols=(1, 6), height=30)
    ws.row_dimensions[3].height = 8

def add_instructions(ws, instructions, start_row=4):
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=6)
    c = ws.cell(row=start_row, column=1, value=instructions)
    c.font      = Font(italic=True, size=9, color="595959")
    c.alignment = left()
    ws.row_dimensions[start_row].height = 30
    return start_row + 1

def create_file1(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("访谈前准备清单")
    set_col_widths(ws, [4, 28, 50, 18, 18])
    add_sheet_header(ws, "【工具一】隐性需求挖掘访谈工具包 - 访谈前准备清单")
    r = add_instructions(ws, "使用说明：进入访谈前，逐一核对以下准备项，全部完成后再出发。浅蓝色单元格为待填写输入区。")

    apply_header(ws, r, [1,2,3,4,5], ["序号","准备事项","具体内容/备注","状态","检查人"])
    ws.row_dimensions[r].height = 20
    r += 1

    items = [
        (1,"背景调研","查阅对接人的岗位和司龄，判断是否经历过与本次采购相关的历史事故","未完成,进行中,已完成",""),
        (2,"组织分析","查阅本次采购在客户组织内部会影响到哪些岗位，列出可能的决策人、使用者、影响者","未完成,进行中,已完成",""),
        (3,"场景式问题准备","准备2-3个具体场景式问题，避免通篇使用总结式提问","未完成,进行中,已完成",""),
        (4,"历史案例回顾","查阅公司内是否有类似项目后来的效果不理想，记录关键细节备用","未完成,进行中,已完成",""),
        (5,"关键决策人确认","确认除对接人外，还需要过哪些人的审批或知情","未完成,进行中,已完成",""),
        (6,"最挑剔人选分析","预判谁最可能挑剔、最抵触本次方案，准备针对性应对策略","未完成,进行中,已完成",""),
        (7,"访谈目标明确","明确本次访谈要验证的核心假设是什么，填写在右侧","未完成,进行中,已完成",""),
    ]

    for i, (num, item, detail, status, checker) in enumerate(items):
        alt = (i % 2 == 0)
        apply_data_row(ws, r, 1, num, "normal", alt)
        apply_data_row(ws, r, 2, item, "normal", alt)
        c = ws.cell(row=r, column=3, value=detail)
        c.font = Font(size=10); c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
        c.alignment = left(); c.border = thin_border()
        ws.row_dimensions[r].height = 18
        apply_data_row(ws, r, 4, status, "input", alt)
        apply_data_row(ws, r, 5, checker, "input", alt)
        r += 1

    dv = DataValidation(type="list", formula1='"未完成,进行中,已完成"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.sqref = f"D6:D{r-1}"

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value="注：状态栏可使用下拉菜单选择：未完成/进行中/已完成")
    c.font = Font(italic=True, size=9, color="595959")
    c.alignment = left()
    ws.freeze_panes = "A5"

    ws2 = wb.create_sheet("核心提问模板")
    set_col_widths(ws2, [4, 22, 50, 22, 22])
    add_sheet_header(ws2, "【工具一】隐性需求挖掘访谈工具包 - 核心提问模板")
    r2 = add_instructions(ws2, "使用说明：以下提问模板按访谈逻辑分类。蓝色字体为参考提问，访谈时根据现场情况灵活选用，不必照本宣科。")

    apply_header(ws2, r2, [1,2,3,4,5], ["序号","提问类型","参考提问模板","使用时机","注意事项"])
    ws2.row_dimensions[r2].height = 20
    r2 += 1

    questions = [
        (1,"打开话题","你们现在是怎么处理这件事的，方便具体说说这个流程吗？","开场破冰","降低对方防御，建立信任感"),
        (2,"打开话题","上一次遇到类似问题的时候，是怎么解决的？","开场破冰","引导具体场景，而非抽象理念"),
        (3,"打开话题","如果这次项目最后特别顺利，对你个人来说，会有什么不一样？","开场破冰","探测个人利益关联度"),
        (4,"探测风险规避","如果这套方案上线之后出了点小问题，一般是谁会最先被问到？","方案讨论阶段","不要直接问"你担心什么""),
        (5,"探测风险规避","公司里以前有没有类似的项目，后来效果不太理想的？","方案讨论阶段","不追问具体细节，观察对方愿不愿意展开讲"),
        (6,"探测决策链","除了您，这件事最后拍板还需要过哪些人？","方案讨论阶段","识别真正的决策者"),
        (7,"探测决策链","真正每天要用这套东西的，主要是哪个团队？","方案讨论阶段","识别使用者角色"),
        (8,"收尾追问","对了，这套系统真正上线之后，你们内部谁会最不适应？","访谈收尾阶段","防御最松弛时的非正式追问"),
    ]

    for i, (num, qtype, template, timing, note) in enumerate(questions):
        alt = (i % 2 == 0)
        apply_data_row(ws2, r2, 1, num, "normal", alt)
        apply_data_row(ws2, r2, 2, qtype, "normal", alt)
        c = ws2.cell(row=r2, column=3, value=template)
        c.font = Font(size=10, bold=True, color="C00000"); c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.alignment = left(); c.border = thin_border(); ws2.row_dimensions[r2].height = 22
        apply_data_row(ws2, r2, 4, timing, "normal", alt)
        apply_data_row(ws2, r2, 5, note, "normal", alt)
        r2 += 1
    ws2.freeze_panes = "A5"

    ws3 = wb.create_sheet("追问技巧提示卡")
    set_col_widths(ws3, [4, 24, 52, 16])
    add_sheet_header(ws3, "【工具一】隐性需求挖掘访谈工具包 - 追问技巧提示卡")
    r3 = add_instructions(ws3, "使用说明：本提示卡用于访谈过程中应对常见情况。将提示卡打印或投影使用，随时对照。")

    apply_header(ws3, r3, [1,2,3,4], ["序号","场景","应对技巧","优先级"])
    ws3.row_dimensions[r3].height = 20
    r3 += 1

    tips = [
        (1,"对方给出笼统回答（如整体还算配合、应该没什么问题）",
         "不要直接质疑，改用具体化追问：具体是哪个环节配合得比较好？把对方的笼统表态拉回具体场景。真话往往藏在具体细节里，不藏在总结句里。","高"),
        (2,"对方出现明显停顿、语气变化",
         "不要当场逼问，先记下这个时间点和上下文，留到复盘或下一次非正式沟通里再回头验证。强行当场追问容易让对方彻底关闭话匣子。","高"),
        (3,"对方岔开话题",
         "不要强行拉回，顺着话题自然过渡，在复盘笔记中记录话题转移的节点和可能原因。","中"),
        (4,"对方给出明显矛盾的回答",
         "不要指出矛盾，而是换一个角度重新提问，让对方在不同表述中自然暴露真实想法。","中"),
        (5,"对方表示应该没什么大问题",
         "语气偏轻、缺乏具体细节支撑时，值得留意。继续追问：方便展开说说吗？真正没问题的表态通常会带一两句具体理由。","高"),
        (6,"对方说我们内部对这个还有点讨论",
         "通常意味着决策链中存在尚未被说服的关键人物。建议主动询问：具体是哪个层面的讨论？技术层/执行层/管理层？","高"),
        (7,"对方说这个我们再看看流程",
         "通常意味着预算审批或内部流程中存在未言明的障碍。建议追问：流程涉及哪些部门？哪个环节最可能卡住？","高"),
    ]

    for i, (num, scene, technique, priority) in enumerate(tips):
        alt = (i % 2 == 0)
        apply_data_row(ws3, r3, 1, num, "normal", alt)
        c = ws3.cell(row=r3, column=2, value=scene)
        c.font = Font(size=10); c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.alignment = left(); c.border = thin_border(); ws3.row_dimensions[r3].height = 52
        c = ws3.cell(row=r3, column=3, value=technique)
        c.font = Font(size=10); c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.alignment = left(); c.border = thin_border()
        priority_color = "C00000" if priority == "高" else "000000"
        c = ws3.cell(row=r3, column=4, value=priority)
        c.font = Font(size=10, bold=True, color=priority_color)
        c.fill = hdr_fill("FFE0E0" if priority == "高" else "E2EFDA")
        c.alignment = center(); c.border = thin_border()
        r3 += 1
    ws3.freeze_panes = "A5"

    ws4 = wb.create_sheet("访谈记录表")
    set_col_widths(ws4, [18, 14, 14, 14, 40])
    add_sheet_header(ws4, "【工具一】隐性需求挖掘访谈工具包 - 访谈记录表")
    r4 = add_instructions(ws4, "使用说明：访谈结束后尽快填写，原始记录不加工。浅蓝色区域为输入区。")

    apply_header(ws4, r4, [1,2,3,4,5], ["访谈主题","被访者姓名","被访者职位","访谈日期","访谈地点"])
    ws4.row_dimensions[r4].height = 20
    for col in range(1, 6):
        c = ws4.cell(row=r4, column=col)
        c.fill = hdr_fill(INPUT_BG); c.border = thin_border()
    r4 += 1
    apply_header(ws4, r4, [1,2,3,4,5], ["访谈时长","访谈人","记录人","访谈形式","被访者公司/部门"])
    ws4.row_dimensions[r4].height = 20
    for col in range(1, 6):
        c = ws4.cell(row=r4, column=col)
        c.fill = hdr_fill(INPUT_BG); c.border = thin_border()
    r4 += 1
    r4 += 1
    apply_section(ws4, r4, 1, "一、被访者背景", 5)
    r4 += 1
    apply_header(ws4, r4, [1,2,3,4,5], ["背景维度","具体内容","备注","",""], bg="E85053")
    r4 += 1
    bg_fields = [
        ("岗位与司龄", "", "在此岗位的工作年限，大约经历"),
        ("过往相关经历", "", "是否经历过与本次采购相关的历史事故/项目"),
        ("汇报线与决策影响力", "", "在本次采购决策中的角色：决策者/使用者/影响者/知情者"),
        ("个人利益关联", "", "本次采购成功与否对其个人的影响"),
    ]
    for i, (f, val, note) in enumerate(bg_fields):
        alt = (i % 2 == 0)
        apply_data_row(ws4, r4, 1, f, "normal", alt)
        c = ws4.cell(row=r4, column=2, value=val)
        c.font = Font(size=10, color="0000FF"); c.fill = hdr_fill(INPUT_BG)
        c.alignment = left(); c.border = thin_border()
        ws4.merge_cells(start_row=r4, start_column=3, end_row=r4, end_column=5)
        c = ws4.cell(row=r4, column=3, value=note)
        c.font = Font(size=9, color="595959")
        c.fill = hdr_fill("F2F2F2"); c.border = thin_border(); c.alignment = left()
        r4 += 1

    r4 += 1
    apply_section(ws4, r4, 1, "二、核心需求信号（访谈中观察到的关键信息）", 5)
    r4 += 1
    apply_header(ws4, r4, [1,2,3,4,5], ["序号","信号类型","原话/观察记录","时间点","是否被追问"], bg="E85053")
    ws4.row_dimensions[r4].height = 20
    r4 += 1
    for i in range(8):
        alt = (i % 2 == 0)
        apply_data_row(ws4, r4, 1, i+1, "normal", alt)
        for col in [2,3,4]:
            c = ws4.cell(row=r4, column=col)
            c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB"); c.border = thin_border(); c.alignment = left()
        apply_data_row(ws4, r4, 5, "是,否", "input", alt)
        r4 += 1

    r4 += 1
    apply_section(ws4, r4, 1, "三、隐性需求初步判断", 5)
    r4 += 1
    assess = [
        ("隐性需求初步判断", "", "访谈后对被访者真实需求的主观判断"),
        ("风险规避偏好", "", "对方是风险偏好型还是风险规避型？有哪些具体信号？"),
        ("关键决策人识别", "", "除被访者外，谁还可能影响最终决策？"),
    ]
    for i, (f, val, note) in enumerate(assess):
        alt = (i % 2 == 0)
        apply_data_row(ws4, r4, 1, f, "normal", alt)
        ws4.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=4)
        c = ws4.cell(row=r4, column=2, value=val)
        c.font = Font(size=10, color="0000FF"); c.fill = hdr_fill(INPUT_BG)
        c.alignment = left(); c.border = thin_border()
        apply_data_row(ws4, r4, 5, note, "normal", alt)
        r4 += 1
    ws4.freeze_panes = "A6"

    ws5 = wb.create_sheet("访谈复盘表")
    set_col_widths(ws5, [4, 22, 38, 20, 20])
    add_sheet_header(ws5, "【工具一】隐性需求挖掘访谈工具包 - 访谈复盘表")
    r5 = add_instructions(ws5, "使用说明：每次访谈结束后24小时内填写。复盘越及时，细节越完整。")

    apply_header(ws5, r5, [1,2,3,4,5], ["序号","复盘维度","具体内容","自我评分(1-5)","改进计划"])
    ws5.row_dimensions[r5].height = 20
    r5 += 1

    reviews = [
        (1,"访谈目标达成情况","本次访谈是否验证了核心假设？哪些假设已验证，哪些还需要进一步验证？"),
        (2,"提问质量","哪些问题效果最好？哪些问题让对方防御增加或话题转移？"),
        (3,"追问技巧运用","遇到笼统回答/停顿/岔开话题时，是否处理得当？"),
        (4,"隐性信号捕捉","发现了哪些当时没当回事、后来觉得重要的信号？"),
        (5,"决策链认知更新","对客户决策链的判断是否需要修正？"),
        (6,"下一步行动","下一次访谈需要重点验证什么？需要补充哪些背景调研？"),
        (7,"情绪与状态","访谈中自己的情绪状态如何？有哪些时刻情绪影响了访谈质量？"),
    ]

    for i, (num, dim, content) in enumerate(reviews):
        alt = (i % 2 == 0)
        apply_data_row(ws5, r5, 1, num, "normal", alt)
        apply_data_row(ws5, r5, 2, dim, "normal", alt)
        c = ws5.cell(row=r5, column=3, value=content)
        c.font = Font(size=10, color="595959"); c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.alignment = left(); c.border = thin_border(); ws5.row_dimensions[r5].height = 36
        c = ws5.cell(row=r5, column=4, value=3)
        c.font = Font(size=10, color="0000FF"); c.fill = hdr_fill(INPUT_BG)
        c.alignment = center(); c.border = thin_border()
        apply_data_row(ws5, r5, 5, "", "input", alt)
        r5 += 1

    dv5 = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws5.add_data_validation(dv5)
    dv5.sqref = f"D6:D{r5-1}"

    r5 += 1
    apply_section(ws5, r5, 1, "被忽略信号记录（本次访谈中出现了但当时没当回事的）", 5)
    r5 += 1
    apply_header(ws5, r5, [1,2,3,4,5], ["信号","出现时间点","当时为什么没重视","现在回看的判断","是否需要跟进"], bg="E85053")
    ws5.row_dimensions[r5].height = 20
    r5 += 1
    for i in range(5):
        alt = (i % 2 == 0)
        for col in range(1, 6):
            c = ws5.cell(row=r5, column=col)
            c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
            c.border = thin_border(); c.alignment = left()
        r5 += 1
    ws5.freeze_panes = "A5"

    wb.save(path)
    print(f"Created: {path}")

def create_file2(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("排错式验证五问")
    set_col_widths(ws, [4, 30, 52])
    add_sheet_header(ws, "【工具二】验证实验设计工具 - 排错式验证五问")
    r = add_instructions(ws, "使用说明：设计任何一次验证实验之前，先回答以下五个问题。排错式思维——先想最坏情况，再想如何排除。")

    apply_header(ws, r, [1,2,3], ["序号","验证问题","填写区"], bg="C00000")
    ws.row_dimensions[r].height = 22
    r += 1

    questions = [
        (1, "这次验证要排除的最坏可能是什么，而不是要证明的最好结果是什么？",
         "最坏可能是：（填写你在本次实验中，最担心被证明错误的那个假设）"),
        (2, "如果专门找最挑剔、最抵触、最不利于假设成立的场景或人选来测试，假设还站得住吗？",
         "最挑剔的场景/人选：\n\n具体描述："),
        (3, "这次验证的最小动作是什么？能不能用一句话、一封邮件、一次非正式沟通完成？",
         "最小动作：\n\n是否可行：是/否，理由："),
        (4, "如果验证结果推翻了原假设，团队有没有心理准备接受，会不会因为已经投入的资源而选择性忽视不利结果？",
         "团队心态评估：\n\n预案：若假设被推翻，下一步是："),
        (5, "这次验证的结果，谁来判断是否通过？判断标准提前写没写清楚？",
         "判断人：\n\n通过标准（提前写清楚）："),
    ]

    for i, (num, question, fill_area) in enumerate(questions):
        alt = (i % 2 == 0)
        apply_data_row(ws, r, 1, num, "normal", alt)
        c = ws.cell(row=r, column=2, value=question)
        c.font = Font(size=10, bold=True, color="C00000")
        c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.alignment = left(); c.border = thin_border()
        ws.row_dimensions[r].height = 50
        c = ws.cell(row=r, column=3, value=fill_area)
        c.font = Font(size=10, color="0000FF")
        c.fill = hdr_fill(INPUT_BG)
        c.alignment = left(); c.border = thin_border()
        r += 1
    ws.freeze_panes = "A5"

    ws2 = wb.create_sheet("最小暴露实验设计表")
    set_col_widths(ws2, [4, 24, 38, 18, 20])
    add_sheet_header(ws2, "【工具二】验证实验设计工具 - 最小暴露实验设计表")
    r2 = add_instructions(ws2, "使用说明：在正式方案提交前，挑出方案里最具争议、最可能触碰隐性需求的一条，单独抽出测试。")

    apply_header(ws2, r2, [1,2,3,4,5], ["序号","实验要素","填写内容","状态","下次验证时间"], bg="C00000")
    ws2.row_dimensions[r2].height = 20
    r2 += 1

    fields = [
        (1,"最具争议的那条方案要点","从完整方案中提炼出最具争议或最可能触碰隐性需求的一点",""),
        (2,"最小验证动作","用什么方式抛出去？（闲聊/邮件/电梯偶遇/正式汇报）",""),
        (3,"测试对象","关键决策人或使用者是谁？",""),
        (4,"为什么选这个人","此人为什么是最挑剔/最可能提出异议的？",""),
        (5,"观察什么","第一反应是什么？（表情/语气/回应内容/沉默时长）",""),
        (6,"判断标准","什么反应算通过？什么反应算不通过？",""),
        (7,"最可能出现的借口/反驳","提前预判对方最常用的推辞是什么",""),
        (8,"应对预案","如果对方反应负面，准备如何回应？",""),
        (9,"验证结果记录","实际观察到的第一反应（原始记录，不加工）",""),
        (10,"结论","假设是否站住：是/否/待继续验证",""),
        (11,"下一步","基于结论，下一步动作是什么？",""),
    ]

    for i, (num, elem, content, status) in enumerate(fields):
        alt = (i % 2 == 0)
        apply_data_row(ws2, r2, 1, num, "normal", alt)
        apply_data_row(ws2, r2, 2, elem, "normal", alt)
        c = ws2.cell(row=r2, column=3, value=content)
        is_result = i >= 8
        c.font = Font(size=10, color="0000FF" if not is_result else "000000")
        c.fill = hdr_fill(INPUT_BG if not is_result else ("FFF0F0" if alt else "FFFFFF"))
        c.alignment = left(); c.border = thin_border()
        ws2.row_dimensions[r2].height = 24
        apply_data_row(ws2, r2, 4, status or "待验证,已完成", "input", alt)
        apply_data_row(ws2, r2, 5, "", "input", alt)
        r2 += 1
    ws2.freeze_panes = "A5"

    ws3 = wb.create_sheet("验证结果记录表")
    set_col_widths(ws3, [18, 16, 16, 16, 30, 16])
    add_sheet_header(ws3, "【工具二】验证实验设计工具 - 验证结果记录表")
    r3 = add_instructions(ws3, "使用说明：每次验证实验完成后填写。原始观察记录不加工，结论单独填写。")

    apply_header(ws3, r3, [1,2,3,4,5,6],
        ["验证目标（要排除的最坏可能）","验证方式（最小动作是什么）","测试对象（是否选择最挑剔场景）",
         "观察到的第一反应（原始记录，不加工）","结论（假设是否站住）","下一步动作"], bg="C00000")
    ws3.row_dimensions[r3].height = 36
    r3 += 1

    for i in range(12):
        alt = (i % 2 == 0)
        for col in range(1, 7):
            c = ws3.cell(row=r3, column=col)
            if col in [1,2,3,4,6]:
                c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
            else:
                c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
            c.border = thin_border(); c.alignment = left()
        r3 += 1
    ws3.freeze_panes = "A5"

    ws4 = wb.create_sheet("验证结论汇总表")
    set_col_widths(ws4, [4, 22, 20, 20, 20, 20, 18])
    add_sheet_header(ws4, "【工具二】验证实验设计工具 - 验证结论汇总表")
    r4 = add_instructions(ws4, "使用说明：汇总所有验证实验的结论，用于向团队汇报和决策参考。")

    apply_header(ws4, r4, [1,2,3,4,5,6,7],
        ["序号","假设/待验证点","验证方式","验证结论","信心度","负责人","更新时间"], bg="C00000")
    ws4.row_dimensions[r4].height = 20
    r4 += 1

    for i in range(10):
        alt = (i % 2 == 0)
        apply_data_row(ws4, r4, 1, i+1, "normal", alt)
        for col in [2,3,4,5,6,7]:
            c = ws4.cell(row=r4, column=col)
            c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
            c.border = thin_border(); c.alignment = left()
        r4 += 1

    r4 += 1
    apply_section(ws4, r4, 1, "验证结论汇总统计", 7)
    r4 += 1

    summary_labels = [
        ("已验证（假设成立）", '=COUNTIF(D6:D15,"成立")'),
        ("已验证（假设不成立）", '=COUNTIF(D6:D15,"不成立")'),
        ("待继续验证", '=COUNTIF(D6:D15,"待继续验证")'),
        ("验证总数", '=COUNTA(D6:D15)'),
    ]
    for label, formula in summary_labels:
        apply_data_row(ws4, r4, 1, label, "normal", False)
        ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=3)
        c = ws4.cell(row=r4, column=4, value=formula)
        c.font = Font(size=11, bold=True, color="C00000")
        c.fill = hdr_fill("FFF0F0"); c.border = thin_border()
        c.alignment = center()
        r4 += 1

    ws4.freeze_panes = "A5"
    wb.save(path)
    print(f"Created: {path}")

def create_file3(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("三维评估表")
    set_col_widths(ws, [4, 24, 14, 14, 14, 20, 20])
    add_sheet_header(ws, "【工具三】需求优先级矩阵工具 - 三维评估表（业务价值/实现成本/责任敞口）")
    r = add_instructions(ws, "使用说明：对每个需求/功能点进行三维评分（1-5分制）。评分标准见下方备注栏。浅蓝色区域为输入区。")

    apply_header(ws, r, [1,2,3,4,5,6,7],
        ["序号","需求/功能描述","业务价值(1-5)","实现成本(1-5)","责任敞口(1-5)","综合评分(业务价值+责任敞口)","优先级建议"], bg="C00000")
    ws.row_dimensions[r].height = 30
    r += 1

    for i in range(15):
        alt = (i % 2 == 0)
        apply_data_row(ws, r, 1, i+1, "normal", alt)
        c = ws.cell(row=r, column=2)
        c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
        c.border = thin_border(); c.alignment = left()
        for col in [3,4,5]:
            c = ws.cell(row=r, column=col)
            c.font = Font(size=11, bold=True, color="0000FF")
            c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
            c.border = thin_border(); c.alignment = center()
        c = ws.cell(row=r, column=6, value=f"=C{r}+E{r}")
        c.font = Font(size=11, bold=True, color="C00000")
        c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.border = thin_border(); c.alignment = center()
        c = ws.cell(row=r, column=7,
            value=f'=IF(AND(C{r}>=4,E{r}>=4),"最高",IF(AND(E{r}>=4,C{r}<4),"次高",IF(AND(C{r}>=4,E{r}<4),"待观察","可延后")))')
        c.font = Font(size=10, color="0000FF")
        c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
        c.border = thin_border(); c.alignment = center()
        r += 1

    dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.sqref = f"C6:E{r-1}"

    r += 1
    apply_section(ws, r, 1, "评分标准说明", 7)
    r += 1
    notes = [
        ("业务价值", "该功能对客户业务指标的实际提升程度，1=几乎无提升，5=核心业务指标显著提升"),
        ("实现成本", "团队完成这个功能所需要的时间与资源投入，1=极低成本，5=需要大量人力和时间"),
        ("责任敞口", "如果这个功能不优先做，客户方相关负责人是否会因此面临具体的解释压力或问责风险。判断提问：如果这个功能三个月内做不出来，谁会被问到？他要怎么解释？"),
    ]
    for i, (dim, note) in enumerate(notes):
        alt = (i % 2 == 0)
        apply_data_row(ws, r, 1, dim, "normal", alt)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(row=r, column=2, value=note)
        c.font = Font(size=9, italic=True, color="595959")
        c.fill = hdr_fill("F2F2F2"); c.border = thin_border()
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 36
        r += 1
    ws.freeze_panes = "A5"

    ws2 = wb.create_sheet("排序建议逻辑说明")
    set_col_widths(ws2, [4, 28, 40, 22])
    add_sheet_header(ws2, "【工具三】需求优先级矩阵工具 - 排序建议逻辑说明")
    r2 = add_instructions(ws2, "使用说明：本说明解释三维评估后的排序逻辑，帮助团队理解优先级判断依据。")

    apply_header(ws2, r2, [1,2,3,4], ["象限","责任敞口","业务价值","排序建议"], bg="C00000")
    ws2.row_dimensions[r2].height = 20
    r2 += 1

    quadrants = [
        ("A象限（最高优先）", "高（>=4分）", "高（>=4分）", "优先级最高，既有说服力也有推动力。立即安排开发资源。"),
        ("B象限（次高优先）", "高（>=4分）", "一般（<4分）", "靠前安排，解决的是让具体的人不被问责的问题。，即便业务价值不突出，也是项目推进的政治基础。"),
        ("C象限（待观察）", "低（<4分）", "高（>=4分）", "容易被无限期拖延。需要主动找一个愿意推动的人，或者主动创造有人问责的场景。"),
        ("D象限（可延后）", "低（<4分）", "低（<4分）", "可以直接砍掉或长期搁置，不必纠结。"),
    ]
    colors = ["E2EFDA", "FDE9D9", "FFF0C0", "F2F2F2"]
    for i, (quad, risk, value, advice) in enumerate(quadrants):
        c = ws2.cell(row=r2, column=1, value=quad)
        c.font = Font(bold=True, color="C00000"); c.fill = hdr_fill(colors[i])
        c.border = thin_border(); c.alignment = center()
        for col, val in [(2, risk), (3, value)]:
            c = ws2.cell(row=r2, column=col, value=val)
            c.font = Font(bold=True); c.fill = hdr_fill(colors[i])
            c.border = thin_border(); c.alignment = center()
        c = ws2.cell(row=r2, column=4, value=advice)
        c.font = Font(size=10); c.fill = hdr_fill(colors[i])
        c.border = thin_border()
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.row_dimensions[r2].height = 40
        r2 += 1

    r2 += 1
    apply_section(ws2, r2, 1, "责任敞口判断关键提问", 4)
    r2 += 1
    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=4)
    c = ws2.cell(row=r2, column=1,
        value="核心提问：如果这个功能三个月内做不出来，谁会被问到？他要怎么解释？\n\n"
              "如果这个问题能被清晰回答 → 责任敞口高（>=4分）\n"
              "如果对方也说不清楚会不会有人追问 → 责任敞口低（<4分）")
    c.font = Font(size=11, bold=True, color="C00000")
    c.fill = hdr_fill("FFE0E0"); c.border = thin_border()
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws2.row_dimensions[r2].height = 50
    r2 += 2
    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=4)
    c = ws2.cell(row=r2, column=1,
        value="注意：责任敞口不等于业务价值。一个功能即使业务价值一般，只要负责它的具体的人会因为没做而被问责，"
              "它就是项目能推进下去的政治基础，应该优先安排。")
    c.font = Font(size=10, italic=True, color="595959")
    c.fill = hdr_fill("F2F2F2"); c.border = thin_border()
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws2.row_dimensions[r2].height = 36
    ws2.freeze_panes = "A5"

    ws3 = wb.create_sheet("优先级评分汇总表")
    set_col_widths(ws3, [4, 28, 12, 12, 12, 12, 14])
    add_sheet_header(ws3, "【工具三】需求优先级矩阵工具 - 优先级评分汇总表")
    r3 = add_instructions(ws3, "使用说明：从「三维评估表」复制需求列表，汇总所有需求的评分和优先级。自动计算排序。")

    apply_header(ws3, r3, [1,2,3,4,5,6,7],
        ["序号","需求描述","业务价值","实现成本","责任敞口","综合得分","优先级排序"], bg="C00000")
    ws3.row_dimensions[r3].height = 20
    r3 += 1

    for i in range(15):
        alt = (i % 2 == 0)
        apply_data_row(ws3, r3, 1, i+1, "normal", alt)
        for col in range(2, 7):
            c = ws3.cell(row=r3, column=col)
            c.fill = hdr_fill(INPUT_BG if col != 6 else ("FFF0F0" if alt else "FFFFFF"))
            c.border = thin_border(); c.alignment = left() if col != 6 else center()
        c = ws3.cell(row=r3, column=7, value=f'=IF(B{r3}="","",RANK(F{r3},$F$6:$F$20,0))')
        c.font = Font(size=10, bold=True, color="C00000")
        c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.border = thin_border(); c.alignment = center()
        r3 += 1
    ws3.freeze_panes = "A5"

    ws4 = wb.create_sheet("决策记录表")
    set_col_widths(ws4, [18, 18, 14, 14, 28, 18])
    add_sheet_header(ws4, "【工具三】需求优先级矩阵工具 - 决策记录表")
    r4 = add_instructions(ws4, "使用说明：记录优先级排序会议的决策结论。蓝色单元格为输入区。")

    apply_header(ws4, r4, [1,2,3,4,5,6], ["决策会议主题","决策日期","参与人","主持人","决策结论","决策依据"])
    ws4.row_dimensions[r4].height = 20
    for col in range(1, 7):
        c = ws4.cell(row=r4, column=col)
        c.fill = hdr_fill(INPUT_BG); c.border = thin_border()
    r4 += 1
    apply_header(ws4, r4, [1,2,3,4,5,6], ["需求范围确认", "", "", "", "", ""], bg="E85053")
    r4 += 1
    for i in range(10):
        alt = (i % 2 == 0)
        apply_data_row(ws4, r4, 1, f"需求项{i+1}", "normal", alt)
        for col in [2,3,4,5,6]:
            c = ws4.cell(row=r4, column=col)
            c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
            c.border = thin_border(); c.alignment = left()
        r4 += 1
    r4 += 1
    apply_section(ws4, r4, 1, "决策争议点记录", 6)
    r4 += 1
    apply_header(ws4, r4, [1,2,3,4,5,6], ["争议需求","分歧方A观点","分歧方B观点","最终决议","决议理由","负责人"], bg="C00000")
    ws4.row_dimensions[r4].height = 20
    r4 += 1
    for i in range(5):
        alt = (i % 2 == 0)
        for col in range(1, 7):
            c = ws4.cell(row=r4, column=col)
            c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
            c.border = thin_border(); c.alignment = left()
        r4 += 1
    ws4.freeze_panes = "A5"
    wb.save(path)
    print(f"Created: {path}")

def create_file4(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("案例记录表")
    set_col_widths(ws, [4, 16, 16, 16, 24, 24, 20, 20])
    add_sheet_header(ws, "【工具四】案例库与复盘工具 - 案例记录表")
    r = add_instructions(ws, "使用说明：每完成一个项目后填写。只记信号和判断逻辑，不记可识别出具体个人身份的敏感细节。")

    apply_header(ws, r, [1,2,3,4,5,6,7,8],
        ["序号","行业","项目规模","项目类型","表层需求（合同/需求文档内容）",
         "事后判断的隐性需求（复盘时才清楚的真实诉求）","被忽略的信号（原话/停顿/反常沉默）",
         "信号出现阶段"], bg="C00000")
    ws.row_dimensions[r].height = 36
    r += 1

    for i in range(20):
        alt = (i % 2 == 0)
        apply_data_row(ws, r, 1, i+1, "normal", alt)
        for col in range(2, 8):
            c = ws.cell(row=r, column=col)
            c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
            c.border = thin_border(); c.alignment = left()
        r += 1
    ws.freeze_panes = "A5"

    ws2 = wb.create_sheet("信号词典积累表")
    set_col_widths(ws2, [4, 28, 40, 22])
    add_sheet_header(ws2, "【工具四】案例库与复盘工具 - 信号词典积累表")
    r2 = add_instructions(ws2, "使用说明：团队根据真实项目案例，持续积累当时没当回事、后来证明关键的具体句子，形成属于团队自己的信号词典。")

    apply_header(ws2, r2, [1,2,3,4], ["序号","信号原话/典型表现","通常含义","建议的后续动作"], bg="C00000")
    ws2.row_dimensions[r2].height = 20
    r2 += 1

    initial_signals = [
        ("我们内部对这个还有点讨论", "通常意味着决策链中存在尚未被说服的关键人物", "建议主动询问：具体是哪个层面的讨论？技术层/执行层/管理层？"),
        ("这个我们再看看流程", "通常意味着预算审批或内部流程中存在未言明的障碍", "建议追问：流程涉及哪些部门？哪个环节最可能卡住？"),
        ("应该没什么大问题", "语气偏轻、缺乏具体细节支撑时，值得留意", "继续追问：方便展开说说吗？真正没问题的表态通常会带一两句具体理由"),
        ("我们以前做过类似的项目", "可能暗示客户对供应商有负面印象，需要谨慎处理", "不直接追问效果，而是通过场景式提问引导客户主动展开"),
        ("你们的方案挺好的，不过……", "典型的委婉拒绝信号，"不过"后面才是真实顾虑", "不要接话夸方案，直接问：您方便说说主要顾虑是什么吗？"),
        ("这个我们领导说了算", "对接人可能不是真正的决策者，或者在用领导做挡箭牌", "试探性询问：领导主要关注哪方面？有没有直接沟通的渠道？"),
        ("最近特别忙，等有空再说", "可能意味着优先级不高，或者优先级判断有分歧", "尝试了解：目前最紧急的事情是什么？这件事对您来说优先级如何？"),
        ("你们价格有点高", "可能只是压价借口，也可能是真实顾虑，需要进一步探测", "不直接讨论价格，而是问：除了价格，还有哪些因素在考虑？"),
    ]

    for i, (signal, meaning, action) in enumerate(initial_signals):
        alt = (i % 2 == 0)
        apply_data_row(ws2, r2, 1, i+1, "normal", alt)
        c = ws2.cell(row=r2, column=2, value=signal)
        c.font = Font(size=10, bold=True, color="C00000")
        c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.border = thin_border(); c.alignment = left()
        ws2.row_dimensions[r2].height = 30
        c = ws2.cell(row=r2, column=3, value=meaning)
        c.font = Font(size=10); c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.border = thin_border(); c.alignment = left()
        c = ws2.cell(row=r2, column=4, value=action)
        c.font = Font(size=10, color="0000FF")
        c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
        c.border = thin_border(); c.alignment = left()
        r2 += 1

    for i in range(10):
        alt = ((len(initial_signals) + i) % 2 == 0)
        apply_data_row(ws2, r2, 1, len(initial_signals)+i+1, "normal", alt)
        for col in [2,3,4]:
            c = ws2.cell(row=r2, column=col)
            c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
            c.border = thin_border(); c.alignment = left()
        r2 += 1
    ws2.freeze_panes = "A5"

    ws3 = wb.create_sheet("复盘分析表")
    set_col_widths(ws3, [4, 24, 38, 20])
    add_sheet_header(ws3, "【工具四】案例库与复盘工具 - 复盘分析表")
    r3 = add_instructions(ws3, "使用说明：每个项目结束后填写。复盘的价值在于找出当初差点被忽略的信号，即便项目成功也要找出被侥幸绕过的风险点。")

    apply_header(ws3, r3, [1,2,3,4], ["复盘维度","具体内容","评分(1-5)","改进行动"], bg="C00000")
    ws3.row_dimensions[r3].height = 20
    r3 += 1

    review_dims = [
        ("项目基本信息", "行业、规模、项目类型（不含可识别身份的具体信息）"),
        ("表层需求", "写进合同或需求文档的内容，是否与实际交付一致？"),
        ("隐性需求识别", "事后想清楚的真实诉求，当初哪些信号指向了这些隐性需求？"),
        ("被忽略的信号", "当初出现过但没被当回事的一句话、一个停顿、一次反常沉默，尽量原话记录"),
        ("信号出现阶段", "被忽略的信号出现在哪个阶段？初次访谈/方案确认/验收/续约前？"),
        ("访谈质量", "访谈中哪些问题效果好？哪些问题让对方防御增加？"),
        ("验证实验有效性", "设计的验证实验是否有效排除或确认了假设？"),
        ("需求优先级判断", "当时的优先级排序是否准确？有没有本应靠前但被延后的需求？"),
        ("决策链判断", "对客户决策链的判断是否准确？实际拍板人是谁？"),
        ("团队协作", "团队在哪些环节配合得好，哪些环节有改进空间？"),
    ]

    for i, (dim, content) in enumerate(review_dims):
        alt = (i % 2 == 0)
        apply_data_row(ws3, r3, 1, dim, "normal", alt)
        c = ws3.cell(row=r3, column=2, value=content)
        c.font = Font(size=10, color="595959"); c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.border = thin_border(); c.alignment = left()
        ws3.row_dimensions[r3].height = 30
        c = ws3.cell(row=r3, column=3)
        c.font = Font(size=11, bold=True, color="0000FF")
        c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
        c.border = thin_border(); c.alignment = center()
        apply_data_row(ws3, r3, 4, "", "input", alt)
        r3 += 1

    dv3 = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws3.add_data_validation(dv3)
    dv3.sqref = f"C6:C{r3-1}"

    r3 += 1
    apply_section(ws3, r3, 1, "成功因素与风险点", 4)
    r3 += 1
    apply_header(ws3, r3, [1,2,3,4], ["类型","描述","对项目的实际影响","经验教训"], bg="E85053")
    ws3.row_dimensions[r3].height = 20
    r3 += 1
    for i in range(6):
        alt = (i % 2 == 0)
        apply_data_row(ws3, r3, 1, "成功因素" if i < 3 else "风险点", "normal", alt)
        for col in [2,3,4]:
            c = ws3.cell(row=r3, column=col)
            c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
            c.border = thin_border(); c.alignment = left()
        r3 += 1
    ws3.freeze_panes = "A5"

    ws4 = wb.create_sheet("经验教训总结表")
    set_col_widths(ws4, [4, 24, 36, 20, 16])
    add_sheet_header(ws4, "【工具四】案例库与复盘工具 - 经验教训总结表")
    r4 = add_instructions(ws4, "使用说明：定期（季度/半年度）汇总复盘结果，形成团队层面的经验教训总结，用于新人培训和知识传承。")

    apply_header(ws4, r4, [1,2,3,4,5], ["序号","经验/教训类别","具体内容","适用场景","更新日期"], bg="C00000")
    ws4.row_dimensions[r4].height = 20
    r4 += 1

    categories = [
        ("隐性需求识别", "客户说"应该没什么大问题"时，往往意味着真正的问题被掩盖在笼统表态之下，需要追问具体场景"),
        ("隐性需求识别", "客户说"内部还有点讨论"时，意味着有未说服的关键人物，需要识别并主动接触"),
        ("风险规避判断", "客户主动提起"以前做过类似项目但效果不好"，是在委婉表达顾虑，不应直接追问细节"),
        ("决策链识别", "说"领导说了算"的对接人可能不是真正决策者，需要找到实际拍板人"),
        ("验证实验设计", "最小暴露实验比完整方案更能获得真实反馈，因为对方没有时间组织防御性回答"),
        ("访谈技巧", "遇到停顿或话题转移时，当场逼问会让对方彻底关闭话匣子，记下来下次再验证"),
        ("需求优先级", "责任敞口高的需求即使业务价值一般也应优先，因为它解决的是有人不被问责的政治基础"),
        ("团队协作", "复盘时必须找出"当初差点被忽略的信号"，成功项目往往掩盖了更多被侥幸绕过的风险"),
    ]

    for i, (cat, content) in enumerate(categories):
        alt = (i % 2 == 0)
        apply_data_row(ws4, r4, 1, i+1, "normal", alt)
        apply_data_row(ws4, r4, 2, cat, "normal", alt)
        c = ws4.cell(row=r4, column=3, value=content)
        c.font = Font(size=10); c.fill = hdr_fill("FFF0F0" if alt else "FFFFFF")
        c.border = thin_border(); c.alignment = left()
        ws4.row_dimensions[r4].height = 30
        c = ws4.cell(row=r4, column=4)
        c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
        c.border = thin_border(); c.alignment = left()
        apply_data_row(ws4, r4, 5, "", "input", alt)
        r4 += 1

    for i in range(10):
        alt = ((len(categories) + i) % 2 == 0)
        apply_data_row(ws4, r4, 1, len(categories)+i+1, "normal", alt)
        for col in range(2, 6):
            c = ws4.cell(row=r4, column=col)
            c.fill = hdr_fill(INPUT_BG if not alt else "E8F5FB")
            c.border = thin_border(); c.alignment = left()
        r4 += 1
    ws4.freeze_panes = "A5"
    wb.save(path)
    print(f"Created: {path}")

if __name__ == "__main__":
    create_file1(f"{OUTPUT_DIR}/01_隐性需求挖掘访谈工具包.xlsx")
    create_file2(f"{OUTPUT_DIR}/02_验证实验设计工具.xlsx")
    create_file3(f"{OUTPUT_DIR}/03_需求优先级矩阵工具.xlsx")
    create_file4(f"{OUTPUT_DIR}/04_案例库与复盘工具.xlsx")
    print("All 4 Excel files created!")

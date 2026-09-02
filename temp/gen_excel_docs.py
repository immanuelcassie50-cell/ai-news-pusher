# -*- coding: utf-8 -*-
"""
跨部门协作剧本杀 · 阶段 4B · 4 个 Excel 模板
- 01-评分表.xlsx
- 02-时间线.xlsx
- 03-风险码清单.xlsx
- 04-觉察点清单.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = r"D:\2026年课程\ai课2026整理\剧本杀\跨部门协作剧本杀\可视化输出\Office物料"
os.makedirs(OUT_DIR, exist_ok=True)

# =================== 通用样式 ===================
INK = "FF1A1A1A"
RED = "FF8B1A1A"
GOLD = "FFB8860B"
BLUE = "FF2E5BFF"
GRAY = "FF666666"
SOFT = "FFF4F0E8"
PAPER = "FFFFFFFF"
LIGHT_RED = "FFFFE8E8"
LIGHT_GOLD = "FFFFF5E0"
LIGHT_BLUE = "FFEEF1FA"
LIGHT_GREEN = "FFEEF8F1"

thin = Side(border_style="thin", color=INK)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# 字体（全部独立对象，禁用 StyleProxy 二次读取）
F_TITLE = Font(name="Microsoft YaHei", size=18, bold=True, color=INK)
F_HEAD_W = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFFFF")
F_H1_W = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFFFF")
F_H1_GOLD = Font(name="Microsoft YaHei", size=14, bold=True, color=GOLD)
F_H1_RED = Font(name="Microsoft YaHei", size=14, bold=True, color=RED)
F_BODY = Font(name="Microsoft YaHei", size=10, color=INK)
F_NUM = Font(name="Courier New", size=10, bold=True, color=RED)
F_NUM_GOLD = Font(name="Courier New", size=10, bold=True, color=GOLD)
F_RED = Font(name="Microsoft YaHei", size=10, bold=True, color=RED)
F_GOLD_BIG = Font(name="Microsoft YaHei", size=16, bold=True, color=GOLD)
F_NOTE = Font(name="Microsoft YaHei", size=9, color=GRAY, italic=True)
F_TINY = Font(name="Microsoft YaHei", size=9, bold=True, color=RED)

FILL_HEAD = PatternFill(start_color=INK, end_color=INK, fill_type="solid")
FILL_SOFT = PatternFill(start_color=SOFT, end_color=SOFT, fill_type="solid")
FILL_RED = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
FILL_GOLD = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
FILL_BLUE = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
FILL_GREEN = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

A_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
A_RGT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def set_cell(ws, row, col, value=None, font=F_BODY, fill=None, align=A_CTR):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font = font
    c.alignment = align
    c.border = BORDER
    if fill is not None:
        c.fill = fill
    return c


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_block(ws, title, subtitle="跨部门协作剧本杀 · v1.0 · 2026-06-07"):
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1)
    c.value = "跨部门协作剧本杀 · CROSS-DEPARTMENT SCRIPT-KILLER"
    c.font = F_TINY
    c.alignment = A_CTR
    ws.merge_cells("A2:H2")
    c = ws.cell(2, 1)
    c.value = title
    c.font = F_TITLE
    c.alignment = A_CTR
    ws.merge_cells("A3:H3")
    c = ws.cell(3, 1)
    c.value = subtitle
    c.font = F_NOTE
    c.alignment = A_CTR
    for r in [1, 2, 3]:
        ws.row_dimensions[r].height = 14 if r != 2 else 22


def write_h1(ws, row, title, fill=FILL_HEAD, font=F_H1_W, span=8):
    end_col = get_column_letter(span)
    ws.merge_cells(f"A{row}:{end_col}{row}")
    set_cell(ws, row, 1, title, font=font, fill=fill, align=A_LFT)


def write_th_row(ws, row, headers, span_start=1):
    for i, h in enumerate(headers):
        col = span_start + i
        set_cell(ws, row, col, h, font=F_HEAD_W, fill=FILL_HEAD, align=A_CTR)


# =================== 01 评分表 ===================
def excel_score_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "评分表"
    set_col_widths(ws, [3, 14, 12, 12, 12, 12, 12, 4])
    write_header_block(ws, "评分表 · 4 维度")

    row = 5
    write_h1(ws, row, "  01  风险码识别表 · 40%")
    row += 1
    write_th_row(ws, row, ["风险码", "09:00", "12:00", "16:00", "责任人", "风险描述", "必失分点"], 1)
    row += 1

    risks = [
        ("R-01", "张建国", "物料 80 万缺口", "80 万分摊失败"),
        ("R-02", "陈伟", "AI 95% 不达标", "91% 35 天 + 误差 ≤3"),
        ("R-03", "赵丽", "客户流失 30%", "4-25 复评基准"),
        ("R-04", "王敏", "5-31 飞行检查", "5 项必过清单"),
        ("R-05", "刘芳", "客户合同 27 天", "5-10 培训达标"),
        ("R-06", "王敏", "监管罚款 500 万", "路线 2 必失分"),
        ("R-07", "PM", "团队协作断裂", "5 VP 协同失败"),
        ("R-08", "PM", "项目延期 4-30", "物料 27 天"),
        ("R-09 ★", "CEO", "★ 核心爆炸点 ★", "拍脑袋 27 天"),
        ("R-10", "张建国", "资源争夺", "80 万分摊失败"),
        ("R-11", "刘芳", "5-10 培训达标", "减范围 30%"),
        ("R-12", "PM", "接班人问题", "9 月副手"),
        ("R-13", "CEO", "战略转型", "B 轮对赌"),
    ]
    for code, owner, desc, must in risks:
        is_star = "★" in code
        fill = FILL_GOLD if is_star else (FILL_SOFT if int(code[2:]) % 2 == 1 else None)
        set_cell(ws, row, 1, code, font=F_NUM if is_star else F_BODY, fill=fill)
        set_cell(ws, row, 2, "□", font=F_BODY, fill=fill)
        set_cell(ws, row, 3, "□", font=F_BODY, fill=fill)
        set_cell(ws, row, 4, "□", font=F_BODY, fill=fill)
        set_cell(ws, row, 5, owner, font=F_BODY, fill=fill)
        set_cell(ws, row, 6, desc, font=F_BODY, fill=fill, align=A_LFT)
        set_cell(ws, row, 7, must, font=F_BODY, fill=fill, align=A_LFT)
        row += 1

    row += 1
    write_h1(ws, row, "  02  4 维度 × 7 角色 评分")
    row += 1
    write_th_row(ws, row, ["角色", "风险码 40%", "协同 25%", "决断 20%", "兑现 15%", "加权总分", ""], 1)
    row += 1
    vps = ["陈伟 CTO", "王敏 CLO", "张建国 CFO", "赵丽 CMO", "刘芳 CHO", "林晓晨 PM", "方远航 CEO"]
    for vp in vps:
        set_cell(ws, row, 1, vp, font=F_NUM)
        for c in range(2, 6):
            set_cell(ws, row, c, "__/100", font=F_BODY)
        set_cell(ws, row, 6, "__分", font=F_RED)
        set_cell(ws, row, 7, "", font=F_BODY)
        row += 1

    row += 1
    write_h1(ws, row, "  03  加分项 · 最多 +20", fill=FILL_GOLD, font=F_H1_GOLD)
    row += 1
    bonuses = [
        ("+10", "R-09 核心爆炸点单独识别 + 处理"),
        ("+5", "5 VP 三轮表态一致（09:00 / 12:00 / 16:00）"),
        ("+5", "5/14 兑现中,出现至少 1 个具体数字"),
    ]
    for score, cond in bonuses:
        set_cell(ws, row, 1, score, font=F_NUM_GOLD)
        ws.merge_cells(f"B{row}:G{row}")
        set_cell(ws, row, 2, cond, font=F_BODY, align=A_LFT)
        for cc in range(3, 8):
            c2 = ws.cell(row, cc)
            c2.border = BORDER
        row += 1
    set_cell(ws, row, 1, "+__", font=Font(name="Courier New", size=12, bold=True, color=GOLD))
    ws.merge_cells(f"B{row}:G{row}")
    set_cell(ws, row, 2, "加分小计", font=F_RED, align=A_RGT)
    for cc in range(3, 8):
        c2 = ws.cell(row, cc)
        c2.border = BORDER
    row += 2

    write_h1(ws, row, "  04  失分项 · 最多 -30", fill=FILL_RED, font=F_H1_RED)
    row += 1
    deductions = [
        ("-10", "CEO 拍脑袋决断,不听 5 VP"),
        ("-10", "PM 12:30 硬窗口不决断"),
        ("-10", "5 VP 互相推诿,把责任推给别的部门"),
    ]
    for score, cond in deductions:
        set_cell(ws, row, 1, score, font=F_NUM)
        ws.merge_cells(f"B{row}:G{row}")
        set_cell(ws, row, 2, cond, font=F_BODY, align=A_LFT)
        for cc in range(3, 8):
            c2 = ws.cell(row, cc)
            c2.border = BORDER
        row += 1
    set_cell(ws, row, 1, "-__", font=Font(name="Courier New", size=12, bold=True, color=RED))
    ws.merge_cells(f"B{row}:G{row}")
    set_cell(ws, row, 2, "失分小计", font=F_RED, align=A_RGT)
    for cc in range(3, 8):
        c2 = ws.cell(row, cc)
        c2.border = BORDER
    row += 2

    write_h1(ws, row, "  05  最终总分")
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    set_cell(ws, row, 1, "最终总分 = 4 维度加权 + 加分项 - 失分项 = __分",
             font=F_GOLD_BIG, fill=FILL_GOLD, align=A_CTR)
    ws.row_dimensions[row].height = 30
    row += 2

    write_h1(ws, row, "  06  培训师总评")
    row += 1
    for label in ["最大亮点", "最大风险", "改进建议", "5-31 预测"]:
        set_cell(ws, row, 1, label, font=F_BODY, fill=FILL_SOFT, align=A_LFT)
        ws.merge_cells(f"B{row}:H{row}")
        for cc in range(2, 9):
            c2 = ws.cell(row, cc)
            c2.border = BORDER
        ws.row_dimensions[row].height = 40
        row += 1

    out = os.path.join(OUT_DIR, "01-评分表.xlsx")
    wb.save(out)
    print(f"[OK] {out}")


# =================== 02 时间线 ===================
def excel_timeline():
    wb = Workbook()
    ws = wb.active
    ws.title = "时间线"
    set_col_widths(ws, [4, 16, 12, 24, 14, 24])
    write_header_block(ws, "时间线 · 49 天 · 4-13 → 5-31")

    row = 5
    write_h1(ws, row, "  01  7 个不可撤回节点", span=6)
    row += 1
    write_th_row(ws, row, ["节点", "日期", "类型", "事件", "责任人", "必失分点"], 1)
    row += 1

    nodes = [
        ("T0", "4-13 周一", "启动会", "5 VP 陈述 13 风险码", "7 角色全员", "装不知道 = 12:00 暴露"),
        ("N1", "4-14 周二", "私下对账", "PM 5 分钟 × 5 VP 收风险", "PM", "未对账 = 决断信息不全"),
        ("N2", "4-25 周六", "复评窗口", "减范围 30% 基准", "5 VP", "不减范围 = 5-10 必失分"),
        ("N3 ★", "4-30 周四", "B 轮对赌", "5 VP 物料必达 27 天", "CEO + 5 VP", "不可达 = 失分"),
        ("N4", "5-10 周日", "客户合同", "HR 培训必过", "刘芳 CHO", "未减范围 = 失分"),
        ("N5", "5-14 周四", "兑现节点", "5 VP 各自本部门承诺", "5 VP", "未兑现 = 5-31 失分"),
        ("N6 ★", "5-31 周六", "飞行检查", "上海网信办 5 项必过", "王敏 CLO", "失分 = 真实水平验证"),
    ]
    for n, d, t, e, o, m in nodes:
        is_star = "★" in n
        fill = FILL_RED if is_star else (FILL_GOLD if n == "T0" else None)
        set_cell(ws, row, 1, n, font=F_NUM, fill=fill)
        set_cell(ws, row, 2, d, font=F_BODY, fill=fill)
        set_cell(ws, row, 3, t, font=F_BODY, fill=fill)
        set_cell(ws, row, 4, e, font=F_BODY, fill=fill, align=A_LFT)
        set_cell(ws, row, 5, o, font=F_BODY, fill=fill)
        set_cell(ws, row, 6, m, font=F_BODY, fill=fill, align=A_LFT)
        row += 1

    row += 1
    write_h1(ws, row, "  02  三轮剧本 · 4 小时主流程", span=6)
    row += 1
    write_th_row(ws, row, ["轮次", "时段", "时间", "动作", "硬窗口", ""], 1)
    row += 1

    rounds = [
        ("第 1 轮", "启动会", "09:00-09:30", "5 VP 公开陈述 5 分钟 × 5", "10:00 结束"),
        ("第 1 轮", "全员质询", "09:30-10:00", "5 VP 互相质询 30 分钟", "10:00 结束"),
        ("第 2 轮", "私下对账", "11:00-12:00", "PM 5 VP × 5 分钟", "12:00 决断"),
        ("第 2 轮", "升级邮件", "12:00-12:30", "PM 4 段式风险汇总", "12:30 决断 ★"),
        ("第 3 轮", "CEO 决断", "16:00-16:45", "5 VP 表态 + CEO 三选一", "16:45 结束"),
        ("第 3 轮", "兑现承诺", "17:00-17:30", "5 VP 5/14 兑现承诺", "17:30 结束"),
    ]
    for r_name, p, t, a, w in rounds:
        set_cell(ws, row, 1, r_name, font=F_NUM)
        set_cell(ws, row, 2, p, font=F_BODY, align=A_LFT)
        set_cell(ws, row, 3, t, font=F_BODY)
        set_cell(ws, row, 4, a, font=F_BODY, align=A_LFT)
        set_cell(ws, row, 5, w, font=F_RED if "★" in w else F_BODY)
        set_cell(ws, row, 6, "", font=F_BODY)
        row += 1

    row += 1
    write_h1(ws, row, "  03  复盘 5 阶段 · 60 分钟", span=6)
    row += 1
    write_th_row(ws, row, ["时段", "阶段", "动作", "学员动作", "", ""], 1)
    row += 1
    debriefs = [
        ("0-10 min", "回顾", "重放 16:00 决断片段", "5 VP 自评"),
        ("10-25 min", "诊断", "一致性矩阵展示", "VP 之间互评"),
        ("25-40 min", "觉察", "引导我以为 vs 原来", "个人觉察发言"),
        ("40-50 min", "承诺", "引导 5/14 我做什么", "5 VP 各写 1 句"),
        ("50-60 min", "落地", "引导 5-31 真实水平验证", "5 VP 各写 1 句"),
    ]
    for t, p, a, s in debriefs:
        set_cell(ws, row, 1, t, font=F_NUM)
        set_cell(ws, row, 2, p, font=F_BODY)
        set_cell(ws, row, 3, a, font=F_BODY, align=A_LFT)
        set_cell(ws, row, 4, s, font=F_BODY, align=A_LFT)
        set_cell(ws, row, 5, "", font=F_BODY)
        set_cell(ws, row, 6, "", font=F_BODY)
        row += 1

    row += 1
    write_h1(ws, row, "  04  倒计时 · 5-31 距今", span=6)
    row += 1
    write_th_row(ws, row, ["距 5-31", "状态", "必做", "", "", ""], 1)
    row += 1
    counts = [
        ("49 天", "启动", "5 VP 公开陈述 13 风险码", None),
        ("48 天", "私下对账", "PM 5 VP × 5 分钟", None),
        ("37 天", "复评窗口", "减范围 30% 基准", None),
        ("32 天 ★", "B 轮对赌", "物料必达 27 天", "red"),
        ("22 天", "客户合同", "HR 培训必过", None),
        ("18 天", "兑现节点", "5 VP 各自本部门承诺", None),
        ("1 天 ★", "飞行检查", "上海网信办 5 项必过", "red"),
        ("0 天", "5-31 验证", "真实水平 = 必过", "gold"),
    ]
    for d, s, m, color in counts:
        fill = FILL_RED if color == "red" else (FILL_GOLD if color == "gold" else None)
        set_cell(ws, row, 1, d, font=F_NUM, fill=fill)
        set_cell(ws, row, 2, s, font=F_BODY, fill=fill)
        ws.merge_cells(f"C{row}:F{row}")
        set_cell(ws, row, 3, m, font=F_BODY, fill=fill, align=A_LFT)
        for cc in range(4, 7):
            c2 = ws.cell(row, cc)
            c2.border = BORDER
            if fill:
                c2.fill = fill
        row += 1

    out = os.path.join(OUT_DIR, "02-时间线.xlsx")
    wb.save(out)
    print(f"[OK] {out}")


# =================== 03 风险码清单 ===================
def excel_risk_list():
    wb = Workbook()
    ws = wb.active
    ws.title = "风险码清单"
    set_col_widths(ws, [4, 12, 10, 22, 14, 22, 14, 18])
    write_header_block(ws, "13 风险码全清单")

    row = 5
    write_h1(ws, row, "  13 风险码 · 7 角色责任")
    row += 1
    write_th_row(ws, row, ["风险码", "颜色", "责任 VP", "风险描述", "触发条件", "应对剧本", "必失分点", "5-31 验证"], 1)
    row += 1

    risks = [
        ("R-01", "金 B8860B", "张建国 CFO", "物料 80 万缺口", "B 轮对赌物料 27 天必达", "5 VP 分摊 16 万 × 5 = 80 万", "80 万分摊失败", "4-30 物料 27 天必达"),
        ("R-02", "绿 1A8E5F", "陈伟 CTO", "AI 95% 不达标", "客户合同 27 天 AI 必过", "AI 91% 35 天训练 + 14 天复评", "91% + 误差 ≤3", "5-25 必达 35 天"),
        ("R-03", "玫红 C71585", "赵丽 CMO", "客户流失 30%", "4-25 复评基准", "减范围 30% 基准", "4-25 复评不达", "5-10 客户合同 27 天"),
        ("R-04", "棕 8B4513", "王敏 CLO", "5-31 飞行检查", "5-31 上海网信办 5 项必过", "5 项必过清单", "5 项不达", "5-31 必过 5 项"),
        ("R-05", "灰 708090", "刘芳 CHO", "客户合同 27 天", "5-10 客户合同培训", "5-10 培训达标 100% 覆盖", "5-10 培训不达", "5-10 培训达标"),
        ("R-06", "司法红 8B1A1A", "王敏 CLO", "监管罚款 500 万", "路线 2 推 14 天", "强烈反对路线 2", "路线 2 必失分", "路线 1 共识"),
        ("R-07", "紫 5C2C8C", "林晓晨 PM", "团队协作断裂", "5 VP 三轮表态冲突", "升级邮件 4 段式", "5 VP 协同失败", "5 VP 三轮一致"),
        ("R-08", "深灰 4A4A4A", "林晓晨 PM", "项目延期 4-30", "物料 27 天不可达", "路线 1 减范围 30%", "物料 27 天失分", "4-30 物料必达"),
        ("R-09 ★", "黑+金", "方远航 CEO", "★ 核心爆炸点 ★", "CEO 拍脑袋 27 天物料", "16:00 听 PM + 5 VP", "CEO 拍脑袋 = 失分", "CEO 选路线 1 = 95"),
        ("R-10", "赭石 A0522D", "张建国 CFO", "资源争夺", "5 VP 部门预算缩", "80 万分摊 = 5 VP 共识", "分摊失败", "5 VP 各 16 万"),
        ("R-11", "钢蓝 4682B4", "刘芳 CHO", "5-10 培训达标", "5-10 客户合同培训", "减范围 30% 培训", "5-10 培训不达", "5-10 培训达标"),
        ("R-12", "深石板 2F4F4F", "林晓晨 PM", "接班人问题", "9 月 PM 副手未定", "5 VP 部门副手", "9 月副手未指定", "5 VP 副手 ✓"),
        ("R-13", "午夜蓝 191970", "方远航 CEO", "战略转型", "B 轮对赌", "B 轮对赌 = 27 天物料", "B 轮对赌失分", "4-30 物料 27 天"),
    ]
    for code, color, vp, desc, trig, act, must, verify in risks:
        is_star = "★" in code
        fill = FILL_GOLD if is_star else None
        set_cell(ws, row, 1, code, font=F_NUM, fill=fill)
        set_cell(ws, row, 2, color, font=F_BODY, fill=fill, align=A_LFT)
        set_cell(ws, row, 3, vp, font=F_BODY, fill=fill)
        set_cell(ws, row, 4, desc, font=F_BODY, fill=fill, align=A_LFT)
        set_cell(ws, row, 5, trig, font=F_BODY, fill=fill, align=A_LFT)
        set_cell(ws, row, 6, act, font=F_BODY, fill=fill, align=A_LFT)
        set_cell(ws, row, 7, must, font=F_RED, fill=fill, align=A_LFT)
        set_cell(ws, row, 8, verify, font=F_BODY, fill=fill, align=A_LFT)
        row += 1

    row += 1
    write_h1(ws, row, "  13 风险码 · 颜色图例")
    row += 1
    legend = [
        ("金 B8860B", "财务", "张建国 CFO"),
        ("绿 1A8E5F", "技术", "陈伟 CTO"),
        ("玫红 C71585", "市场", "赵丽 CMO"),
        ("棕 8B4513", "法务", "王敏 CLO"),
        ("灰 708090", "人事", "刘芳 CHO"),
        ("蓝 2E5BFF", "项目", "林晓晨 PM"),
        ("黑 1A1A1A", "战略", "方远航 CEO"),
    ]
    for color, dept, owner in legend:
        set_cell(ws, row, 1, color, font=F_BODY, align=A_LFT)
        set_cell(ws, row, 2, dept, font=F_BODY)
        ws.merge_cells(f"C{row}:H{row}")
        set_cell(ws, row, 3, owner, font=F_BODY, align=A_LFT)
        for cc in range(4, 9):
            c2 = ws.cell(row, cc)
            c2.border = BORDER
        row += 1

    out = os.path.join(OUT_DIR, "03-风险码清单.xlsx")
    wb.save(out)
    print(f"[OK] {out}")


# =================== 04 觉察点清单 ===================
def excel_awareness():
    wb = Workbook()
    ws = wb.active
    ws.title = "觉察点清单"
    set_col_widths(ws, [4, 10, 16, 32, 14, 16, 18, 4])
    write_header_block(ws, "25 问 · 7 角色 · 觉察点清单")

    row = 5
    write_h1(ws, row, "  5 阶段 × 5 问 = 25 · 必失分点对应")
    row += 1
    write_th_row(ws, row, ["编号", "阶段", "问题", "失分点", "5 VP 互评", "培训师评", "觉察深度"], 1)
    row += 1

    questions = [
        ("01", "启动会", "5 VP 启动会都说了什么？有没有人装不知道？", "装不知道 = 12:00 暴露"),
        ("02", "启动会", "13 个风险码,启动会提到了几个？少 1 个 = 5 分扣分。", "风险码漏识 = -5 分/个"),
        ("03", "启动会", "CEO 拍脑袋 27 天 = R-09 触发,谁察觉了？", "未察觉 R-09 = 失分"),
        ("04", "启动会", "5 VP 私利 = 路线 1/2/3,你支持哪条？", "路线选择 = 协同前提"),
        ("05", "启动会", "如果你是 PM,听完 5 VP 陈述,你的第一反应是什么？", "PM 不汇总 = 决断失败"),
        ("06", "升级邮件", "PM 升级邮件识别了 13 个风险码吗？少几个？", "少 1 = -5 分"),
        ("07", "升级邮件", "PM 私下对账（11:00-12:00）谈了什么？", "未对账 = 信息不全"),
        ("08", "升级邮件", "PM 建议路线 1（推迟 14 天）,5 VP 都同意吗？", "5 VP 不同意 = 协同失败"),
        ("09", "升级邮件", "12:30 硬窗口前,谁没表态？谁被打断了？", "不表态 = 失分"),
        ("10", "升级邮件", "如果你是 PM,你会怎么汇总 5 VP 冲突的私利？", "PM 单方 = 协同失败"),
        ("11", "CEO 决断", "CEO 听 PM 升级邮件了吗？还是拍脑袋决断？", "拍脑袋 = -10 分"),
        ("12", "CEO 决断", "5 VP 最终表态 = 路线 1/2/3,谁变了？为什么？", "变脸 = 一致性失分"),
        ("13", "CEO 决断", "CEO 三选一,选了什么？为什么？", "听 = 95；不听 = 失分"),
        ("14", "CEO 决断", "5/14 兑现承诺,5 VP 都给了什么？", "空话 = 不兑现"),
        ("15", "CEO 决断", "如果你是 CEO,听完 5 VP 表态,你会怎么决断？", "CEO 拍脑袋 = R-09"),
        ("16", "物料承诺", "5 VP 各自本部门承诺 = 路线 1 落地,谁最具体？", "具体数字 = +5"),
        ("17", "物料承诺", "80 万缺口分摊了吗？怎么分？", "分摊失败 = R-10"),
        ("18", "物料承诺", "5-10 减范围 30%,HR 给了什么具体动作？", "空话 = 5-10 失分"),
        ("19", "物料承诺", "9 月 PM 接班,5 VP 部门副手是谁？", "未答 = 9 月失分"),
        ("20", "物料承诺", "5/14 兑现 = 5-25 必达,你相信吗？", "不信 = 协同失败"),
        ("21", "5-31 验证", "5-31 飞行检查,5 项过了几个？", "5 项失分 = 真实水平"),
        ("22", "5-31 验证", "4-30 失分项,5-31 还能补救吗？", "4-30 失分 = 5-31 必失分"),
        ("23", "5-31 验证", "5 VP 协同 = 5 项必过,谁贡献最大？", "贡献最大 = +5"),
        ("24", "5-31 验证", "如果让你重演一遍,你会怎么改？", "重演 = 觉察深度"),
        ("25", "5-31 验证", "这场剧本杀,你最大的觉察是什么？", "觉察 = 复盘价值"),
    ]
    for num, stage, q, must in questions:
        if stage == "启动会":
            fill = FILL_RED
        elif stage == "升级邮件":
            fill = FILL_BLUE
        elif stage == "CEO 决断":
            fill = FILL_GOLD
        elif stage == "物料承诺":
            fill = FILL_GREEN
        else:
            fill = FILL_RED
        set_cell(ws, row, 1, num, font=F_NUM, fill=fill)
        set_cell(ws, row, 2, stage, font=F_BODY, fill=fill)
        set_cell(ws, row, 3, q, font=F_BODY, fill=fill, align=A_LFT)
        set_cell(ws, row, 4, must, font=F_RED, fill=fill, align=A_LFT)
        for c in [5, 6, 7]:
            set_cell(ws, row, c, "", font=F_BODY, fill=fill, align=A_LFT)
        set_cell(ws, row, 8, "", font=F_BODY, fill=fill)
        ws.row_dimensions[row].height = 36
        row += 1

    row += 1
    write_h1(ws, row, "  觉察记录 · 我以为 vs 原来")
    row += 1
    vps = ["陈伟 CTO", "王敏 CLO", "张建国 CFO", "赵丽 CMO", "刘芳 CHO", "林晓晨 PM", "方远航 CEO"]
    for vp in vps:
        set_cell(ws, row, 1, vp, font=F_NUM, fill=FILL_SOFT)
        ws.merge_cells(f"B{row}:C{row}")
        set_cell(ws, row, 2, "我以为：", font=F_BODY, fill=FILL_SOFT, align=A_LFT)
        c2 = ws.cell(row, 3)
        c2.border = BORDER
        c2.fill = FILL_SOFT
        ws.merge_cells(f"D{row}:H{row}")
        for cc in range(4, 9):
            ccc = ws.cell(row, cc)
            ccc.border = BORDER
            ccc.fill = FILL_SOFT
        ws.row_dimensions[row].height = 50
        row += 1
        set_cell(ws, row, 1, vp, font=F_NUM, fill=FILL_SOFT)
        ws.merge_cells(f"B{row}:C{row}")
        set_cell(ws, row, 2, "原来：", font=F_BODY, fill=FILL_SOFT, align=A_LFT)
        c2 = ws.cell(row, 3)
        c2.border = BORDER
        c2.fill = FILL_SOFT
        ws.merge_cells(f"D{row}:H{row}")
        for cc in range(4, 9):
            ccc = ws.cell(row, cc)
            ccc.border = BORDER
            ccc.fill = FILL_SOFT
        ws.row_dimensions[row].height = 50
        row += 1

    out = os.path.join(OUT_DIR, "04-觉察点清单.xlsx")
    wb.save(out)
    print(f"[OK] {out}")


if __name__ == "__main__":
    excel_score_sheet()
    excel_timeline()
    excel_risk_list()
    excel_awareness()
    print("\n=== 4 个 Excel 文档生成完毕 ===")

# -*- coding: utf-8 -*-
"""Inspect reference xlsx files."""
import openpyxl
import sys
import json

files = [
    r"D:\2026年课程\新课开发demo\配套表单和指引-Excel版\表单使用指引.xlsx",
    r"D:\2026年课程\新课开发demo\配套表单和指引-Excel版\配套表单_空表.xlsx",
    r"D:\2026年课程\新课开发demo\配套表单和指引-Excel版\配套表单_填好版.xlsx",
]

for f in files:
    print("=" * 80)
    print("FILE:", f)
    print("=" * 80)
    wb = openpyxl.load_workbook(f, data_only=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[Sheet: {sheet_name}] dim={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=False):
            for cell in row:
                if cell.value is not None:
                    val = str(cell.value)
                    if len(val) > 100:
                        val = val[:100] + "..."
                    print(f"  {cell.coordinate}: {val}")
    print()

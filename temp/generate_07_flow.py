import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "互动流程表"

# Define colors
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
chapter_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
time_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

header_font = Font(name='Microsoft YaHei', bold=True, color="FFFFFF", size=10)
subheader_font = Font(name='Microsoft YaHei', bold=True, size=10)
normal_font = Font(name='Microsoft YaHei', size=9)
title_font = Font(name='Microsoft YaHei', bold=True, size=12)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:G1')
ws['A1'] = '07 · 互动流程表 ·《普通人的积极心理学实操课》'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 25

# Headers
headers = ['章节', '时间段', '时长', '环节类型', '讲师动作', '学员动作', '备注']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Data for each chapter
data = [
    # Chapter 1
    ['第一章', '0:00-0:10', '10min', '开场/引入', '认知自测：发放"幸福假设判断表"，说明"先写第一反应"', '凭直觉判断8句话的真假', '不点评，只收集'],
    ['第一章', '0:10-0:15', '5min', '讲解', '讲述1978年布里克曼彩票研究，引出"享乐适应"', '听讲+思考', '停顿问："这个结果有没有出乎你的意料？"'],
    ['第一章', '0:15-0:25', '10min', '讲解', '破除三个误解：强迫正能量/心灵鸡汤/基因宿命论', '听讲+思考', '用"不是...而是..."的句式'],
    ['第一章', '0:25-0:40', '15min', '概念讲解', 'PERMA五维度讲解', '理解五维度含义', '每个维度用生活例子说明'],
    ['第一章', '0:40-0:55', '15min', '练习', '发放PERMA自评表，引导学员打分', '为自己最近一个月打分，画雷达图', '不要求完美，如实即可'],
    ['第一章', '0:55-1:05', '10min', '练习', '上周PERMA生活扫描练习', '为每维度找一件上周的小事', '找不到的如实写"本周为空"'],
    ['第一章', '1:05-1:15', '10min', '反思', '回到开场认知自测，让学员重新判断', '比较两列差异', '让学员说出改变最大的一个点'],
    ['第一章', '1:15-1:20', '5min', '总结', '总结本章3个关键收获', '记下"今天带走的一句话"', '预告第二章内容'],

    # Chapter 2
    ['第二章', '0:00-0:03', '3min', '开场', '回顾第一章核心：幸福是多维度的，可以主动建设', '回忆', '用2句话快速回顾'],
    ['第二章', '0:03-0:10', '7min', '开场/练习', '10秒"不要想白熊"实验，让学员写"想到了什么"', '执行实验，写下反应', '引出"压制情绪是无效策略"'],
    ['第二章', '0:10-0:25', '15min', '概念讲解', '讲解8种情绪的功能（恐惧/愤怒/悲伤/焦虑/孤独/内疚/羞耻）', '听讲+思考', '用表格呈现，情绪后面给一句话功能'],
    ['第二章', '0:25-0:40', '15min', '练习', '三个场景的情绪功能解码练习', '为小慧/老赵/晓萱的场景解码', '两人一组讨论，然后分享'],
    ['第二章', '0:40-0:55', '15min', '概念讲解', '情绪颗粒度讲解+情绪轮盘介绍', '听讲+获得情绪词汇参考工具', '130+词汇分8大类'],
    ['第二章', '0:55-1:10', '15min', '练习', '精准描述练习：用情绪轮盘描述今天的三种心情', '原始描述→精准词汇→情境', '重点是"比第一反应更精准"'],
    ['第二章', '1:10-1:25', '15min', '概念+练习', '"我注意到我..."句式讲解+身体定位练习', '改写4个原始表述+身体定位', '两者都是创造"观察空间"的工具'],
    ['第二章', '1:25-1:40', '15min', '综合练习', '情绪日记初稿练习', '完整走一遍：事件→原始描述→精准命名→信号→"我注意到"→身体位置', '这是本章最重要的产出'],
    ['第二章', '1:40-1:50', '10min', '总结', '总结本章3个关键收获', '记下"今天带走的一个工具"', '预告第三章内容'],

    # Chapter 3
    ['第三章', '0:00-0:05', '5min', '开场', '回顾第二章核心：情绪是信号，如何精准命名', '回忆', '快速2句话'],
    ['第三章', '0:05-0:15', '10min', '开场/练习', '盖洛普研究数据+注意力分配练习', '快速写3个弱点vs3个优势', '引出"67%的人从未被问过优势是什么"'],
    ['第三章', '0:15-0:35', '20min', '概念讲解', '品格优势vs技能vs性格标签+三个识别信号', '理解三个识别信号', '用"用了之后充电还是放电"判断'],
    ['第三章', '0:35-0:50', '15min', '概念讲解', 'VIA24种品格优势体系介绍', '听讲+了解分类框架', '分6大美德：智慧/勇气/人道/公正/节制/超验'],
    ['第三章', '0:50-1:10', '20min', '练习', 'VIA优势快速自评', '用三个信号判断每种优势的符合度，选出前5', '可联系viacharacter.org做完整评估'],
    ['第三章', '1:10-1:25', '15min', '练习', '优势使用证据采集练习', '为前5优势各找最近一周的例子', '"具体"很重要，不是"我比较善良"'],
    ['第三章', '1:25-1:40', '15min', '练习', '优势再应用练习', '用优势重新设计一件吃力任务的切入角度', '只改变"怎么看待"，不改变任务本身'],
    ['第三章', '1:40-1:50', '10min', '总结', '总结本章3个关键收获，介绍优势档案', '完成优势档案初稿', '预告第四章内容'],

    # Chapter 4
    ['第四章', '0:00-0:05', '5min', '开场', '回顾第三章核心：优势视角，找到内在资源', '回忆', '快速2句话'],
    ['第四章', '0:05-0:15', '10min', '开场/概念', '心理账户比喻+充值频率自查表', '填写自查表：主动充值次数vs被动扣款次数', '引出"你上一次主动充电是什么时候"'],
    ['第四章', '0:15-0:40', '25min', '概念+练习', '感恩练习：机制讲解+标准版跟练', '完成一次完整的"感恩三件事"练习', '"具体事件+为什么发生"，缺一不可'],
    ['第四章', '0:40-0:55', '15min', '练习', '习惯绑定：找到每天必做的事，绑定感恩练习', '为自己设计一个触发点', '"刷牙后"vs"有空时"，后者不可靠'],
    ['第四章', '0:55-1:15', '20min', '概念+练习', '心流讲解+挑战-技能坐标分析+触发器识别', '用坐标图定位日常活动，找到心流触发器', '"进入心流"vs"感到无聊/焦虑"'],
    ['第四章', '1:15-1:30', '15min', '概念+练习', '拓展-建构理论+个人充值菜单设计', '分三档设计：5分钟/20-30分钟/半天', '这是本章最重要的产出'],
    ['第四章', '1:30-1:40', '10min', '总结', '总结本章3个关键收获', '记下3个确定会用的充值来源', '预告第五章内容'],

    # Chapter 5
    ['第五章', '0:00-0:05', '5min', '开场', '回顾第四章核心：主动充值，而不是等大事', '回忆', '快速2句话'],
    ['第五章', '0:05-0:15', '10min', '开场/故事', '讲述"两封邮件"的故事+反思自己的逆境叙事', '回想自己最近的一次挫折反应', '引出"解释方式"的重要性'],
    ['第五章', '0:15-0:30', '15min', '概念讲解', '韧性的三个误解+真正定义（弹簧vs盔甲）', '听讲+重新理解韧性', '关键转变：不是"不受影响"，是"能恢复"'],
    ['第五章', '0:30-0:50', '20min', '概念+练习', '六种认知扭曲讲解+识别练习', '判断6段内心独白的认知扭曲类型', '两人一组讨论，分享最像自己的那种'],
    ['第五章', '0:50-1:15', '25min', '核心练习', 'ABCDE完整讲解+示例跟练', '用老赵的情境完成D和E，然后独立做自己的', '这是本章最重要的练习'],
    ['第五章', '1:15-1:30', '15min', '概念+练习', '创伤后成长讲解+韧性叙事写作', '选择一段过去经历重写叙述', '"这件事在哪个方面成为了我生命的一部分"'],
    ['第五章', '1:30-1:45', '15min', '总结', '总结本章3个关键收获', '记下"下次遇到挫折会用哪个工具"', '预告第六章内容'],

    # Chapter 6
    ['第六章', '0:00-0:05', '5min', '开场', '回顾第五章核心：逆境面前，认知重构让你恢复', '回忆', '快速2句话'],
    ['第六章', '0:05-0:15', '10min', '开场/数据', '91%新年计划失败的数据+意志力vs设计', '反思自己曾经失败的计划', '引出"用设计代替意志力"'],
    ['第六章', '0:15-0:30', '15min', '概念讲解', '习惯的神经机制（基底神经节）+习惯环结构', '理解"习惯不需要意志力"的机制', '触发-行为-奖励，三要素'],
    ['第六章', '0:30-0:50', '20min', '核心练习', 'Tiny Habits讲解+习惯配方练习', '写出3个"习惯配方"', '模板："在我X之后，我立刻做Y"'],
    ['第六章', '0:50-1:15', '25min', '核心练习', '最小幸福例程设计+7天计划', '设计自己的7天最小例程：晨间+日间+晚间+最难那天', '总时长不超过20分钟/天'],
    ['第六章', '1:15-1:30', '15min', '整合', 'PERMA前后对比+工具箱整合', '对比第一章的基准线，选出真正要用的工具', '看到六章之后发生了什么'],
    ['第六章', '1:30-1:50', '20min', '整合', '个人幸福蓝图综合填写', '写出自己的幸福系统描述', '这是本课程的结业成果'],
    ['第六章', '1:50-2:00', '10min', '收尾', '30天行动承诺书', '写下3件承诺做的事+绑定触发', '把承诺书放在每天能看到的地方'],
]

for row_idx, row_data in enumerate(data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = thin_border

        # Color coding by chapter
        chapter = row_data[0]
        if '第一章' in chapter:
            if col_idx == 1:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        elif '第二章' in chapter:
            if col_idx == 1:
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        elif '第三章' in chapter:
            if col_idx == 1:
                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        elif '第四章' in chapter:
            if col_idx == 1:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        elif '第五章' in chapter:
            if col_idx == 1:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        elif '第六章' in chapter:
            if col_idx == 1:
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        # Alternate row colors
        if row_idx % 2 == 0 and '第一章' not in chapter and '第二章' not in chapter and '第三章' not in chapter and '第四章' not in chapter and '第五章' not in chapter and '第六章' not in chapter:
            cell.fill = light_fill

        # Time column
        if col_idx == 3:
            cell.fill = time_fill

# Column widths
col_widths = [10, 12, 8, 10, 35, 30, 25]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Row heights
for row in range(1, 100):
    ws.row_dimensions[row].height = 20

ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 25

# Add legend
ws['I3'] = '环节类型图例'
ws['I3'].font = subheader_font
legend_data = [
    ['开场/引入', '引入主题，激活已有认知'],
    ['讲解', '讲师单向讲授核心概念'],
    ['练习', '学员动手做，包含个人/两人/全班'],
    ['反思', '学员内化，与自身情境连接'],
    ['总结', '提炼要点，预告下章'],
    ['综合练习', '整合多工具的综合性产出练习'],
]

for row_idx, (ltype, desc) in enumerate(legend_data, 4):
    cell = ws.cell(row=row_idx, column=9, value=ltype)
    cell2 = ws.cell(row=row_idx, column=10, value=desc)
    cell.font = Font(name='Microsoft YaHei', size=9, bold=True)
    cell2.font = normal_font

wb.save('D:/新课开发/情绪与心理学/普通人积极心理学实操课/完成课程包/09_成果demo/07_Interaction_Flow.xlsx')
print("07_Interaction_Flow.xlsx created successfully")

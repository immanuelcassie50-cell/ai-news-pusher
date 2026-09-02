# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Create workbook for 填写示例表单
wb = openpyxl.Workbook()

# ===== 学员信息表 =====
ws1 = wb.active
ws1.title = "学员信息表"
ws1.merge_cells("A1:D1")
ws1["A1"] = "学员信息表"
ws1["A1"].font = Font(name="微软雅黑", size=16, bold=True)
ws1["A1"].alignment = Alignment(horizontal="center")

headers1 = ["项目", "填写内容", "说明", "备注"]
for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=col, value=header)
    cell.font = Font(name="微软雅黑", size=12, bold=True)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")

fields = [
    ("姓名", "张明", "必填", "示例学员"),
    ("所属机构/单位", "长江商学院", "必填", "示例单位"),
    ("职位/角色", "高级管理顾问", "选填", ""),
    ("联系电话", "138-0000-1234", "选填", ""),
    ("电子邮箱", "zhangming@company.com", "选填", ""),
    ("报名渠道", "机构推荐", "选填", ""),
    ("学习目标", "深入理解东西方经济发展差异，提升跨文化管理能力", "请简要描述您的学习目标", "")
]

for row, (field, value, note, remark) in enumerate(fields, 4):
    ws1.cell(row=row, column=1, value=field).font = Font(name="微软雅黑", size=11, bold=True)
    ws1.cell(row=row, column=2, value=value)
    ws1.cell(row=row, column=3, value=note).font = Font(name="微软雅黑", size=10, color="808080")
    ws1.cell(row=row, column=4, value=remark).font = Font(name="微软雅黑", size=10, color="808080")

for col in range(1, 5):
    ws1.column_dimensions[get_column_letter(col)].width = 20

# ===== 课前调研问卷 =====
ws2 = wb.create_sheet("课前调研问卷")
ws2.merge_cells("A1:E1")
ws2["A1"] = "课前调研问卷"
ws2["A1"].font = Font(name="微软雅黑", size=16, bold=True)
ws2["A1"].alignment = Alignment(horizontal="center")

ws2["A3"] = "序号"
ws2["B3"] = "问题"
ws2["C3"] = "选项A"
ws2["D3"] = "选项B"
ws2["E3"] = "您的答案"

for col in range(1, 6):
    ws2.cell(row=3, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws2.cell(row=3, column=col).font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

questions = [
    ("您对'大分流'这个概念的了解程度？", "非常了解", "听说过但不太清楚", "B"),
    ("您最关注东西方比较的哪个维度？", "经济发展路径", "制度与文化", "历史事件", "A"),
    ("您参加本次课程的主要目的是？", "学术研究", "管理实践", "个人兴趣", "B"),
    ("您之前是否学习过类似的历史经济课程？", "是，系统学习过", "是，泛泛了解", "C"),
    ("您对课程最期待的收获是？", "理论框架", "实践方法", "案例分析", "B"),
]

for row, (q, *opts) in enumerate(questions, 4):
    ws2.cell(row=row, column=1, value=row-3)
    ws2.cell(row=row, column=2, value=q)
    ws2.cell(row=row, column=3, value=opts[0] if len(opts) > 0 else "")
    ws2.cell(row=row, column=4, value=opts[1] if len(opts) > 1 else "")
    ws2.cell(row=row, column=5, value=opts[2] if len(opts) > 2 else "")

for col in range(1, 6):
    ws2.column_dimensions[get_column_letter(col)].width = 25

# ===== 课堂笔记模板 =====
ws3 = wb.create_sheet("课堂笔记模板")
ws3.merge_cells("A1:F1")
ws3["A1"] = "课堂笔记模板"
ws3["A1"].font = Font(name="微软雅黑", size=16, bold=True)
ws3["A1"].alignment = Alignment(horizontal="center")

headers3 = ["模块", "时间", "核心概念", "个人理解", "疑问/反思", "行动项"]
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col, value=header)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

modules = [
    ("课程导论", "9:00-10:30", "大分流概念、东西方比较框架", "理解历史分叉点的重要性", "为何18世纪后分化加速？", "阅读彭慕兰著作"),
    ("资源禀赋", "10:45-12:00", "土地、劳动力、资本配置", "生态约束对路径的影响", "环境因素的决定性如何？", "分析中国煤炭使用"),
    ("市场整合", "14:00-15:30", "商业网络、市场密度", "理解市场看不见的手", "东西方市场整合差异？", "研究晋商票号制度"),
    ("制度分析", "15:45-17:00", "国家能力、产权保护", "制度对经济的塑造作用", "强国家等于强经济？", "比较英荷制度差异"),
    ("殖民资本", "次日9:00", "原始积累、殖民剥削", "资本主义的阴暗面", "殖民收益如何计量？", "整理印度案例"),
    ("现代启示", "次日10:30", "文明对话、人类命运共同体", "历史照进现实", "如何避免修昔底德陷阱？", "思考中美关系"),
]

for row, module_data in enumerate(modules, 4):
    for col, value in enumerate(module_data, 1):
        ws3.cell(row=row, column=col, value=value)

for col in range(1, 7):
    ws3.column_dimensions[get_column_letter(col)].width = 22

# ===== 四维分析工作表 =====
ws4 = wb.create_sheet("四维分析工作表")
ws4.merge_cells("A1:E1")
ws4["A1"] = "四维分析工作表"
ws4["A1"].font = Font(name="微软雅黑", size=16, bold=True)
ws4["A1"].alignment = Alignment(horizontal="center")

headers4 = ["维度", "东方（中国）", "西方（欧洲）", "异同分析", "个人见解"]
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=col, value=header)
    cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

dimensions = [
    ("经济制度", "中央集权、官营经济、重农抑商", "市场经济、私有财产、自由贸易", "都强调稳定，但实现路径不同", "国家资本主义是否有独特优势？"),
    ("政治体制", "君主专制、科举取士、文官治国", "君主立宪/民主选举、议会制度", "东西方都重视人才选拔，但方式迥异", "科举制是否是早期的精英选拔？"),
    ("文化传统", "儒家伦理、重义轻利、集体主义", "新教伦理、资本积累、个人主义", "对财富态度截然不同", "儒教能否催生资本主义？"),
    ("技术发展", "四大发明、农业技术、丝绸瓷器", "工业革命、机械化、科学技术", "西方技术爆发，中国相对停滞", "李约瑟难题的当代解答"),
]

for row, dim_data in enumerate(dimensions, 4):
    for col, value in enumerate(dim_data, 1):
        ws4.cell(row=row, column=col, value=value)

for col in range(1, 6):
    ws4.column_dimensions[get_column_letter(col)].width = 25

# ===== 课后反思报告 =====
ws5 = wb.create_sheet("课后反思报告")
ws5.merge_cells("A1:D1")
ws5["A1"] = "课后反思报告"
ws5["A1"].font = Font(name="微软雅黑", size=16, bold=True)
ws5["A1"].alignment = Alignment(horizontal="center")

reflection_items = [
    ("最深刻的收获", "彭慕兰的'地理决定论'vs制度决定论的交锋，让我重新审视历史复杂性", "请描述本次课程中最让您印象深刻的知识点或观点", ""),
    ("与预期的差距", "原以为课程偏学术，没想到对管理实践也有直接启发", "与课前预期相比，实际情况如何", ""),
    ("实践应用计划", "将东西方比较框架用于跨文化管理培训课程设计", "计划如何将所学应用到实际工作/研究中", ""),
    ("待深入了解的领域", "英荷东印度公司的治理模式与当代企业治理的比较", "还有哪些方面希望进一步学习", ""),
    ("对课程的建议", "希望增加更多中国本土案例，如晋商、徽商的分析", "对课程内容/形式/组织的建议", ""),
]

row = 3
for item, placeholder, note, remark in reflection_items:
    ws5.cell(row=row, column=1, value=item).font = Font(name="微软雅黑", size=11, bold=True)
    ws5.cell(row=row, column=2, value=placeholder)
    ws5.cell(row=row, column=3, value=note).font = Font(name="微软雅黑", size=10, color="808080")
    row += 1

for col in range(1, 4):
    ws5.column_dimensions[get_column_letter(col)].width = 25

# ===== 学习成果展示模板 =====
ws6 = wb.create_sheet("学习成果展示模板")
ws6.merge_cells("A1:C1")
ws6["A1"] = "学习成果展示模板"
ws6["A1"].font = Font(name="微软雅黑", size=16, bold=True)
ws6["A1"].alignment = Alignment(horizontal="center")

ws6["A3"] = "项目"
ws6["B3"] = "内容"
ws6["C3"] = "说明"

for col in range(1, 4):
    ws6.cell(row=3, column=col).fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    ws6.cell(row=3, column=col).font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

display_items = [
    ("主题标题", "从大分流看跨文化管理：历史教训与现代启示", "概括您的展示主题", ""),
    ("核心观点", "1. 东西方分流始于制度选择差异\n2. 文化基因影响经济行为模式\n3. 文明对话优于文明冲突", "列出2-3个核心观点", ""),
    ("案例/证据", "彭慕兰《大分流》、韦伯《新教伦理》、中国科举制与英国议会制的比较", "支撑观点的案例或数据", ""),
    ("实践意义", "跨国企业需理解文化深层结构，避免用单一管理模式处理多元团队", "对管理实践的启示", ""),
    ("展示时长", "8分钟", "建议5-10分钟", ""),
]

row = 4
for item, placeholder, note, remark in display_items:
    ws6.cell(row=row, column=1, value=item).font = Font(name="微软雅黑", size=11, bold=True)
    ws6.cell(row=row, column=2, value=placeholder)
    ws6.cell(row=row, column=3, value=note).font = Font(name="微软雅黑", size=10, color="808080")
    row += 1

for col in range(1, 4):
    ws6.column_dimensions[get_column_letter(col)].width = 25

wb.save("D:/新课开发/经济学/30_大分流/配套表单和指引/配套表单_填好版.xlsx")
print("配套表单_填好版.xlsx created successfully")

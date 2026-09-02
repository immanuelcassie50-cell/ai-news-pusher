#!/usr/bin/env python3
"""Create Cold War course Excel supporting tables."""

import os
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SKILL_DIR = "C:/Users/Administrator/.claude/skills/Excel表格处理"
TEMPLATE_DIR = "/tmp/xlsx_work"
OUTPUT_BASE = "D:/新课开发/政治学/16_冷战重访-意识形态对抗的政治遗产/Excel"

def copy_template():
    """Copy minimal template to work directory."""
    if os.path.exists(TEMPLATE_DIR):
        shutil.rmtree(TEMPLATE_DIR)
    shutil.copytree(f"{SKILL_DIR}/templates/minimal_xlsx", TEMPLATE_DIR)

def style_header(ws, row_num, col_start, col_end):
    """Apply header style to a range of cells."""
    header_font = Font(name='Microsoft YaHei', bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name='Microsoft YaHei', bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def style_data_cell(cell):
    """Apply data cell style."""
    cell.font = Font(name='Microsoft YaHei', size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border

def set_column_widths(ws, widths):
    """Set column widths."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def create_kcznr_zonglan():
    """Create 课程内容总览.xlsx"""
    wb = Workbook()

    # ===== Sheet 1: 模块总览 =====
    ws1 = wb.active
    ws1.title = "模块总览"

    headers1 = ["模块编号", "模块名称", "核心问题", "关键概念", "课时"]
    data1 = [
        ["01", "什么是冷战", "冷战如何定义？有哪些核心特征？", "冷战定义、四大特征、诊断清单", "4"],
        ["02", "意识形态对抗", "意识形态如何在冷战中发挥作用？", "意识形态矩阵、三条动员路径", "4"],
        ["03", "代理人战争", "为何大国选择代理人战争？", "三角模型、利益分离风险", "3"],
        ["04", "冷战思维延续", "冷战思维如何延续至今？", "技术脱钩、阵营化外交、话语竞争", "3"],
        ["05", '"新冷战"再思考', '"新冷战"概念是否适用？', "八维诊断表、三大陷阱", "3"],
        ["06", "独立思考工具", "如何批判性地分析国际关系话语？", "四步分析法、逻辑谬误、检查清单", "3"],
    ]

    # Write headers
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, 1, 5)

    # Write data
    for row_idx, row_data in enumerate(data1, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(cell)
            if col_idx == 5:  # 课时列居中
                cell.alignment = Alignment(horizontal='center', vertical='center')

    set_column_widths(ws1, [12, 18, 35, 40, 8])
    ws1.row_dimensions[1].height = 25

    # ===== Sheet 2: 知识点清单 =====
    ws2 = wb.create_sheet("知识点清单")

    headers2 = ["模块", "知识点", "掌握程度", "备注"]
    data2 = [
        ["01", "冷战的定义与时间范围", "掌握", "理解冷战的起点与终点争议"],
        ["01", "冷战的四大核心特征", "掌握", "两大阵营、意识形态对抗、代理人战争、核威慑"],
        ["01", "冷战与热战的区别", "理解", "间接对抗、直接战争的不同形式"],
        ["01", "冷战诊断清单的使用", "应用", "用六问诊断表判断冷战的适用性"],
        ["02", "意识形态的定义与功能", "掌握", "意识形态作为政治动员工具"],
        ["02", "资本主义vs共产主义的对立矩阵", "掌握", "市场经济vs计划经济、民主vs专制的核心差异"],
        ["02", "意识形态动员的三条路径", "掌握", "宣传动员、群众动员、制度动员"],
        ["02", "意识形态在冷战中的双重作用", "理解", "既是冲突根源，也是合法性来源"],
        ["03", "代理人战争的定义", "掌握", "大国通过第三方进行的间接战争"],
        ["03", "三角模型的三个顶点", "掌握", "代理人、两大国、当地政府"],
        ["03", "利益分离风险", "理解", "代理人拥有独立利益的风险"],
        ["03", "冷战中典型的代理人战争", "了解", "朝鲜战争、越南战争、阿富汗战争等"],
        ["04", "冷战思维的定义", "掌握", "将世界分为敌我阵营的二元思维"],
        ["04", "技术脱钩的表现", "理解", "芯片禁令、技术标准分裂"],
        ["04", "阵营化外交", "理解", "结盟对抗、势力范围"],
        ["04", "话语竞争", "理解", "不同制度模式的话语权争夺"],
        ["04", "冷战遗产的持续性", "理解", "北约东扩、后苏联秩序"],
        ["05", '"新冷战"概念争议', "掌握", "该概念是否准确反映当前国际形势"],
        ["05", "八维诊断表的使用", "应用", "从八个维度评估新冷战的程度"],
        ["05", "三大陷阱", "理解", "修昔底德陷阱、中等收入陷阱、联盟陷阱"],
        ["05", "当前国际形势与冷战的不同", "理解", "多极化、非意识形态化、全球化"],
        ["06", "四步分析法", "掌握", "描述→诊断→评估→建议"],
        ["06", "常见逻辑谬误", "掌握", "稻草人谬误、虚假两难诉诸权威"],
        ["06", "批判性思维检查清单", "应用", "用于评估国际关系话语的可靠性"],
        ["06", "如何识别偏见性信息源", "应用", "交叉验证、多元来源"],
        ["06", "独立思考的重要性", "理解", "避免二元对立、保持分析中立"],
    ]

    # Write headers
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, 1, 4)

    # Write data
    for row_idx, row_data in enumerate(data2, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(cell)
            if col_idx == 3:  # 掌握程度列居中
                cell.alignment = Alignment(horizontal='center', vertical='center')

    set_column_widths(ws2, [10, 30, 12, 35])
    ws2.row_dimensions[1].height = 25

    # Save
    path = f"{OUTPUT_BASE}/课程内容总览.xlsx"
    wb.save(path)
    print(f"Created: {path}")

def create_lengzhan_zhenduan():
    """Create 冷战诊断工具.xlsx"""
    wb = Workbook()

    # ===== Sheet 1: 六问诊断表 =====
    ws1 = wb.active
    ws1.title = "六问诊断表"

    headers1 = ["问题编号", "诊断问题", "是", "否", "说明"]
    data1 = [
        ["1", "是否存在明确的两大阵营对立？", "□", "□", ""],
        ["2", "意识形态是否为核心驱动力？", "□", "□", ""],
        ["3", "是否存在代理人战争模式？", "□", "□", ""],
        ["4", "是否有军备竞赛或核威慑因素？", "□", "□", ""],
        ["5", "是否有明确的话语/意识形态斗争？", "□", "□", ""],
        ["6", "国内政治是否受国际意识形态斗争影响？", "□", "□", ""],
    ]

    # Write headers
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, 1, 5)

    # Write data
    for row_idx, row_data in enumerate(data1, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(cell)
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in [3, 4]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Summary row
    summary_row = 8
    ws1.cell(row=summary_row, column=1, value="综合判断:")
    ws1.cell(row=summary_row, column=1).font = Font(name='Microsoft YaHei', bold=True, size=10)
    ws1.merge_cells(f'A{summary_row}:E{summary_row}')
    ws1.cell(row=summary_row, column=1).alignment = Alignment(horizontal='left', vertical='center')

    criteria = [
        '6-5个"是" = 高度符合冷战特征',
        '4-3个"是" = 部分符合冷战特征',
        '2及以下 = 不符合冷战特征'
    ]
    for i, c in enumerate(criteria, summary_row + 1):
        ws1.cell(row=i, column=1, value=c)
        ws1.cell(row=i, column=1).font = Font(name='Microsoft YaHei', size=10)
        ws1.merge_cells(f'A{i}:E{i}')
        ws1.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')

    set_column_widths(ws1, [12, 40, 8, 8, 30])
    ws1.row_dimensions[1].height = 25

    # ===== Sheet 2: 八维诊断表 =====
    ws2 = wb.create_sheet("八维诊断表")

    headers2 = ["维度", "诊断问题", "判断(高/中/低/无)", "与冷战的关键差异"]
    data2 = [
        ["1", "意识形态中心性", "", "冷战以意识形态为核心驱动力，当前大国竞争更侧重技术和经济"],
        ["2", "经济体系分离", "", "冷战形成两大平行市场（布雷顿森林体系vs经互会），当前经济相互依存度更高"],
        ["3", "联盟体系对抗", "", "冷战有清晰的军事联盟（北约vs华约），当前盟国关系更为复杂"],
        ["4", "代理人战争模式", "", "冷战有大量代理人战争，当前更多采用混合战争和信息战"],
        ["5", "军备竞赛", "", "冷战有激烈的核军备竞赛，当前主要聚焦高超音速武器和网络战"],
        ["6", "话语体系对立", "", "冷战有资本主义vs共产主义的清晰话语对立，当前更多是发展模式竞争"],
        ["7", "国内政治动员", "", "冷战时期意识形态对国内政治影响更为直接和深入"],
        ["8", "技术体系分化", "", "冷战有技术脱钩（如巴黎统筹委员会），当前主要是芯片和5G领域的竞争"],
    ]

    # Write headers
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, 1, 4)

    # Write data
    for row_idx, row_data in enumerate(data2, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(cell)
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    set_column_widths(ws2, [10, 20, 22, 50])
    ws2.row_dimensions[1].height = 25

    # Save
    path = f"{OUTPUT_BASE}/冷战诊断工具.xlsx"
    wb.save(path)
    print(f"Created: {path}")

def create_anli_fenxi():
    """Create 案例分析表.xlsx"""
    wb = Workbook()

    # ===== Sheet 1: 历史案例 =====
    ws1 = wb.active
    ws1.title = "历史案例"

    headers1 = ["案例", "时间", "类型", "符合冷战特征(是/否)", "关键分析"]
    data1 = [
        ["越南战争", "1955-1975", "代理人战争", "是", "美国支持南越，苏联和中国支持北越，体现了典型的代理人战争模式"],
        ["阿富汗抗苏战争", "1979-1989", "代理人战争", "是", "美国通过巴基斯坦向阿富汗抵抗力量提供武器，苏联深陷战争泥潭"],
        ["古巴导弹危机", "1962", "核威慑", "是", "美苏冷战期间最接近核战争的事件，体现了核威慑的作用"],
        ["匈牙利事件", "1956", "意识形态斗争", "是", "苏联对匈牙利革命的镇压，体现了意识形态对内政的影响"],
        ["马歇尔计划", "1947-1951", "意识形态动员", "是", "美国对西欧的经济援助计划，同时带有遏制共产主义的战略目的"],
        ["柏林墙危机", "1961", "冷战对峙", "是", "东西方在柏林的直接对抗，体现了冷战的两极格局"],
        ["朝鲜战争", "1950-1953", "代理人战争", "是", "冷战初期最重要的代理人战争之一"],
        ["北约与华约成立", "1949/1955", "联盟对抗", "是", "标志着两大军事集团的正式形成"],
    ]

    # Write headers
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, 1, 5)

    # Write data
    for row_idx, row_data in enumerate(data1, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(cell)
            if col_idx == 4:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    set_column_widths(ws1, [18, 15, 15, 20, 45])
    ws1.row_dimensions[1].height = 25

    # ===== Sheet 2: 当代案例 =====
    ws2 = wb.create_sheet("当代案例")

    headers2 = ["案例", "时间", "类型", "符合新冷战特征程度", "关键分析"]
    data2 = [
        ["芯片禁令", "2018-至今", "技术脱钩", "部分符合", "美国限制对华芯片技术出口，体现技术竞争，但尚未形成两大平行体系"],
        ["俄乌冲突", "2022-至今", "混合战争", "部分符合", "包含传统战争、信息战、经济战，但背后驱动力更复杂"],
        ["一带一路", "2013-至今", "制度竞争", "不符合", "更多是经济合作倡议，非冷战式的意识形态输出"],
        ["中美贸易战", "2018-至今", "经济竞争", "部分符合", "包含关税、科技、地缘政治多重因素，但双方经济高度相互依存"],
        ["5G技术竞争", "2019-至今", "技术标准竞争", "部分符合", "华为与西方供应商的竞争，但多数国家未选边站队"],
        ["香港问题", "2019-至今", "制度竞争", "部分符合", "涉及自治与国家安全，但不同于冷战式的意识形态对抗"],
    ]

    # Write headers
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, 1, 5)

    # Write data
    for row_idx, row_data in enumerate(data2, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(cell)
            if col_idx == 4:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    set_column_widths(ws2, [18, 15, 15, 20, 45])
    ws2.row_dimensions[1].height = 25

    # Save
    path = f"{OUTPUT_BASE}/案例分析表.xlsx"
    wb.save(path)
    print(f"Created: {path}")

def create_yanshen_yuedu():
    """Create 延伸阅读.xlsx"""
    wb = Workbook()

    # ===== Sheet 1: 推荐书单 =====
    ws1 = wb.active
    ws1.title = "推荐书单"

    headers1 = ["类别", "书名", "作者", "出版社", "核心观点"]
    data1 = [
        ["冷战史", "冷战史", "约翰·刘易斯·加迪斯", "社科文献出版社", "冷战不是必然的，是一系列决策的结果"],
        ["冷战史", "冷战最终篇", "小约瑟夫·奈", "北京大学出版社", "重新审视冷战的教训与遗产"],
        ["冷战史", "全球冷战", "梅尔文·莱夫勒", "江苏人民出版社", "美苏在全球的争夺与第三世界"],
        ["意识形态", "意识形态与意识形态斗争", "大卫·温斯顿", "中国人民大学出版社", "意识形态在政治中的作用机制"],
        ["意识形态", "意识形态的兴衰", "阿兰·瑞安", "上海译文出版社", "从历史看意识形态的演变"],
        ["国际关系理论", "大国政治的悲剧", "约翰·米尔斯海默", "上海人民出版社", "进攻性现实主义视角下的冷战"],
        ["国际关系理论", "冷战后国际关系理论", "罗伯特·基欧汉", "北京大学出版社", "冷战后国际关系理论的发展"],
        ["案例研究", "越南战争", "陈音箱", "浙江大学出版社", "越南战争的根源与影响"],
        ["案例研究", "阿富汗战争", "克莱夫·琼斯", "上海译文出版社", "苏联入侵阿富汗的教训"],
        ["当代思考", "新冷战？", "亨利·基辛格", "上海人民出版社", "当前大国竞争与冷战的比较"],
        ["当代思考", "美国与中国", "Graham Allison", "Simon & Schuster", "修昔底德陷阱与中美关系"],
        ["思想史", "意识形态的起源", "卢卡奇", "商务印书馆", "意识形态概念的起源与发展"],
    ]

    # Write headers
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, 1, 5)

    # Write data
    for row_idx, row_data in enumerate(data1, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(cell)

    set_column_widths(ws1, [12, 22, 20, 18, 40])
    ws1.row_dimensions[1].height = 25

    # Save
    path = f"{OUTPUT_BASE}/延伸阅读.xlsx"
    wb.save(path)
    print(f"Created: {path}")

def main():
    """Create all Excel files."""
    os.makedirs(OUTPUT_BASE, exist_ok=True)

    create_kcznr_zonglan()
    create_lengzhan_zhenduan()
    create_anli_fenxi()
    create_yanshen_yuedu()

    print("\nAll Excel files created successfully!")

if __name__ == "__main__":
    main()

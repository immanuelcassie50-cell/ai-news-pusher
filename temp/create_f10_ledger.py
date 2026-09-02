# -*- coding: utf-8 -*-
"""
创建 F10 金融纠纷案例台账 Excel 文件
使用 openpyxl 库
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 创建工作簿
wb = Workbook()

# ==================== Sheet 1: 案例台账 ====================
ws1 = wb.active
ws1.title = "\u6848\u4F8B\u53F0\u8D26"

# 定义列配置
columns = [
    ("\u6848\u4F8B\u7F16\u53F7", "A", 12),
    ("\u7EA0\u7EB0\u7C7B\u578B", "B", 14),
    ("\u4EA7\u54C1\u540D\u79F0", "C", 20),
    ("\u8D2D\u4E70\u91D1\u989D(\u4E07)", "D", 14),
    ("\u635F\u5931\u91D1\u989D(\u4E07)", "E", 14),
    ("\u8FFD\u56DE\u91D1\u989D(\u4E07)", "F", 14),
    ("\u8FFD\u56DE\u6BD4\u4F8B", "G", 12),
    ("\u8D2D\u4E70\u65E5\u671F", "H", 14),
    ("\u53D1\u73B0\u95EE\u9898\u65E5\u671F", "I", 14),
    ("\u7EF4\u6743\u5F00\u59CB\u65E5\u671F", "J", 14),
    ("\u7ED3\u6848\u65E5\u671F", "K", 14),
    ("\u7EF4\u6743\u8017\u65F6(\u6708)", "L", 14),
    ("\u9500\u552E\u673A\u6784", "M", 18),
    ("\u6295\u8BC9\u6E20\u9053", "N", 14),
    ("\u6700\u7EC8\u7ED3\u679C", "O", 12),
    ("\u6838\u5FC3\u8FDD\u89C4\u70B9", "P", 25),
    ("\u5173\u952E\u8BC1\u636E", "Q", 25),
    ("\u7ECF\u9A8C\u6559\u8BAD", "R", 30),
    ("\u5907\u6CE8", "S", 20),
]

# 设置表头
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=11, bold=True, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_name, col_letter, col_width in columns:
    cell = ws1[f"{col_letter}1"]
    cell.value = col_name
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    ws1.column_dimensions[col_letter].width = col_width

# 冻结首行
ws1.freeze_panes = "A2"

# 数据验证 - 纠纷类型下拉列表
dispute_type_dv = DataValidation(
    type="list",
    formula1='"\u94F6\u884C\u7406\u8D22,\u57FA\u91D1\u9500\u552E,\u4FDD\u9669\u8BEF\u5BFC,\u975E\u6CD5\u96C6\u8D44"',
    allow_blank=True,
    showDropDown=False
)
ws1.add_data_validation(dispute_type_dv)
dispute_type_dv.add("B2:B1000")

# 投诉渠道下拉列表
complaint_channel_dv = DataValidation(
    type="list",
    formula1='"\u5185\u90E8\u6295\u8BC9,\u76D1\u7BA1\u6295\u8BC9,\u8C03\u89E3,\u8BC9\u8BBC,\u5211\u4E8B\u62A5\u6848"',
    allow_blank=True,
    showDropDown=False
)
ws1.add_data_validation(complaint_channel_dv)
complaint_channel_dv.add("N2:N1000")

# 最终结果下拉列表
result_dv = DataValidation(
    type="list",
    formula1='"\u80DC\u8BC9,\u548C\u89E3,\u8D25\u8BC9,\u672A\u7ED3"',
    allow_blank=True,
    showDropDown=False
)
ws1.add_data_validation(result_dv)
result_dv.add("O2:O1000")

# 样本数据 - 5个金融消费者维权案例
sample_data = [
    {
        "case_id": "FC2024001",
        "dispute_type": "\u94F6\u884C\u7406\u8D22",
        "product_name": "\u67D0\u884C\"\u7A33\u4EAB\"\u7CFB\u5217\u7406\u8D22\u4EA7\u54C1",
        "purchase_amount": 100.0,
        "loss_amount": 45.0,
        "recovered_amount": 38.0,
        "purchase_date": "2021-03-15",
        "discovery_date": "2022-06-20",
        "rights_start_date": "2022-07-01",
        "close_date": "2023-03-15",
        "sales_institution": "\u67D0\u56FD\u6709\u94F6\u884C\u652F\u884C",
        "complaint_channel": "\u8BC9\u8BBC",
        "final_result": "\u548C\u89E3",
        "violation_points": "\u672A\u5145\u5206\u63D0\u793A\u4EA7\u54C1\u98CE\u9669\uFF0C\u98CE\u9669\u8BC4\u7EA7\u4E0E\u5BA2\u6237\u98CE\u9669\u627F\u53D7\u80FD\u529B\u4E0D\u5339\u914D",
        "key_evidence": "\u4EA7\u54C1\u8BF4\u660E\u4E66\u3001\u98CE\u9669\u8BC4\u4F30\u95EE\u5377\u3001\u9500\u552E\u5F55\u97F3\u3001\u5BA2\u6237\u7B7E\u5B57\u6587\u4EF6",
        "lessons": "\u8D2D\u4E70\u524D\u52A1\u5FC5\u5B8C\u6210\u98CE\u9669\u8BC4\u4F30\uFF0C\u8BA4\u771F\u9605\u8BFB\u4EA7\u54C1\u8BF4\u660E\u4E66\uFF0C\u7279\u522B\u662F\u98CE\u9669\u63ED\u793A\u6761\u6B3E",
        "notes": "\u7ECF\u8C03\u89E3\u8FBE\u6210\u548C\u89E3\u534F\u8BAE\uFF0C\u5206\u671F\u5FEB\u8FD8"
    },
    {
        "case_id": "FC2024002",
        "dispute_type": "\u57FA\u91D1\u9500\u552E",
        "product_name": "\u67D0\u660E\u661F\u57FA\u91D1\u7ECF\u7406\u79C1\u52DF\u57FA\u91D1",
        "purchase_amount": 300.0,
        "loss_amount": 180.0,
        "recovered_amount": 60.0,
        "purchase_date": "2020-11-10",
        "discovery_date": "2022-04-01",
        "rights_start_date": "2022-04-15",
        "close_date": "2024-01-20",
        "sales_institution": "\u67D0\u8BC1\u5238\u516C\u53F8",
        "complaint_channel": "\u8BC9\u8BBC",
        "final_result": "\u80DC\u8BC9",
        "violation_points": "\u6269\u5927\u9884\u671F\u6536\u76CA\uFF0C\u9690\u7791\u57FA\u91D1\u5B9E\u9645\u6301\u4ED3\uFF0C\u96C6\u4E2D\u6301\u4ED3\u9AD8\u98CE\u9669\u80A1\u7968",
        "key_evidence": "\u57FA\u91D1\u5408\u540C\u3001\u52DF\u96C6\u8BF4\u660E\u4E66\u3001\u94F6\u884C\u8F6C\u8D26\u8BB0\u5F55\u3001\u57FA\u91D1\u7BA1\u7406\u4EBA\u5BA3\u4F20\u6750\u6599",
        "lessons": "\u8B66\u60D8\u9AD8\u6536\u76CA\u627F\u8BFA\u79C1\u52DF\u57FA\u91D1\uFF0C\u6838\u5B9E\u57FA\u91D1\u7BA1\u7406\u4EBA\u8D44\u8D28\uFF0C\u4E86\u89E3\u8D44\u91D1\u6258\u7BA1\u60C5\u51B5",
        "notes": "\u6CD5\u9662\u5224\u51B3\u57FA\u91D1\u7BA1\u7406\u4EBA\u8D54\u507F\u90E8\u5206\u635F\u5931"
    },
    {
        "case_id": "FC2024003",
        "dispute_type": "\u4FDD\u9669\u8BEF\u5BFC",
        "product_name": "\u5206\u7EA2\u578B\u4EBA\u5BFF\u4FDD\u9669",
        "purchase_amount": 50.0,
        "loss_amount": 8.0,
        "recovered_amount": 7.5,
        "purchase_date": "2019-08-20",
        "discovery_date": "2023-02-10",
        "rights_start_date": "2023-02-28",
        "close_date": "2023-08-30",
        "sales_institution": "\u67D0\u4FDD\u9669\u516C\u53F8\u4EE3\u7406\u4EBA",
        "complaint_channel": "\u76D1\u7BA1\u6295\u8BC9",
        "final_result": "\u548C\u89E3",
        "violation_points": "\u5C06\u4FDD\u9669\u4EA7\u54C1\u6DF7\u6DC6\u4E3A\u94F6\u884C\u5B58\u6B3E\u63A8\u9500\uFF0C\u6269\u5927\u5206\u7EA2\u6536\u76CA\u6F14\u793A",
        "key_evidence": "\u4FDD\u9669\u5408\u540C\u3001\u9500\u552E\u4EBA\u5458\u5BA3\u4F20\u8D44\u6599\u3001\u5FAE\u4FE1\u804A\u5929\u8BB0\u5F55\u3001\u94F6\u884C\u8F6C\u8D26\u51ED\u8BC1",
        "lessons": "\u4FDD\u9669\u4EA7\u54C1\u7B49\u4E8E\u5B58\u6B3E\uFF0C\u8981\u533A\u5206\u4FDD\u8BC1\u6536\u76CA\u4E0E\u6F14\u793A\u6536\u76CA\uFF0C\u6CE8\u610F\u9000\u4FDD\u635F\u5931",
        "notes": "\u5168\u989D\u9000\u8FD8\u4FDD\u8D39\u53CA\u90E8\u5206\u5229\u606F"
    },
    {
        "case_id": "FC2024004",
        "dispute_type": "\u975E\u6CD5\u96C6\u8D44",
        "product_name": "\u67D0P2P\u7F51\u8D37\u5E73\u53F0\u51FA\u501F",
        "purchase_amount": 200.0,
        "loss_amount": 160.0,
        "recovered_amount": 25.0,
        "purchase_date": "2019-06-01",
        "discovery_date": "2020-09-15",
        "rights_start_date": "2020-10-01",
        "close_date": "2023-12-31",
        "sales_institution": "\u67D0P2P\u7F51\u8D37\u5E73\u53F0",
        "complaint_channel": "\u5211\u4E8B\u62A5\u6848",
        "final_result": "\u672A\u7ED3",
        "violation_points": "\u5E73\u53F0\u81EA\u878D\u3001\u865A\u6784\u5019\u6B20\u4EBA\u3001\u8D44\u91D1\u6C60\u8FD0\u4F5C\u3001\u627F\u8BFA\u4FDD\u672C\u4FDD\u606F",
        "key_evidence": "\u51FA\u501F\u5408\u540C\u3001\u5E73\u53F0\u5BA3\u4F20\u8D44\u6599\u3001\u94F6\u884C\u8D44\u91D1\u6D41\u6C34\u3001\u5E73\u53F0\u8FD0\u8425\u6570\u636E",
        "lessons": "\u8FDC\u79BB\u8D44\u91D1\u6C60\u6A21\u5F0F\u5E73\u53F0\uFF0C\u6838\u5B9E\u5E73\u53F0\u662F\u5426\u5B8C\u6210\u5907\u6848\uFF0C\u5173\u6CE8\u51FA\u501F\u8D44\u91D1\u6D41\u5411",
        "notes": "\u5DF2\u8FDB\u5165\u53F8\u6CD5\u8FFD\u5F81\u7A0D\u7A0D\uFF0C\u8D44\u4EA7\u5904\u7F6E\u4E2D"
    },
    {
        "case_id": "FC2024005",
        "dispute_type": "\u94F6\u884C\u7406\u8D22",
        "product_name": "\u67D0\u57CE\u5546\u884C\u7406\u8D22\u8BA1\u5212",
        "purchase_amount": 80.0,
        "loss_amount": 12.0,
        "recovered_amount": 10.8,
        "purchase_date": "2022-01-05",
        "discovery_date": "2023-05-20",
        "rights_start_date": "2023-06-01",
        "close_date": "2023-11-15",
        "sales_institution": "\u67D0\u57CE\u5546\u884C",
        "complaint_channel": "\u8C03\u89E3",
        "final_result": "\u80DC\u8BC9",
        "violation_points": "\u4EA7\u54C1\u51C0\u503C\u5316\u8F6C\u578B\u4E2D\u672A\u5145\u5206\u544A\u77E5\u6295\u8D44\u8005\uFF0C\u4EA7\u54C1\u8FD0\u4F5C\u4E0D\u900F\u660E",
        "key_evidence": "\u4EA7\u54C1\u8BA4\u8D2D\u534F\u8BAE\u3001\u4EA7\u54C1\u8FD0\u4F5C\u62A5\u544A\u3001\u94F6\u884C\u5BF9\u8D26\u5355\u3001\u6C9F\u901A\u8BB0\u5F55",
        "lessons": "\u51C0\u503C\u578B\u4EA7\u54C1\u9700\u5173\u6CE8\u51C0\u503C\u6CE2\u52A8\uFF0C\u4E86\u89E3\u4EA7\u54C1\u6295\u8D44\u7B56\u7565\uFF0C\u4FDD\u6301\u4E0E\u673A\u6784\u6C9F\u901A",
        "notes": "\u94F6\u884C\u8865\u507F\u90E8\u5206\u635F\u5931\uFF0C\u6295\u8D44\u8005\u8D4E\u56DE\u5206\u989D"
    },
]

# 填充数据行
row_num = 2
for data in sample_data:
    ws1[f"A{row_num}"] = data["case_id"]
    ws1[f"B{row_num}"] = data["dispute_type"]
    ws1[f"C{row_num}"] = data["product_name"]
    ws1[f"D{row_num}"] = data["purchase_amount"]
    ws1[f"E{row_num}"] = data["loss_amount"]
    ws1[f"F{row_num}"] = data["recovered_amount"]
    # 追回比例公式
    ws1[f"G{row_num}"] = f"=IF(F{row_num}>0,F{row_num}/E{row_num},0)"
    ws1[f"G{row_num}"].number_format = "0.0%"

    ws1[f"H{row_num}"] = data["purchase_date"]
    ws1[f"I{row_num}"] = data["discovery_date"]
    ws1[f"J{row_num}"] = data["rights_start_date"]
    ws1[f"K{row_num}"] = data["close_date"]
    # 维权耗时月公式
    ws1[f"L{row_num}"] = f'=IF(K{row_num}<>"",DATEDIF(J{row_num},K{row_num},"M"),"")'

    ws1[f"M{row_num}"] = data["sales_institution"]
    ws1[f"N{row_num}"] = data["complaint_channel"]
    ws1[f"O{row_num}"] = data["final_result"]
    ws1[f"P{row_num}"] = data["violation_points"]
    ws1[f"Q{row_num}"] = data["key_evidence"]
    ws1[f"R{row_num}"] = data["lessons"]
    ws1[f"S{row_num}"] = data["notes"]

    # 日期格式
    for col in ["H", "I", "J", "K"]:
        ws1[f"{col}{row_num}"].number_format = "YYYY-MM-DD"

    row_num += 1

# 添加汇总行
summary_row = row_num
ws1[f"A{summary_row}"] = "\u6C47\u603B"

# 汇总公式
ws1[f"D{summary_row}"] = f"=SUM(D2:D{summary_row-1})"
ws1[f"E{summary_row}"] = f"=SUM(E2:E{summary_row-1})"
ws1[f"F{summary_row}"] = f"=SUM(F2:F{summary_row-1})"
ws1[f"G{summary_row}"] = f"=IF(F{summary_row}>0,F{summary_row}/E{summary_row},0)"
ws1[f"G{summary_row}"].number_format = "0.0%"
ws1[f"L{summary_row}"] = f"=AVERAGE(L2:L{summary_row-1})"

# 汇总行样式
summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
summary_font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=11, bold=True)

for col_letter in ["A", "D", "E", "F", "G", "L"]:
    ws1[f"{col_letter}{summary_row}"].font = summary_font
    ws1[f"{col_letter}{summary_row}"].fill = summary_fill

# 设置行高
ws1.row_dimensions[1].height = 30
for r in range(2, summary_row + 1):
    ws1.row_dimensions[r].height = 45

# ==================== 条件格式 - 颜色编码 ====================
# 追回比例颜色: 绿色(>50%), 黄色(20-50%), 红色(<20%)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# 应用颜色到追回比例列 (G2:G6)
for row in range(2, 7):
    recovered = sample_data[row-2]["recovered_amount"]
    loss = sample_data[row-2]["loss_amount"]
    if loss > 0:
        ratio = recovered / loss
        cell = ws1[f"G{row}"]
        if ratio > 0.5:
            cell.fill = green_fill
        elif ratio >= 0.2:
            cell.fill = yellow_fill
        else:
            cell.fill = red_fill

# ==================== Sheet 2: 统计分析 ====================
ws2 = wb.create_sheet(title="\u7EDF\u8BA1\u5206\u6790")

# 标题
ws2["A1"] = "\u91D1\u878D\u7EA0\u7EB0\u6848\u4F8B\u7EDF\u8BA1\u5206\u6790"
ws2["A1"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=16, bold=True)
ws2.merge_cells("A1:F1")

# 总体统计指标
ws2["A3"] = "\u603B\u4F53\u7EDF\u8BA1\u6307\u6807"
ws2["A3"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=12, bold=True)

stats_labels = [
    ("\u6848\u4F8B\u603B\u6570", "=\u6848\u4F8B\u53F0\u8D26!A7"),
    ("\u603B\u8D2D\u4E70\u91D1\u989D(\u4E07)", "=\u6848\u4F8B\u53F0\u8D26!D7"),
    ("\u603B\u635F\u5931\u91D1\u989D(\u4E07)", "=\u6848\u4F8B\u53F0\u8D26!E7"),
    ("\u603B\u8FFD\u56DE\u91D1\u989D(\u4E07)", "=\u6848\u4F8B\u53F0\u8D26!F7"),
    ("\u5E73\u5747\u8FFD\u56DE\u6BD4\u4F8B", "=\u6848\u4F8B\u53F0\u8D26!G7"),
    ("\u5E73\u5747\u7EF4\u6743\u5468\u671F(\u6708)", "=\u6848\u4F8B\u53F0\u8D26!L7"),
]

for i, (label, formula) in enumerate(stats_labels, start=4):
    ws2[f"A{i}"] = label
    ws2[f"B{i}"] = formula
    ws2[f"A{i}"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=11)
    ws2[f"B{i}"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=11, bold=True)
    if "\u6BD4\u4F8B" in label:
        ws2[f"B{i}"].number_format = "0.0%"
    elif "\u91D1\u989D" in label:
        ws2[f"B{i}"].number_format = "#,##0.0"

# 按纠纷类型统计
ws2["A11"] = "\u6309\u7EA0\u7EB0\u7C7B\u578B\u7EDF\u8BA1"
ws2["A11"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=12, bold=True)

ws2["A12"] = "\u7EA0\u7EB0\u7C7B\u578B"
ws2["B12"] = "\u6848\u4F8B\u6570"
ws2["C12"] = "\u603B\u635F\u5931(\u4E07)"
ws2["D12"] = "\u603B\u8FFD\u56DE(\u4E07)"
ws2["E12"] = "\u5E73\u5747\u8FFD\u56DE\u6BD4\u4F8B"

for col in ["A", "B", "C", "D", "E"]:
    ws2[f"{col}12"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=11, bold=True)
    ws2[f"{col}12"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws2[f"{col}12"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=11, bold=True, color="FFFFFF")

dispute_types = ["\u94F6\u884C\u7406\u8D22", "\u57FA\u91D1\u9500\u552E", "\u4FDD\u9669\u8BEF\u5BFC", "\u975E\u6CD5\u96C6\u8D44"]
for i, dtype in enumerate(dispute_types, start=13):
    ws2[f"A{i}"] = dtype
    ws2[f"B{i}"] = f'=COUNTIF(\u6848\u4F8B\u53F0\u8D26!B:B,"{dtype}")'
    ws2[f"C{i}"] = f'=SUMIF(\u6848\u4F8B\u53F0\u8D26!B:B,"{dtype}",\u6848\u4F8B\u53F0\u8D26!E:E)'
    ws2[f"D{i}"] = f'=SUMIF(\u6848\u4F8B\u53F0\u8D26!B:B,"{dtype}",\u6848\u4F8B\u53F0\u8D26!F:F)'
    ws2[f"E{i}"] = f'=IF(C{i}>0,D{i}/C{i},0)'
    ws2[f"E{i}"].number_format = "0.0%"

# 按结果类型统计
ws2["A18"] = "\u6309\u7EF4\u6743\u7ED3\u679C\u7EDF\u8BA1"
ws2["A18"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=12, bold=True)

ws2["A19"] = "\u7ED3\u679C\u7C7B\u578B"
ws2["B19"] = "\u6848\u4F8B\u6570"
ws2["C19"] = "\u5E73\u5747\u7EF4\u6743\u5468\u671F(\u6708)"

for col in ["A", "B", "C"]:
    ws2[f"{col}19"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=11, bold=True)
    ws2[f"{col}19"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws2[f"{col}19"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=11, bold=True, color="FFFFFF")

result_types = ["\u80DC\u8BC9", "\u548C\u89E3", "\u8D25\u8BC9", "\u672A\u7ED3"]
for i, rtype in enumerate(result_types, start=20):
    ws2[f"A{i}"] = rtype
    ws2[f"B{i}"] = f'=COUNTIF(\u6848\u4F8B\u53F0\u8D26!O:O,"{rtype}")'
    ws2[f"C{i}"] = f'=AVERAGEIF(\u6848\u4F8B\u53F0\u8D26!O:O,"{rtype}",\u6848\u4F8B\u53F0\u8D26!L:L)'
    ws2[f"C{i}"].number_format = "0.0"

# 设置列宽
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 16

# 创建图表
# 纠纷类型分布饼图
pie = PieChart()
pie.title = "\u7EA0\u7EB0\u7C7B\u578B\u5206\u5E03"
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True

cats = Reference(ws2, min_col=1, min_row=13, max_row=16)
data = Reference(ws2, min_col=2, min_row=12, max_row=16)
pie.add_data(data, titles_from_data=True)
pie.set_categories(cats)
pie.width = 12
pie.height = 10
ws2.add_chart(pie, "A26")

# 追回比例对比柱状图
bar = BarChart()
bar.title = "\u5404\u7C7B\u578B\u5E73\u5747\u8FFD\u56DE\u6BD4\u4F8B"
bar.x_axis.title = "\u7EA0\u7EB0\u7C7B\u578B"
bar.y_axis.title = "\u8FFD\u56DE\u6BD4\u4F8B"
bar.y_axis.numFmt = "0.0%"

cats = Reference(ws2, min_col=1, min_row=13, max_row=16)
data = Reference(ws2, min_col=5, min_row=12, max_row=16)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.width = 14
bar.height = 10
ws2.add_chart(bar, "J26")

# ==================== Sheet 3: 使用说明 ====================
ws3 = wb.create_sheet(title="\u4F7F\u7528\u8BF4\u660E")

instructions = [
    ("\u91D1\u878D\u7EA0\u7EB0\u6848\u4F8B\u53F0\u8D26\u4F7F\u7528\u8BF4\u660E", Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=16, bold=True)),
    ("", None),
    ("\u4E00\u3001\u529F\u80FD\u6982\u8FF0", Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=14, bold=True)),
    ("\u672C\u53F0\u8D26\u7528\u4E8E\u8BB0\u5F55\u548C\u7BA1\u7406\u91D1\u878D\u6D88\u8D39\u8005\u7EF4\u6743\u6848\u4F8B\uFF0C\u5E2E\u52A9\u8DDF\u8E2A\u4ECE\u8D2D\u4E70\u5230\u7ED3\u6848\u7684\u5168\u6D41\u7A0B\u4FE1\u606F\u3002", None),
    ("", None),
    ("\u4E8C\u3001\u5DE5\u4F5C\u8868\u8BF4\u660E", Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=12, bold=True)),
    ("1. \u6848\u4F8B\u53F0\u8D26\uFF1A\u8BB0\u5F55\u6240\u6709\u6848\u4F8B\u7684\u8BE6\u7EC6\u4FE1\u606F\uFF0C\u5305\u62EC\u7EA0\u7EB0\u7C7B\u578B\u3001\u91D1\u989D\u3001\u65E5\u671F\u3001\u7ED3\u679C\u7B49", None),
    ("2. \u7EDF\u8BA1\u5206\u6790\uFF1A\u81EA\u52A8\u6C47\u603B\u6570\u636E\uFF0C\u63D0\u4F9B\u53EF\u89C6\u5316\u7684\u7EDF\u8BA1\u56FE\u8868\u548C\u5173\u952E\u6307\u6807", None),
    ("3. \u4F7F\u7528\u8BF4\u660E\uFF1A\u672C\u8BF4\u660E\u6587\u6863", None),
    ("", None),
    ("\u4E09\u3001\u4F7F\u7528\u65B9\u6CD5", Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=12, bold=True)),
    ("1. \u65B0\u589E\u6848\u4F8B\uFF1A\u5728\"\u6848\u4F8B\u53F0\u8D26\"\u5DE5\u4F5C\u8868\u4E2D\uFF0C\u4EFF\u7167\u793A\u4F8B\u884C\u683C\u5F0F\uFF0C\u5728\u7A7A\u767D\u884C\u586B\u5165\u65B0\u6848\u4F8B\u4FE1\u606F", None),
    ("2. \u9009\u62E9\u7EA0\u7EB0\u7C7B\u578B\uFF1A\u70B9\u51FB\"\u7EA0\u7EB0\u7C7B\u578B\"\u5217\u5355\u5143\u683C\uFF0C\u4ECE\u4E0B\u62C9\u5217\u8868\u4E2D\u9009\u62E9\uFF08\u94F6\u884C\u7406\u8D22/\u57FA\u91D1\u9500\u552E/\u4FDD\u9669\u8BEF\u5BFC/\u975E\u6CD5\u96C6\u8D44\uFF09", None),
    ("3. \u9009\u62E9\u6295\u8BC9\u6E20\u9053\uFF1A\u70B9\u51FB\"\u6295\u8BC9\u6E20\u9053\"\u5217\u5355\u5143\u683C\uFF0C\u4ECE\u4E0B\u62C9\u5217\u8868\u4E2D\u9009\u62E9", None),
    ("4. \u9009\u62E9\u6700\u7EC8\u7ED3\u679C\uFF1A\u70B9\u51FB\"\u6700\u7EC8\u7ED3\u679C\"\u5217\u5355\u5143\u683C\uFF0C\u4ECE\u4E0B\u62C9\u5217\u8868\u4E2D\u9009\u62E9", None),
    ("5. \u8FFD\u56DE\u6BD4\u4F8B\uFF1A\u7CFB\u7EDF\u81EA\u52A8\u8BA1\u7B97\uFF0C\u65E0\u9700\u624B\u52A8\u8F93\u5165", None),
    ("6. \u7EF4\u6743\u8017\u65F6\u6708\uFF1A\u7CFB\u7EDF\u6839\u636E\u7EF4\u6743\u5F00\u59CB\u65E5\u671F\u548C\u7ED3\u6848\u65E5\u671F\u81EA\u52A8\u8BA1\u7B97", None),
    ("", None),
    ("\u56DB\u3001\u989C\u8272\u7F16\u7801\u8BF4\u660E", Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=12, bold=True)),
    ("\u8FFD\u56DE\u6BD4\u4F8B\u5217\u7684\u989C\u8272\u8868\u793A\u8FFD\u56DE\u6548\u679C\uFF1A", None),
    ("\u7EFF\u8272\uFF1A\u8FFD\u56DE\u6BD4\u4F8B > 50%\uFF0C\u6548\u679C\u826F\u597D", None),
    ("\u9EC4\u8272\uFF1A\u8FFD\u56DE\u6BD4\u4F8B 20%-50%\uFF0C\u6548\u679C\u4E00\u822C", None),
    ("\u7EA2\u8272\uFF1A\u8FFD\u56DE\u6BD4\u4F8B < 20%\uFF0C\u6548\u679C\u8F83\u5DEE", None),
    ("", None),
    ("\u4E94\u3001\u7EDF\u8BA1\u5206\u6790", Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=12, bold=True)),
    ("\u7EDF\u8BA1\u5206\u6790\u5DE5\u4F5C\u8868\u4F1A\u81EA\u52A8\u6C47\u603B\u4EE5\u4E0B\u4FE1\u606F\uFF1A", None),
    ("1. \u603B\u4F53\u6307\u6807\uFF1A\u6848\u4F8B\u603B\u6570\u3001\u603B\u91D1\u989D\u3001\u5E73\u5747\u8FFD\u56DE\u6BD4\u4F8B\u3001\u5E73\u5747\u7EF4\u6743\u5468\u671F", None),
    ("2. \u6309\u7EA0\u7EB0\u7C7B\u578B\u7EDF\u8BA1\uFF1A\u5404\u7C7B\u578B\u6848\u4F8B\u6570\u91CF\u3001\u635F\u5931\u91D1\u989D\u3001\u8FFD\u56DE\u91D1\u989D\u3001\u5E73\u5747\u8FFD\u56DE\u6BD4\u4F8B", None),
    ("3. \u6309\u7EF4\u6743\u7ED3\u679C\u7EDF\u8BA1\uFF1A\u5404\u7ED3\u679C\u7C7B\u578B\u6848\u4F8B\u6570\u91CF\u3001\u5E73\u5747\u7EF4\u6743\u5468\u671F", None),
    ("4. \u56FE\u8868\u5206\u6790\uFF1A\u7EA0\u7EB0\u7C7B\u578B\u5206\u5E03\u997C\u56FE\u3001\u5404\u7C7B\u578B\u8FFD\u56DE\u6BD4\u4F8B\u5BF9\u6BD4\u67F1\u72B6\u56FE", None),
    ("", None),
    ("\u516D\u3001\u6CE8\u610F\u4E8B\u9879", Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=12, bold=True)),
    ("1. \u65E5\u671F\u683C\u5F0F\uFF1A\u5EFA\u8BAE\u4F7F\u7528 YYYY-MM-DD \u683C\u5F0F\u8F93\u5165\u65E5\u671F", None),
    ("2. \u91D1\u989D\u5355\u4F4D\uFF1A\u91D1\u989D\u4EE5\"\u4E07\"\u4E3A\u5355\u4F4D\uFF0C\u8F93\u5165\u6570\u5B57\u5373\u53EF", None),
    ("3. \u672A\u7ED3\u6848\u6848\u4F8B\uFF1A\u7ED3\u6848\u65E5\u671F\u548C\u7EF4\u6743\u8017\u65F6\u6708\u5B57\u6BB5\u53EF\u7559\u7A7A", None),
    ("4. \u6570\u636E\u5907\u4EFD\uFF1A\u5B9A\u671F\u4FDD\u5B58\u6587\u4EF6\uFF0C\u5EFA\u8BAE\u5907\u4EFD\u91CD\u8981\u6570\u636E", None),
    ("", None),
    ("\u4E03\u3001\u8054\u7CFB\u65B9\u5F0F", Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=12, bold=True)),
    ("\u5982\u6709\u7591\u95EE\u6216\u5EFA\u8BAE\uFF0C\u8BF7\u8054\u7CFB\u6CD5\u52A1\u90E8\u95E8\u3002", None),
]

for i, (text, font) in enumerate(instructions, start=1):
    ws3[f"A{i}"] = text
    if font:
        ws3[f"A{i}"].font = font
    else:
        ws3[f"A{i}"].font = Font(name="\u5FAE\u8F6F\u96C5\u9ED1", size=11)
    ws3.merge_cells(f"A{i}:F{i}")

ws3.column_dimensions["A"].width = 80

# 保存文件
output_path = "D:/新课开发/法学/20-金融消费者维权：从理财爆雷到银行纠纷/全流程工具表单/F10_金融纠纷案例台账.xlsx"
wb.save(output_path)
print(f"Excel\u6587\u4EF6\u5DF2\u6210\u529F\u521B\u5EFA: {output_path}")

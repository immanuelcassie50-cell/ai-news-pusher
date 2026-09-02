# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = 'D:/新课开发/职业生涯和画布/VUCA时代，升级情绪力，做自己职场的主角/完整课程包/配套表单和指引-Excel版'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ===== Style helpers =====
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row_num, num_cols, bg='4472C4', fg='FFFFFF'):
    for c in range(1, num_cols+1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = Font(bold=True, color=fg, size=11)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()

def style_title_row(ws, row_num, text, num_cols, font_size=14, bg='2F5496', fg='FFFFFF'):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = Font(bold=True, size=font_size, color=fg)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_section(ws, row_num, text, num_cols):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = Font(bold=True, size=12, color='2F5496')
    cell.fill = PatternFill('solid', fgColor='D9E1F2')
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_data_row(ws, row_num, num_cols):
    for c in range(1, num_cols+1):
        cell = ws.cell(row=row_num, column=c)
        cell.border = thin_border()
        cell.alignment = Alignment(vertical='top', wrap_text=True)

def center_cell(ws, row, col):
    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

# ==============================================================
# FILE 1: 表单使用指引.xlsx
# ==============================================================
def create_guide():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '如何使用本指引'

    set_col_widths(ws, [18, 25, 35, 18, 18, 18])
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20

    style_title_row(ws, 1, 'VUCA时代，升级情绪力 —— 表单使用指引', 6)
    ws.cell(row=2, column=1).value = '课程模块：看清局 · 读懂情绪 · 主角归位 · 转化行动 | 开发者：罗宏伟'
    ws.cell(row=2, column=1).font = Font(size=10, italic=True)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    ws.cell(row=3, column=1).value = '本配套表单包含：表单使用指引 · 配套表单（填好版） · 配套表单（空表）'
    ws.cell(row=3, column=1).font = Font(size=11)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='center')

    ws.row_dimensions[5].height = 18
    style_header_row(ws, 5, 6)
    headers5 = ['标签页', '内容说明', '主要用途', '适用人员', '难度', '建议时间']
    for i, h in enumerate(headers5, 1):
        ws.cell(row=5, column=i).value = h

    guide_data = [
        ('如何使用本指引', '索引 + 使用说明', '了解整个指引的结构', '讲师 + 学员', '基础', '3分钟'),
        ('配套表单（填好版）', '全套表单的参考示例', '讲师展示，学员自学', '讲师 + 学员', '—', '—'),
        ('配套表单（空表）', '全套表单空白模板', '学员独立填写练习', '学员', '各模块不等', '课后练习'),
    ]
    for r, row_data in enumerate(guide_data, 6):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c).value = val
            ws.cell(row=r, column=c).border = thin_border()
            ws.cell(row=r, column=c).alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[r+1].height = 20

    ws.row_dimensions[9].height = 22
    style_section(ws, 9, '填写指南', 6)

    guide_steps = [
        ('第一步', '课程开始前', '通读本指引，了解表单结构'),
        ('第二步', '各模块练习', '使用配套表单_空表 完成课堂练习'),
        ('第三步', '对照参考', '完成后对照 配套表单_填好版 核对自己的思路'),
        ('第四步', '课后复盘', '使用复盘记录表 整理课程收获和行动计划'),
    ]
    ws.row_dimensions[10].height = 20
    style_header_row(ws, 10, 6, bg='8EA9DB')
    step_headers = ['步骤', '时机', '操作']
    for i, h in enumerate(step_headers, 1):
        ws.cell(row=10, column=i).value = h
    ws.merge_cells(start_row=10, start_column=3, end_row=10, end_column=6)
    for i, (s, t, d) in enumerate(guide_steps, 11):
        ws.cell(row=i, column=1).value = s
        ws.cell(row=i, column=2).value = t
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=6)
        ws.cell(row=i, column=3).value = d
        for c in range(1, 7):
            ws.cell(row=i, column=c).border = thin_border()
            ws.cell(row=i, column=c).alignment = Alignment(vertical='center')
        ws.row_dimensions[i].height = 20

    ws.row_dimensions[15].height = 22
    style_section(ws, 15, '情绪日志使用说明', 6)

    log_desc = [
        ('目的', '记录日常工作中的情绪波动，提升对情绪的觉察力和掌控感'),
        ('频率', '每次有明显情绪时填写，建议每天至少记录一次'),
        ('关键', '情绪强度评分 + 身体反应 + 应对方式 是核心三要素'),
        ('转化', '每次记录后，尝试写出：将情绪转化为下一步行动的一句话'),
    ]
    for i, (k, v) in enumerate(log_desc, 16):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=1).border = thin_border()
        ws.cell(row=i, column=1).alignment = Alignment(vertical='top')
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.cell(row=i, column=2).value = v
        ws.cell(row=i, column=2).border = thin_border()
        ws.cell(row=i, column=2).alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[i].height = 22

    ws.row_dimensions[20].height = 22
    style_section(ws, 20, '行动清单使用说明', 6)

    action_desc = [
        ('目的', '将情绪转化结果落地为具体可执行的行动'),
        ('目标', '每个情绪日志可以生成 1-3 条具体行动'),
        ('检视', '完成后在 复盘记录表 中回顾行动完成情况'),
        ('要点', '时间节点要明确，资源支持要具体，完成状态要打勾'),
    ]
    for i, (k, v) in enumerate(action_desc, 21):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=1).border = thin_border()
        ws.cell(row=i, column=1).alignment = Alignment(vertical='top')
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.cell(row=i, column=2).value = v
        ws.cell(row=i, column=2).border = thin_border()
        ws.cell(row=i, column=2).alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[i].height = 22

    ws.row_dimensions[25].height = 22
    style_section(ws, 25, '场景分析使用说明', 6)

    scene_desc = [
        ('时机', '遇到职场重大决策、人际冲突或情绪激动场景时使用'),
        ('VUCA四维度', 'V=易变性、U=不确定性、C=复杂性、A=模糊性'),
        ('主角归位四问', '①我想要什么？②我现在的情绪是什么？③我能做什么？④我的第一步是什么？'),
        ('输出', '形成一个具体的、可执行的转化方案'),
    ]
    for i, (k, v) in enumerate(scene_desc, 26):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=1).border = thin_border()
        ws.cell(row=i, column=1).alignment = Alignment(vertical='top')
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.cell(row=i, column=2).value = v
        ws.cell(row=i, column=2).border = thin_border()
        ws.cell(row=i, column=2).alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[i].height = 22

    ws.row_dimensions[30].height = 22
    style_section(ws, 30, '常见问题 FAQ', 6)

    faq_data = [
        ('Q1: 情绪日志要写多长？', 'A: 简洁为主，核心是触发事件 + 情绪强度 + 身体反应，3-5句话即可'),
        ('Q2: 情绪强度10分是最高吗？', 'A: 是，10分为情绪完全失控状态，建议超过7分时先停下来做深呼吸再填写'),
        ('Q3: 行动清单可以一次写多条吗？', 'A: 可以，建议每次写1-3条，聚焦最重要的行动，写太多反而难以执行'),
        ('Q4: 场景分析在哪里用？', 'A: 遇到职场冲突、重大决策或情绪低落时随时用，尤其适合月度复盘'),
        ('Q5: 填好版和空表有什么区别？', 'A: 填好版是参考示例，帮助理解思路；空表是练习本，需独立完成'),
        ('Q6: 复盘记录多久写一次？', 'A: 建议每周一次，配合行动清单的检视，形成PDCA循环'),
    ]
    ws.row_dimensions[31].height = 20
    style_header_row(ws, 31, 6, bg='8EA9DB')
    ws.cell(row=31, column=1).value = '问题'
    ws.cell(row=31, column=1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=31, start_column=2, end_row=31, end_column=6)
    ws.cell(row=31, column=2).value = '答案'
    ws.cell(row=31, column=2).alignment = Alignment(horizontal='center')
    for i, (q, a) in enumerate(faq_data, 32):
        ws.cell(row=i, column=1).value = q
        ws.cell(row=i, column=1).font = Font(bold=True, size=10)
        ws.cell(row=i, column=1).border = thin_border()
        ws.cell(row=i, column=1).alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.cell(row=i, column=2).value = a
        ws.cell(row=i, column=2).border = thin_border()
        ws.cell(row=i, column=2).alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[i].height = 26

    # Sheet 2: 情绪力四模块概览
    ws2 = wb.create_sheet('四模块概览')
    set_col_widths(ws2, [20, 30, 30, 20])
    ws2.row_dimensions[1].height = 35
    style_title_row(ws2, 1, 'VUCA时代，升级情绪力 —— 四模块概览', 4)

    modules = [
        ('模块一：看清局', '易变性（V）· 不确定性（U）· 复杂性（C）· 模糊性（A）',
         '识别VUCA职场环境，理解为什么传统的稳定思维不再适用，建立对外部环境的敏锐感知'),
        ('模块二：读懂情绪', '情绪识别 · 情绪命名 · 情绪强度评估 · 身体反应觉察',
         '在情绪升起时能够识别它、命名它、评估强度，并觉察身体信号——这是转化的前提'),
        ('模块三：主角归位', '主角归位四问 · 情绪的主人 · 不是受害者的思维练习',
         '通过四问将情绪从被动反应变成主动选择，重新拿回对情绪的掌控权'),
        ('模块四：转化行动', '从情绪到行动 · 行动清单 · 复盘记录 · 持续改进',
         '将情绪能量转化为具体可执行的行动，形成持续改进的闭环'),
    ]
    for i, (title, subtitle, desc) in enumerate(modules, 3):
        ws2.row_dimensions[i].height = 50
        ws2.cell(row=i, column=1).value = title
        ws2.cell(row=i, column=1).font = Font(bold=True, size=12, color='FFFFFF')
        ws2.cell(row=i, column=1).fill = PatternFill('solid', fgColor='2F5496')
        ws2.cell(row=i, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=i, column=1).border = thin_border()

        ws2.cell(row=i, column=2).value = subtitle
        ws2.cell(row=i, column=2).font = Font(bold=True, size=11, color='2F5496')
        ws2.cell(row=i, column=2).fill = PatternFill('solid', fgColor='D9E1F2')
        ws2.cell(row=i, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws2.cell(row=i, column=2).border = thin_border()

        ws2.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
        ws2.cell(row=i, column=3).value = desc
        ws2.cell(row=i, column=3).alignment = Alignment(vertical='center', wrap_text=True)
        ws2.cell(row=i, column=3).border = thin_border()

    # Sheet 3: 表单清单
    ws3 = wb.create_sheet('表单清单')
    set_col_widths(ws3, [25, 20, 15, 15, 30])
    ws3.row_dimensions[1].height = 35
    style_title_row(ws3, 1, 'VUCA情绪力课程 —— 配套表单清单', 5)

    ws3.row_dimensions[3].height = 20
    style_header_row(ws3, 3, 5)
    list_headers = ['表单名称', '表单类型', '难度', '建议时机', '主要内容']
    for i, h in enumerate(list_headers, 1):
        ws3.cell(row=3, column=i).value = h

    forms = [
        ('情绪日志模板', '填好示例', '基础', '日常随时', '触发事件+情绪类型+强度+身体反应+应对+转化'),
        ('行动清单模板', '填好示例', '基础', '情绪记录后', '目标+行动步骤+时间节点+资源+检视方式'),
        ('场景分析模板', '填好示例', '进阶', '重大场景', 'VUCA四维度+情绪反应+四问应用+转化方案'),
        ('情绪日志_空表', '空白模板', '基础', '课后练习', '同上，空表格式方便独立填写'),
        ('行动清单_空表', '空白模板', '基础', '课后练习', '同上，空白版'),
        ('场景分析_空表', '空白模板', '进阶', '课后练习', '同上，空白版'),
        ('复盘记录_空表', '空白模板', '进阶', '周/月复盘', '收获总结+关键突破+改进点+行动计划'),
    ]
    for i, row_data in enumerate(forms, 4):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=i, column=c).value = val
            ws3.cell(row=i, column=c).border = thin_border()
            ws3.cell(row=i, column=c).alignment = Alignment(vertical='center', wrap_text=True)
        ws3.row_dimensions[i].height = 28

    wb.save(os.path.join(OUTPUT_DIR, '表单使用指引.xlsx'))
    print('Created: 表单使用指引.xlsx')

# ==============================================================
# FILE 2: 配套表单_填好版.xlsx
# ==============================================================
def create_filled():
    wb = openpyxl.Workbook()

    # ===== Sheet: 情绪日志模板（填好示例）=====
    ws1 = wb.active
    ws1.title = '情绪日志模板（填好示例）'
    set_col_widths(ws1, [18, 30, 20, 15, 22, 25, 30])
    ws1.row_dimensions[1].height = 35
    style_title_row(ws1, 1, '情绪日志 —— 参考示例（填好版）', 7, font_size=14)

    ws1.row_dimensions[2].height = 20
    ws1.cell(row=2, column=1).value = '课程名称：VUCA时代，升级情绪力，做自己职场的主角 | 开发者：罗宏伟'
    ws1.cell(row=2, column=1).font = Font(size=10, italic=True)
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws1.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    ws1.row_dimensions[4].height = 25
    style_header_row(ws1, 4, 7)
    headers = ['日期时间', '触发事件', '情绪类型', '情绪强度\n（1-10）', '身体反应', '应对方式', '转化结果']
    for i, h in enumerate(headers, 1):
        ws1.cell(row=4, column=i).value = h

    log_examples = [
        ('周一 09:15', '部门例会上，老板当众质疑我的方案，说"没有数据支撑"，但我认为数据在附录里他没看到',
         '愤怒', 7,
         '胸口发紧，脸发热，想反驳但强压着',
         '先深呼吸，在笔记本上写下"冷静，先听完"，强迫自己不在当场反驳',
         '会后主动找老板，发送附录数据，并简短说明"想确认您关心的是数据这块，感谢提示"。老板回复"我看一下"，没有再追问。情绪强度从7降到3。'),
        ('周三 14:30', '同事小王私下跟其他同事说我"工作态度有问题"，消息传到我耳朵里',
         '沮丧', 6,
         '胃部不适，食欲下降，下午一直在想这件事',
         '先出去走了10分钟，然后找了一个信任的同事聊了聊，发现小王可能因为上周项目赶工的事对我有误解',
         '主动找小王当面聊，问"是不是我哪里做得不好让你有这个印象"，小王承认是误会。情绪转化为："下次被误解时，主动约聊而不超过24小时"——写进行动清单。情绪强度从6降到2。'),
        ('周五 11:00', '收到HR通知，说我的晋升评审没通过，理由是"影响力不足"',
         '恐惧', 8,
         '心跳加速，手有点发抖，脑子里反复回放各种可能性',
         '深呼吸三次，离开工位去茶水间，用4-7-8呼吸法让自己平静，然后写下自己的情绪和想法',
         '情绪平稳后，分析"影响力不足"可能指的是跨部门协作。写邮件给老板请求具体反馈，并预约下周1:1。转化：把恐惧转化为"下个月主动参与跨部门项目"的具体行动。情绪强度从8降到4。'),
        ('周四 16:45', '给客户做汇报，演讲中忘了一段内容，停顿了大约10秒',
         '焦虑', 5,
         '手心出汗，声音有点抖，之后一直回想那10秒',
         '演讲结束后没有立刻离场，而是问客户"您有哪些地方想深入了解"，把话题引到互动上',
         '客户反而反馈"你讲得很清晰"。认识到：停顿10秒在客户看来可能没那么严重。情绪转化：把焦虑变成"每次汇报前做3分钟默剧练习"的新习惯。情绪强度从5降到2。'),
    ]
    for i, row_data in enumerate(log_examples, 5):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=i, column=c).value = val
            ws1.cell(row=i, column=c).border = thin_border()
            ws1.cell(row=i, column=c).alignment = Alignment(vertical='top', wrap_text=True)
        ws1.row_dimensions[i].height = 80

    ws1.row_dimensions[10].height = 22
    style_section(ws1, 10, '填好版说明', 7)
    ws1.merge_cells(start_row=11, start_column=1, end_row=11, end_column=7)
    ws1.cell(row=11, column=1).value = '以上为虚构参考案例（小罗的经历），展示了不同情绪类型（愤怒/沮丧/恐惧/焦虑）的记录方式。情绪强度8-10分建议当下先处理情绪，暂不填写；等情绪平稳后再补写。'
    ws1.cell(row=11, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws1.cell(row=11, column=1).font = Font(size=10, italic=True)
    ws1.row_dimensions[11].height = 35

    # ===== Sheet: 行动清单模板（填好示例）=====
    ws2 = wb.create_sheet('行动清单模板（填好示例）')
    set_col_widths(ws2, [5, 22, 30, 20, 20, 22, 18, 15])
    ws2.row_dimensions[1].height = 35
    style_title_row(ws2, 1, '行动清单 —— 参考示例（填好版）', 8, font_size=14)

    ws2.row_dimensions[2].height = 20
    ws2.cell(row=2, column=1).value = '从情绪转化到落地行动的工具 | 开发者：罗宏伟'
    ws2.cell(row=2, column=1).font = Font(size=10, italic=True)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws2.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    ws2.row_dimensions[4].height = 30
    style_header_row(ws2, 4, 8)
    action_headers = ['#', '目标设定', '具体行动步骤', '时间节点', '资源支持', '检视方式', '完成状态', '备注']
    for i, h in enumerate(action_headers, 1):
        ws2.cell(row=4, column=i).value = h

    action_examples = [
        (1, '提升跨部门影响力，避免晋升评审再次因"影响力不足"被拒',
         '①梳理本月需要跨部门协作的3个节点\n②主动联系相关同事了解他们的需求\n③每周四下午主动发送本周进展更新',
         '本月内完成第一轮沟通，下月1日前汇报结果给老板',
         '老板的1:1时间，邮件签名档更新',
         '老板1:1时确认"影响力"具体指什么，收集反馈',
         '已完成 80%', '4/5项已完成，还差1个部门没约到'),
        (2, '改善被当众质疑时的情绪反应，不在当场反驳',
         '①记录每次想反驳的场景\n②用"先回应情绪，再回应事实"的公式\n③准备一句过渡语："这点我回去再核实一下"',
         '本周起每天复盘一次，3周后回顾',
         '笔记本，4-7-8呼吸法练习音频',
         '每周五用情绪日志复盘当周场景',
         '进行中', '第一周记录了3次，第二周降到1次'),
        (3, '减少因误解导致的情绪消耗（沮丧）',
         '①被误解时24小时内主动约聊\n②约聊时先问"是不是我哪里做得不好"\n③不要等对方主动道歉',
         '本月内遇到2次误解都执行',
         '勇气和主动性（心理资源）',
         '情绪日志中记录"转化行动是否执行"',
         '已完成', '两次误解都主动约聊，结果都是误会'),
        (4, '汇报前做好充分准备，减少汇报中的焦虑',
         '①每次汇报前做3分钟默剧练习\n②准备"如果忘词就说：请允许我翻到下一页展开"\n③提前到会议室适应环境',
         '每次汇报前都执行',
         '镜子或手机录音，汇报文档V2版本',
         '汇报后记录：停顿了几秒？感受如何？',
         '已形成习惯', '连续3次汇报没有明显停顿'),
    ]
    for i, row_data in enumerate(action_examples, 5):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=i, column=c).value = val
            ws2.cell(row=i, column=c).border = thin_border()
            ws2.cell(row=i, column=c).alignment = Alignment(vertical='top', wrap_text=True)
            if c == 1:
                ws2.cell(row=i, column=c).alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[i].height = 90

    ws2.row_dimensions[10].height = 22
    style_section(ws2, 10, '行动清单使用要点', 8)
    ws2.merge_cells(start_row=11, start_column=1, end_row=11, end_column=8)
    ws2.cell(row=11, column=1).value = ('好的行动清单：①目标具体可衡量（不是"提升影响力"而是"本月内完成3个部门的第一轮沟通"）'
                                         '②时间节点明确（不是"尽快"而是"本周五前"）'
                                         '③有检视方式（不是"做好"而是"每周五复盘情绪日志"）')
    ws2.cell(row=11, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws2.cell(row=11, column=1).font = Font(size=10, italic=True)
    ws2.row_dimensions[11].height = 45

    # ===== Sheet: 场景分析模板（填好示例）=====
    ws3 = wb.create_sheet('场景分析模板（填好示例）')
    set_col_widths(ws3, [20, 35, 20, 35, 35])
    ws3.row_dimensions[1].height = 35
    style_title_row(ws3, 1, '场景分析 —— 参考示例（填好版）', 5, font_size=14)

    ws3.row_dimensions[2].height = 20
    ws3.cell(row=2, column=1).value = '用VUCA四维度 + 主角归位四问，分析重大职场场景 | 开发者：罗宏伟'
    ws3.cell(row=2, column=1).font = Font(size=10, italic=True)
    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws3.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    scene_examples = [
        {
            '场景': '老板临时加了新任务，要求周五前完成，但我的工作量已经饱和',
            'vuca': 'V（易变）：任务内容随时可能再调整\nU（不确定）：不知道老板是否会继续加任务\nC（复杂）：其他任务和新任务互相依赖，时间紧张\nA（模糊）：老板的真实期待不清楚——是要质量还是要速度？',
            '情绪': '愤怒（6分）：觉得不被尊重，自己的工作量不被看见\n焦虑（5分）：担心完不成影响考核',
            '四问': '①我想要什么？→希望老板认可我的工作量，或者明确优先级\n②我现在的情绪是什么？→愤怒+焦虑，背后的诉求是被尊重和被看见\n③我能做什么？→主动约老板10分钟，说明当前工作量，请求明确优先级\n④我的第一步是什么？→今天下午发一条消息给老板："想约您10分钟，确认一下本周的优先级"',
            '转化': '给老板发了一条消息约10分钟面谈。面谈时没有抱怨，而是说"我手上有A、B、C三个项目，想请您帮我确认优先级"。老板重新排了序，取消了C任务。情绪强度从6降到2。',
        },
        {
            '场景': '年终评审没通过，同事私下议论说我"能力不行"',
            'vuca': 'V（易变）：评审结果可以申诉，也可以接受\nU（不确定）：不知道下次评审是什么时候，也不确定能不能通过\nC（复杂）：同事的议论和评审结果互相强化，形成负面氛围\nA（模糊）："能力不行"这个评价来源不清——是评审标准的问题还是真实问题？',
            '情绪': '沮丧（7分）：觉得自己不被认可，努力没有被看见\n愤怒（4分）：对同事议论感到不满，但不想当面对质',
            '四问': '①我想要什么？→想要一个公正的评价，想要知道真实差距在哪里\n②我现在的情绪是什么？→沮丧为主，愤怒为辅；沮丧背后是渴望被认可\n③我能做什么？→主动找老板要具体反馈，列出自己认为做得好的地方请老板指正\n④我的第一步是什么？→明天约老板1:1，开场白："我想了解我在哪些具体地方有差距"',
            '转化': '约了老板1:1，老板指出两个具体问题（跨部门协作、项目汇报结构化）。情绪从沮丧转化为具体的改进行动：每月主动给老板发一页纸的进展汇报。情绪强度从7降到3。',
        },
    ]

    row = 4
    for idx, ex in enumerate(scene_examples):
        ws3.row_dimensions[row].height = 25
        style_title_row(ws3, row, f'场景 {idx+1}', 5, bg='2F5496', fg='FFFFFF')
        row += 1

        ws3.row_dimensions[row].height = 22
        style_section(ws3, row, '场景描述', 5)
        row += 1
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws3.cell(row=row, column=1).value = ex['场景']
        ws3.cell(row=row, column=1).border = thin_border()
        ws3.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws3.row_dimensions[row].height = 30
        row += 1

        ws3.row_dimensions[row].height = 22
        style_section(ws3, row, 'VUCA四维度分析', 5)
        row += 1
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws3.cell(row=row, column=1).value = ex['vuca']
        ws3.cell(row=row, column=1).border = thin_border()
        ws3.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws3.row_dimensions[row].height = 80
        row += 1

        ws3.row_dimensions[row].height = 22
        style_section(ws3, row, '情绪反应', 5)
        row += 1
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws3.cell(row=row, column=1).value = ex['情绪']
        ws3.cell(row=row, column=1).border = thin_border()
        ws3.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws3.row_dimensions[row].height = 50
        row += 1

        ws3.row_dimensions[row].height = 22
        style_section(ws3, row, '主角归位四问应用', 5)
        row += 1
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws3.cell(row=row, column=1).value = ex['四问']
        ws3.cell(row=row, column=1).border = thin_border()
        ws3.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws3.row_dimensions[row].height = 90
        row += 1

        ws3.row_dimensions[row].height = 22
        style_section(ws3, row, '转化方案', 5)
        row += 1
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws3.cell(row=row, column=1).value = ex['转化']
        ws3.cell(row=row, column=1).border = thin_border()
        ws3.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws3.row_dimensions[row].height = 60
        row += 1

        row += 1

    wb.save(os.path.join(OUTPUT_DIR, '配套表单_填好版.xlsx'))
    print('Created: 配套表单_填好版.xlsx')

# ==============================================================
# FILE 3: 配套表单_空表.xlsx
# ==============================================================
def create_empty():
    wb = openpyxl.Workbook()

    # ===== 情绪日志_空表 =====
    ws1 = wb.active
    ws1.title = '情绪日志_空表'
    set_col_widths(ws1, [18, 30, 20, 15, 22, 25, 30])
    ws1.row_dimensions[1].height = 35
    style_title_row(ws1, 1, '情绪日志 —— 空白模板', 7)

    ws1.row_dimensions[2].height = 20
    ws1.cell(row=2, column=1).value = '课程名称：VUCA时代，升级情绪力，做自己职场的主角 | 开发者：罗宏伟'
    ws1.cell(row=2, column=1).font = Font(size=10, italic=True)
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws1.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    ws1.row_dimensions[3].height = 18
    ws1.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
    ws1.cell(row=3, column=1).value = '使用说明：每次有明显情绪时填写，建议每天至少1次。情绪强度超过7分时，先做深呼吸，等平静后再写。'
    ws1.cell(row=3, column=1).font = Font(size=10, italic=True)
    ws1.cell(row=3, column=1).alignment = Alignment(wrap_text=True, horizontal='center')

    ws1.row_dimensions[5].height = 30
    style_header_row(ws1, 5, 7)
    headers = ['日期时间', '触发事件', '情绪类型\n（焦虑/愤怒/沮丧/恐惧/其他）', '情绪强度\n（1-10）', '身体反应', '应对方式', '转化结果\n（写下一步行动）']
    for i, h in enumerate(headers, 1):
        ws1.cell(row=5, column=i).value = h

    emotion_types = ['焦虑', '愤怒', '沮丧', '恐惧', '其他：___']

    for r in range(6, 26):
        ws1.row_dimensions[r].height = 50
        for c in range(1, 8):
            cell = ws1.cell(row=r, column=c)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if c == 3:
                cell.value = '请选择情绪类型'
                cell.font = Font(size=9, italic=True, color='888888')
            elif c == 4:
                cell.value = '请打分'
                cell.font = Font(size=9, italic=True, color='888888')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif c == 1:
                cell.value = f'第 {r-5} 条'
                cell.font = Font(size=9, color='888888')
                cell.alignment = Alignment(horizontal='center', vertical='top')

    # ===== 行动清单_空表 =====
    ws2 = wb.create_sheet('行动清单_空表')
    set_col_widths(ws2, [5, 22, 32, 20, 20, 22, 15, 18])
    ws2.row_dimensions[1].height = 35
    style_title_row(ws2, 1, '行动清单 —— 空白模板', 8)

    ws2.row_dimensions[2].height = 20
    ws2.cell(row=2, column=1).value = '从情绪转化到落地行动的工具 | 开发者：罗宏伟'
    ws2.cell(row=2, column=1).font = Font(size=10, italic=True)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws2.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    ws2.row_dimensions[3].height = 18
    ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)
    ws2.cell(row=3, column=1).value = '使用说明：每个情绪日志生成1-3条行动，目标要具体可衡量，时间节点要明确，每周检视一次完成情况。'
    ws2.cell(row=3, column=1).font = Font(size=10, italic=True)
    ws2.cell(row=3, column=1).alignment = Alignment(wrap_text=True, horizontal='center')

    ws2.row_dimensions[5].height = 30
    style_header_row(ws2, 5, 8)
    action_headers = ['#', '目标设定\n（具体可衡量）', '具体行动步骤\n（分条写）', '时间节点\n（明确日期）', '资源支持\n（人/物/信息）', '检视方式\n（怎么确认完成）', '完成状态\n（✅/进行中/未开始）', '备注']
    for i, h in enumerate(action_headers, 1):
        ws2.cell(row=5, column=i).value = h

    for r in range(6, 21):
        ws2.row_dimensions[r].height = 60
        for c in range(1, 9):
            cell = ws2.cell(row=r, column=c)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if c == 1:
                cell.value = r - 5
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif c == 7:
                cell.value = '请选择状态'
                cell.font = Font(size=9, italic=True, color='888888')

    # ===== 场景分析_空表 =====
    ws3 = wb.create_sheet('场景分析_空表')
    set_col_widths(ws3, [20, 35])
    ws3.row_dimensions[1].height = 35
    style_title_row(ws3, 1, '场景分析 —— 空白模板', 2)

    ws3.row_dimensions[2].height = 20
    ws3.cell(row=2, column=1).value = '用VUCA四维度 + 主角归位四问，分析重大职场场景 | 开发者：罗宏伟'
    ws3.cell(row=2, column=1).font = Font(size=10, italic=True)
    ws3.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

    sections = [
        ('场景描述', 80),
        ('VUCA四维度分析\nV（易变）：这件事哪些地方在变化？\nU（不确定）：有哪些未知信息？\nC（复杂）：涉及哪些相互关联的因素？\nA（模糊）：哪些信息可以被不同方式解读？', 120),
        ('情绪反应\n记录你在这个场景中的各种情绪，以及每种情绪的强度（1-10分）', 80),
        ('主角归位四问\n①我想要什么？（具体目标）\n②我现在的情绪是什么？（命名情绪+强度）\n③我能做什么？（列出3个可能性）\n④我的第一步是什么？（今天可以做的最小行动）', 130),
        ('转化方案\n结合以上分析，写出一个具体的、可执行的行动方案', 100),
    ]

    row = 4
    for section_name, height in sections:
        ws3.row_dimensions[row].height = 25
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws3.cell(row=row, column=1).value = section_name.split('\n')[0]
        ws3.cell(row=row, column=1).font = Font(bold=True, size=12, color='2F5496')
        ws3.cell(row=row, column=1).fill = PatternFill('solid', fgColor='D9E1F2')
        ws3.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws3.cell(row=row, column=1).border = thin_border()
        row += 1

        if '\n' in section_name:
            hint_text = '\n'.join(section_name.split('\n')[1:])
            ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws3.cell(row=row, column=1).value = hint_text
            ws3.cell(row=row, column=1).font = Font(size=9, italic=True, color='888888')
            ws3.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws3.cell(row=row, column=1).border = thin_border()
        ws3.row_dimensions[row].height = height
        row += 1
        row += 1

    # ===== 复盘记录_空表 =====
    ws4 = wb.create_sheet('复盘记录_空表')
    set_col_widths(ws4, [22, 40])
    ws4.row_dimensions[1].height = 35
    style_title_row(ws4, 1, '复盘记录 —— 空白模板', 2)

    ws4.row_dimensions[2].height = 20
    ws4.cell(row=2, column=1).value = '课程：VUCA时代，升级情绪力，做自己职场的主角 | 开发者：罗宏伟'
    ws4.cell(row=2, column=1).font = Font(size=10, italic=True)
    ws4.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws4.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    ws4.row_dimensions[3].height = 18
    ws4.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws4.cell(row=3, column=1).value = '使用说明：建议每周一次，配合行动清单检视，形成PDCA循环。聚焦：这次做得好的是什么？下次改进什么？'
    ws4.cell(row=3, column=1).font = Font(size=10, italic=True)
    ws4.cell(row=3, column=1).alignment = Alignment(wrap_text=True)

    reflection_sections = [
        ('基本信息\n复盘日期：\n复盘周期：本周/本月/专项', 50),
        ('课程收获总结\n通过这次课程，我最大的3个收获是：\n1.\n2.\n3.', 100),
        ('关键突破点\n在这段时间里，我在情绪管理上最大的突破是：\n\n突破的关键是什么？', 100),
        ('持续改进点\n目前仍存在的情绪困扰或行为模式：\n\n我想在下个周期重点改进的是：', 100),
        ('下一步行动计划\n具体要做什么？\n时间节点是什么？\n需要什么资源或支持？\n如何检视完成情况？', 120),
        ('给他人的一个建议\n如果要把这门课程的收获分享给同事，我会说：', 80),
    ]

    row = 5
    for section_name, height in reflection_sections:
        ws4.row_dimensions[row].height = 25
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws4.cell(row=row, column=1).value = section_name.split('\n')[0]
        ws4.cell(row=row, column=1).font = Font(bold=True, size=12, color='FFFFFF')
        ws4.cell(row=row, column=1).fill = PatternFill('solid', fgColor='2F5496')
        ws4.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws4.cell(row=row, column=1).border = thin_border()
        row += 1

        if '\n' in section_name:
            hint_lines = section_name.split('\n')[1:]
            if hint_lines:
                hint_text = '\n'.join(hint_lines)
                ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                ws4.cell(row=row, column=1).value = hint_text
                ws4.cell(row=row, column=1).font = Font(size=9, italic=True, color='888888')
                ws4.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws4.cell(row=row, column=1).border = thin_border()
        ws4.row_dimensions[row].height = height
        row += 1
        row += 1

    wb.save(os.path.join(OUTPUT_DIR, '配套表单_空表.xlsx'))
    print('Created: 配套表单_空表.xlsx')

# Run all three
create_guide()
create_filled()
create_empty()
print('\nAll 3 files created successfully!')
print('Output path:', OUTPUT_DIR)

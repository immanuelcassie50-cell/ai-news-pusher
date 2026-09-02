#!/usr/bin/env python3
"""
私董会运作方法论 - 系列工具表
生成专业Excel工作簿，包含8个工具表单Sheet
"""

import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

# ========== 样式定义 ==========
# 标题样式
TITLE_FONT = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
SUBTITLE_FONT = Font(name='微软雅黑', size=14, bold=True, color='2F5496')
HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
LABEL_FONT = Font(name='微软雅黑', size=10, bold=True, color='2F5496')
CONTENT_FONT = Font(name='微软雅黑', size=10)
SMALL_FONT = Font(name='微软雅黑', size=9)

# 颜色定义
DARK_BLUE_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
MEDIUM_BLUE_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
ACCENT_BLUE_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
LIGHT_GRAY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

# 边框样式
THIN_BORDER = Border(
    left=Side(style='thin', color='B8CCE4'),
    right=Side(style='thin', color='B8CCE4'),
    top=Side(style='thin', color='B8CCE4'),
    bottom=Side(style='thin', color='B8CCE4')
)
MEDIUM_BORDER = Border(
    left=Side(style='medium', color='2F5496'),
    right=Side(style='medium', color='2F5496'),
    top=Side(style='medium', color='2F5496'),
    bottom=Side(style='medium', color='2F5496')
)

# 对齐方式
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True)


def set_column_widths(ws, widths):
    """设置列宽"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def apply_header_style(cell):
    """应用表头样式"""
    cell.font = HEADER_FONT
    cell.fill = DARK_BLUE_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER


def apply_cell_style(cell, fill=None, font=None, alignment=None, border=None):
    """应用单元格通用样式"""
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def create_title_row(ws, row, title, subtitle, num_cols, start_row=1):
    """创建标题行"""
    # 主标题行
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    title_cell = ws.cell(row=row, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.fill = DARK_BLUE_FILL
    title_cell.alignment = CENTER_ALIGN
    title_cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 35

    # 副标题行
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=num_cols)
    subtitle_cell = ws.cell(row=row+1, column=1, value=subtitle)
    subtitle_cell.font = Font(name='微软雅黑', size=11, italic=True, color='2F5496')
    subtitle_cell.fill = ACCENT_BLUE_FILL
    subtitle_cell.alignment = CENTER_ALIGN
    subtitle_cell.border = THIN_BORDER
    ws.row_dimensions[row+1].height = 25

    return row + 2  # 返回数据开始行


def add_copyright(ws, num_cols):
    """添加版权声明"""
    last_row = ws.max_row + 2
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=num_cols)
    copyright_cell = ws.cell(row=last_row, column=1,
        value="内容来源：《私董会运作方法论©》研发团队第V1.0版，仅供项目现场练习时使用，未经授权严禁对外传播和使用")
    copyright_cell.font = Font(name='微软雅黑', size=9, italic=True, color='808080')
    copyright_cell.alignment = CENTER_ALIGN


# ========== Sheet 1: 七步法流程卡 ==========
def create_sheet1_seven_steps(wb):
    ws = wb.active
    ws.title = "01 七步法流程卡"

    # 设置列宽
    set_column_widths(ws, {
        'A': 8,   # 步骤编号
        'B': 15,  # 步骤名称
        'C': 12,  # 时间分配
        'D': 35,  # 关键动作
        'E': 45,  # 话术要点
        'F': 40,  # 注意事项
    })

    # 创建标题
    data_start = create_title_row(ws, 1,
        "系列工具表",
        "私董会运作方法论-七步法流程卡",
        6)

    # 表头
    headers = ['步骤编号', '步骤名称', '时间分配', '关键动作', '话术要点', '注意事项']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=header)
        apply_header_style(cell)

    # 七步法数据
    seven_steps = [
        {
            'num': '第一步',
            'name': '躬问',
            'time': '15分钟',
            'actions': '''1. 主持人宣布进入"躬问"环节
2. 邀请案主重新陈述核心困惑
3. 确认所有成员是否清晰理解问题''',
            'script': '''"请案主再次向大家陈述，您目前最困扰、最希望解决的核心问题是什么？"
"在案主陈述过程中，请各位静静倾听，不要打断，思考您想要提问的方向。"''',
            'notes': '''- 确保问题陈述清晰、具体
- 避免问题过于模糊或开放
- 记录案主强调的关键词'''
        },
        {
            'num': '第二步',
            'name': '躬思',
            'time': '10分钟',
            'actions': '''1. 所有成员独立思考
2. 在内心形成初步问题清单
3. 记录自己最想提问的3个方向''',
            'script': '''"请各位利用接下来的10分钟，独立思考您最想向案主提问的3个问题。"
"可以写在心电图上，也可以默默在心中酝酿。"''',
            'notes': '''- 保持安静，不交流
- 思考问题的深度而非广度
- 关注"为什么"而非"做什么"'''
        },
        {
            'num': '第三步',
            'name': '躬答',
            'time': '10分钟',
            'actions': '''1. 成员依次提问（每次1-2个）
2. 案主现场即时回答
3. 主持人引导深入追问''',
            'script': '''"现在进入躬答环节，请依次分享您的问题。问题的形式是'您曾经……/您是否……/您觉得……'
"案主，请直接回答，不要解释为什么或试图说服。"''',
            'notes': '''- 提问以"您"开头
- 一次只问一个问题
- 不带建议、不给方案'''
        },
        {
            'num': '第四步',
            'name': '轮流发言',
            'time': '20分钟',
            'actions': '''1. 每人2-3分钟分享洞察
2. 从案主左手边开始顺时针
3. 聚焦建设性建议而非评论''',
            'script': '''"接下来请每位成员分享您对案主问题的洞察或建议，每人2-3分钟。"
"请聚焦于您看到了什么、感受到了什么，而不是评判案主的选择。"''',
            'notes': '''- 不重复已经说过的观点
- 真诚、直接、有建设性
- 避免"但是""不过"等转折'''
        },
        {
            'num': '第五步',
            'name': '专家分享',
            'time': '10分钟',
            'actions': '''1. 行业专家或外部顾问分享
2. 提供专业视角和行业洞察
3. 分享类似问题的最佳实践''',
            'script': '''"有请[专家姓名]从专业角度为我们分享他的洞察和建议。"
"专家的分享可以帮助我们打开思路，看到更多的可能性。"''',
            'notes': '''- 提前与专家沟通背景
- 鼓励跨行业视角
- 记录关键洞察'''
        },
        {
            'num': '第六步',
            'name': '现场观测',
            'time': '10分钟',
            'actions': '''1. 主持人观察全场动态
2. 记录关键洞察和洞见
3. 捕捉非语言信息''',
            'script': '''"请各位稍安勿躁，给我1-2分钟整理今天的讨论要点。"
（主持人在白板上快速记录关键洞察）''',
            'notes': '''- 关注沉默者的反应
- 记录重复出现的观点
- 观察能量变化'''
        },
        {
            'num': '第七步',
            'name': '三赏闭环',
            'time': '10分钟',
            'actions': '''1. 感谢（案主感谢成员）
2. 欣赏（成员相互欣赏）
3. 期待（展望后续行动）''',
            'script': '''"请案主用一个词或一句话总结，今天的讨论对您最大的帮助是什么？"
"请各位用一句话分享，今天您最大的收获是什么？"
"期待我们下次再见。"''',
            'notes': '''- 真诚、不敷衍
- 具体说明感谢的理由
- 约定下次聚会时间'''
        },
    ]

    row = data_start + 1
    for step in seven_steps:
        ws.cell(row=row, column=1, value=step['num']).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2, value=step['name']).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3, value=step['time']).alignment = CENTER_ALIGN
        ws.cell(row=row, column=4, value=step['actions']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=5, value=step['script']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=6, value=step['notes']).alignment = LEFT_ALIGN

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = CONTENT_FONT
            cell.border = THIN_BORDER
            if col in [1, 2, 3]:
                cell.fill = LIGHT_BLUE_FILL
            else:
                cell.fill = WHITE_FILL

        ws.row_dimensions[row].height = 80
        row += 1

    # 添加时间总计行
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    total_cell = ws.cell(row=row, column=1, value='时间总计')
    total_cell.font = LABEL_FONT
    total_cell.fill = MEDIUM_BLUE_FILL
    total_cell.alignment = CENTER_ALIGN
    total_cell.border = THIN_BORDER

    ws.cell(row=row, column=3, value='95分钟').alignment = CENTER_ALIGN
    ws.cell(row=row, column=3).font = LABEL_FONT
    ws.cell(row=row, column=3).fill = MEDIUM_BLUE_FILL
    ws.cell(row=row, column=3).border = THIN_BORDER

    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    note_cell = ws.cell(row=row, column=4, value='注：标准版私董会会议时长约2-2.5小时，其中七步法约95分钟，剩余时间用于茶歇和自由交流')
    note_cell.font = SMALL_FONT
    note_cell.alignment = LEFT_ALIGN
    for col in range(4, 7):
        ws.cell(row=row, column=col).border = THIN_BORDER
        if col > 4:
            ws.cell(row=row, column=col).fill = WHITE_FILL

    add_copyright(ws, 6)
    return wb


# ========== Sheet 2: 催化师话术工具包 ==========
def create_sheet2_facilitator_script(wb):
    ws = wb.create_sheet("02 催化师话术工具包")

    set_column_widths(ws, {
        'A': 15,  # 场景分类
        'B': 50,  # 建议话术
        'C': 50,  # 禁忌话术
        'D': 35,  # 场景说明
    })

    data_start = create_title_row(ws, 1,
        "系列工具表",
        "私董会运作方法论-催化师话术工具包",
        4)

    # 表头
    headers = ['场景分类', '建议话术', '禁忌话术', '场景说明']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=header)
        apply_header_style(cell)

    # 话术数据
    scripts = [
        {
            'category': '开场破冰',
            'good': '''"欢迎各位参加今天的私董会。在开始之前，请允许我简单介绍一下今天的流程……"
"请各位先做一个简短的自我介绍，让大家认识您。"
"今天我们的案主是[姓名]，让我们用热烈的掌声欢迎他/她。"''',
            'bad': '''"大家都认识吧，我就不介绍了。"
"直接开始吧，时间有限。"
"怎么还有人没到？我们先开始吧。"''',
            'note': '首次聚会需要破冰，熟悉成员背景；建立轻松氛围让成员放下防备'
        },
        {
            'category': '邀请案主',
            'good': '''"案主，请用3分钟时间向大家介绍您目前面临的核心挑战。"
"请重点描述这个问题对您意味着什么，给我们一些背景信息。"
"如果用1-10分来衡量这个问题的紧迫性，您会打几分？为什么？"''',
            'bad': '''"案主，您的问题我们大概知道了，简单说一下就行。"
"直接说重点，别讲那么多背景。"
"我们时间紧，您长话短说。"''',
            'note': '给案主充分表达的空间，同时确保信息精准、聚焦'
        },
        {
            'category': '提问环节',
            'good': '''"您的这个问题很好，请案主直接回答，不要解释原因。"
"我注意到您提到[关键词]，能具体解释一下吗？"
"这个问题我想深挖一下：[具体追问]"''',
            'bad': '''"我有个建议给您……"
"您这个问题我不太认同……"
"我以前也遇到过类似情况，当时我是这样处理的……"''',
            'note': '严格遵循"提问-回答"模式，不给建议、不评价、不比较'
        },
        {
            'category': '沉默时刻',
            'good': '''（安静等待，不打破沉默）
"如果需要更多时间思考，我们可以给案主几分钟。"
"沉默是金。请各位利用这段时间深入思考。"
"有没有人愿意先分享一个想法，哪怕只是一个词？"''',
            'bad': '''"没人有问题吗？那我代替大家问一个……"
"看来大家都没什么想法，那我们跳过这个环节吧。"
"怎么都不说话？案主说了那么多，你们就没反应？"''',
            'note': '沉默是深度思考的体现，催化师需耐心等待，不强行推进'
        },
        {
            'category': '情绪处理',
            'good': '''"我看到您（案主）情绪有些激动，请先深呼吸，我们可以暂停一下。"
"案主，您此刻的感受是什么？愿意和我们分享吗？"
"请各位给案主一点时间，我们等他/她准备好再继续。"''',
            'bad': '''"您别激动，有什么话好好说。"
"这点小事至于吗？"
"来来来，我们继续，别耽误时间。"''',
            'note': '情绪是真实反馈的信号，催化师要接纳而非压制'
        },
        {
            'category': '冲突介入',
            'good': '''"我听到两位有不同的观点，这很好。请问您们的分歧在哪里？"
"让我们先把观点亮出来，看看能否找到共识。"
"不同的声音往往是洞见的来源。请各自陈述您的理由。"''',
            'bad': '''"别吵了，我们今天是来帮助案主的，不是来吵架的。"
"您们两个的意见都不对，应该这样……"
"算了，这个话题我们不讨论了，换下一个。"''',
            'note': '冲突可能是洞见的来源，催化师要引导而非压制'
        },
        {
            'category': '收尾三赏',
            'good': '''"请案主用一句话总结，今天的讨论对您最大的帮助是什么？"
"请各位用三个词来形容今天参加私董会的感受。"
"期待[具体时间]我们再相聚，届时请案主分享进展。"''',
            'bad': '''"好了，今天就到这里吧，散会。"
"时间差不多了，大家还有什么要说的吗？没有就结束。"
"案主，记得我们今天说的，别忘了啊。"''',
            'note': '三赏闭环是私董会的精华，收尾要真诚、温暖、有期待感'
        },
    ]

    row = data_start + 1
    fills = [GREEN_FILL, ORANGE_FILL, LIGHT_BLUE_FILL, YELLOW_FILL,
             ACCENT_BLUE_FILL, LIGHT_GRAY_FILL, WHITE_FILL]

    for i, script in enumerate(scripts):
        ws.cell(row=row, column=1, value=script['category']).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2, value=script['good']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=3, value=script['bad']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=4, value=script['note']).alignment = LEFT_ALIGN

        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = CONTENT_FONT
            cell.border = THIN_BORDER
            cell.fill = fills[i % len(fills)]

        ws.row_dimensions[row].height = 90
        row += 1

    add_copyright(ws, 4)
    return wb


# ========== Sheet 3: 提问技术练习表 ==========
def create_sheet3_questioning(wb):
    ws = wb.create_sheet("03 提问技术练习表")

    set_column_widths(ws, {
        'A': 12,  # 提问类型
        'B': 30,  # 目的说明
        'C': 45,  # 经典句式
        'D': 45,  # 练习题
        'E': 45,  # 参考答案
    })

    data_start = create_title_row(ws, 1,
        "系列工具表",
        "私董会运作方法论-提问技术练习表",
        5)

    headers = ['提问类型', '目的说明', '经典句式', '练习题', '参考答案']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=header)
        apply_header_style(cell)

    questions = [
        {
            'type': '事实提问',
            'purpose': '收集客观信息，还原事实真相。通过事实提问，可以了解案主的真实处境，为后续分析提供基础。',
            'sentences': '''1. 请您描述一下具体发生了什么？
2. 这个问题最早是什么时候出现的？
3. 您上次遇到类似情况是什么时候？当时您是如何处理的？
4. 在过去6个月里，这件事进展如何？
5. 有哪些人参与了这件事？各自的角色是什么？''',
            'exercise': '''【练习题】
案主说："我的公司最近业绩下滑很厉害，我不知道该怎么办。"
请针对这个模糊陈述，提出3个事实提问来还原真相。''',
            'answer': '''参考答案：
1. "您说业绩下滑，能告诉我具体下滑了多少吗？与去年同期相比是什么情况？"
2. "这个问题最早是从什么时候开始出现的？当时发生了什么？"
3. "您认为导致业绩下滑的主要原因是什么？有哪些因素？"
4. "公司的核心高管团队目前是什么状况？有没有人离职或变动？"'''
        },
        {
            'type': '感受提问',
            'purpose': '探索情感世界，理解案主内心感受。感受提问帮助案主觉察情绪，看见自己的内心模式。',
            'sentences': '''1. 面对这件事，您的感受是什么？
2. 您现在的心情如何？如果用1-10分打分，您会打几分？
3. 您的家人/团队知道这件事吗？他们怎么看？
4. 这件事让您感到最沮丧的是什么？
5. 如果抛开所有现实顾虑，您真正想要的是什么？''',
            'exercise': '''【练习题】
一位企业创始人说："我把公司当成自己的孩子，但现在我不得不考虑卖掉它。"
请提出3个感受提问，探索他/她的内心世界。''',
            'answer': '''参考答案：
1. "当您说把公司当成孩子时，我感受到您对公司的深厚情感。做出这个决定时，您内心最大的感受是什么？"
2. "如果用一种情绪来形容您此刻的心情，会是什么？为什么？"
3. "卖掉公司对您来说意味着什么？会不会觉得自己背叛了什么？"
4. "您的家人怎么看这个决定？他们支持您吗？"'''
        },
        {
            'type': '意图提问',
            'purpose': '探索深层动机，明确真正目标。通过意图提问，帮助案主区分"想要"和"需要"，看见自己的核心诉求。',
            'sentences': '''1. 您希望通过今天的讨论获得什么？
2. 如果这个问题解决了，对您意味着什么？
3. 您最看重的是什么？是事业、家庭、还是自我实现？
4. 如果不用考虑金钱和时间，您会做什么选择？
5. 您说的"成功"具体是指什么？''',
            'exercise': '''【练习题】
一位企业家说："我想让公司上市，但又担心上市后失去控制权。"
请提出3个意图提问，探索他/她真正的意图。''',
            'answer': '''参考答案：
1. "您说担心失去控制权，对您来说'控制'为什么这么重要？控制意味着什么？"
2. "如果上市后您仍然能主导公司发展方向，这个选择对您还有障碍吗？"
3. "您真正想要的是上市本身，还是上市带来的东西（如资金、声誉、退出）？"
4. "如果这是您最后一次做重大决策，您会怎么选择？为什么？"'''
        },
        {
            'type': '假设提问',
            'purpose': '挑战思维定式，打开新的可能性。假设提问帮助案主跳出现有框架，看到更多解决方案。',
            'sentences': '''1. 如果您现在有无限的资源，这个问题会如何解决？
2. 如果您是旁观者，会给自己什么建议？
3. 如果这个问题明天就消失了，您的生活会有什么不同？
4. 如果您的朋友遇到同样的问题，您会对他说什么？
5. 假设您可以回到过去重新选择，您会改变什么？''',
            'exercise': '''【练习题】
一位传统企业主说："我做了30年传统行业，我不会做互联网，我也不相信那些新模式。"
请提出3个假设提问，挑战他/她的思维框架。''',
            'answer': '''参考答案：
1. "如果互联网只是工具而非替代品，您会如何利用它来服务您现有的业务？"
2. "如果您团队中的年轻人有机会运用互联网思维，您愿意给他们授权去尝试吗？"
3. "假设5年后，您希望公司成为什么样子？那时互联网会是您业务的一部分吗？"
4. "如果您今天做出一个改变来拥抱数字化，您觉得最小的第一步是什么？"'''
        },
    ]

    row = data_start + 1
    fills = [GREEN_FILL, ORANGE_FILL, LIGHT_BLUE_FILL, YELLOW_FILL]

    for i, q in enumerate(questions):
        ws.cell(row=row, column=1, value=q['type']).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2, value=q['purpose']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=3, value=q['sentences']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=4, value=q['exercise']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=5, value=q['answer']).alignment = LEFT_ALIGN

        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = CONTENT_FONT
            cell.border = THIN_BORDER
            if col == 1:
                cell.fill = fills[i]
            elif col == 4 or col == 5:
                cell.fill = LIGHT_GRAY_FILL
            else:
                cell.fill = WHITE_FILL

        ws.row_dimensions[row].height = 120
        row += 1

    add_copyright(ws, 5)
    return wb


# ========== Sheet 4: 私董会质量评估量表 ==========
def create_sheet4_evaluation(wb):
    ws = wb.create_sheet("04 私董会质量评估量表")

    set_column_widths(ws, {
        'A': 18,  # 评估维度
        'B': 40,  # 评分标准
        'C': 10,  # 评分
        'D': 35,  # 综合评价
    })

    data_start = create_title_row(ws, 1,
        "系列工具表",
        "私董会运作方法论-私董会质量评估量表",
        4)

    headers = ['评估维度', '评分标准', '评分(1-5)', '综合评价/改进建议']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=header)
        apply_header_style(cell)

    # 评分标准
    criteria = [
        {
            'dimension': '问题质量',
            'standards': '''优秀(5分)：问题精准聚焦、具体可衡量、真实反映案主核心困惑
良好(4分)：问题比较清晰、有一定深度
合格(3分)：问题基本明确、但不够具体或精准
待改进(2分)：问题模糊、过于开放、难以聚焦讨论
不合格(1分)：问题偏离主题、或案主自己也不清楚要讨论什么''',
            'weight': '权重：20%'
        },
        {
            'dimension': '躬问深度',
            'standards': '''优秀(5分)：通过追问挖掘到问题本质、原因清晰、影响明确
良好(4分)：追问有深度、能找到关键因素
合格(3分)：有一定追问、但深度不够
待改进(2分)：追问停留在表面、浅尝辄止
不合格(1分)：几乎没有追问、问题越讨论越模糊''',
            'weight': '权重：20%'
        },
        {
            'dimension': '场域安全',
            'standards': '''优秀(5分)：所有成员都感到被尊重、安全、可坦诚表达
良好(4分)：大部分成员能敞开发言、有少量保留
合格(3分)：基本安全、但有少数成员发言较少
待改进(2分)：有明显的防御或紧张气氛
不合格(1分)：成员之间有冲突或明显不信任''',
            'weight': '权重：20%'
        },
        {
            'dimension': '成员参与',
            'standards': '''优秀(5分)：全员积极参与、发言均衡、有高质量贡献
良好(4分)：大部分成员参与、有几位贡献突出
合格(3分)：有人积极参与、也有人沉默
待改进(2分)：只有少数人发言、其他人旁观
不合格(1分)：参与度极低、讨论被个别人主导''',
            'weight': '权重：15%'
        },
        {
            'dimension': '三赏真诚',
            'standards': '''优秀(5分)：三赏环节真诚感人、有具体事例、有深度连接
良好(4分)：三赏比较真诚、有一定感动
合格(3分)：三赏形式完整、但略显客套
待改进(2分)：三赏比较敷衍、不够真诚
不合格(1分)：跳过三赏或流于形式''',
            'weight': '权重：15%'
        },
        {
            'dimension': '时间控制',
            'standards': '''优秀(5分)：各环节时间分配合理、准时开始和结束
良好(4分)：基本按时、有小调整
合格(3分)：有些拖沓或仓促、但整体可控
待改进(2分)：严重超时或提前结束
不合格(1分)：完全失控、时间分配混乱''',
            'weight': '权重：10%'
        },
    ]

    row = data_start + 1
    for i, c in enumerate(criteria):
        ws.cell(row=row, column=1, value=c['dimension']).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2, value=c['standards']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=3, value='').alignment = CENTER_ALIGN  # 评分留空
        ws.cell(row=row, column=4, value='').alignment = LEFT_ALIGN  # 改进建议留空

        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = CONTENT_FONT
            cell.border = THIN_BORDER
            if col == 1:
                cell.fill = ACCENT_BLUE_FILL
            elif col == 3:
                cell.fill = YELLOW_FILL
            else:
                cell.fill = WHITE_FILL

        ws.row_dimensions[row].height = 100
        row += 1

    # 综合评分行
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    total_cell = ws.cell(row=row, column=1, value='综合评分')
    total_cell.font = LABEL_FONT
    total_cell.fill = DARK_BLUE_FILL
    total_cell.alignment = CENTER_ALIGN
    total_cell.border = THIN_BORDER

    # 综合评分公式说明
    formula_note = ws.cell(row=row, column=3,
        value='=加权平均')
    formula_note.font = Font(name='微软雅黑', size=9, italic=True, color='808080')
    formula_note.alignment = CENTER_ALIGN
    formula_note.fill = LIGHT_GRAY_FILL
    formula_note.border = THIN_BORDER

    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=4)
    result_cell = ws.cell(row=row, column=4,
        value='评估等级：4.5-5分=优秀 | 4.0-4.4=良好 | 3.0-3.9=合格 | <3分=待改进')
    result_cell.font = SMALL_FONT
    result_cell.alignment = LEFT_ALIGN
    result_cell.fill = WHITE_FILL
    result_cell.border = THIN_BORDER
    row += 2

    # 改进建议区
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    section_title = ws.cell(row=row, column=1, value='改进计划')
    section_title.font = LABEL_FONT
    section_title.fill = MEDIUM_BLUE_FILL
    section_title.alignment = CENTER_ALIGN
    section_title.border = THIN_BORDER
    row += 1

    improvement_labels = ['本次会议最大亮点', '最需改进的地方', '下次会议重点关注']
    for label in improvement_labels:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.fill = ACCENT_BLUE_FILL
        label_cell.alignment = CENTER_ALIGN
        label_cell.border = THIN_BORDER

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        value_cell = ws.cell(row=row, column=2, value='')
        value_cell.fill = WHITE_FILL
        value_cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    add_copyright(ws, 4)
    return wb


# ========== Sheet 5: 私董会小组组建工具 ==========
def create_sheet5_group_building(wb):
    ws = wb.create_sheet("05 私董会小组组建工具")

    set_column_widths(ws, {
        'A': 25,  # 维度
        'B': 35,  # 具体标准
        'C': 20,  # 评估结果
        'D': 20,  # 备注
    })

    data_start = create_title_row(ws, 1,
        "系列工具表",
        "私董会运作方法论-私董会小组组建工具",
        4)

    # Part 1: 成员筛选标准
    ws.merge_cells(start_row=data_start, start_column=1, end_row=data_start, end_column=4)
    part1_title = ws.cell(row=data_start, column=1, value='一、成员筛选标准表（5个维度）')
    part1_title.font = LABEL_FONT
    part1_title.fill = DARK_BLUE_FILL
    part1_title.alignment = CENTER_ALIGN
    part1_title.border = THIN_BORDER

    row = data_start + 1
    headers = ['筛选维度', '具体标准', '评估结果', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)

    criteria = [
        {'dim': '维度一：身份匹配', 'std': '企业创始人、董事长、CEO或同等级别高管；具有决策权', 'eval': '', 'note': '对等才能对话'},
        {'dim': '维度二：发展需求', 'std': '有明确的个人或企业发展问题需要探讨；愿意开放分享', 'eval': '', 'note': '需求明确才有动力'},
        {'dim': '维度三：行业多样性', 'std': '不同行业背景；非直接竞争对手；能带来跨界视角', 'eval': '', 'note': '跨界带来洞见'},
        {'dim': '维度四：信任基础', 'std': '有共同信任的推荐人；愿意签署保密协议；人品口碑良好', 'eval': '', 'note': '信任是底线'},
        {'dim': '维度五：参与承诺', 'std': '能保证持续参与（每次会议、每季度至少一次）；愿意付出会费', 'eval': '', 'note': '承诺决定质量'},
    ]

    row += 1
    for c in criteria:
        ws.cell(row=row, column=1, value=c['dim']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=2, value=c['std']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=3, value=c['eval']).alignment = CENTER_ALIGN
        ws.cell(row=row, column=4, value=c['note']).alignment = CENTER_ALIGN

        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = CONTENT_FONT
            cell.border = THIN_BORDER
            if col == 1:
                cell.fill = ACCENT_BLUE_FILL
            elif col == 3:
                cell.fill = YELLOW_FILL
            else:
                cell.fill = WHITE_FILL
        ws.row_dimensions[row].height = 35
        row += 1

    row += 1

    # Part 2: 保密协议模板
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    part2_title = ws.cell(row=row, column=1, value='二、保密协议模板')
    part2_title.font = LABEL_FONT
    part2_title.fill = DARK_BLUE_FILL
    part2_title.alignment = CENTER_ALIGN
    part2_title.border = THIN_BORDER
    row += 1

    nda_text = '''保密协议

本人/公司（以下简称"承诺人"）自愿加入私董会组织，并承诺遵守以下保密义务：

一、保密内容
在私董会活动中知悉的案主信息、企业情况、商业秘密、个人隐私及其他成员分享的一切信息。

二、保密义务
1. 承诺人不得向任何第三方透露上述保密内容；
2. 承诺人不得利用保密信息为自己或第三方谋取利益；
3. 承诺人不得记录、复制或保存保密信息（经案主同意的除外）；
4. 违反本协议任何条款，承诺人愿承担相应法律责任。

三、协议期限
本协议自签署之日起生效，保密义务永久有效。

四、签署信息
承诺人签名：________________  日期：________________
推荐人签名：________________  日期：________________'''

    ws.merge_cells(start_row=row, start_column=1, end_row=row+15, end_column=4)
    nda_cell = ws.cell(row=row, column=1, value=nda_text)
    nda_cell.font = CONTENT_FONT
    nda_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    nda_cell.fill = LIGHT_GRAY_FILL
    nda_cell.border = THIN_BORDER
    row += 16

    row += 1

    # Part 3: 规则制定清单
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    part3_title = ws.cell(row=row, column=1, value='三、规则制定清单')
    part3_title.font = LABEL_FONT
    part3_title.fill = DARK_BLUE_FILL
    part3_title.alignment = CENTER_ALIGN
    part3_title.border = THIN_BORDER
    row += 1

    headers = ['规则类别', '规则内容', '约定事项', '确认签署']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    rules = [
        ('会议频率', '每月/每季度一次', '具体时间：_____', '□已确认'),
        ('会议时长', '每次2.5-3小时', '迟到处理：_____', '□已确认'),
        ('轮值案主', '按顺序轮流担任', '缺席安排：_____', '□已确认'),
        ('会费标准', '年费_____元', '支付方式：_____', '□已确认'),
        ('请假制度', '提前48小时请假', '缺席限制：每年不超过2次', '□已确认'),
        ('退出机制', '提前30天书面通知', '会费退还：按剩余月份比例', '□已确认'),
        ('保密条款', '永久保密义务', '违规处理：立即退出且不退费', '□已确认'),
    ]

    for r in rules:
        ws.cell(row=row, column=1, value=r[0]).alignment = LEFT_ALIGN
        ws.cell(row=row, column=2, value=r[1]).alignment = LEFT_ALIGN
        ws.cell(row=row, column=3, value=r[2]).alignment = LEFT_ALIGN
        ws.cell(row=row, column=4, value=r[3]).alignment = CENTER_ALIGN

        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = CONTENT_FONT
            cell.border = THIN_BORDER
            if col == 1:
                cell.fill = ACCENT_BLUE_FILL
            elif col == 4:
                cell.fill = GREEN_FILL
            else:
                cell.fill = WHITE_FILL
        ws.row_dimensions[row].height = 25
        row += 1

    row += 1

    # Part 4: 启动仪式流程
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    part4_title = ws.cell(row=row, column=1, value='四、启动仪式流程')
    part4_title.font = LABEL_FONT
    part4_title.fill = DARK_BLUE_FILL
    part4_title.alignment = CENTER_ALIGN
    part4_title.border = THIN_BORDER
    row += 1

    ceremony_headers = ['环节', '内容', '时长', '负责人']
    for col, header in enumerate(ceremony_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    ceremony = [
        ('1. 开场致辞', '召集人介绍私董会理念、目标和运作方式', '10分钟', '召集人'),
        ('2. 成员自我介绍', '每人3分钟，分享企业和个人背景', '25分钟', '全体成员'),
        ('3. 共同愿景讨论', '讨论小组使命、价值观、期望', '20分钟', '引导师'),
        ('4. 规则共创', '共同讨论制定小组规则', '20分钟', '引导师'),
        ('5. 保密协议签署', '逐一签署保密协议', '10分钟', '全体成员'),
        ('6. 启动庆祝', '合影留念、自由交流', '15分钟', '全体成员'),
    ]

    for c in ceremony:
        ws.cell(row=row, column=1, value=c[0]).alignment = LEFT_ALIGN
        ws.cell(row=row, column=2, value=c[1]).alignment = LEFT_ALIGN
        ws.cell(row=row, column=3, value=c[2]).alignment = CENTER_ALIGN
        ws.cell(row=row, column=4, value=c[3]).alignment = CENTER_ALIGN

        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = CONTENT_FONT
            cell.border = THIN_BORDER
            if col == 1:
                cell.fill = LIGHT_BLUE_FILL
            else:
                cell.fill = WHITE_FILL
        ws.row_dimensions[row].height = 30
        row += 1

    add_copyright(ws, 4)
    return wb


# ========== Sheet 6: 私董会运营管理工具 ==========
def create_sheet6_operations(wb):
    ws = wb.create_sheet("06 私董会运营管理工具")

    set_column_widths(ws, {
        'A': 20,  # 项目
        'B': 15,  # 时间/日期
        'C': 30,  # 内容摘要
        'D': 25,  # 关键决策
        'E': 25,  # 行动计划
        'F': 10,  # 责任人
        'G': 10,  # 状态
    })

    data_start = create_title_row(ws, 1,
        "系列工具表",
        "私董会运作方法论-私董会运营管理工具",
        7)

    # Part 1: 会议记录模板
    ws.merge_cells(start_row=data_start, start_column=1, end_row=data_start, end_column=7)
    part1_title = ws.cell(row=data_start, column=1, value='一、会议记录模板')
    part1_title.font = LABEL_FONT
    part1_title.fill = DARK_BLUE_FILL
    part1_title.alignment = CENTER_ALIGN
    part1_title.border = THIN_BORDER

    row = data_start + 1

    # 会议基本信息
    info_labels = ['会议日期：', '会议地点：', '召集人：', '记录人：', '案主：']
    for i, label in enumerate(info_labels):
        ws.cell(row=row, column=i+1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=i+1).fill = ACCENT_BLUE_FILL
        ws.cell(row=row, column=i+1).alignment = RIGHT_ALIGN
        ws.cell(row=row, column=i+1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=i+2, end_row=row, end_column=i+3 if i < 4 else 7)
        val_cell = ws.cell(row=row, column=i+2)
        val_cell.fill = WHITE_FILL
        val_cell.border = THIN_BORDER
    row += 2

    # 出席人员
    ws.cell(row=row, column=1, value='出席人员：').font = LABEL_FONT
    ws.cell(row=row, column=1).fill = ACCENT_BLUE_FILL
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    att_cell = ws.cell(row=row, column=2)
    att_cell.fill = WHITE_FILL
    att_cell.border = THIN_BORDER
    row += 1

    # 缺席人员
    ws.cell(row=row, column=1, value='缺席人员：').font = LABEL_FONT
    ws.cell(row=row, column=1).fill = ACCENT_BLUE_FILL
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    abs_cell = ws.cell(row=row, column=2)
    abs_cell.fill = WHITE_FILL
    abs_cell.border = THIN_BORDER
    row += 2

    # 记录表格
    headers = ['项目', '时间/日期', '内容摘要', '关键决策', '行动计划', '责任人', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    for i in range(1, 11):  # 10行空白记录
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col, value='')
            cell.fill = WHITE_FILL if col % 2 == 0 else LIGHT_GRAY_FILL
            cell.border = THIN_BORDER
        row += 1

    row += 1

    # Part 2: 成员参与跟踪表
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    part2_title = ws.cell(row=row, column=1, value='二、成员参与跟踪表（12个月）')
    part2_title.font = LABEL_FONT
    part2_title.fill = DARK_BLUE_FILL
    part2_title.alignment = CENTER_ALIGN
    part2_title.border = THIN_BORDER
    row += 1

    # 表头：成员姓名 + 12个月
    member_headers = ['成员姓名', '月份1', '月份2', '月份3', '月份4', '月份5', '月份6',
                      '月份7', '月份8', '月份9', '月份10', '月份11', '月份12']
    ws.cell(row=row, column=1, value='成员姓名').font = HEADER_FONT
    ws.cell(row=row, column=1).fill = DARK_BLUE_FILL
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    ws.cell(row=row, column=1).border = THIN_BORDER

    for i in range(1, 13):
        cell = ws.cell(row=row, column=i+1, value=f'月')
        cell.font = HEADER_FONT
        cell.fill = MEDIUM_BLUE_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    row += 1

    # 成员行（预留8个成员位置）
    for i in range(1, 9):
        ws.cell(row=row, column=1, value=f'成员{i}').font = LABEL_FONT
        ws.cell(row=row, column=1).fill = ACCENT_BLUE_FILL
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER

        for col in range(2, 14):
            cell = ws.cell(row=row, column=col, value='□出席')
            cell.alignment = CENTER_ALIGN
            cell.fill = WHITE_FILL
            cell.border = THIN_BORDER
        row += 1

    # 统计行
    ws.cell(row=row, column=1, value='出席次数').font = LABEL_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    ws.cell(row=row, column=1).border = THIN_BORDER

    for col in range(2, 14):
        cell = ws.cell(row=row, column=col, value='0')
        cell.font = LABEL_FONT
        cell.fill = YELLOW_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    row += 2

    # Part 3: 效果评估跟踪表
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    part3_title = ws.cell(row=row, column=1, value='三、效果评估跟踪表')
    part3_title.font = LABEL_FONT
    part3_title.fill = DARK_BLUE_FILL
    part3_title.alignment = CENTER_ALIGN
    part3_title.border = THIN_BORDER
    row += 1

    eval_headers = ['评估维度', '第1次', '第2次', '第3次', '第4次', '第5次', '第6次']
    for col, header in enumerate(eval_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    eval_dims = ['问题解决满意度', '成员信任度', '参与活跃度', '学习成长度', '整体满意度']
    for dim in eval_dims:
        ws.cell(row=row, column=1, value=dim).font = CONTENT_FONT
        ws.cell(row=row, column=1).fill = ACCENT_BLUE_FILL
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER

        for col in range(2, 8):
            cell = ws.cell(row=row, column=col, value='_/10')
            cell.alignment = CENTER_ALIGN
            cell.fill = WHITE_FILL
            cell.border = THIN_BORDER
        row += 1

    row += 1

    # Part 4: 持续改进计划表
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    part4_title = ws.cell(row=row, column=1, value='四、持续改进计划表')
    part4_title.font = LABEL_FONT
    part4_title.fill = DARK_BLUE_FILL
    part4_title.alignment = CENTER_ALIGN
    part4_title.border = THIN_BORDER
    row += 1

    improvement_headers = ['改进项目', '现状问题', '目标状态', '行动计划', '开始日期', '负责人', '完成日期']
    for col, header in enumerate(improvement_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    for i in range(1, 6):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col, value='')
            cell.fill = WHITE_FILL if i % 2 == 0 else LIGHT_GRAY_FILL
            cell.border = THIN_BORDER
        row += 1

    add_copyright(ws, 7)
    return wb


# ========== Sheet 7: 角色扮演场景卡 ==========
def create_sheet7_scenario_cards(wb):
    ws = wb.create_sheet("07 角色扮演场景卡")

    set_column_widths(ws, {
        'A': 8,   # 编号
        'B': 20,  # 场景名称
        'C': 40,  # 背景描述
        'D': 20,  # 问题类型
        'E': 40,  # 关键挑战
        'F': 45,  # 评分要点
    })

    data_start = create_title_row(ws, 1,
        "系列工具表",
        "私董会运作方法论-角色扮演场景卡（10个典型场景）",
        6)

    headers = ['编号', '场景名称', '背景描述', '问题类型', '关键挑战', '评分要点']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=header)
        apply_header_style(cell)

    scenarios = [
        {
            'num': '01',
            'name': '战略转型困境',
            'background': '''一家年营收5亿元的制造业企业，老板考虑是否要从传统制造转型到智能制造。
转型需要投入2亿元，更换大部分设备，同时需要裁员30%。
老板内心矛盾：不做转型可能失去竞争力，做了转型风险巨大且不确定。''',
            'type': '战略决策',
            'challenges': '''1. 如何帮助案主厘清转型的真实动机（是恐惧驱动还是愿景驱动）
2. 如何评估转型风险与不转型的风险
3. 如何平衡股东利益、员工利益和社会责任
4. 如何制定切实可行的转型路径''',
            'criteria': '''- 问题澄清：能否帮助案主找到核心问题（10分）
- 多角度探索：是否充分探讨了转型与不转型的利弊（20分）
- 情感支持：是否给予案主情感上的理解和支持（15分）
- 行动建议：是否产生了可执行的行动计划（25分）
- 三赏质量：收尾环节是否真诚有深度（10分）
- 团队参与：所有成员是否都有贡献（20分）'''
        },
        {
            'num': '02',
            'name': '传承难题',
            'background': '''一家经营30年的家族企业，老板65岁，儿子30岁但在家族企业工作意愿不强。
儿子有自己的创业梦想，想在外面发展。老板希望儿子接班，但不知如何开口。
如果儿子不接班，企业未来何去何从？''',
            'type': '传承规划',
            'challenges': '''1. 如何理解两代人的不同诉求
2. 如何帮助案主（父亲）放下控制欲
3. 如何探讨儿子接班的各种可能性
4. 如何面对可能的"去家族化"选项''',
            'criteria': '''- 关系洞察：能否看见案主家庭关系的深层动态（20分）
- 情感疏导：能否让案主坦然面对传承焦虑（20分）
- 方案探索：是否充分探讨了各种传承可能性（25分）
- 尊重边界：是否尊重儿子独立选择权（15分）
- 行动计划：是否制定了具体的沟通或推进计划（20分）'''
        },
        {
            'num': '03',
            'name': '团队危机',
            'background': '''公司联合创始人兼CTO突然提出离职，要去竞争对手那里。
这位高管掌握公司核心技术，带走了部分核心技术团队。
公司正处于融资关键期，这个变故可能影响融资进程。''',
            'type': '人才危机',
            'challenges': '''1. 如何帮助案主冷静分析高管离职的真实原因
2. 如何评估对公司业务和融资的实际影响
3. 如何处理情感冲击（被背叛的感觉）
4. 如何制定应急方案和长期人才战略''',
            'criteria': '''- 情绪管理：能否帮助案主从情绪中抽离、理性分析（20分）
- 原因探索：是否深入挖掘了离职的真正原因（20分）
- 风险评估：是否全面评估了影响和应对方案（25分）
- 行动导向：是否制定了具体的危机处理计划（25分）
- 预防机制：是否探讨了如何避免类似情况再次发生（10分）'''
        },
        {
            'num': '04',
            'name': '战略分歧',
            'background': '''两位联合创始人对于公司战略方向产生严重分歧。
创始人A主张专注国内市场，稳扎稳打；创始人B主张快速国际化。
两人矛盾已经影响到日常决策和团队氛围，其他高管也开始站队。''',
            'type': '股东矛盾',
            'challenges': '''1. 如何帮助案主跳出一致性冲突，看到共同利益
2. 如何探索第三选择的可能性
3. 如何处理双方的情感因素（被否定感、不安全感）
4. 如何建立长期有效的决策机制''',
            'criteria': '''- 中立立场：能否保持中立，不偏袒任何一方（15分）
- 利益挖掘：能否找到双方的共同利益点（25分）
- 方案探索：是否探讨了多元化的战略路径（25分）
- 机制建设：是否提出了长期决策机制的改善建议（20分）
- 关系修复：是否帮助改善了双方的关系（15分）'''
        },
        {
            'num': '05',
            'name': '空降高管融入',
            'background': '''公司重金从大企业引进一位COO，但三个月过去了，融入情况不理想。
老团队对空降高管有抵触情绪，不配合工作。
COO开始怀疑自己是否应该来这家公司。案主是CEO，需要帮助COO融入。''',
            'type': '管理融合',
            'challenges': '''1. 如何帮助案主理解融入困难的真实原因
2. 如何平衡空降高管的期望与现实
3. 如何处理老团队的抵触情绪
4. 如何建立信任和影响力''',
            'criteria': '''- 原因分析：能否帮助案主深入分析融入困难的原因（25分）
- 双方视角：是否同时考虑了COO和老团队的立场（20分）
- 期望管理：是否帮助调整了各方的期望（20分）
- 行动计划：是否制定了具体的融入支持计划（25分）
- 长期机制：是否探讨了如何避免类似问题（10分）'''
        },
        {
            'num': '06',
            'name': '业务瓶颈',
            'background': '''公司业务连续三年停滞不前，营收一直在2亿元左右徘徊。
尝试过多种新业务拓展，都以失败告终。
团队士气低落，中层骨干开始流失。''',
            'type': '增长瓶颈',
            'challenges': '''1. 如何帮助案主找到业务停滞的根本原因
2. 如何突破"舒适区"，找到新的增长点
3. 如何重振团队士气和信心
4. 如何平衡现有业务优化与新业务探索''',
            'criteria': '''- 根因挖掘：能否找到真正的瓶颈所在（25分）
- 思维破局：是否能帮助案主打破固有思维（25分）
- 信心重建：是否帮助重塑团队信心（20分）
- 行动探索：是否产生了具体的突破方向（20分）
- 风险意识：是否考虑了试错成本和风险控制（10分）'''
        },
        {
            'num': '07',
            'name': '股权纠纷',
            'background': '''早期创业时，三位创始人平分股权，没有明确分工。
随着公司发展，大家贡献差异越来越大，但股权仍然是均分的。
贡献大的创始人感到不公平，想要调整股权，但其他人不愿让步。''',
            'type': '股权治理',
            'challenges': '''1. 如何帮助案主理性看待历史贡献与股权的关系
2. 如何平衡情感因素与理性分析
3. 如何探讨各方都能接受的解决方案
4. 如何避免股权纠纷影响公司发展''',
            'criteria': '''- 历史视角：是否客观看待了股权分配的历史背景（15分）
- 公平探索：是否探讨了多维度的公平定义（25分）
- 方案设计：是否提出了切实可行的调整方案（30分）
- 关系维护：是否考虑了调整后的合作关系（20分）
- 法律意识：是否提示了法律和合规风险（10分）'''
        },
        {
            'num': '08',
            'name': '家庭与事业平衡',
            'background': '''一位女性企业家，创业10年，公司年营收过亿。
但她几乎没有个人生活，每天工作14小时，孩子由保姆带大。
她开始怀疑这一切是否值得，但停下来又怕公司受影响。''',
            'type': '个人成长',
            'challenges': '''1. 如何帮助案主觉察真实的内心需求
2. 如何探索事业与家庭的新平衡模式
3. 如何让案主放下"完美"执念
4. 如何帮助她看见自己的成就和价值''',
            'criteria': '''- 情感洞察：能否深入理解案主的内心挣扎（25分）
- 需求探索：是否帮助案主厘清真正想要的是什么（25分）
- 方案探索：是否探讨了多种平衡可能性（20分）
- 自我认知：是否帮助案主重新认识自己的价值（20分）
- 支持系统：是否探讨了可以寻求支持的方式（10分）'''
        },
        {
            'num': '09',
            'name': '创新阻力',
            'background': '''公司想推动数字化转型，但老员工强烈反对。
老员工担心被新技术取代，在暗中抵制变革。
管理层也分成立场鲜明的两派，公司气氛紧张。''',
            'type': '变革管理',
            'challenges': '''1. 如何理解老员工的恐惧和不安全感
2. 如何平衡变革与稳定的关系
3. 如何找到让各方都能接受的转型节奏
4. 如何处理管理层内部的分歧''',
            'criteria': '''- 共情能力：能否帮助案主理解老员工的立场（20分）
- 利益分析：是否分析了各方的核心利益诉求（25分）
- 方案设计：是否提出了渐进式的变革路径（25分）
- 阻力化解：是否探讨了化解阻力的具体方法（20分）
- 内部统一：是否帮助管理层达成共识（10分）'''
        },
        {
            'num': '10',
            'name': '上市决策',
            'background': '''公司已达到上市基本条件，有投资机构表示愿意投资。
但上市意味着更严格的信息披露，创始人将失去部分控制权。
案主不确定现在是上市的时机，也不确定自己是否真的想要上市。''',
            'type': '资本决策',
            'challenges': '''1. 如何帮助案主厘清上市的真正动机
2. 如何权衡上市的利弊
3. 如何面对上市后的挑战和压力
4. 如何做出符合长期价值的决策''',
            'criteria': '''- 动机探索：是否帮助案主看清上市的真正动机（25分）
- 利弊分析：是否全面分析了上市的利弊（25分）
- 长期视角：是否从长期价值角度思考问题（20分）
- 替代方案：是否探讨了其他融资和发展路径（20分）
- 决策质量：是否帮助案主做出内心踏实的决定（10分）'''
        },
    ]

    row = data_start + 1
    fills = [GREEN_FILL, ORANGE_FILL, LIGHT_BLUE_FILL, YELLOW_FILL,
             ACCENT_BLUE_FILL, LIGHT_GRAY_FILL, WHITE_FILL, GREEN_FILL,
             ORANGE_FILL, LIGHT_BLUE_FILL]

    for i, s in enumerate(scenarios):
        ws.cell(row=row, column=1, value=s['num']).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2, value=s['name']).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3, value=s['background']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=4, value=s['type']).alignment = CENTER_ALIGN
        ws.cell(row=row, column=5, value=s['challenges']).alignment = LEFT_ALIGN
        ws.cell(row=row, column=6, value=s['criteria']).alignment = LEFT_ALIGN

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = CONTENT_FONT
            cell.border = THIN_BORDER
            if col == 1:
                cell.fill = fills[i]
            elif col == 2 or col == 4:
                cell.fill = ACCENT_BLUE_FILL
            else:
                cell.fill = WHITE_FILL

        ws.row_dimensions[row].height = 100
        row += 1

    add_copyright(ws, 6)
    return wb


# ========== Sheet 8: 课后跟进工具 ==========
def create_sheet8_followup(wb):
    ws = wb.create_sheet("08 课后跟进工具")

    set_column_widths(ws, {
        'A': 25,  # 项目
        'B': 40,  # 内容/问题
        'C': 40,  # 回答/计划
        'D': 20,  # 日期
    })

    data_start = create_title_row(ws, 1,
        "系列工具表",
        "私董会运作方法论-课后跟进工具",
        4)

    # Part 1: 个人反思报告模板
    ws.merge_cells(start_row=data_start, start_column=1, end_row=data_start, end_column=4)
    part1_title = ws.cell(row=data_start, column=1, value='一、个人反思报告模板（30天）')
    part1_title.font = LABEL_FONT
    part1_title.fill = DARK_BLUE_FILL
    part1_title.alignment = CENTER_ALIGN
    part1_title.border = THIN_BORDER

    row = data_start + 1

    reflection_labels = [
        ('学习收获', '参加本次私董会，您最大的3个收获是什么？', '请列出并简要说明'),
        ('理念更新', '有哪些旧观念被打破？有哪些新认知形成？', '请具体描述'),
        ('行动计划', '基于本次讨论，您决定采取什么行动？', '请设定明确的目标和时间节点'),
        ('阻力预测', '执行计划中可能遇到什么阻力？如何克服？', '请提前预判并准备应对方案'),
        ('支持需求', '您需要什么支持？谁可以提供帮助？', '请说明需要的具体支持'),
        ('下一步', '未来30天，您最想做的一件事是什么？', '请设定具体、可衡量的目标'),
    ]

    headers = ['反思维度', '引导问题', '您的回答', '完成日期']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    for label, q, note in reflection_labels:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).fill = ACCENT_BLUE_FILL
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER

        ws.cell(row=row, column=2, value=q).font = CONTENT_FONT
        ws.cell(row=row, column=2).fill = LIGHT_GRAY_FILL
        ws.cell(row=row, column=2).alignment = LEFT_ALIGN
        ws.cell(row=row, column=2).border = THIN_BORDER

        ws.cell(row=row, column=3, value='').font = CONTENT_FONT
        ws.cell(row=row, column=3).fill = WHITE_FILL
        ws.cell(row=row, column=3).alignment = LEFT_ALIGN
        ws.cell(row=row, column=3).border = THIN_BORDER

        ws.cell(row=row, column=4, value='').font = CONTENT_FONT
        ws.cell(row=row, column=4).fill = YELLOW_FILL
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN
        ws.cell(row=row, column=4).border = THIN_BORDER

        ws.row_dimensions[row].height = 40
        row += 1

    row += 1

    # Part 2: 行动计划模板
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    part2_title = ws.cell(row=row, column=1, value='二、个人行动计划模板')
    part2_title.font = LABEL_FONT
    part2_title.fill = DARK_BLUE_FILL
    part2_title.alignment = CENTER_ALIGN
    part2_title.border = THIN_BORDER
    row += 1

    action_headers = ['行动项', '具体做法', '衡量标准', '完成时间']
    for col, header in enumerate(action_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    for i in range(1, 6):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col, value='')
            cell.fill = WHITE_FILL if i % 2 == 0 else LIGHT_GRAY_FILL
            cell.border = THIN_BORDER
        row += 1

    row += 1

    # Part 3: 三个月跟进计划
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    part3_title = ws.cell(row=row, column=1, value='三、三个月跟进计划')
    part3_title.font = LABEL_FONT
    part3_title.fill = DARK_BLUE_FILL
    part3_title.alignment = CENTER_ALIGN
    part3_title.border = THIN_BORDER
    row += 1

    followup_headers = ['阶段', '时间节点', '重点跟进事项', '预期成果']
    for col, header in enumerate(followup_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    followup_phases = [
        ('第一周', '第1周结束时', '回顾私董会内容，确定优先级最高的问题', '形成问题清单，选择第一个聚焦课题'),
        ('第一个月', '第30天', '检查行动计划的执行情况，记录遇到的挑战', '完成至少一项具体行动，整理挑战清单'),
        ('第二个月', '第60天', '评估第一项课题的进展，讨论下一步行动', '取得阶段性进展或调整方向'),
        ('第三个月', '第90天', '回顾整体进展，制定下一阶段计划', '完成复盘，准备下一次私董会分享'),
    ]

    fills = [GREEN_FILL, ORANGE_FILL, LIGHT_BLUE_FILL, YELLOW_FILL]
    for i, (phase, time, focus, outcome) in enumerate(followup_phases):
        ws.cell(row=row, column=1, value=phase).font = LABEL_FONT
        ws.cell(row=row, column=1).fill = fills[i]
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER

        ws.cell(row=row, column=2, value=time).font = CONTENT_FONT
        ws.cell(row=row, column=2).fill = WHITE_FILL
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2).border = THIN_BORDER

        ws.cell(row=row, column=3, value=focus).font = CONTENT_FONT
        ws.cell(row=row, column=3).fill = WHITE_FILL
        ws.cell(row=row, column=3).alignment = LEFT_ALIGN
        ws.cell(row=row, column=3).border = THIN_BORDER

        ws.cell(row=row, column=4, value=outcome).font = CONTENT_FONT
        ws.cell(row=row, column=4).fill = WHITE_FILL
        ws.cell(row=row, column=4).alignment = LEFT_ALIGN
        ws.cell(row=row, column=4).border = THIN_BORDER

        ws.row_dimensions[row].height = 40
        row += 1

    row += 1

    # Part 4: 常见问题应对指南
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    part4_title = ws.cell(row=row, column=1, value='四、常见问题应对指南')
    part4_title.font = LABEL_FONT
    part4_title.fill = DARK_BLUE_FILL
    part4_title.alignment = CENTER_ALIGN
    part4_title.border = THIN_BORDER
    row += 1

    faq_headers = ['问题类型', '常见表现', '应对策略', '示例话术']
    for col, header in enumerate(faq_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    faqs = [
        {
            'type': '案主不开放',
            'manifest': '回答简短、不愿深入、回避关键问题',
            'strategy': '1. 接纳沉默，给案主时间
2. 换个角度提问，从侧面包围
3. 请其他成员先分享类似经历
4. 温和地反馈您观察到的现象',
            'example': '"我注意到您对这个问题的回应比较简短，是不是这个问题对您来说特别敏感？如果需要更多时间，我们可以先听听其他成员的想法。"'
        },
        {
            'type': '成员给建议',
            'manifest': '成员直接说"我建议您……"或"您应该……"',
            'strategy': '1. 温和打断，感谢好意
2. 将建议转化为提问
3. 提醒私董会的规则
4. 解释提问与建议的区别',
            'example': '"感谢您的建议。不过在私董会中，我们更希望通过提问来帮助案主自己找到答案。我们可以把这个建议转化为一个问题吗？比如：您有没有想过……"'
        },
        {
            'type': '沉默抗拒',
            'manifest': '某些成员全程不说话，或多次表示"我没问题"',
            'strategy': '1. 尊重选择，不强迫
2. 观察沉默背后的原因
3. 会后私下沟通
4. 调整下次会议的提问方向',
            'example': '"沉默也是一种参与方式。如果您准备好了，任何时候都欢迎分享。如果没有也没关系，您的倾听对案主也是支持。"'
        },
        {
            'type': '情绪失控',
            'manifest': '案主或成员情绪激动，落泪或愤怒',
            'strategy': '1. 暂停环节，给情绪空间
2. 表达理解，不评判
3. 给予时间恢复
4. 决定是否继续或改期',
            'example': '"我看到您此刻情绪很激动。我们可以暂停几分钟，您需要喝点水吗？我相信在场的每位成员都能理解您的感受。"'
        },
        {
            'type': '超时问题',
            'manifest': '某个问题讨论时间过长，影响后续环节',
            'strategy': '1. 温和提醒时间
2. 总结已讨论的内容
3. 询问案主是否需要继续
4. 果断收尾，转入下一环节',
            'example': '"我们在这个问题上已经讨论了15分钟，我观察到您有了很多新的思考。为了保证后续环节的时间，我们可以做个简短总结，然后进入下一个话题吗？"'
        },
        {
            'type': '话题偏离',
            'manifest': '讨论越来越偏离案主原始问题',
            'strategy': '1. 记录偏离的话题
2. 温和拉回主题
3. 说明偏离的原因
4. 询问是否需要专门讨论新话题',
            'example': '"我注意到我们的讨论已经延伸到了[新话题]，这确实是个重要议题。不过为了保证对案主的支持，我们是不是先回到[原始问题]？如果我们时间允许，可以在最后留几分钟讨论这个新话题。"'
        },
    ]

    for faq in faqs:
        ws.cell(row=row, column=1, value=faq['type']).font = LABEL_FONT
        ws.cell(row=row, column=1).fill = ACCENT_BLUE_FILL
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER

        ws.cell(row=row, column=2, value=faq['manifest']).font = CONTENT_FONT
        ws.cell(row=row, column=2).fill = ORANGE_FILL
        ws.cell(row=row, column=2).alignment = LEFT_ALIGN
        ws.cell(row=row, column=2).border = THIN_BORDER

        ws.cell(row=row, column=3, value=faq['strategy']).font = CONTENT_FONT
        ws.cell(row=row, column=3).fill = WHITE_FILL
        ws.cell(row=row, column=3).alignment = LEFT_ALIGN
        ws.cell(row=row, column=3).border = THIN_BORDER

        ws.cell(row=row, column=4, value=faq['example']).font = CONTENT_FONT
        ws.cell(row=row, column=4).fill = LIGHT_GRAY_FILL
        ws.cell(row=row, column=4).alignment = LEFT_ALIGN
        ws.cell(row=row, column=4).border = THIN_BORDER

        ws.row_dimensions[row].height = 80
        row += 1

    add_copyright(ws, 4)
    return wb


# ========== 主函数 ==========
def main():
    # 创建工作簿
    wb = openpyxl.Workbook()

    # 创建所有工作表
    create_sheet1_seven_steps(wb)
    create_sheet2_facilitator_script(wb)
    create_sheet3_questioning(wb)
    create_sheet4_evaluation(wb)
    create_sheet5_group_building(wb)
    create_sheet6_operations(wb)
    create_sheet7_scenario_cards(wb)
    create_sheet8_followup(wb)

    # 保存文件
    output_path = '/c/Users/Administrator/Desktop/私董会/完整课程包/04-工具表单/工具表单-私董会运作方法论.xlsx'
    wb.save(output_path)
    print(f"文件已生成：{output_path}")
    return output_path


if __name__ == '__main__':
    main()

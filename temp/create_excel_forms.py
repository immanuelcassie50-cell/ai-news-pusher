# -*- coding: utf-8 -*-
"""
创建数智时代的品牌创新课程配套表单 (F1-F10)
使用 openpyxl 生成专业格式的 Excel 文件
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter

# 配色方案
HEADER_BG = "264653"  # 深青色背景
HEADER_FONT = "FFFFFF"  # 白色字体
ALT_ROW_BG_1 = "FFFFFF"  # 白色
ALT_ROW_BG_2 = "E8F4F8"  # 浅蓝灰色
BORDER_COLOR = "B0B0B0"  # 边框颜色
INPUT_FONT_COLOR = "0000FF"  # 蓝色输入字体

# 课程标题
COURSE_TITLE = "数智时代的品牌创新"

def create_header_fill():
    """创建深青色表头填充"""
    return PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")

def create_alt_fill_1():
    """创建白色交替行填充"""
    return PatternFill(start_color=ALT_ROW_BG_1, end_color=ALT_ROW_BG_1, fill_type="solid")

def create_alt_fill_2():
    """创建浅蓝灰色交替行填充"""
    return PatternFill(start_color=ALT_ROW_BG_2, end_color=ALT_ROW_BG_2, fill_type="solid")

def create_border():
    """创建细边框"""
    thin = Side(style='thin', color=BORDER_COLOR)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_cell(cell, value):
    """设置表头单元格样式"""
    cell.value = value
    cell.font = Font(name='微软雅黑', size=11, bold=True, color=HEADER_FONT)
    cell.fill = create_header_fill()
    cell.border = create_border()
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_data_cell(cell, value, is_alt_row=False, is_input=False):
    """设置数据单元格样式"""
    cell.value = value
    cell.font = Font(name='微软雅黑', size=10, color=INPUT_FONT_COLOR if is_input else "000000")
    cell.fill = create_alt_fill_2() if is_alt_row else create_alt_fill_1()
    cell.border = create_border()
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_label_cell(cell, value, is_alt_row=False):
    """设置标签单元格样式"""
    cell.value = value
    cell.font = Font(name='微软雅黑', size=10, bold=False, color="000000")
    cell.fill = create_alt_fill_2() if is_alt_row else create_alt_fill_1()
    cell.border = create_border()
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_column_widths(ws, widths):
    """设置列宽"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def add_course_title(ws, num_sheets):
    """添加课程标题到工作簿"""
    pass  # 在每个sheet中单独添加

def create_f1():
    """F1_品牌数智化诊断表.xlsx"""
    wb = Workbook()

    # 删除默认sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ===== Sheet1: 诊断问卷 =====
    ws1 = wb.create_sheet("诊断问卷")
    set_column_widths(ws1, [5, 40, 15, 15, 15, 30])

    # 标题行
    ws1.merge_cells('A1:F1')
    style_header_cell(ws1['A1'], f"{COURSE_TITLE} - 品牌数智化成熟度诊断问卷")
    ws1.row_dimensions[1].height = 30

    # 列标题
    headers = ["序号", "诊断问题", "完全不符合(1分)", "部分符合(2分)", "基本符合(3分)", "完全符合(4分)"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws1.cell(row=2, column=col), header)
    ws1.row_dimensions[2].height = 25

    # 诊断问题数据
    questions = [
        "1. 企业已建立完整的数字化品牌战略规划",
        "2. 品牌数据实现了跨部门整合与共享",
        "3. AI技术被广泛应用于品牌内容创作",
        "4. 建立了实时品牌舆情监测系统",
        "5. 用户画像体系支持精准营销投放",
        "6. 品牌资产管理实现数字化、可视化",
        "7. 建立了品牌体验的线上线下一体化",
        "8. 数据驱动决策成为品牌管理常态"
    ]

    for i, q in enumerate(questions):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws1.cell(row=row, column=1), i+1, is_alt)
        style_label_cell(ws1.cell(row=row, column=2), q, is_alt)
        for col in range(3, 7):
            style_data_cell(ws1.cell(row=row, column=col), "", is_alt, is_input=True)

    # ===== Sheet2: 成熟度评级 =====
    ws2 = wb.create_sheet("成熟度评级")
    set_column_widths(ws2, [20, 25, 25, 25])

    ws2.merge_cells('A1:D1')
    style_header_cell(ws2['A1'], f"{COURSE_TITLE} - 品牌数智化成熟度评级")
    ws2.row_dimensions[1].height = 30

    headers = ["成熟度等级", "分值区间", "特征描述", "改进建议"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws2.cell(row=2, column=col), header)

    levels = [
        ("初始级", "8-14分", "品牌数字化处于萌芽阶段，流程分散、数据孤岛", "建立基础数据收集体系，明确数字化优先级"),
        ("发展级", "15-21分", "已具备一定数字化基础，但应用深度不足", "深化关键环节数字化，建立数据标准"),
        ("成熟级", "22-26分", "数字化体系较为完善，部分环节实现智能化", "推进AI应用集成，优化用户体验"),
        ("领先级", "27-32分", "品牌数智化全面领先，创新能力突出", "持续创新，输出行业最佳实践")
    ]

    for i, (level, score, desc, suggestion) in enumerate(levels):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws2.cell(row=row, column=1), level, is_alt)
        style_data_cell(ws2.cell(row=row, column=2), score, is_alt)
        style_label_cell(ws2.cell(row=row, column=3), desc, is_alt)
        style_label_cell(ws2.cell(row=row, column=4), suggestion, is_alt)

    # ===== Sheet3: 改进建议 =====
    ws3 = wb.create_sheet("改进建议")
    set_column_widths(ws3, [5, 30, 40, 25])

    ws3.merge_cells('A1:D1')
    style_header_cell(ws3['A1'], f"{COURSE_TITLE} - 改进建议清单")
    ws3.row_dimensions[1].height = 30

    headers = ["序号", "改进领域", "具体措施", "优先级"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws3.cell(row=2, column=col), header)

    suggestions = [
        ("1", "数据基础设施", "部署CDP客户数据平台，打通全渠道用户数据", "高"),
        ("2", "AI内容生产", "引入AIGC工具，建立品牌内容素材库与智能生成流程", "高"),
        ("3", "舆情监测", "搭建实时社交舆情监测与预警系统", "中"),
        ("4", "用户洞察", "构建360度用户画像，实现精准个性化推荐", "高"),
        ("5", "团队能力", "开展品牌数智化专项培训，提升团队数字素养", "中"),
    ]

    for i, (num, area, measure, priority) in enumerate(suggestions):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws3.cell(row=row, column=1), num, is_alt)
        style_label_cell(ws3.cell(row=row, column=2), area, is_alt)
        style_label_cell(ws3.cell(row=row, column=3), measure, is_alt)
        style_data_cell(ws3.cell(row=row, column=4), priority, is_alt)

    wb.save("D:/新课开发/经营/系列/05_数智时代的品牌创新/配套表单Excel/F1_品牌数智化诊断表.xlsx")
    print("F1_品牌数智化诊断表.xlsx 创建完成")

def create_f2():
    """F2_用户画像分析表.xlsx"""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ===== Sheet1: 用户画像四问 =====
    ws1 = wb.create_sheet("用户画像四问")
    set_column_widths(ws1, [5, 25, 35, 35])

    ws1.merge_cells('A1:D1')
    style_header_cell(ws1['A1'], f"{COURSE_TITLE} - 用户画像四问")
    ws1.row_dimensions[1].height = 30

    headers = ["序号", "问题维度", "问题内容", "洞察要点"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws1.cell(row=2, column=col), header)

    questions = [
        ("1", "用户是谁", "目标用户的基本属性特征是什么？", "年龄、职业、地域、收入水平"),
        ("2", "用户需求", "用户核心痛点和需求是什么？", "功能需求、情感需求、社交需求"),
        ("3", "用户行为", "用户在哪些场景下与品牌互动？", "触点偏好、使用习惯、决策路径"),
        ("4", "用户价值", "用户的终身价值和对品牌的贡献？", "消费频次、客单价、推荐意愿"),
    ]

    for i, (num, dimension, content, insight) in enumerate(questions):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws1.cell(row=row, column=1), num, is_alt)
        style_label_cell(ws1.cell(row=row, column=2), dimension, is_alt)
        style_label_cell(ws1.cell(row=row, column=3), content, is_alt)
        style_label_cell(ws1.cell(row=row, column=4), insight, is_alt)

    # ===== Sheet2: 画像汇总 =====
    ws2 = wb.create_sheet("画像汇总")
    set_column_widths(ws2, [8, 20, 20, 20, 25, 15])

    ws2.merge_cells('A1:F1')
    style_header_cell(ws2['A1'], f"{COURSE_TITLE} - 用户画像汇总")
    ws2.row_dimensions[1].height = 30

    headers = ["画像ID", "用户类型", "核心特征", "需求痛点", "典型场景", "占比"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws2.cell(row=2, column=col), header)

    personas = [
        ("P001", "品质精英型", "追求高品质、高格调，愿意为溢价付费", "担心品质不达预期", "高端定制消费场景", "15%"),
        ("P002", "理性务实型", "注重性价比，功能导向，货比三家", "担心买贵了", "促销抢购、比价决策", "35%"),
        ("P003", "社交活跃型", "爱分享、重口碑，决策受KOL影响大", "担心踩雷丢脸", "社交电商、直播带货", "30%"),
        ("P004", "忠诚会员型", "高复购、高粘性，注重会员权益", "担心权益缩水", "会员日、积分兑换", "20%"),
    ]

    for i, persona in enumerate(personas):
        row = i + 3
        is_alt = i % 2 == 1
        for col, val in enumerate(persona, 1):
            style_label_cell(ws2.cell(row=row, column=col), val, is_alt)

    # ===== Sheet3: 行为数据 =====
    ws3 = wb.create_sheet("行为数据")
    set_column_widths(ws3, [8, 20, 15, 15, 15, 20])

    ws3.merge_cells('A1:F1')
    style_header_cell(ws3['A1'], f"{COURSE_TITLE} - 用户行为数据")
    ws3.row_dimensions[1].height = 30

    headers = ["用户ID", "用户类型", "月均访问", "平均停留", "转化率", "最近购买"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws3.cell(row=2, column=col), header)

    behaviors = [
        ("U001", "品质精英型", "8次", "5分钟", "12%", "3天前"),
        ("U002", "理性务实型", "15次", "2分钟", "25%", "1天前"),
        ("U003", "社交活跃型", "20次", "8分钟", "18%", "7天前"),
        ("U004", "忠诚会员型", "12次", "4分钟", "30%", "昨天"),
    ]

    for i, behavior in enumerate(behaviors):
        row = i + 3
        is_alt = i % 2 == 1
        for col, val in enumerate(behavior, 1):
            style_data_cell(ws3.cell(row=row, column=col), val, is_alt)

    wb.save("D:/新课开发/经营/系列/05_数智时代的品牌创新/配套表单Excel/F2_用户画像分析表.xlsx")
    print("F2_用户画像分析表.xlsx 创建完成")

def create_f3():
    """F3_品牌内容检核表.xlsx"""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ===== Sheet1: AIGC内容检查 =====
    ws1 = wb.create_sheet("AIGC内容检查")
    set_column_widths(ws1, [5, 35, 15, 15, 30])

    ws1.merge_cells('A1:E1')
    style_header_cell(ws1['A1'], f"{COURSE_TITLE} - AIGC品牌内容质量检查标准")
    ws1.row_dimensions[1].height = 30

    headers = ["序号", "检查维度", "评分标准", "得分", "不合格原因"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws1.cell(row=2, column=col), header)

    criteria = [
        ("1", "品牌调性一致性", "内容风格是否符合品牌VI和调性要求", ""),
        ("2", "信息准确性", "产品参数、价格、促销信息是否准确无误", ""),
        ("3", "原创性检验", "通过AIGC检测工具，原创度需>70%", ""),
        ("4", "合规性审查", "是否包含虚假宣传、绝对化用语等违规内容", ""),
        ("5", "目标受众匹配", "内容是否针对目标用户画像定制", ""),
        ("6", "多模态一致性", "图文、视频、语音是否传递一致信息", ""),
    ]

    for i, (num, dimension, standard, score) in enumerate(criteria):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws1.cell(row=row, column=1), num, is_alt)
        style_label_cell(ws1.cell(row=row, column=2), dimension, is_alt)
        style_label_cell(ws1.cell(row=row, column=3), standard, is_alt)
        style_data_cell(ws1.cell(row=row, column=4), score, is_alt, is_input=True)
        style_data_cell(ws1.cell(row=row, column=5), "", is_alt, is_input=True)

    # ===== Sheet2: 发布前检查 =====
    ws2 = wb.create_sheet("发布前检查")
    set_column_widths(ws2, [5, 30, 15, 30, 15])

    ws2.merge_cells('A1:E1')
    style_header_cell(ws2['A1'], f"{COURSE_TITLE} - 内容发布前检查清单")
    ws2.row_dimensions[1].height = 30

    headers = ["序号", "检查项目", "状态", "备注", "检查人"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws2.cell(row=2, column=col), header)

    checks = [
        ("1", "品牌标识露出规范", "待检查"),
        ("2", "版权素材授权核实", "待检查"),
        ("3", "竞品对比合规性", "待检查"),
        ("4", "法务风险审核", "待检查"),
        ("5", "技术格式适配", "待检查"),
        ("6", "发布时间排期", "待检查"),
    ]

    for i, (num, item, status) in enumerate(checks):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws2.cell(row=row, column=1), num, is_alt)
        style_label_cell(ws2.cell(row=row, column=2), item, is_alt)
        style_data_cell(ws2.cell(row=row, column=3), status, is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=4), "", is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=5), "", is_alt, is_input=True)

    # ===== Sheet3: 效果追踪 =====
    ws3 = wb.create_sheet("效果追踪")
    set_column_widths(ws3, [8, 30, 15, 15, 15, 20])

    ws3.merge_cells('A1:F1')
    style_header_cell(ws3['A1'], f"{COURSE_TITLE} - 内容效果追踪")
    ws3.row_dimensions[1].height = 30

    headers = ["内容ID", "内容标题", "曝光量", "互动率", "转化率", "发布时间"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws3.cell(row=2, column=col), header)

    contents = [
        ("C001", "新品上市推广图文", "50,000", "3.2%", "1.8%", "2024-01-15"),
        ("C002", "春节营销短视频", "120,000", "5.6%", "2.5%", "2024-01-20"),
        ("C003", "AIGC科普系列推文", "35,000", "4.1%", "0.9%", "2024-01-25"),
    ]

    for i, content in enumerate(contents):
        row = i + 3
        is_alt = i % 2 == 1
        for col, val in enumerate(content, 1):
            style_data_cell(ws3.cell(row=row, column=col), val, is_alt)

    wb.save("D:/新课开发/经营/系列/05_数智时代的品牌创新/配套表单Excel/F3_品牌内容检核表.xlsx")
    print("F3_品牌内容检核表.xlsx 创建完成")

def create_f4():
    """F4_品牌资产评估表.xlsx"""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ===== Sheet1: 用户关系诊断 =====
    ws1 = wb.create_sheet("用户关系诊断")
    set_column_widths(ws1, [5, 30, 20, 20, 25])

    ws1.merge_cells('A1:E1')
    style_header_cell(ws1['A1'], f"{COURSE_TITLE} - 用户品牌关系诊断")
    ws1.row_dimensions[1].height = 30

    headers = ["序号", "关系维度", "健康度评分", "行业均值", "差距分析"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws1.cell(row=2, column=col), header)

    relations = [
        ("1", "品牌知名度", "", "75分", ""),
        ("2", "品牌美誉度", "", "68分", ""),
        ("3", "品牌忠诚度", "", "62分", ""),
        ("4", "品牌联想度", "", "55分", ""),
        ("5", "品牌感知质量", "", "70分", ""),
    ]

    for i, relation in enumerate(relations):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws1.cell(row=row, column=1), relation[0], is_alt)
        style_label_cell(ws1.cell(row=row, column=2), relation[1], is_alt)
        style_data_cell(ws1.cell(row=row, column=3), relation[2], is_alt, is_input=True)
        style_data_cell(ws1.cell(row=row, column=4), relation[3], is_alt)
        style_data_cell(ws1.cell(row=row, column=5), relation[4], is_alt, is_input=True)

    # ===== Sheet2: 金字塔评估 =====
    ws2 = wb.create_sheet("金字塔评估")
    set_column_widths(ws2, [20, 35, 35])

    ws2.merge_cells('A1:C1')
    style_header_cell(ws2['A1'], f"{COURSE_TITLE} - 品牌资产金字塔评估")
    ws2.row_dimensions[1].height = 30

    headers = ["金字塔层级", "核心要素", "评估要点"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws2.cell(row=2, column=col), header)

    pyramid = [
        ("顶层", "品牌共鸣", "用户与品牌的情感连接深度、品牌社区活跃度、用户推荐意愿"),
        ("中层", "品牌判断", "感知质量、品牌信誉、品牌联想的一致性和强度"),
        ("底层", "品牌表现", "品牌知名度、品牌再购意愿、价格敏感度"),
    ]

    for i, (level, element, points) in enumerate(pyramid):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws2.cell(row=row, column=1), level, is_alt)
        style_label_cell(ws2.cell(row=row, column=2), element, is_alt)
        style_label_cell(ws2.cell(row=row, column=3), points, is_alt)

    # ===== Sheet3: 综合评分 =====
    ws3 = wb.create_sheet("综合评分")
    set_column_widths(ws3, [25, 20, 20, 25])

    ws3.merge_cells('A1:D1')
    style_header_cell(ws3['A1'], f"{COURSE_TITLE} - 品牌资产综合评分")
    ws3.row_dimensions[1].height = 30

    headers = ["评估维度", "权重", "得分", "加权得分"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws3.cell(row=2, column=col), header)

    scores = [
        ("财务维度", "30%", "", ""),
        ("市场维度", "25%", "", ""),
        ("客户维度", "25%", "", ""),
        ("潜力维度", "20%", "", ""),
    ]

    for i, (dim, weight, score, weighted) in enumerate(scores):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws3.cell(row=row, column=1), dim, is_alt)
        style_data_cell(ws3.cell(row=row, column=2), weight, is_alt)
        style_data_cell(ws3.cell(row=row, column=3), score, is_alt, is_input=True)
        style_data_cell(ws3.cell(row=row, column=4), weighted, is_alt)

    # 总分行
    row = 7
    style_label_cell(ws3.cell(row=row, column=1), "综合得分", True)
    style_data_cell(ws3.cell(row=row, column=2), "100%", True)
    style_data_cell(ws3.cell(row=row, column=3), "", True, is_input=True)
    style_data_cell(ws3.cell(row=row, column=4), "", True)

    wb.save("D:/新课开发/经营/系列/05_数智时代的品牌创新/配套表单Excel/F4_品牌资产评估表.xlsx")
    print("F4_品牌资产评估表.xlsx 创建完成")

def create_f5():
    """F5_触达效果追踪表.xlsx"""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ===== Sheet1: 渠道效果 =====
    ws1 = wb.create_sheet("渠道效果")
    set_column_widths(ws1, [8, 20, 15, 15, 15, 20])

    ws1.merge_cells('A1:F1')
    style_header_cell(ws1['A1'], f"{COURSE_TITLE} - 触达渠道效果分析")
    ws1.row_dimensions[1].height = 30

    headers = ["渠道ID", "渠道名称", "触达人数", "互动率", "转化率", "成本效率"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws1.cell(row=2, column=col), header)

    channels = [
        ("CH01", "微信公众号", "25,000", "4.5%", "2.1%", "高"),
        ("CH02", "抖音短视频", "80,000", "6.2%", "1.8%", "中"),
        ("CH03", "小红书", "35,000", "8.3%", "3.5%", "高"),
        ("CH04", "微博", "15,000", "2.8%", "0.9%", "低"),
    ]

    for i, channel in enumerate(channels):
        row = i + 3
        is_alt = i % 2 == 1
        for col, val in enumerate(channel, 1):
            style_data_cell(ws1.cell(row=row, column=col), val, is_alt)

    # ===== Sheet2: 时间效果 =====
    ws2 = wb.create_sheet("时间效果")
    set_column_widths(ws2, [8, 20, 15, 15, 15, 20])

    ws2.merge_cells('A1:F1')
    style_header_cell(ws2['A1'], f"{COURSE_TITLE} - 触达时间效果分析")
    ws2.row_dimensions[1].height = 30

    headers = ["时段ID", "时间段", "触达人数", "打开率", "转化率", "最优内容类型"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws2.cell(row=2, column=col), header)

    times = [
        ("T01", "早间 7:00-9:00", "12,000", "22%", "1.5%", "资讯类"),
        ("T02", "午间 12:00-14:00", "28,000", "28%", "2.3%", "促销类"),
        ("T03", "晚间 19:00-21:00", "45,000", "35%", "3.1%", "娱乐类"),
        ("T04", "周末全天", "38,000", "30%", "2.8%", "深度内容"),
    ]

    for i, time_slot in enumerate(times):
        row = i + 3
        is_alt = i % 2 == 1
        for col, val in enumerate(time_slot, 1):
            style_data_cell(ws2.cell(row=row, column=col), val, is_alt)

    # ===== Sheet3: 内容效果 =====
    ws3 = wb.create_sheet("内容效果")
    set_column_widths(ws3, [8, 25, 15, 15, 15, 15])

    ws3.merge_cells('A1:F1')
    style_header_cell(ws3['A1'], f"{COURSE_TITLE} - 内容形式效果分析")
    ws3.row_dimensions[1].height = 30

    headers = ["内容ID", "内容形式", "触达人数", "完播率/阅读率", "互动率", "种草转化"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws3.cell(row=2, column=col), header)

    contents = [
        ("CF01", "15秒短视频", "65,000", "45%", "5.2%", "2.1%"),
        ("CF02", "长图图文", "32,000", "68%", "3.8%", "1.5%"),
        ("CF03", "直播带货", "52,000", "N/A", "8.5%", "5.2%"),
        ("CF04", "互动H5", "18,000", "72%", "12.3%", "3.2%"),
    ]

    for i, content in enumerate(contents):
        row = i + 3
        is_alt = i % 2 == 1
        for col, val in enumerate(content, 1):
            style_data_cell(ws3.cell(row=row, column=col), val, is_alt)

    wb.save("D:/新课开发/经营/系列/05_数智时代的品牌创新/配套表单Excel/F5_触达效果追踪表.xlsx")
    print("F5_触达效果追踪表.xlsx 创建完成")

def create_f6():
    """F6_数智化路线图.xlsx"""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ===== Sheet1: 现状评估 =====
    ws1 = wb.create_sheet("现状评估")
    set_column_widths(ws1, [5, 30, 20, 25, 20])

    ws1.merge_cells('A1:E1')
    style_header_cell(ws1['A1'], f"{COURSE_TITLE} - 数智化现状评估")
    ws1.row_dimensions[1].height = 30

    headers = ["序号", "能力领域", "当前成熟度", "差距分析", "改进优先级"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws1.cell(row=2, column=col), header)

    assessment = [
        ("1", "数据采集与整合", "2级", "跨部门数据孤岛严重", "高"),
        ("2", "AI应用能力", "2级", "应用场景有限，未形成规模化", "高"),
        ("3", "用户洞察", "3级", "画像维度需更丰富", "中"),
        ("4", "内容智能化", "2级", "AIGC使用率不足30%", "高"),
        ("5", "触达精准度", "3级", "投放ROI有待提升", "中"),
    ]

    for i, item in enumerate(assessment):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws1.cell(row=row, column=1), item[0], is_alt)
        style_label_cell(ws1.cell(row=row, column=2), item[1], is_alt)
        style_data_cell(ws1.cell(row=row, column=3), item[2], is_alt, is_input=True)
        style_label_cell(ws1.cell(row=row, column=4), item[3], is_alt)
        style_data_cell(ws1.cell(row=row, column=5), item[4], is_alt, is_input=True)

    # ===== Sheet2: 目标设定 =====
    ws2 = wb.create_sheet("目标设定")
    set_column_widths(ws2, [5, 25, 25, 25, 20])

    ws2.merge_cells('A1:E1')
    style_header_cell(ws2['A1'], f"{COURSE_TITLE} - 数智化目标设定")
    ws2.row_dimensions[1].height = 30

    headers = ["序号", "目标领域", "现状", "90天后目标", "衡量指标"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws2.cell(row=2, column=col), header)

    goals = [
        ("1", "数据整合", "数据孤岛", "建立统一数据中台", "数据打通率>80%"),
        ("2", "AI应用", "单一场景", "覆盖5大核心场景", "AI采用率>60%"),
        ("3", "用户洞察", "基础画像", "实时动态画像体系", "画像完整度>90%"),
        ("4", "内容生产", "人工为主", "AIGC占比>50%", "内容产出效率提升3倍"),
    ]

    for i, goal in enumerate(goals):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws2.cell(row=row, column=1), goal[0], is_alt)
        style_label_cell(ws2.cell(row=row, column=2), goal[1], is_alt)
        style_label_cell(ws2.cell(row=row, column=3), goal[2], is_alt)
        style_label_cell(ws2.cell(row=row, column=4), goal[3], is_alt)
        style_data_cell(ws2.cell(row=row, column=5), goal[4], is_alt, is_input=True)

    # ===== Sheet3: 90天计划 =====
    ws3 = wb.create_sheet("90天计划")
    set_column_widths(ws3, [5, 25, 15, 15, 30, 15])

    ws3.merge_cells('A1:F1')
    style_header_cell(ws3['A1'], f"{COURSE_TITLE} - 90天数智化行动计划")
    ws3.row_dimensions[1].height = 30

    headers = ["周次", "行动项目", "开始日期", "结束日期", "关键里程碑", "负责人"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws3.cell(row=2, column=col), header)

    plan = [
        ("第1-2周", "数据现状审计与方案设计", "2024-02-01", "2024-02-15", "完成数据资产梳理报告", "数据负责人"),
        ("第3-4周", "CDP平台选型与部署", "2024-02-16", "2024-02-29", "CDP系统上线试运行", "技术负责人"),
        ("第5-6周", "AIGC工具链整合", "2024-03-01", "2024-03-15", "AIGC工作流打通", "产品负责人"),
        ("第7-8周", "用户洞察体系升级", "2024-03-16", "2024-03-31", "新画像标签体系上线", "运营负责人"),
        ("第9-12周", "全渠道数据贯通与迭代", "2024-04-01", "2024-04-30", "数据中台正式启用", "项目负责人"),
    ]

    for i, week in enumerate(plan):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws3.cell(row=row, column=1), week[0], is_alt)
        style_label_cell(ws3.cell(row=row, column=2), week[1], is_alt)
        style_data_cell(ws3.cell(row=row, column=3), week[2], is_alt, is_input=True)
        style_data_cell(ws3.cell(row=row, column=4), week[3], is_alt, is_input=True)
        style_label_cell(ws3.cell(row=row, column=5), week[4], is_alt)
        style_data_cell(ws3.cell(row=row, column=6), week[5], is_alt, is_input=True)

    wb.save("D:/新课开发/经营/系列/05_数智时代的品牌创新/配套表单Excel/F6_数智化路线图.xlsx")
    print("F6_数智化路线图.xlsx 创建完成")

def create_f7():
    """F7_案例分析表.xlsx"""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ===== Sheet1: 案例背景 =====
    ws1 = wb.create_sheet("案例背景")
    set_column_widths(ws1, [15, 40, 35])

    ws1.merge_cells('A1:C1')
    style_header_cell(ws1['A1'], f"{COURSE_TITLE} - 案例背景")
    ws1.row_dimensions[1].height = 30

    headers = ["背景维度", "内容", "补充说明"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws1.cell(row=2, column=col), header)

    background = [
        ("企业名称", "某知名消费品牌（匿名处理）", "国内头部消费品牌，年营收50亿+"),
        ("行业背景", "快消品行业竞争激烈", "同质化严重，差异化难"),
        ("核心挑战", "品牌老化，Z世代认知度下降", "传统营销打法效果递减"),
        ("转型契机", "2023年启动品牌数智化升级项目", "新任CMO主导变革"),
    ]

    for i, (dim, content, note) in enumerate(background):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws1.cell(row=row, column=1), dim, is_alt)
        style_label_cell(ws1.cell(row=row, column=2), content, is_alt)
        style_label_cell(ws1.cell(row=row, column=3), note, is_alt)

    # ===== Sheet2: 核心问题 =====
    ws2 = wb.create_sheet("核心问题")
    set_column_widths(ws2, [5, 35, 35, 20])

    ws2.merge_cells('A1:D1')
    style_header_cell(ws2['A1'], f"{COURSE_TITLE} - 核心问题分析")
    ws2.row_dimensions[1].height = 30

    headers = ["序号", "问题描述", "影响分析", "紧急程度"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws2.cell(row=2, column=col), header)

    problems = [
        ("1", "用户数据分散在12个系统中，无法形成统一视图", "无法实现精准营销，资源浪费严重", "高"),
        ("2", "内容生产依赖人工，产能瓶颈明显", "爆款率低，内容迭代慢", "高"),
        ("3", "品牌触点割裂，用户体验不连贯", "品牌认知模糊，忠诚度下降", "中"),
        ("4", "决策依赖经验，数据驱动能力弱", "市场反应滞后，错过窗口期", "中"),
    ]

    for i, problem in enumerate(problems):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws2.cell(row=row, column=1), problem[0], is_alt)
        style_label_cell(ws2.cell(row=row, column=2), problem[1], is_alt)
        style_label_cell(ws2.cell(row=row, column=3), problem[2], is_alt)
        style_data_cell(ws2.cell(row=row, column=4), problem[3], is_alt, is_input=True)

    # ===== Sheet3: 解决方案 =====
    ws3 = wb.create_sheet("解决方案")
    set_column_widths(ws3, [5, 30, 35, 25])

    ws3.merge_cells('A1:D1')
    style_header_cell(ws3['A1'], f"{COURSE_TITLE} - 解决方案")
    ws3.row_dimensions[1].height = 30

    headers = ["序号", "解决方案", "实施路径", "关键资源"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws3.cell(row=2, column=col), header)

    solutions = [
        ("1", "构建CDP用户数据中台", "整合全渠道数据，建立统一用户ID体系", "CDP供应商、数据团队"),
        ("2", "引入AIGC内容生产体系", "部署多模态AI工具，建立内容素材库", "AI工具授权、创意团队培训"),
        ("3", "打造全渠道品牌触点矩阵", "线上线下融合，优化用户体验旅程", "IT系统对接、用户运营团队"),
        ("4", "建立数据驱动决策体系", "搭建品牌健康度仪表盘，实现实时洞察", "BI工具、数据分析团队"),
    ]

    for i, solution in enumerate(solutions):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws3.cell(row=row, column=1), solution[0], is_alt)
        style_label_cell(ws3.cell(row=row, column=2), solution[1], is_alt)
        style_label_cell(ws3.cell(row=row, column=3), solution[2], is_alt)
        style_data_cell(ws3.cell(row=row, column=4), solution[3], is_alt, is_input=True)

    # ===== Sheet4: 效果评估 =====
    ws4 = wb.create_sheet("效果评估")
    set_column_widths(ws4, [20, 25, 25, 20])

    ws4.merge_cells('A1:D1')
    style_header_cell(ws4['A1'], f"{COURSE_TITLE} - 效果评估")
    ws4.row_dimensions[1].height = 30

    headers = ["评估维度", "改善前", "改善后", "提升幅度"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws4.cell(row=2, column=col), header)

    results = [
        ("用户数据整合率", "15%", "85%", "+70pp"),
        ("内容生产效率", "100篇/月", "350篇/月", "+250%"),
        ("营销投放ROI", "1.2x", "2.8x", "+133%"),
        ("用户复购率", "18%", "32%", "+14pp"),
        ("品牌NPS评分", "25分", "48分", "+23分"),
    ]

    for i, result in enumerate(results):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws4.cell(row=row, column=1), result[0], is_alt)
        style_data_cell(ws4.cell(row=row, column=2), result[1], is_alt)
        style_data_cell(ws4.cell(row=row, column=3), result[2], is_alt)
        style_data_cell(ws4.cell(row=row, column=4), result[3], is_alt, is_input=True)

    wb.save("D:/新课开发/经营/系列/05_数智时代的品牌创新/配套表单Excel/F7_案例分析表.xlsx")
    print("F7_案例分析表.xlsx 创建完成")

def create_f8():
    """F8_学员评估表.xlsx"""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ===== Sheet1: 课前测试 =====
    ws1 = wb.create_sheet("课前测试")
    set_column_widths(ws1, [8, 20, 15, 15, 35])

    ws1.merge_cells('A1:E1')
    style_header_cell(ws1['A1'], f"{COURSE_TITLE} - 学员课前能力测试")
    ws1.row_dimensions[1].height = 30

    headers = ["学员ID", "姓名", "品牌认知", "数字化基础", "备注"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws1.cell(row=2, column=col), header)

    students = [
        ("S001", "张明", "3分", "2分", "市场部骨干"),
        ("S002", "李华", "2分", "3分", "销售部转岗"),
        ("S003", "王芳", "4分", "2分", "品牌经理"),
        ("S004", "陈伟", "2分", "4分", "IT部门"),
    ]

    for i, student in enumerate(students):
        row = i + 3
        is_alt = i % 2 == 1
        style_data_cell(ws1.cell(row=row, column=1), student[0], is_alt, is_input=True)
        style_label_cell(ws1.cell(row=row, column=2), student[1], is_alt)
        style_data_cell(ws1.cell(row=row, column=3), student[2], is_alt, is_input=True)
        style_data_cell(ws1.cell(row=row, column=4), student[3], is_alt, is_input=True)
        style_data_cell(ws1.cell(row=row, column=5), student[4], is_alt, is_input=True)

    # ===== Sheet2: 课堂表现 =====
    ws2 = wb.create_sheet("课堂表现")
    set_column_widths(ws2, [8, 20, 15, 15, 15, 20])

    ws2.merge_cells('A1:F1')
    style_header_cell(ws2['A1'], f"{COURSE_TITLE} - 学员课堂表现评估")
    ws2.row_dimensions[1].height = 30

    headers = ["学员ID", "姓名", "参与度", "案例分析", "团队协作", "综合评级"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws2.cell(row=2, column=col), header)

    performance = [
        ("S001", "张明", "积极", "优秀", "良好", "A"),
        ("S002", "李华", "一般", "良好", "优秀", "B+"),
        ("S003", "王芳", "非常积极", "优秀", "优秀", "A+"),
        ("S004", "陈伟", "积极", "良好", "良好", "B"),
    ]

    for i, perf in enumerate(performance):
        row = i + 3
        is_alt = i % 2 == 1
        style_data_cell(ws2.cell(row=row, column=1), perf[0], is_alt, is_input=True)
        style_label_cell(ws2.cell(row=row, column=2), perf[1], is_alt)
        style_data_cell(ws2.cell(row=row, column=3), perf[2], is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=4), perf[3], is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=5), perf[4], is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=6), perf[5], is_alt, is_input=True)

    # ===== Sheet3: 课后实践 =====
    ws3 = wb.create_sheet("课后实践")
    set_column_widths(ws3, [8, 20, 20, 20, 20])

    ws3.merge_cells('A1:E1')
    style_header_cell(ws3['A1'], f"{COURSE_TITLE} - 学员课后实践成果")
    ws3.row_dimensions[1].height = 30

    headers = ["学员ID", "姓名", "实践项目", "完成度", "应用效果"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws3.cell(row=2, column=col), header)

    practices = [
        ("S001", "张明", "品牌数智化升级方案", "90%", "已提交管理层"),
        ("S002", "李华", "用户画像分析报告", "85%", "部门内部分享"),
        ("S003", "王芳", "数智化路线图规划", "95%", "纳入年度计划"),
        ("S004", "陈伟", "AIGC工具选型报告", "80%", "技术选型中"),
    ]

    for i, practice in enumerate(practices):
        row = i + 3
        is_alt = i % 2 == 1
        style_data_cell(ws3.cell(row=row, column=1), practice[0], is_alt, is_input=True)
        style_label_cell(ws3.cell(row=row, column=2), practice[1], is_alt)
        style_data_cell(ws3.cell(row=row, column=3), practice[2], is_alt, is_input=True)
        style_data_cell(ws3.cell(row=row, column=4), practice[3], is_alt, is_input=True)
        style_data_cell(ws3.cell(row=row, column=5), practice[4], is_alt, is_input=True)

    # ===== Sheet4: 综合评定 =====
    ws4 = wb.create_sheet("综合评定")
    set_column_widths(ws4, [8, 20, 15, 15, 15, 15, 15])

    ws4.merge_cells('A1:G1')
    style_header_cell(ws4['A1'], f"{COURSE_TITLE} - 学员综合评定")
    ws4.row_dimensions[1].height = 30

    headers = ["学员ID", "姓名", "课前测试", "课堂表现", "课后实践", "综合得分", "等级"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws4.cell(row=2, column=col), header)

    evaluations = [
        ("S001", "张明", "65分", "85分", "90分", "", ""),
        ("S002", "李华", "60分", "75分", "82分", "", ""),
        ("S003", "王芳", "75分", "95分", "92分", "", ""),
        ("S004", "陈伟", "70分", "78分", "78分", "", ""),
    ]

    for i, eval_data in enumerate(evaluations):
        row = i + 3
        is_alt = i % 2 == 1
        style_data_cell(ws4.cell(row=row, column=1), eval_data[0], is_alt, is_input=True)
        style_label_cell(ws4.cell(row=row, column=2), eval_data[1], is_alt)
        style_data_cell(ws4.cell(row=row, column=3), eval_data[2], is_alt)
        style_data_cell(ws4.cell(row=row, column=4), eval_data[3], is_alt)
        style_data_cell(ws4.cell(row=row, column=5), eval_data[4], is_alt)
        style_data_cell(ws4.cell(row=row, column=6), eval_data[5], is_alt, is_input=True)
        style_data_cell(ws4.cell(row=row, column=7), eval_data[6], is_alt, is_input=True)

    wb.save("D:/新课开发/经营/系列/05_数智时代的品牌创新/配套表单Excel/F8_学员评估表.xlsx")
    print("F8_学员评估表.xlsx 创建完成")

def create_f9():
    """F9_培训效果追踪表.xlsx"""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ===== Sheet1: 行为改变追踪 =====
    ws1 = wb.create_sheet("行为改变追踪")
    set_column_widths(ws1, [8, 25, 20, 20, 20])

    ws1.merge_cells('A1:E1')
    style_header_cell(ws1['A1'], f"{COURSE_TITLE} - 培训后行为改变追踪")
    ws1.row_dimensions[1].height = 30

    headers = ["学员ID", "关键行为指标", "培训前", "培训后30天", "改变程度"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws1.cell(row=2, column=col), header)

    behaviors = [
        ("S001", "数据驱动决策频次", "2次/周", "5次/周", "+150%"),
        ("S002", "AIGC工具使用率", "10%", "45%", "+35pp"),
        ("S003", "用户画像应用场景", "1个", "4个", "+3个"),
        ("S004", "跨部门数据共享", "偶尔", "常规化", "显著改善"),
    ]

    for i, behavior in enumerate(behaviors):
        row = i + 3
        is_alt = i % 2 == 1
        style_data_cell(ws1.cell(row=row, column=1), behavior[0], is_alt, is_input=True)
        style_label_cell(ws1.cell(row=row, column=2), behavior[1], is_alt)
        style_data_cell(ws1.cell(row=row, column=3), behavior[2], is_alt, is_input=True)
        style_data_cell(ws1.cell(row=row, column=4), behavior[3], is_alt, is_input=True)
        style_data_cell(ws1.cell(row=row, column=5), behavior[4], is_alt, is_input=True)

    # ===== Sheet2: KPI变化 =====
    ws2 = wb.create_sheet("KPI变化")
    set_column_widths(ws2, [20, 20, 20, 20, 20])

    ws2.merge_cells('A1:E1')
    style_header_cell(ws2['A1'], f"{COURSE_TITLE} - 培训后KPI变化追踪")
    ws2.row_dimensions[1].height = 30

    headers = ["KPI指标", "培训前数值", "培训后30天", "培训后90天", "变化趋势"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws2.cell(row=2, column=col), header)

    kpis = [
        ("营销活动准备周期", "14天", "10天", "7天", "下降50%"),
        ("内容产出效率", "50篇/月", "80篇/月", "120篇/月", "+140%"),
        ("用户洞察报告产出", "1份/季度", "1份/月", "1份/两周", "效率提升6倍"),
        ("数据驱动决策占比", "20%", "40%", "65%", "+45pp"),
    ]

    for i, kpi in enumerate(kpis):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws2.cell(row=row, column=1), kpi[0], is_alt)
        style_data_cell(ws2.cell(row=row, column=2), kpi[1], is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=3), kpi[2], is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=4), kpi[3], is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=5), kpi[4], is_alt, is_input=True)

    # ===== Sheet3: 投资回报 =====
    ws3 = wb.create_sheet("投资回报")
    set_column_widths(ws3, [25, 20, 20, 20, 20])

    ws3.merge_cells('A1:E1')
    style_header_cell(ws3['A1'], f"{COURSE_TITLE} - 培训投资回报计算")
    ws3.row_dimensions[1].height = 30

    headers = ["回报项目", "计算方式", "估算金额", "计算依据", "备注"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws3.cell(row=2, column=col), header)

    roi_items = [
        ("效率提升节省人力", "时间节省 × 人力成本", "15万元/年", "每周节省10小时 × 52周 × 300元/小时", ""),
        ("决策质量提升价值", "避免错误决策损失", "20万元/年", "预估减少3次重大决策失误", ""),
        ("内容产出增量价值", "AIGC增量内容价值", "30万元/年", "每月多产出60篇 × 500元/篇", ""),
        ("培训总投入成本", "培训费用+时间成本", "-8万元", "外部培训5万+内部时间3万", ""),
    ]

    for i, item in enumerate(roi_items):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws3.cell(row=row, column=1), item[0], is_alt)
        style_label_cell(ws3.cell(row=row, column=2), item[1], is_alt)
        style_data_cell(ws3.cell(row=row, column=3), item[2], is_alt, is_input=True)
        style_label_cell(ws3.cell(row=row, column=4), item[3], is_alt)
        style_data_cell(ws3.cell(row=row, column=5), item[4], is_alt, is_input=True)

    # ROI汇总行
    row = 7
    style_label_cell(ws3.cell(row=row, column=1), "预估ROI", True)
    style_data_cell(ws3.cell(row=row, column=2), "(收益-成本)/成本", True)
    style_data_cell(ws3.cell(row=row, column=3), "575%", True, is_input=True)
    style_data_cell(ws3.cell(row=row, column=4), "(65万-8万)/8万", True)
    style_data_cell(ws3.cell(row=row, column=5), "", True, is_input=True)

    wb.save("D:/新课开发/经营/系列/05_数智时代的品牌创新/配套表单Excel/F9_培训效果追踪表.xlsx")
    print("F9_培训效果追踪表.xlsx 创建完成")

def create_f10():
    """F10_工具使用记录表.xlsx"""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ===== Sheet1: 使用记录 =====
    ws1 = wb.create_sheet("使用记录")
    set_column_widths(ws1, [8, 20, 15, 15, 15, 25])

    ws1.merge_cells('A1:F1')
    style_header_cell(ws1['A1'], f"{COURSE_TITLE} - 工具使用记录")
    ws1.row_dimensions[1].height = 30

    headers = ["记录ID", "工具名称", "使用日期", "使用时长", "使用场景", "主要功能"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws1.cell(row=2, column=col), header)

    logs = [
        ("L001", "ChatGPT", "2024-01-15", "2小时", "内容文案撰写", "产品卖点文案生成"),
        ("L002", "Midjourney", "2024-01-16", "1.5小时", "创意素材制作", "品牌主视觉设计"),
        ("L003", "剪映", "2024-01-18", "3小时", "短视频剪辑", "营销视频制作"),
        ("L004", "Tableau", "2024-01-20", "2小时", "数据可视化", "用户数据看板"),
    ]

    for i, log in enumerate(logs):
        row = i + 3
        is_alt = i % 2 == 1
        style_data_cell(ws1.cell(row=row, column=1), log[0], is_alt, is_input=True)
        style_data_cell(ws1.cell(row=row, column=2), log[1], is_alt, is_input=True)
        style_data_cell(ws1.cell(row=row, column=3), log[2], is_alt, is_input=True)
        style_data_cell(ws1.cell(row=row, column=4), log[3], is_alt, is_input=True)
        style_data_cell(ws1.cell(row=row, column=5), log[4], is_alt, is_input=True)
        style_label_cell(ws1.cell(row=row, column=6), log[5], is_alt)

    # ===== Sheet2: 效果反馈 =====
    ws2 = wb.create_sheet("效果反馈")
    set_column_widths(ws2, [8, 20, 15, 15, 30])

    ws2.merge_cells('A1:E1')
    style_header_cell(ws2['A1'], f"{COURSE_TITLE} - 工具效果反馈")
    ws2.row_dimensions[1].height = 30

    headers = ["工具ID", "工具名称", "效率提升", "质量评分", "反馈意见"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws2.cell(row=2, column=col), header)

    feedbacks = [
        ("T01", "ChatGPT", "+40%", "4分/5分", "文案效率大幅提升，需人工润色"),
        ("T02", "Midjourney", "+60%", "3.5分/5分", "创意素材质量不错，版权需注意"),
        ("T03", "剪映", "+50%", "4.5分/5分", "操作简便，功能满足需求"),
        ("T04", "Tableau", "+30%", "4分/5分", "可视化效果好，学习曲线较陡"),
    ]

    for i, feedback in enumerate(feedbacks):
        row = i + 3
        is_alt = i % 2 == 1
        style_data_cell(ws2.cell(row=row, column=1), feedback[0], is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=2), feedback[1], is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=3), feedback[2], is_alt, is_input=True)
        style_data_cell(ws2.cell(row=row, column=4), feedback[3], is_alt, is_input=True)
        style_label_cell(ws2.cell(row=row, column=5), feedback[4], is_alt)

    # ===== Sheet3: 优化建议 =====
    ws3 = wb.create_sheet("优化建议")
    set_column_widths(ws3, [5, 20, 20, 25, 20])

    ws3.merge_cells('A1:E1')
    style_header_cell(ws3['A1'], f"{COURSE_TITLE} - 工具使用优化建议")
    ws3.row_dimensions[1].height = 30

    headers = ["序号", "工具名称", "问题描述", "优化建议", "优先级"]
    for col, header in enumerate(headers, 1):
        style_header_cell(ws3.cell(row=2, column=col), header)

    suggestions = [
        ("1", "ChatGPT", "提示词模板分散", "建立品牌提示词模板库，统一管理", "高"),
        ("2", "Midjourney", "版权风险", "明确商业使用边界，建立审核流程", "高"),
        ("3", "剪映", "团队协作", "建立共享素材库，支持团队协同", "中"),
        ("4", "Tableau", "学习成本", "组织内部培训，编写操作手册", "中"),
    ]

    for i, suggestion in enumerate(suggestions):
        row = i + 3
        is_alt = i % 2 == 1
        style_label_cell(ws3.cell(row=row, column=1), suggestion[0], is_alt)
        style_data_cell(ws3.cell(row=row, column=2), suggestion[1], is_alt, is_input=True)
        style_label_cell(ws3.cell(row=row, column=3), suggestion[2], is_alt)
        style_label_cell(ws3.cell(row=row, column=4), suggestion[3], is_alt)
        style_data_cell(ws3.cell(row=row, column=5), suggestion[4], is_alt, is_input=True)

    wb.save("D:/新课开发/经营/系列/05_数智时代的品牌创新/配套表单Excel/F10_工具使用记录表.xlsx")
    print("F10_工具使用记录表.xlsx 创建完成")

if __name__ == "__main__":
    print("开始创建Excel表单...")
    create_f1()
    create_f2()
    create_f3()
    create_f4()
    create_f5()
    create_f6()
    create_f7()
    create_f8()
    create_f9()
    create_f10()
    print("\n全部10个Excel表单创建完成！")

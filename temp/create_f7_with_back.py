# -*- coding: utf-8 -*-
"""
Create F7_发展对话记录表.xlsx with front and back side content
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# Color scheme
RED_ACCENT = "B81025"
GRAY_DARK = "4A4A4A"
LIGHT_BG = "F5F5F5"
WHITE = "FFFFFF"

def create_border(style='thin'):
    side = Side(style=style, color="000000")
    return Border(left=side, right=side, top=side, bottom=side)

def set_cell_style(cell, font_size=11, bold=False, font_color="000000",
                   bg_color=None, align_h="left", align_v="center",
                   wrap_text=False, border=True, font_name="微软雅黑"):
    cell.font = Font(name=font_name, size=font_size, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap_text)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    if border:
        cell.border = create_border()

def merge_and_style(ws, start_row, start_col, end_row, end_col, value,
                    font_size=11, bold=False, font_color="000000",
                    bg_color=None, align_h="left", align_v="center",
                    wrap_text=False, border=True, font_name="微软雅黑"):
    ws.merge_cells(start_row=start_row, start_column=start_col,
                  end_row=end_row, end_column=end_col)
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = value
    set_cell_style(cell, font_size=font_size, bold=bold, font_color=font_color,
                   bg_color=bg_color, align_h=align_h, align_v=align_v,
                   wrap_text=wrap_text, border=border, font_name=font_name)
    return cell

def set_column_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_f7_development_dialogue_record():
    wb = Workbook()
    ws = wb.active
    ws.title = "发展对话记录"

    # Page setup - A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    # Column widths
    widths = {
        'A': 3, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 16, 'H': 16, 'I': 16, 'J': 16, 'K': 16, 'L': 3
    }
    set_column_widths(ws, widths)

    current_row = 1

    # ==================== FRONT SIDE ====================

    # Title
    merge_and_style(ws, current_row, 2, current_row, 11, "发展对话记录表",
                    font_size=20, bold=True, font_color=RED_ACCENT,
                    bg_color=LIGHT_BG, align_h="center", align_v="center")
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    # Subtitle
    merge_and_style(ws, current_row, 2, current_row, 11, "A4单张 | 每次对话后填写 | 记录和跟进",
                    font_size=11, bold=False, font_color=GRAY_DARK,
                    bg_color=LIGHT_BG, align_h="center", align_v="center")
    ws.row_dimensions[current_row].height = 22
    current_row += 2

    # Section 1: 对话基本信息
    merge_and_style(ws, current_row, 2, current_row, 11, "对话基本信息",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Basic info fields
    info_fields = [
        ("员工姓名:", ""), ("岗位:", ""), ("直接经理:", ""),
        ("对话日期:", ""), ("对话类型:", ""), ("第几次对话:", "")
    ]

    for i in range(0, len(info_fields), 2):
        field1 = info_fields[i]
        ws.cell(row=current_row, column=2).value = field1[0]
        set_cell_style(ws.cell(row=current_row, column=2), font_size=11, bold=True,
                       font_color=GRAY_DARK, bg_color=LIGHT_BG, border=True)
        ws.cell(row=current_row, column=3).value = field1[1]
        set_cell_style(ws.cell(row=current_row, column=3), font_size=11, border=True)

        if i + 1 < len(info_fields):
            field2 = info_fields[i + 1]
            ws.cell(row=current_row, column=5).value = field2[0]
            set_cell_style(ws.cell(row=current_row, column=5), font_size=11, bold=True,
                           font_color=GRAY_DARK, bg_color=LIGHT_BG, border=True)
            ws.cell(row=current_row, column=6).value = field2[1]
            set_cell_style(ws.cell(row=current_row, column=6), font_size=11, border=True)

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    current_row += 1

    # Section 2: 对话摘要
    merge_and_style(ws, current_row, 2, current_row, 11, "对话摘要",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Summary text area
    merge_and_style(ws, current_row, 2, current_row + 4, 11, "",
                    font_size=11, border=True)
    ws.row_dimensions[current_row].height = 100
    current_row += 5

    current_row += 1

    # Section 3: 达成的共识
    merge_and_style(ws, current_row, 2, current_row, 11, "达成的共识",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Consensus fields
    consensus_fields = [
        ("员工的发展目标:", 3),
        ("双方认同的优势:", 2),
        ("双方认同的发展机会:", 2),
        ("下一步行动计划:", 3)
    ]

    for field_name, row_span in consensus_fields:
        merge_and_style(ws, current_row, 2, current_row, 2, field_name,
                        font_size=11, bold=True, font_color=GRAY_DARK,
                        bg_color=LIGHT_BG, align_h="left", border=True)
        merge_and_style(ws, current_row, 3, current_row + row_span - 1, 11, "",
                        font_size=11, border=True)
        ws.row_dimensions[current_row].height = 20 * row_span
        current_row += row_span

    current_row += 1

    # Section 4: 下一步行动
    merge_and_style(ws, current_row, 2, current_row, 11, "下一步行动",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Action table header
    action_headers = ["角色", "行动项", "截止日期"]
    action_cols = [2, 4, 11]
    for idx, header in enumerate(action_headers):
        end_col = action_cols[idx + 1] - 1 if idx + 1 < len(action_cols) else 11
        cell = ws.cell(row=current_row, column=action_cols[idx])
        ws.merge_cells(start_row=current_row, start_column=action_cols[idx],
                       end_row=current_row, end_column=end_col)
        cell.value = header
        set_cell_style(cell, font_size=11, bold=True, font_color=WHITE,
                       bg_color=RED_ACCENT, align_h="center", border=True)
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # Action roles
    action_roles = ["员工行动项", "经理行动项", "需要的支持资源"]
    for role in action_roles:
        ws.merge_cells(start_row=current_row, start_column=2,
                       end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=2)
        cell.value = role
        set_cell_style(cell, font_size=11, bold=True, font_color=GRAY_DARK,
                       bg_color=LIGHT_BG, align_h="left", border=True)

        ws.merge_cells(start_row=current_row, start_column=4,
                       end_row=current_row, end_column=10)
        cell = ws.cell(row=current_row, column=4)
        set_cell_style(cell, font_size=11, border=True)

        ws.cell(row=current_row, column=11).value = ""
        set_cell_style(ws.cell(row=current_row, column=11), font_size=11, border=True)

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    current_row += 1

    # Section 5: 下次检查点
    merge_and_style(ws, current_row, 2, current_row, 11, "下次检查点",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Checkpoint fields
    checkpoint_fields = [
        ("日期:", ""), ("重点关注:", ""), ("届时讨论的问题:", "")
    ]

    for field_name, val in checkpoint_fields:
        ws.cell(row=current_row, column=2).value = field_name
        set_cell_style(ws.cell(row=current_row, column=2), font_size=11, bold=True,
                       font_color=GRAY_DARK, bg_color=LIGHT_BG, border=True)
        ws.merge_cells(start_row=current_row, start_column=3,
                       end_row=current_row, end_column=11)
        cell = ws.cell(row=current_row, column=3)
        cell.value = val
        set_cell_style(cell, font_size=11, border=True)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    front_side_end_row = current_row

    # ==================== BACK SIDE ====================
    current_row += 2

    # Back side title
    merge_and_style(ws, current_row, 2, current_row, 11, "背面内容",
                    font_size=16, bold=True, font_color=WHITE,
                    bg_color=RED_ACCENT, align_h="center", align_v="center")
    ws.row_dimensions[current_row].height = 32
    current_row += 2

    # Section B1: 对话质量自检清单
    merge_and_style(ws, current_row, 2, current_row, 11, "对话质量自检清单",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Checklist phases
    phases = [
        ("对话前准备", 4),
        ("对话中执行", 5),
        ("对话后跟进", 4)
    ]

    for phase_name, row_span in phases:
        merge_and_style(ws, current_row, 2, current_row + row_span - 1, 2, phase_name,
                        font_size=11, bold=True, font_color=GRAY_DARK,
                        bg_color=LIGHT_BG, align_h="center", border=True)
        merge_and_style(ws, current_row, 3, current_row + row_span - 1, 11, "",
                        font_size=10, border=True)
        ws.row_dimensions[current_row].height = 20 * row_span
        current_row += row_span

    current_row += 1

    # Section B2: 常见对话场景应对参考
    merge_and_style(ws, current_row, 2, current_row, 11, "常见对话场景应对参考",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Scenario headers
    scenario_headers = ["场景", "应对要点"]
    scenario_cols = [2, 5]
    for idx, header in enumerate(scenario_headers):
        end_col = scenario_cols[idx + 1] - 1 if idx + 1 < len(scenario_cols) else 11
        cell = ws.cell(row=current_row, column=scenario_cols[idx])
        ws.merge_cells(start_row=current_row, start_column=scenario_cols[idx],
                       end_row=current_row, end_column=end_col)
        cell.value = header
        set_cell_style(cell, font_size=11, bold=True, font_color=WHITE,
                       bg_color=RED_ACCENT, align_h="center", border=True)
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # 6 scenarios
    scenarios = [
        ("员工表现出抵触情绪", "保持冷静,倾听担忧,共同寻找解决方案"),
        ("员工目标过于远大", "分解目标,设定阶段性里程碑"),
        ("员工缺乏自信", "强调过往成就,提供支持资源"),
        ("经理时间紧张", "提前预约,明确议程,高效利用时间"),
        ("意见分歧较大", "聚焦共同目标,寻求第三方视角"),
        ("员工进展顺利", "肯定鼓励,讨论拓展可能性")
    ]

    for scenario_name, guidance in scenarios:
        ws.merge_cells(start_row=current_row, start_column=2,
                       end_row=current_row, end_column=4)
        cell = ws.cell(row=current_row, column=2)
        cell.value = scenario_name
        set_cell_style(cell, font_size=10, bold=True, font_color=GRAY_DARK,
                       bg_color=LIGHT_BG, align_h="left", border=True)

        ws.merge_cells(start_row=current_row, start_column=5,
                       end_row=current_row, end_column=11)
        cell = ws.cell(row=current_row, column=5)
        cell.value = guidance
        set_cell_style(cell, font_size=10, align_h="left", border=True)

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    current_row += 1

    # Section B3: 员工反馈收集方式
    merge_and_style(ws, current_row, 2, current_row, 11, "员工反馈收集方式",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    feedback_methods = [
        ("即时反馈", "对话结束后填写简短反馈表"),
        ("定期调研", "季度匿名调研收集整体满意度"),
        ("非正式沟通", "日常工作中的开放式对话"),
        ("小组讨论", "团队会议中的发展话题讨论")
    ]

    for i, (method, desc) in enumerate(feedback_methods):
        ws.cell(row=current_row, column=2).value = f"{i+1}. {method}"
        set_cell_style(ws.cell(row=current_row, column=2), font_size=11, bold=True,
                       font_color=GRAY_DARK, bg_color=LIGHT_BG, border=True)
        ws.merge_cells(start_row=current_row, start_column=3,
                       end_row=current_row, end_column=11)
        cell = ws.cell(row=current_row, column=3)
        cell.value = desc
        set_cell_style(cell, font_size=11, border=True)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    current_row += 1

    # Section B4: 后续跟进提醒机制
    merge_and_style(ws, current_row, 2, current_row, 11, "后续跟进提醒机制",
                    font_size=14, bold=True, font_color=WHITE,
                    bg_color=GRAY_DARK, align_h="left", align_v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    followup_items = [
        ("周检视", "每周一封邮件确认本周进展"),
        ("月回顾", "每月一次面对面或视频交流"),
        ("季度评估", "每季度正式绩效发展对话"),
        ("年度复盘", "年度总结与下一年计划制定")
    ]

    for i, (item, desc) in enumerate(followup_items):
        ws.cell(row=current_row, column=2).value = f"{i+1}. {item}"
        set_cell_style(ws.cell(row=current_row, column=2), font_size=11, bold=True,
                       font_color=GRAY_DARK, bg_color=LIGHT_BG, border=True)
        ws.merge_cells(start_row=current_row, start_column=3,
                       end_row=current_row, end_column=11)
        cell = ws.cell(row=current_row, column=3)
        cell.value = desc
        set_cell_style(cell, font_size=11, border=True)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # Print area
    ws.print_area = f"A1:L{current_row}"

    output_path = r"D:/新课开发/HR/培训/10_员工发展导向：把学习数据变成成长路径而不是完课记录/Excel工具表单/F7_发展对话记录表.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")

if __name__ == "__main__":
    import os

    output_dir = r"D:/新课开发/HR/培训/10_员工发展导向：把学习数据变成成长路径而不是完课记录/Excel工具表单"
    os.makedirs(output_dir, exist_ok=True)

    create_f7_development_dialogue_record()

    print("F7 Excel file created successfully!")

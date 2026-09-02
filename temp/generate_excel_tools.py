# -*- coding: utf-8 -*-
"""
工具表单Excel生成脚本
将10个工具表单从Markdown转换为Excel格式
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import os

# 配色方案
COLORS = {
    'primary': '2b2d42',      # 深蓝灰 - 表头背景
    'secondary': '6b7280',    # 中灰 - 次要文字
    'accent': 'dc2626',       # 红色 - 强调
    'light': 'e5e7eb',        # 浅灰 - 斑马纹
    'bg': 'f9fafb',          # 白灰 - 背景
    'white': 'FFFFFF',        # 白色
    'yellow': 'fef3c7',       # 浅黄色 - 示例行背景
    'orange': 'fed7aa',       # 橙色 - 高风险
    'green': 'd1fae5',        # 绿色 - 低风险
    'red': 'fee2e2',          # 红色 - 高风险背景
}

def get_border(style='thin'):
    """获取边框样式"""
    side = Side(style=style, color='000000')
    return Border(left=side, right=side, top=side, bottom=side)

def set_header_style(cell):
    """设置表头样式"""
    cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = get_border()

def set_cell_style(cell, is_zebra=False, is_example=False, is_header=False):
    """设置单元格样式"""
    if is_header:
        set_header_style(cell)
        return

    cell.font = Font(name='微软雅黑', size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = get_border()

    if is_example:
        cell.fill = PatternFill(start_color=COLORS['yellow'], end_color=COLORS['yellow'], fill_type='solid')
    elif is_zebra:
        cell.fill = PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid')

def add_dropdown_validation(ws, cell_range, options):
    """添加下拉数据验证"""
    options_str = ','.join(options)
    dv = DataValidation(type='list', formula1=f'"{options_str}"', allow_blank=True)
    dv.error = '请从下拉列表中选择'
    dv.errorTitle = '无效输入'
    ws.add_data_validation(dv)
    dv.add(cell_range)

def create_instruction_sheet(wb, tool_name, usage, instructions):
    """创建说明Sheet"""
    ws = wb.create_sheet(title='使用说明')
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80

    # 标题
    ws['A1'] = tool_name
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=14, color=COLORS['primary'])
    ws.merge_cells('A1:B1')

    row = 3
    # 用途说明
    ws[f'A{row}'] = '工具用途'
    ws[f'A{row}'].font = Font(name='微软雅黑', bold=True, size=11)
    set_header_style(ws[f'A{row}'])
    ws[f'B{row}'] = usage
    ws[f'B{row}'].font = Font(name='微软雅黑', size=10)
    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 40

    row += 1
    ws[f'A{row}'] = '使用说明'
    ws[f'A{row}'].font = Font(name='微软雅黑', bold=True, size=11)
    set_header_style(ws[f'A{row}'])
    ws[f'B{row}'] = instructions
    ws[f'B{row}'].font = Font(name='微软雅黑', size=10)
    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 150

    return ws

def create_tool01():
    """工具01：抵制信号诊断表"""
    wb = Workbook()
    ws = wb.active
    ws.title = '诊断表'

    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15

    # 标题区
    ws['A1'] = '抵制信号诊断表'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 30

    ws['A2'] = '评估日期：'
    ws['B2'] = '__________'
    ws['C2'] = '评估人：'
    ws['D2'] = '__________'
    ws['E2'] = '被评估部门：'
    ws['F2'] = '__________'

    # 维度选择
    ws['A4'] = '诊断维度'
    set_header_style(ws['A4'])
    ws['B4'] = '数量型抵制（担心工作量增加）'
    set_header_style(ws['B4'])
    ws['C4'] = '□ 选中'
    ws['D4'] = '短期型抵制（担心短期利益受损）'
    set_header_style(ws['D4'])
    ws['E4'] = '□ 选中'
    ws['F4'] = '控制型抵制（感到自主权被削弱）'
    set_header_style(ws['F4'])
    ws['G4'] = '□ 选中'

    # 评分表头
    headers = ['序号', '诊断问题', '从未如此(0)', '偶尔如此(1)', '经常如此(2)', '总是如此(3)', '得分']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        set_header_style(cell)

    # 示例数据 - 数量型抵制
    example_data_q1 = [
        ('Q1', '员工在会议中频繁抱怨新系统操作复杂、耗时较长', '', '√', '', '', ''),
        ('Q2', '跨部门协作时，对方以"不熟悉新流程"为由延迟响应', '', '', '√', '', ''),
        ('Q3', '员工提交的工作量明显低于变革前同期水平', '', '√', '', '', ''),
        ('Q4', '员工私下讨论时表达"用旧系统更顺手"等观点', '', '', '√', '', ''),
        ('Q5', '培训出勤率低于预期（实际60%，目标90%）', '', '', '√', '', ''),
        ('Q6', '员工倾向于用旧系统完成工作而非学习新系统', '', '√', '', '', ''),
    ]

    row = 6
    for i, data in enumerate(example_data_q1):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        # 添加得分公式
        score_cell = ws.cell(row=row, column=7)
        score_cell.value = f'=IF(D{row}="√",1,0)+IF(E{row}="√",2,0)+IF(F{row}="√",3,0)'
        score_cell.font = Font(name='微软雅黑', size=10)
        score_cell.border = get_border()
        row += 1

    # 小计行
    ws.cell(row=row, column=2).value = '维度小计'
    ws.cell(row=row, column=2).font = Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=row, column=2).border = get_border()
    ws.cell(row=row, column=7).value = f'=SUM(G6:G{row-1})'
    ws.cell(row=row, column=7).font = Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=row, column=7).border = get_border()
    row += 1

    # 风险等级公式
    score_result_row = row
    ws.cell(row=row, column=2).value = '数量型抵制得分'
    ws.cell(row=row, column=7).value = f'=G{row-1}'
    ws.cell(row=row, column=7).font = Font(name='微软雅黑', bold=True, size=11, color=COLORS['accent'])
    row += 1

    ws.cell(row=row, column=2).value = '风险等级'
    ws.cell(row=row, column=7).value = f'=IF(G{score_result_row}>16,"高风险",IF(G{score_result_row}>8,"中风险","低风险"))'
    ws.cell(row=row, column=7).font = Font(name='微软雅黑', bold=True, size=11)
    row += 2

    # 空白模板提示
    ws.cell(row=row, column=1).value = '--- 以下为空白模板 ---'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', italic=True, size=10, color=COLORS['secondary'])
    ws.merge_cells(f'A{row}:G{row}')
    row += 1

    # 空白模板表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    # 空白行
    for i in range(6):
        is_zebra = (i % 2 == 1)
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.value = ''
            set_cell_style(cell, is_zebra=is_zebra)
        # 得分公式
        score_cell = ws.cell(row=row, column=7)
        score_cell.value = f'=IF(D{row}="√",1,0)+IF(E{row}="√",2,0)+IF(F{row}="√",3,0)'
        score_cell.font = Font(name='微软雅黑', size=10)
        score_cell.border = get_border()
        row += 1

    # 添加下拉验证
    add_dropdown_validation(ws, 'C6:F30', ['√', ''])

    # 综合评级区
    row += 1
    ws.cell(row=row, column=1).value = '综合评级'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:G{row}')
    row += 1

    summary_headers = ['评估项目', '数量型', '短期型', '控制型', '综合得分']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    ws.cell(row=row, column=1).value = '本次得分'
    ws.cell(row=row, column=2).value = '____分'
    ws.cell(row=row, column=3).value = '____分'
    ws.cell(row=row, column=4).value = '____分'
    ws.cell(row=row, column=5).value = '=SUM(B{0}:D{0})'.format(row)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = get_border()
        ws.cell(row=row, column=col).font = Font(name='微软雅黑', size=10)

    # 创建说明Sheet
    usage = '本工具用于系统识别和评估组织成员在变革过程中表现出的抵制信号，通过三个维度的结构化诊断，帮助精准定位阻力来源、评估风险等级，并制定针对性的应对策略。'
    instructions = '''评分标准：
从未如此(0分)、偶尔如此(1分)、经常如此(2分)、总是如此(3分)

风险等级判定：
- 低风险（绿灯）：综合得分0-24分 → 以观察和预防为主
- 中风险（黄灯）：综合得分25-40分 → 需制定针对性干预计划
- 高风险（橙灯）：综合得分41-60分 → 需立即启动深度干预
- 极高风险（红灯）：综合得分61-90分 → 需上级介入并重新评估变革策略

使用时机：
- 变革启动后的第2周、第4周、第8周进行定期诊断
- 遇到重大阻力事件时进行即时诊断
- 变革阶段性评估时使用'''
    create_instruction_sheet(wb, '抵制信号诊断表', usage, instructions)

    output_path = r'D:/新课开发/变革管理/06-考核与激励机制重设计：让新流程有人真心愿意执行/完整课程包/06-工具表单/Excel/工具01-抵制信号诊断表.xlsx'
    wb.save(output_path)
    print(f'已生成: {output_path}')

def create_tool02():
    """工具02：双轨考核设计模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = '双轨考核设计'

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20

    # 标题
    ws['A1'] = '双轨考核设计模板'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 30

    # 权重过渡原则表
    ws['A3'] = '权重过渡原则'
    ws['A3'].font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells('A3:E3')

    headers = ['阶段', '时间节点', '旧指标权重', '新指标权重', '调整触发条件']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        set_header_style(cell)

    transition_data = [
        ('启动期（0-3月）', '第1-3月', '80%', '20%', '基准：全员培训完成率≥90%'),
        ('加速期（4-6月）', '第4-6月', '70%', '30%', '前提：启动期双轨运行无重大异常'),
        ('深化期（7-12月）', '第7-12月', '60%', '40%', '前提：加速期新指标达成率≥70%'),
        ('稳定期（13-18月）', '第13-18月', '50%', '50%', '前提：深化期新指标达成率≥75%'),
        ('完全切换（18月后）', '第19月起', '0%', '100%', '前提：稳定期新指标达成率≥80%'),
    ]

    for i, row_data in enumerate(transition_data):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=5+i, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra)

    row = 11

    # 旧指标保留清单
    ws.cell(row=row, column=1).value = '旧指标保留清单'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    headers = ['序号', '指标名称', '指标定义', '计算方式', '保留理由']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    # 示例数据
    old_indicators = [
        (1, '销售额达成率', '实际销售额/目标销售额×100%', '财务取数', '保障基本业务稳定'),
        (2, '客户满意度', '季度客户NPS评分', '问卷调研', '持续关注服务质量'),
        (3, '成本控制率', '实际成本/预算成本×100%', '财务取数', '防止资源浪费'),
        (4, '团队流失率', '主动离职人数/期末人数×100%', 'HR系统', '维护团队稳定性'),
    ]

    for i, data in enumerate(old_indicators):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 空白行
    for i in range(3):
        is_zebra = True
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.value = ''
            set_cell_style(cell, is_zebra=is_zebra)
        row += 1

    # 权重汇总公式
    ws.cell(row=row, column=2).value = '旧指标小计权重'
    ws.cell(row=row, column=3).value = f'=SUM(C{row-7}:C{row-1})'
    ws.cell(row=row, column=3).font = Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=row, column=3).border = get_border()
    row += 2

    # 新指标引入清单
    ws.cell(row=row, column=1).value = '新指标引入清单'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    headers = ['序号', '指标名称', '指标定义', '计算方式', '引入目的']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    new_indicators = [
        (1, '数字化工具使用率', '员工使用新系统完成关键业务的占比', '系统埋点统计', '推动新工具落地'),
        (2, '新流程合规率', '关键业务按照新SOP执行的比例', '审计抽查+系统记录', '确保新流程真正被执行'),
        (3, '跨部门协作评分', '合作部门对本部门的协作满意度', '季度互评', '促进组织壁垒打破'),
        (4, '创新提案数量', '员工提交的流程改进建议被采纳的数量', '改进系统统计', '激励员工参与变革'),
    ]

    for i, data in enumerate(new_indicators):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 空白行
    for i in range(3):
        is_zebra = True
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.value = ''
            set_cell_style(cell, is_zebra=is_zebra)
        row += 1

    # 权重汇总
    ws.cell(row=row, column=2).value = '新指标小计权重'
    ws.cell(row=row, column=3).value = f'=SUM(C{row-7}:C{row-1})'
    ws.cell(row=row, column=3).font = Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=row, column=3).border = get_border()

    # 安全网设计区
    row += 2
    ws.cell(row=row, column=1).value = '过渡期安全网设计'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    headers = ['保障维度', '机制设计', '适用对象', '有效期']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    safety_net = [
        ('收入兜底', '总体收入不低于变革前90%', '全体员工', '前6个月'),
        ('达标门槛', '新指标达成60%即视为合格', '全体员工', '前12个月'),
        ('缓冲机制', '新指标未达标不扣罚', '全体员工', '前3个月'),
        ('专项奖励', '新指标单项冠军可获得额外奖励', '各岗位排名前20%', '全程'),
    ]

    for i, data in enumerate(safety_net):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 创建说明Sheet
    usage = '本工具用于解决变革过程中的考核过渡难题，帮助企业在保留稳定旧指标的同时，渐进式引入新指标，通过权重动态调整实现平稳过渡。'
    instructions = '''填写步骤：
1. 盘点现状：梳理现有考核指标，识别需要保留的"稳定器"指标
2. 定义目标：明确新业务模式下的核心能力要求，设计新指标
3. 权重规划：根据变革阶段设定双轨权重比例
4. 安全网设计：设计过渡期保障机制，兜底员工基本收益
5. 动态调整：设定权重调整触发条件和时间节点

安全网原则：
- 过渡期保障：员工总体收入不低于变革前的90%
- 达标门槛：新指标达成60%即可视为合格
- 缓冲机制：设置3个月缓冲期，缓冲期内即使新指标未达标也不扣罚'''
    create_instruction_sheet(wb, '双轨考核设计模板', usage, instructions)

    output_path = r'D:/新课开发/变革管理/06-考核与激励机制重设计：让新流程有人真心愿意执行/完整课程包/06-工具表单/Excel/工具02-双轨考核设计模板.xlsx'
    wb.save(output_path)
    print(f'已生成: {output_path}')

def create_tool03():
    """工具03：先行者激励包设计表"""
    wb = Workbook()
    ws = wb.active
    ws.title = '先行者激励'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # 标题
    ws['A1'] = '先行者激励包设计表'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30

    # 先行者等级标准
    ws['A3'] = '先行者等级划分标准'
    ws['A3'].font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells('A3:F3')

    headers = ['等级', '综合得分要求', '预期人数（%）', '激励预算/人']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        set_header_style(cell)

    levels = [
        ('钻石先行者', '≥90分', 'Top 1%', '30,000元'),
        ('金牌先行者', '80-89分', 'Top 5%', '15,000元'),
        ('银牌先行者', '70-79分', 'Top 10%', '8,000元'),
        ('铜牌先行者', '60-69分', 'Top 20%', '3,000元'),
    ]

    for i, data in enumerate(levels):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=5+i, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)

    row = 10

    # 识别维度评分表
    ws.cell(row=row, column=1).value = '先行者识别维度评分表'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    headers = ['员工姓名', '部门', '行为领先度', '绩效领先度', '影响领先度', '创新贡献度', '综合得分', '先行者等级']
    # 重新设置列宽
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    # 示例数据
    employees = [
        ('赵海', '基础研发部', 95, 92, 85, 88),
        ('钱琳', '产品开发部', 88, 90, 80, 75),
        ('孙磊', '测试部', 85, 85, 70, 80),
        ('周婷', '运维部', 78, 82, 75, 0),
    ]

    example_start_row = row
    for i, data in enumerate(employees):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        # 综合得分公式
        score_cell = ws.cell(row=row, column=7)
        score_cell.value = f'=AVERAGE(C{row}:F{row})'
        score_cell.font = Font(name='微软雅黑', size=10)
        score_cell.border = get_border()
        # 等级判定公式
        level_cell = ws.cell(row=row, column=8)
        level_cell.value = f'=IF(G{row}>=90,"钻石先行者",IF(G{row}>=80,"金牌先行者",IF(G{row}>=70,"银牌先行者",IF(G{row}>=60,"铜牌先行者","非先行者"))))'
        level_cell.font = Font(name='微软雅黑', size=10)
        level_cell.border = get_border()
        row += 1

    row += 1

    # 物质激励设计
    ws.cell(row=row, column=1).value = '激励包设计（物质激励）'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    headers = ['激励项目', '钻石先行者', '金牌先行者', '银牌先行者', '铜牌先行者']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    material_incentives = [
        ('一次性奖金', '20,000元', '10,000元', '5,000元', '2,000元'),
        ('专项补贴', '5,000元/季度', '3,000元/季度', '2,000元/季度', '1,000元/季度'),
        ('额外带薪假期', '5天', '3天', '2天', '1天'),
        ('年度旅游基金', '10,000元', '5,000元', '3,000元', '1,000元'),
        ('实物奖励', '高端办公装备', '品牌配件套装', '办公用品礼包', '电影票2张'),
    ]

    for i, data in enumerate(material_incentives):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 创建说明Sheet
    usage = '本工具用于识别和激励在变革过程中率先尝试、率先达标的"先行者"，通过设计系统性的激励包，形成示范效应，带动更多员工主动拥抱变革。'
    instructions = '''先行者定义：
1. 早期采纳者：在变革启动后1个月内率先完成新流程切换的个体
2. 标杆达成者：在新指标考核中持续表现优异的个体
3. 创新贡献者：为新流程优化提出有价值建议的个体

激励包设计原则：
- 即时性：激励需在行为发生后2周内兑现
- 差异化：不同贡献度匹配不同激励档次
- 可叠加：物质激励、精神激励、发展激励可同时享有'''
    create_instruction_sheet(wb, '先行者激励包设计表', usage, instructions)

    output_path = r'D:/新课开发/变革管理/06-考核与激励机制重设计：让新流程有人真心愿意执行/完整课程包/06-工具表单/Excel/工具03-先行者激励包设计表.xlsx'
    wb.save(output_path)
    print(f'已生成: {output_path}')

def create_tool04():
    """工具04：变革导向PBC模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = '变革PBC'

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25

    # 标题
    ws['A1'] = '变革导向PBC模板'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30

    # 基本信息
    ws['A3'] = '基本信息'
    ws['A3'].font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells('A3:F3')

    info_data = [
        ('姓名', ''), ('岗位', ''), ('变革项目', ''),
        ('考核周期', ''), ('直接上级', '')
    ]
    row = 4
    for i, (label, value) in enumerate(info_data):
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=10)
        ws.cell(row=row, column=1).border = get_border()
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=2).border = get_border()
        ws.merge_cells(f'B{row}:F{row}')
        row += 1

    row += 1

    # 业务目标
    ws.cell(row=row, column=1).value = '业务目标（60%）'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    headers = ['序号', '目标名称', '衡量指标', '目标值', '权重', '达成路径']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    # 示例数据
    business_goals = [
        (1, '产线升级如期完成', '升级项目里程碑达成率', '100%按时', '20%', '制定详细项目计划'),
        (2, '新产线产能达标', '产能利用率', '≥85%', '15%', '3个月内完成产能爬坡'),
        (3, '生产效率提升', '人均产出提升率', '同比+20%', '15%', '导入精益生产方法'),
        (4, '产品不良率控制', '不良率', '≤1.5%', '10%', '建立SPC质量监控系统'),
    ]

    for i, data in enumerate(business_goals):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 空白行
    for i in range(2):
        is_zebra = True
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.value = ''
            set_cell_style(cell, is_zebra=is_zebra)
        row += 1

    row += 1

    # 能力发展目标
    ws.cell(row=row, column=1).value = '能力发展目标（20%）'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    headers = ['序号', '能力名称', '当前水平', '目标水平', '提升措施', '完成时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    ability_goals = [
        (1, '智能制造技术应用', '了解（2级）', '熟练应用（4级）', '参加外部培训40学时', '2026年6月'),
        (2, '数据驱动决策', '初级（2级）', '精通（4级）', '每月分析报告10份', '2026年12月'),
        (3, '变革领导力', '基础（2级）', '熟练（4级）', '参加变革管理培训', '2026年9月'),
    ]

    for i, data in enumerate(ability_goals):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 空白行
    for i in range(2):
        is_zebra = True
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.value = ''
            set_cell_style(cell, is_zebra=is_zebra)
        row += 1

    row += 1

    # PBC综合评定表
    ws.cell(row=row, column=1).value = 'PBC综合评定表'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    headers = ['维度', '自评得分', '上级评价', '综合得分', '权重', '加权得分']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    pbc_data = [
        ('业务目标', 88, 90, '=(B{0}+C{0})/2', '60%', '=D{0}*0.6'),
        ('能力发展', 82, 85, '=(B{0}+C{0})/2', '20%', '=D{0}*0.2'),
        ('团队协作', 85, 88, '=(B{0}+C{0})/2', '20%', '=D{0}*0.2'),
        ('变革贡献加分', 0, 0, '—', '—', '+5分'),
    ]

    pbc_start_row = row
    for i, data in enumerate(pbc_data):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(value, str) and '{0}' in value:
                cell.value = value.format(row)
            else:
                cell.value = value
            set_cell_style(cell, is_example=True)
        row += 1

    # 综合得分
    ws.cell(row=row, column=1).value = 'PBC综合得分'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=11)
    ws.cell(row=row, column=6).value = f'=SUM(F{pbc_start_row}:F{row-1})'
    ws.cell(row=row, column=6).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['accent'])
    ws.cell(row=row, column=6).border = get_border()
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = get_border()
    row += 1

    # 评级公式
    ws.cell(row=row, column=1).value = 'PBC评级'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=11)
    ws.cell(row=row, column=6).value = f'=IF(F{row-1}>=90,"A级(卓越)",IF(F{row-1}>=80,"B级(优秀)",IF(F{row-1}>=70,"C级(合格)",IF(F{row-1}>=60,"D级(待改进)","E级(不合格)"))))'
    ws.cell(row=row, column=6).font = Font(name='微软雅黑', bold=True, size=11)
    ws.cell(row=row, column=6).border = get_border()

    # 创建说明Sheet
    usage = '本工具将传统的绩效承诺（PBC）框架与变革管理目标相结合，帮助员工在个人绩效承诺中明确变革相关目标。'
    instructions = '''PBC框架：
变革导向PBC = 业务目标（60%）+ 能力发展（20%）+ 团队协作（20%）

填写流程：
1. 回顾变革要求：明确本年度/季度变革核心目标
2. 分解业务目标：将变革目标转化为个人可衡量的业务指标
3. 设定能力目标：识别变革所需的胜任能力
4. 明确协作承诺：识别跨部门/跨岗位协作需求
5. 对齐上级确认：与上级进行目标对齐对话

评级标准：
- A（卓越）：90-100分
- B（优秀）：80-89分
- C（合格）：70-79分
- D（待改进）：60-69分
- E（不合格）：<60分'''
    create_instruction_sheet(wb, '变革导向PBC模板', usage, instructions)

    output_path = r'D:/新课开发/变革管理/06-考核与激励机制重设计：让新流程有人真心愿意执行/完整课程包/06-工具表单/Excel/工具04-变革导向PBC模板.xlsx'
    wb.save(output_path)
    print(f'已生成: {output_path}')

def create_tool05():
    """工具05：考核固化检查清单"""
    wb = Workbook()
    ws = wb.active
    ws.title = '固化检查清单'

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # 标题
    ws['A1'] = '考核固化检查清单'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30

    # 检查节点信息
    ws['A2'] = '检查节点：'
    ws['B2'] = '_____年_____月_____日（第___个月）'
    ws['A3'] = '检查人：'
    ws['B3'] = '________________'

    # 表头
    row = 5
    headers = ['序号', '检查项', '评定等级', '佐证说明', '主责部门', '配合部门']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    # 示例数据
    check_items = [
        ('1★', '新流程绩效指标已写入年度目标责任书', '已落实', '各部门负责人已签署2026年度目标责任书', '人力资源部', '各业务部门'),
        ('2★', '考核周期与新流程节奏匹配', '已落实', '销售部门由季度考核调整为双月考核', '人力资源部', '销售管理部'),
        ('3★', '考核数据来源已切换至新系统', '部分落实', 'CRM系统新模块已上线', '信息部', '业务运营部'),
        ('4★', '新流程执行率纳入考核权重', '已落实', '市场部已纳入权重15%', '人力资源部', '研发管理部'),
        ('5★', '跨部门协作指标已设置', '已落实', '新增"跨部门问题解决时效"指标', '综合管理部', '各部门'),
        ('6★', '变革试点团队获得差异化激励', '已落实', '试点团队额外获得项目奖金池20%', '财务管理部', '人力资源部'),
        ('7★', '负面行为处罚条款已明确', '已落实', '《员工行为准则》已修订', '法务部', '人力资源部'),
        ('8★', '考核结果应用场景明确', '已落实', '晋升、调薪、调配三挂钩制度已公布', '人力资源部', '综合管理部'),
        ('9', '新流程培训时长纳入培训考核', '已落实', '培训合格证书与转正答辩挂钩', '人力资源部', '各部门'),
        ('10★', '高层管理者变革指标纳入述职', '已落实', '副总裁及以上述职内容权重≥25%', '总经办', '人力资源部'),
    ]

    check_start_row = row
    for i, data in enumerate(check_items):
        is_zebra = (i % 2 == 1)
        is_key = '★' in str(data[0])
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if is_key and col == 2:
                cell.font = Font(name='微软雅黑', size=10, bold=True)
            else:
                set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 综合达成率公式
    ws.cell(row=row, column=2).value = '综合达成率'
    ws.cell(row=row, column=2).font = Font(name='微软雅黑', bold=True, size=11)
    ws.cell(row=row, column=3).value = f'=COUNTIF(C{check_start_row}:C{row-1},"已落实")/COUNTA(C{check_start_row}:C{row-1})'
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=3).font = Font(name='微软雅黑', bold=True, size=11, color=COLORS['accent'])
    ws.cell(row=row, column=3).border = get_border()

    # 添加下拉验证
    add_dropdown_validation(ws, f'C{check_start_row}:C{row}', ['已落实', '部分落实', '未落实', '不适用'])

    # 创建说明Sheet
    usage = '本检查清单用于确认考核制度已完整嵌入新流程各项环节，确保变革成果通过制度固化得以持续生效。'
    instructions = '''评分规则：
每项检查点按"已落实/部分落实/未落实/不适用"四级评定

通过标准：
- 关键项（★标记）必须全部"已落实"
- 总体达成率≥80%方可通过

三阶段检查时机：
- 上线后第1周：基础制度是否就位
- 上线后第1个月：考核是否实际运行
- 上线后第3个月：考核效果评估'''
    create_instruction_sheet(wb, '考核固化检查清单', usage, instructions)

    output_path = r'D:/新课开发/变革管理/06-考核与激励机制重设计：让新流程有人真心愿意执行/完整课程包/06-工具表单/Excel/工具05-考核固化检查清单.xlsx'
    wb.save(output_path)
    print(f'已生成: {output_path}')

def create_tool06():
    """工具06：个人变革行动计划表"""
    wb = Workbook()
    ws = wb.active
    ws.title = '个人行动计划'

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10

    # 标题
    ws['A1'] = '个人变革行动计划表'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 30

    # 基本信息
    ws['A2'] = '填写人：'
    ws['B2'] = '__________'
    ws['C2'] = '部门：'
    ws['D2'] = '__________'
    ws['E2'] = '岗位：'
    ws['F2'] = '__________'
    ws['G2'] = '填写日期：__________'

    row = 4

    # 现状评估
    ws.cell(row=row, column=1).value = '一、个人现状评估'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:G{row}')
    row += 1

    headers = ['评估维度', '现状描述', '变革影响分析']
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    ws.merge_cells(f'C{row}:G{row}')
    row += 1

    assessment_dims = [
        ('现有工作流程', '', ''),
        ('现有能力水平', '', ''),
        ('现有绩效结果', '', ''),
        ('心态与顾虑', '', ''),
        ('现有资源支持', '', ''),
    ]

    for i, data in enumerate(assessment_dims):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra)
        row += 1

    row += 1

    # 行动计划时间表
    ws.cell(row=row, column=1).value = '二、行动计划时间表'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:G{row}')
    row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10

    headers = ['目标', '行动步骤', '开始时间', '完成标志', '所需资源', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    ws.merge_cells(f'F{row}:G{row}')
    row += 1

    # 示例数据
    action_plan = [
        ('目标1', '参加客户分层方法论培训', '6月20日', '培训合格证书', '培训部排课', '✓'),
        ('目标1', '导出存量客户数据，梳理基础信息', '6月25日', '数据表格完整', 'CRM系统权限', '✓'),
        ('目标1', '对照分层标准完成客户分类', '7月1日', '分层结果确认', '区域负责人指导', '◐'),
        ('目标2', '注册BI系统学习账号', '7月1日', '账号激活', 'IT部门开号', '✓'),
        ('目标2', '完成《Power BI入门》课程', '7月31日', '考试成绩≥80分', '在线学习资源', '◐'),
        ('目标3', '参加跨部门协作小组启动会', '6月30日', '会议纪要确认', '项目组邀请', '✓'),
        ('目标3', '与市场部对接人建立周沟通机制', '7月6日', '首周沟通记录', '对接人确认', '◐'),
    ]

    plan_start_row = row
    for i, data in enumerate(action_plan):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 空白行
    for i in range(5):
        is_zebra = True
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.value = ''
            set_cell_style(cell, is_zebra=is_zebra)
        row += 1

    row += 1

    # 复盘记录
    ws.cell(row=row, column=1).value = '三、复盘记录'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:G{row}')
    row += 1

    headers = ['复盘节点', '复盘日期', '目标达成情况', '主要收获', '未达标原因', '下月调整']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    ws.merge_cells(f'F{row}:G{row}')
    row += 1

    for i in range(2):
        is_zebra = (i % 2 == 1)
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.value = ''
            set_cell_style(cell, is_zebra=is_zebra)
        row += 1

    # 添加状态下拉
    add_dropdown_validation(ws, f'G{plan_start_row}:G{row}', ['✓', '◐', '○'])

    # 创建说明Sheet
    usage = '本表用于帮助员工个人系统规划变革过渡期的行动路径，实现从"知道要变"到"知道怎么变"的认知与行动转化。'
    instructions = '''填写时机：
- 变革项目启动后1周内完成初版
- 变革实施中每月复盘更新

目标设定SMART原则：
- Specific（具体）：明确具体要达成什么
- Measurable（可衡量）：有可量化的标准
- Achievable（可达成）：基于资源评估可实现
- Relevant（相关）：与变革目标紧密相关
- Time-bound（有时限）：有明确的完成时间

复盘机制：
每月最后一个周五填写复盘进度'''
    create_instruction_sheet(wb, '个人变革行动计划表', usage, instructions)

    output_path = r'D:/新课开发/变革管理/06-考核与激励机制重设计：让新流程有人真心愿意执行/完整课程包/06-工具表单/Excel/工具06-个人变革行动计划表.xlsx'
    wb.save(output_path)
    print(f'已生成: {output_path}')

def create_tool07():
    """工具07：组织变革影响评估表"""
    wb = Workbook()
    ws = wb.active
    ws.title = '变革影响评估'

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15

    # 标题
    ws['A1'] = '组织变革影响评估表'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30

    # 基本信息
    ws['A2'] = '评估项目：________________'
    ws['A3'] = '评估部门：__________  评估人：__________  评估日期：__________'

    row = 5

    # 工作流程影响评估
    ws.cell(row=row, column=1).value = '一、工作流程影响评估'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:H{row}')
    row += 1

    headers = ['影响项', '影响描述', '影响程度(1-5)', '发生概率(%)', '风险等级', '应对策略', '主责部门']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    # 示例数据
    workflow_impact = [
        ('客户接待流程', '电话接单→全渠道接入→需求洞察', 4, 90, '极高', '提前发布操作手册+双轨过渡期', '销售管理部'),
        ('审批流程', '5级审批压缩为3级', 3, 75, '高', '明确新审批权限清单+授权培训', '综合管理部'),
        ('信息系统操作', 'CRM新增多个模块', 4, 80, '极高', '分模块渐进式上线+操作指南', '信息部'),
    ]

    impact_start_row = row
    for i, data in enumerate(workflow_impact):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 空白行
    for i in range(3):
        is_zebra = True
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.value = ''
            set_cell_style(cell, is_zebra=is_zebra)
        row += 1

    row += 1

    # 风险等级公式
    ws.cell(row=row, column=1).value = '综合风险等级'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=11)
    risk_formula = f'=IF(AND(C{impact_start_row}>=4,D{impact_start_row}>=80),"极高",IF(AND(C{impact_start_row}>=3,D{impact_start_row}>=70),"高","中"))'
    ws.cell(row=row, column=5).value = risk_formula
    ws.cell(row=row, column=5).font = Font(name='微软雅黑', bold=True, size=11)
    ws.cell(row=row, column=5).border = get_border()

    # 添加数据验证
    add_dropdown_validation(ws, f'C{impact_start_row}:C{row}', ['1', '2', '3', '4', '5'])
    add_dropdown_validation(ws, f'D{impact_start_row}:D{row}', ['50', '60', '70', '75', '80', '85', '90', '95'])
    add_dropdown_validation(ws, f'E{impact_start_row}:E{row}', ['低', '中', '高', '极高'])

    # 创建说明Sheet
    usage = '本表用于系统评估变革项目对组织各维度的深层影响，帮助提前识别高风险领域，制定针对性应对策略。'
    instructions = '''影响程度评分：
- 1分=无影响
- 2分=轻微影响（容易适应）
- 3分=中等影响（需要适应期）
- 4分=较大影响（需专门支持）
- 5分=重大影响（需重新配置资源）

风险等级判定矩阵：
- 极高：影响程度≥4 且 发生概率≥80%
- 高：影响程度≥3 且 发生概率≥70%
- 中：影响程度≥2 且 发生概率≥50%
- 低：影响程度≤2 或 发生概率≤50%'''
    create_instruction_sheet(wb, '组织变革影响评估表', usage, instructions)

    output_path = r'D:/新课开发/变革管理/06-考核与激励机制重设计：让新流程有人真心愿意执行/完整课程包/06-工具表单/Excel/工具07-组织变革影响评估表.xlsx'
    wb.save(output_path)
    print(f'已生成: {output_path}')

def create_tool08():
    """工具08：新旧指标对比分析表"""
    wb = Workbook()
    ws = wb.active
    ws.title = '指标对比分析'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # 标题
    ws['A1'] = '新旧指标对比分析表'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 30

    # 基本信息
    ws['A2'] = '分析项目：________________'
    ws['A3'] = '分析部门：__________  分析人：__________  分析日期：__________'

    row = 5

    # 指标对比维度总览
    ws.cell(row=row, column=1).value = '一、指标对比维度总览'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    headers = ['对比维度', '旧指标体系', '新指标体系', '变化幅度', '变化性质']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    comparison_dims = [
        ('指标结构', '单一销售额指标为主（80%权重）', '四维指标并重（业绩/客户/能力/协作）', '大幅调整', '结构性变革'),
        ('考核导向', '结果导向（只看业绩数字）', '过程+结果双导向', '中等调整', '导向优化'),
        ('数据来源', '人工统计上报（Excel填报为主）', '系统自动提取+人工确认', '大幅调整', '机制变革'),
        ('应用场景', '绩效工资计算', '绩效工资+晋升+培训+荣誉四位一体', '显著扩展', '应用深化'),
        ('激励效果', '强激励（提成比例高但方式单一）', '多元激励（物质+发展+荣誉+认可）', '中等调整', '方式优化'),
    ]

    for i, data in enumerate(comparison_dims):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    row += 1

    # 新指标体系详情
    ws.cell(row=row, column=1).value = '二、新指标体系详情'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10

    headers = ['指标维度', '指标名称', '指标定义', '计算公式', '数据来源', '权重', '目标值', '评分标准']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    new_indicators = [
        ('销售业绩（40%）', '销售额达成率', '实际销售额/目标销售额×100%', '(实际/目标)×100%', 'CRM系统', '20%', '≥100%', '达成率×20分'),
        ('销售业绩（40%）', '新客户开发数', '考核期内新增签约客户数量', '统计新签约客户数', 'CRM系统', '10%', '≥15家', '每少1家扣2分'),
        ('客户价值（20%）', '客户满意度指数', '通过标准化问卷调研获取', 'NPS×50%+满意度×50%', '客服系统', '8%', '≥75分', '满意度/100×权重分'),
        ('能力发展（20%）', '客户洞察能力认证', '方法论掌握与应用水平', '认证考试+实战案例', '培训部', '8%', '≥初级', '按认证等级赋分'),
        ('协作贡献（20%）', '跨部门协作评分', '协作质量360度评价', '关联部门平均评分', '协作系统', '10%', '≥80分', '评分/100×权重分'),
    ]

    for i, data in enumerate(new_indicators):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 空白行
    for i in range(5):
        is_zebra = True
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.value = ''
            set_cell_style(cell, is_zebra=is_zebra)
        row += 1

    row += 1

    # 指标切换影响分析
    ws.cell(row=row, column=1).value = '三、指标切换影响分析'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    headers = ['影响维度', '受益群体', '受损群体', '潜在风险', '应对策略']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    impact_analysis = [
        ('收入变化', '客户满意度高的员工（+5-15%）', '单一业绩依赖型员工（-10-25%）', '部分员工收入大幅下降', '设置6个月过渡期保底'),
        ('职业发展', '综合能力强的复合型人才', '单一业绩导向型', '晋升标准重置', '公示新晋升标准'),
        ('工作方式', '善于运用系统工具的年轻员工', '习惯传统方式的中老年员工', '学习曲线导致效率下降', '渐进式切换+充足培训'),
    ]

    for i, data in enumerate(impact_analysis):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 创建说明Sheet
    usage = '本表用于系统诊断旧指标体系的问题，并设计与验证新指标体系的科学性与适用性。'
    instructions = '''填写说明：
1. 对比维度：指标结构、考核导向、数据来源、应用场景、激励效果
2. 问题诊断：旧指标问题需深入根因，避免停留在表象描述
3. 新指标设计原则：每项设计原则需具体可操作

关键成功要素：
1. 问题诊断要深入：找到"为什么不合理"的根因
2. 设计原则要具体：每条原则必须包含"做什么"和"怎么做"
3. 影响分析要全面：分析对收入、发展、关系、文化的影响
4. 切换计划要务实：设置过渡期，避免"休克式"切换'''
    create_instruction_sheet(wb, '新旧指标对比分析表', usage, instructions)

    output_path = r'D:/新课开发/变革管理/06-考核与激励机制重设计：让新流程有人真心愿意执行/完整课程包/06-工具表单/Excel/工具08-新旧指标对比分析表.xlsx'
    wb.save(output_path)
    print(f'已生成: {output_path}')

def create_tool09():
    """工具09：激励效果追踪表"""
    wb = Workbook()
    ws = wb.active
    ws.title = '激励效果追踪'

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    # 标题
    ws['A1'] = '激励效果追踪表'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 30

    # 基本信息
    row = 3
    info_data = [
        ('追踪期间', ''), ('填写部门', ''), ('填写人', ''),
        ('审核人', ''), ('填写日期', '')
    ]
    for i, (label, value) in enumerate(info_data):
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=10)
        ws.cell(row=row, column=1).border = get_border()
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=2).border = get_border()
        row += 1

    row += 1

    # 激励包执行追踪
    ws.cell(row=row, column=1).value = '一、激励包执行追踪'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:G{row}')
    row += 1

    headers = ['激励项目', '预算投入（元）', '实际投入（元）', '执行率', '覆盖人数', '人均投入（元）', '执行进度']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    # 示例数据
    execution_data = [
        ('先行者奖金', 150000, 148500, '=C{0}/B{0}', 28, '=C{0}/E{0}', '已全额发放'),
        ('专项培训机会', 80000, 75000, '=C{0}/B{0}', 35, '=C{0}/E{0}', '培训已完成'),
        ('晋升绿色通道', 50000, 50000, '=C{0}/B{0}', 8, '=C{0}/E{0}', '3人已晋升'),
        ('荣誉表彰', 20000, 18500, '=C{0}/B{0}', 45, '=C{0}/E{0}', '表彰已完成'),
        ('弹性工作安排', 30000, 28000, '=C{0}/B{0}', 22, '=C{0}/E{0}', '持续执行中'),
    ]

    exec_start_row = row
    for i, data in enumerate(execution_data):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(value, str) and '{0}' in value:
                cell.value = value.format(row)
            else:
                cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 合计行
    ws.cell(row=row, column=1).value = '合计'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=row, column=1).border = get_border()
    ws.cell(row=row, column=2).value = f'=SUM(B{exec_start_row}:B{row-1})'
    ws.cell(row=row, column=2).font = Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=row, column=2).border = get_border()
    ws.cell(row=row, column=3).value = f'=SUM(C{exec_start_row}:C{row-1})'
    ws.cell(row=row, column=3).font = Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=row, column=3).border = get_border()
    ws.cell(row=row, column=4).value = f'=C{row}/B{row}'
    ws.cell(row=row, column=4).number_format = '0.0%'
    ws.cell(row=row, column=4).font = Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=row, column=4).border = get_border()
    row += 2

    # 激励效果关键指标
    ws.cell(row=row, column=1).value = '二、激励效果关键指标（KPI）'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:G{row}')
    row += 1

    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12

    headers = ['指标维度', '具体指标', '第一月', '第二月', '第三月', '季度汇总', '目标值', '达成率']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    kpi_data = [
        ('覆盖度', '激励覆盖员工占比', '15%', '22%', '28%', '=AVERAGE(C{0}:E{0})', '30%', '=G{0}/H{0}'),
        ('覆盖度', '部门参与率', '60%', '75%', '90%', '=E{0}', '85%', '=G{0}/H{0}'),
        ('响应度', '激励申请率', '35%', '52%', '68%', '=E{0}', '70%', '=G{0}/H{0}'),
        ('响应度', '激励知晓率', '70%', '85%', '95%', '=E{0}', '90%', '=G{0}/H{0}'),
        ('转化率', '新流程采用率', '20%', '38%', '55%', '=E{0}', '50%', '=G{0}/H{0}'),
        ('转化率', '行为改变持续率', '18%', '35%', '52%', '=E{0}', '50%', '=G{0}/H{0}'),
        ('满意度', '激励公平性评分', '3.8', '4.1', '4.3', '=AVERAGE(C{0}:E{0})', '4.0', '=G{0}/H{0}'),
        ('满意度', '激励有效性评分', '3.5', '3.9', '4.2', '=AVERAGE(C{0}:E{0})', '4.0', '=G{0}/H{0}'),
        ('满意度', '持续参与意愿', '65%', '78%', '88%', '=E{0}', '85%', '=G{0}/H{0}'),
    ]

    kpi_start_row = row
    for i, data in enumerate(kpi_data):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(value, str) and '{0}' in value:
                cell.value = value.format(row)
            else:
                cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        ws.cell(row=row, column=8).number_format = '0.0%'
        row += 1

    # 添加百分比格式
    for r in range(kpi_start_row, row):
        for c in [3, 4, 5, 6, 7]:
            ws.cell(row=r, column=c).number_format = '0%'

    # 创建说明Sheet
    usage = '本表用于系统追踪先行者激励包的执行进度与效果，评估激励措施对推动新流程落地的实际成效。'
    instructions = '''追踪周期：
建议按月度填写，每月初更新上月数据

分析维度：
- 覆盖度：激励覆盖员工占比、部门参与率
- 响应度：激励申请率、激励知晓率
- 转化率：新流程采用率、行为改变持续率
- 满意度：激励公平性评分、激励有效性评分

数据来源：
- 人力资源系统、绩效考核系统、部门上报数据'''
    create_instruction_sheet(wb, '激励效果追踪表', usage, instructions)

    output_path = r'D:/新课开发/变革管理/06-考核与激励机制重设计：让新流程有人真心愿意执行/完整课程包/06-工具表单/Excel/工具09-激励效果追踪表.xlsx'
    wb.save(output_path)
    print(f'已生成: {output_path}')

def create_tool10():
    """工具10：考核制度嵌入检查表"""
    wb = Workbook()
    ws = wb.active
    ws.title = '制度嵌入检查'

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 25

    # 标题
    ws['A1'] = '考核制度嵌入检查表'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30

    # 基本信息
    row = 3
    info_fields = ['流程名称', '检查阶段', '检查周期', '填写人', '审核人', '填写日期']
    for i, field in enumerate(info_fields):
        col = 1 if i % 2 == 0 else 4
        if i % 2 == 0:
            ws.cell(row=row, column=col).value = f'{field}：'
            ws.cell(row=row, column=col).font = Font(name='微软雅黑', bold=True, size=10)
            ws.cell(row=row, column=col+1).value = '__________'
            ws.cell(row=row, column=col+1).border = get_border()
        else:
            ws.cell(row=row, column=col).value = f'{field}：'
            ws.cell(row=row, column=col).font = Font(name='微软雅黑', bold=True, size=10)
            ws.cell(row=row, column=col+1).value = '__________'
            ws.cell(row=row, column=col+1).border = get_border()
            row += 1

    row += 1

    # 考核制度嵌入全流程检查
    ws.cell(row=row, column=1).value = '一、考核制度嵌入全流程检查'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    headers = ['序号', '检查项目', '检查标准', '检查结果', '证据/说明', '不符合项处理']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    # 示例数据
    check_items = [
        (1, '考核指标已更新', '新流程关键动作已转化为可量化考核指标', '通过', '客户响应时效从"2小时内"纳入KPI', '—'),
        (2, '考核权重已调整', '新流程相关指标在个人考核中占比≥15%', '通过', '客服岗位KPI中客户响应时效占比20%', '—'),
        (3, '考核周期已匹配', '考核周期与流程运行周期匹配', '通过', '响应时效指标纳入月度考核', '—'),
        (4, '数据采集机制已建立', '流程执行数据可自动采集至考核系统', '部分通过', '响应时长可自动采集', '待优化：Q3实现自动化'),
        (5, '考核流程已嵌入', '考核评定流程与新流程运行流程同步启动', '通过', '新流程上线同步启动考核', '—'),
        (6, '考核标准已宣贯', '相关人员100%知晓新考核标准', '通过', '已开展3场宣贯会', '—'),
        (7, '考核申诉通道已建立', '员工对考核结果有异议可正式申诉', '通过', '已建立线上申诉通道', '—'),
        (8, '考核结果应用明确', '考核结果与薪酬、晋升、培训等挂钩', '通过', '考核结果应用于季度奖金', '—'),
        (9, '跨部门考核协同', '各部门考核责任已明确', '通过', '跨部门SLA已签订', '—'),
        (10, '主管评估能力达标', '直线主管具备准确评估的能力', '通过', '已完成主管评估技能培训', '—'),
        (11, '考核公平性有保障', '考核标准公开透明，评分标准一致', '通过', '考核制度已在内部平台公示', '—'),
        (12, '特殊情形有预案', '对借调、兼职、休假等有考核规定', '通过', '已制定特殊情形考核细则', '—'),
        (13, '试用期员工有衔接', '试用期员工考核标准与正式员工一致', '通过', '试用期前3个月按80%标准考核', '—'),
        (14, '历史数据可追溯', '能查询员工历史考核数据', '通过', 'HR系统保留36个月历史数据', '—'),
        (15, '系统集成已完成', '考核系统与业务流程系统数据打通', '部分通过', '已实现基础数据联通', '待优化：Q4实现自动化报表'),
        (16, '考核制度版本管理', '考核制度有版本号，能追踪变更历史', '通过', '现行版本V2.3', '—'),
    ]

    check_start_row = row
    for i, data in enumerate(check_items):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    # 添加下拉验证
    add_dropdown_validation(ws, f'D{check_start_row}:D{row-1}', ['通过', '部分通过', '未通过', '不适用'])

    row += 1

    # 各阶段关键里程碑
    ws.cell(row=row, column=1).value = '二、各阶段关键里程碑'
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    headers = ['阶段', '里程碑节点', '计划完成时间', '实际完成时间', '完成状态', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        set_header_style(cell)
    row += 1

    milestones = [
        ('设计阶段', '考核指标设计完成', '2026年1月15日', '2026年1月12日', '提前完成', ''),
        ('设计阶段', '考核权重确定', '2026年1月20日', '2026年1月22日', '略有延迟', ''),
        ('试运行阶段', '宣贯培训完成率100%', '2026年3月15日', '2026年3月18日', '略有延迟', '因业务旺季推迟3天'),
        ('试运行阶段', '模拟考核完成', '2026年3月底', '2026年3月28日', '按时完成', ''),
        ('正式运行阶段', '首次正式考核', '2026年5月初', '2026年5月2日', '提前完成', ''),
        ('正式运行阶段', '第一周期考核复盘', '2026年6月中旬', '2026年6月15日', '按时完成', ''),
    ]

    for i, data in enumerate(milestones):
        is_zebra = (i % 2 == 1)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            set_cell_style(cell, is_zebra=is_zebra, is_example=True)
        row += 1

    add_dropdown_validation(ws, f'E{row-6}:E{row-1}', ['提前完成', '按时完成', '略有延迟', '严重延迟'])

    # 创建说明Sheet
    usage = '本表用于系统检查考核制度是否已完整嵌入新流程的全生命周期，确保变革成果通过制度化手段固化下来。'
    instructions = '''检查时机：
- 流程设计阶段：考核指标设计、权重分配是否合理
- 试运行阶段（2次/月度）：考核制度执行可行性
- 正式运行阶段（每季度1次）：考核效果、制度适配性
- 年度评审：整体制度有效性

关键成功因素：
1. 制度完整性：16项检查项目必须全部覆盖
2. 执行一致性：检查标准要统一
3. 证据支撑：每项检查结果必须有文档或数据支撑
4. 闭环改进：发现的问题必须跟踪改进并验证效果
5. 版本管理：制度变更必须留有记录'''
    create_instruction_sheet(wb, '考核制度嵌入检查表', usage, instructions)

    output_path = r'D:/新课开发/变革管理/06-考核与激励机制重设计：让新流程有人真心愿意执行/完整课程包/06-工具表单/Excel/工具10-考核制度嵌入检查表.xlsx'
    wb.save(output_path)
    print(f'已生成: {output_path}')

def main():
    """主函数：生成所有Excel文件"""
    print('开始生成Excel文件...')
    print('=' * 50)

    create_tool01()
    create_tool02()
    create_tool03()
    create_tool04()
    create_tool05()
    create_tool06()
    create_tool07()
    create_tool08()
    create_tool09()
    create_tool10()

    print('=' * 50)
    print('所有Excel文件生成完成！')

if __name__ == '__main__':
    main()

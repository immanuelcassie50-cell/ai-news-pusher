# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "团队AI诊断总览"

header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
normal_font = Font(name='微软雅黑', size=10)
header_fill = PatternFill(start_color='B81025', end_color='B81025', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws.merge_cells('A1:L1')
ws['A1'] = '星辰科技市场部 AI成熟度诊断报告'
ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='B81025')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:L2')
ws['A2'] = '诊断日期：2024年1月15日    诊断人：张明（市场总监）    团队规模：12人'
ws['A2'].font = Font(name='微软雅黑', size=10, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

headers = ['序号', '姓名', '职位', '类型', 'AI意识', '使用频率', '场景应用', '学习意愿', '综合得分', '成熟度等级', '带教优先级', '备注']
ws.append([])
ws.append(headers)
header_row = ws.max_row
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

team_data = [
    [1, '李雪', '高级内容策划', 'A', 5, 5, 5, 4, 19, 'A类-精通者', '赋能', '内容能力强，AI文案已常态化'],
    [2, '王浩', '数字营销主管', 'A', 5, 4, 5, 5, 19, 'A类-精通者', '赋能', '数据分析背景，学习速度快'],
    [3, '陈思', '品牌主管', 'B', 4, 3, 4, 4, 15, 'B类-尝试者', '加速', '对AI有兴趣，需要场景引导'],
    [4, '刘洋', '活动策划', 'B', 4, 3, 3, 5, 15, 'B类-尝试者', '加速', '执行力强，需要正向反馈'],
    [5, '赵敏', '内容策划', 'B', 3, 4, 3, 4, 14, 'B类-尝试者', '加速', '文字功底好，AI辅助效果明显'],
    [6, '孙悦', '媒介专员', 'C', 3, 2, 2, 3, 10, 'C类-观望者', '陪伴', '担心AI替代工作，需建立信心'],
    [7, '周杰', '投放专员', 'C', 2, 2, 2, 3, 9, 'C类-观望者', '陪伴', '技术背景弱，需要更多支持'],
    [8, '吴婷', '内容专员', 'C', 3, 1, 2, 4, 10, 'C类-观望者', '陪伴', '新入职，对团队方式不熟悉'],
    [9, '郑强', '设计主管', 'C', 3, 2, 2, 2, 9, 'C类-观望者', '陪伴', '视觉设计师，对文字类AI不感兴趣'],
    [10, '黄丽', '市场分析师', 'C', 2, 2, 2, 3, 9, 'C类-观望者', '陪伴', '习惯旧流程，改变需要时间'],
    [11, '林峰', '市场总监助理', 'D', 1, 1, 1, 1, 4, 'D类-拒绝者', '意义', '年龄偏大，对新技术有抵触'],
    [12, '杨帆', '行政专员', 'D', 1, 1, 1, 1, 4, 'D类-拒绝者', '意义', '工作内容与AI关联度低'],
]

type_colors = {'A': 'E8F5E9', 'B': 'FFF8E1', 'C': 'FFF3E0', 'D': 'FFEBEE'}

for row_data in team_data:
    ws.append(row_data)

for row in range(header_row + 1, ws.max_row + 1):
    for col in range(1, 13):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        member_type = ws.cell(row=row, column=4).value
        if member_type in type_colors:
            cell.fill = PatternFill(start_color=type_colors[member_type], end_color=type_colors[member_type], fill_type='solid')

col_widths = [6, 10, 14, 8, 8, 10, 10, 10, 10, 14, 10, 24]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws2 = wb.create_sheet("四类成员分布")
ws2.merge_cells('A1:D1')
ws2['A1'] = '团队AI成熟度四类分布'
ws2['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='B81025')
ws2['A1'].alignment = Alignment(horizontal='center')

ws2.append([])
ws2.append(['类型', '人数', '占比', '代表成员'])
type_header_row = ws2.max_row
for col in range(1, 5):
    cell = ws2.cell(row=type_header_row, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

type_data = [
    ['A类-精通者', 2, '17%', '李雪、王浩'],
    ['B类-尝试者', 3, '25%', '陈思、刘洋、赵敏'],
    ['C类-观望者', 5, '42%', '孙悦、周杰、吴婷、郑强、黄丽'],
    ['D类-拒绝者', 2, '17%', '林峰、杨帆'],
]
for row_data in type_data:
    ws2.append(row_data)
    row = ws2.max_row
    for col in range(1, 5):
        cell = ws2.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if row_data[0].startswith('A'):
            cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        elif row_data[0].startswith('B'):
            cell.fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
        elif row_data[0].startswith('C'):
            cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
        elif row_data[0].startswith('D'):
            cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 30

ws3 = wb.create_sheet("评估维度说明")
ws3.merge_cells('A1:E1')
ws3['A1'] = 'AI成熟度评估维度说明'
ws3['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='B81025')

ws3.append([])
ws3.append(['维度', '权重', '评分标准', '评估方法', '改进建议'])
dim_header_row = ws3.max_row
for col in range(1, 6):
    cell = ws3.cell(row=dim_header_row, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

dim_data = [
    ['AI意识', '25%', '1-5分：对AI价值的认知程度', '问卷+访谈', '通过案例分享提升认知'],
    ['使用频率', '25%', '1-5分：每周使用AI工具的次数', '工具埋点+自报', '从每天5分钟开始养成习惯'],
    ['场景应用', '30%', '1-5分：实际应用到工作的场景数', '工作日志+产出物', '每周解锁1个新场景'],
    ['学习意愿', '20%', '1-5分：主动学习AI的积极性', '访谈+行为观察', '与绩效挂钩，建立激励'],
]
for row_data in dim_data:
    ws3.append(row_data)
    row = ws3.max_row
    for col in range(1, 6):
        cell = ws3.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 25

wb.save('D:/新课开发/领导力/中基层/04带队伍用AI/成果demo/01_Team_AI_Diagnosis.xlsx')
print("01_Team_AI_Diagnosis.xlsx created successfully")

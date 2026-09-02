import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

output_dir = "D:/新课开发/营销/1. AI时代销售突破：从客户洞察到高质量成交/全流程工具表单"
os.makedirs(output_dir, exist_ok=True)
wb = openpyxl.Workbook()
rf = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
gf = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
wf = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
hf = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
bf = Font(name="Microsoft YaHei", size=10)
bs = Border(left=Side(style="thin", color="E8E8E8"), right=Side(style="thin", color="E8E8E8"), top=Side(style="thin", color="E8E8E8"), bottom=Side(style="thin", color="E8E8E8"))
ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
la = Alignment(horizontal="left", vertical="center", wrap_text=True)

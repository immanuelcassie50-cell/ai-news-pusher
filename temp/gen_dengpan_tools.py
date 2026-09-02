from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "D:/新课开发/战略和领导力/登攀者——AI时代的授权赋能领导力/完整课程包/06_工具表单/登攀者课程工具表单集.xlsx"

# Use ASCII-safe quotes inside strings
Q = "'"
DQ = '"'

CLR = {
    "title_bg":   "1F4E79",
    "title_fg":   "FFFFFF",
    "header_bg":  "2E75B6",
    "header_fg":  "FFFFFF",
    "section_bg": "D6E4F0",
    "section_fg": "1F4E79",
    "row_alt":    "EBF3FB",
    "input_bg":   "FFFACD",
    "label_bg":   "F2F2F2",
    "border":     "B8CCE4",
    "green_bg":   "E2EFDA",
    "orange_bg":  "FCE4D6",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, name="Microsoft YaHei"):
    return Font(bold=bold, color=color, size=size, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(color=None, style="thin"):
    if color is None:
        color = CLR["border"]
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, row, texts, bg=None, fg=None, bold=True, heights=None):
    bg = bg or CLR["header_bg"]
    fg = fg or CLR["header_fg"]
    for col, text in enumerate(texts, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=fg)
        c.alignment = align("center", "center", wrap=True)
        c.border = border(CLR["header_bg"])
    if heights:
        ws.row_dimensions[row].height = heights

def apply_title(ws, row, text, ncols, height=32):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(CLR["title_bg"])
    c.font = font(bold=True, color=CLR["title_fg"], size=14)
    c.alignment = align("center", "center")
    ws.row_dimensions[row].height = height

def apply_section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(CLR["section_bg"])
    c.font = font(bold=True, color=CLR["section_fg"], size=11)
    c.alignment = align("left", "center")
    c.border = border(CLR["section_bg"])
    ws.row_dimensions[row].height = 20

def label_cell(ws, row, col, text, bg=None, bold=False, align_h="left"):
    bg = bg or CLR["label_bg"]
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=bold, color="1F4E79")
    c.alignment = align(align_h, "center", wrap=True)
    c.border = border()
    return c

def input_cell(ws, row, col, text="", bg=None):
    bg = bg or CLR["input_bg"]
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(color="000000")
    c.alignment = align("left", "center", wrap=True)
    c.border = border()
    return c

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

wb = Workbook()
wb.remove(wb.active)

# ==============================================================
# SHEET 1: 学员信息表
# ==============================================================
ws1 = wb.create_sheet("学员信息表")
set_col_widths(ws1, [18, 22, 18, 26, 18, 28])
r = 1
apply_title(ws1, r, "登攀者 - AI时代的授权赋能领导力  学员信息表", 6)
r += 1
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws1.cell(row=r, column=1, value="本表用于收集学员基本信息，供讲师了解学员背景，便于教学互动和课后跟进。请认真填写。")
c.fill = fill(CLR["row_alt"])
c.font = font(size=10, color="595959")
c.alignment = align("center", "center")
r += 1

apply_section(ws1, r, "一、基本信息", 6)
r += 1
apply_header(ws1, r, ["姓名", "公司/部门", "职位", "电子邮箱", "联系电话", "备注"])
r += 1
for _ in range(8):
    for col in range(1, 7):
        input_cell(ws1, r, col)
    r += 1

apply_section(ws1, r, "二、AI工具使用经验自评", 6)
r += 1
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws1.cell(row=r, column=1, value="请在符合您情况的选项上打钩：  □ 从未使用  □ 偶尔试用  □ 常规使用  □ 系统化深度使用")
c.fill = fill(CLR["row_alt"])
c.font = font(size=10)
c.alignment = align("left", "center")
ws1.row_dimensions[r].height = 20
r += 1

labels2 = [
    "您目前使用哪些AI工具？（列举）：",
    "您认为AI工具对您工作的帮助程度：",
    "您在工作中最常用AI做什么？",
    "您对AI取代人类工作的看法：",
]
for label in labels2:
    label_cell(ws1, r, 1, label, bold=True)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    input_cell(ws1, r, 2)
    r += 1

apply_section(ws1, r, "三、教练经验自评", 6)
r += 1
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws1.cell(row=r, column=1, value="请在符合您情况的选项上打钩：  □ 无教练经验  □ 接受过教练辅导  □ 有过非正式教练实践  □ 有正式教练认证或丰富经验")
c.fill = fill(CLR["row_alt"])
c.font = font(size=10)
c.alignment = align("left", "center")
ws1.row_dimensions[r].height = 20
r += 1

coach_labels = [
    "您作为管理者做过多少次正式教练对话？",
    "您用过的教练模型有哪些？（如GROW, CLEAR等）",
    "您认为教练的最大价值是什么？",
    "您在教练实践中遇到的最大挑战是什么？",
]
for label in coach_labels:
    label_cell(ws1, r, 1, label, bold=True)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    input_cell(ws1, r, 2)
    r += 1

apply_section(ws1, r, "四、学习目标", 6)
r += 1
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws1.cell(row=r, column=1, value="请描述您参加本次课程最想达成的三个目标：")
c.fill = fill(CLR["label_bg"])
c.font = font(bold=True, color="1F4E79")
c.alignment = align("left", "center")
ws1.row_dimensions[r].height = 18
r += 1
for i in range(1, 4):
    label_cell(ws1, r, 1, "目标{}：".format(i), bold=True)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    input_cell(ws1, r, 2)
    ws1.row_dimensions[r].height = 22
    r += 1

apply_section(ws1, r, "五、课程期望", 6)
r += 1
for label in ["您最希望通过本次课程解决什么问题？", "您对课程有任何特殊需求或顾虑吗？"]:
    label_cell(ws1, r, 1, label, bold=True)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    input_cell(ws1, r, 2)
    ws1.row_dimensions[r].height = 30
    r += 1

ws1.freeze_panes = "A4"
print("Sheet 1 done")

# ==============================================================
# SHEET 2: 课程准备检查表
# ==============================================================
ws2 = wb.create_sheet("课程准备检查表")
set_col_widths(ws2, [5, 28, 42, 12, 24])
r = 1
apply_title(ws2, r, "登攀者 - AI时代的授权赋能领导力  课程准备检查表", 5, height=30)
r += 1
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws2.cell(row=r, column=1, value="课程日期：________________  讲师：________________  场地：________________")
c.fill = fill(CLR["row_alt"])
c.font = font(size=10)
c.alignment = align("center", "center")
ws2.row_dimensions[r].height = 20
r += 1

def checklist_section(ws, r, title, items):
    apply_section(ws, r, title, 5)
    r += 1
    apply_header(ws, r, ["", "准备项", "具体要求/说明", "完成", "备注"])
    r += 1
    for i, (item, desc) in enumerate(items):
        bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
        label_cell(ws, r, 1, str(i+1), bg=bg, align_h="center")
        label_cell(ws, r, 2, item, bg=bg, bold=True)
        label_cell(ws, r, 3, desc, bg=bg)
        c = ws.cell(row=r, column=4, value="□ 是  □ 否")
        c.fill = fill(bg)
        c.font = font(size=10)
        c.alignment = align("center", "center")
        c.border = border()
        input_cell(ws, r, 5, "", bg=bg)
        r += 1
    return r

r = checklist_section(ws2, r, "一、讲师准备项", [
    ("教学PPT", "全套PPT，确认各模块播放正常"),
    ("学员手册", "每人一本，含工作页和应用卡"),
    ("工具表单", "本次课程所需全部表单打印完毕"),
    ("GUIDE/DIRECT模型卡", "每人一张，可裁剪随身携带"),
    ("AI时代六类场景卡", "每组一套，彩色打印"),
    ("教练行动计划工作纸", "每人2张"),
    ("讲师手册", "全套（Day1上午/下午/Day2上午/下午）"),
    ("讲师提示卡", "关键话术提醒卡"),
    ("演练观察清单", "每组1份"),
    ("白板/大白纸", "至少2面，记录关键洞察"),
    ("投影设备", "确认PPT投影正常"),
    ("音响设备", "背景音乐播放正常"),
    ("计时器", "2个，控制演练时间"),
    ("便利贴", "多种颜色，开场活动用"),
    ("签字笔", "每组2支"),
])

r = checklist_section(ws2, r, "二、学员准备项", [
    ("预习资料", "课前发送的阅读材料已发送给学员"),
    ("课前调研问卷", "收回学员信息表，了解背景"),
    ("AI工具经验", "学员自带AI使用经验案例"),
    ("真实挑战", "学员准备一个真实的教练挑战案例"),
])

r = checklist_section(ws2, r, "三、场地准备项", [
    ("座位布置", "U型或小组圈形，便于讨论和演练"),
    ("空间大小", "足够容纳所有学员和活动"),
    ("光线/通风", "良好的光线和通风条件"),
    ("电源插座", "投影设备和讲师设备供电"),
    ("手机信号", "场地手机信号良好（不屏蔽）"),
])

apply_section(ws2, r, "四、应急预案", 5)
r += 1
for label in ["设备故障预案：", "学员情绪应对：", "时间超支处理：", "其他备注："]:
    label_cell(ws2, r, 1, label, bold=True)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    input_cell(ws2, r, 2)
    r += 1

ws2.freeze_panes = "A4"
print("Sheet 2 done")

# ==============================================================
# SHEET 3: 教练观察记录表
# ==============================================================
ws3 = wb.create_sheet("教练观察记录表")
set_col_widths(ws3, [16, 16, 14, 12, 14, 30, 30, 16])
r = 1
apply_title(ws3, r, "登攀者 - AI时代的授权赋能领导力  教练观察记录表", 8, height=30)
r += 1
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = ws3.cell(row=r, column=1, value="观察者：________________  日期：________________  场次：________________")
c.fill = fill(CLR["row_alt"])
c.font = font(size=10)
c.alignment = align("center", "center")
ws3.row_dimensions[r].height = 20
r += 1

apply_section(ws3, r, "基本信息", 8)
r += 1
label_cell(ws3, r, 1, "教练者", bold=True)
label_cell(ws3, r, 2, "", bold=True)
ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
input_cell(ws3, r, 2)
label_cell(ws3, r, 4, "被教练者", bold=True)
ws3.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
input_cell(ws3, r, 5)
label_cell(ws3, r, 7, "模型", bold=True)
input_cell(ws3, r, 8)
r += 1
label_cell(ws3, r, 1, "日期", bold=True)
input_cell(ws3, r, 2)
label_cell(ws3, r, 3, "时间", bold=True)
input_cell(ws3, r, 4)
label_cell(ws3, r, 5, "场景类型", bold=True)
ws3.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
input_cell(ws3, r, 5)
label_cell(ws3, r, 7, "演练序号", bold=True)
input_cell(ws3, r, 8)
r += 1

apply_section(ws3, r, "GUIDE模型观察要点", 8)
r += 1
apply_header(ws3, r, ["步骤", "教练者行为记录", "被教练者回应", "关键瞬间", "评分(1-5)", "", "", ""])
r += 1

guide_steps = [
    ("G - Goal（目标）", "教练是否帮助被教练者明确了自己的目标？目标是否是被教练者自己的？"),
    ("U - Understand（现状）", "教练是否帮助被教练者清晰看到现状？是否有评判或急于分析？"),
    ("I - Insight（洞察）", "被教练者是否产生了新的洞察？教练是否给了足够的沉默和空间？"),
    ("D - Design（行动）", "行动计划是否具体可执行？是否包含双轨（人类深度+AI杠杆）？"),
    ("E - Enable（障碍）", "是否识别了真实障碍？AI相关心智阻碍是否有针对性处理？"),
]
for step, desc in guide_steps:
    label_cell(ws3, r, 1, step, bold=True, bg=CLR["label_bg"])
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    input_cell(ws3, r, 2)
    ws3.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    input_cell(ws3, r, 4)
    input_cell(ws3, r, 6)
    ws3.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    input_cell(ws3, r, 7)
    ws3.row_dimensions[r].height = 28
    r += 1

apply_section(ws3, r, "DIRECT补充观察（业绩型对话）", 8)
r += 1
direct_steps = [
    ("D - Data（收集行为）", "是否收集了具体可观察的行为？是否有归因分析（你的判断在哪里）？"),
    ("I - Impact（影响）", "是否沟通了行为对团队/业务的影响？是否提到长期能力发展影响？"),
    ("R - Requirement（期望）", "是否清晰表达了双维度期望（产出+人类贡献）？"),
    ("E - Explore（探索）", "是否帮助被教练者自己想出解决方案？"),
    ("C - Commitment（承诺）", "承诺是否具体、可追踪？是否包含人类贡献可见化？"),
    ("T - Track（追踪）", "追踪要点是否明确？是否包含人类判断贡献的成长指标？"),
]
for step, desc in direct_steps:
    label_cell(ws3, r, 1, step, bold=True, bg=CLR["green_bg"])
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    input_cell(ws3, r, 2)
    ws3.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    input_cell(ws3, r, 4)
    input_cell(ws3, r, 6)
    ws3.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    input_cell(ws3, r, 7)
    ws3.row_dimensions[r].height = 28
    r += 1

apply_section(ws3, r, "总体评价与反馈", 8)
r += 1
for label in ["本次教练最有效的部分：", "需要加强的部分：", "给教练者的具体建议："]:
    label_cell(ws3, r, 1, label, bold=True)
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    input_cell(ws3, r, 2)
    ws3.row_dimensions[r].height = 28
    r += 1

ws3.freeze_panes = "A4"
print("Sheet 3 done")

# ==============================================================
# SHEET 4: 聆听自我评估表
# ==============================================================
ws4 = wb.create_sheet("聆听自我评估表")
set_col_widths(ws4, [8, 22, 34, 14, 32])
r = 1
apply_title(ws4, r, "登攀者 - AI时代的授权赋能领导力  聆听自我评估表", 5, height=30)
r += 1
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws4.cell(row=r, column=1, value="有意识聆听四层次 - 请根据每次教练对话后的真实感受填写此表，进行自我反思与成长。")
c.fill = fill(CLR["row_alt"])
c.font = font(size=10)
c.alignment = align("center", "center")
ws4.row_dimensions[r].height = 20
r += 1

apply_section(ws4, r, "聆听四层次评估标准", 5)
r += 1
apply_header(ws4, r, ["层次", "名称", "特征描述", "自评(1-10)", "具体表现记录"])
r += 1

listening_levels = [
    ("1", "假装聆听",
     "眼神接触但心不在焉；等待发言而非理解对方；表面点头但没有真正接收信息；教练者已在想下一个问题而非当下。", "", ""),
    ("2", "选择性聆听",
     "听到部分信息并筛选；注意对方说了什么但忽略了他为什么这么说；基于自己的预设进行判断；错过非语言信息。", "", ""),
    ("3", "有意识聆听",
     "全神贯注于对方；注意到语言、语气和非语言信号；听到字面意思和深层含义；在回应前先确认理解。", "", ""),
    ("4", "深度聆听（+AI时代身份信号）",
     "不仅听到对方说的话，还能听到他没说出来的担忧和渴望；在教练对话中能捕捉到AI时代的身份焦虑信号（价值焦虑/过度依赖/无效抵触）。这是最高层次的聆听。", "", ""),
]
for i, (level, name, desc, score, record) in enumerate(listening_levels):
    bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
    if level == "4":
        bg = CLR["green_bg"]
    elif level == "3":
        bg = CLR["section_bg"]
    label_cell(ws4, r, 1, level, bg=bg, align_h="center", bold=True)
    label_cell(ws4, r, 2, name, bg=bg, bold=True)
    label_cell(ws4, r, 3, desc, bg=bg)
    c = ws4.cell(row=r, column=4)
    c.fill = fill(CLR["input_bg"])
    c.font = font(size=12, bold=True)
    c.alignment = align("center", "center")
    c.border = border()
    input_cell(ws4, r, 5, record, bg=bg)
    ws4.row_dimensions[r].height = 40
    r += 1

apply_section(ws4, r, "深度聆听行为自查", 5)
r += 1
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws4.cell(row=r, column=1, value="在最近一次教练对话中，请回想：你是否有以下行为？")
c.fill = fill(CLR["label_bg"])
c.font = font(bold=True, color="1F4E79")
c.alignment = align("left", "center")
r += 1

checklist_items = [
    "对方说话时，我在想接下来要问什么问题",
    "我注意到对方的语气或表情变化",
    "我在回应前先复述并确认了我的理解",
    "我注意到了对方没有说出来的情绪或担忧",
    "我给了对方足够的沉默让他思考",
    "我捕捉到了AI时代的身份焦虑信号（价值焦虑/过度依赖/无效抵触）",
    "我区分了对方的表面陈述和深层渴望",
]
for item in checklist_items:
    c = ws4.cell(row=r, column=1, value="□")
    c.fill = fill(CLR["input_bg"])
    c.font = font(size=12)
    c.alignment = align("center", "center")
    c.border = border()
    ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    c2 = ws4.cell(row=r, column=2, value=item)
    c2.fill = fill(CLR["input_bg"])
    c2.font = font()
    c2.alignment = align("left", "center")
    c2.border = border()
    ws4.row_dimensions[r].height = 18
    r += 1

apply_section(ws4, r, "改进目标设定", 5)
r += 1
for label in ["我的聆听强项：", "我最需要改进的聆听行为：", "我的改进行动计划：", "下次教练对话我要特别关注："]:
    label_cell(ws4, r, 1, label, bold=True)
    ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    input_cell(ws4, r, 2)
    ws4.row_dimensions[r].height = 22
    r += 1

ws4.freeze_panes = "A4"
print("Sheet 4 done")

# ==============================================================
# SHEET 5: 提问技巧练习表
# ==============================================================
ws5 = wb.create_sheet("提问技巧练习表")
set_col_widths(ws5, [10, 20, 36, 36])
r = 1
apply_title(ws5, r, "登攀者 - AI时代的授权赋能领导力  提问技巧练习表", 4, height=30)
r += 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws5.cell(row=r, column=1, value="三种问题类型对比 + AI时代突破性问题设计练习")
c.fill = fill(CLR["row_alt"])
c.font = font(size=10)
c.alignment = align("center", "center")
ws5.row_dimensions[r].height = 20
r += 1

apply_section(ws5, r, "一、三种问题类型对比练习", 4)
r += 1
apply_header(ws5, r, ["类型", "定义", "示例问题", "练习区"])
r += 1

q_types = [
    ("封闭式问题",
     "可以用是/否或简短事实回答的问题。用于确认信息、澄清细节。",
     "例：你上周用了多少次AI工具？\n例：这个目标有具体的完成日期吗？", ""),
    ("开放式问题",
     "无法用简单词汇回答，需要详细阐述的问题。用于探索想法、深化理解。",
     "例：你对这件事的感受是什么？\n例：你觉得接下来可以怎么做？", ""),
    ("突破性问题",
     "帮助被教练者看到盲区、挑战现有框架、激发新洞察的问题。",
     "例：如果AI处理了这部分工作，你的注意力会流向哪里？\n例：你最担心AI取代你的那部分 - 反过来想，那里是不是你真正最有深度的地方？", ""),
]
for i, (qtype, define, example, practice) in enumerate(q_types):
    bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
    if qtype == "突破性问题":
        bg = CLR["green_bg"]
    label_cell(ws5, r, 1, qtype, bg=bg, bold=True, align_h="center")
    label_cell(ws5, r, 2, define, bg=bg)
    label_cell(ws5, r, 3, example, bg=bg)
    input_cell(ws5, r, 4, practice, bg=CLR["input_bg"])
    ws5.row_dimensions[r].height = 52
    r += 1

apply_section(ws5, r, "二、AI时代突破性问题设计", 4)
r += 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws5.cell(row=r, column=1, value="请为以下AI时代挑战场景各设计3个突破性问题：")
c.fill = fill(CLR["label_bg"])
c.font = font(bold=True, color="1F4E79")
c.alignment = align("left", "center")
r += 1
apply_header(ws5, r, ["场景类型", "场景描述", "突破性问题设计", "设计思路"])
r += 1

scenarios = [
    ("价值焦虑型",
     "一名数据分析师，看到AI工具大量出现，开始怀疑自己的价值，问我觉得我还能在这里做多久。",
     "1.\n2.\n3.", ""),
    ("过度依赖型",
     "一名产品经理，用AI大幅提升了产出效率，但说不清楚自己判断的逻辑在哪里。",
     "1.\n2.\n3.", ""),
    ("无效抵触型",
     "一名资深工程师，抵触AI辅助工具，说真正的工程师靠自己。",
     "1.\n2.\n3.", ""),
    ("接受更大责任",
     "一名高潜负责人，说我不是技术出身，AI这块我不行，但项目真正需要的是系统整合能力。",
     "1.\n2.\n3.", ""),
]
for i, (stype, desc, qs, note) in enumerate(scenarios):
    bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
    label_cell(ws5, r, 1, stype, bg=bg, bold=True)
    label_cell(ws5, r, 2, desc, bg=bg)
    input_cell(ws5, r, 3, qs, bg=CLR["input_bg"])
    input_cell(ws5, r, 4, note, bg=CLR["input_bg"])
    ws5.row_dimensions[r].height = 60
    r += 1

apply_section(ws5, r, "三、自我评估", 4)
r += 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws5.cell(row=r, column=1, value="反思：在本次练习中，你在设计哪种类型的问题时最顺手？哪种类型最有挑战？")
c.fill = fill(CLR["label_bg"])
c.font = font(bold=True, color="1F4E79")
c.alignment = align("left", "center")
r += 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
input_cell(ws5, r, 1)
ws5.row_dimensions[r].height = 36
r += 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws5.cell(row=r, column=1, value="教练的核心是问问题而不是给答案 - 今天你在演练中做到了吗？给自己打分（1-10）：________")
c.fill = fill(CLR["row_alt"])
c.font = font()
c.alignment = align("left", "center")
r += 1

ws5.freeze_panes = "A4"
print("Sheet 5 done")

# ==============================================================
# SHEET 6: 发展性反馈练习表
# ==============================================================
ws6 = wb.create_sheet("发展性反馈练习表")
set_col_widths(ws6, [10, 26, 26, 26, 12])
r = 1
apply_title(ws6, r, "登攀者 - AI时代的授权赋能领导力  发展性反馈练习表", 5, height=30)
r += 1
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws6.cell(row=r, column=1, value="发展性反馈 = 帮助被教练者看见自己没意识到的内在资源。四要素结构练习：")
c.fill = fill(CLR["row_alt"])
c.font = font(size=10)
c.alignment = align("center", "center")
r += 1

apply_section(ws6, r, "一、发展性反馈四要素结构", 5)
r += 1
apply_header(ws6, r, ["要素", "内容", "示范话术示例", "我的练习区", "评分"])
r += 1

feedback_elements = [
    ("要素一\n具体观察",
     "描述具体的、可观察的行为，不是印象或评价。",
     "在上周的团队复盘里，你在分析这个问题时停顿了一下，然后说了一个完全不同的角度……", ""),
    ("要素二\n内在资源命名",
     "帮助被教练者看到他没说出来的优势、特质或能力。",
     "……那个停顿，和那个不同的角度，体现的是你在这个领域的专业直觉 - 这是经验积累才会有的。", ""),
    ("要素三\nAI时代价值定位",
     "帮助被教练者在AI时代背景下重新认识自己的独特价值。",
     "……AI可以给你十个分析框架，但它没有你在这个行业里打了五年交道积累的直觉。那个直觉，是你最有价值的东西。", ""),
    ("要素四\n邀请自我探索",
     "不直接给结论，而是邀请被教练者自己感受和探索。",
     "……你感受到这部分吗？你觉得怎么样把它发挥得更充分？", ""),
]
for i, (elem, desc, example, practice) in enumerate(feedback_elements):
    bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
    label_cell(ws6, r, 1, elem, bg=bg, bold=True, align_h="center")
    label_cell(ws6, r, 2, desc, bg=bg)
    label_cell(ws6, r, 3, example, bg=bg)
    input_cell(ws6, r, 4, practice, bg=CLR["input_bg"])
    c = ws6.cell(row=r, column=5)
    c.fill = fill(CLR["input_bg"])
    c.font = font()
    c.alignment = align("center", "center")
    c.border = border()
    ws6.row_dimensions[r].height = 50
    r += 1

apply_section(ws6, r, "二、认同技巧 vs 夸奖练习", 5)
r += 1
apply_header(ws6, r, ["", "夸奖式反馈（避免）", "认同式反馈（练习）", "", ""])
r += 1

comparison = [
    ("例1",
     "你这次做得很好！（评价者视角）",
     "在你刚才说的话里，我听到了一种对这个问题的深度关心，这是做好这件事真正需要的东西。你注意到了吗？（帮助他自己看见）"),
    ("例2",
     "你的报告写得很棒！",
     "我在你身上看到一件事 - 当我们讨论这个问题的时候，你说了一句话，让整个对话改变了方向。那句话来自你三年里和客户之间建立的真实的关系理解。这不是AI能做到的，是你做到的。"),
    ("练习",
     "请写出一个你曾经对下属说过的夸奖式反馈：",
     "请改写成认同式反馈："),
]
for i, (num, bad, good) in enumerate(comparison):
    bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
    label_cell(ws6, r, 1, num, bg=bg, align_h="center", bold=True)
    label_cell(ws6, r, 2, bad, bg=CLR["orange_bg"])
    ws6.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    input_cell(ws6, r, 3, good if "练习" not in num else "", bg=CLR["green_bg"] if "练习" in num else CLR["input_bg"])
    ws6.row_dimensions[r].height = 36
    r += 1

apply_section(ws6, r, "三、AI时代价值定位专项练习", 5)
r += 1
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws6.cell(row=r, column=1, value="场景：一名市场经理因为AI可以批量生成创意框架而沮丧，觉得自己没有独特的创意能力了。请用四要素写出发展性反馈：")
c.fill = fill(CLR["green_bg"])
c.font = font(bold=True, color="1F4E79")
c.alignment = align("left", "center")
ws6.row_dimensions[r].height = 24
r += 1
for label in ["要素一（具体观察）：", "要素二（内在资源）：", "要素三（AI价值定位）：", "要素四（邀请探索）："]:
    label_cell(ws6, r, 1, label, bold=True)
    ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    input_cell(ws6, r, 2)
    ws6.row_dimensions[r].height = 24
    r += 1

ws6.freeze_panes = "A4"
print("Sheet 6 done")

# ==============================================================
# SHEET 7: 教练行动计划追踪表
# ==============================================================
ws7 = wb.create_sheet("教练行动计划追踪表")
set_col_widths(ws7, [16, 16, 14, 28, 28, 14])
r = 1
apply_title(ws7, r, "登攀者 - AI时代的授权赋能领导力  教练行动计划追踪表", 6, height=30)
r += 1
ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws7.cell(row=r, column=1, value="双轨行动计划：人类深度轨道（我要发展的能力）+ AI杠杆轨道（我要更好地使用哪个工具来支持前者）")
c.fill = fill(CLR["green_bg"])
c.font = font(size=10, bold=True)
c.alignment = align("center", "center")
ws7.row_dimensions[r].height = 20
r += 1

apply_section(ws7, r, "被教练者基本信息", 6)
r += 1
row1 = [("被教练者姓名：", ""), ("职位：", ""), ("教练者：", "")]
for i, (label, val) in enumerate(row1):
    col_offset = i * 2 + 1
    label_cell(ws7, r, col_offset, label, bold=True)
    input_cell(ws7, r, col_offset + 1, val)
r += 1
row2 = [("首次对话日期：", ""), ("目标类型：□ GUIDE  □ DIRECT", ""), ("跟进周期：", "")]
for i, (label, val) in enumerate(row2):
    col_offset = i * 2 + 1
    label_cell(ws7, r, col_offset, label, bold=True)
    input_cell(ws7, r, col_offset + 1, val)
r += 1

apply_section(ws7, r, "教练目标设定", 6)
r += 1
label_cell(ws7, r, 1, "被教练者想要达成的核心目标：", bold=True)
ws7.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
input_cell(ws7, r, 2)
ws7.row_dimensions[r].height = 28
r += 1
label_cell(ws7, r, 1, "这个目标在AI时代的特别意义：", bold=True)
ws7.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
input_cell(ws7, r, 2)
ws7.row_dimensions[r].height = 28
r += 1

apply_section(ws7, r, "双轨行动计划", 6)
r += 1
apply_header(ws7, r, ["行动维度", "具体行动描述", "时间节点", "完成情况", "反思与调整", "评分(1-5)"])
r += 1

action_tracks = [
    ("人类深度轨道\n（我要发展的能力）",
     "例：在没有AI工具的情况下，完成一份核心分析报告的判断逻辑构建。", "", "", "", ""),
    ("人类深度轨道", "", "", "", "", ""),
    ("AI杠杆轨道\n（我要更好地使用哪个工具）",
     "例：学会使用AI工具快速生成报告框架，但核心判断由我自己完成。", "", "", "", ""),
    ("AI杠杆轨道", "", "", "", "", ""),
]
for i, (track, action, deadline, done, reflect, score) in enumerate(action_tracks):
    bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
    if "AI杠杆" in track:
        bg = CLR["green_bg"]
    label_cell(ws7, r, 1, track, bg=bg, bold=True, align_h="center")
    input_cell(ws7, r, 2, action, bg=bg)
    input_cell(ws7, r, 3, deadline, bg=bg)
    input_cell(ws7, r, 4, done, bg=bg)
    input_cell(ws7, r, 5, reflect, bg=bg)
    c = ws7.cell(row=r, column=6)
    c.fill = fill(CLR["input_bg"])
    c.font = font()
    c.alignment = align("center", "center")
    c.border = border()
    ws7.row_dimensions[r].height = 36
    r += 1

apply_section(ws7, r, "追踪记录", 6)
r += 1
apply_header(ws7, r, ["日期", "对话形式", "关键进展记录", "挑战与解决方案", "下次计划", ""])
r += 1
for _ in range(6):
    for col in range(1, 6):
        input_cell(ws7, r, col)
    ws7.row_dimensions[r].height = 28
    r += 1

apply_section(ws7, r, "成果评估", 6)
r += 1
for label in ["目标达成情况：", "人类深度发展成果：", "AI杠杆使用改进：", "总体评分与总结："]:
    label_cell(ws7, r, 1, label, bold=True)
    ws7.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    input_cell(ws7, r, 2)
    ws7.row_dimensions[r].height = 28
    r += 1

ws7.freeze_panes = "A4"
print("Sheet 7 done")

# ==============================================================
# SHEET 8: 课程评估表
# ==============================================================
ws8 = wb.create_sheet("课程评估表")
set_col_widths(ws8, [32, 14, 14, 14, 26])
r = 1
apply_title(ws8, r, "登攀者 - AI时代的授权赋能领导力  课程评估表（反应层）", 5, height=30)
r += 1
ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws8.cell(row=r, column=1, value="姓名：________________  公司/部门：________________  填写日期：________________")
c.fill = fill(CLR["row_alt"])
c.font = font(size=10)
c.alignment = align("center", "center")
ws8.row_dimensions[r].height = 20
r += 1
ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws8.cell(row=r, column=1, value="请在每个评分项上选择您真实的感受（1=完全不认同，5=完全认同，N/A=不适用）")
c.fill = fill(CLR["label_bg"])
c.font = font(bold=True, color="1F4E79", size=10)
c.alignment = align("center", "center")
r += 1

def eval_section(ws, r, title, items):
    apply_section(ws, r, title, 5)
    r += 1
    apply_header(ws, r, ["评估项", "1", "2", "3", "备注"])
    r += 1
    for i, item in enumerate(items):
        bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
        label_cell(ws, r, 1, item, bg=bg)
        for col in range(2, 5):
            c = ws.cell(row=r, column=col, value="□")
            c.fill = fill(bg)
            c.font = font(size=12)
            c.alignment = align("center", "center")
            c.border = border()
        input_cell(ws, r, 5, "", bg=bg)
        ws.row_dimensions[r].height = 18
        r += 1
    return r

r = eval_section(ws8, r, "一、课程内容满意度", [
    "课程内容符合我的预期需求",
    "课程内容在AI时代背景下具有实用性",
    "双峰教练视角（人类深度+AI杠杆）对我有启发",
    "GUIDE模型清晰易学",
    "DIRECT模型清晰易学",
    "AI时代六类场景分析有参考价值",
    "课程结构逻辑清晰",
])

r = eval_section(ws8, r, "二、讲师满意度", [
    "讲师对教练示范的质量高",
    "讲师能有效处理课堂中的情绪和挑战性对话",
    "讲师对AI时代主题有真实认知和经验",
    "讲师反馈有深度、有帮助",
    "讲师在演练中的介入时机恰当",
])

r = eval_section(ws8, r, "三、教学方法满意度", [
    "体验式活动（开场体验、场景演练）设计有效",
    "配对演练帮助我实践了教练技巧",
    "三人组演练提供了多角度反馈",
    "反馈复盘环节对我的成长有帮助",
    "课程时间安排合理",
])

apply_section(ws8, r, "四、开放性问题", 5)
r += 1
for q in ["本次课程最有价值的模块是：", "我最希望改进的模块是：",
          "回到工作岗位后，我计划立刻应用的是：", "其他建议或反馈："]:
    ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws8.cell(row=r, column=1, value=q)
    c.fill = fill(CLR["label_bg"])
    c.font = font(bold=True, color="1F4E79")
    c.alignment = align("left", "center")
    r += 1
    ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    input_cell(ws8, r, 1)
    ws8.row_dimensions[r].height = 28
    r += 1

ws8.freeze_panes = "A5"
print("Sheet 8 done")

# ==============================================================
# SHEET 9: 30天跟进表
# ==============================================================
ws9 = wb.create_sheet("30天跟进表")
set_col_widths(ws9, [14, 14, 28, 28, 28, 10])
r = 1
apply_title(ws9, r, "登攀者 - AI时代的授权赋能领导力  30天跟进表", 6, height=30)
r += 1
ws9.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws9.cell(row=r, column=1, value="教练者：________________  被教练者：________________  跟进周期：第一次到第二次")
c.fill = fill(CLR["row_alt"])
c.font = font(size=10)
c.alignment = align("center", "center")
ws9.row_dimensions[r].height = 20
r += 1

apply_section(ws9, r, "教练对话记录（第1次跟进）", 6)
r += 1
apply_header(ws9, r, ["日期", "时间", "对话主要内容记录", "关键洞察/顿悟时刻", "行动计划确认", "时 长"])
r += 1
for col in range(1, 7):
    input_cell(ws9, r, col)
ws9.row_dimensions[r].height = 50
r += 1

apply_section(ws9, r, "关键进展记录", 6)
r += 1
apply_header(ws9, r, ["", "进展描述", "支持证据/具体行为", "AI时代价值体现", "备注", ""])
r += 1
progress_areas = ["人类深度发展方面", "AI杠杆使用改进方面", "关键判断能力的可见化"]
for i, area in enumerate(progress_areas):
    bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
    label_cell(ws9, r, 1, area, bg=bg, bold=True)
    for col in range(2, 6):
        input_cell(ws9, r, col, "", bg=bg)
    ws9.row_dimensions[r].height = 32
    r += 1

apply_section(ws9, r, "挑战与解决方案", 6)
r += 1
apply_header(ws9, r, ["遇到的挑战", "我的应对策略", "效果评估", "学到了什么", "下次如何调整", ""])
r += 1
for _ in range(3):
    for col in range(1, 6):
        input_cell(ws9, r, col)
    ws9.row_dimensions[r].height = 36
    r += 1

apply_section(ws9, r, "下一步计划", 6)
r += 1
for label in ["下一个30天的核心目标：", "具体行动计划（双轨）：", "需要的支持或资源：", "下次跟进时间："]:
    label_cell(ws9, r, 1, label, bold=True)
    ws9.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    input_cell(ws9, r, 2)
    ws9.row_dimensions[r].height = 24 if "行动计划" not in label else 36
    r += 1

ws9.freeze_panes = "A4"
print("Sheet 9 done")

# ==============================================================
# SHEET 10: AI时代场景分析表
# ==============================================================
ws10 = wb.create_sheet("AI时代场景分析表")
set_col_widths(ws10, [8, 18, 22, 20, 20, 26])
r = 1
apply_title(ws10, r, "登攀者 - AI时代的授权赋能领导力  AI时代场景分析表", 6, height=30)
r += 1
ws10.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws10.cell(row=r, column=1, value="六类AI时代教练场景 - 适用模型判断 + 关键引导点 + 演练记录")
c.fill = fill(CLR["green_bg"])
c.font = font(size=10, bold=True)
c.alignment = align("center", "center")
ws10.row_dimensions[r].height = 20
r += 1

apply_section(ws10, r, "六类AI时代教练场景概览", 6)
r += 1
apply_header(ws10, r, ["场景", "类型", "典型信号/语言", "适用模型", "核心引导点", "演练记录"])
r += 1

scenarios10 = [
    ("A", "价值焦虑型",
     "AI也能做这个了...\n我也不知道这有什么意义...\n我觉得我还能在这里做多久？",
     "GUIDE\n（发展型）",
     "帮助看到AI做不到的具体价值；找到只有我才能做到的那部分；反向视角：担心的那部分反倒是最有深度的。", ""),
    ("B", "过度依赖型",
     "我让AI先做...\nAI说...\n说不清楚自己判断的逻辑在哪里。",
     "DIRECT\n（业绩型）",
     "归因分析：你的判断在哪里？长期能力影响：积累深度判断经验的机会正在被替代。重新设计人机协作边界。", ""),
    ("C", "无效抵触型",
     "那些工具没什么用...\n真正的工程师靠自己...",
     "GUIDE或\nDIRECT皆可",
     "温和打开可能性：如果AI帮你处理了X，你节省下来的时间会用来做什么？观察优秀专业人士如何对待新工具。", ""),
    ("D", "接受更大责任",
     "我不是技术出身，AI这块我不行。\n但项目真正需要的是系统整合和利益相关方管理能力。",
     "GUIDE\n（发展型）",
     "帮助区分表面障碍和真实能力；探索人类判断在这件事里的具体体现；爬到人类深度峰。", ""),
    ("E", "克服失望情绪",
     "以前觉得自己有独特的创意能力，现在感觉AI比我强。\n投入度明显下降。",
     "GUIDE\n（发展型）",
     "帮助重新定义创意能力；AI处理的是框架，真正的判断、直觉、关系理解是人类的；探索创意工作中AI做不到的那部分。", ""),
    ("F", "处理业绩缺陷",
     "产出量达标，但核心产出大量是AI直接生成的，判断和修改微乎其微。客户追问时支撑不了。",
     "DIRECT\n（业绩型）",
     "D步骤：具体行为描述+归因分析；I步骤：长期能力影响；R步骤：双维度期望（产出+人类贡献）；C步骤：人类贡献可见化的行动承诺。", ""),
]
for i, (scene, stype, signal, model, guide, record) in enumerate(scenarios10):
    bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
    label_cell(ws10, r, 1, scene, bg=bg, bold=True, align_h="center")
    label_cell(ws10, r, 2, stype, bg=bg, bold=True)
    label_cell(ws10, r, 3, signal, bg=bg)
    label_cell(ws10, r, 4, model, bg=bg, align_h="center")
    label_cell(ws10, r, 5, guide, bg=bg)
    input_cell(ws10, r, 6, record, bg=CLR["input_bg"])
    ws10.row_dimensions[r].height = 60
    r += 1

apply_section(ws10, r, "场景判断练习", 6)
r += 1
practice_scenarios = [
    "一名产品负责人在使用AI工具后效率提升了三倍，但主管问起背后的判断逻辑时，她说不清楚。",
    "一名数据分析师最近在团队复盘时明显变得沉默，说反正AI都能做了，我说的也不重要。",
    "一名工程师拒绝使用任何AI辅助工具，认为真正的能力都是自己练出来的。",
]
for i, desc in enumerate(practice_scenarios):
    bg = CLR["row_alt"] if i % 2 == 0 else "FFFFFF"
    label_cell(ws10, r, 1, "场景{}：".format(i+1), bold=True)
    ws10.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    input_cell(ws10, r, 2, desc, bg=bg)
    ws10.row_dimensions[r].height = 28
    r += 1

apply_section(ws10, r, "我的学习总结", 6)
r += 1
for q in ["通过六类场景的学习，我最大的收获是：", "在真实工作中，我最可能先应用哪一个场景？为什么？"]:
    ws10.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws10.cell(row=r, column=1, value=q)
    c.fill = fill(CLR["label_bg"])
    c.font = font(bold=True, color="1F4E79")
    c.alignment = align("left", "center")
    r += 1
    ws10.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    input_cell(ws10, r, 1)
    ws10.row_dimensions[r].height = 36
    r += 1

ws10.freeze_panes = "A4"
print("Sheet 10 done")

# ==============================================================
# Page setup and tab colors
# ==============================================================
colors = ["1F4E79","2E75B6","2E75B6","4472C4","4472C4","70AD47","70AD47","ED7D31","ED7D31","FFC000"]
for i, ws in enumerate([ws1,ws2,ws3,ws4,ws5,ws6,ws7,ws8,ws9,ws10]):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.tabColor = colors[i]

wb.save(OUTPUT_PATH)
print("\n{}  saved to:\n  {}".format("=" * 50, OUTPUT_PATH))

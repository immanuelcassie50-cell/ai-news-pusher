# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# Sheet 1: 学员痛点
ws1 = wb.active
ws1.title = '学员痛点'
headers1 = ['痛点编号', '痛点描述', '出现场景', '根本原因', '当前应对方式', '课程解决方案']
data1 = [
    ['1', '方向感丧失', '战略规划会议、年度复盘', '成功后的舒适区、缺乏危机感', '等待总部指令、参与更多会议', '通过「志向设立」模块找到个人北极星'],
    ['2', '决策焦虑', '产品策略、资源分配、团队整合', '信息过载、过度分析、害怕承担责任', '反复讨论、拖延决策、让集体决策', '通过「艰难抉择」框架建立决策信心'],
    ['3', '信任危机', '跨部门协作、授权下属、重要人才流失', '角色距离感、绩效导向、缺乏深度对话', '更多团建、增加一对一沟通', '通过「信任资产负债表」重建信任'],
    ['4', '担责困境', '危机处理、失败项目复盘、下属犯错', '保护下属 vs 培养下属的矛盾', '全部自己扛 or 全部推给下属', '通过「责任担当」四象限明确边界'],
]
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
header_font = Font(bold=True)
col_widths1 = [12, 36, 28, 26, 30, 36]
for col, (h, w) in enumerate(zip(headers1, col_widths1), 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    ws1.column_dimensions[get_column_letter(col)].width = w
for row_idx, row_data in enumerate(data1, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws1.cell(row=row_idx, column=col_idx, value=val)

# Sheet 2: 组织需求
ws2 = wb.create_sheet('组织需求')
headers2 = ['需求类型', '具体需求', '业务背景', '期望产出', '优先级']
data2 = [
    ['战略执行', '中高层管理者能够独立承担战略落地责任', '集团战略到执行层的传导不畅', '能说清楚「我的业务单元如何支撑集团战略」', '高'],
    ['人才梯队', '建立稳定的管理人才输送通道', '关键岗位过度依赖个别明星员工', '识别并培养3-5名高潜人才', '高'],
    ['文化传承', '将企业价值观真正落地而非停留在墙上的标语', '价值观培训效果差、言行不一', '能在真实决策场景中体现价值观', '中'],
    ['变革领导力', '在不确定性增加的环境中带领团队持续前进', '业务转型、组织调整、人员优化', '能带领团队穿越至暗时刻', '高'],
]
col_widths2 = [14, 38, 28, 38, 10]
for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    ws2.column_dimensions[get_column_letter(col)].width = w
for row_idx, row_data in enumerate(data2, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws2.cell(row=row_idx, column=col_idx, value=val)

# Sheet 3: 课程目标
ws3 = wb.create_sheet('课程目标')
headers3 = ['层次', '目标描述', '具体行为表现', '评估方式']
data3 = [
    ['记忆', '记住领导力四力模型的核心框架', '能完整说出「志向设立-艰难抉择-信任建立-责任担当」四力模型', '课后测试'],
    ['理解', '能用自己的话解释每个能力的内涵', '能向他人解释「什么是真正的志向设立」', '小组分享'],
    ['应用', '能将四力框架应用于分析真实管理场景', '能用框架分析自己面临的一个两难困境', '案例作业'],
    ['分析', '能识别他人决策背后的四力考量', '能点评课堂案例中领导者的四力表现', '角色扮演'],
    ['评价', '能评估不同决策选项的四力影响', '能在两难选择中做出有意识的权衡', '小组决策模拟'],
    ['创造', '能设计属于自己的领导力实践路径', '能制定个人领导力发展计划并承诺行动', '个人行动计划'],
]
col_widths3 = [12, 38, 38, 20]
for col, (h, w) in enumerate(zip(headers3, col_widths3), 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    ws3.column_dimensions[get_column_letter(col)].width = w
for row_idx, row_data in enumerate(data3, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws3.cell(row=row_idx, column=col_idx, value=val)

out = 'D:/新课开发/领导力/01-领导力重生：AI时代真正稀缺的领导者内核/成果demo/KnowledgeShareEvent_Demo/01_Needs_Analysis.xlsx'
wb.save(out)
print(f'Saved to {out}')
import os
print(f'File size: {os.path.getsize(out)} bytes')

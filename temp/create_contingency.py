import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "触发条件与应急方案"

header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
data_font = Font(name='微软雅黑', size=10)
title_font = Font(name='微软雅黑', bold=True, size=14)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws.merge_cells('A1:H1')
ws['A1'] = '华东制造集团ERP升级项目 - 触发条件与应急方案'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:H2')
ws['A2'] = '版本：V1.0    编制人：李晓峰    更新日期：2024年7月20日'
ws['A2'].font = Font(name='微软雅黑', size=10, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')
ws.row_dimensions[2].height = 20

headers = ['假设编号', '假设描述', '触发条件', '触发指标', '监测责任人', '应急方案', '应急权限人', '备注']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[4].height = 35

contingencies = [
    ['H001', '现有SAP R/3系统数据可以完整迁移至S/4HANA', '数据迁移测试失败率超过1%', '迁移测试错误率 > 1%', '陈建国', '1.立即停止迁移，分析失败原因\n2.评估数据修复可能性和周期\n3.如无法修复，启动数据重新清洗流程\n4.向项目总监汇报，评估对项目进度的影响', '张明华', '需要准备数据修复工具和回滚方案'],
    ['H002', '现有服务器配置可以满足S/4HANA性能要求', '压力测试响应时间超过标准值的150%', '页面响应时间 > 7.5秒', '刘建华', '1.分析性能瓶颈（CPU/内存/磁盘IO）\n2.如可通过优化解决，进行系统调优\n3.如需硬件升级，提交采购申请\n4.评估对上线计划的影响', '张明华', '提前与采购部沟通可能的硬件采购流程'],
    ['H003', '与MES系统的接口可以在6周内完成开发', '接口开发周期超过8周', '接口开发实际工时 > 480小时', '刘建华', '1.评估是否可以简化接口方案\n2.协调更多开发资源支持\n3.如影响核心功能，向客户说明并协商范围调整\n4.启动备选接口方案（文件交换模式）', '李晓峰', '已准备备选方案（文件交换模式）'],
    ['H005', '各子公司愿意放弃本地流程，接受集团统一流程', '任何一家子公司明确表示抵制统一流程', '子公司总经理会议纪要明确反对', '赵敏', '1.立即向项目总监汇报\n2.组织专项沟通会，了解抵制原因\n3.评估是否需要调整统一流程方案\n4.如无法协调，启动范围变更流程', '王建业', '需要在启动会后尽快完成子公司访谈'],
    ['H006', '关键用户可以投入足够时间配合实施', '关键用户实际投入时间低于承诺的80%', '关键用户月度工时 < 承诺工时的80%', '李晓峰', '1.与客户方项目经理协商调整关键用户\n2.评估是否需要延期关键里程碑\n3.向客户方高层反馈，请求支持协调\n4.考虑增加外部顾问驻场时间', '陈永强', '建立关键用户工时跟踪机制'],
    ['H007', '现有物料编码体系可以标准化而不需要重新编码', '数据质量分析显示不匹配率超过20%', '物料数据不匹配率 > 20%', '陈建国', '1.评估重新编码的工作量和周期\n2.与客户方协商编码方案（渐进式 vs 一次性）\n3.申请额外预算和时间\n4.制定数据清洗详细计划', '张明华', '已初步分析，可能存在30%不匹配'],
    ['H010', '预算800万可以覆盖全部实施范围', '月度成本超出预算10%以上', '月度成本偏差 > 10%', '李晓峰', '1.立即分析超支原因\n2.识别可削减的范围或质量\n3.与客户方协商变更（减少范围或增加预算）\n4.如无法达成，启动合同变更流程', '张明华', '建立成本预警机制，每月跟踪'],
    ['H013', '12个月的实施周期足够完成全部功能', '详细工作量评估显示需要超过12个月', '评估工作量 > 5760人天', '李晓峰', '1.识别可以削减的非核心功能\n2.评估增加资源的可能性（成本影响）\n3.与客户方协商调整项目范围或周期\n4.制定分批上线方案', '王建业', '已有延期预警，需尽快启动评估'],
    ['H014', '业务调研可以在6周内完成', '调研实际用时超过8周', '调研实际用时 > 40个工作日', '赵敏', '1.分析延期原因，识别耽误的关键路径\n2.评估是否可以压缩后续设计阶段\n3.与客户方协商调整里程碑\n4.增加资源投入以加快进度', '李晓峰', '已出现延期，需立即采取行动'],
    ['H019', '客户方项目对接人员有足够的决策权限', '日常决策需要升级超过3次/周', '周决策升级次数 > 3次', '李晓峰', '1.与陈永强明确日常决策权限边界\n2.建立明确的升级机制和流程\n3.如权限限制影响进度，正式报告项目总监\n4.必要时安排与王建业的直接沟通', '张明华', '已在启动会发现此问题'],
]

for row_idx, row_data in enumerate(contingencies, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row_idx].height = 80

column_widths = [12, 30, 30, 20, 12, 40, 15, 20]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

output_path = 'D:/CC/新课开发/工作手册/假设管理：项目经理的风险前置手册/完整课程包/07-成果demo/触发条件与应急方案-演示.xlsx'
wb.save(output_path)
print(f"OK: {output_path}")

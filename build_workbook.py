from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = Workbook()

# Define styles
header_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
header_font = Font(name="Microsoft YaHei", bold=True, size=11)
data_font = Font(name="Microsoft YaHei", size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_header(ws, row, values):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def set_data_row(ws, row, values, even_row=False):
    fill_color = "FFFFFF" if not even_row else "DCE6F1"
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

# Sheet 1: 学员信息
ws1 = wb.active
ws1.title = "学员信息"
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 20
ws1.column_dimensions['F'].width = 25
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 35
ws1['A1'] = "学员信息"
ws1['A1'].font = Font(name="Microsoft YaHei", bold=True, size=14)
set_header(ws1, 2, ["序号", "姓名", "部门/岗位", "带教经验年限", "授课经验", "备注"])
for i in range(3, 11):
    ws1.row_dimensions[i].height = 25
    set_data_row(ws1, i, [str(i-2), "", "", "", "", ""], even_row=(i%2==0))
ws1.freeze_panes = 'A3'

# Sheet 2: 练习记录
ws2 = wb.create_sheet("练习记录")
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 30
ws2.column_dimensions['F'].width = 15
ws2.column_dimensions['G'].width = 20
ws2.row_dimensions[1].height = 30
ws2.row_dimensions[2].height = 35
ws2['A1'] = "练习记录"
ws2['A1'].font = Font(name="Microsoft YaHei", bold=True, size=14)
set_header(ws2, 2, ["环节名称", "练习小组", "学员姓名", "场景描述（是否具体）", "点评内容（是否具体）", "讲师点评时间", "备注"])
for i in range(3, 11):
    ws2.row_dimensions[i].height = 25
    set_data_row(ws2, i, ["", "", "", "", "", "", ""], even_row=(i%2==0))
ws2.freeze_panes = 'A3'

# Sheet 3: 行动卡
ws3 = wb.create_sheet("行动卡")
ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 25
ws3.column_dimensions['E'].width = 25
ws3.column_dimensions['F'].width = 15
ws3.column_dimensions['G'].width = 20
ws3.column_dimensions['H'].width = 25
ws3.row_dimensions[1].height = 30
ws3.row_dimensions[2].height = 35
ws3['A1'] = "行动卡"
ws3['A1'].font = Font(name="Microsoft YaHei", bold=True, size=14)
set_header(ws3, 2, ["序号", "学员姓名", "计划带的真实场景", "计划怎么演", "计划怎么点", "计划实施时间", "实际完成情况", "反思与总结"])
for i in range(3, 11):
    ws3.row_dimensions[i].height = 25
    set_data_row(ws3, i, ["", "", "", "", "", "", "", ""], even_row=(i%2==0))
ws3.freeze_panes = 'A3'

# Sheet 4: 讲师自检
ws4 = wb.create_sheet("讲师自检")
ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 55
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 25
ws4.row_dimensions[1].height = 30
ws4.row_dimensions[2].height = 35
ws4['A1'] = "讲师自检"
ws4['A1'].font = Font(name="Microsoft YaHei", bold=True, size=14)
set_header(ws4, 2, ["检查项目", "检查内容", "完成状态", "备注"])

check_items = [
    ("课前准备", "提前想好一个真实场景用于方法论传递"),
    ("课前准备", "读过应急工具包至少三个场景"),
    ("课前准备", "确认场地允许学员自由走动"),
    ("课前准备", "准备好收尾用的空白卡片"),
    ("课前准备", "在心里过了想演换点四步和三个不"),
    ("课前准备", "计时器准备就绪"),
    ("课程导入", "方法论讲解清晰"),
    ("互动环节", "互动环节把控到位"),
    ("收尾环节", "收尾环节完整"),
]
for i, (a, b) in enumerate(check_items, 3):
    ws4.row_dimensions[i].height = 25
    set_data_row(ws4, i, [a, b, "", ""], even_row=(i%2==0))
ws4.freeze_panes = 'A3'

# Sheet 5: 时间控制
ws5 = wb.create_sheet("时间控制")
ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 15
ws5.column_dimensions['C'].width = 15
ws5.column_dimensions['D'].width = 12
ws5.column_dimensions['E'].width = 12
ws5.column_dimensions['F'].width = 10
ws5.column_dimensions['G'].width = 25
ws5.row_dimensions[1].height = 30
ws5.row_dimensions[2].height = 35
ws5['A1'] = "时间控制"
ws5['A1'].font = Font(name="Microsoft YaHei", bold=True, size=14)
set_header(ws5, 2, ["环节", "开始时间", "结束时间", "实际时长", "建议时长", "是否超时", "备注"])

time_items = [
    ("想：认知导入（为什么要带教）", "", "", "", "15分钟", "", ""),
    ("演：示范展示（完整带教演示）", "", "", "", "30分钟", "", ""),
    ("换：练习转换（学员实操演练）", "", "", "", "45分钟", "", ""),
    ("点：点评升华（方法论提炼）", "", "", "", "30分钟", "", ""),
    ("课程导入", "", "", "", "10分钟", "", ""),
    ("方法论讲解", "", "", "", "20分钟", "", ""),
    ("学员练习", "", "", "", "60分钟", "", ""),
    ("综合点评", "", "", "", "30分钟", "", ""),
]
for i, row_data in enumerate(time_items, 3):
    ws5.row_dimensions[i].height = 25
    set_data_row(ws5, i, row_data, even_row=(i%2==0))
ws5.freeze_panes = 'A3'

# Sheet 6: 点评话术
ws6 = wb.create_sheet("点评话术")
ws6.column_dimensions['A'].width = 15
ws6.column_dimensions['B'].width = 15
ws6.column_dimensions['C'].width = 50
ws6.column_dimensions['D'].width = 35
ws6.row_dimensions[1].height = 30
ws6.row_dimensions[2].height = 35
ws6['A1'] = "点评话术"
ws6['A1'].font = Font(name="Microsoft YaHei", bold=True, size=14)
set_header(ws6, 2, ["场景", "点评类型", "话术模板", "使用说明"])

comment_items = [
    ("通用", "追问具体", "具体是哪一次？当时发生了什么？", "当学员场景描述过于抽象时，追问细节"),
    ("通用", "肯定做对的地方", "你刚才XX这个地方做得很好，因为……", "肯定具体行为，帮助学员建立信心"),
    ("通用", "指出可以调整的地方", "如果换成XX方式，会不会更好？", "用建议的方式提出，不要直接否定"),
    ("通用", "三句话点评模板", "刚才你做了XX（第1句），这里做得很好（第2句），下次如果注意XX会更好（第3句）", "适用于所有点评场景"),
    ("通用", "应急情况点评", "没关系，我们每个人都是这样过来的", "当学员表现紧张或不自信时使用"),
]
for i, row_data in enumerate(comment_items, 3):
    ws6.row_dimensions[i].height = 30
    set_data_row(ws6, i, row_data, even_row=(i%2==0))
ws6.freeze_panes = 'A3'

# Sheet 7: 应急话术
ws7 = wb.create_sheet("应急话术")
ws7.column_dimensions['A'].width = 8
ws7.column_dimensions['B'].width = 22
ws7.column_dimensions['C'].width = 22
ws7.column_dimensions['D'].width = 40
ws7.column_dimensions['E'].width = 30
ws7.row_dimensions[1].height = 30
ws7.row_dimensions[2].height = 35
ws7['A1'] = "应急话术"
ws7['A1'].font = Font(name="Microsoft YaHei", bold=True, size=14)
set_header(ws7, 2, ["场景编号", "场景描述", "表现", "应对话术", "注意事项"])

emergency_items = [
    ("1", "开场对比环节学员反应冷淡", "学员不积极发言，气氛沉闷", "我来分享一个我的经历……（讲师先示范一个失败的场景）", "不要硬抛问题，先用自己的经历暖场"),
    ("2", "学员讲的场景太抽象", "学员说就是沟通有问题等笼统描述", "能不能说具体一点？比如，上周有没有发生过类似的事？", "用上周等具体时间框架引导"),
    ("3", "学员讲成了表演", "学员讲述像是背台词，缺乏真实细节", "停，我感觉到你在讲道理，能不能用一个真实的例子？", "温柔打断，引导回归真实"),
    ("4", "学员讲成了正确的废话", "学员说要耐心、要倾听等正确但无内容的话", "这些道理我们都懂，具体当时你说了什么、做了什么？", "追问具体行为动作"),
    ("5", "没人愿意先上台", "沉默，无人主动", "我看到有人在跃跃欲试，这样，我找一位看起来准备好的（眼神接触）", "不要等，主动点将，讲师是导演"),
    ("6", "学员演练效果不理想", "学员紧张或方法完全不对", "没关系，这就是练习的意义。我们再来一次，这次我给你做个示范", "先肯定勇气，再给方法"),
    ("7", "时间严重超时", "某个环节占用过多时间", "我们先到这里，总结一下你刚才学到的……（强制收尾）", "讲师要果断，不要让一个学员拖累全班"),
    ("8", "学员质疑与案例教学的区别", "学员问这和普通案例分析有什么区别", "好问题。区别在于，我们不是分析别人的案例，而是带你用自己的场景练一遍", "强调自己的场景和练一遍"),
    ("9", "讲师自己担心示范讲不好", "讲师紧张，怕出丑", "记住：你的价值不是表演完美，而是陪他练一遍", "降低对自己的要求，聚焦陪伴价值"),
]
for i, row_data in enumerate(emergency_items, 3):
    ws7.row_dimensions[i].height = 30
    set_data_row(ws7, i, row_data, even_row=(i%2==0))
ws7.freeze_panes = 'A3'

# Sheet 8: 质量追踪
ws8 = wb.create_sheet("质量追踪")
ws8.column_dimensions['A'].width = 15
ws8.column_dimensions['B'].width = 35
ws8.column_dimensions['C'].width = 20
ws8.column_dimensions['D'].width = 35
ws8.row_dimensions[1].height = 30
ws8.row_dimensions[2].height = 35
ws8['A1'] = "质量追踪"
ws8['A1'].font = Font(name="Microsoft YaHei", bold=True, size=14)
set_header(ws8, 2, ["追踪维度", "具体指标", "评估结果", "改进建议"])

quality_items = [
    ("课程内容", "方法论讲解清晰度"),
    ("课程内容", "方法论逻辑是否完整"),
    ("课程内容", "案例选择是否贴近实际"),
    ("时间把控", "各环节时间分配是否合理"),
    ("时间把控", "是否有超时情况"),
    ("学员参与度", "练习环节参与率"),
    ("学员参与度", "学员互动积极性"),
    ("讲师表现", "点评是否具体有效"),
    ("讲师表现", "现场把控能力"),
    ("学员收获", "学员是否能说清带教步骤"),
    ("学员收获", "学员是否有清晰的行动计划"),
    ("综合评分", "学员整体满意度（10分制）"),
]
for i, (a, b) in enumerate(quality_items, 3):
    ws8.row_dimensions[i].height = 25
    set_data_row(ws8, i, [a, b, "", ""], even_row=(i%2==0))
ws8.freeze_panes = 'A3'

# Save
output_dir = "D:/2026年课程/培训师带教/完整课程包/07-全流程工具表单"
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "工具表单.xlsx")
wb.save(output_path)
print(f"Saved to {output_path}")
